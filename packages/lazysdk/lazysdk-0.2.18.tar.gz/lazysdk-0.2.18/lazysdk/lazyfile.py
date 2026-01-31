#!/usr/bin/env python3
# coding = utf8
"""
@ Author : ZeroSeeker
@ e-mail : zeroseeker@foxmail.com
@ GitHub : https://github.com/ZeroSeeker
@ Gitee : https://gitee.com/ZeroSeeker
"""
import copy

from . import lazypath
import collections
import subprocess
import requests
import platform
import datetime
import openpyxl
import showlog
import time
import xlrd
import json
import sys
import os
# import zipfile
from rich.progress import Progress
from requests import exceptions

headers_default = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:68.0) Gecko/20100101 Firefox/68.0"}


def delete(
        file,
        postfix: str = None,
        path: str = None,
):
    """
    删除文件
    """
    if path:
        # 如果指定了路径，就加上路径
        file_dir = os.path.join(path, file)
    else:
        # 如果没指定路径，就直接使用文件名
        file_dir = file

    if postfix:
        # 如果指定了后缀名，就加上后缀名
        file_dir = f'{file}.{postfix}'
    else:
        # 如果没指定后缀名，就忽略
        pass
    os.remove(file_dir)


def get_suffix(
        file_name: str
):
    return file_name.split('.')[-1]  # 获取文件后缀名


def file_rename(
        before_name: str,  # 原名称
        after_name: str  # 重命名名称
):
    """
    重命名文件，注意是完整文件路径
    :param before_name:
    :param after_name:
    :return:
    """
    try:
        os.rename(before_name, after_name)
        return True
    except Exception as e:
        print(e)
        print('rename file fail\r\n')
        return False


def open_folder(
        folder: str = os.path.dirname(os.path.abspath(__file__))  # 默认为当前路径
):
    """
    功能：打开路径
    """
    if platform.system() == 'Windows':
        os.startfile(folder)  # Windows上打开文件
    else:
        subprocess.check_call(['open', folder])  # 非Windows上打开文件


def download(
        url,
        filename: str = None,
        suffix_name: str = None,
        headers: dict = None,
        path: str = None,
        proxies=None,
        size_limit: int = None,
        range_start: int = None,
        range_end: int = None,
        overwrite: bool = False,
        verify: bool = True
):
    """
    实现文件下载功能，可指定url、文件名、后缀名、请求头、文件保存路径
    :param url:
    :param filename:文件名
    :param suffix_name:后缀名
    :param headers:请求头
    :param path:文件保存路径
    :param proxies:代理
    :param size_limit:尺寸限制
    :param range_start:开始位置
    :param range_end:结束位置
    :param overwrite: 覆盖
    :param verify: 验证证书
    :return:
    """
    if not headers:
        headers_local = headers_default
    else:
        headers_local = headers

    if not range_start and not range_end:
        range_start = 0
        range_info = None
    elif range_start is not None and range_end is None:
        range_info = 'bytes=%d-' % range_start  # 从这里向后
    elif range_start is None and range_end is not None:
        range_start = 0
        range_info = 'bytes=0-%d' % range_end
    else:
        range_info = 'bytes=%d-%d' % (range_start, range_end)

    if not range_info:
        pass
    else:
        headers_local['Range'] = range_info
    # 获取文件的基本信息
    response = requests.get(
        url=url,
        headers=headers_local,
        stream=True,
        proxies=proxies,
        verify=verify
    )
    # print("response:", response.headers)
    total_length = response.headers.get('content-length', '0')  # 文件大小
    content_type = response.headers.get('content-type')  # 文件类型
    content_disposition = response.headers.get('content-disposition')  # 文件名及类型
    filename_default = 'unknown_' + str(time.time())
    if content_disposition is not None:
        content_dispositions = content_disposition.replace(' ', '').split(';')
        for each_content_disposition in content_dispositions:
            if 'filename=' in each_content_disposition:
                each_content_disposition_split = each_content_disposition.split(sep='=', maxsplit=1)  # 只拆分一次，防止有多个=影响
                filename_default_full = each_content_disposition_split[1].replace('"', '')
                filename_default = filename_default_full[:filename_default_full.rfind('.')]  # 解析文件名
                suffix_name = filename_default_full[filename_default_full.rfind('.')+1:]  # 解析文件后缀
            else:
                pass
    else:
        pass
    # print("filename_default:", filename_default)
    if suffix_name is None:
        # 尝试自动获取文件后缀名
        suffix_name = content_type.split('/')[1]

    if filename is None:
        download_file_name = str(filename_default) + "." + str(suffix_name)
    else:
        if filename.endswith(suffix_name):
            # 如果文件名中已存在后缀名，则不重复添加
            download_file_name = filename
        else:
            download_file_name = str(filename) + "." + str(suffix_name)

    if path is None:
        path_local = download_file_name
    else:
        if os.path.exists(path):
            pass
        else:
            os.makedirs(path)
        path_local = os.path.join(path, download_file_name)

    lazypath.make_path(lazypath.file_path(path_local))

    if not range_start:
        temp_size = 0  # 已经下载文件大小
    else:
        temp_size = range_start + 0  # 已经下载文件大小
    chunk_size = 1024 * 1024  # 分割文件大小，字节B
    total_size = int(total_length)  # 文件总大小
    # total_size_mb = round(total_size / (1024 * 1024), 2)  # 换算到MB的文件大小
    # 添加文件大小控制，跳过下载超大文件
    if size_limit is None:
        pass
    else:
        if total_size > size_limit:
            return
        else:
            pass

    is_finish = False

    if overwrite and os.path.exists(path_local):
        delete(file=path_local)
    elif not overwrite and os.path.exists(path_local):
        # 这里将对已存在的文件名重新命名
        rename_count = 1
        while True:
            download_file_name_rename = f'{download_file_name.split(".")[0]}({rename_count}).{download_file_name.split(".")[1]}'
            if path:
                path_local_rename = os.path.join(path, download_file_name_rename)
            else:
                path_local_rename = download_file_name_rename
            if os.path.exists(path_local_rename):
                rename_count += 1
                continue
            else:
                break
        path_local = copy.deepcopy(path_local_rename)
    else:
        pass

    with open(path_local, "ab") as f:  # wb新建文件，a追加
        with Progress() as progress:
            task = progress.add_task(description="[red]Downloading...", total=total_size)
            for chunk in response.iter_content(chunk_size=chunk_size):
                try:
                    if not chunk:
                        if temp_size >= total_size:
                            is_finish = True
                        # else:
                        #     is_finish = False
                        break
                    else:
                        f.write(chunk)
                        f.flush()
                        progress.update(task, advance=chunk_size)
                        if temp_size >= total_size:
                            is_finish = True
                        else:
                            # is_finish = False
                            temp_size += chunk_size
                except:
                    showlog.error('')
    print("\n  ==> 文件已全部下载完成，保存位置:", path_local)
    res_dict = {
        'file_dir': path_local,
        'is_finish': is_finish,
        'temp_size': temp_size
    }
    return res_dict


def safe_download(
        url,
        filename=None,
        suffix_name=None,
        headers=None,
        path="download",
        proxies=None,
        size_limit=None,
        range_start=None,
        range_end=None,
        verify: bool = False
):
    while True:
        try:
            download_response = download(
                url=url,
                filename=filename,
                suffix_name=suffix_name,
                headers=headers,
                path=path,
                proxies=proxies,
                size_limit=size_limit,
                range_start=range_start,
                range_end=range_end,
                overwrite=False,
                verify=verify
            )
            if download_response.get('is_finish') is True:
                local_file_dir = download_response.get('file_dir')
                return local_file_dir
            else:
                print(':( 下载中断')
                range_start = download_response.get('temp_size')
                time.sleep(1)
                print('将继续下载（断点续传）...')
        except:
            print(':( 下载中断，将重新下载...')
            time.sleep(1)


def read(
        file,
        postfix: str = None,
        path: str = None,
        json_auto: bool = False,
        read_lines: bool = False
):
    """
    读取文件
    json_auto：json格式自动转换
    """
    if path:
        # 如果指定了路径，就加上路径
        file_dir = os.path.join(path, file)
    else:
        # 如果没指定路径，就直接使用文件名
        file_dir = file

    if postfix:
        # 如果指定了后缀名，就加上后缀名
        file_dir = f'{file}.{postfix}'
    else:
        # 如果没指定后缀名，就忽略
        pass

    if not os.path.exists(file_dir):
        return

    if read_lines:
        with open(file=file_dir, mode='r', encoding='utf-8') as f:
            content = f.readlines()
    else:
        f = open(file=file_dir, mode='r', encoding='utf-8')
        content = f.read()

    if content:
        if json_auto:
            if isinstance(content, str):
                return json.loads(content)
            else:
                json_content = list()
                for each_line in content:
                    json_content.append(json.loads(each_line))
                return json_content
        else:
            return content
    else:
        return content


def save(
        file,
        content,
        suffix: str = None,
        path: str = None,
        overwrite: bool = True,
        encoding: str = 'utf-8'
) -> str:
    """
    保存文件
    """
    if path:
        # 如果指定了路径，就加上路径
        lazypath.make_path(path)
        file_dir = os.path.join(path, file)
    else:
        # 如果没指定路径，就直接使用文件名
        file_dir = file

    if suffix:
        # 如果指定了后缀名，就加上后缀名
        file_dir = f'{file_dir}.{suffix}'
    else:
        # 如果没指定后缀名，就忽略
        pass

    if overwrite is True:
        write_mode = "w"  # 覆盖
    else:
        write_mode = "a"  # 追加

    f = open(
        file=file_dir,
        mode=write_mode,
        encoding=encoding
    )
    f.write(content)
    f.close()
    return file_dir


def read_(_source_file):
    # >>读取数据【方式：一次性全部读取】------------------------------------------------------------
    data = xlrd.open_workbook(_source_file)  # 打开表
    res = list()
    # table = data.sheets()[0]  # 默认使用第一张表格
    for table in data.sheets():
        nrows = table.nrows  # 获取行数
        ncols = table.ncols  # 获取列数
        for inrows in range(nrows):
            res_temp = dict()
            for incols in range(ncols):
                res_temp.update({table.cell(0, incols).value: table.cell(inrows, incols).value})
            # print(res_temp)
            res.append(res_temp)
    return res


def read_txt(text_name, path=None):
    try:
        if path is None:
            f = open("%s.txt" %text_name, encoding='utf-8')
        else:
            f = open("%s/%s.txt" % (path, text_name), encoding='utf-8')
        res = f.read()
        return res
    except:
        return


def read_file(file_name, suffix_name, path=None):
    try:
        if path is None:
            f = open("\%s.%s" % (file_name, suffix_name))
        else:
            f = open("\%s\%s.%s" % (path, file_name, suffix_name))
        res = f.read()
        return res
    except:
        return


def dir_file_list(file_dir):
    file_list = list()
    for root, dirs, files in os.walk(file_dir):
        # print(root)  # 当前目录路径
        # print(dirs)  # 当前路径下所有子目录
        # print(files)  # 当前路径下所有非目录子文件
        file_list.extend(files)
    return file_list


def get_file_size(file_dir):
    # 获取文件大小
    size_byte = os.path.getsize(file_dir)  # 字节
    size_kb = int(size_byte / 1024)
    size_mb = round(size_byte / (1024 * 1024), 2)
    size_gb = round(size_byte / (1024 * 1024 * 1024), 2)
    if size_gb < 1:
        if size_mb < 1:
            size_str = str(size_kb) + ' KB'
        else:
            size_str = str(size_mb) + ' MB'
    else:
        size_str = str(size_gb) + ' GB'
    temp_dict = {
        'size_str': size_str,
        'size_byte': size_byte
    }
    return temp_dict


class DateEncoder(json.JSONEncoder):
    # 处理json.dumps中会出现的datetime格式无法转换问题：json.dumps(each, cls=DateEncoder)
    def default(self, obj):
        if isinstance(obj, datetime.datetime):
            return obj.strftime("%Y-%m-%d %H:%M:%S")
        else:
            return json.JSONEncoder.default(self, obj)


def save_list(
        file_name,
        list_data,
        split_by='\n',
        postfix="txt",
        path=None,
        overwrite=False
):
    """
    将list按照行存储到文件中，对于datetime类型自动转换为日期时间的格式
    """
    if path is None:
        file_dir = "%s.%s" % (file_name, postfix)
    else:
        lazypath.make_path(path)
        file_dir = os.path.join(path, f'{file_name}.{postfix}')

    if overwrite is True:

        write_mode = "w"  # 覆盖
    else:
        write_mode = "a"  # 追加

    f = open(file_dir, write_mode, encoding='utf-8')
    for each in list_data:

        if isinstance(each, collections.OrderedDict):
            f.write(str(json.dumps(each, cls=DateEncoder)))
        else:
            f.write(str(each))
        f.write(split_by)
    f.close()
    return file_dir


def dict_write07excel(path, table_name, sheet_name, data):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = sheet_name
    if len(data) > 0:
        row_num = 1
        col_num = 1
        for key in data[0]:
            sheet.cell(row=row_num, column=col_num, value=key)
            col_num += 1
        row_num += 1

        for each in data:
            col_num = 1
            for key in each:
                sheet.cell(row=row_num, column=col_num, value=each.get(key))
                col_num += 1
            row_num += 1

        wb.save("%s/%s.xlsx" % (path, table_name))
        showlog.info("导出数据成功！共计%s条数据" % len(data))


def save_bytes(
        bytes_content: bytes,
        file: str
):
    """
    保存bytes为文件
    """
    with open(file, 'wb') as f:
        f.write(bytes_content)
    f.close()
    return os.path.join(sys.path[0], file)


def get_file_info(file_dir):
    """
    获取文件信息
    """
    from lazysdk import lazytime
    file_path = os.path.dirname(file_dir)  # 获取文件路径
    file_name = os.path.basename(file_dir)  # 获取不含路径信息的文件名
    file_name_pure, file_suffix = os.path.splitext(file_name)  # 获取不含后缀名的文件名和后缀名
    return {
        'name': file_name,  # 获取不含路径信息的文件名
        'path': file_path,
        'pure_name': file_name_pure,
        'suffix': file_suffix,
        'size': get_file_size(file_dir=file_dir),
        'create_timestamp': os.path.getctime(file_dir),
        'create_time': lazytime.get_timestamp2datetime(os.path.getctime(file_dir)),
        'modified_timestamp': os.path.getmtime(file_dir),
        'modified_time': lazytime.get_timestamp2datetime(os.path.getmtime(file_dir)),
    }


def zip_file(
        source_file: str,
        zip_file: str = None
) -> str:
    """
    压缩文件，此方法只适用于压缩单个文件，不适合压缩文件夹
    """
    import zipfile
    file_path = os.path.dirname(source_file)
    file_name = os.path.basename(source_file)
    file_name_pure, file_suffix = os.path.splitext(file_name)
    if not zip_file:
        zip_file = os.path.join(file_path, f'{file_name_pure}.zip')
    my_zip = zipfile.ZipFile(
        file=zip_file,  # 打开一个对象，也就是目标文件名
        mode='w'  # 写入模式
    )
    my_zip.write(
        filename=source_file,  # 压缩目标文件路径
        arcname=os.path.basename(source_file),  # 内部文件名
        compress_type=zipfile.ZIP_DEFLATED
    )
    my_zip.close()
    return zip_file


def zip_path(
        source_path
) -> str:
    """
    压缩文件夹，并在原来的同级目录输出，返回zip文件路径
    """
    import shutil
    return shutil.make_archive(
        root_dir=source_path,
        base_name=source_path,
        format='zip'
    )


def make_zip(
        file: str
) -> str:
    """
    创建压缩文件，自动判断源是文件还是路径
    """
    if os.path.isdir(file):
        return zip_path(file)
    elif os.path.isfile(file):
        return zip_file(file)
    else:
        return ''


def unzip(
        file
) -> list:
    """
    解压文件，会在源文件的同级目录增加一个路径，以存放解压后的文件
    返回的是一个已解压的所有文件路径的list
    """
    import zipfile
    import locale
    file_path = os.path.dirname(file)
    file_name = os.path.basename(file)
    file_name_pure, file_suffix = os.path.splitext(file_name)
    encoding = locale.getpreferredencoding()
    my_zip = zipfile.ZipFile(
        file=file,
        mode='r'
    )
    unzip_path = os.path.join(file_path, file_name_pure)
    while True:
        if os.path.exists(unzip_path):
            unzip_path = f'{unzip_path}_unpack'
        else:
            break

    # 生成一个源文件名对照正确文件名的字典
    name_dict = dict()
    for each in my_zip.filelist:
        each_filename = copy.deepcopy(each.filename)
        if each_filename[0:8] == '__MACOSX':
            continue  # 屏蔽mac下的__MACOSX路径
        else:
            if each_filename.isascii():
                pass   # 如果isascii=True，则不改编码
            else:
                try:
                    each_filename = each_filename.encode('437').decode(encoding)  # 尝试使用437编码先解码，再编码
                except:
                    pass
            name_dict[each.filename] = each_filename

    unzip_names = list()
    for each_name in my_zip.namelist():
        new_name = name_dict.get(each_name)
        if new_name:
            new_name_dir = os.path.join(unzip_path, new_name)
            # print("unzip_path:", unzip_path)
            # print("file_path:", file_path)
            extract_name = my_zip.extract(
                member=each_name,
                # path=file_path
            )
            if extract_name == new_name_dir:
                # print(extract_name)
                unzip_names.append(new_name_dir)
            else:
                # print(extract_name, '-->', new_name_dir)
                os.renames(old=extract_name, new=new_name_dir)  # 对乱码文件名重命名
                if os.path.isfile(new_name_dir):
                    unzip_names.append(new_name_dir)
        else:
            continue
    return unzip_names


def get_stream(
        url,
        verify=True,
        **kwargs
):
    """
    以流式传输方式发起请求，适用于下载文件时，当文件过大，不会立即下载。
    当把get函数的stream参数设置成True时，它不会立即开始下载，当你使用iter_content或iter_lines遍历内容或访问内容属性时才开始下载。需要注意一点：文件没有下载之前，它也需要保持连接。
    iter_content：一块一块的遍历要下载的内容
    iter_lines：一行一行的遍历要下载的内容
    """
    try:
        response = requests.get(
            url=url,
            verify=verify,
            stream=True,
            **kwargs
        )
    except exceptions.ConnectionError:
        raise exceptions.ConnectionError(f"Could not reach host. Are you offline?")
    return response
