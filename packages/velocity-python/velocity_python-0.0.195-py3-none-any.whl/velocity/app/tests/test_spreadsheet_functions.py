import unittest
import base64
from io import BytesIO
from openpyxl import load_workbook
from velocity.misc.export import (
    extract,
    autosize_columns,
    create_spreadsheet,
    get_downloadable_spreadsheet,
)


class TestSpreadsheetFunctions(unittest.TestCase):

    def test_extract(self):
        """Test extracting values from a dictionary based on a list of keys."""
        data = {"name": "Alice", "age": 30, "city": "Wonderland"}
        keys = ["name", "city", "nonexistent_key"]
        result = extract(data, keys)
        self.assertEqual(result, ["Alice", "Wonderland", None])

    def test_autosize_columns(self):
        """Test that the columns are autosized based on the content."""
        buffer = BytesIO()
        headers = ["Column1", "Column2"]
        rows = [["Short", "This is a longer text that should set the column width"]]

        # Create a workbook and worksheet
        create_spreadsheet(headers, rows, buffer)
        buffer.seek(0)
        workbook = load_workbook(buffer)
        worksheet = workbook.active

        # Run autosize_columns on the loaded worksheet
        autosize_columns(worksheet)

        # Check if the column widths were adjusted properly
        col1_width = worksheet.column_dimensions["A"].width
        col2_width = worksheet.column_dimensions["B"].width
        self.assertGreater(col2_width, col1_width)

    def test_create_spreadsheet_basic(self):
        """Test creating a spreadsheet with headers and rows."""
        buffer = BytesIO()
        headers = ["Header1", "Header2"]
        rows = [["Row1-Col1", "Row1-Col2"], ["Row2-Col1", "Row2-Col2"]]

        create_spreadsheet(headers, rows, buffer)
        buffer.seek(0)
        workbook = load_workbook(buffer)
        worksheet = workbook.active

        # Verify headers and rows
        self.assertEqual(worksheet["A1"].value, "Header1")
        self.assertEqual(worksheet["B1"].value, "Header2")
        self.assertEqual(worksheet["A2"].value, "Row1-Col1")
        self.assertEqual(worksheet["B2"].value, "Row1-Col2")
        self.assertEqual(worksheet["A3"].value, "Row2-Col1")
        self.assertEqual(worksheet["B3"].value, "Row2-Col2")

    def test_create_spreadsheet_with_styles_and_merge(self):
        """Test creating a spreadsheet with custom styles and merged cells."""
        buffer = BytesIO()
        headers = ["Header1", "Header2"]
        rows = [["Row1-Col1", "Row1-Col2"]]
        styles = {"A1": "col_header", "B1": "col_header"}
        merge = ["A1:B1"]

        create_spreadsheet(headers, rows, buffer, styles=styles, merge=merge)
        buffer.seek(0)
        workbook = load_workbook(buffer)
        worksheet = workbook.active

        # Verify merged cells and styles
        self.assertTrue(worksheet.merged_cells.ranges)
        self.assertEqual(str(worksheet.merged_cells.ranges[0]), "A1:B1")
        self.assertEqual(worksheet["A1"].style, "col_header")
        self.assertEqual(worksheet["B1"].style, "col_header")

    def test_create_spreadsheet_with_freeze_panes_and_dimensions(self):
        """Test creating a spreadsheet with freeze panes and custom row/column dimensions."""
        buffer = BytesIO()
        headers = ["Header1", "Header2"]
        rows = [["Row1-Col1", "Row1-Col2"]]
        dimensions = {"rows": {1: 25}, "columns": {"A": 20, "B": 30}}

        create_spreadsheet(
            headers, rows, buffer, freeze_panes="A2", dimensions=dimensions
        )
        buffer.seek(0)
        workbook = load_workbook(buffer)
        worksheet = workbook.active

        # Verify freeze panes and custom dimensions
        self.assertEqual(worksheet.freeze_panes, "A2")
        self.assertEqual(worksheet.row_dimensions[1].height, 25)
        self.assertEqual(worksheet.column_dimensions["A"].width, 20)
        self.assertEqual(worksheet.column_dimensions["B"].width, 30)

    def test_get_downloadable_spreadsheet(self):
        """Test generating a downloadable spreadsheet encoded in base64."""
        headers = ["Header1", "Header2"]
        rows = [["Row1-Col1", "Row1-Col2"], ["Row2-Col1", "Row2-Col2"]]

        # Generate the base64-encoded spreadsheet
        encoded_spreadsheet = get_downloadable_spreadsheet(headers, rows)
        decoded_data = base64.b64decode(encoded_spreadsheet)

        # Load the spreadsheet from the decoded data and verify its content
        buffer = BytesIO(decoded_data)
        workbook = load_workbook(buffer)
        worksheet = workbook.active

        # Verify headers and rows
        self.assertEqual(worksheet["A1"].value, "Header1")
        self.assertEqual(worksheet["B1"].value, "Header2")
        self.assertEqual(worksheet["A2"].value, "Row1-Col1")
        self.assertEqual(worksheet["B2"].value, "Row1-Col2")
        self.assertEqual(worksheet["A3"].value, "Row2-Col1")
        self.assertEqual(worksheet["B3"].value, "Row2-Col2")


if __name__ == "__main__":
    unittest.main()
