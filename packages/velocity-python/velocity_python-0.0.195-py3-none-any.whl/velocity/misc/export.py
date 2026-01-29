from typing import List, Dict
from io import BytesIO
import base64
import openpyxl
from openpyxl.styles import NamedStyle, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter


def extract(d: dict, keys: List[str]) -> List:
    """Extract values from a dictionary based on a list of keys."""
    return [d.get(key) for key in keys]


def autosize_columns(ws, fixed: Dict[str, float] = {}):
    """Autosize columns in the worksheet based on content length."""
    for col in ws.columns:
        max_length = 0
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                continue
        adjusted_width = (max_length + 2) * 1.2
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = fixed.get(col_letter, adjusted_width)


def create_spreadsheet(
    headers: List[str],
    rows: List[List],
    fileorbuffer,
    styles: Dict[str, str] = {},
    merge: List[str] = [],
    formats: Dict[str, str] = {},
    named_styles: List[NamedStyle] = [],
    freeze_panes: str = "A2",
    dimensions: dict = None,
    auto_size: bool = True,
):
    """Create an Excel spreadsheet with specified headers, rows, and styles."""
    wb = openpyxl.Workbook()
    ws = wb.active

    # Define default named styles
    def get_named_styles():
        named_styles = {
            "col_header": NamedStyle(
                name="col_header",
                font=Font(bold=True),
                border=Border(bottom=Side(style="medium", color="000000")),
            ),
            "sum_total": NamedStyle(
                name="sum_total",
                border=Border(bottom=Side(style="double", color="000000")),
            ),
            "sub_total": NamedStyle(
                name="sub_total",
                font=Font(bold=True),
                border=Border(bottom=Side(style="thin", color="000000")),
            ),
            "bold": NamedStyle(name="bold", font=Font(bold=True)),
            "align_right": NamedStyle(
                name="align_right",
                font=Font(bold=True),
                border=Border(top=Side(style="thin", color="000000")),
                alignment=Alignment(horizontal="right", vertical="center"),
            ),
            "align_left": NamedStyle(
                name="align_left",
                font=Font(bold=True),
                border=Border(top=Side(style="thin", color="000000")),
                alignment=Alignment(horizontal="left", vertical="center"),
            ),
            "align_right_double": NamedStyle(
                name="align_right_double",
                font=Font(bold=True),
                border=Border(top=Side(style="double", color="000000")),
                alignment=Alignment(horizontal="right", vertical="center"),
            ),
            "align_left_double": NamedStyle(
                name="align_left_double",
                font=Font(bold=True),
                border=Border(top=Side(style="double", color="000000")),
                alignment=Alignment(horizontal="left", vertical="center"),
            ),
        }
        return named_styles

    # Add default and user-defined styles
    local_styles = get_named_styles()
    for style in named_styles:
        local_styles[style.name] = style
    for style in local_styles.values():
        wb.add_named_style(style)

    # Add headers and rows
    ws.append(headers)
    for row in rows:
        ws.append(row)

    # Set freeze panes
    ws.freeze_panes = freeze_panes

    # Auto-size columns if enabled
    if auto_size:
        autosize_columns(ws, fixed={})

    # Set row and column dimensions if provided
    if dimensions:
        for key, val in dimensions.get("rows", {}).items():
            ws.row_dimensions[key].height = val
        for key, val in dimensions.get("columns", {}).items():
            ws.column_dimensions[key].width = val

    # Apply cell styles, merges, and formats
    for cell, style_name in styles.items():
        if style_name in local_styles:
            ws[cell].style = local_styles[style_name]
    for cell_range in merge:
        ws.merge_cells(cell_range)
    for cell, format_code in formats.items():
        ws[cell].number_format = format_code

    # Save workbook to the provided file or buffer
    wb.save(fileorbuffer)


def get_downloadable_spreadsheet(
    headers: List[str],
    rows: List[List],
    styles: Dict[str, str] = {},
    merge: List[str] = [],
    formats: Dict[str, str] = {},
    named_styles: List[NamedStyle] = [],
    freeze_panes: str = "A2",
    dimensions: dict = None,
    auto_size: bool = True,
) -> str:
    """Generate a downloadable spreadsheet encoded in base64."""
    buffer = BytesIO()
    create_spreadsheet(
        headers,
        rows,
        buffer,
        styles,
        merge,
        formats,
        named_styles,
        freeze_panes,
        dimensions,
        auto_size,
    )
    return base64.b64encode(buffer.getvalue()).decode()
