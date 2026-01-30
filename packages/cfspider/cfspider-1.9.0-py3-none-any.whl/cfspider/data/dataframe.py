"""
DataFrame - 轻量级数据处理类
"""

import json
import csv
from typing import List, Dict, Any, Callable, Union, Optional, Tuple
from copy import deepcopy


class DataFrame:
    """
    轻量级 DataFrame 实现
    支持链式操作，与爬虫无缝集成
    """
    
    def __init__(self, data: List[Dict[str, Any]] = None):
        """
        创建 DataFrame
        
        Args:
            data: 字典列表，如 [{"name": "A", "price": 100}, ...]
        """
        self._data = data if data is not None else []
        self._columns = self._infer_columns()
    
    def _infer_columns(self) -> List[str]:
        """推断列名"""
        if not self._data:
            return []
        cols = set()
        for row in self._data:
            cols.update(row.keys())
        return list(cols)
    
    @property
    def columns(self) -> List[str]:
        """列名列表"""
        return self._columns.copy()
    
    @property
    def rows(self) -> int:
        """行数"""
        return len(self._data)
    
    @property
    def shape(self) -> Tuple[int, int]:
        """(行数, 列数)"""
        return (len(self._data), len(self._columns))
    
    def __len__(self) -> int:
        return len(self._data)
    
    def __iter__(self):
        return iter(self._data)
    
    def __getitem__(self, key):
        if isinstance(key, str):
            # 获取列
            return [row.get(key) for row in self._data]
        elif isinstance(key, int):
            # 获取行
            return self._data[key]
        elif isinstance(key, slice):
            # 切片
            return DataFrame(self._data[key])
        elif isinstance(key, list):
            # 多列选择
            return self.select(*key)
        raise TypeError(f"不支持的索引类型: {type(key)}")
    
    def __repr__(self) -> str:
        return self._format_table()
    
    def _format_table(self, max_rows: int = 20) -> str:
        """格式化为表格字符串"""
        if not self._data:
            return "DataFrame(空)"
        
        # 计算列宽
        cols = self._columns[:10]  # 最多显示10列
        widths = {}
        for col in cols:
            values = [str(row.get(col, ""))[:30] for row in self._data[:max_rows]]
            widths[col] = max(len(col), max(len(v) for v in values) if values else 0)
        
        # 构建表格
        lines = []
        
        # 顶部边框
        top = "┌" + "┬".join("─" * (widths[c] + 2) for c in cols) + "┐"
        lines.append(top)
        
        # 表头
        header = "│" + "│".join(f" {c:^{widths[c]}} " for c in cols) + "│"
        lines.append(header)
        
        # 分隔线
        sep = "├" + "┼".join("─" * (widths[c] + 2) for c in cols) + "┤"
        lines.append(sep)
        
        # 数据行
        display_data = self._data[:max_rows]
        for row in display_data:
            values = [str(row.get(c, ""))[:30] for c in cols]
            line = "│" + "│".join(f" {v:<{widths[cols[i]]}} " for i, v in enumerate(values)) + "│"
            lines.append(line)
        
        # 省略提示
        if len(self._data) > max_rows:
            lines.append(f"│ ... 共 {len(self._data)} 行，显示前 {max_rows} 行 ...".ljust(sum(widths.values()) + len(cols) * 3) + "│")
        
        # 底部边框
        bottom = "└" + "┴".join("─" * (widths[c] + 2) for c in cols) + "┘"
        lines.append(bottom)
        
        # 添加维度信息
        lines.append(f"\n[{len(self._data)} 行 x {len(self._columns)} 列]")
        
        return "\n".join(lines)
    
    # ========== 过滤 ==========
    
    def filter(self, func: Callable = None, **conditions) -> 'DataFrame':
        """
        过滤数据
        
        Args:
            func: 过滤函数，接收行字典，返回 bool
            **conditions: 列条件，如 price=lambda x: x > 100
        
        Examples:
            df.filter(lambda row: row["price"] > 100)
            df.filter(price=lambda x: x > 100)
            df.filter(stock=lambda x: x is not None)
        """
        result = self._data.copy()
        
        if func:
            result = [row for row in result if func(row)]
        
        for col, cond in conditions.items():
            if callable(cond):
                result = [row for row in result if cond(row.get(col))]
            else:
                result = [row for row in result if row.get(col) == cond]
        
        return DataFrame(result)
    
    # ========== 转换 ==========
    
    def transform(self, **transforms) -> 'DataFrame':
        """
        转换列值
        
        Args:
            **transforms: 列转换，如 price=float, name=str.strip
        
        Examples:
            df.transform(price=float)
            df.transform(price=lambda x: x * 0.9)
            df.transform(name=str.strip, price=lambda x: float(x.replace("$", "")))
        """
        result = deepcopy(self._data)
        
        for col, func in transforms.items():
            for row in result:
                if col in row and row[col] is not None:
                    try:
                        row[col] = func(row[col])
                    except:
                        pass
        
        return DataFrame(result)
    
    # ========== 排序 ==========
    
    def sort(self, by: Union[str, List[str]], desc: Union[bool, List[bool]] = False) -> 'DataFrame':
        """
        排序
        
        Args:
            by: 排序列名
            desc: 是否降序
        
        Examples:
            df.sort("price")
            df.sort("price", desc=True)
            df.sort(["category", "price"], desc=[False, True])
        """
        if isinstance(by, str):
            by = [by]
            desc = [desc] if isinstance(desc, bool) else desc
        
        def sort_key(row):
            values = []
            for i, col in enumerate(by):
                val = row.get(col)
                # 处理 None
                if val is None:
                    val = "" if isinstance(val, str) else float('inf')
                values.append(val)
            return values
        
        reverse = desc[0] if isinstance(desc, list) else desc
        result = sorted(self._data, key=sort_key, reverse=reverse)
        
        return DataFrame(result)
    
    # ========== 选择/删除列 ==========
    
    def select(self, *columns) -> 'DataFrame':
        """
        选择列
        
        Examples:
            df.select("name", "price")
        """
        result = [{col: row.get(col) for col in columns} for row in self._data]
        return DataFrame(result)
    
    def drop(self, *columns) -> 'DataFrame':
        """
        删除列
        
        Examples:
            df.drop("temp_column")
        """
        keep = [c for c in self._columns if c not in columns]
        return self.select(*keep)
    
    # ========== 添加列 ==========
    
    def add(self, name: str, value: Union[Any, Callable]) -> 'DataFrame':
        """
        添加列
        
        Args:
            name: 列名
            value: 值或函数
        
        Examples:
            df.add("total", lambda row: row["price"] * row["quantity"])
            df.add("source", "爬虫")
        """
        result = deepcopy(self._data)
        
        for row in result:
            if callable(value):
                row[name] = value(row)
            else:
                row[name] = value
        
        return DataFrame(result)
    
    # ========== 去重 ==========
    
    def unique(self, *columns) -> 'DataFrame':
        """
        去重
        
        Examples:
            df.unique()
            df.unique("name")
            df.unique("name", "category")
        """
        if not columns:
            columns = self._columns
        
        seen = set()
        result = []
        
        for row in self._data:
            key = tuple(row.get(c) for c in columns)
            if key not in seen:
                seen.add(key)
                result.append(row)
        
        return DataFrame(result)
    
    # ========== 缺失值处理 ==========
    
    def dropna(self, *columns) -> 'DataFrame':
        """
        删除空值行
        
        Examples:
            df.dropna()
            df.dropna("price")
        """
        if not columns:
            columns = self._columns
        
        result = [
            row for row in self._data
            if all(row.get(c) is not None for c in columns)
        ]
        
        return DataFrame(result)
    
    def fillna(self, **values) -> 'DataFrame':
        """
        填充空值
        
        Examples:
            df.fillna(price=0, stock="未知")
        """
        result = deepcopy(self._data)
        
        for row in result:
            for col, val in values.items():
                if row.get(col) is None:
                    row[col] = val
        
        return DataFrame(result)
    
    # ========== 分组聚合 ==========
    
    def group(self, *by) -> 'GroupedDataFrame':
        """
        分组
        
        Examples:
            df.group("category").agg(count=("name", "count"), avg=("price", "mean"))
        """
        return GroupedDataFrame(self._data, list(by))
    
    # ========== 合并 ==========
    
    def join(self, other: 'DataFrame', on: str, how: str = "inner") -> 'DataFrame':
        """
        合并两个 DataFrame
        
        Args:
            other: 另一个 DataFrame
            on: 合并键
            how: 合并方式 (inner, left, right, outer)
        """
        result = []
        other_dict = {row[on]: row for row in other._data if on in row}
        
        for row in self._data:
            key = row.get(on)
            if key in other_dict:
                merged = {**row, **other_dict[key]}
                result.append(merged)
            elif how in ("left", "outer"):
                result.append(row.copy())
        
        if how in ("right", "outer"):
            self_keys = {row.get(on) for row in self._data}
            for row in other._data:
                if row.get(on) not in self_keys:
                    result.append(row.copy())
        
        return DataFrame(result)
    
    def concat(self, other: 'DataFrame') -> 'DataFrame':
        """
        连接两个 DataFrame
        """
        return DataFrame(self._data + other._data)
    
    # ========== 切片 ==========
    
    def head(self, n: int = 5) -> 'DataFrame':
        """取前 N 行"""
        return DataFrame(self._data[:n])
    
    def tail(self, n: int = 5) -> 'DataFrame':
        """取后 N 行"""
        return DataFrame(self._data[-n:])
    
    # ========== 保存 ==========
    
    def save(self, path: str, **options) -> None:
        """
        保存到文件
        
        Args:
            path: 文件路径，根据后缀自动识别格式
        
        Examples:
            df.save("output.csv")
            df.save("output.json")
            df.save("output.xlsx")
        """
        ext = path.rsplit(".", 1)[-1].lower()
        
        if ext == "csv":
            self.to_csv(path, **options)
        elif ext == "json":
            self.to_json(path, **options)
        elif ext in ("xlsx", "xls"):
            self.to_excel(path, **options)
        elif ext == "parquet":
            self.to_parquet(path, **options)
        else:
            raise ValueError(f"不支持的文件格式: {ext}")
    
    def to_csv(self, path: str, encoding: str = "utf-8") -> None:
        """保存为 CSV"""
        with open(path, "w", newline="", encoding=encoding) as f:
            if self._data:
                writer = csv.DictWriter(f, fieldnames=self._columns)
                writer.writeheader()
                writer.writerows(self._data)
    
    def to_json(self, path: str, indent: int = 2, encoding: str = "utf-8") -> None:
        """保存为 JSON"""
        with open(path, "w", encoding=encoding) as f:
            json.dump(self._data, f, indent=indent, ensure_ascii=False)
    
    def to_excel(self, path: str, sheet_name: str = "Sheet1") -> None:
        """保存为 Excel（需要 openpyxl）"""
        try:
            from openpyxl import Workbook
        except ImportError:
            raise ImportError("保存 Excel 需要安装 openpyxl: pip install openpyxl")
        
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # 写入表头
        for col, name in enumerate(self._columns, 1):
            ws.cell(row=1, column=col, value=name)
        
        # 写入数据
        for row_idx, row in enumerate(self._data, 2):
            for col_idx, col_name in enumerate(self._columns, 1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(col_name))
        
        wb.save(path)
    
    def to_parquet(self, path: str) -> None:
        """保存为 Parquet（需要 pyarrow）"""
        try:
            import pyarrow as pa
            import pyarrow.parquet as pq
        except ImportError:
            raise ImportError("保存 Parquet 需要安装 pyarrow: pip install pyarrow")
        
        table = pa.Table.from_pylist(self._data)
        pq.write_table(table, path)
    
    # ========== 转换 ==========
    
    def to_list(self) -> List[Dict[str, Any]]:
        """转为字典列表"""
        return deepcopy(self._data)
    
    def to_dict(self) -> Dict[str, List[Any]]:
        """转为列字典"""
        return {col: [row.get(col) for row in self._data] for col in self._columns}
    
    def to_pandas(self):
        """转为 pandas DataFrame"""
        try:
            import pandas as pd
        except ImportError:
            raise ImportError("需要安装 pandas: pip install pandas")
        return pd.DataFrame(self._data)
    
    def to_polars(self):
        """转为 polars DataFrame"""
        try:
            import polars as pl
        except ImportError:
            raise ImportError("需要安装 polars: pip install polars")
        return pl.DataFrame(self._data)
    
    # ========== 类方法 ==========
    
    @classmethod
    def from_pandas(cls, df) -> 'DataFrame':
        """从 pandas DataFrame 创建"""
        return cls(df.to_dict(orient="records"))
    
    @classmethod
    def from_polars(cls, df) -> 'DataFrame':
        """从 polars DataFrame 创建"""
        return cls(df.to_dicts())


class GroupedDataFrame:
    """分组后的 DataFrame"""
    
    def __init__(self, data: List[Dict], by: List[str]):
        self._data = data
        self._by = by
        self._groups = self._make_groups()
    
    def _make_groups(self) -> Dict[tuple, List[Dict]]:
        """创建分组"""
        groups = {}
        for row in self._data:
            key = tuple(row.get(col) for col in self._by)
            if key not in groups:
                groups[key] = []
            groups[key].append(row)
        return groups
    
    def agg(self, **aggregations) -> DataFrame:
        """
        聚合
        
        Args:
            **aggregations: 聚合定义，如 count=("name", "count"), avg=("price", "mean")
        
        Examples:
            df.group("category").agg(
                count=("name", "count"),
                avg_price=("price", "mean"),
                total=("stock", "sum")
            )
        """
        result = []
        
        for key, rows in self._groups.items():
            row_result = dict(zip(self._by, key))
            
            for name, (col, func) in aggregations.items():
                values = [r.get(col) for r in rows if r.get(col) is not None]
                
                if func == "count":
                    row_result[name] = len(values)
                elif func == "sum":
                    row_result[name] = sum(values) if values else 0
                elif func == "mean":
                    row_result[name] = sum(values) / len(values) if values else None
                elif func == "min":
                    row_result[name] = min(values) if values else None
                elif func == "max":
                    row_result[name] = max(values) if values else None
                elif func == "first":
                    row_result[name] = values[0] if values else None
                elif func == "last":
                    row_result[name] = values[-1] if values else None
                elif callable(func):
                    row_result[name] = func(values)
            
            result.append(row_result)
        
        return DataFrame(result)

