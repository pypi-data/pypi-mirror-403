"""
IO 模块 - 统一的数据读取入口
"""

import json
import csv
import os
from typing import List, Dict, Any, Union

from .dataframe import DataFrame


def read(source: Any, pick: Dict[str, str] = None, **options) -> DataFrame:
    """
    统一读取数据入口
    
    Args:
        source: 数据来源，可以是:
            - 文件路径: "data.csv", "data.json", "data.xlsx"
            - URL: "https://..." (自动爬取)
            - URL列表: ["url1", "url2"] (批量爬取)
            - 字典列表: [{"a": 1}, {"a": 2}]
            - pandas DataFrame
            - polars DataFrame
        pick: 爬取时的数据提取规则
        **options: 其他选项
    
    Returns:
        DataFrame
    
    Examples:
        # 从文件读取
        df = read("data.csv")
        df = read("data.json")
        
        # 从 URL 爬取
        df = read("https://example.com", pick={"title": "h1", "price": ".price"})
        
        # 批量爬取
        df = read(["url1", "url2"], pick={"title": "h1"})
        
        # 从字典列表创建
        df = read([{"name": "A", "price": 100}])
    """
    # 1. 字典列表
    if isinstance(source, list) and source and isinstance(source[0], dict):
        return DataFrame(source)
    
    # 2. 空列表
    if isinstance(source, list) and not source:
        return DataFrame([])
    
    # 3. URL 列表（爬取）
    if isinstance(source, list) and source and isinstance(source[0], str):
        if source[0].startswith(("http://", "https://")):
            return _read_urls(source, pick, **options)
        else:
            # 可能是文件路径列表
            dfs = [read(path, **options) for path in source]
            result = dfs[0]
            for df in dfs[1:]:
                result = result.concat(df)
            return result
    
    # 4. 字符串
    if isinstance(source, str):
        # 4.1 URL
        if source.startswith(("http://", "https://")):
            return _read_url(source, pick, **options)
        
        # 4.2 文件路径
        if os.path.exists(source):
            return _read_file(source, **options)
        
        raise FileNotFoundError(f"文件不存在: {source}")
    
    # 5. pandas DataFrame
    if hasattr(source, "to_dict") and hasattr(source, "columns"):
        return DataFrame.from_pandas(source)
    
    # 6. polars DataFrame
    if hasattr(source, "to_dicts"):
        return DataFrame.from_polars(source)
    
    raise TypeError(f"不支持的数据源类型: {type(source)}")


def _read_file(path: str, **options) -> DataFrame:
    """读取文件"""
    ext = path.rsplit(".", 1)[-1].lower()
    
    if ext == "csv":
        return read_csv(path, **options)
    elif ext == "json":
        return read_json(path, **options)
    elif ext in ("xlsx", "xls"):
        return read_excel(path, **options)
    elif ext == "parquet":
        return read_parquet(path, **options)
    else:
        raise ValueError(f"不支持的文件格式: {ext}")


def read_csv(path: str, encoding: str = "utf-8", **options) -> DataFrame:
    """
    读取 CSV 文件
    
    Args:
        path: 文件路径
        encoding: 编码
    """
    data = []
    with open(path, "r", encoding=encoding, newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            # 尝试转换数值
            parsed_row = {}
            for k, v in row.items():
                if v == "":
                    parsed_row[k] = None
                else:
                    try:
                        # 尝试转为整数
                        if "." not in v:
                            parsed_row[k] = int(v)
                        else:
                            parsed_row[k] = float(v)
                    except (ValueError, TypeError):
                        parsed_row[k] = v
            data.append(parsed_row)
    return DataFrame(data)


def read_json(path: str, encoding: str = "utf-8", **options) -> DataFrame:
    """
    读取 JSON 文件
    
    Args:
        path: 文件路径
        encoding: 编码
    """
    with open(path, "r", encoding=encoding) as f:
        data = json.load(f)
    
    # 确保是列表
    if isinstance(data, dict):
        # 可能是 {"data": [...]} 格式
        for key in ("data", "items", "results", "records"):
            if key in data and isinstance(data[key], list):
                data = data[key]
                break
        else:
            data = [data]
    
    return DataFrame(data)


def read_excel(path: str, sheet_name: str = None, **options) -> DataFrame:
    """
    读取 Excel 文件
    
    Args:
        path: 文件路径
        sheet_name: 工作表名称
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("读取 Excel 需要安装 openpyxl: pip install openpyxl")
    
    wb = load_workbook(path, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return DataFrame([])
    
    headers = [str(h) if h else f"col_{i}" for i, h in enumerate(rows[0])]
    data = []
    
    for row in rows[1:]:
        data.append(dict(zip(headers, row)))
    
    return DataFrame(data)


def read_parquet(path: str, **options) -> DataFrame:
    """
    读取 Parquet 文件
    
    Args:
        path: 文件路径
    """
    try:
        import pyarrow.parquet as pq
    except ImportError:
        raise ImportError("读取 Parquet 需要安装 pyarrow: pip install pyarrow")
    
    table = pq.read_table(path)
    data = table.to_pylist()
    return DataFrame(data)


def _read_url(url: str, pick: Dict[str, str] = None, **options) -> DataFrame:
    """爬取单个 URL"""
    try:
        from .. import api as cfspider_api
    except ImportError:
        raise ImportError("爬取功能需要 cfspider 主模块")
    
    # 获取代理配置
    cf_proxies = options.pop("cf_proxies", None)
    uuid = options.pop("uuid", None)
    headers = options.pop("headers", None)
    
    # 发送请求
    response = cfspider_api.get(
        url,
        cf_proxies=cf_proxies,
        uuid=uuid,
        headers=headers,
        **options
    )
    
    # 如果没有提取规则，返回原始内容
    if not pick:
        try:
            # 尝试解析为 JSON
            data = response.json()
            if isinstance(data, list):
                return DataFrame(data)
            return DataFrame([data])
        except:
            return DataFrame([{"url": url, "content": response.text}])
    
    # 使用提取规则
    row = {"url": url}
    for name, selector in pick.items():
        try:
            if hasattr(response, "find"):
                row[name] = response.find(selector)
            else:
                # 简单的 CSS 选择器支持
                from bs4 import BeautifulSoup
                soup = BeautifulSoup(response.text, "html.parser")
                elem = soup.select_one(selector)
                row[name] = elem.get_text(strip=True) if elem else None
        except:
            row[name] = None
    
    return DataFrame([row])


def _read_urls(urls: List[str], pick: Dict[str, str] = None, **options) -> DataFrame:
    """批量爬取多个 URL"""
    try:
        from .. import api as cfspider_api
    except ImportError:
        raise ImportError("爬取功能需要 cfspider 主模块")
    
    # 获取配置
    cf_proxies = options.pop("cf_proxies", None)
    uuid = options.pop("uuid", None)
    concurrency = options.pop("concurrency", 5)
    delay = options.pop("delay", 0)
    progress = options.pop("progress", True)
    
    results = []
    
    # 简单的串行爬取（可以后续优化为并发）
    for i, url in enumerate(urls):
        if progress:
            print(f"\r爬取进度: {i+1}/{len(urls)}", end="", flush=True)
        
        try:
            response = cfspider_api.get(
                url,
                cf_proxies=cf_proxies,
                uuid=uuid,
                **options
            )
            
            if not pick:
                try:
                    data = response.json()
                    if isinstance(data, dict):
                        data["url"] = url
                        results.append(data)
                    else:
                        results.append({"url": url, "content": response.text})
                except:
                    results.append({"url": url, "content": response.text})
            else:
                row = {"url": url}
                for name, selector in pick.items():
                    try:
                        if hasattr(response, "find"):
                            row[name] = response.find(selector)
                        else:
                            from bs4 import BeautifulSoup
                            soup = BeautifulSoup(response.text, "html.parser")
                            elem = soup.select_one(selector)
                            row[name] = elem.get_text(strip=True) if elem else None
                    except:
                        row[name] = None
                results.append(row)
                
        except Exception as e:
            results.append({"url": url, "error": str(e)})
        
        # 延迟
        if delay and i < len(urls) - 1:
            import time
            if isinstance(delay, (list, tuple)):
                import random
                time.sleep(random.uniform(delay[0], delay[1]))
            else:
                time.sleep(delay)
    
    if progress:
        print()  # 换行
    
    return DataFrame(results)

