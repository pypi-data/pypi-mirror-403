"""Functions for shaping and formatting spreadsheets."""

import logging
import warnings
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet
from typeguard import typechecked

from bfb_delivery.lib.constants import (
    ADDRESS_COLUMN_WIDTH,
    BOX_TYPE_COLOR_MAP,
    COLUMN_NAME_MAP,
    COMBINED_ROUTES_COLUMNS,
    FILE_DATE_FORMAT,
    FORMATTED_ROUTES_COLUMNS,
    MANIFEST_DATE_FORMAT,
    NOTES_COLUMN_WIDTH,
    PROTEIN_BOX_TYPES,
    SPLIT_ROUTE_COLUMNS,
    BoxType,
    CellColors,
    Columns,
    DocStrings,
)
from bfb_delivery.lib.formatting.data_cleaning import (
    format_and_validate_data,
    format_column_names,
)
from bfb_delivery.lib.formatting.utils import (
    get_book_one_drivers,
    get_extra_notes,
    get_phone_number,
    map_columns,
    set_row_height_of_wrapped_cell,
)
from bfb_delivery.lib.utils import get_friday

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

# Silences warning for in-place operations on copied df slices.
pd.options.mode.copy_on_write = True


# TODO: Use Pandera.
# https://github.com/crickets-and-comb/bfb_delivery/issues/80
# TODO: Switch to or allow CSVs instead of Excel files.
# https://github.com/crickets-and-comb/bfb_delivery/issues/81
@typechecked
def split_chunked_route(  # noqa: D103
    input_path: Path | str,
    output_dir: Path | str,
    output_filename: str,
    n_books: int,
    book_one_drivers_file: str,
    date: str,
) -> list[Path]:
    if n_books <= 0:
        raise ValueError("n_books must be greater than 0.")
    # TODO: Make this accept input_path only as Path? Or only as str to simplify?
    # https://github.com/crickets-and-comb/bfb_delivery/issues/55
    input_path = Path(input_path)
    date = date if date else get_friday(fmt=MANIFEST_DATE_FORMAT)

    chunked_sheet: pd.DataFrame = pd.read_excel(input_path)
    chunked_sheet.columns = format_column_names(columns=chunked_sheet.columns.to_list())
    map_columns(df=chunked_sheet, column_name_map=COLUMN_NAME_MAP, invert_map=False)
    format_and_validate_data(df=chunked_sheet, columns=SPLIT_ROUTE_COLUMNS + [Columns.DRIVER])
    chunked_sheet.sort_values(by=[Columns.DRIVER, Columns.STOP_NO], inplace=True)
    # TODO: Validate columns? (Use Pandera?)
    # https://github.com/crickets-and-comb/bfb_delivery/issues/80

    drivers = list(chunked_sheet[Columns.DRIVER].unique())
    driver_count = len(drivers)
    if driver_count < n_books:
        raise ValueError(
            "n_books must be less than or equal to the number of drivers: "
            f"driver_count: {driver_count}, n_books: {n_books}."
        )

    output_dir = Path(output_dir) if output_dir else Path(input_path).parent
    base_output_filename = (
        f"split_workbook_{datetime.now().strftime(FILE_DATE_FORMAT)}.xlsx"
        if output_filename == ""
        else output_filename
    )
    output_dir.mkdir(parents=True, exist_ok=True)

    split_workbook_paths: list[Path] = []
    driver_sets = _get_driver_sets(
        drivers=drivers, n_books=n_books, book_one_drivers_file=book_one_drivers_file
    )

    logger.info(f"Writing split chunked workbooks to {output_dir.resolve()}")
    for i, driver_set in enumerate(driver_sets):
        i_file_name = f"{base_output_filename.split('.')[0]}_{i + 1}.xlsx"
        split_workbook_path: Path = output_dir / i_file_name
        split_workbook_paths.append(split_workbook_path)

        driver_set_df = chunked_sheet[chunked_sheet[Columns.DRIVER].isin(driver_set)]
        driver_set_df.sort_values(by=[Columns.DRIVER, Columns.STOP_NO], inplace=True)

        with pd.ExcelWriter(split_workbook_path) as writer:
            for driver_name, data in driver_set_df.groupby(Columns.DRIVER):
                data[SPLIT_ROUTE_COLUMNS].to_excel(
                    writer, sheet_name=f"{date} {driver_name}", index=False
                )

    split_workbook_paths = [path.resolve() for path in split_workbook_paths]

    return split_workbook_paths


split_chunked_route.__doc__ = DocStrings.SPLIT_CHUNKED_ROUTE.api_docstring


@typechecked
def create_manifests(  # noqa: D103
    input_dir: Path | str, output_dir: Path | str, output_filename: str, extra_notes_file: str
) -> Path:
    output_filename = (
        f"final_manifests_{datetime.now().strftime(FILE_DATE_FORMAT)}.xlsx"
        if output_filename == ""
        else output_filename
    )

    combined_route_workbook_path = combine_route_tables(
        input_dir=input_dir, output_dir=output_dir, output_filename=""
    )

    formatted_manifest_path = format_combined_routes(
        input_path=combined_route_workbook_path,
        output_dir=output_dir,
        output_filename=output_filename,
        extra_notes_file=extra_notes_file,
    )

    return formatted_manifest_path


create_manifests.__doc__ = DocStrings.CREATE_MANIFESTS.api_docstring


@typechecked
def combine_route_tables(  # noqa: D103
    input_dir: Path | str, output_dir: Path | str, output_filename: str
) -> Path:
    input_dir = Path(input_dir)
    paths = list(input_dir.glob("*.csv"))

    output_dir = Path(output_dir) if output_dir else paths[0].parent
    output_filename = (
        f"combined_routes_{datetime.now().strftime(FILE_DATE_FORMAT)}.xlsx"
        if output_filename == ""
        else output_filename
    )
    output_path = output_dir / output_filename
    output_dir.mkdir(parents=True, exist_ok=True)

    logger.info(f"Writing combined routes to {output_path.resolve()}")
    with pd.ExcelWriter(output_path) as writer:
        for path in sorted(paths):
            route_df = pd.read_csv(path)
            map_columns(df=route_df, column_name_map=COLUMN_NAME_MAP, invert_map=True)
            route_df.sort_values(by=[Columns.STOP_NO], inplace=True)
            route_df[COMBINED_ROUTES_COLUMNS].to_excel(
                writer, sheet_name=path.stem, index=False
            )

    return output_path.resolve()


combine_route_tables.__doc__ = DocStrings.COMBINE_ROUTE_TABLES.api_docstring


@typechecked
def format_combined_routes(  # noqa: D103
    input_path: Path | str,
    output_dir: Path | str,
    output_filename: str,
    extra_notes_file: str,
) -> Path:
    input_path = Path(input_path)
    output_dir = Path(output_dir) if output_dir else input_path.parent
    output_filename = (
        f"formatted_routes_{datetime.now().strftime(FILE_DATE_FORMAT)}.xlsx"
        if output_filename == ""
        else output_filename
    )
    output_path = Path(output_dir) / output_filename

    output_dir.mkdir(parents=True, exist_ok=True)

    extra_notes_df = get_extra_notes(file_path=extra_notes_file)

    wb = Workbook()
    wb.remove(wb["Sheet"])
    with pd.ExcelFile(input_path) as xls:
        for sheet_idx, sheet_name in enumerate(sorted(xls.sheet_names)):

            route_df = pd.read_excel(xls, sheet_name)

            # TODO: Use Pandera?
            # https://github.com/crickets-and-comb/bfb_delivery/issues/80
            route_df.columns = format_column_names(columns=route_df.columns.to_list())
            format_and_validate_data(df=route_df, columns=COMBINED_ROUTES_COLUMNS)
            route_df.sort_values(by=[Columns.STOP_NO], inplace=True)

            agg_dict = _aggregate_route_data(df=route_df, extra_notes_df=extra_notes_df)

            _make_manifest_sheet(
                wb=wb,
                agg_dict=agg_dict,
                route_df=route_df,
                sheet_name=str(sheet_name),
                sheet_idx=sheet_idx,
            )

    # TODO: Can check cell values, though. (Maybe read dataframe from start row?)
    # https://github.com/crickets-and-comb/bfb_delivery/issues/62
    logger.info(f"Writing formatted routes to {output_path.resolve()}")
    wb.save(output_path)

    return output_path.resolve()


format_combined_routes.__doc__ = DocStrings.FORMAT_COMBINED_ROUTES.api_docstring


@typechecked
def _get_driver_sets(
    drivers: list[str], n_books: int, book_one_drivers_file: str
) -> list[list[str]]:
    """Split drivers into n_books sets."""
    drivers = sorted(drivers)
    drivers = _move_book_one_drivers_to_front(
        drivers=drivers, book_one_drivers_file=book_one_drivers_file
    )
    driver_sets = _split_driver_list(drivers=drivers, n_books=n_books)
    driver_sets = _group_numbered_drivers(driver_sets=driver_sets)
    driver_sets = [sorted(driver_set) for driver_set in driver_sets]

    return driver_sets


@typechecked
def _move_book_one_drivers_to_front(
    drivers: list[str], book_one_drivers_file: str
) -> list[str]:
    """Move book one drivers to the front of the list."""
    book_one_drivers = get_book_one_drivers(file_path=book_one_drivers_file)
    drivers = [d for d in book_one_drivers if d in drivers] + [
        d for d in drivers if d not in book_one_drivers
    ]

    return drivers


@typechecked
def _split_driver_list(drivers: list[str], n_books: int) -> list[list[str]]:
    """Split drivers into n_books sets, in order passed."""
    driver_sets = []

    len_drivers = len(drivers)
    start_idx = 0
    inc = len_drivers // n_books
    remainder = len_drivers % n_books

    while start_idx < len_drivers:
        end_idx = start_idx + inc
        if remainder > 0:
            end_idx += 1
            remainder -= 1
        end_idx = end_idx if end_idx < len_drivers else len_drivers

        driver_sets.append(drivers[start_idx:end_idx])
        start_idx = end_idx

    return driver_sets


@typechecked
def _group_numbered_drivers(driver_sets: list[list[str]]) -> list[list[str]]:
    """Merge drivers with numbers into a single set."""
    driver_sets_updated = driver_sets.copy()

    updated = True
    while updated:
        driver_sets_copy = driver_sets_updated.copy()
        for i, driver_set in enumerate(driver_sets_updated):
            driver_set_i = driver_set.copy()
            numbered_drivers = [d for d in driver_set if "#" in d]

            for d in numbered_drivers:
                driver_name = d.split("#")[0].strip()

                i_plus_1 = i + 1
                if i_plus_1 < len(driver_sets_updated):
                    for j, driver_set_j in enumerate(
                        driver_sets_updated[i_plus_1:], start=i_plus_1
                    ):
                        matching_drivers = [d for d in driver_set_j if driver_name in d]
                        driver_set_i += matching_drivers
                        driver_sets_updated[j] = [
                            d for d in driver_set_j if d not in matching_drivers
                        ]
            driver_sets_updated[i] = driver_set_i

        updated = driver_sets_copy != driver_sets_updated

    driver_sets_updated = [driver_set for driver_set in driver_sets_updated if driver_set]

    return driver_sets_updated


@typechecked
def _aggregate_route_data(df: pd.DataFrame, extra_notes_df: pd.DataFrame) -> dict[str, Any]:
    """Aggregate data for a single route.

    Args:
        df: The route data to aggregate.
        extra_notes_df: Extra notes to include in the manifest if tagged.

    Returns:
        Dictionary of aggregated data.
    """
    df = df.copy()

    df[Columns.BOX_TYPE] = df[Columns.BOX_TYPE].str.upper().str.strip()
    box_types = df[Columns.BOX_TYPE].unique()
    extra_box_types = set(box_types) - set(BoxType)
    if extra_box_types:
        raise ValueError(f"Invalid box type in route data: {extra_box_types}")

    extra_notes_list = []
    used_tags = []
    for _, row in extra_notes_df.iterrows():
        for notes in df[Columns.NOTES]:
            if row["tag"].upper() in notes.upper() and row["tag"] not in used_tags:
                used_tags.append(row["tag"])
                note = "* " + row["tag"].replace("*", "").strip() + ": " + row["note"]
                extra_notes_list.append(note)

    agg_dict = {
        "box_counts": df.groupby(Columns.BOX_TYPE)[Columns.ORDER_COUNT].sum().to_dict(),
        "total_box_count": df[Columns.ORDER_COUNT].sum(),
        "protein_box_count": (
            df[df[Columns.BOX_TYPE].isin(PROTEIN_BOX_TYPES)][Columns.ORDER_COUNT].sum()
        ),
        "neighborhoods": df[Columns.NEIGHBORHOOD].unique().tolist(),
        "extra_notes": extra_notes_list,
    }

    for box_type in BoxType:
        if box_type.value not in agg_dict["box_counts"]:
            agg_dict["box_counts"][box_type.value] = 0

    return agg_dict


@typechecked
def _make_manifest_sheet(
    wb: Workbook, agg_dict: dict, route_df: pd.DataFrame, sheet_name: str, sheet_idx: int
) -> None:
    """Create a manifest sheet."""
    ws = wb.create_sheet(title=str(sheet_name), index=sheet_idx)
    _add_header_row(ws=ws)
    neighborhoods_row_number = _add_aggregate_block(
        ws=ws, agg_dict=agg_dict, sheet_name=sheet_name
    )
    df_start_row = _write_data_to_sheet(ws=ws, df=route_df)
    _auto_adjust_column_widths(ws=ws, df_start_row=df_start_row)
    _word_wrap_columns(ws=ws)
    _merge_and_wrap_neighborhoods(ws=ws, neighborhoods_row_number=neighborhoods_row_number)
    _append_extra_notes(ws=ws, extra_notes=agg_dict["extra_notes"])

    # TODO: Set print_area (Use calculate_dimensions)
    # TODO: set_printer_settings(paper_size, orientation)
    # https://github.com/crickets-and-comb/bfb_delivery/issues/82


@typechecked
def _add_header_row(ws: Worksheet) -> None:
    """Append a reusable formatted row to the worksheet."""
    font = Font(bold=True)
    alignment_left = Alignment(horizontal="left")
    alignment_right = Alignment(horizontal="right")
    fill = PatternFill(
        start_color=CellColors.HEADER, end_color=CellColors.HEADER, fill_type="solid"
    )

    driver_support_phone = get_phone_number("driver_support")
    recipient_support_phone = get_phone_number("recipient_support")
    formatted_row = [
        {
            "value": f"DRIVER SUPPORT: {driver_support_phone}",
            "font": font,
            "alignment": alignment_left,
            "fill": fill,
        },
        {"value": "", "font": font, "alignment": None, "fill": fill},
        {"value": "", "font": font, "alignment": None, "fill": fill},
        {
            "value": f"RECIPIENT SUPPORT: {recipient_support_phone}",
            "font": font,
            "alignment": alignment_right,
            "fill": fill,
        },
        {"value": "", "font": font, "alignment": None, "fill": fill},
        {
            "value": "PLEASE SHRED MANIFEST AFTER COMPLETING ROUTE.",
            "font": font,
            "alignment": alignment_right,
            "fill": fill,
        },
    ]

    for col_idx, col_data in enumerate(formatted_row, start=1):
        cell_value = col_data["value"] if isinstance(col_data["value"], str) else ""
        cell = ws.cell(row=1, column=col_idx, value=cell_value)
        if col_data["font"]:
            if isinstance(col_data["font"], Font):
                cell.font = col_data["font"]
        if col_data["alignment"]:
            if isinstance(col_data["alignment"], Alignment):
                cell.alignment = col_data["alignment"]
        if col_data["fill"]:
            if isinstance(col_data["fill"], PatternFill):
                cell.fill = col_data["fill"]

    return


@typechecked
def _add_aggregate_block(ws: Worksheet, agg_dict: dict, sheet_name: str) -> int:
    """Append left and right aggregation blocks to the worksheet row by row."""
    date = str(sheet_name).split(" ")[0]
    driver_name = " ".join(str(sheet_name).split(" ")[1:])

    # TODO: Yeah, let's use an enum for box types since the manifest is a contract.
    # https://github.com/crickets-and-comb/bfb_delivery/issues/78
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    alignment_left = Alignment(horizontal="left")
    alignment_right = Alignment(horizontal="right")
    bold_font = Font(bold=True)

    left_block = _get_left_block(date=date, driver_name=driver_name, agg_dict=agg_dict)
    right_block = _get_right_block(thin_border=thin_border, agg_dict=agg_dict)

    start_row = ws.max_row + 1
    neighborhoods_row_number = 0
    for i, (left_row, right_row) in enumerate(
        zip(left_block, right_block, strict=True), start=start_row
    ):
        for col_idx, cell_definition in enumerate(left_row, start=1):
            cell = ws.cell(row=i, column=col_idx, value=cell_definition["value"])
            cell.font = bold_font
            cell.alignment = alignment_left
            if cell_definition["value"] and cell_definition["value"].startswith(
                "Neighborhoods"
            ):
                neighborhoods_row_number = i

        for col_idx, cell_definition in enumerate(right_row, start=5):
            cell = ws.cell(row=i, column=col_idx, value=cell_definition["value"])
            cell.font = bold_font
            cell.alignment = alignment_right
            if isinstance(cell_definition["fill"], PatternFill):
                cell.fill = cell_definition["fill"]
            if isinstance(cell_definition["border"], Border):
                cell.border = cell_definition["border"]

    return neighborhoods_row_number


def _get_left_block(
    date: str, driver_name: str, agg_dict: dict
) -> list[list[dict[str, None]] | list[dict[str, str]]]:
    left_block = [
        [{"value": None}],
        [{"value": f"Date: {date}"}],
        [{"value": None}],
        [{"value": f"Driver: {driver_name}"}],
        [{"value": None}],
        [{"value": f"Neighborhoods: {', '.join(agg_dict['neighborhoods'])}"}],
        [{"value": None}],
    ]

    return left_block


def _get_right_block(thin_border: Border, agg_dict: dict) -> list[list[dict]]:
    right_block = [
        [{"value": None, "fill": None, "border": None}],
        [
            {
                "value": "BASIC",
                "fill": PatternFill(
                    start_color=CellColors.BASIC,
                    end_color=CellColors.BASIC,
                    fill_type="solid",
                ),
                "border": thin_border,
            },
            {
                "value": agg_dict["box_counts"].get("BASIC", 0),
                "fill": None,
                "border": thin_border,
            },
        ],
        [
            {
                "value": "LA",
                "fill": PatternFill(
                    start_color=CellColors.LA, end_color=CellColors.LA, fill_type="solid"
                ),
                "border": thin_border,
            },
            {
                "value": agg_dict["box_counts"].get("LA", 0),
                "fill": None,
                "border": thin_border,
            },
        ],
        [
            {
                "value": "GF",
                "fill": PatternFill(
                    start_color=CellColors.GF, end_color=CellColors.GF, fill_type="solid"
                ),
                "border": thin_border,
            },
            {
                "value": agg_dict["box_counts"].get("GF", 0),
                "fill": None,
                "border": thin_border,
            },
        ],
        [
            {
                "value": "VEGAN",
                "fill": PatternFill(
                    start_color=CellColors.VEGAN,
                    end_color=CellColors.VEGAN,
                    fill_type="solid",
                ),
                "border": thin_border,
            },
            {
                "value": agg_dict["box_counts"].get("VEGAN", 0),
                "fill": None,
                "border": thin_border,
            },
        ],
        [
            {"value": "TOTAL BOX COUNT=", "fill": None, "border": None},
            {"value": agg_dict["total_box_count"], "fill": None, "border": None},
        ],
        [
            {"value": "PROTEIN COUNT=", "fill": None, "border": None},
            {"value": agg_dict["protein_box_count"], "fill": None, "border": None},
        ],
    ]

    return right_block


@typechecked
def _write_data_to_sheet(ws: Worksheet, df: pd.DataFrame) -> int:
    """Write and format the dataframe itself."""
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    header_font = Font(bold=True)

    box_type_col_idx = df.columns.get_loc(Columns.BOX_TYPE)

    start_row = ws.max_row + 1
    for r_idx, row in enumerate(
        dataframe_to_rows(df[FORMATTED_ROUTES_COLUMNS], index=False, header=True),
        start=start_row,
    ):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border
            if r_idx == start_row:
                cell.font = header_font
                cell.alignment = Alignment(horizontal="left")

            if c_idx == box_type_col_idx and r_idx > start_row:
                box_type = str(value)
                fill_color = BOX_TYPE_COLOR_MAP.get(box_type)
                if fill_color:
                    cell.fill = PatternFill(
                        start_color=fill_color, end_color=fill_color, fill_type="solid"
                    )

    return start_row


@typechecked
def _auto_adjust_column_widths(ws: Worksheet, df_start_row: int) -> None:
    """Auto-adjust column widths to fit the dataframe."""
    for col in ws.columns:
        max_length = 0

        col_letter = col[0].column_letter
        padding_scalar = 1
        for cell in col:
            if cell.row >= df_start_row:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)) * padding_scalar)
                except Exception as e:
                    warnings.warn(f"Error while adjusting column widths: {e}", stacklevel=2)

        ws.column_dimensions[col_letter].width = max(8, round(max_length))

    return


@typechecked
def _word_wrap_columns(ws: Worksheet) -> None:
    """Word wrap the notes column, and set width."""
    start_row = 10
    end_row = ws.max_row
    _word_wrap_column(
        ws=ws,
        start_row=start_row,
        end_row=end_row,
        col_letter="C",
        width=ADDRESS_COLUMN_WIDTH,
    )
    _word_wrap_column(
        ws=ws, start_row=start_row, end_row=end_row, col_letter="E", width=NOTES_COLUMN_WIDTH
    )

    return


@typechecked
def _word_wrap_column(
    ws: Worksheet, start_row: int, end_row: int, col_letter: str, width: float
) -> None:
    """Word wrap column, and set width."""
    ws.column_dimensions[col_letter].width = width
    for row in ws[f"{col_letter}{start_row}:{col_letter}{end_row}"]:
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    return


@typechecked
def _merge_and_wrap_neighborhoods(ws: Worksheet, neighborhoods_row_number: int) -> None:
    """Merge the neighborhoods cell and wrap the text."""
    start_col = 1
    end_col = 3
    ws.merge_cells(
        start_row=neighborhoods_row_number,
        start_column=start_col,
        end_row=neighborhoods_row_number,
        end_column=end_col,
    )
    cell = ws.cell(row=neighborhoods_row_number, column=start_col)
    cell.alignment = Alignment(wrap_text=True, horizontal="left", vertical="top")

    set_row_height_of_wrapped_cell(cell=cell)

    return


@typechecked
def _append_extra_notes(ws: Worksheet, extra_notes: list[str]) -> None:
    """Append extra notes to the worksheet in the leftmost column.

    Places notes in column A and merges across all columns (A-F) with text wrapping.
    """
    start_row = ws.max_row + 2
    start_col = 1
    end_col = 6
    for i, note in enumerate(extra_notes, start=start_row):
        cell = ws.cell(row=i, column=start_col, value=note)
        cell.alignment = Alignment(wrap_text=True, horizontal="left", vertical="top")
        ws.merge_cells(start_row=i, start_column=start_col, end_row=i, end_column=end_col)

        set_row_height_of_wrapped_cell(cell=cell)

    return
