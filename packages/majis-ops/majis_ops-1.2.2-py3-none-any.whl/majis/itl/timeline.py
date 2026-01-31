"""MAJIS ITL timeline sub-module."""

import warnings
from collections import defaultdict, namedtuple
from pathlib import Path

from numpy import datetime64
from openpyxl import Workbook, load_workbook
from packaging.version import Version

from planetary_coverage import datetime
from planetary_coverage.events import EventsDict, EventsList
from planetary_coverage.html import table

from ..misc import fmt_datetime
from ..misc.events import flatten, flatten_events
from ..misc.evf import ref_key
from ..misc.time import fmt_timedelta
from .reader import read_itl

DEFAULT_TEMPLATE = Path(__file__).parent / 'timeline.xlsm'

BINNING = {
    '1': 'No Binning',
    '2': 'Binning x2',
    '4': 'Binning x4',
}

ITL_MAPPING = {
    'OBS_NAME': ('OBS_NAME', str),
    'start_angle': ('START_ANGLE', float),  # °
    'start_scan_speed': ('START_SCAN_SPEED', float),  # °/s
    'stop_scan_speed': ('STOP_SCAN_SPEED', float),  # °/s
    'Scanner step per frame': ('SYNCHRONOUS', float),  # (-3, 0, 3)
    'stop_angle': ('STOP_ANGLE', float),  # °
    'First CU_frame start (UTC)': ('t_start', fmt_datetime),  # YYYY-MM-DDThh:mm:ss.msZ
    'Last CU_frame stop (UTC)': ('t_end', fmt_datetime),  # YYYY-MM-DDThh:mm:ss.msZ
    'cu_trep_ms': ('CU_TREP', lambda s: int(str(s).replace('ms', ''))),  # ms
    'spatial_binning': ('BINNING', lambda i: BINNING[str(i)]),
    'nb_cu_frames_tot': ('CU_FRAME', int),
    'ppe': ('PPE', int),
    'Start Row VI': ('START_ROW_VIS', int),
    'prime': ('PRIME', lambda b: 'MAJIS' if b else 'other'),
    'Comments': ('COMMENTS', lambda s: s if s else None),
    'ITL name': ('ITL', lambda p: Path(p).name if p else None),
}

ChangeLog = namedtuple('Log', 'version,date,author,change')


class TimelineChangeLog:
    """MAJIS timeline change log."""

    def __init__(self, template: Workbook, sheet_name='change log'):
        self.log = self.read_changelog(list(template[sheet_name])[1:])

    def __str__(self):
        return '\n\n'.join(
            f'>>> {str(log.version):>6} | {log.date} | {log.author}\n{log.change}'
            for log in self.log
        )

    def __repr__(self):
        return str(self)

    def __getitem__(self, i):
        return self.log[i]

    def __iter__(self):
        return iter(self.log)

    def _repr_html_(self):
        return table(self.log, header=['Version', 'Date', 'Author', 'Changes'])

    def read_changelog(self, rows: list) -> (list, list):
        """Read changelog changes grouped by version, date and author."""
        log = defaultdict(list)

        # Group release changes
        key = None, None, None
        for row in rows:
            version, date, change, author = (cell.value for cell in row[:4])

            if change and change != '…':
                key = (
                    Version(str(version)) if version else key[0],
                    str(date.date()) if date else key[1],
                    str(author) if author else key[2],
                )

                log[key].extend(change.splitlines())

        # Reverse changes order (latest first) and add line returns between changes
        return [ChangeLog(*key, '\n'.join(log[key])) for key in reversed(log)]


class Timeline:
    """MAJIS timeline from template.

    You can either append an existing timeline template
    or use the default one.

    Warning
    -------
    Some extensions (`Data Validation`) are not included
    into the exported file. This is a openpyxl limitation.

    """

    sheet_name = 'Timeline'

    def __init__(
        self,
        observations: str | Path | EventsList | EventsDict | None = None,
        timeline: str | Path | None = None,
        ca_ref: str | dict | None = None,
        refs: dict | str | list | None = None,
    ):
        self.fname = Path(timeline) if timeline else DEFAULT_TEMPLATE
        self._load_data()

        # Set global reference time w.r.t. C/A
        self.ca_ref = ca_ref

        if observations:
            self.append(observations, refs=refs)

    def __repr__(self):
        return f"{self.__class__.__name__}('{self.fname}')"

    def __len__(self) -> int:
        return len(
            [
                obs
                for obs in self._timeline[self.fields['OBS_NAME']][self.header :]
                if obs.value
            ]
        )

    def __getitem__(self, item: str | int | tuple) -> list | dict | str | int | float:
        if isinstance(item, str):
            if item in self.fields:
                return [
                    col.value
                    for col in self._timeline[self.fields[item]][
                        self.header : self.header + len(self)
                    ]
                ]
            raise KeyError(f'Unknown `{item}` key')

        if isinstance(item, int):
            if 1 <= item <= len(self):
                return {
                    key: cell.value
                    for key, cell in zip(
                        self.fields, self._timeline[self.header + item], strict=True
                    )
                    if cell.value and not str(cell.value).startswith('=')
                }
            raise IndexError(f'Invalid index: {item} not between 1 and {len(self)}')

        if isinstance(item, tuple):
            name, i = item
            return self[name][i - 1]

        raise TypeError(
            f'Only `str`, `int` or `str, int` are accepted (`{type(item).__name__}` provided)'
        )

    def __setitem__(self, item: (str, int), value: str | int | float) -> None:
        name, i = item
        self._timeline[self.fields[name] + str(i + self.header)] = value

    def _repr_html_(self):
        return table(
            [
                [cell.value for cell in row]
                for row in self._timeline[f'{self.header + 1}:{self.header + len(self)}']
            ],
            header=list(self.fields),
        )

    def _load_data(self) -> None:
        """Load timeline data."""
        # Disable openpyxl warning about unsupported extension (`Data validation`)
        with warnings.catch_warnings():
            warnings.simplefilter('ignore')
            self._template = load_workbook(self.fname, keep_vba=True)

        # Load Timeline spreadsheet
        self._timeline = self._template[self.sheet_name]

        # Load Timeline header fields mapping
        self.fields = {cell.value: cell.column_letter for cell in self._timeline[1]}

        # Load template changelog
        if 'template change log' in self._template.sheetnames:
            sheet_name = 'template change log'
        else:
            sheet_name = 'change log'

        self.log = TimelineChangeLog(self._template, sheet_name=sheet_name)

    @property
    def version(self):
        """Template version from changelog."""
        return self.log[0].version

    @property
    def header(self):
        """Timeline header size."""
        return 2 if self.version < Version('2.0') else 3

    def append(
        self,
        observations: str | Path | EventsList | EventsDict,
        refs: dict | str | list | None = None,
    ) -> None:
        """Append ITL observations blocks."""
        if isinstance(observations, str | Path):
            observations = read_itl(observations, refs=refs)

        for i, obs in enumerate(flatten(observations), start=len(self) + 1):
            # Required fields
            for field, (key, fmt) in ITL_MAPPING.items():
                self[field, i] = fmt(obs[key])

            # Additional fields
            self['Mirror Flag', i] = (
                'ENABLE'
                if float(obs['START_SCAN_SPEED']) or float(obs['STOP_SCAN_SPEED'])
                else 'DISABLE'
            )

            self['First CU_frame start wrt C/A', i] = self._fmt_timedelta(obs.start)
            self['Last CU_frame stop wrt C/A', i] = self._fmt_timedelta(obs.stop)

    @property
    def ca_ref(self) -> datetime:
        """C/A reference for relative time."""
        return self._ca_ref

    @ca_ref.setter
    def ca_ref(self, ca_ref):
        """Set relative time w.r.t. C/A reference.

        Warning
        -------
        All the previous relative values will be recomputed if this value is changed.

        """
        if ca_ref:
            _, self._ca_ref = ref_key(ca_ref)

            # Recompute relative values for all observation w.r.t. C/A reference
            for i in range(1, len(self) + 1):
                self['First CU_frame start wrt C/A', i] = self._fmt_timedelta(
                    self['First CU_frame start (UTC)', i]
                )
                self['Last CU_frame stop wrt C/A', i] = self._fmt_timedelta(
                    self['Last CU_frame stop (UTC)', i]
                )
        else:
            self._ca_ref = None

    def _fmt_timedelta(self, t: str | datetime64) -> str | None:
        """Compute relative time w.r.t. C/A reference."""
        return fmt_timedelta(datetime(t) - self.ca_ref) if self.ca_ref else None

    def save(self, fout: str | Path | None = None) -> Path:
        """Save MAJIS timeline."""
        if fout:
            fout = Path(fout)
        elif self.fname != DEFAULT_TEMPLATE:
            fout = self.fname
        else:
            raise FileExistsError(
                'Can not overwrite default template. Please provide and export filename.'
            )

        self._template.save(fout)

        return fout

    @property
    def science(self) -> TimelineChangeLog:
        """Science change log.

        Raises
        ------
        ValueError:
            Only available for template ≥ 2.0

        """
        if self.version < Version('2.0'):
            raise ValueError(
                f'Science changelog is only available in template ≥ 2.0 (current: `{self.version}`)'
            )

        return TimelineChangeLog(self._template, sheet_name='science change log')


@flatten_events
def save_itl_xlsm(
    fout: str | Path | None,
    *events: EventsList | EventsDict,
    ca_ref: str | dict | None = None,
    overlap: bool = False,
    timeline: str | Path | None = None,
) -> Path:
    """Save ITL events to XLSM timeline.

    If a timeline is provided but no explicit output file
    the output file will be same in the original timeline.

    """
    return Timeline(events, timeline=timeline, ca_ref=ca_ref).save(fout)
