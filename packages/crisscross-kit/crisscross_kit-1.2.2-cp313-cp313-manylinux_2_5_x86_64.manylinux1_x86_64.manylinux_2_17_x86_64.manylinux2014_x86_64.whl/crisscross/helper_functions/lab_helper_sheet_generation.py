from collections import defaultdict, OrderedDict
import numpy as np
from colorama import Fore
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
import math
import matplotlib.pyplot as plt
import platform
import os
import pandas as pd
from tqdm import tqdm
from math import floor
import seaborn as sns

from crisscross.helper_functions import create_dir_if_empty
from crisscross.helper_functions.simple_plate_visuals import visualize_plate_with_color_labels

# consistent figure formatting between mac, windows and linux
if platform.system() == 'Darwin':
    plt.rcParams.update({'font.sans-serif': 'Helvetica'})
elif platform.system() == 'Windows':
    plt.rcParams.update({'font.sans-serif': 'Arial'})
else:
    plt.rcParams.update({'font.sans-serif': 'DejaVu Sans'}) # should work with linux


def next_excel_column_name(n):
    """Given a 0-based index, return the Excel-style column name."""
    result = ""
    while n >= 0:
        result = chr(n % 26 + ord('A')) + result
        n = n // 26 - 1
    return result

red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

def apply_box_border(ws, top_left, top_right, bottom_left, bottom_right, style='thick'):
    """
    Applies a thick border to an excel sheet surrounding the specified cells.
    :param ws: Excel worsheet object.
    :param top_left: Top left cell.
    :param top_right: Top right cell.
    :param bottom_left: Bottom left cell.
    :param bottom_right: Bottom right cell.
    :param style: Border style to be applied.
    :return: N/A, applied in-place.
    """

    selected_border = Side(border_style=style)

    # Top row
    for cell in ws[f"{top_left}:{top_right}"][0]:
        cell.border = Border(top=selected_border)
    # Bottom row
    for cell in ws[f"{bottom_left}:{bottom_right}"][0]:
        cell.border = Border(bottom=selected_border)
    # Left column
    for cell in ws[f"{top_left}:{bottom_left}"]:
        cell[0].border = Border(left=selected_border)
    # Right column
    for cell in ws[f"{top_right}:{bottom_right}"]:
        cell[0].border = Border(right=selected_border)

    # Top-left corner
    ws[top_left].border = Border(top=selected_border, left=selected_border)
    # Top-right corner
    ws[top_right].border = Border(top=selected_border, right=selected_border)
    # Bottom-left corner
    ws[bottom_left].border = Border(bottom=selected_border, left=selected_border)
    # Bottom-right corner
    ws[bottom_right].border = Border(bottom=selected_border, right=selected_border)

def adjust_column_width(ws):
    """
    Adjusts the column width of an excel sheet based on the maximum length of the content in each column.
    :param ws: Excel sheet object.
    :return: N/A, adjusted in-place.
    """
    # Adjust the column width based on the maximum length of the content in each column
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter

        # Find the maximum length of the content in the column
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        # Set the column width to the max_length
        adjusted_width = max_length
        ws.column_dimensions[column_letter].width = adjusted_width

def find_cell_by_value(ws, search_value):
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            if cell.value == search_value:
                return cell.coordinate
    return None

def prepare_master_mix_sheet(slat_dict, echo_sheet=None, reference_handle_volume=150, reference_handle_concentration=500,
                             slat_mixture_volume=50, unique_transfer_volume_plates=None, workbook=None,
                             handle_mix_ratio=10, split_core_staple_pools=False):
    """
    Prepares a 'master mix' sheet to be used for combining slat mixtures with scaffold
    and core staples into the final slat mixture.

    :param slat_dict: Dictionary of slats with slat names as keys and slat objects as values.
    :param echo_sheet: Exact list of commands sent to the Echo robot for this group of slats.
    :param reference_handle_volume: Reference staple volume for each handle in a pool in nL (this refers to the control handles plate).
    :param reference_handle_concentration: Reference staple concentration used for the core staples in uM (this refers to the control handles plate).
    All concentration values will be referenced to this value.
    :param slat_mixture_volume: Reaction volume (in uL) for a single slat annealing mixture.  Can be set to 'max' to use up all available handle mix.
    :param handle_mix_ratio: Ratio of handle mix concentration to scaffold concentration (default is 10).
    :param unique_transfer_volume_plates: Plates that have special non-standard volumes.
    This will be ignored if the echo sheet is provided with the exact details.
    :param workbook: The workbook to which the new excel sheet should be added.
    :param split_core_staple_pools: If True, the core staples will be assumed to have been split into 4 pools (S0, S1, S3, S4).
    :return: Workbook with new sheet included.
    """

    slat_count = len(slat_dict)
    slat_concentration_distribution = []
    handle_volume_distribution = []

    # calculate the minimum handle concentration for each slat (from the echo pools)

    # preference is to use echo sheet for calculation if available (as this has exceptions applied which won't be apparent in the slat dictionary)
    if echo_sheet is not None:
        for slat_name in slat_dict.keys():
            echo_indices = echo_sheet[echo_sheet['Component'].str.contains(fr"{slat_name}_", na=False)].index
            total_handle_mix_volume = sum(echo_sheet.loc[echo_indices]['Transfer Volume'].values)
            slat_concentration_nM = 1000 * (reference_handle_concentration * reference_handle_volume) / total_handle_mix_volume
            slat_concentration_distribution.append(slat_concentration_nM)
            handle_volume_distribution.append(total_handle_mix_volume)
    else:
        for slat in slat_dict.values():
            total_handle_mix_volume = 0
            for handle in list(slat.H2_handles.values()) + list(slat.H5_handles.values()):
                if handle['plate'] in unique_transfer_volume_plates:
                    total_handle_mix_volume += unique_transfer_volume_plates[handle['plate']]
                else:
                    total_handle_mix_volume += reference_handle_volume * int(reference_handle_concentration / handle['concentration'])
            slat_concentration_nM = 1000 * (reference_handle_concentration * reference_handle_volume) / total_handle_mix_volume
            slat_concentration_distribution.append(slat_concentration_nM)
            handle_volume_distribution.append(total_handle_mix_volume)

    min_handle_mix_conc = min(slat_concentration_distribution)
    max_handle_mix_conc = max(slat_concentration_distribution)

    print(Fore.BLUE + f'Info: Lowest handle mixture concentration: {round(min_handle_mix_conc, 1)}nM ({max(handle_volume_distribution)/1000} μl total), highest handle mixture concentration: {round(max_handle_mix_conc, 1)}nM ({min(handle_volume_distribution)/1000} μl total).' + Fore.RESET)

    if (max_handle_mix_conc - min_handle_mix_conc) / min_handle_mix_conc > 0.2:
        print(Fore.MAGENTA + f'Warning: The handle mixtures generated have a wide concentration range. '
                             f'You could save on some staples by splitting the master mix preparation '
                             f'into two or more batches.' + Fore.RESET)
    if workbook is not None:
        wb = workbook
    else:
        wb = Workbook()

    ws = wb.create_sheet("Slat Folding & Master Mix")

    core_staple_mix_count = 4 if split_core_staple_pools else 1

    # Titles and formatting
    ws['A1'] = 'Single Slat Folding Quantities (can prepare directly or follow master mix details below)'
    ws['A1'].font = Font(bold=True)
    ws['A2'] = 'H2/H5 Handles from Echo [individual handle] (nM)'
    ws['A3'] = 'P8064 scaffold (nM)'

    if not split_core_staple_pools:
        ws['A4'] = 'Core staples pool (each, nM)'
    else:
        ws['A4'] = 'Core staples pool S0 (each, nM)'
        ws['A5'] = 'Core staples pool S1 (each, nM)'
        ws['A6'] = 'Core staples pool S3 (each, nM)'
        ws['A7'] = 'Core staples pool S4 (each, nM)'

    ws[f'A{4+core_staple_mix_count}'] = 'TEF (X)'
    ws[f'A{5+core_staple_mix_count}'] = 'MgCl2 (mM)'
    ws[f'A{6+core_staple_mix_count}'] = 'UPW (deionized water)'
    ws[f'A{7+core_staple_mix_count}'] = 'Total volume (µL)'
    ws[f'A{8+core_staple_mix_count}'] = 'Total amount (pmol)'

    # Standard concentration values
    ws['B1'] = 'Stock Concentration'
    ws['B2'] = round(min_handle_mix_conc, 1)
    ws['B3'] = 1062
    if not split_core_staple_pools:
        ws['B4'] = 3937
    else:
        ws['B4'] = 10000
        ws['B5'] = 10000
        ws['B6'] = 10000
        ws['B7'] = 10000

    ws[f'B{4+core_staple_mix_count}'] = 10
    ws[f'B{5+core_staple_mix_count}'] = 1000
    ws[f'B2'].fill = orange_fill
    ws[f'B3'].fill = red_fill

    # Standard target concentrations
    ws['C1'] = 'Final Concentration'
    ws['C2'] = 50 * handle_mix_ratio
    ws['C3'] = 50
    ws['C4'] = 500
    if core_staple_mix_count:
        ws['C5'] = 500
        ws['C6'] = 500
        ws['C7'] = 500

    ws[f'C{4+core_staple_mix_count}'] = 1
    ws[f'C{5+core_staple_mix_count}'] = 6

    # Calculations for single slat volumes
    ws['D1'] = 'Amount to Add (µL)'
    for cell_col in range(2,6+core_staple_mix_count):
        ws[f'D{cell_col}'] = f"=round(C{cell_col}*D${7+core_staple_mix_count}/B{cell_col},2)"

    ws[f'D{6+core_staple_mix_count}'] = f'=D{7+core_staple_mix_count}-sum(D2:D{5+core_staple_mix_count})'

    if slat_mixture_volume == "max":
        # selects the max reaction volume that reduces staple pool waste, assuming echo transfers only 75% of expected
        # 0.75 * total femtomoles of slats (based on min conc)/reaction final conc (typically 500 nM), rounded down to the nearest multiple of 5
        ws[f'D{7+core_staple_mix_count}'] = math.floor(0.75 * min_handle_mix_conc * total_handle_mix_volume/ws['C2'].value/5/1000) * 5
        print(Fore.BLUE + f'Info: You selected "max", so the slat mixtures will use all the staples in your pooled handle mixtures.  The output slat mixture volume will be {ws["D8"].value}μl.' + Fore.RESET)
    else:
        ws[f'D{7+core_staple_mix_count}'] = slat_mixture_volume # simplest solution: user decides, defaults to 50 µL

    ws[f'D{8+core_staple_mix_count}'] = f'=D{7+core_staple_mix_count}*C3/1000'

    ws['F2'].fill = red_fill
    ws['G2'] = 'Cells with this shading should be adjusted to match the actual values in your experiment.'
    ws['F3'].fill = orange_fill
    ws['G3'] = 'Cells with this shading contain an average or minimum value for a group of slats - if further precision is required for each slat, this needs to be changed.'

    apply_box_border(ws, 'A1', 'D1', f'A{8+core_staple_mix_count}', f'D{8+core_staple_mix_count}')

    # Calculations for master mix
    ws[f'A{10+core_staple_mix_count}'] = 'Master Mix Preparation (prepare once)'
    ws[f'A{10+core_staple_mix_count}'].font = Font(bold=True)
    ws[f'B{10+core_staple_mix_count}'] = 'Count or Volume (µL)'
    ws[f'A{11+core_staple_mix_count}'] = 'Number of slats (with a buffer of 3 extra slats)'

    ws[f'A{12+core_staple_mix_count}'] = 'P8064 scaffold'

    if not split_core_staple_pools:
        ws[f'A{13+core_staple_mix_count}'] = 'Core staples pool'
    else:
        ws[f'A{13+core_staple_mix_count}'] = 'Core staples pool S0'
        ws[f'A{14+core_staple_mix_count}'] = 'Core staples pool S1'
        ws[f'A{15+core_staple_mix_count}'] = 'Core staples pool S3'
        ws[f'A{16+core_staple_mix_count}'] = 'Core staples pool S4'

    ws[f'A{13+(2*core_staple_mix_count)}'] = 'TEF'
    ws[f'A{14+(2*core_staple_mix_count)}'] = 'MgCl2'
    ws[f'A{15+(2*core_staple_mix_count)}'] = 'UPW (deionized water)'
    ws[f'A{16+(2*core_staple_mix_count)}'] = 'Total volume'

    ws[f'B{11+core_staple_mix_count}'] = slat_count + 3

    for component_index, cell_col in enumerate(range(12+core_staple_mix_count, 16+(2*core_staple_mix_count))):
        ws[f'B{cell_col}'] = f"=D{3+component_index}*B{11+core_staple_mix_count}"

    ws[f'B{16+(2*core_staple_mix_count)}'] = f'=SUM(B{12+core_staple_mix_count}:B{15+(2*core_staple_mix_count)})'

    apply_box_border(ws, f'A{10+core_staple_mix_count}', f'B{10+core_staple_mix_count}', f'A{16+(2*core_staple_mix_count)}', f'B{16+(2*core_staple_mix_count)}')

    special_border = Side(border_style='mediumDashDotDot')
    thick_border = Side(border_style='thick')

    ws[f'A{11+core_staple_mix_count}'].border = Border(top=special_border, left=thick_border, bottom=special_border)
    ws[f'B{11+core_staple_mix_count}'].border = Border(top=special_border, right=thick_border, bottom=special_border)

    # # Calculations for master mix + handle mix
    ws[f'A{18+(2*core_staple_mix_count)}'] = 'Final Slat Mixture (prepare once for each slat)'
    ws[f'A{18+(2*core_staple_mix_count)}'].font = Font(bold=True)
    ws[f'B{18+(2*core_staple_mix_count)}'] = 'Volume (µL)'

    ws[f'A{19+(2*core_staple_mix_count)}'] = 'Master Mix'
    ws[f'A{20+(2*core_staple_mix_count)}'] = 'Slat Handle Mixture'
    ws[f'A{21+(2*core_staple_mix_count)}'] = 'Total volume'

    ws[f'B{19+(2*core_staple_mix_count)}'] = f'=SUM(B{12+core_staple_mix_count}:B{15+(2*core_staple_mix_count)})/B{11+core_staple_mix_count}'
    ws[f'B{20+(2*core_staple_mix_count)}'] = '=D2'
    ws[f'B{21+(2*core_staple_mix_count)}'] = f'=B{19+(2*core_staple_mix_count)}+B{20+(2*core_staple_mix_count)}'

    apply_box_border(ws, f'A{18+(2*core_staple_mix_count)}', f'B{18+(2*core_staple_mix_count)}', f'A{21+(2*core_staple_mix_count)}', f'B{21+(2*core_staple_mix_count)}')

    adjust_column_width(ws)

    return wb

def prepare_peg_purification_sheet(slat_dict, groups_per_layer=2, max_slat_concentration_uM=2,
                                   slat_mixture_volume=50, workbook=None,
                                   echo_sheet=None, special_slat_groups=None, peg_concentration=2):
    """
    Prepares standard instructions for combining and purifying slat mixtures using PEG purification.  Also prepares lists of slat groups as a reference for when in the lab.
    :param slat_dict: Dictionary of slats with slat names as keys and slat objects as values.
    :param groups_per_layer: Number of PEG groups to use per crisscross layer.  You might want to adjust this if you have too many slats together in one group.
    :param max_slat_concentration_uM: Maximum concentration of slats in a combined PEG mixture (in UM) before a warning is triggered.
    :param slat_mixture_volume: Reaction volume (in uL) for a single slat annealing mixture.
    :param workbook: The workbook to which the new excel sheet should be added.
    :param echo_sheet: Exact list of commands sent to the Echo robot for this group of slats.
    :param special_slat_groups: IDs of slats that should be separated from the general slat groups and placed in their own group.
    :param peg_concentration: PEG concentration (in terms of X) to be used as the stock solution for the purification step.
    :return: Workbook with new sheet included.
    """
    if workbook is not None:
        wb = workbook
    else:
        wb = Workbook()

    ws = wb.create_sheet("PEG Purification")

    layer_groups = defaultdict(list)
    full_data_groups = OrderedDict()

    for slat in slat_dict.values():
        if slat.ID not in layer_groups[slat.layer]:
            layer_groups[slat.layer].append(slat.ID)

    # Assigns slats to a group based on the number of groups per layer
    for layer, slats in layer_groups.items():
        if len(slats) < groups_per_layer:
            full_data_groups[f'L{layer}-ALL'] = {'IDs': slats}
        else:
            start_point = 0
            number_jump = len(slats) // groups_per_layer
            for i in range(1, groups_per_layer+1):
                full_data_groups[f'L{layer}-G{i}'] = {'IDs': slats[start_point:start_point+number_jump]}
                start_point += number_jump

    if special_slat_groups is not None:
        for special_group, slats in special_slat_groups.items():
            for slat in slats:
                for standard_group in full_data_groups:
                    if slat in full_data_groups[standard_group]['IDs']:
                        full_data_groups[standard_group]['IDs'].remove(slat)
            full_data_groups[special_group] = {'IDs': slats}

    # cleaning up empty groups
    keys_to_remove = []
    for position, group in enumerate(full_data_groups.keys()):
        if len(full_data_groups[group]['IDs']) == 0:  # this means that all slats in this group were moved to a special group and so this group is now empty
            keys_to_remove.append(group)
    for key in keys_to_remove:
        del full_data_groups[key]

    # Titles and formatting
    ws['A1'] = 'PEG Purification'
    ws['A1'].font = Font(bold=True)
    ws['A2'] = 'Step 0: Combine all slats into groups (slat IDs for each group beneath table)'
    ws['A3'] = 'Volume extracted from each well (µl)'
    ws['A4'] = '# of slats'
    ws['A5'] = 'Total volume expected (µL)'
    ws['A6'] = 'Scaffold Concentration per slat (nM)'
    ws['A7'] = 'Expected total origami amount (pmol, all)'
    ws['A8'] = 'Original Mg conc (mM)'
    ws['A9'] = 'Target Final Mg conc (mM)'
    ws['A10'] = 'Step 1: Add 1M Mg'
    ws['A11'] = 'Amount of 1M Mg to add (µl)'
    ws['A12'] = f'Step 2: Add {peg_concentration}X PEG'
    ws['A13'] = f'Amount of {peg_concentration}X PEG to add (µl)'
    ws['A14'] = 'Final volume (DNA + Mg +PEG combined) (µl)'

    ws['A15'] = 'Step 3: SPIN FOR 30 MINS AT 16KG, RT'
    ws['A16'] = 'Step 4: REMOVE SUPERNATANT AND ADD 150ul of RESUS1'
    ws['A17'] = 'Step 5: SPIN AGAIN FOR 30 MINS AT 16KG, RT'
    ws['A18'] = 'Step 6: REMOVE SUPERNATANT AND RESUSPEND IN RESUS2 AS BELOW'
    ws['A19'] = 'Desired final concentration (nM, per slat)'
    ws['A20'] = 'Resuspend with Resus2 to achieve target concentration for each slat (µl)'
    ws['A21'] = 'Expected total slat concentration (µM)'
    ws['A22'] = 'Step 7: SHAKE AT 33C FOR 1 HOUR AT 1000RPM, THEN NANODROP'
    ws['A23'] = 'Final Nanodrop (1x dilution - ng/µl dsDNA)'
    ws['A24'] = 'Average slat molecular weight (Da)'
    ws['A25'] = 'Total concentration from Nanodrop (µM)'
    ws['A26'] = 'Estimated concentration of each individual slat (nM)'
    ws['A27'] = 'Total amount of each slat (pmol)'
    ws['A28'] = 'Total origami (pmol)'
    ws['A29'] = 'PEG Yield (%)'

    # merge and center
    for cell in ['A2', 'A10', 'A12', 'A15', 'A16', 'A17', 'A18', 'A22']:
        ws.merge_cells(f'{cell}:{next_excel_column_name(len(full_data_groups))}{cell[1:]}')
        ws[cell].alignment = Alignment(horizontal='center', vertical='center')
        ws[cell].font = Font(bold=True)

    # fills in the equations and data for each group
    for position, group in enumerate(full_data_groups.keys()):
        column = next_excel_column_name(position+1)
        # block 1 - slat counts and volumes
        full_data_groups[group][f'{column}1'] = group
        try:
            coordinate = find_cell_by_value(wb["Slat Folding & Master Mix"], 'Total volume (µL)')
            full_data_groups[group][f'{column}3'] = wb["Slat Folding & Master Mix"][f"D{coordinate[1:]}"].value
        except KeyError:
            print("No Slat Folding & Master Mix sheet detected. Defaulting to 50 µL reaction volume...")
            full_data_groups[group][f'{column}3'] = slat_mixture_volume

        full_data_groups[group][f'{column}4'] = len(full_data_groups[group]['IDs'])
        full_data_groups[group][f'{column}5'] = f"={column}3*{column}4"
        full_data_groups[group][f'{column}6'] = 50
        ws[f'{column}6'].fill = orange_fill

        # block 2 - magnesium addition
        full_data_groups[group][f'{column}7'] = f"={column}5*{column}6/1000"
        full_data_groups[group][f'{column}8'] = 6
        ws[f'{column}8'].fill = orange_fill
        full_data_groups[group][f'{column}9'] = 10 * (peg_concentration/(peg_concentration-1))
        full_data_groups[group][f'{column}11'] = f"=round(({column}9-{column}8)*{column}5/(1000-{column}9),2)"
        full_data_groups[group][f'{column}13'] = f"=({column}11 + {column}5)/({peg_concentration}-1)"
        full_data_groups[group][f'{column}14'] = f"={column}11 + {column}5 + {column}13"
        ws[f'{column}14'].font = Font(bold=True)

        # block 3 - resuspension calculations
        full_data_groups[group][f'{column}19'] = 100
        ws[f'{column}19'].fill = blue_fill
        full_data_groups[group][f'{column}20'] = f"=ROUND((({column}7/{column}4)/{column}19)*1000,1)"
        full_data_groups[group][f'{column}21'] = f"=({column}19 * {column}4)/1000"
        rule = CellIsRule(operator='greaterThan', formula=[f'{max_slat_concentration_uM}'], fill=red_fill)
        ws.conditional_formatting.add(f'{column}21', rule)
        ws[f'{column}23'].fill = green_fill

        # block 4 - MW and concentration calculations
        mw_total = 0
        for id in full_data_groups[group]['IDs']:
            mw_total += slat_dict[id].get_molecular_weight()
        full_data_groups[group][f'{column}24'] = mw_total / len(full_data_groups[group]['IDs'])

        full_data_groups[group][f'{column}25'] = f'=round(({column}23*1000)/{column}24,2)'
        full_data_groups[group][f'{column}26'] = f'=round({column}25/{column}4*1000,2)'
        full_data_groups[group][f'{column}27'] = f'=round({column}26*{column}20/1000,2)'
        full_data_groups[group][f'{column}28'] = f'={column}27*{column}4'
        full_data_groups[group][f'{column}29'] = f'={column}28/{column}7*100'

    # sidebar definitions
    sidebar_col_start = next_excel_column_name(position+3)
    sidebar_col_2 = next_excel_column_name(position + 4)
    ws[f'{sidebar_col_start}3'].fill = red_fill
    ws[f'{sidebar_col_2}3'] = 'If these cells are red, then your slat mixture is over the 2µM limit - there is a high chance the mixture will aggregate.'
    ws[f'{sidebar_col_start}5'].fill = blue_fill
    ws[f'{sidebar_col_2}5'] = 'Change the target concentration of these cells to reduce the concentration below the 2µM limit.'

    ws[f'{sidebar_col_start}4'].fill = orange_fill
    ws[f'{sidebar_col_2}4'] = 'If for any reason slat folding conditions are changed, make sure to update these cells.'

    ws[f'{sidebar_col_start}6'].fill = green_fill
    ws[f'{sidebar_col_2}6'] = 'Fill these cells with your nanodrop values.  It is suggested to dilute by 10 for nanodrop if you have more than 5 slats in one mixture.'

    # fills out the data for each marked cell
    for _, group in full_data_groups.items():
        for cell, value in group.items():
            if cell != 'IDs':
                ws[cell] = value

    # resus 1/2 handy values
    ws['A31'] = 'Resus1/2 Buffer Components'
    ws['A31'].font = Font(bold=True)
    ws['A32'] = '10X TEF'
    ws['A33'] = 'MgCl2 (mM)'
    ws['A34'] = 'UPW (deionized water)'
    ws['A35'] = 'Total Volume'
    ws['B31'] = 'Stock'
    ws['C31'] = 'Resus 1'
    ws['D31'] = 'Resus 2'

    ws['B32'] = 10
    ws['B33'] = 1000
    ws['C35'] = 2000
    ws['D35'] = 2000

    ws['C32'] = '=1*C35/B32'
    ws['D32'] = '=1*D35/B32'

    ws['C33'] = '=ROUND(C35*20/B33,2)'
    ws['D33'] = '=ROUND(D35*10/B33,2)'

    ws['C34'] = '=C35 - C33 - C32'
    ws['D34'] = '=D35 - D33 - D32'

    apply_box_border(ws, 'A31', 'D31', 'A35', 'D35')
    apply_box_border(ws, 'A1', f'{column}1', 'A29', f'{column}29')

    # slat group components and values
    ws['A37'] = 'Slat Group Components'
    ws.merge_cells(f'A37:D37')
    ws['A37'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A37'].font = Font(bold=True)

    ws['A38'] = 'Group Name'
    ws['A38'].font = Font(bold=True)
    ws['B38'] = 'Slat ID'
    ws['B38'].font = Font(bold=True)
    ws['C38'] = 'Slat Well'
    ws['C38'].font = Font(bold=True)
    ws['D38'] = 'Plate Name'
    ws['D38'].font = Font(bold=True)

    indexer = 39
    divider = Side(border_style='thick')
    for group in full_data_groups.keys():
        for id in full_data_groups[group]['IDs']:
            ws[f'A{indexer}'] = group
            ws[f'B{indexer}'] = id
            # if echo data available, can also point towards the exact plate wells
            if echo_sheet is not None:
                echo_index = echo_sheet[echo_sheet['Component'].str.contains(fr"{id}_h\d_", na=False)].index[0]
                ws[f'C{indexer}'] = echo_sheet.loc[echo_index]['Destination Well']
                ws[f'D{indexer}'] = echo_sheet.loc[echo_index]['Destination Plate Name']
            indexer += 1
        for cell in ws[f"A{indexer-1}:D{indexer-1}"][0]:
            cell.border = Border(bottom=divider)

    adjust_column_width(ws)

    return wb

def prepare_all_standard_sheets(slat_dict, save_filepath, reference_single_handle_volume=150,
                                reference_single_handle_concentration=500,
                                slat_mixture_volume=50,
                                peg_groups_per_layer=2,
                                peg_concentration=2,
                                echo_sheet=None,
                                max_slat_concentration_uM=2,
                                unique_transfer_volume_plates=None,
                                special_slat_groups=None,
                                handle_mix_ratio=10,
                                split_core_staple_pools=False
                                ):
    """
    Prepares a series of excel sheets to aid lab assembler while preparing and purifying slat mixtures.
    :param slat_dict: Dictionary of slats to be assembled (each item in the dict is a Slat Object containing all 64 handles in place)
    :param save_filepath: Output file path for the combined excel workbook
    :param reference_single_handle_volume: Reference staple volume for each handle in a pool in nL (this refers to the control handles plate).
    :param reference_single_handle_concentration: Reference staple concentration used for the core staples in uM (this refers to the control handles plate).
    All concentration values will be referenced to this value.
    :param slat_mixture_volume: Reaction volume (in uL) for a single slat annealing mixture. Can be set to 'max' to use up all available handle mix.
    :param peg_groups_per_layer: Number of PEG groups to use per crisscross layer.  You might want to adjust this if you have too many slats together in one group.
    :param peg_concentration: PEG concentration (in terms of X) to be used as the stock solution for the purification step.
    :param echo_sheet: Exact echo commands to use as a reference for calculating slat concentrations.
    :param max_slat_concentration_uM: Maximum concentration of slats in a combined PEG mixture (in UM) before a warning is triggered.
    :param unique_transfer_volume_plates: Plates that have special non-standard volumes.  This will be ignored if the echo sheet is provided with the exact details.
    :param special_slat_groups: IDs of slats that should be separated from the general slat groups and placed in their own group.
    :param handle_mix_ratio: Ratio of handle mix concentration to scaffold concentration (default is 10).
    :param split_core_staple_pools: If True, the core staples will be assumed to have been split into 4 pools (S0, S1, S3, S4).
    :return: N/A, file saved directly to disk.
    """


    wb = Workbook()
    wb.remove(wb["Sheet"])

    clean_slat_dict =  {k:v for k,v in slat_dict.items() if v.phantom_parent is None}

    # prepares slat assembly mixture details
    prepare_master_mix_sheet(clean_slat_dict, echo_sheet, reference_single_handle_volume, reference_single_handle_concentration,
                             slat_mixture_volume, unique_transfer_volume_plates, wb, handle_mix_ratio=handle_mix_ratio,
                             split_core_staple_pools=split_core_staple_pools)

    # prepares slat purification details
    prepare_peg_purification_sheet(clean_slat_dict, peg_groups_per_layer, max_slat_concentration_uM, slat_mixture_volume,
                                   wb, peg_concentration=peg_concentration, echo_sheet=echo_sheet, special_slat_groups=special_slat_groups)
    wb.save(save_filepath)


def prepare_liquid_handle_plates_multiple_files(output_directory, file_list=None, extract_all_from_folder=None,
                                                target_concentration_uM=1000, volume_cap_ul=120,
                                                target_concentration_per_plate=None,
                                                max_commands_per_file=None,
                                                plot_distribution_per_plate=True, plate_size='384'):
    """
    Generates resuspension maps for all provided DNA spec files.
    :param output_directory: Output folder to save results.
    :param file_list: Specific list of filepaths to assess.
    :param extract_all_from_folder: Alternatively, specify a folder and all excel sheets will be extracted from the folder.
    :param target_concentration_uM: Target concentration for plate resuspension.
    :param volume_cap_ul: Maximum volume to resuspend (after which the volume is kept constant and the concentration is raised instead).
    :param target_concentration_per_plate: Set to a dictionary of concentrations per plate if different concentrations are desired for different plates.
    :param max_commands_per_file: Maximum commands that can be taken in by the liquid handler in one go.  If this is exceeded, files are split into different components.
    :param plot_distribution_per_plate: If true, generate a volume distribution plot for each plate.
    :param plate_size: Specify the size of the plate to generate.
    :return: N/A
    """

    if file_list is None and extract_all_from_folder is None:
        raise ValueError("You must provide either a list of files or a folder to extract files from.")

    if file_list is None or len(file_list) == 0:
        file_list = [os.path.join(extract_all_from_folder, f) for f in os.listdir(extract_all_from_folder) if f.endswith('.xlsx') and not f.startswith('~$')]


    for target_file in tqdm(file_list, total=len(file_list), desc='Processing files...'):
        prepare_liquid_handler_plate_resuspension_map(target_file, output_directory, target_concentration_uM=target_concentration_uM,
                                                      volume_cap_ul=volume_cap_ul, max_commands_per_file=max_commands_per_file,
                                                      target_concentration_per_plate=target_concentration_per_plate,
                                                      plot_distribution_per_plate=plot_distribution_per_plate, plate_size=plate_size)


def prepare_liquid_handler_plate_resuspension_map(filename, output_directory, target_concentration_uM=1000,
                                                  volume_cap_ul=120, max_commands_per_file=None,
                                                  plot_distribution_per_plate=True, target_concentration_per_plate=None,
                                                  plate_size='384'):
    """
    Generates a visual plate map and resuspension instructions for an entire plate of DNA oligos.
    The amount of DNA per well should be specified in an excel file using the standard IDT format.
    :param filename: Excel file containing DNA spec sheets (can contain multiples plates per sheet).
    :param output_directory: Output folder to save results.
    :param target_concentration_uM: Target concentration for plate resuspension.
    :param volume_cap_ul: Maximum volume to resuspend (after which the volume is kept constant and the concentration is raised instead).
    :param max_commands_per_file: Maximum commands that can be taken in by the liquid handler in one go.  If this is exceeded, files are split into different components.
    :param plot_distribution_per_plate: If true, generate a volume distribution plot for each plate.
    :param target_concentration_per_plate: Set to a dictionary of concentrations per plate if different concentrations are desired for different plates.
    :param plate_size: Specify the size of the plate to generate.
    :return: Distribution of volumes generated from the specified file.
    """

    # prepare output folders
    map_directory = os.path.join(output_directory, 'maps')
    script_directory = os.path.join(output_directory, 'scripts')
    create_dir_if_empty(output_directory, map_directory, script_directory)

    min_volume = np.inf
    max_volume = 0
    volume_dist = []

    spec_df = pd.read_excel(filename, sheet_name=None, engine='calamine')['Plate Specs']
    # get all unique names in the 'Plate Name' column
    plate_names = spec_df['Plate Name'].unique()

    # run through all available plates
    for plate_name in plate_names:
        # extract information
        selected_plate_df = spec_df[spec_df['Plate Name'] == plate_name]
        nmole_values = selected_plate_df['nmoles'].values
        plate_wells = selected_plate_df['Well Position'].values

        if target_concentration_per_plate and plate_name in target_concentration_per_plate:
            plate_target_conc = target_concentration_per_plate[plate_name]
        else:
            plate_target_conc = target_concentration_uM

        volume_required = (nmole_values / plate_target_conc) * 1000

        # cap all volumes to selected ceiling
        volume_required[volume_required > volume_cap_ul] = volume_cap_ul

        visual_dict = {}
        numeric_dict = {}

        max_volume = max(max_volume, np.max(volume_required))
        min_volume = min(min_volume, np.min(volume_required))

        volume_dist.extend(volume_required.tolist())

        for well_index, vol in enumerate(volume_required):
            well_name = plate_wells[well_index]
            if well_name[1] == '0':
                well_name = well_name[0] + well_name[2]
            visual_dict[well_name] = '#40E0D0'

            # how to use the floor function in python
            numeric_dict[well_name] = floor(vol * 10) / 10

        visualize_plate_with_color_labels(plate_size, visual_dict,
                                          direct_show=False,
                                          well_label_dict=numeric_dict,
                                          plate_title=f'{plate_name} (numbers = μl of UPW to add to achieve a target concentration of {plate_target_conc}μM)',
                                          save_file=f'{plate_name}_resuspension_map', save_folder=map_directory)

        # convert the well position and volumes to a pandas df and save to csv file
        output_volume_df = pd.DataFrame.from_dict(numeric_dict, orient='index', columns=['Volume'])
        output_volume_df = output_volume_df.reset_index()
        output_volume_df.columns = ['Destination', 'Volume']
        output_volume_df['Rack'] = 1
        output_volume_df['Source'] = 1
        output_volume_df['Rack 2'] = 1
        output_volume_df['Tool'] = 2
        output_volume_df = output_volume_df[['Rack', 'Source', 'Rack 2', 'Destination', 'Volume', 'Tool']]

        if max_commands_per_file and len(output_volume_df) > max_commands_per_file:
            # export in groups of the specified max command size
            for index, i in enumerate(range(0, len(output_volume_df), max_commands_per_file)):
                output_volume_df.iloc[i:i + max_commands_per_file].to_csv(
                    os.path.join(script_directory, f'{plate_name}_resuspension_volumes_{index + 1}.csv'),
                    header=['Rack', 'Source', 'Rack', 'Destination', 'Volume', 'Tool'], index=False)
        else:
            output_volume_df.to_csv(os.path.join(script_directory, f'{plate_name}_resuspension_volumes.csv'),
                                    header=['Rack', 'Source', 'Rack', 'Destination', 'Volume', 'Tool'], index=False)

        if plot_distribution_per_plate:
            visualize_plate_volume_distribution(os.path.join(map_directory, f'{plate_name}_volume_distribution.png'), volume_dist, plate_target_conc)

    return volume_dist

def visualize_plate_volume_distribution(file_output, volume_dist, target_concentration_uM):
    """
    Visualizes the distribution of resuspension volumes for a given list.
    :param file_output: Output filename
    :param volume_dist: List of volumes (in ul)
    :param target_concentration_uM: Target concentration used for this particular plate
    :return: N/A
    """

    fig, ax = plt.subplots(figsize=(12, 8))

    # Overlay bar plot (make sure alpha is higher than KDE plot)
    sns.histplot(volume_dist, ax=ax, bins=40, kde=True, line_kws={'linewidth': 6})

    plt.xlabel('Volume (μl)', fontsize=20)
    plt.ylabel('Frequency', fontsize=20)
    plt.title(f'Distribution of resuspension volumes (target concentration {target_concentration_uM}μM,\n max vol {max(volume_dist)}μl, min vol {min(volume_dist)}μl)', fontsize=22)
    plt.xticks(fontsize=18)
    plt.yticks(fontsize=18)
    plt.tight_layout()
    plt.savefig(file_output, dpi=300)
    plt.close(fig)
