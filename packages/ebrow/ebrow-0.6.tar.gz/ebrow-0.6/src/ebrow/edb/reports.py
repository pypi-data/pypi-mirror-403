"""
******************************************************************************

    Echoes Data Browser (Ebrow) is a data navigation and report generation
    tool for Echoes.
    Echoes is a RF spectrograph for SDR devices designed for meteor scatter
    Both copyright (C) 2018-2025
    Giuseppe Massimo Bertani gm_bertani(a)yahoo.it

    This program is free software: you can redistribute it and/or modify
    it under the terms of the GNU General Public License as published by
    the Free Software Foundation, version 3 of the License.

    This program is distributed in the hope that it will be useful,
    but WITHOUT ANY WARRANTY; without even the implied warranty of
    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
    GNU General Public License for more details.

    You should have received a copy of the GNU General Public License
    along with this program.  If not, http://www.gnu.org/copyleft/gpl.html

*******************************************************************************

"""
import os
import sys
import string
from datetime import datetime
from pathlib import Path
from shutil import copytree, copy2
from PyQt5.QtGui import QPixmap
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

from .logprint import print
from .rtsconfig import RTSconfig
from .utilities import toTypePoints, toChars


class Report:

    def __init__(self, parent, ui, settings):
        self.TAB_COVERAGE = 0
        self.TAB_SITEINFOS = 1
        self.TAB_DAILY = 2
        self.TAB_HOURLY = 3
        self.TAB_10MIN = 4
        self.TAB_RMOB = 5
        self.TAB_CHRONO = 6
        self.TAB_RTS = 7
        self.TAB_TOTAL = 8

        self.DIV_SIDEBAR = 0
        self.DIV_DAILY = 1
        self.DIV_HOURLY = 2
        self.DIV_10MIN = 3
        self.DIV_RMOB = 4
        self.DIV_CHRONO = 5
        self.DIV_RTS = 6
        self.DIV_INDEX = 7
        self.DIV_TOTAL = 8

        self._sections = list(range(self.TAB_SITEINFOS, self.TAB_TOTAL))
        self._ui = ui
        self._parent = parent
        self._settings = settings
        self._dataSource = None
        self._classFilter = None
        self._RMOBfile = None
        self._itemSelected = None
        self._defaultPreface = None

        self._dirPath = None
        self._assetsPath = None
        self._imgPath = None
        self._pageHeader = None
        self._pageFooter = None
        self._pageHeaderRmob = None
        self._totalDays = 0
        self._totUnfiltered = 0
        self._stationLogo = None
        self._events = list()
        self._statImgs = list()
        self._echoesVer = "0.51++"

        self._rdf = [None] * self.TAB_TOTAL  # report dataframes (-->XSLX)
        self._div = [None] * self.DIV_TOTAL  # report divs (-->HTML)

        # pushbuttons events

        # TBD: unuseful for now
        self._ui.pbRefresh_3.hide()
        # self._ui.pbRefresh_3.clicked.connect(self.updateTabReport)

        self._ui.pbReportXLSX.clicked.connect(self._reportXLSX)
        self._ui.pbReportHTML.clicked.connect(self._reportHTML)
        self._ui.pbRecallDefault.clicked.connect(self._recallDefaultPrefaceText)
        self._ui.pbRecallStored.clicked.connect(self._recallDefaultStoredText)
        self._ui.pbStoreNew.clicked.connect(self._savePrefaceText)

        # filter checkboxes toggling
        self._ui.chkOverdense_3.clicked.connect(self._setClassFilter)
        self._ui.chkUnderdense_3.clicked.connect(self._setClassFilter)
        self._ui.chkFakeRfi_3.clicked.connect(self._setClassFilter)
        self._ui.chkFakeEsd_3.clicked.connect(self._setClassFilter)
        self._ui.chkFakeCar1_3.clicked.connect(self._setClassFilter)
        self._ui.chkFakeCar2_3.clicked.connect(self._setClassFilter)
        self._ui.chkFakeSat_3.clicked.connect(self._setClassFilter)
        self._ui.chkFakeLong_3.clicked.connect(self._setClassFilter)
        self._ui.chkAll_3.clicked.connect(self._toggleCheckAll)

        # exclusions checkboxes toggling
        self._ui.chkSiteInfosExc.clicked.connect(self._toggleSection)
        self._ui.chkDailyExc.clicked.connect(self._toggleSection)
        self._ui.chkRMOBexc.clicked.connect(self._toggleSection)
        self._ui.chkHourlyExc.clicked.connect(self._toggleSection)
        self._ui.chk10minExc.clicked.connect(self._toggleSection)
        self._ui.chkChronoExc.clicked.connect(self._toggleSection)
        self._ui.chkExpEvExc.clicked.connect(self._toggleSection)
        self._ui.chkExpStTabExc.clicked.connect(self._toggleSection)
        self._ui.chkExpStGrpExc.clicked.connect(self._toggleSection)
        self._ui.chkSetupExc.clicked.connect(self._toggleSection)

        # autoexport controls events
        self._ui.sbAEminLast.valueChanged.connect(lambda val: self._settings.writeSetting('aeMinLast', val))
        self._ui.chkScreenshot.toggled.connect(lambda checked: self._settings.writeSetting('aeScreenshot', checked))
        self._ui.chkPowerPlot.toggled.connect(lambda checked: self._settings.writeSetting('aePowerPlot', checked))
        self._ui.chk2Dplot.toggled.connect(lambda checked: self._settings.writeSetting('ae2Dplot', checked))
        self._ui.chk3Dplot.toggled.connect(lambda checked: self._settings.writeSetting('ae3Dplot', checked))
        self._ui.chkDetails.toggled.connect(lambda checked: self._settings.writeSetting('aeDetails', checked))
        self._ui.chkNoComments.toggled.connect(lambda checked: self._settings.writeSetting('aeNoComments', checked))


    def _recallDefaultPrefaceText(self):
        preface = self._defaultPreface
        self._ui.pteReportPreface.setPlainText(preface)

    def _recallDefaultStoredText(self):
        preface = self._settings.readSettingAsString("preface")
        if preface != "":
            preface = preface.replace("<br>", "\n")
        self._ui.pteReportPreface.setPlainText(preface)

    def _savePrefaceText(self):
        prefaceDef = self._ui.pteReportPreface.toPlainText()
        self._settings.writeSetting("preface", prefaceDef)

    def updateTabReport(self):
        self._parent.busy(True)
        self._ui.chkSiteInfosExc.setChecked(self._settings.readSettingAsBool('siteInfosExc'))
        self._ui.chkDailyExc.setChecked(self._settings.readSettingAsBool('dailyExc'))
        self._ui.chkRMOBexc.setChecked(self._settings.readSettingAsBool('RMOBexc'))
        self._ui.chkHourlyExc.setChecked(self._settings.readSettingAsBool('hourlyExc'))
        self._ui.chk10minExc.setChecked(self._settings.readSettingAsBool('tenMinExc'))
        self._ui.chkChronoExc.setChecked(self._settings.readSettingAsBool('chronoExc'))
        self._ui.chkExpEvExc.setChecked(self._settings.readSettingAsBool('expEvExc'))
        self._ui.chkExpStTabExc.setChecked(self._settings.readSettingAsBool('stTabExc'))
        self._ui.chkExpStGrpExc.setChecked(self._settings.readSettingAsBool('stGrpExc'))
        self._ui.chkSetupExc.setChecked(self._settings.readSettingAsBool('setupExc'))

        self._ui.sbAEminLast.setValue(self._settings.readSettingAsInt('aeMinLast'))
        self._ui.chkScreenshot.setChecked(self._settings.readSettingAsBool('aeScreenshot'))
        self._ui.chkPowerPlot.setChecked(self._settings.readSettingAsBool('aePowerPlot'))
        self._ui.chk2Dplot.setChecked(self._settings.readSettingAsBool('ae2Dplot'))
        self._ui.chk3Dplot.setChecked(self._settings.readSettingAsBool('ae3Dplot'))
        self._ui.chkDetails.setChecked(self._settings.readSettingAsBool('aeDetails'))
        self._ui.chkNoComments.setChecked(self._settings.readSettingAsBool('aeNoComments'))

        self._dataSource = self._parent.dataSource
        self._parent.updateStatusBar("Gathering informations for report...")
        self._getClassFilter()

        # creates a default preface

        self._defaultPreface = f'''
        Radio meteor reception report produced by station [station], located in [city], [country]
        on days between [dateFrom] and [dateTo], generated by [ebrow].
        
        A total of [unclassified] events have been received.
        
        The tables that can be recalled through the side menu <i>Tables</i> include
        the chronological of the events with the related details and counts, 
        which are represented with different time resolutions or daily by classification.
        The counts reported regard only the events satisfying the class filters set when
        this report has been generated. These filters are reported above each table.
        
        In addition to the tables mentioned above - which are automatically generated by the program - this
        report includes a set of screens, graphs and tables that have been exported manually 
        by the operator. This set is divided into two galleries, depending if they
        concern individuals events or statistical counts.
        
        Regarding the latter, keep in mind that exports can have been done in any previous
        moment before the report generation, so they could have been generated with different filters applied.
        For this reason, their comments always include the related filter settings.
        
        '''
        storedPreface = self._settings.readSettingAsString("preface")

        if storedPreface != "":
            preface = storedPreface
        else:
            preface = self._defaultPreface

        self._ui.pteReportPreface.setPlainText(preface)
        if self._parent.isBatchReport and not self._parent.isReporting:
            # when self generated, ignores the personalized preface from ebrow.ini
            # and uses always a self generated one
            self._ui.pteReportPreface.setPlainText(self._defaultPreface)
            self._reportHTML()

        if self._parent.isBatchXLSX and not self._parent.isReporting:
            # when self generated, ignores the personalized preface from ebrow.ini
            # and uses always a self generated one
            self._ui.pteReportPreface.setPlainText(self._defaultPreface)
            self._reportXLSX()

        self._parent.busy(False)

    def _prepareTables(self):
        self._parent.busy(True)

        cfgRev = self._dataSource.getCfgRevisions()[-1]
        self._echoesVer = self._dataSource.getEchoesVersion(cfgRev)

        self._parent.updateProgressBar(0, 8)
        if self._rdf[self.TAB_COVERAGE] is None:
            self._buildCoverageTable()

        self._parent.updateProgressBar(1, 8)
        if self._rdf[self.TAB_SITEINFOS] is None:
            self._buildSiteInfosTable()

        self._parent.updateProgressBar(2, 8)
        if self._rdf[self.TAB_DAILY] is None:
            self._buildDailyCountsTable()

        self._parent.updateProgressBar(3, 8)
        if self._rdf[self.TAB_RMOB] is None:
            self._buildRMOBtable()

        self._parent.updateProgressBar(4, 8)
        if self._rdf[self.TAB_HOURLY] is None:
            self._buildHourlyCountsTable()

        self._parent.updateProgressBar(5, 8)
        if self._rdf[self.TAB_10MIN] is None:
            self._buildTenMinCountsTable()

        self._parent.updateProgressBar(6, 8)
        if self._rdf[self.TAB_CHRONO] is None:
            self._buildChronologicalTable()

        self._parent.updateProgressBar(7, 8)
        if self._rdf[self.TAB_RTS] is None:
            self._buildSetupTable()

        self._parent.updateProgressBar()

        self._parent.busy(False)

    def _toggleSection(self):
        # self._sections.append(self.TAB_COVERAGE) # always included
        for (key, chk) in [('siteInfosExc', self._ui.chkSiteInfosExc),
                           ('RMOBexc', self._ui.chkRMOBexc),
                           ('dailyExc', self._ui.chkDailyExc),
                           ('hourlyExc', self._ui.chkHourlyExc),
                           ('tenMinExc', self._ui.chk10minExc),
                           ('chronoExc', self._ui.chkChronoExc),
                           ('expEvExc', self._ui.chkExpEvExc),
                           ('stTabExc', self._ui.chkExpStTabExc),
                           ('stGrpExc', self._ui.chkExpStGrpExc),
                           ('setupExc', self._ui.chkSetupExc)]:
            if not chk.isChecked():
                self._settings.writeSetting(key, False)
            else:
                self._settings.writeSetting(key, True)

    def _getClassFilter(self):
        self._parent.busy(True)
        self._classFilter = self._settings.readSettingAsString('classFilterReport')
        idx = 0
        for tag in self._parent.filterTags:
            isCheckTrue = tag in self._classFilter
            if idx < len(self._parent.filterCheckReport):
                self._parent.filterCheckReport[idx].setChecked(isCheckTrue)
                idx += 1
        self._parent.updateStatusBar("Filtering report data by {}".format(self._classFilter))
        self._parent.busy(False)

    def _setClassFilter(self):
        self._parent.busy(True)
        classFilter = ''
        idx = 0
        for check in self._parent.filterCheckReport:
            if check.isChecked():
                classFilter += self._parent.filterTags[idx] + ','
            idx += 1

        if classFilter != '':
            classFilter = classFilter[0:-1]  # discards latest comma+space

        self._classFilter = classFilter
        self._settings.writeSetting('classFilterReport', self._classFilter)
        self._parent.updateStatusBar("Filtering report data by classification: {}".format(self._classFilter))
        self._parent.busy(False)

    def _toggleCheckAll(self):
        self._ui.chkOverdense_3.setChecked(self._ui.chkAll_3.isChecked())
        self._ui.chkUnderdense_3.setChecked(self._ui.chkAll_3.isChecked())
        self._ui.chkFakeRfi_3.setChecked(self._ui.chkAll_3.isChecked())
        self._ui.chkFakeEsd_3.setChecked(self._ui.chkAll_3.isChecked())
        self._ui.chkFakeCar1_3.setChecked(self._ui.chkAll_3.isChecked())
        self._ui.chkFakeCar2_3.setChecked(self._ui.chkAll_3.isChecked())
        self._ui.chkFakeSat_3.setChecked(self._ui.chkAll_3.isChecked())
        self._ui.chkFakeLong_3.setChecked(self._ui.chkAll_3.isChecked())
        self._setClassFilter()

    def _fillImportList(self):

        # scanning exported event-related files
        shotDir = Path(self._parent.exportDir, 'events')
        fileList = os.listdir(shotDir)

        # sort the filenames by date then by ID
        toSort = list()
        for fileName in fileList:
            fields = fileName.split('_')
            patchedFileName = "{}_{}${}".format(fields[-2], fields[-1], fileName)
            toSort.append(patchedFileName)
        toSort.sort()
        fileList = []
        for patchedFileName in toSort:
            fields = patchedFileName.split('$')
            fileList.append(fields[1])

        # create a list of event-related files (screenshots/graphs/comments)
        self._events = []
        detFile = None
        powFile = None
        img2dFile = None
        img3dFile = None
        commFile = None

        for fileName in fileList:
            img = Path(fileName)
            # looking for optional auxiliary files
            detFile = None
            powFile = None
            img2dFile = None
            img3dFile = None
            commDetailsFile = None
            commPowerFile = None
            commImg2dFile = None
            commImg3dFile = None

            if img.name.startswith('autoshot'):
                imgFile = fileName

                comments = 'comments_' + fileName
                comments = comments.replace('png', 'txt')
                commPath = shotDir / Path(comments)
                if commPath.exists():
                    commImgFile = commPath.name

                details = fileName.replace('autoshot', 'details')
                details = details.replace('png', 'csv')
                detPath = shotDir / Path(details)
                if detPath.exists():
                    detFile = detPath.name

                comments = 'comments_' + details
                comments = comments.replace('csv', 'txt')
                commPath = shotDir / Path(comments)
                if commPath.exists():
                    commDetailsFile = commPath.name

                power = fileName.replace('autoshot', 'power_profile')
                powPath = shotDir / Path(power)
                if powPath.exists():
                    powFile = powPath.name

                comments = 'comments_' + power
                comments = comments.replace('png', 'txt')
                commPath = shotDir / Path(comments)
                if commPath.exists():
                    commPowerFile = commPath.name

                img2d = fileName.replace('autoshot', 'image2d')
                img2dPath = shotDir / Path(img2d)
                if img2dPath.exists():
                    img2dFile = img2dPath.name

                comments = 'comments_' + img2d
                comments = comments.replace('png', 'txt')
                commPath = shotDir / Path(comments)
                if commPath.exists():
                    commImg2dFile = commPath.name

                img3d = fileName.replace('autoshot', 'image3d')
                img3dPath = shotDir / Path(img3d)
                if img3dPath.exists():
                    img3dFile = img3dPath.name

                comments = 'comments_' + img3d
                comments = comments.replace('png', 'txt')
                commPath = shotDir / Path(comments)
                if commPath.exists():
                    commImg3dFile = commPath.name

                htmlFile = imgFile.replace('.png', '.html')
                evTuple = (htmlFile, imgFile, commImgFile, detFile, commDetailsFile, powFile, commPowerFile, img2dFile, commImg2dFile, img3dFile, commImg3dFile)
                self._events.append(evTuple)

        # scanning exported statistic files (tables/graphs/comments)
        statDir = Path(self._parent.exportDir, 'statistics')
        fileList = os.listdir(statDir)
        fileList.sort()

        # create a list of event-related files
        self._statImgs = []

        for fileName in fileList:
            dataFile = Path(fileName)
            # statistic graph
            if dataFile.suffix == '.png' and fileName.startswith('stat-'):
                comments = fileName.replace('stat-', 'comments-')
                comments = comments.replace('png', 'txt')
                commPath = statDir / Path(comments)
                commFile = None
                if commPath.exists():
                    commFile = commPath.name

                htmlFile = fileName.replace('.png', '.html')
                self._statImgs.append((dataFile, htmlFile, commFile))

        self._statTabs = []
        for fileName in fileList:
            dataFile = Path(fileName)
            # statistic table
            if dataFile.suffix == '.csv' and fileName.startswith('stat-'):
                comments = fileName.replace('stat-', 'comments-')
                comments = comments.replace('csv', 'txt')
                commPath = statDir / Path(comments)
                commFile = None
                if commPath.exists():
                    commFile = commPath.name

                htmlFile = fileName.replace('.csv', '.html')
                self._statTabs.append((dataFile, htmlFile, commFile))

    def _buildCoverageTable(self):
        self._parent.updateStatusBar("Gathering time coverage data...")

        # first row contains current UTC date and time
        utc = datetime.utcnow()
        (fromStr, toStr) = self._settings.coverage()

        # note: this is not a real uptime
        df = self._dataSource.getASpartialFrame(self._parent.fromDate, self._parent.toDate)
        deltaMins = df['delta_min'].sum()
        coveredIDs = self._parent.covID

        self._stationLogo = self._settings.readSettingAsString('logoPath')
        if self._stationLogo == ':/logo':
            self._stationLogo = 'logo.png'

        # main header
        rdfCols = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
        rdfDict = dict()
        rdfDict['1'] = ['', 'Meteor scatter detection report', '']

        # section 1 report coverage, cannot be excluded
        # splittenPreface = toMultiline(self._ui.pteReportPreface.toPlainText());
        # print("splittenPreface={}".format(splittenPreface))
        # rdfDict['2'] = ['Preface', splittenPreface]
        
        self._savePrefaceText()
        prefaceDef = self._ui.pteReportPreface.toPlainText()
        preface = self._formatPreface(prefaceDef)

        rdfDict['2'] = ['Preface:', preface]
        rdfDict['3'] = ['Created by:', 'Echoes Data Browser', 'version:', self._parent.version, ' on ',
                        utc.strftime("%Y %m %d"), utc.strftime("%H:%M:%S")]
        rdfDict['4'] = ['Covered period from: ', fromStr, ' to: ', toStr, 'total events:', coveredIDs]
        rdfDict['5'] = ['Complessive uptime: ', deltaMins, ' minutes']
        rdfDict['6'] = ['Daily summary plot: ']
        rdf = pd.DataFrame.from_dict(rdfDict, orient='index', columns=rdfCols)
        self._rdf[self.TAB_COVERAGE] = rdf

    def _buildSiteInfosTable(self):
        self._parent.updateStatusBar("Gathering site informations...")
        if not self._ui.chkSiteInfosExc.isChecked():
            rdfCols = ['A', 'B', 'C', 'D', 'E', 'F']
            rdfDict = dict()
            for rowNr in range(1, len(rdfCols)):
                rdfDict[str(rowNr)] = len(rdfCols) * [' ']
            # main header

            # section 2 site infos
            rdfDict['1'] = ['Site infos']
            rdfDict['2'] = ['Station name: ', self._settings.readSettingAsString('stationName'),
                            self._stationLogo]
            rdfDict['3'] = ['Owner data: ', self._settings.readSettingAsString('owner')]
            rdfDict['4'] = ['Country: ', self._settings.readSettingAsString('country'), 'City: ',
                            self._settings.readSettingAsString('city')]
            rdfDict['5'] = ['Longitude: ', self._settings.readSettingAsString('longitude'), 'Latitude: ',
                            self._settings.readSettingAsString('latitude')]
            rdfDict['6'] = ['Altitude: ', self._settings.readSettingAsString('altitude')]
            rdfDict['7'] = ['Antenna type: ', self._settings.readSettingAsString('antenna'),
                            'Azimuth: ', str(self._settings.readSettingAsString('antAzimuth')) + '&deg',
                            'Elevation: ', str(self._settings.readSettingAsString('antElevation')) + '&deg']

            rdfDict['8'] = ['Preamplifier: ', self._settings.readSettingAsString('preamplifier')]
            rdfDict['9'] = ['Receiver: ', self._settings.readSettingAsString('receiver')]
            rdfDict['10'] = ['Frequencies: ', self._settings.readSettingAsString('frequencies')]
            rdfDict['11'] = ['Computer type: ', self._settings.readSettingAsString('computer')]
            rdfDict['12'] = ['Software: ', 'Echoes v.'+str(self._echoesVer) + ' + Ebrow v.'+str(self._parent.version)]
            rdfDict['13'] = ['e-mail: ', self._settings.readSettingAsString('email')]
            rdfDict['14'] = ['Notes: ', self._settings.readSettingAsString('notes')]

            rdf = pd.DataFrame.from_dict(rdfDict, orient='index', columns=rdfCols)
        else:
            rdf = None
        self._rdf[self.TAB_SITEINFOS] = rdf

    def _buildRMOBtable(self):

        if not self._ui.chkRMOBexc.isChecked():
            self._parent.updateStatusBar("Gathering RMOB countings...")
            # section 3 hourly counts by month (RMOB data)
            dfMonth, heatmapFileName, bargraphFileName = self._parent.getRMOBdata(sendOk=False)

            # in order to merge the daily counts table with the page header, their column names
            # must become a data row and column names become standardized numbers

            colNames = dfMonth.columns.values.tolist()  # gets the column names
            dfMonth.columns = range(dfMonth.columns.size)  # reset column names

            # creates a dataframe containing the column names as single row
            dfCols = pd.DataFrame([colNames], columns=list(range(0, len(colNames))))

            # the dailyDf index contains the dates, they must become a data column
            dates = dfMonth.index.values.tolist()
            dfMonth.insert(0, 'Day', dates)
            dfMonth.columns = range(dfMonth.columns.size)  # reset column names

            # here creates the page header dataframe
            columns = len(dfMonth.columns)

            # section 4 hourly counts by month, RMOB format
            # now joins the three dataframes
            fullDf = pd.concat([dfCols, dfMonth], ignore_index=True)
            fullDf.columns = list(string.ascii_uppercase + string.ascii_lowercase)[
                             :fullDf.columns.size]  # replace column names with letter
            fullDf.loc[0, 'B':] = fullDf.loc[0, 'A':].shift(periods=1, fill_value='')
            fullDf.loc[0, 'A'] = 'DAY'
            title = pd.DataFrame([['Hourly counts by month, RMOB format', '', '', '', '', '', '', '', '', '', '', '',
                                   '', '', '', '', '', '', '', '', '', '', '', '', '']], columns=fullDf.columns)
            fullDf = pd.concat([title, fullDf], ignore_index=True)

            # RMOBimgName = self._settings.readSettingAsString("RMOBfilePrefix") + ".jpg"
            # RMOBfullPath = Path(self._parent.exportDir, "statistics", RMOBimgName)

            # fullDf.loc[2, 'B'] = str(RMOBfullPath.absolute())

            # nex index must start from 1, not zero
            fullDf.index = fullDf.index + 1
        else:
            fullDf = None

        self._rdf[self.TAB_RMOB] = fullDf

    def _buildDailyCountsTable(self):
        if not self._ui.chkDailyExc.isChecked():
            self._parent.updateStatusBar("Gathering daily counts...")

            df = self._dataSource.getADpartialFrame(self._parent.fromDate, self._parent.toDate)

            dailyDf, rawDf, bgDf = self._dataSource.dailyCountsByClassification(df, self._classFilter,
                                                                   self._parent.fromDate, self._parent.toDate,
                                                                   totalRow=True, totalColumn=True)

            # in order to merge the daily counts table with the page header, their column names
            # must become a data row and column names become standardized numbers

            colNames = ['Day'] + dailyDf.columns.values.tolist()  # gets the column names
            dailyDf.columns = range(dailyDf.columns.size)  # reset column names

            # creates a dataframe containing the column names as single row
            dfCols = pd.DataFrame([colNames], columns=list(range(0, len(colNames))))

            # the dailyDf index contains the dates, they must become a data column
            dates = dailyDf.index.values.tolist()
            dailyDf.insert(0, 'Day', dates)
            dailyDf.reset_index(inplace=True, drop=True)
            dailyDf.columns = range(dailyDf.columns.size)  # reset column names

            # here creates the page header dataframe
            columns = len(dailyDf.columns)
            rdfCols = list(range(0, columns))
            rdfDict = dict()
            rdfDict['1'] = len(rdfCols) * [' ']

            # section 3 daily counts by classifications

            rdf = pd.DataFrame.from_dict(rdfDict, orient='index', columns=rdfCols)

            # now joins the three dataframes
            fullDf = pd.concat([rdf, dfCols, dailyDf], ignore_index=True)

            fullDf.columns = list(string.ascii_uppercase +
                                  string.ascii_lowercase)[:fullDf.columns.size]  # replace column names with letters

            # the column name 'DAY' is prepended for correct columns/headers alignment
            fullDf.loc[0, 'A'] = 'Daily counts by classification'
            # nex index must start from 1, not zero
            fullDf.index = fullDf.index + 1
        else:
            fullDf = None

        self._rdf[self.TAB_DAILY] = fullDf

    def _buildHourlyCountsTable(self):
        if not self._ui.chkHourlyExc.isChecked():
            self._parent.updateStatusBar("Gathering hourly counts...")
            df = self._dataSource.getADpartialFrame(self._parent.fromDate, self._parent.toDate)

            hourlyDf, rawDf, sbDf = self._dataSource.makeCountsDf(df, self._parent.fromDate, self._parent.toDate, dtRes='h',
                                                     filters=self._classFilter, totalRow=True, totalColumn=True)

            # in order to merge the daily counts table with the page header, their column names
            # must become a data row and column names become standardized numbers
            colNames = hourlyDf.columns.values.tolist()  # gets the column names
            hourlyDf.columns = range(hourlyDf.columns.size)  # reset column names

            # creates a dataframe containing the column names as single row
            dfCols = pd.DataFrame([colNames], columns=list(range(0, len(colNames))))

            # the dailyDf index contains the dates, they must become a data column
            dates = hourlyDf.index.values.tolist()
            hourlyDf.insert(0, 'Day', dates)
            hourlyDf.columns = range(hourlyDf.columns.size)  # reset column names

            # here creates the page header dataframe
            columns = len(hourlyDf.columns)
            rdfCols = list(range(0, columns))
            rdfDict = dict()
            rdfDict['1'] = len(rdfCols) * [' ']

            # section 3 Hourly counts
            rdf = pd.DataFrame.from_dict(rdfDict, orient='index', columns=rdfCols)

            # now joins the three dataframes
            fullDf = pd.concat([rdf, dfCols, hourlyDf], ignore_index=True)

            fullDf.columns = list(string.ascii_uppercase)[:fullDf.columns.size]  # replace column names with letters
            # the column name 'DAY' is prepended for correct columns/headers alignment
            fullDf.loc[0, 'A'] = 'Hourly counts'
            fullDf.loc[1, 'B':] = fullDf.loc[1, 'A':].shift(periods=1, fill_value=' ')
            fullDf.loc[1, 'A'] = 'utc_date'
            # nex index must start from 1, not zero
            fullDf.index = fullDf.index + 1
        else:
            fullDf = None
        self._rdf[self.TAB_HOURLY] = fullDf

    def _buildTenMinCountsTable(self):
        if not self._ui.chk10minExc.isChecked():
            self._parent.updateStatusBar("Gathering 10min counts...")
            df = self._dataSource.getADpartialFrame(self._parent.fromDate, self._parent.toDate)

            tenDf, rawDf, sbDf = self._dataSource.makeCountsDf(df, self._parent.fromDate, self._parent.toDate, dtRes='10T',
                                                  filters=self._classFilter)
            tenDf = self._dataSource.splitAndStackDataframe(tenDf, maxColumns=24)

            # in order to merge the daily counts table with the page header, their column names
            # must become a data row and column names become standardized numbers
            colNames = tenDf.columns.values.tolist()  # gets the column names
            tenDf.columns = range(tenDf.columns.size)  # reset column names

            # creates a dataframe containing the column names as single row
            # dfCols = pd.DataFrame([colNames], columns=list(range(0, len(colNames))))

            # the dailyDf index contains the dates, they must become a data column
            # dates = tenDf.index.values.tolist()
            # tenDf.insert(0, 'Day', dates)
            # tenDf.columns = range(tenDf.columns.size)  # reset column names

            # here creates the page header dataframe
            columns = len(tenDf.columns)
            rdfCols = list(range(0, columns))
            rdfDict = dict()
            rdfDict['1'] = len(rdfCols) * [' ']

            # section 3 Counts by 10 min intervals
            rdf = pd.DataFrame.from_dict(rdfDict, orient='index', columns=rdfCols)

            # now joins the three dataframes

            fullDf = pd.concat([rdf, tenDf], ignore_index=True)

            fullDf.columns = list(string.ascii_uppercase + string.ascii_lowercase)[
                             :fullDf.columns.size]  # replace column names with letters
            # the column name 'DAY' is prepended for correct columns/headers alignment
            # fullDf.loc[3, 'A':] = fullDf.loc[3, 'A':].shift(periods=1, fill_value='DAY')
            # nex index must start from 1, not zero
            # fullDf.index = fullDf.index + 1
            fullDf.reset_index(drop=True)
            fullDf.index = fullDf.index + 1
            rows = len(fullDf.index) + 1
            days = int(rows / 6)

            fullDf.iloc[0, 0] = 'Counts by 10 min intervals'

            # fullDf.drop(columns='INDEX', inplace=True)
        else:
            fullDf = None
        self._rdf[self.TAB_10MIN] = fullDf

    def _buildChronologicalTable(self):

        if not self._ui.chkChronoExc.isChecked():
            self._parent.updateStatusBar("Getting chronological events table...")
            df = self._parent.dataSource.getADpartialCompositeFrame(self._parent.fromDate, self._parent.toDate,
                                                                    filters=self._classFilter)
            self._rdf[self.TAB_CHRONO] = df

    def _buildSetupTable(self):
        if not self._ui.chkSetupExc.isChecked():
            self._parent.updateStatusBar("Getting Echoes RTS setup table...")
            revList = self._dataSource.getCfgRevisions()
            # considers only the latest revision
            rev = revList[-1]
            rts = RTSconfig(self._parent, self._ui, self._settings)
            self._rdf[self.TAB_RTS] = rts.getAllTables(rev)

       
    def _reportXLSX(self):
        self._rdf = [None] * self.TAB_TOTAL  # report dataframes (-->XSLX)
        self._div = [None] * self.DIV_TOTAL  # report divs (-->HTML)
        self._prepareTables()
        self._parent.isReporting = True
        self._parent.busy(True)
        self._savePrefaceText()
        # preface = self._formatPreface(prefaceDef)
        # create a excel writer object
        defaultFilename = "report.xlsx"
        self._parent.updateStatusBar("Saving XLSX report file")
        with pd.ExcelWriter(defaultFilename) as writer:
            # use to_excel function and specify the sheet_name and index
            # to store the dataframe in specified sheet
            # fullRdf.to_excel(writer, sheet_name="Report", index=False)
            self._parent.updateStatusBar("Adding Time Coverage sheet to XLSX...")
            self._rdf[self.TAB_COVERAGE].to_excel(writer, sheet_name="Time coverage", index=False, header=None)

            if not self._ui.chkSiteInfosExc.isChecked():
                self._parent.updateStatusBar("Adding Site Informations sheet to XLSX...")
                self._rdf[self.TAB_SITEINFOS].to_excel(writer, sheet_name="Site information", index=False, header=None)

            if not self._ui.chkDailyExc.isChecked():
                self._parent.updateStatusBar("Adding Daily Counts sheet to XLSX...")
                self._rdf[self.TAB_DAILY].to_excel(writer, sheet_name="Daily counts", index=False, header=None)

            if not self._ui.chkHourlyExc.isChecked():
                self._parent.updateStatusBar("Adding Hourly Counts sheet to XLSX...")
                self._rdf[self.TAB_HOURLY].to_excel(writer, sheet_name="Hourly counts", index=False, header=None)

            if not self._ui.chk10minExc.isChecked():
                self._parent.updateStatusBar("Adding 10min Counts sheet to XLSX...")
                self._rdf[self.TAB_10MIN].to_excel(writer, sheet_name="10min counts", index=False, header=None)

            if not self._ui.chkRMOBexc.isChecked():
                self._parent.updateStatusBar("Adding Current RMOB sheet to XLSX...")
                self._rdf[self.TAB_RMOB].to_excel(writer, sheet_name="current month's RMOB", index=False)

            if not self._ui.chkChronoExc.isChecked():
                self._parent.updateStatusBar("Adding Chronological table sheet to XLSX...")
                self._rdf[self.TAB_CHRONO].to_excel(writer, sheet_name="chronological events table", index=False)

            if not self._ui.chkSetupExc.isChecked():
                self._parent.updateStatusBar("Adding RTS setup tables sheets to XLSX...")
                sheets = ['RTS device configuration',
                          'RTS FFT configuration',
                          'RTS Output configuration',
                          'RTS Storage configuration',
                          'RTS Preferences'
                          ]
                idx = 0
                for table in self._rdf[self.TAB_RTS]:
                    sheet = sheets[idx]
                    table.to_excel(writer, sheet_name=sheet, index=False)
                    idx += 1
        # reopening the file with openpyxl to add formats and images
        wb = load_workbook(defaultFilename)
               
        # Set font and alignment on all the filled cells
        family=self._settings.readSettingAsString('fontFamily')
        dataFont = Font(name=family, size=14, bold=False)
        dataAlignment = Alignment(horizontal="center", vertical="center")
        imagesAlignment = Alignment(horizontal="center", vertical="bottom")
        headingFont = Font(name=family, size=16, bold=True)
        rowHeadAlignment = Alignment(horizontal="right", vertical="center")
        columnHeadAlignment = dataAlignment
        titleFont = Font(name=family, size=20, bold=True)
        titleAlignment = Alignment(horizontal="center", vertical="center")
        
         
        
        #------------------------------------------------------------------------------------
        ws = wb['Time coverage']
        
        # ebrow logo
        assetsPath = Path(self._parent.appPath) / "assets" / "img"
        img = Image(assetsPath / "logo.png")
        img.width = 128
        img.height = 128
        ws.add_image(img, 'A1')
        
        # station logo
        img = Image(self._stationLogo)
        img.width = 128
        img.height = 128
        ws.add_image(img, 'C1')
        
        # daily summary plot
        summary = self._parent.getSummaryPlot(self._classFilter)
        img = Image(summary)
        ws.add_image(img, 'B6')
        ws['B6'].alignment = imagesAlignment

        for col in ws.iter_cols(min_col=1, max_col=ws.max_column, min_row=1, max_row=ws.max_row):
            for cell in col:
                if cell.column_letter in ['A', 'C', 'E']:
                    defaultFont = headingFont
                    defaultAlignment = rowHeadAlignment
                else:
                    defaultFont = dataFont
                    defaultAlignment = dataAlignment
                cell.font = defaultFont
                cell.alignment = defaultAlignment

        # title cell     
        ws['B1'].font = titleFont
        ws['B1'].alignment = titleAlignment

        # formatting preface
        ws['B2'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)  
        ws['B2'].font = Font(italic=True)

        ws.row_dimensions[1].height = toTypePoints(128)   
        ws.row_dimensions[6].height = toTypePoints(128)

        for letter in range(ord('A'), ord('G') + 1):
            colName = chr(letter)
            ws.column_dimensions[colName].width = toChars(256)
            
        ws.column_dimensions['B'].width = toChars(640)           # wider due to summary plot

        #------------------------------------------------------------------------------------
        ws = wb['Site information']
        subtitleFont = Font(name="Arial", size=18, bold=True)
        subtitleAlignment = Alignment(horizontal="center", vertical="center")

        for col in ws.iter_cols(min_col=1, max_col=ws.max_column, min_row=1, max_row=ws.max_row):
            for cell in col:
                if cell.column_letter in ['A', 'C', 'E']:
                    defaultFont = headingFont
                    defaultAlignment = rowHeadAlignment
                else:
                    defaultFont = dataFont
                    defaultAlignment = dataAlignment
                cell.font = defaultFont
                cell.alignment = defaultAlignment

        # sheet title cell
        ws.merge_cells("A1:D1")
        ws['A1'].font = subtitleFont
        ws['A1'].alignment = subtitleAlignment

        # station logo
        img = Image(self._stationLogo)
        img.width = 128
        img.height = 128
        ws['C2'].value = None
        ws.add_image(img, 'C2')

        val = ws['D7'].value
        val = val.replace('&deg', '°')
        ws['D7'].value = val

        val = ws['F7'].value
        val = val.replace('&deg', '°')
        ws['F7'].value = val

        for letter in range(ord('A'), ord('G') + 1):
            colName = chr(letter)
            ws.column_dimensions[colName].width = toChars(300)

        ws.row_dimensions[1].height = toTypePoints(128)
        ws.row_dimensions[2].height = toTypePoints(128)
        #------------------------------------------------------------------------------------
        ws = wb['Daily counts']
        for letter in range(ord('A'), ord('G') + 1):
            colName = chr(letter)
            ws.column_dimensions[colName].width = toChars(128)
            
            
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.row == 2:
                    cell.font = headingFont
                    cell.alignment = columnHeadAlignment    
                else:
                    cell.font = dataFont
                    cell.alignment = dataAlignment
                    
        # sheet title cell
        ws.merge_cells("A1:D1")  
        ws['A1'].font = subtitleFont
        ws['A1'].alignment = subtitleAlignment
        
        ws.row_dimensions[1].height = toTypePoints(128)

        #------------------------------------------------------------------------------------
        ws = wb['Hourly counts']
        for letter in range(ord('A'), ord('Z') + 1):
            colName = chr(letter)
            ws.column_dimensions[colName].width = toChars(128)
            
            
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.row == 2:
                    cell.font = headingFont
                    cell.alignment = columnHeadAlignment    
                else:
                    cell.font = dataFont
                    cell.alignment = dataAlignment
        # sheet title cell
        ws.merge_cells("A1:Z1")  
        ws['A1'].font = subtitleFont
        ws['A1'].alignment = subtitleAlignment    
        ws.row_dimensions[1].height = toTypePoints(128)


        #------------------------------------------------------------------------------------
        ws = wb["current month's RMOB"]
        
        ws.delete_rows(1) # remove the first row with column letters
        
        for letter in range(ord('A'), ord('Y') + 1):
            colName = chr(letter)
            ws.column_dimensions[colName].width = toChars(128)
             
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.row == 2:
                    cell.font = headingFont
                    cell.alignment = columnHeadAlignment    
                else:
                    cell.font = dataFont
                    cell.alignment = dataAlignment
        # sheet title cell
        ws.merge_cells("A1:Z1")  
        ws['A1'].font = subtitleFont
        ws['A1'].alignment = subtitleAlignment    
        ws.row_dimensions[1].height = toTypePoints(128)
        
        # RMOB summary image
        self._RMOBfile = self._settings.readSettingAsString('RMOBfilePrefix') + ".jpg"
        RMOBimgFilePath = Path(self._parent.workingDir) / self._RMOBfile
        img = Image(RMOBimgFilePath)
        img.width = 700
        img.height = 220
        ws.add_image(img, 'B36')
        ws.row_dimensions[36].height = toTypePoints(220) 
        ws['B36'].alignment = imagesAlignment
        
        #------------------------------------------------------------------------------------
        ws = wb['chronological events table']
        ws.insert_rows(1)
        
        for colNum in range(1, 31):  # Da 1 a 31 per coprire da A ad AD AF
            colName = get_column_letter(colNum)
            ws.column_dimensions[colName].width = toChars(128)
            
        # columns with file names are wider
        ws.column_dimensions['AE'].width = toChars(300)
        ws.column_dimensions['AF'].width = toChars(300)
            
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.row == 2:
                    cell.font = headingFont
                    cell.alignment = columnHeadAlignment    
                else:
                    cell.font = dataFont
                    cell.alignment = dataAlignment
        # sheet title cell
        ws.merge_cells("A1:D1")  
        ws['A1'].font = subtitleFont
        ws['A1'].alignment = subtitleAlignment    
        ws['A1'].value = ws.title
        ws.row_dimensions[1].height = toTypePoints(128)
                
        #------------------------------------------------------------------------------------
        # RTS configuration tables
        for ws in wb.worksheets:
            if 'RTS' in ws.title:
                ws.insert_rows(1)
                
                for letter in range(ord('A'), ord('P') + 1):
                    colName = chr(letter)
                    ws.column_dimensions[colName].width = toChars(256)
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    for cell in row:
                        if cell.row == 2:
                            cell.font = headingFont
                            cell.alignment = columnHeadAlignment    
                        else:
                            cell.font = dataFont
                            cell.alignment = dataAlignment
                            
                ws.merge_cells("A1:D1")  
                ws['A1'].font = subtitleFont
                ws['A1'].alignment = subtitleAlignment    
                ws['A1'].value = ws.title
           
           
        #------------------------------------------------------------------------------------ 
        # final makeup
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    cell.border = Border()    # Remove all borders
           
            # define the print area to cover all non-empty cells
            min_cell, max_cell = sheet.calculate_dimension().split(':')
            area = f"{min_cell}:{max_cell}"  
            sheet.print_area = area
           
            sheet.page_setup.fitToWidth = 1
            sheet.page_setup.fitToHeight = 1
            # sheet.page_setup.paperSize = PageSetup.PAPERSIZE_A4  # Imposta il formato carta su A4 (optional)
            sheet.page_setup.orientation = "landscape"
           
        wb.save(defaultFilename)
        self._parent.isReporting = False
        self._parent.busy(False)

    def _reportHTML(self):
        self._parent.busy(True)
        self._rdf = [None] * self.TAB_TOTAL  # report dataframes (-->XSLX)
        self._div = [None] * self.DIV_TOTAL  # report divs (-->HTML)
        self._parent.isReporting = True

        self._savePrefaceText()
        prefaceDef = self._ui.pteReportPreface.toPlainText()
        preface = self._formatPreface(prefaceDef)

        if self._settings.readSettingAsBool('aeScreenshot') or self._settings.readSettingAsBool('aePowerPlot') or self._settings.readSettingAsBool('ae2Dplot') or self._settings.readSettingAsBool('ae3Dplot') or self._settings.readSettingAsBool('aeDetails'):
            self._parent.tabScreenshots.autoExport(self._classFilter)
        self._parent.updateProgressBar(0)
        self._prepareTables()
        self._parent.updateStatusBar("Saving HTML report files")
        now = datetime.utcnow()
        dirName = now.strftime("report_%Y-%m-%d_%H-%M-%S")
        self._dirPath = Path(".") / Path(dirName)
        self._assetsPath = self._dirPath / "assets"

        # self._cssPath = assetsPath / "css"
        # self._jsPath = assetsPath / "js"
        # cssPath.mkdir(mode=0o644, parents=True, exist_ok=True)
        # imgPath.mkdir(mode=0o644, parents=True, exist_ok=True)
        # jsPath.mkdir(mode=0o644, parents=True, exist_ok=True)

        # copies the html assets directory tree with its files
        self._imgPath = self._assetsPath / "img"
        srcAssetsPath = Path(self._parent.appPath) / "assets" / "img"
        print("Copying from {} to {}".format(srcAssetsPath, self._assetsPath))
        copytree(srcAssetsPath, self._imgPath)
        self._cssPath = self._assetsPath / "css"
        srcAssetsPath = Path(self._parent.appPath) / "assets" / "css"
        print("Copying from {} to {}".format(srcAssetsPath, self._assetsPath))
        copytree(srcAssetsPath, self._cssPath)

        # the site logo is copied into report's assets img
        if self._stationLogo != 'logo.png' and self._stationLogo is not None and os.path.exists(self._stationLogo):
            copy2(self._stationLogo, self._imgPath)

        # if the logo is the default one, there is no need to copy
        # since its file is already present under assets/

        df = self._rdf[self.TAB_DAILY]
        if df is None:
            self._parent.updateStatusBar("Cannot create HTML reports, all sections are excluded")
            self._parent.isReporting = False
            self._parent.busy(False)
            return

        self._totalDays = len(df.index) - 3
        if self._parent.covID < 2 or self._totalDays < 2:
            message = "Cannot generate a report covering less than 2 days"
            self._parent.updateStatusBar(message)
            if not self._parent.isBatchReport and not self._parent.isBatchXLSX:
                self._parent.infoMessage("Report generation", message)
            self._parent.isReporting = False
            sys.exit()

        df = self._rdf[self.TAB_COVERAGE]
        self._totUnfiltered = df.loc['4', 'F']

        rmobCSSpatch = '<link rel="stylesheet" href="assets/css/rmob.css">'
        headerModel = '''<!DOCTYPE html>
<html lang="en">
    <head>
        <title>Echoes Report</title>
        <!-- <link rel="stylesheet" href="assets/css/reset.css"> -->
        <link rel="stylesheet" href="assets/css/style.css">
        <link rel="stylesheet" href="assets/css/table.css">
        {0}
        <style>
            @import url('https://fonts.googleapis.com/css2?family=Titillium+Web&display=swap');
        </style>
    </head>
    <body onload="onLoadFunc()">
        <div id="loader"></div>
        <div id="container">
            <div class="info clearfix">
                <img src="assets/img/logo.png" alt="Echoes">
                <h1>Meteor scatter detection report</h1>
                <p>Created with <i>Echoes Data Browser</i> release {1}</p>
            </div> <!-- info -->
            <div id="content" class="animate-bottom">
'''

        self._pageHeader = headerModel.format('', self._parent.version)
        self._pageHeaderRmob = headerModel.format(rmobCSSpatch, self._parent.version)

        self._pageFooter = '''\t\t\t<div id="footer">
                <div class="box">
                    <strong>Echoes</strong> Reporting System - <a href="https://www.gabb.it/echoes" title="Echoes Web Site">Web Site</a> -
                     <a href="https://sourceforge.net/projects/echoes/" title="Find Echoes on Sourceforge">Find Echoes on Sourceforge</a> - 
                     <a href="https://www.facebook.com/gmbertani/" title="Find Echoes on Facebook">Find Echoes on Facebook</a>
                </div><!--box-->
            </div> <!-- footer -->
        </div> <!-- container -->
        <script>
            var tout;
            
            function onLoadFunc() {
                tout = setTimeout(showPage, 3000);
            }
            
            function showPage() {
                document.getElementById("loader").style.display = "none";
                document.getElementById("content").style.display = "flex";
            }
        </script>
    </body>
</html>'''

        self._fillImportList()
        self._parent.updateStatusBar("Embedding exported event files")
        for evTup in self._events:
            self._buildEventHtml(evTup)

        self._parent.updateStatusBar("Embedding exported statistic graphs")
        for stTup in self._statImgs:
            self._buildStatGraphsHtml(stTup)
        self._parent.updateStatusBar("Embedding exported statistic tables")
        for tabTup in self._statTabs:
            self._buildStatTablesHtml(tabTup)

        self._buildIndexHtml('index.html')
        self._buildDailyHtml('daily.html')
        self._buildHourlyHtml('hourly.html')
        self._build10minHtml('10min.html')
        self._buildRMOBhtml('rmob.html')
        self._buildChronologicalHtml('chrono.html')
        self._buildRTShtml('setup.html')
        self._parent.isReporting = False
        self._parent.updateStatusBar("Report ready under {}".format(self._dirPath))
        self._parent.busy(False)

    def _buildIndexHtml(self, outputFile):
        self._parent.updateStatusBar("Creating index.html")
        summary = self._parent.getSummaryPlot(self._classFilter)
        srcImgFilePath = Path(self._parent.workingDir) / summary
        copy2(srcImgFilePath, self._imgPath)
        relImgPath = self._imgPath.relative_to(self._dirPath)
        summaryPath = relImgPath / summary
        summaryPath = summaryPath.as_posix()

        prefaceDef = self._ui.pteReportPreface.toPlainText()
        self._settings.writeSetting("preface", prefaceDef)
        preface = self._formatPreface(prefaceDef)
        preface = preface.replace('\n', '<br>')

        self._div[self.DIV_INDEX] = '''\t\t\t<div class="box">
                <h2>Preface</h2>
                <p>{0}</p>
                <hr>
                <h2>Daily Summary plot</h2>
'''.format(preface)
        self._div[self.DIV_INDEX] += self._activeFiltersHtml()
        self._div[self.DIV_INDEX] += '''<img src = "{0}" title = "Daily summary plot">
            </div><!--box-->
        </div><!--content-->
'''.format(summaryPath)

        self._buildSidebarHtml(outputFile)

        htmlDoc = self._pageHeader
        htmlDoc += self._div[self.DIV_INDEX]
        htmlDoc += self._div[self.DIV_SIDEBAR]
        htmlDoc += self._pageFooter

        outFilePath = self._dirPath / outputFile
        with open(outFilePath, 'w') as f:
            f.write(htmlDoc)

    def _buildChronologicalHtml(self, outputFile):
        """
        Being a very long and wide table, it must be stored in a frame
        to allow browsing via scrollbars
        """
        if self._ui.chkChronoExc.isChecked():
            return

        self._parent.updateStatusBar("Creating chrono.html")
        iframeHeader = '''<!DOCTYPE html>
<html lang="en">
    <head>
        <title>Chronological event\'s table</title>
        <!-- <link rel="stylesheet" href="assets/css/reset.css"> -->
        <link rel="stylesheet" href="assets/css/style.css">
        <link rel="stylesheet" href="assets/css/table.css">
        <style>
            @import url('https://fonts.googleapis.com/css2?family=Titillium+Web&display=swap');
        </style>
    </head>
    <body>
        <div id="container">
            <div class="info clearfix">
                <h2>Chronological event\'s table</h2>
            </div> <!-- info -->
            <div id="frametable">
'''

        df = self._rdf[self.TAB_CHRONO]
        frameHtml = iframeHeader
        frameHtml += df.to_html(index=False)
        frameHtml += self._pageFooter

        tableIFrame = "table-" + outputFile
        outFilePath = self._dirPath / tableIFrame
        with open(outFilePath, 'w') as f:
            f.write(frameHtml)

        self._buildSidebarHtml(outputFile)

        self._div[self.DIV_CHRONO] = '''\t\t\t\t<iframe src="{}" title="chrono table frame" height="2000"></iframe>
            </div><!--content-->\n
'''.format(tableIFrame)

        htmlDoc = self._pageHeader
        htmlDoc += self._div[self.DIV_CHRONO]
        htmlDoc += self._div[self.DIV_SIDEBAR]
        htmlDoc += self._pageFooter

        outFilePath = self._dirPath / outputFile
        with open(outFilePath, 'w') as f:
            f.write(htmlDoc)

    def _buildDailyHtml(self, outputFile):
        if self._ui.chkDailyExc.isChecked():
            return

        self._parent.updateStatusBar("Creating daily.html")
        df = self._rdf[self.TAB_DAILY]
        totals = df.iloc[-1, -1]

        self._div[self.DIV_DAILY] = '''\t\t\t<div class="box">
                <h2 class="graph">Daily statistic counts by classifications</h2>
                <p>Total filtered events displayed: <strong>{0}</strong> of <strong>{1}</strong> unfiltered.</p>
                <h3>1 day resolution</h3>            
'''.format(totals, self._totUnfiltered)
        self._div[self.DIV_DAILY] += self._activeFiltersHtml()
        self._div[self.DIV_DAILY] += '''\t\t\t\t<table class="blueTable">
                    <thead>
                        <tr>
                            <th>UTC date</th>
'''
        if self._isChecked('OVER') == 'checked':
            self._div[self.DIV_DAILY] += '\t\t\t\t\t\t\t<th>Overdense</th>\n'
        if self._isChecked('UNDER') == 'checked':
            self._div[self.DIV_DAILY] += '\t\t\t\t\t\t\t<th>Underdense</th>\n'
        if self._isChecked('FAKE RFI') == 'checked':
            self._div[self.DIV_DAILY] += '\t\t\t\t\t\t\t<th>Fake RFI</th>\n'
        if self._isChecked('FAKE ESD') == 'checked':
            self._div[self.DIV_DAILY] += '\t\t\t\t\t\t\t<th>Fake ESD</th>\n'
        if self._isChecked('FAKE CAR1') == 'checked':
            self._div[self.DIV_DAILY] += '\t\t\t\t\t\t\t<th>Fake Carrier 1</th>\n'
        if self._isChecked('FAKE CAR2') == 'checked':
            self._div[self.DIV_DAILY] += '\t\t\t\t\t\t\t<th>FaKe Carrier 2</th>\n'
        if self._isChecked('FAKE SAT') == 'checked':
            self._div[self.DIV_DAILY] += '\t\t\t\t\t\t\t<th>Fake Saturation</th>\n'
        if self._isChecked('FAKE LONG') == 'checked':
            self._div[self.DIV_DAILY] += '\t\t\t\t\t\t\t<th>Fake Long</th>\n'
        self._div[self.DIV_DAILY] += '\t\t\t\t\t\t\t<th>Daily total</th>\n'
        self._div[self.DIV_DAILY] += '''\t\t\t\t\t\t</tr>
                    </thead>
                    <tbody>
'''
        for row in range(2, len(df.index)):
            self._div[self.DIV_DAILY] += '\t\t\t\t\t\t<tr>\n'
            for col in range(0, len(df.columns)):
                self._div[self.DIV_DAILY] += '\t\t\t\t\t\t\t<td>{}</td>\n'.format(df.iloc[row, col])
            self._div[self.DIV_DAILY] += '\t\t\t\t\t\t</tr>\n'

        self._div[self.DIV_DAILY] += '''\t\t\t\t\t</tbody>
                </table>
            </div><!--box-->
        </div><!--content-->
'''
        self._buildSidebarHtml(outputFile)

        htmlDoc = self._pageHeader
        htmlDoc += self._div[self.DIV_DAILY]
        htmlDoc += self._div[self.DIV_SIDEBAR]
        htmlDoc += self._pageFooter

        outFilePath = self._dirPath / outputFile
        with open(outFilePath, 'w') as f:
            f.write(htmlDoc)

    def _buildHourlyHtml(self, outputFile):
        if self._ui.chkHourlyExc.isChecked():
            return

        self._parent.updateStatusBar("Creating hourly.html")
        df = self._rdf[self.TAB_HOURLY]

        totals = df.iloc[-1, -1]

        self._div[self.DIV_HOURLY] = '''\t\t\t<div class="box">
                    <h2 class="graph">Daily statistic counts</h2>
                      <p>Total filtered events displayed: <strong>{0}</strong> of <strong>{1}</strong> unfiltered.</p>
                    <h3>1 hour resolution</h3>            
'''.format(totals, self._totUnfiltered)

        self._div[self.DIV_HOURLY] += self._activeFiltersHtml()
        self._div[self.DIV_HOURLY] += '''\t\t\t\t<table class="blueTable">
                        <thead>
                            <tr>
'''
        for col in range(0, len(df.columns)):
            self._div[self.DIV_HOURLY] += '\t\t\t\t\t\t\t<th>{}</th>\n'.format(df.iloc[1, col])
        self._div[self.DIV_HOURLY] += '''\t\t\t\t\t\t</tr>
                        </thead>
                        <tbody>
'''
        for row in range(2, len(df.index)):
            self._div[self.DIV_HOURLY] += '\t\t\t\t\t\t<tr>\n'
            for col in range(0, len(df.columns)):
                self._div[self.DIV_HOURLY] += '\t\t\t\t\t\t\t<td>{}</td>\n'.format(df.iloc[row, col])
            self._div[self.DIV_HOURLY] += '\t\t\t\t\t\t</tr>\n'
        self._div[self.DIV_HOURLY] += '''\t\t\t\t\t</tbody>
                    </table>
                </div><!--box-->
            </div><!--content-->
'''
        self._buildSidebarHtml(outputFile)

        htmlDoc = self._pageHeader
        htmlDoc += self._div[self.DIV_HOURLY]
        htmlDoc += self._div[self.DIV_SIDEBAR]
        htmlDoc += self._pageFooter

        outFilePath = self._dirPath / outputFile
        with open(outFilePath, 'w') as f:
            f.write(htmlDoc)

    def _build10minHtml(self, outputFile):
        if self._ui.chk10minExc.isChecked():
            return

        self._parent.updateStatusBar("Creating 10min.html")
        df = self._rdf[self.TAB_10MIN]
        totals = df.iloc[-1, -1]

        self._div[self.DIV_10MIN] = '''\t\t\t<div class="box">
                <h2 class="graph">Daily statistic counts</h2>
                  <p>Total filtered events displayed: <strong>{0}</strong> of <strong>{1}</strong> unfiltered.</p>
                <h3>10 min resolution</h3>            
'''.format(totals, self._totUnfiltered)

        self._div[self.DIV_10MIN] += self._activeFiltersHtml()
        self._div[self.DIV_10MIN] += '\t\t\t\t<table class="blueTable">\n'

        for row in range(1, len(df.index)):
            if row % (self._totalDays + 2) == 1:
                # heading row
                self._div[self.DIV_10MIN] += '''\t\t\t\t\t<thead>
                        <tr>
'''
                for col in range(0, len(df.columns)):
                    self._div[self.DIV_10MIN] += '\t\t\t\t\t\t\t<th>{}</th>\n'.format(df.iloc[row, col])
                self._div[self.DIV_10MIN] += '''\t\t\t\t\t\t</tr>
                    </thead>
'''
            else:
                if row % (self._totalDays + 2) == 2:
                    # begin data body
                    self._div[self.DIV_10MIN] += '\t\t\t\t\t<tbody>\n'

                # data row
                self._div[self.DIV_10MIN] += '\t\t\t\t\t\t<tr>\n'
                for col in range(0, len(df.columns)):
                    self._div[self.DIV_10MIN] += '\t\t\t\t\t\t\t<td>{}</td>\n'.format(df.iloc[row, col])
                self._div[self.DIV_10MIN] += '\t\t\t\t\t\t</tr>\n'

                if row % (self._totalDays + 2) == self._totalDays:
                    # end data body
                    self._div[self.DIV_10MIN] += '\t\t\t\t\t<tbody>\n'

        self._div[self.DIV_10MIN] += '''\t\t\t\t</table>
            </div><!--box-->
        </div><!--content-->
'''
        self._buildSidebarHtml(outputFile)

        htmlDoc = self._pageHeader
        htmlDoc += self._div[self.DIV_10MIN]
        htmlDoc += self._div[self.DIV_SIDEBAR]
        htmlDoc += self._pageFooter

        outFilePath = self._dirPath / outputFile
        with open(outFilePath, 'w') as f:
            f.write(htmlDoc)

    def _buildRMOBhtml(self, outputFile):
        if self._ui.chkRMOBexc.isChecked():
            return

        self._parent.updateStatusBar("Creating rmob.html")
        df = self._rdf[self.TAB_RMOB]
        self._RMOBfile = self._settings.readSettingAsString('RMOBfilePrefix') + ".jpg"
        srcImgFilePath = Path(self._parent.workingDir) / self._RMOBfile
        copy2(srcImgFilePath, self._imgPath)
        relImgPath = self._imgPath.relative_to(self._dirPath)
        RMOBimgPath = relImgPath / self._RMOBfile
        RMOBimgPath = RMOBimgPath.as_posix()

        self._div[self.DIV_RMOB] = '''\t\t\t<div class="box">
                <h2>RMOB summary picture</h2>
                    <img src = "{0}" title = "RMOB summary">
                <h2 class="graph">Current month's Daily statistic counts - RMOB format</h2>           
'''.format(RMOBimgPath)
        self._div[self.DIV_RMOB] += '''\t\t\t\t<table class="blueTable">
                        <thead>
                            <tr>
'''
        for col in range(0, len(df.columns)):
            self._div[self.DIV_RMOB] += '\t\t\t\t\t\t\t<th>{}</th>\n'.format(df.iloc[1, col])
        self._div[self.DIV_RMOB] += '''\t\t\t\t\t\t</tr>
                        </thead>
                        <tbody>
'''
        for row in range(2, len(df.index)):
            self._div[self.DIV_RMOB] += '\t\t\t\t\t\t<tr>\n'
            for col in range(0, len(df.columns)):
                self._div[self.DIV_RMOB] += '\t\t\t\t\t\t\t<td>{}</td>\n'.format(df.iloc[row, col])
            self._div[self.DIV_RMOB] += '\t\t\t\t\t\t</tr>\n'
        self._div[self.DIV_RMOB] += '''\t\t\t\t\t</tbody>
                    </table>

            </div><!--box-->
        </div><!--content-->
'''

        self._buildSidebarHtml(outputFile)

        htmlDoc = self._pageHeaderRmob
        htmlDoc += self._div[self.DIV_RMOB]
        htmlDoc += self._div[self.DIV_SIDEBAR]
        htmlDoc += self._pageFooter

        outFilePath = self._dirPath / outputFile
        with open(outFilePath, 'w') as f:
            f.write(htmlDoc)

    def _buildRTShtml(self, outputFile):
        if self._ui.chkSetupExc.isChecked():
            return

        self._parent.updateStatusBar("Creating setup.html")
        dfList = self._rdf[self.TAB_RTS]

        cfgFile = self._dataSource.stem()
        if 'snapshot_' in cfgFile:
            cfgFile = cfgFile.replace('snapshot_', '')

        self._div[self.DIV_RTS] = '''\t\t\t<div class="box">
                <h2 class="graph">Echoes settings</h2>
                <p>Settings stored in configuration file <strong>{0}.rts</strong></p>
'''.format(cfgFile)

        titles = ['\t\t\t\t\t<h3>Device settings</h3>\n',
                  '\t\t\t\t\t<h3>FFT settings</h3>\n',
                  '\t\t\t\t\t<h3>Output settings</h3>\n',
                  '\t\t\t\t\t<h3>Storage settings</h3>\n',
                  '\t\t\t\t\t<h3>Preferences</h3>\n']

        idx = 0

        for df in dfList:
            self._div[self.DIV_RTS] += titles[idx]
            self._div[self.DIV_RTS] += '''\t\t\t\t\t<table class="blueTable">
                        <thead>
                            <tr>
'''
            # header
            for col in df.columns:
                self._div[self.DIV_RTS] += '\t\t\t\t\t\t\t<th>{}</th>\n'.format(col)
            self._div[self.DIV_RTS] += '''\t\t\t\t\t\t</tr>
                            </thead>
                            <tbody>
'''
            # data
            for row in range(0, len(df.index)):
                self._div[self.DIV_RTS] += '\t\t\t\t\t\t<tr>\n'
                for col in range(0, len(df.columns)):
                    self._div[self.DIV_RTS] += '\t\t\t\t\t\t\t<td>{}</td>\n'.format(df.iloc[row, col])
                self._div[self.DIV_RTS] += '\t\t\t\t\t\t</tr>\n'
            self._div[self.DIV_RTS] += '''\t\t\t\t\t</tbody>
                    </table>
                    <hr>
'''
            idx += 1

        self._div[self.DIV_RTS] += '''\t\t\t</div><!--box-->
        </div><!--content-->        
'''

        self._buildSidebarHtml(outputFile)

        htmlDoc = self._pageHeader
        htmlDoc += self._div[self.DIV_RTS]
        htmlDoc += self._div[self.DIV_SIDEBAR]
        htmlDoc += self._pageFooter

        outFilePath = self._dirPath / outputFile
        with open(outFilePath, 'w') as f:
            f.write(htmlDoc)

    def _buildEventHtml(self, evTuple):
        if self._ui.chkExpEvExc.isChecked():
            return
        (htmlFile, imgFile, commImgFile, detFile, commDetailsFile, powFile, commPowerFile, img2dFile, commImg2dFile, img3dFile, commImg3dFile) = evTuple
        srcFilePath = Path(self._parent.exportDir) / "events"

        srcImgFilePath = srcFilePath / imgFile
        copy2(srcImgFilePath, self._imgPath)

        srcPowFilePath = None
        if powFile is not None:
            srcPowFilePath = srcFilePath / powFile
            copy2(srcPowFilePath, self._imgPath)

        srcImg2dFilePath = None
        if img2dFile is not None:
            srcImg2dFilePath = srcFilePath / img2dFile
            copy2(srcImg2dFilePath, self._imgPath)

        srcImg3dFilePath = None
        if img3dFile is not None:
            srcImg3dFilePath = srcFilePath / img3dFile
            copy2(srcImg3dFilePath, self._imgPath)

        relImgPath = self._imgPath.relative_to(self._dirPath)
        imgFilePath = relImgPath / imgFile

        # comments and details files remain under the export directory
        # to be integrated into the page as HTML, while images are copied
        # under page's assets directory, so the report directory can
        # be moved or compressed without caring of external resources

        # reads comments if present
        commentImg = "\n"
        if commImgFile is not None:
            commImgFilePath = srcFilePath / commImgFile
            with open(commImgFilePath, 'r') as f:
                for line in f.readlines():
                    line = '\t\t\t\t\t' + line.strip() + '<br>\n'
                    commentImg += line
        commentDetails = "\n"
        if commDetailsFile is not None:
            commDetailsFilePath = srcFilePath / commDetailsFile
            with open(commDetailsFilePath, 'r') as f:
                for line in f.readlines():
                    line = '\t\t\t\t\t' + line.strip() + '<br>\n'
                    commentDetails += line
        commentPower = "\n"
        if commPowerFile is not None:
            commImgFilePath = srcFilePath / commPowerFile
            with open(commImgFilePath, 'r') as f:
                for line in f.readlines():
                    line = '\t\t\t\t\t' + line.strip() + '<br>\n'
                    commentPower += line
        commentImg2d = "\n"
        if commImg2dFile is not None:
            commImg2dFilePath = srcFilePath / commImg2dFile
            with open(commImg2dFilePath, 'r') as f:
                for line in f.readlines():
                    line = '\t\t\t\t\t' + line.strip() + '<br>\n'
                    commentImg2d += line
        commentImg3d = "\n"
        if commImg3dFile is not None:
            commImg3dFilePath = srcFilePath / commImg3dFile
            with open(commImg3dFilePath, 'r') as f:
                for line in f.readlines():
                    line = '\t\t\t\t\t' + line.strip() + '<br>\n'
                    commentImg3d += line

        # reads the event details table if present
        df = None
        if detFile is not None:
            detFilePath = srcFilePath / detFile
            # reads the csv in a dataframe
            df = pd.read_csv(detFilePath, sep=self._settings.dataSeparator())

        self._buildSidebarHtml(imgFile)

        dt, id = imgFile[-20:-4].split('_')
        htmlDoc = self._pageHeader
        htmlDoc += '''\t\t\t\t<div class="box">
                    <h2 class="graph">Events gallery, daily id#{0}</h2>
                    <hr>
                    <h3>Screenshot</h3>
                        <p>{1}</p>
                        <img src="{2}" alt="{3}">
                    <hr>
'''.format(id, commentImg, imgFilePath.as_posix(), imgFilePath.name)

        if powFile is not None:
            powFilePath = relImgPath / powFile
            htmlDoc += '''\t\t\t\t\t<h3>Total power profile</h3>
                        <p>{0}</p>
                        <img src="{1}" alt="{2}">
                    <hr>
'''.format(commentPower, powFilePath.as_posix(), powFilePath.name)

        if img2dFile is not None:
            img2dFilePath = relImgPath / img2dFile
            htmlDoc += '''\t\t\t\t\t<h3>2D data plotting</h3>
                        <p>{0}</p>
                        <img src="{1}" alt="{2}">
                    <hr>
'''.format(commentImg2d, img2dFilePath.as_posix(), img2dFilePath.name)

        if img3dFile is not None:
            img3dFilePath = relImgPath / img3dFile
            htmlDoc += '''\t\t\t\t\t<h3>3D data plotting</h3>
                        <p>{0}</p>
                        <img src="{1}" alt="{2}">
                    <hr>
'''.format(commentImg3d, img3dFilePath.as_posix(), img3dFilePath.name)

        if df is not None:
            htmlDoc += '''\t\t\t\t\t<h3>Event details</h3>
                        <p>{0}</p>
'''.format(commentDetails)
            htmlDoc += '''\t\t\t\t\t<table class="blueTable" style="table-layout: fixed;word-wrap: break-word;">
                            <thead>
                                <tr>
'''
            for col in df.columns:
                if "Unnamed: 0" in col:
                    col = " "
                htmlDoc += '\t\t\t\t\t\t\t\t<th>{}</th>\n'.format(col)
            htmlDoc += '''\t\t\t\t\t\t\t\t\t</tr>
                        </thead>
                        <tbody>
'''
            for row in range(0, len(df.index)):
                htmlDoc += '\t\t\t\t\t\t\t<tr>\n'
                for col in range(0, len(df.columns)):
                    if col == 0:
                        htmlDoc += '\t\t\t\t\t\t\t\t<td><strong>{}</strong></td>\n'.format(df.iloc[row, 0])
                    else:
                        htmlDoc += '\t\t\t\t\t\t\t\t<td>{}</td>\n'.format(df.iloc[row, col])
                htmlDoc += '\t\t\t\t\t\t\t</tr>\n'
            htmlDoc += '''\t\t\t\t\t\t</tbody>
                    </table>
                </div><!--box-->
            </div><!--content-->
'''
        else:
            htmlDoc += '''\t\t\t\t</div><!--box-->
            </div><!--content-->
'''
        htmlDoc += self._div[self.DIV_SIDEBAR]
        htmlDoc += self._pageFooter

        outFilePath = self._dirPath / htmlFile
        with open(outFilePath, 'w') as f:
            f.write(htmlDoc)

    def _buildStatTablesHtml(self, stTup):
        if self._ui.chkExpStTabExc.isChecked():
            return

        # reads the comments file if existing
        comment = "\n"
        if stTup[2] is not None:
            txtFile = str(stTup[2])
            txtFilePath = Path(self._parent.exportDir) / "statistics" / txtFile
            with open(txtFilePath, 'r') as f:
                for line in f.readlines():
                    line = '\t\t\t\t\t' + line.strip() + '<br>\n'
                    comment += line
            comment += '\t\t\t\t\t'

        # reads the csv in a dataframe
        tableFile = str(stTup[0])
        outputFile = str(stTup[1])
        srcTabFilePath = Path(self._parent.exportDir) / "statistics" / tableFile
        df = pd.read_csv(srcTabFilePath, sep=self._settings.dataSeparator())

        self._buildSidebarHtml(outputFile)

        # builds the page
        htmlDoc = self._pageHeader
        htmlDoc += '''\t\t\t\t<div class="box">
                            <h2 class="data">Statistic data gallery</h2>
                            <p>{0}</p>
'''.format(comment)

        htmlDoc += '''\t\t\t\t<table class="blueTable">
                                <thead>
                                    <tr>
'''
        for col in df.columns:
            if "Unnamed: 0" in col:
                col = " "
            htmlDoc += '\t\t\t\t\t\t\t<th>{}</th>\n'.format(col)
        htmlDoc += '''\t\t\t\t\t\t</tr>
                                </thead>
                            <tbody>
'''
        for row in range(0, len(df.index)):
            htmlDoc += '\t\t\t\t\t\t<tr>\n'
            for col in range(0, len(df.columns)):
                htmlDoc += '\t\t\t\t\t\t\t<td>{}</td>\n'.format(df.iloc[row, col])
            htmlDoc += '\t\t\t\t\t\t</tr>\n'
        htmlDoc += '''\t\t\t\t\t</tbody>
                            </table>
                        </div><!--box-->
                    </div><!--content-->
        '''

        htmlDoc += self._div[self.DIV_SIDEBAR]
        htmlDoc += self._pageFooter

        outFilePath = self._dirPath / outputFile
        with open(outFilePath, 'w') as f:
            f.write(htmlDoc)

    def _buildStatGraphsHtml(self, stTup):
        if self._ui.chkExpStGrpExc.isChecked():
            return
        imageFile = str(stTup[0])
        outputFile = str(stTup[1])
        srcImgFilePath = Path(self._parent.exportDir) / "statistics" / imageFile
        copy2(srcImgFilePath, self._imgPath)
        relImgPath = self._imgPath.relative_to(self._dirPath)
        imgFilePath = relImgPath / imageFile
        comment = "\n"
        if stTup[2] is not None:
            txtFile = str(stTup[2])
            txtFilePath = Path(self._parent.exportDir) / "statistics" / txtFile
            with open(txtFilePath, 'r') as f:
                for line in f.readlines():
                    line = '\t\t\t\t\t' + line.strip() + '<br>\n'
                    comment += line
            comment += '\t\t\t\t\t'
        self._buildSidebarHtml(imageFile)
        htmlDoc = self._pageHeader
        htmlDoc += '''\t\t\t\t<div class="box">
                    <h2 class="graph">Statistic data gallery</h2>
                    <p>{0}</p>
                    <img src="{1}" alt="{2}">
                </div>
            </div><!--content-->
'''.format(comment, imgFilePath.as_posix(), imageFile)

        htmlDoc += self._div[self.DIV_SIDEBAR]
        htmlDoc += self._pageFooter

        outFilePath = self._dirPath / outputFile
        with open(outFilePath, 'w') as f:
            f.write(htmlDoc)

    def _buildSidebarHtml(self, currentFile):
        self._div[self.DIV_SIDEBAR] = '\t\t\t<div id="sidebar">\n'

        df = self._rdf[self.TAB_COVERAGE]
        if df is None:
            self._parent.updateStatusBar("skipping excluded section")
            return

        if currentFile != 'index.html':
            divBoxBack = '''\t\t\t\t\t<div class="box index">
                        <a href="index.html" title="Back to index" class="current"><strong>Back to Index</strong></a>
                    </div><!-- box -->
'''
        else:
            divBoxBack = ''

        divBoxDate = '''\t\t\t\t<div class="box date">
                    <h2 class="calendar">Time span</h2>
                    <dl>
                        <dt>From</dt>        
                        <dd>{0}</dd>
                        <dt>To</dt>
                        <dd>{1}</dd>
                        <dt>Total</dt>
                        <dd>{2} days</dd>
                        <dt>Uptime</dt>
                        <dd>{3} min.</dd>
                    </dl>
                </div><!--date-->
'''.format(df.loc['4', 'B'], df.loc['4', 'D'], self._totalDays, df.loc['5', 'B'])

        df = self._rdf[self.TAB_SITEINFOS]

        origLogoPath = df.loc['2', 'C']
        origLogo = QPixmap(origLogoPath)
        destPath = self._imgPath / origLogoPath
        origLogoName = destPath.name
        origLogo.save(origLogoName)
        logoPath = Path("assets/img") / origLogoName

        divSiteInfo = '''\t\t\t\t<div class="box">
                    <h2 class="site">Site information</h2>
                    <img src="{0}" alt="Site logo not found" width="128">  
                    <dl>
                        <dt>Station name</dt>
                        <dd>{1}</dd>
                        <dt>Owner</dt>
                        <dd>{2}</dd>
                        <dt>Country</dt>
                        <dd>{3}</dd>
                        <dt>City</dt>
                        <dd>{4}</dd>
                        <dt>Longitude</dt>
                        <dd>{5}</dd>
                        <dt>Latitude</dt>
                        <dd>{6}</dd>
                        <dt>Altitude</dt>
                        <dd>{7}m</dd>
                    </dl>
                </div><!-- box -->
'''.format(
            logoPath.as_posix(),  # logo
            df.loc['2', 'B'],  # station name
            df.loc['3', 'B'],  # owner
            df.loc['4', 'B'], df.loc['4', 'D'],  # country, city
            df.loc['5', 'B'], df.loc['5', 'D'],  # longitude, latitude
            df.loc['6', 'B']  # altitude
        )

        divReceiver = '''\t\t\t\t<div class="box">
                    <h2 class="radio">Receiver details</h2>
                    <dl>
                        <dt>Antenna type</dt>
                        <dd>{0}</dd>
                        <dt>Azimuth</dt>
                        <dd>{1}</dd>
                        <dt>Elevation</dt>
                        <dd>{2}</dd>
                        <dt>Preamplifier</dt>
                        <dd>{3}</dd>
                        <dt>Receiver</dt>
                        <dd>{4}</dd>
                        <dt>Frequencies</dt>
                        <dd>{5} Hz</dd>
                        <dt>Computer type</dt>
                        <dd>{6}</dd>
                        <dt>Software</dt>
                        <dd>{7}</dd>
                        <dt>e-mail</dt>
                        <dd>{8}</dd>
                        <dt>Notes:</dt>
                        <dd>{9}</dd>
                    </dl>
                </div><!-- box -->
'''.format(df.loc['7', 'B'], df.loc['7', 'D'], df.loc['7', 'F'],
           df.loc['8', 'B'], 
           df.loc['9', 'B'], 
           df.loc['10', 'B'],
           df.loc['11', 'B'],
           df.loc['12', 'B'], 
           df.loc['13', 'B'], 
           df.loc['14', 'B'])

        divTables = '''\t\t\t\t<div class="box">
                    <h2 class="data">Tables</h2>
                    <ul>
'''
        if not self._ui.chkDailyExc.isChecked():
            divTables += '\t\t\t\t\t\t<li><a href="daily.html" title="Daily counts by classification" {}>Daily ' \
                         'counts by classifications</a></li>\n' \
                .format('class="current"' if currentFile == 'daily.html' else '')
        if not self._ui.chkHourlyExc.isChecked():
            divTables += '\t\t\t\t\t\t<li><a href="hourly.html" title="Hourly counts by day" {}>Hourly counts by ' \
                         'day</a></li>\n' \
                .format('class="current"' if currentFile == 'hourly.html' else '')
        if not self._ui.chk10minExc.isChecked():
            divTables += '\t\t\t\t\t\t<li><a href="10min.html" title="10 minutes counts by day" {}>10 minutes counts ' \
                         'by day</a></li>\n '.format('class="current"' if currentFile == '10min.html' else '')
        if not self._ui.chkRMOBexc.isChecked():
            divTables += '\t\t\t\t\t\t<li><a href="rmob.html"  title="Current month\'s RMOB counts" {}>Current ' \
                         'month\'s RMOB counts</a></li>\n '.format(
                'class="current"' if currentFile == 'rmob.html' else '')
        if not self._ui.chkChronoExc.isChecked():
            divTables += '\t\t\t\t\t\t<li><a href="chrono.html"  title="Event\'s chronological table" {}>' \
                         'Event\'s chronological table</a></li>\n '.format(
                'class="current"' if currentFile == 'chrono.html' else '')
        if not self._ui.chkSetupExc.isChecked():
            divTables += '\t\t\t\t\t\t<li><a href="setup.html"  title="Echoes setup tables" {}>' \
                         'Echoes setup tables (RTS)</a></li>\n '.format(
                'class="current"' if currentFile == 'setup.html' else '')
        divTables += '''
                    </ul>
                </div><!-- box -->
'''

        divEvents = '''\t\t\t\t<div class="box">
                    <h2 class="events">Exported events gallery</h2>
                    <ul>
'''

        if self._ui.chkExpEvExc.isChecked():
            divEvents += '\t\t\t\t\t\t<li>None</li>\n'
        else:
            for evTuple in self._events:
                (htmlFile, imgFile, commImgFile, detFile, commDetailsFile, powFile, commPowerFile, img2dFile, commImg2dFile, img3dFile, commImg3dFile) = evTuple
                dt, id = imgFile[-20:-4].split('_')
                isCurrent = 'class="current"' if currentFile == imgFile else ''
                divEvents += '\t\t\t\t\t\t<li><a href="{0}" title="{1}" {4}>Event#{2}, {3}</a></li>\n'.format(
                    htmlFile, imgFile, id, dt, isCurrent)

        divEvents += '''
                    </ul>
                </div><!-- box -->
'''

        divStats = '''\t\t\t\t<div class="box">
                    <h2 class="graph">Exported statistics data gallery</h2>
                    <ul>
'''
        if self._ui.chkExpStTabExc.isChecked():
            divStats += '\t\t\t\t\t\t<li>None</li>\n'
        else:
            for stTup in self._statImgs:
                imgName = str(stTup[0])
                htmlName = str(stTup[1])
                fields = imgName.split('-')
                kind = fields[-1].split('.')[0]
                title = fields[1].replace('_', ' ')
                if len(fields) >= 5:
                    resolution = fields[2].replace('_', ' ')
                    prog = fields[3]
                else:
                    resolution = ''
                    prog = fields[2]
                isCurrent = 'class="current"' if currentFile == imgName else ''
                divStats += '\t\t\t\t\t\t<li><a href="{0}" title="{1}" {6}>{2} {3} {4} #{5}</a></li>\n'.format(
                    htmlName, imgName, kind, title, resolution, prog, isCurrent)

            for tabTup in self._statTabs:
                tabName = str(tabTup[0])
                htmlName = str(tabTup[1])
                fields = tabName.split('-')
                kind = fields[-1].split('.')[0]
                title = fields[1].replace('_', ' ')
                if len(fields) >= 5:
                    resolution = fields[2].replace('_', ' ')
                    prog = fields[3]
                else:
                    resolution = ''
                    prog = fields[2]

                isCurrent = 'class="current"' if currentFile == tabName else ''
                divStats += '\t\t\t\t\t\t<li><a href="{0}" title="{1}" {6}>{2} {3} {4} #{5}</a></li>\n'.format(
                    htmlName, tabName, kind, title, resolution, prog, isCurrent)

        divStats += '''
                    </ul>
                </div><!-- box -->
'''

        self._div[self.DIV_SIDEBAR] += divBoxBack + '\n'
        self._div[self.DIV_SIDEBAR] += divBoxDate + '\n'
        self._div[self.DIV_SIDEBAR] += divTables + '\n'
        self._div[self.DIV_SIDEBAR] += divEvents + '\n'
        self._div[self.DIV_SIDEBAR] += divStats + '\n'
        self._div[self.DIV_SIDEBAR] += divSiteInfo + '\n'
        self._div[self.DIV_SIDEBAR] += divReceiver + '\n'
        self._div[self.DIV_SIDEBAR] += '\t\t\t</div><!--sidebar-->\n'

    def _isChecked(self, filt):
        if filt in self._classFilter:
            return 'checked'
        return ''

    def _filterDescription(self, filt):
        descDict = dict()
        descDict['OVER'] = 'Overdense'
        descDict['UNDER'] = 'Underdense'
        descDict['FAKE RFI'] = 'Fake RFI'
        descDict['FAKE ESD'] = 'Fake ESD'
        descDict['FAKE CAR1'] = 'Fake Carrier 1'
        descDict['FAKE CAR2'] = 'Fake Carrier 2'
        descDict['FAKE SAT'] = 'Fake Saturation'
        descDict['FAKE LONG'] = 'Fake Long'
        return descDict[filt]

    def _activeFiltersHtml(self):
        html = '''\t\t\t\t<h4>Active filters:</h4>
                <ol>
                    <li> 
                        <input type="checkbox" id="od" name="od" value="Overdense events" {0} onclick="return false;">
                        <label for="od">Overdense events</label>
                    </li>
                    <li> 
                        <input type="checkbox" id="ud" name="ud" value="Underdense events" {1} onclick="return false;">
                        <label for="ud">Underdense events</label>
                    </li>
                    <li> 
                        <input type="checkbox" id="rfi" name="rfi" value="Fake events (RFI)" {2} onclick="return false;">
                        <label for="rfi">Fake events (RFI)</label>
                    </li>
                    <li> 
                        <input type="checkbox" id="esd" name="esd" value="Fake events (ESD)" {2} onclick="return false;">
                        <label for="esd">Fake events (ESD)</label>
                    </li>
                    <li> 
                        <input type="checkbox" id="car1" name="car1" value="Fake events (Carrier type 1)" {3} onclick="return false;">
                        <label for="car1">Fake events (Carrier type 1)</label>
                    </li>
                    <li> 
                        <input type="checkbox" id="car2" name="car2" value="Fake events (Carrier type 2)" {4} onclick="return false;">
                        <label for="car2">Fake events (Carrier type 2)</label>
                    </li>
                    <li> 
                        <input type="checkbox" id="sat" name="sat" value="Fake events (RX saturation)" {5} onclick="return false;">
                        <label for="car2">Fake events (RX saturation)</label>
                    </li>
                    <li> 
                        <input type="checkbox" id="long" name="long" value="Fake events (too long duration)" {6} onclick="return false;">
                        <label for="car2">Fake events (too long duration)</label>
                    </li>
                </ol>
'''.format(self._isChecked('OVER'), self._isChecked('UNDER'), self._isChecked('FAKE RFI'),
           self._isChecked('FAKE ESD'), self._isChecked('FAKE CAR1'),
           self._isChecked('FAKE CAR2'), self._isChecked('FAKE SAT'), self._isChecked('FAKE LONG'))
        return html

    def _formatPreface(self, prefaceDef: str):
        # replacing macros [] in prefaceDef

        station = self._settings.readSettingAsString("stationName")
        country = self._settings.readSettingAsString("country")
        city = self._settings.readSettingAsString("city")
        dateFrom = self._settings.readSettingAsString("dateFrom")
        dateTo =self._settings.readSettingAsString("dateTo")
        unclassified = self._dataSource.totalsUnclassified(
            self._settings.readSettingAsString("dateFrom"),
            self._settings.readSettingAsString("dateTo"))

        preface = prefaceDef.replace("[station]", station)

        preface = preface.replace("[country]", country)
        preface = preface.replace("[city]", city)
        preface = preface.replace("[dateFrom]", dateFrom)
        preface = preface.replace("[dateTo]", dateTo)
        preface = preface.replace("[unclassified]", str(unclassified))
        preface = preface.replace("[ebrow]", f"Ebrow v.{self._parent.version}")

        return preface
