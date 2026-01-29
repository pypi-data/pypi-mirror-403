import numpy as np
import openpyxl

from steam_sdk.builders.BuilderProteCCT import BuilderProteCCT
from steam_sdk.data.TemplateProteCCT import get_template_ProteCCT_input_sheet
from steam_sdk.parsers.ParserExcel import read_row, write2Excel
from steam_sdk.utils.compare_two_parameters import compare_two_parameters


class ParserProteCCT:
    """
        Class with methods to read/write ProteCCT information from/to other programs
    """

    def __init__(self, builder_protecct: BuilderProteCCT = BuilderProteCCT(flag_build=False)):
        """
            Initialization using a BuilderProteCCT object containing ProteCCT parameter structure
        """

        self.builder_protecct: BuilderProteCCT = builder_protecct


    def readFromExcel(self, file_name: str, verbose: bool = True):
        """
            *** Function that reads an Excel file defining a ProteCCT input file and imports it in a BuilderProteCCT object ***

            :param file_name: Name of the file to read
            :type file_name: str
            :param verbose: Flag indicating whether additional information should be displayed
            :type verbose: str

            :return: None
        """

        # Unpack variables
        builder_protecct = self.builder_protecct

        ##File must be whole eos string
        workbookVariables = openpyxl.load_workbook(file_name, data_only=True)

        #Inputs
        worksheetInputs = workbookVariables['Inputs']
        lastAttribute = worksheetInputs.cell(1, 2).value
        for i in range(1, worksheetInputs.max_row+1):
            # builder_protecct.variablesInputs[str(worksheetInputs.cell(i, 2).value)] = str(worksheetInputs.cell(i, 1).value)
            attribute = worksheetInputs.cell(i, 2).value
            try:
                if attribute is None:
                    if worksheetInputs.cell(i, 3).value is not None:
                        values = read_row(worksheetInputs, i)
                        values = np.array([k for k in values if(str(k))])
                        current = builder_protecct.getAttribute(builder_protecct.Inputs, lastAttribute)
                        current = np.vstack((current, values))
                        builder_protecct.setAttribute(builder_protecct.Inputs, lastAttribute, current)
                    else:
                        continue
                elif type(builder_protecct.getAttribute(builder_protecct.Inputs, attribute)) == np.ndarray:
                    lastAttribute = attribute
                    values = read_row(worksheetInputs, i)
                    values = np.array([k for k in values if(str(k))])
                    builder_protecct.setAttribute(builder_protecct.Inputs, attribute, values)
                else:
                    value = worksheetInputs.cell(i, 3).value
                    builder_protecct.setAttribute(builder_protecct.Inputs, attribute, value)
            except TypeError as e:
                if attribute in builder_protecct.Inputs.__annotations__: raise e
                if attribute == 'None' or attribute == None: continue
                if verbose: print("Error with attribute: {}, continuing.".format(attribute))

        return builder_protecct


    def writeProtecct2Excel(self, full_path_file_name: str, verbose: bool = False):
        """
        ** Writes a ProteCCT input file **

        :param full_path_file_name:
        :param verbose:
        :return:
        """

        # Import template for ProteCCT input file
        template_ProteCCT_input_sheet = get_template_ProteCCT_input_sheet()

        # Define optional variables, which will be written only if present in the dataclass
        optional_variables_input_sheet = []

        name_sheet = 'Inputs'  # This defines the name of the sheet and also of the variable group
        ProteCCT_input_sheet = []
        for row in template_ProteCCT_input_sheet:
            name, description = row[0], row[2]

            # If the row defines an empty row, or a title row, leave it unchanged
            if name is None:
                ProteCCT_input_sheet.append(row)  # Leave the row unchanged
                continue  # stop treating this row: go to the next row

            # Get value of the current parameter
            value = self.builder_protecct.getAttribute(name_sheet, name)

            # Skip optional variables if they have 0 elements
            if name in optional_variables_input_sheet:
                if isinstance(value, list) and len(value) == 0:
                    print('Variable {} is optional and has 0 elements, hence it will be skipped.'.format(name))
                    continue
                elif isinstance(value, np.ndarray) and value.shape[0] == 0:
                    print('Variable {} is optional and has 0 elements, hence it will be skipped.'.format(name))
                    continue
                elif len(value) == 1 and value is None:
                    print('Variable {} is optional and has 0 elements, hence it will be skipped.'.format(name))
                    continue

            # Assign value to the variable sheet
            ProteCCT_input_sheet.append([name, value, description])
            # TODO exception handling

        # Write ProteCCT Excel input file
        write2Excel(name_file=full_path_file_name, name_sheets='Inputs', listOf_variables_values_descriptions=[ProteCCT_input_sheet], verbose=verbose)


#######################  Helper functions - START  #######################
def CompareProteCCTParameters(fileA, fileB, max_relative_error=1E-5, verbose=True):
    """
        Compare all the variables imported from two ProteCCT Excel input files
    """

    Diff = False

    pp_a = ParserProteCCT(BuilderProteCCT(flag_build=False))
    pp_a.readFromExcel(fileA, verbose=False)
    pp_b = ParserProteCCT(BuilderProteCCT(flag_build=False))
    pp_b.readFromExcel(fileB, verbose=False)
    print("Starting Comparison of A: ({}) and B: ({})".format(fileA, fileB))

    ## Check Inputs
    for attribute in pp_a.builder_protecct.Inputs.__annotations__:
        last_diff = compare_two_parameters(pp_a.builder_protecct.getAttribute("Inputs", attribute),
                                      pp_b.builder_protecct.getAttribute("Inputs", attribute),
                                      attribute, max_relative_error, flag_difference_messages=True, verbose=verbose)
        if last_diff:
            Diff = True

    if Diff == False:
        print("Files {} and {} are equal.".format(fileA, fileB))
        return True
    else:
        return False
#######################  Helper functions - END  #######################