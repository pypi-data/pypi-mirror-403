import os
from pathlib import Path
from typing import Dict

import numpy as np
import datetime
import shutil
import openpyxl
import json
import ruamel.yaml

from steam_sdk.builders.BuilderLEDET import BuilderLEDET
from steam_sdk.data.DataLEDET import LEDETInputs, LEDETOptions, LEDETPlots, LEDETVariables
from steam_sdk.data.TemplateLEDET import get_template_LEDET_inputs_sheet, get_template_LEDET_options_sheet, get_template_LEDET_plots_sheet, get_template_LEDET_variables_sheet
from steam_sdk.parsers.ParserMap2d import ParserMap2dFile, ParserMap2dData
from steam_sdk.parsers.ParserExcel import read_row, write2Excel
from steam_sdk.parsers.ParserYAML import dict_to_yaml
from steam_sdk.utils.NumpyEncoder import NumpyEncoder
from steam_sdk.utils.rgetattr import rgetattr
from steam_sdk.utils.sgetattr import rsetattr
from steam_sdk.utils.compare_two_parameters import compare_two_parameters
from steam_sdk.utils.MatrixOperations import multiply_column_by_value


class ParserLEDET:
    """
        Class with methods to read/write LEDET information from/to other programs
    """

    def __init__(self, builder_ledet: BuilderLEDET):
        """
            Initialization using a BuilderLEDET object containing LEDET parameter structure
        """

        self.builder_ledet: BuilderLEDET = builder_ledet


    def readFromExcel(self, file_name: str, verbose: bool = True):
        '''
            *** Function that reads an Excel file defining a LEDET input file and imports it in a BuilderLEDET object ***

            :param file_name: Name of the file to read
            :type file_name: str
            :param verbose: Flag indicating whether additional information should be displayed
            :type verbose: str

            :return: None
        '''

        # Unpack variables
        builder_ledet = self.builder_ledet

        ##File must be whole eos string
        workbookVariables = openpyxl.load_workbook(file_name, data_only=True)

        #Inputs
        worksheetInputs = workbookVariables['Inputs']
        lastAttribute = worksheetInputs.cell(1, 2).value
        list_exceptions_to_ignore = ['STEAM-LEDET website', 'STEAM-LEDET manual', 'STEAM-SDK website', '[1]', '[2]', '[3]', '[4]']
        for i in range(1, worksheetInputs.max_row+1):
            # builder_ledet.variablesInputs[str(worksheetInputs.cell(i, 2).value)] = str(worksheetInputs.cell(i, 1).value)
            attribute = worksheetInputs.cell(i, 2).value
            try:
                if (attribute == None):
                    if worksheetInputs.cell(i, 1).value in list_exceptions_to_ignore:
                        continue
                    if worksheetInputs.cell(i, 3).value is not None:
                        values = read_row(worksheetInputs, i)
                        values = np.array([k for k in values if(str(k))])
                        current = builder_ledet.getAttribute(builder_ledet.Inputs, lastAttribute)
                        current = np.vstack((current, values))
                        builder_ledet.setAttribute(builder_ledet.Inputs, lastAttribute, current)
                    else:
                        continue
                elif (type(builder_ledet.getAttribute(builder_ledet.Inputs, attribute)) == np.ndarray):
                    lastAttribute = attribute
                    values = read_row(worksheetInputs, i)
                    values = np.array([k for k in values if(str(k))])
                    builder_ledet.setAttribute(builder_ledet.Inputs, attribute, values)
                else:
                    value = worksheetInputs.cell(i, 3).value
                    builder_ledet.setAttribute(builder_ledet.Inputs, attribute, value)
            except TypeError as e:
                if attribute in builder_ledet.Inputs.__annotations__: raise e
                if attribute=='None' or attribute==None: continue
                if verbose: print("Error with attribute: {}, continuing.".format(attribute))
        #Options
        worksheetOptions = workbookVariables['Options']
        for i in range(1, worksheetOptions.max_row+1):
            # builder_ledet.variablesOptions[str(worksheetOptions.cell(i, 2).value)] = str(worksheetOptions.cell(i, 1).value)
            attribute = worksheetOptions.cell(i, 2).value
            try:
                if (type(builder_ledet.getAttribute(builder_ledet.Options, attribute)) == np.ndarray):
                    values = read_row(worksheetOptions, i)
                    values = np.array([k for k in values if(str(k))])
                    builder_ledet.setAttribute(builder_ledet.Options, attribute, values)
                else:
                    value = worksheetOptions.cell(i, 3).value
                    builder_ledet.setAttribute(builder_ledet.Options, attribute, value)
            except TypeError as e:
                if attribute in builder_ledet.Options.__annotations__: raise e
                if attribute == 'None' or attribute == None: continue
                if verbose: print("Error with attribute: {}, continuing.".format(attribute))
        #Plots
        worksheetPlots = workbookVariables['Plots']
        for i in range(1, worksheetPlots.max_row+1):
            # builder_ledet.variablesPlots[str(worksheetPlots.cell(i, 2).value)] = str(worksheetPlots.cell(i, 1).value)
            attribute = worksheetPlots.cell(i, 2).value
            try:
                if (type(builder_ledet.getAttribute(builder_ledet.Plots, attribute)) == np.ndarray):
                    values = read_row(worksheetPlots, i, St=True)[2:]
                    values = np.array([k for k in values if(str(k))])
                    builder_ledet.setAttribute(builder_ledet.Plots, attribute, values)
                else:
                    try:
                        value = worksheetPlots.cell(i, 3).value
                    except:
                        value = ''
                    builder_ledet.setAttribute(builder_ledet.Plots, attribute, value)
            except TypeError as e:
                if attribute == 'None' or attribute == None: continue
                if verbose: print("Error with attribute: {}, continuing.".format(attribute))
        # Variables
        try:
            worksheetVariables = workbookVariables['Variables']
            for i in range(1, worksheetVariables.max_row+1):
                # builder_ledet.variablesVariables[str(worksheetVariables.cell(i, 2).value)] = str(worksheetVariables.cell(i, 1).value)
                attribute = worksheetVariables.cell(i, 2).value
                try:
                    if (type(builder_ledet.getAttribute(builder_ledet.Variables, attribute)) == np.ndarray):
                        if attribute != 'typeVariableToSaveTxt':  values = read_row(worksheetVariables, i, St = True)[2:]
                        else:  values = read_row(worksheetVariables, i)
                        values = np.array([k for k in values if(str(k))])
                        builder_ledet.setAttribute(builder_ledet.Variables, attribute, values)
                    else:
                        value = worksheetVariables.cell(i, 3).value
                        builder_ledet.setAttribute(builder_ledet.Variables, attribute, value)
                except TypeError as e:
                    if attribute in builder_ledet.Variables.__annotations__: raise e
                    if attribute == 'None' or attribute == None: continue
                    if verbose: print("Error with attribute: {}, continuing.".format(attribute))
        except:
            pass
            print("Error while reading Variables. Please check!")

        return builder_ledet


    def writeLedet2Excel(self, full_path_file_name: str, verbose: bool = False):
        '''
        ** Writes a LEDET input file **

        :param full_path_file_name:
        :param verbose:
        :return:
        '''

        self._expand_scalar_to_array(verbose)

        # Import templates for LEDET input file sheets
        template_LEDET_inputs_sheet    = get_template_LEDET_inputs_sheet()
        template_LEDET_options_sheet   = get_template_LEDET_options_sheet()
        template_LEDET_plots_sheet     = get_template_LEDET_plots_sheet()
        template_LEDET_variables_sheet = get_template_LEDET_variables_sheet()

        # Define optional variables, which will be written only if present in the dataclass
        optional_variables_input_sheet  = _getOptionalVariables_input()
        optional_variables_options_sheet = _getOptionalVariables_options()

        ### Inputs sheet
        name_sheet_inputs = 'Inputs'  # This defines the name of the sheet and also of the variable group
        LEDET_inputs_sheet = []
        for row in template_LEDET_inputs_sheet:
            name, description = row[0], row[2]

            # If the row defines an empty row, or a title row, leave it unchanged
            if name == None:
                LEDET_inputs_sheet.append(row)  # Leave the row unchanged
                continue  # stop treating this row: go to the next row

            # Get value of the current parameter
            value = self.builder_ledet.getAttribute(name_sheet_inputs, name)

            # Skip optional variables if they have 0 elements, or if they are all NaN
            if name in optional_variables_input_sheet:
                if isinstance(value, list) and len(value) == 0:
                    if verbose: print('Variable {} is optional and has 0 elements, hence it will be skipped.'.format(name))
                    continue
                elif isinstance(value, np.ndarray) and value.shape[0] == 0:
                    if verbose: print('Variable {} is optional and has 0 elements, hence it will be skipped.'.format(name))
                    continue
                elif value is None:
                    if verbose: print('Variable {} is optional and has 0 elements, hence it will be skipped.'.format(name))
                    continue
                elif isinstance(value, np.ndarray) and np.all(np.isnan(value.astype(float))):
                    if verbose: print('Variable {} is optional and has only NaN elements, hence it will be skipped.'.format(name))
                    continue

            # Assign value to the variable sheet
            LEDET_inputs_sheet.append([name, value, description])

        ### Options sheet
        name_sheet_options = 'Options'  # This defines the name of the sheet and also of the variable group
        LEDET_options_sheet = []
        for row in template_LEDET_options_sheet:
            name, description = row[0], row[2]

            # If the row defines an empty row, or a title row, leave it unchanged
            if name == None:
                LEDET_options_sheet.append(row)  # Leave the row unchanged
                continue  # stop treating this row: go to the next row

            # Get value of the current parameter
            value = self.builder_ledet.getAttribute(name_sheet_options, name)

            # Skip optional variables if they have 0 elements
            if name in optional_variables_options_sheet:
                if isinstance(value, list) and len(value) == 0:
                    if verbose: print('Variable {} is optional and has 0 elements, hence it will be skipped.'.format(name))
                    continue
                elif isinstance(value, np.ndarray) and value.shape[0] == 0:
                    if verbose: print('Variable {} is optional and has 0 elements, hence it will be skipped.'.format(name))
                    continue
                elif value is None:
                    if verbose: print('Variable {} is optional and has 0 elements, hence it will be skipped.'.format(name))
                    continue

            # Assign value to the variable sheet
            LEDET_options_sheet.append([name, value, description])

        ### Plots sheet
        name_sheet_plots = 'Plots'  # This defines the name of the sheet and also of the variable group
        LEDET_plots_sheet = []
        for row in template_LEDET_plots_sheet:
            name, description = row[0], row[2]

            # If the row defines an empty row, or a title row, leave it unchanged
            if name == None:
                LEDET_plots_sheet.append(row)  # Leave the row unchanged
                continue  # stop treating this row: go to the next row

            # Assign value to the variable sheet
            value = self.builder_ledet.getAttribute(name_sheet_plots, name)
            LEDET_plots_sheet.append([name, value, description])

        ### Variables sheet
        name_sheet_variables = 'Variables'  # This defines the name of the sheet and also of the variable group
        LEDET_variables_sheet = []
        for row in template_LEDET_variables_sheet:
            name, description = row[0], row[2]

            # If the row defines an empty row, or a title row, leave it unchanged
            if name == None:
                LEDET_variables_sheet.append(row)  # Leave the row unchanged
                continue  # stop treating this row: go to the next row

            # Assign value to the variable sheet
            value = self.builder_ledet.getAttribute(name_sheet_variables, name)
            LEDET_variables_sheet.append([name, value, description])

        # Write LEDET Excel input file
        write2Excel(name_file=full_path_file_name,
                    name_sheets=[name_sheet_inputs, name_sheet_options, name_sheet_plots, name_sheet_variables],
                    listOf_variables_values_descriptions=[LEDET_inputs_sheet, LEDET_options_sheet, LEDET_plots_sheet, LEDET_variables_sheet],
                    verbose=verbose)


    def write2json(self, full_path_file_name: str, verbose: bool = True):
        '''
            *** Function that writes a LEDET input file as a .json file ***

            Function to write a LEDET input file containing "Inputs", "Options", "Plots", and "Variables" variables

            :param full_path_file_name: String defining the name of the LEDET input file to be written
            :type full_path_file_name: string
            :param verbose: flag that determines whether the output are printed
            :type verbose: bool

            :return: None
        '''

        # Unpack variables
        builder_ledet = self.builder_ledet

        # Define optional variables, which will be written only if present in the dataclass
        optional_variables_input_sheet = _getOptionalVariables_input()
        optional_variables_options_sheet = _getOptionalVariables_options()

        # Write LEDET data into a dictionary
        ledet_data_dict = {
            **builder_ledet.Inputs.__dict__,
            **builder_ledet.Options.__dict__,
            **builder_ledet.Plots.__dict__,
            **builder_ledet.Variables.__dict__}  # to add program-specific variables

        # Skip optional variables if they have 0 elements, or if they are all NaN
        for name, value in ledet_data_dict.copy().items():
            if (name in optional_variables_input_sheet) or (name in optional_variables_options_sheet):
                if isinstance(value, list) and len(value) == 0:
                    if verbose: print('Variable {} is optional and has 0 elements, hence it will be skipped.'.format(name))
                    del ledet_data_dict[name]
                elif isinstance(value, np.ndarray) and value.shape[0] == 0:
                    if verbose: print('Variable {} is optional and has 0 elements, hence it will be skipped.'.format(name))
                    del ledet_data_dict[name]
                elif value is None:
                    if verbose: print('Variable {} is optional and has 0 elements, hence it will be skipped.'.format(name))
                    del ledet_data_dict[name]
                elif isinstance(value, np.ndarray) and np.all(np.isnan(value)):
                    if verbose: print('Variable {} is optional and has only NaN elements, hence it will be skipped.'.format(name))
                    del ledet_data_dict[name]

        # Write output .json file
        with open(full_path_file_name, 'w') as outfile:
            json.dump(ledet_data_dict, outfile, cls=NumpyEncoder, indent=4)

        # Display time stamp
        currentDT = datetime.datetime.now()
        if verbose:
            print(' ')
            print('Time stamp: ' + str(currentDT))
            print('New file ' + full_path_file_name + ' generated.')
        return


    def write2yaml(self, full_path_file_name: str, verbose: bool = True):
        '''
            *** Function that writes a LEDET input file as a .yaml file ***

            Function to write a LEDET input file containing "Inputs", "Options", "Plots", and "Variables" variables

            :param full_path_file_name: String defining the name of the LEDET input file to be written
            :type full_path_file_name: string
            :param verbose: flag that determines whether the output are printed
            :type verbose: bool

            :return: None
        '''

        # Unpack variables
        builder_ledet = self.builder_ledet

        # Define optional variables, which will be written only if present in the dataclass
        optional_variables_input_sheet = _getOptionalVariables_input()
        optional_variables_options_sheet = _getOptionalVariables_options()

        # Write LEDET data into a dictionary
        ledet_data_dict = {
            **builder_ledet.Inputs.__dict__,
            **builder_ledet.Options.__dict__,
            **builder_ledet.Plots.__dict__,
            **builder_ledet.Variables.__dict__}  # to add program-specific variables

        # Skip optional variables if they have 0 elements, or if they are all NaN
        for name, value in ledet_data_dict.copy().items():
            if (name in optional_variables_input_sheet) or (name in optional_variables_options_sheet):
                if isinstance(value, list) and len(value) == 0:
                    if verbose: print('Variable {} is optional and has 0 elements, hence it will be skipped.'.format(name))
                    del ledet_data_dict[name]
                elif isinstance(value, np.ndarray) and value.shape[0] == 0:
                    if verbose: print('Variable {} is optional and has 0 elements, hence it will be skipped.'.format(name))
                    del ledet_data_dict[name]
                elif value is None:
                    if verbose: print('Variable {} is optional and has 0 elements, hence it will be skipped.'.format(name))
                    del ledet_data_dict[name]
                elif isinstance(value, np.ndarray) and np.all(np.isnan(value)):
                    if verbose: print('Variable {} is optional and has only NaN elements, hence it will be skipped.'.format(name))
                    del ledet_data_dict[name]

        # Write output .yaml file
        dict_to_yaml(ledet_data_dict, full_path_file_name, list_exceptions=[])

        # Display time stamp
        currentDT = datetime.datetime.now()
        if verbose:
            print(' ')
            print('Time stamp: ' + str(currentDT))
            print('New file ' + full_path_file_name + ' generated.')
        return


    def read_from_json(self, file_name: str, verbose: bool = True):
        '''
            *** Function that reads a .json file defining a LEDET input file and imports it in a BuilderLEDET object ***

            :param file_name: Name of the file to read
            :type file_name: str
            :param verbose: Flag indicating whether additional information should be displayed
            :type verbose: str

            :return: None
        '''

        # Read file as a dictionary
        with open(file_name) as json_file:
            dict_data = json.load(json_file)

        # Assign dictionary keys to the four LEDET dataclasses
        self._assign_dict_to_LEDET_dataclasses(dict_data)

        if verbose:
            print(f'File {file_name} successfully read.')


    def read_from_yaml(self, file_name: str, verbose: bool = True):
        '''
            *** Function that reads a .yaml file defining a LEDET input file and imports it in a BuilderLEDET object ***

            :param file_name: Name of the file to read
            :type file_name: str
            :param verbose: Flag indicating whether additional information should be displayed
            :type verbose: str

            :return: None
        '''

        # Read file as a dictionary
        yaml = ruamel.yaml.YAML(typ='safe', pure=True)
        with open(file_name) as yaml_file:
            dict_data = yaml.load(yaml_file)

        # Assign dictionary keys to the four LEDET dataclasses
        self._assign_dict_to_LEDET_dataclasses(dict_data)

        if verbose:
            print(f'File {file_name} successfully read.')


    #######################  Helper methods - START  #######################
    def _expand_scalar_to_array(self, verbose: bool = False):
        ''' Method adds as many elements as the number of half-turns to selected variables '''
        # TODO: Method can be made more general

        list_variables_to_expand = [
            'sim3D_f_cooling_down',
            'sim3D_f_cooling_up',
            'sim3D_f_cooling_left',
            'sim3D_f_cooling_right',
            ]

        for var_name in list_variables_to_expand:
            var_value = rgetattr(self.builder_ledet.Inputs, var_name)
            if isinstance(var_value, (int, float)):
                if verbose: print(f'{var_name} is a scalar. The same value will be written for all half-turns.')
            elif isinstance(var_value, (np.ndarray, list)) and len(var_value) == 1:
                if verbose: print(f'{var_name} is an array or a list of one element. The same value will be written for all half-turns.')
            else:
                continue  # do nothing
            # If the code reaches this line, assign the scalar value to all half-turns
            rsetattr(self.builder_ledet.Inputs, var_name, np.array([var_value] * int(np.sum(self.builder_ledet.Inputs.nT))))


    def _assign_dict_to_LEDET_dataclasses(self, dict_data: Dict):
        '''
        Assign the keys of a flat dictionary to the four LEDET dataclasses of a BuilderLEDET object
        :param dict_data:
        :return:
        '''
        for key, value in dict_data.items():
            if key in LEDETInputs.__annotations__:
                self.builder_ledet.setAttribute(self.builder_ledet.Inputs, key, value)
            if key in LEDETOptions.__annotations__:
                self.builder_ledet.setAttribute(self.builder_ledet.Options, key, value)
            if key in LEDETPlots.__annotations__:
                self.builder_ledet.setAttribute(self.builder_ledet.Plots, key, value)
            if key in LEDETVariables.__annotations__:
                self.builder_ledet.setAttribute(self.builder_ledet.Variables, key, value)

    #######################  Helper methods - END  #######################


#######################  Helper functions - START  #######################
def _getOptionalVariables_input():
    # Define optional variables, which will be written only if present in the dataclass
    optional_variables_input_sheet = [
        'dcore_inGroup',
        'dfilamentary_inGroup',
        'Jc_ref_NbTi_inGroup','C0_NbTi_inGroup','alpha_NbTi_inGroup', 'beta_NbTi_inGroup', 'gamma_NbTi_inGroup',
        'alpha_Nb3Sn_inGroup',
        'f_scaling_Jc_BSCCO2212_inGroup',
        'overwrite_f_internalVoids_inGroup', 'overwrite_f_externalVoids_inGroup',
        'f_RRR1_Cu_inGroup', 'f_RRR2_Cu_inGroup', 'f_RRR3_Cu_inGroup',
        'RRR1_Cu_inGroup', 'RRR2_Cu_inGroup', 'RRR3_Cu_inGroup',
        'R_EE_power', 'R_EE_initial_energy', 'R_EE_max_energy',
        ]
    return optional_variables_input_sheet


def _getOptionalVariables_options():
    optional_variables_options_sheet = [
        'fieldMapNumber',
        'selfMutualInductanceFileNumber',
        'flag_calculateMagneticField',
        'flag_controlInductiveVoltages',
        'flag_controlMagneticField',
        'flag_controlBoundaryTemperatures',
        'Jc_SC_max',
        'deltaB_PC'
        ]
    return optional_variables_options_sheet


def CompareLEDETParameters(fileA: str, fileB: str, max_relative_error: float=1E-5, verbose=False):
    '''
        Compare all the variables imported from two LEDET Excel input files
        Returns True if the two files contain LEDET parameters that differ by less than a certain relative error. Returns False otherwise.
    '''

    pl_a = ParserLEDET(BuilderLEDET(flag_build=False))
    pl_a.readFromExcel(fileA, verbose=False)
    pl_b = ParserLEDET(BuilderLEDET(flag_build=False))
    pl_b.readFromExcel(fileB, verbose=False)
    print("Starting Comparison of A: ({}) and B: ({})".format(fileA, fileB))

    Diff = check_for_differences(pl_a, pl_b, max_relative_error, verbose=verbose)

    if Diff == False:
        print(f'Files {fileA} and {fileB} are equal: they contain LEDET parameters that differ by a relative error lower than {max_relative_error}.')
        return True
    else:
        return False


def check_for_differences(pl_a, pl_b, max_relative_error, verbose):
    Diff = False

    ## Check Inputs
    for attribute in pl_a.builder_ledet.Inputs.__annotations__:
        last_diff = compare_two_parameters(pl_a.builder_ledet.getAttribute("Inputs", attribute),
                                      pl_b.builder_ledet.getAttribute("Inputs", attribute),
                                      attribute, max_relative_error, flag_difference_messages=True, verbose=verbose)
        if last_diff: Diff = True

    ## Check Options
    for attribute in pl_a.builder_ledet.Options.__annotations__:
        last_diff = compare_two_parameters(pl_a.builder_ledet.getAttribute("Options", attribute),
                                      pl_b.builder_ledet.getAttribute("Options", attribute),
                                      attribute, max_relative_error, flag_difference_messages=True, verbose=verbose)
        if last_diff: Diff = True

    ## Check Plots
    for attribute in pl_a.builder_ledet.Plots.__annotations__:
        last_diff = compare_two_parameters(pl_a.builder_ledet.getAttribute("Plots", attribute),
                                      pl_b.builder_ledet.getAttribute("Plots", attribute),
                                      attribute, max_relative_error, flag_difference_messages=True, verbose=verbose)
        if last_diff: Diff = True

    ## Check Variables
    for attribute in pl_a.builder_ledet.Variables.__annotations__:
        last_diff = compare_two_parameters(pl_a.builder_ledet.getAttribute("Variables", attribute),
                                      pl_b.builder_ledet.getAttribute("Variables", attribute),
                                      attribute, max_relative_error, flag_difference_messages=True, verbose=verbose)
        if last_diff: Diff = True

    return Diff



def copy_map2d(magnet_name: str, map2d_file_name: str, output_path: str, flagIron: int, flagSelfField: int,
               suffix: str = '', suffix_map2d_set: str = '', verbose: bool = False):
    ''' rename and copy '''

    # Checks if path exists
    if output_path != '' and not os.path.isdir(output_path):
        print("Output folder {} does not exist. Making it now".format(output_path))
        Path(output_path).mkdir(parents=True)

    ### new naming of file (depending on magnet properties)

    if flagIron == 1:
        suffix_iron = "_WithIron"
    else:
        suffix_iron = "_NoIron"

    if flagSelfField == 1:
        suffix_self = "_WithSelfField"
    else:
        suffix_self = "_NoSelfField"

    suffix_complete = suffix + suffix_iron + suffix_self + suffix_map2d_set
    file_name_output = magnet_name + suffix_complete + ".map2d"

    # Paths
    output_path_full = os.path.join(output_path, file_name_output)

    # Copy
    shutil.copy2(map2d_file_name, output_path_full)

    if verbose:
        print('File {} copied to {}.'.format(map2d_file_name, output_path_full))

    return file_name_output


def copy_modified_map2d_ribbon_cable(magnet_name: str, map2d_file_name: str, output_path: str, geometry_ribbon_cable,
                                     flagIron: int, flagSelfField: int, list_flag_ribbon: list, suffix: str = '',
                                     suffix_map2d_set: str = '', verbose: bool = False):
    ''' #TODO '''

    ### Check if output path exists

    if output_path != '' and not os.path.isdir(output_path):
        print("Output folder {} does not exist. Making it now".format(output_path))
        Path(output_path).mkdir(parents=True)

    ### naming of file (depending on magnet properties)

    if flagIron == 1:
        suffix_iron = "_WithIron"
    else:
        suffix_iron = "_NoIron"

    if flagSelfField == 1:
        suffix_self = "_WithSelfField"
    else:
        suffix_self = "_NoSelfField"

    suffix_complete = suffix + suffix_iron + suffix_self + suffix_map2d_set
    file_name_output = magnet_name + suffix_complete + ".map2d"

    # Parse to ParserMap2d
    NewMap2d = ParserMap2dFile(map2dFile=Path(map2d_file_name)).modify_map2d_ribbon_cable(
        geometry_ribbon_cable=geometry_ribbon_cable, list_flag_ribbon=list_flag_ribbon)
    # Multiply 3rd,4th and 7th column to not have SI-Units
    multiply_column_by_value(matrix=NewMap2d, column_number=3, multiplication_factor=1000)
    multiply_column_by_value(matrix=NewMap2d, column_number=4, multiplication_factor=1000)
    multiply_column_by_value(matrix=NewMap2d, column_number=7, multiplication_factor=1000000)
    # Create .map2d file from ParserMap2d_
    ParserMap2dData(map2d_input=NewMap2d, output_folder_path=Path(output_path),
                    physical_quantity='magnetic_flux_density').create_map2d_file_from_matrix(file_name=file_name_output)

    return file_name_output
#######################  Helper functions - END  #######################