"""
CLI interface for excel-to-sql using Typer.
"""

from pathlib import Path
from typer import Typer, Option, Exit
from rich.console import Console
from rich.table import Table
import pandas as pd

from excel_to_sql.entities.project import Project
from excel_to_sql.entities.excel_file import ExcelFile
from excel_to_sql.entities.dataframe import DataFrame
from excel_to_sql.__version__ import __version__

app = Typer(
    name="excel-to-sql",
    help="Excel to SQL CLI - Import Excel files to SQL and export back",
    no_args_is_help=True,
    add_completion=False,
)

console = Console()


# ──────────────────────────────────────────────────────────────
# Command: INIT
# ──────────────────────────────────────────────────────────────


@app.command()
def init(
    db_path: str = Option("data/excel-to-sql.db", "--db-path", help="Path to SQLite database"),
) -> None:
    """Initialize project structure and database."""
    console.print("[bold cyan]Initializing excel-to-sql project...[/bold cyan]")

    project = Project()
    project.initialize()

    console.print("[green]OK[/green] Project initialized successfully")
    console.print(f"  Location: {project.root}")
    console.print(f"  Database: {project.database.path}")
    console.print(f"  Imports:  {project.imports_dir}")
    console.print(f"  Exports:  {project.exports_dir}")


# ──────────────────────────────────────────────────────────────
# Command: IMPORT
# ──────────────────────────────────────────────────────────────


@app.command("import")
def import_cmd(
    excel_path: str = Option(..., "--file", "-f", help="Path to Excel file"),
    type: str = Option(..., "--type", "-t", help="Type configuration from mappings"),
    force: bool = Option(False, "--force", help="Re-import even if content unchanged"),
) -> None:
    """Import an Excel file into the database."""
    import sys
    import traceback

    path = Path(excel_path)

    try:
        # 1. Validate file exists
        if not path.exists():
            console.print(f"[red]Error:[/red] File not found: {excel_path}")
            raise Exit(1)

        # 2. Validate file extension
        if path.suffix.lower() not in {".xlsx", ".xls"}:
            console.print(f"[red]Error:[/red] Not an Excel file: {excel_path}")
            raise Exit(1)

        # 3. Load project
        try:
            project = Project.from_current_directory()
        except Exception:
            console.print("[red]Error:[/red] Not an excel-to-sql project")
            console.print("  Run 'excel-to-sql init' first")
            raise Exit(1)

        # 4. Validate type exists
        if type not in project.mappings:
            console.print(f"[red]Error:[/red] Unknown type: '{type}'")
            available_types = ", ".join(project.list_types())
            console.print(f"  Available types: {available_types}")
            raise Exit(1)

        mapping = project.mappings[type]

        # 5. Load Excel file
        console.print(f"[bold cyan]Reading {excel_path}...[/bold cyan]")

        excel_file = ExcelFile(path)

        # Validate file is readable
        if not excel_file.validate():
            console.print("[red]Error:[/red] Invalid or corrupted Excel file")
            raise Exit(1)

        # Get content hash
        content_hash = excel_file.content_hash
        console.print(f"  Content hash: {content_hash[:16]}...")

        # 6. Check if already imported
        if not force and project.database.is_imported(content_hash):
            console.print("[yellow]Already imported[/yellow]")
            console.print("  Use --force to re-import")

            # Show previous import details
            history = project.database.query(
                "SELECT * FROM _import_history WHERE content_hash = :hash",
                {"hash": content_hash}
            )
            if len(history) > 0:
                record = history.iloc[0]
                console.print(f"  Imported: {record['imported_at']}")
                console.print(f"  Rows: {record['rows_imported']}")
            raise Exit(0)

        # 7. Read data
        console.print("[dim]Loading data...[/dim]")
        raw_df = excel_file.read()
        initial_rows = len(raw_df)
        console.print(f"  Rows loaded: {initial_rows}")

        # 8. Clean data
        console.print("[dim]Cleaning data...[/dim]")
        df_wrapper = DataFrame(raw_df)
        df_wrapper.clean()
        cleaned_rows = len(df_wrapper.to_pandas())
        rows_removed = initial_rows - cleaned_rows
        console.print(f"  Rows removed (empty): {rows_removed}")

        # 9. Apply mapping
        console.print("[dim]Applying mappings...[/dim]")
        df_wrapper.apply_mapping(mapping)
        final_df = df_wrapper.to_pandas()
        final_rows = len(final_df)
        console.print(f"  Columns: {len(final_df.columns)}")
        console.print(f"  Rows: {final_rows}")

        # 10. Import to database
        console.print("[dim]Importing to database...[/dim]")
        table = project.database.get_table(mapping["target_table"])
        stats = table.upsert(final_df, primary_key=mapping["primary_key"])
        console.print(f"  Inserted: {stats['inserted']}")
        console.print(f"  Updated: {stats['updated']}")

        # 11. Record import in history
        # If using --force and content_hash exists, delete old record first
        if force and project.database.is_imported(content_hash):
            project.database.execute(
                "DELETE FROM _import_history WHERE content_hash = :hash",
                {"hash": content_hash}
            )

        project.database.record_import(
            file_name=excel_file.name,
            file_path=str(path.absolute()),
            content_hash=content_hash,
            file_type=type,
            rows_imported=stats["inserted"] + stats["updated"],
            rows_skipped=0,
            status="success",
        )

        # 12. Display summary
        console.print()
        console.print("[green]OK[/green] Import completed successfully")

        summary_table = Table(show_header=True, title="Import Summary")
        summary_table.add_column("Metric", style="cyan")
        summary_table.add_column("Value", style="green")

        summary_table.add_row("File", excel_file.name)
        summary_table.add_row("Type", type)
        summary_table.add_row("Table", mapping["target_table"])
        summary_table.add_row("Rows inserted", str(stats["inserted"]))
        summary_table.add_row("Rows updated", str(stats["updated"]))
        summary_table.add_row("Total rows", str(final_rows))
        summary_table.add_row("Content hash", content_hash[:16] + "...")

        console.print(summary_table)

    except FileNotFoundError:
        console.print(f"[red]Error:[/red] File not found: {excel_path}")
        raise Exit(1)

    except ValueError as e:
        console.print(f"[red]Error:[/red] {e}")
        raise Exit(1)

    except Exit:
        # Re-raise Exit exceptions (they're intentional)
        raise

    except Exception as e:
        console.print(f"[red]Error:[/red] Import failed")
        console.print(f"  {e}")
        if "--debug" in sys.argv:
            console.print(traceback.format_exc())
        raise Exit(1)


# ──────────────────────────────────────────────────────────────
# Command: EXPORT
# ──────────────────────────────────────────────────────────────


@app.command("export")
def export_cmd(
    output: str = Option(..., "--output", "-o", help="Output Excel file path"),
    table: str = Option(None, "--table", help="Export entire table"),
    query: str = Option(None, "--query", help="Custom SQL query"),
) -> None:
    """Export data from database to Excel."""
    # Validation
    if not table and not query:
        console.print("[red]Error:[/red] Must specify --table or --query")
        raise Exit(1)

    if table and query:
        console.print("[red]Error:[/red] Cannot specify both --table and --query")
        raise Exit(1)

    try:
        # Load project
        project = Project.from_current_directory()
    except Exception:
        console.print("[red]Error:[/red] Not an excel-to-sql project")
        console.print("[dim]Run 'excel-to-sql init' to initialize[/dim]")
        raise Exit(1)

    console.print(f"[bold cyan]Exporting to {output}...[/bold cyan]")

    try:
        # Execute export
        if table:
            # Export table
            console.print(f"  Table: {table}")
            try:
                df = project.database.export_table(table)
                source_desc = f"table '{table}'"
            except ValueError as e:
                console.print(f"[red]Error:[/red] {e}")
                raise Exit(1)
        else:
            # Export query
            console.print(f"  Query: {query[:50]}...")
            # Validate query starts with SELECT
            if not query.strip().upper().startswith("SELECT"):
                console.print("[red]Error:[/red] Query must start with SELECT")
                raise Exit(1)

            try:
                df = project.database.query(query)
                source_desc = "custom query"
            except Exception as e:
                console.print(f"[red]Error:[/red] Invalid SQL query")
                console.print(f"[dim]{e}[/dim]")
                raise Exit(1)

        # Check for empty results
        if len(df) == 0:
            console.print("[yellow]Warning:[/yellow] No data to export")
            raise Exit(0)

        # Create output directory if needed
        output_path = Path(output)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        # Write to Excel with formatting
        with pd.ExcelWriter(str(output_path), engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Sheet1')

            # Apply formatting
            worksheet = writer.sheets['Sheet1']

            # Bold headers
            for cell in worksheet[1]:
                cell.font = cell.font.copy(bold=True)

            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter

                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)  # Cap at 50
                worksheet.column_dimensions[column_letter].width = adjusted_width

            # Freeze header row
            worksheet.freeze_panes = 'A2'

        # Record export history
        project.database.record_export(
            table_name=table,
            query=query,
            output_path=str(output_path),
            row_count=len(df)
        )

        # Display summary
        console.print("")
        console.print("[bold green]OK[/bold green] Export completed successfully")
        console.print("")
        summary_table = Table(title="Export Summary")
        summary_table.add_column("Metric", style="cyan")
        summary_table.add_column("Value", style="green")

        summary_table.add_row("Source", source_desc)
        summary_table.add_row("Output", str(output_path))
        summary_table.add_row("Rows", str(len(df)))
        summary_table.add_row("Columns", str(len(df.columns)))

        # Get file size
        file_size = output_path.stat().st_size
        if file_size > 1024 * 1024:
            size_str = f"{file_size / (1024 * 1024):.2f} MB"
        elif file_size > 1024:
            size_str = f"{file_size / 1024:.2f} KB"
        else:
            size_str = f"{file_size} bytes"

        summary_table.add_row("File size", size_str)

        console.print(summary_table)

    except Exit:
        raise
    except Exception as e:
        console.print(f"[red]Error:[/red] Export failed")
        console.print(f"[dim]{e}[/dim]")
        raise Exit(1)


# ──────────────────────────────────────────────────────────────
# Command: STATUS
# ──────────────────────────────────────────────────────────────


@app.command()
def status() -> None:
    """Show import history."""
    try:
        # Load project
        project = Project.from_current_directory()
    except Exception:
        console.print("[red]Error:[/red] Not an excel-to-sql project")
        console.print("[dim]Run 'excel-to-sql init' to initialize[/dim]")
        raise Exit(1)

    # Get import history
    history = project.database.get_import_history()

    # Handle empty history
    if len(history) == 0:
        console.print("[dim]No imports yet[/dim]")
        console.print("")
        console.print("[dim]Run 'excel-to-sql import --file <file> --type <type>' to start importing[/dim]")
        return

    # Create Rich table for display
    table = Table(title="Import History")
    table.add_column("Date", style="cyan", no_wrap=False)
    table.add_column("File", style="green")
    table.add_column("Type", style="yellow")
    table.add_column("Rows", style="magenta", justify="right")
    table.add_column("Status", style="blue")

    # Add rows to table
    for _, row in history.iterrows():
        # Format date (remove seconds for cleaner display)
        date_str = str(row["imported_at"]).split(".")[0]  # Remove microseconds if present
        if "T" in date_str:
            date_str = date_str.replace("T", " ")

        # Style status based on value
        status = str(row["status"])
        status_style = "green" if status == "success" else "red"

        table.add_row(
            date_str,
            str(row["file_name"]),
            str(row["file_type"]),
            str(row["rows_imported"]),
            f"[{status_style}]{status}[/{status_style}]"
        )

    # Display table
    console.print("")
    console.print(table)

    # Display statistics
    total_imports = len(history)
    total_rows = history["rows_imported"].sum()
    total_skipped = history["rows_skipped"].sum()

    # Count successful imports
    successful = len(history[history["status"] == "success"])
    success_rate = (successful / total_imports * 100) if total_imports > 0 else 0

    # Get last import date
    last_import = history.iloc[0]["imported_at"]
    last_import_str = str(last_import).split(".")[0]
    if "T" in last_import_str:
        last_import_str = last_import_str.replace("T", " ")

    # Display statistics
    console.print("")
    console.print(f"[bold]Statistics:[/bold]")
    console.print(f"  Total imports: {total_imports}")
    console.print(f"  Total rows: {total_rows}")
    console.print(f"  Total skipped: {total_skipped}")
    console.print(f"  Success rate: {success_rate:.1f}%")
    console.print(f"  Last import: {last_import_str}")


# ──────────────────────────────────────────────────────────────
# Command: VERSION
# ──────────────────────────────────────────────────────────────


@app.command()
def version() -> None:
    """Show version information."""
    console.print(f"[bold cyan]excel-to-sql[/bold cyan] version [bold green]{__version__}[/bold green]")
    console.print("")
    console.print("[dim]GitHub:[/dim] https://github.com/wareflowx/excel-to-sql")
    console.print("[dim]PyPI:[/dim] https://pypi.org/project/excel-to-sql/")
    console.print("")
    console.print(f"[dim]Python: {__import__('sys').version.split()[0]}[/dim]")


# ──────────────────────────────────────────────────────────────
# Command: MAGIC (Auto-Pilot)
# ──────────────────────────────────────────────────────────────


@app.command()
def magic(
    data_path: str = Option(".", "--data", "-d", help="Path to directory containing Excel files"),
    output_path: str = Option(".excel-to-sql", "--output", "-o", help="Output directory for generated mappings"),
    dry_run: bool = Option(False, "--dry-run", help="Analyze files without generating configuration"),
    interactive: bool = Option(False, "--interactive", "-i", help="Interactive mode with guided configuration"),
) -> None:
    """Auto-pilot: Automatically detect patterns and generate configuration."""
    import sys
    from pathlib import Path
    from typing import Dict, Any, List
    from openpyxl import load_workbook
    from rich.panel import Panel
    from rich.progress import Progress, BarColumn, TextColumn
    from rich.text import Text
    from rich.table import Table
    from rich.live import Live
    from rich.align import Align

    # Import Auto-Pilot components
    try:
        from excel_to_sql.auto_pilot.detector import PatternDetector
        from excel_to_sql.auto_pilot.header_detector import HeaderDetector
        from excel_to_sql.auto_pilot.quality import QualityScorer
        from excel_to_sql.ui.interactive import InteractiveWizard
    except ImportError as e:
        console.print(f"[red]Error:[/red] {e}")
        console.print("[dim]This feature requires the auto_pilot module[/dim]")
        raise Exit(1)

    # Display styled header
    header_text = Text("AUTO-PILOT MODE", style="bold cyan")
    version_info = Text("Intelligent Excel to SQLite Configuration", style="dim")

    header = Panel(
        Text.assemble(
            header_text, "\n", version_info
        ),
        title="excel-to-sql",
        border_style="cyan",
        padding=(1, 2),
    )
    console.print(header)
    console.print("")

    # Initialize detectors
    detector = PatternDetector()
    header_detector = HeaderDetector()

    # Find Excel files
    data_dir = Path(data_path)
    if not data_dir.exists():
        console.print(f"[red]Error:[/red] Directory not found: {data_path}")
        raise Exit(1)

    excel_files = list(data_dir.glob("*.xlsx")) + list(data_dir.glob("*.xls"))

    if not excel_files:
        console.print(f"[yellow]Warning:[/yellow] No Excel files found in {data_path}")
        console.print("[dim]Place .xlsx or .xls files in the directory and try again[/dim]")
        raise Exit(1)

    # Display files found panel
    files_panel = Panel(
        f"Found [bold green]{len(excel_files)}[/bold green] Excel file(s) in [cyan]{data_path}[/cyan]",
        title="Files Discovered",
        border_style="green",
        padding=(0, 1),
    )
    console.print(files_panel)
    console.print("")

    # Process with progress bar
    all_results: Dict[str, Any] = {}

    with console.status("[bold]Analyzing Excel files...", spinner="dots") as status:
        for excel_file in excel_files:
            status.update(f"Processing {excel_file.name}...")

            try:
                # Load workbook to get sheet names
                wb = load_workbook(excel_file, read_only=True)
                sheet_names = wb.sheetnames
                wb.close()

                for sheet_name in sheet_names:
                    try:
                        # Read sheet with automatic header detection
                        df = header_detector.read_excel_with_header_detection(excel_file, sheet_name)
                        table_name = excel_file.stem.lower()

                        # Skip empty sheets
                        if len(df) == 0:
                            continue

                        # Detect patterns
                        patterns = detector.detect_patterns(df, table_name)
                        all_results[f"{excel_file.stem}/{sheet_name}"] = {
                            "file": str(excel_file),
                            "sheet": sheet_name,
                            "table_name": table_name,
                            "patterns": patterns,
                            "row_count": len(df),
                            "column_count": len(df.columns),
                        }

                    except Exception as e:
                        console.print(f"  [red]Error analyzing {sheet_name}:[/red] {e}")

            except Exception as e:
                console.print(f"[red]Error processing {excel_file.name}:[/red] {e}")

    # Interactive mode
    if interactive:
        console.print("")
        console.print("[bold cyan]Starting Interactive Mode...[/bold cyan]")
        console.print("")

        # Initialize quality scorer
        scorer = QualityScorer()

        # Prepare patterns and quality dictionaries for wizard
        patterns_dict: Dict[str, Dict[str, Any]] = {}
        quality_dict: Dict[str, Dict[str, Any]] = {}

        for key, result in all_results.items():
            table_name = result["table_name"]
            patterns_dict[table_name] = result["patterns"]

            # Generate quality report
            try:
                df = header_detector.read_excel_with_header_detection(result["file"], result["sheet"])
                quality_report = scorer.generate_quality_report(df, table_name)
                quality_dict[table_name] = quality_report
            except Exception:
                # Default quality report if analysis fails
                quality_dict[table_name] = {
                    "score": 100,
                    "grade": "A",
                    "issues": []
                }

        # Launch interactive wizard
        wizard = InteractiveWizard(console)
        wizard_result = wizard.run_interactive_mode(
            excel_files,
            patterns_dict,
            quality_dict,
            Path(output_path)
        )

        # Check if user wants to save configuration
        if wizard_result.get("action") == "save":
            # Generate configuration based on user choices
            output_dir = Path(output_path)
            output_dir.mkdir(parents=True, exist_ok=True)
            mappings_file = output_dir / "mappings.json"

            # Build mappings configuration from accepted files
            mappings_config: Dict[str, Any] = {"mappings": {}}

            for file_data in wizard_result.get("files_data", []):
                if file_data.get("skipped"):
                    continue

                file_path_str = file_data["file_path"]
                file_path = Path(file_path_str)
                table_name = file_path.stem

                # Get patterns for this file
                patterns = patterns_dict.get(table_name, {})

                # Build column mappings
                column_mappings = {}
                try:
                    df = header_detector.read_excel_with_header_detection(file_path)
                    for col in df.columns:
                        col_type = _infer_sql_type(df[col])
                        column_mappings[str(col)] = {
                            "target": str(col),
                            "type": col_type,
                            "required": False,
                            "default": None,
                        }
                except Exception:
                    pass

                # Build value mappings from accepted transformations
                value_mappings = []
                for trans in file_data.get("accepted_transformations", []):
                    if trans["type"] == "value_mapping":
                        value_mappings.append({
                            "column": trans["column"],
                            "mappings": trans.get("mappings", {}),
                        })

                # Build validation rules
                validation_rules = []
                pk = patterns.get("primary_key")
                if pk:
                    validation_rules.append({
                        "column": pk,
                        "type": "unique",
                        "params": {},
                        "message": f"{pk} must be unique",
                        "severity": "error",
                    })

                # Build reference validations
                reference_validations = []
                for fk in patterns.get("foreign_keys", []):
                    reference_validations.append({
                        "column": fk["column"],
                        "reference_table": fk["ref_table"],
                        "reference_column": fk.get("ref_column", "id"),
                    })

                # Build metadata
                metadata = {
                    "row_count": len(pd.read_excel(file_path)),
                    "column_count": len(pd.read_excel(file_path).columns),
                    "detection_confidence": patterns.get("confidence", 0.0),
                    "auto_generated": True,
                    "interactive_mode": True,
                    "primary_key_detected": pk is not None,
                    "has_value_mappings": len(value_mappings) > 0,
                    "has_foreign_keys": len(reference_validations) > 0,
                }

                # Build complete type mapping
                type_mapping = {
                    "target_table": table_name,
                    "primary_key": [pk] if pk else [],
                    "column_mappings": column_mappings,
                    "value_mappings": value_mappings,
                    "calculated_columns": [],
                    "validation_rules": validation_rules,
                    "reference_validations": reference_validations,
                    "hooks": [],
                    "tags": ["auto-generated", "interactive"],
                    "metadata": metadata,
                }

                mappings_config["mappings"][table_name] = type_mapping

            # Save to file
            import json
            with open(mappings_file, "w", encoding="utf-8") as f:
                json.dump(mappings_config, f, indent=2, ensure_ascii=False)

            # Display success message
            console.print("")
            console.print("[bold green]Configuration saved successfully![/bold green]")
            console.print(f"  Location: [cyan]{mappings_file}[/cyan]")

        return  # Exit interactive mode

    # Display results in file cards
    console.print("")
    console.print("[bold cyan]Detection Results[/bold cyan]")
    console.print("")

    for key, result in sorted(all_results.items()):
        patterns = result["patterns"]
        pk = patterns.get("primary_key")

        # Create file card
        file_info = Text.assemble(
            f"[bold cyan]{result['table_name'].title()}[/bold cyan]\n",
            f"File: [dim]{result['file']}[/dim]\n",
            f"Sheet: [cyan]{result['sheet']}[/cyan]\n",
            f"Rows: [green]{result['row_count']:,}[/green]   ",
            f"Cols: [blue]{result['column_count']}[/blue]\n"
        )

        if pk:
            file_info += f"PK: [bold green]{pk}[/bold green]\n"
        else:
            file_info += f"PK: [yellow]Not detected[/yellow]\n"

        if patterns.get("value_mappings"):
            mappings_text = ", ".join(patterns["value_mappings"].keys())
            file_info += f"Value Maps: [green]{mappings_text}[/green]\n"

        if patterns.get("foreign_keys"):
            fk_text = ", ".join([f"{fk['column']}->{fk['ref_table']}" for fk in patterns["foreign_keys"]])
            file_info += f"Foreign Keys: [cyan]{fk_text}[/cyan]\n"

        if patterns.get("split_fields"):
            split_text = ", ".join(patterns["split_fields"])
            file_info += f"Split Fields: [yellow]{split_text}[/yellow]\n"

        file_info += f"Confidence: [bold]{patterns['confidence']:.0%}[/bold]"

        file_card = Panel(
            file_info,
            border_style="blue",
            padding=(0, 2),
            title_align="left",
        )
        console.print(file_card)
        console.print("")

    # Display summary table
    if all_results:
        summary_title = Text.assemble(
            "[bold cyan]DETECTION SUMMARY[/bold cyan]\n",
            f"[dim]{len(all_results)} table(s) analyzed[/dim]"
        )
        console.print(Panel(summary_title, border_style="cyan", padding=(0, 1)))
        console.print("")

        table = Table(show_header=True, header_style="bold cyan", title_style="cyan", show_lines=True)
        table.add_column("Table", style="cyan", width=20)
        table.add_column("Rows", justify="right", style="green")
        table.add_column("Primary Key", style="bold green", width=15)
        table.add_column("Value Maps", justify="center", width=12)
        table.add_column("FKs", justify="center", width=8)
        table.add_column("Score", justify="right", width=8)

        for key, result in sorted(all_results.items()):
            patterns = result["patterns"]
            pk = patterns.get("primary_key") or "[dim]-[/dim]"
            value_maps = "[green]OK[/green]" if patterns.get("value_mappings") else "-"
            fks = str(len(patterns.get("foreign_keys", []))) if patterns.get("foreign_keys") else "-"
            score = f"[cyan]{patterns.get('confidence', 0):.0%}[/cyan]"

            table.add_row(
                result["table_name"],
                f"{result['row_count']:,}",
                pk,
                value_maps,
                fks,
                score,
            )

        console.print(table)
        console.print("")

    # Generate configuration
    if not dry_run and all_results:
        output_dir = Path(output_path)
        output_dir.mkdir(parents=True, exist_ok=True)

        mappings_file = output_dir / "mappings.json"

        # Show generation progress
        with console.status("[bold]Generating configuration...", spinner="dots2") as status:
            status.update("Building mappings structure...")

            # Build mappings configuration
            mappings_config: Dict[str, Any] = {"mappings": {}}

            for key, result in all_results.items():
                patterns = result["patterns"]
                table_name = result["table_name"]

                # Build column mappings
                column_mappings = {}
                try:
                    df = header_detector.read_excel_with_header_detection(result["file"], result["sheet"])
                    for col in df.columns:
                        col_type = _infer_sql_type(df[col])
                        column_mappings[str(col)] = {
                            "target": str(col),
                            "type": col_type,
                            "required": False,
                            "default": None,
                        }
                except Exception:
                    pass

                # Build value mappings
                value_mappings = []
                for col, mappings in patterns.get("value_mappings", {}).items():
                    value_mappings.append({
                        "column": col,
                        "mappings": mappings,
                    })

                # Build validation rules
                validation_rules = []
                pk = patterns.get("primary_key")
                if pk:
                    validation_rules.append({
                        "column": pk,
                        "type": "unique",
                        "params": {},
                        "message": f"{pk} must be unique",
                        "severity": "error",
                    })

                # Build reference validations
                reference_validations = []
                for fk in patterns.get("foreign_keys", []):
                    reference_validations.append({
                        "column": fk["column"],
                        "reference_table": fk["ref_table"],
                        "reference_column": fk.get("ref_column", "id"),
                    })

                # Build metadata
                metadata = {
                    "row_count": result["row_count"],
                    "column_count": result["column_count"],
                    "detection_confidence": patterns.get("confidence", 0.0),
                    "auto_generated": True,
                    "primary_key_detected": pk is not None,
                    "has_value_mappings": len(value_mappings) > 0,
                    "has_foreign_keys": len(reference_validations) > 0,
                    "has_split_fields": patterns.get("split_fields") is not None,
                }

                # Build complete type mapping
                type_mapping = {
                    "target_table": table_name,
                    "primary_key": [pk] if pk else [],
                    "column_mappings": column_mappings,
                    "value_mappings": value_mappings,
                    "calculated_columns": [],
                    "validation_rules": validation_rules,
                    "reference_validations": reference_validations,
                    "hooks": [],
                    "tags": ["auto-generated"],
                    "metadata": metadata,
                }

                mappings_config["mappings"][table_name] = type_mapping

            status.update(f"Saving to {mappings_file}...")

            # Save to file
            import json
            with open(mappings_file, "w", encoding="utf-8") as f:
                json.dump(mappings_config, f, indent=2, ensure_ascii=False)

        # Display success panel
        success_panel = Panel(
            Text.assemble(
                "[bold green]Configuration generated successfully![/bold green]\n\n",
                f"Location: [cyan]{mappings_file}[/cyan]\n",
                f"Tables: [green]{len(mappings_config['mappings'])}[/green]\n",
                f"Total Rows: [green]{sum(r['row_count'] for r in all_results.values()):,}[/green]\n"
            ),
            title="OK SUCCESS",
            border_style="green",
            padding=(1, 2),
        )
        console.print(success_panel)
        console.print("")
        console.print("[dim]Next steps:[/dim]")
        console.print("  1. Review the generated configuration")
        console.print("  2. Run: [cyan]excel-to-sql import --file <file> --type <table>[/cyan]")
        console.print("")

    elif dry_run:
        dry_panel = Panel(
            "[bold yellow]Dry-Run Mode[/bold yellow]\n\nConfiguration was analyzed but not generated.\n\nOmit [cyan]--dry-run[/cyan] to generate configuration files.",
            title="!",
            border_style="yellow",
            padding=(1, 2),
        )
        console.print(dry_panel)
        console.print("")

    # Final completion message
    console.print("[bold green]Auto-Pilot Analysis Complete![/bold green]")


def _infer_sql_type(series) -> str:
    """Infer SQL type from pandas Series."""
    import pandas as pd

    dtype = series.dtype

    if pd.api.types.is_datetime64_any_dtype(dtype):
        return "date"
    elif pd.api.types.is_integer_dtype(dtype):
        return "integer"
    elif pd.api.types.is_float_dtype(dtype):
        return "float"
    elif pd.api.types.is_bool_dtype(dtype):
        return "boolean"
    else:
        return "string"


# ──────────────────────────────────────────────────────────────
# Command: CONFIG
# ──────────────────────────────────────────────────────────────


@app.command()
def config(
    add_type: str = Option(None, "--add-type", help="Add new type configuration"),
    table: str = Option(None, "--table", help="Target table name"),
    pk: str = Option(None, "--pk", help="Primary key column(s), comma-separated for composite"),
    file: str = Option(None, "--file", help="Excel file to auto-detect columns (optional with --add-type)"),
    list_all: bool = Option(False, "--list", help="List all mappings"),
    show: str = Option(None, "--show", help="Show specific mapping details"),
    remove: str = Option(None, "--remove", help="Remove mapping"),
    validate: bool = Option(False, "--validate", help="Validate all mappings"),
) -> None:
    """Manage configuration mappings."""
    try:
        # Load project
        project = Project.from_current_directory()
    except Exception:
        console.print("[red]Error:[/red] Not an excel-to-sql project")
        console.print("[dim]Run 'excel-to-sql init' to initialize[/dim]")
        raise Exit(1)

    # Route to appropriate subcommand
    if add_type:
        _config_add_type(project, add_type, table, pk, file)
    elif list_all:
        _config_list(project)
    elif show:
        _config_show(project, show)
    elif remove:
        _config_remove(project, remove)
    elif validate:
        _config_validate(project)
    else:
        # Show help by default
        console.print("[yellow]Usage:[/yellow]")
        console.print("  excel-to-sql config --add-type <name> --table <table> --pk <columns> [--file <excel>]")
        console.print("  excel-to-sql config --list")
        console.print("  excel-to-sql config --show <type>")
        console.print("  excel-to-sql config --remove <type>")
        console.print("  excel-to-sql config --validate")


def _config_add_type(project, add_type: str, table: str, pk: str, file: str) -> None:
    """Add a new type mapping."""
    # Validate required parameters
    if not table:
        console.print("[red]Error:[/red] --table is required when using --add-type")
        raise Exit(1)

    if not pk:
        console.print("[red]Error:[/red] --pk is required when using --add-type")
        raise Exit(1)

    # Check if type already exists
    if add_type in project.list_types():
        console.print(f"[red]Error:[/red] Type '{add_type}' already exists")
        console.print("[dim]Use --show {add_type} to view, or --remove {add_type} to delete[/dim]")
        raise Exit(1)

    # Parse primary key (comma-separated for composite)
    primary_key = [col.strip() for col in pk.split(",")]

    # Build column mappings
    if file:
        # Auto-detect from Excel file
        console.print(f"[bold cyan]Auto-detecting columns from {file}...[/bold cyan]")

        try:
            detected_columns = project.auto_detect_columns(file)
            column_mappings = {}
            for col_name, col_type in detected_columns.items():
                column_mappings[col_name] = {
                    "target": col_name.lower().replace(" ", "_"),
                    "type": col_type
                }

            console.print(f"  Detected {len(column_mappings)} columns")
        except Exception as e:
            console.print(f"[red]Error:[/red] Failed to read Excel file")
            console.print(f"[dim]{e}[/dim]")
            raise Exit(1)
    else:
        # Create empty mapping (user will edit manually)
        column_mappings = {}
        console.print("[yellow]Note:[/yellow] No --file specified. You'll need to edit config/mappings.json to add column mappings.")

    # Create the mapping
    project.add_mapping(
        type_name=add_type,
        table_name=table,
        primary_key=primary_key,
        column_mappings=column_mappings
    )

    console.print("")
    console.print(f"[bold green]OK[/bold green] Created mapping for type '{add_type}'")
    console.print("")
    console.print(f"  Table: {table}")
    console.print(f"  Primary Key: {', '.join(primary_key)}")
    console.print(f"  Columns: {len(column_mappings)}")
    console.print("")
    if file:
        console.print("[dim]Edit config/mappings.json to adjust column types and targets[/dim]")
    else:
        console.print("[dim]Edit config/mappings.json to add column mappings[/dim]")


def _config_list(project) -> None:
    """List all mappings."""
    types = project.list_types()

    # Filter out internal types (starting with _)
    user_types = [t for t in types if not t.startswith("_")]

    if len(user_types) == 0:
        console.print("[dim]No mappings configured[/dim]")
        console.print("[dim]Use --add-type to create one[/dim]")
        return

    # Create table
    table = Table(title="Configured Mappings")
    table.add_column("Type", style="cyan")
    table.add_column("Table", style="green")
    table.add_column("Primary Key", style="yellow")
    table.add_column("Columns", style="magenta", justify="right")

    for type_name in sorted(user_types):
        mapping = project.get_mapping(type_name)

        if mapping:
            target_table = mapping.get("target_table", "N/A")
            primary_key = mapping.get("primary_key", [])
            column_mappings = mapping.get("column_mappings", {})

            pk_str = ", ".join(primary_key) if isinstance(primary_key, list) else str(primary_key)
            col_count = len(column_mappings)

            table.add_row(type_name, target_table, pk_str, str(col_count))

    console.print("")
    console.print(table)
    console.print("")
    console.print(f"Total: {len(user_types)} mapping(s)")


def _config_show(project, type_name: str) -> None:
    """Show details for a specific mapping."""
    mapping = project.get_mapping(type_name)

    if not mapping:
        console.print(f"[red]Error:[/red] Type '{type_name}' not found")
        console.print("[dim]Use --list to see all configured types[/dim]")
        raise Exit(1)

    console.print("")
    console.print(f"[bold cyan]Mapping: {type_name}[/bold cyan]")
    console.print("")

    # Basic info
    console.print(f"[bold]Target Table:[/bold] {mapping.get('target_table', 'N/A')}")
    console.print(f"[bold]Primary Key:[/bold] {', '.join(mapping.get('primary_key', []))}")
    console.print("")

    # Column mappings
    column_mappings = mapping.get("column_mappings", {})

    if len(column_mappings) > 0:
        table = Table(title="Column Mappings")
        table.add_column("Source", style="cyan")
        table.add_column("Target", style="green")
        table.add_column("Type", style="yellow")
        table.add_column("Required", style="magenta")

        for source, config in column_mappings.items():
            target = config.get("target", "N/A")
            col_type = config.get("type", "string")
            required = config.get("required", False)
            required_str = "[green]Yes[/green]" if required else "[dim]No[/dim]"

            table.add_row(source, target, col_type, required_str)

        console.print(table)
    else:
        console.print("[yellow]No column mappings defined[/yellow]")
        console.print("[dim]Edit config/mappings.json to add column mappings[/dim]")


def _config_remove(project, type_name: str) -> None:
    """Remove a mapping."""
    # Confirm removal
    console.print(f"[yellow]Remove mapping '{type_name}'?[/yellow]")
    console.print("[dim]This will delete the mapping configuration.[/dim]")

    # For now, just remove it (in a real CLI, you'd ask for confirmation)
    # Since this is non-interactive, we'll just do it
    if project.remove_mapping(type_name):
        console.print("")
        console.print(f"[bold green]OK[/bold green] Removed mapping '{type_name}'")
    else:
        console.print("")
        console.print(f"[red]Error:[/red] Type '{type_name}' not found")
        raise Exit(1)


def _config_validate(project) -> None:
    """Validate all mappings."""
    console.print("[bold cyan]Validating mappings...[/bold cyan]")
    console.print("")

    errors = project.validate_mappings()
    mappings = project.mappings

    # Count user mappings (excluding internal)
    user_mappings = [t for t in mappings.keys() if not t.startswith("_")]
    total = len(user_mappings)

    if len(errors) == 0:
        console.print(f"[bold green]OK[/bold green] All {total} mapping(s) are valid")
    else:
        console.print(f"[bold red]X[/bold red] Found {len(errors)} error(s)")
        console.print("")

        # Create error table
        table = Table()
        table.add_column("Type", style="cyan")
        table.add_column("Error", style="red")

        for error in errors:
            table.add_row(error["type"], error["error"])

        console.print(table)


# ──────────────────────────────────────────────────────────────
# MAIN ENTRY POINT
# ──────────────────────────────────────────────────────────────

if __name__ == "__main__":
    app()
