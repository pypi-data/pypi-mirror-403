"""
Auto-Fix Engine for Auto-Pilot Mode.

Automatically corrects common data quality issues with user confirmation.
"""

from typing import Dict, List, Any, Optional
from pathlib import Path
from datetime import datetime
import shutil
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


class AutoFixer:
    """
    Automatically corrects common data quality issues.

    Applies safe, reversible fixes to data quality issues identified
    by the Recommendations Engine. Creates backups before modification.
    """

    # Maximum number of backups to keep per file
    MAX_BACKUPS = 5

    # French to English code mappings for common WMS systems
    FRENCH_CODE_MAPPINGS = {
        "ENTRÉE": "inbound",
        "ENTREE": "inbound",
        "SORTIE": "outbound",
        "TRANSFERT": "transfer",
        "AJUSTEMENT": "adjustment",
        "INVENTAIRE": "inventory",
        "ACTIF": "active",
        "INACTIF": "inactive",
        "EN_ATTENTE": "pending",
        "ANNULÉ": "cancelled",
        "ANNULE": "cancelled",
        "VALIDÉ": "validated",
        "VALIDE": "validated",
    }

    def __init__(self, backup_dir: Optional[Path] = None) -> None:
        """
        Initialize the AutoFixer.

        Args:
            backup_dir: Directory for backups. Defaults to .excel-to-sql/backups/
        """
        self.backup_dir = backup_dir or Path(".excel-to-sql/backups")
        self.fixes_applied: List[Dict[str, Any]] = []

    def apply_auto_fixes(
        self,
        df: pd.DataFrame,
        file_path: Path,
        sheet_name: str,
        recommendations: List[Dict[str, Any]],
        dry_run: bool = False
    ) -> Dict[str, Any]:
        """
        Apply all auto-fixable recommendations.

        Args:
            df: DataFrame to fix
            file_path: Path to Excel file
            sheet_name: Name of sheet
            recommendations: List of recommendations from RecommendationEngine
            dry_run: If True, preview changes without applying

        Returns:
            Dictionary with fix results and statistics
        """
        self.fixes_applied = []
        result = {
            "fixes_applied": [],
            "backup_path": None,
            "rows_modified": 0,
            "total_fixes": 0,
            "dry_run": dry_run,
            "status": "success"
        }

        # Filter for auto-fixable recommendations only
        auto_fixable = [r for r in recommendations if r.get("auto_fix", False)]

        if not auto_fixable:
            result["message"] = "No auto-fixable issues found"
            return result

        # Create backup unless dry-run
        if not dry_run:
            backup_path = self._create_backup(file_path)
            result["backup_path"] = str(backup_path)

        # Apply each auto-fix
        df_fixed = df.copy()

        for recommendation in auto_fixable:
            fix_result = self._apply_fix(
                df_fixed,
                recommendation,
                dry_run
            )

            if fix_result:
                self.fixes_applied.append(fix_result)
                result["fixes_applied"].append(fix_result)
                result["rows_modified"] += fix_result.get("rows_affected", 0)

                # Update DataFrame with fixes if not dry-run
                if not dry_run and fix_result.get("df_fixed") is not None:
                    df_fixed = fix_result["df_fixed"]

        # Save fixed file unless dry-run
        if not dry_run and result["rows_modified"] > 0:
            self._save_fixed_file(df_fixed, file_path, sheet_name)

        result["total_fixes"] = len(self.fixes_applied)
        result["status"] = "success"

        return result

    def _apply_fix(
        self,
        df: pd.DataFrame,
        recommendation: Dict[str, Any],
        dry_run: bool
    ) -> Optional[Dict[str, Any]]:
        """
        Apply a single fix based on recommendation type.

        Args:
            df: DataFrame to fix
            recommendation: Recommendation dictionary
            dry_run: If True, preview only

        Returns:
            Fix result dictionary or None
        """
        issue_type = recommendation.get("issue_type", "")
        column = recommendation.get("column")

        if issue_type == "null_values":
            return self._fix_null_values(df, column, recommendation, dry_run)
        elif issue_type == "missing_default":
            return self._fix_missing_default(df, column, recommendation, dry_run)
        elif issue_type == "future_dates":
            # Not auto-fixable
            return None
        elif issue_type == "duplicate_pk":
            # Not auto-fixable
            return None
        elif issue_type == "invalid_values":
            # Not auto-fixable
            return None

        return None

    def _fix_null_values(
        self,
        df: pd.DataFrame,
        column: str,
        recommendation: Dict[str, Any],
        dry_run: bool
    ) -> Dict[str, Any]:
        """Fix null values by applying default value."""
        if column not in df.columns:
            return None

        default_value = recommendation.get("suggested_default", "NULL")
        null_count = df[column].isna().sum()

        if null_count == 0:
            return None

        df_fixed = df.copy()

        # Apply default value
        if default_value == "NULL":
            # Keep as null (or could set to a special value)
            pass
        elif default_value == "0":
            df_fixed[column] = df_fixed[column].fillna(0)
        elif default_value == "CURRENT_TIMESTAMP":
            df_fixed[column] = df_fixed[column].fillna(pd.Timestamp.now())
        else:
            # String default
            df_fixed[column] = df_fixed[column].fillna(default_value)

        return {
            "type": "null_values",
            "column": column,
            "default_applied": default_value,
            "rows_affected": int(null_count),
            "status": "applied" if not dry_run else "preview",
            "df_fixed": df_fixed if not dry_run else None
        }

    def _fix_missing_default(
        self,
        df: pd.DataFrame,
        column: str,
        recommendation: Dict[str, Any],
        dry_run: bool
    ) -> Dict[str, Any]:
        """Fix missing default values."""
        if column not in df.columns:
            return None

        default_value = recommendation.get("suggested_default", "NULL")

        # Count rows that would benefit (nulls)
        null_count = df[column].isna().sum()

        if null_count == 0:
            return None

        df_fixed = df.copy()

        # Apply default value
        if default_value == "0":
            df_fixed[column] = df_fixed[column].fillna(0)
        elif default_value == "CURRENT_TIMESTAMP":
            df_fixed[column] = df_fixed[column].fillna(pd.Timestamp.now())
        else:
            # String default
            df_fixed[column] = df_fixed[column].fillna(default_value)

        return {
            "type": "missing_default",
            "column": column,
            "default_applied": default_value,
            "rows_affected": int(null_count),
            "status": "applied" if not dry_run else "preview",
            "df_fixed": df_fixed if not dry_run else None
        }

    def _fix_french_codes(
        self,
        df: pd.DataFrame,
        column: str,
        dry_run: bool
    ) -> Optional[Dict[str, Any]]:
        """
        Fix French codes by replacing with English equivalents.

        This is called when value mappings are detected in patterns.
        """
        if column not in df.columns:
            return None

        df_fixed = df.copy()
        rows_affected = 0

        # Count replacements
        for french, english in self.FRENCH_CODE_MAPPINGS.items():
            count = (df_fixed[column] == french).sum()
            if count > 0:
                df_fixed[column] = df_fixed[column].replace(french, english)
                rows_affected += count

        if rows_affected == 0:
            return None

        return {
            "type": "french_codes",
            "column": column,
            "rows_affected": rows_affected,
            "status": "applied" if not dry_run else "preview",
            "df_fixed": df_fixed if not dry_run else None
        }

    def _fix_split_fields(
        self,
        df: pd.DataFrame,
        split_columns: List[str],
        target_column: str,
        dry_run: bool
    ) -> Optional[Dict[str, Any]]:
        """
        Fix split fields by combining with COALESCE.

        Creates a new column that takes the first non-null value
        from the split columns.
        """
        if not split_columns or target_column in df.columns:
            return None

        # Check if split columns exist
        existing_cols = [col for col in split_columns if col in df.columns]
        if not existing_cols:
            return None

        df_fixed = df.copy()

        # Create COALESCE column
        df_fixed[target_column] = df_fixed[existing_cols[0]]
        for col in existing_cols[1:]:
            df_fixed[target_column] = df_fixed[target_column].fillna(df_fixed[col])

        rows_affected = len(df_fixed)

        return {
            "type": "split_fields",
            "column": target_column,
            "source_columns": existing_cols,
            "rows_affected": rows_affected,
            "status": "applied" if not dry_run else "preview",
            "df_fixed": df_fixed if not dry_run else None
        }

    def _create_backup(self, file_path: Path) -> Path:
        """
        Create backup of Excel file before modification.

        Args:
            file_path: Path to file to backup

        Returns:
            Path to backup file
        """
        # Create backup directory
        self.backup_dir.mkdir(parents=True, exist_ok=True)

        # Generate backup filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_name = f"{file_path.stem}_{timestamp}.xlsx.bak"
        backup_path = self.backup_dir / backup_name

        # Copy file to backup location
        shutil.copy2(file_path, backup_path)

        # Clean up old backups (keep only MAX_BACKUPS)
        self._cleanup_old_backups(file_path.stem)

        return backup_path

    def _cleanup_old_backups(self, file_stem: str) -> None:
        """
        Remove old backups, keeping only MAX_BACKUPS most recent.

        Args:
            file_stem: Stem of the filename (without extension)
        """
        # Find all backups for this file
        pattern = f"{file_stem}_*.xlsx.bak"
        backups = sorted(self.backup_dir.glob(pattern), reverse=True)

        # Remove old backups beyond MAX_BACKUPS
        for old_backup in backups[self.MAX_BACKUPS:]:
            old_backup.unlink()

    def _save_fixed_file(
        self,
        df: pd.DataFrame,
        file_path: Path,
        sheet_name: str
    ) -> None:
        """
        Save fixed DataFrame back to Excel file.

        Args:
            df: Fixed DataFrame
            file_path: Original file path
            sheet_name: Sheet name to update
        """
        # Load workbook to preserve formatting
        wb = load_workbook(file_path)

        # Remove old sheet if exists
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

        # Create new sheet with data
        ws = wb.create_sheet(sheet_name)

        # Write data
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # Save
        wb.save(file_path)
        wb.close()

    def restore_backup(self, backup_path: Path, original_path: Path) -> bool:
        """
        Restore a file from backup.

        Args:
            backup_path: Path to backup file
            original_path: Path where to restore

        Returns:
            True if restore successful, False otherwise
        """
        try:
            if not backup_path.exists():
                return False

            # Copy backup to original location
            shutil.copy2(backup_path, original_path)
            return True
        except Exception:
            return False

    def get_backups(self, file_stem: str) -> List[Path]:
        """
        Get list of backups for a file.

        Args:
            file_stem: Stem of filename

        Returns:
            List of backup paths sorted by date (newest first)
        """
        pattern = f"{file_stem}_*.xlsx.bak"
        backups = sorted(self.backup_dir.glob(pattern), reverse=True)
        return backups[:self.MAX_BACKUPS]

    def get_fix_summary(self) -> Dict[str, Any]:
        """
        Get summary of fixes applied.

        Returns:
            Dictionary with fix statistics
        """
        return {
            "total_fixes": len(self.fixes_applied),
            "rows_modified": sum(f.get("rows_affected", 0) for f in self.fixes_applied),
            "fixes_by_type": self._group_fixes_by_type(),
            "fixes_by_column": self._group_fixes_by_column()
        }

    def _group_fixes_by_type(self) -> Dict[str, int]:
        """Group fixes by type."""
        groups = {}
        for fix in self.fixes_applied:
            fix_type = fix.get("type", "unknown")
            groups[fix_type] = groups.get(fix_type, 0) + 1
        return groups

    def _group_fixes_by_column(self) -> Dict[str, int]:
        """Group fixes by column."""
        groups = {}
        for fix in self.fixes_applied:
            column = fix.get("column", "unknown")
            groups[column] = groups.get(column, 0) + 1
        return groups
