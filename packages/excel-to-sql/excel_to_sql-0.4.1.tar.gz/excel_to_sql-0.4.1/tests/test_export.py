"""
Tests for export command.
"""

import pytest
from typer.testing import CliRunner
from pathlib import Path
import tempfile
import shutil
import gc
import pandas as pd

from excel_to_sql.cli import app
from excel_to_sql.entities.project import Project


def rmtree_with_retry(path):
    """Remove directory tree with retry for Windows file locks."""
    for _ in range(5):
        try:
            shutil.rmtree(path, ignore_errors=True)
            return
        except PermissionError:
            gc.collect()
            import time
            time.sleep(0.1)
    try:
        shutil.rmtree(path, ignore_errors=True)
    except:
        pass


class TestExportCommand:
    """Tests for export command."""

    @pytest.fixture
    def runner(self):
        """Create a CLI runner for testing."""
        return CliRunner()

    @pytest.fixture
    def temp_project(self, monkeypatch):
        """Create a temporary project with data to export."""
        import os

        temp = Path(tempfile.mkdtemp())

        # Add a .git marker to make it a valid project root
        (temp / ".git").mkdir()

        # Mock Project.from_current_directory to return our temp project
        def mock_from_current_directory():
            return Project(root=temp)

        monkeypatch.setattr(
            "excel_to_sql.entities.project.Project.from_current_directory",
            mock_from_current_directory
        )

        # Initialize project
        project = Project(root=temp)
        project.initialize()

        # Add products mapping
        project.add_mapping(
            type_name="products",
            table_name="products",
            primary_key=["id"],
            column_mappings={
                "ID": {"target": "id", "type": "integer"},
                "Name": {"target": "name", "type": "string"},
                "Price": {"target": "price", "type": "float"},
            },
        )

        # Create imports directory
        project.imports_dir.mkdir(parents=True, exist_ok=True)

        # Create and import data
        data = pd.DataFrame(
            {
                "ID": [1, 2, 3],
                "Name": ["Product A", "Product B", "Product C"],
                "Price": [10.50, 20.00, 30.99],
            }
        )

        excel_path = project.imports_dir / "products.xlsx"
        data.to_excel(excel_path, index=False)

        # Import the data
        runner = CliRunner()
        result = runner.invoke(
            app, ["import", "--file", str(excel_path), "--type", "products"]
        )

        assert result.exit_code == 0

        yield project

        # Cleanup
        project.database.dispose()
        gc.collect()
        rmtree_with_retry(temp)

    def test_export_table_success(self, runner, temp_project, tmp_path):
        """Test exporting a table successfully."""
        output_file = tmp_path / "export.xlsx"

        result = runner.invoke(
            app, ["export", "--table", "products", "--output", str(output_file)]
        )

        assert result.exit_code == 0
        assert "Export completed successfully" in result.stdout
        assert "Export Summary" in result.stdout

        # Verify file was created
        assert output_file.exists()

        # Verify data
        df = pd.read_excel(output_file)
        assert len(df) == 3
        assert "name" in df.columns
        assert df["name"].tolist() == ["Product A", "Product B", "Product C"]

    def test_export_query_success(self, runner, temp_project, tmp_path):
        """Test exporting a custom query successfully."""
        output_file = tmp_path / "export.xlsx"

        result = runner.invoke(
            app,
            [
                "export",
                "--query",
                "SELECT * FROM products WHERE price > 15",
                "--output",
                str(output_file)
            ]
        )

        assert result.exit_code == 0
        assert "Export completed successfully" in result.stdout

        # Verify file was created
        assert output_file.exists()

        # Verify data (only products with price > 15)
        df = pd.read_excel(output_file)
        assert len(df) == 2  # Product B (20.00) and Product C (30.99)

    def test_export_table_not_found(self, runner, temp_project, tmp_path):
        """Test exporting a non-existent table."""
        output_file = tmp_path / "export.xlsx"

        result = runner.invoke(
            app, ["export", "--table", "nonexistent", "--output", str(output_file)]
        )

        assert result.exit_code == 1
        assert "does not exist" in result.stdout

    def test_export_missing_both_table_and_query(self, runner, temp_project, tmp_path):
        """Test export without --table or --query."""
        output_file = tmp_path / "export.xlsx"

        result = runner.invoke(
            app, ["export", "--output", str(output_file)]
        )

        assert result.exit_code == 1
        assert "Must specify --table or --query" in result.stdout

    def test_export_both_table_and_query(self, runner, temp_project, tmp_path):
        """Test export with both --table and --query (should fail)."""
        output_file = tmp_path / "export.xlsx"

        result = runner.invoke(
            app,
            [
                "export",
                "--table", "products",
                "--query", "SELECT * FROM products",
                "--output", str(output_file)
            ]
        )

        assert result.exit_code == 1
        assert "Cannot specify both --table and --query" in result.stdout

    def test_export_invalid_query(self, runner, temp_project, tmp_path):
        """Test export with invalid SQL query."""
        output_file = tmp_path / "export.xlsx"

        result = runner.invoke(
            app,
            [
                "export",
                "--query",
                "INVALID SQL QUERY",
                "--output",
                str(output_file)
            ]
        )

        assert result.exit_code == 1
        assert "must start with SELECT" in result.stdout

    def test_export_not_a_project(self, runner, tmp_path):
        """Test export when not in an excel-to-sql project."""
        empty_dir = tmp_path / "empty_project"
        empty_dir.mkdir()

        # Mock from_current_directory to raise an exception
        def mock_error():
            raise ValueError("Not a project")

        import excel_to_sql.cli
        original_from_current = excel_to_sql.cli.Project.from_current_directory
        excel_to_sql.cli.Project.from_current_directory = mock_error

        try:
            output_file = tmp_path / "export.xlsx"
            result = runner.invoke(
                app, ["export", "--table", "products", "--output", str(output_file)]
            )

            assert result.exit_code == 1
            assert "Not an excel-to-sql project" in result.stdout
        finally:
            excel_to_sql.cli.Project.from_current_directory = original_from_current

    def test_export_creates_output_directory(self, runner, temp_project, tmp_path):
        """Test that export creates output directory if it doesn't exist."""
        # Create a path with non-existent subdirectories
        output_file = tmp_path / "deep" / "nested" / "export.xlsx"

        result = runner.invoke(
            app, ["export", "--table", "products", "--output", str(output_file)]
        )

        assert result.exit_code == 0
        assert output_file.exists()

    def test_export_empty_table(self, runner, temp_project, tmp_path):
        """Test exporting a table with no data."""
        # Create an empty table
        temp_project.database.execute(
            "CREATE TABLE empty_table (id INTEGER, name TEXT)"
        )

        output_file = tmp_path / "export.xlsx"

        result = runner.invoke(
            app, ["export", "--table", "empty_table", "--output", str(output_file)]
        )

        assert result.exit_code == 0
        assert "No data to export" in result.stdout

    def test_export_applies_formatting(self, runner, temp_project, tmp_path):
        """Test that export applies Excel formatting."""
        output_file = tmp_path / "export.xlsx"

        result = runner.invoke(
            app, ["export", "--table", "products", "--output", str(output_file)]
        )

        assert result.exit_code == 0

        # Load with openpyxl to check formatting
        from openpyxl import load_workbook

        wb = load_workbook(output_file)
        ws = wb.active

        # Check that headers are bold
        for cell in ws[1]:
            assert cell.font.bold is True

        # Check that freeze panes is set
        assert ws.freeze_panes == "A2"

    def test_export_records_history(self, runner, temp_project, tmp_path):
        """Test that export records history in _export_history table."""
        output_file = tmp_path / "export.xlsx"

        # Export
        result = runner.invoke(
            app, ["export", "--table", "products", "--output", str(output_file)]
        )

        assert result.exit_code == 0

        # Check history was recorded
        history = temp_project.database.query("SELECT * FROM _export_history")
        assert len(history) == 1
        assert history.iloc[0]["table_name"] == "products"
        assert history.iloc[0]["row_count"] == 3
        assert str(output_file) in history.iloc[0]["output_path"]

    def test_export_query_empty_results(self, runner, temp_project, tmp_path):
        """Test export query that returns no results."""
        output_file = tmp_path / "export.xlsx"

        result = runner.invoke(
            app,
            [
                "export",
                "--query",
                "SELECT * FROM products WHERE price > 1000",
                "--output",
                str(output_file)
            ]
        )

        assert result.exit_code == 0
        assert "No data to export" in result.stdout

    def test_export_displays_file_size(self, runner, temp_project, tmp_path):
        """Test that export summary includes file size."""
        output_file = tmp_path / "export.xlsx"

        result = runner.invoke(
            app, ["export", "--table", "products", "--output", str(output_file)]
        )

        assert result.exit_code == 0
        assert "File size" in result.stdout
        # File size should have a number and unit (bytes, KB, or MB)
        assert any(unit in result.stdout for unit in ["bytes", "KB", "MB"])
