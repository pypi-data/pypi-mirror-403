"""
Tests for HeaderDetector in auto_pilot module.
"""

import pytest
import pandas as pd
from pathlib import Path
import tempfile
import shutil
import gc

from excel_to_sql.auto_pilot.header_detector import HeaderDetector


def rmtree_with_retry(path):
    """Remove directory tree with retry for Windows file locks."""
    for _ in range(5):
        try:
            shutil.rmtree(path, ignore_errors=True)
            return
        except PermissionError:
            gc.collect()
            import time
            time.sleep(0.1)
    try:
        shutil.rmtree(path, ignore_errors=True)
    except:
        pass


class TestHeaderDetector:
    """Tests for HeaderDetector class."""

    @pytest.fixture
    def detector(self):
        """Create a HeaderDetector instance."""
        return HeaderDetector()

    @pytest.fixture
    def temp_dir(self):
        """Create a temporary directory for test files."""
        temp = Path(tempfile.mkdtemp())
        yield temp
        gc.collect()
        rmtree_with_retry(temp)

    # ──────────────────────────────────────────────────────────────
    # Tests for detect_header_row
    # ──────────────────────────────────────────────────────────────

    def test_detect_header_row_french_headers(self, detector, temp_dir):
        """Test detecting header row with French column names (Issue #42)."""
        # Create Excel file with French headers like in the issue
        # Write without headers so row 0 contains the header strings
        data = [
            ["No. du produit", "Nom du produit", "Description", "Classe Produit", "État", "Configuration"],
            ["2725", "SHOEI MENTONNIERE UNIV", "SHOEI MENTONNIERE UNIV", "FG", "Actif", "Configuration"],
            ["7353", "SHOEI CACHENEZ XR1000GMSUP", "SHOEI CACHENEZ XR1000GMSUP", "FG", "Actif", "Configuration"],
        ]
        excel_path = temp_dir / "french_headers.xlsx"
        # Write using openpyxl directly to have full control
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        for row in data:
            ws.append(row)
        wb.save(excel_path)
        wb.close()

        # Should detect row 0 as header
        header_row = detector.detect_header_row(excel_path)
        assert header_row == 0

        # Verify reading with detected header works correctly
        df_read = detector.read_excel_with_header_detection(excel_path)
        assert "No. du produit" in df_read.columns
        assert "Nom du produit" in df_read.columns
        assert "Description" in df_read.columns
        assert len(df_read) == 2

    def test_detect_header_row_english_headers(self, detector, temp_dir):
        """Test detecting header row with English column names."""
        data = [
            ["Product ID", "Product Name", "Description", "Category", "Status"],
            ["1", "Widget A", "A widget", "Hardware", "Active"],
            ["2", "Widget B", "Another widget", "Hardware", "Inactive"],
        ]
        excel_path = temp_dir / "english_headers.xlsx"
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        for row in data:
            ws.append(row)
        wb.save(excel_path)
        wb.close()

        header_row = detector.detect_header_row(excel_path)
        assert header_row == 0

    def test_detect_header_row_with_offset(self, detector, temp_dir):
        """Test detecting header row when it's not on the first row."""
        # Create Excel with empty first row, then headers
        data = [
            ["", "", ""],  # Empty row
            ["Product ID", "Name", "Status"],  # Header row
            ["1", "Product A", "Active"],
            ["2", "Product B", "Inactive"],
        ]
        excel_path = temp_dir / "offset_header.xlsx"
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        for row in data:
            ws.append(row)
        wb.save(excel_path)
        wb.close()

        # Should still detect row 0 as header (first non-empty row)
        # because the empty row doesn't contain header keywords
        # For this test, we expect row 0 since it doesn't match keywords
        header_row = detector.detect_header_row(excel_path)
        # The empty row won't match, so it will fall back to row 1
        # But our algorithm checks rows in order, so if row 1 has 2+ matches, it returns 1
        assert header_row == 1

    def test_detect_header_row_no_keywords(self, detector, temp_dir):
        """Test detection when file has no recognizable header keywords."""
        # Create Excel with generic data that doesn't match header patterns
        data = [
            ["A", "B", "C"],
            ["1", "2", "3"],
            ["4", "5", "6"],
        ]
        excel_path = temp_dir / "no_keywords.xlsx"
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        for row in data:
            ws.append(row)
        wb.save(excel_path)
        wb.close()

        # Should default to 0 when no clear header is found
        header_row = detector.detect_header_row(excel_path)
        assert header_row == 0

    # ──────────────────────────────────────────────────────────────
    # Tests for normalize_column_name
    # ──────────────────────────────────────────────────────────────

    def test_normalize_column_name_french_accents(self, detector):
        """Test normalizing French column names with accents."""
        assert detector.normalize_column_name("État") == "etat"
        assert detector.normalize_column_name("Catégorie") == "categorie"
        assert detector.normalize_column_name("Numéro") == "numero"

    def test_normalize_column_name_special_characters(self, detector):
        """Test normalizing column names with special characters."""
        assert detector.normalize_column_name("No. du produit") == "no_produit"
        assert detector.normalize_column_name("Product #1") == "product_1"
        assert detector.normalize_column_name("Price ($)") == "price"

    def test_normalize_column_name_multi_word(self, detector):
        """Test normalizing multi-word column names."""
        assert detector.normalize_column_name("Product Category Name") == "product_category_name"
        assert detector.normalize_column_name("Nom du produit") == "nom_produit"
        assert detector.normalize_column_name("Catégorie de produit #1") == "categorie_produit_1"

    def test_normalize_column_name_stop_words(self, detector):
        """Test that French stop words are removed."""
        assert detector.normalize_column_name("Le nom du produit") == "nom_produit"
        assert detector.normalize_column_name("La catégorie de produit") == "categorie_produit"

    def test_normalize_column_name_english(self, detector):
        """Test normalizing English column names."""
        assert detector.normalize_column_name("Product Name") == "product_name"
        assert detector.normalize_column_name("customer_id") == "customer_id"
        assert detector.normalize_column_name("Order Date") == "order_date"

    # ──────────────────────────────────────────────────────────────
    # Tests for read_excel_with_header_detection
    # ──────────────────────────────────────────────────────────────

    def test_read_with_header_detection_french(self, detector, temp_dir):
        """Test reading Excel file with French headers using auto-detection."""
        # Create file similar to issue #42
        data = [
            ["No. du produit", "Nom du produit", "Description", "État"],
            ["2725", "SHOEI MENTONNIERE", "Description 1", "Actif"],
            ["7353", "SHOEI CACHENEZ", "Description 2", "Actif"],
        ]
        excel_path = temp_dir / "produits.xlsx"
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        for row in data:
            ws.append(row)
        wb.save(excel_path)
        wb.close()

        # Read with header detection
        result = detector.read_excel_with_header_detection(excel_path)

        # Verify headers are correctly detected
        assert "No. du produit" in result.columns
        assert "Nom du produit" in result.columns
        assert len(result) == 2

    def test_read_with_header_detection_normalize(self, detector, temp_dir):
        """Test reading Excel with column name normalization."""
        data = [
            ["No. du produit", "Nom du produit", "État"],
            ["1", "Product A", "Actif"],
        ]
        excel_path = temp_dir / "test.xlsx"
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        for row in data:
            ws.append(row)
        wb.save(excel_path)
        wb.close()

        # Read with normalization enabled
        result = detector.read_excel_with_header_detection(excel_path, normalize_columns=True)

        # Verify column names are normalized
        assert "no_produit" in result.columns
        assert "nom_produit" in result.columns
        assert "etat" in result.columns
        assert len(result) == 1

    def test_read_with_header_detection_multiple_sheets(self, detector, temp_dir):
        """Test reading specific sheet with header detection."""
        # Create Excel with multiple sheets
        from openpyxl import Workbook, load_workbook

        excel_path = temp_dir / "multi_sheet.xlsx"
        wb = Workbook()

        # First sheet
        ws1 = wb.active
        ws1.title = "Products"
        for row in [["Product ID", "Name"], ["1", "A"], ["2", "B"]]:
            ws1.append(row)

        # Second sheet
        ws2 = wb.create_sheet("Orders")
        for row in [["Order ID", "Date"], ["101", "2024-01-01"], ["102", "2024-01-02"]]:
            ws2.append(row)

        wb.save(excel_path)
        wb.close()

        # Read first sheet
        df1 = detector.read_excel_with_header_detection(excel_path, sheet_name="Products")
        assert "Product ID" in df1.columns
        assert len(df1) == 2

        # Read second sheet
        df2 = detector.read_excel_with_header_detection(excel_path, sheet_name="Orders")
        assert "Order ID" in df2.columns
        assert len(df2) == 2

    # ──────────────────────────────────────────────────────────────
    # Tests for _is_header_row (private method)
    # ──────────────────────────────────────────────────────────────

    def test_is_header_row_with_keywords(self, detector):
        """Test _is_header_row identifies header rows with keywords."""
        df = pd.DataFrame([
            ["Product ID", "Product Name", "Category", "Status"],
            ["1", "Widget A", "Hardware", "Active"],
        ])

        # First row should be identified as header
        assert detector._is_header_row(df, 0, min_matches=2) is True
        # Second row should not be identified as header
        assert detector._is_header_row(df, 1, min_matches=2) is False

    def test_is_header_row_without_keywords(self, detector):
        """Test _is_header_row with rows that have no header keywords."""
        df = pd.DataFrame([
            ["A", "B", "C"],
            ["D", "E", "F"],
        ])

        # Neither row should be identified as header
        assert detector._is_header_row(df, 0, min_matches=2) is False
        assert detector._is_header_row(df, 1, min_matches=2) is False

    # ──────────────────────────────────────────────────────────────
    # Integration test for Issue #42 scenario
    # ──────────────────────────────────────────────────────────────

    def test_issue_42_scenario(self, detector, temp_dir):
        """
        Integration test for Issue #42.
        Tests the exact scenario described in the issue: French headers not being detected.
        """
        # Create the exact data from Issue #42
        headers = ["No. du produit", "Nom du produit", "Description", "Classe Produit",
                   "Catégorie de produit #1", "Catégorie de produit #2", "Catégorie de produit #3",
                   "État", "Configuration", "EAN Alternatif"]

        data_rows = [
            ["2725", "SHOEI MENTONNIERE UNIV", "SHOEI MENTONNIERE UNIV", "FG",
             "Pilote", "CASQUES", "DIVERS ACCESS.CASQUES-LUN", "Actif", "Configuration", ""],
            ["7353", "SHOEI CACHENEZ XR1000GMSUP", "SHOEI CACHENEZ XR1000GMSUP", "FG",
             "Pilote", "CASQUES", "DIVERS ACCESS.CASQUES-LUN", "Actif", "Configuration", ""],
        ]

        # Create Excel file using openpyxl
        excel_path = temp_dir / "produits_issue_42.xlsx"
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        # Add headers as first row
        ws.append(headers)
        # Add data rows
        for row in data_rows:
            ws.append(row)
        wb.save(excel_path)
        wb.close()

        # Detect header row
        header_row = detector.detect_header_row(excel_path)
        assert header_row == 0, "Header should be detected on row 0"

        # Read with header detection
        result = detector.read_excel_with_header_detection(excel_path)

        # Verify all French headers are correctly detected
        assert "No. du produit" in result.columns
        assert "Nom du produit" in result.columns
        assert "Description" in result.columns
        assert "Classe Produit" in result.columns
        assert "Catégorie de produit #1" in result.columns
        assert "Catégorie de produit #2" in result.columns
        assert "Catégorie de produit #3" in result.columns
        assert "État" in result.columns
        assert "Configuration" in result.columns
        assert "EAN Alternatif" in result.columns

        # Verify no "Unnamed" columns (the bug being fixed)
        unnamed_cols = [col for col in result.columns if str(col).startswith("Unnamed")]
        assert len(unnamed_cols) == 0, f"Found 'Unnamed' columns: {unnamed_cols}"

        # Verify data is correct
        assert len(result) == 2
        assert str(result["No. du produit"].iloc[0]) == "2725"
        assert result["Nom du produit"].iloc[1] == "SHOEI CACHENEZ XR1000GMSUP"

        # Test with normalized column names
        result_normalized = detector.read_excel_with_header_detection(excel_path, normalize_columns=True)
        assert "no_produit" in result_normalized.columns
        assert "nom_produit" in result_normalized.columns
        assert "etat" in result_normalized.columns
