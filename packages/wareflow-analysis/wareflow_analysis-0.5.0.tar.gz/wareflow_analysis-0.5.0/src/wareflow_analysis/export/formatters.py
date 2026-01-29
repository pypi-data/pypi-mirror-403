"""Excel Formatters module for styling and formatting Excel sheets.

This module provides predefined styles and formatting utilities for Excel sheets.
"""

from typing import Dict, Any
from numbers import Number

try:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError:
    raise ImportError(
        "openpyxl is required for Excel export. "
        "Install it with: pip install openpyxl"
    )


class ExcelFormatter:
    """Excel formatter with predefined styles and formatting methods."""

    # Predefined styles
    HEADER_STYLE = {
        "font": Font(bold=True, color="FFFFFF", size=11),
        "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border": Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        ),
    }

    NUMBER_STYLE = {
        "font": Font(size=11),
        "alignment": Alignment(horizontal="right", vertical="center"),
        "number_format": "#,##0",
    }

    PERCENTAGE_STYLE = {
        "font": Font(size=11),
        "alignment": Alignment(horizontal="right", vertical="center"),
        "number_format": "0.0",
    }

    WARNING_STYLE = {
        "font": Font(bold=True, color="FFFFFF", size=11),
        "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "alignment": Alignment(horizontal="left", vertical="center"),
    }

    TITLE_STYLE = {
        "font": Font(bold=True, size=14, color="4472C4"),
        "alignment": Alignment(horizontal="left", vertical="center"),
    }

    def __init__(self):
        """Initialize Excel formatter."""
        pass

    def apply_style_to_cell(self, sheet: Worksheet, row: int, col: int, style: Dict[str, Any]) -> None:
        """Apply a style dictionary to a specific cell.

        Args:
            sheet: Worksheet object
            row: Row number (1-indexed)
            col: Column number (1-indexed)
            style: Dictionary of style attributes
        """
        cell = sheet.cell(row=row, column=col)

        if "font" in style:
            cell.font = style["font"]
        if "fill" in style:
            cell.fill = style["fill"]
        if "alignment" in style:
            cell.alignment = style["alignment"]
        if "border" in style:
            cell.border = style["border"]
        if "number_format" in style:
            cell.number_format = style["number_format"]

    def apply_header_style_to_row(self, sheet: Worksheet, row: int, num_cols: int) -> None:
        """Apply header style to an entire row.

        Args:
            sheet: Worksheet object
            row: Row number to style (1-indexed)
            num_cols: Number of columns to style
        """
        for col in range(1, num_cols + 1):
            self.apply_style_to_cell(sheet, row, col, self.HEADER_STYLE)

    def format_summary_table(
        self,
        sheet: Worksheet,
        start_row: int,
        data: Dict[str, Any],
    ) -> int:
        """Format a summary table with labels and values.

        Args:
            sheet: Worksheet object
            start_row: Starting row number (1-indexed)
            data: Dictionary with labels as keys and values as values

        Returns:
            Number of rows written
        """
        current_row = start_row

        for label, value in data.items():
            # Label
            label_cell = sheet.cell(row=current_row, column=1, value=label)
            label_cell.font = Font(bold=True, size=11)

            # Value
            value_cell = sheet.cell(row=current_row, column=2, value=value)

            # Format based on value type
            if isinstance(value, Number) and not isinstance(value, bool):
                value_cell.number_format = "#,##0"
                value_cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                value_cell.alignment = Alignment(horizontal="left", vertical="center")

            current_row += 1

        return current_row - start_row

    def format_data_table(
        self,
        sheet: Worksheet,
        start_row: int,
        headers: list,
        data_rows: list,
    ) -> int:
        """Format a data table with headers and rows.

        Args:
            sheet: Worksheet object
            start_row: Starting row number (1-indexed)
            headers: List of header values
            data_rows: List of lists with data values

        Returns:
            Number of rows written (including header)
        """
        # Write and format headers
        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=start_row, column=col, value=header)
            self.apply_style_to_cell(sheet, start_row, col, self.HEADER_STYLE)

        # Write data rows
        current_row = start_row + 1
        for row_data in data_rows:
            for col, value in enumerate(row_data, start=1):
                cell = sheet.cell(row=current_row, column=col, value=value)

                # Auto-detect formatting based on value type
                if isinstance(value, Number) and not isinstance(value, bool):
                    # Check if it's a percentage (0-100)
                    if 0 <= value <= 100 and col == len(headers):
                        # Last column might be percentage
                        cell.number_format = "0.0"
                    else:
                        cell.number_format = "#,##0"
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

            current_row += 1

        return current_row - start_row

    def add_warning_row(
        self,
        sheet: Worksheet,
        row: int,
        warning_text: str,
    ) -> None:
        """Add a warning row with distinctive styling.

        Args:
            sheet: Worksheet object
            row: Row number (1-indexed)
            warning_text: Warning message to display
        """
        cell = sheet.cell(row=row, column=1, value=warning_text)
        self.apply_style_to_cell(sheet, row, 1, self.WARNING_STYLE)

    def add_title(
        self,
        sheet: Worksheet,
        row: int,
        title: str,
        col: int = 1,
    ) -> None:
        """Add a title row with formatting.

        Args:
            sheet: Worksheet object
            row: Row number (1-indexed)
            title: Title text
            col: Column number (default: 1)
        """
        cell = sheet.cell(row=row, column=col, value=title)
        cell.font = Font(bold=True, size=14, color="4472C4")
        cell.alignment = Alignment(horizontal="left", vertical="center")

    def add_blank_row(self, sheet: Worksheet, row: int) -> None:
        """Add a blank row for spacing.

        Args:
            sheet: Worksheet object
            row: Row number to make blank (1-indexed)
        """
        # This is implicit - just returns the next row number
        pass

    def apply_borders_to_range(
        self,
        sheet: Worksheet,
        start_row: int,
        start_col: int,
        end_row: int,
        end_col: int,
    ) -> None:
        """Apply borders to a range of cells.

        Args:
            sheet: Worksheet object
            start_row: Starting row (1-indexed)
            start_col: Starting column (1-indexed)
            end_row: Ending row (1-indexed)
            end_col: Ending column (1-indexed)
        """
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                sheet.cell(row=row, column=col).border = border
