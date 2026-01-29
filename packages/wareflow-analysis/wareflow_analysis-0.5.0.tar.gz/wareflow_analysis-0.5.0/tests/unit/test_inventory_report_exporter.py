"""Unit tests for Inventory Report Exporter module."""

import pytest
from pathlib import Path
from datetime import datetime

from wareflow_analysis.export.reports.inventory_report import InventoryReportExporter


class TestInventoryReportExporter:
    """Test suite for InventoryReportExporter class."""

    @pytest.fixture
    def sample_inventory_results(self):
        """Create sample inventory analysis results.

        Returns:
            Dictionary with inventory analysis results
        """
        return {
            "total_products": 150,
            "by_category": [
                {"category": "Pilote", "count": 100, "percentage": 66.7},
                {"category": "Accessoires", "count": 30, "percentage": 20.0},
                {"category": "Moto", "count": 20, "percentage": 13.3},
            ],
            "by_status": [
                {"status": "Actif", "count": 140, "percentage": 93.3},
                {"status": "Inactif", "count": 10, "percentage": 6.7},
            ],
            "issues": {
                "missing_ean": 50,
                "missing_description": 5,
                "missing_name": 0,
            }
        }

    @pytest.fixture
    def perfect_inventory_results(self):
        """Create inventory results with no issues.

        Returns:
            Dictionary with perfect inventory analysis results
        """
        return {
            "total_products": 100,
            "by_category": [
                {"category": "Pilote", "count": 100, "percentage": 100.0},
            ],
            "by_status": [
                {"status": "Actif", "count": 100, "percentage": 100.0},
            ],
            "issues": {
                "missing_ean": 0,
                "missing_description": 0,
                "missing_name": 0,
            }
        }

    def test_exporter_initialization(self):
        """Test exporter initialization."""
        exporter = InventoryReportExporter()
        assert exporter is not None
        assert exporter.builder is None
        assert exporter.formatter is None

    def test_export_creates_file(self, tmp_path, sample_inventory_results):
        """Test that export creates an Excel file."""
        exporter = InventoryReportExporter()
        output_path = tmp_path / "inventory_report.xlsx"

        exporter.export(sample_inventory_results, output_path)

        assert output_path.exists()

    def test_export_creates_three_sheets(self, tmp_path, sample_inventory_results):
        """Test that export creates the expected sheets."""
        from openpyxl import load_workbook

        exporter = InventoryReportExporter()
        output_path = tmp_path / "inventory_report.xlsx"

        exporter.export(sample_inventory_results, output_path)

        # Load and verify sheets
        wb = load_workbook(output_path)
        sheet_names = wb.sheetnames

        assert "Résumé" in sheet_names
        assert "Par Catégorie" in sheet_names
        assert "Par État" in sheet_names
        assert len(sheet_names) == 3

    def test_summary_sheet_content(self, tmp_path, sample_inventory_results):
        """Test that summary sheet contains correct data."""
        from openpyxl import load_workbook

        exporter = InventoryReportExporter()
        output_path = tmp_path / "inventory_report.xlsx"

        exporter.export(sample_inventory_results, output_path)

        wb = load_workbook(output_path)
        sheet = wb["Résumé"]

        # Verify title exists
        assert sheet["A1"].value is not None
        assert "ANALYSE D'INVENTAIRE" in str(sheet["A1"].value)

        # Verify total products is present
        found_total = False
        for row in sheet.iter_rows(min_row=2, max_row=10, min_col=1, max_col=2):
            if row[0].value == "Total produits":
                assert row[1].value == 150
                found_total = True
                break

        assert found_total, "Total produits not found in summary sheet"

    def test_category_sheet_content(self, tmp_path, sample_inventory_results):
        """Test that category sheet contains correct data."""
        from openpyxl import load_workbook

        exporter = InventoryReportExporter()
        output_path = tmp_path / "inventory_report.xlsx"

        exporter.export(sample_inventory_results, output_path)

        wb = load_workbook(output_path)
        sheet = wb["Par Catégorie"]

        # Verify headers (row 2, after title)
        assert sheet["A2"].value == "category"
        assert sheet["B2"].value == "count"
        assert sheet["C2"].value == "percentage"

        # Verify first data row
        assert sheet["A3"].value == "Pilote"
        assert sheet["B3"].value == 100

    def test_status_sheet_content(self, tmp_path, sample_inventory_results):
        """Test that status sheet contains correct data."""
        from openpyxl import load_workbook

        exporter = InventoryReportExporter()
        output_path = tmp_path / "inventory_report.xlsx"

        exporter.export(sample_inventory_results, output_path)

        wb = load_workbook(output_path)
        sheet = wb["Par État"]

        # Verify headers
        assert sheet["A2"].value == "status"
        assert sheet["B2"].value == "count"
        assert sheet["C2"].value == "percentage"

        # Verify first data row
        assert sheet["A3"].value == "Actif"
        assert sheet["B3"].value == 140

    def test_issues_displayed_in_summary(self, tmp_path, sample_inventory_results):
        """Test that quality issues are displayed in summary."""
        from openpyxl import load_workbook

        exporter = InventoryReportExporter()
        output_path = tmp_path / "inventory_report.xlsx"

        exporter.export(sample_inventory_results, output_path)

        wb = load_workbook(output_path)
        sheet = wb["Résumé"]

        # Look for warning text in the sheet
        sheet_content = []
        for row in sheet.iter_rows(values_only=True):
            sheet_content.extend([str(cell) if cell is not None else "" for cell in row])

        sheet_text = " ".join(sheet_content)

        # Verify warnings are present
        assert "50" in sheet_text  # Missing EAN count
        assert "5" in sheet_text   # Missing description count

    def test_perfect_inventory_no_warnings(self, tmp_path, perfect_inventory_results):
        """Test that perfect inventory shows no warnings."""
        from openpyxl import load_workbook

        exporter = InventoryReportExporter()
        output_path = tmp_path / "inventory_report.xlsx"

        exporter.export(perfect_inventory_results, output_path)

        wb = load_workbook(output_path)
        sheet = wb["Résumé"]

        # Look for success message
        sheet_content = []
        for row in sheet.iter_rows(values_only=True):
            sheet_content.extend([str(cell) if cell is not None else "" for cell in row])

        sheet_text = " ".join(sheet_content)

        # Should show success message
        assert "Aucun problème" in sheet_text or "qualité" in sheet_text.lower()

    def test_export_with_empty_category_list(self, tmp_path):
        """Test export with empty category list."""
        exporter = InventoryReportExporter()
        output_path = tmp_path / "inventory_empty.xlsx"

        results = {
            "total_products": 100,
            "by_category": [],
            "by_status": [
                {"status": "Actif", "count": 100, "percentage": 100.0}
            ],
            "issues": {
                "missing_ean": 0,
                "missing_description": 0,
                "missing_name": 0,
            }
        }

        # Should not raise error
        exporter.export(results, output_path)
        assert output_path.exists()

    def test_export_creates_directory_if_needed(self, tmp_path):
        """Test that export creates output directory if it doesn't exist."""
        exporter = InventoryReportExporter()

        # Create path with non-existent subdirectories
        output_path = tmp_path / "reports" / "2025" / "inventory.xlsx"

        results = {
            "total_products": 10,
            "by_category": [{"category": "Test", "count": 10, "percentage": 100.0}],
            "by_status": [{"status": "Actif", "count": 10, "percentage": 100.0}],
            "issues": {"missing_ean": 0, "missing_description": 0, "missing_name": 0}
        }

        exporter.export(results, output_path)

        assert output_path.exists()
        assert output_path.parent.exists()

    def test_timestamp_in_title(self, tmp_path, sample_inventory_results):
        """Test that summary sheet title includes current date."""
        from openpyxl import load_workbook

        exporter = InventoryReportExporter()
        output_path = tmp_path / "inventory_report.xlsx"

        exporter.export(sample_inventory_results, output_path)

        wb = load_workbook(output_path)
        sheet = wb["Résumé"]

        title = str(sheet["A1"].value)
        current_year = str(datetime.now().year)

        # Verify current year is in title
        assert current_year in title
