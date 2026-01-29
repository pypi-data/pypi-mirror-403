"""ABC Classification Report Exporter.

Exports ABC classification analysis results to formatted Excel workbook.
"""

from pathlib import Path
from datetime import datetime
import pandas as pd

from wareflow_analysis.export.excel_builder import ExcelBuilder
from wareflow_analysis.export.formatters import ExcelFormatter


class ABCReportExporter:
    """Export ABC classification results to Excel."""

    def __init__(self):
        """Initialize ABC report exporter."""
        self.builder = None
        self.formatter = None

    def export(self, results: dict, output_path: Path) -> None:
        """Export ABC analysis results to Excel file.

        Args:
            results: Dictionary with ABC analysis results
            output_path: Path where to save the Excel file
        """
        # Initialize builder and formatter
        self.builder = ExcelBuilder()
        self.formatter = ExcelFormatter()

        # Create summary sheet
        self._create_summary_sheet(results)

        # Create classification sheet
        self._create_classification_sheet(results)

        # Remove default empty sheet if exists
        self.builder.remove_default_sheet()

        # Save the workbook
        self.builder.save(output_path)

    def _create_summary_sheet(self, results: dict) -> None:
        """Create the summary sheet with ABC overview statistics.

        Args:
            results: ABC analysis results
        """
        sheet_name = "Résumé ABC"

        classification = results.get("classification", [])
        summary = results.get("summary", {})

        if not classification:
            # No data - show empty state
            self.builder.add_sheet_with_key_value(
                sheet_name=sheet_name,
                data={
                    "Total Products": 0,
                    "Class A": 0,
                    "Class B": 0,
                    "Class C": 0,
                },
                title=f"ANALYSE ABC - {datetime.now().strftime('%Y-%m-%d')}",
            )
            return

        # Calculate statistics
        total_products = summary.get("total_products", 0)
        total_picks = summary.get("total_picks", 0)

        class_a = summary.get("class_a", {})
        class_b = summary.get("class_b", {})
        class_c = summary.get("class_c", {})

        # Add summary data
        summary_data = {
            "Date de l'analyse": datetime.now().strftime('%Y-%m-%d %H:%M'),
            "Produits analysés": total_products,
            "---": "---",
            "Classe A": f"{class_a.get('count', 0):,} produits ({class_a.get('percentage', 0):.1f}%)",
            "  Picks A": f"{class_a.get('picks', 0):,} ({class_a.get('percentage', 0):.1f}%)",
            "---": "---",
            "Classe B": f"{class_b.get('count', 0):,} produits ({class_b.get('percentage', 0):.1f}%)",
            "  Picks B": f"{class_b.get('picks', 0):,} ({class_b.get('percentage', 0):.1f}%)",
            "---": "---",
            "Classe C": f"{class_c.get('count', 0):,} produits ({class_c.get('percentage', 0):.1f}%)",
            "  Picks C": f"{class_c.get('picks', 0):,} ({class_c.get('percentage', 0):.1f}%)",
            "---": "---",
            "Total picks": f"{total_picks:,}",
        }

        self.builder.add_sheet_with_key_value(
            sheet_name=sheet_name,
            data=summary_data,
            title=f"ANALYSE ABC - {datetime.now().strftime('%Y-%m-%d')}",
        )

        # Auto-adjust columns
        self.builder.auto_adjust_columns(sheet_name)

    def _create_classification_sheet(self, results: dict) -> None:
        """Create detailed classification sheet with all products.

        Args:
            results: ABC analysis results
        """
        from openpyxl.styles import PatternFill

        sheet_name = "Classification ABC"
        classification = results.get("classification", [])

        if not classification:
            return

        # Convert to DataFrame for easier handling
        df = pd.DataFrame(classification)

        # Add title
        title = f"Détail Classification ABC - {len(classification)} produits"

        # Add sheet with DataFrame
        self.builder.add_sheet_from_dataframe(sheet_name, df, title=title)

        # Get the sheet for formatting
        sheet = self.builder.workbook[sheet_name]

        # Apply color coding by ABC class
        # Class A: Green, Class B: Orange, Class C: Red
        for row_idx, row_data in enumerate(df.itertuples(index=False), start=3):  # Start after title (row 1) and header (row 2)
            abc_class = row_data[-1]  # abc_class is last column

            if abc_class == "A":
                fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif abc_class == "B":
                fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            elif abc_class == "C":
                fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            else:
                continue

            # Apply color to entire row
            for col_idx, _ in enumerate(row_data):
                cell = sheet.cell(row=row_idx, column=col_idx + 1)  # +1 because of title
                cell.fill = fill

        # Auto-adjust columns
        self.builder.auto_adjust_columns(sheet_name)

        # Freeze header row
        self.builder.freeze_panes(sheet_name, row=3)
