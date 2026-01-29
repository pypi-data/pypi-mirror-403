"""Excel Builder module for creating Excel workbooks.

This module provides a simple interface for creating Excel workbooks
with formatted data using openpyxl.
"""

from pathlib import Path
from typing import Any, Dict, List, Union
import pandas as pd

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
except ImportError:
    raise ImportError(
        "openpyxl is required for Excel export. "
        "Install it with: pip install openpyxl"
    )


class ExcelBuilder:
    """Excel workbook builder with formatting capabilities."""

    def __init__(self):
        """Initialize a new Excel builder."""
        self.workbook = Workbook()
        self.default_sheet = self.workbook.active
        # Remove default sheet if creating new sheets
        self.default_sheet.title = "Sheet1"

    def create_workbook(self) -> None:
        """Create a new Excel workbook."""
        self.workbook = Workbook()
        self.default_sheet = self.workbook.active
        self.default_sheet.title = "Sheet1"

    def add_sheet_from_dict(
        self,
        sheet_name: str,
        data: List[Dict[str, Any]],
        title: str = None,
    ) -> None:
        """Add a sheet from a list of dictionaries.

        Args:
            sheet_name: Name for the new sheet
            data: List of dictionaries with column names as keys
            title: Optional title to add at the top of the sheet
        """
        if not data:
            # Create empty sheet with headers only
            sheet = self.workbook.create_sheet(title=sheet_name)
            return

        # Convert to DataFrame for easier handling
        df = pd.DataFrame(data)

        # Create sheet
        sheet = self.workbook.create_sheet(title=sheet_name)

        # Add title if provided
        if title:
            sheet.cell(row=1, column=1, value=title)
            sheet.cell(row=1, column=1).font = Font(bold=True, size=14)
            start_row = 2
        else:
            start_row = 1

        # Write headers
        for col_idx, column in enumerate(df.columns, start=1):
            cell = sheet.cell(row=start_row, column=col_idx, value=column)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Write data
        for row_idx, row_data in enumerate(df.itertuples(index=False), start=start_row + 1):
            for col_idx, value in enumerate(row_data, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=value)

    def add_sheet_from_dataframe(
        self,
        sheet_name: str,
        df: pd.DataFrame,
        title: str = None,
    ) -> None:
        """Add a sheet from a pandas DataFrame.

        Args:
            sheet_name: Name for the new sheet
            df: DataFrame with data
            title: Optional title to add at the top of the sheet
        """
        if df.empty:
            # Create empty sheet
            sheet = self.workbook.create_sheet(title=sheet_name)
            return

        # Create sheet
        sheet = self.workbook.create_sheet(title=sheet_name)

        # Add title if provided
        if title:
            sheet.cell(row=1, column=1, value=title)
            sheet.cell(row=1, column=1).font = Font(bold=True, size=14)
            start_row = 2
        else:
            start_row = 1

        # Write headers
        for col_idx, column in enumerate(df.columns, start=1):
            cell = sheet.cell(row=start_row, column=col_idx, value=column)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Write data
        for row_idx, row_data in enumerate(df.itertuples(index=False), start=start_row + 1):
            for col_idx, value in enumerate(row_data, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=value)

    def add_sheet_with_key_value(
        self,
        sheet_name: str,
        data: Dict[str, Any],
        title: str = None,
    ) -> None:
        """Add a sheet from a dictionary (key-value pairs).

        Args:
            sheet_name: Name for the new sheet
            data: Dictionary with key-value pairs
            title: Optional title to add at the top of the sheet
        """
        sheet = self.workbook.create_sheet(title=sheet_name)

        start_row = 1
        if title:
            sheet.cell(row=1, column=1, value=title)
            sheet.cell(row=1, column=1).font = Font(bold=True, size=14)
            start_row = 3
            sheet.cell(row=2, column=1).value = ""  # Empty row for spacing

        # Write key-value pairs
        for row_idx, (key, value) in enumerate(data.items(), start=start_row):
            sheet.cell(row=row_idx, column=1, value=key)
            sheet.cell(row=row_idx, column=1).font = Font(bold=True)

            # Handle different value types
            if isinstance(value, (int, float)):
                sheet.cell(row=row_idx, column=2, value=value)
                sheet.cell(row=row_idx, column=2).number_format = "#,##0"
            elif isinstance(value, str):
                sheet.cell(row=row_idx, column=2, value=value)
            else:
                sheet.cell(row=row_idx, column=2, value=str(value))

    def auto_adjust_columns(self, sheet_name: str = None, max_width: int = 50) -> None:
        """Auto-adjust column widths based on content.

        Args:
            sheet_name: Name of sheet to adjust (None = all sheets)
            max_width: Maximum column width in characters
        """
        sheets_to_adjust = (
            [self.workbook[sheet_name]] if sheet_name else self.workbook.worksheets
        )

        for sheet in sheets_to_adjust:
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter

                for cell in column:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass

                # Set width with some padding
                adjusted_width = min(max_length + 2, max_width)
                sheet.column_dimensions[column_letter].width = adjusted_width

    def freeze_panes(self, sheet_name: str, row: int = 2, col: int = 1) -> None:
        """Freeze panes for a sheet.

        Args:
            sheet_name: Name of sheet
            row: Row number to freeze above (1-indexed)
            col: Column number to freeze left of (1-indexed)
        """
        if sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            sheet.freeze_panes = sheet.cell(row=row, column=col)

    def remove_default_sheet(self) -> None:
        """Remove the default 'Sheet' if empty."""
        if "Sheet1" in self.workbook.sheetnames:
            sheet = self.workbook["Sheet1"]
            # Only remove if empty (no data beyond row 1)
            if sheet.max_row == 1 and sheet.max_column == 1:
                self.workbook.remove(sheet)

    def save(self, filepath: Union[str, Path]) -> None:
        """Save the workbook to a file.

        Args:
            filepath: Path where to save the Excel file
        """
        filepath = Path(filepath)
        filepath.parent.mkdir(parents=True, exist_ok=True)
        self.workbook.save(filepath)
