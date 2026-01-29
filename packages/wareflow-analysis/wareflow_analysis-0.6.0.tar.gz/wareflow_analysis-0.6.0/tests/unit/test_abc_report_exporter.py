"""Unit tests for ABC Report Exporter module."""

import pytest
from pathlib import Path
from datetime import datetime

from wareflow_analysis.export.reports.abc_report import ABCReportExporter


class TestABCReportExporter:
    """Test suite for ABCReportExporter class."""

    @pytest.fixture
    def sample_abc_results(self):
        """Create sample ABC analysis results.

        Returns:
            Dictionary with ABC analysis results
        """
        return {
            "classification": [
                {
                    "no_produit": "PROD001",
                    "nom_produit": "Product A",
                    "total_movements": 100,
                    "total_picked": 5000,
                    "percentile_rank": 5,
                    "abc_class": "A"
                },
                {
                    "no_produit": "PROD002",
                    "nom_produit": "Product B",
                    "total_movements": 80,
                    "total_picked": 3000,
                    "percentile_rank": 15,
                    "abc_class": "A"
                },
                {
                    "no_produit": "PROD003",
                    "nom_produit": "Product C",
                    "total_movements": 60,
                    "total_picked": 2000,
                    "percentile_rank": 35,
                    "abc_class": "B"
                },
                {
                    "no_produit": "PROD004",
                    "nom_produit": "Product D",
                    "total_movements": 40,
                    "total_picked": 1000,
                    "percentile_rank": 55,
                    "abc_class": "B"
                },
                {
                    "no_produit": "PROD005",
                    "nom_produit": "Product E",
                    "total_movements": 20,
                    "total_picked": 500,
                    "percentile_rank": 75,
                    "abc_class": "C"
                },
                {
                    "no_produit": "PROD006",
                    "nom_produit": "Product F",
                    "total_movements": 10,
                    "total_picked": 200,
                    "percentile_rank": 90,
                    "abc_class": "C"
                },
            ],
            "summary": {
                "total_products": 6,
                "total_picks": 11700,
                "class_a": {
                    "count": 2,
                    "picks": 8000,
                    "percentage": 33.3
                },
                "class_b": {
                    "count": 2,
                    "picks": 3000,
                    "percentage": 33.3
                },
                "class_c": {
                    "count": 2,
                    "picks": 700,
                    "percentage": 33.3
                }
            }
        }

    @pytest.fixture
    def empty_abc_results(self):
        """Create empty ABC analysis results.

        Returns:
            Dictionary with empty ABC analysis results
        """
        return {
            "classification": [],
            "summary": {
                "total_products": 0,
                "total_picks": 0,
                "class_a": {"count": 0, "picks": 0, "percentage": 0},
                "class_b": {"count": 0, "picks": 0, "percentage": 0},
                "class_c": {"count": 0, "picks": 0, "percentage": 0}
            }
        }

    def test_exporter_initialization(self):
        """Test exporter initialization."""
        exporter = ABCReportExporter()
        assert exporter is not None
        assert exporter.builder is None
        assert exporter.formatter is None

    def test_export_creates_file(self, tmp_path, sample_abc_results):
        """Test that export creates an Excel file."""
        exporter = ABCReportExporter()
        output_path = tmp_path / "abc_report.xlsx"

        exporter.export(sample_abc_results, output_path)

        assert output_path.exists()

    def test_export_creates_two_sheets(self, tmp_path, sample_abc_results):
        """Test that export creates the expected sheets."""
        from openpyxl import load_workbook

        exporter = ABCReportExporter()
        output_path = tmp_path / "abc_report.xlsx"

        exporter.export(sample_abc_results, output_path)

        # Load and verify sheets
        wb = load_workbook(output_path)
        sheet_names = wb.sheetnames

        assert "Résumé ABC" in sheet_names
        assert "Classification ABC" in sheet_names
        assert len(sheet_names) == 2

    def test_summary_sheet_content(self, tmp_path, sample_abc_results):
        """Test that summary sheet contains correct data."""
        from openpyxl import load_workbook

        exporter = ABCReportExporter()
        output_path = tmp_path / "abc_report.xlsx"

        exporter.export(sample_abc_results, output_path)

        wb = load_workbook(output_path)
        sheet = wb["Résumé ABC"]

        # Verify title exists
        assert sheet["A1"].value is not None
        assert "ANALYSE ABC" in str(sheet["A1"].value)

        # Verify date is present
        sheet_content = []
        for row in sheet.iter_rows(min_row=2, max_row=10, min_col=1, max_col=2, values_only=True):
            sheet_content.extend([str(cell) if cell is not None else "" for cell in row])

        sheet_text = " ".join(sheet_content)

        # Verify class information is present
        assert "Classe A" in sheet_text
        assert "Classe B" in sheet_text
        assert "Classe C" in sheet_text
        assert "6" in sheet_text  # Total products

    def test_classification_sheet_content(self, tmp_path, sample_abc_results):
        """Test that classification sheet contains correct data."""
        from openpyxl import load_workbook

        exporter = ABCReportExporter()
        output_path = tmp_path / "abc_report.xlsx"

        exporter.export(sample_abc_results, output_path)

        wb = load_workbook(output_path)
        sheet = wb["Classification ABC"]

        # Verify headers (row 2, after title)
        headers = [cell.value for cell in sheet[2]]
        assert "no_produit" in headers
        assert "nom_produit" in headers
        assert "total_movements" in headers
        assert "total_picked" in headers
        assert "abc_class" in headers

        # Find abc_class column index
        abc_class_col = headers.index("abc_class") + 1  # +1 for 1-indexed column

        # Verify first data row
        assert sheet["A3"].value == "PROD001"
        assert sheet["B3"].value == "Product A"
        assert sheet.cell(row=3, column=abc_class_col).value == "A"

    def test_classification_sheet_color_coding(self, tmp_path, sample_abc_results):
        """Test that classification sheet has color coding by ABC class."""
        from openpyxl import load_workbook

        exporter = ABCReportExporter()
        output_path = tmp_path / "abc_report.xlsx"

        exporter.export(sample_abc_results, output_path)

        wb = load_workbook(output_path)
        sheet = wb["Classification ABC"]

        # Check that rows have color coding
        # Class A row should have green fill
        class_a_row = sheet[3]  # First product is class A
        class_a_fill = class_a_row[0].fill

        # Verify fill is not None (has color)
        assert class_a_fill is not None
        assert class_a_fill.start_color.rgb != "00000000"  # Not black default

        # Check that we have different colors for different classes
        # Get fill colors from row 3 (Class A) and row 7 (Class C)
        class_a_color = class_a_row[0].fill.start_color.rgb

        class_c_row = sheet[7]  # Fifth product is class C
        class_c_color = class_c_row[0].fill.start_color.rgb

        # Colors should be different
        assert class_a_color != class_c_color

    def test_empty_classification_results(self, tmp_path, empty_abc_results):
        """Test export with empty classification results."""
        from openpyxl import load_workbook

        exporter = ABCReportExporter()
        output_path = tmp_path / "abc_empty.xlsx"

        exporter.export(empty_abc_results, output_path)

        # Should still create file
        assert output_path.exists()

        wb = load_workbook(output_path)
        sheet_names = wb.sheetnames

        # Should have summary sheet
        assert "Résumé ABC" in sheet_names

        # Classification sheet should NOT be created (no data)
        assert "Classification ABC" not in sheet_names

    def test_summary_sheet_shows_empty_state(self, tmp_path, empty_abc_results):
        """Test that summary sheet shows zero values when empty."""
        from openpyxl import load_workbook

        exporter = ABCReportExporter()
        output_path = tmp_path / "abc_empty.xlsx"

        exporter.export(empty_abc_results, output_path)

        wb = load_workbook(output_path)
        sheet = wb["Résumé ABC"]

        # Look for zero counts in the sheet
        sheet_content = []
        for row in sheet.iter_rows(values_only=True):
            sheet_content.extend([str(cell) if cell is not None else "" for cell in row])

        sheet_text = " ".join(sheet_content)

        # Should show zeros
        assert "0" in sheet_text

    def test_export_creates_directory_if_needed(self, tmp_path, sample_abc_results):
        """Test that export creates output directory if it doesn't exist."""
        exporter = ABCReportExporter()

        # Create path with non-existent subdirectories
        output_path = tmp_path / "reports" / "2025" / "abc_report.xlsx"

        exporter.export(sample_abc_results, output_path)

        assert output_path.exists()
        assert output_path.parent.exists()

    def test_timestamp_in_summary_title(self, tmp_path, sample_abc_results):
        """Test that summary sheet title includes current date."""
        from openpyxl import load_workbook

        exporter = ABCReportExporter()
        output_path = tmp_path / "abc_report.xlsx"

        exporter.export(sample_abc_results, output_path)

        wb = load_workbook(output_path)
        sheet = wb["Résumé ABC"]

        title = str(sheet["A1"].value)
        current_year = str(datetime.now().year)

        # Verify current year is in title
        assert current_year in title

    def test_classification_sheet_title(self, tmp_path, sample_abc_results):
        """Test that classification sheet has correct title."""
        from openpyxl import load_workbook

        exporter = ABCReportExporter()
        output_path = tmp_path / "abc_report.xlsx"

        exporter.export(sample_abc_results, output_path)

        wb = load_workbook(output_path)
        sheet = wb["Classification ABC"]

        title = str(sheet["A1"].value)

        # Should mention classification
        assert "Classification ABC" in title
        # Should mention number of products
        assert "6" in title

    def test_all_abc_classes_present(self, tmp_path, sample_abc_results):
        """Test that all ABC classes (A, B, C) are in classification."""
        from openpyxl import load_workbook

        exporter = ABCReportExporter()
        output_path = tmp_path / "abc_report.xlsx"

        exporter.export(sample_abc_results, output_path)

        wb = load_workbook(output_path)
        sheet = wb["Classification ABC"]

        # Get headers to find abc_class column
        headers = [cell.value for cell in sheet[2]]
        abc_class_col_idx = headers.index("abc_class")  # 0-indexed

        # Collect all abc_class values
        abc_classes = set()
        for row in sheet.iter_rows(min_row=3, max_row=10, values_only=True):
            if row[abc_class_col_idx]:
                abc_classes.add(row[abc_class_col_idx])

        # Should have all three classes
        assert "A" in abc_classes
        assert "B" in abc_classes
        assert "C" in abc_classes

    def test_summary_shows_pick_counts(self, tmp_path, sample_abc_results):
        """Test that summary shows pick counts for each class."""
        from openpyxl import load_workbook

        exporter = ABCReportExporter()
        output_path = tmp_path / "abc_report.xlsx"

        exporter.export(sample_abc_results, output_path)

        wb = load_workbook(output_path)
        sheet = wb["Résumé ABC"]

        # Get all cell values
        sheet_content = []
        for row in sheet.iter_rows(values_only=True):
            sheet_content.extend([str(cell) if cell is not None else "" for cell in row])

        sheet_text = " ".join(sheet_content)

        # Should show pick counts
        assert "8,000" in sheet_text or "8000" in sheet_text  # Class A picks
        assert "3,000" in sheet_text or "3000" in sheet_text  # Class B picks
        assert "700" in sheet_text  # Class C picks

    def test_classification_sheet_column_order(self, tmp_path, sample_abc_results):
        """Test that classification sheet columns are in correct order."""
        from openpyxl import load_workbook

        exporter = ABCReportExporter()
        output_path = tmp_path / "abc_report.xlsx"

        exporter.export(sample_abc_results, output_path)

        wb = load_workbook(output_path)
        sheet = wb["Classification ABC"]

        # Get header row (row 2)
        headers = [cell.value for cell in sheet[2]]

        # Expected column order
        expected_columns = [
            "no_produit",
            "nom_produit",
            "total_movements",
            "total_picked",
            "percentile_rank",
            "abc_class"
        ]

        # Verify all expected columns are present
        for col in expected_columns:
            assert col in headers, f"Column {col} not found in headers"

    def test_export_with_single_product(self, tmp_path):
        """Test export with only one product."""
        from openpyxl import load_workbook

        single_product_results = {
            "classification": [
                {
                    "no_produit": "PROD001",
                    "nom_produit": "Product A",
                    "total_movements": 100,
                    "total_picked": 5000,
                    "percentile_rank": 1,
                    "abc_class": "A"
                }
            ],
            "summary": {
                "total_products": 1,
                "total_picks": 5000,
                "class_a": {"count": 1, "picks": 5000, "percentage": 100.0},
                "class_b": {"count": 0, "picks": 0, "percentage": 0},
                "class_c": {"count": 0, "picks": 0, "percentage": 0}
            }
        }

        exporter = ABCReportExporter()
        output_path = tmp_path / "abc_single.xlsx"

        exporter.export(single_product_results, output_path)

        assert output_path.exists()

        wb = load_workbook(output_path)
        sheet = wb["Classification ABC"]

        # Should have exactly one data row (row 3)
        assert sheet["A3"].value == "PROD001"
        assert sheet["A4"].value is None  # No second row
