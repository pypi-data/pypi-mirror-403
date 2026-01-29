"""
Data Analyzer Module for Multi-Task Optimization Experiments

This module provides a comprehensive analysis and visualization pipeline for
multi-task optimization experiments, including metric calculation, statistical
comparison tables, convergence plots, runtime analysis, and Pareto front visualization.

Classes:
    MetricResults: Dataclass for storing metric calculation results
    TableConfig: Dataclass for table generation configuration
    PlotConfig: Dataclass for plot generation configuration
    DataAnalyzer: Main class for data analysis pipeline

Author: Jiangtao Shen
Email: j.shen5@exeter.ac.uk
Date: 2025.10.10
Version: 2.1
"""

import os
import pickle
import shutil
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple, Union, Callable
from dataclasses import dataclass, field
from enum import Enum

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font
from scipy import stats
from tqdm import tqdm

# Import from project modules
from ddmtolab.Methods.metrics import IGD, HV, GD, IGDp, FR, CV, DeltaP, Spread, Spacing
from ddmtolab.Methods.Algo_Methods.algo_utils import nd_sort


# =============================================================================
# Enums and Constants
# =============================================================================

class OptimizationDirection(Enum):
    """Optimization direction enumeration."""
    MINIMIZE = "minimize"
    MAXIMIZE = "maximize"


class TableFormat(Enum):
    """Output table format enumeration."""
    EXCEL = "excel"
    LATEX = "latex"


class StatisticType(Enum):
    """Statistical measure type enumeration."""
    MEAN = "mean"
    MEDIAN = "median"
    MAX = "max"
    MIN = "min"


# Default color palette for plots
DEFAULT_COLORS = [
    '#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd',
    '#8c564b', '#e377c2', '#7f7f7f', '#bcbd22', '#17becf'
]

# Default markers for plots
DEFAULT_MARKERS = ['o', 's', '^', 'v', 'D', 'p', '*', 'h', '<', '>']


# =============================================================================
# Data Classes
# =============================================================================

@dataclass
class ScanResult:
    """
    Result of scanning a data directory.

    :no-index:

    Attributes
    ----------
    algorithms : List[str]
        Sorted list of algorithm names found in the directory.
    problems : List[str]
        Sorted list of problem names extracted from filenames.
    runs : int
        Number of independent runs per algorithm-problem combination.
    data_path : Path
        Path to the scanned data directory.
    """
    algorithms: List[str]
    problems: List[str]
    runs: int
    data_path: Path


@dataclass
class MetricResults:
    """
    Container for all metric calculation results.

    :no-index:

    Attributes
    ----------
    metric_values : Dict[str, Dict[str, Dict[int, List[np.ndarray]]]]
        Nested dictionary storing metric values per generation.
        Structure: metric_values[algorithm][problem][run] = List[np.ndarray]
        where each np.ndarray contains metric values per generation for each task.
    best_values : Dict[str, Dict[str, Dict[int, List[float]]]]
        Nested dictionary storing final best metric values.
        Structure: best_values[algorithm][problem][run] = List[float]
        where each float is the final best value for each task.
    objective_values : Dict[str, Dict[str, Dict[int, List[np.ndarray]]]]
        Nested dictionary storing original objective values.
        Structure: objective_values[algorithm][problem][run] = List[np.ndarray]
        where each np.ndarray has shape (n_solutions, n_objectives).
    runtime : Dict[str, Dict[str, Dict[int, float]]]
        Nested dictionary storing runtime in seconds.
        Structure: runtime[algorithm][problem][run] = float
    max_nfes : Dict[str, Dict[str, List[int]]]
        Nested dictionary storing maximum number of function evaluations.
        Structure: max_nfes[algorithm][problem] = List[int] (per task)
    metric_name : Optional[str]
        Name of the metric used (e.g., 'IGD', 'HV', or None for single-objective).
    """
    metric_values: Dict[str, Dict[str, Dict[int, Any]]]
    best_values: Dict[str, Dict[str, Dict[int, List[float]]]]
    objective_values: Dict[str, Dict[str, Dict[int, List[np.ndarray]]]]
    runtime: Dict[str, Dict[str, Dict[int, float]]]
    max_nfes: Dict[str, Dict[str, List[int]]]
    metric_name: Optional[str]


@dataclass
class TableConfig:
    """
    Configuration for table generation.

    :no-index:

    Attributes
    ----------
    table_format : TableFormat
        Output format (EXCEL or LATEX).
    statistic_type : StatisticType
        Type of statistic to display (MEAN, MEDIAN, MAX, MIN).
    significance_level : float
        P-value threshold for statistical significance testing.
        Default: 0.05
    rank_sum_test : bool
        Whether to perform Wilcoxon rank-sum test.
        Default: True
    save_path : Path
        Directory path to save output tables.
    """
    table_format: TableFormat = TableFormat.EXCEL
    statistic_type: StatisticType = StatisticType.MEAN
    significance_level: float = 0.05
    rank_sum_test: bool = True
    save_path: Path = Path('./Results')


@dataclass
class PlotConfig:
    """
    Configuration for plot generation.

    :no-index:

    Attributes
    ----------
    figure_format : str
        Output figure format (e.g., 'pdf', 'png', 'svg').
        Default: 'pdf'
    statistic_type : StatisticType
        Type of statistic for selecting representative run.
    log_scale : bool
        Whether to use logarithmic scale for y-axis.
        Default: False
    show_pf : bool
        Whether to show true Pareto front in ND solution plots.
        Default: True
    show_nd : bool
        Whether to filter and show only non-dominated solutions.
        Default: True
    save_path : Path
        Directory path to save output figures.
    colors : List[str]
        Color palette for plotting algorithms.
    markers : List[str]
        Marker styles for plotting algorithms.
    """
    figure_format: str = 'pdf'
    statistic_type: StatisticType = StatisticType.MEAN
    log_scale: bool = False
    show_pf: bool = True
    show_nd: bool = True
    save_path: Path = Path('./Results')
    colors: List[str] = field(default_factory=lambda: DEFAULT_COLORS.copy())
    markers: List[str] = field(default_factory=lambda: DEFAULT_MARKERS.copy())


@dataclass
class ComparisonResult:
    """
    Result of statistical comparison between algorithms.

    :no-index:

    Attributes
    ----------
    symbol : str
        Comparison symbol: '+' (better), '-' (worse), '=' (no significant difference).
    p_value : Optional[float]
        P-value from statistical test, or None if test not performed.
    """
    symbol: str
    p_value: Optional[float] = None


@dataclass
class ComparisonCounts:
    """
    Aggregated comparison counts for an algorithm.

    :no-index:

    Attributes
    ----------
    plus : int
        Number of significantly better results.
    minus : int
        Number of significantly worse results.
    equal : int
        Number of statistically equivalent results.
    """
    plus: int = 0
    minus: int = 0
    equal: int = 0


# =============================================================================
# Utility Functions
# =============================================================================

class DataUtils:
    """
    Utility class for data loading and processing operations.
    """

    @staticmethod
    def load_pickle(file_path: Path) -> Dict[str, Any]:
        """
        Load and return a Python object from a pickle file.

        Parameters
        ----------
        file_path : Path
            Path to the pickle file.

        Returns
        -------
        Dict[str, Any]
            Unpickled Python object (typically a dictionary containing
            'all_objs', 'runtime', 'max_nfes' keys).

        Raises
        ------
        FileNotFoundError
            If the pickle file does not exist.
        pickle.UnpicklingError
            If the file cannot be unpickled.
        """
        with open(file_path, 'rb') as f:
            return pickle.load(f)

    @staticmethod
    def load_reference(
            settings: Dict[str, Any],
            problem: str,
            task_identifier: Union[str, int],
            M: int,
            D: Optional[int] = None,
            C: int = 0
    ) -> Optional[np.ndarray]:
        """
        Load reference data (Pareto Front or reference point) for a specific problem and task.

        Parameters
        ----------
        settings : Dict[str, Any]
            Dictionary containing problem configurations and reference definitions.
            Expected keys:

            - problem (str): Contains task definitions
            - 'n_ref' (int, optional): Number of reference points (default: 10000)
            - 'ref_path' (str, optional): Path to reference files (default: './MOReference')

        problem : str
            Name of the problem (e.g., "DTLZ1", "DTLZ2").
        task_identifier : Union[str, int]
            Task identifier - either task name (str like "T1") or index (int like 0).
        M : int
            Number of objectives (required).
        D : int, optional
            Number of decision variables (dimension).
        C : int, optional
            Number of constraints (default: 0).

        Returns
        -------
        Optional[np.ndarray]
            Reference data with shape (n_points, M), or None if not available.

        Notes
        -----
        Supports three types of reference definitions:

        1. Callable: Function that returns reference data

           - Must accept parameter N (number of reference points)
           - Must accept parameter M (number of objectives)
           - May optionally accept parameters D, C
           - Example signatures: ``func(N, M)``, ``func(N, M, D)``, ``func(N, M, D, C)``

        2. String: File path to .npy or .csv reference file
        3. Array-like: Direct reference data (list, tuple, np.ndarray)

        If 'all_tasks' key is present instead of individual task keys, the same
        reference data will be used for all tasks.
        """
        # Convert task index to task name if necessary
        task_name = f"T{task_identifier + 1}" if isinstance(task_identifier, int) else task_identifier

        # Check if problem exists in settings
        if problem not in settings:
            print(f"Warning: Problem '{problem}' not found in settings")
            return None

        problem_settings = settings[problem]

        # Check if task exists for this problem
        if task_name in problem_settings:
            ref_definition = problem_settings[task_name]
        elif 'all_tasks' in problem_settings:
            # Use the same reference for all tasks
            ref_definition = problem_settings['all_tasks']
        else:
            print(f"Warning: Task '{task_name}' and 'all_tasks' not found for problem '{problem}'")
            return None

        # Case 1: Callable function
        if callable(ref_definition):
            N = settings.get('n_ref', 10000)

            try:
                import inspect
                sig = inspect.signature(ref_definition)
                params = list(sig.parameters.keys())
                num_params = len(params)

                if num_params == 2:
                    # func(N, M)
                    return ref_definition(N, M)
                elif num_params == 3:
                    # func(N, M, D)
                    if D is None:
                        print(f"Warning: D not provided for {problem}_{task_name}, using 0")
                        D = 0
                    return ref_definition(N, M, D)
                elif num_params >= 4:
                    # func(N, M, D, C)
                    if D is None:
                        print(f"Warning: D not provided for {problem}_{task_name}, using 0")
                        D = 0
                    return ref_definition(N, M, D, C)
                else:
                    print(
                        f"Warning: Unexpected number of parameters ({num_params}) for reference function {problem}_{task_name}")
                    return None

            except Exception as e:
                print(f"Warning: Failed to call reference function for {problem}_{task_name}: {e}")
                return None

        # Case 2: String (file path or file name)
        elif isinstance(ref_definition, str):
            return DataUtils._load_reference_from_file(
                settings,
                ref_definition,
                problem,
                task_name
            )

        # Case 3: Array-like (list, tuple, numpy array)
        elif isinstance(ref_definition, (list, tuple, np.ndarray)):
            reference = np.array(ref_definition)
            # Ensure it's at least 2D
            if reference.ndim == 1:
                reference = reference.reshape(1, -1)
            return reference

        else:
            print(f"Warning: Unknown reference definition type for {problem}_{task_name}: {type(ref_definition)}")
            return None

    @staticmethod
    def _load_reference_from_file(
            settings: Dict[str, Any],
            ref_definition: str,
            problem: str,
            task_name: str
    ) -> Optional[np.ndarray]:
        """
        Load reference data from file.

        Parameters
        ----------
        settings : Dict[str, Any]
            Settings dictionary containing 'ref_path'.
        ref_definition : str
            File path or filename.
        problem : str
            Problem name for alternative path construction.
        task_name : str
            Task name for alternative path construction.

        Returns
        -------
        Optional[np.ndarray]
            Loaded reference data or None if loading fails.
        """
        ref_path = settings.get('ref_path', './MOReference')

        # Construct full path
        if not os.path.isabs(ref_definition):
            full_path = os.path.join(ref_path, ref_definition)
        else:
            full_path = ref_definition

        # Try to load the file
        try:
            if full_path.endswith('.npy'):
                return np.load(full_path)
            elif full_path.endswith('.csv'):
                return np.loadtxt(full_path, delimiter=',')
            else:
                print(f"Warning: Unsupported file format for '{full_path}'")
                return None
        except FileNotFoundError:
            # Try alternative naming conventions
            base_name = f"{problem}_{task_name}_ref"

            for ext in ['.npy', '.csv']:
                alt_path = os.path.join(ref_path, base_name + ext)
                if os.path.exists(alt_path):
                    try:
                        if ext == '.npy':
                            return np.load(alt_path)
                        else:
                            return np.loadtxt(alt_path, delimiter=',')
                    except Exception as e:
                        print(f"Error loading file '{alt_path}': {e}")

            print(f"Warning: File not found: '{full_path}'")
            return None
        except Exception as e:
            print(f"Error loading reference data from file '{full_path}': {e}")
            return None

    @staticmethod
    def get_metric_direction(metric_name: Optional[str]) -> OptimizationDirection:
        """
        Determine optimization direction based on metric type (Version 2 - More maintainable).

        Parameters
        ----------
        metric_name : Optional[str]
            Name of the metric or None for single-objective.

        Returns
        -------
        OptimizationDirection
            MINIMIZE or MAXIMIZE based on the metric's sign attribute.
        """
        if metric_name is None:
            return OptimizationDirection.MINIMIZE

        # Metric sign mapping (based on your code)
        # sign = -1 means minimize, sign = 1 means maximize
        metric_signs = {
            'IGD': -1,  # Inverted Generational Distance (minimize)
            'HV': 1,  # Hypervolume (maximize)
            'IGDp': -1,  # IGD+ (minimize)
            'GD': -1,  # Generational Distance (minimize)
            'DeltaP': -1,  # Delta_p (minimize)
            'Spacing': -1,  # Spacing (minimize)
            'Spread': -1,  # Spread (minimize)
            'FR': 1,  # Feasibility Rate (maximize)
            'CV': -1,  # Constraint Violation (minimize)
        }

        if metric_name not in metric_signs:
            raise ValueError(f'Unsupported metric: {metric_name}')

        sign = metric_signs[metric_name]
        return OptimizationDirection.MAXIMIZE if sign == 1 else OptimizationDirection.MINIMIZE


# =============================================================================
# Statistics Module
# =============================================================================

class StatisticsCalculator:
    """
    Class for statistical calculations and hypothesis testing.
    """

    @staticmethod
    def calculate_statistic(
            data: List[float],
            statistic_type: StatisticType
    ) -> Tuple[float, Optional[float]]:
        """
        Calculate a statistical measure and optional standard deviation from data.

        Parameters
        ----------
        data : List[float]
            List of numeric values to compute statistics from.
        statistic_type : StatisticType
            Type of statistic to calculate (MEAN, MEDIAN, MAX, MIN).

        Returns
        -------
        Tuple[float, Optional[float]]
            Tuple of (statistic_value, std_value).
            std_value is only returned for MEAN, None otherwise.
            Returns (np.nan, np.nan) for empty data.
        """
        if len(data) == 0:
            return np.nan, np.nan

        if statistic_type == StatisticType.MEAN:
            stat_value = np.mean(data)
            std_value = np.std(data, ddof=1) if len(data) > 1 else 0.0
            return stat_value, std_value
        elif statistic_type == StatisticType.MEDIAN:
            return np.median(data), None
        elif statistic_type == StatisticType.MAX:
            return np.max(data), None
        elif statistic_type == StatisticType.MIN:
            return np.min(data), None
        else:
            return np.nan, np.nan

    @staticmethod
    def perform_rank_sum_test(
            algo_data: List[float],
            base_data: List[float],
            significance_level: float = 0.05,
            direction: OptimizationDirection = OptimizationDirection.MINIMIZE
    ) -> ComparisonResult:
        """
        Perform Wilcoxon rank-sum test to compare two algorithms.

        Parameters
        ----------
        algo_data : List[float]
            Data from the algorithm being tested.
        base_data : List[float]
            Data from the baseline algorithm.
        significance_level : float, optional
            P-value threshold for significance (default: 0.05).
        direction : OptimizationDirection, optional
            Optimization direction (MINIMIZE or MAXIMIZE).

        Returns
        -------
        ComparisonResult
            Result containing comparison symbol and p-value.
            Symbol: '+' (better), '-' (worse), '=' (no significant difference).
        """
        if len(algo_data) == 0 or len(base_data) == 0:
            return ComparisonResult(symbol='=', p_value=None)

        try:
            _, p_value = stats.ranksums(algo_data, base_data)

            if p_value < significance_level:
                algo_median = np.median(algo_data)
                base_median = np.median(base_data)

                if direction == OptimizationDirection.MINIMIZE:
                    symbol = '+' if algo_median < base_median else '-'
                else:
                    symbol = '+' if algo_median > base_median else '-'
            else:
                symbol = '='

            return ComparisonResult(symbol=symbol, p_value=p_value)
        except Exception:
            return ComparisonResult(symbol='=', p_value=None)

    @staticmethod
    def collect_task_data(
            all_best_values: Dict[str, Dict[str, Dict[int, List[float]]]],
            algo: str,
            prob: str,
            task_idx: int
    ) -> List[float]:
        """
        Collect non-NaN values from all runs for a specific algorithm-problem-task combination.

        Parameters
        ----------
        all_best_values : Dict[str, Dict[str, Dict[int, List[float]]]]
            Nested dictionary containing best metric values.
        algo : str
            Algorithm name.
        prob : str
            Problem name.
        task_idx : int
            Task index (0-based).

        Returns
        -------
        List[float]
            List of non-NaN metric values from all runs.
        """
        data = []
        for run in all_best_values[algo][prob].keys():
            value = all_best_values[algo][prob][run][task_idx]
            if not np.isnan(value):
                data.append(value)
        return data

    @staticmethod
    def select_representative_run(
            all_best_values: Dict[str, Dict[str, Dict[int, List[float]]]],
            algo: str,
            prob: str,
            task_idx: int,
            statistic_type: StatisticType
    ) -> Optional[int]:
        """
        Select a representative run based on the specified statistic type.

        Parameters
        ----------
        all_best_values : Dict[str, Dict[str, Dict[int, List[float]]]]
            Nested dictionary containing best metric values.
        algo : str
            Algorithm name.
        prob : str
            Problem name.
        task_idx : int
            Task index (0-based).
        statistic_type : StatisticType
            Type of statistic (MEAN returns None as all runs are used).

        Returns
        -------
        Optional[int]
            Run number of the representative run, or None if MEAN or no valid data.
        """
        if statistic_type == StatisticType.MEAN:
            return None

        # Collect final values from all runs
        final_values = []
        runs = []

        for run in all_best_values[algo][prob].keys():
            value = all_best_values[algo][prob][run][task_idx]
            if not np.isnan(value):
                final_values.append(value)
                runs.append(run)

        if len(final_values) == 0:
            return None

        final_values = np.array(final_values)
        runs = np.array(runs)

        if statistic_type == StatisticType.MEDIAN:
            target_value = np.median(final_values)
            idx = np.argmin(np.abs(final_values - target_value))
        elif statistic_type == StatisticType.MAX:
            idx = np.argmax(final_values)
        elif statistic_type == StatisticType.MIN:
            idx = np.argmin(final_values)
        else:
            return None

        return runs[idx]


# =============================================================================
# Table Generator Module
# =============================================================================

class TableGenerator:
    """
    Class for generating comparison tables in Excel and LaTeX formats.
    """

    def __init__(self, config: TableConfig):
        """
        Initialize TableGenerator with configuration.

        Parameters
        ----------
        config : TableConfig
            Configuration object for table generation.
        """
        self.config = config

    def generate(
            self,
            all_best_values: Dict[str, Dict[str, Dict[int, List[float]]]],
            algorithm_order: List[str],
            metric_name: Optional[str] = None
    ) -> Union[pd.DataFrame, str]:
        """
        Generate comparison table with statistical analysis.

        Parameters
        ----------
        all_best_values : Dict[str, Dict[str, Dict[int, List[float]]]]
            Nested dictionary containing best metric values.
            Structure: all_best_values[algorithm][problem][run] = List[float]
        algorithm_order : List[str]
            List of algorithm names in display order.
            The last algorithm is treated as the baseline for comparisons.
        metric_name : Optional[str], optional
            Metric name to determine optimization direction.

        Returns
        -------
        Union[pd.DataFrame, str]
            DataFrame for Excel format, LaTeX string for LaTeX format.
        """
        # Extract problems and determine task count
        # problems = sorted(all_best_values[algorithm_order[0]].keys())
        problems = sorted(all_best_values[algorithm_order[0]].keys(),
                          key=lambda x: int(''.join(filter(str.isdigit, x))) if any(c.isdigit() for c in x) else x)

        # Determine optimization direction
        direction = DataUtils.get_metric_direction(metric_name)

        # Generate data rows
        rows, comparison_counts = self._generate_data_rows(all_best_values, algorithm_order, problems, direction)

        # Generate and save table
        if self.config.table_format == TableFormat.EXCEL:
            return self._generate_excel_table(rows, algorithm_order, comparison_counts, direction)
        else:
            return self._generate_latex_table(rows, algorithm_order, comparison_counts, direction)

    def _generate_data_rows(
            self,
            all_best_values: Dict[str, Dict[str, Dict[int, List[float]]]],
            algorithm_order: List[str],
            problems: List[str],
            direction: OptimizationDirection
    ) -> Tuple[List[Dict[str, Any]], Dict[str, ComparisonCounts]]:
        """
        Generate table rows with formatted metric values and statistical comparisons.

        Parameters
        ----------
        all_best_values : Dict
            Best metric values dictionary.
        algorithm_order : List[str]
            Algorithm display order.
        problems : List[str]
            List of problem names.
        direction : OptimizationDirection
            Optimization direction.

        Returns
        -------
        Tuple[List[Dict[str, Any]], Dict[str, ComparisonCounts]]
            Tuple of (rows, comparison_counts).
        """
        base_algo = algorithm_order[-1]
        rows = []
        comparison_counts = {algo: ComparisonCounts() for algo in algorithm_order[:-1]}

        for prob in problems:
            # Get the number of tasks of the problem
            first_algo = algorithm_order[0]
            first_run = list(all_best_values[first_algo][prob].keys())[0]
            num_tasks = len(all_best_values[first_algo][prob][first_run])

            for task_idx in range(num_tasks):
                row = {'Problem': prob, 'Task': task_idx + 1}

                # Collect baseline data
                base_data = StatisticsCalculator.collect_task_data(
                    all_best_values, base_algo, prob, task_idx
                )

                for algo in algorithm_order:
                    algo_data = StatisticsCalculator.collect_task_data(
                        all_best_values, algo, prob, task_idx
                    )

                    stat_value, std_value = StatisticsCalculator.calculate_statistic(
                        algo_data, self.config.statistic_type
                    )

                    symbol = ''
                    if self.config.rank_sum_test and algo != base_algo:
                        result = StatisticsCalculator.perform_rank_sum_test(
                            algo_data, base_data,
                            self.config.significance_level, direction
                        )
                        symbol = result.symbol

                        # Update comparison counts
                        if algo in comparison_counts:
                            if symbol == '+':
                                comparison_counts[algo].plus += 1
                            elif symbol == '-':
                                comparison_counts[algo].minus += 1
                            else:
                                comparison_counts[algo].equal += 1

                    cell_content = self._format_cell_content(stat_value, std_value, symbol)
                    row[algo] = cell_content

                rows.append(row)

        return rows, comparison_counts

    def _format_cell_content(
            self,
            stat_value: float,
            std_value: Optional[float],
            symbol: str
    ) -> str:
        """
        Format a table cell with statistic value, optional std deviation, and comparison symbol.

        Parameters
        ----------
        stat_value : float
            Statistical value.
        std_value : Optional[float]
            Standard deviation (or None).
        symbol : str
            Comparison symbol.

        Returns
        -------
        str
            Formatted cell content string.
        """
        if np.isnan(stat_value):
            return 'N/A'

        if self.config.table_format == TableFormat.EXCEL:
            if self.config.statistic_type == StatisticType.MEAN:
                cell_content = f"{stat_value:.4e}({std_value:.2e})"
            else:
                cell_content = f"{stat_value:.4e}"

            if symbol:
                cell_content += f" {symbol}"
        else:
            # LaTeX format
            if self.config.statistic_type == StatisticType.MEAN:
                stat_str = f"{stat_value:.4e}".replace('e-', 'e$-$')
                std_str = f"{std_value:.2e}".replace('e-', 'e$-$')
                cell_content = f"{stat_str}({std_str})"
            else:
                stat_str = f"{stat_value:.4e}".replace('e-', 'e$-$')
                cell_content = stat_str

            if symbol:
                symbol_map = {'+': '~$+$', '-': '~$-$', '=': '~='}
                cell_content += symbol_map.get(symbol, '')

        return cell_content

    def _find_best_value_in_row(
            self,
            row: Dict[str, Any],
            algorithm_order: List[str],
            direction: OptimizationDirection
    ) -> Optional[str]:
        """
        Find the algorithm with the best performance in a table row.

        Parameters
        ----------
        row : Dict[str, Any]
            Dictionary mapping algorithm names to formatted cell values.
        algorithm_order : List[str]
            List of algorithm names.
        direction : OptimizationDirection
            Optimization direction.

        Returns
        -------
        Optional[str]
            Name of the best-performing algorithm or None.
        """
        best_val = None
        best_algo = None

        for algo in algorithm_order:
            cell = row[algo]
            if cell != 'N/A':
                try:
                    val_str = cell.split('(')[0].replace('e$-$', 'e-')
                    val = float(val_str)

                    if direction == OptimizationDirection.MINIMIZE:
                        if best_val is None or val < best_val:
                            best_val = val
                            best_algo = algo
                    else:
                        if best_val is None or val > best_val:
                            best_val = val
                            best_algo = algo
                except Exception:
                    pass

        return best_algo

    def _generate_excel_table(
            self,
            rows: List[Dict[str, Any]],
            algorithm_order: List[str],
            comparison_counts: Dict[str, ComparisonCounts],
            direction: OptimizationDirection
    ) -> pd.DataFrame:
        """
        Generate and save a formatted Excel table.

        Parameters
        ----------
        rows : List[Dict[str, Any]]
            Table row data.
        algorithm_order : List[str]
            Algorithm display order.
        comparison_counts : Dict[str, ComparisonCounts]
            Comparison result counts.
        direction : OptimizationDirection
            Optimization direction.

        Returns
        -------
        pd.DataFrame
            DataFrame containing the table data.
        """
        # Append summary row
        if self.config.rank_sum_test:
            summary_row = {'Problem': '+/-/=', 'Task': ''}
            for algo in algorithm_order[:-1]:
                counts = comparison_counts[algo]
                summary_row[algo] = f"{counts.plus}/{counts.minus}/{counts.equal}"
            summary_row[algorithm_order[-1]] = 'Base'
            rows.append(summary_row)

        # Create DataFrame
        df = pd.DataFrame(rows)
        columns = ['Problem', 'Task'] + algorithm_order
        df = df[columns]

        # Save and format
        save_dir = Path(self.config.save_path)
        save_dir.mkdir(parents=True, exist_ok=True)
        output_file = save_dir / f'results_table_{self.config.statistic_type.value}.xlsx'
        df.to_excel(output_file, index=False)

        # Apply Excel formatting
        self._apply_excel_formatting(output_file, df, algorithm_order, direction)

        print(f"Excel table saved to: {output_file}")
        return df

    def _apply_excel_formatting(
            self,
            output_file: Path,
            df: pd.DataFrame,
            algorithm_order: List[str],
            direction: OptimizationDirection
    ) -> None:
        """
        Apply formatting to Excel workbook.

        Parameters
        ----------
        output_file : Path
            Path to Excel file.
        df : pd.DataFrame
            DataFrame for row count reference.
        algorithm_order : List[str]
            Algorithm names.
        direction : OptimizationDirection
            Optimization direction for best value highlighting.
        """
        wb = load_workbook(output_file)
        ws = wb.active

        # Define styles
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        normal_font = Font(name='Times New Roman', size=11)
        bold_font = Font(name='Times New Roman', size=11, bold=True)

        # Apply formatting and auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = normal_font

                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except Exception:
                    pass

            ws.column_dimensions[column_letter].width = max_length + 2

        # Bold the best value in each data row
        num_data_rows = len(df) - (1 if self.config.rank_sum_test else 0)

        for row_idx in range(2, num_data_rows + 2):
            best_val = None
            best_col = None

            for col_idx, algo in enumerate(algorithm_order, start=3):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell_value = cell.value

                if cell_value and cell_value != 'N/A':
                    try:
                        val_str = str(cell_value).split('(')[0].strip()
                        val = float(val_str)

                        if direction == OptimizationDirection.MINIMIZE:
                            if best_val is None or val < best_val:
                                best_val = val
                                best_col = col_idx
                        else:
                            if best_val is None or val > best_val:
                                best_val = val
                                best_col = col_idx
                    except Exception:
                        pass

            if best_col is not None:
                ws.cell(row=row_idx, column=best_col).font = bold_font

        wb.save(output_file)

    def _generate_latex_table(
            self,
            rows: List[Dict[str, Any]],
            algorithm_order: List[str],
            comparison_counts: Dict[str, ComparisonCounts],
            direction: OptimizationDirection
    ) -> str:
        """
        Generate and save a LaTeX-formatted table.

        Parameters
        ----------
        rows : List[Dict[str, Any]]
            Table row data.
        algorithm_order : List[str]
            Algorithm display order.
        comparison_counts : Dict[str, ComparisonCounts]
            Comparison result counts.
        direction : OptimizationDirection
            Optimization direction.

        Returns
        -------
        str
            LaTeX table string.
        """
        df = pd.DataFrame(rows)

        # Build table structure
        num_cols = len(algorithm_order) + 2
        col_format = '|'.join(['c'] * num_cols)
        col_format = '|' + col_format + '|'

        # Initialize LaTeX table
        latex_str = "\\begin{table*}[htbp]\n"
        latex_str += "\\renewcommand{\\arraystretch}{1.2}\n"
        latex_str += "\\centering\n"
        latex_str += "\\caption{Your caption here}\n"
        latex_str += "\\label{tab:results}\n"
        latex_str += "\\resizebox{1.0\\textwidth}{!}{\n"
        latex_str += f"\\begin{{tabular}}{{{col_format}}}\n"
        latex_str += "\\hline\n"

        # Header row
        header = "Problem & Task & " + " & ".join(algorithm_order) + " \\\\\n"
        latex_str += header
        latex_str += "\\hline\n"

        # Data rows
        for _, row in df.iterrows():
            best_algo = self._find_best_value_in_row(row, algorithm_order, direction)

            row_str = f"{row['Problem']} & {row['Task']}"
            for algo in algorithm_order:
                cell = row[algo]
                if algo == best_algo:
                    cell = f"\\textbf{{{cell}}}"
                row_str += f" & {cell}"
            row_str += " \\\\\n"
            latex_str += row_str
            latex_str += "\\hline\n"

        # Summary row
        if self.config.rank_sum_test:
            summary_str = "\\multicolumn{2}{|c|}{+/$-$/=}"
            for algo in algorithm_order[:-1]:
                counts = comparison_counts[algo]
                summary_str += f" & {counts.plus}/{counts.minus}/{counts.equal}"
            summary_str += " & Base \\\\\n"
            latex_str += summary_str
            latex_str += "\\hline\n"

        latex_str += "\\end{tabular}}\n"
        latex_str += "\\end{table*}\n"

        # Save to file
        save_dir = Path(self.config.save_path)
        save_dir.mkdir(parents=True, exist_ok=True)
        output_file = save_dir / f'results_table_{self.config.statistic_type.value}.tex'
        with open(output_file, 'w') as f:
            f.write(latex_str)
        print(f"LaTeX table saved to: {output_file}")

        return latex_str


# =============================================================================
# Plot Generator Module
# =============================================================================

class PlotGenerator:
    """
    Class for generating various visualization plots.
    """

    def __init__(self, config: PlotConfig):
        """
        Initialize PlotGenerator with configuration.

        Parameters
        ----------
        config : PlotConfig
            Configuration object for plot generation.
        """
        self.config = config

    def plot_convergence_curves(
            self,
            metric_values: Dict[str, Dict[str, Dict[int, Any]]],
            best_values: Dict[str, Dict[str, Dict[int, List[float]]]],
            max_nfes: Dict[str, Dict[str, List[int]]],
            algorithm_order: List[str],
            metric_name: Optional[str] = None
    ) -> None:
        """
        Generate and save convergence curve plots for all algorithms, problems, and tasks.

        Parameters
        ----------
        metric_values : Dict[str, Dict[str, Dict[int, Any]]]
            Metric values per generation.
            Structure: metric_values[algorithm][problem][run] = List[np.ndarray]
        best_values : Dict[str, Dict[str, Dict[int, List[float]]]]
            Best metric values for representative run selection.
        max_nfes : Dict[str, Dict[str, List[int]]]
            Maximum number of function evaluations per task.
            Structure: max_nfes[algorithm][problem] = List[int]
        algorithm_order : List[str]
            List of algorithm names to plot.
        metric_name : Optional[str], optional
            Metric name for y-axis label.

        Returns
        -------
        None
            Saves figures to disk.
        """
        problems = sorted(metric_values[algorithm_order[0]].keys())
        save_dir = Path(self.config.save_path)
        save_dir.mkdir(parents=True, exist_ok=True)

        for prob in problems:
            first_run_data = best_values[algorithm_order[0]][prob][1]
            num_tasks = len(first_run_data)

            for task_idx in range(num_tasks):
                fig = self._create_convergence_figure(
                    num_tasks, metric_values, best_values, max_nfes,
                    algorithm_order, prob, task_idx, metric_name
                )

                if num_tasks == 1:
                    output_file = save_dir / f'{prob}.{self.config.figure_format}'
                else:
                    output_file = save_dir / f'{prob}-Task{task_idx + 1}.{self.config.figure_format}'

                fig.savefig(output_file, dpi=300, bbox_inches='tight')
                plt.close(fig)

        print(f"All convergence plots saved to: {save_dir}")

    def _create_convergence_figure(
            self,
            num_tasks: int,
            metric_values: Dict,
            best_values: Dict,
            max_nfes: Dict,
            algorithm_order: List[str],
            prob: str,
            task_idx: int,
            metric_name: Optional[str]
    ) -> plt.Figure:
        """
        Create a single convergence curve figure.

        Parameters
        ----------
        num_tasks : int
            Total number of tasks.
        metric_values : Dict
            Metric values dictionary.
        best_values : Dict
            Best values dictionary.
        max_nfes : Dict
            Max NFEs dictionary.
        algorithm_order : List[str]
            Algorithm order.
        prob : str
            Problem name.
        task_idx : int
            Task index.
        metric_name : Optional[str]
            Metric name for label.

        Returns
        -------
        plt.Figure
            Matplotlib figure object.
        """
        fig, ax = plt.subplots(figsize=(5, 3.5))

        for idx, algo in enumerate(algorithm_order):
            selected_run = StatisticsCalculator.select_representative_run(
                best_values, algo, prob, task_idx, self.config.statistic_type
            )

            curve = self._get_convergence_curve(metric_values, algo, prob, task_idx, selected_run)

            if len(curve) == 0:
                continue

            nfes = max_nfes[algo][prob][task_idx]
            x = np.linspace(0, nfes, len(curve))
            marker_interval = max(1, len(curve) // 10)

            ax.plot(
                x, curve, label=algo,
                color=self.config.colors[idx % len(self.config.colors)],
                marker=self.config.markers[idx % len(self.config.markers)],
                markevery=marker_interval,
                markersize=5, linewidth=1.5, linestyle='-', alpha=0.8
            )

        if self.config.log_scale:
            ax.set_yscale('log')

        # Apply scientific notation for large values
        self._apply_scientific_notation(ax)

        y_label = metric_name if metric_name is not None else 'Objective Value'
        ax.set_xlabel('NFEs', fontsize=10)
        ax.set_ylabel(y_label, fontsize=10)

        title = f'{prob}' if num_tasks == 1 else f'{prob} - Task {task_idx + 1}'
        ax.set_title(title, fontsize=10)
        ax.tick_params(axis='both', which='major', labelsize=10)
        ax.legend(loc='best', fontsize=10)
        ax.grid(True, alpha=0.2, linestyle='-')

        fig.tight_layout()
        return fig

    def _get_convergence_curve(
            self,
            metric_values: Dict,
            algo: str,
            prob: str,
            task_idx: int,
            run: Optional[int]
    ) -> np.ndarray:
        """
        Extract convergence curve for a specific configuration.

        Parameters
        ----------
        metric_values : Dict
            Metric values dictionary.
        algo : str
            Algorithm name.
        prob : str
            Problem name.
        task_idx : int
            Task index.
        run : Optional[int]
            Specific run number (None for mean across runs).

        Returns
        -------
        np.ndarray
            Convergence curve values.
        """
        if run is not None:
            return np.array(metric_values[algo][prob][run][task_idx])
        else:
            all_curves = []
            for r in metric_values[algo][prob].keys():
                curve = np.array(metric_values[algo][prob][r][task_idx])
                if len(curve) > 0:
                    all_curves.append(curve)

            if len(all_curves) == 0:
                return np.array([])

            min_len = min(len(c) for c in all_curves)
            truncated_curves = [c[:min_len] for c in all_curves]
            return np.mean(truncated_curves, axis=0)

    def _apply_scientific_notation(self, ax: plt.Axes) -> None:
        """
        Apply scientific notation to axes if values exceed threshold.

        Parameters
        ----------
        ax : plt.Axes
            Matplotlib axes object.
        """
        threshold = 1000
        xmax = ax.get_xlim()[1]
        ymax = ax.get_ylim()[1]

        for axis_name, axis, lim in [('x', ax.xaxis, xmax), ('y', ax.yaxis, ymax)]:
            if lim > threshold:
                formatter = plt.matplotlib.ticker.ScalarFormatter(useMathText=True)
                formatter.set_powerlimits((0, 0))
                axis.set_major_formatter(formatter)
                ax.ticklabel_format(style='sci', axis=axis_name, scilimits=(0, 0))

    def plot_runtime(
            self,
            runtime: Dict[str, Dict[str, Dict[int, float]]],
            algorithm_order: List[str]
    ) -> None:
        """
        Generate and save a bar plot showing average runtime comparison.

        Parameters
        ----------
        runtime : Dict[str, Dict[str, Dict[int, float]]]
            Runtime dictionary.
            Structure: runtime[algorithm][problem][run] = float (seconds)
        algorithm_order : List[str]
            List of algorithm names in display order.

        Returns
        -------
        None
            Saves figure to disk.
        """
        problems = sorted(runtime[algorithm_order[0]].keys(),
                          key=lambda x: int(''.join(filter(str.isdigit, x))) if any(c.isdigit() for c in x) else x)
        save_dir = Path(self.config.save_path)
        save_dir.mkdir(parents=True, exist_ok=True)

        fig, ax = plt.subplots(figsize=(6, 3.5))

        n_algorithms = len(algorithm_order)
        n_problems = len(problems)
        bar_width = 0.8 / n_algorithms
        x_groups = np.arange(n_problems)

        for idx, algo in enumerate(algorithm_order):
            means = []
            stds = []

            for prob in problems:
                runtimes = [runtime[algo][prob][run] for run in runtime[algo][prob].keys()]
                means.append(np.mean(runtimes))

                # Only calculate std if there are at least 2 data points
                if len(runtimes) > 1:
                    stds.append(np.std(runtimes, ddof=1))
                else:
                    stds.append(0.0)  # No error bar for single data point

            x_offset = x_groups + (idx - n_algorithms / 2 + 0.5) * bar_width

            ax.bar(
                x_offset, means, bar_width,
                yerr=stds, label=algo,
                color=self.config.colors[idx % len(self.config.colors)],
                alpha=0.8, capsize=4,
                error_kw={'linewidth': 1.2, 'ecolor': 'black', 'alpha': 0.6}
            )

        ax.set_ylabel('Runtime (s)', fontsize=12)
        ax.set_xticks(x_groups)
        ax.set_xticklabels(problems, fontsize=12)
        ax.tick_params(axis='both', which='major', labelsize=10)
        ax.legend(loc='best', fontsize=10, framealpha=0.7)
        ax.grid(True, axis='y', alpha=0.3, linestyle='-')

        fig.tight_layout()

        output_file = save_dir / f'runtime_comparison.{self.config.figure_format}'
        fig.savefig(output_file, dpi=300, bbox_inches='tight')
        plt.close(fig)

        print(f"Runtime plot saved to: {output_file}")

    def plot_nd_solutions(
            self,
            best_values: Dict[str, Dict[str, Dict[int, List[float]]]],
            objective_values: Dict[str, Dict[str, Dict[int, List[np.ndarray]]]],
            algorithm_order: List[str],
            settings: Optional[Dict[str, Any]] = None
    ) -> None:
        """
        Generate and save non-dominated solution plots.

        Parameters
        ----------
        best_values : Dict[str, Dict[str, Dict[int, List[float]]]]
            Best values for representative run selection.
        objective_values : Dict[str, Dict[str, Dict[int, List[np.ndarray]]]]
            Original objective values.
            Structure: objective_values[algorithm][problem][run] = List[np.ndarray]
            where each np.ndarray has shape (n_solutions, n_objectives).
        algorithm_order : List[str]
            List of algorithm names.
        settings : Optional[Dict[str, Any]], optional
            Problem settings for loading true Pareto fronts.

        Returns
        -------
        None
            Saves figures to disk.
        """
        nd_folder = Path(self.config.save_path) / 'ND_Solutions'
        nd_folder.mkdir(parents=True, exist_ok=True)

        problems = list(objective_values[algorithm_order[0]].keys())

        for algo in algorithm_order:
            for prob in problems:
                first_run = list(objective_values[algo][prob].keys())[0]
                n_tasks = len(objective_values[algo][prob][first_run])

                for task_idx in range(n_tasks):
                    first_run_objs = objective_values[algo][prob][first_run][task_idx]
                    n_objectives = first_run_objs.shape[1]

                    if n_objectives <= 1:
                        continue

                    selected_run = StatisticsCalculator.select_representative_run(
                        best_values, algo, prob, task_idx, self.config.statistic_type
                    )

                    if selected_run is None:
                        selected_run = 1

                    objectives = objective_values[algo][prob][selected_run][task_idx]

                    if objectives.shape[0] == 0:
                        continue

                    # Filter non-dominated solutions if requested
                    if self.config.show_nd:
                        front_no, _ = nd_sort(objectives, objectives.shape[0])
                        nd_solutions = objectives[front_no == 1]
                    else:
                        nd_solutions = objectives

                    # Load true Pareto front if requested
                    true_pf = None
                    if self.config.show_pf and settings is not None:
                        true_pf = DataUtils.load_reference(settings, prob, task_idx, M=n_objectives)

                    # Create appropriate plot based on number of objectives
                    fig = self._create_nd_plot(nd_solutions, true_pf, n_objectives, n_tasks, prob, task_idx, algo)

                    # Save figure
                    if n_tasks == 1:
                        filename = f'{prob}-{algo}.{self.config.figure_format}'
                    else:
                        filename = f'{prob}-Task{task_idx + 1}-{algo}.{self.config.figure_format}'

                    fig.savefig(nd_folder / filename, dpi=300)
                    plt.close(fig)

        print(f"All non-dominated solutions plots saved to: {nd_folder}\n")

    def _create_nd_plot(
            self,
            nd_solutions: np.ndarray,
            true_pf: Optional[np.ndarray],
            n_objectives: int,
            n_tasks: int,
            prob: str,
            task_idx: int,
            algo: str
    ) -> plt.Figure:
        """
        Create a non-dominated solution plot.

        Parameters
        ----------
        nd_solutions : np.ndarray
            Non-dominated solutions array with shape (n_solutions, n_objectives).
        true_pf : Optional[np.ndarray]
            True Pareto front array.
        n_objectives : int
            Number of objectives.
        n_tasks : int
            Total number of tasks.
        prob : str
            Problem name.
        task_idx : int
            Task index.
        algo : str
            Algorithm name.

        Returns
        -------
        plt.Figure
            Matplotlib figure object.
        """
        fig = plt.figure(figsize=(4.5, 3.5))

        if n_objectives == 2:
            ax = fig.add_subplot(111)

            if true_pf is not None and true_pf.shape[1] == 2:
                sort_idx = np.argsort(true_pf[:, 0])
                sorted_pf = true_pf[sort_idx]
                ax.scatter(sorted_pf[:, 0], sorted_pf[:, 1],
                           c='gray', s=2, linewidth=0.1, label='True PF', zorder=1)

            ax.scatter(nd_solutions[:, 0], nd_solutions[:, 1],
                       c='dodgerblue', s=60, alpha=0.8, edgecolors='black',
                       linewidth=0.8, label='ND Solutions', zorder=2)

            ax.set_xlabel('$f_1$', fontsize=12)
            ax.set_ylabel('$f_2$', fontsize=12)
            ax.grid(True, alpha=0.2, linestyle='-')

        elif n_objectives == 3:
            ax = fig.add_subplot(111, projection='3d')

            if true_pf is not None and true_pf.shape[1] == 3:
                ax.scatter(true_pf[:, 0], true_pf[:, 1], true_pf[:, 2],
                           c='gray', s=4, alpha=0.2, label='True PF', zorder=1, depthshade=True)

            ax.scatter(nd_solutions[:, 0], nd_solutions[:, 1], nd_solutions[:, 2],
                       c='dodgerblue', s=60, alpha=0.8, edgecolors='black',
                       linewidth=0.8, label='ND Solutions', zorder=2, depthshade=True)

            ax.set_xlabel('$f_1$', fontsize=12)
            ax.set_ylabel('$f_2$', fontsize=12)
            ax.set_zlabel('$f_3$', fontsize=12)

            ax.view_init(elev=20, azim=60)

        else:
            # Parallel coordinates for many-objective
            ax = fig.add_subplot(111)

            for i in range(nd_solutions.shape[0]):
                ax.plot(range(n_objectives), nd_solutions[i, :],
                        'b-', alpha=0.3, linewidth=0.8)

            ax.set_xlabel('Objective Index', fontsize=10)
            ax.set_ylabel('Value', fontsize=10)
            ax.set_xticks(range(n_objectives))
            ax.set_xticklabels([rf'$f_{{{i + 1}}}$' for i in range(n_objectives)])
            ax.grid(True, alpha=0.3, linestyle='--')

        title = f'{prob} - {algo}' if n_tasks == 1 else f'{prob} - Task{task_idx + 1} - {algo}'
        plt.title(title, fontsize=10)
        plt.tight_layout()

        return fig


# =============================================================================
# Main Data Analyzer Class
# =============================================================================

class DataAnalyzer:
    """
    Main class for comprehensive data analysis and visualization of multi-task optimization experiments.

    This class provides a complete pipeline for:

    - Scanning data directories to detect algorithms, problems, and runs
    - Calculating performance metrics (IGD, HV, or objective values)
    - Generating statistical comparison tables (Excel or LaTeX)
    - Creating convergence curve plots
    - Visualizing runtime comparisons
    - Plotting non-dominated solutions

    Attributes
    ----------
    data_path : Path
        Path to the data directory containing experiment results.
    settings : Optional[Dict[str, Any]]
        Problem settings including reference definitions and metric configuration.
    algorithm_order : Optional[List[str]]
        Custom ordering of algorithms for display.
    table_config : TableConfig
        Configuration for table generation.
    plot_config : PlotConfig
        Configuration for plot generation.
    """

    def __init__(
            self,
            data_path: Union[str, Path] = './Data',
            settings: Optional[Dict[str, Any]] = None,
            algorithm_order: Optional[List[str]] = None,
            save_path: Union[str, Path] = './Results',
            table_format: str = 'excel',
            figure_format: str = 'pdf',
            statistic_type: str = 'mean',
            significance_level: float = 0.05,
            rank_sum_test: bool = True,
            log_scale: bool = False,
            show_pf: bool = True,
            show_nd: bool = True,
            best_so_far: bool = True,
            clear_results: bool = True
    ):
        """
        Initialize DataAnalyzer with configuration parameters.

        Parameters
        ----------
        data_path : Union[str, Path], optional
            Path to data directory containing algorithm subdirectories.
            Each subdirectory should contain pickle files named: ALGO_problem_run.pkl
            Default: './Data'
        settings : Optional[Dict[str, Any]], optional
            Problem settings dictionary containing:

            - Problem names as keys (e.g., 'P1', 'P2')
            - Task definitions as nested dictionaries
            - 'metric': str ('IGD' or 'HV')
            - 'ref_path': str (path to reference files)
            - 'n_ref': int (number of reference points)

            Default: None (single-objective mode)
        algorithm_order : Optional[List[str]], optional
            Custom ordering of algorithms for display.
            The last algorithm is used as baseline for statistical tests.
            Default: None (alphabetical order)
        save_path : Union[str, Path], optional
            Directory path to save all output files.
            Default: './Results'
        table_format : str, optional
            Output table format: 'excel' or 'latex'.
            Default: 'excel'
        figure_format : str, optional
            Output figure format: 'pdf', 'png', 'svg', etc.
            Default: 'pdf'
        statistic_type : str, optional
            Type of statistic: 'mean', 'median', 'max', 'min'.
            Default: 'mean'
        significance_level : float, optional
            P-value threshold for statistical significance testing.
            Default: 0.05
        rank_sum_test : bool, optional
            Whether to perform Wilcoxon rank-sum test.
            Default: True
        log_scale : bool, optional
            Whether to use logarithmic scale for convergence plot y-axis.
            Default: False
        show_pf : bool, optional
            Whether to show true Pareto front in ND solution plots.
            Default: True
        show_nd : bool, optional
            Whether to filter and show only non-dominated solutions.
            Default: True
        best_so_far : bool, optional
            Whether to use best-so-far metric values.
            Default: True
        clear_results : bool, optional
            Whether to clear existing results folder before analysis.
            Default: True
        """
        self.data_path = Path(data_path)
        self.settings = settings
        self.algorithm_order = algorithm_order
        self.best_so_far = best_so_far
        self.clear_results = clear_results

        # Parse enums
        stat_type = StatisticType(statistic_type)
        tbl_format = TableFormat(table_format)

        # Initialize configurations
        self.table_config = TableConfig(
            table_format=tbl_format,
            statistic_type=stat_type,
            significance_level=significance_level,
            rank_sum_test=rank_sum_test,
            save_path=Path(save_path)
        )

        self.plot_config = PlotConfig(
            figure_format=figure_format,
            statistic_type=stat_type,
            log_scale=log_scale,
            show_pf=show_pf,
            show_nd=show_nd,
            save_path=Path(save_path)
        )

        # Internal state
        self._scan_result: Optional[ScanResult] = None
        self._metric_results: Optional[MetricResults] = None

    def scan_data(self) -> ScanResult:
        """
        Scan the data directory to detect algorithms, problems, run counts.

        Returns
        -------
        ScanResult
            Dataclass containing:

            - algorithms: List[str] - Sorted list of algorithm names
            - problems: List[str] - Sorted list of problem names
            - runs: int - Number of independent runs
            - data_path: Path - Path to scanned directory

        Raises
        ------
        FileNotFoundError
            If data_path does not exist.
        ValueError
            If no algorithm directories or pickle files found.
        """
        algorithms = []
        problems = []
        runs_dict = {}

        for algo_dir in [d for d in self.data_path.iterdir() if d.is_dir()]:
            algo = algo_dir.name
            algorithms.append(algo)
            runs_dict[algo] = {}

            for pkl in algo_dir.glob('*.pkl'):
                parts = pkl.stem.split('_')
                if len(parts) >= 3:
                    prob = '_'.join(parts[1:-1])
                    runs_dict[algo].setdefault(prob, []).append(pkl)

                    if prob not in problems:
                        problems.append(prob)

        algorithms.sort()
        problems.sort(key=lambda x: int(''.join(filter(str.isdigit, x))) if any(c.isdigit() for c in x) else x)
        # problems.sort()

        first_algo = algorithms[0]
        first_prob = problems[0]
        runs = len(runs_dict[first_algo][first_prob])

        print(f"Found {len(algorithms)} algorithms: {algorithms}")
        print(f"Found {len(problems)} problems: {problems}")
        print(f"Run times: {runs}")

        self._scan_result = ScanResult(
            algorithms=algorithms,
            problems=problems,
            runs=runs,
            data_path=self.data_path
        )

        return self._scan_result

    def calculate_metrics(self) -> MetricResults:
        """
        Calculate metric values for all algorithms, problems, and runs.

        Returns
        -------
        MetricResults
            Dataclass containing all computed metrics:

            - metric_values: Metric values per generation
            - best_values: Final best metric values
            - objective_values: Original objective values
            - runtime: Runtime in seconds
            - max_nfes: Maximum function evaluations
            - metric_name: Name of metric used

        Raises
        ------
        RuntimeError
            If scan_data() has not been called.
        """
        if self._scan_result is None:
            self.scan_data()

        scan = self._scan_result
        algo_order = self.algorithm_order if self.algorithm_order else scan.algorithms
        metric_name = self.settings.get('metric') if self.settings else None

        # Initialize storage dictionaries
        all_values = {algo: {prob: {} for prob in scan.problems} for algo in algo_order}
        all_values_best_so_far = {algo: {prob: {} for prob in scan.problems} for algo in algo_order}
        all_best_values = {algo: {prob: {} for prob in scan.problems} for algo in algo_order}
        original_objective_values = {algo: {prob: {} for prob in scan.problems} for algo in algo_order}
        all_runtime = {algo: {prob: {} for prob in scan.problems} for algo in algo_order}
        all_max_nfes = {algo: {prob: None for prob in scan.problems} for algo in algo_order}

        total = len(algo_order) * len(scan.problems) * scan.runs
        pbar = tqdm(total=total, desc="Calculating metric values", dynamic_ncols=False, delay=0.2)

        for algo in algo_order:
            for prob in scan.problems:
                for run in range(1, scan.runs + 1):
                    pkl_file = f"{algo}_{prob}_{run}.pkl"
                    pkl_path = self.data_path / algo / pkl_file

                    data = DataUtils.load_pickle(pkl_path)
                    metric_values, metric_values_best_bs = self._get_single_run_metric_value(data, prob)

                    all_values[algo][prob][run] = metric_values
                    all_values_best_so_far[algo][prob][run] = metric_values_best_bs

                    last_vals = [
                        np.asarray(task_arr).ravel()[-1] if len(task_arr) > 0 else np.nan
                        for task_arr in metric_values_best_bs
                    ]
                    all_best_values[algo][prob][run] = last_vals

                    last_objs = [data['all_objs'][t][-1] for t in range(len(data['all_objs']))]
                    original_objective_values[algo][prob][run] = last_objs

                    all_runtime[algo][prob][run] = data['runtime']

                    if all_max_nfes[algo][prob] is None:
                        all_max_nfes[algo][prob] = data['max_nfes']

                    pbar.update(1)

        pbar.close()

        selected = all_values_best_so_far if self.best_so_far else all_values

        self._metric_results = MetricResults(
            metric_values=selected,
            best_values=all_best_values,
            objective_values=original_objective_values,
            runtime=all_runtime,
            max_nfes=all_max_nfes,
            metric_name=metric_name
        )

        return self._metric_results

    def _get_single_run_metric_value(
            self,
            data: Dict[str, Any],
            prob: str
    ) -> Tuple[List[np.ndarray], List[np.ndarray]]:
        """
        Calculate metric values for a single run.

        Parameters
        ----------
        data : Dict[str, Any]
            Loaded pickle data containing 'all_objs' key.
        prob : str
            Problem name for loading references.

        Returns
        -------
        Tuple[List[np.ndarray], List[np.ndarray]]
            Tuple of (metric_values, metric_values_best_so_far).
            Each is a list of arrays, one per task.
        """
        all_decs = data['all_decs']
        all_objs = data['all_objs']
        all_cons = data.get('all_cons', None)
        n_tasks = len(all_objs)
        n_gens_per_task = [len(all_objs[t]) for t in range(n_tasks)]

        metric_values = [np.zeros((n_gens_per_task[t], 1)) for t in range(n_tasks)]
        metric_values_best_so_far = [np.zeros((n_gens_per_task[t], 1)) for t in range(n_tasks)]

        for t in range(n_tasks):
            task_key = f'T{t + 1}'
            best_so_far = None

            reference = None

            if self.settings is not None and n_gens_per_task[t] > 0:

                M = all_objs[t][0].shape[1]
                D = all_decs[t][0].shape[1]
                C = all_cons[t][0].shape[1] if all_cons is not None else 0

                reference = DataUtils.load_reference(
                    self.settings,
                    prob,
                    task_key,
                    M=M,
                    D=D,
                    C=C
                )

            for gen in range(n_gens_per_task[t]):
                objs_tgen = all_objs[t][gen]
                cons_tgen = all_cons[t][gen] if all_cons is not None else None
                M = objs_tgen.shape[1]

                if M == 1:
                    metric_value = np.min(objs_tgen[:, 0])
                    sign = -1
                else:
                    if self.settings is None:
                        raise ValueError('Multi-objective metric calculation requires settings parameter')

                    metric_name = self.settings.get('metric')

                    if metric_name == 'IGD':
                        metric_instance = IGD()
                        metric_value = metric_instance.calculate(objs_tgen, reference)
                        sign = metric_instance.sign
                    elif metric_name == 'HV':
                        metric_instance = HV()
                        metric_value = metric_instance.calculate(objs_tgen, reference)
                        sign = metric_instance.sign
                    elif metric_name == 'IGDp':
                        metric_instance = IGDp()
                        metric_value = metric_instance.calculate(objs_tgen, reference)
                        sign = metric_instance.sign
                    elif metric_name == 'GD':
                        metric_instance = GD()
                        metric_value = metric_instance.calculate(objs_tgen, reference)
                        sign = metric_instance.sign
                    elif metric_name == 'DeltaP':
                        metric_instance = DeltaP()
                        metric_value = metric_instance.calculate(objs_tgen, reference)
                        sign = metric_instance.sign
                    elif metric_name == 'Spacing':
                        metric_instance = Spacing()
                        metric_value = metric_instance.calculate(objs_tgen)
                        sign = metric_instance.sign
                    elif metric_name == 'Spread':
                        metric_instance = Spread()
                        metric_value = metric_instance.calculate(objs_tgen, reference)
                        sign = metric_instance.sign
                    elif metric_name == 'FR':
                        if cons_tgen is None:
                            raise ValueError('FR metric requires constraint data, but all_cons is not available')
                        metric_instance = FR()
                        metric_value = metric_instance.calculate(cons_tgen)
                        sign = metric_instance.sign
                    elif metric_name == 'CV':
                        if cons_tgen is None:
                            raise ValueError('CV metric requires constraint data, but all_cons is not available')
                        metric_instance = CV()
                        metric_value = metric_instance.calculate(cons_tgen)
                        sign = metric_instance.sign
                    else:
                        raise ValueError(f'Unsupported metric: {metric_name}')

                metric_values[t][gen, 0] = metric_value

                if best_so_far is None:
                    best_so_far = metric_value
                else:
                    if sign == -1:
                        best_so_far = min(best_so_far, metric_value)
                    else:
                        best_so_far = max(best_so_far, metric_value)

                metric_values_best_so_far[t][gen, 0] = best_so_far

        return metric_values, metric_values_best_so_far

    def generate_tables(self) -> Union[pd.DataFrame, str]:
        """
        Generate comparison tables with statistical analysis.

        Returns
        -------
        Union[pd.DataFrame, str]
            DataFrame for Excel format, LaTeX string for LaTeX format.

        Raises
        ------
        RuntimeError
            If calculate_metrics() has not been called.
        """
        if self._metric_results is None:
            self.calculate_metrics()

        algo_order = self.algorithm_order if self.algorithm_order else self._scan_result.algorithms

        table_gen = TableGenerator(self.table_config)
        return table_gen.generate(
            self._metric_results.best_values,
            algo_order,
            self._metric_results.metric_name
        )

    def generate_convergence_plots(self) -> None:
        """
        Generate and save convergence curve plots.

        Returns
        -------
        None
            Saves figures to disk at configured save_path.

        Raises
        ------
        RuntimeError
            If calculate_metrics() has not been called.
        """
        if self._metric_results is None:
            self.calculate_metrics()

        algo_order = self.algorithm_order if self.algorithm_order else self._scan_result.algorithms

        plot_gen = PlotGenerator(self.plot_config)
        plot_gen.plot_convergence_curves(
            self._metric_results.metric_values,
            self._metric_results.best_values,
            self._metric_results.max_nfes,
            algo_order,
            self._metric_results.metric_name
        )

    def generate_runtime_plots(self) -> None:
        """
        Generate and save runtime comparison bar plots.

        Returns
        -------
        None
            Saves figure to disk at configured save_path.

        Raises
        ------
        RuntimeError
            If calculate_metrics() has not been called.
        """
        if self._metric_results is None:
            self.calculate_metrics()

        algo_order = self.algorithm_order if self.algorithm_order else self._scan_result.algorithms

        plot_gen = PlotGenerator(self.plot_config)
        plot_gen.plot_runtime(self._metric_results.runtime, algo_order)

    def generate_nd_solution_plots(self) -> None:
        """
        Generate and save non-dominated solution visualization plots.

        Returns
        -------
        None
            Saves figures to disk at configured save_path/ND_Solutions/.

        Raises
        ------
        RuntimeError
            If calculate_metrics() has not been called.
        """
        if self._metric_results is None:
            self.calculate_metrics()

        algo_order = self.algorithm_order if self.algorithm_order else self._scan_result.algorithms

        plot_gen = PlotGenerator(self.plot_config)
        plot_gen.plot_nd_solutions(
            self._metric_results.best_values,
            self._metric_results.objective_values,
            algo_order,
            self.settings
        )

    def run(self) -> MetricResults:
        """
        Execute the complete analysis pipeline.

        This method runs all analysis steps in sequence:

        1. Clear existing results (if configured)
        2. Scan data directory
        3. Calculate metrics
        4. Generate statistical tables
        5. Generate convergence plots
        6. Generate runtime plots
        7. Generate non-dominated solution plots

        Returns
        -------
        MetricResults
            Complete metric results from the analysis.
        """
        print("=" * 60)
        print('🚀🚀🚀 Starting Data Analysis Pipeline! 🚀🚀🚀')
        print("=" * 60)

        # Step 0: Clear results folder if requested
        if self.clear_results:
            results_path = self.table_config.save_path
            if results_path.exists():
                print(f'\n♻️  Clearing existing results folder: {results_path}')
                shutil.rmtree(results_path)
            results_path.mkdir(parents=True, exist_ok=True)

        # Step 1: Scan data
        print('\n🔍 Scanning data directory...')
        self.scan_data()

        # Step 2: Calculate metrics
        print('\n📊 Calculating metric values...')
        self.calculate_metrics()

        # Step 3: Generate tables
        print('\n📋 Generating statistical tables...')
        self.generate_tables()

        # Step 4: Plot convergence curves
        print('\n📈 Plotting convergence curves...')
        self.generate_convergence_plots()

        # Step 5: Plot runtime
        print('\n⏱️ Plotting runtime comparison...')
        self.generate_runtime_plots()

        # Step 6: Plot non-dominated solutions
        print('\n🎯 Plotting non-dominated solutions...')
        self.generate_nd_solution_plots()

        print("=" * 60)
        print('🎉🎉🎉 Data Analysis Completed! 🎉🎉🎉')
        print("=" * 60)

        return self._metric_results


# =============================================================================
# Module Entry Point and Usage Examples
# =============================================================================

if __name__ == '__main__':
    """
    Usage Examples for DataAnalyzer Module
    ======================================

    This module provides a comprehensive analysis pipeline for multi-task 
    optimization experiments. Below are various usage patterns.


    Example 1: Quick Start - Full Pipeline
    --------------------------------------
    Run complete analysis with default settings::

        from data_analyzer import DataAnalyzer

        analyzer = DataAnalyzer(data_path='./Data')
        results = analyzer.run()


    Example 2: Multi-Objective Optimization with Custom Settings
    ------------------------------------------------------------
    Analyze multi-objective results with IGD metric::

        from data_analyzer import DataAnalyzer

        # Define problem settings with Pareto front references
        SETTINGS = {
            'metric': 'IGD',
            'ref_path': './MOReference',
            'n_ref': 10000,
            'P1': {
                'T1': 'P1_T1_ref.npy',
                'T2': 'P1_T2_ref.npy',
            },
            'P2': {
                'T1': lambda n, m: generate_pf(n, m),  # Callable reference
            }
        }

        analyzer = DataAnalyzer(
            data_path='./Data',
            settings=SETTINGS,
            save_path='./Results',
            table_format='latex',
            figure_format='pdf'
        )
        results = analyzer.run()


    Example 3: Step-by-Step Analysis
    --------------------------------
    Execute individual analysis steps for fine-grained control::

        from data_analyzer import DataAnalyzer

        analyzer = DataAnalyzer(
            data_path='./Data',
            settings=SETTINGS,
            algorithm_order=['NSGA-II', 'MOEA/D', 'MyAlgo'],  # Last is baseline
            clear_results=False
        )

        # Step 1: Scan data directory
        scan_result = analyzer.scan_data()
        print(f"Found algorithms: {scan_result.algorithms}")
        print(f"Found problems: {scan_result.problems}")

        # Step 2: Calculate metrics
        metric_results = analyzer.calculate_metrics()

        # Step 3: Generate only specific outputs
        analyzer.generate_tables()           # Statistical comparison tables
        analyzer.generate_convergence_plots() # Convergence curves
        analyzer.generate_runtime_plots()     # Runtime bar charts
        analyzer.generate_nd_solution_plots() # Pareto front visualizations


    Example 4: Custom Table Generation
    ----------------------------------
    Generate tables with specific statistical settings::

        from data_analyzer import (
            DataAnalyzer, TableGenerator, TableConfig, 
            TableFormat, StatisticType
        )

        # Create custom table configuration
        table_config = TableConfig(
            table_format=TableFormat.LATEX,
            statistic_type=StatisticType.MEDIAN,
            significance_level=0.01,
            rank_sum_test=True,
            save_path=Path('./CustomResults')
        )

        # Use with analyzer
        analyzer = DataAnalyzer(data_path='./Data', settings=SETTINGS)
        analyzer.scan_data()
        analyzer.calculate_metrics()

        # Generate table with custom config
        table_gen = TableGenerator(table_config)
        latex_table = table_gen.generate(
            analyzer._metric_results.best_values,
            algorithm_order=['Algo1', 'Algo2', 'Baseline'],
            metric_name='IGD'
        )


    Example 5: Custom Plot Generation
    ---------------------------------
    Create plots with specific visual settings::

        from data_analyzer import DataAnalyzer, PlotGenerator, PlotConfig, StatisticType

        # Create custom plot configuration
        plot_config = PlotConfig(
            figure_format='png',
            statistic_type=StatisticType.MEDIAN,
            log_scale=True,
            show_pf=True,
            show_nd=True,
            save_path=Path('./Figures'),
            colors=['#E41A1C', '#377EB8', '#4DAF4A'],  # Custom colors
            markers=['o', 's', '^']
        )

        analyzer = DataAnalyzer(data_path='./Data', settings=SETTINGS)
        analyzer.calculate_metrics()

        # Generate plots with custom config
        plot_gen = PlotGenerator(plot_config)
        plot_gen.plot_convergence_curves(
            analyzer._metric_results.metric_values,
            analyzer._metric_results.best_values,
            analyzer._metric_results.max_nfes,
            algorithm_order=['Algo1', 'Algo2'],
            metric_name='IGD'
        )


    Example 6: Access Raw Results
    -----------------------------
    Access computed metrics for custom analysis::

        from data_analyzer import DataAnalyzer

        analyzer = DataAnalyzer(data_path='./Data', settings=SETTINGS)
        results = analyzer.run()

        # Access metric values
        # Structure: results.metric_values[algo][problem][run][task_idx]
        algo1_p1_run1_task0 = results.metric_values['Algo1']['P1'][1][0]

        # Access best values
        # Structure: results.best_values[algo][problem][run] = [task0_val, task1_val, ...]
        best_vals = results.best_values['Algo1']['P1'][1]

        # Access objective values (Pareto solutions)
        # Structure: results.objective_values[algo][problem][run][task_idx] = np.ndarray
        pareto_solutions = results.objective_values['Algo1']['P1'][1][0]

        # Access runtime
        runtime_seconds = results.runtime['Algo1']['P1'][1]

        # Access max NFEs per task
        max_nfes_list = results.max_nfes['Algo1']['P1']


    Example 7: Using Utility Classes Directly
    -----------------------------------------
    Use statistics and data utilities independently::

        from data_analyzer import (
            StatisticsCalculator, DataUtils, 
            StatisticType, OptimizationDirection
        )
        import numpy as np

        # Calculate statistics
        data = [1.2, 1.5, 1.1, 1.3, 1.4]
        mean, std = StatisticsCalculator.calculate_statistic(data, StatisticType.MEAN)

        # Perform statistical comparison
        algo_data = [1.0, 1.1, 0.9, 1.2]
        base_data = [2.0, 2.1, 1.9, 2.2]
        result = StatisticsCalculator.perform_rank_sum_test(
            algo_data, base_data,
            significance_level=0.05,
            direction=OptimizationDirection.MINIMIZE
        )
        print(f"Comparison: {result.symbol}, p-value: {result.p_value}")

        # Load reference data
        reference = DataUtils.load_reference(
            settings=SETTINGS,
            problem='P1',
            task_identifier=0,  # or 'T1'
            n_objectives=2
        )


    Data Directory Structure
    ------------------------
    Expected directory structure for input data::

        ./Data/
        ├── Algorithm1/
        │   ├── Algorithm1_Problem1_1.pkl
        │   ├── Algorithm1_Problem1_2.pkl
        │   ├── Algorithm1_Problem2_1.pkl
        │   └── ...
        ├── Algorithm2/
        │   ├── Algorithm2_Problem1_1.pkl
        │   └── ...
        └── ...

    Each .pkl file should contain a dictionary with keys:

    - 'all_objs': List[List[np.ndarray]] - Objectives per task per generation
    - 'runtime': float - Total runtime in seconds
    - 'max_nfes': List[int] - Max function evaluations per task


    Output Structure
    ----------------
    Generated output files::

        ./Results/
        ├── results_table_mean.xlsx      # or .tex for LaTeX
        ├── Problem1.pdf                 # Convergence plot (single task)
        ├── Problem2-Task1.pdf           # Convergence plot (multi-task)
        ├── Problem2-Task2.pdf
        ├── runtime_comparison.pdf       # Runtime bar chart
        └── ND_Solutions/
            ├── Problem1-Algorithm1.pdf  # Pareto front plot
            ├── Problem1-Algorithm2.pdf
            └── ...
    """

    # Demo: Run analysis with sample configuration
    print("DataAnalyzer Module - Demo Run")
    print("=" * 50)

    # Example configuration (modify paths as needed)
    analyzer = DataAnalyzer(
        data_path='./Data',
        save_path='./Results',
        table_format='excel',
        figure_format='pdf',
        statistic_type='mean',
        significance_level=0.05,
        rank_sum_test=True,
        log_scale=False,
        show_pf=True,
        show_nd=True,
        clear_results=True
    )

    # Run complete analysis pipeline
    results = analyzer.run()