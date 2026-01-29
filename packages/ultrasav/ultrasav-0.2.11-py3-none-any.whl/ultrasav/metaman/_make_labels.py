"""
Make Labels Module (v2 - Polars Backend)
=========================================
A high-performance utility module for transforming Excel files containing label mappings 
into Python dictionaries using Polars for vectorized operations.

This module provides the make_labels function to:
- Read column labels and value labels from Excel sheets using Polars
- Generate Python dictionary files for SPSS metadata labeling
- Support customizable sheet names, column names, and output formatting

Version: 2.0.0
Dependencies: polars, pathlib
"""

import polars as pl
from pathlib import Path


# =============================================================================
# Type Aliases (for clarity)
# =============================================================================
ColumnLabelsDict = dict[str, str]
ValueLabelsDict = dict[str, dict[int | float | str, str]]


# =============================================================================
# Excel Reading (adapted from def_read_files.py)
# =============================================================================

def _read_excel_sheet(file_path: Path, sheet_name: str, engine: str = "calamine") -> pl.DataFrame:
    """
    Read a single Excel sheet into a Polars DataFrame with engine fallback.
    
    Parameters
    ----------
    file_path : Path
        Path to the Excel file
    sheet_name : str
        Name of the sheet to read
    engine : str, default "calamine"
        Excel engine to use. Falls back to openpyxl, then xlsx2csv on failure.
    
    Returns
    -------
    pl.DataFrame
        The sheet data as a Polars DataFrame
    """
    file_str = str(file_path)
    engines = [engine, "openpyxl", "xlsx2csv"]
    
    last_error: Exception | None = None
    for eng in engines:
        try:
            return pl.read_excel(file_str, sheet_name=sheet_name, engine=eng)
        except Exception as e:
            last_error = e
            continue
    
    raise ValueError(f"Failed to read sheet '{sheet_name}' from '{file_path}'. Last error: {last_error}")


def _get_sheet_names(file_path: Path) -> list[str]:
    """
    Get all sheet names from an Excel file.
    
    Uses Polars' sheet_id=0 to load all sheets as a dict, then extracts keys.
    Falls back to fastexcel or openpyxl if needed.
    
    Parameters
    ----------
    file_path : Path
        Path to the Excel file
    
    Returns
    -------
    list[str]
        List of sheet names
    """
    file_str = str(file_path)
    
    # Primary: Use Polars with sheet_id=0 (returns {sheetname: DataFrame, ...} dict)
    engines = ["calamine", "openpyxl", "xlsx2csv"]
    for engine in engines:
        try:
            sheets_dict = pl.read_excel(file_str, sheet_id=0, engine=engine)
            return list(sheets_dict.keys())
        except Exception:
            continue
    
    # Fallback: Try fastexcel directly (calamine's Python binding)
    try:
        import fastexcel
        excel_file = fastexcel.read_excel(file_str)
        return excel_file.sheet_names
    except (ImportError, Exception):
        pass
    
    # Last resort: openpyxl directly
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True, data_only=True)
        names = wb.sheetnames
        wb.close()
        return names
    except Exception as e:
        raise ValueError(
            f"Failed to read sheet names from '{file_path}'. "
            f"Ensure the file exists and is a valid Excel file. Error: {e}"
        )


# =============================================================================
# Data Cleaning Helpers (Pure Functions)
# =============================================================================

def _clean_variable_column(df: pl.DataFrame, col_name: str) -> pl.DataFrame:
    """
    Clean variable name column: cast to string, mark nulls/empty as null.
    
    Parameters
    ----------
    df : pl.DataFrame
        Input DataFrame
    col_name : str
        Name of the variable column to clean
    
    Returns
    -------
    pl.DataFrame
        DataFrame with cleaned variable column
    """
    return df.with_columns(
        pl.col(col_name)
        .cast(pl.Utf8)
        .replace("", None)
        .alias(col_name)
    )


def _clean_label_column(df: pl.DataFrame, col_name: str) -> pl.DataFrame:
    """
    Clean label column: cast to string, strip whitespace, convert null to empty string.
    
    Parameters
    ----------
    df : pl.DataFrame
        Input DataFrame
    col_name : str
        Name of the label column to clean
    
    Returns
    -------
    pl.DataFrame
        DataFrame with cleaned label column
    """
    return df.with_columns(
        pl.col(col_name)
        .cast(pl.Utf8)
        .fill_null("")
        .str.strip_chars()
        .alias(col_name)
    )


def _filter_valid_rows(df: pl.DataFrame, required_col: str) -> pl.DataFrame:
    """
    Filter rows where the required column is not null.
    
    Parameters
    ----------
    df : pl.DataFrame
        Input DataFrame
    required_col : str
        Column that must not be null
    
    Returns
    -------
    pl.DataFrame
        Filtered DataFrame
    """
    return df.filter(pl.col(required_col).is_not_null())


# =============================================================================
# Value Conversion (Vectorized)
# =============================================================================

def _convert_values_vectorized(df: pl.DataFrame, value_col: str) -> pl.DataFrame:
    """
    Convert value column to appropriate types: int > float > string (vectorized).
    
    This uses a vectorized approach:
    1. Cast to string first
    2. Try to cast to Float64
    3. Check if float is integer-like, cast to Int64 where possible
    4. Keep as string where numeric conversion failed
    
    Parameters
    ----------
    df : pl.DataFrame
        Input DataFrame
    value_col : str
        Name of the value column to convert
    
    Returns
    -------
    pl.DataFrame
        DataFrame with converted value column (as Object/mixed type via struct)
    """
    # First, ensure string representation for processing
    df = df.with_columns(
        pl.col(value_col).cast(pl.Utf8).str.strip_chars().alias("_val_str")
    )
    
    # Try numeric conversion
    df = df.with_columns(
        pl.col("_val_str").cast(pl.Float64, strict=False).alias("_val_float")
    )
    
    # Check if float values are actually integers
    df = df.with_columns(
        (pl.col("_val_float") == pl.col("_val_float").floor()).alias("_is_int")
    )
    
    # Create integer version where applicable
    df = df.with_columns(
        pl.when(pl.col("_is_int") & pl.col("_val_float").is_not_null())
        .then(pl.col("_val_float").cast(pl.Int64))
        .otherwise(None)
        .alias("_val_int")
    )
    
    return df


def _extract_typed_value(row: dict) -> int | float | str | None:
    """
    Extract the properly typed value from a row with conversion columns.
    
    Parameters
    ----------
    row : dict
        Row dictionary with _val_int, _val_float, _val_str columns
    
    Returns
    -------
    int | float | str | None
        The value in its appropriate type
    """
    if row.get("_val_int") is not None:
        return int(row["_val_int"])
    elif row.get("_val_float") is not None:
        return float(row["_val_float"])
    elif row.get("_val_str") is not None and row["_val_str"] != "":
        return str(row["_val_str"])
    return None


# =============================================================================
# Duplicate Detection (Vectorized)
# =============================================================================

def _count_duplicates(df: pl.DataFrame, cols: list[str]) -> int:
    """
    Count the number of duplicate groups (groups with more than 1 row).
    
    Parameters
    ----------
    df : pl.DataFrame
        Input DataFrame
    cols : list[str]
        Columns to group by for duplicate detection
    
    Returns
    -------
    int
        Number of duplicate groups
    """
    counts = df.group_by(cols).agg(pl.len().alias("_count"))
    return counts.filter(pl.col("_count") > 1).height


# =============================================================================
# Column Labels Processing
# =============================================================================

def _validate_column_labels_schema(df: pl.DataFrame, var_col: str, label_col: str) -> None:
    """
    Validate that required columns exist in column labels DataFrame.
    
    Parameters
    ----------
    df : pl.DataFrame
        Input DataFrame
    var_col : str
        Expected variable column name
    label_col : str
        Expected label column name
    
    Raises
    ------
    ValueError
        If required columns are missing
    """
    if var_col not in df.columns or label_col not in df.columns:
        raise ValueError(f"Column labels sheet must have '{var_col}' and '{label_col}' columns")


def _process_column_labels_df(
    df: pl.DataFrame,
    var_col: str,
    label_col: str
) -> tuple[ColumnLabelsDict, int, int, int]:
    """
    Process column labels DataFrame into dictionary (pure function).
    
    Parameters
    ----------
    df : pl.DataFrame
        Raw column labels DataFrame
    var_col : str
        Variable column name
    label_col : str
        Label column name
    
    Returns
    -------
    tuple[ColumnLabelsDict, int, int, int]
        (column_labels_dict, initial_rows, removed_rows, duplicate_count)
    """
    initial_rows = df.height
    
    # Clean columns
    df = _clean_variable_column(df, var_col)
    df = _clean_label_column(df, label_col)
    
    # Filter valid rows
    df_cleaned = _filter_valid_rows(df, var_col)
    removed = initial_rows - df_cleaned.height
    
    # Check duplicates
    duplicate_count = _count_duplicates(df_cleaned, [var_col])
    
    # Build dictionary (vectorized extraction)
    result: ColumnLabelsDict = {}
    for row in df_cleaned.select([var_col, label_col]).iter_rows(named=True):
        result[row[var_col]] = row[label_col]
    
    return result, initial_rows, removed, duplicate_count


# =============================================================================
# Value Labels Processing
# =============================================================================

def _validate_value_labels_schema(
    df: pl.DataFrame,
    var_col: str,
    value_col: str,
    label_col: str
) -> None:
    """
    Validate that required columns exist in value labels DataFrame.
    
    Parameters
    ----------
    df : pl.DataFrame
        Input DataFrame
    var_col : str
        Expected variable column name
    value_col : str
        Expected value column name
    label_col : str
        Expected label column name
    
    Raises
    ------
    ValueError
        If required columns are missing
    """
    required = [var_col, value_col, label_col]
    missing = [col for col in required if col not in df.columns]
    if missing:
        raise ValueError(f"Value labels sheet missing columns: {missing}")


def _process_value_labels_df(
    df: pl.DataFrame,
    var_col: str,
    value_col: str,
    label_col: str
) -> tuple[ValueLabelsDict, int, int, int]:
    """
    Process value labels DataFrame into nested dictionary (pure function).
    
    Parameters
    ----------
    df : pl.DataFrame
        Raw value labels DataFrame
    var_col : str
        Variable column name
    value_col : str
        Value column name
    label_col : str
        Label column name
    
    Returns
    -------
    tuple[ValueLabelsDict, int, int, int]
        (value_labels_dict, initial_rows, removed_rows, duplicate_count)
    """
    initial_rows = df.height
    
    # Clean variable column
    df = _clean_variable_column(df, var_col)
    
    # Clean label column
    df = _clean_label_column(df, label_col)
    
    # Filter rows with valid variable and value
    df = df.filter(
        pl.col(var_col).is_not_null() & 
        pl.col(value_col).is_not_null()
    )
    
    # Convert values to appropriate types (vectorized)
    df = _convert_values_vectorized(df, value_col)
    
    # Filter out rows where conversion resulted in null
    df_cleaned = df.filter(
        pl.col("_val_int").is_not_null() | 
        pl.col("_val_float").is_not_null() |
        (pl.col("_val_str").is_not_null() & (pl.col("_val_str") != ""))
    )
    
    removed = initial_rows - df_cleaned.height
    
    # Check duplicates on variable + original value string (before type conversion)
    duplicate_count = _count_duplicates(df_cleaned, [var_col, "_val_str"])
    
    # Build nested dictionary
    # Group by variable first for efficiency
    result: ValueLabelsDict = {}
    
    # Select only needed columns for iteration
    select_cols = [var_col, label_col, "_val_int", "_val_float", "_val_str"]
    
    for row in df_cleaned.select(select_cols).iter_rows(named=True):
        variable = row[var_col]
        label = row[label_col]
        value = _extract_typed_value(row)
        
        if value is not None:
            if variable not in result:
                result[variable] = {}
            result[variable][value] = label
    
    return result, initial_rows, removed, duplicate_count


# =============================================================================
# Output Formatting (Pure Functions)
# =============================================================================

def _format_column_dict_lines(
    data: ColumnLabelsDict,
    dict_name: str,
    quote_style: str,
    indent: str
) -> list[str]:
    """
    Format column labels dictionary as Python code lines.
    
    Parameters
    ----------
    data : ColumnLabelsDict
        Column labels dictionary
    dict_name : str
        Name for the dictionary variable
    quote_style : str
        Quote style for values
    indent : str
        Indentation string
    
    Returns
    -------
    list[str]
        Lines of Python code
    """
    lines = [f"{dict_name} = {{"]
    for variable, label in data.items():
        lines.append(f"{indent}'{variable}': {quote_style}{label}{quote_style},")
    lines.append("}")
    return lines


def _format_value_dict_lines(
    data: ValueLabelsDict,
    dict_name: str,
    quote_style: str,
    indent: str
) -> list[str]:
    """
    Format value labels dictionary as Python code lines.
    
    Parameters
    ----------
    data : ValueLabelsDict
        Value labels dictionary
    dict_name : str
        Name for the dictionary variable
    quote_style : str
        Quote style for label values
    indent : str
        Indentation string
    
    Returns
    -------
    list[str]
        Lines of Python code
    """
    lines = [f"\n\n{dict_name} = {{"]
    for variable, value_dict in data.items():
        lines.append(f"{indent}'{variable}': {{")
        for value, label in value_dict.items():
            key_repr = f"'{value}'" if isinstance(value, str) else str(value)
            lines.append(f"{indent}{indent}{key_repr}: {quote_style}{label}{quote_style},")
        lines.append(f"{indent}}},")
    lines.append("}")
    return lines


def _save_combined_output(
    col_labels: ColumnLabelsDict,
    val_labels: ValueLabelsDict,
    output_path: str,
    col_dict_name: str,
    value_dict_name: str,
    col_quote_style: str,
    value_quote_style: str,
    indent: str,
    encoding: str
) -> Path:
    """
    Save both dictionaries to a single Python file.
    
    Parameters
    ----------
    col_labels : ColumnLabelsDict
        Column labels dictionary
    val_labels : ValueLabelsDict
        Value labels dictionary
    output_path : str
        Output file path
    col_dict_name : str
        Name for column labels dictionary
    value_dict_name : str
        Name for value labels dictionary
    col_quote_style : str
        Quote style for column labels
    value_quote_style : str
        Quote style for value labels
    indent : str
        Indentation string
    encoding : str
        File encoding
    
    Returns
    -------
    Path
        Path to the saved file
    """
    output_path_obj = Path(output_path)
    output_path_obj.parent.mkdir(parents=True, exist_ok=True)
    
    all_lines: list[str] = []
    all_lines.extend(_format_column_dict_lines(col_labels, col_dict_name, col_quote_style, indent))
    all_lines.extend(_format_value_dict_lines(val_labels, value_dict_name, value_quote_style, indent))
    
    with open(output_path_obj, "w", encoding=encoding) as file:
        file.write("\n".join(all_lines))
    
    return output_path_obj


# =============================================================================
# Summary Helpers
# =============================================================================

def _get_longest_labels(col_labels: ColumnLabelsDict, top_n: int = 3) -> list[tuple[str, str, int]]:
    """
    Get variables with the longest labels.
    
    Parameters
    ----------
    col_labels : ColumnLabelsDict
        Column labels dictionary
    top_n : int, default 3
        Number of top results to return
    
    Returns
    -------
    list[tuple[str, str, int]]
        List of (variable, label, length) tuples
    """
    non_empty = [(var, label, len(label)) for var, label in col_labels.items() if label]
    return sorted(non_empty, key=lambda x: x[2], reverse=True)[:top_n]


def _get_top_value_label_counts(val_labels: ValueLabelsDict, top_n: int = 5) -> list[tuple[str, int]]:
    """
    Get variables with the most value labels.
    
    Parameters
    ----------
    val_labels : ValueLabelsDict
        Value labels dictionary
    top_n : int, default 5
        Number of top results to return
    
    Returns
    -------
    list[tuple[str, int]]
        List of (variable, count) tuples
    """
    counts = [(var, len(labels)) for var, labels in val_labels.items()]
    return sorted(counts, key=lambda x: x[1], reverse=True)[:top_n]


# =============================================================================
# Main Function
# =============================================================================

def make_labels(
    input_path: str,
    output_path: str | None = None,
    col_label_sheet: str = 'col_label',
    value_label_sheet: str = 'value_label',
    col_dict_name: str = 'user_column_labels',
    value_dict_name: str = 'user_variable_value_labels',
    col_quote_style: str = "'''",
    value_quote_style: str = "'''",
    indent: str = "    ",
    encoding: str = 'utf-8',
    col_variable_column: str = 'variable',
    col_label_column: str = 'label',
    value_variable_column: str = 'variable',
    value_value_column: str = 'value',
    value_label_column: str = 'label',
    verbose: bool = False
) -> tuple[ColumnLabelsDict, ValueLabelsDict]:
    """
    Transform Excel file with two sheets into a single Python file containing both 
    column labels and value labels dictionaries.
    
    This v2 version uses Polars for high-performance vectorized operations.
    
    Parameters
    ----------
    input_path : str
        Path to input Excel file with two sheets: one for column labels, one for value labels
    output_path : str, optional
        Path where the output Python file will be saved. If None, no file is written (default: None)
    col_label_sheet : str, optional
        Name of the sheet containing column labels (default: 'col_label')
    value_label_sheet : str, optional
        Name of the sheet containing value labels (default: 'value_label')
    col_dict_name : str, optional
        Name of the column labels dictionary in output (default: 'user_column_labels')
    value_dict_name : str, optional
        Name of the value labels dictionary in output (default: 'user_variable_value_labels')
    col_quote_style : str, optional
        Quote style for column labels in output (default: triple quotes "'''")
    value_quote_style : str, optional
        Quote style for value labels in output (default: triple quotes "'''")
    indent : str, optional
        Indentation for dictionary items (default: 4 spaces)
    encoding : str, optional
        File encoding (default: 'utf-8')
    col_variable_column : str, optional
        Name of the column containing variable names in col_label sheet (default: 'variable')
    col_label_column : str, optional
        Name of the column containing labels in col_label sheet (default: 'label')
    value_variable_column : str, optional
        Name of the column containing variable names in value_label sheet (default: 'variable')
    value_value_column : str, optional
        Name of the column containing values in value_label sheet (default: 'value')
    value_label_column : str, optional
        Name of the column containing labels in value_label sheet (default: 'label')
    verbose : bool, optional
        Whether to print progress messages (default: False)
    
    Returns
    -------
    tuple[ColumnLabelsDict, ValueLabelsDict]
        Tuple containing (column_labels_dict, value_labels_dict)
    
    Examples
    --------
    >>> # Basic usage
    >>> col_labels, val_labels = make_labels(
    ...     input_path="label_mapping.xlsx",
    ...     output_path="label_mapping.py"
    ... )
    
    >>> # Return only (no file output)
    >>> col_labels, val_labels = make_labels(
    ...     input_path="label_mapping.xlsx"
    ... )
    
    >>> # Custom configuration
    >>> col_labels, val_labels = make_labels(
    ...     input_path="mappings.xlsx",
    ...     output_path="all_labels.py",
    ...     col_label_sheet="columns",
    ...     value_label_sheet="values",
    ...     col_dict_name="column_labels",
    ...     value_dict_name="value_labels"
    ... )
    """
    
    def _print(msg: str) -> None:
        """Conditional print helper."""
        if verbose:
            print(msg)
    
    try:
        _print("=" * 60)
        _print("COMBINED LABEL MAKER (v2 - Polars) - Starting Processing")
        _print("=" * 60)
        _print(f"Input file: {input_path}")
        _print(f"Output file: {output_path if output_path else 'None (return only)'}")
        
        # Validate input file
        file_path = Path(input_path)
        if not file_path.exists():
            raise FileNotFoundError(f"Input file not found: {file_path}")
        
        if file_path.suffix.lower() not in ['.xlsx', '.xls']:
            raise ValueError(f"Input must be an Excel file, got: {file_path.suffix}")
        
        # Get available sheets
        available_sheets = _get_sheet_names(file_path)
        _print(f"\nAvailable sheets: {available_sheets}")
        
        # Validate required sheets
        if col_label_sheet not in available_sheets:
            raise ValueError(f"Sheet '{col_label_sheet}' not found. Available: {available_sheets}")
        if value_label_sheet not in available_sheets:
            raise ValueError(f"Sheet '{value_label_sheet}' not found. Available: {available_sheets}")
        
        # Process column labels
        _print("\n📊 Processing Column Labels Sheet...")
        df_col = _read_excel_sheet(file_path, col_label_sheet)
        _validate_column_labels_schema(df_col, col_variable_column, col_label_column)
        
        column_labels, col_initial, col_removed, col_dups = _process_column_labels_df(
            df_col, col_variable_column, col_label_column
        )
        
        _print(f"  Found {col_initial} rows")
        if col_removed > 0:
            _print(f"  ⚠ Removed {col_removed} rows with null variable names")
        if col_dups > 0:
            _print(f"  ⚠ Warning: Found {col_dups} duplicate variable names")
        empty_labels = sum(1 for label in column_labels.values() if label == "")
        _print(f"  ✓ Processed {len(column_labels)} column labels ({empty_labels} empty)")
        
        # Process value labels
        _print("\n📊 Processing Value Labels Sheet...")
        df_val = _read_excel_sheet(file_path, value_label_sheet)
        _validate_value_labels_schema(df_val, value_variable_column, value_value_column, value_label_column)
        
        value_labels, val_initial, val_removed, val_dups = _process_value_labels_df(
            df_val, value_variable_column, value_value_column, value_label_column
        )
        
        _print(f"  Found {val_initial} rows")
        if val_removed > 0:
            _print(f"  ⚠ Removed {val_removed} invalid rows")
        if val_dups > 0:
            _print(f"  ⚠ Warning: Found {val_dups} duplicate variable-value pairs")
        total_labels = sum(len(labels) for labels in value_labels.values())
        _print(f"  ✓ Processed {len(value_labels)} variables with {total_labels} total value labels")
        
        # Save output if path provided
        if output_path:
            saved_path = _save_combined_output(
                column_labels, value_labels, output_path,
                col_dict_name, value_dict_name,
                col_quote_style, value_quote_style,
                indent, encoding
            )
            _print(f"\n✓ Combined dictionaries saved to: {saved_path}")
        
        # Print summary
        _print("\n" + "=" * 60)
        _print("PROCESSING SUMMARY")
        _print("=" * 60)
        _print(f"Column Labels Dictionary: {len(column_labels)} variables")
        _print(f"Value Labels Dictionary: {len(value_labels)} variables")
        
        # Show longest labels
        if column_labels:
            longest = _get_longest_labels(column_labels)
            if longest:
                _print(f"\nVariables with longest labels:")
                for i, (var, label, length) in enumerate(longest, 1):
                    preview = label[:50] + "..." if len(label) > 50 else label
                    _print(f"  {i}. {var}: {length} chars - '{preview}'")
        
        # Show top value label counts
        if value_labels:
            top_counts = _get_top_value_label_counts(value_labels)
            _print(f"\nTop 5 variables by value label count:")
            for i, (var, count) in enumerate(top_counts, 1):
                _print(f"  {i}. {var}: {count} labels")
        
        _print("=" * 60)
        if output_path:
            _print("🎉 Combined label transformation and file save completed successfully!")
        else:
            _print("🎉 Combined label transformation completed successfully! (No file saved)")
        
        return column_labels, value_labels
        
    except Exception as e:
        _print(f"\n❌ Error during transformation: {str(e)}")
        raise


__all__ = ["make_labels"]
