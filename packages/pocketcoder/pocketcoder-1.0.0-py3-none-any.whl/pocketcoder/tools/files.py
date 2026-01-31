"""
File utilities for PocketCoder.

Handles file reading, validation, and path resolution.
"""

from __future__ import annotations

import fnmatch
import subprocess
import sys
from pathlib import Path


def read_file(path: Path, start_line: int = None, end_line: int = None) -> str:
    """
    Read file content as text or extract from binary formats.

    v2.5.1: Added start_line/end_line for reading specific ranges (gearbox).

    Supports:
    - Text files (.py, .js, .txt, .md, etc.)
    - PDF files (.pdf) - requires pdfplumber
    - Excel files (.xlsx, .xls) - requires openpyxl
    - Word files (.docx) - requires python-docx

    Args:
        path: Path to file
        start_line: First line to read (1-indexed, optional)
        end_line: Last line to read (1-indexed, inclusive, optional)

    Returns:
        File content as string (or markdown for binary formats)

    Raises:
        FileNotFoundError: If file doesn't exist
        PermissionError: If file isn't readable
    """
    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")

    ext = path.suffix.lower()

    # Binary formats - use specialized readers (no line range support)
    if ext == '.pdf':
        return _read_pdf(path)
    elif ext in ('.xlsx', '.xls'):
        return _read_excel(path)
    elif ext == '.docx':
        return _read_docx(path)
    elif ext in ('.png', '.jpg', '.jpeg', '.gif', '.webp', '.bmp'):
        return _read_image(path)

    # Text files
    content = path.read_text()

    # v2.5.1: If line range specified, extract only those lines
    if start_line is not None or end_line is not None:
        lines = content.split('\n')
        total = len(lines)
        start = max(1, start_line or 1) - 1  # Convert to 0-indexed
        end = min(total, end_line or total)
        selected = lines[start:end]
        return f"[Lines {start + 1}-{end} of {total}]\n" + '\n'.join(selected)

    # Default: full file with smart mode for large files
    return _smart_read(content, path)


# =============================================================================
# Smart Read for Large Files (v2.2.0)
# =============================================================================

SMART_READ_THRESHOLD = 200  # Lines threshold for smart mode


def _smart_read(content: str, path: Path) -> str:
    """
    Smart read: show structure for large files instead of full content.

    v2.2.0: Saves context tokens by showing:
    - First 50 lines
    - Structure (def/class/import lines with line numbers)
    - Last 20 lines

    Args:
        content: Full file content
        path: File path (for extension detection)

    Returns:
        Full content if small, or smart summary if large
    """
    import re

    lines = content.splitlines()

    # Small file - return full content
    if len(lines) <= SMART_READ_THRESHOLD:
        return content

    # Large file - show structure
    ext = path.suffix.lower()

    # Patterns for structure extraction by language
    patterns = {
        '.py': r'^(def |class |import |from |@)',
        '.js': r'^(function |class |const |let |var |import |export )',
        '.ts': r'^(function |class |const |let |var |import |export |interface |type )',
        '.java': r'^(public |private |protected |class |interface |import )',
        '.go': r'^(func |type |package |import )',
        '.rs': r'^(fn |struct |enum |impl |use |mod )',
        '.rb': r'^(def |class |module |require |include )',
        '.php': r'^(function |class |namespace |use )',
    }

    pattern = patterns.get(ext, r'^(def |class |function |import )')

    # Extract structure
    structure = []
    for i, line in enumerate(lines, 1):
        if re.match(pattern, line.strip()):
            structure.append(f"{i:4d}: {line.rstrip()}")

    # Build smart output
    parts = [
        f"# File: {path.name} ({len(lines)} lines) — SMART READ MODE",
        f"# Showing: first 50 + structure + last 20 lines",
        "",
        "# === FIRST 50 LINES ===",
        "\n".join(lines[:50]),
        "",
        f"# === STRUCTURE ({len(structure)} definitions) ===",
        "\n".join(structure[:50]) if structure else "# (no definitions found)",
        "",
        "# === LAST 20 LINES ===",
        "\n".join(lines[-20:]),
    ]

    return "\n".join(parts)


# =============================================================================
# Binary File Readers (lazy import)
# =============================================================================

def _read_pdf(path: Path) -> str:
    """Read PDF file using pdfplumber."""
    try:
        import pdfplumber
    except ImportError:
        return (
            f"❌ Cannot read PDF: pdfplumber not installed\n"
            f"💡 Install: pip install pdfplumber\n\n"
            f"After installing, try reading the file again."
        )

    try:
        with pdfplumber.open(path) as pdf:
            pages = []
            for i, page in enumerate(pdf.pages, 1):
                text = page.extract_text()
                if text:
                    pages.append(f"## Page {i}\n\n{text}")

            if not pages:
                return f"PDF file '{path.name}' has no extractable text (may be scanned/image-only)"

            content = "\n\n".join(pages)

            # Limit size
            if len(content) > 50000:
                content = content[:50000] + f"\n\n... (truncated, {len(pdf.pages)} pages total)"

            return f"# {path.name}\n\n{content}"
    except Exception as e:
        return f"❌ Error reading PDF: {e}"


def _read_excel(path: Path) -> str:
    """Read Excel file using openpyxl."""
    try:
        import openpyxl
    except ImportError:
        return (
            f"❌ Cannot read Excel: openpyxl not installed\n"
            f"💡 Install: pip install openpyxl\n\n"
            f"After installing, try reading the file again."
        )

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        sheets = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []

            for row in ws.iter_rows(max_row=1000, values_only=True):  # Limit rows
                # Skip empty rows
                if all(cell is None for cell in row):
                    continue
                row_str = " | ".join(str(cell) if cell is not None else "" for cell in row)
                rows.append(row_str)

            if rows:
                # Create markdown table header
                header = rows[0] if rows else ""
                separator = " | ".join("---" for _ in header.split(" | "))
                table = f"| {header} |\n| {separator} |\n"
                table += "\n".join(f"| {row} |" for row in rows[1:50])  # Limit displayed rows

                if len(rows) > 50:
                    table += f"\n\n... ({len(rows)} rows total, showing first 50)"

                sheets.append(f"## Sheet: {sheet_name}\n\n{table}")

        if not sheets:
            return f"Excel file '{path.name}' is empty"

        return f"# {path.name}\n\n" + "\n\n".join(sheets)
    except Exception as e:
        return f"❌ Error reading Excel: {e}"


def _read_docx(path: Path) -> str:
    """Read Word document using python-docx."""
    try:
        from docx import Document
    except ImportError:
        return (
            f"❌ Cannot read Word document: python-docx not installed\n"
            f"💡 Install: pip install python-docx\n\n"
            f"After installing, try reading the file again."
        )

    try:
        doc = Document(path)
        paragraphs = []

        for para in doc.paragraphs:
            text = para.text.strip()
            if text:
                # Detect headings by style
                if para.style and 'Heading' in para.style.name:
                    level = para.style.name.replace('Heading', '').strip()
                    try:
                        level = int(level)
                    except ValueError:
                        level = 1
                    paragraphs.append(f"{'#' * level} {text}")
                else:
                    paragraphs.append(text)

        if not paragraphs:
            return f"Word document '{path.name}' is empty"

        content = "\n\n".join(paragraphs)

        # Limit size
        if len(content) > 50000:
            content = content[:50000] + "\n\n... (truncated)"

        return f"# {path.name}\n\n{content}"
    except Exception as e:
        return f"❌ Error reading Word document: {e}"


def _read_image(path: Path) -> str:
    """
    Handle image files.

    For now, returns info about the image.
    Vision support requires model that supports images.
    """
    try:
        # Try to get image info with Pillow if available
        from PIL import Image
        img = Image.open(path)
        return (
            f"📷 Image: {path.name}\n"
            f"   Format: {img.format}\n"
            f"   Size: {img.width}x{img.height}\n"
            f"   Mode: {img.mode}\n\n"
            f"💡 To extract text from image, OCR is needed.\n"
            f"   Install: pip install pytesseract\n"
            f"   Also need Tesseract: brew install tesseract (macOS)\n\n"
            f"💡 For vision analysis, use a model that supports images (GPT-4o, Claude)."
        )
    except ImportError:
        return (
            f"📷 Image: {path.name}\n\n"
            f"Cannot read image details: Pillow not installed\n"
            f"💡 Install: pip install Pillow\n\n"
            f"For OCR text extraction:\n"
            f"💡 Install: pip install pytesseract Pillow\n"
            f"Also need Tesseract: brew install tesseract (macOS)"
        )
    except Exception as e:
        return f"❌ Error reading image: {e}"


def is_binary(path: Path) -> bool:
    """
    Check if file is binary (not text).

    Uses heuristics:
    - Presence of null bytes
    - High ratio of non-printable characters

    Args:
        path: Path to file

    Returns:
        True if file appears to be binary
    """
    try:
        with open(path, "rb") as f:
            chunk = f.read(8192)

            # Null byte = definitely binary
            if b"\x00" in chunk:
                return True

            # High ratio of non-printable = probably binary
            # (excluding tab, newline, carriage return)
            if chunk:
                non_printable = sum(
                    1 for b in chunk if b < 32 and b not in (9, 10, 13)
                )
                if non_printable / len(chunk) > 0.3:
                    return True

            return False

    except Exception:
        # On error, assume binary (safer)
        return True


def is_ignored(path: Path, gitignore_patterns: list[str] | None = None) -> bool:
    """
    Check if file matches .gitignore patterns.

    Args:
        path: Path to file
        gitignore_patterns: List of patterns (loads from .gitignore if None)

    Returns:
        True if file should be ignored
    """
    if gitignore_patterns is None:
        gitignore_patterns = load_gitignore(path.parent)

    # Always ignore common patterns
    always_ignore = [
        ".git",
        "__pycache__",
        "*.pyc",
        ".env",
        "node_modules",
        ".venv",
        "venv",
    ]

    all_patterns = always_ignore + gitignore_patterns

    for pattern in all_patterns:
        # Match against filename
        if fnmatch.fnmatch(path.name, pattern):
            return True
        # Match against relative path
        try:
            rel_path = path.relative_to(Path.cwd())
            if fnmatch.fnmatch(str(rel_path), pattern):
                return True
        except ValueError:
            pass

    return False


def load_gitignore(directory: Path) -> list[str]:
    """
    Load .gitignore patterns from directory and parents.

    Args:
        directory: Directory to start from

    Returns:
        List of gitignore patterns
    """
    patterns = []

    # Walk up to find .gitignore files
    current = directory.resolve()
    root = Path(current.anchor)

    while current != root:
        gitignore = current / ".gitignore"
        if gitignore.exists():
            try:
                for line in gitignore.read_text().splitlines():
                    line = line.strip()
                    if line and not line.startswith("#"):
                        patterns.append(line)
            except Exception:
                pass
        current = current.parent

    return patterns


def resolve_path(filename: str, base: Path | None = None) -> Path:
    """
    Resolve filename to absolute path with fuzzy search.

    If exact path not found, searches for file by name in project.

    Args:
        filename: Filename or relative path
        base: Base directory (defaults to cwd)

    Returns:
        Absolute path (may not exist if not found)
    """
    if base is None:
        base = Path.cwd()

    path = Path(filename)

    # 1. Absolute path — return as is
    if path.is_absolute():
        return path

    # 2. Exact relative path exists — return it
    exact_path = (base / path).resolve()
    if exact_path.exists():
        return exact_path

    # 3. Fuzzy search: find file by name in project
    search_name = path.name  # Just the filename without dirs
    found = find_file_in_project(search_name, base)

    if found:
        return found

    # 4. Not found — return original path (will error later)
    return exact_path


def find_file_in_project(filename: str, base: Path, max_depth: int = 5) -> Path | None:
    """
    Find file by name recursively in project directory.

    Args:
        filename: Filename to search for
        base: Base directory to search from
        max_depth: Maximum directory depth to search

    Returns:
        Path to first matching file, or None
    """
    # Skip common directories
    skip_dirs = {".git", "__pycache__", "node_modules", ".venv", "venv", ".tox", "dist", "build"}

    def search(directory: Path, depth: int) -> Path | None:
        if depth > max_depth:
            return None

        try:
            for item in directory.iterdir():
                # Skip hidden and ignored directories
                if item.is_dir():
                    if item.name.startswith(".") or item.name in skip_dirs:
                        continue
                    # Recurse
                    result = search(item, depth + 1)
                    if result:
                        return result
                elif item.is_file() and item.name == filename:
                    return item
        except PermissionError:
            pass

        return None

    return search(base, 0)


def get_file_info(path: Path) -> dict:
    """
    Get information about a file.

    Args:
        path: Path to file

    Returns:
        Dict with file info (exists, size, lines, is_binary, etc.)
    """
    info = {
        "path": str(path),
        "name": path.name,
        "exists": path.exists(),
    }

    if path.exists():
        stat = path.stat()
        info["size"] = stat.st_size
        info["mtime"] = stat.st_mtime
        info["is_binary"] = is_binary(path)

        if not info["is_binary"]:
            try:
                content = path.read_text()
                info["lines"] = len(content.splitlines())
            except Exception:
                info["lines"] = 0

    return info


def find_file(filename: str, base: Path | None = None) -> tuple[bool, str]:
    """
    Find file location in project.

    Args:
        filename: Filename to search for
        base: Base directory (defaults to cwd)

    Returns:
        Tuple of (found, message with path or error)
    """
    if base is None:
        base = Path.cwd()

    found = find_file_in_project(filename, base)

    if found:
        # Return relative path for readability
        try:
            rel_path = found.relative_to(base)
            return True, f"Found: {rel_path}"
        except ValueError:
            return True, f"Found: {found}"
    else:
        return False, f"File not found: {filename}"


def open_file(path: Path) -> tuple[bool, str]:
    """
    Open file in default application.

    Args:
        path: Path to file

    Returns:
        Tuple of (success, message)
    """
    if not path.exists():
        return False, f"File not found: {path}"

    try:
        if sys.platform == "darwin":
            # macOS
            subprocess.run(["open", str(path)], check=True)
        elif sys.platform == "win32":
            # Windows
            subprocess.run(["start", "", str(path)], shell=True, check=True)
        else:
            # Linux
            subprocess.run(["xdg-open", str(path)], check=True)
        return True, f"Opened {path.name}"
    except Exception as e:
        return False, f"Failed to open {path.name}: {e}"


# =============================================================================
# NEW TOOLS: list_files, search_files, write_file, glob_files
# =============================================================================

# Patterns to ignore when listing/searching
IGNORE_PATTERNS = {
    "node_modules", "__pycache__", ".git", "dist", "build",
    "target", ".venv", "venv", ".tox", ".pytest_cache",
    ".mypy_cache", ".coverage", "htmlcov", ".egg-info",
    ".idea", ".vscode", ".cache", "*.pyc", "*.pyo"
}

DEFAULT_LIMIT = 100
MAX_LINE_LENGTH = 500
MAX_SEARCH_RESULTS = 100


def _should_ignore(path: Path) -> bool:
    """Check if path should be ignored."""
    for pattern in IGNORE_PATTERNS:
        if pattern.startswith("*"):
            if path.name.endswith(pattern[1:]):
                return True
        elif path.name == pattern or pattern in path.parts:
            return True
    return False


def list_files(
    path: str = ".",
    recursive: bool = False,
    limit: int = DEFAULT_LIMIT,
    base: Path | None = None
) -> tuple[bool, str]:
    """
    List files in directory with tree structure.

    Args:
        path: Directory path (relative or absolute)
        recursive: Include subdirectories
        limit: Max files to return
        base: Base directory (defaults to cwd)

    Returns:
        Tuple of (success, formatted_output)
    """
    if base is None:
        base = Path.cwd()

    target = Path(path)
    if not target.is_absolute():
        target = (base / path).resolve()

    if not target.exists():
        return False, f"Directory not found: {path}"

    if not target.is_dir():
        return False, f"Not a directory: {path}"

    file_count = [0]  # Use list to allow modification in nested function

    def format_tree(directory: Path, prefix: str = "", depth: int = 0) -> str:
        """Format directory as tree."""
        if file_count[0] >= limit:
            return ""

        output = []

        try:
            items = sorted(
                [i for i in directory.iterdir() if not _should_ignore(i)],
                key=lambda x: (not x.is_dir(), x.name.lower())
            )
        except PermissionError:
            return ""

        for i, item in enumerate(items):
            if file_count[0] >= limit:
                break

            is_last = i == len(items) - 1
            connector = "└── " if is_last else "├── "

            if item.is_dir():
                output.append(f"{prefix}{connector}{item.name}/")
                if recursive and depth < 10:
                    extension = "    " if is_last else "│   "
                    subtree = format_tree(item, prefix + extension, depth + 1)
                    if subtree:
                        output.append(subtree)
            else:
                output.append(f"{prefix}{connector}{item.name}")
                file_count[0] += 1

        return "\n".join(filter(None, output))

    # Build output
    try:
        rel_path = target.relative_to(base)
        header = f"{rel_path}/" if str(rel_path) != "." else "./"
    except ValueError:
        header = f"{target}/"

    tree = format_tree(target)
    hit_limit = file_count[0] >= limit
    footer = f"\n[{file_count[0]} files" + (", truncated]" if hit_limit else "]")

    return True, f"{header}\n{tree}{footer}"


def search_files(
    pattern: str,
    path: str = ".",
    include: str | None = None,
    base: Path | None = None
) -> tuple[bool, str]:
    """
    Search for pattern in file contents (grep).

    Args:
        pattern: Regex pattern to search
        path: Directory to search in
        include: Glob pattern to filter files (e.g. "*.py")
        base: Base directory (defaults to cwd)

    Returns:
        Tuple of (success, formatted_output)
    """
    import re

    if base is None:
        base = Path.cwd()

    target = Path(path)
    if not target.is_absolute():
        target = (base / path).resolve()

    if not target.exists():
        return False, f"Path not found: {path}"

    # Validate regex
    try:
        regex = re.compile(pattern)
    except re.error as e:
        return False, f"Invalid regex pattern: {e}"

    matches = []

    def search_file(filepath: Path):
        """Search single file for pattern."""
        if len(matches) >= MAX_SEARCH_RESULTS:
            return

        try:
            content = filepath.read_text(errors='ignore')
            for line_num, line in enumerate(content.splitlines(), 1):
                if regex.search(line):
                    matches.append({
                        "file": filepath,
                        "line_num": line_num,
                        "line": line[:MAX_LINE_LENGTH] + ("..." if len(line) > MAX_LINE_LENGTH else ""),
                        "mtime": filepath.stat().st_mtime
                    })
                    if len(matches) >= MAX_SEARCH_RESULTS:
                        return
        except (PermissionError, UnicodeDecodeError, OSError):
            pass

    def search_dir(directory: Path):
        """Recursively search directory."""
        if len(matches) >= MAX_SEARCH_RESULTS:
            return

        try:
            for item in directory.iterdir():
                if _should_ignore(item):
                    continue

                if item.is_dir():
                    search_dir(item)
                elif item.is_file():
                    # Apply include filter if specified
                    if include and not fnmatch.fnmatch(item.name, include):
                        continue
                    # Skip binary files
                    if is_binary(item):
                        continue
                    search_file(item)
        except PermissionError:
            pass

    if target.is_file():
        search_file(target)
    else:
        search_dir(target)

    if not matches:
        return True, f"No matches found for pattern: {pattern}"

    # Sort by modification time (newest first)
    matches.sort(key=lambda x: x["mtime"], reverse=True)

    # Format output
    truncated = len(matches) >= MAX_SEARCH_RESULTS
    output_lines = [f"Found {len(matches)} matches" + (" (truncated)" if truncated else "")]

    current_file = None
    for match in matches:
        try:
            rel_path = match["file"].relative_to(base)
        except ValueError:
            rel_path = match["file"]

        if current_file != match["file"]:
            current_file = match["file"]
            output_lines.append(f"\n{rel_path}:")

        output_lines.append(f"  Line {match['line_num']}: {match['line']}")

    return True, "\n".join(output_lines)


def write_file(
    path: str,
    content: str,
    base: Path | None = None,
    create_dirs: bool = True
) -> tuple[bool, str]:
    """
    Write content to file (create or overwrite).

    Args:
        path: File path (relative or absolute)
        content: Content to write
        base: Base directory (defaults to cwd)
        create_dirs: Create parent directories if needed

    Returns:
        Tuple of (success, message_with_diff)
    """
    import difflib
    import html

    if base is None:
        base = Path.cwd()

    target = Path(path)
    if not target.is_absolute():
        target = (base / path).resolve()

    # Security: check if inside project
    try:
        target.relative_to(base)
    except ValueError:
        return False, f"Cannot write outside project directory: {path}"

    # Clean content (remove markdown fences, unescape HTML)
    content = _clean_content(content)

    # Check if file exists
    exists = target.exists()
    old_content = ""

    if exists:
        try:
            old_content = target.read_text()
        except Exception:
            pass

    # Create parent directories
    if create_dirs and not target.parent.exists():
        try:
            target.parent.mkdir(parents=True, exist_ok=True)
        except Exception as e:
            return False, f"Cannot create directory: {e}"

    # Generate diff for preview
    diff = _generate_diff(path, old_content, content)

    # Write file
    try:
        target.write_text(content)

        action = "Modified" if exists else "Created"
        try:
            rel_path = target.relative_to(base)
        except ValueError:
            rel_path = target

        return True, f"{action} {rel_path}\n\n{diff}"

    except Exception as e:
        return False, f"Failed to write file: {e}"


def _clean_content(content: str) -> str:
    """Clean content from markdown fences and HTML entities."""
    import html

    lines = content.split("\n")

    # Remove markdown code fence at start
    if lines and lines[0].startswith("```"):
        lines = lines[1:]

    # Remove markdown code fence at end
    if lines and lines[-1].strip() == "```":
        lines = lines[:-1]

    content = "\n".join(lines)

    # Unescape HTML entities
    content = html.unescape(content)

    return content


def _generate_diff(filename: str, old: str, new: str) -> str:
    """Generate unified diff for preview."""
    import difflib

    if not old:
        # New file - show content preview
        lines = new.splitlines()
        preview = lines[:20]
        if len(lines) > 20:
            preview.append(f"... (+{len(lines) - 20} more lines)")
        return "```\n" + "\n".join(preview) + "\n```"

    # Generate unified diff
    diff = difflib.unified_diff(
        old.splitlines(keepends=True),
        new.splitlines(keepends=True),
        fromfile=f"a/{filename}",
        tofile=f"b/{filename}",
        lineterm=""
    )

    diff_lines = list(diff)
    if not diff_lines:
        return "(no changes)"

    return "```diff\n" + "".join(diff_lines) + "\n```"


def glob_files(
    pattern: str,
    path: str = ".",
    base: Path | None = None
) -> tuple[bool, str]:
    """
    Find files matching glob pattern.

    Args:
        pattern: Glob pattern (e.g. "*.py", "**/*.ts")
        path: Directory to search in
        base: Base directory (defaults to cwd)

    Returns:
        Tuple of (success, formatted_output)
    """
    if base is None:
        base = Path.cwd()

    target = Path(path)
    if not target.is_absolute():
        target = (base / path).resolve()

    if not target.exists():
        return False, f"Path not found: {path}"

    if not target.is_dir():
        return False, f"Not a directory: {path}"

    # Find matching files
    matches = []
    try:
        for match in target.glob(pattern):
            if _should_ignore(match):
                continue
            if match.is_file():
                matches.append(match)
                if len(matches) >= DEFAULT_LIMIT:
                    break
    except Exception as e:
        return False, f"Glob error: {e}"

    if not matches:
        return True, f"No files matching: {pattern}"

    # Sort by modification time (newest first)
    matches.sort(key=lambda x: x.stat().st_mtime, reverse=True)

    # Format output
    output_lines = [f"Found {len(matches)} files matching '{pattern}'"]

    for match in matches:
        try:
            rel_path = match.relative_to(base)
        except ValueError:
            rel_path = match
        output_lines.append(f"  {rel_path}")

    truncated = len(matches) >= DEFAULT_LIMIT
    if truncated:
        output_lines.append("  (truncated)")

    return True, "\n".join(output_lines)
