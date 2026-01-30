"""Formula evaluation engine with external library integration."""

# ============================================================================
# Imports
# ============================================================================

import logging
from pathlib import Path
from typing import Any

from messy_xlsx.exceptions import (
    CircularReferenceError,
    FormulaError,
    UnsupportedFunctionError,
)
from messy_xlsx.formulas.config import (
    CircularRefStrategy,
    FormulaConfig,
    FormulaEvaluationMode,
)


# ============================================================================
# Config
# ============================================================================

log = logging.getLogger(__name__)


# ============================================================================
# Core
# ============================================================================

class FormulaEngine:
    """Formula evaluation engine with library fallback chain."""

    def __init__(self, config: FormulaConfig | None = None):
        """Initialize engine."""
        self.config            = config or FormulaConfig()
        self._cache: dict[str, Any] = {}
        self._eval_stack: list[str] = []
        self._custom_functions: dict[str, callable] = {}

        self._init_xlcalculator()
        self._init_formulas()

        self._xlcalc_model     = None
        self._xlcalc_evaluator = None
        self._formulas_model   = None
        self._current_file: Path | None = None

    def _init_xlcalculator(self) -> None:
        """Try to initialize xlcalculator."""
        try:
            from xlcalculator import Evaluator, ModelCompiler

            self.xlcalc_available = True
            self._ModelCompiler   = ModelCompiler
            self._Evaluator       = Evaluator
            log.debug("xlcalculator available")
        except ImportError:
            self.xlcalc_available = False
            log.debug("xlcalculator not installed")

    def _init_formulas(self) -> None:
        """Try to initialize formulas library."""
        try:
            import formulas

            self.formulas_available = True
            self._formulas          = formulas
            log.debug("formulas library available")
        except ImportError:
            self.formulas_available = False
            log.debug("formulas library not installed")

    @property
    def is_available(self) -> bool:
        """Check if any formula evaluation library is available."""
        return self.xlcalc_available or self.formulas_available

    def load_workbook(self, file_path: Path | str) -> None:
        """Load workbook for formula evaluation."""
        file_path = Path(file_path)

        if self._current_file == file_path:
            return

        self._current_file = file_path
        self._cache.clear()

        if self.xlcalc_available:
            try:
                compiler              = self._ModelCompiler()
                self._xlcalc_model    = compiler.read_and_parse_archive(str(file_path))
                self._xlcalc_evaluator = self._Evaluator(self._xlcalc_model)
                log.debug(f"Loaded {file_path} with xlcalculator")
            except Exception as e:
                log.warning(f"xlcalculator failed to load {file_path}: {e}")
                self._xlcalc_model     = None
                self._xlcalc_evaluator = None

        if self.formulas_available:
            try:
                self._formulas_model = (
                    self._formulas.ExcelModel().loads(str(file_path)).finish()
                )
                log.debug(f"Loaded {file_path} with formulas library")
            except Exception as e:
                log.warning(f"formulas library failed to load {file_path}: {e}")
                self._formulas_model = None

    def evaluate(
        self,
        sheet: str,
        row: int,
        col: int,
        cached_value: Any = None,
    ) -> Any:
        """Evaluate a cell's formula."""
        if self.config.mode == FormulaEvaluationMode.DISABLED:
            return cached_value

        if self.config.mode == FormulaEvaluationMode.CACHED_ONLY:
            return cached_value

        if self.config.mode == FormulaEvaluationMode.CACHED_WITH_FALLBACK:
            if cached_value is not None:
                return cached_value

        from openpyxl.utils import get_column_letter

        col_letter = get_column_letter(col)
        cell_ref   = f"{sheet}!{col_letter}{row}"

        if cell_ref in self._cache:
            return self._cache[cell_ref]

        if cell_ref in self._eval_stack:
            cycle = self._eval_stack + [cell_ref]
            return self._handle_circular_ref(cycle, cached_value)

        if len(self._eval_stack) > self.config.max_depth:
            raise FormulaError(
                f"Evaluation depth exceeded ({self.config.max_depth})",
                cell_ref = cell_ref,
            )

        self._eval_stack.append(cell_ref)

        try:
            result             = self._evaluate_formula(cell_ref)
            self._cache[cell_ref] = result
            return result
        finally:
            self._eval_stack.pop()

    def _evaluate_formula(self, cell_ref: str) -> Any:
        """Evaluate formula using available libraries."""
        if self._xlcalc_evaluator is not None:
            try:
                result = self._xlcalc_evaluator.evaluate(cell_ref)
                return result
            except Exception as e:
                error_str = str(e).lower()
                if "not supported" in error_str or "unknown function" in error_str:
                    log.debug(f"xlcalculator: unsupported function in {cell_ref}")
                else:
                    log.debug(f"xlcalculator failed for {cell_ref}: {e}")

        if self._formulas_model is not None:
            try:
                self._formulas_model.calculate()
                parts = cell_ref.split("!")
                if len(parts) == 2:
                    sheet, cell = parts
                    book        = list(self._formulas_model.books.values())[0]
                    if sheet in book:
                        if cell in book[sheet]:
                            return book[sheet][cell].value
            except Exception as e:
                error_str = str(e).lower()
                if "not supported" in error_str:
                    log.debug(f"formulas: unsupported function in {cell_ref}")
                else:
                    log.debug(f"formulas failed for {cell_ref}: {e}")

        if self.config.raise_on_unsupported:
            raise UnsupportedFunctionError("UNKNOWN", cell_ref)

        return self.config.unsupported_value

    def _handle_circular_ref(
        self,
        cycle: list[str],
        cached_value: Any,
    ) -> Any:
        """Handle circular reference based on strategy."""
        if self.config.circular_strategy == CircularRefStrategy.ERROR:
            raise CircularReferenceError(
                "Circular reference detected",
                cycle = cycle,
            )

        if self.config.circular_strategy == CircularRefStrategy.RETURN_CACHED:
            if cached_value is not None:
                return cached_value
            raise CircularReferenceError(
                "Circular reference with no cached value",
                cycle = cycle,
            )

        if self.config.circular_strategy == CircularRefStrategy.ITERATE:
            cell_ref = cycle[-1]
            if cell_ref in self._cache:
                return self._cache[cell_ref]
            if cached_value is not None:
                return cached_value
            return 0

        return cached_value

    def register_function(self, name: str, func: callable) -> None:
        """Register a custom Excel function."""
        self._custom_functions[name.upper()] = func

        if self.xlcalc_available:
            try:
                from xlcalculator.xlfunctions import xl

                xl.register()(func)
            except Exception as e:
                log.warning(f"Failed to register {name} with xlcalculator: {e}")

        if self.formulas_available:
            try:
                functions = self._formulas.get_functions()
                functions[name.upper()] = func
            except Exception as e:
                log.warning(f"Failed to register {name} with formulas: {e}")

    def clear_cache(self) -> None:
        """Clear evaluation cache."""
        self._cache.clear()
        self._eval_stack.clear()
