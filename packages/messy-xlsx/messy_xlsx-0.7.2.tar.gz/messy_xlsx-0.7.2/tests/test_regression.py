"""Regression tests for previously found bugs."""

import openpyxl
import pytest

from messy_xlsx import MessyWorkbook, SheetConfig


class TestRegressionBugs:
    """Tests for specific bugs found in production."""

    def test_header_detection_with_metadata_rows(self, temp_dir):
        """Regression: Headers after metadata rows should be detected."""
        file_path = temp_dir / "metadata_header.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Company Report"])
        ws.append(["Generated: 2024-01-01"])
        ws.append([])
        ws.append(["Name", "Value"])
        ws.append(["Alice", 100])
        wb.save(file_path)
        wb.close()

        with MessyWorkbook(file_path) as mwb:
            structure = mwb.get_structure()

            # Should detect row 4 as header
            assert structure.header_row == 4
            assert structure.header_confidence >= 0.7

    def test_european_numbers_not_corrupted(self, temp_dir):
        """Regression: European numbers should parse correctly."""
        file_path = temp_dir / "european.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Amount"])
        ws.append(["1.234,56"])  # European format
        wb.save(file_path)
        wb.close()

        config = SheetConfig(locale="de_DE")

        with MessyWorkbook(file_path, sheet_config=config) as mwb:
            df = mwb.to_dataframe()

            # Should parse as 1234.56 (column name is lowercase)
            import pandas as pd
            if pd.api.types.is_numeric_dtype(df["amount"]):
                assert df.iloc[0]["amount"] == pytest.approx(1234.56)

    def test_merged_cells_dont_crash(self, temp_dir):
        """Regression: Merged cells should not crash parser."""
        file_path = temp_dir / "merged_regression.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Merged"
        ws.merge_cells("A1:C1")
        ws.append(["X", "Y", "Z"])
        ws.append([1, 2, 3])
        wb.save(file_path)
        wb.close()

        with MessyWorkbook(file_path) as mwb:
            df = mwb.to_dataframe()
            assert df is not None

    def test_hidden_rows_excluded_by_default(self, temp_dir):
        """Regression: Hidden rows should be excluded."""
        file_path = temp_dir / "hidden.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["A", "B"])
        ws.append([1, 2])
        ws.append([3, 4])  # Will be hidden
        ws.append([5, 6])

        # Hide row 3
        ws.row_dimensions[3].hidden = True

        wb.save(file_path)
        wb.close()

        config = SheetConfig(include_hidden=False)

        with MessyWorkbook(file_path, sheet_config=config) as mwb:
            df = mwb.to_dataframe()
            # Should exclude hidden row
            assert len(df) <= 3
