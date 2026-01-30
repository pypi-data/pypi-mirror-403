"""Tests for formula evaluation mode configurations."""

import openpyxl
import pytest

from messy_xlsx import MessyWorkbook
from messy_xlsx.formulas import FormulaConfig, FormulaEvaluationMode, CircularRefStrategy


class TestFormulaEvaluationModes:
    """Test all formula evaluation modes."""

    @pytest.mark.parametrize("mode", [
        FormulaEvaluationMode.DISABLED,
        FormulaEvaluationMode.CACHED_ONLY,
        FormulaEvaluationMode.CACHED_WITH_FALLBACK,
        FormulaEvaluationMode.ALWAYS_EVALUATE,
    ])
    def test_formula_modes(self, temp_dir, mode):
        """Test all formula evaluation modes."""
        file_path = temp_dir / f"formulas_{mode.value}.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["A", "B", "Sum"])
        ws["A2"] = 10
        ws["B2"] = 20
        ws["C2"] = "=A2+B2"
        wb.save(file_path)
        wb.close()

        config = FormulaConfig(mode=mode)

        with MessyWorkbook(file_path, formula_config=config) as mwb:
            df = mwb.to_dataframe()
            assert df is not None
            assert len(df) == 1


class TestCircularReferenceStrategies:
    """Test circular reference handling strategies."""

    @pytest.mark.parametrize("strategy", [
        CircularRefStrategy.ERROR,
        CircularRefStrategy.RETURN_CACHED,
        CircularRefStrategy.ITERATE,
    ])
    def test_circular_strategies(self, temp_dir, strategy):
        """Test all circular reference strategies."""
        file_path = temp_dir / f"circular_{strategy.value}.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "=B1+1"
        ws["B1"] = "=A1+1"
        wb.save(file_path)
        wb.close()

        config = FormulaConfig(
            mode=FormulaEvaluationMode.CACHED_ONLY,
            circular_strategy=strategy
        )

        with MessyWorkbook(file_path, formula_config=config) as mwb:
            # Should handle based on strategy
            df = mwb.to_dataframe()
            assert df is not None


class TestUnsupportedFunctionHandling:
    """Test handling of unsupported functions."""

    def test_raise_on_unsupported_true(self, temp_dir):
        """Test raising error on unsupported functions."""
        file_path = temp_dir / "unsupported_raise.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Function"])
        ws["A2"] = "=UNKNOWNFUNC(123)"
        wb.save(file_path)
        wb.close()

        config = FormulaConfig(
            mode=FormulaEvaluationMode.CACHED_ONLY,
            raise_on_unsupported=True
        )

        with MessyWorkbook(file_path, formula_config=config) as mwb:
            # Should handle gracefully with cached values
            df = mwb.to_dataframe()
            assert df is not None

    def test_raise_on_unsupported_false(self, temp_dir):
        """Test placeholder on unsupported functions."""
        file_path = temp_dir / "unsupported_placeholder.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Function"])
        ws["A2"] = "=UNKNOWNFUNC(123)"
        wb.save(file_path)
        wb.close()

        config = FormulaConfig(
            mode=FormulaEvaluationMode.CACHED_ONLY,
            raise_on_unsupported=False,
            unsupported_value="N/A"
        )

        with MessyWorkbook(file_path, formula_config=config) as mwb:
            df = mwb.to_dataframe()
            assert df is not None
