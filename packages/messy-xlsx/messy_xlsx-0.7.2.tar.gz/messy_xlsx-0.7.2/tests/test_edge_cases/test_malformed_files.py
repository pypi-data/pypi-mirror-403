"""Tests for edge cases with malformed or unusual files."""

import tempfile
from pathlib import Path

import openpyxl
import pytest

from messy_xlsx import MessyWorkbook, SheetConfig
from messy_xlsx.exceptions import FileError, FormatError, MessyXlsxError


class TestEmptyFiles:
    """Test handling of empty files."""

    def test_empty_xlsx_file(self, temp_dir):
        """Test parsing completely empty XLSX file."""
        empty_file = temp_dir / "empty.xlsx"
        empty_file.write_bytes(b"")

        with pytest.raises(MessyXlsxError):
            MessyWorkbook(empty_file)

    def test_empty_sheet(self, temp_dir):
        """Test parsing XLSX with empty sheet."""
        file_path = temp_dir / "empty_sheet.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Empty"
        # Don't add any data
        wb.save(file_path)
        wb.close()

        with MessyWorkbook(file_path) as mwb:
            df = mwb.to_dataframe("Empty")
            assert len(df) == 0 or df.empty


class TestHeaderOnlyFiles:
    """Test files with only headers, no data."""

    def test_header_only_file(self, temp_dir):
        """Test file with headers but no data rows."""
        file_path = temp_dir / "header_only.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Name", "Age", "City"])
        wb.save(file_path)
        wb.close()

        from messy_xlsx.models import SheetConfig
        with MessyWorkbook(file_path) as mwb:
            # Disable normalization and sanitization to preserve original column names
            config = SheetConfig(auto_detect=False, header_rows=1, normalize=False, sanitize_column_names=False)
            df = mwb.to_dataframe(config=config)
            assert len(df) == 0
            assert list(df.columns) == ["Name", "Age", "City"]


class TestDataOnlyFiles:
    """Test files with data but no headers."""

    def test_data_only_no_header(self, temp_dir):
        """Test file with data but no clear header row."""
        file_path = temp_dir / "data_only.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([1, 2, 3])
        ws.append([4, 5, 6])
        ws.append([7, 8, 9])
        wb.save(file_path)
        wb.close()

        config = SheetConfig(header_detection_mode="manual", skip_rows=0, header_rows=0)

        with MessyWorkbook(file_path, sheet_config=config) as mwb:
            df = mwb.to_dataframe()
            assert len(df) == 3


class TestAllBlankRows:
    """Test files with all blank rows."""

    def test_all_blank_rows(self, temp_dir):
        """Test file with only blank rows."""
        file_path = temp_dir / "all_blank.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        for _ in range(10):
            ws.append([None, None, None])
        wb.save(file_path)
        wb.close()

        with MessyWorkbook(file_path) as mwb:
            df = mwb.to_dataframe()
            assert len(df) == 0 or df.empty


class TestExtremelyWideFiles:
    """Test files with many columns."""

    def test_1000_columns(self, temp_dir):
        """Test file with 1000 columns."""
        file_path = temp_dir / "wide.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active

        # Header row
        headers = [f"Col_{i}" for i in range(1000)]
        ws.append(headers)

        # Data row
        data = list(range(1000))
        ws.append(data)

        wb.save(file_path)
        wb.close()

        with MessyWorkbook(file_path) as mwb:
            df = mwb.to_dataframe()
            assert len(df.columns) == 1000
            assert len(df) == 1


class TestSingleCellFile:
    """Test file with just one cell."""

    def test_single_cell(self, temp_dir):
        """Test file with only one cell."""
        file_path = temp_dir / "single.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "OnlyCell"
        wb.save(file_path)
        wb.close()

        with MessyWorkbook(file_path) as mwb:
            structure = mwb.get_structure()
            assert structure.data_start_row == 1
            assert structure.data_end_row == 1


class TestCorruptedFiles:
    """Test handling of corrupted files."""

    def test_corrupted_zip(self, temp_dir):
        """Test corrupted ZIP/XLSX file."""
        file_path = temp_dir / "corrupted.xlsx"
        file_path.write_bytes(b"PK\x03\x04" + b"corrupted data" * 100)

        with pytest.raises(MessyXlsxError):
            MessyWorkbook(file_path)

    def test_not_a_zip(self, temp_dir):
        """Test file that claims to be XLSX but isn't."""
        file_path = temp_dir / "fake.xlsx"
        file_path.write_text("This is not an Excel file")

        with pytest.raises(MessyXlsxError):
            MessyWorkbook(file_path)


class TestSpecialCharacters:
    """Test files with special characters."""

    def test_unicode_headers(self, temp_dir):
        """Test headers with unicode characters."""
        file_path = temp_dir / "unicode.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["名前", "年齢", "都市", "Straße", "café"])
        ws.append(["太郎", 30, "東京", "Berlin", "Paris"])
        wb.save(file_path)
        wb.close()

        # Disable sanitization to preserve Unicode headers
        config = SheetConfig(sanitize_column_names=False)
        with MessyWorkbook(file_path, sheet_config=config) as mwb:
            df = mwb.to_dataframe()
            assert len(df) == 1
            assert "名前" in df.columns

    def test_special_char_data(self, temp_dir):
        """Test data with newlines and special characters."""
        file_path = temp_dir / "special.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Text"])
        ws.append(["Line1\nLine2\nLine3"])
        ws.append(["Tab\tSeparated"])
        ws.append(["Quote\"Inside"])
        wb.save(file_path)
        wb.close()

        with MessyWorkbook(file_path) as mwb:
            df = mwb.to_dataframe()
            assert len(df) == 3
