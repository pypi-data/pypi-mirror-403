"""Unit tests for MultiSheetParser."""

import tempfile
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from messy_xlsx.multi_sheet import (
    MultiSheetParser,
    MultiSheetOptions,
    SheetInfo,
    read_all_sheets,
    analyze_excel,
)


# ============================================================================
# Fixtures
# ============================================================================

@pytest.fixture
def temp_dir():
    """Create a temporary directory for test files."""
    with tempfile.TemporaryDirectory() as tmpdir:
        yield Path(tmpdir)


@pytest.fixture
def multi_sheet_xlsx(temp_dir):
    """Create an XLSX with multiple sheet types."""
    file_path = temp_dir / "multi_sheet.xlsx"

    wb = openpyxl.Workbook()

    # Sheet 1: Simple data
    ws1 = wb.active
    ws1.title = "Sales"
    ws1.append(["Product", "Price", "Quantity"])
    ws1.append(["Widget A", 10.50, 100])
    ws1.append(["Widget B", 20.00, 50])
    ws1.append(["Widget C", 15.75, 75])

    # Sheet 2: Empty sheet
    ws2 = wb.create_sheet("Empty")
    # Leave empty

    # Sheet 3: Data with metadata header
    ws3 = wb.create_sheet("Report")
    ws3.append(["Printed Date: 2025-01-10"])
    ws3.append(["Page 1 of 1"])
    ws3.append([])  # Empty row
    ws3.append(["Name", "Department", "Salary"])
    ws3.append(["Alice", "Engineering", 75000])
    ws3.append(["Bob", "Sales", 65000])

    # Sheet 4: Pivot-like table
    ws4 = wb.create_sheet("Summary")
    ws4.append(["Row Labels", "Sum of Sales"])
    ws4.append(["North", 50000])
    ws4.append(["South", 45000])
    ws4.append(["Grand Total", 95000])

    wb.save(file_path)
    wb.close()

    return file_path


@pytest.fixture
def mixed_types_xlsx(temp_dir):
    """Create an XLSX with mixed type columns."""
    file_path = temp_dir / "mixed_types.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    ws.append(["ID", "Code", "Value"])
    ws.append(["001", "A1", 100])
    ws.append(["002", 42, 200])  # Code has int mixed with string
    ws.append(["003", "B3", 300])

    wb.save(file_path)
    wb.close()

    return file_path


@pytest.fixture
def messy_columns_xlsx(temp_dir):
    """Create an XLSX with messy column names."""
    file_path = temp_dir / "messy_columns.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    ws.append(["P/O No.", "Balance Qty.", "Ship To.", "ETD. Date"])
    ws.append(["PO001", 100, "NYC", "2025-01-15"])
    ws.append(["PO002", 200, "LA", "2025-01-20"])

    wb.save(file_path)
    wb.close()

    return file_path


@pytest.fixture
def metadata_variations_xlsx(temp_dir):
    """Create XLSX with various metadata patterns."""
    file_path = temp_dir / "metadata.xlsx"

    wb = openpyxl.Workbook()

    # Sheet with "Generated on" metadata
    ws1 = wb.active
    ws1.title = "Generated"
    ws1.append(["Generated on: 2025-01-10 14:30"])
    ws1.append(["Col1", "Col2"])
    ws1.append(["A", "B"])

    # Sheet with "Report Date" metadata
    ws2 = wb.create_sheet("ReportDate")
    ws2.append(["Report Date: 2025-01-10"])
    ws2.append([])
    ws2.append(["X", "Y", "Z"])
    ws2.append([1, 2, 3])

    # Sheet with "As of" metadata
    ws3 = wb.create_sheet("AsOf")
    ws3.append(["As of: December 2024"])
    ws3.append(["Category", "Amount"])
    ws3.append(["Sales", 1000])

    wb.save(file_path)
    wb.close()

    return file_path


@pytest.fixture
def duplicate_columns_xlsx(temp_dir):
    """Create XLSX with duplicate column names."""
    file_path = temp_dir / "duplicates.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    ws.append(["Name", "Value", "Value", "Name"])
    ws.append(["A", 1, 2, "X"])
    ws.append(["B", 3, 4, "Y"])

    wb.save(file_path)
    wb.close()

    return file_path


@pytest.fixture
def small_sheet_xlsx(temp_dir):
    """Create XLSX with very small sheets."""
    file_path = temp_dir / "small.xlsx"

    wb = openpyxl.Workbook()

    # Single cell sheet
    ws1 = wb.active
    ws1.title = "SingleCell"
    ws1.append(["Only one cell"])

    # Single row sheet
    ws2 = wb.create_sheet("SingleRow")
    ws2.append(["A", "B", "C"])

    # Valid small sheet
    ws3 = wb.create_sheet("Valid")
    ws3.append(["X", "Y"])
    ws3.append([1, 2])
    ws3.append([3, 4])

    wb.save(file_path)
    wb.close()

    return file_path


# ============================================================================
# Tests: Sheet Analysis
# ============================================================================

class TestSheetAnalysis:
    """Test sheet analysis functionality."""

    def test_analyze_detects_empty_sheet(self, multi_sheet_xlsx):
        """Test that empty sheets are detected."""
        infos = analyze_excel(multi_sheet_xlsx)

        empty_sheet = next(i for i in infos if i.name == "Empty")
        assert empty_sheet.is_empty is True
        assert empty_sheet.skip_reason is not None

    def test_analyze_detects_pivot_table(self, multi_sheet_xlsx):
        """Test that pivot tables are detected."""
        infos = analyze_excel(multi_sheet_xlsx)

        pivot_sheet = next(i for i in infos if i.name == "Summary")
        assert pivot_sheet.is_pivot is True
        assert pivot_sheet.skip_reason == "Pivot table"

    def test_analyze_detects_data_sheets(self, multi_sheet_xlsx):
        """Test that data sheets are identified."""
        infos = analyze_excel(multi_sheet_xlsx)

        sales_sheet = next(i for i in infos if i.name == "Sales")
        assert sales_sheet.is_empty is False
        assert sales_sheet.is_pivot is False
        assert sales_sheet.skip_reason is None

    def test_analyze_detects_header_row_with_metadata(self, multi_sheet_xlsx):
        """Test header detection skips metadata rows."""
        infos = analyze_excel(multi_sheet_xlsx)

        report_sheet = next(i for i in infos if i.name == "Report")
        # Should detect headers at row 3 (0-indexed), skipping metadata
        assert report_sheet.header_row == 3

    def test_analyze_returns_sheet_info(self, multi_sheet_xlsx):
        """Test that analysis returns SheetInfo objects."""
        infos = analyze_excel(multi_sheet_xlsx)

        assert len(infos) == 4
        for info in infos:
            assert isinstance(info, SheetInfo)
            assert isinstance(info.name, str)
            assert isinstance(info.row_count, int)
            assert isinstance(info.col_count, int)


class TestMetadataDetection:
    """Test metadata row detection."""

    def test_detects_printed_date(self, multi_sheet_xlsx):
        """Test detection of 'Printed Date' metadata."""
        infos = analyze_excel(multi_sheet_xlsx)
        report = next(i for i in infos if i.name == "Report")
        assert report.header_row == 3

    def test_detects_generated_on(self, metadata_variations_xlsx):
        """Test detection of 'Generated on' metadata."""
        infos = analyze_excel(metadata_variations_xlsx)
        sheet = next(i for i in infos if i.name == "Generated")
        assert sheet.header_row == 1

    def test_detects_report_date(self, metadata_variations_xlsx):
        """Test detection of 'Report Date' metadata."""
        infos = analyze_excel(metadata_variations_xlsx)
        sheet = next(i for i in infos if i.name == "ReportDate")
        assert sheet.header_row == 2

    def test_detects_as_of(self, metadata_variations_xlsx):
        """Test detection of 'As of' metadata."""
        infos = analyze_excel(metadata_variations_xlsx)
        sheet = next(i for i in infos if i.name == "AsOf")
        assert sheet.header_row == 1


# ============================================================================
# Tests: Parsing
# ============================================================================

class TestParsing:
    """Test sheet parsing functionality."""

    def test_parse_all_skips_empty_and_pivot(self, multi_sheet_xlsx):
        """Test that parse_all skips empty and pivot sheets."""
        sheets = read_all_sheets(multi_sheet_xlsx)

        assert "Sales" in sheets
        assert "Report" in sheets
        assert "Empty" not in sheets
        assert "Summary" not in sheets

    def test_parse_all_returns_dataframes(self, multi_sheet_xlsx):
        """Test that parse_all returns DataFrames."""
        sheets = read_all_sheets(multi_sheet_xlsx)

        for name, df in sheets.items():
            assert isinstance(df, pd.DataFrame)
            assert len(df) > 0

    def test_parse_respects_header_detection(self, multi_sheet_xlsx):
        """Test that parsing uses detected header row."""
        sheets = read_all_sheets(multi_sheet_xlsx)

        report = sheets["Report"]
        # Should have correct column names from row 3 (sanitized to lowercase)
        assert "name" in report.columns
        assert "department" in report.columns
        assert "salary" in report.columns
        # Should not have metadata in data
        assert not any("Printed" in str(v) for v in report["name"])

    def test_parse_specific_sheet(self, multi_sheet_xlsx):
        """Test parsing a specific sheet by name."""
        parser = MultiSheetParser(multi_sheet_xlsx)
        df = parser.parse_sheet("Sales")

        assert isinstance(df, pd.DataFrame)
        assert len(df) == 3
        # Column names are sanitized to lowercase by default
        assert "product" in df.columns


class TestColumnCleaning:
    """Test column name cleaning."""

    def test_cleans_special_characters(self, messy_columns_xlsx):
        """Test that special characters are cleaned from column names."""
        sheets = read_all_sheets(messy_columns_xlsx)
        df = sheets["Data"]

        # Check all columns are alphanumeric + underscore
        for col in df.columns:
            assert col.replace("_", "").isalnum(), f"Column '{col}' has invalid chars"

    def test_cleans_slashes_and_dots(self, messy_columns_xlsx):
        """Test specific character replacements."""
        sheets = read_all_sheets(messy_columns_xlsx)
        df = sheets["Data"]

        # P/O No. should become p_o_no (lowercase, no slashes/dots)
        assert any("p" in col and "o" in col for col in df.columns)
        # No slashes or dots in column names
        for col in df.columns:
            assert "/" not in col
            assert "." not in col

    def test_handles_duplicate_columns(self, duplicate_columns_xlsx):
        """Test that duplicate column names are made unique."""
        sheets = read_all_sheets(duplicate_columns_xlsx)
        df = sheets["Data"]

        # All column names should be unique
        assert len(df.columns) == len(set(df.columns))

    def test_can_disable_column_cleaning(self, messy_columns_xlsx):
        """Test that column cleaning can be disabled."""
        sheets = read_all_sheets(
            messy_columns_xlsx,
            clean_column_names=False,
        )
        df = sheets["Data"]

        # Original column names preserved
        assert "P/O No." in df.columns


class TestTypeConsistency:
    """Test type consistency enforcement."""

    def test_mixed_types_converted_to_string(self, mixed_types_xlsx):
        """Test that mixed type columns become all strings."""
        sheets = read_all_sheets(mixed_types_xlsx)
        df = sheets["Data"]

        # Code column had mixed int/str, should all be strings now (column name is lowercase)
        code_types = set(type(v).__name__ for v in df["code"].dropna())
        assert len(code_types) == 1, f"Expected single type, got {code_types}"

    def test_pure_numeric_stays_numeric(self, multi_sheet_xlsx):
        """Test that pure numeric columns stay numeric."""
        sheets = read_all_sheets(multi_sheet_xlsx)
        df = sheets["Sales"]

        # Price and Quantity should be numeric (column names are lowercase)
        assert df["price"].dtype in ["float64", "int64"]
        assert df["quantity"].dtype in ["float64", "int64"]

    def test_can_disable_type_consistency(self, mixed_types_xlsx):
        """Test that type consistency can be disabled."""
        sheets = read_all_sheets(
            mixed_types_xlsx,
            ensure_type_consistency=False,
        )
        df = sheets["Data"]

        # Mixed types may still exist (column name is lowercase)
        code_types = set(type(v).__name__ for v in df["code"].dropna())
        # Could be mixed or not depending on pandas inference
        assert len(code_types) >= 1


class TestSmallSheets:
    """Test handling of very small sheets."""

    def test_skips_single_cell_sheet(self, small_sheet_xlsx):
        """Test that single-cell sheets are skipped."""
        sheets = read_all_sheets(small_sheet_xlsx)
        assert "SingleCell" not in sheets

    def test_skips_single_row_sheet(self, small_sheet_xlsx):
        """Test that single-row sheets are skipped."""
        sheets = read_all_sheets(small_sheet_xlsx)
        assert "SingleRow" not in sheets

    def test_keeps_valid_small_sheet(self, small_sheet_xlsx):
        """Test that valid small sheets are kept."""
        sheets = read_all_sheets(small_sheet_xlsx)
        assert "Valid" in sheets
        assert len(sheets["Valid"]) == 2


# ============================================================================
# Tests: Options
# ============================================================================

class TestOptions:
    """Test MultiSheetOptions functionality."""

    def test_explicit_sheet_list(self, multi_sheet_xlsx):
        """Test parsing only specified sheets."""
        sheets = read_all_sheets(
            multi_sheet_xlsx,
            sheets=["Sales"],
        )

        assert "Sales" in sheets
        assert "Report" not in sheets

    def test_include_pivots_option(self, multi_sheet_xlsx):
        """Test including pivot tables when skip_pivots=False."""
        sheets = read_all_sheets(
            multi_sheet_xlsx,
            skip_pivots=False,
        )

        assert "Summary" in sheets

    def test_custom_min_rows(self, small_sheet_xlsx):
        """Test custom minimum row threshold."""
        # With min_rows=1, single row sheet should be included
        sheets = read_all_sheets(
            small_sheet_xlsx,
            min_rows=1,
        )

        # SingleRow has 1 row of data (header becomes columns)
        # It may still be excluded if it has no data rows after header
        assert "Valid" in sheets

    def test_custom_sheet_filter(self, multi_sheet_xlsx):
        """Test custom sheet filter function."""
        def only_sales(info: SheetInfo) -> bool:
            return "Sales" in info.name

        sheets = read_all_sheets(
            multi_sheet_xlsx,
            sheet_filter=only_sales,
        )

        assert "Sales" in sheets
        assert len(sheets) == 1


# ============================================================================
# Tests: Edge Cases
# ============================================================================

class TestEdgeCases:
    """Test edge cases and error handling."""

    def test_all_sheets_empty(self, temp_dir):
        """Test file where all sheets are empty."""
        file_path = temp_dir / "all_empty.xlsx"
        wb = openpyxl.Workbook()
        wb.active.title = "Empty1"
        wb.create_sheet("Empty2")
        wb.save(file_path)
        wb.close()

        sheets = read_all_sheets(file_path)
        assert len(sheets) == 0

    def test_all_sheets_pivot(self, temp_dir):
        """Test file where all sheets look like pivots."""
        file_path = temp_dir / "all_pivot.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pivot1"
        ws.append(["Row Labels", "Sum of X"])
        ws.append(["A", 100])
        ws.append(["Grand Total", 100])
        wb.save(file_path)
        wb.close()

        sheets = read_all_sheets(file_path)
        assert len(sheets) == 0

    def test_sheet_with_all_null_columns(self, temp_dir):
        """Test sheet with columns that are entirely null."""
        file_path = temp_dir / "null_cols.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.append(["A", "B", "C"])
        ws.append([1, None, 2])
        ws.append([3, None, 4])
        wb.save(file_path)
        wb.close()

        sheets = read_all_sheets(file_path)
        df = sheets["Data"]
        # Column B should be dropped (all null)
        assert "B" not in df.columns

    def test_numeric_looking_headers(self, temp_dir):
        """Test that numeric-looking rows aren't mistaken for headers."""
        file_path = temp_dir / "numeric_first.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.append(["Name", "Value"])
        ws.append(["Item1", 100])
        ws.append(["Item2", 200])
        wb.save(file_path)
        wb.close()

        sheets = read_all_sheets(file_path)
        df = sheets["Data"]
        # Should correctly identify row 0 as headers (column names are lowercase)
        assert "name" in df.columns
        assert len(df) == 2


class TestParserClass:
    """Test MultiSheetParser class directly."""

    def test_parser_initialization(self, multi_sheet_xlsx):
        """Test parser initialization."""
        parser = MultiSheetParser(multi_sheet_xlsx)
        assert parser.file_path == multi_sheet_xlsx

    def test_parser_with_options(self, multi_sheet_xlsx):
        """Test parser with custom options."""
        options = MultiSheetOptions(
            skip_pivots=False,
            clean_column_names=False,
        )
        parser = MultiSheetParser(multi_sheet_xlsx, options)

        sheets = parser.parse_all()
        assert "Summary" in sheets  # Pivot included

    def test_parser_analyze_sheets(self, multi_sheet_xlsx):
        """Test analyze_sheets method."""
        parser = MultiSheetParser(multi_sheet_xlsx)
        infos = parser.analyze_sheets()

        assert len(infos) == 4
        names = [i.name for i in infos]
        assert "Sales" in names
        assert "Empty" in names
        assert "Report" in names
        assert "Summary" in names

    def test_unsupported_file_type(self, temp_dir):
        """Test error on unsupported file type."""
        file_path = temp_dir / "test.txt"
        file_path.write_text("not an excel file")

        with pytest.raises(ValueError, match="Unsupported file type"):
            MultiSheetParser(file_path)
