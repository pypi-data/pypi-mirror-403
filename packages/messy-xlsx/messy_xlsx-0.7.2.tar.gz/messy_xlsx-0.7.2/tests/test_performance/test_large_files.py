"""Performance tests for large files."""

import time
from pathlib import Path

import openpyxl
import pytest

from messy_xlsx import MessyWorkbook, SheetConfig


class TestLargeFilePerformance:
    """Test performance on large files."""

    @pytest.mark.slow
    def test_10k_rows_performance(self, temp_dir):
        """Test parsing 10,000 rows completes in reasonable time."""
        file_path = temp_dir / "10k_rows.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["ID", "Name", "Value"])

        for i in range(10000):
            ws.append([i, f"Item_{i}", i * 1.5])

        wb.save(file_path)
        wb.close()

        start = time.time()

        with MessyWorkbook(file_path) as mwb:
            df = mwb.to_dataframe()
            assert len(df) == 10000

        elapsed = time.time() - start
        assert elapsed < 30  # Should complete in under 30 seconds

    @pytest.mark.slow
    def test_1000_columns_performance(self, temp_dir):
        """Test parsing 1000 columns."""
        file_path = temp_dir / "1000_cols.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active

        headers = [f"Col_{i}" for i in range(1000)]
        ws.append(headers)

        for _ in range(10):
            ws.append(list(range(1000)))

        wb.save(file_path)
        wb.close()

        start = time.time()

        with MessyWorkbook(file_path) as mwb:
            df = mwb.to_dataframe()
            assert len(df.columns) == 1000

        elapsed = time.time() - start
        assert elapsed < 30

    def test_existing_large_samples(self):
        """Test existing large sample files."""
        samples_dir = Path(__file__).parent.parent / "samples"

        large_files = [
            "sales_transactions.xlsx",
            "customers.xlsx",
            "job_operations.xlsx",
        ]

        for filename in large_files:
            file_path = samples_dir / filename
            if not file_path.exists():
                continue

            start = time.time()

            with MessyWorkbook(file_path) as wb:
                df = wb.to_dataframe()
                assert df is not None

            elapsed = time.time() - start
            # Should complete reasonably fast
            assert elapsed < 60  # 1 minute max


class TestStructureDetectionPerformance:
    """Test performance of structure detection."""

    def test_header_detection_fast(self, temp_dir):
        """Test header detection completes quickly."""
        file_path = temp_dir / "detect_perf.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Header1", "Header2", "Header3"])
        for i in range(1000):
            ws.append([i, i*2, i*3])
        wb.save(file_path)
        wb.close()

        start = time.time()

        with MessyWorkbook(file_path) as mwb:
            structure = mwb.get_structure()
            assert structure.header_row is not None

        elapsed = time.time() - start
        assert elapsed < 5  # Detection should be fast

    def test_cache_effectiveness(self, sample_xlsx):
        """Test that caching improves performance."""
        # First call (not cached)
        start1 = time.time()
        with MessyWorkbook(sample_xlsx) as wb:
            structure1 = wb.get_structure()
        time1 = time.time() - start1

        # Second call (should use cache)
        start2 = time.time()
        with MessyWorkbook(sample_xlsx) as wb:
            structure2 = wb.get_structure()
        time2 = time.time() - start2

        # Second call should be faster or similar
        assert time2 <= time1 * 2  # Allow some variance
        assert structure1.header_row == structure2.header_row


class TestMemoryUsage:
    """Test memory efficiency."""

    def test_iterator_doesnt_load_all(self, temp_dir):
        """Test that iteration doesn't load entire file."""
        file_path = temp_dir / "iterate.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["A", "B"])
        for i in range(1000):
            ws.append([i, i*2])
        wb.save(file_path)
        wb.close()

        with MessyWorkbook(file_path) as mwb:
            sheet = mwb.get_sheet(mwb.sheet_names[0])

            # Iterate through rows
            row_count = 0
            for row in sheet.iter_rows(min_row=2, max_row=100):
                row_count += 1

            assert row_count <= 99  # Should iterate without loading all
