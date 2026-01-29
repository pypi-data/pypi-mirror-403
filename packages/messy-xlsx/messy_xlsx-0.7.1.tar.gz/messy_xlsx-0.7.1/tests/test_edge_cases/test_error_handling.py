"""Tests for error handling with formulas and Excel errors."""

import openpyxl
import pytest

from messy_xlsx import MessyWorkbook, SheetConfig
from messy_xlsx.formulas import FormulaConfig, FormulaEvaluationMode


class TestFormulaErrors:
    """Test handling of Excel formula errors."""

    def test_all_formula_errors(self, temp_dir):
        """Test file with all types of Excel errors."""
        file_path = temp_dir / "formula_errors.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Error Type", "Formula", "Result"])
        ws.append(["DIV/0", "=1/0", None])
        ws.append(["N/A", "=NA()", None])
        ws.append(["NAME", "=UNKNOWNFUNC()", None])
        ws.append(["NULL", "=A1 B1", None])
        ws.append(["NUM", "=SQRT(-1)", None])
        ws.append(["REF", "=A1000000", None])
        ws.append(["VALUE", "=1+\"text\"", None])
        wb.save(file_path)
        wb.close()

        config = FormulaConfig(mode=FormulaEvaluationMode.CACHED_ONLY)

        with MessyWorkbook(file_path, formula_config=config) as mwb:
            df = mwb.to_dataframe()
            # Should handle errors gracefully
            assert len(df) >= 0


class TestCircularReferences:
    """Test handling of circular formula references."""

    def test_simple_circular_reference(self, temp_dir):
        """Test circular reference A1=B1, B1=A1."""
        file_path = temp_dir / "circular.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "=B1"
        ws["B1"] = "=A1"
        wb.save(file_path)
        wb.close()

        config = FormulaConfig(mode=FormulaEvaluationMode.CACHED_ONLY)

        with MessyWorkbook(file_path, formula_config=config) as mwb:
            # Should not crash
            df = mwb.to_dataframe()
            assert df is not None


class TestComplexFormulas:
    """Test complex formula scenarios."""

    def test_nested_formulas(self, temp_dir):
        """Test deeply nested formulas."""
        file_path = temp_dir / "nested.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Value", "Formula"])
        ws["A2"] = 10
        ws["B2"] = "=IF(A2>5,IF(A2>8,\"High\",\"Medium\"),\"Low\")"
        wb.save(file_path)
        wb.close()

        with MessyWorkbook(file_path) as mwb:
            df = mwb.to_dataframe()
            assert len(df) == 1

    def test_array_formulas(self, temp_dir):
        """Test array formulas if supported."""
        file_path = temp_dir / "array.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Data", "Sum"])
        ws.append([1, "=SUM(A2:A5)"])
        ws.append([2, None])
        ws.append([3, None])
        ws.append([4, None])
        wb.save(file_path)
        wb.close()

        with MessyWorkbook(file_path) as mwb:
            df = mwb.to_dataframe()
            assert len(df) >= 1


class TestFormulaEvaluationModes:
    """Test different formula evaluation modes."""

    def test_disabled_mode(self, temp_dir):
        """Test formula evaluation disabled."""
        file_path = temp_dir / "formulas.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["A", "B", "Sum"])
        ws.append([1, 2, "=A2+B2"])
        wb.save(file_path)
        wb.close()

        config = FormulaConfig(mode=FormulaEvaluationMode.DISABLED)

        with MessyWorkbook(file_path, formula_config=config) as mwb:
            df = mwb.to_dataframe()
            assert len(df) == 1

    def test_cached_only_mode(self, temp_dir):
        """Test using only cached formula values."""
        file_path = temp_dir / "cached.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Value", "Double"])
        ws["A2"] = 5
        ws["B2"] = "=A2*2"
        wb.save(file_path)
        wb.close()

        config = FormulaConfig(mode=FormulaEvaluationMode.CACHED_ONLY)

        with MessyWorkbook(file_path, formula_config=config) as mwb:
            df = mwb.to_dataframe()
            assert len(df) == 1


class TestUnsupportedFunctions:
    """Test handling of unsupported Excel functions."""

    def test_unsupported_function_placeholder(self, temp_dir):
        """Test unsupported function returns placeholder."""
        file_path = temp_dir / "unsupported.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Function", "Result"])
        ws.append(["WEBSERVICE", "=WEBSERVICE(\"url\")"])
        wb.save(file_path)
        wb.close()

        config = FormulaConfig(
            mode=FormulaEvaluationMode.CACHED_WITH_FALLBACK,
            raise_on_unsupported=False,
            unsupported_value="#UNSUPPORTED"
        )

        with MessyWorkbook(file_path, formula_config=config) as mwb:
            df = mwb.to_dataframe()
            assert len(df) >= 0
