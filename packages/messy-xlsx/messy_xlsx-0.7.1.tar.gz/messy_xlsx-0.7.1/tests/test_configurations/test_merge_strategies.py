"""Tests for merged cell handling strategies."""

import openpyxl
import pytest

from messy_xlsx import MessyWorkbook, SheetConfig


class TestMergeStrategies:
    """Test different merge strategies."""

    @pytest.mark.parametrize("strategy", ["fill", "skip", "first_only"])
    def test_merge_strategies(self, temp_dir, strategy):
        """Test all merge strategies."""
        file_path = temp_dir / f"merged_{strategy}.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active

        # Create merged cells
        ws["A1"] = "Merged Header"
        ws.merge_cells("A1:C1")

        ws.append(["Col1", "Col2", "Col3"])
        ws.append([1, 2, 3])
        ws.append([4, 5, 6])

        wb.save(file_path)
        wb.close()

        config = SheetConfig(merge_strategy=strategy)

        with MessyWorkbook(file_path, sheet_config=config) as mwb:
            df = mwb.to_dataframe()
            assert df is not None

    def test_heavily_merged_file_fill_strategy(self, temp_dir):
        """Test fill strategy on heavily merged file."""
        file_path = temp_dir / "heavy_merge.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active

        ws["A1"] = "Title"
        ws.merge_cells("A1:E1")

        ws.append(["H1", "H2", "H3", "H4", "H5"])

        # Merged data cells
        ws["A3"] = "Region A"
        ws.merge_cells("A3:A5")
        ws["B3"] = 100
        ws["B4"] = 200
        ws["B5"] = 300

        wb.save(file_path)
        wb.close()

        config = SheetConfig(merge_strategy="fill")

        with MessyWorkbook(file_path, sheet_config=config) as mwb:
            df = mwb.to_dataframe()
            assert df is not None

    def test_skip_merged_cells(self, merged_cells_xlsx):
        """Test skip strategy ignores merged cells."""
        config = SheetConfig(merge_strategy="skip")

        with MessyWorkbook(merged_cells_xlsx, sheet_config=config) as wb:
            df = wb.to_dataframe()
            assert df is not None
