"""Tests for v0.3.0 fixes.

Tests for:
1. Date parsing fix - integers not converted to dates
2. BytesIO/file-like object support
3. Normalization toggle
4. Pandas deprecation warnings
"""

import io
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from messy_xlsx import MessyWorkbook, SheetConfig


class TestDateParsingFix:
    """Tests for the aggressive date parsing fix."""

    def test_integer_column_not_converted_to_date(self, temp_dir):
        """Integer columns like Total_Transactions should stay as integers."""
        file_path = temp_dir / "sales_summary.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Date", "Total_Transactions", "Unique_Customers", "Revenue"])
        ws.append(["2024-01-01", 211, 142, 15000.50])
        ws.append(["2024-01-02", 198, 128, 14200.75])
        ws.append(["2024-01-03", 256, 175, 18500.25])
        wb.save(file_path)
        wb.close()

        with MessyWorkbook(file_path) as mwb:
            df = mwb.to_dataframe()

            # Total_Transactions should be integer, not datetime (column names are lowercase)
            assert df["total_transactions"].dtype in (int, "int64", "Int64", float, "float64")
            assert df["total_transactions"].iloc[0] == 211

            # Unique_Customers should be integer, not datetime
            assert df["unique_customers"].dtype in (int, "int64", "Int64", float, "float64")
            assert df["unique_customers"].iloc[0] == 142

    def test_count_column_not_converted_to_date(self, temp_dir):
        """Columns with 'count' in the name should not be converted."""
        file_path = temp_dir / "counts.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["item_count", "order_count", "user_count"])
        ws.append([100, 50, 25])
        ws.append([200, 75, 35])
        wb.save(file_path)
        wb.close()

        with MessyWorkbook(file_path) as mwb:
            df = mwb.to_dataframe()

            for col in ["item_count", "order_count", "user_count"]:
                assert df[col].dtype in (int, "int64", "Int64", float, "float64")
                assert not pd.api.types.is_datetime64_any_dtype(df[col])

    def test_actual_date_column_is_converted(self, temp_dir):
        """Columns with date-like names should still be converted."""
        file_path = temp_dir / "dates.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["order_date", "created_timestamp", "value"])
        ws.append(["2024-01-01", "2024-01-01 10:00:00", 100])
        ws.append(["2024-01-02", "2024-01-02 11:00:00", 200])
        wb.save(file_path)
        wb.close()

        with MessyWorkbook(file_path) as mwb:
            df = mwb.to_dataframe()

            # Text date columns should be converted to datetime
            assert pd.api.types.is_datetime64_any_dtype(df["order_date"])
            assert pd.api.types.is_datetime64_any_dtype(df["created_timestamp"])

    def test_explicit_timestamp_hint_converts_numeric(self, temp_dir):
        """Explicit TIMESTAMP hint should convert numeric values to dates."""
        file_path = temp_dir / "explicit_dates.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["serial_date", "value"])
        ws.append([45292, 100])  # Excel serial date for 2024-01-01
        ws.append([45293, 200])  # Excel serial date for 2024-01-02
        wb.save(file_path)
        wb.close()

        config = SheetConfig(
            type_hints={"serial_date": "TIMESTAMP"}
        )

        with MessyWorkbook(file_path, sheet_config=config) as mwb:
            df = mwb.to_dataframe()

            # With explicit TIMESTAMP hint, should convert
            assert pd.api.types.is_datetime64_any_dtype(df["serial_date"])


class TestBytesIOSupport:
    """Tests for BytesIO/file-like object support."""

    def test_read_from_bytesio(self, temp_dir):
        """Should be able to read from BytesIO object."""
        # Create a file first
        file_path = temp_dir / "test.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Name", "Value"])
        ws.append(["Alice", 100])
        ws.append(["Bob", 200])
        wb.save(file_path)
        wb.close()

        # Read as bytes and create BytesIO
        with open(file_path, "rb") as f:
            content = f.read()

        buffer = io.BytesIO(content)

        # Should work with BytesIO
        with MessyWorkbook(buffer) as mwb:
            df = mwb.to_dataframe()
            assert len(df) == 2
            # Column names are sanitized by default (lowercased)
            assert "name" in df.columns
            assert "value" in df.columns

    def test_read_from_bytesio_with_filename_hint(self, temp_dir):
        """Should accept filename hint for format detection."""
        file_path = temp_dir / "test.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["A", "B"])
        ws.append([1, 2])
        wb.save(file_path)
        wb.close()

        with open(file_path, "rb") as f:
            content = f.read()

        buffer = io.BytesIO(content)

        with MessyWorkbook(buffer, filename="data.xlsx") as mwb:
            assert mwb.format_type == "xlsx"

    def test_bytesio_repr(self, temp_dir):
        """Repr should work for BytesIO input."""
        file_path = temp_dir / "test.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["A", "B"])
        wb.save(file_path)
        wb.close()

        with open(file_path, "rb") as f:
            content = f.read()

        # Without filename hint
        buffer = io.BytesIO(content)
        with MessyWorkbook(buffer) as mwb:
            assert "<stream>" in repr(mwb)

        # With filename hint
        buffer = io.BytesIO(content)
        with MessyWorkbook(buffer, filename="myfile.xlsx") as mwb:
            assert "myfile.xlsx" in repr(mwb)

    def test_bytesio_file_path_is_none(self, temp_dir):
        """file_path property should be None for BytesIO input."""
        file_path = temp_dir / "test.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["A", "B"])
        wb.save(file_path)
        wb.close()

        with open(file_path, "rb") as f:
            content = f.read()

        buffer = io.BytesIO(content)

        with MessyWorkbook(buffer) as mwb:
            assert mwb.file_path is None
            assert mwb.source is buffer

    def test_bytesio_multiple_reads(self, temp_dir):
        """Should be able to read multiple times from same BytesIO."""
        file_path = temp_dir / "multisheet.xlsx"

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1.append(["A", "B"])
        ws1.append([1, 2])

        ws2 = wb.create_sheet("Sheet2")
        ws2.append(["C", "D"])
        ws2.append([3, 4])

        wb.save(file_path)
        wb.close()

        with open(file_path, "rb") as f:
            content = f.read()

        buffer = io.BytesIO(content)

        with MessyWorkbook(buffer) as mwb:
            df1 = mwb.to_dataframe(sheet="Sheet1")
            df2 = mwb.to_dataframe(sheet="Sheet2")

            assert len(df1) == 1
            assert len(df2) == 1
            # Column names are sanitized by default (lowercased)
            assert "a" in df1.columns
            assert "c" in df2.columns


class TestNormalizationToggle:
    """Tests for normalization toggle options."""

    def test_normalize_disabled(self, temp_dir):
        """normalize=False should disable all normalization."""
        file_path = temp_dir / "raw.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Number", "Date"])
        ws.append(["  1,234.56  ", "2024-01-01"])
        wb.save(file_path)
        wb.close()

        # Also disable sanitization to preserve original column names
        config = SheetConfig(normalize=False, sanitize_column_names=False)

        with MessyWorkbook(file_path, sheet_config=config) as mwb:
            df = mwb.to_dataframe()

            # Should be raw string values
            assert df["Number"].iloc[0] == "  1,234.56  "  # Whitespace preserved
            assert df["Date"].iloc[0] == "2024-01-01"  # Not converted to datetime

    def test_normalize_dates_disabled(self, temp_dir):
        """normalize_dates=False should prevent date conversion."""
        file_path = temp_dir / "dates.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["date_column"])
        ws.append(["2024-01-01"])
        ws.append(["2024-01-02"])
        wb.save(file_path)
        wb.close()

        config = SheetConfig(normalize_dates=False)

        with MessyWorkbook(file_path, sheet_config=config) as mwb:
            df = mwb.to_dataframe()

            # Date strings should remain as strings (object or StringDtype)
            assert df["date_column"].dtype == object or isinstance(df["date_column"].dtype, pd.StringDtype)
            assert df["date_column"].iloc[0] == "2024-01-01"

    def test_normalize_numbers_disabled(self, temp_dir):
        """normalize_numbers=False should prevent number parsing."""
        file_path = temp_dir / "numbers.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["amount"])
        ws.append(["1,234.56"])
        ws.append(["$2,345.67"])
        wb.save(file_path)
        wb.close()

        config = SheetConfig(normalize_numbers=False)

        with MessyWorkbook(file_path, sheet_config=config) as mwb:
            df = mwb.to_dataframe()

            # Number strings should remain as strings (object or StringDtype)
            assert df["amount"].dtype == object or isinstance(df["amount"].dtype, pd.StringDtype)

    def test_normalize_whitespace_disabled(self, temp_dir):
        """normalize_whitespace=False should preserve whitespace."""
        file_path = temp_dir / "whitespace.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["text"])
        ws.append(["  hello  world  "])
        wb.save(file_path)
        wb.close()

        config = SheetConfig(normalize_whitespace=False)

        with MessyWorkbook(file_path, sheet_config=config) as mwb:
            df = mwb.to_dataframe()

            # Whitespace should be preserved
            # Note: Leading/trailing spaces may still be stripped by other processes
            # This test ensures the whitespace normalizer is skipped

    def test_all_normalization_enabled_by_default(self, temp_dir):
        """Default config should enable all normalization."""
        file_path = temp_dir / "data.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["date", "amount"])
        ws.append(["2024-01-01", "1,234.56"])
        wb.save(file_path)
        wb.close()

        config = SheetConfig()

        assert config.normalize is True
        assert config.normalize_dates is True
        assert config.normalize_numbers is True
        assert config.normalize_whitespace is True


class TestPandasDeprecationFixes:
    """Tests to ensure no deprecation warnings with pandas 2.x."""

    def test_no_infer_datetime_format_warning(self, temp_dir):
        """Should not produce infer_datetime_format deprecation warning."""
        import warnings

        file_path = temp_dir / "dates.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["date"])
        ws.append(["2024-01-01"])
        ws.append(["2024/01/02"])
        ws.append(["Jan 3, 2024"])
        wb.save(file_path)
        wb.close()

        with warnings.catch_warnings(record=True) as w:
            warnings.simplefilter("always")

            with MessyWorkbook(file_path) as mwb:
                df = mwb.to_dataframe()

            # Filter for the specific deprecation
            datetime_warnings = [
                warning for warning in w
                if "infer_datetime_format" in str(warning.message)
            ]
            assert len(datetime_warnings) == 0, f"Found warnings: {datetime_warnings}"


class TestColumnNamePatternMatching:
    """Tests for column name pattern matching in date detection."""

    @pytest.mark.parametrize("column_name,should_be_date", [
        # Date-like names (should be converted if values look like dates)
        ("date", True),
        ("created_at", True),
        ("updated_time", True),
        ("birth_date", True),
        ("start_date", True),
        ("end_date", True),
        # Non-date names (should NOT be converted even if values are in range)
        ("count", False),
        ("total", False),
        ("quantity", False),
        ("transaction_id", False),
        ("user_count", False),
        ("order_total", False),
        ("item_qty", False),
        ("price", False),
        ("amount", False),
        ("score", False),
        ("rank", False),
        ("age", False),
    ])
    def test_column_name_date_detection(self, temp_dir, column_name, should_be_date):
        """Test that column names affect date detection for numeric values."""
        file_path = temp_dir / f"test_{column_name}.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([column_name])
        ws.append([100])  # Value in Excel date range
        ws.append([200])
        ws.append([300])
        wb.save(file_path)
        wb.close()

        with MessyWorkbook(file_path) as mwb:
            df = mwb.to_dataframe()

            is_datetime = pd.api.types.is_datetime64_any_dtype(df[column_name])

            if should_be_date:
                # For date-like columns with numeric values in range, may or may not convert
                # depending on other heuristics - just check it doesn't crash
                pass
            else:
                # For non-date columns, should NOT be datetime
                assert not is_datetime, f"Column '{column_name}' was incorrectly converted to datetime"
