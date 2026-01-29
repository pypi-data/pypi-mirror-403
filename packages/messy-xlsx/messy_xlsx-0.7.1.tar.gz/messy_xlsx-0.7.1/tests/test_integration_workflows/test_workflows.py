"""Integration tests for common workflows."""

from pathlib import Path

import openpyxl
import pytest

from messy_xlsx import MessyWorkbook, SheetConfig, read_excel


class TestEndToEndWorkflows:
    """Test complete workflows from file to processed data."""

    def test_simple_read_workflow(self, sample_xlsx):
        """Test simplest workflow: read_excel()."""
        df = read_excel(sample_xlsx)

        assert df is not None
        assert len(df) > 0
        assert len(df.columns) > 0

    def test_multisheet_workflow(self, temp_dir):
        """Test multi-sheet workbook workflow."""
        file_path = temp_dir / "multisheet.xlsx"

        wb = openpyxl.Workbook()

        # Sheet 1
        ws1 = wb.active
        ws1.title = "Sales"
        ws1.append(["Product", "Amount"])
        ws1.append(["A", 100])
        ws1.append(["B", 200])

        # Sheet 2
        ws2 = wb.create_sheet("Inventory")
        ws2.append(["Product", "Stock"])
        ws2.append(["A", 50])
        ws2.append(["B", 75])

        wb.save(file_path)
        wb.close()

        with MessyWorkbook(file_path) as mwb:
            # Get all sheets
            assert len(mwb.sheet_names) == 2

            # Parse all sheets
            dfs = mwb.to_dataframes()
            assert len(dfs) == 2
            assert "Sales" in dfs
            assert "Inventory" in dfs

            # Parse specific sheet
            sales_df = mwb.to_dataframe("Sales")
            assert len(sales_df) == 2

    def test_structure_then_parse_workflow(self, messy_xlsx):
        """Test analyzing structure before parsing."""
        with MessyWorkbook(messy_xlsx) as wb:
            # 1. Analyze structure
            structure = wb.get_structure()

            # 2. Create config based on analysis
            config = SheetConfig(
                skip_rows=structure.suggested_skip_rows,
                locale=structure.detected_locale
            )

            # 3. Parse with config
            df = wb.to_dataframe(config=config)

            assert df is not None

    def test_cell_access_workflow(self, sample_xlsx):
        """Test accessing individual cells."""
        with MessyWorkbook(sample_xlsx) as wb:
            # Access by sheet/row/col
            cell1 = wb.get_cell("Data", 1, 1)
            assert cell1 is not None

            # Access by reference
            cell2 = wb.get_cell_by_ref("Data!A1")
            assert cell2 is not None

            # Both should return same value
            assert cell1.value == cell2.value

    def test_normalization_workflow(self, temp_dir):
        """Test full normalization pipeline."""
        file_path = temp_dir / "normalize.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Name", "Amount", "Date"])
        ws.append(["  Alice  ", "$1,234.56", "2024-01-15"])
        ws.append(["Bob", "  NA  ", "2024-01-16"])
        wb.save(file_path)
        wb.close()

        with MessyWorkbook(file_path) as mwb:
            df = mwb.to_dataframe()

            # Whitespace should be cleaned
            # Column names are sanitized by default (lowercased)
            assert df.iloc[0]["name"] == "Alice"

            # Missing values should be standardized
            import pandas as pd
            assert pd.isna(df.iloc[1]["amount"])


class TestConfigurationChaining:
    """Test chaining configuration options."""

    def test_combined_config_options(self, messy_xlsx):
        """Test multiple config options together."""
        config = SheetConfig(
            auto_detect=True,
            header_detection_mode="smart",
            merge_strategy="fill",
            locale="auto",
            evaluate_formulas=True,
            include_hidden=False
        )

        with MessyWorkbook(messy_xlsx, sheet_config=config) as wb:
            df = wb.to_dataframe()
            assert df is not None

    def test_type_hints_and_renames(self, temp_dir):
        """Test type hints with column renames."""
        file_path = temp_dir / "config_chain.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["id", "name", "amount"])
        ws.append(["00123", "Item", "1234.56"])
        wb.save(file_path)
        wb.close()

        config = SheetConfig(
            type_hints={"id": "VARCHAR"},  # Preserve leading zeros
            column_renames={"id": "customer_id", "name": "item_name"}
        )

        with MessyWorkbook(file_path, sheet_config=config) as mwb:
            df = mwb.to_dataframe()
            assert "customer_id" in df.columns
            assert "item_name" in df.columns


class TestErrorRecovery:
    """Test recovery from errors."""

    def test_continue_after_sheet_error(self, temp_dir):
        """Test processing continues after one sheet fails."""
        file_path = temp_dir / "mixed_sheets.xlsx"

        wb = openpyxl.Workbook()

        # Good sheet
        ws1 = wb.active
        ws1.title = "Good"
        ws1.append(["A", "B"])
        ws1.append([1, 2])

        # Another good sheet
        ws2 = wb.create_sheet("AlsoGood")
        ws2.append(["X", "Y"])
        ws2.append([3, 4])

        wb.save(file_path)
        wb.close()

        with MessyWorkbook(file_path) as mwb:
            dfs = mwb.to_dataframes()
            assert len(dfs) >= 1  # At least one sheet should parse
