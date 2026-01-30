"""Data models for messy-xlsx."""

# ============================================================================
# Imports
# ============================================================================

from dataclasses import dataclass, field
from typing import Any


# ============================================================================
# Format Models
# ============================================================================

@dataclass
class FormatInfo:
    """Information about detected file format."""

    format_type: str
    confidence: float = 1.0
    version: str | None = None
    encoding: str | None = None
    has_macros: bool = False
    is_encrypted: bool = False
    is_compressed: bool = False


# ============================================================================
# Structure Models
# ============================================================================

@dataclass
class StructureInfo:
    """Results from structure analysis of a sheet."""

    data_start_row: int
    data_end_row: int
    data_start_col: int
    data_end_col: int
    header_row: int | None
    header_rows_count: int
    header_confidence: float
    metadata_rows: list[int] = field(default_factory=list)
    metadata_type: str = "unknown"
    merged_ranges: list[tuple[int, int, int, int]] = field(default_factory=list)
    merged_in_headers: bool = False
    merged_in_data: bool = False
    hidden_rows: list[int] = field(default_factory=list)
    hidden_columns: list[int] = field(default_factory=list)
    detected_locale: str = "en_US"
    decimal_separator: str = "."
    thousands_separator: str = ","
    num_tables: int = 1
    table_ranges: list[dict[str, Any]] = field(default_factory=list)
    blank_rows: list[int] = field(default_factory=list)
    inconsistent_columns: bool = False
    has_formulas: bool = False
    suggested_skip_rows: int = 0
    suggested_skip_footer: int = 0
    suggested_overrides: dict[str, Any] = field(default_factory=dict)


@dataclass
class TableInfo:
    """Information about a detected table within a sheet."""

    start_row: int
    end_row: int
    start_col: int
    end_col: int
    has_header: bool = True
    header_row: int | None = None
    confidence: float = 1.0
    title_rows: list[int] = field(default_factory=list)

    @property
    def row_count(self) -> int:
        """Number of data rows (excluding header)."""
        header_offset = 1 if self.has_header else 0
        return self.end_row - self.start_row + 1 - header_offset

    @property
    def column_count(self) -> int:
        """Number of columns."""
        return self.end_col - self.start_col + 1

    def to_range_string(self) -> str:
        """Convert to Excel range notation (e.g., 'A1:F100')."""
        from openpyxl.utils import get_column_letter

        start_col_letter = get_column_letter(self.start_col)
        end_col_letter = get_column_letter(self.end_col)
        return f"{start_col_letter}{self.start_row}:{end_col_letter}{self.end_row}"


# ============================================================================
# Configuration Models
# ============================================================================

@dataclass
class SheetConfig:
    """Configuration options for parsing a sheet."""

    skip_rows: int = 0
    header_rows: int = 1
    skip_footer: int = 0
    cell_range: str | None = None
    column_renames: dict[str, str] = field(default_factory=dict)
    type_hints: dict[str, str] = field(default_factory=dict)
    auto_detect: bool = True
    include_hidden: bool = False
    merge_strategy: str = "fill"
    locale: str | None = None
    evaluate_formulas: bool = True
    drop_regex: str | None = None
    drop_conditions: list[dict[str, Any]] = field(default_factory=list)

    # Header detection configuration
    header_detection_mode: str = "smart"
    header_confidence_threshold: float = 0.7
    header_fallback: str = "first_row"
    multi_row_headers: bool = False
    header_patterns: list[str] | None = None

    # Normalization configuration
    normalize: bool = True  # Master switch for all normalization
    normalize_dates: bool = True  # Convert date-like columns to datetime
    normalize_numbers: bool = True  # Convert number-like strings to numeric
    normalize_whitespace: bool = True  # Clean whitespace in text columns

    # Column name sanitization (ON by default for BigQuery compatibility)
    sanitize_column_names: bool = True


# ============================================================================
# Cell Models
# ============================================================================

@dataclass
class CellValue:
    """Represents a single cell's value with metadata."""

    value: Any
    formula: str | None = None
    is_merged: bool = False
    is_hidden: bool = False
    data_type: str = "empty"
    original_format: str | None = None

    @property
    def is_formula(self) -> bool:
        """True if cell contains a formula."""
        return self.formula is not None

    @property
    def is_error(self) -> bool:
        """True if cell contains an Excel error value."""
        if isinstance(self.value, str):
            return self.value.startswith("#") and self.value.endswith("!")
        return False

    def __repr__(self) -> str:
        if self.formula:
            return f"CellValue({self.value!r}, formula={self.formula!r})"
        return f"CellValue({self.value!r})"
