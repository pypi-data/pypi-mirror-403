"""XLSX/XLSM file handler using fastexcel with openpyxl fallback."""

# ============================================================================
# Imports
# ============================================================================

import fastexcel
import numpy as np
import openpyxl
import pandas as pd

from messy_xlsx.exceptions        import FileError, FormatError
from messy_xlsx.parsing.base_handler import FileSource, FormatHandler, ParseOptions


# ============================================================================
# Constants
# ============================================================================

EXCEL_ERRORS = {
    "#DIV/0!",
    "#N/A",
    "#NAME?",
    "#NULL!",
    "#NUM!",
    "#REF!",
    "#VALUE!",
    "#GETTING_DATA",
}


# ============================================================================
# Helpers
# ============================================================================

def _is_fileobj(source: FileSource) -> bool:
    return hasattr(source, "read")


def _reset_buffer(source: FileSource) -> None:
    if hasattr(source, "seek"):
        source.seek(0)


def _get_file_desc(source: FileSource) -> str:
    return "<stream>" if _is_fileobj(source) else str(source)


def _read_file_content(source: FileSource) -> bytes:
    """Read content from file-like object for fastexcel."""
    _reset_buffer(source)
    return source.read()


# ============================================================================
# Core
# ============================================================================

class XLSXHandler(FormatHandler):
    """Handler for XLSX and XLSM files."""

    def can_handle(self, format_type: str) -> bool:
        return format_type in ("xlsx", "xlsm")

    def parse(
        self,
        file_source: FileSource,
        sheet:       str | None,
        options:     ParseOptions,
    ) -> pd.DataFrame:
        """Parse XLSX/XLSM file to DataFrame."""
        needs_openpyxl = (
            options.merge_strategy != "skip"
            or options.ignore_hidden
            or options.cell_range
        )

        if needs_openpyxl:
            return self._parse_openpyxl(file_source, sheet, options)
        return self._parse_fastexcel(file_source, sheet, options)


    # -------------------------------------------------------------------------
    # Fastexcel (fast path)
    # -------------------------------------------------------------------------

    def _parse_fastexcel(
        self,
        file_source: FileSource,
        sheet:       str | None,
        options:     ParseOptions,
    ) -> pd.DataFrame:
        """Fast path using fastexcel."""
        is_fileobj = _is_fileobj(file_source)
        file_desc  = _get_file_desc(file_source)

        if is_fileobj:
            _reset_buffer(file_source)

        try:
            content    = _read_file_content(file_source) if is_fileobj else file_source
            excel_file = fastexcel.read_excel(content)
        except PermissionError as e:
            raise FileError(f"Permission denied: {file_desc}", file_path=file_desc, operation="open") from e
        except Exception as e:
            raise FormatError(f"Cannot open Excel file: {e}", file_path=file_desc) from e

        try:
            sheet_idx = self._resolve_sheet_index(excel_file, sheet, file_desc)

            df = excel_file.load_sheet(
                sheet_idx,
                skip_rows  = options.skip_rows,
                header_row = None,
            ).to_pandas()

            if df.empty:
                return pd.DataFrame()

            df = self._apply_options(df, options)
            return self._clean_excel_data(df, options)

        except FormatError:
            raise
        except Exception as e:
            raise FormatError(f"Error reading Excel file: {e}", file_path=file_desc) from e

    def _resolve_sheet_index(self, excel_file, sheet: str | None, file_desc: str) -> int:
        if not sheet:
            return 0
        if sheet not in excel_file.sheet_names:
            raise FormatError(f"Sheet '{sheet}' not found", file_path=file_desc, detected_format="xlsx")
        return excel_file.sheet_names.index(sheet)


    # -------------------------------------------------------------------------
    # Openpyxl (fallback for advanced features)
    # -------------------------------------------------------------------------

    def _parse_openpyxl(
        self,
        file_source: FileSource,
        sheet:       str | None,
        options:     ParseOptions,
    ) -> pd.DataFrame:
        """Fallback path using openpyxl for advanced features."""
        read_only  = options.merge_strategy == "skip"
        is_fileobj = _is_fileobj(file_source)
        file_desc  = _get_file_desc(file_source)

        if is_fileobj:
            _reset_buffer(file_source)

        try:
            wb = openpyxl.load_workbook(
                file_source,
                read_only  = read_only,
                data_only  = options.data_only,
                keep_links = False,
            )
        except PermissionError as e:
            raise FileError(f"Permission denied: {file_desc}", file_path=file_desc, operation="open") from e
        except Exception as e:
            raise FormatError(f"Cannot open Excel file: {e}", file_path=file_desc) from e

        try:
            ws = self._resolve_worksheet(wb, sheet, file_desc)

            if options.merge_strategy != "skip" and not read_only:
                self._handle_merged_cells(ws, options.merge_strategy)

            data = self._read_worksheet(ws, options)
            if not data:
                return pd.DataFrame()

            df = pd.DataFrame(data)

            if options.skip_rows > 0 and not options.cell_range:
                df = df.iloc[options.skip_rows:]

            df = self._apply_options(df, options)
            return self._clean_excel_data(df, options)

        finally:
            wb.close()

    def _resolve_worksheet(self, wb, sheet: str | None, file_desc: str):
        if not sheet:
            return wb.active
        if sheet not in wb.sheetnames:
            raise FormatError(f"Sheet '{sheet}' not found", file_path=file_desc, detected_format="xlsx")
        return wb[sheet]

    def _read_worksheet(self, ws, options: ParseOptions) -> list[list]:
        """Read worksheet data using openpyxl."""
        if options.cell_range:
            return self._read_cell_range(ws, options.cell_range)
        return self._read_full_sheet(ws, options.ignore_hidden)

    def _read_cell_range(self, ws, cell_range: str) -> list[list]:
        try:
            rows_iter = ws[cell_range]
            if not isinstance(rows_iter[0], tuple):
                rows_iter = [rows_iter]

            data = []
            for row in rows_iter:
                if not isinstance(row, tuple):
                    row = [row]
                data.append([cell.value for cell in row])
            return data

        except Exception as e:
            raise FormatError(f"Invalid cell range: {cell_range}", detected_format="xlsx") from e

    def _read_full_sheet(self, ws, ignore_hidden: bool) -> list[list]:
        hidden_rows, hidden_cols = set(), set()

        if ignore_hidden:
            hidden_rows = {r for r, d in ws.row_dimensions.items() if d.hidden}
            hidden_cols = {c for c, d in ws.column_dimensions.items() if d.hidden}

        data = []
        for row_idx, row in enumerate(ws.iter_rows(values_only=False), start=1):
            if row_idx in hidden_rows:
                continue

            row_values = [
                cell.value for cell in row
                if cell.column_letter not in hidden_cols
            ]
            data.append(row_values)

        return data

    def _handle_merged_cells(self, ws, strategy: str) -> None:
        """Handle merged cells according to strategy."""
        for merged_range in list(ws.merged_cells.ranges):
            top_left = ws.cell(merged_range.min_row, merged_range.min_col).value
            ws.unmerge_cells(str(merged_range))

            if strategy == "fill":
                for r in range(merged_range.min_row, merged_range.max_row + 1):
                    for c in range(merged_range.min_col, merged_range.max_col + 1):
                        ws.cell(r, c).value = top_left

            elif strategy == "first_only":
                for r in range(merged_range.min_row, merged_range.max_row + 1):
                    for c in range(merged_range.min_col, merged_range.max_col + 1):
                        if r != merged_range.min_row or c != merged_range.min_col:
                            ws.cell(r, c).value = None


    # -------------------------------------------------------------------------
    # Common
    # -------------------------------------------------------------------------

    def _apply_options(self, df: pd.DataFrame, options: ParseOptions) -> pd.DataFrame:
        """Apply skip_footer and header options."""
        if options.skip_footer > 0:
            df = df.iloc[:-options.skip_footer]

        if options.header_rows > 0 and len(df) >= options.header_rows:
            df, columns = self._generate_column_names(df, options.header_rows)
            df.columns  = columns
            df = df.reset_index(drop=True)
        else:
            df.columns = [f"col_{i}" for i in range(len(df.columns))]

        return df

    def _clean_excel_data(self, df: pd.DataFrame, options: ParseOptions) -> pd.DataFrame:
        """Replace Excel errors and custom NA values with NaN."""
        na_values = EXCEL_ERRORS | set(options.na_values or [])

        for idx in range(len(df.columns)):
            series = df.iloc[:, idx]
            if series.dtype == object:
                df.iloc[:, idx] = series.map(lambda x: np.nan if x in na_values else x)

        return df


    # -------------------------------------------------------------------------
    # Public API
    # -------------------------------------------------------------------------

    def get_sheet_names(self, file_source: FileSource) -> list[str]:
        """Get list of sheet names."""
        is_fileobj = _is_fileobj(file_source)
        file_desc  = _get_file_desc(file_source)

        if is_fileobj:
            _reset_buffer(file_source)

        try:
            content    = _read_file_content(file_source) if is_fileobj else file_source
            excel_file = fastexcel.read_excel(content)
            return excel_file.sheet_names
        except Exception:
            pass

        # Fallback to openpyxl
        try:
            if is_fileobj:
                _reset_buffer(file_source)
            wb     = openpyxl.load_workbook(file_source, read_only=True)
            sheets = wb.sheetnames
            wb.close()
            return sheets
        except PermissionError as e:
            raise FileError(f"Permission denied: {file_desc}", file_path=file_desc, operation="get_sheets") from e
        except Exception as e:
            raise FormatError(f"Cannot read sheet names: {e}", file_path=file_desc) from e

    def validate(self, file_source: FileSource) -> tuple[bool, str | None]:
        """Validate that file can be parsed."""
        is_fileobj = _is_fileobj(file_source)

        if is_fileobj:
            _reset_buffer(file_source)

        try:
            content = _read_file_content(file_source) if is_fileobj else file_source
            fastexcel.read_excel(content)
            return True, None
        except Exception as e:
            return False, str(e)
