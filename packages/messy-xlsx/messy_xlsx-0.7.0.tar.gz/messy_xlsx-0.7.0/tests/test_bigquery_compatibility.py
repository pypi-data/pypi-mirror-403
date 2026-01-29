"""Tests to ensure output is BigQuery import compatible."""

import re
from pathlib import Path

import numpy as np
import pandas as pd
import pyarrow as pa
import pytest

from messy_xlsx import MessyWorkbook, SheetConfig, sanitize_column_name


# Find all sample files
SAMPLES_DIR = Path(__file__).parent / "samples"
SAMPLE_FILES = list(SAMPLES_DIR.glob("*.xlsx")) + list(SAMPLES_DIR.glob("*.csv"))


def check_bigquery_compatible(df: pd.DataFrame) -> list[str]:
    """
    Check if DataFrame is BigQuery compatible.

    Returns list of issues found.
    """
    issues = []

    for col in df.columns:
        col_str = str(col)

        # Check column name is valid identifier (letters, numbers, underscore, starting with letter/underscore)
        if not re.match(r"^[a-zA-Z_][a-zA-Z0-9_]*$", col_str):
            issues.append(f"Column '{col}' has invalid name for BigQuery (must be alphanumeric + underscore)")

        # Check for mixed types in object columns
        if df[col].dtype == object:
            non_null = df[col].dropna()
            if len(non_null) > 0:
                # Get unique types (excluding NoneType)
                types = set()
                has_nan_float = False

                for val in non_null:
                    if isinstance(val, float) and np.isnan(val):
                        has_nan_float = True
                    elif val is not None:
                        types.add(type(val).__name__)

                # Check for np.nan in object columns (should be None)
                if has_nan_float:
                    issues.append(f"Column '{col}' has np.nan in object column (should be None for BigQuery)")

                # Check for mixed types (allow str only, or numeric only)
                if len(types) > 1:
                    # Allow int/float mix (both are numeric)
                    numeric_types = {"int", "float", "int64", "float64"}
                    if not types.issubset(numeric_types):
                        issues.append(f"Column '{col}' has mixed types: {types}")

                # Check for unhashable/complex types that BigQuery can't handle
                for val in non_null:
                    if isinstance(val, (list, dict, set, tuple)):
                        issues.append(f"Column '{col}' contains {type(val).__name__} (not BigQuery compatible)")
                        break

    return issues


class TestBigQueryCompatibility:
    """Ensure DataFrame output is BigQuery compatible."""

    @pytest.mark.parametrize("sample_file", SAMPLE_FILES, ids=lambda f: f.name)
    def test_sample_file_bq_compatible(self, sample_file):
        """All sample files should produce BigQuery-compatible output."""
        try:
            with MessyWorkbook(sample_file) as mwb:
                # Test each sheet
                for sheet_name in mwb.sheet_names:
                    df = mwb.to_dataframe(sheet=sheet_name)

                    if df.empty:
                        continue

                    issues = check_bigquery_compatible(df)

                    # Filter known issues that are acceptable
                    # Column names with spaces are common in real files
                    critical_issues = [
                        i for i in issues
                        if "np.nan in object column" in i  # This is the critical one
                        or "contains list" in i
                        or "contains dict" in i
                    ]

                    assert critical_issues == [], \
                        f"BigQuery critical issues in {sample_file.name} sheet '{sheet_name}': {critical_issues}"

        except Exception as e:
            # Some files might have issues - log but don't fail
            pytest.skip(f"Could not process {sample_file.name}: {e}")

    def test_simple_xlsx_is_bq_compatible(self, tmp_path):
        """Simple XLSX should produce BQ-compatible output."""
        import openpyxl

        file_path = tmp_path / "simple.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["name", "value", "date"])
        ws.append(["Alice", 100, "2024-01-01"])
        ws.append(["Bob", 200, "2024-01-02"])
        ws.append([None, None, None])  # Empty row
        wb.save(file_path)

        with MessyWorkbook(file_path) as mwb:
            df = mwb.to_dataframe()

        issues = check_bigquery_compatible(df)
        # Only check critical issues
        critical = [i for i in issues if "np.nan" in i or "contains" in i]
        assert critical == [], f"BigQuery compatibility issues: {critical}"

    def test_missing_values_use_none_not_nan(self, tmp_path):
        """String columns should use None, not np.nan for missing values."""
        import openpyxl

        file_path = tmp_path / "missing.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["name", "value"])
        ws.append(["Alice", 100])
        ws.append(["NA", 200])  # NA should become None
        ws.append(["", 300])   # Empty should become None
        ws.append([None, 400])
        wb.save(file_path)

        with MessyWorkbook(file_path) as mwb:
            df = mwb.to_dataframe()

        # Check string column doesn't have np.nan
        issues = check_bigquery_compatible(df)
        nan_issues = [i for i in issues if "np.nan" in i]
        assert nan_issues == [], f"Should not have np.nan in object columns: {nan_issues}"

    def test_excel_errors_produce_none_not_nan(self, tmp_path):
        """Excel errors like #DIV/0! should become None, not np.nan in string columns."""
        import openpyxl

        file_path = tmp_path / "errors.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["label", "value"])
        ws.append(["#DIV/0!", 100])  # Error value as string
        ws.append(["#N/A", 200])
        ws.append(["Normal", 300])
        wb.save(file_path)

        with MessyWorkbook(file_path) as mwb:
            df = mwb.to_dataframe()

        issues = check_bigquery_compatible(df)
        nan_issues = [i for i in issues if "np.nan" in i]
        assert nan_issues == [], f"Excel errors should become None, not np.nan: {nan_issues}"

    def test_no_complex_types_in_output(self, tmp_path):
        """Output should not contain lists, dicts, or other complex types."""
        import openpyxl

        file_path = tmp_path / "simple.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["id", "name"])
        ws.append([1, "Alice"])
        ws.append([2, "Bob"])
        wb.save(file_path)

        with MessyWorkbook(file_path) as mwb:
            df = mwb.to_dataframe()

        for col in df.columns:
            for val in df[col].dropna():
                assert not isinstance(val, (list, dict, set, tuple)), \
                    f"Column {col} contains complex type: {type(val)}"


class TestColumnNameNormalization:
    """Test that column names can be made BigQuery compatible."""

    def test_detect_invalid_column_names_when_sanitization_off(self, tmp_path):
        """Should detect column names that are invalid for BigQuery when sanitization is disabled."""
        import openpyxl

        file_path = tmp_path / "names.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["valid_name", "Also Valid", "123invalid", "has-dash", "has space"])
        ws.append([1, 2, 3, 4, 5])
        wb.save(file_path)

        # With sanitization OFF
        config = SheetConfig(sanitize_column_names=False)
        with MessyWorkbook(file_path, sheet_config=config) as mwb:
            df = mwb.to_dataframe()

        issues = check_bigquery_compatible(df)
        name_issues = [i for i in issues if "invalid name" in i]

        # Should detect at least some invalid names
        assert len(name_issues) >= 3, f"Should detect invalid column names, found: {name_issues}"


class TestSanitizeColumnNameFunction:
    """Test the sanitize_column_name utility function."""

    def test_basic_transformation(self):
        """Test basic column name sanitization."""
        assert sanitize_column_name("First Name") == "first_name"
        assert sanitize_column_name("Amount ($)") == "amount"
        assert sanitize_column_name("Sales (Q1-2024)") == "sales_q1_2024"

    def test_camel_case_to_snake_case(self):
        """camelCase and PascalCase should be converted to snake_case."""
        assert sanitize_column_name("firstName") == "first_name"
        assert sanitize_column_name("lastName") == "last_name"
        assert sanitize_column_name("FirstName") == "first_name"
        assert sanitize_column_name("createdAt") == "created_at"
        assert sanitize_column_name("updatedAt") == "updated_at"
        assert sanitize_column_name("isActive") == "is_active"
        assert sanitize_column_name("hasPermission") == "has_permission"

    def test_camel_case_with_acronyms(self):
        """Acronyms in camelCase should be handled correctly."""
        assert sanitize_column_name("XMLParser") == "xml_parser"
        assert sanitize_column_name("getHTTPResponse") == "get_http_response"
        assert sanitize_column_name("parseJSON") == "parse_json"
        assert sanitize_column_name("userID") == "user_id"
        assert sanitize_column_name("apiURL") == "api_url"

    def test_already_snake_case(self):
        """Already snake_case names should remain unchanged."""
        assert sanitize_column_name("first_name") == "first_name"
        assert sanitize_column_name("created_at") == "created_at"
        assert sanitize_column_name("is_active") == "is_active"

    def test_reserved_words_get_prefix(self):
        """BigQuery reserved words should get 'col_' prefix."""
        assert sanitize_column_name("select") == "col_select"
        assert sanitize_column_name("from") == "col_from"
        assert sanitize_column_name("where") == "col_where"
        assert sanitize_column_name("order") == "col_order"
        assert sanitize_column_name("group") == "col_group"
        assert sanitize_column_name("count") == "col_count"
        assert sanitize_column_name("sum") == "col_sum"
        assert sanitize_column_name("date") == "col_date"
        assert sanitize_column_name("time") == "col_time"
        assert sanitize_column_name("timestamp") == "col_timestamp"
        assert sanitize_column_name("table") == "col_table"
        assert sanitize_column_name("index") == "col_index"

    def test_reserved_words_case_insensitive(self):
        """Reserved word detection should be case-insensitive."""
        assert sanitize_column_name("SELECT") == "col_select"
        assert sanitize_column_name("Select") == "col_select"
        assert sanitize_column_name("FROM") == "col_from"
        assert sanitize_column_name("Date") == "col_date"

    def test_reserved_word_as_part_of_name(self):
        """Reserved words as part of a larger name should NOT get prefix."""
        assert sanitize_column_name("select_all") == "select_all"
        assert sanitize_column_name("from_date") == "from_date"
        assert sanitize_column_name("order_id") == "order_id"
        assert sanitize_column_name("date_created") == "date_created"
        assert sanitize_column_name("user_count") == "user_count"

    def test_digit_prefix(self):
        """Names starting with digits get 'col_' prefix."""
        assert sanitize_column_name("123-ID") == "col_123_id"
        assert sanitize_column_name("1st Place") == "col_1st_place"
        assert sanitize_column_name("2024 Revenue") == "col_2024_revenue"

    def test_special_characters(self):
        """Special characters are replaced with underscores."""
        assert sanitize_column_name("email@domain") == "email_domain"
        assert sanitize_column_name("price#1") == "price_1"
        assert sanitize_column_name("user.name") == "user_name"
        assert sanitize_column_name("field/value") == "field_value"

    def test_unicode_characters(self):
        """Unicode characters should be removed (not BigQuery compatible)."""
        assert sanitize_column_name("café") == "caf"
        assert sanitize_column_name("naïve") == "na_ve"
        assert sanitize_column_name("日本語") == "unnamed"

    def test_consecutive_underscores(self):
        """Consecutive underscores should be collapsed."""
        assert sanitize_column_name("a__b") == "a_b"
        assert sanitize_column_name("a___b___c") == "a_b_c"
        assert sanitize_column_name("first  name") == "first_name"

    def test_leading_trailing_underscores(self):
        """Leading/trailing underscores should be stripped."""
        assert sanitize_column_name("_name_") == "name"
        assert sanitize_column_name("__id__") == "id"
        assert sanitize_column_name("  name  ") == "name"

    def test_empty_and_none(self):
        """Empty strings and None should return 'unnamed'."""
        assert sanitize_column_name(None) == "unnamed"
        assert sanitize_column_name("") == "unnamed"
        assert sanitize_column_name("   ") == "unnamed"
        assert sanitize_column_name("nan") == "unnamed"
        assert sanitize_column_name("NaN") == "unnamed"

    def test_lowercase_conversion(self):
        """All names should be converted to lowercase."""
        assert sanitize_column_name("UPPERCASE") == "uppercase"
        assert sanitize_column_name("MixedCase") == "mixed_case"  # camelCase -> snake_case
        assert sanitize_column_name("camelCase") == "camel_case"  # camelCase -> snake_case

    def test_max_length_truncation(self):
        """Names exceeding max_length should be truncated."""
        long_name = "a" * 350
        result = sanitize_column_name(long_name)
        assert len(result) == 300
        assert result == "a" * 300

        # With custom max_length
        result = sanitize_column_name(long_name, max_length=10)
        assert len(result) == 10

    def test_truncation_strips_trailing_underscore(self):
        """Truncation should not leave trailing underscores."""
        name = "a" * 299 + "_b"  # 301 chars, truncates to 300
        result = sanitize_column_name(name, max_length=300)
        assert not result.endswith("_")

    def test_bigquery_pattern_compliance(self):
        """All results should match BigQuery's column name pattern."""
        test_cases = [
            "First Name", "Amount ($)", "123-ID", "Sales (Q1-2024)",
            "email@domain.com", "user.name", "UPPERCASE", "  spaces  ",
            "_underscore_", "normal_name", "αβγ", "日本語",
        ]
        pattern = re.compile(r"^[a-z_][a-z0-9_]*$")

        for name in test_cases:
            result = sanitize_column_name(name)
            if result != "unnamed":  # unnamed is also valid
                assert pattern.match(result), f"'{name}' -> '{result}' doesn't match BigQuery pattern"


class TestSanitizeColumnNamesConfig:
    """Test the sanitize_column_names config option."""

    def test_sanitization_enabled_by_default(self, tmp_path):
        """Sanitization should be enabled by default."""
        import openpyxl

        file_path = tmp_path / "test.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["First Name", "Amount ($)", "123-ID"])
        ws.append(["Alice", 100, "A001"])
        wb.save(file_path)

        with MessyWorkbook(file_path) as mwb:
            df = mwb.to_dataframe()

        assert list(df.columns) == ["first_name", "amount", "col_123_id"]

    def test_sanitization_can_be_disabled(self, tmp_path):
        """Sanitization can be disabled via config."""
        import openpyxl

        file_path = tmp_path / "test.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["First Name", "Amount ($)", "123-ID"])
        ws.append(["Alice", 100, "A001"])
        wb.save(file_path)

        config = SheetConfig(sanitize_column_names=False)
        with MessyWorkbook(file_path, sheet_config=config) as mwb:
            df = mwb.to_dataframe()

        # Original names preserved
        assert list(df.columns) == ["First Name", "Amount ($)", "123-ID"]

    def test_user_renames_take_precedence(self, tmp_path):
        """User-specified column_renames should override sanitization."""
        import openpyxl

        file_path = tmp_path / "test.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["First Name", "Amount ($)"])
        ws.append(["Alice", 100])
        wb.save(file_path)

        config = SheetConfig(
            sanitize_column_names=True,
            column_renames={"first_name": "user_name"}  # Rename after sanitization
        )
        with MessyWorkbook(file_path, sheet_config=config) as mwb:
            df = mwb.to_dataframe()

        assert "user_name" in df.columns
        assert "amount" in df.columns

    def test_duplicate_column_handling(self, tmp_path):
        """Duplicate column names after sanitization should get suffixes."""
        import openpyxl

        file_path = tmp_path / "test.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        # These will all become "name" after sanitization
        ws.append(["Name", "NAME", "name"])
        ws.append(["Alice", "Bob", "Charlie"])
        wb.save(file_path)

        with MessyWorkbook(file_path) as mwb:
            df = mwb.to_dataframe()

        # Should have unique names
        assert len(set(df.columns)) == 3
        assert "name" in df.columns
        assert "name_1" in df.columns
        assert "name_2" in df.columns

    def test_all_columns_match_bigquery_pattern(self, tmp_path):
        """All sanitized columns should match BigQuery's pattern."""
        import openpyxl

        file_path = tmp_path / "test.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["First Name", "Amount ($)", "123-ID", "日本語", "café", "email@domain"])
        ws.append(["Alice", 100, "A001", "test", "test", "test"])
        wb.save(file_path)

        with MessyWorkbook(file_path) as mwb:
            df = mwb.to_dataframe()

        pattern = re.compile(r"^[a-z_][a-z0-9_]*$")
        for col in df.columns:
            assert pattern.match(col), f"Column '{col}' doesn't match BigQuery pattern"

    def test_sanitization_works_with_normalization_disabled(self, tmp_path):
        """Sanitization should work even when normalization is disabled."""
        import openpyxl

        file_path = tmp_path / "test.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["First Name", "Amount"])
        ws.append(["Alice", "100"])
        wb.save(file_path)

        config = SheetConfig(normalize=False, sanitize_column_names=True)
        with MessyWorkbook(file_path, sheet_config=config) as mwb:
            df = mwb.to_dataframe()

        # Columns should still be sanitized
        assert list(df.columns) == ["first_name", "amount"]

    def test_camelcase_columns_in_real_file(self, tmp_path):
        """camelCase columns from real files should be converted to snake_case."""
        import openpyxl

        file_path = tmp_path / "camel.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["firstName", "lastName", "emailAddress", "createdAt"])
        ws.append(["Alice", "Smith", "alice@example.com", "2024-01-01"])
        ws.append(["Bob", "Jones", "bob@example.com", "2024-01-02"])
        wb.save(file_path)

        with MessyWorkbook(file_path) as mwb:
            df = mwb.to_dataframe()

        expected_columns = ["first_name", "last_name", "email_address", "created_at"]
        assert list(df.columns) == expected_columns

    def test_reserved_word_columns_in_real_file(self, tmp_path):
        """Reserved word columns should get 'col_' prefix for easy querying."""
        import openpyxl

        file_path = tmp_path / "reserved.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["select", "from", "date", "count", "order"])
        ws.append(["A", "B", "2024-01-01", 10, 1])
        ws.append(["C", "D", "2024-01-02", 20, 2])
        wb.save(file_path)

        with MessyWorkbook(file_path) as mwb:
            df = mwb.to_dataframe()

        expected_columns = ["col_select", "col_from", "col_date", "col_count", "col_order"]
        assert list(df.columns) == expected_columns

    def test_mixed_camelcase_and_reserved_words(self, tmp_path):
        """Test combination of camelCase conversion and reserved word handling."""
        import openpyxl

        file_path = tmp_path / "mixed.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        # Mix of camelCase, reserved words, spaces, and normal names
        ws.append(["userId", "First Name", "select", "orderDate", "count"])
        ws.append([1, "Alice", "A", "2024-01-01", 100])
        wb.save(file_path)

        with MessyWorkbook(file_path) as mwb:
            df = mwb.to_dataframe()

        expected = ["user_id", "first_name", "col_select", "order_date", "col_count"]
        assert list(df.columns) == expected


def convert_to_arrow(df: pd.DataFrame) -> tuple[pa.Table | None, str | None]:
    """
    Try to convert DataFrame to PyArrow Table.

    This simulates what BigQuery does internally when loading data.
    Returns (table, None) on success, (None, error_message) on failure.
    """
    try:
        # BigQuery uses Arrow format internally
        # This will fail if there are incompatible types
        table = pa.Table.from_pandas(df)
        return table, None
    except Exception as e:
        return None, str(e)


def validate_arrow_schema_for_bq(table: pa.Table) -> list[str]:
    """
    Validate that Arrow schema is BigQuery compatible.

    BigQuery supports: INT64, FLOAT64, BOOL, STRING, BYTES, DATE, DATETIME, TIME, TIMESTAMP
    """
    issues = []

    # Map Arrow types to BigQuery compatibility
    bq_compatible_types = {
        pa.int8(), pa.int16(), pa.int32(), pa.int64(),
        pa.uint8(), pa.uint16(), pa.uint32(), pa.uint64(),
        pa.float16(), pa.float32(), pa.float64(),
        pa.bool_(),
        pa.string(), pa.large_string(), pa.utf8(), pa.large_utf8(),
        pa.binary(), pa.large_binary(),
        pa.date32(), pa.date64(),
    }

    for field in table.schema:
        field_type = field.type

        # Check for timestamp types (compatible)
        if pa.types.is_timestamp(field_type):
            continue

        # Check for duration (not directly supported)
        if pa.types.is_duration(field_type):
            issues.append(f"Column '{field.name}' has duration type (not BQ compatible)")
            continue

        # Check for nested types
        if pa.types.is_list(field_type):
            issues.append(f"Column '{field.name}' has list type (requires REPEATED mode in BQ)")
            continue

        if pa.types.is_struct(field_type):
            issues.append(f"Column '{field.name}' has struct type (requires RECORD mode in BQ)")
            continue

        if pa.types.is_map(field_type):
            issues.append(f"Column '{field.name}' has map type (not directly supported in BQ)")
            continue

    return issues


class TestPyArrowConversion:
    """Test that output can be converted to PyArrow (BigQuery's internal format)."""

    @pytest.mark.parametrize("sample_file", SAMPLE_FILES, ids=lambda f: f.name)
    def test_sample_converts_to_arrow(self, sample_file):
        """All sample files should convert to PyArrow successfully."""
        try:
            with MessyWorkbook(sample_file) as mwb:
                for sheet_name in mwb.sheet_names:
                    df = mwb.to_dataframe(sheet=sheet_name)

                    if df.empty:
                        continue

                    # Try to convert to Arrow
                    table, error = convert_to_arrow(df)

                    assert error is None, \
                        f"Failed to convert {sample_file.name} sheet '{sheet_name}' to Arrow: {error}"

                    # Validate schema
                    schema_issues = validate_arrow_schema_for_bq(table)

                    # Filter out acceptable issues (nested types might be intentional)
                    critical_issues = [i for i in schema_issues if "not BQ compatible" in i]

                    assert critical_issues == [], \
                        f"Arrow schema issues in {sample_file.name}: {critical_issues}"

        except Exception as e:
            pytest.skip(f"Could not process {sample_file.name}: {e}")

    def test_arrow_roundtrip(self, tmp_path):
        """Data should survive Arrow roundtrip without corruption."""
        import openpyxl

        file_path = tmp_path / "roundtrip.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["name", "value", "active"])
        ws.append(["Alice", 100, True])
        ws.append(["Bob", 200.5, False])
        ws.append([None, None, None])
        wb.save(file_path)

        with MessyWorkbook(file_path) as mwb:
            df_original = mwb.to_dataframe()

        # Convert to Arrow and back
        table, error = convert_to_arrow(df_original)
        assert error is None, f"Arrow conversion failed: {error}"

        df_roundtrip = table.to_pandas()

        # Verify data integrity
        assert len(df_roundtrip) == len(df_original)
        assert list(df_roundtrip.columns) == list(df_original.columns)

    def test_parquet_export(self, tmp_path):
        """Data should be exportable to Parquet (BigQuery's preferred format)."""
        import openpyxl

        file_path = tmp_path / "export.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["id", "name", "amount"])
        ws.append([1, "Alice", 100.50])
        ws.append([2, "Bob", 200.75])
        wb.save(file_path)

        with MessyWorkbook(file_path) as mwb:
            df = mwb.to_dataframe()

        # Export to Parquet (BigQuery's preferred format)
        parquet_path = tmp_path / "output.parquet"
        table, error = convert_to_arrow(df)
        assert error is None

        # Write to Parquet
        import pyarrow.parquet as pq
        pq.write_table(table, parquet_path)

        # Read back and verify
        table_read = pq.read_table(parquet_path)
        df_read = table_read.to_pandas()

        assert len(df_read) == len(df)
        assert list(df_read.columns) == list(df.columns)
