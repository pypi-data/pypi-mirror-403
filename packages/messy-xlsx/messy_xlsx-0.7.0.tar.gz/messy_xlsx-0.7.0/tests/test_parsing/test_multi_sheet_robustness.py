"""Robustness tests for MultiSheetParser based on real-world messy files."""

import tempfile
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from messy_xlsx.multi_sheet import (
    MultiSheetParser,
    read_all_sheets,
    analyze_excel,
)


# ============================================================================
# Fixtures - Real-world patterns from messy files
# ============================================================================

@pytest.fixture
def temp_dir():
    """Create a temporary directory for test files."""
    with tempfile.TemporaryDirectory() as tmpdir:
        yield Path(tmpdir)


@pytest.fixture
def plant_column_xlsx(temp_dir):
    """
    Simulate Plant column issue: "1&3" (string) mixed with 3 (int).
    This was causing PyArrow errors in BigQuery uploads.
    """
    file_path = temp_dir / "plant_column.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    ws.append(["ID", "Name", "Plant"])
    ws.append(["001", "Item A", "1&3"])  # String
    ws.append(["002", "Item B", 3])       # Int - this caused mixed types
    ws.append(["003", "Item C", "1&3"])
    ws.append(["004", "Item D", 3])
    ws.append(["005", "Item E", "1"])     # String that looks numeric

    wb.save(file_path)
    wb.close()

    return file_path


@pytest.fixture
def status_column_xlsx(temp_dir):
    """
    Simulate status columns with mixed text/numbers.
    Track Process Date had: "ON HOLD", "X", "BUY OFF", "CMM", 0
    """
    file_path = temp_dir / "status_column.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    ws.append(["ID", "Status", "Track_Date"])
    ws.append(["001", "OPEN", "ON HOLD"])
    ws.append(["002", "CLOSED", "X"])
    ws.append(["003", "PENDING", 0])        # Numeric mixed with strings
    ws.append(["004", "OPEN", "BUY OFF"])
    ws.append(["005", "CLOSED", "CMM"])

    wb.save(file_path)
    wb.close()

    return file_path


@pytest.fixture
def metadata_with_number_xlsx(temp_dir):
    """
    Simulate metadata row that has both text and a number.
    Row 0 had: "Printed Date : 22-10-2025" in col 0, 45952 in col 14
    """
    file_path = temp_dir / "metadata_number.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    # Metadata row with text and number
    row1 = ["Printed Date : 22-10-2025"] + [None] * 13 + [45952]
    ws.append(row1)

    # Actual headers
    ws.append(["JO_No", "Type", "Date", "Customer", "Status"])

    # Data
    ws.append(["J001", "MP", "2025-01-10", "ACME", "OPEN"])
    ws.append(["J002", "FA", "2025-01-11", "CORP", "CLOSED"])

    wb.save(file_path)
    wb.close()

    return file_path


@pytest.fixture
def trailing_spaces_xlsx(temp_dir):
    """
    Simulate column names with trailing/extra spaces.
    Real file had: "PASS DUE ", "PROCESS ", "Balance  Qty."
    """
    file_path = temp_dir / "trailing_spaces.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    ws.append(["Name ", "Value  ", "  Status", "Balance  Qty."])
    ws.append(["A", 100, "OK", 50])
    ws.append(["B", 200, "FAIL", 25])

    wb.save(file_path)
    wb.close()

    return file_path


@pytest.fixture
def deep_header_xlsx(temp_dir):
    """
    Simulate headers at various depths (not row 0).
    Data 21.10pm had headers at row 3, DATA at row 1.
    """
    file_path = temp_dir / "deep_header.xlsx"

    wb = openpyxl.Workbook()

    # Sheet with header at row 3
    ws1 = wb.active
    ws1.title = "DeepHeader"
    ws1.append(["Customer P/O Status"])
    ws1.append(["Page -1 of 1"])
    ws1.append([])  # Empty row
    ws1.append(["PO_No", "Item", "Qty", "Price"])  # Actual headers
    ws1.append(["PO001", "Widget", 10, 99.99])
    ws1.append(["PO002", "Gadget", 5, 149.99])

    # Sheet with header at row 1
    ws2 = wb.create_sheet("ShallowHeader")
    ws2.append(["Report Date: 2025-01-10"])
    ws2.append(["ID", "Name", "Value"])  # Headers at row 1
    ws2.append(["001", "Alpha", 100])
    ws2.append(["002", "Beta", 200])

    wb.save(file_path)
    wb.close()

    return file_path


@pytest.fixture
def sparse_columns_xlsx(temp_dir):
    """
    Simulate sheets with entirely null columns in between data.
    Some columns like "PASS DUE RESULT" were entirely empty.
    """
    file_path = temp_dir / "sparse_columns.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    ws.append(["ID", "Name", "Empty1", "Value", "Empty2", "Status"])
    ws.append(["001", "A", None, 100, None, "OK"])
    ws.append(["002", "B", None, 200, None, "OK"])
    ws.append(["003", "C", None, 300, None, "FAIL"])

    wb.save(file_path)
    wb.close()

    return file_path


@pytest.fixture
def date_in_data_xlsx(temp_dir):
    """
    Simulate data rows that start with dates (shouldn't be mistaken for headers).
    """
    file_path = temp_dir / "date_data.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    ws.append(["Date", "Event", "Value"])
    ws.append(["2025-01-01", "Sale", 1000])
    ws.append(["2025-01-02", "Return", -50])
    ws.append(["2025-01-03", "Sale", 2000])

    wb.save(file_path)
    wb.close()

    return file_path


@pytest.fixture
def concatenated_key_xlsx(temp_dir):
    """
    Simulate concatenated key fields like "TP04820011024-404 REV C.05".
    Column 0 in Data 21.10pm had these compound keys.
    """
    file_path = temp_dir / "concat_key.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    ws.append(["CompositeKey", "PO_No", "Item", "Part_No"])
    ws.append(["TP04820011024-404 REV C.05", "TP0482", "001", "1024-404 REV C.0"])
    ws.append(["1800006680021002554 REV G2", "180000668", "002", "1002554 REV G"])
    ws.append(["886-11285432-OP002C208S88002 REV 0110", "886-11285432-OP", "002", "C208S88002 REV 01"])

    wb.save(file_path)
    wb.close()

    return file_path


@pytest.fixture
def numeric_string_column_xlsx(temp_dir):
    """
    Simulate columns with numeric-looking strings that should stay strings.
    Mat / Std Part Status had "0" and text like "MAT PR24070163 - 12.07 IN"
    """
    file_path = temp_dir / "numeric_string.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    ws.append(["ID", "Mat_Status", "Qty"])
    ws.append(["001", "0", 100])
    ws.append(["002", "MAT PR24070163 - 12.07 IN", 50])
    ws.append(["003", "0", 75])
    ws.append(["004", "MAT PR25030467 20.03 - IN 26.03", 25])

    wb.save(file_path)
    wb.close()

    return file_path


# ============================================================================
# Tests: Mixed Type Handling
# ============================================================================

class TestMixedTypeHandling:
    """Test handling of mixed types that cause PyArrow errors."""

    def test_plant_column_string_int_mix(self, plant_column_xlsx):
        """Test that '1&3' and 3 become consistent types."""
        sheets = read_all_sheets(plant_column_xlsx)
        df = sheets["Data"]

        # All Plant values should be same type (column name is lowercase)
        types = set(type(v).__name__ for v in df["plant"].dropna())
        assert len(types) == 1, f"Mixed types found: {types}"

    def test_status_column_text_zero_mix(self, status_column_xlsx):
        """Test that 'ON HOLD', 'X', and 0 become consistent types."""
        sheets = read_all_sheets(status_column_xlsx)
        df = sheets["Data"]

        # Track_Date should have consistent types (column name is lowercase)
        types = set(type(v).__name__ for v in df["track_date"].dropna())
        assert len(types) == 1, f"Mixed types found: {types}"

    def test_numeric_string_preserved(self, numeric_string_column_xlsx):
        """Test that '0' stays string when mixed with text."""
        sheets = read_all_sheets(numeric_string_column_xlsx)
        df = sheets["Data"]

        # Mat_Status should all be strings (not partial float conversion, column name is lowercase)
        types = set(type(v).__name__ for v in df["mat_status"].dropna())
        assert len(types) == 1
        assert "str" in types, "Should preserve as strings"


# ============================================================================
# Tests: Header Detection Edge Cases
# ============================================================================

class TestHeaderDetectionEdgeCases:
    """Test header detection with tricky patterns."""

    def test_metadata_with_number_skipped(self, metadata_with_number_xlsx):
        """Test metadata row with number in different column is skipped."""
        infos = analyze_excel(metadata_with_number_xlsx)
        info = infos[0]

        # Should detect headers at row 1, not row 0
        assert info.header_row == 1

    def test_deep_header_row_3(self, deep_header_xlsx):
        """Test detection of headers at row 3."""
        infos = analyze_excel(deep_header_xlsx)
        deep = next(i for i in infos if i.name == "DeepHeader")

        assert deep.header_row == 3

    def test_shallow_header_row_1(self, deep_header_xlsx):
        """Test detection of headers at row 1."""
        infos = analyze_excel(deep_header_xlsx)
        shallow = next(i for i in infos if i.name == "ShallowHeader")

        assert shallow.header_row == 1

    def test_date_data_not_mistaken_for_headers(self, date_in_data_xlsx):
        """Test that date rows aren't mistaken for metadata/headers."""
        sheets = read_all_sheets(date_in_data_xlsx)
        df = sheets["Data"]

        # Should have correct headers (column names are lowercase)
        assert "date" in df.columns
        assert "event" in df.columns
        # Should have all data rows
        assert len(df) == 3


# ============================================================================
# Tests: Column Name Cleaning
# ============================================================================

class TestColumnNameCleaning:
    """Test column name sanitization."""

    def test_trailing_spaces_stripped(self, trailing_spaces_xlsx):
        """Test that trailing spaces in column names are handled."""
        sheets = read_all_sheets(trailing_spaces_xlsx)
        df = sheets["Data"]

        # No spaces in column names
        for col in df.columns:
            assert " " not in col, f"Column '{col}' has spaces"

    def test_double_spaces_handled(self, trailing_spaces_xlsx):
        """Test 'Balance  Qty.' with double space becomes valid name."""
        sheets = read_all_sheets(trailing_spaces_xlsx)
        df = sheets["Data"]

        # Should have a balance_qty column (spaces replaced, lowercase)
        balance_cols = [c for c in df.columns if "balance" in c]
        assert len(balance_cols) == 1

    def test_concatenated_keys_preserved(self, concatenated_key_xlsx):
        """Test that complex keys in data are preserved."""
        sheets = read_all_sheets(concatenated_key_xlsx)
        df = sheets["Data"]

        # CompositeKey column should exist and have data (column name is lowercase)
        assert "compositekey" in df.columns
        assert "TP04820011024-404 REV C.05" in df["compositekey"].values


# ============================================================================
# Tests: Sparse Data Handling
# ============================================================================

class TestSparseDataHandling:
    """Test handling of sparse/null columns."""

    def test_empty_columns_dropped(self, sparse_columns_xlsx):
        """Test that entirely empty columns are dropped."""
        sheets = read_all_sheets(sparse_columns_xlsx)
        df = sheets["Data"]

        # Empty1 and Empty2 should be dropped
        assert "Empty1" not in df.columns
        assert "Empty2" not in df.columns

    def test_data_columns_preserved(self, sparse_columns_xlsx):
        """Test that columns with data are preserved."""
        sheets = read_all_sheets(sparse_columns_xlsx)
        df = sheets["Data"]

        # Column names are lowercase
        assert "id" in df.columns
        assert "name" in df.columns
        assert "value" in df.columns
        assert "status" in df.columns


# ============================================================================
# Tests: Data Integrity
# ============================================================================

class TestDataIntegrity:
    """Test that data values are preserved correctly."""

    def test_numeric_values_preserved(self, plant_column_xlsx):
        """Test numeric values aren't corrupted."""
        sheets = read_all_sheets(plant_column_xlsx)
        df = sheets["Data"]

        # ID should still be readable (column name is lowercase)
        assert "001" in df["id"].values or 1 in df["id"].values

    def test_special_characters_in_data_preserved(self, concatenated_key_xlsx):
        """Test that special chars in data (not headers) are preserved."""
        sheets = read_all_sheets(concatenated_key_xlsx)
        df = sheets["Data"]

        # Part numbers with dashes and spaces should be intact (column name is lowercase)
        assert any("-" in str(v) for v in df["part_no"].values)

    def test_row_count_correct_after_header_skip(self, deep_header_xlsx):
        """Test correct row count after skipping metadata."""
        sheets = read_all_sheets(deep_header_xlsx)

        deep = sheets["DeepHeader"]
        assert len(deep) == 2  # 2 data rows after header at row 3

        shallow = sheets["ShallowHeader"]
        assert len(shallow) == 2  # 2 data rows after header at row 1


# ============================================================================
# Tests: BigQuery Compatibility
# ============================================================================

class TestBigQueryCompatibility:
    """Test that output is BigQuery-compatible."""

    def test_no_mixed_types_any_column(self, plant_column_xlsx):
        """Verify no column has mixed types."""
        sheets = read_all_sheets(plant_column_xlsx)

        for name, df in sheets.items():
            for col in df.columns:
                if df[col].dtype == object:
                    types = set(type(v).__name__ for v in df[col].dropna())
                    assert len(types) <= 1, f"{name}.{col} has mixed types: {types}"

    def test_column_names_alphanumeric(self, trailing_spaces_xlsx):
        """Verify column names are alphanumeric + underscore only."""
        sheets = read_all_sheets(trailing_spaces_xlsx)

        for name, df in sheets.items():
            for col in df.columns:
                clean = col.replace("_", "")
                assert clean.isalnum(), f"Column '{col}' invalid for BigQuery"

    def test_no_leading_numbers_in_columns(self, temp_dir):
        """Test columns starting with numbers get prefixed."""
        file_path = temp_dir / "numeric_headers.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.append(["1stCol", "2ndCol", "Name"])
        ws.append(["a", "b", "c"])
        wb.save(file_path)
        wb.close()

        sheets = read_all_sheets(file_path)
        df = sheets["Data"]

        for col in df.columns:
            assert not col[0].isdigit(), f"Column '{col}' starts with digit"


# ============================================================================
# Tests: Multiple Sheets Combined
# ============================================================================

class TestMultipleSheetsCombined:
    """Test parsing files with multiple problematic sheets."""

    def test_all_sheets_consistent_types(self, deep_header_xlsx):
        """Test all sheets in file have consistent types."""
        sheets = read_all_sheets(deep_header_xlsx)

        for name, df in sheets.items():
            for col in df.columns:
                if df[col].dtype == object:
                    types = set(type(v).__name__ for v in df[col].dropna())
                    assert len(types) <= 1, f"{name}.{col} mixed: {types}"

    def test_all_sheets_clean_columns(self, deep_header_xlsx):
        """Test all sheets have clean column names."""
        sheets = read_all_sheets(deep_header_xlsx)

        for name, df in sheets.items():
            for col in df.columns:
                assert " " not in col
                assert "/" not in col
                assert "." not in col
