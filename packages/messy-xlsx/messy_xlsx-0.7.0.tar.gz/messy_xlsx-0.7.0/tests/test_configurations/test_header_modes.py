"""Tests for header detection mode configurations."""

from pathlib import Path

import pytest

from messy_xlsx import MessyWorkbook, SheetConfig


# Get a few sample files for matrix testing
SAMPLES_DIR = Path(__file__).parent.parent / "samples"
TEST_FILES = list(SAMPLES_DIR.glob("*.xlsx"))[:10]  # First 10 files


class TestHeaderDetectionModes:
    """Test header detection mode variations."""

    @pytest.mark.parametrize("mode", ["auto", "manual", "smart"])
    def test_header_mode_on_simple_file(self, sample_xlsx, mode):
        """Test all header modes on simple file."""
        config = SheetConfig(
            auto_detect=True,
            header_detection_mode=mode
        )

        with MessyWorkbook(sample_xlsx, sheet_config=config) as wb:
            df = wb.to_dataframe()
            assert df is not None
            assert len(df.columns) > 0

    @pytest.mark.parametrize("sample_file", TEST_FILES, ids=lambda f: f.name)
    @pytest.mark.parametrize("mode", ["auto", "smart"])
    def test_auto_modes_on_samples(self, sample_file, mode):
        """Test auto and smart modes on sample files."""
        config = SheetConfig(
            auto_detect=True,
            header_detection_mode=mode
        )

        with MessyWorkbook(sample_file, sheet_config=config) as wb:
            df = wb.to_dataframe()
            assert df is not None

    def test_manual_mode_ignores_detection(self, messy_xlsx):
        """Test manual mode ignores auto-detection."""
        config = SheetConfig(
            skip_rows=1,
            header_rows=1,
            header_detection_mode="manual",
            auto_detect=False
        )

        with MessyWorkbook(messy_xlsx, sheet_config=config) as wb:
            df = wb.to_dataframe()
            assert df is not None


class TestConfidenceThresholds:
    """Test header confidence threshold variations."""

    @pytest.mark.parametrize("threshold", [0.5, 0.7, 0.9, 1.0])
    def test_confidence_thresholds(self, sample_xlsx, threshold):
        """Test different confidence thresholds."""
        config = SheetConfig(
            auto_detect=True,
            header_detection_mode="auto",
            header_confidence_threshold=threshold
        )

        with MessyWorkbook(sample_xlsx, sheet_config=config) as wb:
            df = wb.to_dataframe()
            assert df is not None

    def test_low_confidence_fallback(self, temp_dir):
        """Test fallback when confidence is low."""
        file_path = temp_dir / "low_conf.xlsx"

        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        # All numeric rows (hard to detect header)
        ws.append([1, 2, 3])
        ws.append([4, 5, 6])
        ws.append([7, 8, 9])
        wb.save(file_path)
        wb.close()

        config = SheetConfig(
            auto_detect=True,
            header_detection_mode="auto",
            header_confidence_threshold=0.9,
            header_fallback="first_row"
        )

        with MessyWorkbook(file_path, sheet_config=config) as mwb:
            df = mwb.to_dataframe()
            assert df is not None


class TestHeaderPatterns:
    """Test header pattern matching."""

    def test_pattern_matching_boost(self, temp_dir):
        """Test patterns boost header confidence."""
        file_path = temp_dir / "patterns.xlsx"

        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Customer Name", "Invoice Date", "Amount"])
        ws.append(["Alice", "2024-01-01", 1000])
        wb.save(file_path)
        wb.close()

        config = SheetConfig(
            auto_detect=True,
            header_patterns=[r".*name.*", r".*date.*", r".*amount.*"]
        )

        with MessyWorkbook(file_path, sheet_config=config) as mwb:
            structure = mwb.get_structure()
            # Patterns should boost confidence
            assert structure.header_confidence >= 0.7

    def test_no_pattern_match(self, temp_dir):
        """Test when patterns don't match."""
        file_path = temp_dir / "no_match.xlsx"

        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["A", "B", "C"])
        ws.append([1, 2, 3])
        wb.save(file_path)
        wb.close()

        config = SheetConfig(
            auto_detect=True,
            header_patterns=[r".*customer.*", r".*invoice.*"]
        )

        with MessyWorkbook(file_path, sheet_config=config) as mwb:
            df = mwb.to_dataframe()
            assert df is not None


class TestHeaderFallbacks:
    """Test header fallback strategies."""

    def test_first_row_fallback(self, temp_dir):
        """Test first_row fallback strategy."""
        file_path = temp_dir / "ambiguous.xlsx"

        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([1, 2, 3])
        ws.append([4, 5, 6])
        wb.save(file_path)
        wb.close()

        config = SheetConfig(
            auto_detect=True,
            header_detection_mode="auto",
            header_fallback="first_row"
        )

        with MessyWorkbook(file_path, sheet_config=config) as mwb:
            df = mwb.to_dataframe()
            assert df is not None

    def test_none_fallback(self, temp_dir):
        """Test none fallback (no headers)."""
        file_path = temp_dir / "no_header.xlsx"

        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([1, 2, 3])
        ws.append([4, 5, 6])
        wb.save(file_path)
        wb.close()

        config = SheetConfig(
            auto_detect=True,
            header_detection_mode="auto",
            header_confidence_threshold=0.99,
            header_fallback="none"
        )

        with MessyWorkbook(file_path, sheet_config=config) as mwb:
            df = mwb.to_dataframe()
            assert df is not None
