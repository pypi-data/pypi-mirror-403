"""
Pytest configuration and fixtures for messy-xlsx tests.
"""

import tempfile
from pathlib import Path

import openpyxl
import pytest


@pytest.fixture
def temp_dir():
    """Create a temporary directory for test files."""
    with tempfile.TemporaryDirectory() as tmpdir:
        yield Path(tmpdir)


@pytest.fixture
def sample_xlsx(temp_dir):
    """Create a simple XLSX file for testing."""
    file_path = temp_dir / "sample.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    # Add headers
    ws.append(["Name", "Age", "City"])

    # Add data
    ws.append(["Alice", 30, "New York"])
    ws.append(["Bob", 25, "Los Angeles"])
    ws.append(["Charlie", 35, "Chicago"])

    wb.save(file_path)
    wb.close()

    return file_path


@pytest.fixture
def messy_xlsx(temp_dir):
    """Create a messy XLSX file with common issues."""
    file_path = temp_dir / "messy.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"

    # Title rows (metadata)
    ws.append(["Monthly Sales Report"])
    ws.append(["Generated: 2024-01-15"])
    ws.append([])  # Blank row

    # Headers
    ws.append(["Product", "Q1 Sales", "Q2 Sales", "Total"])

    # Data
    ws.append(["Widget A", 1000, 1200, "=B5+C5"])
    ws.append(["Widget B", 800, 950, "=B6+C6"])
    ws.append([])  # Blank row separator
    ws.append(["Grand Total", "=SUM(B5:B6)", "=SUM(C5:C6)", "=SUM(D5:D6)"])

    wb.save(file_path)
    wb.close()

    return file_path


@pytest.fixture
def european_xlsx(temp_dir):
    """Create XLSX with European number format."""
    file_path = temp_dir / "european.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    # Headers
    ws.append(["Item", "Price", "Quantity"])

    # Data with European formatting (stored as text to preserve format)
    ws.append(["Product 1", "1.234,56", "100"])
    ws.append(["Product 2", "2.345,67", "200"])

    wb.save(file_path)
    wb.close()

    return file_path


@pytest.fixture
def multi_table_xlsx(temp_dir):
    """Create XLSX with multiple tables on same sheet."""
    file_path = temp_dir / "multi_table.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    # First table
    ws.append(["Table 1: Sales"])
    ws.append(["Product", "Amount"])
    ws.append(["A", 100])
    ws.append(["B", 200])

    # Separator (2+ blank rows)
    ws.append([])
    ws.append([])

    # Second table
    ws.append(["Table 2: Inventory"])
    ws.append(["Product", "Stock"])
    ws.append(["A", 50])
    ws.append(["B", 75])

    wb.save(file_path)
    wb.close()

    return file_path


@pytest.fixture
def merged_cells_xlsx(temp_dir):
    """Create XLSX with merged cells."""
    file_path = temp_dir / "merged.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    # Merged header
    ws["A1"] = "Sales Report"
    ws.merge_cells("A1:C1")

    # Sub-headers
    ws.append(["Region", "Q1", "Q2"])

    # Data
    ws.append(["North", 100, 150])
    ws.append(["South", 200, 250])

    wb.save(file_path)
    wb.close()

    return file_path
