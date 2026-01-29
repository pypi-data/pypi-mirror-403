"""MessyWorkbook - Main entry point for parsing Excel files."""

# ============================================================================
# Imports
# ============================================================================

import io
from pathlib import Path
from typing import Any, BinaryIO

import openpyxl
import pandas as pd

from messy_xlsx.cache import get_structure_cache
from messy_xlsx.detection.format_detector import FormatDetector
from messy_xlsx.detection.structure_analyzer import StructureAnalyzer
from messy_xlsx.exceptions import FileError, FormatError
from messy_xlsx.formulas.config import FormulaConfig, FormulaEvaluationMode
from messy_xlsx.formulas.engine import FormulaEngine
from messy_xlsx.models import CellValue, SheetConfig, StructureInfo
from messy_xlsx.normalization.pipeline import NormalizationPipeline
from messy_xlsx.parsing.base_handler import ParseOptions
from messy_xlsx.parsing.handler_registry import HandlerRegistry
from messy_xlsx.sheet import MessySheet


# ============================================================================
# Core
# ============================================================================

class MessyWorkbook:
    """Main entry point for parsing Excel files."""

    def __init__(
        self,
        file_path_or_buffer: str | Path | BinaryIO,
        sheet_config: SheetConfig | None = None,
        formula_config: FormulaConfig | None = None,
        filename: str | None = None,
    ):
        """Open an Excel file for parsing.

        Args:
            file_path_or_buffer: Path to file, or file-like object (BytesIO, etc.)
            sheet_config: Configuration for parsing sheets
            formula_config: Configuration for formula evaluation
            filename: Optional filename hint when using file-like objects (for format detection)
        """
        self._sheet_config   = sheet_config or SheetConfig()
        self._formula_config = formula_config or FormulaConfig()

        self._detector      = FormatDetector()
        self._registry      = HandlerRegistry()
        self._analyzer      = StructureAnalyzer(get_structure_cache())
        self._formula_engine = FormulaEngine(self._formula_config)

        # Determine if input is file path or file-like object
        self._is_fileobj = hasattr(file_path_or_buffer, "read")

        if self._is_fileobj:
            self._file_path = None
            self._file_buffer = file_path_or_buffer
            self._filename_hint = filename

            # Detect format from file object
            self._format_info = self._detector.detect(file_path_or_buffer, filename=filename)

            # Reset buffer position after detection
            if hasattr(file_path_or_buffer, "seek"):
                file_path_or_buffer.seek(0)
        else:
            self._file_path = Path(file_path_or_buffer)
            self._file_buffer = None
            self._filename_hint = None

            if not self._file_path.exists():
                raise FileError(
                    f"File not found: {self._file_path}",
                    file_path = str(self._file_path),
                )

            self._format_info = self._detector.detect(self._file_path)

        if self._format_info.format_type == "unknown":
            file_desc = self._filename_hint or self._file_path or "<stream>"
            raise FormatError(
                f"Unknown file format: {file_desc}",
                file_path = str(file_desc),
            )

        # Validate extension matches detected format for Excel files
        # This catches files with .xlsx extension but different content
        if not self._is_fileobj:
            file_ext = self._file_path.suffix.lower()
            excel_extensions = {".xlsx", ".xlsm", ".xltx", ".xltm", ".xls"}
            if file_ext in excel_extensions and self._format_info.format_type not in ("xlsx", "xlsm", "xls", "xltx", "xltm"):
                raise FormatError(
                    f"File extension {file_ext} suggests Excel format, but content is {self._format_info.format_type}",
                    file_path=str(self._file_path),
                    detected_format=self._format_info.format_type,
                )

        # Get sheet names and validate file is readable
        source = self._file_buffer if self._is_fileobj else self._file_path
        self._sheet_names = self._registry.get_sheet_names(source)

        # Validate that the file is actually readable (not just format-detected)
        # This catches corrupted files that pass format detection but can't be opened
        if self._format_info.format_type in ("xlsx", "xlsm", "xltx", "xltm", "xls"):
            is_valid, error = self._registry.validate(source, self._format_info.format_type)
            if not is_valid:
                file_desc = self._filename_hint or self._file_path or "<stream>"
                raise FormatError(
                    f"File appears corrupted or invalid: {error}",
                    file_path=str(file_desc),
                    detected_format=self._format_info.format_type,
                )

        self._sheets: dict[str, MessySheet] = {}

        if self._formula_config.mode != FormulaEvaluationMode.DISABLED:
            if self._formula_engine.is_available:
                try:
                    source = self._file_buffer if self._is_fileobj else self._file_path
                    self._formula_engine.load_workbook(source)
                except Exception:
                    pass

        self._wb: openpyxl.Workbook | None = None

    @property
    def file_path(self) -> Path | None:
        """Path to the Excel file, or None if reading from buffer."""
        return self._file_path

    @property
    def source(self) -> Path | BinaryIO:
        """The source file path or buffer."""
        return self._file_buffer if self._is_fileobj else self._file_path

    @property
    def sheet_names(self) -> list[str]:
        """List of sheet names in the workbook."""
        return self._sheet_names.copy()

    @property
    def format_type(self) -> str:
        """Detected file format (xlsx, xls, csv, etc.)."""
        return self._format_info.format_type

    def get_sheet(self, name: str | None = None) -> MessySheet:
        """Get a sheet by name."""
        if name is None:
            name = self._sheet_names[0]

        if name not in self._sheet_names:
            file_desc = self._file_path or self._filename_hint or "<stream>"
            raise FormatError(
                f"Sheet '{name}' not found",
                file_path = str(file_desc),
            )

        if name not in self._sheets:
            self._sheets[name] = MessySheet(self, name)

        return self._sheets[name]

    def to_dataframe(
        self,
        sheet: str | None = None,
        config: SheetConfig | None = None,
    ) -> pd.DataFrame:
        """Convert a sheet to a pandas DataFrame."""
        sheet_name = sheet or self._sheet_names[0]
        return self._parse_sheet(sheet_name, config)

    def to_dataframes(
        self,
        config: SheetConfig | None = None,
    ) -> dict[str, pd.DataFrame]:
        """Convert all sheets to DataFrames."""
        result = {}
        for name in self._sheet_names:
            try:
                result[name] = self._parse_sheet(name, config)
            except Exception as e:
                pass
        return result

    def get_structure(self, sheet: str | None = None) -> StructureInfo:
        """Get detected structure for a sheet."""
        sheet_name = sheet or self._sheet_names[0]
        return self._analyze_structure(sheet_name)

    def get_cell(
        self,
        sheet: str,
        row: int,
        col: int,
    ) -> CellValue:
        """Get a single cell value."""
        self._ensure_workbook()

        ws   = self._wb[sheet]
        cell = ws.cell(row, col)

        cached_value = cell.value

        formula    = None
        is_formula = False
        if hasattr(cell, "data_type") and cell.data_type == "f":
            is_formula = True
            if hasattr(cell, "value") and isinstance(cell.value, str):
                if cell.value.startswith("="):
                    formula = cell.value

        if is_formula and self._formula_config.mode != FormulaEvaluationMode.DISABLED:
            try:
                cached_value = self._formula_engine.evaluate(
                    sheet, row, col, cached_value
                )
            except Exception:
                pass

        data_type = self._get_data_type(cached_value)

        is_merged = self._is_cell_merged(ws, row, col)

        is_hidden = self._is_cell_hidden(ws, row, col)

        return CellValue(
            value           = cached_value,
            formula         = formula,
            is_merged       = is_merged,
            is_hidden       = is_hidden,
            data_type       = data_type,
            original_format = cell.number_format if hasattr(cell, "number_format") else None,
        )

    def get_cell_by_ref(self, ref: str) -> CellValue:
        """Get a cell by A1-style reference."""
        from messy_xlsx.utils import cell_ref_to_coords

        sheet, row, col = cell_ref_to_coords(ref)
        sheet           = sheet or self._sheet_names[0]
        return self.get_cell(sheet, row, col)

    def _parse_sheet(
        self,
        sheet: str,
        config: SheetConfig | None = None,
    ) -> pd.DataFrame:
        """Parse a sheet to DataFrame with normalization."""
        config = config or self._sheet_config

        # Skip structure analysis for CSV/TSV/TXT - these formats don't support openpyxl
        if config.auto_detect and self.format_type not in ("csv", "tsv", "txt"):
            structure        = self._analyze_structure(sheet, config)
            effective_config = self._apply_structure_detection(config, structure)
        else:
            effective_config = config

        parse_options = ParseOptions(
            skip_rows      = effective_config.skip_rows,
            header_rows    = effective_config.header_rows,
            skip_footer    = effective_config.skip_footer,
            merge_strategy = effective_config.merge_strategy,
            ignore_hidden  = not effective_config.include_hidden,
            cell_range     = effective_config.cell_range,
            data_only      = True,
        )

        # Reset buffer position before parsing
        if self._is_fileobj and hasattr(self._file_buffer, "seek"):
            self._file_buffer.seek(0)

        df = self._registry.parse(
            self.source,
            sheet   = sheet,
            options = parse_options,
        )

        # Skip normalization if disabled
        if not effective_config.normalize:
            # Still apply sanitization if enabled
            if effective_config.sanitize_column_names:
                df = self._sanitize_columns(df)
            if effective_config.column_renames:
                df = df.rename(columns=effective_config.column_renames)
            return df

        pipeline = NormalizationPipeline(
            decimal_separator   = None,
            thousands_separator = None,
        )

        type_hints = effective_config.type_hints.copy()

        # Build skip_steps based on config
        skip_steps = []
        if not effective_config.normalize_whitespace:
            skip_steps.append("whitespace")
        if not effective_config.normalize_numbers:
            skip_steps.append("numbers")
        if not effective_config.normalize_dates:
            skip_steps.append("dates")

        df = pipeline.normalize(df, semantic_hints=type_hints, skip_steps=skip_steps)

        # Sanitize column names if requested
        if effective_config.sanitize_column_names:
            df = self._sanitize_columns(df)

        # Apply user renames (user overrides take precedence)
        if effective_config.column_renames:
            df = df.rename(columns=effective_config.column_renames)

        return df

    def _analyze_structure(self, sheet: str, config: SheetConfig | None = None) -> StructureInfo:
        """Analyze sheet structure."""
        header_patterns = config.header_patterns if config else None
        # Reset buffer position before analysis
        if self._is_fileobj and hasattr(self._file_buffer, "seek"):
            self._file_buffer.seek(0)
        return self._analyzer.analyze(self.source, sheet, header_patterns=header_patterns)

    def _apply_structure_detection(
        self,
        config: SheetConfig,
        structure: StructureInfo,
    ) -> SheetConfig:
        """Merge user config with detected structure."""
        from .exceptions import StructureError

        # Determine skip_rows and header_rows based on detection mode
        skip_rows   = config.skip_rows
        header_rows = config.header_rows

        if config.header_detection_mode == "auto":
            # Trust detection if confidence >= threshold
            if (
                structure.header_row is not None
                and structure.header_confidence >= config.header_confidence_threshold
            ):
                skip_rows   = max(0, structure.header_row - 1)
                header_rows = structure.header_rows_count
            else:
                # Fallback logic
                if config.header_fallback == "first_row":
                    skip_rows   = 0
                    header_rows = 1
                elif config.header_fallback == "none":
                    header_rows = 0
                elif config.header_fallback == "error":
                    raise StructureError(
                        f"No header detected with sufficient confidence "
                        f"(found: {structure.header_confidence:.2f}, "
                        f"required: {config.header_confidence_threshold:.2f})"
                    )

        elif config.header_detection_mode == "smart":
            # Use detection unless user explicitly overrode
            if (
                config.skip_rows == 0
                and structure.header_row is not None
                and structure.header_confidence >= config.header_confidence_threshold
            ):
                skip_rows   = max(0, structure.header_row - 1)
                header_rows = structure.header_rows_count
            else:
                skip_rows   = config.skip_rows
                header_rows = config.header_rows

        # "manual" mode: just use config values (no changes)

        return SheetConfig(
            skip_rows             = skip_rows,
            header_rows           = header_rows,
            skip_footer           = config.skip_footer if config.skip_footer > 0 else structure.suggested_skip_footer,
            cell_range            = config.cell_range,
            column_renames        = config.column_renames,
            type_hints            = config.type_hints,
            auto_detect           = False,
            include_hidden        = config.include_hidden,
            merge_strategy        = config.merge_strategy,
            locale                = config.locale or structure.detected_locale,
            evaluate_formulas     = config.evaluate_formulas,
            drop_regex            = config.drop_regex,
            drop_conditions       = config.drop_conditions,
            header_detection_mode = config.header_detection_mode,
            header_confidence_threshold = config.header_confidence_threshold,
            header_fallback       = config.header_fallback,
            multi_row_headers     = config.multi_row_headers,
            header_patterns       = config.header_patterns,
            # Normalization options
            normalize             = config.normalize,
            normalize_dates       = config.normalize_dates,
            normalize_numbers     = config.normalize_numbers,
            normalize_whitespace  = config.normalize_whitespace,
            sanitize_column_names = config.sanitize_column_names,
        )

    def _sanitize_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """Sanitize column names for BigQuery compatibility."""
        from .utils import sanitize_column_name

        new_columns = []
        seen: dict[str, int] = {}

        for col in df.columns:
            clean = sanitize_column_name(col)

            # Handle duplicates by appending counter
            if clean in seen:
                seen[clean] += 1
                clean = f"{clean}_{seen[clean]}"
            else:
                seen[clean] = 0

            new_columns.append(clean)

        df.columns = new_columns
        return df

    def _ensure_workbook(self) -> None:
        """Ensure openpyxl workbook is loaded."""
        if self._wb is None:
            # Reset buffer position before loading
            if self._is_fileobj and hasattr(self._file_buffer, "seek"):
                self._file_buffer.seek(0)

            # Load with data_only=False to preserve formula information
            # Note: read_only=True is incompatible with some features (merged_cells)
            self._wb = openpyxl.load_workbook(
                self.source,
                read_only = False,
                data_only = False,
            )

    def _get_data_type(self, value: Any) -> str:
        """Determine data type string for a value."""
        if value is None:
            return "empty"
        if isinstance(value, bool):
            return "boolean"
        if isinstance(value, (int, float)):
            return "number"
        if isinstance(value, str):
            if value.startswith("#") and value.endswith("!"):
                return "error"
            return "text"
        if hasattr(value, "date"):
            return "date"
        return "text"

    def _is_cell_merged(self, ws, row: int, col: int) -> bool:
        """Check if cell is part of a merged range."""
        try:
            for merged_range in ws.merged_cells.ranges:
                if (
                    merged_range.min_row <= row <= merged_range.max_row
                    and merged_range.min_col <= col <= merged_range.max_col
                ):
                    return True
        except Exception:
            pass
        return False

    def _is_cell_hidden(self, ws, row: int, col: int) -> bool:
        """Check if cell is in a hidden row or column."""
        try:
            if row in ws.row_dimensions and ws.row_dimensions[row].hidden:
                return True
            from openpyxl.utils import get_column_letter
            col_letter = get_column_letter(col)
            if col_letter in ws.column_dimensions and ws.column_dimensions[col_letter].hidden:
                return True
        except Exception:
            pass
        return False

    def close(self) -> None:
        """Close the workbook and release resources."""
        if self._wb is not None:
            self._wb.close()
            self._wb = None

    def __enter__(self) -> "MessyWorkbook":
        return self

    def __exit__(self, exc_type, exc_val, exc_tb) -> None:
        self.close()

    def __repr__(self) -> str:
        if self._file_path:
            name = self._file_path.name
        elif self._filename_hint:
            name = self._filename_hint
        else:
            name = "<stream>"
        return f"MessyWorkbook({name!r}, sheets={self._sheet_names})"
