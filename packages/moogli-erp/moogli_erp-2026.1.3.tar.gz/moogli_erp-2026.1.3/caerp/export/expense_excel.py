"""
Tools used to export datas in xls format
"""

import itertools
import logging
from string import ascii_uppercase

from openpyxl.styles import Color
from sqla_inspect.excel import XlsWriter

from caerp.compute.math_utils import integer_to_amount
from caerp.export.excel import (
    BOLD_CELL,
    FOOTER_CELL,
    HEADER_STYLE,
    LARGE_FOOTER_CELL,
    NUMBER_CELL,
    TITLE_STYLE,
)
from caerp.models.base import DBSESSION
from caerp.models.expense.sheet import BaseExpenseLine
from caerp.models.expense.types import ExpenseKmType, ExpenseTelType, ExpenseType
from caerp.utils import strings

log = logging.getLogger(__name__)


# A, B, C, ..., AA, AB, AC, ..., ZZ
ASCII_UPPERCASE = list(ascii_uppercase) + list(
    "".join(duple)
    for duple in itertools.combinations_with_replacement(ascii_uppercase, 2)
)


class Column:
    """
    A column object
    """

    def __init__(self, label, index=None, last_index=None):
        self.label = label
        self.code = ""
        self.ht = 0
        self.force_visible = False
        self.additional_cell_nb = None
        self.style = None
        self.set_index(index, last_index)

    def set_index(self, start, end=None):
        self.start = start
        self.end = end or start

    @property
    def start_index(self):
        return self.start

    @property
    def end_index(self):
        return self.end

    def reset_ht(self):
        self.ht = 0


class StaticColumn(Column):
    """
    A static column object representing static datas representation
    """

    static = True

    def __init__(
        self,
        label,
        key,
        formatter=None,
        style=None,
        nb_col=None,
        index=None,
        last_index=None,
    ):
        Column.__init__(self, label, index, last_index)
        self.key = key
        self.formatter = formatter
        self.style = style
        self.additional_cell_nb = nb_col

    def get_val(self, line):
        val = getattr(line, self.key, "")
        if self.formatter is not None and val:
            val = self.formatter(val)
        return val


class TypedColumn(Column):
    static = False

    def __init__(self, expense_type, label=None, index=None, last_index=None):
        if label is None:
            label = expense_type.label

        Column.__init__(self, label)
        self.id = expense_type.id
        self.code = expense_type.code
        self.set_index(index, last_index)

    def get_val(self, line):
        if hasattr(line, "ht"):
            val = integer_to_amount(line.ht)
        else:
            val = integer_to_amount(line.total)
        # Le total_ht s'affiche en bas de page en mode calculée
        self.ht += integer_to_amount(line.total_ht)
        return val


EXPENSEKM_COLUMNS = [
    StaticColumn(
        key="date",
        label="Date",
        index=1,
    ),
    StaticColumn(key="vehicle", label="Type de véhicule", index=2, last_index=3),
    StaticColumn(
        key="start",
        label="Lieu de départ",
        index=4,
        last_index=5,
    ),
    StaticColumn(key="end", label="Lieu d'arrivée", index=6, last_index=7),
    StaticColumn(
        key="description", label="Description/Mission", index=8, last_index=10
    ),
    StaticColumn(
        formatter=integer_to_amount, key="km", label="Nombre de kms", index=11
    ),
    StaticColumn(
        formatter=integer_to_amount,
        key="total",
        label="Indemnités",
        index=12,
        style=NUMBER_CELL,
    ),
]


class XlsExpense(XlsWriter):
    """
    Xls exporter of an expensesheet object

    Provide two sheets : the expenses and the kilometric datas
    """

    title = "NDF"

    def __init__(self, expensesheet):
        XlsWriter.__init__(self)
        self.model = expensesheet
        self.columns = self.get_columns()
        self.index = 2

    def get_merged_cells(self, start, end):
        """
        returned merged cells of the current line index
        """
        self.worksheet.merge_cells(
            start_row=self.index, end_row=self.index, start_column=start, end_column=end
        )
        cell = self.worksheet.cell(self.index, start)
        return cell

    def get_tel_column(self):
        """
        Return the columns associated to telephonic expenses
        """
        teltype = ExpenseTelType.query().first()
        col = None
        if teltype:
            # Tel expenses should be visible
            col = TypedColumn(teltype, label="Téléphonie")
            if teltype.initialize:
                col.force_visible = True
        return col

    def get_km_column(self):
        """
        Return the columns associated to km expenses
        """
        kmtype = ExpenseKmType.query().first()
        col = None
        if kmtype:
            col = TypedColumn(
                kmtype,
                label="Frais de déplacement",
            )
        return col

    def get_disabled_types_columns(self):
        """ """
        types = []
        for line in self.model.lines:
            type_ = line.expense_type
            if not type_.active and type_.type != "expensetel":
                if type_.id not in types:
                    types.append(type_.id)
                    yield TypedColumn(
                        type_,
                        label="%s (ce type de dépense n'existe plus)" % (type_.label),
                    )

    def get_columns(self):
        """
        Retrieve all columns and define a global column attribute
        :param internal: are we asking columns for internal expenses
        """
        columns = []
        # Add the two first columns
        columns.append(StaticColumn(label="Date", key="date"))
        columns.append(StaticColumn(label="Description", key="description", nb_col=3))

        # Telephonic fees are only available as internal expenses
        tel_column = self.get_tel_column()
        if tel_column is not None:
            columns.append(tel_column)

        km_column = self.get_km_column()
        if km_column:
            columns.append(km_column)
            kmtype_code = km_column.code
        else:
            kmtype_code = None

        type_ids = [
            i[0]
            for i in DBSESSION()
            .query(BaseExpenseLine.type_id)
            .filter_by(sheet_id=self.model.id)
            .distinct()
            .order_by(BaseExpenseLine.type)
        ]
        commontypes = ExpenseType.query().filter(
            ExpenseType.id.in_(type_ids), ExpenseType.type == "expense"
        )

        for type_ in commontypes:
            # Here's a hack to allow to group km fee types and displacement
            # fees
            if kmtype_code is not None and type_.code != kmtype_code:
                columns.append(TypedColumn(type_))

        columns.extend(self.get_disabled_types_columns())

        # Add the last columns
        columns.append(
            StaticColumn(label="Total HT", key="total_ht", formatter=integer_to_amount)
        )
        columns.append(
            StaticColumn(label="Tva", key="tva", formatter=integer_to_amount)
        )
        columns.append(
            StaticColumn(
                label="Total",
                key="total",
                formatter=integer_to_amount,
                style=NUMBER_CELL,
            )
        )

        # We set the appropriate letter to each column
        current_index = 1
        for col in columns:
            index = current_index
            additional_cell_nb = col.additional_cell_nb
            if additional_cell_nb:
                last_index = index + additional_cell_nb
                current_index += additional_cell_nb + 1
            else:
                last_index = index
                current_index += 1
            col.set_index(index, last_index)
        return columns

    def _write_inline(self, label, value):
        """
        Write inline label value on 3 - 3 columns

        :param str label: The label
        :param str value: The value
        """
        cell = self.get_merged_cells(1, 4)
        cell.value = label
        cell = self.get_merged_cells(5, 10)
        cell.value = value
        self.index += 1

    def write_company(self):
        """
        write the company code in the header
        """
        name = self.model.company.name
        self._write_inline("Enseigne", name)

    def write_code(self):
        """
        write the company code in the header
        """
        code = self.model.company.code_compta
        if not code:
            code = "Code non renseigné"
        self._write_inline("Code analytique de l'enseigne", code)

    def write_user(self):
        """
        write the username in the header
        """
        self._write_inline("Nom de l'entrepreneur", self.model.user.label)

    def write_period(self):
        """
        write the period in the header
        """
        period = "{0} {1}".format(strings.month_name(self.model.month), self.model.year)
        self._write_inline("Période de la demande", period)

    def write_title(self):
        """
        write the title in the header
        """
        self._write_inline("Titre", self.model.title if self.model.title else "")

    def write_number(self):
        """
        write the expense sheet id in the header
        """
        if self.model.status != "valid":
            number = "Ce document n'a pas été validé"
        else:
            number = self.model.official_number
        self._write_inline("Numéro de pièce", number)

    def write_global_total_ht(self):
        """
        Write the total ht in the upper part of the sheet
        """
        self._write_inline(
            "Total HT",
            "%s €" % strings.format_amount(self.model.total_ht, grouping=False),
        )

    def write_global_total_tva(self):
        """
        Write the total tva in the upper part of the sheet
        """
        self._write_inline(
            "Total TVA",
            "%s €" % strings.format_amount(self.model.total_tva, grouping=False),
        )

    def write_global_total_ttc(self):
        """
        Write the total tva in the upper part of the sheet
        """
        self._write_inline(
            "Total TTC",
            "%s €" % strings.format_amount(self.model.total, grouping=False),
        )

    def get_column_cell(self, column):
        """
        Return the cell corresponding to a given column
        """
        index = column.start_index
        last_index = column.end_index
        return self.get_merged_cells(index, last_index)

    def write_table_header(self, columns):
        """
        Write the table's header and its subheader
        """
        for column in columns:
            cell = self.get_column_cell(column)
            cell.style = HEADER_STYLE
            cell.value = column.label
        self.index += 1
        for column in columns:
            cell = self.get_column_cell(column)
            cell.style = BOLD_CELL
            cell.value = column.code
        self.index += 1

    def get_formatted_cell_val(self, line, column):
        """
        For a given expense line, check if a value should be provided in the
        given column
        """
        val = ""

        if line.expense_type is not None and column.static:
            val = column.get_val(line)

        return val

    def get_cell_val(self, line, column, by_id=True):
        """
        For a given expense line, check if a value should be provided in the
        given column

        :param obj line: a expense line object
        :param dict column: a dict describing a column
        :param bool by_id: Should the match be done by id
        :return: a value if the the given line is form the type of column ''
        """
        val = ""
        # Première passe, on essaye de retrouver le type de dépense par id
        if by_id:
            if column.id == line.expense_type.id:
                val = column.get_val(line)

        # Deuxième passe, on essaye de retrouver le type de dépense par code
        else:
            if column.code == line.expense_type.code:
                val = column.get_val(line)

        return val

    def set_col_width(self, col_letter, width, force=False):
        """
        Set the width of a given column

        :param str col_letter: the letter for the column
        :param int width: The width of the given column
        :param bool force: force the display of the column
        """
        col_dim = self.worksheet.column_dimensions.get(col_letter)
        if col_dim:
            if (
                col_dim.width
                in (
                    -1,
                    None,
                )
                or force
            ):
                if width == 0:
                    col_dim.hidden = True
                else:
                    col_dim.width = width
                    col_dim.hidden = False

    def write_table(self, columns, lines):
        """
        write a table with headers and content
        :param columns: list of dict
        :params lines: list of models to be written
        """
        self.write_table_header(columns)
        for line in lines:
            got_value = False

            for column in columns:
                cell = self.get_column_cell(column)

                if column.static:
                    # On récupère les valeurs pour les colonnes fixes
                    value = self.get_formatted_cell_val(
                        line,
                        column,
                    )
                else:
                    # On récupère les valeurs pour les colonnes spécifiques à
                    # chaque type de données

                    # Première passe on essaye de remplir les colonnes pour la
                    # ligne de dépense données en fonction de l'id du type de
                    # dépense associé
                    value = self.get_cell_val(line, column, by_id=True)
                    if value:
                        got_value = True

                cell.value = value
                if column.style:
                    cell.style = column.style

            # Deuxième passe, on a rempli aucune case pour cette ligne on va
            # essayer de remplir les colonnes en recherchant le type de dépense
            # par code
            if not got_value:
                for column in columns:
                    cell = self.get_column_cell(column)

                    if not column.static and not got_value:
                        value = self.get_cell_val(
                            line,
                            column,
                            by_id=False,
                        )
                        if value:
                            got_value = True

                        cell.value = value

                        if column.style:
                            cell.style = column.style

            self.index += 1

        self.write_table_footer(columns, lines)

    def write_table_footer(self, columns, lines):
        """
        Write table footer (total, ht, tva, km)

        :param list columns: The columns as described in the static vars here
        above
        :param list lines: The lines presented in this table
        """
        for column in columns:
            cell = self.get_column_cell(column)
            cell.style = FOOTER_CELL

            if not column.static:
                value = column.ht
                cell.value = value

                if value == 0 and not column.force_visible:
                    col_width = 0
                else:
                    col_width = 13
                self.set_col_width(column.start_index, col_width)

            elif column.key == "description":
                cell.value = "Totaux"

            elif column.key == "total_ht":
                cell.value = integer_to_amount(
                    sum([getattr(line, "total_ht", 0) for line in lines])
                )

            elif column.key == "tva":
                cell.value = integer_to_amount(
                    sum([getattr(line, "total_tva", 0) for line in lines])
                )

            elif column.key == "km":
                cell.value = integer_to_amount(
                    sum([getattr(line, "km", 0) for line in lines])
                )

            elif column.key == "total":
                cell.value = integer_to_amount(sum([line.total for line in lines]))

        self.index += 4

    def write_expense_table(self, category):
        """
        write expenses tables for the given category
        """
        lines = [line for line in self.model.lines if line.category == category]
        kmlines = [lin for lin in self.model.kmlines if lin.category == category]
        lines.extend(kmlines)
        self.write_table(self.columns, lines)
        self.index += 2

        for column in self.columns:
            column.reset_ht()

    def write_full_line(self, txt, start=1, end=10):
        """
        Write a full line, merging cells
        """
        cell = self.get_merged_cells(start, end)
        cell.value = txt
        self.index += 1
        return cell

    def write_internal_expenses(self):
        """
        write the internal expense table to the current worksheet
        """
        txt = "FRAIS (dépenses directes liées au fonctionnement)"
        cell = self.write_full_line(txt)
        self.set_color(cell, Color.Crimson)
        self.write_expense_table("1")

    def write_activity_expenses(self):
        """
        write the activity expense table to the current worksheet
        """
        txt = "ACHATS (dépenses concernant directement l'activité auprès \
de vos clients)"
        cell = self.write_full_line(txt)
        self.set_color(cell, Color.Crimson)
        self.write_expense_table("2")

    def write_total(self):
        """
        write the final total
        """
        cell = self.get_merged_cells(1, 4)
        cell.value = "Total des dépenses professionnelles à payer"
        cell.style = LARGE_FOOTER_CELL
        cell = self.get_merged_cells(5, 5)
        cell.value = integer_to_amount(self.model.total)
        cell.style = LARGE_FOOTER_CELL
        self.index += 2

    def write_accord(self):
        """
        Write the endline
        """
        cell = self.get_merged_cells(2, 5)
        cell.value = "Accord après vérification"
        self.index += 1
        self.worksheet.merge_cells(
            start_row=self.index,
            end_row=self.index + 4,
            start_column=1,
            end_column=4,
        )

    def write_vehicle_information(self):
        user = self.model.user
        columns = [
            StaticColumn(
                label="Puissance fiscale : ", key="ps_label", index=1, last_index=2
            ),
            StaticColumn(
                label="",
                index=3,
                key="fiscal_power",
            ),
            StaticColumn(
                key="r_label",
                label="Plaque : ",
                index=6,
            ),
            StaticColumn(
                label="",
                index=7,
                key="registration",
            ),
        ]
        for column in columns:
            cell = self.get_column_cell(column)
            if column.key == "fiscal_power":
                cell.value = user.vehicle_fiscal_power
            elif column.key == "registration":
                cell.value = user.vehicle_registration
            else:
                cell.value = column.label

    def write_km_book(self):
        """
        Write the km book associated to this expenses
        """
        self.index = 3
        user = self.model.user
        title = "Tableau de bord kilométrique de {0} {1}".format(
            user.lastname, user.firstname
        )
        cell = self.write_full_line(title)
        cell.style = TITLE_STYLE

        # index has already been increased
        row_dim = self.worksheet.row_dimensions[self.index - 1]
        row_dim.height = 30
        self.index += 1
        self.write_vehicle_information()
        self.index += 3
        self.write_table(EXPENSEKM_COLUMNS, self.model.kmlines)

    def render(self):
        """
        Return the current excel export as a String buffer (BytesIO)
        """
        cell = self.write_full_line("Notes de dépenses")

        cell.style = TITLE_STYLE
        # index has already been increased
        row_dim = self.worksheet.row_dimensions[self.index - 1]
        row_dim.height = 30
        self.index += 2

        self.write_number()
        self.write_company()
        self.write_code()
        self.write_user()
        self.write_period()
        self.write_title()
        self.write_global_total_ht()
        self.write_global_total_tva()
        self.write_global_total_ttc()
        self.index += 1
        self.write_internal_expenses()
        self.write_activity_expenses()
        self.write_total()
        self.write_accord()

        #        # We set a width to all columns that have no width set (-1)
        #        for let in ASCII_UPPERCASE:
        #            self.set_col_width(let, 13)
        #
        self.worksheet = self.book.create_sheet()
        self.worksheet.title = "Journal de bord"
        self.write_km_book()

        for let in ASCII_UPPERCASE:
            col_dim = self.worksheet.column_dimensions.get(let)
            if col_dim:
                col_dim.width = 13

        return self.save_book()
