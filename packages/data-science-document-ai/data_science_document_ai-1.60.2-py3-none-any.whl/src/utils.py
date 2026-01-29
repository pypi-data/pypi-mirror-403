"""Utility functions for data science projects."""
import asyncio
import hashlib
import io  # type: ignore
import json
import os
import pickle
from datetime import datetime
from typing import Any, Dict, List, Literal, Optional

import httpx
import numpy as np
import openpyxl
import pandas as pd
from google.cloud import documentai_v1beta3 as docu_ai_beta
from pypdf import PdfReader, PdfWriter

from src.io import bq_logs, get_storage_client, logger


def get_pdf_page_count(pdf_bytes):
    """Get the number of pages in a PDF document efficiently.

    Args:
        pdf_bytes (bytes): The PDF content as bytes.

    Returns:
        int: The number of pages in the PDF.
    """
    reader = PdfReader(io.BytesIO(pdf_bytes))
    return len(reader.pages)


async def get_data_set_schema_from_docai(
    schema_client, project_id=None, location=None, processor_id=None, name=None
):
    """Get the existing processor schema.

    Args:
        schema_client (documentai_v1beta3.DocumentServiceClient): The schema client.
        project_id (str): The ID of the project.
        location (str): The location of the project.
        processor_id (str): The id of the processor.
        name (str, optional): The name of the dataset schema. Defaults to None.
        The format is "projects/{project_id}/locations/{location}/processors/{processor_id}/dataset/datasetSchema".

    Returns:
        documentai.DatasetSchema: The schema of the dataset.
    """
    # Check if the name is provided as per the format
    if name and not name.endswith("datasetSchema"):
        name = f"{name}/dataset/datasetSchema"

    # Initialize request argument(s)
    if name:
        request = docu_ai_beta.GetDatasetSchemaRequest(
            name=name, visible_fields_only=True
        )
    else:
        request = docu_ai_beta.GetDatasetSchemaRequest(
            name=schema_client.dataset_schema_path(project_id, location, processor_id),
            visible_fields_only=True,
        )

    # Make the request
    response = await schema_client.get_dataset_schema(request=request)

    return response.document_schema


def get_processor_name(
    params, input_doc_type, version: Literal["stable", "beta"] = "stable"
):
    """Access models based on the environment and isBetaTest."""
    g_ai_project_id = params["models_project_id"]

    doctype_models_list = params["model_config"][version][input_doc_type]
    selected_model_idx = params["model_selector"][version][input_doc_type]
    processor_id = doctype_models_list[selected_model_idx]["id"]

    processor_name = (
        f"projects/{g_ai_project_id}/locations/eu/processors/{processor_id}"
    )
    logger.info(f"Processor: {processor_name}")

    logger.info(f"Processor ID for {input_doc_type}: {processor_id}")

    return processor_name


async def validate_based_on_schema(params, extracted_raw_data, processor_name):
    """Validate the extracted data based on the schema."""
    # Get the schema of a processor and select only the entity types
    schema_response = get_data_set_schema(params, processor_name)
    # schema_response.document_schema.entity_types contains 2 elements:
    # One for entities at document level
    # One for entities at line item level
    schemas = schema_response.entity_types
    schema_header_fields = schemas[0].properties

    result = dict(extracted_raw_data)
    for data_field in schema_header_fields:
        if "once" in data_field.occurrence_type.name.lower():
            if data_field.name in result:
                result[data_field.name] = result[data_field.name][0]

    # Exclude the fields that are not in the schema.This is to avoid submitting additional fields from General AI to PAW
    result = {
        key: value
        for key, value in result.items()
        if key in [_entity.name for _entity in schema_header_fields]
    }

    return result


def store_json_in_gcs(
    params, document_id, json_data, folder_path="docai_entity_storage/"
):
    """Store a JSON object in a Google Cloud Storage bucket.

    Args:
        params (dict): The parameters dictionary.
        document_id (str): The document ID.
        json_data (dict): The JSON data to be stored.
        folder_path (str): The folder path in the GCS bucket. Default is "docai_entity_storage/".
    """
    try:
        storage_client = get_storage_client(params)
        bucket = storage_client.bucket(params.get("doc_ai_bucket_name"))
        full_object_name = folder_path + document_id
        blob = bucket.blob(full_object_name)

        # Convert dict to JSON string if needed
        json_string = (
            json.dumps(json_data) if isinstance(json_data, dict) else json_data
        )
        blob.upload_from_string(json_string, content_type="application/json")

        logger.info(
            f"JSON object stored successfully in gs://{params.get('doc_ai_bucket_name')}/{full_object_name}"  # noqa
        )

    except Exception as e:
        logger.error(f"Error storing JSON object in GCS: {e}")


# Execute synchronous functions in the background
async def run_background_tasks(
    params,
    doc_id,
    docType,
    extracted_data,
    store_data,
    processor_version,
    mime_type,
    elapsed_time=None,
    page_count=None,
):
    """
    Run background tasks asynchronously.

    Args:
        params (dict): The parameters dictionary.
        doc_id (str): The document ID.
        docType (str): The document type code.
        extracted_data (dict): The extracted data from the document.
        store_data: The data to store in GCS.
        processor_version: The processor version used to extract the data.
        mime_type: The MIME type of the document.
        elapsed_time: The time taken to process the document.
        page_count (int, optional): The number of pages in the document.

    Returns:
        None
    """
    loop = asyncio.get_running_loop()

    await loop.run_in_executor(None, store_json_in_gcs, params, doc_id, store_data)

    # Use the passed page_count or default to 0 if not provided
    if page_count is None:
        page_count = 0

    # Log the request in BigQuery
    await loop.run_in_executor(
        None,
        bq_logs,
        [
            {
                "session_id": params["session_id"],
                "upload_date": datetime.utcnow().isoformat(),
                "doc_id": doc_id,
                "documentTypeCode": docType,
                "status": "processed",
                "response": json.dumps(extracted_data),
                "processor_version": processor_version,
                "page_count": page_count,
                "mime_type": mime_type,
                "elapsed_time": elapsed_time,
            }
        ],
        params,
    )


def get_excel_sheets(file_content, mime_type):
    """Get the sheet names from the Excel file.

    Args:
        file_content (bytes): The content of the Excel file.
        mime_type (str): The MIME type of the file.

    Returns:
        sheets (list): The list of sheet names.
        openpyxl.Workbook: The workbook
    """
    file_stream = io.BytesIO(file_content)
    if "spreadsheet" in mime_type:
        workbook = openpyxl.load_workbook(file_stream, data_only=True)
        sheets = [
            sheet_name
            for sheet_name in workbook.sheetnames
            if workbook[sheet_name].sheet_state == "visible"
        ]
    else:
        workbook = pd.read_excel(file_stream, sheet_name=None)
        # Select only the sheets that are not empty
        sheets = [sheet for sheet in workbook.keys() if not workbook[sheet].empty]

    return sheets, workbook


def generate_schema_structure(params, input_doc_type):
    """
    Generate the schema placeholder and the JSON response structure.

    Args:
        params (dict): Parameters dictionary.
        input_doc_type (str): Document type to select the appropriate schema.

    Returns:
        dict: The response schema structure.
    """
    # Get the processor name and the Doc Ai schema of the processor
    processor_name = get_processor_name(params, input_doc_type)
    schema = get_data_set_schema(params, processor_name)

    def build_schema(entity):
        return {
            "type": "OBJECT",
            "properties": {
                prop.name: {
                    "type": "string",
                    "nullable": True,
                    "description": prop.description,
                }
                for prop in entity.properties
            },
            "required": [
                prop.name
                for prop in entity.properties
                if "REQUIRED" in prop.occurrence_type.name
            ],
        }

    # Build the response schema structure for the header fields
    response_schema = build_schema(
        next(
            e
            for e in schema.entity_types
            if e.name == "custom_extraction_document_type"
        )
    )

    # Build the child schemas
    child_schemas = {
        entity.name: {"type": "ARRAY", "items": build_schema(entity)}
        for entity in schema.entity_types
        if entity.name != "custom_extraction_document_type"
    }

    # Attach child schemas to the parent schema
    response_schema["properties"].update(child_schemas)

    # TODO: expand or remove this workaround after testing
    if input_doc_type in ["finalMbL", "draftMbl"]:
        response_schema["properties"]["plug"] = {
            "description": "",
            "nullable": True,
            "type": "string",
        }

    return response_schema


def get_hash_of_data(data):
    """Generate a hash for data."""
    sha256_hash = hashlib.sha256()
    if data:
        if isinstance(data, bytes):
            sha256_hash.update(data)
        else:
            sha256_hash.update(str(data).encode("utf-8"))
        return sha256_hash.hexdigest()
    return None


async def cache_on_disk(func, **kwargs):
    """Cache function result if the arguments are the same. Enableable via environment variable CACHE."""
    if os.getenv("CACHE") != "enabled":
        return await func(**kwargs)
    os.makedirs("cache", exist_ok=True)
    func_name = func.__name__
    serialized_kwargs = {k: get_hash_of_data(v) for k, v in kwargs.items()}
    serialized_kwargs_str = json.dumps(serialized_kwargs, sort_keys=True)
    unique_key = get_hash_of_data(serialized_kwargs_str)
    cache_file = os.path.join("cache", f"{func_name}_{unique_key}.pkl")

    # Try retrieving cached result
    if os.path.exists(cache_file):
        with open(cache_file, "rb") as f:
            cached_data = f.read()
            return pickle.loads(cached_data)

    # Execute the function and cache the result
    result = await func(**kwargs)
    with open(cache_file, "wb") as f:
        f.write(pickle.dumps(result))
    return result


async def update_response_schema_from_docai(params, schema_client):
    params["docai_schema_dict"] = params.get("docai_schema_dict", {})
    for version in params["model_config"]:
        for input_doc_type in params["model_config"][version]:
            processor_name = get_processor_name(params, input_doc_type, version)
            # Get schema
            schema = await get_data_set_schema_from_docai(
                schema_client, name=processor_name
            )
            params["docai_schema_dict"].update({processor_name: schema})


def get_data_set_schema(params, processor_name):
    return params["docai_schema_dict"][processor_name]


def extract_top_pages(pdf_bytes, num_pages=4):
    """Extract the top pages from a PDF document."""
    reader = PdfReader(io.BytesIO(pdf_bytes))
    writer = PdfWriter()

    for page_num in range(min(num_pages, len(reader.pages))):
        writer.add_page(reader.pages[page_num])

    output = io.BytesIO()
    writer.write(output)

    return output.getvalue()


async def get_tms_mappings(
    input_list: List[str], embedding_type: str, llm_ports: Optional[List[str]] = None
) -> Dict[str, Any]:
    """Get TMS mappings for the given values.

    Args:
        input_list (list[str]): List of strings to get embeddings for.
        embedding_type (str): Type of embedding to use
         (e.g., "container_types", "ports", "depots", "lineitems", "terminals").
        llm_ports (list[str], optional): List of LLM ports to use. Defaults to None.

    Returns:
        dict or string: A dictionary or a string with the mapping results.
    """
    base_url = (
        "http://0.0.0.0:8080/"
        if os.getenv("CLUSTER") is None
        else "http://tms-mappings.api.svc.cluster.local./"
    )

    # Ensure clean inputs
    if not input_list:
        return {}

    # Ensure input_list is a list
    if not isinstance(input_list, list):
        input_list = [input_list]

    # Always send a dict with named keys
    payload = {embedding_type: input_list}

    if llm_ports:
        payload["llm_ports"] = llm_ports if isinstance(llm_ports, list) else [llm_ports]

    # Make the POST request to the TMS mappings API
    url = f"{base_url}{embedding_type}"

    # Use a timeout so the code doesn't hang forever
    timeout = httpx.Timeout(60.0, connect=10.0)

    async with httpx.AsyncClient(timeout=timeout) as client:
        try:
            response = await client.post(url, json=payload)
            response.raise_for_status()

            # Structure expected: {"response": {"data": {"desc1": "code1", "desc2": "code2"}}}
            return response.json().get("response", {}).get("data", {})

        except httpx.HTTPStatusError as exc:
            logger.error(
                f"Error from TMS mappings API: {exc.response.status_code} - {exc.response.text}"
            )
            return {}


async def batch_fetch_all_mappings(container_types, terminals, depots):
    """Batch fetch all mappings for container types, terminals, and depots."""
    # run batch calls concurrently
    results = await asyncio.gather(
        get_tms_mappings(list(container_types), "container_types"),
        get_tms_mappings(list(terminals), "terminals"),
        get_tms_mappings(list(depots), "depots"),
    )

    batch_container_map, batch_terminal_map, batch_depot_map = results

    # Convert lists of tuples to dicts if necessary
    return (
        dict(batch_container_map or {}),
        dict(batch_terminal_map or {}),
        dict(batch_depot_map or {}),
    )


def transform_schema_strings(schema):
    """
    Recursively transforms a schema dictionary, replacing all "type": "STRING"
    definitions with a new object containing "value" and "page_number" fields.
    It preserves 'nullable' and 'description' fields by moving them to the
    new 'value' property.

    Args:
        schema (dict): The input schema dictionary.

    Returns:
        dict: The transformed schema dictionary.
    """
    if not isinstance(schema, dict):
        return schema

    schema_type = schema.get("type")
    if not schema_type:
        return schema

    # Base case: STRING → OBJECT (only if not already transformed)
    if schema_type.upper() == "STRING":
        return {
            "type": "OBJECT",
            "properties": {
                "value": {
                    "type": "STRING",
                    "nullable": schema.get("nullable", False),
                    "description": schema.get("description", ""),
                },
                "page_number": {
                    "type": "STRING",
                    "description": "Number of a page where the value was found in the document starting from 0.",
                },
            },
            "required": [],
        }

    # Skip already transformed OBJECT (has both 'value' & 'page_number')
    if (
        schema_type.upper() == "OBJECT"
        and "properties" in schema
        and {"value", "page_number"}.issubset(schema["properties"].keys())
    ):
        return schema

    # Recursive case for OBJECT
    if schema_type.upper() == "OBJECT" and "properties" in schema:
        new_schema = schema.copy()
        new_schema["properties"] = {
            k: transform_schema_strings(v) for k, v in schema["properties"].items()
        }
        return new_schema

    # Recursive case for ARRAY
    if schema_type.upper() == "ARRAY" and "items" in schema:
        new_schema = schema.copy()
        new_schema["items"] = transform_schema_strings(schema["items"])
        return new_schema

    return schema


def estimate_page_count(sheet):
    """Assuming a page is 10 columns x 50 rows."""
    if hasattr(sheet, "shape"):
        pg_cnt = sheet.shape[0] * sheet.shape[1]
    elif hasattr(sheet, "max_row"):
        pg_cnt = sheet.max_column * sheet.max_row
    else:
        return None
    return np.ceil(pg_cnt / 500)


def split_pdf_into_chunks(file_content: bytes, chunk_size: int = 1):
    """Split PDF into smaller page chunks."""
    pdf = PdfReader(io.BytesIO(file_content))
    total_pages = len(pdf.pages)

    # TODO: update the chunk_size based on doc length. However, it breaks the page number extraction logic.
    for i in range(0, total_pages, chunk_size):
        writer = PdfWriter()
        for j in range(i, min(i + chunk_size, total_pages)):
            writer.add_page(pdf.pages[j])

        buffer = io.BytesIO()
        writer.write(buffer)
        buffer.seek(0)

        yield buffer.getvalue()
