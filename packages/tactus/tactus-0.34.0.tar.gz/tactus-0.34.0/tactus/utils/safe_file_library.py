"""Safe file I/O libraries for Lua sandbox.

Provides sandboxed file operations restricted to the working directory.
All paths are validated to prevent directory traversal attacks.

Supported formats:
- File: Raw text read/write
- Csv: CSV with automatic header handling
- Tsv: Tab-separated values
- Json: JSON read/write
- Parquet: Apache Parquet (via pyarrow)
- Hdf5: HDF5 datasets (via h5py)
- Excel: Excel spreadsheets (via openpyxl)
"""

import csv
import json
import logging
import os
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)


class PathValidator:
    """Validates file paths are within allowed base directory."""

    def __init__(self, base_path: str):
        """
        Initialize path validator.

        Args:
            base_path: The base directory that all file operations are restricted to.
        """
        self.base_path = os.path.realpath(base_path)

    def validate(self, filepath: str) -> str:
        """
        Validate and resolve a file path.

        Args:
            filepath: Relative or absolute file path to validate.

        Returns:
            Resolved absolute path if valid.

        Raises:
            PermissionError: If path is outside the base directory.
        """
        # Join with base path and resolve to absolute
        resolved = os.path.realpath(os.path.join(self.base_path, filepath))

        # Check if resolved path is within base directory
        # Allow exact match (base_path itself) or paths that start with base_path + separator
        if resolved != self.base_path and not resolved.startswith(self.base_path + os.sep):
            raise PermissionError(f"Access denied: path outside working directory: {filepath}")

        return resolved


def create_safe_file_library(base_path: str) -> Dict[str, Any]:
    """
    Create raw text file operations library.

    Args:
        base_path: Base directory for file operations.

    Returns:
        Dictionary of file operation functions.
    """
    validator = PathValidator(base_path)

    def read(filepath: str) -> str:
        """Read entire file as text."""
        path = validator.validate(filepath)
        with open(path, "r", encoding="utf-8") as f:
            return f.read()

    def write(filepath: str, content: str) -> None:
        """Write text to file."""
        path = validator.validate(filepath)
        # Ensure parent directory exists
        os.makedirs(os.path.dirname(path), exist_ok=True) if os.path.dirname(path) else None
        with open(path, "w", encoding="utf-8") as f:
            f.write(content)

    def exists(filepath: str) -> bool:
        """Check if file exists."""
        try:
            path = validator.validate(filepath)
            return os.path.exists(path)
        except PermissionError:
            return False

    return {"read": read, "write": write, "exists": exists}


class LuaList:
    """
    Wrapper for Python lists that works better with Lua via lupa.

    Provides both 0-indexed access (Python style) and a len() method
    that can be called from Lua.
    """

    def __init__(self, data: List):
        self._data = data

    def __getitem__(self, index):
        # Lua method access comes through __getitem__ with string keys
        if isinstance(index, str):
            # Handle method access
            if index == "len":
                return self.len
            elif index == "get":
                return self.get
            else:
                raise KeyError(f"Unknown method: {index}")
        # Lua numbers are floats, convert to int for indexing
        if isinstance(index, float):
            index = int(index)
        return self._data[index]

    def __len__(self):
        return len(self._data)

    def __iter__(self):
        return iter(self._data)

    def len(self):
        """Return length - callable from Lua as data:len()."""
        return len(self._data)

    def get(self, index):
        """Alternative access method - data:get(0) instead of data[0]."""
        if isinstance(index, float):
            index = int(index)
        return self._data[index]


def create_safe_csv_library(base_path: str) -> Dict[str, Any]:
    """
    Create CSV file operations library.

    Args:
        base_path: Base directory for file operations.

    Returns:
        Dictionary of CSV operation functions.
    """
    validator = PathValidator(base_path)

    def read(filepath: str) -> LuaList:
        """Read CSV file, returning list of dictionaries with headers as keys."""
        path = validator.validate(filepath)
        with open(path, "r", encoding="utf-8", newline="") as f:
            reader = csv.DictReader(f)
            return LuaList(list(reader))

    def write(filepath: str, data: List[Dict], options: Optional[Dict] = None) -> None:
        """Write list of dictionaries to CSV file."""
        path = validator.validate(filepath)

        # Convert Lua table options to Python dict if needed
        if options and hasattr(options, "items"):
            options = dict(options.items())
        options = options or {}
        headers = options.get("headers")

        # Convert headers from Lua table to Python list if needed
        if headers and hasattr(headers, "values"):
            headers = list(headers.values())

        # Convert Lua table to Python list if needed
        if hasattr(data, "values"):
            data = list(data.values())

        if not headers and data:
            # Get headers from first row
            first_row = data[0]
            if hasattr(first_row, "keys"):
                headers = list(first_row.keys())
            elif isinstance(first_row, dict):
                headers = list(first_row.keys())
            else:
                raise ValueError("Cannot determine headers from data")

        # Ensure parent directory exists
        os.makedirs(os.path.dirname(path), exist_ok=True) if os.path.dirname(path) else None

        with open(path, "w", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            for row in data:
                # Convert Lua table to dict if needed
                if hasattr(row, "items"):
                    row = dict(row.items())
                writer.writerow(row)

    return {"read": read, "write": write}


def create_safe_tsv_library(base_path: str) -> Dict[str, Any]:
    """
    Create TSV (tab-separated values) file operations library.

    Args:
        base_path: Base directory for file operations.

    Returns:
        Dictionary of TSV operation functions.
    """
    validator = PathValidator(base_path)

    def read(filepath: str) -> LuaList:
        """Read TSV file, returning list of dictionaries with headers as keys."""
        path = validator.validate(filepath)
        with open(path, "r", encoding="utf-8", newline="") as f:
            reader = csv.DictReader(f, delimiter="\t")
            return LuaList(list(reader))

    def write(filepath: str, data: List[Dict], options: Optional[Dict] = None) -> None:
        """Write list of dictionaries to TSV file."""
        path = validator.validate(filepath)

        # Convert Lua table options to Python dict if needed
        if options and hasattr(options, "items"):
            options = dict(options.items())
        options = options or {}
        headers = options.get("headers")

        # Convert headers from Lua table to Python list if needed
        if headers and hasattr(headers, "values"):
            headers = list(headers.values())

        # Convert Lua table to Python list if needed
        if hasattr(data, "values"):
            data = list(data.values())

        if not headers and data:
            first_row = data[0]
            if hasattr(first_row, "keys"):
                headers = list(first_row.keys())
            elif isinstance(first_row, dict):
                headers = list(first_row.keys())
            else:
                raise ValueError("Cannot determine headers from data")

        # Ensure parent directory exists
        os.makedirs(os.path.dirname(path), exist_ok=True) if os.path.dirname(path) else None

        with open(path, "w", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=headers, delimiter="\t")
            writer.writeheader()
            for row in data:
                if hasattr(row, "items"):
                    row = dict(row.items())
                writer.writerow(row)

    return {"read": read, "write": write}


def create_safe_json_library(base_path: str) -> Dict[str, Any]:
    """
    Create JSON file operations library.

    Args:
        base_path: Base directory for file operations.

    Returns:
        Dictionary of JSON operation functions.
    """
    validator = PathValidator(base_path)

    def read(filepath: str) -> Any:
        """Read JSON file and return parsed data."""
        path = validator.validate(filepath)
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)

    def write(filepath: str, data: Any, options: Optional[Dict] = None) -> None:
        """Write data to JSON file."""
        path = validator.validate(filepath)

        # Convert Lua table options to Python dict if needed
        if options and hasattr(options, "items"):
            options = dict(options.items())
        options = options or {}
        indent = options.get("indent", 2)

        # Convert Lua tables to Python dicts/lists
        data = _lua_to_python(data)

        # Ensure parent directory exists
        os.makedirs(os.path.dirname(path), exist_ok=True) if os.path.dirname(path) else None

        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=indent, default=str)

    return {"read": read, "write": write}


def create_safe_parquet_library(base_path: str) -> Dict[str, Any]:
    """
    Create Parquet file operations library.

    Args:
        base_path: Base directory for file operations.

    Returns:
        Dictionary of Parquet operation functions.
    """
    import pyarrow as pa
    import pyarrow.parquet as pq

    validator = PathValidator(base_path)

    def read(filepath: str) -> LuaList:
        """Read Parquet file, returning list of dictionaries."""
        path = validator.validate(filepath)
        table = pq.read_table(path)
        return LuaList(table.to_pylist())

    def write(filepath: str, data: List[Dict]) -> None:
        """Write list of dictionaries to Parquet file."""
        path = validator.validate(filepath)

        # Convert Lua tables to Python
        data = _lua_to_python(data)

        # Ensure parent directory exists
        os.makedirs(os.path.dirname(path), exist_ok=True) if os.path.dirname(path) else None

        table = pa.Table.from_pylist(data)
        pq.write_table(table, path)

    return {"read": read, "write": write}


def create_safe_hdf5_library(base_path: str) -> Dict[str, Any]:
    """
    Create HDF5 file operations library.

    Args:
        base_path: Base directory for file operations.

    Returns:
        Dictionary of HDF5 operation functions.
    """
    import h5py
    import numpy as np

    validator = PathValidator(base_path)

    def read(filepath: str, dataset: str) -> List[Any]:
        """Read dataset from HDF5 file."""
        path = validator.validate(filepath)
        with h5py.File(path, "r") as f:
            return f[dataset][:].tolist()

    def write(filepath: str, dataset: str, data: List) -> None:
        """Write data to HDF5 dataset."""
        path = validator.validate(filepath)

        # Convert Lua tables to Python
        data = _lua_to_python(data)

        # Ensure parent directory exists
        os.makedirs(os.path.dirname(path), exist_ok=True) if os.path.dirname(path) else None

        with h5py.File(path, "a") as f:
            if dataset in f:
                del f[dataset]
            f.create_dataset(dataset, data=np.array(data))

    def list_datasets(filepath: str) -> List[str]:
        """List all datasets in HDF5 file."""
        path = validator.validate(filepath)
        datasets = []
        with h5py.File(path, "r") as f:

            def visitor(name, obj):
                if isinstance(obj, h5py.Dataset):
                    datasets.append(name)

            f.visititems(visitor)
        return datasets

    return {"read": read, "write": write, "list": list_datasets}


def create_safe_excel_library(base_path: str) -> Dict[str, Any]:
    """
    Create Excel file operations library.

    Args:
        base_path: Base directory for file operations.

    Returns:
        Dictionary of Excel operation functions.
    """
    from openpyxl import Workbook, load_workbook

    validator = PathValidator(base_path)

    def read(filepath: str, options: Optional[Dict] = None) -> LuaList:
        """Read Excel file, returning list of dictionaries."""
        path = validator.validate(filepath)

        # Convert Lua table options to Python dict if needed
        if options and hasattr(options, "items"):
            options = dict(options.items())
        options = options or {}
        sheet_name = options.get("sheet")

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return LuaList([])

        # First row is headers
        headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
        return LuaList([dict(zip(headers, row)) for row in rows[1:]])

    def write(filepath: str, data: List[Dict], options: Optional[Dict] = None) -> None:
        """Write list of dictionaries to Excel file."""
        path = validator.validate(filepath)

        # Convert Lua table options to Python dict if needed
        if options and hasattr(options, "items"):
            options = dict(options.items())
        options = options or {}
        sheet_name = options.get("sheet", "Sheet1")

        # Convert Lua tables to Python
        data = _lua_to_python(data)

        # Ensure parent directory exists
        os.makedirs(os.path.dirname(path), exist_ok=True) if os.path.dirname(path) else None

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        if data:
            headers = list(data[0].keys())
            ws.append(headers)
            for row in data:
                ws.append([row.get(h) for h in headers])

        wb.save(path)

    def sheets(filepath: str) -> List[str]:
        """List sheet names in Excel file."""
        path = validator.validate(filepath)
        wb = load_workbook(path, read_only=True)
        return wb.sheetnames

    return {"read": read, "write": write, "sheets": sheets}


def _lua_to_python(obj: Any) -> Any:
    """
    Recursively convert Lua table-like objects to Python dicts/lists.

    Args:
        obj: Object to convert (may be Lua table or Python object).

    Returns:
        Python dict, list, or original value.
    """
    # Check if it's a Lua table (has values() or items() method)
    if hasattr(obj, "items"):
        # Could be a dict-like Lua table
        try:
            items = list(obj.items())
            # Check if it's array-like (all integer keys starting from 1)
            if items and all(isinstance(k, (int, float)) for k, v in items):
                keys = [int(k) for k, v in items]
                if keys == list(range(1, len(keys) + 1)):
                    # It's an array-like table, convert to list
                    return [_lua_to_python(obj[k]) for k in range(1, len(keys) + 1)]
            # It's a dict-like table
            return {k: _lua_to_python(v) for k, v in items}
        except (TypeError, AttributeError):
            pass

    if hasattr(obj, "values") and not isinstance(obj, (dict, str)):
        try:
            return [_lua_to_python(v) for v in obj.values()]
        except (TypeError, AttributeError):
            pass

    if isinstance(obj, dict):
        return {k: _lua_to_python(v) for k, v in obj.items()}

    if isinstance(obj, list):
        return [_lua_to_python(item) for item in obj]

    return obj
