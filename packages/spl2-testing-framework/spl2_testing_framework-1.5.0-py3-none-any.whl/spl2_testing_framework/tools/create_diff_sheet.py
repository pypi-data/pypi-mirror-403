# Copyright 2026 Splunk Inc.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.


import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment
import json, os
from typing import Union, List, Dict
import logging

_LOGGER = logging.getLogger(__name__)


class ComparisonSheet:
    """Class to create a comparison Excel sheet between expected and actual data."""

    def __init__(self):
        self.actual_json_list = []
        self.expected_json_list = []
        self.expected_path = "comparison_box_test/expected.json"
        self.actual_path = "comparison_box_test/actual.json"
        self.output_path = "comparison_box_test/comparison.xlsx"
        os.makedirs("comparison_box_test", exist_ok=True)

    def create_comparison_sheet(self, output_result, box_test):
        """
        1. Creates a comparison_box_test directory.
        2. Then creates actual.json, expected.json and comparison.xlsx sheet in that folder.

        Parameters:
            output_result (list): The output result got from the execution.
            box_test (object): Box test object.

        """
        self.actual_json_list.append(output_result[0])
        self.expected_json_list.append(box_test.output[0])
        with open(self.expected_path, "w") as f:
            json.dump([r.__dict__ for r in self.expected_json_list], f, indent=2)
        with open(self.actual_path, "w") as f:
            json.dump([r.__dict__ for r in self.actual_json_list], f, indent=2)
        self._create_comparison_sheet()

    def _create_comparison_sheet(self):
        """
        Creates a comparison Excel sheet between expected and actual data.
        """

        _LOGGER.info("Starting comparison generation...")

        expected_list = self._load_json(self.expected_path)
        actual_list = self._load_json(self.actual_path)
        df = self._compare_events(expected_list, actual_list)
        df.to_excel(self.output_path, index=False)

        self._apply_excel_styles(self.output_path)

        _LOGGER.info(f"✅ Comparison sheet created at: {self.output_path}")

    def _load_json(self, json_input: Union[str, list]) -> List[Dict]:
        """Load JSON data from a file path or return list directly."""
        if isinstance(json_input, str):
            with open(json_input) as f:
                return json.load(f)
        return json_input

    def _compare_events(self, list1: List[Dict], list2: List[Dict]) -> pd.DataFrame:
        """Compare expected and actual events, returning a formatted DataFrame."""

        self.fields = []
        self.expected_vals = []
        self.actual_vals = []
        self.problems = []
        self.solutions = []

        for idx, (event_expected, event_actual) in enumerate(
            zip(list1, list2), start=1
        ):
            dict_expected = event_expected.get("data", {})
            dict_actual = event_actual.get("data", {})

            all_fields = sorted(set(dict_expected.keys()) | set(dict_actual.keys()))
            equal_fields = []

            for field in all_fields:
                exp_val = dict_expected.get(field)
                act_val = dict_actual.get(field)

                if exp_val is None and act_val is not None:
                    self._add_row(
                        field,
                        exp_val,
                        act_val,
                        problem="Extracted additionally from SPL2 pipeline",
                        solution="Possible to ignore",
                    )
                elif act_val is None and exp_val is not None:
                    self._add_row(
                        field,
                        exp_val,
                        act_val,
                        problem="Missing from SPL2 pipeline. Need to fix.",
                        solution="",
                    )
                elif exp_val != act_val:
                    self._add_row(
                        field,
                        exp_val,
                        act_val,
                        problem="Value mismatch",
                        solution="Check this in SPL2 pipeline and fix if needed.",
                    )
                else:
                    if field == "_raw":
                        self._add_row(field, exp_val, act_val)
                    else:
                        equal_fields.append(field)

            if equal_fields:
                self._add_row(
                    field="Equal fields for this event",
                    exp_val=", ".join(equal_fields),
                    act_val="",
                )

            self._add_row(
                field="All Expected Fields", exp_val=dict_expected, act_val=""
            )
            self._add_row(field="All Actual Fields", exp_val=dict_actual, act_val="")
            self._add_row(field="", exp_val="", act_val="")

        return pd.DataFrame(
            {
                "Field": self.fields,
                "Expected": self.expected_vals,
                "Actual": self.actual_vals,
                "Problem": self.problems,
                "Solution": self.solutions,
            }
        )

    def _add_row(self, field, exp_val, act_val, problem="", solution=""):
        """Append a single comparison row to the running lists."""
        self.fields.append(field)
        self.expected_vals.append(exp_val)
        self.actual_vals.append(act_val)
        self.problems.append(problem)
        self.solutions.append(solution)

    def _apply_excel_styles(self, output_path: str) -> None:
        """Apply Excel formatting and highlighting."""
        wb = load_workbook(output_path)
        ws = wb.active

        diff_fill = PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid"
        )
        missing_fill = PatternFill(
            start_color="FF9999", end_color="FF9999", fill_type="solid"
        )
        equal_fill = PatternFill(
            start_color="CCFFCC", end_color="CCFFCC", fill_type="solid"
        )

        thin_border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )

        alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        for r, row in enumerate(
            ws.iter_rows(
                min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column
            ),
            start=1,
        ):
            ws.row_dimensions[r].height = 15

            field_name = row[0].value
            expected_val = row[1].value if len(row) > 1 else None
            actual_val = row[2].value if len(row) > 2 else None

            for cell in row:
                cell.border = thin_border
                cell.alignment = alignment

            if r > 1:  # skip header row
                if expected_val is None and actual_val not in (None, ""):
                    row[1].fill = missing_fill
                elif actual_val is None and expected_val not in (None, ""):
                    row[2].fill = missing_fill
                elif expected_val != actual_val and field_name not in (
                    "",
                    "Equal fields for this event",
                    "All Expected Fields",
                    "All Actual Fields",
                ):
                    row[1].fill = diff_fill
                    row[2].fill = diff_fill

                if field_name in (
                    "All Expected Fields",
                    "All Actual Fields",
                    "Equal fields for this event",
                ):
                    for cell in row:
                        cell.fill = equal_fill

        for col in ["A", "B", "C", "D", "E"]:
            ws.column_dimensions[col].width = 50
        ws.freeze_panes = "B2"

        wb.save(output_path)
