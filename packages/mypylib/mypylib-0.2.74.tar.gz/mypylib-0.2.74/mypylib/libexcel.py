import multiprocessing
import ssl
import csv
import math
import datetime
import json
import os
import pathlib
import pickle
import random
import time
from datetime import timedelta, date

import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import requests
import talib
import pandas as pd

request_headers = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.95 Safari/537.36'}

opendays = []

tail_trim = {'0': 0,
             '1': -0.000001,
             '2': -0.000002,
             '3': -0.000003,
             '4': -0.000004,
             # '5': 0,
             '6': 0.000004,
             '7': 0.000003,
             '8': 0.000002,
             '9': 0.000001}


def parse_date_time(date_string, time_string):
    if time_string is not None and '.' in time_string:
        if '/' in date_string:
            timestamp = datetime.datetime.strptime(f'{date_string} {time_string}', '%Y/%m/%d %H:%M:%S.%f')
        else:
            timestamp = datetime.datetime.strptime(f'{date_string} {time_string}', '%Y-%m-%d %H:%M:%S.%f')
    else:
        if '/' in date_string:
            if time_string is None:
                timestamp = datetime.datetime.strptime(f'{date_string}', '%Y/%m/%d')
            else:
                timestamp = datetime.datetime.strptime(f'{date_string} {time_string}', '%Y/%m/%d %H:%M:%S')
        else:
            if time_string is None:
                timestamp = datetime.datetime.strptime(f'{date_string}', '%Y-%m-%d')
            else:
                timestamp = datetime.datetime.strptime(f'{date_string} {time_string}', '%Y-%m-%d %H:%M:%S')
    return timestamp


def date_range(start_date, end_date):
    for n in range(int((end_date - start_date).days)):
        yield start_date + timedelta(n)


def rename_column(org):
    x = org
    x = x.rstrip()
    x = x.strip()
    x = x.replace(' ', '')
    x = x.replace('(', '')
    x = x.replace(')', '')
    x = x.replace('%', '')
    # print(f'rename function: org: [{org}] [{x}]')
    return x


def str_to_float(x):
    return float(x.replace(',', ''))


def trim_price(x):
    # print(f'x={x}')
    try:
        c = x[-1]
        # print(f'c=[{c}] [{tail_trim[c]}]')
        xx = x.split('.')
        if len(xx) > 1 and len(x.split('.')[1]) == 6:
            # print(f'trim: {tail_trim[x[-1]]}')
            return float(x) + tail_trim[x[-1]]
        else:
            return float(x)
    except Exception as e:
        print(e, x, x[-1])


def get_day_trade_candidates(output_file_path, target_date):
    OTC_url_format = 'https://www.tpex.org.tw/web/stock/trading/' \
                     'intraday_trading/intraday_trading_list_print.php?' \
                     'l=zh-tw&d={}/{:02d}/{:02d}&stock_code=&s=0,asc,1'
    SEM_url_format = 'https://www.twse.com.tw/exchangeReport/' \
                     'TWTB4U?response=html&date={}{:02d}{:02d}&selectType=All'
    if isinstance(target_date, str):
        today = parse_date_time(target_date, None)
    elif isinstance(target_date, datetime.datetime):
        today = target_date
    SEM_url = SEM_url_format.format(today.year, today.month, today.day)
    print(SEM_url)
    table = pd.read_html(SEM_url)
    df = table[0]
    df.columns = df.columns.droplevel()
    if '證券代號' not in df.columns:
        df = table[1]
        df.columns = df.columns.droplevel()
    df['證券代號'] = df['證券代號'].astype('str')
    mask = df['證券代號'].str.len() == 4
    df = df.loc[mask]
    df = df[['證券代號', '證券名稱', '暫停現股賣出後現款買進當沖註記']]
    df['暫停現股賣出後現款買進當沖註記'] = df['暫停現股賣出後現款買進當沖註記'].apply(lambda x: False if x == 'Y' else True)
    OCT_url = OTC_url_format.format(today.year - 1911, today.month, today.day)
    print(OCT_url)
    table = pd.read_html(OCT_url)
    df1 = table[0]
    df1.columns = df1.columns.droplevel()
    df1['證券代號'] = df1['證券代號'].astype('str')
    mask = df1['證券代號'].str.len() == 4
    df1 = df1.loc[mask]
    df1 = df1[['證券代號', '證券名稱', '暫停現股賣出後現款買進當沖註記']]
    df1['暫停現股賣出後現款買進當沖註記'] = df1['暫停現股賣出後現款買進當沖註記'].apply(lambda x: False if x == '＊' else True)
    all_df = pd.concat([df, df1])
    all_df.rename({'證券代號': 'symbol'}, axis=1, inplace=True)
    all_df.rename({'證券名稱': 'name'}, axis=1, inplace=True)
    all_df.rename({'暫停現股賣出後現款買進當沖註記': 'DayTrade'}, axis=1, inplace=True)
    all_df = all_df.set_index('symbol')
    ret = all_df.to_dict('index')
    all_df.to_csv(output_file_path)
    return ret


def load_opendays():
    if len(opendays) != 0:
        return opendays

    today = datetime.datetime.today()
    sod = today - datetime.timedelta(days=1000)

    if os.name == 'nt':
        unix_sod_struct = today.timetuple()
        today_str = int(time.mktime(unix_sod_struct))
        unix_today_struct = sod.timetuple()
        sod_str = int(time.mktime(unix_today_struct))
    else:
        today_str = today.strftime('%s')
        sod_str = sod.strftime('%s')

    symbol = 2330
    url = f'https://ws.api.cnyes.com/ws/api/v1/charting/history?resolution=D&symbol=TWS:{symbol}:STOCK&from={today_str}&to={sod_str}'
    # print(url)
    r = requests.get(url, headers=request_headers)
    jdata = json.loads(r.content)
    # print(r.content)
    for t in jdata['data']['t']:
        opendays.append(datetime.datetime.utcfromtimestamp(t))
    return opendays


def find_last_number_day(date: datetime, last_number_day):
    load_opendays()

    bool_found = False
    index = 0
    for index, open_day in enumerate():
        if date == open_day:
            bool_found = True
            break

    if not bool_found:
        raise f'find_last_number_day: {date} not found'

    index -= last_number_day

    print(f'find_last_number_day {date}, {last_number_day}, {opendays[index]}')

    return opendays[index]


def calculate_call_percentage(buy, sell, fee=0.001425, tax=0.003):
    return (sell * (1 - fee) - buy * (1 + fee + tax)) / buy


def convert_to_str(input):
    if type(input) is str:
        return input.rstrip(' ').lstrip(' ').rstrip(' ')

    if type(input) is int:
        return str(input)

    return input


def load_cell(s, r, c):
    x = s.cell(row=r, column=c)
    x.value = convert_to_str(x.value)
    return x


def load_value(s: object, r: object, c: object) -> object:
    return load_cell(s, r, c).value


def set_cell(s, r, c, value, hyper_link=None, number_format=None):
    cell = s.cell(row=r, column=c)
    cell.value = value
    if hyper_link is not None:
        cell.hyperlink = hyper_link
    if number_format is not None:
        cell.number_format = number_format


def column_idx(s):
    ret = 0
    s = s.upper()
    for i in range(0, len(s)):
        ret = ret * 26 + ord(s[i]) - ord('A') + 1
    return ret


def conv(s):
    try:
        s = float(s)
    except ValueError:
        pass
    return s


def calculate_percentage(a, b):
    return round((b - a) * 100.0 / a, 2)


def get_history(symbol, start_of_day=datetime.datetime(2019, 11, 1), always_get_latest=False, skip_sod_check=False):
    """
    :param always_get_latest: always get the latest one
    :rtype: object
    :param symbol: 股票代號
    :param start_of_day: 從哪一天開始
    :return: (Time, Open, Close, High, Low, Volume,
        Volume MA3, MA3, slop 3,
        Volume MA5, MA3, slop 5,
        Volume MA7, MA3, slop 7,
        Volume MA20, MA20, slop 20,
        BBands Up, BBands middle20, BBands lower20, Bandwidth, level,
        MACD, SIGNAL, HIST)
    """
    adj_start_of_day = start_of_day - datetime.timedelta(days=40)

    if os.name == 'nt':
        # unix_sod_struct = (start_of_day - datetime.timedelta(days=200)).timetuple()
        unix_sod_struct = adj_start_of_day.timetuple()
        unix_sod = int(time.mktime(unix_sod_struct))
        unix_today_struct = datetime.date.today().timetuple()
        unix_today = int(time.mktime(unix_today_struct))
    else:
        # unix_sod = (adj_start_of_day - datetime.timedelta(days=200)).strftime("%s")
        unix_sod = adj_start_of_day.strftime('%s')
        # unix_today = datetime.date.today().strftime("%s")
        unix_today = datetime.datetime.now().strftime("%s")

    if len(symbol) < 4:
        print(f'padding symbol {symbol} to 00{symbol}')
        symbol = f'00{symbol}'

    if not os.path.isdir('history'):
        os.mkdir('history')
    filepath = f'history/{symbol}.csv'

    if always_get_latest:
        if os.path.isfile(filepath):
            st_mtime = datetime.datetime.fromtimestamp(pathlib.Path(filepath).stat().st_mtime)
            # 當有 always_get_latest 的時候，要下載最新的資料。
            # 如果 mtime 不是今天，表示是舊的資料，則砍掉重新下載。
            if st_mtime.date() != datetime.datetime.now().date():
                os.unlink(filepath)
            # 如果 mtime 是今天，但是時間卻是 15:00 以前下載的，也是一樣砍掉重新下載。因為今天盤後的資料已經有了。
            elif st_mtime.time() < datetime.time(hour=15, minute=0):
                pass
                # os.unlink(filepath)

    if os.path.isfile(filepath):
        with open(filepath, 'r') as f:
            history = []
            rows = csv.reader(f)
            for idx, r in enumerate(rows):
                if idx == 0:
                    Date = datetime.datetime.strptime(r[0], '%Y/%m/%d')
                    if not skip_sod_check and Date > adj_start_of_day:
                        break
                history.append([conv(s) for s in r])
        if len(history) == 0:
            os.unlink(filepath)
        else:
            return history

    # url = f'https://ws.api.cnyes.com/charting/api/v1/history?resolution=D&symbol=TWS:{symbol}:STOCK&from={unix_today}&to={unix_sod}&quote=1'
    # url = f'https://ws.api.cnyes.com/ws/api/v1/charting/history?symbol=TWS:{symbol}:STOCK&resolution=D&quote=1&from={unix_today}&to={unix_sod}'
    url = f'https://ws.api.cnyes.com/ws/api/v1/charting/history?resolution=D&symbol=TWS:{symbol}:STOCK&from={unix_today}&to={unix_sod}'

    print(url)
    r = requests.get(url, headers=request_headers)
    # print(r.content)
    data = json.loads(r.content)
    if data['statusCode'] != 200:
        print("HTTP request error!")
        return None

    history = []

    for t in data['data']['t']:
        # print(datetime.datetime.utcfromtimestamp(t).strftime('%Y-%m-%d'))
        pass

    T = data['data']['t']
    O = data['data']['o']
    H = data['data']['h']
    L = data['data']['l']
    C = data['data']['c']
    V = data['data']['v']

    closed = []
    volume = []

    for t, o, h, l, c, v in zip(T, O, H, L, C, V):
        t = datetime.datetime.utcfromtimestamp(t).strftime('%Y/%m/%d')
        history.insert(0, [t, o, h, l, c, v])
        # print(f'here: {history}')
        closed.insert(0, float(c))
        volume.insert(0, float(v))

    if len(closed) == 0:
        return None

    if True:
        closed = np.array(closed)
        volume = np.array(volume)
        upper20, middle20, lower20 = talib.BBANDS(closed, timeperiod=20, nbdevup=2, nbdevdn=2)

        mav3 = talib.MA(volume, timeperiod=3)
        ma3 = talib.MA(closed, timeperiod=3)

        mav5 = talib.MA(volume, timeperiod=5)
        ma5 = talib.MA(closed, timeperiod=5)

        mav7 = talib.MA(volume, timeperiod=7)
        ma7 = talib.MA(closed, timeperiod=7)

        mav20 = talib.MA(volume, timeperiod=20)
        ma20 = talib.MA(closed, timeperiod=20)

        macd, macdsignal, macdhist = talib.MACD(closed,
                                                fastperiod=12,
                                                slowperiod=26,
                                                signalperiod=9)

        for i in range(0, len(history)):
            # 3日：均量、均線、斜率
            history[i].append(round(mav3[i], 3))  # 3日均量
            history[i].append(round(ma3[i], 3))  # 3日均線
            if np.isnan(ma3[i - 1]):  # 3日斜率
                history[i].append(float('nan'))
            else:
                history[i].append(calculate_percentage(ma3[i - 1], ma3[i]))

            # 5日：均量、均線、斜率
            history[i].append(round(mav5[i], 3))  # 5日均量
            history[i].append(round(ma5[i], 3))  # 5日均線
            if np.isnan(ma5[i - 1]):  # 5日斜率
                history[i].append(float('nan'))
            else:
                history[i].append(calculate_percentage(ma5[i - 1], ma5[i]))

            # 7日：均量、均線、斜率
            history[i].append(round(mav7[i], 3))  # 7日均量
            history[i].append(round(ma7[i], 3))  # 7日均線
            if np.isnan(ma7[i - 1]):  # 7日斜率
                history[i].append(float('nan'))
            else:
                history[i].append(calculate_percentage(ma7[i - 1], ma7[i]))

            # 20日：均量、均線、斜率
            history[i].append(round(mav20[i], 3))  # 20日均量
            history[i].append(round(ma20[i], 3))  # 20日均線
            if np.isnan(ma20[i - 1]):  # 20日斜率
                history[i].append(float('nan'))
            else:
                history[i].append(calculate_percentage(ma20[i - 1], ma20[i]))

            history[i].append(round(upper20[i], 3))  # 布林上軌
            history[i].append(round(middle20[i], 3))  # 布林中線
            history[i].append(round(lower20[i], 3))  # 布林下軌

            #
            # 帶寬
            # 位階
            #
            if np.isnan(upper20[i]):
                history[i].append(float('nan'))
                history[i].append(float('nan'))
            else:
                bandwidth = round(calculate_percentage(lower20[i], upper20[i]))
                # print(f'lower20:{lower20[i]} upper20:{upper20[i]} bandwidth:{bandwidth}')
                history[i].append(bandwidth)

                close = history[i][4]
                unit = (upper20[i] - middle20[i]) / 10.0
                level = 0 if unit == 0 else round((close - middle20[i]) / unit)
                # print(f'close {close} unit {unit} upper20 {upper20[i]} middle20 {middle20[i]} level {level}')
                history[i].append(level)

            # MACD
            history[i].append(macd[i])
            history[i].append(macdsignal[i])
            history[i].append(macdhist[i])

    with open(filepath, 'w', newline='') as f:
        csv.writer(f).writerows(history)

    return get_history(symbol, start_of_day, always_get_latest=always_get_latest, skip_sod_check=True)


def build_date_policy(sheet, date_column, policy_column, start_row):
    for i in range(start_row + 1, sheet.max_row + 1):

        if date_column != 0:
            date = load_value(sheet, i, date_column)
            if date is not None:
                last_date = date
            else:
                set_cell(sheet, i, date_column, last_date)

        if policy_column != 0:
            policy = load_value(sheet, i, policy_column)
            if policy is not None:
                last_policy = policy
            else:
                set_cell(sheet, i, policy_column, last_policy)


def build_stock_price(sheet, date_column, symbol_column, start_row, then_price_column, now_price_column,
                      percentage_column):
    for i in range(start_row + 1, sheet.max_row + 1):
        symbol = load_value(sheet, i, symbol_column)
        if symbol is None:
            continue
        date: datetime.datetime = load_value(sheet, i, date_column)
        if date is None:
            continue
        if type(date) == str:
            if '/' in date:
                date = datetime.datetime.strptime(date, '%Y/%m/%d')
            elif '-' in date:
                date = datetime.datetime.strptime(date, '%Y-%m-%d')
            else:
                print(f'Unknown date format {date}')
                exit(0)
        date = datetime.datetime(year=date.year, month=date.month, day=date.day)

        print(f'date:{date} symbol:{symbol}')
        history = get_history(symbol, date, always_get_latest=True)

        if history is None:
            print(f'Cannot get data of {symbol}')
            continue

        if len(history) == 0:
            print(f'Length of data of {symbol} is zero')
            continue

        sod_found = False
        base_price = 0

        trade_on_open = True

        for idx, h in enumerate(history):
            # hdate = datetime.datetime.fromtimestamp(h[0])
            hdate = datetime.datetime.strptime(h[0], '%Y/%m/%d')
            # if (hdate.year, hdate.month, hdate.day) == (date.year, date.month, date.day):
            if sod_found is not True and hdate >= date:
                print('found the date:', date)
                sod_found = True

            if sod_found is not True:
                continue

            if base_price == 0:
                base_price = h[1] if trade_on_open else h[4]
                set_cell(sheet, i, then_price_column, base_price)

        now_price = h[4]

        set_cell(sheet, i, now_price_column, now_price)
        set_cell(sheet, i, percentage_column, 0 if base_price == 0 else (now_price - base_price) / base_price)


# Verified.
def build_stock_info(sheet, symbol_column, start_row, start_column):
    """
    :param sheet: 要填入的sheet
    :param symbol_column:股票編號欄
    :param start_row:從哪一列開始
    :param start_column:從哪一欄開始
    """
    wb_obj = openpyxl.load_workbook('名稱代號.xlsx')
    name_sheet = wb_obj['名稱代號']

    # name, 上市櫃, 產業, 50/100成分股, 有股期, 成長股
    stock_info = {}
    for i in range(2, name_sheet.max_row + 1):
        symbol = load_value(name_sheet, i, 1)
        if symbol is None:
            continue
        name = load_value(name_sheet, i, 2)
        market = load_value(name_sheet, i, 3)
        category = load_value(name_sheet, i, 4)
        attr = load_value(name_sheet, i, 5)
        future = load_value(name_sheet, i, 6)
        groth = load_value(name_sheet, i, 7)

        stock_info[symbol] = (name, market, category, attr, future, groth)

    wb_obj.close()

    last_date = ''
    last_policy = ''

    item_index = 1
    item_index_row = 0

    set_cell(sheet, start_row, start_column + 0, '名稱')
    set_cell(sheet, start_row, start_column + 1, '上市櫃')
    set_cell(sheet, start_row, start_column + 2, '產業')
    set_cell(sheet, start_row, start_column + 3, '指數股')
    set_cell(sheet, start_row, start_column + 4, '有股期')
    set_cell(sheet, start_row, start_column + 5, '自分類')

    for i in range(start_row + 1, sheet.max_row + 1):
        symbol = load_value(sheet, i, symbol_column)

        if symbol in stock_info.keys():
            stock = stock_info[symbol]
        else:
            continue

        # set_cell(sheet, i, column_idx('A'), i - start_row)
        set_cell(sheet, i, start_column + 0, stock[0])
        set_cell(sheet, i, start_column + 1, stock[1])
        set_cell(sheet, i, start_column + 2, stock[2])
        set_cell(sheet, i, start_column + 3, stock[3])
        set_cell(sheet, i, start_column + 4, stock[4])
        set_cell(sheet, i, start_column + 5, stock[5])

    return start_column + 6


def build_1mk_from_shioaji(sheet, date_column, symbol_column, start_row, start_column, days, seconds_array: list):

    for row in range(days):
        for j in range(len(seconds_array)):
            set_cell(sheet, start_row, start_column + len(seconds_array) * row + j, f'第{row}天: {seconds_array[j]}')

    for row in range(start_row + 1, sheet.max_row + 1):
        symbol = load_value(sheet, row, symbol_column)
        if symbol is None:
            continue
        date: datetime.datetime = load_value(sheet, row, date_column)
        if date is None:
            continue
        if type(date) == str:
            if '/' in date:
                date = datetime.datetime.strptime(date, '%Y/%m/%d')
            elif '-' in date:
                date = datetime.datetime.strptime(date, '%Y-%m-%d')
            else:
                print(f'Unknown date format {date}')
                exit(0)
        date = datetime.datetime(year=date.year, month=date.month, day=date.day)

        print(f'{date} {symbol}')

        str_date = date.strftime('%Y-%m-%d')

        path_shioaji_ticks = f'shioaji_ticks/{symbol}'
        if not os.path.isdir(path_shioaji_ticks):
            continue
        files = os.listdir(path_shioaji_ticks)
        files.sort()

        for file_index, file in enumerate(files):
            if file[0] != '2':
                continue
            if str_date == file.split('.')[0]:

                for k in range(days):

                    this_seconds_array = seconds_array.copy()
                    this_prices_array = []

                    path_day = f'shioaji_ticks/{symbol}/{files[file_index + k]}'
                    df = pd.read_csv(path_day)
                    d_tick = df.to_dict('index')

                    print(f'=> {path_day}')

                    try:
                        for idx in d_tick:

                            timestamp = datetime.datetime.strptime(d_tick[idx]['ts'], '%Y-%m-%d %H:%M:%S.%f')
                            delta_seconds = (timestamp - timestamp.replace(hour=9, minute=0, second=0, microsecond=0)).seconds

                            if len(this_seconds_array) != 0:
                                if delta_seconds > this_seconds_array[0]:
                                    this_prices_array.append(d_tick[idx]['close'])
                                    this_seconds_array.pop(0)
                            else:
                                break

                        print(this_prices_array)

                        for j in range(len(seconds_array)):
                            set_cell(sheet, row, start_column + len(seconds_array) * k + j, f'{this_prices_array[j]}')

                    except Exception:
                        continue
                break




# 這部分程式已經太舊了，因為是使用群益的一分K資料
def build_1mk_profit(sheet, date_column, symbol_column, start_row, start_column,
                     call=True,
                     tax=0.003,
                     fee=0.001425):
    for i in range(start_row + 1, sheet.max_row + 1):
        # for i in range(start_row + 1, 100):
        symbol = load_value(sheet, i, symbol_column)
        date = load_value(sheet, i, date_column)

        if symbol is None or date is None:
            continue

        pickle_file = f'kpickle/K_{symbol}.pickle'

        if os.path.isfile(pickle_file) is False:
            continue

        start_date = date
        print(f'\033[1;33m{symbol} {start_date}\033[0m')

        with open(pickle_file, 'rb') as f:
            rows = pickle.load(f)
            day_found = False
            last_day_close_price = 0
            days_found = 0
            price_reached = False
            open_high_second_day = False
            for r in rows:
                d = r[0]
                # print(f'here:{d}')

                if day_found is not True and d >= start_date:
                    print(f'{start_date} day found')
                    day_found = True
                    set_cell(sheet, i, start_column, 0)

                if day_found is True:
                    if d.hour == 9 and d.minute == 1:
                        days_found = days_found + 1
                        open_price = r[4]

                        if days_found == 2:  # 第二天
                            if open_price > last_day_close_price:
                                print(f'Open {open_price} > last day close price {last_day_close_price}')
                                open_high_second_day = True

                    if d.hour == 13 and d.minute == 30:
                        last_day_close_price = r[4]

                    if days_found == 2:  # the second day. We will simulate the trading
                        if not open_high_second_day:
                            continue

                        current_price = r[4]

                        if not price_reached and current_price <= (last_day_close_price + open_price) / 2:
                            set_cell(sheet, i, start_column + 0, last_day_close_price)
                            set_cell(sheet, i, start_column + 1, open_price)
                            set_cell(sheet, i, start_column + 2, (last_day_close_price + open_price) / 2)
                            set_cell(sheet, i, start_column + 3, 'Yes')
                            set_cell(sheet, i, start_column + 5, 0.0)
                            set_cell(sheet, i, start_column + 6, 0.0)
                            print(
                                f'price reached: Last: {last_day_close_price} Open: {open_price} half: {(last_day_close_price + open_price) / 2} current: {current_price}')
                            price_reached = True
                            base_price = current_price

                        if price_reached and open_price != 0:
                            profit_rate = (base_price - current_price) / open_price
                            # if profit_rate > 0.02:

                            current_profit_rate = load_value(sheet, i, start_column + 5)
                            if current_profit_rate is None:
                                current_profit_rate = 0
                            if profit_rate > current_profit_rate:
                                set_cell(sheet, i, start_column + 5, profit_rate)
                                set_cell(sheet, i, start_column + 6, current_price)
                                print(f'{profit_rate}')

                            if d.hour == 13 and d.minute == 30:
                                set_cell(sheet, i, start_column + 9, profit_rate)


# 這部分程式已經太舊了，因為是使用群益的一分K資料
def build_1mk_chart1(sheet, date_column, symbol_column, start_row,
                     dst_dir,
                     url_column=None,
                     base_url=None,
                     days=1):
    for i in range(start_row + 1, sheet.max_row + 1):
        symbol = load_value(sheet, i, symbol_column)
        date = load_value(sheet, i, date_column)

        if symbol is None or date is None:
            continue

        pickle_file = f'kpickle/K_{symbol}.pickle'

        if os.path.isfile(pickle_file) is False:
            continue

        start_date = date

        with open(pickle_file, 'rb') as f:
            rows = pickle.load(f)
            day_found = False
            days_found = 0
            first_k = False
            x = []
            y = []
            x_index = 0
            half_price = 0
            for r in rows:
                d = r[0]
                # print(f'here:{d}')

                if day_found is not True and d >= start_date:
                    print(f'{start_date} day found')
                    day_found = True

                if day_found is True:
                    if d.hour == 9 and d.minute == 1:
                        days_found = days_found + 1
                        if days_found > days:
                            break
                        if days_found == 2:
                            half_price = (r[4] + y[-5]) / 2

                    # x.append(d)
                    x.append(x_index)
                    x_index = x_index + 1
                    y.append(r[4])

                    if d.hour == 13 and d.minute == 30:
                        x.append(x_index)
                        x_index = x_index + 1
                        x.append(x_index)
                        x_index = x_index + 1
                        x.append(x_index)
                        x_index = x_index + 1
                        x.append(x_index)
                        x_index = x_index + 1
                        y.append(None)
                        y.append(None)
                        y.append(None)
                        y.append(None)

            xa = np.array(x)
            ya = np.array(y)
            plt.plot(xa, ya)
            plt.axvline(x=int(x_index / 2), color='red')
            x = [x[0], x[-1]]
            y = [half_price, half_price]
            xa = np.array(x)
            ya = np.array(y)
            plt.plot(xa, ya, color='red')

            # plt.plot(ya)
            # plt.show()

        filename = f'{dst_dir}/{symbol}-{start_date.year}-{start_date.month:02d}-{start_date.day:02d}.png'
        print(filename)
        plt.savefig(filename)
        plt.close()

        if url_column is not None:
            set_cell(sheet, i, url_column, filename, f'{base_url}/{filename}')


# 這部分程式已經太舊了，因為是使用群益的一分K資料
def build_1mk_chart(sheet, date_column, symbol_column, start_row, dst_dir, days=1):
    for i in range(start_row + 1, sheet.max_row + 1):
        symbol = load_value(sheet, i, symbol_column)
        date = load_value(sheet, i, date_column)

        if symbol is None or date is None:
            continue

        pickle_file = f'kpickle/K_{symbol}.pickle'

        if os.path.isfile(pickle_file) is False:
            continue

        start_date = date

        fig, axs = plt.subplots(2, 2)
        with open(pickle_file, 'rb') as f:
            rows = pickle.load(f)
            day_found = False
            days_found = 0
            first_k = False
            for r in rows:
                d = r[0]
                # print(f'here:{d}')

                if day_found is not True and d >= start_date:
                    print(f'{start_date} day found')
                    day_found = True

                if day_found is True:
                    if d.hour == 9 and d.minute == 1:
                        x = []
                        y = []
                        first_k = True
                        days_found = days_found + 1
                        if days_found > days:
                            break

                    minute_index = d.hour * 60 + d.minute
                    if 9 * 60 + 1 < minute_index <= 13 * 60 + 30:
                        if minute_index % 5 == 0 or minute_index % 10 == 0 or first_k is True:
                            x.append(d)
                            y.append(r[4])

                    if minute_index == 13 * 60 + 30:
                        xa = np.array(x)
                        ya = np.array(y)

                        if days_found == 1:
                            axs[0, 0].plot(xa, ya)
                            axs[0, 0].xaxis.set_visible(False)
                        if days_found == 2:
                            axs[0, 1].plot(xa, ya)
                            axs[0, 1].xaxis.set_visible(False)
                        if days_found == 3:
                            axs[1, 0].plot(xa, ya)
                            axs[1, 0].xaxis.set_visible(False)
                        if days_found == 4:
                            axs[1, 1].plot(xa, ya)
                            axs[1, 1].xaxis.set_visible(False)

        filename = f'{dst_dir}/{symbol}-{start_date.year}-{start_date.month:02d}-{start_date.day:02d}.png'
        print(filename)
        plt.savefig(filename)
        plt.close()


# 這部分程式已經太舊了，因為是使用群益的一分K資料
def build_1mk_data(sheet, date_column, symbol_column, start_row, dst_sheet, days=1):
    print(f'max_row: {sheet.max_row}')
    for i in range(start_row + 1, sheet.max_row + 1):
        # for i in range(start_row + 1, 100):
        symbol = load_value(sheet, i, symbol_column)
        date = load_value(sheet, i, date_column)

        if symbol is None or date is None:
            continue

        start_date = date
        # TODO: 要是真正交易日
        end_date = start_date + datetime.timedelta(days)

        print(f'\033[1;33m{symbol} {start_date}\033[0m')

        data_file = f'kpickle/K_{symbol}.pickle'

        current_column = 1

        set_cell(dst_sheet, i - start_row, current_column, symbol)
        current_column = current_column + 1

        if os.path.isfile(data_file) is False:
            continue
        with open(data_file, 'rb') as f:
            rows = pickle.load(f)
            day_found = False
            days_found = 0
            for r in rows:
                d = r[0]

                if False and day_found is True and d > end_date:
                    print('finished')
                    break

                if day_found is not True and d >= start_date:
                    print(f'{start_date} day found')
                    day_found = True

                if day_found is True:

                    first_k = False
                    if d.hour == 9 and d.minute == 1:
                        days_found = days_found + 1
                        if days_found > days:
                            break
                        current_column = current_column + 1
                        set_cell(dst_sheet, i - start_row, current_column, d)
                        current_column = current_column + 1
                        first_k = True

                    minute_index = d.hour * 60 + d.minute
                    if minute_index <= 13 * 60 + 30:
                        if minute_index % 5 == 0 or minute_index % 10 == 0 or first_k is True:
                            set_cell(dst_sheet, i - start_row, current_column, r[4])
                            current_column = current_column + 1


# 這部分程式已經太舊了，因為是使用群益的一分K資料
def build_1mk(sheet, date_column, symbol_column, start_row, start_column, minute_array, next_day=False):
    minute_title_set = False

    for i in range(start_row + 1, sheet.max_row + 1):
        symbol = load_value(sheet, i, symbol_column)
        date = load_value(sheet, i, date_column)

        print(f'\033[1;33m{symbol}\033[0m')

        if next_day is True:
            print(f'this day: {date}')
            date = datetime.datetime(year=date.year, month=date.month, day=date.day, hour=9, minute=1)
            date = date + datetime.timedelta(days=1)

        if symbol is None or date is None:
            continue

        pickle_file = f'kpickle/K_{symbol}.pickle'

        with open(pickle_file, 'rb') as f:
            rows = pickle.load(f)

            column = start_column
            day_found = False
            base_price = 0
            call_out = False
            put_out = False
            for r in rows:
                d = r[0]
                # print(f'here:{d}')

                if day_found is True:
                    # print(f'{d.hour}:{d.minute}')
                    if (d.year, d.month, d.day) != (date.year, date.month, date.day) or (
                            d.hour == 13 and d.minute > 30):
                        print('finished')
                        break

                # if day_found is not True and (d.year, d.month, d.day) == (date.year, date.month, date.day):
                if day_found is not True and d >= date:
                    date = d
                    print(f'{d} day found')
                    day_found = True

                if day_found is True:

                    if minute_title_set is False:
                        set_cell(sheet, start_row, column, f'{d.hour}:{d.minute:02d}')

                    if d.minute in minute_array:
                        ret = r[3]
                        # print(f'{r[3]} {ret}')
                        set_cell(sheet, i, column, ret)
                        column = column + 1

        minute_title_set = True


def build_simulation(sheet,
                     date_column,
                     symbol_column,
                     name_column,
                     industry_column,
                     future_column,
                     market_column,
                     slop3_column,
                     slop5_column,
                     slop7_column,
                     slop20_column,
                     width_column,
                     level_column,
                     open_high_column,
                     open_highest_column,
                     open_low_column,
                     red_k_column,
                     percentage_column,
                     start_row,
                     cate_column,
                     category,
                     base_price_column,
                     profit_column,
                     next_day_OHLC_column,
                     dealer_column,
                     limit_up_touch_column,
                     limit_up_close_column,
                     next_day_column,
                     number_targets_to_trade,
                     seconds_to_percentage_column,
                     days_of_cost,
                     price_limitation=[],
                     future_only=0,
                     exclusive_industry=[],
                     exclusive_market=[],
                     slop_filter=lambda slop3, slop5, slop7, slop20: True,
                     width_filter=lambda x: True,
                     level_filter=lambda x: True,
                     iterations=1,
                     top_ones=False,
                     fix_base_price=False,
                     open_high_only=False,
                     open_highest_only=False,
                     open_low_only=False,
                     trade_on_red_k_only=False,
                     percentage_filter=lambda x: True,
                     next_day_OHLC_filter=lambda Open, High, Low, Close: True,
                     profit_adjustment=lambda profit: profit,
                     dealer_filter=lambda buy1, sell1, net1, buy2, sell2, net2, buy3, sell3, net3: True,
                     bool_sort_by_seconds=False,
                     check_day_trade=True):
    all_data = {}

    # load all data
    for i in range(start_row + 1, sheet.max_row + 1):
        symbol = load_value(sheet, i, symbol_column)
        if symbol is None:
            continue
        name = load_value(sheet, i, name_column)

        future = load_value(sheet, i, future_column)
        print(f'{symbol} {name} {future}')

        cate = load_value(sheet, i, cate_column)
        if cate != category:
            print(f'{symbol} {name} 跳過種類 {cate} in {category}')
            continue

        market = load_value(sheet, i, market_column)
        if market in exclusive_market:
            print(f'{symbol} {name} 跳過排除的市場 {market} in {exclusive_market}')
            continue

        industry = load_value(sheet, i, industry_column)
        if industry in exclusive_industry:
            print(f'{symbol} {name} 跳過排除的產業 {industry} in {exclusive_industry}')
            continue

        if future_only == 1:
            if future != '有股期':
                print(f'{symbol} {name} 跳過沒有股期的')
                continue
        elif future_only == -1:
            if future == '有股期':
                print(f'{symbol} {name} 跳過有股期的')
                continue
        else:
            pass

        next_day_open = load_value(sheet, i, next_day_OHLC_column + 0)
        next_day_high = load_value(sheet, i, next_day_OHLC_column + 1)
        next_day_low = load_value(sheet, i, next_day_OHLC_column + 2)
        next_day_close = load_value(sheet, i, next_day_OHLC_column + 3)

        if next_day_open is None:
            print(f'{symbol} {name} 跳過沒有隔日開盤價的')
            continue

        next_day_open = float(next_day_open)
        next_day_high = float(next_day_high)
        next_day_low = float(next_day_low)
        next_day_close = float(next_day_close)

        if not next_day_OHLC_filter(next_day_open, next_day_high, next_day_low, next_day_close):
            print(f'{symbol} {name} 跳過 next_day_OHLC_filter 不合格的 ')
            continue

        slop3 = load_value(sheet, i, slop3_column)
        slop3 = -9999 if slop3 is None else float(slop3)

        slop5 = load_value(sheet, i, slop5_column)
        slop5 = -9999 if slop5 is None else float(slop5)

        slop7 = load_value(sheet, i, slop7_column)
        slop7 = -9999 if slop7 is None else float(slop7)

        slop20 = load_value(sheet, i, slop20_column)
        slop20 = -9999 if slop20 is None else float(slop20)

        if not slop_filter(slop3, slop5, slop7, slop20):
            print(f'{symbol} {name} 跳過 slop_filter 不合格的')
            continue

        width = load_value(sheet, i, width_column)
        width = -9999 if width is None else float(width)
        if not width_filter(width):
            print(f'{symbol} {name} 跳過 width_filter 不合格的')
            continue

        level = load_value(sheet, i, level_column)
        level = -9999 if level is None else float(level)
        if not level_filter(level):
            print(f'{symbol} {name} 跳過 level_filter 不合格的')
            continue

        open_high = True if load_value(sheet, i, open_high_column) == 'Yes' else False

        open_highest = True if load_value(sheet, i, open_highest_column) == 'Yes' else False

        open_low = True if load_value(sheet, i, open_low_column) == 'Yes' else False

        red_k = True if load_value(sheet, i, red_k_column) == 'Yes' else False

        if True:
            buy1 = load_value(sheet, i, dealer_column + 0)
            sell1 = load_value(sheet, i, dealer_column + 1)
            net1 = load_value(sheet, i, dealer_column + 2)
            buy2 = load_value(sheet, i, dealer_column + 3)
            sell2 = load_value(sheet, i, dealer_column + 4)
            net2 = load_value(sheet, i, dealer_column + 5)
            buy3 = load_value(sheet, i, dealer_column + 6)
            sell3 = load_value(sheet, i, dealer_column + 7)
            net3 = load_value(sheet, i, dealer_column + 8)

            # print(f'{symbol} {name} {buy1} {sell1} {net1} {buy2} {sell2} {net2} {buy3} {sell3} {net3}')
            if buy1 is not None and not dealer_filter(buy1, sell1, net1, buy2, sell2, net2, buy3, sell3, net3):
                print(f'{symbol} {name} 跳過 dealer_filter 不合格的')
                continue

        limit_up_touch = True if load_value(sheet, i, limit_up_touch_column) == 'Yes' else False
        limit_up_close = True if load_value(sheet, i, limit_up_close_column) == 'Yes' else False

        percentage = load_value(sheet, i, percentage_column)
        if percentage is None:
            print(f'{symbol} {name} 跳過沒有 percentage 的')
            continue
        percentage = float(percentage)

        if not percentage_filter(percentage):
            print(f'{symbol} {name} 跳過 percentage_filter 不合格的')
            continue

        base_price = load_value(sheet, i, base_price_column)
        if base_price is None:
            print(f'{symbol} {name} 跳過沒有 base_price 的')
            continue
        base_price = float(base_price)

        if len(price_limitation) != 0:
            skip = True
            for x in price_limitation:
                if x[0] <= base_price <= x[1]:
                    skip = False
                    break
            if skip:
                print(f'{symbol} {name} 超過價格定義的 {price_limitation}')
                continue

        if open_high_only and not open_high:
            print(f'{symbol} {name} 跳過沒有開高股票')
            continue

        if open_highest_only and not open_highest:
            print(f'{symbol} {name} 跳過沒有高於昨日最高股票')
            continue

        if open_low_only and not open_low:
            print(f'{symbol} {name} 跳過沒有低於昨日收盤股票')
            continue

        if trade_on_red_k_only and not red_k:
            print(f'{symbol} {name} 跳過昨天不是紅K的股票')
            continue

        seconds_to_percentage = load_value(sheet, i, seconds_to_percentage_column)
        if seconds_to_percentage is not None:
            seconds_to_percentage = float(seconds_to_percentage)
        else:
            seconds_to_percentage = 86400

        if bool_sort_by_seconds and seconds_to_percentage == 86400:
            print(f'{symbol} {name} 跳過沒有達到某個 % 數的股票')
            continue

        profit = load_value(sheet, i, profit_column)

        if profit is None:
            print(f'{symbol} {name} 跳過沒有 profit 的')
            continue

        profit = profit_adjustment(profit)

        date = load_value(sheet, i, date_column)

        if date not in all_data.keys():
            all_data[date] = []

            if check_day_trade:

                next_day = load_value(sheet, i, next_day_column)
                next_day = parse_date_time(next_day, None)

                print(f'\033[1;33mNext day: {next_day}\033[0m')
                day_trade_file = f'day_trade/{next_day.strftime("%Y-%m-%d")}.csv'
                if not os.path.isfile(day_trade_file):
                    get_day_trade_candidates(day_trade_file, next_day)
                    time.sleep(3)
                df = pd.read_csv(day_trade_file, dtype=str)
                df['symbol'] = df['symbol'].astype('str')
                df = df.set_index('symbol')
                day_trade_dict = df.to_dict('index')

        if check_day_trade:
            if symbol not in day_trade_dict.keys():
                continue
            if day_trade_dict[symbol]['DayTrade'] == 'False':
                print(f'{symbol} {name} 跳過無法當沖標地')
                continue

        all_data[date].append((date,
                               symbol,
                               name,
                               base_price,
                               profit,
                               limit_up_touch,
                               limit_up_close,
                               seconds_to_percentage))

    if len(all_data) == 0:
        return []

    # do simulation
    trade_records = []
    for i in range(iterations):

        total_profit = 0

        profit_trade_times = 0
        total_trade_times = 0

        max_cost = 0
        max_limit_up_touched_times = 0
        trade_record = []
        cost_history = []

        for date in all_data:

            if bool_sort_by_seconds:
                # print(all_data[date])
                all_data[date].sort(key=lambda x: x[-1])
                # print(all_data[date])

            items = len(all_data[date])
            if items == 0:
                continue

            targets = random.sample(range(items), items if items < number_targets_to_trade else number_targets_to_trade)

            if top_ones:
                targets = list(range(0, len(targets)))

            print(f'{date} {targets}')

            cost = 0
            limit_up_close_times = 0
            for x in targets:
                array = all_data[date]

                (this_date, symbol, name, base_price, profit, limit_up_touch, limit_up_close, seconds_to_percentage) = array[x]

                if limit_up_close:
                    limit_up_close_times += 1

                if fix_base_price:
                    base_price = 2000

                print(f'\t{symbol} {base_price} {profit:.2f} {base_price * profit * 1000:.0f}')

                cost = cost + base_price * 1000

                this_profit = base_price * profit * 1000

                total_profit = total_profit + this_profit

                total_trade_times = total_trade_times + 1
                if profit > 0:
                    profit_trade_times = profit_trade_times + 1

                trade_record.append((this_date, symbol, name, base_price, profit, int(this_profit), int(total_profit),
                                     limit_up_touch, limit_up_close, seconds_to_percentage))

            cost_history.insert(0, cost)

            if len(cost_history) > days_of_cost:
                cost_history.pop()

            current_cost = sum(cost_history)
            max_cost = current_cost if current_cost > max_cost else max_cost
            max_limit_up_touched_times = limit_up_close_times if limit_up_close_times > max_limit_up_touched_times else max_limit_up_touched_times

        print(f'{date} {targets} {profit:.0f} {total_profit:.0f} {max_cost} {total_profit / max_cost:.1f}')
        print(f'Win Rate: {profit_trade_times / total_trade_times * 100.0:.2f}%')

        # 最前面那個是總結 summary
        trade_record.insert(0, (
            total_profit, total_trade_times, profit_trade_times, profit_trade_times / total_trade_times, max_cost,
            total_profit / max_cost, max_limit_up_touched_times))

        trade_records.append(trade_record)

    # print(trade_records)
    return trade_records


def build_simulation_total_result(sheet, trade_records):
    if len(trade_records) == 0:
        return
    for j in range(1, len(trade_records[0])):
        set_cell(sheet, 1 + j, 1, trade_records[0][j][0])

    for i in range(len(trade_records)):
        trade_record = trade_records[i]
        for j in range(1, len(trade_record)):
            set_cell(sheet, 1 + j, 2 + i, trade_record[j][6])


def build_simulation_result(sheet, trade_records):
    column_index = 1
    row = 1
    set_cell(sheet, row, column_index + 0, '總獲利')
    set_cell(sheet, row, column_index + 1, '總交易次數')
    set_cell(sheet, row, column_index + 2, '獲利次數')
    set_cell(sheet, row, column_index + 3, '勝率')
    set_cell(sheet, row, column_index + 4, '成本')
    set_cell(sheet, row, column_index + 5, '報酬率')
    set_cell(sheet, row, column_index + 6, '單日最多漲停家數')

    row = row + 1

    for i in range(len(trade_records)):
        trade_record = trade_records[i]

        set_cell(sheet, row, column_index + 0, trade_record[0][0])
        set_cell(sheet, row, column_index + 1, trade_record[0][1])
        set_cell(sheet, row, column_index + 2, trade_record[0][2])
        set_cell(sheet, row, column_index + 3, trade_record[0][3])
        set_cell(sheet, row, column_index + 4, trade_record[0][4])
        set_cell(sheet, row, column_index + 5, trade_record[0][5])
        set_cell(sheet, row, column_index + 6, trade_record[0][6])

        row = row + 1

    column_index = 10
    for i in range(len(trade_records)):
        row = 1

        set_cell(sheet, row, column_index + 0, '總獲利')
        set_cell(sheet, row, column_index + 1, '總交易次數')
        set_cell(sheet, row, column_index + 2, '獲利次數')
        set_cell(sheet, row, column_index + 3, '勝率')
        set_cell(sheet, row, column_index + 4, '成本')
        set_cell(sheet, row, column_index + 5, '報酬率')
        set_cell(sheet, row, column_index + 6, '單日最多漲停家數')

        row = row + 1

        trade_record = trade_records[i]

        set_cell(sheet, row, column_index + 0, trade_record[0][0])
        set_cell(sheet, row, column_index + 1, trade_record[0][1])
        set_cell(sheet, row, column_index + 2, trade_record[0][2])
        set_cell(sheet, row, column_index + 3, trade_record[0][3])
        set_cell(sheet, row, column_index + 4, trade_record[0][4])
        set_cell(sheet, row, column_index + 5, trade_record[0][5])
        set_cell(sheet, row, column_index + 5, trade_record[0][6])

        row = row + 2

        set_cell(sheet, row, column_index + 0, '日期')
        set_cell(sheet, row, column_index + 1, '代號')
        set_cell(sheet, row, column_index + 2, '名稱')
        set_cell(sheet, row, column_index + 3, '價格')
        set_cell(sheet, row, column_index + 4, '獲利率')
        set_cell(sheet, row, column_index + 5, '獲利')
        set_cell(sheet, row, column_index + 6, '累積獲利')
        set_cell(sheet, row, column_index + 7, '觸漲停')
        set_cell(sheet, row, column_index + 8, '收漲停')
        set_cell(sheet, row, column_index + 9, '秒數')

        row = row + 1

        for j in range(1, len(trade_record)):
            set_cell(sheet, row, column_index + 0, trade_record[j][0])
            set_cell(sheet, row, column_index + 1, trade_record[j][1])
            set_cell(sheet, row, column_index + 2, trade_record[j][2])
            set_cell(sheet, row, column_index + 3, trade_record[j][3])
            set_cell(sheet, row, column_index + 4, trade_record[j][4])
            set_cell(sheet, row, column_index + 5, trade_record[j][5])
            set_cell(sheet, row, column_index + 6, trade_record[j][6])
            set_cell(sheet, row, column_index + 7, trade_record[j][7])
            set_cell(sheet, row, column_index + 8, trade_record[j][8])
            set_cell(sheet, row, column_index + 9, trade_record[j][9])

            row = row + 1

        column_index = column_index + 11


def build_days(sheet, date_column, symbol_column, start_row, start_column,
               days_array, gap=0,
               call=True,
               build_profit=True,
               trade_on_open=True,  # False or trade on close
               next_day=False,
               last_day=False,
               tax=0.003,
               fee=0.001425):
    last_day_column = 0
    current_column = start_column
    set_cell(sheet, start_row, current_column, '基準價')
    current_column = current_column + 1
    if call is True:
        dir_str = '多'
    else:
        dir_str = '空'

    for d in days_array:
        set_cell(sheet, start_row, current_column, f'{d} 日 {dir_str}')
        current_column = current_column + 1 + gap

    if last_day is True:
        set_cell(sheet, start_row, current_column, f'最近日 {dir_str}')
        last_day_column = current_column

    set_cell(sheet, start_row, last_day_column + 1, '月線斜率')
    set_cell(sheet, start_row, last_day_column + 2, '帶寬')
    set_cell(sheet, start_row, last_day_column + 3, '位階')
    set_cell(sheet, start_row, last_day_column + 4, '高於收盤')
    set_cell(sheet, start_row, last_day_column + 5, '高於前日最高')
    set_cell(sheet, start_row, last_day_column + 6, '低於收盤')
    set_cell(sheet, start_row, last_day_column + 7, '低於前日最低')
    set_cell(sheet, start_row, last_day_column + 8, '紅K')
    set_cell(sheet, start_row, last_day_column + 9, '漲跌幅')
    set_cell(sheet, start_row, last_day_column + 10, '三日均線斜率')
    set_cell(sheet, start_row, last_day_column + 11, '五日均線斜率')
    set_cell(sheet, start_row, last_day_column + 12, '七日均線斜率')
    set_cell(sheet, start_row, last_day_column + 13, '廿日均線斜率')
    set_cell(sheet, start_row, last_day_column + 14, '次日碰漲停')
    set_cell(sheet, start_row, last_day_column + 15, '次日收漲停')
    set_cell(sheet, start_row, last_day_column + 16, '次交易日日期')

    for i in range(start_row + 1, sheet.max_row + 1):
        symbol = load_value(sheet, i, symbol_column)
        if symbol is None:
            continue
        date: datetime.datetime = load_value(sheet, i, date_column)
        if date is None:
            continue
        if type(date) == str:
            if '/' in date:
                date = datetime.datetime.strptime(date, '%Y/%m/%d')
            elif '-' in date:
                date = datetime.datetime.strptime(date, '%Y-%m-%d')
            else:
                print(f'Unknown date format {date}')
                exit(0)

        date = datetime.datetime(year=date.year, month=date.month, day=date.day)

        print(f'date:{date} symbol:{symbol}')
        history = get_history(symbol, date - datetime.timedelta(days=50), always_get_latest=True)

        if history is None:
            print(f'Cannot get data of {symbol}')
            continue

        if len(history) == 0:
            print(f'Length of data of {symbol} is zero')
            continue

        sod_found = False
        day_index = 0
        current_column = start_column
        base_price = 0

        if next_day is True:
            date = date + datetime.timedelta(days=1)

        limit_up = 0
        limit_down = 0
        for idx, h in enumerate(history):
            # hdate = datetime.datetime.fromtimestamp(h[0])
            hdate = datetime.datetime.strptime(h[0], '%Y/%m/%d')
            # if (hdate.year, hdate.month, hdate.day) == (date.year, date.month, date.day):
            if sod_found is not True and hdate >= date:
                print('found the date:', date)
                sod_found = True

            if sod_found is not True:
                continue

            #
            # 某次更新之後，get_history() 的傳回值變成這樣
            # 為了怕未來更新影響到相容性，列在這邊，如果有出錯誤就知道有改動到了
            #
            (Date, Open, High, Low, Close, Volume,
             VMA3, MA3, MA3_slop,
             VMA5, MA5, MA5_slop,
             VMA7, MA7, MA7_slop,
             VMA20, MA20, MA20_slop,
             BBand_up, BBand_middle, BBand_low, Bandwidth, Level,
             MACD, SIGNAL, HIST) = h

            if base_price == 0:
                base_price = h[1] if trade_on_open else h[4]
                set_cell(sheet, i, current_column, base_price)
                current_column = current_column + 1

                # 這部分比較tricky
                # h[-6] h[-2] h[-1] 分別是 月線斜率、帶寬、位階
                # base_price == 0 表示是下單日期的當天
                # 但是 月線斜率、帶寬、位階 必須使用前一天的以作判斷，不能用這一天的資料判斷，不然就是使用未來的資料。
                # 以下的寫法是錯誤的
                # set_cell(sheet, i, last_day_column + 1, h[-6])
                # set_cell(sheet, i, last_day_column + 2, h[-2])
                # set_cell(sheet, i, last_day_column + 3, h[-1])
                set_cell(sheet, i, last_day_column + 1, history[idx - 1][-6])
                set_cell(sheet, i, last_day_column + 2, history[idx - 1][-2])
                set_cell(sheet, i, last_day_column + 3, history[idx - 1][-1])

                # h[0]: Date
                # h[1]: Open
                # h[2]: High
                # h[3]: Low
                # h[4]: Close

                # 4: 高於收盤
                set_cell(sheet, i, last_day_column + 4, 'Yes' if base_price > history[idx - 1][4] else 'No')
                # 5: 高於前日最高
                # set_cell(sheet, i, last_day_column + 5, 'Yes' if base_price > history[idx - 1][2] else 'No')
                set_cell(sheet, i, last_day_column + 5, (base_price - history[idx - 1][2]) / base_price)
                # 6: 低於收盤
                set_cell(sheet, i, last_day_column + 6, 'Yes' if base_price < history[idx - 1][4] else 'No')
                # 7: 低於前日最低
                set_cell(sheet, i, last_day_column + 7, 'Yes' if base_price < history[idx - 1][3] else 'No')
                # 8: 紅K
                set_cell(sheet, i, last_day_column + 8, 'Yes' if history[idx - 1][4] > history[idx - 1][1] else 'No')
                # 9: 漲跌幅
                set_cell(sheet, i, last_day_column + 9, (history[idx - 1][4] - history[idx - 1][1]) / history[idx - 1][1])
                # 10: 三日均線斜率
                set_cell(sheet, i, last_day_column + 10, history[idx - 1][8])
                # 11: 五日均線斜率
                set_cell(sheet, i, last_day_column + 11, history[idx - 1][11])
                # 12: 七日均線斜率
                set_cell(sheet, i, last_day_column + 12, history[idx - 1][14])
                # 13: 廿日均線斜率
                set_cell(sheet, i, last_day_column + 13, history[idx - 1][17])

                yesterday_close = history[idx - 1][4]

                limit_up, limit_down = get_limit_up_and_down_price(yesterday_close)

                print(f'{hdate} {symbol} yesterday close: {yesterday_close} today high: {High} close: {Close} limit up: {limit_up}')
                if High >= limit_up:
                    set_cell(sheet, i, last_day_column + 14, 'Yes')
                else:
                    set_cell(sheet, i, last_day_column + 14, 'No')

                if Close >= limit_up:
                    set_cell(sheet, i, last_day_column + 15, 'Yes')
                else:
                    set_cell(sheet, i, last_day_column + 15, 'No')

                set_cell(sheet, i, last_day_column + 16, Date)

            if base_price == 0:
                continue

            day_index = day_index + 1
            if day_index not in days_array:
                continue

            if build_profit is True:
                if False:
                    if call is True:
                        profit_rate = (-base_price * (1 + fee) + h[4] * (1 - fee - tax)) / base_price
                    else:
                        profit_rate = (base_price * (1 - fee - tax) - h[4] * (1 + fee)) / base_price
                else:
                    if call is True:
                        # 作多：如果碰到漲停就出場 (High >= limit_up)
                        sell_price = limit_up if h[2] >= limit_up else h[4]
                        profit_rate = (-base_price * (1 + fee) + sell_price * (1 - fee - tax)) / base_price
                    else:
                        # 做空：
                        # 如果碰到跌停，那就出場
                        # 如果接近漲停 -2 * ticks，停損
                        #
                        if h[3] <= limit_down:
                            buy_price = h[3]
                        elif h[2] >= limit_up:
                            buy_price = limit_up - 2 * get_current_price_tick(limit_up)
                        else:
                            buy_price = h[4]

                        profit_rate = (base_price * (1 - fee - tax) - buy_price * (1 + fee)) / base_price
                set_cell(sheet, i, current_column, profit_rate, number_format=openpyxl.styles.numbers.FORMAT_PERCENTAGE_00)
            else:
                set_cell(sheet, i, current_column, h[4])

            current_column = current_column + 1 + gap

        if last_day is True:
            print(f'last day: {h}')
            if build_profit is True and base_price != 0:
                if call is True:
                    profit_rate = (-base_price * (1 + fee) + h[4] * (1 - fee - tax)) / base_price
                else:
                    profit_rate = (base_price * (1 - fee - tax) - h[4] * (1 + fee)) / base_price
                set_cell(sheet, i, last_day_column, profit_rate,
                         number_format=openpyxl.styles.numbers.FORMAT_PERCENTAGE_00)
            else:
                set_cell(sheet, i, last_day_column, h[4])


def build_ohlc(sheet, date_column, symbol_column, start_row, start_column, days, days_before=0, bbands=True):
    for i in range(start_row + 1, sheet.max_row + 1):
        symbol = load_value(sheet, i, symbol_column)
        if symbol is None:
            continue
        date: datetime.datetime = load_value(sheet, i, date_column)
        if date is None:
            continue
        if type(date) == str:
            if '/' in date:
                date = datetime.datetime.strptime(date, '%Y/%m/%d')
            elif '-' in date:
                date = datetime.datetime.strptime(date, '%Y-%m-%d')
            else:
                print(f'Unknown date format {date}')
                exit(0)
        date = datetime.datetime(year=date.year, month=date.month, day=date.day)

        if days_before != 0:
            date = find_last_number_day(date, days_before)

        history = get_history(symbol, date, always_get_latest=True)

        if history is None:
            print(f'Cannot get data of {symbol}')
            continue

        if len(history) == 0:
            print(f'Length of data of {symbol} is zero')
            continue

        sod_found = False
        day_index = -days_before
        current_column = start_column
        title_written = False

        for h in history:

            (Date, Open, High, Low, Close, Volume,
             VMA3, MA3, MA3_slop,
             VMA5, MA5, MA5_slop,
             VMA7, MA7, MA7_slop,
             VMA20, MA20, MA20_slop,
             BBand_up, BBand_middle, BBand_low, Bandwidth, Level,
             MACD, SIGNAL, HIST) = h

            # print(h, type(h))
            hdate = datetime.datetime.strptime(h[0], '%Y/%m/%d')
            if (hdate.year, hdate.month, hdate.day) == (date.year, date.month, date.day):
                print('found the date:', date)
                sod_found = True

            if sod_found is not True:
                continue

            if day_index >= (days + days_before):
                break

            set_cell(sheet, start_row, current_column + 0, f'開 {day_index}')
            set_cell(sheet, start_row, current_column + 1, f'高 {day_index}')
            set_cell(sheet, start_row, current_column + 2, f'低 {day_index}')
            set_cell(sheet, start_row, current_column + 3, f'收 {day_index}')
            set_cell(sheet, start_row, current_column + 4, f'量 {day_index}')
            if bbands is True:
                set_cell(sheet, start_row, current_column + 5, f'均量 {day_index}')
                set_cell(sheet, start_row, current_column + 6, f'MA斜率 {day_index}')
                set_cell(sheet, start_row, current_column + 7, f'帶寬 {day_index}')
                set_cell(sheet, start_row, current_column + 8, f'位階 {day_index}')

            set_cell(sheet, i, current_column + 0, Open)
            set_cell(sheet, i, current_column + 1, High)
            set_cell(sheet, i, current_column + 2, Low)
            set_cell(sheet, i, current_column + 3, Close)
            set_cell(sheet, i, current_column + 4, Volume)
            if bbands is True:
                set_cell(sheet, i, current_column + 5, VMA20)
                set_cell(sheet, i, current_column + 6, MA20_slop)
                set_cell(sheet, i, current_column + 7, Bandwidth)
                set_cell(sheet, i, current_column + 8, Level)
                current_column = current_column + 4

            current_column = current_column + 5
            day_index = day_index + 1


def build_column_ranking_each_day(sheet, date_column, ranking_column, start_row):
    last_day = datetime.datetime(1970, 1, 1)
    ranking = 1
    for i in range(start_row + 1, sheet.max_row + 1):
        date = load_value(sheet, i, date_column)
        if date is None:
            continue

        if last_day != date:
            ranking = 1
            last_day = date

        set_cell(sheet, i, ranking_column, ranking)
        ranking = ranking + 1


def build_dealer_downloader(date=datetime.datetime(year=2020, month=1, day=1), target_directory='自營商歷史資料'):
    if not os.path.isdir(target_directory):
        os.mkdir(target_directory)

    day_delta = datetime.timedelta(days=1)
    while date < datetime.datetime.now():
        print(date)

        # time.sleep(30)

        date = date + day_delta

        #
        # 上櫃
        #
        otc_dealer_buy_csv_path = f'{target_directory}/{date.year}{date.month:02d}{date.day:02d}_OTC_buy.csv'
        otc_dealer_sell_csv_path = f'{target_directory}/{date.year}{date.month:02d}{date.day:02d}_OTC_sell.csv'
        if os.path.isfile(otc_dealer_buy_csv_path) is False:
            url = f"https://www.tpex.org.tw/web/stock/3insti/dealer_trading/dealtr_hedge_result.php?l=zh-tw&t=D&type=buy&d={date.year - 1911}/{date.month:02d}/{date.day:02d}"
            print(url)

            r = requests.get(url, headers=request_headers)
            jdata = json.loads(r.content)

            if len(jdata['aaData']) == 0:
                continue

            with open(otc_dealer_buy_csv_path, 'w') as f:
                f.write('\ufeff')
                f.write(',,,(自行買賣),(自行買賣),(自行買賣),(避險),(避險),(避險),買賣超\n')
                f.write(',,,買進,賣出,買賣超,買進,賣出,買賣超,(仟股)\n')
                f.write(',,,,,(仟股),,,(仟股),,\n')
                # print(jdata)

                for x in jdata['aaData']:
                    # print(x)
                    string = '\t'.join(x)
                    string = string.replace(',', '')
                    string = string.replace('\t', ',')
                    f.write(string + '\n')

        if os.path.isfile(otc_dealer_sell_csv_path) is False:
            url = f"https://www.tpex.org.tw/web/stock/3insti/dealer_trading/dealtr_hedge_result.php?l=zh-tw&t=D&type=sell&d={date.year - 1911}/{date.month:02d}/{date.day:02d}"
            print(url)

            r = requests.get(url, headers=request_headers)
            jdata = json.loads(r.content)
            with open(otc_dealer_sell_csv_path, 'w') as f:
                f.write('\ufeff')
                f.write(',,,(自行買賣),(自行買賣),(自行買賣),(避險),(避險),(避險),買賣超\n')
                f.write(',,,買進,賣出,買賣超,買進,賣出,買賣超,(仟股)\n')
                f.write(',,,,,(仟股),,,(仟股),,\n')

                for x in jdata['aaData']:
                    # print(x)
                    string = '\t'.join(x)
                    string = string.replace(',', '')
                    string = string.replace('\t', ',')
                    f.write(string + '\n')

        #
        # 上市
        #
        dealer_csv_path = f'{target_directory}/{date.year}{date.month:02d}{date.day:02d}.csv'
        if os.path.isfile(dealer_csv_path) is False:
            url = f'https://www.twse.com.tw/fund/TWT43U?response=csv&date={date.year}{date.month:02d}{date.day:02d}'
            print(url)

            r = requests.get(url, headers=request_headers)
            with open(dealer_csv_path + '_', 'wb') as f:
                f.write(r.content)

            with open(dealer_csv_path, 'w') as f:
                f.write('\ufeff')

            time.sleep(10)

            os.system(f'/usr/bin/iconv -f BIG5-2003 -t UTF-8 {dealer_csv_path}_ >> {dealer_csv_path}')
            os.unlink(dealer_csv_path + '_')


#
# 根據 date_column 當中的值去建立第二天 ticks 當中的資料，包括：
#
# 1. 在第幾分鐘拉到 percentage。這部分資料需要先用 ShioajiPythonTrade/ticks_downloader.py 下載下來，放在 ../shioaji_ticks 的目錄裡面
#    這邊的程式才能找得到。目前使用soft line ../shioaji_ticks -> ../ShioajiPythonTrade/shioaji_ticks
#
# 2. 是否碰到漲停，或是是否碰到跌停
#
def build_ticks_related_info(sheet,
                             date_column,
                             symbol_column,
                             start_row,
                             start_column,
                             percentage,
                             greater=True,
                             next_day=True):
    set_cell(sheet, start_row, start_column + 0, '秒數')

    for i in range(start_row + 1, sheet.max_row + 1):
        print(f'idx = {i}')
        symbol = load_value(sheet, i, symbol_column)
        if symbol is None:
            continue
        date: datetime.datetime = load_value(sheet, i, date_column)
        if date is None:
            continue
        if type(date) == str:
            if '/' in date:
                date = datetime.datetime.strptime(date, '%Y/%m/%d')
            elif '-' in date:
                date = datetime.datetime.strptime(date, '%Y-%m-%d')
            else:
                print(f'Unknown date format {date}')
                exit(0)
        date = datetime.datetime(year=date.year, month=date.month, day=date.day)

        if next_day:
            days = 1
        else:
            days = 0

        found = False
        for day in date_range(date + timedelta(days=days), date + timedelta(days=days + 21)):
            date_str = day.strftime('%Y-%m-%d')
            tick_file = f'../shioaji_ticks/{symbol}/{date_str}.csv'

            if os.path.isfile(tick_file):
                found = True
                break

        if not found:
            continue

        ticks = csv.DictReader(open(tick_file))

        found_threshold = False
        first_tick = True
        for tick in ticks:
            if first_tick:
                first_tick = False
                Open = float(tick['close'])
                current_tick = get_current_price_tick(Open)
                threshold = Open * (1 + percentage)
                threshold = price_ticks_offset(threshold, 0)
                print(f'開盤：{Open} 觸價：{threshold}')

            if greater:
                if float(tick['close']) >= threshold:
                    found_threshold = True
                    break
            else:
                if float(tick['close']) <= threshold:
                    found_threshold = True
                    break

        if found_threshold:
            date_str, time_str = tick['ts'].split(' ')
            then = parse_date_time(date_str, time_str)
            begin = datetime.datetime.strptime(f'{date_str.split(" ")[0]} 09:00:00', '%Y-%m-%d %H:%M:%S')
            seconds = (then - begin).seconds
        else:
            seconds = 86400

        # print(seconds)
        set_cell(sheet, i, start_column + 0, seconds)


def build_dealer_info(sheet,
                      date_column,
                      symbol_column,
                      start_row,
                      start_column,
                      dealer_history_days=1):
    if False:
        set_cell(sheet, start_row - 1, start_column + 0, '自營商(自行買賣)')
        set_cell(sheet, start_row - 1, start_column + 3, '自營商(避險)')
        set_cell(sheet, start_row - 1, start_column + 6, '自營商')

    set_cell(sheet, start_row, start_column + 0, '自行 買進股數')
    set_cell(sheet, start_row, start_column + 1, '自行 賣出股數')
    set_cell(sheet, start_row, start_column + 2, '自行 買賣超股數')
    set_cell(sheet, start_row, start_column + 3, '避險 買進股數')
    set_cell(sheet, start_row, start_column + 4, '避險 賣出股數')
    set_cell(sheet, start_row, start_column + 5, '避險 買賣超股數')
    set_cell(sheet, start_row, start_column + 6, '買進股數')
    set_cell(sheet, start_row, start_column + 7, '賣出股數')
    set_cell(sheet, start_row, start_column + 8, '買賣超股數')

    for i in range(start_row + 1, sheet.max_row + 1):
        symbol = load_value(sheet, i, symbol_column)
        if symbol is None:
            continue
        date: datetime.datetime = load_value(sheet, i, date_column)
        if date is None:
            continue
        if type(date) == str:
            if '/' in date:
                date = datetime.datetime.strptime(date, '%Y/%m/%d')
            elif '-' in date:
                date = datetime.datetime.strptime(date, '%Y-%m-%d')
            else:
                print(f'Unknown date format {date}')
                exit(0)
        date = datetime.datetime(year=date.year, month=date.month, day=date.day)

        if dealer_history_days != 1:
            start_count_date = find_last_number_day(date, dealer_history_days)
            print(f'count from: {start_count_date.year}/{start_count_date.month:02d}/{start_count_date.day:02d}')

        print(f'symbol: {symbol}')

        #
        # 上櫃
        #

        otc_dealer_buy_csv_path = f'自營商歷史資料/{date.year}{date.month:02d}{date.day:02d}_OTC_buy.csv'
        otc_dealer_sell_csv_path = f'自營商歷史資料/{date.year}{date.month:02d}{date.day:02d}_OTC_sell.csv'

        #
        # Local record not found.
        # Download it from www.tpex.org.tw
        #
        if os.path.isfile(otc_dealer_buy_csv_path) is False:
            url = f"https://www.tpex.org.tw/web/stock/3insti/dealer_trading/dealtr_hedge_result.php?l=zh-tw&t=D&type=buy&d={date.year - 1911}/{date.month:02d}/{date.day:02d}"
            print(url)

            r = requests.get(url, headers=request_headers)
            jdata = json.loads(r.content)
            with open(otc_dealer_buy_csv_path, 'w') as f:
                f.write('\ufeff')
                f.write(',,,(自行買賣),(自行買賣),(自行買賣),(避險),(避險),(避險),買賣超\n')
                f.write(',,,買進,賣出,買賣超,買進,賣出,買賣超,(仟股)\n')
                f.write(',,,,,(仟股),,,(仟股),,\n')
                # print(jdata)

                for x in jdata['aaData']:
                    # print(x)
                    string = '\t'.join(x)
                    string = string.replace(',', '')
                    string = string.replace('\t', ',')
                    f.write(string + '\n')

        #
        # File should exists since here. Open it.
        #
        if os.path.isfile(otc_dealer_buy_csv_path) is True:
            print(otc_dealer_buy_csv_path + ' exits')
            with open(otc_dealer_buy_csv_path) as f:
                rows = csv.reader(f)
                for r in rows:
                    if len(r) < 10:
                        continue
                    if r[1] == symbol:
                        print(f'found {symbol}')
                        #
                        # 0. 流水號
                        # 1. symbol
                        # 2. 中文名稱
                        # 3. 自行買賣 買進
                        # 4. 自行買賣 賣出
                        # 5. 自行買賣 買賣超
                        # 6. 避險 買進
                        # 7. 避險 賣出
                        # 8. 避險 買賣超
                        # 9. 總買賣超
                        #
                        set_cell(sheet, i, start_column + 0, str_to_float(r[3]))
                        set_cell(sheet, i, start_column + 1, str_to_float(r[4]))
                        set_cell(sheet, i, start_column + 2, str_to_float(r[5]))

                        set_cell(sheet, i, start_column + 3, str_to_float(r[6]))
                        set_cell(sheet, i, start_column + 4, str_to_float(r[7]))
                        set_cell(sheet, i, start_column + 5, str_to_float(r[8]))

                        set_cell(sheet, i, start_column + 6, str_to_float(r[3]) + str_to_float(r[6]))
                        set_cell(sheet, i, start_column + 7, str_to_float(r[4]) + str_to_float(r[7]))
                        set_cell(sheet, i, start_column + 8, str_to_float(r[9]))

        if os.path.isfile(otc_dealer_sell_csv_path) is False:
            url = f"https://www.tpex.org.tw/web/stock/3insti/dealer_trading/dealtr_hedge_result.php?l=zh-tw&t=D&type=sell&d={date.year - 1911}/{date.month:02d}/{date.day:02d}"
            print(url)
            r = requests.get(url, headers=request_headers)
            jdata = json.loads(r.content)
            with open(otc_dealer_sell_csv_path, 'w') as f:
                f.write('\ufeff')
                f.write(',,,(自行買賣),(自行買賣),(自行買賣),(避險),(避險),(避險),買賣超\n')
                f.write(',,,買進,賣出,買賣超,買進,賣出,買賣超,(仟股)\n')
                f.write(',,,,,(仟股),,,(仟股),,\n')

                for x in jdata['aaData']:
                    # print(x)
                    string = '\t'.join(x)
                    string = string.replace(',', '')
                    string = string.replace('\t', ',')
                    f.write(string + '\n')

        if os.path.isfile(otc_dealer_sell_csv_path) is True:
            print(otc_dealer_sell_csv_path + ' exits')
            with open(otc_dealer_sell_csv_path) as f:
                rows = csv.reader(f)
                for r in rows:
                    if len(r) < 10:
                        continue
                    if r[1] == symbol:
                        print(f'found {symbol}')

                        set_cell(sheet, i, start_column + 0, str_to_float(r[3]))
                        set_cell(sheet, i, start_column + 1, str_to_float(r[4]))
                        set_cell(sheet, i, start_column + 2, str_to_float(r[5]))
                        set_cell(sheet, i, start_column + 3, str_to_float(r[6]))
                        set_cell(sheet, i, start_column + 4, str_to_float(r[7]))
                        set_cell(sheet, i, start_column + 5, str_to_float(r[8]))
                        set_cell(sheet, i, start_column + 6, str_to_float(r[3]) + str_to_float(r[6]))
                        set_cell(sheet, i, start_column + 7, str_to_float(r[4]) + str_to_float(r[7]))
                        set_cell(sheet, i, start_column + 8, str_to_float(r[9]))

        #
        # 上市
        #
        dealer_csv_path = f'自營商歷史資料/{date.year}{date.month:02d}{date.day:02d}.csv'

        if os.path.isfile(dealer_csv_path) is False:
            url = f'https://www.twse.com.tw/fund/TWT43U?response=csv&date={date.year}{date.month:02d}{date.day:02d}'
            print(url)
            r = requests.get(url, headers=request_headers)
            with open(dealer_csv_path + '_', 'wb') as f:
                f.write(r.content)

            os.system(f'/usr/bin/iconv -f BIG5-2003 -t UTF-8 {dealer_csv_path}_ > {dealer_csv_path}')
            os.unlink(dealer_csv_path + '_')

        if os.path.isfile(dealer_csv_path) is True:
            print(dealer_csv_path + ' exits')
            with open(dealer_csv_path) as f:
                rows = csv.reader(f)
                for r in rows:
                    if len(r) == 0:
                        continue
                    if r[0].rstrip(' ') == symbol:
                        # print(f'found {symbol}')

                        set_cell(sheet, i, start_column + 0, str_to_float(r[2]))
                        set_cell(sheet, i, start_column + 1, str_to_float(r[3]))
                        set_cell(sheet, i, start_column + 2, str_to_float(r[4]))
                        set_cell(sheet, i, start_column + 3, str_to_float(r[5]))
                        set_cell(sheet, i, start_column + 4, str_to_float(r[6]))
                        set_cell(sheet, i, start_column + 5, str_to_float(r[7]))
                        set_cell(sheet, i, start_column + 6, str_to_float(r[8]))
                        set_cell(sheet, i, start_column + 7, str_to_float(r[9]))
                        set_cell(sheet, i, start_column + 8, str_to_float(r[10]))

    return start_column + 9


def build_win_rate(sheet, symbol_column, profit_column, start_row):
    collections = {}

    for i in range(start_row + 1, sheet.max_row + 1):
        symbol = load_value(sheet, i, symbol_column)
        if symbol is None:
            continue
        profit = load_value(sheet, i, profit_column)

        if profit is None:
            continue

        if symbol not in collections.keys():
            collections[symbol] = [0, 0]

        if profit > 0:
            collections[symbol][0] = collections[symbol][0] + 1

        collections[symbol][1] = collections[symbol][1] + 1

    results = []
    for key in collections:
        win, total = collections[key]
        results.append((win / total, total, key))

    results.sort()

    print(results)


def get_current_price_tick(price, down=False):
    # print(f'get_current_price_tick: {price}')
    if down:
        if price <= 10:
            return 0.01
        if 10 < price <= 50:
            return 0.05
        if 50 < price <= 100:
            return 0.1
        if 100 < price <= 500:
            return 0.5
        if 500 < price <= 1000:
            return 1
        if price > 1000:
            return 5
    else:
        if price < 10:
            return 0.01
        if 10 <= price < 50:
            return 0.05
        if 50 <= price < 100:
            return 0.1
        if 100 <= price < 500:
            return 0.5
        if 500 <= price < 1000:
            return 1
        if price >= 1000:
            return 5


def price_ticks_offset(price, ticks):
    current_tick = get_current_price_tick(price)
    # print(f'current tick: {current_tick}')
    price = math.floor(price / current_tick)
    price *= current_tick
    price = round(price, 3)
    # print(f'normalized price: {price}')
    if ticks == 0:
        return price
    step = 1 if ticks > 0 else -1
    for i in range(0, ticks, step):
        current_tick = get_current_price_tick(price, down=True if step == -1 else False)
        # print(i, price, current_tick)
        price += current_tick * step
    return round(price, 3)


def get_limit_up_and_down_price(price):
    limit_up = price * 1.1
    tick = get_current_price_tick(limit_up)
    limit_up = limit_up // tick
    limit_up = limit_up * tick

    limit_down = price * 0.9
    tick = get_current_price_tick(limit_down)
    limit_down = math.ceil(limit_down / tick)
    limit_down = limit_down * tick

    return round(limit_up, 3), round(limit_down, 3)


def list_fullhistory(human_readable_name, file_after, df_filter):
    def rename_column(org):
        x = org
        x = x.rstrip()
        x = x.strip()
        x = x.replace(' ', '')
        x = x.replace('(', '')
        x = x.replace(')', '')
        x = x.replace('%', '')
        # print(f'rename function: org: [{org}] [{x}]')
        return x

    human_readable_name = '開盤當沖'
    output_csv_file = f'{human_readable_name}.csv'
    output_excel_file = f'{human_readable_name}.xlsx'

    files = os.listdir('fullhistory')
    files.sort()

    first_record = True

    if os.path.isfile(output_csv_file):
        os.unlink(output_csv_file)

    for file in files:

        if file[0:2] != '20':
            continue

        if not os.path.isfile(f'fullhistory/{file}'):
            continue

        if file_after is not None:
            if file <= file_after:
                continue

        # print(file)

        cache_dir = 'fullhistory/cache'
        if not os.path.isdir(cache_dir):
            os.mkdir(cache_dir)
        cache_file = f'{cache_dir}/{file.split(".")[0]}.csv'
        if os.path.isfile(cache_file):
            print(f'cache found, load {cache_file}')
            df = pd.read_csv(cache_file)
        else:
            df = pd.read_excel(f'fullhistory/{file}', index_col=False)
            df.rename(columns=rename_column, inplace=True)
            df.to_csv(cache_file)
        df.reset_index(drop=True, inplace=True)

        # 過濾一些不要的標的
        if True:
            df['代號'] = df['代號'].astype('str')
            mask = df['代號'].str.len() == 4
            df = df.loc[mask]
            df['細產業分類'] = df['細產業分類'].astype('str')
            mask = df['細產業分類'].str.len() >= 4
            df = df.loc[mask]

        df = df_filter(df)

        df['日期'] = f'{file[0:4]}/{file[4:6]}/{file[6:8]}'
        df['種類'] = human_readable_name

        df = df[['日期', '代號', '種類']]
        df.to_csv(output_csv_file, mode='a', header=first_record)
        first_record = False

    wb = openpyxl.Workbook()
    ws = wb.active
    with open(output_csv_file, encoding='utf8') as f:
        reader = csv.reader(f, delimiter=',')
        for row in reader:
            ws.append(row)
    wb.save(output_excel_file)

    return output_csv_file
