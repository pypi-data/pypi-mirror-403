from django.http import HttpRequest, HttpResponse, JsonResponse
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User
from django.contrib.auth import authenticate
from django.shortcuts import render
from django import forms
import json
import rpc.methods as rpc_methods
from rpc.methods import ServiceException, _check_project_permission
from inspect import getmembers, isfunction
from resources.helpers import (
    handle_uploaded_file,
    UnknownFileType,
    ResourceError,
    ConcurrencyError,
)
from resources import models
from typing import Union
from jama import settings
import re
import os
from glob import glob
import hashlib
from django.core.files.base import File
from django.db import transaction
from pathlib import Path
from shutil import rmtree
import base64
from .fileresponse import RangedFileResponse
from django.views.decorators.gzip import gzip_page
import logging
from functools import lru_cache
from resources.tasks import (
    exif_task,
    iiif_task,
    ocr_task,
    update_data_from_xlsx_rows,
    hls_task,
)
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import IllegalCharacterError
import tempfile
from rpc.signals import rpc_success_signal


importlib = __import__("importlib")

logger = logging.getLogger(__name__)


def _silent_rmdir(dir_path: str):
    rmtree(dir_path, ignore_errors=True)


def _silent_remove(file_path: str):
    try:
        os.remove(file_path)
    except FileNotFoundError:
        pass


def _file_hash256(file_path: str) -> str:
    hsh = hashlib.sha256()
    with open(file_path, "rb") as f:
        for chnk in iter(lambda: f.read(8192), b""):
            hsh.update(chnk)
    return hsh.hexdigest()


def _collection_from_origin_dir(
    origin_dir: str, project: models.Project
) -> Union[models.Collection, None]:
    # root dir always first
    previous_dir = project.root_collection
    for dir_name in origin_dir.split("/"):
        dir_name = dir_name.strip()
        if not dir_name:
            continue
        dir_name = dir_name.strip()
        if dir_name:
            previous_dir, created = models.Collection.objects.get_or_create(
                project=project, title=dir_name, parent=previous_dir
            )
            # this was soft-deleted. Un-delete !
            if not created and previous_dir.deleted_at:
                previous_dir.deleted_at = None
                previous_dir.save()
    return previous_dir


def _join_parts(
    parts_dir: str, destination: str, sha256sum: str, total_parts: int
) -> bool:
    with open(destination, "wb") as f:
        for path in sorted(glob("{}/{}-*.part".format(parts_dir, total_parts))):
            with open(path, "rb") as s:
                f.write(s.read())

    for path in glob("{}/{}-*.part".format(parts_dir, total_parts)):
        _silent_remove(path)
    if _file_hash256(destination) != sha256sum:
        return False
    return True


def _get_user_from_request(request: HttpRequest) -> Union[User, None]:
    """
    Tries to find a user given the request. In this order:

    - try to find session user
    - try to find user associated with X-api-key header
    - try to find user corresponding to Authorization header
    """
    if request.user.is_authenticated:
        logger.debug(f"user {request.user.username} is authenticated")
        return request.user
    if "X-Api-Key" in request.headers:
        try:
            return User.objects.get(
                apikey__key=request.headers["X-Api-Key"], apikey__active=True
            )
        except User.DoesNotExist:
            logger.warning("X-Api-Key was given but user was not found")
            return None
    if "Authorization" in request.headers:
        try:
            auth_u, auth_p = (
                base64.decodebytes(request.headers["Authorization"][6:].encode("utf-8"))
                .decode("utf-8")
                .split(":")
            )
            user = authenticate(username=auth_u, password=auth_p)
            if not user:
                logger.warning(
                    "Authorization was given in request headers but user was not authenticated"
                )
            return user
        except KeyError:
            logger.warning("Authorization was malformed")
            return None


def _get_project_from_request(request: HttpRequest) -> Union[models.Project, None]:
    if "X-Project" in request.headers:
        return models.Project.objects.filter(pk=request.headers["X-Project"]).first()
    return None


@lru_cache(None, typed=True)
def _get_rpc_methods() -> dict:
    methods = {}
    for fn_name, _ in getmembers(rpc_methods, isfunction):
        if fn_name[0:1] == "_":
            continue
        methods[fn_name] = getattr(rpc_methods, fn_name)
    for app_name in settings.AUTO_REGISTER_APPS:
        try:
            rpc_module = importlib.import_module("{}.rpc.methods".format(app_name))
            for fn_name, _ in getmembers(rpc_module, isfunction):
                if fn_name[0:1] == "_":
                    continue
                methods[fn_name] = getattr(rpc_module, fn_name)
        except ModuleNotFoundError:
            continue
    return methods


@gzip_page
@transaction.atomic
@csrf_exempt
def rpc(request: HttpRequest) -> HttpResponse:
    req_id = None
    if request.method == "GET":
        return render(request, "docs.html", {"methods": _get_rpc_methods()})
    if request.method != "POST":
        return HttpResponse("Method Not Allowed", status=405)
    try:
        user = _get_user_from_request(request)
        if not user:
            return HttpResponse("Forbidden", status=403)
        methods = _get_rpc_methods()
        json_data = json.loads(request.body)
        method = json_data["method"]
        if method not in methods.keys():
            logger.warning(f'user {user} called unknown method "{method}"')
            return JsonResponse(
                {
                    "result": None,
                    "error": "no such method",
                    "id": req_id,
                }
            )
        params = json_data["params"] if "params" in json_data else None
        if not params:
            params = []
        params.insert(0, user)  # always add user to method call
        req_id = json_data["id"]
        result = methods[method](*params)
        logger.info(
            'User {}({}) called "{}" with params {}'.format(
                user.username, user.pk, method, params[1:]
            )
        )
        params[0] = user.username
        rpc_success_signal.send_robust(sender=method, params=params, result=result)
        return JsonResponse({"result": result, "error": None, "id": req_id})
    except json.JSONDecodeError:
        logger.warning("could not decode json request")
        return HttpResponse("Bad Request", status=400)
    except KeyError as err:
        logger.warning("client data dict key error")
        logger.exception(err)
        return HttpResponse("Bad Request", status=400)
    except TypeError as err:
        logger.warning("client data type error")
        logger.exception(err)
        return HttpResponse("Bad Request", status=400)
    except ValueError as err:
        logger.warning("client data warning error")
        logger.exception(err)
        return HttpResponse("Bad Request", status=400)
    except ServiceException as e:
        return JsonResponse(
            {
                "result": None,
                "error": getattr(e, "message", repr(e)),
                "id": req_id,
            }
        )


@csrf_exempt
@require_http_methods(["POST"])
def upload_partial(request: HttpRequest) -> HttpResponse:
    user = _get_user_from_request(request)
    if not user:
        return HttpResponse("Forbidden", status=403)
    project = _get_project_from_request(request)
    if not project:
        return HttpResponse("Bad Request", status=400)
    try:
        _check_project_permission(user, project, "file.upload")
    except ServiceException:
        return HttpResponse("Forbidden", status=403)
    # optional header, used to create collections
    origin_dir = request.headers.get("X-origin-dir", None)

    try:
        file_hash = request.headers["X-file-hash"]
        if re.match("[A-Fa-f0-9]{64}", file_hash) is None:
            raise ValueError
        file_name = base64.b64decode(request.headers["X-file-name"]).decode("utf-8")
        _, f_extension = os.path.splitext(file_name)
        if not f_extension:
            return HttpResponse("Type de fichier inconnu", status=400)
        chunk_number, total_chunks = request.headers["X-file-chunk"].split("/")
        file_instance = models.File.objects.get(hash=file_hash, project=project)
        # Rise from your grave !
        if file_instance.deleted_at:
            file_instance.deleted_at = None
            file_instance.title = file_name
            file_instance.original_name = file_name
            file_instance.save()
            logger.info(
                "User {}({}) respawned file {}".format(
                    user.username, user.pk, file_instance.pk
                )
            )
        if origin_dir:
            collection = _collection_from_origin_dir(origin_dir, project)
            collection.resources.add(file_instance)
            logger.info(
                "User {}({}) added file({}) to collection({})".format(
                    user.username, user.pk, file_instance.pk, collection.pk
                )
            )
        return HttpResponse(file_instance.id, status=200)
    except (KeyError, ValueError):
        return HttpResponse(status=400)
    except models.File.DoesNotExist:
        pass

    partials_dir = "{}/{}-{}".format(settings.PARTIAL_UPLOADS_DIR, user.pk, file_hash)
    lock_file_path = "{}/.lock".format(partials_dir)
    if os.path.isfile(lock_file_path):
        return HttpResponse(status=202)
    os.makedirs(partials_dir, exist_ok=True)

    # We don't want to store an incomplete chunk.
    # (prevent client disconnect)
    try:
        if len(request.body) == int(request.headers["Content-Length"]):
            partial_destination = "{}/{}-{}.part".format(
                partials_dir, total_chunks, chunk_number.zfill(len(total_chunks))
            )
            with open(partial_destination, "wb") as f:
                f.write(request.body)
        else:
            return HttpResponse("Content-Length does not match body size", status=400)
    except KeyError:  # Content-length header is mandatory
        return HttpResponse("Length Required", status=411)

    # All chunks are complete, time to assemble
    f_name = "{}/complete{}".format(partials_dir, f_extension)
    if len(glob("{}/{}-*.part".format(partials_dir, total_chunks))) == int(
        total_chunks
    ) or os.path.exists(f_name):
        Path(lock_file_path).touch()  # lock dir

        if not os.path.exists(f_name):
            checksum_ok = _join_parts(partials_dir, f_name, file_hash, total_chunks)
        elif file_hash == _file_hash256(f_name):
            checksum_ok = True
        else:
            checksum_ok = False
        try:
            with open(f_name, "rb") as f:
                try:
                    file_id = None
                    with transaction.atomic():
                        file_id = handle_uploaded_file(
                            File(f),
                            project,
                            force_file_name=file_name,
                        )
                        f.close()
                        _silent_rmdir(partials_dir)
                        logger.info(
                            "User {}({}) handled file {}".format(
                                user.username, user.pk, file_id
                            )
                        )
                        if origin_dir:
                            collection = _collection_from_origin_dir(
                                origin_dir, project
                            )
                            collection.resources.add(
                                models.File.objects.get(id=file_id)
                            )
                            logger.info(
                                "User {}({}) added file({}) to collection({})".format(
                                    user.username, user.pk, file_id, collection.pk
                                )
                            )
                    iiif_task(file_id)
                    exif_task(file_id)
                    ocr_task(file_id)
                    hls_task(file_id)
                    if not checksum_ok:
                        return HttpResponse(file_id, status=210)  # content different
                    else:
                        return HttpResponse(file_id, status=200)
                except UnknownFileType:
                    _silent_rmdir(partials_dir)
                    return HttpResponse("Type de fichier inconnu", status=400)
                except ResourceError:
                    _silent_rmdir(partials_dir)
                    return HttpResponse(
                        "Impossible d'enregistrer la ressource", status=400
                    )
                except ConcurrencyError:
                    # Concurrency problem with clients sending multiple chunks in parallel.
                    # Nothing to do here, work has already been done.
                    return HttpResponse("Too Many Requests", status=429)
        # another race condition
        except FileNotFoundError:
            return HttpResponse("Too Many Requests", status=429)
    return HttpResponse(status=202)


def force_download(
    request: HttpRequest, file_id: int
) -> Union[HttpResponse, RangedFileResponse]:
    file = models.File.objects.filter(id=file_id).first()
    if not file:
        return HttpResponse("Not Found", 404)
    try:
        if not settings.DEBUG:  # bypass auth in dev
            user = _get_user_from_request(request)
            if not user:
                return HttpResponse("Forbidden", status=403)
            _check_project_permission(user, file.project, "file.download_source")
    except ServiceException:
        return HttpResponse("Forbidden", status=403)
    return RangedFileResponse(
        request,
        open(file.local_path(), "rb"),
        filename=file.new_filename,
        # content_type=file.file_type.mime, # Firefox is bad
    )


def force_download_public(
    request: HttpRequest, file_hash: str
) -> Union[HttpResponse, RangedFileResponse]:
    file = models.File.objects.filter(hash=file_hash).first()
    if not file:
        return HttpResponse("Not Found", 404)
    return RangedFileResponse(
        request,
        open(file.local_path(), "rb"),
        filename=file.new_filename,
        # content_type=file.file_type.mime, # Firefox is bad
    )


def serve_file_public(
    request: HttpRequest, file_hash: str
) -> Union[HttpResponse, RangedFileResponse]:
    file = models.File.objects.filter(hash=file_hash).first()
    if not file:
        return HttpResponse("Not Found", 404)
    return RangedFileResponse(
        request,
        open(file.local_path(), "rb"),
        filename=file.new_filename,
        content_type=file.file_type.mime,  # Firefox is bad
    )


@csrf_exempt
@require_http_methods(["POST"])
def upload_metas_xls(request: HttpRequest) -> HttpResponse:
    data_rows = []
    user = _get_user_from_request(request)
    if not user:
        return HttpResponse("Forbidden", status=403)

    class MetasXSLXUploadForm(forms.Form):
        file = forms.FileField()

    form = MetasXSLXUploadForm(request.POST, request.FILES)
    if form.is_valid():
        with tempfile.NamedTemporaryFile(suffix=".xlsx") as tmp:
            with open(tmp.name, "wb+") as destination:
                for chunk in request.FILES["file"].chunks():
                    destination.write(chunk)
            wb = load_workbook(filename=tmp)
            ws = wb.active
            keys = []
            for row in ws.iter_rows(values_only=True):
                if not keys:
                    keys = list(row)
                else:
                    data = dict(zip(keys, list(row)))
                    data_rows.append(data)
            user_task = models.UserTask.objects.create(
                owner=user, description="Traitement de fichier XLS de métadonnées"
            )
            update_data_from_xlsx_rows(user.pk, data_rows, user_task.pk)
            return HttpResponse("OK", status=200)
    return HttpResponse("Bad Request", status=400)


def download_metas_xls(request: HttpRequest, collection_id):
    collection_instance = models.Collection.objects.filter(pk=collection_id).first()
    if not collection_instance:
        return HttpResponse("Not Found", status=404)
    if not settings.DEBUG:  # bypass auth in dev
        user = _get_user_from_request(request)
        if not user:
            return HttpResponse("Forbidden", status=403)
        _check_project_permission(user, collection_instance.project, "collection.read")

    wb = Workbook()
    ws = wb.active
    row_number = 2  # /!\ Index starts at 1. Leave the first row for headers.
    for r in collection_instance.yield_resource_data_for_export():
        col_number = 1
        for value in r.values():
            cell = ws.cell(row=row_number, column=col_number)
            try:
                cell.value = value
            except IllegalCharacterError:
                logger.warning(
                    f"Could not add value of type {type(value)} to cell: {value}"
                )
            col_number = col_number + 1
        row_number = row_number + 1
    col_number = 1
    for key in r.keys():
        cell = ws.cell(row=1, column=col_number)
        cell.value = key
        col_number = col_number + 1

    with tempfile.NamedTemporaryFile(suffix=".xlsx") as tmp:
        wb.save(tmp.name)
        return RangedFileResponse(
            request,
            open(tmp.name, "rb"),
            filename=f"collection_{collection_id}.xlsx",
            content_type="application/vnd.ms-excel",
        )
