from pathlib import Path
from openpyxl import load_workbook, Workbook
import pandas as pd
from typing import Optional, Union

from ..path_manager import sanitize_filename, make_fullpath
from .._core import get_logger


_LOGGER = get_logger("Excel Handler")


__all__ = [
    "find_excel_files",
    "unmerge_and_split_excel",
    "unmerge_and_split_from_directory",
    "validate_excel_schema",
    "vertical_merge_transform_excel",
    "horizontal_merge_transform_excel"
]


def find_excel_files(
    directory: Union[str, Path],
    *,
    extensions: tuple[str, ...] = (".xlsx", ".xls"),
    exclude_temp: bool = True
) -> list[Path]:
    """
    Returns a list of Excel file Paths in the specified directory.

    Parameters:
        directory (str | Path): Directory to search.
        extensions (tuple[str, ...]): Valid Excel file extensions (default: .xlsx, .xls).
        exclude_temp (bool): Whether to exclude files that start with '~'.

    Returns:
        list[Path]: List of Excel file paths matching criteria.
    """
    input_path = make_fullpath(directory)

    if not input_path.is_dir():
        _LOGGER.error(f"Directory not found: {input_path}")
        raise NotADirectoryError()

    excel_files = [
        f for f in input_path.iterdir()
        if f.is_file()
        and f.suffix.lower() in extensions
        and (not f.name.startswith('~') if exclude_temp else True)
    ]
    
    if not excel_files:
        _LOGGER.error(f"No valid Excel files found in directory: {input_path}")
        raise FileNotFoundError()

    return excel_files


def unmerge_and_split_excel(filepath: Union[str,Path]) -> None:
    """
    Processes a single Excel file:
      - Unmerges all merged cells (vertical and horizontal), fills each merged region with the top-left cell value.
      - Splits each sheet into a separate Excel file.
      - Saves all results in the same directory as the input file.

    Parameters:
        filepath (str | Path): Full path to the Excel file to process.
    """
    file_path = make_fullpath(filepath)
    wb = load_workbook(file_path)
    base_dir = file_path.parent
    base_name = file_path.stem

    total_output_files = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        new_wb = Workbook()
        new_ws = new_wb.active
        new_ws.title = sheet_name # type: ignore

        # Copy all cell values
        for row in ws.iter_rows():
            for cell in row:
                new_ws.cell(row=cell.row, column=cell.column, value=cell.value) # type: ignore

        # Fill and unmerge merged regions
        for merged_range in list(ws.merged_cells.ranges):
            min_row, min_col, max_row, max_col = (
                merged_range.min_row, merged_range.min_col,
                merged_range.max_row, merged_range.max_col
            )
            value = ws.cell(row=min_row, column=min_col).value
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    new_ws.cell(row=row, column=col, value=value) # type: ignore

        # Construct flat output file name
        sanitized_sheet_name = sanitize_filename(sheet_name)
        output_filename = f"{base_name}_{sanitized_sheet_name}.xlsx"
        output_path = base_dir / output_filename
        new_wb.save(output_path)

        total_output_files += 1

    _LOGGER.info(f"Processed file: {file_path} into {total_output_files} output file(s).")
    return None


def unmerge_and_split_from_directory(input_dir: Union[str,Path], output_dir: Union[str,Path]) -> None:
    """
    Processes all Excel files in the input directory:
      - Unmerges all merged cells (vertical and horizontal), fills each merged region with the top-left cell value,
      - Splits each sheet into separate Excel files.
      - Saves all results into the output directory.

    Parameters:
        input_dir (str | Path): Directory containing Excel files to process.
        output_dir (str | Path): Directory to save processed Excel files.
    """
    global_input_path = make_fullpath(input_dir)
    global_output_path = make_fullpath(output_dir, make=True)
    
    excel_files = find_excel_files(global_input_path)

    total_output_files = 0

    for file_path in excel_files:
        wb = load_workbook(file_path)
        base_name = file_path.stem

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            new_wb = Workbook()
            new_ws = new_wb.active
            new_ws.title = sheet_name # type: ignore

            # Copy all cell values
            for row in ws.iter_rows():
                for cell in row:
                    new_ws.cell(row=cell.row, column=cell.column, value=cell.value) # type: ignore

            # Fill and unmerge merged regions
            for merged_range in list(ws.merged_cells.ranges):
                min_row, min_col, max_row, max_col = (
                    merged_range.min_row, merged_range.min_col,
                    merged_range.max_row, merged_range.max_col
                )
                value = ws.cell(row=min_row, column=min_col).value
                for row in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        new_ws.cell(row=row, column=col, value=value) # type: ignore

            # Construct flat output file name
            sanitized_sheet_name = sanitize_filename(sheet_name)
            output_filename = f"{base_name}_{sanitized_sheet_name}.xlsx"
            output_path = global_output_path / output_filename
            new_wb.save(output_path)

            total_output_files += 1

    _LOGGER.info(f"Processed {len(excel_files)} input Excel file(s) with a total of {total_output_files} output Excel file(s).")
    return None


def validate_excel_schema(
    target_dir: Union[str,Path],
    expected_columns: list[str],
    strict: bool = False
) -> None:
    """
    Validates that each Excel file in a directory conforms to the expected column schema. Only the first worksheet of each file is analyzed.
    
    Parameters:
        target_dir (str | Path): Path to the directory containing Excel files.
        expected_columns (list[str]): List of expected column names.
        strict (bool): If True, columns must match exactly (names and order).
                      If False, columns must contain at least all expected names.
    """
    invalid_files: dict[str, str] = {}
    expected_set = set(expected_columns)
    
    target_path = make_fullpath(target_dir)
    excel_paths = find_excel_files(target_path)
    
    for file in excel_paths:
        try:
            # Using first worksheet
            wb = load_workbook(file, read_only=True)
            ws = wb.active

            header = [cell.value for cell in next(ws.iter_rows(max_row=1))] # type: ignore

            # Change 2: Detailed reason-finding logic
            if strict:
                if header != expected_columns:
                    header_set = set(header)
                    reason_parts = []
                    missing = sorted(list(expected_set - header_set)) # type: ignore
                    extra = sorted(list(header_set - expected_set)) # type: ignore
                    
                    if missing:
                        reason_parts.append(f"Missing: {missing}")
                    if extra:
                        reason_parts.append(f"Extra: {extra}")
                    if not missing and not extra:
                        reason_parts.append("Incorrect column order")
                    
                    invalid_files[file.name] = ". ".join(reason_parts)
            else:
                header_set = set(header)
                if not expected_set.issubset(header_set):
                    missing_cols = sorted(list(expected_set - header_set)) # type: ignore
                    reason = f"Missing required columns: {missing_cols}"
                    invalid_files[file.name] = reason

        except Exception as e:
            _LOGGER.error(f"Error processing '{file}': {e}")
            invalid_files[file.name] = f"File could not be read. Error: {e}"
    
    valid_excel_number = len(excel_paths) - len(invalid_files)
    _LOGGER.info(f"{valid_excel_number} out of {len(excel_paths)} excel files conform to the schema.")
    
    # Change 3: Updated print loop to show the reason
    if invalid_files:
        _LOGGER.warning(f"{len(invalid_files)} excel files are invalid:")
        for file_name, reason in invalid_files.items():
            print(f"  - {file_name}: {reason}")

    return None


def vertical_merge_transform_excel(
    target_dir: Union[str,Path],
    csv_filename: str,
    output_dir: Union[str,Path],
    target_columns: Optional[list[str]] = None,
    rename_columns: Optional[list[str]] = None
) -> None:
    """
    Merges multiple Excel files in a directory vertically and saves as a single CSV file.

    Constraints:
    - Only the first worksheet of each Excel file is processed.
    - All Excel files must have either the same column names or a common subset if `target_columns` is provided.
    - If `rename_columns` is provided, it must match the length of `target_columns` (if used) or the original columns. Names match by position.

    Parameters:
        target_dir (str | Path): Directory containing Excel files.
        csv_filename (str): Output CSV filename.
        output_dir (str | Path): Directory to save the output CSV file.
        target_columns (list[str] | None): Columns to select from each Excel file.
        rename_columns (list[str] | None): Optional renaming for columns. Position-based matching.
    """
    target_path = make_fullpath(target_dir)
    excel_files = find_excel_files(target_path)
    
    # sanitize filename
    csv_filename = sanitize_filename(csv_filename)
    # make output directory
    output_path = make_fullpath(output_dir, make=True)

    csv_filename = csv_filename if csv_filename.endswith('.csv') else f"{csv_filename}.csv"
    csv_path = output_path / csv_filename

    dataframes = []
    for file in excel_files:
        df = pd.read_excel(file, engine='openpyxl')

        if target_columns is not None:
            missing = [col for col in target_columns if col not in df.columns]
            if missing:
                _LOGGER.error(f"Invalid columns in {file.name}: {missing}")
                raise ValueError()
            df = df[target_columns]

        dataframes.append(df)

    merged_df = pd.concat(dataframes, ignore_index=True)

    if rename_columns is not None:
        expected_len = len(target_columns if target_columns is not None else merged_df.columns)
        if len(rename_columns) != expected_len:
            _LOGGER.error("Length of 'rename_columns' must match the selected columns")
            raise ValueError()
        merged_df.columns = rename_columns

    merged_df.to_csv(csv_path, index=False, encoding='utf-8')
    _LOGGER.info(f"Merged {len(dataframes)} excel files into '{csv_filename}'.")


def horizontal_merge_transform_excel(
    target_dir: Union[str,Path],
    csv_filename: str,
    output_dir: Union[str,Path],
    drop_columns: Optional[list[str]] = None,
    skip_duplicates: bool = False
) -> None:
    """
    Horizontally concatenates Excel files (first sheet of each) by aligning rows and expanding columns. 
    Then saves the result as a .csv file.

    Constraints:
    - All Excel files must have the same number of rows, or shorter ones will be padded with empty rows.
    - Only the first sheet in each Excel file is used.
    - Columns in `drop_columns` will be excluded from the result.
    - If `skip_duplicates` is False, duplicate columns are suffixed with "_copy", "_copy2", etc.
      If True, only the first occurrence of each column name is kept.

    Parameters:
        target_dir (str | Path): Directory containing Excel files.
        csv_filename (str): Name of the output CSV file.
        output_dir (str | Path): Directory to save the output CSV file.
        drop_columns (list[str] | None): Columns to exclude from each file before merging.
        skip_duplicates (bool): Whether to skip duplicate columns or rename them.
    """
    target_path = make_fullpath(target_dir)
    excel_files = find_excel_files(target_path)
    
    # sanitize filename
    csv_filename = sanitize_filename(csv_filename)
    # make directory
    output_path = make_fullpath(output_dir, make=True)

    csv_filename = csv_filename if csv_filename.endswith('.csv') else f"{csv_filename}.csv"
    csv_path = output_path / csv_filename

    dataframes = []
    max_rows = 0

    for file in excel_files:
        df = pd.read_excel(file, engine='openpyxl')

        if drop_columns is not None:
            df = df.drop(columns=[col for col in drop_columns if col in df.columns])

        max_rows = max(max_rows, len(df))
        dataframes.append(df)

    padded_dataframes = []
    for df in dataframes:
        padded_df = df.reindex(range(max_rows)).reset_index(drop=True)
        padded_dataframes.append(padded_df)

    merged_df = pd.concat(padded_dataframes, axis=1)

    duplicate_columns = merged_df.columns[merged_df.columns.duplicated()].tolist()
    
    if duplicate_columns:
        _LOGGER.warning(f"Duplicate columns: {duplicate_columns}")

    if skip_duplicates:
        merged_df = merged_df.loc[:, ~merged_df.columns.duplicated()]
    else:
        seen = {}
        new_cols = []
        for col in merged_df.columns:
            base_col = col
            count = seen.get(base_col, 0)
            if count:
                while f"{base_col}_copy{count}" in seen:
                    count += 1
                col = f"{base_col}_copy{count}"
            seen[col] = count + 1
            new_cols.append(col)
        merged_df.columns = new_cols

    merged_df.to_csv(csv_path, index=False, encoding='utf-8')

    _LOGGER.info(f"Merged {len(excel_files)} Excel files into '{csv_filename}'.")
    
