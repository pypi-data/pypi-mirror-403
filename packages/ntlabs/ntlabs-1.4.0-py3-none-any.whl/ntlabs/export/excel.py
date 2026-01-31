"""
Excel generation utilities.

Provides Excel generation using openpyxl.
"""

import io
import logging
from datetime import datetime
from typing import Any

from .styles import ExcelStyle, ExcelTheme, get_excel_theme

logger = logging.getLogger(__name__)

# Type alias for dataframe-like objects
DataFrameLike = Any  # pandas.DataFrame


class ExcelGenerator:
    """
    Excel workbook generator.

    Example:
        generator = ExcelGenerator(
            title="Financial Report",
            style=ExcelStyle.FINANCIAL,
        )

        excel_bytes = generator.generate(
            sheets={
                "Summary": summary_data,
                "Details": details_data,
            },
            metadata={"generated_at": datetime.now()}
        )
    """

    def __init__(
        self,
        title: str = "Workbook",
        author: str | None = None,
        style: ExcelStyle = ExcelStyle.DEFAULT,
        theme: ExcelTheme | None = None,
    ):
        """
        Initialize Excel generator.

        Args:
            title: Workbook title
            author: Workbook author
            style: Pre-defined style
            theme: Custom theme (overrides style)
        """
        self.title = title
        self.author = author
        self.theme = theme or get_excel_theme(style)

    def generate(
        self,
        sheets: dict[str, DataFrameLike | list[dict] | list[list]],
        metadata: dict[str, Any] | None = None,
    ) -> bytes:
        """
        Generate Excel workbook from data.

        Args:
            sheets: Dictionary mapping sheet names to data
                   (DataFrame, list of dicts, or list of lists)
            metadata: Workbook metadata

        Returns:
            Excel file bytes
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            from openpyxl.utils import get_column_letter
        except ImportError as err:
            raise ImportError(
                "openpyxl is required for Excel generation. "
                "Install it with: pip install openpyxl"
            ) from err

        wb = Workbook()

        # Remove default sheet
        default_sheet = wb.active
        wb.remove(default_sheet)

        # Define styles
        header_fill = PatternFill(
            start_color=self.theme.header_fill,
            end_color=self.theme.header_fill,
            fill_type="solid",
        )
        header_font = Font(
            name=self.theme.header_font_name,
            size=self.theme.header_font_size,
            bold=self.theme.header_bold,
            color=self.theme.header_font,
        )
        alt_row_fill = PatternFill(
            start_color=self.theme.alternating_row,
            end_color=self.theme.alternating_row,
            fill_type="solid",
        )
        border = Border(
            left=Side(style="thin", color=self.theme.border_color),
            right=Side(style="thin", color=self.theme.border_color),
            top=Side(style="thin", color=self.theme.border_color),
            bottom=Side(style="thin", color=self.theme.border_color),
        )

        for sheet_name, data in sheets.items():
            ws = wb.create_sheet(title=sheet_name[:31])  # Excel limit

            # Convert data to rows
            rows = self._data_to_rows(data)
            if not rows:
                continue

            # Write rows
            for row_idx, row in enumerate(rows, start=1):
                for col_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border

                    # Header row styling
                    if row_idx == 1:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center")
                    else:
                        # Alternating rows
                        if row_idx % 2 == 0:
                            cell.fill = alt_row_fill

                        # Apply number formats
                        if isinstance(value, (int, float)):
                            cell.number_format = "#,##0.00"
                        elif isinstance(value, datetime):
                            cell.number_format = self.theme.datetime_format

            # Auto-width columns
            if self.theme.auto_width:
                for col_idx in range(1, len(rows[0]) + 1):
                    col_letter = get_column_letter(col_idx)
                    max_length = 0

                    for row in rows:
                        if col_idx <= len(row):
                            cell_value = str(row[col_idx - 1] or "")
                            max_length = max(max_length, len(cell_value))

                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[col_letter].width = max(
                        adjusted_width, self.theme.default_width
                    )

            # Freeze header row
            if self.theme.freeze_header and len(rows) > 1:
                ws.freeze_panes = ws["A2"]

        # Set workbook properties
        wb.properties.title = self.title
        if self.author:
            wb.properties.creator = self.author
        if metadata:
            for key, value in metadata.items():
                if hasattr(wb.properties, key):
                    setattr(wb.properties, key, str(value))

        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()

    def _data_to_rows(
        self, data: DataFrameLike | list[dict] | list[list]
    ) -> list[list]:
        """Convert various data formats to list of lists."""
        # Check if it's a pandas DataFrame
        if hasattr(data, "iterrows") and hasattr(data, "columns"):
            # It's a DataFrame
            headers = list(data.columns)
            rows = [headers]
            for _, row in data.iterrows():
                rows.append(list(row))
            return rows

        # List of dicts
        if data and isinstance(data[0], dict):
            headers = list(data[0].keys())
            rows = [headers]
            for item in data:
                rows.append([item.get(h) for h in headers])
            return rows

        # List of lists (assume first row is headers)
        if data and isinstance(data[0], (list, tuple)):
            return [list(row) for row in data]

        return []

    def generate_single_sheet(
        self,
        data: DataFrameLike | list[dict] | list[list],
        sheet_name: str = "Sheet1",
        metadata: dict[str, Any] | None = None,
    ) -> bytes:
        """
        Generate Excel with single sheet.

        Convenience method for single-sheet workbooks.
        """
        return self.generate(
            sheets={sheet_name: data},
            metadata=metadata,
        )


async def generate_excel(
    data: DataFrameLike | list[dict] | list[list] | dict[str, Any],
    title: str = "Export",
    style: ExcelStyle = ExcelStyle.DEFAULT,
    author: str | None = None,
    metadata: dict[str, Any] | None = None,
) -> bytes:
    """
    Generate Excel from data.

    Convenience function for quick Excel generation.

    Args:
        data: Data to export (DataFrame, list of dicts, or dict of sheets)
        title: Workbook title
        style: Excel style
        author: Workbook author
        metadata: Workbook metadata

    Returns:
        Excel file bytes
    """
    generator = ExcelGenerator(
        title=title,
        author=author,
        style=style,
    )

    # If data is a dict with string keys, treat as multiple sheets
    if isinstance(data, dict) and all(isinstance(k, str) for k in data.keys()):
        # Check if it looks like sheet data vs a single dict
        first_value = next(iter(data.values()), None)
        if isinstance(first_value, (list, tuple)) or hasattr(first_value, "iterrows"):
            return generator.generate(sheets=data, metadata=metadata)

    # Single sheet
    return generator.generate_single_sheet(data, metadata=metadata)
