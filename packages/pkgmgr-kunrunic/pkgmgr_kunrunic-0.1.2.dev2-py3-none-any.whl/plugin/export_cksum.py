#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import argparse
import os
import re
import subprocess
import sys
import time
from copy import copy

try:
    import openpyxl
except Exception:
    openpyxl = None

from pkgmgr import config
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def _load_pkg_yaml(pkg_dir):
    if not pkg_dir:
        return None
    return os.path.join(pkg_dir, "pkg.yaml")


def _resolve_pkg_dir(pkg_id, config_path=None):
    try:
        if config_path:
            main_cfg = config.load_main(path=config_path, allow_interactive=False)
        else:
            main_cfg = config.load_main(allow_interactive=False)
    except Exception:
        return None
    release_root = main_cfg.get("pkg_release_root")
    if not release_root:
        return None
    return os.path.abspath(os.path.expanduser(os.path.join(release_root, str(pkg_id))))


def _group_release_paths(pkg_dir, releases):
    grouped = {}
    for entry in releases:
        if entry is None:
            continue
        rel = str(entry)
        target = rel if os.path.isabs(rel) else os.path.join(pkg_dir, rel)
        target = os.path.abspath(os.path.expanduser(target))
        if not os.path.exists(target):
            print("[export_cksum] skip missing: %s" % target)
            continue
        try:
            relpath = os.path.relpath(target, pkg_dir)
        except Exception:
            relpath = os.path.basename(target)
        if relpath.startswith(".."):
            print("[export_cksum] skip outside pkg_dir: %s" % target)
            continue
        parts = relpath.split(os.sep, 1)
        if len(parts) == 2:
            root, subrel = parts[0], parts[1]
        else:
            root, subrel = parts[0], ""
        grouped.setdefault(root, []).append((target, subrel))
    return grouped


def _collect_files(root_dir, entries):
    files = set()
    for target, subrel in entries:
        if subrel:
            base_dir = os.path.join(root_dir, subrel)
        else:
            base_dir = target
        if os.path.isfile(base_dir):
            files.add(base_dir)
            continue
        if not os.path.isdir(base_dir):
            continue
        for base, _, names in os.walk(base_dir):
            for name in names:
                abspath = os.path.join(base, name)
                if os.path.isfile(abspath):
                    files.add(abspath)
    return files


def _cksum(path):
    try:
        out = subprocess.check_output(["cksum", path], stderr=subprocess.STDOUT)
    except Exception as e:
        print("[export_cksum] cksum failed: %s (%s)" % (path, str(e)))
        return None
    line = out.decode("utf-8", errors="replace").strip()
    parts = line.split()
    if len(parts) < 3:
        print("[export_cksum] invalid cksum output: %s" % line)
        return None
    return parts[0], parts[1], " ".join(parts[2:])


def _normalize_excel_template(excel_arg, pkg_dir):
    path_template = excel_arg
    if os.sep not in path_template:
        export_dir = os.path.join(pkg_dir, "export")
        path_template = os.path.join(export_dir, path_template)
    if not path_template.lower().endswith(".xlsx"):
        path_template = path_template + ".xlsx"
    return path_template


def _next_version(path_template):
    dir_path = os.path.dirname(path_template) or "."
    name_template = os.path.basename(path_template)
    regex = re.escape(name_template)
    regex = regex.replace(re.escape("{YYYYMMDD}"), r"\d{8}")
    regex = regex.replace(re.escape("{date}"), r"\d{8}")
    regex = regex.replace(re.escape("{version}"), r"v(\d+)")
    regex = "^" + regex + "$"
    max_ver = 0
    if os.path.isdir(dir_path):
        for name in os.listdir(dir_path):
            m = re.match(regex, name)
            if not m:
                continue
            try:
                ver = int(m.group(1))
            except Exception:
                continue
            if ver > max_ver:
                max_ver = ver
    return max_ver + 1


def _resolve_excel_path(excel_arg, pkg_dir):
    date_str = time.strftime("%Y%m%d", time.localtime())
    template_path = _normalize_excel_template(excel_arg, pkg_dir)
    if "{version}" in template_path:
        version = _next_version(template_path)
    else:
        version = None
    output_path = template_path
    output_path = output_path.replace("{YYYYMMDD}", date_str).replace("{date}", date_str)
    if version is not None:
        output_path = output_path.replace("{version}", "v%d" % version)
    return output_path


def _default_styles():
    base_font = Font(name="맑은 고딕", size=11)
    bold_font = Font(name="맑은 고딕", size=11, bold=True)
    header_fill = PatternFill("solid", fgColor="92D050")
    sub_fill = PatternFill("solid", fgColor="FCE4D6")
    medium = Side(style="medium")
    hair = Side(style="hair")
    return {
        "base_font": base_font,
        "bold_font": bold_font,
        "header_fill": header_fill,
        "sub_fill": sub_fill,
        "medium": medium,
        "hair": hair,
    }


def _border(left=None, right=None, top=None, bottom=None):
    return Border(left=left, right=right, top=top, bottom=bottom)


def _init_sheet(ws, sheet_name, release_path, ensure_format):
    if ensure_format:
        styles = _default_styles()
        ws.merge_cells("B2:E2")
        ws.merge_cells("B3:E3")
        if "B4:C4" not in ws.merged_cells:
            ws.merge_cells("B4:C4")
        for col in range(2, 6):
            cell = ws.cell(row=2, column=col)
            cell.fill = styles["header_fill"]
            cell.font = styles["bold_font"]
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _border(styles["medium"], styles["medium"], styles["medium"], styles["medium"])
        for col in range(2, 6):
            cell = ws.cell(row=3, column=col)
            cell.fill = styles["sub_fill"]
            cell.font = styles["base_font"]
            cell.alignment = Alignment(horizontal="center", vertical="center")
            left = styles["medium"] if col == 2 else styles["hair"]
            right = styles["medium"] if col == 5 else styles["hair"]
            cell.border = _border(left, right, styles["medium"], styles["hair"])
        for col in range(2, 6):
            cell = ws.cell(row=4, column=col)
            cell.fill = styles["sub_fill"]
            cell.font = styles["base_font"]
            cell.alignment = Alignment(horizontal="center", vertical="center")
            left = styles["medium"] if col == 2 else styles["hair"]
            right = styles["medium"] if col == 5 else styles["hair"]
            cell.border = _border(left, right, styles["hair"], styles["hair"])
        ws.row_dimensions[2].height = 17.25
        ws.column_dimensions["B"].width = 16.75
        ws.column_dimensions["C"].width = 12.75
        ws.column_dimensions["D"].width = 45.875
        ws.column_dimensions["E"].width = 12.75
        ws.column_dimensions["F"].width = 13.0
    ws.cell(row=2, column=2, value=sheet_name)
    ws.cell(row=3, column=2, value=release_path)
    ws.cell(row=4, column=2, value="Check Sum")
    ws.cell(row=4, column=4, value="File Name")
    ws.cell(row=4, column=5, value="비고")


def _apply_table_border(ws, start_row, end_row, ensure_format):
    if not ensure_format:
        return
    styles = _default_styles()
    for row in range(start_row, end_row + 1):
        for col in range(2, 6):
            cell = ws.cell(row=row, column=col)
            left = styles["medium"] if col == 2 else styles["hair"]
            right = styles["medium"] if col == 5 else styles["hair"]
            top = styles["hair"] if row >= 4 else None
            if row == end_row:
                bottom = styles["medium"]
            elif row == 4:
                bottom = styles["hair"]
            else:
                bottom = None
            cell.border = _border(left, right, top, bottom)


def _configure_page(ws, end_row, template_ws=None):
    if template_ws is not None:
        ws.print_area = template_ws.print_area
        ws.page_setup.orientation = template_ws.page_setup.orientation
        ws.page_setup.paperSize = template_ws.page_setup.paperSize
        ws.page_setup.fitToWidth = template_ws.page_setup.fitToWidth
        ws.page_setup.fitToHeight = template_ws.page_setup.fitToHeight
        ws.sheet_view.view = template_ws.sheet_view.view
        ws.sheet_view.zoomScale = template_ws.sheet_view.zoomScale
        ws.sheet_view.zoomScaleNormal = template_ws.sheet_view.zoomScaleNormal
        ws.sheet_view.showGridLines = template_ws.sheet_view.showGridLines
        return
    ws.print_area = "A1:F%d" % max(end_row, 5)
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = 9
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.sheet_view.view = "pageBreakPreview"
    ws.sheet_view.zoomScale = 100
    ws.sheet_view.zoomScaleNormal = 100
    ws.sheet_view.showGridLines = False


def _copy_style(src, dest):
    dest.font = copy(src.font)
    dest.border = copy(src.border)
    dest.fill = copy(src.fill)
    dest.number_format = copy(src.number_format)
    dest.protection = copy(src.protection)
    dest.alignment = copy(src.alignment)


def _write_sheet(ws, rows, ensure_format, template_ws=None):
    if ws.max_row >= 5:
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=2, max_col=5):
            for cell in row:
                cell.value = None
    style_row = 5 if ws.max_row >= 5 else None
    style_cells = {}
    if style_row:
        for col in range(2, 6):
            style_cells[col] = ws.cell(row=style_row, column=col)
    template_height = None
    if template_ws is not None:
        template_height = template_ws.row_dimensions[5].height
    styles = _default_styles() if ensure_format else None
    for idx, (cksum, size, path) in enumerate(rows, 5):
        b = ws.cell(row=idx, column=2, value=cksum)
        c = ws.cell(row=idx, column=3, value=size)
        d = ws.cell(row=idx, column=4, value=path)
        e = ws.cell(row=idx, column=5, value="")
        if style_cells:
            _copy_style(style_cells[2], b)
            _copy_style(style_cells[3], c)
            _copy_style(style_cells[4], d)
            _copy_style(style_cells[5], e)
        if ensure_format and styles:
            b.border = _border(styles["medium"], styles["hair"], styles["hair"], None)
            c.border = _border(styles["hair"], styles["hair"], styles["hair"], None)
            d.border = _border(styles["hair"], styles["hair"], styles["hair"], None)
            e.border = _border(styles["hair"], styles["medium"], styles["hair"], None)
            for cell in (b, c, d, e):
                cell.font = styles["base_font"]
                cell.alignment = Alignment(vertical="center")
        if template_height is not None:
            ws.row_dimensions[idx].height = template_height
    end_row = max(5, 4 + len(rows))
    _apply_table_border(ws, 4, end_row, ensure_format)
    if ensure_format:
        blank_row = end_row + 1
        for col in range(2, 6):
            cell = ws.cell(row=blank_row, column=col)
            cell.value = None
            cell.border = Border()
            cell.fill = PatternFill(fill_type=None)
    _configure_page(ws, end_row + 1, template_ws=template_ws)


def main(argv=None):
    argv = argv if argv is not None else sys.argv[1:]
    parser = argparse.ArgumentParser(description="Export cksum results into an Excel template.")
    parser.add_argument("--config", help="pkgmgr main config path")
    parser.add_argument("--pkg-id", required=True, help="pkg id (resolved via pkg_release_root)")
    parser.add_argument(
        "--excel",
        required=True,
        help="output xlsx path (supports {YYYYMMDD}/{date} and {version})",
    )
    parser.add_argument("--template", help="xlsx template path (optional)")
    args = parser.parse_args(argv)

    if openpyxl is None:
        print("[export_cksum] openpyxl is required (pip install openpyxl)")
        return 1

    pkg_dir = None
    config_path = args.config or os.environ.get("PKGMGR_CONFIG")
    if args.pkg_id:
        pkg_dir = _resolve_pkg_dir(args.pkg_id, config_path=config_path)
    pkg_yaml = _load_pkg_yaml(pkg_dir)
    if not pkg_yaml:
        print("[export_cksum] pkg.yaml not specified; use --pkg-id")
        return 1
    pkg_yaml = os.path.abspath(os.path.expanduser(pkg_yaml))
    pkg_dir = os.path.dirname(pkg_yaml)

    pkg_cfg = config.load_pkg_config(pkg_yaml)
    releases = (pkg_cfg.get("include") or {}).get("releases") or []
    grouped = _group_release_paths(pkg_dir, releases)
    if not grouped:
        print("[export_cksum] no release entries found in %s" % pkg_yaml)
        return 1

    excel_path = _resolve_excel_path(args.excel, pkg_dir)
    template_path = args.template or excel_path
    template_available = template_path and os.path.exists(template_path)
    if template_available:
        wb = openpyxl.load_workbook(template_path)
        print("[export_cksum] template loaded: %s" % template_path)
    else:
        wb = openpyxl.Workbook()
        for name in list(wb.sheetnames):
            wb.remove(wb[name])
        print("[export_cksum] template not found; creating new workbook: %s" % excel_path)
    template_ws = wb[wb.sheetnames[0]] if wb.sheetnames else None
    template_title = template_ws.title if template_ws is not None else None
    roots = []
    for root, entries in sorted(grouped.items()):
        roots.append(root)
        root_dir = os.path.join(pkg_dir, root)
        files = _collect_files(root_dir, entries)
        rows = []
        for path in files:
            if not os.path.isfile(path):
                continue
            try:
                relpath = os.path.relpath(path, root_dir)
            except Exception:
                relpath = os.path.basename(path)
            cksum_row = _cksum(path)
            if not cksum_row:
                continue
            rows.append((cksum_row[0], cksum_row[1], relpath))
        rows.sort(key=lambda r: r[2])

        ensure_format = not template_available
        if root in wb.sheetnames:
            ws = wb[root]
        elif template_ws is not None:
            ws = wb.copy_worksheet(template_ws)
            ws.title = root
        else:
            ws = wb.create_sheet(title=root)
        _init_sheet(ws, root, "PKG Release 경로 작성 필요", ensure_format)
        _write_sheet(ws, rows, ensure_format, template_ws=template_ws)
        print("[export_cksum] sheet=%s rows=%d" % (root, len(rows)))
    if template_title and template_title not in roots and template_title in wb.sheetnames:
        wb.remove(wb[template_title])

    out_path = excel_path
    if not out_path.lower().endswith(".xlsx"):
        out_path = out_path + ".xlsx"
    out_dir = os.path.dirname(os.path.abspath(out_path))
    if out_dir and not os.path.exists(out_dir):
        os.makedirs(out_dir)
    wb.save(out_path)
    print("[export_cksum] wrote %s" % out_path)
    return 0


if __name__ == "__main__":
    sys.exit(main())
