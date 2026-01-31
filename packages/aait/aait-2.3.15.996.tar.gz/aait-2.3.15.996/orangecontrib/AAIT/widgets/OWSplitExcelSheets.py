import os
import sys
import openpyxl
from openpyxl import Workbook

import Orange.data
from Orange.data import Table, Domain, StringVariable
from Orange.widgets.utils.signals import Input, Output
from Orange.widgets.settings import Setting
from AnyQt.QtWidgets import QApplication, QCheckBox, QPushButton

if "site-packages/Orange/widgets" in os.path.dirname(os.path.abspath(__file__)).replace("\\", "/"):
    from Orange.widgets.orangecontrib.AAIT.utils import base_widget, thread_management
    from Orange.widgets.orangecontrib.AAIT.utils.initialize_from_ini import apply_modification_from_python_file
else:
    from orangecontrib.AAIT.utils import base_widget, thread_management
    from orangecontrib.AAIT.utils.initialize_from_ini import apply_modification_from_python_file


@apply_modification_from_python_file(filepath_original_widget=__file__)
class OWSplitExcelSheets(base_widget.BaseListWidget):

    name = "Split Excel Sheets"
    description = "Split Excel files into one file per sheet."
    category = "AAIT - TOOLBOX"
    icon = "icons/splitexcelsheets.png"
    priority = 1070
    want_control_area = False

    gui = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "designer/owsplitexcelsheets.ui"
    )

    recursive = Setting("True")
    selected_column_name = Setting("path")

    class Inputs:
        data = Input("Data", Orange.data.Table)

    class Outputs:
        data = Output("Generated files", Orange.data.Table)

    @Inputs.data
    def set_data(self, in_data):
        self.data = in_data
        if self.data:
            self.var_selector.add_variables(self.data.domain)
            self.var_selector.select_variable_by_name(self.selected_column_name)
        if self.autorun:
            self.run()
        if in_data is None:
            self.Outputs.data.send(None)

    def __init__(self):
        super().__init__()

        self.setFixedWidth(480)
        self.setFixedHeight(500)

        self.data = None
        self.thread = None
        self.autorun = True

        self.checkBox_recursive = self.findChild(QCheckBox, "checkBox_recursive")
        self.pushButton_run = self.findChild(QPushButton, "pushButton_run")

        self.checkBox_recursive.setChecked(self.recursive == "True")
        self.checkBox_recursive.stateChanged.connect(self.on_recursive_changed)
        self.pushButton_run.clicked.connect(self.run)

        self.post_initialized()

    def on_recursive_changed(self, state):
        self.recursive = "True" if state else "False"

    def run(self):
        self.error("")
        self.warning("")

        if self.thread is not None:
            self.thread.safe_quit()

        if self.data is None or len(self.data) == 0:
            self.Outputs.data.send(None)
            return

        # Verification of in_data
        if not self.selected_column_name in self.data.domain:
            self.warning(f'Previously selected column "{self.selected_column_name}" does not exist in your data.')
            return

        if not isinstance(self.data.domain[self.selected_column_name], StringVariable):
            self.error('You must select a text variable.')
            return

        self.progressBarInit()

        self.thread = thread_management.Thread(
            self.split_excel_worker
        )

        self.thread.progress.connect(self.handle_progress)
        self.thread.result.connect(self.handle_result)
        self.thread.finish.connect(self.handle_finish)
        self.thread.start()

    def handle_progress(self, value: float):
        self.progressBarSet(value)

    def handle_result(self, result):
        try:
            if not result:
                self.Outputs.data.send(None)
                return

            domain = Domain([], metas=[StringVariable("path")])
            table = Table.from_numpy(
                domain,
                X=[[] for _ in result],
                metas=result
            )

            self.Outputs.data.send(table)

        except Exception as e:
            self.error(f"Processing error: {e}")
            self.Outputs.data.send(None)

    def handle_finish(self):
        self.progressBarFinished()

    def post_initialized(self):
        pass

    @staticmethod
    def is_excel_already_processed(filepath):
        name_no_ext = os.path.splitext(os.path.basename(filepath))[0]
        parent = os.path.dirname(filepath)
        target_dir = os.path.join(parent, f"{name_no_ext}_sheets")

        if not os.path.isdir(target_dir):
            return False

        return any(f.lower().endswith(".xlsx") for f in os.listdir(target_dir))

    def split_excel_worker(self, progress_callback):
        excel_ext = (".xlsx", ".xlsm", ".xls", ".xlsb")
        all_files = []

        # Filtre et non traitement des dossiers déjà nommé "_sheets" : siffixe que l'on ajoute quadn on crée le dossier
        # pour déterminer si le fichier a été traité
        target_files = []
        for row in self.data.metas:

            filepath = str(row[0]).replace("\\", "/")

            if "_sheets" in filepath:
                continue

            if os.path.isfile(filepath) and filepath.lower().endswith(excel_ext):
                target_files.append(filepath)

        total = len(target_files)
        if total == 0:
            return []

        # 2. Traitement des fichiers sains
        for idx, filepath in enumerate(target_files, start=1):
            name_no_ext = os.path.splitext(os.path.basename(filepath))[0]
            parent = os.path.dirname(filepath)
            target_dir = os.path.join(parent, f"{name_no_ext}_sheets")

            # Si le dossier de sortie existe déjà, on récupère son contenu pour la sortie
            # sans retraiter le fichier source.
            if self.is_excel_already_processed(filepath):
                if os.path.exists(target_dir):
                    for f in os.listdir(target_dir):
                        if f.lower().endswith(".xlsx"):
                            all_files.append([os.path.join(target_dir, f).replace("\\", "/")])
                progress_callback(idx / total * 100)
                continue

            # Traitement réel du fichier
            os.makedirs(target_dir, exist_ok=True)

            try:
                # Mode read_only=True pour économiser la RAM sur les gros fichiers
                # data_only=True pour extraire les valeurs et non les formules
                wb_source = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

                for sheet_name in wb_source.sheetnames:
                    safe_name = "".join(c if c.isalnum() or c in (' ', '-', '_') else "_" for c in sheet_name)
                    out_path = os.path.join(target_dir, f"{safe_name}.xlsx")

                    # Création du nouveau classeur pour la feuille individuelle
                    wb_dest = Workbook()
                    ws_dest = wb_dest.active
                    ws_dest.title = sheet_name

                    ws_source = wb_source[sheet_name]

                    has_data = False
                    # On itère sur les lignes de la source
                    for row in ws_source.rows:
                        row_values = [cell.value for cell in row]
                        # On vérifie si la ligne contient au moins une donnée
                        if any(v is not None for v in row_values):
                            ws_dest.append(row_values)
                            has_data = True

                    if has_data:
                        wb_dest.save(out_path)
                        all_files.append([out_path.replace("\\", "/")])

                wb_source.close()

            except Exception as e:
                print(f"Skipping {filepath} due to error: {e}")
                continue

            progress_callback(idx / total * 100)

        return all_files


if __name__ == "__main__":
    app = QApplication(sys.argv)
    w = OWSplitExcelSheets()
    w.show()
    app.exec()