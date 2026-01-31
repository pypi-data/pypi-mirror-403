"""Shared SQL storage helpers for multiple adapters."""

from __future__ import annotations

import json
from abc import ABC, abstractmethod
from collections.abc import Mapping, Sequence
from contextlib import AbstractContextManager, closing
from datetime import datetime
from pathlib import Path
from typing import Any, cast

from evalvault.domain.entities import (
    EvaluationRun,
    FeedbackSummary,
    MetricScore,
    MultiTurnConversationRecord,
    MultiTurnRunRecord,
    MultiTurnTurnResult,
    RunClusterMap,
    RunClusterMapInfo,
    SatisfactionFeedback,
    TestCaseResult,
)


class SQLQueries:
    """Generates SQL statements with adapter-specific placeholders."""

    def __init__(
        self,
        *,
        placeholder: str = "?",
        metric_name_column: str = "metric_name",
        test_case_returning_clause: str = "",
        feedback_returning_clause: str = "",
    ) -> None:
        self.placeholder = placeholder
        self.metric_name_column = metric_name_column
        self._test_case_returning = test_case_returning_clause
        self._feedback_returning = feedback_returning_clause

    def _values(self, count: int) -> str:
        return ", ".join([self.placeholder] * count)

    def insert_run(self) -> str:
        values = self._values(14)
        return f"""
        INSERT INTO evaluation_runs (
            run_id, dataset_name, dataset_version, model_name,
            started_at, finished_at, total_tokens, total_cost_usd,
            pass_rate, metrics_evaluated, thresholds, langfuse_trace_id,
            metadata, retrieval_metadata
        ) VALUES ({values})
        """

    def insert_test_case(self) -> str:
        values = self._values(12)
        query = f"""
        INSERT INTO test_case_results (
            run_id, test_case_id, tokens_used, latency_ms,
            cost_usd, trace_id, started_at, finished_at,
            question, answer, contexts, ground_truth
        ) VALUES ({values})
        """
        if self._test_case_returning:
            query = f"{query.strip()} {self._test_case_returning}"
        return query

    def insert_metric_score(self) -> str:
        values = self._values(5)
        return f"""
        INSERT INTO metric_scores (
            result_id, {self.metric_name_column}, score, threshold, reason
        ) VALUES ({values})
        """

    def insert_multiturn_run(self) -> str:
        values = self._values(12)
        return f"""
        INSERT INTO multiturn_runs (
            run_id, dataset_name, dataset_version, model_name,
            started_at, finished_at, conversation_count, turn_count,
            metrics_evaluated, drift_threshold, summary, metadata
        ) VALUES ({values})
        """

    def insert_multiturn_conversation(self) -> str:
        values = self._values(7)
        return f"""
        INSERT INTO multiturn_conversations (
            run_id, conversation_id, turn_count, drift_score, drift_threshold,
            drift_detected, summary
        ) VALUES ({values})
        """

    def insert_multiturn_turn(self) -> str:
        values = self._values(8)
        query = f"""
        INSERT INTO multiturn_turn_results (
            run_id, conversation_id, turn_id, turn_index, role,
            passed, latency_ms, metadata
        ) VALUES ({values})
        """
        if self._test_case_returning:
            query = f"{query.strip()} {self._test_case_returning}"
        return query

    def insert_multiturn_metric_score(self) -> str:
        values = self._values(4)
        return f"""
        INSERT INTO multiturn_metric_scores (
            turn_result_id, metric_name, score, threshold
        ) VALUES ({values})
        """

    def insert_cluster_map(self) -> str:
        values = self._values(7)
        return f"""
        INSERT INTO run_cluster_maps (
            run_id, map_id, test_case_id, cluster_id, source, metadata, created_at
        ) VALUES ({values})
        """

    def insert_feedback(self) -> str:
        values = self._values(7)
        query = f"""
        INSERT INTO satisfaction_feedback (
            run_id, test_case_id, satisfaction_score, thumb_feedback, comment, rater_id, created_at
        ) VALUES ({values})
        """
        if self._feedback_returning:
            query = f"{query.strip()} {self._feedback_returning}"
        return query

    def select_feedback_by_run(self) -> str:
        return f"""
        SELECT id, run_id, test_case_id, satisfaction_score, thumb_feedback, comment, rater_id, created_at
        FROM satisfaction_feedback
        WHERE run_id = {self.placeholder}
        ORDER BY created_at DESC
        """

    def select_run(self) -> str:
        return f"""
        SELECT run_id, dataset_name, dataset_version, model_name,
               started_at, finished_at, total_tokens, total_cost_usd,
               pass_rate, metrics_evaluated, thresholds, langfuse_trace_id,
               metadata, retrieval_metadata
        FROM evaluation_runs
        WHERE run_id = {self.placeholder}
        """

    def select_test_case_results(self) -> str:
        return f"""
        SELECT id, test_case_id, tokens_used, latency_ms, cost_usd,
               trace_id, started_at, finished_at, question, answer,
               contexts, ground_truth
        FROM test_case_results
        WHERE run_id = {self.placeholder}
        ORDER BY id
        """

    def select_metric_scores(self) -> str:
        return f"""
        SELECT {self.metric_name_column} AS metric_name, score, threshold, reason
        FROM metric_scores
        WHERE result_id = {self.placeholder}
        ORDER BY id
        """

    def select_multiturn_run(self) -> str:
        return f"""
        SELECT run_id, dataset_name, dataset_version, model_name,
               started_at, finished_at, conversation_count, turn_count,
               metrics_evaluated, drift_threshold, summary, metadata, created_at
        FROM multiturn_runs
        WHERE run_id = {self.placeholder}
        """

    def select_multiturn_conversations(self) -> str:
        return f"""
        SELECT run_id, conversation_id, turn_count, drift_score, drift_threshold,
               drift_detected, summary
        FROM multiturn_conversations
        WHERE run_id = {self.placeholder}
        ORDER BY id
        """

    def select_multiturn_turn_results(self) -> str:
        return f"""
        SELECT id, run_id, conversation_id, turn_id, turn_index, role,
               passed, latency_ms, metadata
        FROM multiturn_turn_results
        WHERE run_id = {self.placeholder}
        ORDER BY id
        """

    def select_multiturn_metric_scores(self) -> str:
        return f"""
        SELECT turn_result_id, metric_name, score, threshold
        FROM multiturn_metric_scores
        WHERE turn_result_id = {self.placeholder}
        ORDER BY id
        """

    def select_cluster_map(self) -> str:
        return f"""
        SELECT test_case_id, cluster_id, source, map_id, created_at, metadata
        FROM run_cluster_maps
        WHERE run_id = {self.placeholder} AND map_id = {self.placeholder}
        ORDER BY test_case_id
        """

    def select_cluster_map_latest(self) -> str:
        return f"""
        SELECT map_id, source, created_at
        FROM run_cluster_maps
        WHERE run_id = {self.placeholder}
        ORDER BY created_at DESC
        LIMIT 1
        """

    def select_cluster_map_sets(self) -> str:
        return f"""
        SELECT map_id, source, created_at, COUNT(*) AS item_count
        FROM run_cluster_maps
        WHERE run_id = {self.placeholder}
        GROUP BY map_id, source, created_at
        ORDER BY created_at DESC
        """

    def update_run_metadata(self) -> str:
        return f"""
        UPDATE evaluation_runs
        SET metadata = {self.placeholder}
        WHERE run_id = {self.placeholder}
        """

    def delete_run(self) -> str:
        return f"DELETE FROM evaluation_runs WHERE run_id = {self.placeholder}"

    def delete_cluster_map(self) -> str:
        return f"DELETE FROM run_cluster_maps WHERE run_id = {self.placeholder} AND map_id = {self.placeholder}"

    def list_runs_base(self) -> str:
        return "SELECT run_id FROM evaluation_runs WHERE 1=1"

    def list_runs_ordering(self) -> str:
        return f" ORDER BY started_at DESC LIMIT {self.placeholder} OFFSET {self.placeholder}"

    def upsert_regression_baseline(self) -> str:
        raise NotImplementedError("Override in subclass")

    def select_regression_baseline(self) -> str:
        return f"""
        SELECT baseline_key, run_id, dataset_name, branch, commit_sha, metadata,
               created_at, updated_at
        FROM regression_baselines
        WHERE baseline_key = {self.placeholder}
        """


class BaseSQLStorageAdapter(ABC):
    """Shared serialization and SQL helpers for DB-API based adapters."""

    def __init__(self, queries: SQLQueries) -> None:
        self.queries = queries

    # Connection helpers -------------------------------------------------

    @abstractmethod
    def _connect(self) -> Any:
        """Return a new DB-API compatible connection."""
        raise NotImplementedError

    def _get_connection(self) -> AbstractContextManager[Any]:
        conn = self._connect()
        if conn is None:
            raise RuntimeError("Database connection not available")
        return closing(conn)

    def _fetch_lastrowid(self, cursor) -> int:
        return cursor.lastrowid

    def _execute(
        self,
        conn: Any,
        query: str,
        params: Sequence[object] | Mapping[str, object] | None = None,
    ) -> Any:
        if params is None:
            return conn.execute(query)
        return conn.execute(query, params)

    # CRUD helpers -------------------------------------------------------

    def save_run(self, run: EvaluationRun) -> str:
        with self._get_connection() as conn:
            self._execute(conn, self.queries.insert_run(), self._run_params(run))

            for result in run.results:
                result_id = self._insert_test_case(conn, run.run_id, result)
                for metric in result.metrics:
                    self._execute(
                        conn,
                        self.queries.insert_metric_score(),
                        self._metric_params(result_id, metric),
                    )

            conn.commit()
            return run.run_id

    def save_multiturn_run(
        self,
        run: MultiTurnRunRecord,
        conversations: list[MultiTurnConversationRecord],
        turn_results: list[MultiTurnTurnResult],
        *,
        metric_thresholds: dict[str, float] | None = None,
    ) -> str:
        with self._get_connection() as conn:
            self._execute(
                conn, self.queries.insert_multiturn_run(), self._multiturn_run_params(run)
            )

            for conversation in conversations:
                self._execute(
                    conn,
                    self.queries.insert_multiturn_conversation(),
                    self._multiturn_conversation_params(conversation),
                )

            for turn in turn_results:
                cursor = self._execute(
                    conn,
                    self.queries.insert_multiturn_turn(),
                    self._multiturn_turn_params(run.run_id, turn),
                )
                turn_result_id = self._fetch_lastrowid(cursor)
                for metric_name, score in (turn.metrics or {}).items():
                    threshold = None
                    if metric_thresholds and metric_name in metric_thresholds:
                        threshold = metric_thresholds[metric_name]
                    self._execute(
                        conn,
                        self.queries.insert_multiturn_metric_score(),
                        self._multiturn_metric_params(
                            turn_result_id, metric_name, score, threshold
                        ),
                    )

            conn.commit()
            return run.run_id

    def _insert_test_case(self, conn, run_id: str, result: TestCaseResult) -> int:
        cursor = self._execute(
            conn,
            self.queries.insert_test_case(),
            self._test_case_params(run_id, result),
        )
        return self._fetch_lastrowid(cursor)

    def get_run(self, run_id: str) -> EvaluationRun:
        with self._get_connection() as conn:
            cursor = self._execute(conn, self.queries.select_run(), (run_id,))
            run_row = cursor.fetchone()
            if not run_row:
                raise KeyError(f"Run not found: {run_id}")

            result_rows = self._execute(
                conn, self.queries.select_test_case_results(), (run_id,)
            ).fetchall()

            results = [self._row_to_test_case(conn, row) for row in result_rows]

            return EvaluationRun(
                run_id=run_row["run_id"],
                dataset_name=run_row["dataset_name"],
                dataset_version=run_row["dataset_version"],
                model_name=run_row["model_name"],
                started_at=self._deserialize_datetime(run_row["started_at"]) or datetime.now(),
                finished_at=self._deserialize_datetime(run_row["finished_at"]),
                total_tokens=run_row["total_tokens"],
                total_cost_usd=self._maybe_float(run_row["total_cost_usd"]),
                results=results,
                metrics_evaluated=self._deserialize_json(run_row["metrics_evaluated"]) or [],
                thresholds=self._deserialize_json(run_row["thresholds"]) or {},
                langfuse_trace_id=run_row["langfuse_trace_id"],
                tracker_metadata=self._deserialize_json(run_row["metadata"]) or {},
                retrieval_metadata=self._deserialize_json(run_row["retrieval_metadata"]) or {},
            )

    def list_runs(
        self,
        limit: int = 100,
        offset: int = 0,
        dataset_name: str | None = None,
        model_name: str | None = None,
    ) -> list[EvaluationRun]:
        with self._get_connection() as conn:
            query = self.queries.list_runs_base()
            params: list[Any] = []

            if dataset_name:
                query += f" AND dataset_name = {self.queries.placeholder}"
                params.append(dataset_name)

            if model_name:
                query += f" AND model_name = {self.queries.placeholder}"
                params.append(model_name)

            query += self.queries.list_runs_ordering()
            params.extend([limit, offset])

            cursor = self._execute(conn, query, params)
            run_ids = [row["run_id"] for row in cursor.fetchall()]

        return [self.get_run(run_id) for run_id in run_ids]

    def delete_run(self, run_id: str) -> bool:
        with self._get_connection() as conn:
            cursor = self._execute(conn, self.queries.delete_run(), (run_id,))
            deleted = (cursor.rowcount or 0) > 0
            conn.commit()
            return deleted

    def update_run_metadata(self, run_id: str, metadata: dict[str, Any]) -> None:
        payload = self._serialize_json(metadata)
        with self._get_connection() as conn:
            self._execute(conn, self.queries.update_run_metadata(), (payload, run_id))
            conn.commit()

    def save_run_cluster_map(
        self,
        run_id: str,
        mapping: dict[str, str],
        source: str | None = None,
        map_id: str | None = None,
        metadata: dict[str, Any] | None = None,
    ) -> str:
        if not mapping:
            raise ValueError("Cluster map is empty")
        if map_id is None:
            from uuid import uuid4

            map_id = str(uuid4())
        created_at = self._serialize_datetime(datetime.now())
        metadata_payload = self._serialize_json(metadata)
        with self._get_connection() as conn:
            self._execute(conn, self.queries.delete_cluster_map(), (run_id, map_id))
            for test_case_id, cluster_id in mapping.items():
                self._execute(
                    conn,
                    self.queries.insert_cluster_map(),
                    (
                        run_id,
                        map_id,
                        test_case_id,
                        cluster_id,
                        source,
                        metadata_payload,
                        created_at,
                    ),
                )
            conn.commit()
        return map_id

    def get_run_cluster_map(self, run_id: str, map_id: str | None = None) -> RunClusterMap | None:
        with self._get_connection() as conn:
            source: str | None = None
            created_at: datetime | None = None
            metadata: dict[str, Any] | None = None
            if map_id is None:
                latest_row = self._execute(
                    conn, self.queries.select_cluster_map_latest(), (run_id,)
                ).fetchone()
                if not latest_row:
                    return None
                map_id = self._row_value(latest_row, "map_id")
                source = self._row_value(latest_row, "source")
                created_at = self._deserialize_datetime(self._row_value(latest_row, "created_at"))
                if map_id is None:
                    return None

            rows = self._execute(
                conn, self.queries.select_cluster_map(), (run_id, map_id)
            ).fetchall()
            if not rows:
                return None

            mapping: dict[str, str] = {}
            for row in rows:
                test_case_id = self._row_value(row, "test_case_id")
                cluster_id = self._row_value(row, "cluster_id")
                row_source = self._row_value(row, "source")
                row_created_at = self._row_value(row, "created_at")
                row_metadata = self._row_value(row, "metadata")
                if row_source and not source:
                    source = row_source
                if row_created_at and created_at is None:
                    created_at = self._deserialize_datetime(row_created_at)
                if metadata is None and row_metadata not in (None, ""):
                    metadata = self._deserialize_json(row_metadata)
                if test_case_id and cluster_id:
                    mapping[str(test_case_id)] = str(cluster_id)

            if not mapping:
                return None
            return RunClusterMap(
                map_id=str(map_id),
                mapping=mapping,
                source=source,
                created_at=created_at,
                metadata=metadata,
            )

    def list_run_cluster_maps(self, run_id: str) -> list[RunClusterMapInfo]:
        with self._get_connection() as conn:
            rows = self._execute(conn, self.queries.select_cluster_map_sets(), (run_id,)).fetchall()
            results: list[RunClusterMapInfo] = []
            for row in rows:
                results.append(
                    RunClusterMapInfo(
                        map_id=str(self._row_value(row, "map_id") or ""),
                        source=self._row_value(row, "source"),
                        created_at=self._deserialize_datetime(self._row_value(row, "created_at")),
                        item_count=int(self._row_value(row, "item_count") or 0),
                    )
                )
            return results

    def delete_run_cluster_map(self, run_id: str, map_id: str) -> int:
        with self._get_connection() as conn:
            cursor = self._execute(conn, self.queries.delete_cluster_map(), (run_id, map_id))
            deleted = cursor.rowcount if cursor.rowcount is not None else 0
            conn.commit()
            return deleted

    def save_feedback(self, feedback: SatisfactionFeedback) -> str:
        created_at = feedback.created_at or datetime.now()
        with self._get_connection() as conn:
            cursor = self._execute(
                conn,
                self.queries.insert_feedback(),
                (
                    feedback.run_id,
                    feedback.test_case_id,
                    feedback.satisfaction_score,
                    feedback.thumb_feedback,
                    feedback.comment,
                    feedback.rater_id,
                    self._serialize_datetime(created_at),
                ),
            )
            feedback_id = self._fetch_lastrowid(cursor)
            conn.commit()
            return str(feedback_id)

    def list_feedback(self, run_id: str) -> list[SatisfactionFeedback]:
        with self._get_connection() as conn:
            rows = self._execute(conn, self.queries.select_feedback_by_run(), (run_id,)).fetchall()
            return [self._row_to_feedback(row) for row in rows]

    def get_feedback_summary(self, run_id: str) -> FeedbackSummary:
        feedbacks = self.list_feedback(run_id)
        latest: dict[tuple[str, str | None], SatisfactionFeedback] = {}
        for feedback in feedbacks:
            key = (feedback.test_case_id, feedback.rater_id)
            current = latest.get(key)
            if current is None:
                latest[key] = feedback
                continue
            current_time = current.created_at or datetime.min
            feedback_time = feedback.created_at or datetime.min
            if feedback_time >= current_time:
                latest[key] = feedback

        effective = [
            feedback
            for feedback in latest.values()
            if feedback.satisfaction_score is not None or feedback.thumb_feedback in {"up", "down"}
        ]
        scores = [
            feedback.satisfaction_score
            for feedback in effective
            if feedback.satisfaction_score is not None
        ]
        thumbs = [
            feedback.thumb_feedback
            for feedback in effective
            if feedback.thumb_feedback in {"up", "down"}
        ]
        avg_score = sum(scores) / len(scores) if scores else None
        thumb_up_rate = None
        if thumbs:
            thumb_up_rate = thumbs.count("up") / len(thumbs)
        return FeedbackSummary(
            avg_satisfaction_score=avg_score,
            thumb_up_rate=thumb_up_rate,
            total_feedback=len(effective),
        )

    def set_regression_baseline(
        self,
        baseline_key: str,
        run_id: str,
        *,
        dataset_name: str | None = None,
        branch: str | None = None,
        commit_sha: str | None = None,
        metadata: dict[str, Any] | None = None,
    ) -> None:
        now = self._serialize_datetime(datetime.now())
        with self._get_connection() as conn:
            self._execute(
                conn,
                self.queries.upsert_regression_baseline(),
                (
                    baseline_key,
                    run_id,
                    dataset_name,
                    branch,
                    commit_sha,
                    self._serialize_json(metadata),
                    now,
                    now,
                ),
            )
            conn.commit()

    def get_regression_baseline(self, baseline_key: str) -> dict[str, Any] | None:
        with self._get_connection() as conn:
            row = self._execute(
                conn,
                self.queries.select_regression_baseline(),
                (baseline_key,),
            ).fetchone()
            if not row:
                return None
            return {
                "baseline_key": self._row_value(row, "baseline_key"),
                "run_id": str(self._row_value(row, "run_id")),
                "dataset_name": self._row_value(row, "dataset_name"),
                "branch": self._row_value(row, "branch"),
                "commit_sha": self._row_value(row, "commit_sha"),
                "metadata": self._deserialize_json(self._row_value(row, "metadata")),
                "created_at": self._row_value(row, "created_at"),
                "updated_at": self._row_value(row, "updated_at"),
            }

    # Serialization helpers --------------------------------------------

    def _run_params(self, run: EvaluationRun) -> Sequence[Any]:
        return (
            run.run_id,
            run.dataset_name,
            run.dataset_version,
            run.model_name,
            self._serialize_datetime(run.started_at),
            self._serialize_datetime(run.finished_at),
            run.total_tokens,
            run.total_cost_usd,
            run.pass_rate,
            self._serialize_json(run.metrics_evaluated),
            self._serialize_json(run.thresholds),
            run.langfuse_trace_id,
            self._serialize_json(run.tracker_metadata),
            self._serialize_json(run.retrieval_metadata),
        )

    def _test_case_params(self, run_id: str, result: TestCaseResult) -> Sequence[Any]:
        return (
            run_id,
            result.test_case_id,
            result.tokens_used,
            result.latency_ms,
            result.cost_usd,
            result.trace_id,
            self._serialize_datetime(result.started_at),
            self._serialize_datetime(result.finished_at),
            result.question,
            result.answer,
            self._serialize_contexts(result.contexts),
            result.ground_truth,
        )

    def _metric_params(self, result_id: int, metric: MetricScore) -> Sequence[Any]:
        return (
            result_id,
            metric.name,
            metric.score,
            metric.threshold,
            metric.reason,
        )

    def _multiturn_run_params(self, run: MultiTurnRunRecord) -> Sequence[Any]:
        return (
            run.run_id,
            run.dataset_name,
            run.dataset_version,
            run.model_name,
            self._serialize_datetime(run.started_at),
            self._serialize_datetime(run.finished_at),
            run.conversation_count,
            run.turn_count,
            self._serialize_json(run.metrics_evaluated),
            run.drift_threshold,
            self._serialize_json(run.summary),
            self._serialize_json(run.metadata),
        )

    def _multiturn_conversation_params(
        self, conversation: MultiTurnConversationRecord
    ) -> Sequence[Any]:
        return (
            conversation.run_id,
            conversation.conversation_id,
            conversation.turn_count,
            conversation.drift_score,
            conversation.drift_threshold,
            int(conversation.drift_detected),
            self._serialize_json(conversation.summary),
        )

    def _multiturn_turn_params(self, run_id: str, turn: MultiTurnTurnResult) -> Sequence[Any]:
        return (
            run_id,
            turn.conversation_id,
            turn.turn_id,
            turn.turn_index,
            turn.role,
            int(turn.passed),
            turn.latency_ms,
            self._serialize_json(turn.metadata),
        )

    def _multiturn_metric_params(
        self,
        turn_result_id: int,
        metric_name: str,
        score: float,
        threshold: float | None,
    ) -> Sequence[Any]:
        return (turn_result_id, metric_name, score, threshold)

    def _row_to_test_case(self, conn, row) -> TestCaseResult:
        result_id = row["id"]
        metrics = self._fetch_metric_scores(conn, result_id)
        return TestCaseResult(
            test_case_id=row["test_case_id"],
            metrics=metrics,
            tokens_used=row["tokens_used"],
            latency_ms=row["latency_ms"],
            cost_usd=self._maybe_float(row["cost_usd"]),
            trace_id=row["trace_id"],
            started_at=self._deserialize_datetime(row["started_at"]) or datetime.now(),
            finished_at=self._deserialize_datetime(row["finished_at"]),
            question=row["question"],
            answer=row["answer"],
            contexts=self._deserialize_contexts(row["contexts"]),
            ground_truth=row["ground_truth"],
        )

    def _row_to_feedback(self, row) -> SatisfactionFeedback:
        feedback_id = self._row_value(row, "id")
        run_id = self._row_value(row, "run_id")
        test_case_id = self._row_value(row, "test_case_id")
        created_at = self._deserialize_datetime(self._row_value(row, "created_at"))
        return SatisfactionFeedback(
            feedback_id=str(feedback_id or ""),
            run_id=str(run_id or ""),
            test_case_id=str(test_case_id or ""),
            satisfaction_score=self._maybe_float(self._row_value(row, "satisfaction_score")),
            thumb_feedback=self._row_value(row, "thumb_feedback"),
            comment=self._row_value(row, "comment"),
            rater_id=self._row_value(row, "rater_id"),
            created_at=created_at,
        )

    def _fetch_metric_scores(self, conn, result_id: int) -> list[MetricScore]:
        rows = self._execute(conn, self.queries.select_metric_scores(), (result_id,)).fetchall()
        metric_column = self.queries.metric_name_column
        return [
            MetricScore(
                name=self._resolve_metric_name(row, metric_column),
                score=self._maybe_float(self._row_value(row, "score")) or 0.0,
                threshold=self._maybe_float(self._row_value(row, "threshold")) or 0.7,
                reason=self._row_value(row, "reason"),
            )
            for row in rows
        ]

    def _resolve_metric_name(self, row, fallback_column: str) -> str:
        name = self._row_value(row, "metric_name")
        if name is None and fallback_column != "metric_name":
            name = self._row_value(row, fallback_column)
        return name or ""

    def _serialize_datetime(self, value: datetime | None) -> str | None:
        return value.isoformat() if value else None

    def _deserialize_datetime(self, value: Any) -> datetime | None:
        if value is None:
            return None
        if isinstance(value, datetime):
            return value
        return datetime.fromisoformat(value)

    def _serialize_json(self, value: Any) -> str | None:
        if value is None:
            return None
        return json.dumps(value, ensure_ascii=False)

    def _deserialize_json(self, value: Any) -> Any:
        if value in (None, ""):
            return None
        if isinstance(value, str):
            return json.loads(value)
        return value

    def _serialize_contexts(self, contexts: list[str] | None) -> str | None:
        if not contexts:
            return None
        return json.dumps(contexts, ensure_ascii=False)

    def _deserialize_contexts(self, value: Any) -> list[str] | None:
        data = self._deserialize_json(value)
        if data is None:
            return None
        if isinstance(data, list):
            return data
        return [data]

    def _maybe_float(self, value: Any) -> float | None:
        if value is None:
            return None
        return float(value)

    def _row_value(self, row: Any, key: str) -> Any:
        if isinstance(row, dict):
            return row.get(key)
        try:
            return row[key]
        except (KeyError, TypeError, IndexError):
            return None

    def _row_to_mapping(self, row: Any) -> dict[str, Any]:
        if row is None:
            return {}
        if isinstance(row, dict):
            return dict(row)
        if hasattr(row, "keys"):
            keys = row.keys()
            return {key: row[key] for key in keys}
        try:
            return dict(row)
        except Exception:
            return {}

    def _coerce_excel_value(self, value: Any, *, force_json: bool = False) -> Any:
        if force_json:
            payload = self._deserialize_json(value)
            if payload is None:
                return None
            return json.dumps(payload, ensure_ascii=False)
        if isinstance(value, (dict, list)):
            return json.dumps(value, ensure_ascii=False)
        if isinstance(value, datetime):
            return value.isoformat()
        if isinstance(value, bytes):
            return value.decode("utf-8", errors="replace")
        return value

    def _normalize_rows(
        self,
        rows: Sequence[Any],
        *,
        json_columns: set[str] | None = None,
    ) -> list[dict[str, Any]]:
        json_columns = json_columns or set()
        normalized: list[dict[str, Any]] = []
        for row in rows:
            payload = self._row_to_mapping(row)
            for key, value in payload.items():
                payload[key] = self._coerce_excel_value(
                    value,
                    force_json=key in json_columns,
                )
            normalized.append(payload)
        return normalized

    def export_run_to_excel(self, run_id: str, output_path) -> Path:
        from openpyxl import Workbook

        from evalvault.domain.metrics.registry import get_metric_spec_map

        output = Path(output_path)
        output.parent.mkdir(parents=True, exist_ok=True)

        placeholder = self.queries.placeholder

        with self._get_connection() as conn:
            run_row = self._execute(conn, self.queries.select_run(), (run_id,)).fetchone()
            if not run_row:
                raise KeyError(f"Run not found: {run_id}")

            run_rows = self._normalize_rows(
                [run_row],
                json_columns={"metrics_evaluated", "thresholds", "metadata", "retrieval_metadata"},
            )

            test_case_rows = self._execute(
                conn,
                (
                    "SELECT id, run_id, test_case_id, tokens_used, latency_ms, cost_usd, trace_id, "
                    "started_at, finished_at, question, answer, contexts, ground_truth "
                    f"FROM test_case_results WHERE run_id = {placeholder} ORDER BY id"
                ),
                (run_id,),
            ).fetchall()
            test_case_payloads = self._normalize_rows(
                test_case_rows,
                json_columns={"contexts"},
            )

            metric_rows = self._execute(
                conn,
                (
                    "SELECT m.result_id, t.test_case_id, m."
                    f"{self.queries.metric_name_column} AS metric_name, m.score, m.threshold, m.reason "
                    "FROM metric_scores m JOIN test_case_results t ON m.result_id = t.id "
                    f"WHERE t.run_id = {placeholder} ORDER BY m.id"
                ),
                (run_id,),
            ).fetchall()
            metric_payloads = self._normalize_rows(metric_rows)

            run_prompt_rows = self._execute(
                conn,
                (
                    "SELECT run_id, prompt_set_id, created_at FROM run_prompt_sets "
                    f"WHERE run_id = {placeholder} ORDER BY created_at DESC"
                ),
                (run_id,),
            ).fetchall()
            run_prompt_payloads = self._normalize_rows(run_prompt_rows)
            prompt_set_ids = [row.get("prompt_set_id") for row in run_prompt_payloads if row]

            prompt_sets_payloads: list[dict[str, Any]] = []
            prompt_set_item_payloads: list[dict[str, Any]] = []
            prompt_payloads: list[dict[str, Any]] = []

            if prompt_set_ids:
                placeholders = ", ".join([placeholder] * len(prompt_set_ids))
                prompt_set_rows = self._execute(
                    conn,
                    (
                        "SELECT prompt_set_id, name, description, metadata, created_at "
                        f"FROM prompt_sets WHERE prompt_set_id IN ({placeholders})"
                    ),
                    prompt_set_ids,
                ).fetchall()
                prompt_sets_payloads = self._normalize_rows(
                    prompt_set_rows,
                    json_columns={"metadata"},
                )

                item_rows = self._execute(
                    conn,
                    (
                        "SELECT id, prompt_set_id, prompt_id, role, item_order, metadata "
                        f"FROM prompt_set_items WHERE prompt_set_id IN ({placeholders})"
                    ),
                    prompt_set_ids,
                ).fetchall()
                prompt_set_item_payloads = self._normalize_rows(
                    item_rows,
                    json_columns={"metadata"},
                )

                prompt_ids = [row.get("prompt_id") for row in prompt_set_item_payloads if row]
                if prompt_ids:
                    prompt_placeholders = ", ".join([placeholder] * len(prompt_ids))
                    prompt_rows = self._execute(
                        conn,
                        (
                            "SELECT prompt_id, name, kind, content, checksum, source, notes, metadata, created_at "
                            f"FROM prompts WHERE prompt_id IN ({prompt_placeholders})"
                        ),
                        prompt_ids,
                    ).fetchall()
                    prompt_payloads = self._normalize_rows(
                        prompt_rows,
                        json_columns={"metadata"},
                    )

            feedback_rows = self._execute(
                conn,
                (
                    "SELECT id, run_id, test_case_id, satisfaction_score, thumb_feedback, comment, rater_id, created_at "
                    f"FROM satisfaction_feedback WHERE run_id = {placeholder} ORDER BY created_at DESC"
                ),
                (run_id,),
            ).fetchall()
            feedback_payloads = self._normalize_rows(feedback_rows)

            cluster_rows = self._execute(
                conn,
                (
                    "SELECT run_id, map_id, test_case_id, cluster_id, source, metadata, created_at "
                    f"FROM run_cluster_maps WHERE run_id = {placeholder} ORDER BY created_at DESC"
                ),
                (run_id,),
            ).fetchall()
            cluster_payloads = self._normalize_rows(cluster_rows, json_columns={"metadata"})

            stage_event_rows = self._execute(
                conn,
                (
                    "SELECT id, run_id, stage_id, parent_stage_id, stage_type, stage_name, status, "
                    "attempt, started_at, finished_at, duration_ms, input_ref, output_ref, attributes, "
                    "metadata, trace_id, span_id FROM stage_events "
                    f"WHERE run_id = {placeholder} ORDER BY id"
                ),
                (run_id,),
            ).fetchall()
            stage_event_payloads = self._normalize_rows(
                stage_event_rows,
                json_columns={"attributes", "metadata"},
            )

            stage_metric_rows = self._execute(
                conn,
                (
                    "SELECT id, run_id, stage_id, metric_name, score, threshold, evidence "
                    f"FROM stage_metrics WHERE run_id = {placeholder} ORDER BY id"
                ),
                (run_id,),
            ).fetchall()
            stage_metric_payloads = self._normalize_rows(
                stage_metric_rows, json_columns={"evidence"}
            )

            report_rows = self._execute(
                conn,
                (
                    "SELECT report_id, run_id, experiment_id, report_type, format, content, metadata, created_at "
                    f"FROM analysis_reports WHERE run_id = {placeholder} ORDER BY created_at DESC"
                ),
                (run_id,),
            ).fetchall()
            report_payloads = self._normalize_rows(report_rows, json_columns={"metadata"})

            pipeline_rows = self._execute(
                conn,
                (
                    "SELECT result_id, intent, query, run_id, pipeline_id, profile, tags, metadata, "
                    "is_complete, duration_ms, final_output, node_results, started_at, finished_at, created_at "
                    f"FROM pipeline_results WHERE run_id = {placeholder} ORDER BY created_at DESC"
                ),
                (run_id,),
            ).fetchall()
            pipeline_payloads = self._normalize_rows(
                pipeline_rows,
                json_columns={"tags", "metadata", "final_output", "node_results"},
            )

        summary_rows: list[dict[str, Any]] = []
        run_payload = run_rows[0] if run_rows else {}
        custom_metric_rows: list[dict[str, Any]] = []
        run_metadata = self._deserialize_json(run_payload.get("metadata")) if run_payload else None
        if isinstance(run_metadata, dict):
            custom_snapshot = run_metadata.get("custom_metric_snapshot")
            if isinstance(custom_snapshot, dict):
                entries = custom_snapshot.get("metrics")
                if isinstance(entries, list):
                    for entry in entries:
                        if isinstance(entry, dict):
                            row = dict(entry)
                            row["schema_version"] = custom_snapshot.get("schema_version")
                            custom_metric_rows.append(row)
        if custom_metric_rows:
            custom_metric_rows = self._normalize_rows(
                custom_metric_rows,
                json_columns={"inputs", "rules"},
            )
        prompt_set_id = None
        prompt_set_name = None
        if run_prompt_payloads:
            prompt_set_id = run_prompt_payloads[0].get("prompt_set_id")
        if prompt_sets_payloads:
            prompt_set_name = prompt_sets_payloads[0].get("name")
        summary_rows.append(
            {
                "run_id": run_payload.get("run_id"),
                "dataset_name": run_payload.get("dataset_name"),
                "model_name": run_payload.get("model_name"),
                "started_at": run_payload.get("started_at"),
                "finished_at": run_payload.get("finished_at"),
                "total_test_cases": len(test_case_payloads),
                "total_tokens": run_payload.get("total_tokens"),
                "total_cost_usd": run_payload.get("total_cost_usd"),
                "pass_rate": run_payload.get("pass_rate"),
                "metrics_evaluated": run_payload.get("metrics_evaluated"),
                "prompt_set_id": prompt_set_id,
                "prompt_set_name": prompt_set_name,
            }
        )

        metric_summary_rows: list[dict[str, Any]] = []
        metrics_index: dict[str, dict[str, Any]] = {}
        for row in metric_payloads:
            metric_name = row.get("metric_name")
            if not metric_name:
                continue
            entry = metrics_index.setdefault(
                metric_name,
                {"metric_name": metric_name, "count": 0, "score_sum": 0.0, "pass_count": 0},
            )
            score = row.get("score")
            threshold = row.get("threshold")
            if isinstance(score, (int, float)):
                entry["count"] += 1
                entry["score_sum"] += float(score)
                if isinstance(threshold, (int, float)) and score >= threshold:
                    entry["pass_count"] += 1

        metric_spec_map = get_metric_spec_map()
        for entry in metrics_index.values():
            count = entry["count"] or 0
            spec = metric_spec_map.get(entry["metric_name"])
            metric_summary_rows.append(
                {
                    "metric_name": entry["metric_name"],
                    "avg_score": (entry["score_sum"] / count) if count else None,
                    "pass_rate": (entry["pass_count"] / count) if count else None,
                    "samples": count,
                    "source": spec.source if spec else None,
                }
            )

        sheet_order: list[tuple[str, list[dict[str, Any]], list[str]]] = [
            (
                "Summary",
                summary_rows,
                [
                    "run_id",
                    "dataset_name",
                    "model_name",
                    "started_at",
                    "finished_at",
                    "total_test_cases",
                    "total_tokens",
                    "total_cost_usd",
                    "pass_rate",
                    "metrics_evaluated",
                    "prompt_set_id",
                    "prompt_set_name",
                ],
            ),
            (
                "Run",
                run_rows,
                [
                    "run_id",
                    "dataset_name",
                    "dataset_version",
                    "model_name",
                    "started_at",
                    "finished_at",
                    "total_tokens",
                    "total_cost_usd",
                    "pass_rate",
                    "metrics_evaluated",
                    "thresholds",
                    "langfuse_trace_id",
                    "metadata",
                    "retrieval_metadata",
                    "created_at",
                ],
            ),
            (
                "TestCases",
                test_case_payloads,
                [
                    "id",
                    "run_id",
                    "test_case_id",
                    "tokens_used",
                    "latency_ms",
                    "cost_usd",
                    "trace_id",
                    "started_at",
                    "finished_at",
                    "question",
                    "answer",
                    "contexts",
                    "ground_truth",
                ],
            ),
            (
                "MetricScores",
                metric_payloads,
                ["result_id", "test_case_id", "metric_name", "score", "threshold", "reason"],
            ),
            (
                "MetricsSummary",
                metric_summary_rows,
                ["metric_name", "avg_score", "pass_rate", "samples", "source"],
            ),
            (
                "CustomMetrics",
                custom_metric_rows,
                [
                    "schema_version",
                    "metric_name",
                    "source",
                    "description",
                    "evaluation_method",
                    "inputs",
                    "output",
                    "evaluation_process",
                    "rules",
                    "notes",
                    "implementation_path",
                    "implementation_hash",
                ],
            ),
            (
                "RunPromptSets",
                run_prompt_payloads,
                ["run_id", "prompt_set_id", "created_at"],
            ),
            (
                "PromptSets",
                prompt_sets_payloads,
                ["prompt_set_id", "name", "description", "metadata", "created_at"],
            ),
            (
                "PromptSetItems",
                prompt_set_item_payloads,
                ["id", "prompt_set_id", "prompt_id", "role", "item_order", "metadata"],
            ),
            (
                "Prompts",
                prompt_payloads,
                [
                    "prompt_id",
                    "name",
                    "kind",
                    "content",
                    "checksum",
                    "source",
                    "notes",
                    "metadata",
                    "created_at",
                ],
            ),
            (
                "Feedback",
                feedback_payloads,
                [
                    "id",
                    "run_id",
                    "test_case_id",
                    "satisfaction_score",
                    "thumb_feedback",
                    "comment",
                    "rater_id",
                    "created_at",
                ],
            ),
            (
                "ClusterMaps",
                cluster_payloads,
                [
                    "run_id",
                    "map_id",
                    "test_case_id",
                    "cluster_id",
                    "source",
                    "metadata",
                    "created_at",
                ],
            ),
            (
                "StageEvents",
                stage_event_payloads,
                [
                    "id",
                    "run_id",
                    "stage_id",
                    "parent_stage_id",
                    "stage_type",
                    "stage_name",
                    "status",
                    "attempt",
                    "started_at",
                    "finished_at",
                    "duration_ms",
                    "input_ref",
                    "output_ref",
                    "attributes",
                    "metadata",
                    "trace_id",
                    "span_id",
                ],
            ),
            (
                "StageMetrics",
                stage_metric_payloads,
                ["id", "run_id", "stage_id", "metric_name", "score", "threshold", "evidence"],
            ),
            (
                "AnalysisReports",
                report_payloads,
                [
                    "report_id",
                    "run_id",
                    "experiment_id",
                    "report_type",
                    "format",
                    "content",
                    "metadata",
                    "created_at",
                ],
            ),
            (
                "PipelineResults",
                pipeline_payloads,
                [
                    "result_id",
                    "intent",
                    "query",
                    "run_id",
                    "pipeline_id",
                    "profile",
                    "tags",
                    "metadata",
                    "is_complete",
                    "duration_ms",
                    "final_output",
                    "node_results",
                    "started_at",
                    "finished_at",
                    "created_at",
                ],
            ),
        ]

        workbook = Workbook()
        default_sheet = workbook.active
        if default_sheet is not None:
            workbook.remove(default_sheet)
        for sheet_name, rows, columns in sheet_order:
            worksheet = cast(Any, workbook.create_sheet(title=sheet_name))
            worksheet.append(columns)
            for row in rows:
                worksheet.append([row.get(column) for column in columns])

        workbook.save(output)
        return output

    def export_analysis_results_to_excel(self, run_id: str, output_path) -> Path:
        from openpyxl import Workbook

        output = Path(output_path)
        output.parent.mkdir(parents=True, exist_ok=True)
        placeholder = self.queries.placeholder

        with self._get_connection() as conn:
            analysis_rows = self._execute(
                conn,
                (
                    "SELECT analysis_id, run_id, analysis_type, result_data, created_at "
                    f"FROM analysis_results WHERE run_id = {placeholder} ORDER BY created_at DESC"
                ),
                (run_id,),
            ).fetchall()
            analysis_payloads = self._normalize_rows(
                analysis_rows,
                json_columns={"result_data"},
            )

            report_rows = self._execute(
                conn,
                (
                    "SELECT report_id, run_id, experiment_id, report_type, format, content, metadata, created_at "
                    f"FROM analysis_reports WHERE run_id = {placeholder} ORDER BY created_at DESC"
                ),
                (run_id,),
            ).fetchall()
            report_payloads = self._normalize_rows(report_rows, json_columns={"metadata"})

            pipeline_rows = self._execute(
                conn,
                (
                    "SELECT result_id, intent, query, run_id, pipeline_id, profile, tags, metadata, "
                    "is_complete, duration_ms, final_output, node_results, started_at, finished_at, created_at "
                    f"FROM pipeline_results WHERE run_id = {placeholder} ORDER BY created_at DESC"
                ),
                (run_id,),
            ).fetchall()
            pipeline_payloads = self._normalize_rows(
                pipeline_rows,
                json_columns={"tags", "metadata", "final_output", "node_results"},
            )

        sheet_order: list[tuple[str, list[dict[str, Any]], list[str]]] = [
            (
                "AnalysisResults",
                analysis_payloads,
                ["analysis_id", "run_id", "analysis_type", "result_data", "created_at"],
            ),
            (
                "AnalysisReports",
                report_payloads,
                [
                    "report_id",
                    "run_id",
                    "experiment_id",
                    "report_type",
                    "format",
                    "content",
                    "metadata",
                    "created_at",
                ],
            ),
            (
                "PipelineResults",
                pipeline_payloads,
                [
                    "result_id",
                    "intent",
                    "query",
                    "run_id",
                    "pipeline_id",
                    "profile",
                    "tags",
                    "metadata",
                    "is_complete",
                    "duration_ms",
                    "final_output",
                    "node_results",
                    "started_at",
                    "finished_at",
                    "created_at",
                ],
            ),
        ]

        workbook = Workbook()
        default_sheet = workbook.active
        if default_sheet is not None:
            workbook.remove(default_sheet)
        for sheet_name, rows, columns in sheet_order:
            worksheet = cast(Any, workbook.create_sheet(title=sheet_name))
            worksheet.append(columns)
            for row in rows:
                worksheet.append([row.get(column) for column in columns])

        workbook.save(output)
        return output

    def export_multiturn_run_to_excel(self, run_id: str, output_path) -> Path:
        from openpyxl import Workbook

        output = Path(output_path)
        output.parent.mkdir(parents=True, exist_ok=True)
        placeholder = self.queries.placeholder

        with self._get_connection() as conn:
            run_row = self._execute(conn, self.queries.select_multiturn_run(), (run_id,)).fetchone()
            if not run_row:
                raise KeyError(f"Multiturn run not found: {run_id}")

            run_rows = self._normalize_rows(
                [run_row],
                json_columns={"metrics_evaluated", "summary", "metadata"},
            )

            conversation_rows = self._execute(
                conn, self.queries.select_multiturn_conversations(), (run_id,)
            ).fetchall()
            conversation_payloads = self._normalize_rows(
                conversation_rows,
                json_columns={"summary"},
            )

            turn_rows = self._execute(
                conn, self.queries.select_multiturn_turn_results(), (run_id,)
            ).fetchall()
            turn_payloads = self._normalize_rows(
                turn_rows,
                json_columns={"metadata"},
            )

            metric_rows = self._execute(
                conn,
                (
                    "SELECT m.turn_result_id, t.conversation_id, t.turn_id, t.turn_index, "
                    "m.metric_name, m.score, m.threshold "
                    "FROM multiturn_metric_scores m "
                    "JOIN multiturn_turn_results t ON m.turn_result_id = t.id "
                    f"WHERE t.run_id = {placeholder} ORDER BY m.id"
                ),
                (run_id,),
            ).fetchall()
            metric_payloads = self._normalize_rows(metric_rows)

        sheet_order: list[tuple[str, list[dict[str, Any]], list[str]]] = [
            (
                "MultiTurnRun",
                run_rows,
                [
                    "run_id",
                    "dataset_name",
                    "dataset_version",
                    "model_name",
                    "started_at",
                    "finished_at",
                    "conversation_count",
                    "turn_count",
                    "metrics_evaluated",
                    "drift_threshold",
                    "summary",
                    "metadata",
                    "created_at",
                ],
            ),
            (
                "MultiTurnConversations",
                conversation_payloads,
                [
                    "run_id",
                    "conversation_id",
                    "turn_count",
                    "drift_score",
                    "drift_threshold",
                    "drift_detected",
                    "summary",
                ],
            ),
            (
                "MultiTurnTurns",
                turn_payloads,
                [
                    "id",
                    "run_id",
                    "conversation_id",
                    "turn_id",
                    "turn_index",
                    "role",
                    "passed",
                    "latency_ms",
                    "metadata",
                ],
            ),
            (
                "MultiTurnTurnMetrics",
                metric_payloads,
                [
                    "turn_result_id",
                    "conversation_id",
                    "turn_id",
                    "turn_index",
                    "metric_name",
                    "score",
                    "threshold",
                ],
            ),
        ]

        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        for sheet_name, rows, headers in sheet_order:
            sheet = workbook.create_sheet(title=sheet_name)
            sheet.append(headers)
            for row in rows:
                sheet.append([self._row_value(row, header) for header in headers])

        workbook.save(output)
        return output
