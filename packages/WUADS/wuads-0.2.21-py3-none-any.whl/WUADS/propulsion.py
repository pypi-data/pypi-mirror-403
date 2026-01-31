import logging
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from .flight_conditions import FlightConditions
import numpy as np
logger = logging.getLogger(__name__)



class engine:
    """
    Contains all analyses and variables required to analyze propulsion performance
    """

    def __init__(self, n_engines=2, max_thrust=0, sfc=0):
        """
        Initializes the number of engines with the specified number of engines.

        :param int n_engines: Number of engines
        """

        self.engine_type = ''
        self.altitude_ref = []  # Altitudes used for the sfc and thrust tables
        self.mach_ref = []  # Mach number used for the sfc and thrust tables
        self.thrust_input = []  # Input values for thrust
        self.sfc_input = []  # Input values for sfc
        self.components = {}
        self.n_engines = n_engines  # Number of engines
        self.max_thrust = max_thrust
        self.sfc = sfc

    def analyze_performance(self, height, mach, thrust_required=None):
        return


class turbofan(engine):
    """
    Represents a generic, commercial aviation turbofan engine

    Scales results based off the values for the cfm56_7b24 (737-800's engine)
    """

    def __init__(self, h_cruise, mach_cruise, n_engines=2, thrust_sea_level=None,
                 sfc_sea_level=None, thrust_cruise=None, sfc_cruise=None, engine_data_file=None):
        """
        Scales the results based off the input conditions at both sea level and cruise

        inputs:
        :param float h_cruise: altitude at cruise (ft)
        :param float mach_cruise: mach number at cruise
        :param int n_engines: Number of engines
        :param float thrust_sea_level: maximum thrust at sea level, lbs
        :param float sfc_sea_level: specific fuel consumption at sea level
        :param float thrust_cruise: maximum thrust at cruise conditions, lbs
        :param float sfc_cruise: specific fuel consuption at cruise, 1/hrs
        """
        super().__init__(n_engines)

        # Values used to scale the standard performance data
        self.thrust_sea_level = None
        self.thrust_cruise = None
        self.sfc_sea_level = None
        self.sfc_cruise = None
        self.engine_type = None
        self.engine_data_file = None

        self.engine_type = 'turbofan'
        # Basic cfm56-7b values
        # self.thrust_input = [[24200, 21016, 19105, 17004, 15284, 13819, 12418, 10508, 8152, 5413],
        #                 [16004, 14597, 13954, 12293, 11335, 10505, 9932, 9165, 8271, 7185],
        #                 [9788, 9074, 8425, 7842, 7326, 6909, 6685, 6426, 6153, 5585],
        #                 [8155, 7522, 6915, 6459, 6165, 6002, 5937, 5741, 5480, 5121],
        #                 [6214, 5750, 5320, 4988, 4757, 4625, 4515, 4363, 4178, 3932],
        #                 [4054, 3516, 3113, 2819, 2605, 2443, 2305, 2162, 1987, 1752]]
        #
        # self.sfc_input = [[0.370, 0.416, 0.474, 0.503, 0.624, 0.694, 0.786, 0.948, 1.214, 1.769],
        #              [0.362, 0.382, 0.413, 0.467, 0.527, 0.592, 0.674, 0.760, 0.869, 1.046],
        #              [0.294, 0.333, 0.367, 0.408, 0.449, 0.505, 0.541, 0.615, 0.670, 0.769],
        #              [0.291, 0.324, 0.353, 0.390, 0.425, 0.462, 0.509, 0.564, 0.627, 0.703],
        #              [0.273, 0.333, 0.359, 0.383, 0.419, 0.466, 0.498, 0.563, 0.614, 0.661],
        #              [0.224, 0.260, 0.296, 0.331, 0.367, 0.406, 0.448, 0.495, 0.548, 0.609]]
        #
        # self.mach_ref = [0, .1, .2, .3, .4, .5, .6, .7, .8, .9]
        # self.altitude_ref = [0, 15000, 31000, 35000, 41000, 50000]

        self.thrust_input = [[  22164,	19964,	17930,	16245,	14852,	13700,	12760,	12001,	11405,	10931,	10532],
                              [  15241,	13827,	12705,	11973,	11554,	11417,	11041,	10452,	9982,	9598,	9269],
                              [  10185,	9241,	8489,	7999,	7726,	7630,	7695,	7902,	8249,	8259,	7991],
                              [  6587,	5974,	5489,	5171,	4995,	4932,	4974,	5104,	5334,	5661,	6077],
                              [  4113,	3730,	3426,	3229,	3119,	3078,	3105,	3185,	3328,	3535,	3792],
                              [  2546,	2310,	2122,	2001,	1932,	1906,	1924,	1973,	2062,	2190,	2350],
                              [  1579,	1432,	1316,	1240,	1198,	1182,	1193,	1223,	1278,	1358,	1457],
                              [  980,	889,	816,	769,    743,	734,	739, 	759,	793, 	842,    904]
        ]
        self.sfc_input = [[ 0.365,	0.406,	0.452,	0.500,	0.550,	0.601,	0.652,	0.704,	0.756,	0.808,	0.860],
                           [ 0.345,	0.385,	0.431,	0.481,	0.532,	0.584,	0.632,	0.678,	0.723,	0.768,	0.815],
                           [ 0.326,	0.362,	0.407,	0.453,	0.501,	0.550,	0.599,	0.647,	0.696,	0.739,	0.781],
                           [ 0.305,	0.340,	0.382,	0.425,	0.470,	0.516,	0.562,	0.607,	0.653,	0.698,	0.746],
                           [ 0.293,	0.326,	0.366,	0.408,	0.451,	0.495,	0.539,	0.583,	0.626,	0.670,	0.716],
                           [ 0.293,	0.327,	0.366,	0.408,	0.451,	0.495,	0.539,	0.583,	0.626,	0.670,	0.716],
                           [ 0.293,	0.327,	0.366,	0.408,	0.451,	0.495,	0.539,	0.583,	0.626,	0.670,	0.716],
                           [ 0.295,	0.328,	0.368,	0.410,	0.453,	0.497,	0.542,	0.585,	0.629,	0.673,	0.719]
        ]
        self.altitude_ref = [0, 10000, 20000, 30000, 40000, 50000, 60000, 70000]
        self.mach_ref = [0, .1, .2, .3, .4, .5, .6, .7, .8, .9, 1]

        if engine_data_file:
            self.engine_data_file = engine_data_file
            self.load_data_file(engine_data_file)

        self.scale_performance(h_cruise, mach_cruise, n_engines=n_engines, thrust_sea_level=thrust_sea_level,
                               thrust_cruise=thrust_cruise,
                               sfc_cruise=sfc_cruise, sfc_sea_level=sfc_sea_level)

    def analyze_performance(self, height, mach, thrust_required=None):
        """
        Find the maximum available thrust and specific fuel consumption at an input altitude, mach number,
        and required thrust


        :param float height: Altitude (ft)
        :param float mach: Mach Number
        :param float thrust_required: Required thrust (lbs)


        :return float sfc: specific fuel consumption at input conditions (s^-1)
        :return float max_thrust: Maximum thrust at input conditions (lbs)

        """
        if mach > self.mach_ref[-1]:  # TODO edit interpolation to include mach and altitudes outside input list
            mach = self.mach_ref[-1] - .001

        # Find values to interpolate between

        # altitude
        if height <= self.altitude_ref[0]:
            x_indx_low = 0
            x_indx_high = 1
            x1 = self.altitude_ref[0]
            x2 = self.altitude_ref[1]
        elif height >= self.altitude_ref[-1]:
            x_indx_low = len(self.altitude_ref) - 2
            x_indx_high = len(self.altitude_ref) - 1
            x1 = self.altitude_ref[-2]
            x2 = self.altitude_ref[-1]
        else:
            for i in range(1, len(self.altitude_ref)):
                if height < self.altitude_ref[i]:
                    x_indx_low = i - 1
                    x_indx_high = i
                    x1 = self.altitude_ref[i - 1]
                    x2 = self.altitude_ref[i]
                    break

        # Mach
        if mach <= self.mach_ref[0]:
            y_indx_low = 0
            y_indx_high = 1
            y1 = self.mach_ref[0]
            y2 = self.mach_ref[1]
        elif mach >= self.mach_ref[-1]:
            y_indx_low = len(self.mach_ref) - 2
            y_indx_high = len(self.mach_ref) - 1
            y1 = self.mach_ref[-2]
            y2 = self.mach_ref[-1]
        else:
            for i in range(1, len(self.mach_ref)):
                if mach < self.mach_ref[i]:
                    y_indx_low = i - 1
                    y_indx_high = i
                    y1 = self.mach_ref[i - 1]
                    y2 = self.mach_ref[i]
                    break

        f11 = self.sfc_input[x_indx_low][y_indx_low]
        f12 = self.sfc_input[x_indx_low][y_indx_high]
        f21 = self.sfc_input[x_indx_high][y_indx_low]
        f22 = self.sfc_input[x_indx_high][y_indx_high]

        sfc = self._weighted_bilinear(x1, x2, y1, y2, f11, f12, f21, f22, height, mach)

        # Now find the max thrust
        f11 = self.thrust_input[x_indx_low][y_indx_low]
        f12 = self.thrust_input[x_indx_low][y_indx_high]
        f21 = self.thrust_input[x_indx_high][y_indx_low]
        f22 = self.thrust_input[x_indx_high][y_indx_high]

        max_thrust = self._weighted_bilinear(x1, x2, y1, y2, f11, f12, f21, f22, height, mach)
        max_thrust *= self.n_engines

        # Scale results for partial power (if required thrust is input)
        if thrust_required:
            thrust_percent = thrust_required / max_thrust
            # Raymer partial power correction
            if thrust_percent > .2:
                sfc *= .1 / thrust_percent + .24 / (thrust_percent) ** .8 + .66 * thrust_percent ** .8 + .1 * mach * (
                        1 / thrust_percent - thrust_percent)
            else:
                sfc *= 1.5

        return sfc, max_thrust

    def scale_performance(self, h_cruise, mach_cruise, n_engines=2, thrust_sea_level=None,
                          sfc_sea_level=None, thrust_cruise=None, sfc_cruise=None):

        if h_cruise > self.altitude_ref[-1]:
            # TODO make this work outside bounds
            logging.warning(
                f'Input cruise altitude is higher than the standard engine bounds ({self.altitude_ref[-1]} ft), could not scale engine data. Please select a different engine or lower the cruise altitude',
                UserWarning)
            return

        self.thrust_sea_level = thrust_sea_level
        self.thrust_cruise = thrust_cruise
        self.sfc_sea_level = sfc_sea_level
        self.sfc_cruise = sfc_cruise

        # Set scale values
        if thrust_sea_level:
            thrust_scale_sl = thrust_sea_level / self.thrust_input[0][0]
        else:
            thrust_scale_sl = 1

        if sfc_sea_level:
            sfc_scale_sl = sfc_sea_level / self.sfc_input[0][0]
        else:
            sfc_scale_sl = 1

        sfc_ref, thrust_cruise_ref = self.analyze_performance(h_cruise, mach_cruise)
        thrust_cruise_ref *= 1 / self.n_engines  # Single engine correction
        if thrust_cruise:
            thrust_scale_cruise = thrust_cruise / thrust_cruise_ref
        else:
            thrust_scale_cruise = 1

        if sfc_cruise:
            sfc_scale_cruise = sfc_cruise / sfc_ref
        else:
            sfc_scale_cruise = 1

        # Scale all values
        for i in range(len(self.thrust_input)):
            for j in range(len(self.thrust_input[i])):
                t_scaled_sea_level = self.thrust_input[i][j] * thrust_scale_sl
                t_scaled_cruise = self.thrust_input[i][j] * thrust_scale_cruise
                sfc_scaled_sl = self.sfc_input[i][j] * sfc_scale_sl
                sfc_scaled_cruise = self.sfc_input[i][j] * sfc_scale_cruise

                if thrust_sea_level and not thrust_cruise:
                    self.thrust_input[i][j] = t_scaled_sea_level
                elif thrust_cruise and not thrust_sea_level:
                    self.thrust_input[i][j] = t_scaled_cruise
                elif thrust_cruise and thrust_sea_level:
                    self.thrust_input[i][j] = (h_cruise - self.altitude_ref[i]) / h_cruise * t_scaled_sea_level + \
                                              self.altitude_ref[i] / h_cruise * t_scaled_cruise

                if sfc_sea_level and not sfc_cruise:
                    self.sfc_input[i][j] = sfc_scaled_sl
                elif sfc_cruise and not sfc_sea_level:
                    self.sfc_input[i][j] = sfc_scaled_cruise
                elif sfc_cruise and sfc_sea_level:
                    self.sfc_input[i][j] = (h_cruise - self.altitude_ref[i]) / h_cruise * sfc_scaled_sl + \
                                           self.altitude_ref[i] / h_cruise * sfc_scaled_cruise

        self.max_thrust = self.thrust_input[0][0]

    @staticmethod
    def _weighted_bilinear(x1, x2, y1, y2, f11, f12, f21, f22, x, y):
        """
        Performs weighted bilinear interpolation for a given set of points.

        :param float x1: Lower bound for the x-coordinate.
        :param float x2: Upper bound for the x-coordinate.
        :param float y1: Lower bound for the y-coordinate.
        :param float y2: Upper bound for the y-coordinate.
        :param float f11: Value at (x1, y1).
        :param float f12: Value at (x1, y2).
        :param float f21: Value at (x2, y1).
        :param float f22: Value at (x2, y2).
        :param float x: Target x-coordinate for interpolation.
        :param float y: Target y-coordinate for interpolation.

        :return: Interpolated value at (x, y).
        :rtype: float
        """
        # Weighted Bilinear interpolation
        w11 = ((x2 - x) * (y2 - y)) / ((x2 - x1) * (y2 - y1))
        w12 = ((x2 - x) * (y - y1)) / ((x2 - x1) * (y2 - y1))
        w21 = ((x - x1) * (y2 - y)) / ((x2 - x1) * (y2 - y1))
        w22 = ((x - x1) * (y - y1)) / ((x2 - x1) * (y2 - y1))

        f = w11 * f11 + w12 * f12 + w21 * f21 + w22 * f22
        return f

    def write_data_file(self, file_name):
        # Output data file to correct location
        thrust_df = pd.DataFrame(self.thrust_input, columns=self.mach_ref, index=self.altitude_ref)
        sfc_df = pd.DataFrame(self.sfc_input, columns=self.mach_ref, index=self.altitude_ref)
        with pd.ExcelWriter(file_name, engine="openpyxl") as writer:
            thrust_df.to_excel(writer, sheet_name="Thrust", startrow=2)
            sfc_df.to_excel(writer, sheet_name="sfc", startrow=2)

        # Fix formatting
        wb = load_workbook(file_name)

        col_headers = self.mach_ref
        row_headers = self.thrust_input
        for sheet_name in ["Thrust", "sfc"]:
            ws = wb[sheet_name]

            # Overall title (row 1, merged across all columns including row header column)
            max_col = len(col_headers) + 1  # +1 for row headers column
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
            title_cell = ws.cell(row=1, column=1)
            if sheet_name == "Thrust":
                sheet_name += ' (lbs)'
            else:
                sheet_name += ' (lbm/hr)/lbs'
            title_cell.value = sheet_name
            title_cell.font = Font(size=16, bold=True)
            title_cell.alignment = Alignment(horizontal="center", vertical="center")

            # Column headers description (row 3, merged across data columns only, leaving row header column out)
            ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=max_col)
            col_desc_cell = ws.cell(row=2, column=2)
            col_desc_cell.value = 'Mach Number'
            col_desc_cell.alignment = Alignment(horizontal="center", vertical="center")
            col_desc_cell.font = Font(italic=True, size=11)

            # Row headers description (put near top-left, above row headers - cell A4)
            row_desc_cell = ws.cell(row=3, column=1)
            row_desc_cell.value = "Altitude (ft)"
            row_desc_cell.font = Font(italic=True, size=11)
            row_desc_cell.alignment = Alignment(horizontal="left", vertical="center")

            wb.save(file_name)

    def load_data_file(self, file_name):
        thrust_df = pd.read_excel(file_name, sheet_name="Thrust", skiprows=2, header=0, index_col=0)
        sfc_df = pd.read_excel(file_name, sheet_name="sfc", skiprows=2, header=0, index_col=0)

        self.altitude_ref = thrust_df.index.to_list()
        self.mach_ref = thrust_df.columns.tolist()
        self.thrust_input = thrust_df.values.tolist()
        self.sfc_input = sfc_df.values.tolist()


class propeller(engine):
    """
    Standard piston powered propeller engine
    """

    #engine_type = 'propeller'

    def __init__(self, n_engines=1, horse_power=None, fuel_consumption_rate=None):
        super().__init__(n_engines)
        self.horse_power = horse_power
        self.current_horse_power = horse_power
        self.fuel_consumption_rate = fuel_consumption_rate
        self.current_fuel_consumption_rate = fuel_consumption_rate
        self.engine_type = 'propeller'

    def analyze_performance(self, height, mach, thrust_required=None):


        # Flight conditions and velocity (ft/s)
        fc = FlightConditions(height, mach)
        V = mach * fc.a  # ft/s

        #efficiency calculation
        nu_propeller = self.prop_efficiency(mach)

        #altitude horsepower correction equation 14.5 from Nicolai
        rho = fc.rho  # air density at altitude (slugs/ft^3 if imperial)
        rho_sl = 0.0023769  # sea-level density, slugs/ft^3
        Bhp_sl = self.horse_power

        Bhp_h = Bhp_sl * (rho / rho_sl - (1 - rho / rho_sl) / 7.75)  # available hp at altitude
        total_horse_power = self.n_engines * Bhp_h

        P_generated = total_horse_power * 550  # horsepower to ft-lb/s
        P_available = nu_propeller * P_generated

        fuel_density = 6.2  # lbs/gal (Avgas density adjusted slightly for temperature differences)
        total_fuel_consumption = self.n_engines * self.current_fuel_consumption_rate  # gal/hr
        fuel_mass_flow_rate_lb_per_hr = total_fuel_consumption * fuel_density  # lbs/hr

        # Handle very low or zero velocity case (static thrust approximation)
        if V <= 1e-3 or np.isnan(V):
            # Estimate static thrust as empirical factor * total horsepower

            static_thrust_factor = 2.5  # lbf per horsepower (typical GA propeller estimate)
            max_thrust = total_horse_power * static_thrust_factor
            sfc = fuel_mass_flow_rate_lb_per_hr / (max_thrust + 1e-6)  # Avoid div by zero
        else:
            max_thrust = P_available / V  # lbf
            # thrust calculated from power and velocity
            thrust = max_thrust
            sfc = fuel_mass_flow_rate_lb_per_hr / thrust  # lb/(lbf*hr)

        self.max_thrust = max_thrust


        return sfc, max_thrust

    # alernative method from section 3.1: https://www.fzt.haw-hamburg.de/pers/Scholz/transfer/Airport2030_TN_Propeller-Efficiency_13-08-12_SLZ.pdf
    @staticmethod
    def prop_efficiency(mach):
        # nu_max = 0.9  # can adjust
        #
        # if mach <= 0.1:
        #     nu_propeller = 10 * mach * nu_max
        #
        #
        # elif .1 < mach <= 0.7:
        #     nu_propeller = nu_max
        #
        # elif .7 < mach < 0.85:
        #     nu_propeller = 10 * mach * nu_max * (1 - (mach - 0.7) / 3)
        #
        # else:
        #     nu_propeller = max(0.0, 10 * mach * nu_max * (
        #                 1 - (mach - 0.7) / 3))  # prevent negative (just sets to 0 if negative)
        nu_propeller = .9

        return nu_propeller

