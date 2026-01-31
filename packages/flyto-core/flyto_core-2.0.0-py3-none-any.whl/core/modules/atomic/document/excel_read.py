"""
Excel Read Module
Read data from Excel files (xlsx, xls)
"""
import logging
import os
from typing import Any, Dict, List, Optional

from ...registry import register_module
from ...schema import compose, presets


logger = logging.getLogger(__name__)


@register_module(
    module_id='excel.read',
    version='1.0.0',
    category='document',
    subcategory='excel',
    tags=['excel', 'spreadsheet', 'read', 'xlsx', 'data', 'path_restricted'],
    label='Read Excel',
    label_key='modules.excel.read.label',
    description='Read data from Excel files (xlsx, xls)',
    description_key='modules.excel.read.description',
    icon='Table',
    color='#217346',

    # Connection types
    input_types=['file_path'],
    output_types=['array', 'object'],
    can_connect_to=['array.*', 'data.*', 'database.*'],
    can_receive_from=['file.*', 'data.*', 'api.*', 'flow.*', 'start'],

    # Execution settings
    timeout_ms=60000,
    retryable=False,
    concurrent_safe=True,

    # Security settings
    requires_credentials=False,
    handles_sensitive_data=True,
    required_permissions=['filesystem.read', 'filesystem.write'],

    params_schema=compose(
        presets.EXCEL_PATH(placeholder='/path/to/data.xlsx'),
        presets.EXCEL_SHEET(),
        presets.EXCEL_HEADER_ROW(),
        presets.EXCEL_RANGE(),
        presets.EXCEL_AS_DICT(),
    ),
    output_schema={
        'data': {
            'type': 'array',
            'description': 'Extracted data rows'
        ,
                'description_key': 'modules.excel.read.output.data.description'},
        'headers': {
            'type': 'array',
            'description': 'Column headers'
        ,
                'description_key': 'modules.excel.read.output.headers.description'},
        'row_count': {
            'type': 'number',
            'description': 'Number of data rows'
        ,
                'description_key': 'modules.excel.read.output.row_count.description'},
        'sheet_names': {
            'type': 'array',
            'description': 'All sheet names in the workbook'
        ,
                'description_key': 'modules.excel.read.output.sheet_names.description'}
    },
    examples=[
        {
            'title': 'Read entire sheet',
            'title_key': 'modules.excel.read.examples.basic.title',
            'params': {
                'path': '/tmp/data.xlsx',
                'as_dict': True
            }
        }
    ],
    author='Flyto2 Team',
    license='MIT'
)
async def excel_read(context: Dict[str, Any]) -> Dict[str, Any]:
    """Read data from Excel file"""
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required for Excel reading. Install with: pip install openpyxl")

    params = context['params']
    path = params['path']
    sheet_name = params.get('sheet')
    header_row = params.get('header_row', 1)
    cell_range = params.get('range')
    as_dict = params.get('as_dict', True)

    # Validate file exists
    if not os.path.exists(path):
        raise FileNotFoundError(f"Excel file not found: {path}")

    # Open workbook
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet_names = wb.sheetnames

    # Select sheet
    if sheet_name:
        if sheet_name not in sheet_names:
            raise ValueError(f"Sheet not found: {sheet_name}")
        ws = wb[sheet_name]
    else:
        ws = wb.active

    # Determine range to read
    if cell_range:
        cells = ws[cell_range]
    else:
        cells = ws.iter_rows()

    # Read data
    all_rows: List[List[Any]] = []
    for row in cells:
        row_data = [cell.value for cell in row]
        all_rows.append(row_data)

    # Extract headers
    headers: List[str] = []
    data_rows: List[Any] = []

    if header_row > 0 and len(all_rows) >= header_row:
        headers = [str(h) if h else f"col_{i}" for i, h in enumerate(all_rows[header_row - 1])]
        data_rows = all_rows[header_row:]
    else:
        data_rows = all_rows

    # Convert to dicts if requested
    if as_dict and headers:
        data = []
        for row in data_rows:
            row_dict = {}
            for i, val in enumerate(row):
                key = headers[i] if i < len(headers) else f"col_{i}"
                row_dict[key] = val
            data.append(row_dict)
    else:
        data = data_rows

    wb.close()

    logger.info(f"Read Excel: {path} ({len(data)} rows)")

    return {
        'ok': True,
        'data': data,
        'headers': headers,
        'row_count': len(data),
        'sheet_names': sheet_names,
        'active_sheet': ws.title
    }
