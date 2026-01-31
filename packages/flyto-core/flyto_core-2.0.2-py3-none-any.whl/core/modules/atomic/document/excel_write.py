"""
Excel Write Module
Write data to Excel files (xlsx)
"""
import logging
import os
from pathlib import Path
from typing import Any, Dict, List, Optional

from ...registry import register_module
from ...schema import compose, presets


logger = logging.getLogger(__name__)


@register_module(
    module_id='excel.write',
    version='1.0.0',
    category='document',
    subcategory='excel',
    tags=['excel', 'spreadsheet', 'write', 'xlsx', 'export', 'path_restricted'],
    label='Write Excel',
    label_key='modules.excel.write.label',
    description='Write data to Excel files (xlsx)',
    description_key='modules.excel.write.description',
    icon='Table',
    color='#217346',

    # Connection types
    input_types=['array', 'object'],
    output_types=['file_path'],
    can_connect_to=['file.*', 'email.*'],
    can_receive_from=['file.*', 'data.*', 'api.*', 'flow.*', 'start'],

    # Execution settings
    timeout_ms=60000,
    retryable=False,
    concurrent_safe=True,

    # Security settings
    requires_credentials=False,
    handles_sensitive_data=True,
    required_permissions=['filesystem.read', 'filesystem.write'],

    params_schema=compose(
        presets.EXCEL_PATH(placeholder='/path/to/output.xlsx'),
        presets.EXCEL_DATA(),
        presets.EXCEL_HEADERS(),
        presets.EXCEL_SHEET_NAME(),
        presets.EXCEL_AUTO_WIDTH(),
    ),
    output_schema={
        'path': {
            'type': 'string',
            'description': 'Path to the created Excel file'
        ,
                'description_key': 'modules.excel.write.output.path.description'},
        'row_count': {
            'type': 'number',
            'description': 'Number of data rows written'
        ,
                'description_key': 'modules.excel.write.output.row_count.description'},
        'size': {
            'type': 'number',
            'description': 'File size in bytes'
        ,
                'description_key': 'modules.excel.write.output.size.description'}
    },
    examples=[
        {
            'title': 'Write data to Excel',
            'title_key': 'modules.excel.write.examples.basic.title',
            'params': {
                'path': '/tmp/output.xlsx',
                'data': [
                    {'name': 'Alice', 'age': 30},
                    {'name': 'Bob', 'age': 25}
                ]
            }
        }
    ],
    author='Flyto2 Team',
    license='MIT'
)
async def excel_write(context: Dict[str, Any]) -> Dict[str, Any]:
    """Write data to Excel file"""
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl is required for Excel writing. Install with: pip install openpyxl")

    params = context['params']
    path = params['path']
    data = params['data']
    headers = params.get('headers')
    sheet_name = params.get('sheet_name', 'Sheet1')
    auto_width = params.get('auto_width', True)

    # Validate data
    if not isinstance(data, list):
        raise ValueError("Data must be an array")

    if not data:
        raise ValueError("Data cannot be empty")

    # Ensure output directory exists
    Path(os.path.dirname(path)).mkdir(parents=True, exist_ok=True)

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Determine headers
    if not headers:
        if isinstance(data[0], dict):
            headers = list(data[0].keys())
        else:
            headers = [f"Column {i+1}" for i in range(len(data[0]))]

    # Write headers
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # Track column widths
    col_widths: Dict[int, int] = {i: len(str(h)) for i, h in enumerate(headers, 1)}

    # Write data
    for row_idx, row_data in enumerate(data, 2):
        if isinstance(row_data, dict):
            for col_idx, header in enumerate(headers, 1):
                value = row_data.get(header, '')
                ws.cell(row=row_idx, column=col_idx, value=value)
                if auto_width:
                    col_widths[col_idx] = max(col_widths.get(col_idx, 0), len(str(value)))
        elif isinstance(row_data, (list, tuple)):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
                if auto_width:
                    col_widths[col_idx] = max(col_widths.get(col_idx, 0), len(str(value)))

    # Adjust column widths
    if auto_width:
        for col, width in col_widths.items():
            ws.column_dimensions[get_column_letter(col)].width = min(width + 2, 50)

    # Save workbook
    wb.save(path)
    wb.close()

    file_size = os.path.getsize(path)

    logger.info(f"Wrote Excel: {path} ({len(data)} rows)")

    return {
        'ok': True,
        'path': path,
        'row_count': len(data),
        'size': file_size
    }
