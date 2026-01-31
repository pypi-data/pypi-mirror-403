"""File analyzer for enterprise mode.

Analyzes files before upload to extract metadata like token counts,
language detection, and file type statistics.
"""

from __future__ import annotations

import asyncio
import warnings
from collections import Counter
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path
from typing import Any

from loguru import logger

# Suppress cryptography deprecation warning from pypdf
warnings.filterwarnings("ignore")
warnings.filterwarnings("ignore", category=DeprecationWarning, module="cryptography")
warnings.filterwarnings(
    "ignore", category=DeprecationWarning, module="pypdf._crypt_providers._cryptography"
)


class FileAnalyzer:
    """Analyzes files to extract metadata and statistics."""

    def __init__(self, model_encoding: str = "cl100k_base"):
        """Initialize FileAnalyzer with a specific tokenizer encoding.

        Args:
            model_encoding: The encoding to use for token counting (default: cl100k_base for GPT-4).
        """
        self.encoding = None
        try:
            import tiktoken

            self.encoding = tiktoken.get_encoding(model_encoding)
        except ImportError:
            logger.debug("tiktoken not available, using fallback token estimation")
        except Exception as e:
            logger.warning(f"Failed to load tiktoken encoding {model_encoding}: {e}")

    def _read_pdf(self, path: Path) -> str:
        """Read text content from a PDF file."""
        text = ""
        try:
            from pypdf import PdfReader

            reader = PdfReader(path)
            for page in reader.pages:
                extracted = page.extract_text()
                if extracted:
                    text += extracted + "\n"
        except ImportError:
            logger.debug("pypdf not available, skipping PDF content extraction")
        except Exception as e:
            logger.warning(f"Failed to read PDF {path}: {e}")
        return text

    def _read_docx(self, path: Path) -> str:
        """Read text content from a DOCX file."""
        text = ""
        try:
            from docx import Document

            doc = Document(path)
            text = "\n".join([p.text for p in doc.paragraphs])
        except ImportError:
            logger.debug("python-docx not available, skipping DOCX content extraction")
        except Exception as e:
            logger.warning(f"Failed to read DOCX {path}: {e}")
        return text

    def _read_xlsx(self, path: Path) -> str:
        """Read text content from an XLSX file."""
        text = ""
        try:
            from openpyxl import load_workbook

            wb = load_workbook(path, read_only=True, data_only=True)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    row_text = " ".join([str(cell) for cell in row if cell is not None])
                    if row_text:
                        text += row_text + "\n"
        except ImportError:
            logger.debug("openpyxl not available, skipping XLSX content extraction")
        except Exception as e:
            logger.warning(f"Failed to read XLSX {path}: {e}")
        return text

    def _read_text(self, path: Path) -> str:
        """Read text content from a text file."""
        try:
            return path.read_text(encoding="utf-8", errors="ignore")
        except Exception as e:
            logger.warning(f"Failed to read text file {path}: {e}")
            return ""

    def _get_file_content(self, path: Path) -> str:
        """Get text content from a file based on its extension."""
        suffix = path.suffix.lower()
        if suffix == ".pdf":
            return self._read_pdf(path)
        elif suffix == ".docx":
            return self._read_docx(path)
        elif suffix == ".xlsx":
            return self._read_xlsx(path)
        elif suffix in [
            ".txt",
            ".md",
            ".py",
            ".json",
            ".yaml",
            ".yml",
            ".xml",
            ".html",
            ".css",
            ".js",
            ".ts",
            ".tsx",
            ".jsx",
            ".sh",
            ".toml",
            ".ini",
            ".cfg",
            ".env",
            ".csv",
        ]:
            return self._read_text(path)
        else:
            # Try reading as text for unknown extensions
            return self._read_text(path)

    def _count_tokens(self, text: str) -> int:
        """Count tokens in text using tiktoken or fallback estimation."""
        if not text:
            return 0
        if self.encoding:
            try:
                return len(self.encoding.encode(text))
            except Exception:
                pass

        # Fallback: approximate tokens (0.75 words per token rule of thumb)
        return int(len(text.split()) / 0.75)

    def _detect_language(self, text: str) -> str:
        """Detect the language of the text."""
        if not text or len(text.strip()) < 10:
            return "unknown"
        try:
            from langdetect import DetectorFactory, detect

            DetectorFactory.seed = 0  # Ensure deterministic detection
            # Detect on first 2000 chars for performance
            return detect(text[:2000])
        except ImportError:
            logger.debug("langdetect not available, skipping language detection")
            return "unknown"
        except Exception:
            return "unknown"

    def _analyze_single_file(self, path: Path) -> dict[str, Any]:
        """Analyze a single file and return its metadata."""
        content = self._get_file_content(path)

        tokens = self._count_tokens(content)
        language = self._detect_language(content)
        file_type = path.suffix.lower().lstrip(".") or "unknown"

        return {
            "tokens": tokens,
            "language": language,
            "type": file_type,
            "path": str(path),
            "name": path.name,
            "size": path.stat().st_size if path.exists() else 0,
        }

    async def analyze_files(self, file_paths: list[Path]) -> dict[str, Any]:
        """Analyze a list of files and return statistics.

        Args:
            file_paths: List of paths to files to analyze.

        Returns:
            Dictionary containing:
            - count_per_type: Dict[str, int]
            - language_statistics: Dict[str, str] (percentage)
            - average_size_in_tokens: int
            - more_than_10k_tokens_count: int
            - total_tokens: int
            - files: List of individual file info
        """
        loop = asyncio.get_running_loop()

        # Filter existing files
        existing_files = [p for p in file_paths if p.exists() and p.is_file()]

        if not existing_files:
            return {
                "count_per_type": {},
                "language_statistics": {},
                "average_size_in_tokens": 0,
                "more_than_10k_tokens_count": 0,
                "total_tokens": 0,
                "files": [],
            }

        results = []
        with ThreadPoolExecutor() as pool:
            tasks = [
                loop.run_in_executor(pool, self._analyze_single_file, path)
                for path in existing_files
            ]
            results = await asyncio.gather(*tasks)

        # Aggregate results
        count_per_type: Counter = Counter()
        language_counts: Counter = Counter()
        total_tokens = 0
        more_than_10k = 0
        total_files = len(results)

        for res in results:
            count_per_type[res["type"]] += 1
            if res["language"] != "unknown":
                language_counts[res["language"]] += 1

            total_tokens += res["tokens"]
            if res["tokens"] > 10000:
                more_than_10k += 1

        # Calculate language statistics
        language_statistics = {}
        total_langs = sum(language_counts.values())
        if total_langs > 0:
            for lang, count in language_counts.items():
                percentage = (count / total_langs) * 100
                language_statistics[lang] = f"{percentage:.1f}%"

        avg_tokens = total_tokens / total_files if total_files > 0 else 0

        return {
            "count_per_type": dict(count_per_type),
            "language_statistics": language_statistics,
            "average_size_in_tokens": round(avg_tokens),
            "more_than_10k_tokens_count": more_than_10k,
            "total_tokens": total_tokens,
            "files": results,
        }

    async def analyze_single(self, file_path: str | Path) -> dict[str, Any]:
        """Analyze a single file.

        Args:
            file_path: Path to the file to analyze.

        Returns:
            Dictionary with file metadata (tokens, language, type, path, name, size)
        """
        path = Path(file_path) if isinstance(file_path, str) else file_path
        if not path.exists() or not path.is_file():
            return {
                "tokens": 0,
                "language": "unknown",
                "type": "unknown",
                "path": str(path),
                "name": path.name,
                "size": 0,
                "error": "File not found",
            }

        loop = asyncio.get_running_loop()
        with ThreadPoolExecutor() as pool:
            result = await loop.run_in_executor(pool, self._analyze_single_file, path)
        return result
