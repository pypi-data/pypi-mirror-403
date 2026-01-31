import os
from pathlib import Path
import shutil
import subprocess
import platform
import tempfile
import datetime
from typing import Tuple

WHITE_LIKE_COLORS = [
    "00000000",
    "FFFFFFFF",
    "FFFFFF00",
]


def evaluate_excel(excel_path: Path | str):
    """
    Evaluate Python code that manipulates an Excel file using xlwings.
    """
    if platform.system() == "Linux":
        # Use LibreOffice for Linux
        evaluate_excel_libre(excel_path)
        return
    else:
        # Use xlwings for Windows and MacOS (assuming Excel is installed)
        import xlwings  # type: ignore

        excel_app = xlwings.App(visible=False)
        excel_book = excel_app.books.open(excel_path)
        excel_book.save()
        excel_book.close()
        excel_app.quit()


def evaluate_excel_libre(excel_path: Path | str) -> None:
    """
    Force‑recalculate in place under Linux using LibreOffice.
    Raises subprocess.CalledProcessError if soffice exits abnormally.
    """
    tmp_outdir = tempfile.mkdtemp(prefix="lo_convert_")
    cmd = [
        "soffice",
        "--headless",
        "--nologo",
        "--nofirststartwizard",
        "--norestore",
        "--calc",
        "--convert-to",
        "xlsx",
        "--outdir",
        tmp_outdir,
        os.path.abspath(excel_path),
    ]
    lo_home = Path(tempfile.mkdtemp(prefix="lo_profile_"))
    env = dict(os.environ, HOME=str(lo_home))
    try:
        subprocess.run(
            cmd,
            check=True,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.STDOUT,
            text=True,
            env=env,
        )
        # Determine the converted file name (same base name, .xlsx extension)
        base_name = os.path.splitext(os.path.basename(excel_path))[0] + ".xlsx"
        converted_path = os.path.join(tmp_outdir, base_name)
        # Overwrite the original file with the converted one
        shutil.move(converted_path, excel_path)
    finally:
        # Clean up the temp folder
        shutil.rmtree(tmp_outdir, ignore_errors=True)
        pass


def excel_to_str_repr(excel_path: Path | str, evaluate_formulas=False) -> str:
    from openpyxl import load_workbook

    # Load workbook twice: data_only=True to get the evaluated values,
    # and data_only=False to get the formulas and styles.
    if evaluate_formulas:
        evaluate_excel(excel_path)

    wb_evaluated = load_workbook(excel_path, data_only=True)
    wb_raw = load_workbook(excel_path, data_only=False)

    result = []

    for sheet_name in wb_evaluated.sheetnames:
        sheet_evaluated = wb_evaluated[sheet_name]
        sheet_raw = wb_raw[sheet_name]

        sheet_result = f"Sheet Name: {sheet_name}"
        result.append(sheet_result)

        for row_evaluated, row_raw in zip(
            sheet_evaluated.iter_rows(), sheet_raw.iter_rows()
        ):
            is_row_empty = True

            for cell_evaluated, cell_raw in zip(row_evaluated, row_raw):
                is_default_background = True
                style = []

                if (
                    cell_raw.fill.start_color.index != "00000000"
                    and type(cell_raw.fill.start_color.rgb) is str
                    and cell_raw.fill.start_color.rgb not in WHITE_LIKE_COLORS
                ):
                    is_default_background = False
                    style.append(f"bg:{cell_raw.fill.start_color.rgb}")
                if (
                    cell_raw.font.color
                    and cell_raw.font.color.index != 1
                    and type(cell_raw.font.color.rgb) is str
                ):
                    style.append(f"color:{cell_raw.font.color.rgb}")
                if cell_raw.font.bold:
                    style.append("bold")
                if cell_raw.font.italic:
                    style.append("italic")
                if cell_raw.font.underline:
                    style.append("underline")

                display_value = cell_evaluated.value
                if cell_raw.data_type == "f":
                    cell_raw_val = cell_raw.value
                    if type(cell_raw_val) is not str:
                        cell_raw_val = cell_raw.value.text  # type: ignore
                    display_value = f"{cell_raw_val} -> {cell_evaluated.value}"

                coords = cell_evaluated.coordinate

                if display_value is None and not is_default_background:
                    # If cell is empty but has background color, still include it
                    result.append(f"{coords}: null [{', '.join(style)}]")
                    is_row_empty = False
                elif display_value:
                    style_str = f" [{', '.join(style)}]" if style else ""
                    result.append(f"{coords}: {display_value}{style_str}")
                    is_row_empty = False
            if not is_row_empty:
                result.append("")  # Newline after each row

    return "\n".join(result)


def transform_value(v):
    if isinstance(v, (int, float)):
        v = round(float(v), 2)
    elif isinstance(v, datetime.time):
        v = str(v)[:-3]
    elif isinstance(v, datetime.datetime):
        v = round(
            (v - datetime.datetime(1899, 12, 30)).days
            + (v - datetime.datetime(1899, 12, 30)).seconds / 86400.0,
            0,
        )
    elif isinstance(v, str):
        try:
            v = round(float(v), 2)
        except ValueError:
            pass
    return v


def compare_fill_color(fill1, fill2):
    fgColor1 = fill1.fgColor.rgb if fill1.fgColor else None
    fgColor2 = fill2.fgColor.rgb if fill2.fgColor else None
    bgColor1 = fill1.bgColor.rgb if fill1.bgColor else None
    bgColor2 = fill2.bgColor.rgb if fill2.bgColor else None
    return fgColor1 == fgColor2 and bgColor1 == bgColor2


def compare_font_color(font1, font2):
    # UNSURE if this is actually correct.
    if font1.color and font2.color:
        return font1.color.rgb == font2.color.rgb
    return font1.color is None and font2.color is None


def col_name2num(name):
    """Convert an Excel column name to a column number"""
    num = 0
    for c in name:
        num = num * 26 + (ord(c.upper()) - ord("A") + 1)
    return num


def parse_cell_range(range_str):
    start_cell, end_cell = range_str.split(":")
    start_col, start_row = "", ""
    for char in start_cell:
        if char.isdigit():
            start_row += char
        else:
            start_col += char
    end_col, end_row = "", ""
    for char in end_cell:
        if char.isdigit():
            end_row += char
        else:
            end_col += char
    return (col_name2num(start_col), int(start_row)), (
        col_name2num(end_col),
        int(end_row),
    )


def generate_cell_names(range_str):
    from openpyxl.utils import get_column_letter

    if ":" not in range_str:
        return [range_str]
    (start_col, start_row), (end_col, end_row) = parse_cell_range(range_str)
    columns = [get_column_letter(i) for i in range(start_col, end_col + 1)]
    return [f"{col}{row}" for col in columns for row in range(start_row, end_row + 1)]


def compare_excel_cells(
    ground_truth_path: str, output_path: str, answer_position: str, is_CF: bool = False
) -> Tuple[bool, str]:
    from openpyxl import load_workbook

    wb_gt = load_workbook(ground_truth_path, data_only=True)
    wb_out = load_workbook(output_path, data_only=True)
    sheet_ranges = answer_position.split(",")
    for sheet_range in sheet_ranges:
        if "!" in sheet_range:
            sheet_name, cell_range = sheet_range.split("!")
            sheet_name = sheet_name.strip("'")
        else:
            sheet_name = wb_gt.sheetnames[0]
            cell_range = sheet_range
        if sheet_name not in wb_out.sheetnames:
            return False, f"Worksheet '{sheet_name}' not found in output workbook."
        ws_gt = wb_gt[sheet_name]
        ws_out = wb_out[sheet_name]
        cell_names = generate_cell_names(cell_range)
        for cell_name in cell_names:
            cell_gt = ws_gt[cell_name]
            cell_out = ws_out[cell_name]
            if not transform_value(cell_gt.value) == transform_value(cell_out.value):
                return (
                    False,
                    f"Value mismatch at {cell_name}: expected {cell_gt.value}, got {cell_out.value}",
                )
            if is_CF:
                if not compare_fill_color(cell_gt.fill, cell_out.fill):
                    return False, f"Fill color mismatch at {cell_name}"
                if not compare_font_color(cell_gt.font, cell_out.font):
                    return False, f"Font color mismatch at {cell_name}"
    return True, "All comparisons passed."
