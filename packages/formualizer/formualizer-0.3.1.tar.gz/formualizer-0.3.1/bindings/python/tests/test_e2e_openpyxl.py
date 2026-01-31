from pathlib import Path

import pytest

import formualizer as fz

try:
    import openpyxl  # type: ignore
except Exception:  # pragma: no cover - allow skipping if not present in dev env
    openpyxl = None

pytestmark = pytest.mark.skipif(openpyxl is None, reason="openpyxl not installed")


# The extension module name configured by maturin


def make_wb(tmp: Path) -> Path:
    p = tmp / "e2e.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    # Values
    ws["A1"] = 1
    ws["A2"] = 2
    ws["A3"] = 3
    # Simple formula
    ws["B1"] = "=SUM(A1:A3)"
    # Conditionals
    ws["C1"] = "=IF(B1>3, B1*2, 0)"
    # SUMIFS-like (if supported) else basic SUM as placeholder
    ws["D1"] = "=SUM(A1:A3)"
    wb.save(p)
    return p


def test_openpyxl_roundtrip(tmp_path: Path):
    # Prepare XLSX via openpyxl
    xlsx_path = make_wb(tmp_path)

    # Load the actual workbook from disk using Calamine adapter
    wb = fz.load_workbook(str(xlsx_path), strategy="eager_all")

    # Evaluate and check values
    assert wb.evaluate_cell("Sheet1", 1, 2) == 6.0
    assert wb.evaluate_cell("Sheet1", 1, 3) == 12.0
    assert wb.evaluate_cell("Sheet1", 1, 4) == 6.0


def test_batch_values_and_formulas():
    wb = fz.Workbook()
    s = wb.sheet("Data")

    s.set_values_batch(
        1,
        1,
        2,
        3,
        [
            [1, 2, 3],
            [4, 5, 6],
        ],
    )

    s.set_formulas_batch(
        1,
        4,
        2,
        1,
        [
            ["=SUM(A1:C1)"],
            ["=SUM(A2:C2)"],
        ],
    )

    # Evaluate the formula cells
    assert wb.evaluate_cell("Data", 1, 4) == 6.0
    assert wb.evaluate_cell("Data", 2, 4) == 15.0

    # Check that formulas were stored correctly
    forms = s.get_formulas(fz.RangeAddress("Data", 1, 4, 2, 4))
    assert forms == [["SUM(A1:C1)"], ["SUM(A2:C2)"]]


def test_load_workbook_from_disk(tmp_path: Path):
    """Test loading an actual XLSX file from disk with Calamine."""
    # Create a workbook with openpyxl
    xlsx_path = tmp_path / "test_workbook.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    # Add some values
    ws["A1"] = 100
    ws["A2"] = 200
    ws["A3"] = 300

    # Add formulas
    ws["B1"] = "=A1*2"
    ws["B2"] = "=A2+A3"
    ws["B3"] = "=SUM(A1:A3)"

    # Save the workbook
    wb.save(xlsx_path)

    # Load with formualizer
    fz_wb = fz.load_workbook(str(xlsx_path))

    # Check values were loaded (raw values from workbook)
    sheet = fz_wb.sheet("Data")
    assert sheet.get_cell(1, 1).value == 100.0
    assert sheet.get_cell(2, 1).value == 200.0
    assert sheet.get_cell(3, 1).value == 300.0

    # Check formulas were loaded and can be evaluated
    assert fz_wb.evaluate_cell("Data", 1, 2) == 200.0  # A1*2
    assert fz_wb.evaluate_cell("Data", 2, 2) == 500.0  # A2+A3
    assert fz_wb.evaluate_cell("Data", 3, 2) == 600.0  # SUM(A1:A3)

    # Test using classmethod directly
    fz_wb2 = fz.Workbook.load_path(str(xlsx_path), strategy="eager_all")
    assert fz_wb2.evaluate_cell("Data", 1, 1) == 100.0


def test_formula_evaluation_types():
    """Test that formula evaluation returns the correct types."""
    wb = fz.Workbook()
    s = wb.sheet("Test")

    # Set up integer values
    s.set_value(1, 1, 10)
    s.set_value(2, 1, 20)
    s.set_value(3, 1, 30)

    # Set up various formulas
    s.set_formula(1, 2, "=A1+A2")  # Simple addition
    s.set_formula(2, 2, "=A1*2")  # Multiplication
    s.set_formula(3, 2, "=A1/2")  # Division (should return float)
    s.set_formula(4, 2, "=SUM(A1:A3)")  # SUM function
    s.set_formula(5, 2, "=AVERAGE(A1:A3)")  # AVERAGE (definitely float)

    # Check the types and values through evaluation
    add_result = wb.evaluate_cell("Test", 1, 2)
    assert add_result == 30.0

    mult_result = wb.evaluate_cell("Test", 2, 2)
    assert mult_result == 20.0

    div_result = wb.evaluate_cell("Test", 3, 2)
    assert isinstance(div_result, float)
    assert div_result == 5.0

    sum_result = wb.evaluate_cell("Test", 4, 2)
    assert sum_result == 60.0

    avg_result = wb.evaluate_cell("Test", 5, 2)
    assert isinstance(avg_result, float)
    assert avg_result == 20.0
