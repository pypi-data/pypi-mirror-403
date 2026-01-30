#!/bin/python
# -*- coding: utf-8 -*-
"""
table_to_lists.py

Author: GrimAndGreedy
License: MIT
"""

import sys
import csv
import json
from io import StringIO
import argparse
from typing import Tuple, Iterable, Optional
import dill as pickle
import os
import logging

logger = logging.getLogger('picker_log')

def read_file_content(file_path: str) -> str:
    """ Read lines from file. """
    logger.info("function: read_file_content (table_to_list_of_lists.py)")
    with open(file_path, 'r') as file:
        return file.read()

def strip_whitespace(item: Iterable) -> Iterable:
    """ Strip whitespace from string or from list of strings. """
    logger.info("function: strip_whitespace (table_to_list_of_lists.py)")
    if isinstance(item, list):
        return [strip_whitespace(sub_item) for sub_item in item]
    elif isinstance(item, str):
        return item.strip()
    else:
        return item


def xlsx_to_list(file_name: str, sheet_number:int = 0, extract_formulae: bool = False, first_row_is_header: bool = True):
    import pandas as pd
    from openpyxl import load_workbook
    # wb = load_workbook(filename=input_arg, read_only=True)
    # values or formulae
    if not os.path.exists(file_name):
        return [], [], []
    wb = load_workbook(filename=file_name, read_only=True, data_only=not extract_formulae)
    
    if not isinstance(sheet_number, int): sheet_number = 0
    sheet_number = max(0, min(sheet_number, len(wb.sheetnames)-1))
    ws = wb.worksheets[sheet_number]
    
    # Read data and formulas from the sheet
    table_data = []
    # table_data = [[cell for cell in row] for row in ws.iter_rows(min_row=1, values_only=False)]
    # table_data = [[cell.value for cell in row] for row in ws.iter_rows(min_row=1, values_only=False)]
    table_data = [[cell if cell != None else "" for cell in row] for row in ws.iter_rows(min_row=1, values_only=True)]
    header = []
    # header = [cell for cell in list(ws.iter_rows(values_only=True))[0]]  # Assuming the first row is the header
    if first_row_is_header and len(table_data) > 1:
        header = table_data[0]
        table_data = table_data[1:]
    else:
        header = []
    #
    # for row in ws.iter_rows(min_row=2, values_only=True):  # Skip the header row
    #     row_data = []
    #     for cell in row:
    #         if isinstance(cell, str) and '=' in cell:  # Check if it's a formula
    #             row_data.append(cell)
    #         else:
    #             row_data.append(str(cell))
    #     table_data.append(row_data)
    
    return table_data, header, wb.sheetnames

def ods_to_list(filename: str, sheet_number: int = 0, extract_formulas: bool = False, first_row_is_header: bool = True):
    from odf.opendocument import load
    from odf import table, text

    from odf.namespaces import TABLENS
    # Load the ODS file
    doc = load(filename)

    sheets = doc.spreadsheet.getElementsByType(table.Table)
    sheet_names = [s.attributes.get((TABLENS, 'name')) for s in sheets]


    # Get the sheet by index
    sheet = doc.spreadsheet.getElementsByType(table.Table)[sheet_number]

    data = []
    for row in sheet.getElementsByType(table.TableRow):
        row_data = []
        for cell in row.getElementsByType(table.TableCell):
            if extract_formulas:
                formula = cell.attributes.get((TABLENS, 'formula'))
                if formula is not None:
                    row_data.append(formula)
                    continue  # Skip extracting value if formula found
            
            # Extract value (as text) from <text:p> elements
            cell_text = ""
            for p in cell.getElementsByType(text.P):
                cell_text += str(p.firstChild) if p.firstChild is not None else ""
            row_data.append(cell_text)
        data.append(row_data)
    if first_row_is_header and len(data) > 1:
        header = data[0]
        data = data[1:]
    else:
        header = []

    return data, header, sheet_names

def ods_to_list_old(file_name: str, sheet_number:int = 0, extract_formulae: bool = False, first_row_is_header: bool = True):
    try:
        import pandas as pd
        ef = pd.ExcelFile(file_name)
        sheets = ef.sheet_names
        sheet_number = max(0, min(sheet_number, len(sheets)-1))
        df = pd.read_excel(file_name, engine='odf', sheet_name=sheet_number)
        # if sheet_number < len(sheets):
        #     df = pd.read_excel(input_arg, engine='odf', sheet_name=sheet_number)
        # else:
        #     df = pd.read_excel(input_arg, engine='odf')
        table_data = df.values.tolist()
        table_data = [[x if not pd.isna(x) else "" for x in row] for row in table_data]
        
        try:
            header = list(df.columns)
        except:
            header = []

        if not first_row_is_header and header:
            table_data = [header] + table_data
            header = []

        return table_data, header, sheets
    except Exception as e:
        print(f"Error loading ODS file: {e}")
        return [], [], []


def table_to_list(
    input_arg: str,
    delimiter:str='\t',
    file_type:Optional[str]=None,
    sheet_number:int = 0,
    first_row_is_header:bool = True,

) -> Tuple[list[list[str]], list[str], list[str]]:
    """ 
    Convert data string to list. The input_arg
    Currently accepts: csv, tsv, json, xlsx, ods


    input_arg: filename


    returns:
        items: list[list[str]]
        header: list[str]
        sheets: list[str]
    
    """
    logger.info("function: table_to_list (table_to_list_of_lists.py)")
    table_data = []

    def parse_csv_like2(data:str, delimiter:str) -> list[list[str]]:
        """ Convert value-separated data (e.g., CSV or TSV) to list of lists. """
        logger.info("function: parse_csv_like (table_to_list_of_lists.py)")

        try:
            # reader = csv.reader(StringIO(data), delimiter=delimiter)
            reader = csv.reader(StringIO(data), dialect='unix')
            return [row for row in reader]
        except Exception as e:
            print(f"Error reading CSV-like input: {e}")
            return []
    def parse_csv_like(data:str, delimiter: str=" "):
        import re
        def split_columns(line):
            # Define the regex pattern to match quoted strings and split by whitespace
            # pattern = r"(?:'[^']*'|[^'\s]+)"
            pattern = r"(?:\"[^\"]*\"|'[^']*'|[^'\s]+)"
            
            # Find all matches using the defined pattern
            columns = re.findall(pattern, line)
    
            return columns

        lines = data.strip().split('\n')
        result = []
        
        for line in lines:
            result.append(split_columns(line))
        
        return result

    def csv_string_to_list(csv_string:str, first_row_is_header: bool = True) -> list[list[str]]:
        """ Convert csv string to list of lists using csv.reader. """
        logger.info("function: csv_string_to_list (table_to_list_of_lists.py)")
        f = StringIO(csv_string)
        reader = csv.reader(f, skipinitialspace=True)
        table_data = [row for row in reader]
        if first_row_is_header and len(table_data) > 1:
            header = table_data[0]
            table_data = table_data[1:]
        else:
            header = []
        return table_data, header
            

    if input_arg == '--stdin':
        input_data = sys.stdin.read()
    elif input_arg == '--stdin2':
        input_count = int(sys.stdin.readline())
        input_data = "\n".join([sys.stdin.readline() for i in range(input_count)])
        sys.stdin.flush()
        # sys.stdin.close()
        # sys.stdin = open('/dev/tty', 'r')
    elif file_type == 'csv' or delimiter in [',']:
        try:
            if input_arg == '--stdin':
                input_data = sys.stdin.read()
            elif input_arg == '--stdin2':
                input_count = int(sys.stdin.readline())
                input_data = "\n".join([sys.stdin.readline().strip() for i in range(input_count)])
            else:
                input_data = read_file_content(input_arg)

            table_data, header = csv_string_to_list(input_data, first_row_is_header)
            table_data = strip_whitespace(table_data)
            header = strip_whitespace([header])[0]
            # table_data = parse_csv_like(input_data, ",")
            return table_data, header, []
        except Exception as e:
            print(f"Error reading CSV/TSV input: {e}")
            return [], [], []

    elif file_type == 'tsv':
        try:
            if input_arg == '--stdin':
                input_data = sys.stdin.read()
            elif input_arg == '--stdin2':
                input_count = int(sys.stdin.readline())
                input_data = "\n".join([sys.stdin.readline().strip() for i in range(input_count)])
            else:
                input_data = read_file_content(input_arg)
            
            # Adjust delimiter for TSV or CSV
            if file_type == 'tsv' or delimiter == '\t':
                delimiter = '\t'
            else:
                delimiter = ','
            
            table_data = parse_csv_like(input_data, delimiter)
            table_data = strip_whitespace(table_data)
            return table_data, [], []
        except Exception as e:
            print(f"Error reading CSV/TSV input: {e}")
            return [], [], []

    elif file_type == 'json':
        try:
            if input_arg == '--stdin':
                input_data = sys.stdin.read()
            elif input_arg == '--stdin2':
                input_count = int(sys.stdin.readline())
                input_data = "\n".join([sys.stdin.readline() for i in range(input_count)])
            else:
                input_data = read_file_content(input_arg)
            
            table_data = json.loads(input_data)
            return table_data, [], []
        except json.JSONDecodeError as e:
            print(f"Error decoding JSON input: {e}")
            return [], [], []
        except FileNotFoundError as e:
            print(f"File not found: {e}")
            return [], [], []

    elif file_type == 'xlsx':
        extract_formulae = False
        return xlsx_to_list(input_arg, sheet_number, extract_formulae, first_row_is_header)

    elif file_type == 'ods':
        extract_formulae = False
        return ods_to_list(input_arg, sheet_number, extract_formulae, first_row_is_header)
    elif file_type == 'pkl':
        with open(os.path.expandvars(os.path.expanduser(input_arg)), 'rb') as f:
            loaded_data = pickle.load(f)
        items = loaded_data["items"] if "items" in loaded_data else []
        header = loaded_data["header"] if "header" in loaded_data else []
        return items, header, []

    else:
        input_data = read_file_content(input_arg)
    
    table_data = parse_csv_like(input_data, delimiter)
    if first_row_is_header and len(table_data) > 1:
        header = table_data[0]
        table_data = table_data[1:]
    else:
        header = []


    return table_data, header, []

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Convert table to list of lists.')
    parser.add_argument('-i', dest='file', help='File containing the table to be converted')
    parser.add_argument('--stdin', action='store_true', help='Table passed on stdin')
    parser.add_argument('--stdin2', action='store_true', help='Table passed on stdin')
    parser.add_argument('-d', dest='delimiter', default='\t', help='Delimiter for rows in the table (default: tab)')
    parser.add_argument('-t', dest='file_type', choices=['tsv', 'csv', 'json', 'xlsx', 'ods'], help='Type of file (tsv, csv, json, xlsx, ods)')
    
    args = parser.parse_args()
    
    if args.file:
        input_arg = args.file
    elif args.stdin:
        input_arg = '--stdin'
    elif args.stdin2:
        input_arg = '--stdin2'
    else:
        print("Error: Please provide input file or use --stdin option.")
        sys.exit(1)
    
    table_data = table_to_list(input_arg, args.delimiter, args.file_type)
    # print(table_data)

    len(table_data[0])
    for row in table_data:
        if len(row) != len(table_data[0]):
            print(len(row), row)
