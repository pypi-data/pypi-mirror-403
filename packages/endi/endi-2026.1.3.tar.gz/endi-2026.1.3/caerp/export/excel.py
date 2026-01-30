import logging
from typing import Optional

import openpyxl
import io

from openpyxl.styles.numbers import (
    FORMAT_NUMBER_00,
)
from openpyxl.utils import get_column_letter
from zope.interface import (
    implementer,
)
from openpyxl.styles import (
    Color,
    fills,
    NamedStyle,
    PatternFill,
    Font,
)

from caerp.export.accounting_spreadsheet import CellContent
from caerp.interfaces import IExporter
from caerp.export.utils import write_file_to_request

logger = logging.getLogger(__name__)

Color.LightCyan = "FFE0FFFF"
Color.LightCoral = "FFF08080"
Color.LightGreen = "FF90EE90"
Color.Crimson = "FFDC143C"
Color.header = "FFD9EDF7"
Color.footer = "FFFCF8E3"
Color.highlight = "FFEFFFEF"
EXCEL_NUMBER_FORMAT = "#,##0.00"
EXCEL_NUMBER_FORMAT_WO_DECIMALS = "#,##0"


def get_number_format(decimal_places=0, percentage=False) -> str:
    """
    Build excel number formatting string

    >>> get_number_format()
    '#,##0'

    >>> get_number_format(1)
    '#,##0.0'

    >>> get_number_format(1, True)
    '#,##0.0%'

    :param decimal_places:
    :param percentage:
    :return:
    """
    ret = EXCEL_NUMBER_FORMAT_WO_DECIMALS
    if decimal_places > 0:
        ret += f".{decimal_places*'0'}"
    if percentage:
        ret += "%"

    return ret


TITLE_STYLE = NamedStyle(font=Font(size=16, bold=True), name="caerp-title")
HEADER_STYLE = NamedStyle(
    font=Font(bold=True),
    fill=PatternFill(fill_type=fills.FILL_SOLID, start_color=Color(rgb=Color.header)),
    name="caerp-header",
)
BOLD_CELL = NamedStyle(font=Font(bold=True), name="caerp-bold-cell")
NUMBER_CELL = NamedStyle(name="caerp-number-cell")
HIGHLIGHTED_ROW_STYLE = NamedStyle(
    font=Font(bold=True),
    fill=PatternFill(
        fill_type=fills.FILL_SOLID, start_color=Color(rgb=Color.highlight)
    ),
    name="caerp-highlight-cell",
)

FOOTER_CELL = NamedStyle(
    font=Font(bold=True),
    fill=PatternFill(fill_type=fills.FILL_SOLID, start_color=Color(rgb=Color.footer)),
    number_format=FORMAT_NUMBER_00,
    name="caerp-footer-cell",
)
LARGE_FOOTER_CELL = NamedStyle(
    font=Font(bold=True, size=16),
    fill=PatternFill(fill_type=fills.FILL_SOLID, start_color=Color(rgb=Color.footer)),
    number_format=FORMAT_NUMBER_00,
    name="caerp-large-footer-cell",
)


@implementer(IExporter)
class XlsExporter:
    """
    Those options are implemented :

    row-level : highlight, hidden
    document-level: decimal_places
    """

    title = "Export"

    WIDE_COLUMN_WIDTH = 50

    def __init__(self, worksheet=None, options=None, **kw):
        options = options or {}
        self.default_decimal_places = int(options.get("decimal_places", "2"))
        self._init_styles()
        if worksheet is None:
            self.book = openpyxl.workbook.Workbook()
            self.worksheet = self.book.active
            self.worksheet.title = self.title
        else:
            self.worksheet = worksheet
            self.book = worksheet.parent
        self.options = kw
        self.current_line = 1

    def _init_styles(self):
        self.default_cell_style = NUMBER_CELL.__copy__()
        self.highlight_cell_style = HIGHLIGHTED_ROW_STYLE.__copy__()

        number_format = get_number_format(self.default_decimal_places)

        self.default_cell_style.number_format = number_format
        self.highlight_cell_style.number_format = number_format

    def add_title(self, title, width):
        self.worksheet.merge_cells(
            start_row=self.current_line,
            end_row=self.current_line,
            start_column=1,
            end_column=width - 1,
        )
        cell = self.worksheet.cell(row=self.current_line, column=1)
        cell.value = title
        cell.style = TITLE_STYLE
        row_dim = self.worksheet.row_dimensions[self.current_line]
        row_dim.height = 20
        self.current_line += 1

    def add_breakline(self):
        self.current_line += 1

    def _add_row(self, labels, styles=None):
        for col_index, label in enumerate(labels):
            cell = self.worksheet.cell(row=self.current_line, column=col_index + 1)
            self._fill_cell(cell, label, styles)

    def _fill_cell(self, cell, label, style=None):
        cell.value = label
        if style:
            cell.style = style
        if (
            isinstance(label, CellContent)
            and label.style_variant == CellContent.STYLE_VARIANT_PERCENTAGE
        ):
            cell.number_format = get_number_format(
                self.default_decimal_places, percentage=True
            )

    def add_headers(self, labels):
        self._add_row(labels, HEADER_STYLE)
        self.current_line += 1

    def set_column_options(
        self, column_index: int, column_style_name=Optional[str]
    ) -> None:
        """Sets column options/styles"""
        column_letter = get_column_letter(column_index + 1)
        if column_style_name == "wide_column":
            self.worksheet.column_dimensions[
                column_letter
            ].width = self.WIDE_COLUMN_WIDTH

    def add_row(self, labels, options=None):
        options = options or {}
        if options.get("highlight", False):
            styles = self.highlight_cell_style
        else:
            styles = self.default_cell_style
        if options.get("hidden", False):
            self.worksheet.row_dimensions[self.current_line].hidden = True

        self._add_row(labels, styles)
        self.current_line += 1

    def save_book(self, f_buf=None):
        """
        Return a file buffer containing the resulting xls

        :param obj f_buf: A file buffer supporting the write and seek
        methods
        """
        if f_buf is None:
            f_buf = io.BytesIO()
        self.book.save(f_buf)
        f_buf.seek(0)
        return f_buf

    def render(self, f_buf=None):
        """
        Definitely render the workbook

        :param obj f_buf: A file buffer supporting the write and seek
        methods
        """
        if f_buf is None:
            f_buf = io.BytesIO()

        return self.save_book(f_buf)


def make_excel_view(filename_builder, factory):
    """
    Build an excel view of a model
    :param filename_builder: a callable that take the request as arg and
        return a filename
    :param factory: the Xls factory that should be used to wrap the
        request context the factory should provide a render method
        returning a file like object
    """

    def _view(request):
        """
        the dynamically built view object
        """
        filename = filename_builder(request)
        result = factory(request.context).render()
        request = write_file_to_request(request, filename, result)
        return request.response

    return _view
