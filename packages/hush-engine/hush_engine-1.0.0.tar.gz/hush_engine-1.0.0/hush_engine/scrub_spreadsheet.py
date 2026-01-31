#!/usr/bin/env python3
"""
CLI tool for scrubbing spreadsheets (CSV, XLSX)
Usage: python scrub_spreadsheet.py <input_file> [output_file]
"""

import sys
import argparse
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from typing import Dict, List, Tuple

from detectors import PIIDetector
from anonymizers import SpreadsheetAnonymizer


class SpreadsheetScrubber:
    """
    High-level interface for scrubbing spreadsheet files
    """

    def __init__(self, sample_size: int = 20, confidence_threshold: float = 0.5):
        """
        Initialize the scrubber

        Args:
            sample_size: Number of rows to sample per column for detection
            confidence_threshold: Minimum confidence for PII detection
        """
        self.detector = PIIDetector()
        self.anonymizer = SpreadsheetAnonymizer()
        self.sample_size = sample_size
        self.confidence_threshold = confidence_threshold

    def analyze_columns(self, df: pd.DataFrame) -> Dict[str, Dict]:
        """
        Analyze DataFrame columns for PII

        Args:
            df: Input DataFrame

        Returns:
            Dictionary mapping column names to detection info
        """
        results = {}

        for column in df.columns:
            # Skip columns with all null values
            non_null_values = df[column].dropna()
            if len(non_null_values) == 0:
                continue

            # Sample values from the column
            sample = non_null_values.sample(
                n=min(self.sample_size, len(non_null_values)),
                random_state=42
            )

            # Analyze concatenated sample
            sample_text = " | ".join(str(val) for val in sample)
            entities = self.detector.analyze_text(sample_text)

            # Filter by confidence
            high_confidence_entities = [
                e for e in entities
                if e.confidence >= self.confidence_threshold
            ]

            if high_confidence_entities:
                # Determine primary entity type (most common)
                entity_types = [e.entity_type for e in high_confidence_entities]
                primary_type = max(set(entity_types), key=entity_types.count)

                # Calculate PII density (what % of sample has PII)
                pii_count = len(high_confidence_entities)
                density = pii_count / len(sample)

                results[column] = {
                    'primary_type': primary_type,
                    'entity_types': list(set(entity_types)),
                    'entities': high_confidence_entities,
                    'density': density,
                    'sample_size': len(sample)
                }

        return results

    def scrub_dataframe(
        self,
        df: pd.DataFrame,
        column_analysis: Dict[str, Dict] = None,
        auto_detect: bool = True
    ) -> Tuple[pd.DataFrame, Dict]:
        """
        Scrub a DataFrame

        Args:
            df: Input DataFrame
            column_analysis: Pre-computed column analysis (optional)
            auto_detect: Automatically detect and scrub all PII columns

        Returns:
            Tuple of (scrubbed_df, analysis_results)
        """
        # Analyze columns if not provided
        if column_analysis is None and auto_detect:
            column_analysis = self.analyze_columns(df)

        if not column_analysis:
            return df.copy(), {}

        # Build entity map for anonymizer
        entity_map = {}
        for column, info in column_analysis.items():
            entity_map[column] = info['entity_types']

        # Anonymize
        scrubbed_df = self.anonymizer.anonymize_dataframe(df, entity_map)

        return scrubbed_df, column_analysis

    def scrub_file(
        self,
        input_path: str,
        output_path: str = None,
        preserve_formatting: bool = True
    ) -> Dict:
        """
        Scrub a CSV or XLSX file

        Args:
            input_path: Path to input file
            output_path: Path to output file (default: adds _scrubbed suffix)
            preserve_formatting: For XLSX, try to preserve cell formatting

        Returns:
            Dictionary with scrubbing results
        """
        input_file = Path(input_path)

        # Set default output path
        if output_path is None:
            output_path = str(input_file.parent / f"{input_file.stem}_scrubbed{input_file.suffix}")

        # Determine file type
        file_ext = input_file.suffix.lower()

        if file_ext == '.csv':
            return self._scrub_csv(input_path, output_path)
        elif file_ext in ['.xlsx', '.xls']:
            return self._scrub_xlsx(input_path, output_path, preserve_formatting)
        else:
            raise ValueError(f"Unsupported file type: {file_ext}. Use .csv or .xlsx")

    def _scrub_csv(self, input_path: str, output_path: str) -> Dict:
        """Scrub a CSV file"""
        # Read CSV
        df = pd.read_csv(input_path)

        print(f"Loaded CSV: {df.shape[0]} rows, {df.shape[1]} columns")

        # Analyze and scrub
        column_analysis = self.analyze_columns(df)
        scrubbed_df, _ = self.scrub_dataframe(df, column_analysis)

        # Write CSV
        scrubbed_df.to_csv(output_path, index=False)

        return {
            'input': input_path,
            'output': output_path,
            'rows': df.shape[0],
            'columns': df.shape[1],
            'pii_columns': len(column_analysis),
            'column_analysis': column_analysis
        }

    def _scrub_xlsx(self, input_path: str, output_path: str, preserve_formatting: bool) -> Dict:
        """Scrub an XLSX file"""
        # Read Excel
        df = pd.read_excel(input_path, engine='openpyxl')

        print(f"Loaded XLSX: {df.shape[0]} rows, {df.shape[1]} columns")

        # Analyze and scrub
        column_analysis = self.analyze_columns(df)
        scrubbed_df, _ = self.scrub_dataframe(df, column_analysis)

        if preserve_formatting:
            # Try to preserve formatting by loading workbook
            try:
                wb = load_workbook(input_path)
                ws = wb.active

                # Update cell values while preserving formatting
                # Map DataFrame columns to Excel columns
                for col_idx, column_name in enumerate(scrubbed_df.columns, start=1):
                    for row_idx, value in enumerate(scrubbed_df[column_name], start=2):  # start=2 to skip header
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.value = value

                wb.save(output_path)
                print("  → Preserved cell formatting")

            except Exception as e:
                print(f"  → Warning: Could not preserve formatting: {e}")
                # Fallback to standard pandas write
                scrubbed_df.to_excel(output_path, index=False, engine='openpyxl')
        else:
            scrubbed_df.to_excel(output_path, index=False, engine='openpyxl')

        return {
            'input': input_path,
            'output': output_path,
            'rows': df.shape[0],
            'columns': df.shape[1],
            'pii_columns': len(column_analysis),
            'column_analysis': column_analysis
        }


def main():
    """CLI entry point"""
    parser = argparse.ArgumentParser(
        description='Scrub PII from spreadsheets (CSV, XLSX)',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Scrub a CSV file
  python scrub_spreadsheet.py customers.csv

  # Scrub an Excel file with custom output
  python scrub_spreadsheet.py data.xlsx scrubbed_data.xlsx

  # Adjust sampling and confidence
  python scrub_spreadsheet.py data.csv --sample-size 50 --confidence 0.7
        """
    )

    parser.add_argument('input', help='Input spreadsheet file (.csv or .xlsx)')
    parser.add_argument('output', nargs='?', help='Output file (default: input_scrubbed.ext)')
    parser.add_argument(
        '--sample-size',
        type=int,
        default=20,
        help='Number of rows to sample per column (default: 20)'
    )
    parser.add_argument(
        '--confidence',
        type=float,
        default=0.5,
        help='Minimum confidence for PII detection (default: 0.5)'
    )
    parser.add_argument(
        '--no-formatting',
        action='store_true',
        help='Do not preserve Excel cell formatting'
    )

    args = parser.parse_args()

    # Validate input file
    if not Path(args.input).exists():
        print(f"Error: Input file not found: {args.input}", file=sys.stderr)
        return 1

    try:
        scrubber = SpreadsheetScrubber(
            sample_size=args.sample_size,
            confidence_threshold=args.confidence
        )

        print(f"Scrubbing spreadsheet: {args.input}")
        print(f"Sample size: {args.sample_size}, Confidence threshold: {args.confidence}\n")

        print("Step 1/3: Loading file...")
        result = scrubber.scrub_file(
            args.input,
            args.output,
            preserve_formatting=not args.no_formatting
        )

        print(f"\nStep 2/3: Analyzing columns for PII...")
        for column, info in result['column_analysis'].items():
            print(f"  → Column '{column}': {info['primary_type']} "
                  f"(density: {info['density']:.1%}, types: {', '.join(info['entity_types'])})")

        print(f"\nStep 3/3: Anonymizing...")
        print(f"  → Scrubbed {result['pii_columns']} columns with PII")
        print(f"  → Saved to: {result['output']}")

        print("\n" + "=" * 60)
        print("✓ Scrubbing complete!")
        print("=" * 60)
        print(f"  Total rows: {result['rows']}")
        print(f"  Total columns: {result['columns']}")
        print(f"  PII columns: {result['pii_columns']}")

        return 0

    except Exception as e:
        print(f"\n✗ Error: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())
