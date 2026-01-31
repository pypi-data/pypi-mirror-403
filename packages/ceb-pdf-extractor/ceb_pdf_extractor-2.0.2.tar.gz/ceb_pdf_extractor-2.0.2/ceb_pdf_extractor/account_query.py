"""账户交易明细查询模板 - 专用于账户交易明细查询PDF"""

import fitz
import os
import re
from collections import defaultdict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter


class PDFTableExtractor_AccountQuery:
    """账户交易明细查询PDF表格数据提取器"""
    def __init__(self, pdf_path, target_font="FangSong", target_size=10.0,
                 size_tolerance=0.2, row_tolerance=13.359375, col_tolerance=30.0,
                 output_mode="multiple"):  # 新增output_mode参数
        self.pdf_path = os.path.abspath(pdf_path)
        self.pdf_filename = os.path.splitext(os.path.basename(self.pdf_path))[0]
        self.output_dir = os.path.join(os.path.dirname(self.pdf_path),
                                      f"光大银行pdf转excel({self.pdf_filename})")
        self.doc = fitz.open(self.pdf_path)

        # 输出模式："multiple"为多个Excel文件，"single"为单个总Excel文件
        self.output_mode = output_mode

        # 提取参数
        self.target_font = target_font
        self.target_size = target_size
        self.size_tolerance = size_tolerance
        self.row_tolerance = row_tolerance  # 13.359375作为行容差
        self.col_tolerance = col_tolerance

        # 账户交易明细查询配置
        self.customer_info_fields = {
            '账户名称': [r'账户名称[：:]\s*([^\s]+)', r'户名[：:]\s*([^\s]+)'],
            '系统账号': [r'系统账号[：:]\s*([^\s]+)'],
            '查询起止日期': [r'查询起止日期[：:]\s*([^\s]+)'],
            '转出笔数': [r'转出笔数[：:]\s*(\d+)'],
            '存入笔数': [r'存入笔数[：:]\s*(\d+)'],
            '账户类型': [r'账户类型[：:]\s*([^\s]+)'],
            '客户账号': [r'客户账号[：:]\s*([^\s]+)', r'账号[：:]\s*([^\s]+)'],
            '交易总笔数': [r'交易总笔数[：:]\s*(\d+)'],
            '转出金额': [r'转出金额[：:]\s*([\d,\.]+)'],
            '存入金额': [r'存入金额[：:]\s*([\d,\.]+)']
        }

        self.supplementary_columns = [
            '账户名称', '系统账号', '查询起止日期', '转出笔数', '存入笔数',
            '账户类型', '客户账号', '交易总笔数', '转出金额', '存入金额'
        ]

        self.report_title = '光大银行PDF对账单提取报告（账户交易明细查询版）'

        # 存储数据
        self.customer_data = defaultdict(list)
        self.all_data_pages = []  # 存储所有页面的数据（按原顺序）
        self.extracted_files = []
        self.skipped_pages = []

        self._create_output_dir()

    def _create_output_dir(self):
        """创建输出目录"""
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)
            print(f"✅ 创建输出目录: {self.output_dir}")

    def _extract_customer_info(self, page):
        """从页面提取客户信息（简化版：只检查账户名称）"""
        full_text = page.get_text("text")
        info = {}

        for field_name, patterns in self.customer_info_fields.items():
            value = ''
            for pattern in patterns:
                match = re.search(pattern, full_text)
                if match:
                    if match.groups():
                        value = match.group(1).strip()
                        if value:
                            break
                        else:
                            value = ' '
                            break
                    else:
                        value = match.group(0).strip()
                        break
                if not value:
                    value = ' '
            info[field_name] = value

        # 简化：只检查账户名称是否非空/非空格
        account_name = info.get('账户名称', '').strip()
        if not account_name:
            return None

        return info

    def _has_no_transaction_details(self, page):
        """检查是否存在交易明细"""
        full_text = page.get_text("text")
        return "不存在交易明细" in full_text \
               or "无符合条件的开户记录" in full_text \
               or "无明细" in full_text

    def _extract_table_cells(self, page):
        """提取页面中的表格单元格数据 - 从"借贷"开始提取"""
        text_dict = page.get_text("dict")
        cells = []

        # 先找到"借贷"的位置
        loan_y_start = None
        for block in text_dict.get("blocks", []):
            if block["type"] == 0:  # 文本块
                for line in block.get("lines", []):
                    for span in line.get("spans", []):
                        if "借贷" in span["text"]:
                            loan_y_start = span["bbox"][1]  # y_start坐标
                            print(f"    找到'借贷'位置，y_start: {loan_y_start}")
                            break
                    if loan_y_start is not None:
                        break
                if loan_y_start is not None:
                    break

        # 如果没有找到"借贷"，则从页面开始提取
        if loan_y_start is None:
            loan_y_start = 0
            print(f"    未找到'借贷'，从页面开始提取")

        # 提取"借贷"及其之后的表格数据
        for block in text_dict.get("blocks", []):
            if block["type"] == 0:  # 文本块
                for line in block.get("lines", []):
                    for span in line.get("spans", []):
                        # 匹配指定字体和大小的文本，且位置在"借贷"之后（包含"借贷"本身）
                        font_match = self.target_font.lower() in span["font"].lower() if self.target_font else True
                        size_match = abs(span["size"] - self.target_size) <= self.size_tolerance

                        if font_match and size_match:
                            text = span["text"].strip()
                            # 只提取y_start大于等于"借贷"位置的文本
                            if text and span["bbox"][1] >= loan_y_start:
                                cells.append({
                                    "text": text,
                                    "bbox": span["bbox"],
                                    "x_center": (span["bbox"][0] + span["bbox"][2]) / 2,
                                    "y_center": (span["bbox"][1] + span["bbox"][3]) / 2,
                                    "y_start": span["bbox"][1],
                                    "y_end": span["bbox"][3],
                                })

        print(f"    提取到 {len(cells)} 个单元格")
        return cells

    def _group_cells_into_rows(self, cells):
        """将单元格按行分组 - 改进的合并算法，避免生成空行"""
        if not cells:
            return []

        # 按x_center分组，找出同一列上的单元格
        columns = {}
        for cell in cells:
            # 找到最接近的列位置
            x_center = cell["x_center"]
            found_column = False

            for col_x in columns.keys():
                if abs(x_center - col_x) <= self.col_tolerance:
                    columns[col_x].append(cell)
                    found_column = True
                    break

            if not found_column:
                columns[x_center] = [cell]

        print(f"    识别出 {len(columns)} 列")

        # 对每一列的单元格按y_center排序
        for col_x, col_cells in columns.items():
            col_cells.sort(key=lambda c: c["y_center"])

        # 合并同一列中y中心坐标差接近row_tolerance的连续单元格
        merged_cells = []

        for col_x, col_cells in columns.items():
            if not col_cells:
                continue

            i = 0
            while i < len(col_cells):
                current_cell = col_cells[i]
                merged_text = current_cell["text"]
                # 初始合并的y中心坐标
                merged_y_center = current_cell["y_center"]
                # 记录合并单元格的边界
                merged_y_start = current_cell["y_start"]
                merged_y_end = current_cell["y_end"]

                # 尝试合并后续单元格
                j = i + 1
                while j < len(col_cells):
                    next_cell = col_cells[j]

                    # 计算当前合并单元格的y中心坐标与下一个单元格y中心坐标的差值
                    y_diff = next_cell["y_center"] - merged_y_center

                    # 如果差值接近row_tolerance（允许±20%的误差）
                    if abs(y_diff - self.row_tolerance) <= self.row_tolerance * 0.2:
                        # 合并文本
                        merged_text += next_cell["text"]
                        # 更新合并后的y中心坐标为下一个单元格的y中心坐标
                        merged_y_center = next_cell["y_center"]
                        # 更新合并单元格的底部边界
                        merged_y_end = next_cell["y_end"]
                        j += 1
                    else:
                        # 差值不符合条件，停止合并
                        break

                # 计算合并后单元格的最终y中心坐标
                final_y_center = (merged_y_start + merged_y_end) / 2

                # 添加合并后的单元格
                merged_cells.append({
                    "text": merged_text,
                    "bbox": [current_cell["bbox"][0], merged_y_start,
                            current_cell["bbox"][2], merged_y_end],
                    "x_center": col_x,
                    "y_center": final_y_center,  # 使用计算后的最终y中心坐标
                    "y_start": merged_y_start,
                })

                i = j  # 跳过已合并的单元格

        # 现在按y_center分组行 - 使用更宽松的容差
        merged_cells.sort(key=lambda c: c["y_center"])
        rows = []

        for cell in merged_cells:
            placed = False

            # 查找是否已经有相近y_center的行
            for row in rows:
                if row:
                    # 计算当前行所有单元格的平均y_center
                    row_y_centers = [c["y_center"] for c in row]
                    avg_y_center = sum(row_y_centers) / len(row_y_centers)

                    # 使用更宽松的容差（1.5倍行容差）
                    if abs(cell["y_center"] - avg_y_center) <= self.row_tolerance * 1.5:
                        row.append(cell)
                        placed = True
                        break

            # 如果没有找到相近的行，创建新行
            if not placed:
                rows.append([cell])

        # 每行按x_center排序
        for row in rows:
            row.sort(key=lambda c: c["x_center"])

        print(f"    分组为 {len(rows)} 行")
        return rows

    def _detect_columns_from_header(self, rows):
        """从表头行检测列位置"""
        if not rows:
            return []

        header_row = rows[0]
        header_x_positions = [cell["x_center"] for cell in header_row]
        header_x_positions.sort()

        column_positions = []
        for x in header_x_positions:
            if not column_positions:
                column_positions.append(x)
            else:
                min_distance = min(abs(x - pos) for pos in column_positions)
                if min_distance <= self.col_tolerance:
                    closest_idx = min(range(len(column_positions)),
                                      key=lambda i: abs(x - column_positions[i]))
                    column_positions[closest_idx] = (column_positions[closest_idx] + x) / 2
                else:
                    column_positions.append(x)

        column_positions.sort()
        return column_positions

    def _assign_cells_to_matrix(self, rows, column_positions):
        """将单元格分配到矩阵中，自动填补空单元格"""
        if not rows or not column_positions:
            return []

        max_cols = len(column_positions)
        matrix = []

        for row_cells in rows:
            matrix_row = [""] * max_cols

            for cell in row_cells:
                # 找到最接近的列
                closest_idx = 0
                min_distance = float('inf')

                for i, col_x in enumerate(column_positions):
                    distance = abs(cell["x_center"] - col_x)
                    if distance < min_distance:
                        min_distance = distance
                        closest_idx = i

                # 如果距离在容差范围内，分配到该列
                if min_distance <= self.col_tolerance * 2:
                    matrix_row[closest_idx] = cell["text"]

            matrix.append(matrix_row)

        return matrix

    def _compact_matrix(self, matrix):
        """压缩矩阵：如果某个格子为空，下面的格子自动往上顶"""
        if not matrix or len(matrix) < 2:
            return matrix

        # 获取最大列数
        max_cols = max(len(row) for row in matrix)

        # 转置矩阵，按列处理
        transposed = []
        for col_idx in range(max_cols):
            column = []
            for row_idx in range(len(matrix)):
                if col_idx < len(matrix[row_idx]):
                    column.append(matrix[row_idx][col_idx])
                else:
                    column.append("")
            transposed.append(column)

        # 对每一列进行压缩：去除空值，保持顺序
        compacted_transposed = []
        for column in transposed:
            # 移除空字符串
            non_empty = [cell for cell in column if cell and str(cell).strip()]
            # 保持原顺序，空位用空字符串填充
            compacted_column = non_empty + [""] * (len(column) - len(non_empty))
            compacted_transposed.append(compacted_column)

        # 转置回来
        compacted_matrix = []
        num_rows = len(matrix)
        for row_idx in range(num_rows):
            row = []
            for col_idx in range(len(compacted_transposed)):
                if row_idx < len(compacted_transposed[col_idx]):
                    row.append(compacted_transposed[col_idx][row_idx])
                else:
                    row.append("")
            compacted_matrix.append(row)

        return compacted_matrix

    def _cells_to_matrix(self, cells):
        """将单元格列表转换为矩阵，并进行压缩"""
        if not cells:
            return []

        rows = self._group_cells_into_rows(cells)
        if not rows:
            return []

        column_positions = self._detect_columns_from_header(rows)
        if not column_positions:
            return []

        matrix = self._assign_cells_to_matrix(rows, column_positions)

        # 压缩矩阵：空单元格自动往上顶
        compacted_matrix = self._compact_matrix(matrix)

        return compacted_matrix

    def scan_pages(self):
        """扫描所有页面，提取表格数据"""
        # 用于跟踪当前处理的账户信息
        current_customer_info = None

        for page_num in range(len(self.doc)):
            page = self.doc[page_num]
            print(f"\n处理第 {page_num + 1} 页...")

            # 检查页面是否包含"不存在交易明细"，直接跳过
            if self._has_no_transaction_details(page):
                print(f"  第{page_num + 1}页: 不存在交易明细，跳过")
                self.skipped_pages.append(page_num + 1)
                continue

            # 提取账户信息
            customer_info = self._extract_customer_info(page)

            # 如果没有提取到账户信息，使用上一页的信息
            if not customer_info and current_customer_info:
                customer_info = current_customer_info.copy()
                print(f"  第{page_num + 1}页: 使用上一页的账户信息")
            elif customer_info:
                current_customer_info = customer_info.copy()
                print(f"  第{page_num + 1}页: 提取到账户信息")

            # 提取表格数据（从"借贷"开始）
            cells = self._extract_table_cells(page)
            if not cells:
                print(f"  第{page_num + 1}页: 无表格数据，跳过")
                continue

            # 转换为矩阵
            matrix = self._cells_to_matrix(cells)
            if not matrix or len(matrix) == 0:
                print(f"  第{page_num + 1}页: 无有效表格数据，跳过")
                continue

            # 存储原始页面数据（保持原始顺序）
            page_data = {
                "page_num": page_num + 1,
                "customer_info": customer_info,
                "matrix": matrix,
                "has_customer_info": bool(customer_info)
            }
            self.all_data_pages.append(page_data)

            # 按账户分组数据（用于multiple模式）
            if customer_info:
                customer_name = customer_info.get('账户名称', '未知').strip()
                customer_key = customer_name  # 简化：只使用账户名称

                print(f"  第{page_num + 1}页: 账户名称 - {customer_name}")
                print(f"    提取表格: {len(matrix)}行 × {len(matrix[0])}列")

                # 初始化该账户的数据（保留表头）
                if customer_key not in self.customer_data:
                    self.customer_data[customer_key] = {
                        "pages": [],
                        "total_rows": 0
                    }

                # 添加页面数据（包含该页的customer_info）
                self.customer_data[customer_key]["pages"].append({
                    "page_num": page_num + 1,
                    "customer_info": customer_info,
                    "matrix": matrix,
                    "has_header": True
                })
                self.customer_data[customer_key]["total_rows"] += len(matrix)
            elif current_customer_info:
                # 如果没有提取到账户信息，但之前有账户，则追加到当前账户
                customer_name = current_customer_info.get('账户名称', '未知').strip()
                customer_key = customer_name

                if customer_key in self.customer_data:
                    print(f"  第{page_num + 1}页: 追加到当前账户 {customer_name}")
                    self.customer_data[customer_key]["pages"].append({
                        "page_num": page_num + 1,
                        "customer_info": current_customer_info,
                        "matrix": matrix,
                        "has_header": False
                    })
                    self.customer_data[customer_key]["total_rows"] += len(matrix)

        return self.customer_data

    def _remove_duplicate_headers(self, all_rows):
        """移除重复的表头行"""
        if not all_rows or len(all_rows) < 2:
            return all_rows

        # 假设第一行是表头
        header = all_rows[0]
        result = [header]

        # 从第二行开始检查，移除与表头完全相同的行
        for i in range(1, len(all_rows)):
            if all_rows[i] != header:
                result.append(all_rows[i])

        return result

    def _apply_excel_format(self, worksheet, data_rows):
        """应用Excel格式设置"""
        if not data_rows:
            return

        max_row = len(data_rows)
        max_col = max(len(row) for row in data_rows) if data_rows else 0

        if max_row == 0 or max_col == 0:
            return

        no_border = Border(
            left=Side(style='none'),
            right=Side(style='none'),
            top=Side(style='none'),
            bottom=Side(style='none')
        )

        for col in range(1, max_col + 1):
            max_length = 0
            for row in range(1, max_row + 1):
                cell_value = worksheet.cell(row=row, column=col).value
                if cell_value:
                    content = str(cell_value)
                    chinese_count = sum(1 for char in content if '\u4e00' <= char <= '\u9fff')
                    length = len(content) + chinese_count
                    max_length = max(max_length, length)

            if max_length > 0:
                column_letter = get_column_letter(col)
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
                cell.number_format = '@'
                cell.border = no_border

        if max_row > 0:
            for col in range(1, max_col + 1):
                header_cell = worksheet.cell(row=1, column=col)
                header_cell.font = Font(bold=True)

    def generate_multiple_excel_files(self):
        """生成多个Excel文件（按账户名称）"""
        if not self.customer_data:
            print("\n⚠️ 未找到有效的客户数据")
            return []

        excel_files_info = []

        for customer_key, account_data in self.customer_data.items():
            print(f"\n📝 处理账户: {customer_key}")

            print(f"   包含 {len(account_data['pages'])} 页数据")
            print(f"   总数据行数: {account_data['total_rows']}")

            # 使用账户名称作为文件名
            safe_name = re.sub(r'[\\/*?:"<>|]', "_", customer_key)
            filename = f"{safe_name}.xlsx"
            filepath = os.path.join(self.output_dir, filename)

            # 处理文件名冲突
            counter = 1
            original_filename = filename
            while os.path.exists(filepath):
                name_without_ext = os.path.splitext(original_filename)[0]
                ext = os.path.splitext(original_filename)[1]
                filename = f"{name_without_ext}_{counter}{ext}"
                filepath = os.path.join(self.output_dir, filename)
                counter += 1

            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "对账单数据"
                print(f"✅ 创建Excel文件: {os.path.basename(filepath)}")
            except Exception as e:
                print(f"❌ 创建Excel文件时出错: {e}")
                continue

            # 合并所有页面的表格数据
            all_rows = []
            header_written = False
            original_header = None

            for i, page_data in enumerate(account_data["pages"]):
                page_num = page_data["page_num"]
                customer_info = page_data["customer_info"]
                matrix = page_data["matrix"]
                print(f"   合并第{page_num}页: {len(matrix)}行数据")

                if not matrix:
                    continue

                # 构建补充信息值列表（使用该页的customer_info）
                supplementary_values = []
                for col in self.supplementary_columns:
                    value = customer_info.get(col, '').strip()
                    supplementary_values.append(value if value else '')

                if i == 0 and len(matrix) > 0:
                    # 第一页：保留完整数据（包括表头）
                    original_header = matrix[0]
                    new_header = self.supplementary_columns + original_header
                    all_rows.append(new_header)
                    header_written = True

                    # 添加数据行（每行都补充该页的客户信息）
                    for j in range(1, len(matrix)):
                        new_row = supplementary_values + matrix[j]
                        all_rows.append(new_row)
                else:
                    # 后续页面：直接追加（假设已处理表头）
                    # 跳过可能是表头的行
                    data_start = 1 if header_written and len(matrix) > 0 and matrix[0] == original_header else 0

                    for j in range(data_start, len(matrix)):
                        new_row = supplementary_values + matrix[j]
                        all_rows.append(new_row)

            # 去重重复表头（如果存在）
            if len(all_rows) > 1:
                header = all_rows[0]
                result_rows = [header]
                for i in range(1, len(all_rows)):
                    if all_rows[i] != header:
                        result_rows.append(all_rows[i])
                all_rows = result_rows

            # 写入数据到Excel
            if all_rows and len(all_rows) > 0:
                for row in all_rows:
                    ws.append(row)

                # 应用格式
                self._apply_excel_format(ws, all_rows)

                # 保存Excel文件
                wb.save(filepath)

                # 记录文件信息
                data_rows_count = len(all_rows) - 1  # 减去表头行
                excel_files_info.append({
                    "filename": filename,
                    "filepath": filepath,
                    "customer_name": customer_key,
                    "total_pages": len(account_data["pages"]),
                    "total_rows": data_rows_count
                })
                self.extracted_files.append(filepath)

                print(f"✅ 保存Excel文件: {filename}")
                print(f"   最终写入数据: {data_rows_count}行")

        return excel_files_info

    def generate_single_excel_file(self):
        """生成单个总Excel文件（按原PDF顺序）"""
        if not self.all_data_pages:
            print("\n⚠️ 未找到有效的页面数据")
            return None

        print(f"\n📝 生成单个总Excel文件（按原PDF顺序）")
        print(f"   总页面数: {len(self.all_data_pages)}")

        # 生成文件名
        filename = f"{self.pdf_filename}_总数据.xlsx"
        filepath = os.path.join(self.output_dir, filename)

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "所有数据"
            print(f"✅ 创建总Excel文件: {os.path.basename(filepath)}")
        except Exception as e:
            print(f"❌ 创建Excel文件时出错: {e}")
            return None

        # 跟踪当前账户信息
        current_customer_info = None
        header_written = False
        total_data_rows = 0

        # 按原始页面顺序处理数据
        for page_data in self.all_data_pages:
            page_num = page_data["page_num"]
            customer_info = page_data["customer_info"]
            matrix = page_data["matrix"]

            print(f"   处理第{page_num}页: {len(matrix)}行数据")

            if not matrix:
                continue

            # 如果没有账户信息但之前有，使用之前的账户信息
            if not customer_info and current_customer_info:
                customer_info = current_customer_info
            elif customer_info:
                current_customer_info = customer_info

            if not customer_info:
                print(f"   第{page_num}页: 无账户信息，跳过")
                continue

            # 构建补充信息值列表（使用该页的customer_info）
            supplementary_values = []
            for col in self.supplementary_columns:
                value = customer_info.get(col, '').strip()
                supplementary_values.append(value if value else '')

            # 写入表头（只写一次）
            if not header_written and len(matrix) > 0:
                original_header = matrix[0]
                new_header = self.supplementary_columns + original_header
                ws.append(new_header)
                header_written = True
                print(f"   ✅ 已写入表头")

            # 写入数据行
            data_start = 1 if header_written else 0
            for i in range(data_start, len(matrix)):
                # 跳过可能是表头的行（如果已经写过表头）
                if header_written and i == 0 and matrix[i] == original_header:
                    continue

                new_row = supplementary_values + matrix[i]
                ws.append(new_row)
                total_data_rows += 1

        # 应用格式
        if header_written:
            # 获取所有行数据以应用格式
            all_rows = list(ws.iter_rows(values_only=True))
            if all_rows:
                self._apply_excel_format(ws, all_rows)

        # 保存Excel文件
        wb.save(filepath)
        self.extracted_files.append(filepath)

        print(f"\n✅ 总Excel文件保存完成")
        print(f"   总数据行数: {total_data_rows}")
        print(f"   文件位置: {filepath}")

        return {
            "filename": filename,
            "filepath": filepath,
            "total_pages": len(self.all_data_pages),
            "total_rows": total_data_rows
        }

    def _clean_excel_files(self):
        """清理Excel文件中的空格和不可见字符"""
        if not self.extracted_files:
            return

        cleaned_files = []
        for excel_file in self.extracted_files:
            try:
                print(f"\n🔧 正在清理文件: {os.path.basename(excel_file)}")
                wb = load_workbook(excel_file)
                ws = wb.active

                headers = []
                for col in range(1, ws.max_column + 1):
                    header = ws.cell(row=1, column=col).value
                    if isinstance(header, str):
                        headers.append(header.strip())
                    else:
                        headers.append(str(header) if header is not None else "")

                amount_columns = []
                for idx, header in enumerate(headers, start=1):
                    if isinstance(header, str) and "额" in header:
                        amount_columns.append(idx)

                for row in range(1, ws.max_row + 1):
                    for col in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row, column=col)
                        value = cell.value

                        if isinstance(value, str):
                            # 清理空格和不可见字符
                            cleaned_value = value.strip()
                            cleaned_value = re.sub(r'[\u200b\u200c\u200d\uFEFF\u00A0]+', '', cleaned_value)

                            # 金额列去掉千分位逗号
                            if col in amount_columns and row > 1:
                                cleaned_value = cleaned_value.replace(',', '')

                            cell.value = cleaned_value

                wb.save(excel_file)
                cleaned_files.append(excel_file)
                print(f"  ✅ 清理完成")

            except Exception as e:
                print(f"  ❌ 清理文件失败 {excel_file}: {e}")
                continue

        return cleaned_files

    def _generate_report(self, excel_files_info):
        """生成处理报告"""
        if not excel_files_info:
            print("\n⚠️ 没有生成任何Excel文件")
            return

        # 统计信息
        if self.output_mode == "multiple":
            customer_names = set()
            total_data_rows = 0

            for file_info in excel_files_info:
                customer_names.add(file_info['customer_name'])
                total_data_rows += file_info['total_rows']

            # 计算万条数据（保留三位小数）
            total_data_wan = round(total_data_rows / 10000, 3)

            # 生成报告内容
            report_content = f"共{len(customer_names)}个客户，{total_data_wan}万条数据"
        else:
            total_data_rows = excel_files_info.get('total_rows', 0)
            total_data_wan = round(total_data_rows / 10000, 3)
            report_content = f"单个总文件，{total_data_wan}万条数据"

        # 报告文件名
        report_filename = f"清洗报告（{self.pdf_filename}）.txt"
        report_path = os.path.join(self.output_dir, report_filename)

        # 写入报告
        with open(report_path, 'w', encoding='utf-8') as f:
            f.write(report_content)

        print(f"\n📊 清洗报告已生成: {report_path}")
        print(f"📋 报告内容: {report_content}")

    def process(self):
        """执行完整的处理流程"""
        try:
            print(f"🚀 开始处理PDF文件: {os.path.basename(self.pdf_path)}")
            print(f"📂 输出目录: {self.output_dir}")
            print(f"📝 输出模式: {'多个Excel文件（按账户名称）' if self.output_mode == 'multiple' else '单个总Excel文件'}")

            # 扫描所有页面提取数据
            self.scan_pages()

            if not self.all_data_pages:
                print("\n⚠️ 未找到有效的页面数据")
                return {"提取的文件数": 0, "跳过的页数": len(self.skipped_pages), "状态": "失败"}

            print(f"\n✅ 扫描完成，找到 {len(self.all_data_pages)} 个有效页面")

            # 根据输出模式选择生成方式
            if self.output_mode == "multiple":
                if not self.customer_data:
                    print("\n⚠️ 未找到有效的客户数据")
                    return {"提取的文件数": 0, "跳过的页数": len(self.skipped_pages), "状态": "失败"}

                print(f"📊 找到 {len(self.customer_data)} 个客户账户")
                excel_files_info = self.generate_multiple_excel_files()
            else:
                excel_files_info = self.generate_single_excel_file()

            if excel_files_info:
                # 清理Excel文件（去空格、不可见字符）
                self._clean_excel_files()
                # 生成处理报告
                self._generate_report(excel_files_info)

            # 关闭文档
            self.doc.close()

            # 打印最终统计信息
            print("\n" + "=" * 70)
            print("✅ 处理完成！")
            if self.output_mode == "multiple":
                print(f"📁 生成Excel文件数: {len(excel_files_info)}")
            else:
                print(f"📁 生成总Excel文件: 1个")
            print(f"⏭️  跳过的页面数: {len(self.skipped_pages)}")
            print(f"📂 输出目录: {self.output_dir}")
            print("=" * 70)

            return excel_files_info

        except Exception as e:
            print(f"\n❌ 处理PDF文件时出错: {e}")
            import traceback
            traceback.print_exc()
            return {"error": str(e), "状态": "失败"}


def main():
    """主函数"""
    print("=" * 60)
    print("光大银行PDF对账单提取工具 - 账户交易明细查询版")
    print("专用于：账户交易明细查询PDF")
    print("=" * 60)

    # 选择输出模式
    print("\n请选择输出模式:")
    print("1. 生成多个Excel文件（按账户名称）")
    print("2. 生成单个总Excel文件（按原PDF顺序）")

    mode_choice = input("\n请选择模式 (输入1或2): ").strip()

    if mode_choice == "1":
        output_mode = "multiple"
        mode_desc = "多个Excel文件（按账户名称）"
    elif mode_choice == "2":
        output_mode = "single"
        mode_desc = "单个总Excel文件"
    else:
        print("❌ 无效选择，默认使用模式1")
        output_mode = "multiple"
        mode_desc = "多个Excel文件（按账户名称）"

    print(f"\n📝 已选择: {mode_desc}")

    # 获取路径
    input_path = input("\n请输入PDF文件或文件夹位置：").strip()

    if not os.path.exists(input_path):
        print(f"❌ 路径不存在: {input_path}")
        return

    # 处理文件或文件夹
    if os.path.isfile(input_path):
        if not input_path.lower().endswith('.pdf'):
            print(f"❌ 输入的文件不是PDF格式: {input_path}")
            return

        print(f"\n📄 处理单个PDF文件: {os.path.basename(input_path)}")
        print(f"📝 输出模式: {mode_desc}")
        print("-" * 60)

        extractor = PDFTableExtractor_AccountQuery(input_path, output_mode=output_mode)
        result = extractor.process()

        if result and not isinstance(result, dict):
            print(f"\n{'=' * 60}")
            print(f"✅ 处理完成！")
            print(f"📁 输出目录: {extractor.output_dir}")
            print(f"{'=' * 60}")

    elif os.path.isdir(input_path):
        print(f"\n📁 批处理文件夹: {os.path.basename(input_path)}")
        print(f"📝 输出模式: {mode_desc}")
        print("-" * 60)

        pdf_files = []
        for root, dirs, files in os.walk(input_path):
            for file in files:
                if file.lower().endswith('.pdf'):
                    pdf_files.append(os.path.join(root, file))

        if not pdf_files:
            print(f"⚠️  在文件夹 {input_path} 中未找到PDF文件")
            return

        print(f"📁 找到 {len(pdf_files)} 个PDF文件需要处理")
        print("-" * 60)

        success_count = 0
        fail_count = 0

        for i, pdf_file in enumerate(pdf_files, 1):
            print(f"\n📊 处理进度: {i}/{len(pdf_files)}")
            print(f"📄 处理文件: {os.path.basename(pdf_file)}")

            try:
                extractor = PDFTableExtractor_AccountQuery(pdf_file, output_mode=output_mode)
                result = extractor.process()

                if result and not isinstance(result, dict):
                    success_count += 1
                    print(f"✅ 处理成功: {os.path.basename(pdf_file)}")
                else:
                    fail_count += 1
                    print(f"❌ 处理失败: {os.path.basename(pdf_file)}")
            except Exception as e:
                fail_count += 1
                print(f"❌ 处理异常: {os.path.basename(pdf_file)} - {str(e)}")

        print(f"\n{'=' * 60}")
        print("📋 批处理完成！")
        print(f"{'=' * 60}")
        print(f"✅ 处理成功: {success_count} 个文件")
        print(f"❌ 处理失败: {fail_count} 个文件")
        print(f"{'=' * 60}")

    else:
        print(f"❌ 输入路径既不是文件也不是文件夹: {input_path}")


# 添加单独的函数供外部调用
def process_pdf_to_multiple_excel(pdf_path):
    """处理PDF生成多个Excel文件（按账户名称）"""
    extractor = PDFTableExtractor_AccountQuery(pdf_path, output_mode="multiple")
    return extractor.process()


def process_pdf_to_single_excel(pdf_path):
    """处理PDF生成单个总Excel文件（按原PDF顺序）"""
    extractor = PDFTableExtractor_AccountQuery(pdf_path, output_mode="single")
    return extractor.process()