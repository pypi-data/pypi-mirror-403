"""个人版PDF表格提取器 - 专门处理光大银行个人版对账单"""
import fitz
import os
import re
import sys
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook


class PDFTableExtractor_Personal:
    """光大银行个人版PDF表格数据提取器"""

    # 个人版固定配置
    PERSONAL_CONFIG = {
        'customer_info_fields': {
            '客户姓名': [r'客户姓名[：:]([^\s\n]+)'],
            '客户账号': [r'客户账号[：:]\s*([\d]+)'],
            '对账日期': [r'对账日期[：:]\s*([^\s\n]+)'],
            '发卡/折机构': [r'发卡/折机构[：:]\s*(.+?)(?=\s+[^\s:]+[：:]|$)'],
            '打印时间': [r'打印时间[：:]\s*(.+)'],
            '系统账号': [r'系统账号[：:]\s*([\d]+)'],
            '币种': [r'币种[：:]\s*([^\s\n]+)'],
            '钞汇标志': [r'钞汇标志[：:]\s*([^\s\n]+)']
        },
        'supplementary_columns': [
            '客户姓名', '客户账号', '对账日期', '发卡/折机构',
            '打印时间', '系统账号', '币种', '钞汇标志'
        ],
        'report_title': '光大银行PDF对账单提取报告（个人版）',
        'default_params': {
            'target_size': 6.0,
            'size_tolerance': 0.2,
            'row_tolerance': 5.0,
            'col_tolerance': 30.0
        }
    }

    def __init__(self, pdf_path, export_type="split"):
        """
        初始化提取器
        :param pdf_path: PDF文件路径
        :param export_type: 导出类型，可选 "split"（分表） 或 "merged"（总表）
        """
        self.pdf_path = os.path.abspath(pdf_path)
        self.pdf_filename = os.path.splitext(os.path.basename(self.pdf_path))[0]
        self.export_type = export_type  # 保存导出类型
        self.output_dir = os.path.join(os.path.dirname(self.pdf_path), f"光大银行pdf转excel({self.pdf_filename})")
        self.doc = fitz.open(self.pdf_path)

        # 使用个人版固定配置
        config = self.PERSONAL_CONFIG
        self.target_size = config['default_params']['target_size']
        self.size_tolerance = config['default_params']['size_tolerance']
        self.row_tolerance = config['default_params']['row_tolerance']
        self.col_tolerance = config['default_params']['col_tolerance']
        self.customer_info_fields = config['customer_info_fields']
        self.supplementary_columns = config['supplementary_columns']
        self.report_title = config['report_title']

        # 存储按客户分组的数据
        self.customer_data = defaultdict(list)

        # 新增：存储原始页面数据（保持顺序）
        self.all_data_pages = []

        # 统计信息
        self.extracted_files = []
        self.skipped_pages = []

        self._create_output_dir()

    """创建输出目录"""
    def _create_output_dir(self):
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)
            print(f"✅ 创建输出目录: {self.output_dir}")

    """从页面提取客户信息"""
    def _extract_customer_info(self, page):
        full_text = page.get_text("text")

        # 使用字典存储所有提取的信息
        info = {}

        # 根据配置提取所有字段
        for field_name, patterns in self.customer_info_fields.items():
            value = ''
            for pattern in patterns:
                match = re.search(pattern, full_text)
                if match:
                    # 获取匹配的组，如果有分组则取第一个分组，否则取整个匹配
                    if match.groups():
                        value = match.group(1).strip()
                        if value:
                            break
                        else:
                            value = ' '
                            break
                    else:
                        value = match.group(0).strip()
                        break
                if not value:
                    value = ' '
            info[field_name] = value

        # 检查关键字段是否都存在（只需检查客户姓名）
        if '客户姓名' in info and info['客户姓名'].strip():
            return info
        else:
            return None

    """检查"不存在交易明细"字样"""
    def _has_no_transaction_details(self, page):
        full_text = page.get_text("text")
        return "不存在交易明细" in full_text \
               or "无符合条件的开户记录" in full_text \
               or "交易日期" not in full_text \
               or "无明细" in full_text

    """提取页面中的表格单元格数据"""
    def _extract_table_cells(self, page):
        text_dict = page.get_text("dict")
        cells = []
        for block in text_dict.get("blocks", []):
            if block["type"] == 0:  # 文本块
                for line in block.get("lines", []):
                    for span in line.get("spans", []):
                        # 匹配指定字体大小的文本
                        if abs(span["size"] - self.target_size) <= self.size_tolerance:
                            text = span["text"].strip()
                            if text:  # 非空文本
                                cells.append({
                                    "text": text,
                                    "bbox": span["bbox"],
                                    "x_center": (span["bbox"][0] + span["bbox"][2]) / 2,
                                    "y_center": (span["bbox"][1] + span["bbox"][3]) / 2,
                                    "y_start": span["bbox"][1],
                                })
        return cells

    """将单元格列表转换为矩阵"""
    def _cells_to_matrix(self, cells):
        if not cells:
            return []

        # 将单元格按行分组
        rows = self._group_cells_into_rows(cells)

        if not rows:
            return []

        # 确定列位置
        column_positions = self._detect_columns_from_header(rows)

        if not column_positions:
            return []

        # 为每行分配单元格
        matrix = self._assign_cells_with_header_constraint(rows, column_positions)

        return matrix

    def _group_cells_into_rows(self, cells):
        cells.sort(key=lambda c: c["y_center"])

        rows = []
        current_row = [cells[0]]
        current_y = cells[0]["y_center"]

        for cell in cells[1:]:
            if abs(cell["y_center"] - current_y) <= self.row_tolerance:
                current_row.append(cell)
            else:
                current_row.sort(key=lambda c: c["x_center"])
                rows.append(current_row)
                current_row = [cell]
                current_y = cell["y_center"]

        if current_row:
            current_row.sort(key=lambda c: c["x_center"])
            rows.append(current_row)

        return rows

    def _detect_columns_from_header(self, rows):
        if not rows:
            return []

        header_row = rows[0]
        header_x_positions = [cell["x_center"] for cell in header_row]
        header_x_positions.sort()

        column_positions = []
        for x in header_x_positions:
            if not column_positions:
                column_positions.append(x)
            else:
                min_distance = min(abs(x - pos) for pos in column_positions)
                if min_distance <= self.col_tolerance:
                    closest_idx = min(range(len(column_positions)),
                                      key=lambda i: abs(x - column_positions[i]))
                    column_positions[closest_idx] = (column_positions[closest_idx] + x) / 2
                else:
                    column_positions.append(x)

        column_positions.sort()
        return column_positions

    def _assign_cells_with_header_constraint(self, rows, column_positions):
        max_cols = len(column_positions)
        matrix = []

        for row_idx, row_cells in enumerate(rows):
            matrix_row = [""] * max_cols
            sorted_cells = sorted(row_cells, key=lambda c: c["x_center"])

            for cell in sorted_cells:
                closest_col_idx = self._find_closest_column(cell["x_center"], column_positions)

                if abs(cell["x_center"] - column_positions[closest_col_idx]) > self.col_tolerance * 2:
                    continue

                self._place_cell_in_column(cell, closest_col_idx, matrix_row, max_cols)

            if row_idx == 0:
                matrix_row = self._ensure_header_completeness(row_cells, matrix_row, column_positions)

            matrix.append(matrix_row)

        return matrix

    def _find_closest_column(self, x_center, column_positions):
        distances = [abs(x_center - pos) for pos in column_positions]
        return distances.index(min(distances))

    def _place_cell_in_column(self, cell, target_col_idx, matrix_row, max_cols):
        if not matrix_row[target_col_idx]:
            matrix_row[target_col_idx] = cell["text"]
            return

        for i in range(target_col_idx + 1, max_cols):
            if not matrix_row[i]:
                matrix_row[i] = cell["text"]
                return

        for i in range(target_col_idx - 1, -1, -1):
            if not matrix_row[i]:
                matrix_row[i] = cell["text"]
                return

        matrix_row[target_col_idx] += " " + cell["text"]

    def _ensure_header_completeness(self, header_cells, matrix_row, column_positions):
        assigned_texts = [text for text in matrix_row if text]
        all_header_texts = [cell["text"] for cell in header_cells]
        unassigned_texts = [text for text in all_header_texts if text not in assigned_texts]

        if not unassigned_texts:
            return matrix_row

        for header_text in unassigned_texts:
            header_cell = next((c for c in header_cells if c["text"] == header_text), None)
            if not header_cell:
                continue

            closest_col_idx = self._find_closest_column(header_cell["x_center"], column_positions)

            if not matrix_row[closest_col_idx]:
                matrix_row[closest_col_idx] = header_text
            else:
                empty_cols = [i for i, text in enumerate(matrix_row) if not text]
                if empty_cols:
                    distances = [abs(closest_col_idx - i) for i in empty_cols]
                    nearest_empty_idx = empty_cols[distances.index(min(distances))]
                    matrix_row[nearest_empty_idx] = header_text

        return matrix_row

    """合并单单元格行（在处理矩阵前进行）"""
    def _merge_single_cell_rows_in_matrix(self, matrix):
        """
        新的合并策略：如果某行第4列为空，则将该行所有数据接到上一行数据后
        """
        if not matrix or len(matrix) < 2:
            return matrix

        rows_to_delete = []

        # 从最后一行向前检查（排除表头行）
        for row_idx in range(len(matrix) - 1, 0, -1):
            row = matrix[row_idx]

            # 检查第4列是否为空（索引为3，因为从0开始）
            # 注意：如果矩阵列数少于4，则跳过
            if len(row) >= 4:
                # 检查第4列是否为空或只有空白
                col_4_value = row[3] if 3 < len(row) else ''
                if not col_4_value or str(col_4_value).strip() == '':
                    # 获取上一行
                    prev_row = matrix[row_idx - 1]

                    # 将当前行所有列的数据合并到上一行对应列
                    for col_idx in range(min(len(row), len(prev_row))):
                        current_value = row[col_idx]
                        prev_value = prev_row[col_idx]

                        # 如果当前单元格有值
                        if current_value and str(current_value).strip():
                            # 如果上一行同列为空，直接赋值
                            if not prev_value or str(prev_value).strip() == '':
                                prev_row[col_idx] = current_value
                            else:
                                # 如果上一行同列也有值，用分号连接
                                prev_row[col_idx] = f"{str(prev_value)} {str(current_value)}"

                    # 标记当前行要删除
                    rows_to_delete.append(row_idx)

        # 删除标记的行（从后往前删除）
        for row_idx in sorted(rows_to_delete, reverse=True):
            matrix.pop(row_idx)

        return matrix

    """扫描所有页面，提取表格数据"""
    def scan_pages(self):
        # 用于跟踪当前处理的客户信息
        current_customer_info = None

        for page_num in range(len(self.doc)):
            page = self.doc[page_num]

            # 检查页面是否包含"不存在交易明细"
            if self._has_no_transaction_details(page):
                print(f"  第{page_num + 1}页: 不存在交易明细，跳过")
                self.skipped_pages.append(page_num + 1)
                continue

            # 提取客户信息
            customer_info = self._extract_customer_info(page)

            # 如果没有提取到客户信息，使用上一页的信息
            if not customer_info and current_customer_info:
                customer_info = current_customer_info.copy()
                print(f"  第{page_num + 1}页: 使用上一页的客户信息")
            elif customer_info:
                current_customer_info = customer_info.copy()
                print(f"  第{page_num + 1}页: 提取到客户信息")

            if not customer_info:
                print(f"  第{page_num + 1}页: 未找到客户姓名，跳过")
                continue

            # 获取客户姓名
            customer_name = customer_info.get('客户姓名', '未知')

            print(f"  第{page_num + 1}页: 客户姓名: {customer_name}")

            # 提取表格数据
            cells = self._extract_table_cells(page)

            if not cells:
                print(f"    未找到表格数据")
                continue

            # 转换为矩阵
            matrix = self._cells_to_matrix(cells)

            if matrix:
                # 存储原始页面数据（保持原始顺序）
                page_data = {
                    "page_num": page_num + 1,
                    "customer_info": customer_info,
                    "matrix": matrix,
                    "has_customer_info": bool(customer_info)
                }
                self.all_data_pages.append(page_data)

                # 按客户姓名分组数据（用于分表导出）
                customer_key = customer_name
                self.customer_data[customer_key].append({
                    "page_num": page_num + 1,
                    "customer_info": customer_info,
                    "matrix": matrix,
                    "rows": len(matrix),
                    "cols": len(matrix[0]) if matrix else 0,
                    "is_header_page": True
                })

                print(f"    提取表格: {len(matrix)}行 × {len(matrix[0])}列")

        return self.customer_data

    """创建Excel文件"""
    def create_excel_file(self, customer_info, filepath):
        try:
            # 创建新的工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "对账单数据"

            print(f"✅ 创建新的Excel文件: {os.path.basename(filepath)}")
            return wb, ws, filepath

        except Exception as e:
            print(f"❌ 创建Excel文件时出错: {e}")
            return None, None, None

    """应用Excel格式设置（无边框风格）"""
    def _apply_excel_format(self, worksheet, data_rows):
        if not data_rows:
            return

        max_row = len(data_rows)
        max_col = max(len(row) for row in data_rows) if data_rows else 0

        if max_row == 0 or max_col == 0:
            return

        # 创建一个无边框的样式
        no_border = Border(
            left=Side(style='none'),
            right=Side(style='none'),
            top=Side(style='none'),
            bottom=Side(style='none')
        )

        # 设置列宽
        for col in range(1, max_col + 1):
            max_length = 0

            for row in range(1, max_row + 1):
                cell_value = worksheet.cell(row=row, column=col).value
                if cell_value:
                    # 计算内容长度（中文字符按2个宽度计算）
                    content = str(cell_value)
                    chinese_count = sum(1 for char in content if '\u4e00' <= char <= '\u9fff')
                    length = len(content) + chinese_count
                    max_length = max(max_length, length)

            # 设置合适的列宽
            if max_length > 0:
                column_letter = get_column_letter(col)
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

        # 设置统一的单元格格式（无边框，文本格式）
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = worksheet.cell(row=row, column=col)
                # 设置垂直和水平居中
                cell.alignment = Alignment(
                    vertical='center',
                    horizontal='center',
                    wrap_text=True
                )
                # 设置为文本格式，避免科学计数法
                cell.number_format = '@'
                # 移除边框
                cell.border = no_border

        # 设置表头行的字体
        if max_row > 0:
            for col in range(1, max_col + 1):
                header_cell = worksheet.cell(row=1, column=col)
                header_cell.font = Font(bold=True)

    """处理客户数据并生成Excel文件（带补充信息列）- 分表版本"""
    def process_customer_data_split(self):
        """按客户姓名导出分表"""
        if not self.customer_data:
            print("\n⚠️ 未找到有效的客户数据")
            return []

        excel_files_info = []

        for customer_key, page_data_list in self.customer_data.items():
            print(f"\n📝 处理客户: {customer_key}")
            print(f"   包含 {len(page_data_list)} 页数据")

            # 生成安全的文件名
            safe_name = re.sub(r'[\\/*?:"<>|]', "_", customer_key)
            filename = f"{safe_name}.xlsx"
            filepath = os.path.join(self.output_dir, filename)

            # 如果文件已存在，添加序号
            counter = 1
            original_filename = filename
            while os.path.exists(filepath):
                name_without_ext = os.path.splitext(original_filename)[0]
                ext = os.path.splitext(original_filename)[1]
                filename = f"{safe_name}_{counter}{ext}"
                filepath = os.path.join(self.output_dir, filename)
                counter += 1

            # 创建Excel文件
            wb, ws, filepath = self.create_excel_file({}, filepath)
            if not wb:
                print(f"   ❌ 创建Excel文件失败")
                continue

            # 初始化变量
            all_rows = []
            header_written = False
            original_header = None

            # 遍历页面数据，合并矩阵
            for i, page_data in enumerate(page_data_list):
                page_num = page_data["page_num"]
                customer_info = page_data["customer_info"]  # 使用该页的customer_info
                matrix = page_data["matrix"]

                print(f"\n   处理第{page_num}页数据:")

                if not matrix or len(matrix) == 0:
                    print(f"    无表格数据，跳过")
                    continue

                # 构建该页的补充信息值列表
                supplementary_values = []
                for col in self.supplementary_columns:
                    value = customer_info.get(col, '')
                    # 确保所有值都是字符串
                    if value is None:
                        value = ''
                    else:
                        value = str(value)
                    supplementary_values.append(value)

                if len(matrix) > 1:
                    # 对数据行（排除表头）应用合并逻辑
                    data_rows = matrix[1:]  # 排除表头行
                    merged_data_rows = self._merge_single_cell_rows_in_matrix(data_rows)

                    if i == 0:  # 第一页：添加表头+数据
                        original_header = matrix[0]
                        # 创建新的表头：补充信息列 + 原表格表头
                        new_header = self.supplementary_columns + original_header
                        all_rows.append(new_header)
                        header_written = True

                        # 添加数据行（每行都补充该页的客户信息）
                        for data_row in merged_data_rows:
                            new_row = supplementary_values + data_row
                            all_rows.append(new_row)
                        print(f"    添加: 表头 + {len(merged_data_rows)}行数据")
                    else:  # 后续页面：只添加数据行
                        # 跳过可能是表头的行
                        data_start = 1 if header_written and len(matrix) > 0 and matrix[0] == original_header else 0

                        for data_row in merged_data_rows:
                            new_row = supplementary_values + data_row
                            all_rows.append(new_row)
                        print(f"    添加: {len(merged_data_rows)}行数据")
                else:
                    # 如果只有一行（可能只是表头）
                    if i == 0 and not header_written:
                        original_header = matrix[0]
                        new_header = self.supplementary_columns + original_header
                        all_rows.append(new_header)
                        header_written = True
                        print(f"    添加: {len(matrix)}行（表头）")
                    else:
                        # 后续页面只有一行，可能是表头，跳过
                        print(f"    跳过: 只有1行（可能是表头）")

            if not all_rows:
                print(f"   ❌ 无有效数据，跳过")
                continue

            # 去重重复表头（如果存在）
            if len(all_rows) > 1:
                header = all_rows[0]
                result_rows = [header]
                for i in range(1, len(all_rows)):
                    if all_rows[i] != header:
                        result_rows.append(all_rows[i])
                all_rows = result_rows

            print(f"\n   总计合并行数: {len(all_rows)} (表头1行 + {len(all_rows) - 1 if all_rows else 0}行数据)")

            # 写入数据到Excel
            if all_rows:
                for row in all_rows:
                    ws.append(row)

                # 应用格式（无边框风格）
                self._apply_excel_format(ws, all_rows)

                # 保存Excel文件
                wb.save(filepath)

                excel_files_info.append({
                    "filename": filename,
                    "filepath": filepath,
                    "customer_name": customer_key,
                    "total_pages": len(page_data_list),
                    "total_rows": len(all_rows) - 1 if len(all_rows) > 0 else 0,  # 减去表头行
                })
                self.extracted_files.append(filepath)

                print(f"✅ 保存Excel文件: {filename}")
                print(f"   客户: {customer_key}")

        return excel_files_info

    """处理客户数据并生成Excel文件（带补充信息列）- 总表版本"""
    def process_customer_data_merged(self):
        """导出总表（所有客户数据在一个文件中）"""
        if not self.all_data_pages:
            print("\n⚠️ 未找到有效的页面数据")
            return []

        print(f"\n📝 开始合并所有客户数据到总表...")

        # 生成总表文件名
        filename = f"总表_{self.pdf_filename}.xlsx"
        filepath = os.path.join(self.output_dir, filename)

        # 创建Excel文件
        wb, ws, filepath = self.create_excel_file({}, filepath)
        if not wb:
            print(f"   ❌ 创建总表Excel文件失败")
            return []

        # 初始化变量
        all_rows = []
        header_written = False
        current_customer_info = None
        original_header = None

        # 按原始页面顺序处理数据
        for page_data in self.all_data_pages:
            page_num = page_data["page_num"]
            customer_info = page_data["customer_info"]
            matrix = page_data["matrix"]

            print(f"   处理第{page_num}页数据")

            # 如果没有客户信息但之前有，使用之前的客户信息
            if not customer_info and current_customer_info:
                customer_info = current_customer_info
            elif customer_info:
                current_customer_info = customer_info

            if not customer_info:
                print(f"   第{page_num}页: 无客户信息，跳过")
                continue

            if not matrix or len(matrix) == 0:
                print(f"   第{page_num}页: 无表格数据，跳过")
                continue

            # 构建补充信息值列表
            supplementary_values = []
            for col in self.supplementary_columns:
                value = customer_info.get(col, '')
                if value is None:
                    value = ''
                else:
                    value = str(value)
                supplementary_values.append(value)

            if len(matrix) > 1:
                # 对数据行（排除表头）应用合并逻辑
                data_rows = matrix[1:]  # 排除表头行
                merged_data_rows = self._merge_single_cell_rows_in_matrix(data_rows)

                if not header_written:
                    # 写入表头（只写一次）
                    original_header = matrix[0]
                    new_header = self.supplementary_columns + original_header
                    all_rows.append(new_header)
                    header_written = True
                    print(f"   ✅ 已写入表头")

                # 添加数据行（每行都补充该页的客户信息）
                for data_row in merged_data_rows:
                    new_row = supplementary_values + data_row
                    all_rows.append(new_row)
                print(f"   添加: {len(merged_data_rows)}行数据")
            else:
                # 如果只有一行（可能只是表头）
                if not header_written:
                    original_header = matrix[0]
                    new_header = self.supplementary_columns + original_header
                    all_rows.append(new_header)
                    header_written = True
                    print(f"   添加: 表头行")

        print(f"\n   总计行数: {len(all_rows)} (表头1行 + {len(all_rows) - 1 if all_rows else 0}行数据)")

        # 写入数据到Excel
        if all_rows:
            # 写入所有行
            for row in all_rows:
                ws.append(row)

            # 应用格式（无边框风格）
            self._apply_excel_format(ws, all_rows)

            # 保存Excel文件
            wb.save(filepath)

            excel_files_info = [{
                "filename": filename,
                "filepath": filepath,
                "total_customers": len(self.customer_data),
                "total_rows": len(all_rows) - 1 if len(all_rows) > 0 else 0,
                "export_type": "merged"
            }]
            self.extracted_files.append(filepath)

            print(f"✅ 保存总表Excel文件: {filename}")
            print(f"   包含 {len(self.customer_data)} 个客户的数据")

        return excel_files_info

    """生成清洗报告"""
    def _generate_clean_report(self, excel_files_info):
        """
        生成简单的清洗报告
        格式：共XX个客户，XX个账户，XX万条数据
        """
        if not excel_files_info:
            print("\n⚠️ 没有生成任何Excel文件")
            return

        # 根据导出类型生成不同的报告
        if self.export_type == "merged":
            # 总表报告
            total_data_rows = 0
            total_customers = 0

            for file_info in excel_files_info:
                total_data_rows += file_info.get('total_rows', 0)
                total_customers = file_info.get('total_customers', 0)

            # 计算万条数据（保留三位小数）
            total_data_wan = round(total_data_rows / 10000, 3)

            # 生成报告内容
            report_content = f"共{total_customers}个客户，{total_data_wan}万条数据，已合并到总表"
        else:
            # 分表报告
            customer_names = set()
            total_data_rows = 0

            for file_info in excel_files_info:
                if 'customer_name' in file_info:
                    customer_names.add(file_info['customer_name'])
                    total_data_rows += file_info.get('total_rows', 0)

            # 计算万条数据（保留三位小数）
            total_data_wan = round(total_data_rows / 10000, 3)

            # 生成报告内容
            report_content = f"共{len(customer_names)}个客户，{total_data_wan}万条数据"

        # 报告文件名（根据导出类型不同）
        export_type_str = "总表" if self.export_type == "merged" else "分表"
        report_filename = f"清洗报告({export_type_str})_{self.pdf_filename}.txt"
        report_path = os.path.join(self.output_dir, report_filename)

        # 写入报告
        with open(report_path, 'w', encoding='utf-8') as f:
            f.write(report_content)

        print(f"\n📊 清洗报告已生成: {report_path}")
        print(f"📋 报告内容: {report_content}")

    """清理Excel文件"""
    def _clean_excel_files(self, excel_files):
        """
        清理Excel文件：删除所有单元格中的空格和不可见字符，并对"额"字段去除逗号
        """
        cleaned_files = []

        for excel_file in excel_files:
            try:
                print(f"🔧 正在清理文件: {os.path.basename(excel_file)}")

                # 读取Excel文件
                wb = load_workbook(excel_file)
                ws = wb.active

                # 获取表头
                headers = []
                for col in range(1, ws.max_column + 1):
                    header = ws.cell(row=1, column=col).value
                    if isinstance(header, str):
                        headers.append(header.strip())
                    else:
                        headers.append(str(header) if header is not None else "")

                # 标记包含"额"字的列
                amount_columns = []
                for idx, header in enumerate(headers, start=1):
                    if isinstance(header, str) and "额" in header:
                        amount_columns.append(idx)

                # 处理每个单元格
                for row in range(1, ws.max_row + 1):
                    for col in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row, column=col)
                        value = cell.value

                        if isinstance(value, str):
                            # 1. 删除数据前后的空格（使用strip()）
                            cleaned_value = value.strip()

                            # 2. 删除所有不可见字符（零宽空格等）
                            cleaned_value = re.sub(r'[\u200b\u200c\u200d\uFEFF\u00A0]+', '', cleaned_value)

                            # 3. 如果是"额"字段，去除逗号
                            if col in amount_columns and row > 1:  # 表头行不处理
                                cleaned_value = cleaned_value.replace(',', '')

                            cell.value = cleaned_value

                # 保存清理后的文件
                wb.save(excel_file)
                cleaned_files.append(excel_file)
                print(f"  ✅ 清理完成")

            except Exception as e:
                print(f"  ❌ 清理文件失败 {excel_file}: {e}")
                continue

        return cleaned_files

    """执行完整的处理流程"""
    def process(self):
        try:
            print(f"🚀 开始处理PDF文件: {os.path.basename(self.pdf_path)}")
            print(f"📋 导出类型: {'总表' if self.export_type == 'merged' else '分表'}")

            # 扫描所有页面提取数据
            customer_data = self.scan_pages()

            if not customer_data:
                print("\n⚠️ 未找到有效的客户数据")
                return {"提取的文件数": 0, "跳过的页数": len(self.skipped_pages), "状态": "失败"}

            print(f"\n✅ 找到 {len(customer_data)} 个客户")

            excel_files_info = []

            # 根据导出类型选择处理方式
            if self.export_type == "merged":
                # 导出总表
                excel_files_info = self.process_customer_data_merged()
            else:
                # 导出分表
                excel_files_info = self.process_customer_data_split()

            if excel_files_info:
                # 获取所有生成的Excel文件
                excel_files = [info['filepath'] for info in excel_files_info]

                # 清理Excel文件
                print("\n🧹 开始清理Excel文件...")
                cleaned_files = self._clean_excel_files(excel_files)

                if cleaned_files:
                    print(f"\n✅ 清理完成，已清理 {len(cleaned_files)} 个文件")

                    # 生成清洗报告
                    self._generate_clean_report(excel_files_info)

                    print(f"\n📋 Excel文件已清理并保存，位置: {self.output_dir}")

            # 关闭文档
            self.doc.close()

            # 打印最终统计信息
            print("\n" + "=" * 70)
            print("✅ 处理完成！")
            if excel_files_info:
                if self.export_type == "merged":
                    print(f"📁 生成总表文件数: {len(excel_files_info)}")
                else:
                    print(f"📁 生成分表文件数: {len(excel_files_info)}")
            print(f"⏭️  跳过的页面数: {len(self.skipped_pages)}")
            print(f"📂 输出目录: {self.output_dir}")
            print("=" * 70)

            return excel_files_info  # 返回文件信息

        except Exception as e:
            print(f"\n❌ 处理PDF文件时出错: {e}")
            import traceback
            traceback.print_exc()
            return {"error": str(e), "状态": "失败"}


def main():
    """主函数 - 直接运行个人版提取器"""
    print("=" * 70)
    print("光大银行个人版PDF对账单提取工具")
    print("=" * 70)

    # 选择导出类型
    print("\n📋 请选择导出类型:")
    print("1. 按客户姓名导出分表（每个客户一个Excel文件）")
    print("2. 导出总表（所有客户数据合并到一个Excel文件）")

    export_choice = input("\n请选择导出类型 (输入1或2): ").strip()

    if export_choice == "1":
        export_type = "split"
        export_type_name = "分表"
    elif export_choice == "2":
        export_type = "merged"
        export_type_name = "总表"
    else:
        print("❌ 输入无效，默认使用分表导出")
        export_type = "split"
        export_type_name = "分表"

    # 获取PDF文件路径
    if len(sys.argv) > 1:
        pdf_path = sys.argv[1]
    else:
        pdf_path = input("\n请输入PDF文件位置：").strip()

    # 验证路径是否存在
    if not os.path.exists(pdf_path):
        print(f"❌ 路径不存在: {pdf_path}")
        print("请检查路径是否正确")
        return

    # 检查是否是PDF文件
    if not pdf_path.lower().endswith('.pdf'):
        print(f"❌ 输入的文件不是PDF格式: {pdf_path}")
        return

    print(f"\n📄 开始处理PDF文件: {os.path.basename(pdf_path)}")
    print(f"📁 导出类型: {export_type_name}")
    print("-" * 70)

    # 创建个人版提取器并处理
    extractor = PDFTableExtractor_Personal(pdf_path, export_type)
    result = extractor.process()

    if isinstance(result, list) and len(result) > 0:
        print(f"\n{'=' * 70}")
        print(f"✅ 处理完成！")
        print(f"📁 输出目录: {extractor.output_dir}")
        print(f"{'=' * 70}")
    else:
        print(f"\n{'=' * 70}")
        print(f"❌ 处理失败！")
        print(f"{'=' * 70}")