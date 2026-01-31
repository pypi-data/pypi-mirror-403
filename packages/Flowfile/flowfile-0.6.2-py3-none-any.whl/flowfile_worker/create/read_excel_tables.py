import gc
from collections.abc import Generator

import polars as pl
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from flowfile_worker.create.utils import create_pl_df_type_save


def raw_data_openpyxl(
    file_path: str,
    sheet_name: str = None,
    min_row: int = None,
    max_row: int = None,
    min_col: int = None,
    max_col: int = None,
) -> Generator[list, None, None]:
    workbook: Workbook = load_workbook(file_path, data_only=True, read_only=True)
    sheet_name = workbook.sheetnames[0] if sheet_name is None else sheet_name
    sheet: Worksheet = workbook[sheet_name]
    yield from sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True)
    workbook.close()
    del workbook
    gc.collect()


def df_from_openpyxl(
    file_path: str,
    sheet_name: str = None,
    min_row: int = None,
    max_row: int = None,
    min_col: int = None,
    max_col: int = None,
    has_headers: bool = True,
) -> pl.DataFrame:
    data_iterator = raw_data_openpyxl(
        file_path=file_path, sheet_name=sheet_name, min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
    )
    raw_data = list(data_iterator)
    if len(raw_data) > 0:
        if has_headers:
            columns = []
            for i, col in enumerate(raw_data[0]):
                if col is None:
                    col = f"_unnamed_column_{i}"
                elif not isinstance(col, str):
                    col = str(col)
                columns.append(col)
            columns = ensure_unique(columns)
            df = create_pl_df_type_save(raw_data[1:])
            renames = {o: n for o, n in zip(df.columns, columns, strict=False)}
            df = df.rename(renames)

        else:
            df = create_pl_df_type_save(raw_data)
        return df
    else:
        return pl.DataFrame()


def ensure_unique(lst: list[str]) -> list[str]:
    """
    Ensures that all elements in the input list are unique by appending
    a version number (e.g., '_v1') to duplicates. It continues adding
    version numbers until all items in the list are unique.

    Args:
        lst (List[str]): A list of strings that may contain duplicates.

    Returns:
        List[str]: A new list where all elements are unique.
    """
    seen = {}
    result = []

    for item in lst:
        if item in seen:
            # Increment the version and append the version number
            seen[item] += 1
            new_item = f"{item}_v{seen[item]}"
            # Ensure the new item is unique by checking for conflicts
            while new_item in seen:
                seen[new_item] += 1
                new_item = f"{item}_v{seen[item]}"
            result.append(new_item)
            seen[new_item] = 1  # Mark the new unique item as seen
        else:
            result.append(item)
            seen[item] = 1  # First occurrence of the item

    return result


def df_from_calamine_xlsx(file_path: str, sheet_name: str, start_row: int = 0, end_row: int = 0) -> pl.DataFrame:
    read_options = {}
    if start_row > 0:
        read_options["header_row"] = start_row
    if end_row > 0:
        read_options["n_rows"] = end_row - start_row
    return pl.read_excel(
        source=file_path, engine="calamine", sheet_name=sheet_name, read_options=read_options, raise_if_empty=False
    )
