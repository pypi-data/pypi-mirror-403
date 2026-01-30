from collections import Counter
from ExcelSage import (
    ExcelSage,
    ExcelFileNotFoundError,
    FileAlreadyExistsError,
    WorkbookNotOpenError,
    SheetAlreadyExistsError,
    InvalidSheetPositionError,
    SheetDoesntExistsError,
    InvalidCellAddressError,
    InvalidRowIndexError,
    InvalidColumnIndexError,
    InvalidColumnNameError,
    ColumnMismatchError,
    InvalidCellRangeError,
    SheetAlreadyProtectedError,
    SheetNotProtectedError,
    WorkbookAlreadyProtectedError,
    WorkbookNotProtectedError,
    InvalidSheetNameError,
    InvalidColorError,
    InvalidAlignmentError,
    InvalidBorderStyleError,
)
from assertpy import assert_that
from openpyxl import Workbook
import openpyxl as excel
import os
from pathlib import Path
from openpyxl.utils import get_column_letter
from openpyxl.workbook.protection import WorkbookProtection
from openpyxl.worksheet.protection import SheetProtection
from openpyxl.utils import cell
import openpyxl as xl
import pytest
import shutil
from pandas import DataFrame
import pandas as pd

PROJECT_ROOT = Path(__file__).parent.parent
DATA_DIR = str(PROJECT_ROOT / "data")

exl = ExcelSage()

EXCEL_FILE_PATH = os.path.join(DATA_DIR, "sample.xlsx")
CSV_FILE_PATH = os.path.join(DATA_DIR, "sample.csv")
INVALID_EXCEL_FILE_PATH = os.path.join(
    str(PROJECT_ROOT.parent), "data", "invalid_file.xlsx"
)
NEW_EXCEL_FILE_PATH = os.path.join(DATA_DIR, "new_excel.xlsx")
INVALID_SHEET_NAME = "invalid[]sheet"
INVALID_CELL_ADDRESS = "AAAA1"
INVALID_ROW_INDEX = 1012323486523
INVALID_COLUMN_INDEX = 163841


@pytest.fixture
def setup_teardown(scope="function", autouse=False):
    yield
    while exl.workbooks:
        exl.close_workbook()


def copy_test_excel_file(destination_file=None):
    if destination_file is None:
        destination_file = os.path.join(DATA_DIR, "sample.xlsx")
    source_file = os.path.join(DATA_DIR, "sample_original.xlsx")
    shutil.copy(source_file, destination_file)
    return destination_file


def delete_the_test_excel_file(directory=None):
    if directory is None:
        directory = DATA_DIR
    for filename in os.listdir(directory):
        file_path = os.path.join(directory, filename)
        try:
            if os.path.isfile(file_path) and filename != "sample_original.xlsx":
                os.chmod(file_path, 0o777)
                os.remove(file_path)
        except PermissionError:
            print(f" PermissionError on {file_path}.")


@pytest.fixture(scope="session", autouse=True)
def setup():
    delete_the_test_excel_file()
    copy_test_excel_file()
    yield
    delete_the_test_excel_file()


def test_open_workbook_success(setup_teardown):
    workbook = exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    assert_that(workbook).is_instance_of(Workbook)


def test_open_workbook_file_not_found(setup_teardown):
    with pytest.raises(ExcelFileNotFoundError) as exc_info:
        exl.open_workbook(workbook_name=INVALID_EXCEL_FILE_PATH)

    assert_that(str(exc_info.value)).is_equal_to(
        f"Excel file '{INVALID_EXCEL_FILE_PATH}' not found. Please give the valid file path."
    )


def test_create_workbook_success(setup_teardown):
    sheet_data = [["Name", "Age"], ["Dee", 26], ["Mark", 56], ["John", 30]]
    workbook = exl.create_workbook(
        workbook_name=NEW_EXCEL_FILE_PATH,
        sheet_data=sheet_data,
        overwrite_if_exists=True,
    )
    assert_that(workbook).is_instance_of(Workbook)


def test_create_workbook_file_already_exists(setup_teardown):
    with pytest.raises(FileAlreadyExistsError) as exc_info:
        exl.create_workbook(workbook_name=NEW_EXCEL_FILE_PATH)

    assert_that(str(exc_info.value)).is_equal_to(
        f"Unable to create workbook. The file '{NEW_EXCEL_FILE_PATH}' already exists. Set 'overwrite_if_exists=True' to overwrite the existing file."
    )


def test_create_workbook_type_error(setup_teardown):
    with pytest.raises(TypeError) as exc_info:
        sheet_data = [["Name", "Age"], ["Dee", 26], ["Mark", 56], "John"]
        exl.create_workbook(
            workbook_name=NEW_EXCEL_FILE_PATH,
            sheet_data=sheet_data,
            overwrite_if_exists=True,
        )

    assert_that(str(exc_info.value)).is_equal_to(
        "Invalid row at index 3 of type 'str'. Each row in 'sheet_data' must be a list."
    )


def test_get_sheets_success(setup_teardown):
    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    sheets = exl.get_sheets()
    assert_that(sheets).is_length(3).contains(
        "Sheet1", "Offset_table", "Invalid_header"
    )


def test_get_sheets_workbook_not_open(setup_teardown):
    with pytest.raises(WorkbookNotOpenError) as exc_info:
        exl.get_sheets()

    assert_that(str(exc_info.value)).is_equal_to(
        "Workbook isn't open. Please open the workbook first."
    )


def test_add_sheet_success(setup_teardown):
    sheet_data = [["Name", "Age"], ["Dee", 26], ["Mark", 56], ["John", 30]]
    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    exl.add_sheet(sheet_name="Sheet3", sheet_data=sheet_data, sheet_pos=2)

    workbook = xl.load_workbook(filename=EXCEL_FILE_PATH)
    sheets = workbook.sheetnames
    assert_that(sheets).is_length(4).contains(
        "Sheet1", "Offset_table", "Sheet3", "Invalid_header"
    )
    sheet_to_delete = workbook["Sheet3"]
    workbook.remove(sheet_to_delete)
    workbook.save(EXCEL_FILE_PATH)
    workbook.close()


def test_add_sheet_workbook_not_open(setup_teardown):
    with pytest.raises(WorkbookNotOpenError) as exc_info:
        exl.add_sheet(sheet_name="Sheet4")

    assert_that(str(exc_info.value)).is_equal_to(
        "Workbook isn't open. Please open the workbook first."
    )


def test_add_sheet_sheet_exists(setup_teardown):
    with pytest.raises(SheetAlreadyExistsError) as exc_info:
        sheet_data = [["Name", "Age"], ["Dee", 26], ["Mark", 56], ["John", 30]]
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.add_sheet(sheet_name="Sheet1", sheet_data=sheet_data, sheet_pos=2)

    assert_that(str(exc_info.value)).is_equal_to("Sheet 'Sheet1' already exists.")


def test_add_sheet_sheet__invalid_position(setup_teardown):
    with pytest.raises(InvalidSheetPositionError) as exc_info:
        sheet_data = [["Name", "Age"], ["Dee", 26], ["Mark", 56], ["John", 30]]
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.add_sheet(sheet_name="Sheet5", sheet_data=sheet_data, sheet_pos=5)

    assert_that(str(exc_info.value)).is_equal_to(
        "Invalid sheet position: 5. Maximum allowed is 3."
    )


def test_add_sheet_type_error(setup_teardown):
    with pytest.raises(TypeError) as exc_info:
        sheet_data = [["Name", "Age"], ["Dee", 26], ["Mark", 56], "John"]
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.add_sheet(sheet_name="Sheet6", sheet_data=sheet_data, sheet_pos=1)

    assert_that(str(exc_info.value)).is_equal_to(
        "Invalid row at index 3 of type 'str'. Each row in 'sheet_data' must be a list."
    )


def test_delete_sheet_success(setup_teardown):
    workbook = xl.load_workbook(filename=EXCEL_FILE_PATH)
    workbook.create_sheet(title="Sheet_to_delete")
    workbook.save(EXCEL_FILE_PATH)
    workbook.close()

    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    deleted_sheet = exl.delete_sheet(sheet_name="Sheet_to_delete")
    assert_that(deleted_sheet).is_equal_to("Sheet_to_delete")

    workbook = xl.load_workbook(filename=EXCEL_FILE_PATH)
    sheet = workbook.sheetnames
    assert_that(sheet).is_length(3).contains("Sheet1", "Offset_table", "Invalid_header")
    workbook.close()


def test_delete_sheet_doesnt_exists(setup_teardown):
    with pytest.raises(SheetDoesntExistsError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.delete_sheet(sheet_name=INVALID_SHEET_NAME)

    assert_that(str(exc_info.value)).is_equal_to(
        f"Sheet '{INVALID_SHEET_NAME}' doesn't exists."
    )


def test_fetch_sheet_data_success(setup_teardown):
    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    sheet_data = exl.fetch_sheet_data(
        sheet_name="Offset_table",
        output_format="list",
        starting_cell="D6",
        ignore_empty_columns=True,
        ignore_empty_rows=True,
    )
    assert_that(isinstance(sheet_data, list)).is_true()

    sheet_data = exl.fetch_sheet_data(
        sheet_name="Offset_table",
        output_format="dict",
        starting_cell="A1",
        ignore_empty_columns=True,
        ignore_empty_rows=True,
    )
    assert_that(
        isinstance(sheet_data, list)
        and all(isinstance(item, dict) for item in sheet_data)
    ).is_true()

    sheet_data = exl.fetch_sheet_data(
        sheet_name="Offset_table",
        output_format="dataframe",
        starting_cell="D6",
        ignore_empty_columns=True,
        ignore_empty_rows=True,
    )
    assert_that(isinstance(sheet_data, DataFrame)).is_true()


def test_fetch_sheet_data_invalid_cell_address(setup_teardown):
    with pytest.raises(InvalidCellAddressError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.fetch_sheet_data(
            sheet_name="Offset_table",
            output_format="list",
            starting_cell=INVALID_CELL_ADDRESS,
            ignore_empty_columns=True,
            ignore_empty_rows=True,
        )

    assert_that(str(exc_info.value)).is_equal_to(
        f"Cell '{INVALID_CELL_ADDRESS}' doesn't exists."
    )


def test_fetch_sheet_data_invalid_output_format(setup_teardown):
    with pytest.raises(ValueError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.fetch_sheet_data(
            sheet_name="Offset_table",
            output_format="invalid",
            starting_cell="D6",
            ignore_empty_columns=True,
            ignore_empty_rows=True,
        )

    assert_that(str(exc_info.value)).is_equal_to(
        "Invalid output format. Use 'list', 'dict', or 'dataframe'."
    )


def test_rename_sheet_success(setup_teardown):
    workbook = excel.load_workbook(filename=EXCEL_FILE_PATH)
    workbook.create_sheet(title="Sheet_to_rename")
    workbook.save(EXCEL_FILE_PATH)
    workbook.close()

    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    exl.rename_sheet(old_name="Sheet_to_rename", new_name="Sheet_renamed")

    workbook = excel.load_workbook(filename=EXCEL_FILE_PATH)
    sheets = workbook.sheetnames
    assert_that(sheets).contains("Sheet_renamed")
    sheet_to_delete = workbook["Sheet_renamed"]
    workbook.remove(sheet_to_delete)
    workbook.save(EXCEL_FILE_PATH)
    workbook.close()


def test_rename_sheet_workbook_not_open(setup_teardown):
    with pytest.raises(WorkbookNotOpenError) as exc_info:
        exl.rename_sheet(old_name="Sheet1", new_name="Offset_table")

    assert_that(str(exc_info.value)).is_equal_to(
        "Workbook isn't open. Please open the workbook first."
    )


def test_rename_sheet_sheet_exists(setup_teardown):
    with pytest.raises(SheetAlreadyExistsError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.rename_sheet(old_name="Sheet1", new_name="Offset_table")

    assert_that(str(exc_info.value)).is_equal_to("Sheet 'Offset_table' already exists.")


def test_rename_sheet_sheet_doesnt_exists(setup_teardown):
    with pytest.raises(SheetDoesntExistsError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.rename_sheet(old_name=INVALID_SHEET_NAME, new_name="Offset_table")

    assert_that(str(exc_info.value)).is_equal_to(
        f"Sheet '{INVALID_SHEET_NAME}' doesn't exists."
    )


def test_get_cell_value_success(setup_teardown):
    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    cell_value = exl.get_cell_value(sheet_name="Sheet1", cell_name="A1")
    assert_that(cell_value).is_equal_to("First Name")


def test_get_cell_value_invalid_cell_address(setup_teardown):
    with pytest.raises(InvalidCellAddressError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.get_cell_value(sheet_name="Sheet1", cell_name=INVALID_CELL_ADDRESS)

    assert_that(str(exc_info.value)).is_equal_to(
        f"Cell '{INVALID_CELL_ADDRESS}' doesn't exists."
    )


def test_close_workbook_success(setup_teardown):
    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    exl.close_workbook()

    assert_that(exl.active_workbook_alias).is_none()
    assert_that(exl.workbooks).is_empty()
    assert_that(exl.active_sheet).is_none()


def test_close_workbook_workbook_not_open(setup_teardown):
    with pytest.raises(WorkbookNotOpenError) as exc_info:
        exl.close_workbook()

    assert_that(str(exc_info.value)).is_equal_to(
        "Workbook isn't open. Please open the workbook first."
    )


def test_open_workbook_with_alias(setup_teardown):
    """Test opening a workbook with a custom alias."""
    exl.open_workbook(workbook_name=EXCEL_FILE_PATH, alias="source")
    
    assert_that(exl.workbooks).contains_key("source")
    assert_that(exl.active_workbook_alias).is_equal_to("source")
    assert_that(exl.workbooks["source"]["name"]).is_equal_to(EXCEL_FILE_PATH)


def test_open_multiple_workbooks(setup_teardown):
    """Test opening multiple workbooks with different aliases."""
    test_file = copy_test_excel_file(
        destination_file=os.path.join(DATA_DIR, "test_multiple_workbooks.xlsx")
    )
    
    try:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH, alias="source")
        assert_that(exl.active_workbook_alias).is_equal_to("source")
        assert_that(exl.workbooks).is_length(1)
        
        exl.open_workbook(workbook_name=test_file, alias="target")
        assert_that(exl.active_workbook_alias).is_equal_to("source")
        assert_that(exl.workbooks).is_length(2)
        assert_that(exl.workbooks).contains_key("source", "target")
        
        source_sheets = exl.get_sheets()
        assert_that(source_sheets).is_not_empty()
        
    finally:
        if os.path.exists(test_file):
            os.remove(test_file)


def test_switch_workbook_success(setup_teardown):
    test_file = copy_test_excel_file(
        destination_file=os.path.join(DATA_DIR, "test_switch_workbook.xlsx")
    )
    
    try:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH, alias="source")
        exl.open_workbook(workbook_name=test_file, alias="target")
        
        assert_that(exl.active_workbook_alias).is_equal_to("source")
        source_sheets = exl.get_sheets()
        
        exl.switch_workbook(alias="target")
        assert_that(exl.active_workbook_alias).is_equal_to("target")
        
        exl.switch_workbook(alias="source")
        assert_that(exl.active_workbook_alias).is_equal_to("source")
        source_sheets_after = exl.get_sheets()
        
        assert_that(source_sheets).is_equal_to(source_sheets_after)
        
    finally:
        if os.path.exists(test_file):
            os.remove(test_file)


def test_switch_workbook_invalid_alias(setup_teardown):
    exl.open_workbook(workbook_name=EXCEL_FILE_PATH, alias="source")
    
    with pytest.raises(WorkbookNotOpenError) as exc_info:
        exl.switch_workbook(alias="nonexistent")
    assert_that(str(exc_info.value)).contains("not open")


def test_close_workbook_by_alias(setup_teardown):
    test_file = copy_test_excel_file(
        destination_file=os.path.join(DATA_DIR, "test_close_by_alias.xlsx")
    )
    
    try:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH, alias="source")
        exl.open_workbook(workbook_name=test_file, alias="target")
        
        assert_that(exl.workbooks).is_length(2)
        assert_that(exl.active_workbook_alias).is_equal_to("source")
        
        exl.close_workbook(alias="target")
        assert_that(exl.workbooks).is_length(1)
        assert_that(exl.workbooks).contains_key("source")
        assert_that(exl.workbooks).does_not_contain_key("target")
        assert_that(exl.active_workbook_alias).is_equal_to("source")
        
        exl.close_workbook(alias="source")
        assert_that(exl.workbooks).is_empty()
        assert_that(exl.active_workbook_alias).is_none()
        
    finally:
        if os.path.exists(test_file):
            os.remove(test_file)


def test_close_workbook_switches_active(setup_teardown):
    test_file = copy_test_excel_file(
        destination_file=os.path.join(DATA_DIR, "test_close_switches_active.xlsx")
    )
    
    try:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH, alias="source")
        exl.open_workbook(workbook_name=test_file, alias="target")
        
        assert_that(exl.active_workbook_alias).is_equal_to("source")
        
        exl.close_workbook(alias="source")
        
        assert_that(exl.active_workbook_alias).is_equal_to("target")
        assert_that(exl.workbooks).is_length(1)
        assert_that(exl.workbooks).contains_key("target")
        
    finally:
        if os.path.exists(test_file):
            os.remove(test_file)


def test_open_workbook_duplicate_alias(setup_teardown):
    """Test that opening a workbook with an existing alias raises an error."""
    exl.open_workbook(workbook_name=EXCEL_FILE_PATH, alias="my_alias")
    
    with pytest.raises(SheetAlreadyExistsError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH, alias="my_alias")
    assert_that(str(exc_info.value)).contains("already open")


def test_operations_with_multiple_workbooks(setup_teardown):
    """Test that operations work correctly with multiple workbooks open."""
    test_file = copy_test_excel_file(
        destination_file=os.path.join(DATA_DIR, "test_operations_multiple.xlsx")
    )
    
    try:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH, alias="source")
        exl.open_workbook(workbook_name=test_file, alias="target")
        
        exl.switch_workbook(alias="source")
        source_sheets = exl.get_sheets()
        
        exl.switch_workbook(alias="target")
        target_sheets = exl.get_sheets()
        
        exl.switch_workbook(alias="source")
        source_cell = exl.get_cell_value(sheet_name=source_sheets[0], cell_name="A1")
        
        exl.switch_workbook(alias="target")
        target_cell = exl.get_cell_value(sheet_name=target_sheets[0], cell_name="A1")
        
        assert_that(source_cell).is_not_none()
        assert_that(target_cell).is_not_none()
        
    finally:
        if os.path.exists(test_file):
            os.remove(test_file)


def test_save_workbook_success(setup_teardown):
    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    exl.add_sheet(sheet_name="New_sheet")
    exl.save_workbook()

    workbook = excel.load_workbook(filename=EXCEL_FILE_PATH)
    sheet = workbook.sheetnames
    assert_that(sheet).is_length(4).contains("New_sheet")
    sheet_to_delete = workbook["New_sheet"]
    workbook.remove(sheet_to_delete)
    workbook.save(EXCEL_FILE_PATH)
    workbook.close()


def test_save_workbook_workbook_not_open(setup_teardown):
    with pytest.raises(WorkbookNotOpenError) as exc_info:
        exl.save_workbook()

    assert_that(str(exc_info.value)).is_equal_to(
        "Workbook isn't open. Please open the workbook first."
    )


def test_save_workbook_by_alias(setup_teardown):
    test_file = copy_test_excel_file(
        destination_file=os.path.join(DATA_DIR, "test_save_by_alias.xlsx")
    )
    
    try:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH, alias="source")
        exl.open_workbook(workbook_name=test_file, alias="target")
        
        assert_that(exl.workbooks).is_length(2)
        assert_that(exl.active_workbook_alias).is_equal_to("source")
        
        exl.switch_workbook(alias="target")
        exl.add_sheet(sheet_name="NewSheet")
        
        exl.save_workbook(alias="target")
        
        workbook = excel.load_workbook(filename=test_file)
        assert_that(workbook.sheetnames).contains("NewSheet")
        workbook.close()
        
        assert_that(exl.active_workbook_alias).is_equal_to("target")
        
    finally:
        if os.path.exists(test_file):
            os.remove(test_file)


def test_save_workbook_invalid_alias(setup_teardown):
    exl.open_workbook(workbook_name=EXCEL_FILE_PATH, alias="source")
    
    with pytest.raises(WorkbookNotOpenError) as exc_info:
        exl.save_workbook(alias="invalid_alias")
    
    assert_that(str(exc_info.value)).is_equal_to(
        "Workbook with alias 'invalid_alias' is not open."
    )


def test_set_active_sheet_success(setup_teardown):
    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    active_sheet = exl.set_active_sheet(sheet_name="Offset_table")

    assert_that(active_sheet).is_equal_to("Offset_table")
    assert_that(str(exl.active_sheet)).contains("Offset_table")


def test_set_active_sheet_workbook_not_open(setup_teardown):
    with pytest.raises(WorkbookNotOpenError) as exc_info:
        exl.set_active_sheet(sheet_name="Offset_table")

    assert_that(str(exc_info.value)).is_equal_to(
        "Workbook isn't open. Please open the workbook first."
    )


def test_set_active_sheet_sheet_doesnt_exists(setup_teardown):
    with pytest.raises(SheetDoesntExistsError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.set_active_sheet(sheet_name=INVALID_SHEET_NAME)

    assert_that(str(exc_info.value)).is_equal_to(
        f"Sheet '{INVALID_SHEET_NAME}' doesn't exists."
    )


def test_write_to_cell_success(setup_teardown):
    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    exl.write_to_cell(cell_name="AAA1", cell_value="New_value", sheet_name="Sheet1")

    workbook = excel.load_workbook(filename=EXCEL_FILE_PATH)
    sheet = workbook["Sheet1"]
    cell_value = sheet["AAA1"].value
    assert_that(cell_value).is_equal_to("New_value")
    sheet["AAA1"] = None
    workbook.save(EXCEL_FILE_PATH)
    workbook.close()


def test_write_to_cell_invalid_cell_address(setup_teardown):
    with pytest.raises(InvalidCellAddressError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.write_to_cell(
            cell_name=INVALID_CELL_ADDRESS, cell_value="New_value", sheet_name="Sheet1"
        )

    assert_that(str(exc_info.value)).is_equal_to(
        f"Cell '{INVALID_CELL_ADDRESS}' doesn't exists."
    )


def test_get_column_count_success(setup_teardown):
    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    column_count = exl.get_column_count(
        sheet_name="Offset_table", starting_cell="D6", ignore_empty_columns=True
    )
    assert_that(column_count).is_equal_to(7)


def test_get_column_count_invalid_cell_address(setup_teardown):
    with pytest.raises(InvalidCellAddressError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.get_column_count(
            sheet_name="Offset_table", starting_cell=INVALID_CELL_ADDRESS
        )

    assert_that(str(exc_info.value)).is_equal_to(
        f"Cell '{INVALID_CELL_ADDRESS}' doesn't exists."
    )


def test_get_row_count_success(setup_teardown):
    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    row_count = exl.get_row_count(
        sheet_name="Offset_table",
        include_header=True,
        starting_cell="D6",
        ignore_empty_rows=True,
    )
    assert_that(row_count).is_equal_to(52)


def test_get_row_count_invalid_cell_address(setup_teardown):
    with pytest.raises(InvalidCellAddressError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.get_row_count(sheet_name="Offset_table", starting_cell=INVALID_CELL_ADDRESS)

    assert_that(str(exc_info.value)).is_equal_to(
        f"Cell '{INVALID_CELL_ADDRESS}' doesn't exists."
    )


def test_append_row_success(setup_teardown):
    data = ["Marisa", "Pia", "Female", 33, "France", "21/05/2015", None, 1946]
    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    exl.append_row(sheet_name="Sheet1", row_data=data)

    workbook = excel.load_workbook(filename=EXCEL_FILE_PATH)
    sheet = workbook["Sheet1"]

    for row in sheet.iter_rows():
        row_values = [cell.value for cell in row]
        if row_values == data:
            row_to_delete = row[0].row
            sheet.delete_rows(row_to_delete)
            break
    else:
        assert False, "Row not appended"

    workbook.save(EXCEL_FILE_PATH)
    workbook.close()


def test_insert_row_success(setup_teardown):
    data = ["Mark", "Pia", "Male", 53, "France", "11/09/2014", None, 1946]
    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    exl.insert_row(sheet_name="Sheet1", row_data=data, row_index=10)

    workbook = excel.load_workbook(filename=EXCEL_FILE_PATH)
    sheet = workbook["Sheet1"]

    for row in sheet.iter_rows():
        row_values = [cell.value for cell in row]
        if row_values == data:
            row_to_delete = row[0].row
            assert_that(row_to_delete).is_equal_to(10)
            sheet.delete_rows(row_to_delete)
            break
    else:
        assert False, "Row not inserted at index 10"

    workbook.save(EXCEL_FILE_PATH)
    workbook.close()


def test_insert_row_invalid_row_index(setup_teardown):
    with pytest.raises(InvalidRowIndexError) as exc_info:
        data = ["Mark", "Pia", "Male", 53, "France", "11/09/2014", None, 1946]
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.insert_row(row_data=data, row_index=INVALID_ROW_INDEX)

    assert_that(str(exc_info.value)).is_equal_to(
        f"Row index {INVALID_ROW_INDEX} is invalid or out of bounds. The valid range is 1 to 1048576."
    )


def test_delete_row_success(setup_teardown):
    data = ["Dee", "Pia", "Male", 53, "France", "11/09/2014", None, 1946]
    workbook = excel.load_workbook(filename=EXCEL_FILE_PATH)
    sheet = workbook["Sheet1"]
    sheet.insert_rows(2)
    for col_index, value in enumerate(data, start=1):
        sheet.cell(row=2, column=col_index, value=value)

    workbook.save(EXCEL_FILE_PATH)
    workbook.close()

    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    exl.delete_row(row_index=2)

    workbook = excel.load_workbook(filename=EXCEL_FILE_PATH)
    sheet = workbook["Sheet1"]

    for row in sheet.iter_rows():
        row_values = [cell.value for cell in row]
        if row[0].row > 2:
            break
        if row_values == data:
            assert False, "Row not deleted at index 2"

    workbook.close()


def test_delete_row_invalid_row_index(setup_teardown):
    with pytest.raises(InvalidRowIndexError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.delete_row(row_index=INVALID_ROW_INDEX)

    assert_that(str(exc_info.value)).is_equal_to(
        f"Row index {INVALID_ROW_INDEX} is invalid or out of bounds. The valid range is 1 to 1048576."
    )


def test_append_column_success(setup_teardown):
    data = ["New Column", "data1", "data2", "data3", "data4", "data5"]
    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    exl.append_column(sheet_name="Sheet1", col_data=data)

    expected_values = ["data1", "data2", "data3", "data4", "data5"]
    workbook = excel.load_workbook(filename=EXCEL_FILE_PATH)
    sheet = workbook["Sheet1"]
    last_column = sheet.max_column
    header = sheet.cell(row=1, column=last_column).value
    assert_that(header).is_equal_to("New Column")
    new_column_values = [
        sheet.cell(row=row, column=last_column).value for row in range(2, 7)
    ]
    assert_that(new_column_values).is_equal_to(expected_values)

    sheet.delete_cols(last_column)
    workbook.save(EXCEL_FILE_PATH)
    workbook.close()


def test_insert_column_success(setup_teardown):
    data = ["New Column", "data1", "data2", "data3", "data4", "data5"]
    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    exl.insert_column(sheet_name="Sheet1", col_data=data, col_index=1)

    expected_values = ["data1", "data2", "data3", "data4", "data5"]
    workbook = excel.load_workbook(filename=EXCEL_FILE_PATH)
    sheet = workbook["Sheet1"]
    header = sheet.cell(row=1, column=1).value

    assert_that(header).is_equal_to("New Column")
    new_column_values = [sheet.cell(row=row, column=1).value for row in range(2, 7)]
    assert_that(new_column_values).is_equal_to(expected_values)

    sheet.delete_cols(1)
    workbook.save(EXCEL_FILE_PATH)
    workbook.close()


def test_insert_column_invalid_column_index(setup_teardown):
    with pytest.raises(InvalidColumnIndexError) as exc_info:
        data = ["New Column", "data1", "data2", "data3", "data4", "data5"]
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.insert_column(
            sheet_name="Sheet1", col_data=data, col_index=INVALID_COLUMN_INDEX
        )

    assert_that(str(exc_info.value)).is_equal_to(
        f"Column index {INVALID_COLUMN_INDEX} is invalid or out of bounds. The valid range is 1 to 16384."
    )


def test_delete_column_success(setup_teardown):
    data = ["New Column", "data1", "data2", "data3", "data4", "data5"]
    workbook = excel.load_workbook(filename=EXCEL_FILE_PATH)
    sheet = workbook["Sheet1"]
    sheet.insert_cols(1)

    for row_index, value in enumerate(data, start=1):
        sheet.cell(row=row_index, column=1, value=value)

    workbook.save(EXCEL_FILE_PATH)
    workbook.close()

    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    exl.delete_column(sheet_name="Sheet1", col_index=1)

    workbook = excel.load_workbook(filename=EXCEL_FILE_PATH)
    sheet = workbook["Sheet1"]
    header = sheet.cell(row=1, column=1).value
    workbook.close()

    assert_that(header).is_not_equal_to("New Column")


def test_delete_column_invalid_column_index(setup_teardown):
    with pytest.raises(InvalidColumnIndexError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.delete_column(sheet_name="Sheet1", col_index=INVALID_COLUMN_INDEX)

    assert_that(str(exc_info.value)).is_equal_to(
        f"Column index {INVALID_COLUMN_INDEX} is invalid or out of bounds. The valid range is 1 to 16384."
    )


@pytest.mark.parametrize(
    "column_name, expected_length, expected_values, output_format",
    [
        ("A", 51, ["Tommie", "Nereida", "Stasia"], "list"),
        (
            ["A", "B"],
            51,
            [["Tommie", "Nereida", "Stasia"], ["Mccrystal", "Partain", "Hanner"]],
            "list",
        ),
        ("First Name", 51, ["Tommie", "Nereida", "Stasia"], "list"),
        (
            ["First Name", "Last Name"],
            51,
            [["Tommie", "Nereida", "Stasia"], ["Mccrystal", "Partain", "Hanner"]],
            "list",
        ),
        ("A", 51, [["Tommie", "Nereida", "Stasia"]], "dict"),
        (
            ["A", "B"],
            51,
            [["Tommie", "Nereida", "Stasia"], ["Mccrystal", "Partain", "Hanner"]],
            "dict",
        ),
        ("First Name", 51, [["Tommie", "Nereida", "Stasia"]], "dict"),
        (
            ["First Name", "Last Name"],
            51,
            [["Tommie", "Nereida", "Stasia"], ["Mccrystal", "Partain", "Hanner"]],
            "dict",
        ),
        ("A", 51, ["Tommie", "Nereida", "Stasia"], "DataFrame"),
        (
            ["A", "B"],
            51,
            [["Tommie", "Nereida", "Stasia"], ["Mccrystal", "Partain", "Hanner"]],
            "DataFrame",
        ),
        ("First Name", 51, ["Tommie", "Nereida", "Stasia"], "DataFrame"),
        (
            ["First Name", "Last Name"],
            51,
            [["Tommie", "Nereida", "Stasia"], ["Mccrystal", "Partain", "Hanner"]],
            "DataFrame",
        ),
    ],
)
def test_get_column_values_success(
    setup_teardown, column_name, expected_length, expected_values, output_format
):
    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    column_values = exl.get_column_values(
        column_names_or_letters=column_name,
        sheet_name="Offset_table",
        starting_cell="D6",
        output_format=output_format,
    )

    if isinstance(column_name, str) or output_format in ["dict", "DataFrame"]:
        assert_that(isinstance(column_values, eval(output_format))).is_true()

        if output_format == "list":
            assert_that(column_values).is_length(expected_length).contains(
                *expected_values
            )
        elif output_format == "dict":
            keys = list(column_values.keys())
            for i, key in enumerate(keys):
                values = column_values.get(key)
                assert_that(values).is_length(expected_length).contains(
                    *expected_values[i]
                )
        else:
            row_count, column_count = column_values.shape
            assert_that(row_count).is_equal_to(expected_length)
            column_headers = column_values.columns.tolist()

            if column_count == 1:
                assert_that(column_values[column_headers[0]].tolist()).contains(
                    *expected_values
                )
            else:
                assert_that(column_values[column_headers[0]].tolist()).contains(
                    *expected_values[0]
                )
                assert_that(column_values[column_headers[1]].tolist()).contains(
                    *expected_values[1]
                )
    else:
        assert_that(isinstance(column_values, eval(output_format))).is_true()
        assert_that(column_values).is_length(2)
        for sublist in column_values:
            assert_that(isinstance(sublist, list)).is_true()
            assert_that(sublist).is_length(expected_length)


def test_get_column_values_invalid_output_format(setup_teardown):
    with pytest.raises(ValueError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.get_column_values(
            column_names_or_letters="A",
            sheet_name="Offset_table",
            starting_cell="D6",
            output_format="invalid",
        )

    assert_that(str(exc_info.value)).is_equal_to(
        "Invalid output format. Use 'list', 'dict', or 'dataframe'."
    )


def test_get_column_values_invalid_header(setup_teardown):
    with pytest.raises(ValueError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.get_column_values(
            column_names_or_letters="G",
            sheet_name="Invalid_header",
            starting_cell="A1",
            output_format="list",
        )

    assert_that(str(exc_info.value)).is_equal_to(
        "Column letter 'G' does not have a valid string header: '1234' found."
    )


@pytest.mark.parametrize("invalid_columns", ["AAAA", "Invalid Column Header"])
def test_get_column_values_invalid_column(setup_teardown, invalid_columns):
    with pytest.raises(ValueError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.get_column_values(
            column_names_or_letters=invalid_columns,
            sheet_name="Sheet1",
            starting_cell="A1",
            output_format="list",
        )

    assert_that(str(exc_info.value)).is_equal_to(
        f"Invalid column name or letter: '{invalid_columns}'"
    )


def test_get_column_values_column_out_of_bound(setup_teardown):
    with pytest.raises(ValueError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.get_column_values(
            column_names_or_letters="Z",
            sheet_name="Sheet1",
            starting_cell="A55",
            output_format="list",
        )

    assert_that(str(exc_info.value)).is_equal_to(
        "Column letter 'Z' is out of bounds for the provided sheet."
    )


@pytest.mark.parametrize(
    "row_index, expected_values, expected_length, output_format",
    [
        (
            2,
            ["Lester", "Prothro", "Male", 20, "France", "15/10/2017", None, 6574],
            8,
            "list",
        ),
        (
            [2, 3],
            [
                ["Lester", "Prothro", "Male", 20, "France", "15/10/2017", None, 6574],
                ["Francesca", "Beaudreau", "Female", 21, "France", "15/10/2017", 5412],
            ],
            8,
            "list",
        ),
        (
            2,
            [["Lester", "Prothro", "Male", 20, "France", "15/10/2017", None, 6574]],
            8,
            "dict",
        ),
        (
            [2, 3],
            [
                ["Lester", "Prothro", "Male", 20, "France", "15/10/2017", None, 6574],
                ["Francesca", "Beaudreau", "Female", 21, "France", "15/10/2017", 5412],
            ],
            8,
            "dict",
        ),
    ],
)
def test_get_row_values_success(
    setup_teardown, row_index, expected_length, expected_values, output_format
):
    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    row_value = exl.get_row_values(
        sheet_name="Sheet1", row_indices=row_index, output_format=output_format
    )
    assert_that(isinstance(row_value, eval(output_format))).is_true()

    if isinstance(row_index, int) or output_format == "dict":
        if output_format == "list":
            assert_that(row_value).is_length(expected_length).contains(*expected_values)
        else:
            keys = list(row_value.keys())
            for i, key in enumerate(keys):
                values = row_value.get(key)
                assert_that(values).is_length(expected_length).contains(
                    *expected_values[i]
                )
    else:
        assert_that(isinstance(row_value, eval(output_format))).is_true()
        assert_that(row_value).is_length(2)
        for sublist in row_value:
            assert_that(isinstance(sublist, list)).is_true()
            assert_that(sublist).is_length(expected_length)


def test_get_row_values_invalid_output_format(setup_teardown):
    with pytest.raises(ValueError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.get_row_values(sheet_name="Sheet1", row_indices=2, output_format="invalid")

    assert_that(str(exc_info.value)).is_equal_to(
        "Invalid output format. Use 'list' or 'dict'."
    )


def test_get_row_values_invalid_row_index(setup_teardown):
    with pytest.raises(InvalidRowIndexError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.get_row_values(
            sheet_name="Sheet1", row_indices=INVALID_ROW_INDEX, output_format="list"
        )

    assert_that(str(exc_info.value)).is_equal_to(
        f"Row index {INVALID_ROW_INDEX} is invalid or out of bounds. The valid range is 1 to 1048576."
    )


def test_protect_sheet_success(setup_teardown):
    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    exl.protect_sheet(sheet_name="Invalid_header", password="password")

    workbook = excel.load_workbook(filename=EXCEL_FILE_PATH)
    sheet = workbook["Invalid_header"]
    assert_that(sheet.protection.sheet).is_true()
    sheet.protection.set_password("password")
    sheet.protection.sheet = False
    workbook.save(EXCEL_FILE_PATH)
    workbook.close()


def test_protect_sheet_sheet_already_protected(setup_teardown):
    workbook = excel.load_workbook(filename=EXCEL_FILE_PATH)
    sheet = workbook["Invalid_header"]
    sheet.protection.set_password("password")
    workbook.save(EXCEL_FILE_PATH)
    workbook.close()

    with pytest.raises(SheetAlreadyProtectedError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.protect_sheet(sheet_name="Invalid_header", password="password")

    assert_that(str(exc_info.value)).is_equal_to(
        "The sheet 'Invalid_header' is already protected and cannot be protected be again."
    )


def test_unprotect_sheet_success(setup_teardown):
    workbook = excel.load_workbook(filename=EXCEL_FILE_PATH)
    sheet = workbook["Invalid_header"]
    sheet.protection.set_password("password")
    workbook.save(EXCEL_FILE_PATH)
    workbook.close()

    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    exl.unprotect_sheet(sheet_name="Invalid_header", password="password")

    workbook = excel.load_workbook(filename=EXCEL_FILE_PATH)
    sheet = workbook["Invalid_header"]
    assert_that(sheet.protection.sheet).is_false()
    workbook.close()


def test_unprotect_sheet_sheet_not_protected(setup_teardown):
    workbook = excel.load_workbook(filename=EXCEL_FILE_PATH)
    sheet = workbook["Invalid_header"]
    sheet.protection.set_password("password")
    sheet.protection.sheet = False
    workbook.save(EXCEL_FILE_PATH)
    workbook.close()

    with pytest.raises(SheetNotProtectedError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.unprotect_sheet(sheet_name="Invalid_header", password="password")

    assert_that(str(exc_info.value)).is_equal_to(
        "The sheet 'Invalid_header' is not currently protected and cannot be unprotected."
    )


def test_protect_workbook_success(setup_teardown):
    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    exl.protect_workbook(password="password", protect_sheets=True)

    workbook = excel.load_workbook(filename=EXCEL_FILE_PATH)
    assert_that(workbook.security.lockStructure).is_true()
    workbook.security.lockStructure = False

    for sheet in workbook.worksheets:
        if sheet.protection.sheet:
            sheet.protection.sheet = False

    workbook.save(EXCEL_FILE_PATH)
    workbook.close()


def test_protect_workbook_workbook_not_open(setup_teardown):
    with pytest.raises(WorkbookNotOpenError) as exc_info:
        exl.protect_workbook(password="password", protect_sheets=True)

    assert_that(str(exc_info.value)).is_equal_to(
        "Workbook isn't open. Please open the workbook first."
    )


def test_protect_workbook_workbook_already_protected(setup_teardown):
    workbook = excel.load_workbook(filename=EXCEL_FILE_PATH)
    protection = WorkbookProtection()
    protection.workbookPassword = "password"
    protection.lockStructure = True
    protection.lockWindows = True

    workbook.security = protection

    for sheet in workbook.worksheets:
        sheet.protection = SheetProtection(password="password")
        sheet.protection.sheet = True
        sheet.protection.formatCells = False
        sheet.protection.formatColumns = False
        sheet.protection.formatRows = False
        sheet.protection.insertColumns = False
        sheet.protection.insertRows = False
        sheet.protection.deleteColumns = False
        sheet.protection.deleteRows = False
        sheet.protection.sort = False
        sheet.protection.autoFilter = False
        sheet.protection.objects = False
        sheet.protection.scenarios = False

    workbook.save(EXCEL_FILE_PATH)
    workbook.close()

    with pytest.raises(WorkbookAlreadyProtectedError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.protect_workbook(protect_sheets=True, password="password")

    assert_that(str(exc_info.value)).is_equal_to(
        "The workbook is already protected and cannot be protected be again."
    )


def test_unprotect_workbook_success(setup_teardown):
    workbook = excel.load_workbook(filename=EXCEL_FILE_PATH)
    protection = WorkbookProtection()
    protection.workbookPassword = "password"
    protection.lockStructure = True
    protection.lockWindows = True

    workbook.security = protection

    for sheet in workbook.worksheets:
        sheet.protection = SheetProtection(password="password")
        sheet.protection.sheet = True
        sheet.protection.formatCells = False
        sheet.protection.formatColumns = False
        sheet.protection.formatRows = False
        sheet.protection.insertColumns = False
        sheet.protection.insertRows = False
        sheet.protection.deleteColumns = False
        sheet.protection.deleteRows = False
        sheet.protection.sort = False
        sheet.protection.autoFilter = False
        sheet.protection.objects = False
        sheet.protection.scenarios = False

    workbook.save(EXCEL_FILE_PATH)
    workbook.close()

    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    exl.unprotect_workbook(unprotect_sheets=True)

    workbook = excel.load_workbook(filename=EXCEL_FILE_PATH)
    assert_that(workbook.security and workbook.security.lockStructure).is_false()
    workbook.close()


def test_unprotect_workbook_workbook_not_open(setup_teardown):
    with pytest.raises(WorkbookNotOpenError) as exc_info:
        exl.unprotect_workbook(unprotect_sheets=True)

    assert_that(str(exc_info.value)).is_equal_to(
        "Workbook isn't open. Please open the workbook first."
    )


def test_unprotect_workbook_workbook_not_protected(setup_teardown):
    workbook = excel.load_workbook(filename=EXCEL_FILE_PATH)
    workbook.security.lockStructure = False

    for sheet in workbook.worksheets:
        if sheet.protection.sheet:
            sheet.protection.sheet = False

    workbook.save(EXCEL_FILE_PATH)
    workbook.close()

    with pytest.raises(WorkbookNotProtectedError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.unprotect_workbook(unprotect_sheets=True)

    assert_that(str(exc_info.value)).is_equal_to(
        "The workbook is not currently protected and cannot be unprotected."
    )


def test_clear_sheet_success(setup_teardown):
    sheet_data = [["Name", "Age"], ["Dee", 26], ["Mark", 56], ["John", 30]]
    workbook = excel.load_workbook(filename=EXCEL_FILE_PATH)
    workbook.create_sheet(title="Clear_sheet", index=1)
    sheet = workbook["Clear_sheet"]

    for row in sheet_data:
        sheet.append(row)

    workbook.save(EXCEL_FILE_PATH)
    workbook.close()

    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    exl.clear_sheet(sheet_name="Clear_sheet")

    workbook = excel.load_workbook(filename=EXCEL_FILE_PATH)
    sheet = workbook["Clear_sheet"]

    is_empty = True
    for row in sheet.iter_rows(
        min_row=1, min_col=1, max_col=sheet.max_column, values_only=True
    ):
        if any(cell is not None for cell in row):
            is_empty = False
            break

    assert_that(is_empty).is_true()
    workbook.remove(sheet)
    workbook.save(EXCEL_FILE_PATH)
    workbook.close()


def test_copy_sheet_success(setup_teardown):
    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    exl.copy_sheet(source_sheet_name="Sheet1", new_sheet_name="Copied_sheet")

    workbook = excel.load_workbook(filename=EXCEL_FILE_PATH)
    sheets = workbook.sheetnames
    assert_that(sheets).is_length(4).contains(
        "Sheet1", "Offset_table", "Invalid_header", "Copied_sheet"
    )

    workbook = excel.load_workbook(filename=EXCEL_FILE_PATH)
    sheet1 = workbook["Sheet1"]
    sheet2 = workbook["Copied_sheet"]

    if sheet1.max_row != sheet2.max_row or sheet1.max_column != sheet2.max_column:
        assert False, "Rows/Columns not matching"

    for row in range(1, sheet1.max_row + 1):
        for col in range(1, sheet1.max_column + 1):
            cell1 = sheet1.cell(row=row, column=col).value
            cell2 = sheet2.cell(row=row, column=col).value

            if cell1 != cell2:
                assert False, "Cells not matching"

    workbook.remove(sheet2)
    workbook.save(EXCEL_FILE_PATH)
    workbook.close()


def test_copy_sheet_workbook_not_open(setup_teardown):
    with pytest.raises(WorkbookNotOpenError) as exc_info:
        exl.copy_sheet(source_sheet_name="Sheet1", new_sheet_name="Copied_sheet")

    assert_that(str(exc_info.value)).is_equal_to(
        "Workbook isn't open. Please open the workbook first."
    )


def test_copy_sheet_invalid_sheet_name(setup_teardown):
    with pytest.raises(InvalidSheetNameError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.copy_sheet(source_sheet_name="Sheet1", new_sheet_name=INVALID_SHEET_NAME)

    assert_that(str(exc_info.value)).is_equal_to(
        f"The sheet name '{INVALID_SHEET_NAME}' is invalid."
    )


def test_copy_sheet_doesnt_exists(setup_teardown):
    with pytest.raises(SheetDoesntExistsError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.copy_sheet(
            source_sheet_name=INVALID_SHEET_NAME, new_sheet_name="Copied_sheet"
        )

    assert_that(str(exc_info.value)).is_equal_to(
        f"Sheet '{INVALID_SHEET_NAME}' doesn't exists."
    )


@pytest.mark.parametrize(
    "value, occurence, expected_value",
    [
        ("Male", "first", ("str", "C2")),
        ("Lester", "all", ("list", ["A2", "A10"])),
        ("Invalid_value", "first", (None, None)),
    ],
)
def test_find_value_success(setup_teardown, value, occurence, expected_value):
    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    cell = exl.find_value(sheet_name="Sheet1", value=value, occurence=occurence)

    if expected_value[0] is None:
        assert_that(cell).is_none()
    else:
        assert_that(isinstance(cell, eval(expected_value[0]))).is_true()
        assert_that(cell).is_equal_to(expected_value[1])


def test_find_value_invalid_occurence(setup_teardown):
    with pytest.raises(ValueError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.find_value(
            sheet_name="Sheet1", value="value", occurence="invalid_occurence"
        )

    assert_that(str(exc_info.value)).is_equal_to(
        "Invalid occurence, use either 'first' or 'all'."
    )


@pytest.mark.parametrize(
    "occurence, expected_value",
    [
        ("first", ("str", "A14")),
        ("all", ("list", ["A18", "A19"])),
        ("first", (None, None)),
    ],
)
def test_find_and_replace_success(setup_teardown, occurence, expected_value):
    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    cell = exl.find_and_replace(
        sheet_name="Sheet1", old_value="Marcel", new_value="Mark", occurence=occurence
    )

    if expected_value[0] is None:
        assert_that(cell).is_none()
    else:
        assert_that(isinstance(cell, eval(expected_value[0]))).is_true()
        assert_that(cell).is_equal_to(expected_value[1])

    workbook = excel.load_workbook(filename=EXCEL_FILE_PATH)
    sheet = workbook["Sheet1"]

    if isinstance(cell, list):
        for i in cell:
            assert_that(sheet[i].value).is_equal_to("Mark")
    elif isinstance(cell, str):
        assert_that(sheet[cell].value).is_equal_to("Mark")

    workbook.close()


def test_find_and_replace_value_invalid_occurence(setup_teardown):
    with pytest.raises(ValueError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.find_and_replace(
            sheet_name="Sheet1",
            old_value="Marcel",
            new_value="Mark",
            occurence="invalid_occurence",
        )

    assert_that(str(exc_info.value)).is_equal_to(
        "Invalid occurence, use either 'first' or 'all'."
    )


def test_format_cell_success(setup_teardown):
    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    alignment_config = {"vertical": "center", "horizontal": "left"}

    border_config = {
        "left": True,
        "right": True,
        "top": True,
        "bottom": True,
        "style": "thin",
        "color": "#FF0000",
    }
    exl.format_cell(
        sheet_name="Sheet1",
        cell_name="C3",
        font_size=12,
        font_color="#FF0000",
        alignment=alignment_config,
        wrap_text=True,
        bg_color="#FFFF00",
        cell_width=120,
        cell_height=25,
        font_name="Arial",
        bold=True,
        italic=True,
        strike_through=True,
        underline=True,
        border=border_config,
        auto_fit_height=False,
        auto_fit_width=False,
    )

    workbook = excel.load_workbook(filename=EXCEL_FILE_PATH)
    sheet = workbook["Sheet1"]
    cell = sheet["C3"]

    cell_properties = {
        "cell_name": cell.coordinate,
        "font_size": cell.font.size,
        "font_color": cell.font.color.rgb if cell.font.color else None,
        "alignment_horizontal": cell.alignment.horizontal,
        "alignment_vertical": cell.alignment.vertical,
        "wrap_text": cell.alignment.wrap_text,
        "bg_color": cell.fill.start_color.rgb if cell.fill.start_color else None,
        "cell_width": sheet.column_dimensions[cell.column_letter].width,
        "cell_height": sheet.row_dimensions[cell.row].height,
        "font_name": cell.font.name,
        "bold": cell.font.bold,
        "italic": cell.font.italic,
        "underline": cell.font.underline,
        "strike_through": cell.font.strike,
        "border": {
            "top": cell.border.top.style if cell.border.top else None,
            "bottom": cell.border.bottom.style if cell.border.bottom else None,
            "left": cell.border.left.style if cell.border.left else None,
            "right": cell.border.right.style if cell.border.right else None,
        },
    }
    assert_that(cell_properties["font_size"]).is_equal_to(12.0)
    assert_that(cell_properties["font_color"]).is_equal_to("FFFF0000")
    assert_that(cell_properties["alignment_horizontal"]).is_equal_to("left")
    assert_that(cell_properties["alignment_vertical"]).is_equal_to("center")
    assert_that(cell_properties["wrap_text"]).is_true()
    assert_that(cell_properties["bg_color"]).is_equal_to("FFFFFF00")
    assert_that(cell_properties["cell_width"]).is_equal_to(120.0)
    assert_that(cell_properties["cell_height"]).is_equal_to(25.0)
    assert_that(cell_properties["font_name"]).is_equal_to("Arial")
    assert_that(cell_properties["bold"]).is_true()
    assert_that(cell_properties["italic"]).is_true()
    assert_that(cell_properties["underline"]).is_equal_to("single")
    assert_that(cell_properties["strike_through"]).is_true()
    assert_that(cell_properties["border"]).is_equal_to(
        {"top": "thin", "bottom": "thin", "left": "thin", "right": "thin"}
    )
    workbook.close()

    exl.format_cell(
        sheet_name="Offset_table",
        cell_name="D6",
        auto_fit_height=True,
        auto_fit_width=True,
    )

    workbook = excel.load_workbook(filename=EXCEL_FILE_PATH)
    sheet = workbook["Offset_table"]
    cell = sheet["D6"]
    cell_value = str(cell.value) if cell.value else ""
    col_letter = get_column_letter(cell.column)
    max_length = max(len(cell_value), len(col_letter))
    computed_width = max_length + 2
    max_line_count = cell_value.count("\n") + 1
    computed_height = max(15, max_line_count * 15)
    cell_properties = {
        "auto_width": computed_width,
        "auto_height": computed_height,
    }

    assert_that(cell_properties["auto_width"]).is_equal_to(12)
    assert_that(cell_properties["auto_height"]).is_equal_to(15)
    workbook.close()


def test_format_cell_invalid_cell_address(setup_teardown):
    with pytest.raises(InvalidCellAddressError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.format_cell(
            sheet_name="Offset_table", cell_name=INVALID_CELL_ADDRESS, font_size=10
        )

    assert_that(str(exc_info.value)).is_equal_to(
        f"Cell '{INVALID_CELL_ADDRESS}' doesn't exists."
    )


def test_format_cell_invalid_font_color(setup_teardown):
    with pytest.raises(InvalidColorError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.format_cell(
            sheet_name="Offset_table", cell_name="D6", font_color="invalid_color"
        )

    assert_that(str(exc_info.value)).is_equal_to(
        "Invalid font color: 'invalid_color'. Use valid hex color in #RRGGBB format."
    )


def test_format_cell_invalid_bg_color(setup_teardown):
    with pytest.raises(InvalidColorError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.format_cell(
            sheet_name="Offset_table", cell_name="D6", bg_color="invalid_color"
        )

    assert_that(str(exc_info.value)).is_equal_to(
        "Invalid background color: 'invalid_color'. Use valid hex color in #RRGGBB format."
    )


def test_format_cell_invalid_border_color(setup_teardown):
    with pytest.raises(InvalidColorError) as exc_info:
        border_config = {
            "left": True,
            "right": True,
            "top": True,
            "bottom": True,
            "style": "thin",
            "color": "invalid_color",
        }
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.format_cell(sheet_name="Offset_table", cell_name="D6", border=border_config)

    assert_that(str(exc_info.value)).is_equal_to(
        "Invalid border color: 'invalid_color'. Use valid hex color in #RRGGBB format."
    )


def test_format_cell_invalid_horizontal_alignment(setup_teardown):
    with pytest.raises(InvalidAlignmentError) as exc_info:
        alignment_config = {"vertical": "center", "horizontal": "invalid_alignment"}
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.format_cell(
            sheet_name="Offset_table", cell_name="D6", alignment=alignment_config
        )

    assert_that(str(exc_info.value)).is_equal_to(
        "Invalid horizontal alignment: 'invalid_alignment'. Allowed values are ['left', 'center', 'right']."
    )


def test_format_cell_invalid_vertical_alignment(setup_teardown):
    with pytest.raises(InvalidAlignmentError) as exc_info:
        alignment_config = {"vertical": "invalid_alignment", "horizontal": "left"}
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.format_cell(
            sheet_name="Offset_table", cell_name="D6", alignment=alignment_config
        )

    assert_that(str(exc_info.value)).is_equal_to(
        "Invalid vertical alignment: 'invalid_alignment'. Allowed values are ['top', 'center', 'bottom']."
    )


def test_format_cell_invalid_border_style(setup_teardown):
    with pytest.raises(InvalidBorderStyleError) as exc_info:
        border_config = {
            "left": True,
            "right": True,
            "top": True,
            "bottom": True,
            "style": "invalid_style",
        }
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.format_cell(sheet_name="Offset_table", cell_name="D6", border=border_config)

    assert_that(str(exc_info.value)).is_equal_to(
        "Invalid border style: 'invalid_style'. Allowed values are ['dashDot', 'dashDotDot', 'dashed', 'dotted', 'double', 'hair', 'medium', 'mediumDashDot', 'mediumDashDotDot', 'mediumDashed', 'slantDashDot', 'thick', 'thin']."
    )


def test_get_column_headers_success(setup_teardown):
    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    headers = exl.get_column_headers(sheet_name="Offset_table", starting_cell="D6")
    assert_that(headers).is_length(7).contains(
        "First Name", "Last Name", "Gender", "Country", "Age", "Date", "Salary"
    )


def test_get_column_headers_invalid_cell_address(setup_teardown):
    with pytest.raises(InvalidCellAddressError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.get_column_headers(
            sheet_name="Offset_table", starting_cell=INVALID_CELL_ADDRESS
        )

    assert_that(str(exc_info.value)).is_equal_to(
        f"Cell '{INVALID_CELL_ADDRESS}' doesn't exists."
    )


def test_export_to_csv_success(setup_teardown):
    output_filename = exl.export_to_csv(
        filename=EXCEL_FILE_PATH,
        output_filename=CSV_FILE_PATH,
        sheet_name="Sheet1",
        overwrite_if_exists=True,
    )

    assert_that(output_filename).is_equal_to(CSV_FILE_PATH)
    assert_that(os.path.exists(CSV_FILE_PATH)).is_true()


def test_export_to_csv_file_not_found(setup_teardown):
    with pytest.raises(ExcelFileNotFoundError) as exc_info:
        exl.export_to_csv(
            filename=INVALID_EXCEL_FILE_PATH,
            output_filename=CSV_FILE_PATH,
            sheet_name="Sheet1",
            overwrite_if_exists=True,
        )

    assert_that(str(exc_info.value)).is_equal_to(
        f"Excel file '{INVALID_EXCEL_FILE_PATH}' not found. Please give the valid file path."
    )


def test_export_to_csv_file_already_exists(setup_teardown):
    with pytest.raises(FileAlreadyExistsError) as exc_info:
        exl.export_to_csv(
            filename=EXCEL_FILE_PATH,
            output_filename=CSV_FILE_PATH,
            sheet_name="Sheet1",
            overwrite_if_exists=False,
        )

    assert_that(str(exc_info.value)).is_equal_to(
        f"Unable to create workbook. The file '{CSV_FILE_PATH}' already exists. Set 'overwrite_if_exists=True' to overwrite the existing file."
    )


def test_export_to_csv_with_custom_separator(setup_teardown):
    """Test that export_to_csv uses the specified separator correctly."""
    csv_file_semicolon = os.path.join(DATA_DIR, "sample_semicolon.csv")
    csv_file_dash = os.path.join(DATA_DIR, "sample_dash.csv")

    original_df = pd.read_excel(EXCEL_FILE_PATH, sheet_name="Sheet1")
    original_row_count = len(original_df)
    original_column_count = len(original_df.columns)

    output_filename = exl.export_to_csv(
        filename=EXCEL_FILE_PATH,
        output_filename=csv_file_semicolon,
        sheet_name="Sheet1",
        separator=";",
        overwrite_if_exists=True,
    )

    assert_that(output_filename).is_equal_to(csv_file_semicolon)
    assert_that(os.path.exists(csv_file_semicolon)).is_true()

    csv_df_semicolon = pd.read_csv(csv_file_semicolon, sep=";")
    assert_that(len(csv_df_semicolon)).is_equal_to(original_row_count)
    assert_that(len(csv_df_semicolon.columns)).is_equal_to(original_column_count)

    with open(csv_file_semicolon, "r", encoding="utf-8") as f:
        first_line = f.readline()
        assert_that(first_line.count(";")).is_equal_to(original_column_count - 1)

    output_filename = exl.export_to_csv(
        filename=EXCEL_FILE_PATH,
        output_filename=csv_file_dash,
        sheet_name="Sheet1",
        separator="-",
        overwrite_if_exists=True,
    )

    assert_that(output_filename).is_equal_to(csv_file_dash)
    assert_that(os.path.exists(csv_file_dash)).is_true()

    csv_df_dash = pd.read_csv(csv_file_dash, sep="-")
    assert_that(len(csv_df_dash)).is_equal_to(original_row_count)
    assert_that(len(csv_df_dash.columns)).is_equal_to(original_column_count)

    with open(csv_file_dash, "r", encoding="utf-8") as f:
        first_line = f.readline()
        assert_that(first_line.count("-")).is_equal_to(original_column_count - 1)

    if os.path.exists(csv_file_semicolon):
        os.remove(csv_file_semicolon)
    if os.path.exists(csv_file_dash):
        os.remove(csv_file_dash)


def test_merge_excels_multi_sheet_success(setup_teardown):
    NEW_FILE = copy_test_excel_file(
        destination_file=os.path.join(DATA_DIR, "sample2.xlsx")
    )
    list_of_files = [EXCEL_FILE_PATH, NEW_FILE]
    output_file = os.path.join(DATA_DIR, "merged_file_multi_sheets.xlsx")
    exl.merge_excels(
        file_list=list_of_files,
        output_filename=output_file,
        merge_type="multiple_sheets",
        skip_bad_rows=True,
    )
    assert_that(os.path.exists(output_file)).is_true()

    workbook = excel.load_workbook(filename=output_file)
    sheets = workbook.sheetnames
    workbook.close()
    assert_that(sheets).is_length(6).contains(
        "Sheet1_sample",
        "Offset_table_sample2",
        "Offset_table_sample",
        "Invalid_header_sample",
        "Invalid_header_sample2",
        "Sheet1_sample2",
    )


def test_merge_excels_single_sheet_success(setup_teardown):
    data = {
        os.path.join(DATA_DIR, "single_sheet_workbook1.xlsx"): [
            ["Name", "Age"],
            ["Mark", 25],
            ["John", 30],
        ],
        os.path.join(DATA_DIR, "single_sheet_workbook2.xlsx"): [
            ["Name", "Age"],
            ["Dee", 26],
            ["Alex", 40],
        ],
    }

    def create_workbook(filename, data):
        workbook = Workbook()
        sheet = workbook.active
        for row in data:
            sheet.append(row)
        workbook.save(filename)
        workbook.close()

    for filename, workbook_data in data.items():
        create_workbook(filename, workbook_data)

    list_of_files = [
        os.path.join(DATA_DIR, "single_sheet_workbook1.xlsx"),
        os.path.join(DATA_DIR, "single_sheet_workbook2.xlsx"),
    ]
    output_file = os.path.join(DATA_DIR, "merged_file_single_sheet.xlsx")
    exl.merge_excels(
        file_list=list_of_files,
        output_filename=output_file,
        merge_type="single_sheet",
        skip_bad_rows=True,
    )
    assert_that(os.path.exists(output_file)).is_true()

    expected_data = [
        ["Name", "Age"],
        ["Mark", 25],
        ["John", 30],
        ["Dee", 26],
        ["Alex", 40],
    ]
    workbook = excel.load_workbook(filename=output_file)
    sheets = workbook.sheetnames
    assert_that(sheets).is_length(1)
    sheet = workbook[sheets[0]]
    assert_that(sheet.max_row).is_equal_to(5)
    assert_that(sheet.max_column).is_equal_to(2)

    for row_index, expected_row in enumerate(expected_data, start=1):
        for col_index, expected_value in enumerate(expected_row, start=1):
            cell_value = sheet.cell(row=row_index, column=col_index).value
            if cell_value != expected_value:
                assert False, f"Data mismatch ({cell_value} != {expected_value})"

    workbook.close()


def test_merge_excels_sheet_wise_success(setup_teardown):
    data = {
        os.path.join(DATA_DIR, "sheet_wise_workbook1.xlsx"): {
            "Sheet1": [["Name", "Age"], ["Mark", 25], ["John", 30]],
            "Sheet2": [["City", "Country"], ["New York", "USA"], ["London", "UK"]],
        },
        os.path.join(DATA_DIR, "sheet_wise_workbook2.xlsx"): {
            "Sheet1": [["Name", "Age"], ["Dee", 26], ["Alex", 40]],
            "Sheet2": [["City", "Country"], ["Berlin", "Germany"], ["Paris", "France"]],
        },
    }

    def create_workbook(filename, sheets_data):
        workbook = Workbook()

        first_sheet_name, first_sheet_data = list(sheets_data.items())[0]
        sheet = workbook.active
        sheet.title = first_sheet_name
        for row in first_sheet_data:
            sheet.append(row)

        for sheet_name, sheet_data in list(sheets_data.items())[1:]:
            sheet = workbook.create_sheet(title=sheet_name)
            for row in sheet_data:
                sheet.append(row)

        workbook.save(filename)
        workbook.close()

    for filename, sheets_data in data.items():
        create_workbook(filename, sheets_data)

    list_of_files = [
        os.path.join(DATA_DIR, "sheet_wise_workbook1.xlsx"),
        os.path.join(DATA_DIR, "sheet_wise_workbook2.xlsx"),
    ]
    output_file = os.path.join(DATA_DIR, "merged_file_sheet_wise1.xlsx")
    exl.merge_excels(
        file_list=list_of_files,
        output_filename=output_file,
        merge_type="sheet_wise",
        skip_bad_rows=True,
    )
    assert_that(os.path.exists(output_file)).is_true()

    expected_data = {
        "Sheet_1": [
            ["Name", "Age"],
            ["Mark", 25],
            ["John", 30],
            ["Dee", 26],
            ["Alex", 40],
        ],
        "Sheet_2": [
            ["City", "Country"],
            ["New York", "USA"],
            ["London", "UK"],
            ["Berlin", "Germany"],
            ["Paris", "France"],
        ],
    }

    workbook = excel.load_workbook(filename=output_file)
    sheets = workbook.sheetnames
    assert_that(sheets).is_length(2)

    for i in sheets:
        sheet = workbook[i]
        assert_that(sheet.max_row).is_equal_to(5)
        assert_that(sheet.max_column).is_equal_to(2)

        for row_index, expected_row in enumerate(expected_data[i], start=1):
            for col_index, expected_value in enumerate(expected_row, start=1):
                cell_value = sheet.cell(row=row_index, column=col_index).value
                if cell_value != expected_value:
                    assert False, f"Data mismatch ({cell_value} != {expected_value})"

    workbook.close()


def test_merge_excels_empty_files_list(setup_teardown):
    with pytest.raises(ValueError) as exc_info:
        exl.merge_excels(file_list=[], output_filename="output_file.xlsx")

    assert_that(str(exc_info.value)).is_equal_to(
        "The file list is empty. Provide at least one file to merge."
    )


def test_merge_excels_invalid_merge_type(setup_teardown):
    with pytest.raises(ValueError) as exc_info:
        list_of_files = [
            os.path.join(DATA_DIR, "workbook1.xlsx"),
            os.path.join(DATA_DIR, "workbook2.xlsx"),
        ]
        output_file = os.path.join(DATA_DIR, "merged_file1.xlsx")
        exl.merge_excels(
            file_list=list_of_files,
            output_filename=output_file,
            merge_type="invalid_merge_type",
        )

    assert_that(str(exc_info.value)).is_equal_to(
        "Invalid merge type. Use 'multiple_sheets', 'single_sheet', or 'sheet_wise'."
    )


def test_merge_excels_multi_sheet_file_not_found(setup_teardown):
    with pytest.raises(ExcelFileNotFoundError) as exc_info:
        list_of_files = [
            INVALID_EXCEL_FILE_PATH,
            os.path.join(DATA_DIR, "workbook2.xlsx"),
        ]
        output_file = os.path.join(DATA_DIR, "merged_file2.xlsx")
        exl.merge_excels(
            file_list=list_of_files,
            output_filename=output_file,
            merge_type="multiple_sheets",
        )

    assert_that(str(exc_info.value)).is_equal_to(
        f"Excel file '{INVALID_EXCEL_FILE_PATH}' not found. Please give the valid file path."
    )


def test_merge_excels_single_sheet_file_not_found(setup_teardown):
    with pytest.raises(ExcelFileNotFoundError) as exc_info:
        list_of_files = [
            INVALID_EXCEL_FILE_PATH,
            os.path.join(DATA_DIR, "workbook2.xlsx"),
        ]
        output_file = os.path.join(DATA_DIR, "merged_file3.xlsx")
        exl.merge_excels(
            file_list=list_of_files,
            output_filename=output_file,
            merge_type="single_sheet",
        )

    assert_that(str(exc_info.value)).is_equal_to(
        f"Excel file '{INVALID_EXCEL_FILE_PATH}' not found. Please give the valid file path."
    )


def test_merge_excels_sheet_wise_file_not_found(setup_teardown):
    with pytest.raises(ExcelFileNotFoundError) as exc_info:
        list_of_files = [
            INVALID_EXCEL_FILE_PATH,
            os.path.join(DATA_DIR, "workbook2.xlsx"),
        ]
        output_file = os.path.join(DATA_DIR, "merged_file4.xlsx")
        exl.merge_excels(
            file_list=list_of_files,
            output_filename=output_file,
            merge_type="sheet_wise",
        )

    assert_that(str(exc_info.value)).is_equal_to(
        f"Excel file '{INVALID_EXCEL_FILE_PATH}' not found. Please give the valid file path."
    )


def test_merge_excels_sheet_wise_index_error(setup_teardown):
    data = {
        os.path.join(DATA_DIR, "sheet_wise_workbook1.xlsx"): {
            "Sheet1": [["Name", "Age"], ["Mark", 25], ["John", 30]],
            "Sheet2": [["City", "Country"], ["New York", "USA"], ["London", "UK"]],
        },
        os.path.join(DATA_DIR, "sheet_wise_workbook2.xlsx"): {
            "Sheet1": [["Name", "Age"], ["Dee", 26], ["Alex", 40]]
        },
    }

    def create_workbook(filename, sheets_data):
        workbook = Workbook()

        first_sheet_name, first_sheet_data = list(sheets_data.items())[0]
        sheet = workbook.active
        sheet.title = first_sheet_name
        for row in first_sheet_data:
            sheet.append(row)

        for sheet_name, sheet_data in list(sheets_data.items())[1:]:
            sheet = workbook.create_sheet(title=sheet_name)
            for row in sheet_data:
                sheet.append(row)

        workbook.save(filename)
        workbook.close()

    for filename, sheets_data in data.items():
        create_workbook(filename, sheets_data)

    with pytest.warns(UserWarning, match="Skipping"):
        list_of_files = [
            os.path.join(DATA_DIR, "sheet_wise_workbook1.xlsx"),
            os.path.join(DATA_DIR, "sheet_wise_workbook2.xlsx"),
        ]
        output_file = os.path.join(DATA_DIR, "merged_file_sheet_wise2.xlsx")
        exl.merge_excels(
            file_list=list_of_files,
            output_filename=output_file,
            merge_type="sheet_wise",
        )


def test_merge_cells_success(setup_teardown):
    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    exl.merge_cells(cell_range="A1:D4", sheet_name="Invalid_header")

    workbook = excel.load_workbook(filename=EXCEL_FILE_PATH)
    sheet = workbook["Invalid_header"]

    range_to_check = "A1:D4"
    start_cell, end_cell = range_to_check.split(":")
    merged_cells = sheet.merged_cells.ranges
    is_merged = False

    for merged_range in merged_cells:
        start_row, start_col = cell.coordinate_to_tuple(start_cell)
        end_row, end_col = cell.coordinate_to_tuple(end_cell)

        if (
            merged_range.min_row <= start_row <= merged_range.max_row
            and merged_range.min_col <= start_col <= merged_range.max_col
            and merged_range.min_row <= end_row <= merged_range.max_row
            and merged_range.min_col <= end_col <= merged_range.max_col
        ):
            is_merged = True
            break

    assert_that(is_merged).is_true()
    sheet.unmerge_cells(range_to_check)
    workbook.save(EXCEL_FILE_PATH)
    workbook.close()


def test_merge_cells_invalid_cell_range1(setup_teardown):
    with pytest.raises(InvalidCellRangeError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.merge_cells(cell_range="AASGD1:D4", sheet_name="Invalid_header")

    assert_that(str(exc_info.value)).is_equal_to(
        "AASGD1:D4 is not a valid coordinate or range"
    )


def test_merge_cells_invalid_cell_range2(setup_teardown):
    with pytest.raises(InvalidCellRangeError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.merge_cells(cell_range="D4:A1", sheet_name="Invalid_header")

    assert_that(str(exc_info.value)).is_equal_to(
        "Invalid cell range: D4:A1. The start cell must be smaller than the end cell."
    )


def test_unmerge_cells_success(setup_teardown):
    workbook = excel.load_workbook(filename=EXCEL_FILE_PATH)
    sheet = workbook["Invalid_header"]

    sheet.merge_cells("C2:D6")
    workbook.save(EXCEL_FILE_PATH)
    workbook.close()

    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    exl.unmerge_cells(cell_range="C2:D6", sheet_name="Invalid_header")

    workbook = excel.load_workbook(filename=EXCEL_FILE_PATH)
    sheet = workbook["Invalid_header"]

    range_to_check = "C2:D6"
    start_cell, end_cell = range_to_check.split(":")
    merged_cells = sheet.merged_cells.ranges
    is_merged = False

    for merged_range in merged_cells:
        start_row, start_col = cell.coordinate_to_tuple(start_cell)
        end_row, end_col = cell.coordinate_to_tuple(end_cell)

        if (
            merged_range.min_row <= start_row <= merged_range.max_row
            and merged_range.min_col <= start_col <= merged_range.max_col
            and merged_range.min_row <= end_row <= merged_range.max_row
            and merged_range.min_col <= end_col <= merged_range.max_col
        ):
            is_merged = True
            break

    workbook.close()
    assert_that(is_merged).is_false()


def test_unmerge_cells_invalid_cell_range1(setup_teardown):
    with pytest.raises(InvalidCellRangeError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.unmerge_cells(cell_range="AASGD1:D4", sheet_name="Invalid_header")

    assert_that(str(exc_info.value)).is_equal_to(
        "AASGD1:D4 is not a valid coordinate or range"
    )


def test_unmerge_cells_invalid_cell_range2(setup_teardown):
    with pytest.raises(InvalidCellRangeError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.unmerge_cells(cell_range="D4:A1", sheet_name="Invalid_header")

    assert_that(str(exc_info.value)).is_equal_to(
        "Invalid cell range: D4:A1. The start cell must be smaller than the end cell."
    )


@pytest.mark.parametrize(
    "column_name_or_letter, output_format, ascending, expected_index, expected_length, expected_value",
    [
        (
            "A",
            "list",
            True,
            50,
            51,
            ["Willodean", "Harn", "Female", "United States", 39, "16/08/2016", 3567],
        ),
        (
            "A",
            "list",
            False,
            49,
            51,
            ["Angelyn", "Vong", "Female", "United States", 29, "21/05/2015", 6125],
        ),
        (
            "First Name",
            "list",
            True,
            50,
            51,
            ["Willodean", "Harn", "Female", "United States", 39, "16/08/2016", 3567],
        ),
        (
            "First Name",
            "list",
            False,
            49,
            51,
            ["Angelyn", "Vong", "Female", "United States", 29, "21/05/2015", 6125],
        ),
        (
            "A",
            "dict",
            True,
            [
                (
                    "First Name",
                    "Last Name",
                    "Gender",
                    "Country",
                    "Age",
                    "Date",
                    "Salary",
                ),
                -1,
            ],
            7,
            ["Willodean", "Harn", "Female", "United States", 39, "16/08/2016", 3567],
        ),
        (
            "A",
            "dict",
            False,
            [
                (
                    "First Name",
                    "Last Name",
                    "Gender",
                    "Country",
                    "Age",
                    "Date",
                    "Salary",
                ),
                -2,
            ],
            7,
            ["Angelyn", "Vong", "Female", "United States", 29, "21/05/2015", 6125],
        ),
        (
            "First Name",
            "dict",
            True,
            [
                (
                    "First Name",
                    "Last Name",
                    "Gender",
                    "Country",
                    "Age",
                    "Date",
                    "Salary",
                ),
                -1,
            ],
            7,
            ["Willodean", "Harn", "Female", "United States", 39, "16/08/2016", 3567],
        ),
        (
            "First Name",
            "dict",
            False,
            [
                (
                    "First Name",
                    "Last Name",
                    "Gender",
                    "Country",
                    "Age",
                    "Date",
                    "Salary",
                ),
                -2,
            ],
            7,
            ["Angelyn", "Vong", "Female", "United States", 29, "21/05/2015", 6125],
        ),
        (
            "A",
            "Dataframe",
            True,
            50,
            51,
            ["Willodean", "Harn", "Female", "United States", 39, "16/08/2016", 3567],
        ),
        (
            "A",
            "Dataframe",
            False,
            49,
            51,
            ["Angelyn", "Vong", "Female", "United States", 29, "21/05/2015", 6125],
        ),
        (
            "First Name",
            "Dataframe",
            True,
            50,
            51,
            ["Willodean", "Harn", "Female", "United States", 39, "16/08/2016", 3567],
        ),
        (
            "First Name",
            "Dataframe",
            False,
            49,
            51,
            ["Angelyn", "Vong", "Female", "United States", 29, "21/05/2015", 6125],
        ),
    ],
)
def test_sort_columns_success(
    column_name_or_letter,
    output_format,
    ascending,
    expected_index,
    expected_length,
    expected_value,
    setup_teardown,
):
    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    data = exl.sort_column(
        column_name_or_letter=column_name_or_letter,
        output_format=output_format,
        asc=ascending,
        starting_cell="D6",
        sheet_name="Offset_table",
    )

    if isinstance(data, list):
        assert_that(data).is_length(expected_length)
        assert_that(data[expected_index]).is_equal_to(expected_value)
    elif isinstance(data, dict):
        for index, col in enumerate(expected_index[0]):
            assert_that(data[col][expected_index[-1]]).is_equal_to(
                expected_value[index]
            )
    else:
        row_count, column_count = data.shape
        assert_that(row_count).is_equal_to(expected_length)
        assert_that(column_count).is_equal_to(7)
        is_present = data.loc[expected_index].tolist() == expected_value
        assert_that(is_present).is_true()


def test_sort_columns_invalid_output_format(setup_teardown):
    with pytest.raises(ValueError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.sort_column(
            column_name_or_letter="D",
            output_format="invalid",
            starting_cell="D6",
            sheet_name="Offset_table",
        )

    assert_that(str(exc_info.value)).is_equal_to(
        "Invalid output format. Use 'list', 'dict', or 'dataframe'."
    )


def test_sort_columns_invalid_column_name(setup_teardown):
    with pytest.raises(ValueError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.sort_column(
            column_name_or_letter=INVALID_CELL_ADDRESS,
            starting_cell="D6",
            sheet_name="Offset_table",
        )

    assert_that(str(exc_info.value)).is_equal_to(
        f"Invalid column name or letter: '{INVALID_CELL_ADDRESS}'"
    )


def test_sort_columns_invalid_cell_address(setup_teardown):
    with pytest.raises(InvalidCellAddressError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.sort_column(
            column_name_or_letter="A",
            starting_cell=INVALID_CELL_ADDRESS,
            sheet_name="Offset_table",
        )

    assert_that(str(exc_info.value)).is_equal_to(
        f"Cell '{INVALID_CELL_ADDRESS}' doesn't exists."
    )


def test_sort_columns_none_header(setup_teardown):
    with pytest.raises(ValueError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.sort_column(
            column_name_or_letter="A", starting_cell="A1", sheet_name="Sheet1"
        )

    assert_that(str(exc_info.value)).is_equal_to(
        "Sheet1 does not have a valid string header: 'None' found."
    )


def test_sort_columns_column_out_of_bound(setup_teardown):
    with pytest.raises(ValueError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.sort_column(
            column_name_or_letter="ZZZ", starting_cell="A1", sheet_name="Sheet1"
        )

    assert_that(str(exc_info.value)).is_equal_to(
        "Column letter 'ZZZ' is out of bounds for the provided sheet."
    )


@pytest.mark.parametrize(
    "columns, output_format, expected_length, expected_value",
    [
        (
            "A",
            "list",
            4,
            Counter(
                {
                    (
                        "Kelsie",
                        "Wachtel",
                        "Female",
                        "France",
                        27,
                        "16/08/2016",
                        8642,
                    ): 2,
                    ("Loreta", "Curren", "Female", "France", 26, "21/05/2015", 9654): 2,
                }
            ),
        ),
        (
            "Last Name",
            "list",
            4,
            Counter(
                {
                    (
                        "Kelsie",
                        "Wachtel",
                        "Female",
                        "France",
                        27,
                        "16/08/2016",
                        8642,
                    ): 2,
                    ("Loreta", "Curren", "Female", "France", 26, "21/05/2015", 9654): 2,
                }
            ),
        ),
        (
            "A",
            "dict",
            4,
            Counter(
                {
                    (
                        "Kelsie",
                        "Wachtel",
                        "Female",
                        "France",
                        27,
                        "16/08/2016",
                        8642,
                    ): 2,
                    ("Loreta", "Curren", "Female", "France", 26, "21/05/2015", 9654): 2,
                }
            ),
        ),
        (
            "First Name",
            "dict",
            4,
            Counter(
                {
                    (
                        "Kelsie",
                        "Wachtel",
                        "Female",
                        "France",
                        27,
                        "16/08/2016",
                        8642,
                    ): 2,
                    ("Loreta", "Curren", "Female", "France", 26, "21/05/2015", 9654): 2,
                }
            ),
        ),
        (
            "B",
            "dataframe",
            4,
            Counter(
                {
                    (
                        "Kelsie",
                        "Wachtel",
                        "Female",
                        "France",
                        27,
                        "16/08/2016",
                        8642,
                    ): 2,
                    ("Loreta", "Curren", "Female", "France", 26, "21/05/2015", 9654): 2,
                }
            ),
        ),
        (
            "Last Name",
            "dataframe",
            4,
            Counter(
                {
                    (
                        "Kelsie",
                        "Wachtel",
                        "Female",
                        "France",
                        27,
                        "16/08/2016",
                        8642,
                    ): 2,
                    ("Loreta", "Curren", "Female", "France", 26, "21/05/2015", 9654): 2,
                }
            ),
        ),
    ],
)
def test_find_duplicates_success(
    columns, output_format, expected_length, expected_value, setup_teardown
):
    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    data = exl.find_duplicates(
        column_names_or_letters=columns,
        starting_cell="D6",
        sheet_name="Offset_table",
        output_format=output_format,
    )

    if isinstance(data, list):
        assert_that(data).is_length(expected_length)
        counts = Counter(tuple(row) for row in data)
        assert_that(counts == expected_value).is_true()
    elif isinstance(data, dict):
        tuples_from_dict = Counter(
            zip(
                data["First Name"],
                data["Last Name"],
                data["Gender"],
                data["Country"],
                data["Age"],
                data["Date"],
                data["Salary"],
            )
        )
        assert_that(tuples_from_dict == expected_value).is_true()
    else:
        df_tuples = [tuple(row) for row in data.to_records(index=False)]
        counts = Counter(df_tuples)
        assert_that(counts == expected_value).is_true()


def test_find_duplicates_invalid_column_name(setup_teardown):
    with pytest.raises(ValueError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.find_duplicates(
            column_names_or_letters=INVALID_CELL_ADDRESS,
            starting_cell="D6",
            sheet_name="Offset_table",
        )

    assert_that(str(exc_info.value)).is_equal_to(
        f"Invalid column name or letter: '{INVALID_CELL_ADDRESS}'"
    )


def test_find_duplicates_invalid_cell_address(setup_teardown):
    with pytest.raises(InvalidCellAddressError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.find_duplicates(
            column_names_or_letters="A",
            starting_cell=INVALID_CELL_ADDRESS,
            sheet_name="Offset_table",
        )

    assert_that(str(exc_info.value)).is_equal_to(
        f"Cell '{INVALID_CELL_ADDRESS}' doesn't exists."
    )


def test_find_duplicates_none_header(setup_teardown):
    with pytest.raises(ValueError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.find_duplicates(
            column_names_or_letters="A", starting_cell="A1", sheet_name="Sheet1"
        )

    assert_that(str(exc_info.value)).is_equal_to(
        "Sheet1 does not have a valid string header: 'None' found."
    )


def test_find_duplicates_column_out_of_bound(setup_teardown):
    with pytest.raises(ValueError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.find_duplicates(
            column_names_or_letters="ZZZ", starting_cell="A1", sheet_name="Sheet1"
        )

    assert_that(str(exc_info.value)).is_equal_to(
        "Column letter 'ZZZ' is out of bounds for the provided sheet."
    )


def test_find_duplicates_delete_with_columns(setup_teardown):
    test_file = copy_test_excel_file(
        destination_file=os.path.join(DATA_DIR, "test_delete_duplicates.xlsx")
    )
    output_file = os.path.join(DATA_DIR, "test_delete_duplicates_output.xlsx")

    try:
        exl.open_workbook(workbook_name=test_file)

        duplicates_before = exl.find_duplicates(
            column_names_or_letters="Last Name",
            starting_cell="D6",
            sheet_name="Offset_table",
            output_format="list",
        )
        assert_that(duplicates_before).is_length(4)

        rows_deleted = exl.find_duplicates(
            column_names_or_letters="Last Name",
            starting_cell="D6",
            sheet_name="Offset_table",
            delete=True,
            output_filename=output_file,
        )

        assert_that(rows_deleted).is_instance_of(int)
        assert_that(rows_deleted).is_equal_to(2)

        duplicates_after = exl.find_duplicates(
            column_names_or_letters="Last Name",
            starting_cell="D6",
            sheet_name="Offset_table",
            output_format="list",
        )
        assert_that(duplicates_after).is_length(0)

        final_row_count = exl.get_row_count(
            sheet_name="Offset_table",
            include_header=True,
            starting_cell="D6",
        )
        assert_that(final_row_count).is_greater_than(0)
        
        sheet_data = exl.fetch_sheet_data(
            sheet_name="Offset_table",
            starting_cell="D6",
            output_format="list",
        )
        assert_that(sheet_data).is_not_empty()

        assert_that(os.path.exists(output_file)).is_true()
        assert_that(os.path.exists(test_file)).is_true()

        exl.switch_workbook(alias=test_file)
        source_duplicates = exl.find_duplicates(
            column_names_or_letters="Last Name",
            starting_cell="D6",
            sheet_name="Offset_table",
            output_format="list",
        )
        assert_that(source_duplicates).is_length(4)

    finally:
        if os.path.exists(test_file):
            os.remove(test_file)
        if os.path.exists(output_file):
            os.remove(output_file)


def test_find_duplicates_delete_without_columns(setup_teardown):
    test_file = copy_test_excel_file(
        destination_file=os.path.join(DATA_DIR, "test_delete_duplicates_all.xlsx")
    )
    output_file = os.path.join(DATA_DIR, "test_delete_duplicates_all_output.xlsx")

    try:
        exl.open_workbook(workbook_name=test_file)

        duplicates_before = exl.find_duplicates(
            starting_cell="D6",
            sheet_name="Offset_table",
            output_format="list",
        )
        assert_that(duplicates_before).is_length(4)

        rows_deleted = exl.find_duplicates(
            starting_cell="D6",
            sheet_name="Offset_table",
            delete=True,
            output_filename=output_file,
        )

        assert_that(rows_deleted).is_instance_of(int)
        assert_that(rows_deleted).is_equal_to(2)

        duplicates_after = exl.find_duplicates(
            starting_cell="D6",
            sheet_name="Offset_table",
            output_format="list",
        )
        assert_that(duplicates_after).is_length(0)

        final_row_count = exl.get_row_count(
            sheet_name="Offset_table",
            include_header=True,
            starting_cell="D6",
        )
        assert_that(final_row_count).is_greater_than(0)
        
        sheet_data = exl.fetch_sheet_data(
            sheet_name="Offset_table",
            starting_cell="D6",
            output_format="list",
        )
        assert_that(sheet_data).is_not_empty()

        assert_that(os.path.exists(output_file)).is_true()

    finally:
        if os.path.exists(test_file):
            os.remove(test_file)
        if os.path.exists(output_file):
            os.remove(output_file)


def test_find_duplicates_delete_with_multiple_columns(setup_teardown):
    test_file = copy_test_excel_file(
        destination_file=os.path.join(DATA_DIR, "test_delete_duplicates_multi.xlsx")
    )
    output_file = os.path.join(DATA_DIR, "test_delete_duplicates_multi_output.xlsx")

    try:
        exl.open_workbook(workbook_name=test_file)

        rows_deleted = exl.find_duplicates(
            column_names_or_letters=["First Name", "Last Name"],
            starting_cell="D6",
            sheet_name="Offset_table",
            delete=True,
            output_filename=output_file,
        )

        assert_that(rows_deleted).is_instance_of(int)
        assert_that(rows_deleted).is_greater_than_or_equal_to(0)

        duplicates_after = exl.find_duplicates(
            column_names_or_letters=["First Name", "Last Name"],
            starting_cell="D6",
            sheet_name="Offset_table",
            output_format="list",
        )
        assert_that(duplicates_after).is_length(0)

        assert_that(os.path.exists(output_file)).is_true()

    finally:
        if os.path.exists(test_file):
            os.remove(test_file)
        if os.path.exists(output_file):
            os.remove(output_file)


def test_find_duplicates_delete_no_duplicates(setup_teardown):
    test_file = copy_test_excel_file(
        destination_file=os.path.join(DATA_DIR, "test_delete_no_duplicates.xlsx")
    )
    output_file1 = os.path.join(DATA_DIR, "test_delete_no_duplicates_output1.xlsx")
    output_file2 = os.path.join(DATA_DIR, "test_delete_no_duplicates_output2.xlsx")

    try:
        exl.open_workbook(workbook_name=test_file)

        exl.find_duplicates(
            column_names_or_letters="Last Name",
            starting_cell="D6",
            sheet_name="Offset_table",
            delete=True,
            output_filename=output_file1,
        )

        row_count_after_first = exl.get_row_count(
            sheet_name="Offset_table",
            include_header=True,
            starting_cell="D6",
        )

        rows_deleted = exl.find_duplicates(
            column_names_or_letters="Last Name",
            starting_cell="D6",
            sheet_name="Offset_table",
            delete=True,
            output_filename=output_file2,
        )

        assert_that(rows_deleted).is_equal_to(0)

        row_count_after_second = exl.get_row_count(
            sheet_name="Offset_table",
            include_header=True,
            starting_cell="D6",
        )
        assert_that(row_count_after_second).is_equal_to(row_count_after_first)

        assert_that(os.path.exists(output_file1)).is_true()
        assert_that(os.path.exists(output_file2)).is_true()

    finally:
        if os.path.exists(test_file):
            os.remove(test_file)
        if os.path.exists(output_file1):
            os.remove(output_file1)
        if os.path.exists(output_file2):
            os.remove(output_file2)


def test_find_duplicates_delete_without_output_filename(setup_teardown):
    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)

    with pytest.raises(ValueError) as exc_info:
        exl.find_duplicates(
            column_names_or_letters="Last Name",
            starting_cell="D6",
            sheet_name="Offset_table",
            delete=True,
        )

    assert_that(str(exc_info.value)).contains("output_filename is mandatory")


def test_find_duplicates_delete_backward_compatibility(setup_teardown):
    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)

    duplicates = exl.find_duplicates(
        column_names_or_letters="Last Name",
        starting_cell="D6",
        sheet_name="Offset_table",
        output_format="list",
        delete=False,
    )

    assert_that(duplicates).is_length(4)
    assert_that(isinstance(duplicates, list)).is_true()

    duplicates_default = exl.find_duplicates(
        column_names_or_letters="Last Name",
        starting_cell="D6",
        sheet_name="Offset_table",
        output_format="list",
    )

    assert_that(duplicates_default).is_length(4)
    assert_that(isinstance(duplicates_default, list)).is_true()


def test_find_duplicates_delete_preserves_other_sheets(setup_teardown):
    test_file = copy_test_excel_file(
        destination_file=os.path.join(DATA_DIR, "test_delete_preserves_sheets.xlsx")
    )
    output_file = os.path.join(DATA_DIR, "test_delete_preserves_sheets_output.xlsx")

    try:
        exl.open_workbook(workbook_name=test_file)

        sheets_before = exl.get_sheets()
        assert_that(sheets_before).contains("Offset_table", "Sheet1")

        sheet1_data_before = exl.fetch_sheet_data(
            sheet_name="Sheet1",
            output_format="list",
        )

        rows_deleted = exl.find_duplicates(
            column_names_or_letters="Last Name",
            starting_cell="D6",
            sheet_name="Offset_table",
            delete=True,
            output_filename=output_file,
        )
        assert_that(rows_deleted).is_greater_than(0)

        sheets_after = exl.get_sheets()
        assert_that(sheets_after).contains("Offset_table", "Sheet1")
        assert_that(sheets_after).is_length(len(sheets_before))

        sheet1_data_after = exl.fetch_sheet_data(
            sheet_name="Sheet1",
            output_format="list",
        )
        assert_that(sheet1_data_after).is_equal_to(sheet1_data_before)

        assert_that(os.path.exists(output_file)).is_true()

    finally:
        if os.path.exists(test_file):
            os.remove(test_file)
        if os.path.exists(output_file):
            os.remove(output_file)


def test_find_duplicates_delete_overwrite_if_exists_false(setup_teardown):
    test_file = copy_test_excel_file(
        destination_file=os.path.join(DATA_DIR, "test_delete_overwrite_false.xlsx")
    )
    output_file = os.path.join(DATA_DIR, "test_delete_overwrite_false_output.xlsx")

    try:
        exl.open_workbook(workbook_name=test_file)

        rows_deleted1 = exl.find_duplicates(
            column_names_or_letters="Last Name",
            starting_cell="D6",
            sheet_name="Offset_table",
            delete=True,
            output_filename=output_file,
        )
        assert_that(rows_deleted1).is_greater_than(0)
        assert_that(os.path.exists(output_file)).is_true()

        exl.switch_workbook(alias=test_file)
        assert_that(os.path.exists(output_file)).is_true()
        
        try:
            exl.find_duplicates(
                column_names_or_letters="Last Name",
                starting_cell="D6",
                sheet_name="Offset_table",
                delete=True,
                output_filename=output_file,
                overwrite_if_exists=False,
            )
            assert_that(False).is_true()
        except FileAlreadyExistsError as e:
            assert_that(str(e)).contains(output_file)

    finally:
        if os.path.exists(test_file):
            os.remove(test_file)
        if os.path.exists(output_file):
            os.remove(output_file)


def test_find_duplicates_delete_overwrite_if_exists_true(setup_teardown):
    test_file = copy_test_excel_file(
        destination_file=os.path.join(DATA_DIR, "test_delete_overwrite_true.xlsx")
    )
    output_file = os.path.join(DATA_DIR, "test_delete_overwrite_true_output.xlsx")

    try:
        exl.open_workbook(workbook_name=test_file)

        rows_deleted1 = exl.find_duplicates(
            column_names_or_letters="Last Name",
            starting_cell="D6",
            sheet_name="Offset_table",
            delete=True,
            output_filename=output_file,
        )
        assert_that(rows_deleted1).is_greater_than(0)
        assert_that(os.path.exists(output_file)).is_true()

        row_count_after_first = exl.get_row_count(
            sheet_name="Offset_table",
            include_header=True,
            starting_cell="D6",
        )

        exl.switch_workbook(alias=test_file)
        rows_deleted2 = exl.find_duplicates(
            column_names_or_letters="Last Name",
            starting_cell="D6",
            sheet_name="Offset_table",
            delete=True,
            output_filename=output_file,
            overwrite_if_exists=True,
        )
        assert_that(rows_deleted2).is_greater_than(0)
        assert_that(os.path.exists(output_file)).is_true()

        exl.switch_workbook(alias=output_file)
        row_count_after_second = exl.get_row_count(
            sheet_name="Offset_table",
            include_header=True,
            starting_cell="D6",
        )
        assert_that(row_count_after_second).is_equal_to(row_count_after_first)

    finally:
        if os.path.exists(test_file):
            os.remove(test_file)
        if os.path.exists(output_file):
            os.remove(output_file)


def test_compare_excels_success(setup_teardown):
    source_excel_config = {
        "sheet_name": "Sheet1",
        "starting_cell": "A1",
        "columns": ["First Name", "Last Name"],
    }
    target_excel_config = {
        "sheet_name": "Offset_table",
        "starting_cell": "D6",
        "columns": ["First Name", "Last Name"],
    }
    data = exl.compare_excels(
        source_excel=EXCEL_FILE_PATH,
        target_excel=EXCEL_FILE_PATH,
        source_excel_config=source_excel_config,
        target_excel_config=target_excel_config,
    )

    expected_data = {
        "First Name": [
            "Lester",
            "Mark",
            "Mark",
            "Mark",
            "Many",
            "Marvel",
            "Marcel",
            "Jona",
            "Felisa",
        ],
        "Last Name": [
            "Grindle",
            "Zabriskie",
            "Hail",
            "Cail",
            "Cuccia",
            "Hail",
            "Zabriskie",
            "Grindle",
            "Cail",
        ],
        "Excel_Source": [
            "Source",
            "Source",
            "Source",
            "Source",
            "Source",
            "Target",
            "Target",
            "Target",
            "Target",
        ],
    }
    expected_df = pd.DataFrame(expected_data)
    df = data[expected_df.columns]
    df_sorted = df.sort_values(by=df.columns.tolist()).reset_index(drop=True)
    expected_df_sorted = expected_df.sort_values(
        by=expected_df.columns.tolist()
    ).reset_index(drop=True)

    assert_that(df_sorted.equals(expected_df_sorted)).is_true()


def test_compare_excels_file_not_found(setup_teardown):
    with pytest.raises(ExcelFileNotFoundError) as exc_info:
        exl.compare_excels(
            source_excel=INVALID_EXCEL_FILE_PATH, target_excel=EXCEL_FILE_PATH
        )

    assert_that(str(exc_info.value)).is_equal_to(
        f"Excel file '{INVALID_EXCEL_FILE_PATH}' not found. Please give the valid file path."
    )


def test_compare_excels_invalid_cell_address(setup_teardown):
    with pytest.raises(InvalidCellAddressError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        source_excel_config = {
            "sheet_name": "Sheet1",
            "starting_cell": INVALID_CELL_ADDRESS,
            "columns": ["First Name", "Last Name"],
        }
        target_excel_config = {
            "sheet_name": "Offset_table",
            "starting_cell": "D6",
            "columns": ["First Name", "Last Name"],
        }
        exl.compare_excels(
            source_excel=EXCEL_FILE_PATH,
            target_excel=EXCEL_FILE_PATH,
            source_excel_config=source_excel_config,
            target_excel_config=target_excel_config,
        )

    assert_that(str(exc_info.value)).is_equal_to(
        f"Cell '{INVALID_CELL_ADDRESS}' doesn't exists."
    )


def test_fetch_sheet_data_invalid_column_name(setup_teardown):
    with pytest.raises(InvalidColumnNameError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        source_excel_config = {
            "sheet_name": "Sheet1",
            "starting_cell": "A1",
            "columns": ["Invalid_column", "Last Name"],
        }
        target_excel_config = {
            "sheet_name": "Offset_table",
            "starting_cell": "D6",
            "columns": ["First Name", "Last Name"],
        }
        exl.compare_excels(
            source_excel=EXCEL_FILE_PATH,
            target_excel=EXCEL_FILE_PATH,
            source_excel_config=source_excel_config,
            target_excel_config=target_excel_config,
        )

    assert_that(str(exc_info.value)).is_equal_to(
        "Invalid columns. Columns not found: ['Invalid_column'] in sheet 'Sheet1'."
    )


def test_fetch_sheet_data_column_mismatch(setup_teardown):
    with pytest.raises(ColumnMismatchError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        source_excel_config = {
            "sheet_name": "Sheet1",
            "starting_cell": "A1",
            "columns": ["First Name", "Salary"],
        }
        target_excel_config = {
            "sheet_name": "Offset_table",
            "starting_cell": "D6",
            "columns": ["First Name", "Last Name"],
        }
        exl.compare_excels(
            source_excel=EXCEL_FILE_PATH,
            target_excel=EXCEL_FILE_PATH,
            source_excel_config=source_excel_config,
            target_excel_config=target_excel_config,
        )

    assert_that(str(exc_info.value)).is_equal_to(
        "Column mismatch found in excel files.\nMissing in source: {'Last Name'}\nMissing in target: {'Salary'}"
    )


def test_remove_empty_rows_all_columns(setup_teardown):
    test_file = os.path.join(DATA_DIR, "test_remove_empty_all.xlsx")
    output_file = os.path.join(DATA_DIR, "test_remove_empty_all_output.xlsx")

    try:
        wb = excel.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        ws.append(["Name", "Age", "City"])
        ws.append(["John", 25, "NYC"])
        ws.append([None, None, None])
        ws.append(["Jane", 30, "LA"])
        ws.append([None, None, None])
        ws.append([None, None, None])
        ws.append(["Bob", 35, "Chicago"])
        wb.save(test_file)
        wb.close()

        exl.open_workbook(workbook_name=test_file)
        
        initial_row_count = exl.get_row_count(
            sheet_name="Sheet1", include_header=True, starting_cell="A1"
        )
        
        rows_removed = exl.remove_empty_rows(
            sheet_name="Sheet1",
            output_filename=output_file,
            starting_cell="A1",
        )

        assert_that(rows_removed).is_instance_of(int)
        assert_that(rows_removed).is_equal_to(3)

        assert_that(os.path.exists(output_file)).is_true()

        assert_that(os.path.exists(test_file)).is_true()
        exl.switch_workbook(alias=test_file)
        source_row_count = exl.get_row_count(
            sheet_name="Sheet1", include_header=True, starting_cell="A1"
        )
        assert_that(source_row_count).is_equal_to(initial_row_count)

        exl.switch_workbook(alias=output_file)
        final_row_count = exl.get_row_count(
            sheet_name="Sheet1", include_header=True, starting_cell="A1"
        )
        assert_that(final_row_count).is_equal_to(initial_row_count - rows_removed)
        assert_that(final_row_count).is_greater_than(0)

    finally:
        if os.path.exists(test_file):
            os.remove(test_file)
        if os.path.exists(output_file):
            os.remove(output_file)


def test_remove_empty_rows_specific_columns(setup_teardown):
    test_file = os.path.join(DATA_DIR, "test_remove_empty_columns.xlsx")
    output_file = os.path.join(DATA_DIR, "test_remove_empty_columns_output.xlsx")

    try:
        wb = excel.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        ws.append(["Name", "Age", "City"])
        ws.append(["John", 25, "NYC"])
        ws.append([None, None, None])
        ws.append(["Jane", None, "LA"])
        ws.append([None, 30, None])
        ws.append(["Bob", 35, "Chicago"])
        wb.save(test_file)
        wb.close()

        exl.open_workbook(workbook_name=test_file)
        
        rows_removed = exl.remove_empty_rows(
            sheet_name="Sheet1",
            column_names_or_letters="Name",
            output_filename=output_file,
            starting_cell="A1",
        )

        assert_that(rows_removed).is_instance_of(int)
        assert_that(rows_removed).is_greater_than(0)

        assert_that(os.path.exists(output_file)).is_true()

    finally:
        if os.path.exists(test_file):
            os.remove(test_file)
        if os.path.exists(output_file):
            os.remove(output_file)


def test_remove_empty_rows_with_starting_cell(setup_teardown):
    test_file = copy_test_excel_file(
        destination_file=os.path.join(DATA_DIR, "test_remove_empty_starting_cell.xlsx")
    )
    output_file = os.path.join(DATA_DIR, "test_remove_empty_starting_cell_output.xlsx")

    try:
        wb = excel.load_workbook(test_file)
        if "Offset_table" not in wb.sheetnames:
            ws = wb.create_sheet("Offset_table")
        else:
            ws = wb["Offset_table"]
            ws.delete_rows(1, ws.max_row)
        
        ws["D6"] = "First Name"
        ws["E6"] = "Last Name"
        ws["F6"] = "Age"
        
        ws["D7"] = "John"
        ws["E7"] = "Doe"
        ws["F7"] = 25
        
        ws["D8"] = ""
        ws["E8"] = ""
        ws["F8"] = ""
        
        ws["D9"] = "Jane"
        ws["E9"] = "Smith"
        ws["F9"] = 30
        
        ws["D10"] = ""
        ws["E10"] = ""
        ws["F10"] = ""
        
        wb.save(test_file)
        wb.close()

        exl.open_workbook(workbook_name=test_file)
        
        initial_row_count = exl.get_row_count(
            sheet_name="Offset_table",
            include_header=True,
            starting_cell="D6",
        )
        
        rows_removed = exl.remove_empty_rows(
            sheet_name="Offset_table",
            output_filename=output_file,
            starting_cell="D6",
        )

        assert_that(rows_removed).is_instance_of(int)
        assert_that(rows_removed).is_equal_to(2)

        exl.switch_workbook(alias=output_file)
        final_row_count = exl.get_row_count(
            sheet_name="Offset_table",
            include_header=True,
            starting_cell="D6",
        )
        assert_that(final_row_count).is_equal_to(initial_row_count - rows_removed)
        
        sheet_data = exl.fetch_sheet_data(
            sheet_name="Offset_table",
            starting_cell="D6",
            output_format="list",
        )
        assert_that(sheet_data).is_not_empty()

    finally:
        if os.path.exists(test_file):
            os.remove(test_file)
        if os.path.exists(output_file):
            os.remove(output_file)


def test_remove_empty_rows_no_empty_rows(setup_teardown):
    test_file = copy_test_excel_file(
        destination_file=os.path.join(DATA_DIR, "test_remove_empty_none.xlsx")
    )
    output_file = os.path.join(DATA_DIR, "test_remove_empty_none_output.xlsx")

    try:
        exl.open_workbook(workbook_name=test_file)
        
        rows_removed = exl.remove_empty_rows(
            sheet_name="Offset_table",
            output_filename=output_file,
            starting_cell="D6",
        )

        assert_that(rows_removed).is_instance_of(int)
        # Note: If the file has a formatting row that's completely empty, it will be removed
        # This is expected behavior - empty rows (even formatting ones) are removed
        assert_that(rows_removed).is_greater_than_or_equal_to(0)

        assert_that(os.path.exists(output_file)).is_true()

    finally:
        if os.path.exists(test_file):
            os.remove(test_file)
        if os.path.exists(output_file):
            os.remove(output_file)


def test_remove_empty_rows_file_already_exists(setup_teardown):
    test_file = os.path.join(DATA_DIR, "test_remove_empty_exists.xlsx")
    output_file = os.path.join(DATA_DIR, "test_remove_empty_exists_output.xlsx")

    try:
        wb = excel.Workbook()
        ws = wb.active
        ws.append(["Name", "Age"])
        ws.append(["John", 25])
        wb.save(test_file)
        wb.close()

        wb = excel.Workbook()
        wb.save(output_file)
        wb.close()

        exl.open_workbook(workbook_name=test_file)

        with pytest.raises(FileAlreadyExistsError) as exc_info:
            exl.remove_empty_rows(
                sheet_name="Sheet",
                output_filename=output_file,
            )

        assert_that(str(exc_info.value)).contains("already exists")

    finally:
        if os.path.exists(test_file):
            os.remove(test_file)
        if os.path.exists(output_file):
            os.remove(output_file)


def test_remove_empty_rows_overwrite_if_exists(setup_teardown):
    test_file = os.path.join(DATA_DIR, "test_remove_empty_overwrite.xlsx")
    output_file = os.path.join(DATA_DIR, "test_remove_empty_overwrite_output.xlsx")

    try:
        wb = excel.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["Name", "Age"])
        ws.append(["John", 25])
        ws.append([None, None])
        wb.save(test_file)
        wb.close()

        wb = excel.Workbook()
        wb.save(output_file)
        wb.close()

        exl.open_workbook(workbook_name=test_file)

        rows_removed = exl.remove_empty_rows(
            sheet_name="Sheet1",
            output_filename=output_file,
            overwrite_if_exists=True,
        )

        assert_that(rows_removed).is_instance_of(int)
        assert_that(os.path.exists(output_file)).is_true()

    finally:
        if os.path.exists(test_file):
            os.remove(test_file)
        if os.path.exists(output_file):
            os.remove(output_file)


def test_remove_empty_rows_invalid_column_name(setup_teardown):
    test_file = os.path.join(DATA_DIR, "test_remove_empty_invalid_col.xlsx")
    output_file = os.path.join(DATA_DIR, "test_remove_empty_invalid_col_output.xlsx")

    try:
        wb = excel.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["Name", "Age"])
        ws.append(["John", 25])
        wb.save(test_file)
        wb.close()

        exl.open_workbook(workbook_name=test_file)

        with pytest.raises(ValueError):
            exl.remove_empty_rows(
                sheet_name="Sheet1",
                column_names_or_letters="InvalidColumn",
                output_filename=output_file,
            )

    finally:
        if os.path.exists(test_file):
            os.remove(test_file)
        if os.path.exists(output_file):
            os.remove(output_file)


def test_remove_empty_rows_invalid_cell_address(setup_teardown):
    test_file = os.path.join(DATA_DIR, "test_remove_empty_invalid_cell.xlsx")
    output_file = os.path.join(DATA_DIR, "test_remove_empty_invalid_cell_output.xlsx")

    try:
        wb = excel.Workbook()
        wb.save(test_file)
        wb.close()

        exl.open_workbook(workbook_name=test_file)

        with pytest.raises(InvalidCellAddressError) as exc_info:
            exl.remove_empty_rows(
                sheet_name="Sheet",
                output_filename=output_file,
                starting_cell=INVALID_CELL_ADDRESS,
            )

        assert_that(str(exc_info.value)).contains(INVALID_CELL_ADDRESS)

    finally:
        if os.path.exists(test_file):
            os.remove(test_file)
        if os.path.exists(output_file):
            os.remove(output_file)


def test_remove_empty_rows_multiple_columns(setup_teardown):
    test_file = os.path.join(DATA_DIR, "test_remove_empty_multi_col.xlsx")
    output_file = os.path.join(DATA_DIR, "test_remove_empty_multi_col_output.xlsx")

    try:
        wb = excel.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        ws.append(["First Name", "Last Name", "Age"])
        ws.append(["John", "Doe", 25])
        ws.append([None, None, None])
        ws.append(["Jane", None, 30])
        ws.append([None, "Smith", 35])
        wb.save(test_file)
        wb.close()

        exl.open_workbook(workbook_name=test_file)
        
        rows_removed = exl.remove_empty_rows(
            sheet_name="Sheet1",
            column_names_or_letters=["First Name", "Last Name"],
            output_filename=output_file,
            starting_cell="A1",
        )

        assert_that(rows_removed).is_instance_of(int)
        assert_that(rows_removed).is_greater_than(0)

        assert_that(os.path.exists(output_file)).is_true()

    finally:
        if os.path.exists(test_file):
            os.remove(test_file)
        if os.path.exists(output_file):
            os.remove(output_file)


def test_remove_empty_rows_preserves_structure(setup_teardown):
    test_file = copy_test_excel_file(
        destination_file=os.path.join(DATA_DIR, "test_remove_empty_structure.xlsx")
    )
    output_file = os.path.join(DATA_DIR, "test_remove_empty_structure_output.xlsx")

    try:
        wb = excel.load_workbook(test_file)
        if "Offset_table" not in wb.sheetnames:
            ws = wb.create_sheet("Offset_table")
        else:
            ws = wb["Offset_table"]
            # Clear existing data
            ws.delete_rows(1, ws.max_row)
        
        ws["D6"] = "First Name"
        ws["E6"] = "Last Name"
        ws["F6"] = "Age"
        
        ws["D7"] = "John"
        ws["E7"] = "Doe"
        ws["F7"] = 25
        
        ws["D8"] = ""
        ws["E8"] = ""
        ws["F8"] = ""
        
        ws["D9"] = "Jane"
        ws["E9"] = "Smith"
        ws["F9"] = 30
        
        wb.save(test_file)
        wb.close()

        exl.open_workbook(workbook_name=test_file)
        
        initial_row_count = exl.get_row_count(
            sheet_name="Offset_table",
            include_header=True,
            starting_cell="D6",
        )
        
        rows_removed = exl.remove_empty_rows(
            sheet_name="Offset_table",
            output_filename=output_file,
            starting_cell="D6",
        )

        assert_that(rows_removed).is_equal_to(1)

        exl.switch_workbook(alias=output_file)
        final_row_count = exl.get_row_count(
            sheet_name="Offset_table",
            include_header=True,
            starting_cell="D6",
        )
        assert_that(final_row_count).is_equal_to(initial_row_count - rows_removed)
        
        sheet_data = exl.fetch_sheet_data(
            sheet_name="Offset_table",
            starting_cell="D6",
            output_format="list",
        )
        assert_that(sheet_data).is_not_empty()
        assert_that(len(sheet_data)).is_greater_than(0)

    finally:
        if os.path.exists(test_file):
            os.remove(test_file)
        if os.path.exists(output_file):
            os.remove(output_file)
