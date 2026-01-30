import re
import os
import warnings
from robot.api import logger
from robot.api.deco import keyword, not_keyword
import pandas as pd
from pandas import DataFrame
from pathlib import Path
import openpyxl as excel
from openpyxl import Workbook
from openpyxl.workbook.protection import WorkbookProtection
from openpyxl.worksheet.protection import SheetProtection
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string, range_boundaries
from typing import Any, Dict, List, Optional, Tuple, Union

__version__ = "1.2.0"


class ExcelError(Exception):
    """Base exception class for all Excel-related errors."""

    def __init__(self, message: str):
        self.message = message
        super().__init__(self.message)


class WorkbookNotProtectedError(ExcelError):
    def __init__(
        self,
        message: str = "The workbook is not currently protected and cannot be unprotected.",
    ):
        super().__init__(message)


class WorkbookAlreadyProtectedError(ExcelError):
    def __init__(
        self,
        message: str = "The workbook is already protected and cannot be protected be again.",
    ):
        super().__init__(message)


class ColumnMismatchError(ExcelError):
    def __init__(self, message: str):
        super().__init__(message)


class InvalidCellRangeError(ExcelError):
    def __init__(self, message: str):
        super().__init__(message)


class WorkbookNotOpenError(ExcelError):
    def __init__(
        self, message: str = "Workbook isn't open. Please open the workbook first."
    ):
        super().__init__(message)


class InvalidColumnNameError(ExcelError):
    def __init__(self, sheet: str, columns: List[str]):
        self.columns = columns
        self.sheet = sheet
        message = f"Invalid columns. Columns not found: {self.columns} in sheet '{self.sheet}'."
        super().__init__(message)


class InvalidColumnIndexError(ExcelError):
    def __init__(self, col_index: int):
        self.col_index = col_index
        message = f"Column index {col_index} is invalid or out of bounds. The valid range is 1 to 16384."
        super().__init__(message)


class InvalidRowIndexError(ExcelError):
    def __init__(self, row_index: int):
        self.row_index = row_index
        message = f"Row index {row_index} is invalid or out of bounds. The valid range is 1 to 1048576."
        super().__init__(message)


class FileAlreadyExistsError(ExcelError):
    def __init__(self, file_name: str):
        self.file_name = file_name
        message = f"Unable to create workbook. The file '{self.file_name}' already exists. Set 'overwrite_if_exists=True' to overwrite the existing file."
        super().__init__(message)


class InvalidColorError(ExcelError):
    def __init__(self, color_type: str, color: str):
        self.color = color
        self.color_type = color_type
        message = f"Invalid {color_type} color: '{self.color}'. Use valid hex color in #RRGGBB format."
        super().__init__(message)


class InvalidBorderStyleError(ExcelError):
    def __init__(self, border_style: str, allowed_styles: List[str]):
        self.border_style = border_style
        self.allowed_styles = allowed_styles
        message = f"Invalid border style: '{self.border_style}'. Allowed values are {self.allowed_styles}."
        super().__init__(message)


class InvalidAlignmentError(ExcelError):
    def __init__(
        self, alignment_type: str, alignment_value: str, allowed_values: List[str]
    ):
        self.alignment_type = alignment_type
        self.alignment_value = alignment_value
        self.allowed_values = allowed_values
        message = f"Invalid {self.alignment_type} alignment: '{self.alignment_value}'. Allowed values are {self.allowed_values}."
        super().__init__(message)


class InvalidSheetNameError(ExcelError):
    def __init__(self, sheet_name: str):
        self.sheet_name = sheet_name
        message = f"The sheet name '{self.sheet_name}' is invalid."
        super().__init__(message)


class SheetAlreadyProtectedError(ExcelError):
    def __init__(self, sheet_name: str):
        self.sheet_name = sheet_name
        message = f"The sheet '{self.sheet_name}' is already protected and cannot be protected be again."
        super().__init__(message)


class SheetNotProtectedError(ExcelError):
    def __init__(self, sheet_name: str):
        self.sheet_name = sheet_name
        message = f"The sheet '{self.sheet_name}' is not currently protected and cannot be unprotected."
        super().__init__(message)


class ExcelFileNotFoundError(ExcelError):
    def __init__(self, file_name: str):
        self.file_name = file_name
        message = (
            f"Excel file '{file_name}' not found. Please give the valid file path."
        )
        super().__init__(message)


class SheetAlreadyExistsError(ExcelError):
    def __init__(self, sheet_name: str):
        self.sheet_name = sheet_name
        message = f"Sheet '{self.sheet_name}' already exists."
        super().__init__(message)


class SheetDoesntExistsError(ExcelError):
    def __init__(self, sheet_name: str):
        self.sheet_name = sheet_name
        message = f"Sheet '{self.sheet_name}' doesn't exists."
        super().__init__(message)


class InvalidCellAddressError(ExcelError):
    def __init__(self, cell_name: str):
        self.cell_name = cell_name
        message = f"Cell '{self.cell_name}' doesn't exists."
        super().__init__(message)


class InvalidSheetPositionError(ExcelError):
    def __init__(self, position: int, max_position: int):
        self.position = position
        self.max_position = max_position
        message = f"Invalid sheet position: {self.position}. Maximum allowed is {self.max_position}."
        super().__init__(message)


class ExcelSage:
    """
    ExcelSage is a robust and user-friendly tool designed to streamline and enhance Excel file operations using Python.
    It provides a comprehensive set of functions for managing workbooks, manipulating sheets, and handling data efficiently within Excel files.
    With built-in validation, exception handling, and logging, the library ensures smooth interactions with Excel documents while maintaining high levels of reliability.

    Key features include:
    - *Workbook Management*: Open, save, and close workbooks with ease.
    - *Sheet Handling*: Create, delete, rename, and set active sheets with detailed error checking to prevent common issues.
    - *Data Retrieval*: Fetch data from sheets in various formats such as lists, dictionaries, or pandas DataFrames, making data analysis simple and efficient.
    - *Cell Operations*: Access and manipulate individual cell values, with built-in error handling for invalid cell references.
    - *Formatting and Customization*: Modify cell properties like fonts, alignment, borders, and more to match your desired styling requirements.

    Whether you're working on automating Excel reports, processing large datasets, or simply interacting with Excel files programmatically, this library provides a flexible and intuitive interface to get the job done.
    """

    ROBOT_LIBRARY_SCOPE = "GLOBAL"
    ROBOT_LIBRARY_VERSION = __version__
    VALID_BORDER_STYLES = [
        "dashDot",
        "dashDotDot",
        "dashed",
        "dotted",
        "double",
        "hair",
        "medium",
        "mediumDashDot",
        "mediumDashDotDot",
        "mediumDashed",
        "slantDashDot",
        "thick",
        "thin",
    ]

    VALID_HORIZONTAL_ALIGNMENTS = ["left", "center", "right"]
    VALID_VERTICAL_ALIGNMENTS = ["top", "center", "bottom"]

    def __init__(self) -> None:
        self.workbooks = {}
        self.active_workbook_alias = None
        self.active_sheet = None

    @not_keyword
    def __get_active_workbook(self) -> Workbook:
        """Helper method to get the currently active workbook."""
        if self.active_workbook_alias is None:
            raise WorkbookNotOpenError()
        if self.active_workbook_alias not in self.workbooks:
            raise WorkbookNotOpenError(
                f"Active workbook alias '{self.active_workbook_alias}' not found in open workbooks."
            )
        return self.workbooks[self.active_workbook_alias]["workbook"]

    @not_keyword
    def __get_active_workbook_name(self) -> str:
        """Helper method to get the currently active workbook name."""
        if self.active_workbook_alias is None:
            raise WorkbookNotOpenError()
        if self.active_workbook_alias not in self.workbooks:
            raise WorkbookNotOpenError(
                f"Active workbook alias '{self.active_workbook_alias}' not found in open workbooks."
            )
        return self.workbooks[self.active_workbook_alias]["name"]

    @not_keyword
    def __get_active_sheet_name(self, sheet_name: Optional[str] = None) -> str:
        """Helper method to get the currently active sheet name."""
        active_workbook = self.__get_active_workbook()

        self.__argument_type_checker({"sheet_name": [sheet_name, str, None]})

        if sheet_name is None:
            if self.active_sheet is None:
                if not active_workbook.sheetnames:
                    raise Exception("No sheets found in the workbook.")
                sheet_name = active_workbook.sheetnames[0]
            else:
                sheet_name = self.active_sheet.title

        if sheet_name not in active_workbook.sheetnames:
            raise SheetDoesntExistsError(sheet_name)
        return sheet_name

    @not_keyword
    def __argument_type_checker(self, arg_list: Dict[str, List[Any]]) -> None:
        """Helper method to check the type of the arguments."""
        for arg_name, value in arg_list.items():
            if isinstance(value[1], tuple):
                expected_type_names = "', or '".join(t.__name__ for t in value[1])
            else:
                expected_type_names = value[1].__name__

            if len(value) == 3:
                if value[0] is not None and not isinstance(value[0], value[1]):
                    raise TypeError(
                        f"'{arg_name}' must be a '{expected_type_names}', got '{type(value[0]).__name__}'"
                    )
            else:
                if not isinstance(value[0], value[1]):
                    raise TypeError(
                        f"'{arg_name}' must be a '{expected_type_names}', got '{type(value[0]).__name__}'"
                    )

    @keyword
    def open_workbook(
        self, workbook_name: str, alias: Optional[str] = None, **kwargs
    ) -> Workbook:
        """
        The ``Open Workbook`` keyword opens an Excel file by its name, checks if the file exists, and raises an ``ExcelFileNotFoundError`` if it doesn't. It uses openpyxl's ``load_workbook`` to load the workbook, allowing additional options via ``**kwargs``. Once the workbook is opened, it is stored with an optional alias and set as the active workbook if it's the first one opened. The keyword returns the loaded workbook object for further use.

        If no alias is provided, the workbook name (file path) will be used as the alias. If an alias is provided and already exists, it will raise a ``SheetAlreadyExistsError``.

        *Examples*
        | ***** Settings *****
        | Library    ExcelSage
        |
        | ***** Test Cases *****
        | Example
        |   Open Workbook     workbook_name=\\path\\to\\excel\\file.xlsx
        |   Open Workbook     workbook_name=\\path\\to\\excel\\file.xlsx     alias=source
        |   Open Workbook     workbook_name=\\path\\to\\excel\\file2.xlsx     alias=target
        |   Open Workbook     workbook_name=\\path\\to\\excel\\file.xlsx     read_only=False     keep_vba=True     rich_text=False
        """
        if not os.path.exists(workbook_name):
            raise ExcelFileNotFoundError(workbook_name)
        self.__argument_type_checker({"workbook_name": [workbook_name, str]})

        if alias is None:
            alias = workbook_name

        if alias in self.workbooks:
            raise SheetAlreadyExistsError(
                f"A workbook with alias '{alias}' is already open. Use a different alias or close it first."
            )

        workbook = excel.load_workbook(filename=workbook_name, **kwargs)

        self.workbooks[alias] = {"workbook": workbook, "name": workbook_name}

        if self.active_workbook_alias is None:
            self.active_workbook_alias = alias

        logger.info(
            f"Workbook '{workbook_name}' opened successfully with alias '{alias}'!"
        )
        return workbook

    @keyword
    def create_workbook(
        self,
        workbook_name: str,
        overwrite_if_exists: bool = False,
        sheet_data: List[List[Any]] = None,
        alias: Optional[str] = None,
    ) -> Workbook:
        """
        The ``Create Workbook`` keyword creates a new Excel workbook with the option to write data into the first sheet during the creation process. It also includes an option to overwrite the file if needed.

        *Examples*
        | ***** Settings *****
        | Library    ExcelSage
        |
        | ***** Keywords *****
        | Prepare Sheet Data
        |   ${row_header}    Create List    Name    Age   Salary
        |   ${row1}    Create List    John    30   1,000
        |   ${row2}    Create List    Mark    25   2,000
        |   ${row3}    Create List    Sam    56   3,000
        |   ${sheet_data}    Create List    ${row_header}   ${row1}    ${row2}    ${row3}
        |   RETURN  ${sheet_data}
        |
        | ***** Test Cases *****
        | Example
        |   ${data}     Prepare Sheet Data
        |   Create Workbook     workbook_name=\\path\\to\\excel\\file.xlsx   overwrite_if_exists=True    sheet_data=${data}
        |   ${all_sheets}   Get Sheets
        |   Rename Sheet    old_name=${all_sheets}[0]    new_name=NewSheet
        |   Close Workbook
        """
        self.__argument_type_checker({"workbook_name": [workbook_name, str]})

        if not overwrite_if_exists and os.path.exists(workbook_name):
            raise FileAlreadyExistsError(workbook_name)

        workbook = Workbook()
        sheet = workbook.active

        if sheet_data:
            for index, row in enumerate(sheet_data):
                if not isinstance(row, list):
                    raise TypeError(
                        f"Invalid row at index {index} of type '{type(row).__name__}'. Each row in 'sheet_data' must be a list."
                    )
                sheet.append(row)

        workbook.save(workbook_name)
        workbook.close()

        if alias is None:
            alias = workbook_name

        if alias in self.workbooks:
            raise SheetAlreadyExistsError(
                f"A workbook with alias '{alias}' is already open. Use a different alias or close it first."
            )

        loaded_workbook = excel.load_workbook(filename=workbook_name)

        self.workbooks[alias] = {"workbook": loaded_workbook, "name": workbook_name}

        if self.active_workbook_alias is None:
            self.active_workbook_alias = alias

        logger.info(
            f"Workbook '{workbook_name}' created and opened with alias '{alias}' for further use."
        )
        return loaded_workbook

    @keyword
    def get_sheets(self) -> List[str]:
        """
        The ``Get Sheets`` keyword returns a list of all sheet names in the active workbook. It  raises a ``WorkbookNotOpenError`` if no workbook is currently active. If the workbook is open, the keyword retrieves and returns the list of sheet names from the active workbook.

        *Examples*
        | ***** Settings *****
        | Library    ExcelSage
        |
        | ***** Test Cases *****
        | Example
        |   Open Workbook     workbook_name=\\path\\to\\excel\\file.xlsx
        |   ${all_sheets}     Get Sheets
        """
        active_workbook = self.__get_active_workbook()
        logger.info(
            f"Sheets in currently opened workbook {active_workbook.sheetnames}."
        )
        return active_workbook.sheetnames

    @keyword
    def add_sheet(
        self,
        sheet_name: str,
        sheet_pos: Optional[int] = None,
        sheet_data: Optional[List[List[Any]]] = None,
    ) -> str:
        """
        The ``Add Sheet`` keyword adds a new sheet to the active workbook. It first checks if the workbook is open, raising a ``WorkbookNotOpenError`` if it's not.

        It then checks if the ``sheet_name`` already exists in the workbook, raising a ``SheetAlreadyExistsError`` if it does. If a ``sheet_pos`` is provided, the keyword ensures that it is within the valid range of sheet positions. If the position is invalid, an ``InvalidSheetPositionError`` is raised, indicating the position and the maximum allowed value.

        *Examples*
        | ***** Settings *****
        | Library    ExcelSage
        |
        | ***** Keywords *****
        | Prepare Sheet Data
        |   ${row_header}    Create List    Name    Age   Salary
        |   ${row1}    Create List    John    30   1,000
        |   ${row2}    Create List    Mark    25   2,000
        |   ${row3}    Create List    Sam    56   3,000
        |   ${sheet_data}    Create List    ${row_header}   ${row1}    ${row2}    ${row3}
        |   RETURN  ${sheet_data}
        |
        | ***** Test Cases *****
        | Example
        |   Open Workbook     workbook_name=\\path\\to\\excel\\file.xlsx
        |   ${newly_added_sheet}     Add Sheet     sheet_name=Sheet1     sheet_pos=1
        |   ${newly_added_sheet}     Add Sheet     sheet_name=Sheet2    sheet_data=${sheet_data}
        """
        active_workbook = self.__get_active_workbook()

        self.__argument_type_checker(
            {
                "sheet_name": [sheet_name, str, None],
                "sheet_data": [sheet_data, list, None],
                "sheet_pos": [sheet_pos, int, None],
            }
        )

        if sheet_name in active_workbook.sheetnames:
            raise SheetAlreadyExistsError(sheet_name)

        if sheet_pos is not None and (
            sheet_pos < 0 or sheet_pos > len(active_workbook.sheetnames)
        ):
            raise InvalidSheetPositionError(sheet_pos, len(active_workbook.sheetnames))

        active_workbook.create_sheet(title=sheet_name, index=sheet_pos)
        sheet = active_workbook[sheet_name]

        if sheet_data:
            for index, row in enumerate(sheet_data):
                if not isinstance(row, list):
                    raise TypeError(
                        f"Invalid row at index {index} of type '{type(row).__name__}'. Each row in 'sheet_data' must be a list."
                    )
                sheet.append(row)

        active_workbook.save(self.__get_active_workbook_name())
        logger.info(f"Sheet '{sheet_name}' added successfully")
        return sheet_name

    @keyword
    def delete_sheet(self, sheet_name: Optional[str] = None) -> str:
        """
        The ``Delete Sheet`` keyword removes a specified sheet from the active workbook. If no sheet name is provided, the currently active sheet is used. It retrieves the sheet to delete by name, removes it from the workbook, saves the workbook with the changes and returns the sheet name.

        *Examples*
        | ***** Settings *****
        | Library    ExcelSage
        |
        | ***** Test Cases *****
        | Example
        |   Open Workbook     workbook_name=\\path\\to\\excel\\file.xlsx
        |   ${deleted_sheet}     Delete Sheet     sheet_name=Sheet1
        """
        sheet_name = self.__get_active_sheet_name(sheet_name)
        active_workbook = self.__get_active_workbook()
        sheet_to_delete = active_workbook[sheet_name]
        active_workbook.remove(sheet_to_delete)
        active_workbook.save(self.__get_active_workbook_name())
        logger.info(f"Sheet '{sheet_name}' deleted successfully")
        return sheet_name

    @keyword
    def fetch_sheet_data(
        self,
        sheet_name: Optional[str] = None,
        ignore_empty_rows: bool = False,
        ignore_empty_columns: bool = False,
        starting_cell: str = "A1",
        output_format: str = "list",
    ) -> Union[List[Any], Dict[Any, Any], DataFrame]:
        """
        The ``Fetch Sheet Data`` keyword retrieves data from a specified sheet in the active workbook. If no sheet
        name is provided, it defaults to the active sheet. The keyword takes an optional ``output_format`` parameter,
        which can be ``list``, ``dict``, or ``dataframe``, specifying the desired format for the returned data. The
        sheet data is read using ``pandas.read_excel()`` and returned in the specified format: - ``list``: The data
        is returned as a list of lists. - ``dict``: The data is returned as a list of dictionaries, with each row
        represented as a dictionary. - ``dataframe``: The data is returned as a pandas DataFrame.

        If an invalid format is provided, a ``ValueError`` is raised.

        *Examples*
        | ***** Settings *****
        | Library    ExcelSage
        |
        | ***** Test Cases *****
        | Example
        |   Open Workbook     workbook_name=\\path\\to\\excel\\file.xlsx
        |   ${fecthed_data}     Fetch Sheet Data     output_format=dataframe    starting_cell=C10   ignore_empty_rows=True
        |   ${fetched_data}     Fetch Sheet Data     sheet_name=Sheet1     output_format=dataframe      ignore_empty_columns=True
        """
        sheet_name = self.__get_active_sheet_name(sheet_name)
        self.__argument_type_checker(
            {
                "output_format": [output_format, str],
                "ignore_empty_columns": [ignore_empty_columns, bool],
                "ignore_empty_rows": [ignore_empty_rows, bool],
                "starting_cell": [starting_cell, str],
            }
        )

        if output_format.lower().strip() not in ["list", "dict", "dataframe"]:
            raise ValueError(
                "Invalid output format. Use 'list', 'dict', or 'dataframe'."
            )

        try:
            range_boundaries(starting_cell)
        except ValueError:
            raise InvalidCellAddressError(starting_cell)

        active_workbook = self.__get_active_workbook()
        sheet = active_workbook[sheet_name]
        data = sheet[starting_cell : sheet.dimensions.split(":")[-1]]
        data_list = [[cell.value for cell in row] for row in data]

        new_data = []
        valid_row_found = False

        for row in data_list:
            if not valid_row_found:
                if len(set(row)) == 1 and None in set(row):
                    continue
                else:
                    valid_row_found = True
            new_data.append(row)

        headers = new_data[0] if new_data else None
        data = new_data[1:] if len(new_data) > 1 else []

        df = pd.DataFrame(data, columns=headers)

        if ignore_empty_rows:
            df.dropna(how="all", inplace=True)

        if ignore_empty_columns:
            df.dropna(axis=1, how="all", inplace=True)

        if output_format == "list":
            return df.values.tolist()
        elif output_format == "dict":
            return df.to_dict(orient="records")
        elif output_format == "dataframe":
            return df.reset_index(drop=True)

    @keyword
    def rename_sheet(self, old_name: str, new_name: str) -> None:
        """
        The ``Rename Sheet`` keyword renames a sheet in the active workbook from ``old_name`` to ``new_name``. It
        first checks if the workbook is open, raising a ``WorkbookNotOpenError`` if it's not.

        The keyword checks if the ``old_name`` exists in the workbook. If it does not, it raises a
        ``SheetDoesntExistsError``. It also verifies that the ``new_name`` does not already exist in the workbook; if
        it does, a ``SheetAlreadyExistsError`` is raised.

        *Examples*
        | ***** Settings *****
        | Library    ExcelSage
        |
        | ***** Test Cases *****
        | Example
        |   Open Workbook     workbook_name=\\path\\to\\excel\\file.xlsx
        |   ${renamed_sheet}     Rename Sheet     old_name=Sheet1     new_name=New_Sheet
        """
        active_workbook = self.__get_active_workbook()

        self.__argument_type_checker(
            {"old_name": [old_name, str], "new_name": [new_name, str]}
        )

        if old_name not in active_workbook.sheetnames:
            raise SheetDoesntExistsError(old_name)

        if new_name in active_workbook.sheetnames:
            raise SheetAlreadyExistsError(new_name)

        sheet = active_workbook[old_name]
        sheet.title = new_name
        active_workbook.save(self.__get_active_workbook_name())
        logger.info(f"Sheet '{old_name}' renamed to '{new_name}'")
        return new_name

    @keyword
    def get_cell_value(self, cell_name: str, sheet_name: Optional[str] = None) -> Any:
        """
        The ``Get Cell Value`` keyword retrieves the value of a specified cell from a given sheet in the active
        workbook. If no sheet name is provided, it defaults to the currently active sheet.

        It then accesses the specified sheet and attempts to fetch the value of the given cell. If the cell is empty,
        it logs this information and returns ``None``. If the cell value exists, it returns the value.

        If a ``ValueError`` is raised (for example, if the cell address is invalid), the keyword raises an
        ``InvalidCellAddressError``.

        *Examples*
        | ***** Settings *****
        | Library    ExcelSage
        |
        | ***** Test Cases *****
        | Example
        |   Open Workbook     workbook_name=\\path\\to\\excel\\file.xlsx
        |   ${cell_value}     Get Cell Value     cell_name=A1     sheet_name=Sheet1
        |   ${cell_value}     Get Cell Value     cell_name=B10
        """
        sheet_name = self.__get_active_sheet_name(sheet_name)
        self.__argument_type_checker({"cell_name": [cell_name, str]})
        active_workbook = self.__get_active_workbook()
        sheet = active_workbook[sheet_name]

        try:
            cell_value = sheet[cell_name].value
            logger.info(f"Cell {cell_name} value is {cell_name}.")
            return cell_value
        except ValueError:
            raise InvalidCellAddressError(cell_name)

    @keyword
    def close_workbook(self, alias: Optional[str] = None) -> None:
        """
        The ``Close Workbook`` keyword is responsible for closing a workbook. By default, it closes the currently active workbook.
        If an alias is provided, it closes the workbook with that alias. After closing, the workbook is removed from the open workbooks dictionary.
        If no workbook is open or the specified alias doesn't exist, it raises a ``WorkbookNotOpenError``.

        If the closed workbook was the active one, the first remaining workbook (if any) becomes the new active workbook.

        *Examples*
        | ***** Settings *****
        | Library    ExcelSage
        |
        | ***** Test Cases *****
        | Example
        |   Open Workbook     workbook_name=\\path\\to\\excel\\file.xlsx     alias=source
        |   Open Workbook     workbook_name=\\path\\to\\excel\\file2.xlsx     alias=target
        |   Close Workbook    # Closes the active workbook
        |   Close Workbook    alias=source    # Closes the workbook with alias 'source'
        """
        if alias is None:
            if self.active_workbook_alias is None:
                raise WorkbookNotOpenError()
            alias = self.active_workbook_alias
        if alias not in self.workbooks:
            raise WorkbookNotOpenError(f"Workbook with alias '{alias}' is not open.")

        self.workbooks[alias]["workbook"].close()

        del self.workbooks[alias]

        if self.active_workbook_alias == alias:
            if self.workbooks:
                self.active_workbook_alias = next(iter(self.workbooks))
                logger.info(
                    f"Switched active workbook to '{self.active_workbook_alias}'"
                )
            else:
                self.active_workbook_alias = None
                self.active_sheet = None

        logger.info(f"Workbook with alias '{alias}' closed successfully!")

    @keyword
    def switch_workbook(self, alias: str) -> None:
        """
        The ``Switch Workbook`` keyword switches the active workbook to the one specified by the alias.
        The alias must correspond to an already open workbook. If the alias doesn't exist, it raises a ``WorkbookNotOpenError``.

        *Examples*
        | ***** Settings *****
        | Library    ExcelSage
        |
        | ***** Test Cases *****
        | Example
        |   Open Workbook     workbook_name=\\path\\to\\excel\\file.xlsx     alias=source
        |   Open Workbook     workbook_name=\\path\\to\\excel\\file2.xlsx     alias=target
        |   Switch Workbook    alias=target    # Switch to 'target' workbook
        |   ${sheets}    Get Sheets    # Gets sheets from 'target' workbook
        """
        self.__argument_type_checker({"alias": [alias, str]})

        if alias not in self.workbooks:
            raise WorkbookNotOpenError(
                f"Workbook with alias '{alias}' is not open. Available aliases: {list(self.workbooks.keys())}"
            )

        self.active_workbook_alias = alias
        self.active_sheet = None
        logger.info(f"Switched to workbook with alias '{alias}'")

    @keyword
    def save_workbook(self, alias: Optional[str] = None) -> None:
        """
        The ``Save Workbook`` keyword saves a workbook. By default, it saves the currently active workbook.
        If an alias is provided, it saves the workbook with that alias. It first checks if there is a workbook
        open, raising a ``WorkbookNotOpenError`` if no workbook is available or if the specified alias doesn't exist.
        If a workbook is open, the keyword saves it to the file specified, ensuring that any changes made to the
        workbook are persisted. This keyword does not return anything, as it simply saves the workbook.

        *Examples*
        | ***** Settings *****
        | Library    ExcelSage
        |
        | ***** Test Cases *****
        | Example
        |   Open Workbook     workbook_name=\\path\\to\\excel\\file.xlsx     alias=source
        |   Open Workbook     workbook_name=\\path\\to\\excel\\file2.xlsx     alias=target
        |   Save Workbook     # Saves the active workbook
        |   Save Workbook     alias=source    # Saves the workbook with alias 'source'
        """
        if alias is None:
            if self.active_workbook_alias is None:
                raise WorkbookNotOpenError()
            alias = self.active_workbook_alias
        if alias not in self.workbooks:
            raise WorkbookNotOpenError(f"Workbook with alias '{alias}' is not open.")

        workbook = self.workbooks[alias]["workbook"]
        workbook_name = self.workbooks[alias]["name"]

        workbook.save(workbook_name)
        logger.info(f"Workbook '{workbook_name}' saved successfully!")

    @keyword
    def set_active_sheet(self, sheet_name: str) -> str:
        """
        The ``Set Active Sheet`` keyword sets a specified sheet as the active sheet in the workbook. If no workbook
        is currently open, it raises a ``WorkbookNotOpenError``.

        Next, it checks if the `sheet_name` exists in the workbook. If the sheet is not found,
        a ``SheetDoesntExistsError`` is raised.  This allows subsequent operations to be performed on the newly
        selected active sheet.

        *Examples*
        | ***** Settings *****
        | Library    ExcelSage
        |
        | ***** Test Cases *****
        | Example
        |   Open Workbook     workbook_name=\\path\\to\\excel\\file.xlsx
        |   ${activae_sheet_name}     Set Active Sheet     sheet_name=Sheet1
        """
        self.__argument_type_checker({"sheet_name": [sheet_name, str, None]})
        active_workbook = self.__get_active_workbook()

        if sheet_name not in active_workbook.sheetnames:
            raise SheetDoesntExistsError(sheet_name)

        self.active_sheet = active_workbook[sheet_name]
        logger.info(f"Sheet '{sheet_name}' set as active.")
        return sheet_name

    @keyword
    def write_to_cell(
        self, cell_name: str, cell_value: Union[str, int, float, bool, type(None)], sheet_name: Optional[str] = None
    ) -> None:
        """
        The ``Write To Cell`` keyword writes a specified value into a cell in the active workbook. It first checks if
        a valid ``sheet_name`` is provided; if not, it defaults to the currently active sheet.

        Once the input is validated, it retrieves the sheet by name and attempts to write the value into the
        specified cell. If successful, the workbook is saved to persist the changes.

        If the provided ``cell_name`` is invalid, a ``ValueError`` is raised, which is caught and re-raised as an
        ``InvalidCellAddressError``.

        *Examples*
        | ***** Settings *****
        | Library    ExcelSage
        |
        | ***** Test Cases *****
        | Example
        |   Open Workbook     workbook_name=\\path\\to\\excel\\file.xlsx
        |   Write To Cell     cell_name=A1     cell_value=Test Data     sheet_name=Sheet1
        """
        sheet_name = self.__get_active_sheet_name(sheet_name)
        self.__argument_type_checker(
            {"cell_name": [cell_name, str], "cell_value": [cell_value, (str, int, float, bool, type(None))]}
        )
        active_workbook = self.__get_active_workbook()
        sheet = active_workbook[sheet_name]

        try:
            sheet[cell_name] = cell_value
            active_workbook.save(self.__get_active_workbook_name())
            logger.info(
                f"Written '{cell_value}' to {cell_name} in sheet '{sheet_name}'."
            )
        except ValueError:
            raise InvalidCellAddressError(cell_name)

    @keyword
    def get_column_count(
        self,
        starting_cell: str = "A1",
        ignore_empty_columns: bool = False,
        sheet_name: Optional[str] = None,
    ) -> int:
        """
        The ``Get Column Count`` keyword retrieves the total number of columns in a specified sheet from the active workbook.
        If no sheet name is provided, it defaults to the currently active sheet.

        *Examples*
        | ***** Settings *****
        | Library    ExcelSage
        |
        | ***** Test Cases *****
        | Example
        |   Open Workbook     workbook_name=\\path\\to\\excel\\file.xlsx
        |   ${column_count}     Get Column Count    starting_cell=C10   ignore_empty_columns=True
        """
        sheet_name = self.__get_active_sheet_name(sheet_name)
        self.__argument_type_checker(
            {
                "starting_cell": [starting_cell, str],
                "ignore_empty_columns": [ignore_empty_columns, bool],
            }
        )

        try:
            range_boundaries(starting_cell)
        except ValueError:
            raise InvalidCellAddressError(starting_cell)

        start_col_letter = "".join(filter(str.isalpha, starting_cell))
        start_row = int("".join(filter(str.isdigit, starting_cell)))
        start_col_index = column_index_from_string(start_col_letter)

        active_workbook = self.__get_active_workbook()
        sheet = active_workbook[sheet_name]
        headers_range = sheet.iter_rows(
            min_row=start_row,
            max_row=start_row,
            min_col=start_col_index,
            values_only=True,
        )
        headers = next(headers_range)

        df = pd.DataFrame([headers])

        if ignore_empty_columns:
            df.dropna(axis=1, how="all", inplace=True)

        column_count = df.shape[1]
        logger.info(f"Column count in sheet {sheet_name} is {column_count}.")
        return column_count

    @keyword
    def get_row_count(
        self,
        sheet_name: Optional[str] = None,
        starting_cell: str = "A1",
        include_header: bool = False,
        ignore_empty_rows: bool = False,
    ) -> int:
        """
        The ``Get Row Count`` keyword retrieves the total number of rows in a specified sheet from the active
        workbook. If no sheet name is provided, it defaults to the currently active sheet.

        If the exclude_header flag is set to True, the keyword reduces the row count by 1 to exclude the header row,
        ensuring that the result is never negative.

        *Examples*
        | ***** Settings *****
        | Library    ExcelSage
        |
        | ***** Test Cases *****
        | Example
        |   Open Workbook     workbook_name=\\path\\to\\excel\\file.xlsx
        |   ${row_count}     Get Row Count      ignore_empty_rows=True      starting_cell=C10       include_header=True
        """
        sheet_name = self.__get_active_sheet_name(sheet_name)
        self.__argument_type_checker(
            {
                "starting_cell": [starting_cell, str],
                "include_header": [include_header, bool],
                "ignore_empty_rows": [ignore_empty_rows, bool],
            }
        )

        try:
            range_boundaries(starting_cell)
        except ValueError:
            raise InvalidCellAddressError(starting_cell)

        start_col_letter = "".join(filter(str.isalpha, starting_cell))
        start_row = int("".join(filter(str.isdigit, starting_cell)))
        start_col_index = column_index_from_string(start_col_letter)

        active_workbook = self.__get_active_workbook()
        sheet = active_workbook[sheet_name]
        headers_range = sheet.iter_rows(
            min_row=start_row,
            max_row=start_row,
            min_col=start_col_index,
            values_only=True,
        )
        headers = next(headers_range)

        data = sheet[starting_cell : sheet.dimensions.split(":")[-1]]
        data_list = [[cell.value for cell in row] for row in data]

        new_data = []
        valid_row_found = False

        for row in data_list:
            if not valid_row_found:
                if len(set(row)) == 1 and None in set(row):
                    continue
                else:
                    valid_row_found = True
            new_data.append(row)

        headers = new_data[0] if new_data else None
        data = new_data[1:] if len(new_data) > 1 else []
        df = pd.DataFrame(data, columns=headers)

        if ignore_empty_rows:
            df.dropna(axis=0, how="all", inplace=True)

        row_count = df.shape[0]

        if include_header and headers:
            row_count += 1

        logger.info(f"Row count in sheet {sheet_name} is {row_count}.")

        return row_count

    @keyword
    def append_row(self, row_data: List[Any], sheet_name: Optional[str] = None) -> None:
        """
        The ``Append Row`` keyword appends a new row of data to the specified sheet in the active workbook. If no ``sheet_name`` is provided, it defaults to the currently active sheet.

        Once validated, the row data is appended to the sheet, which automatically places the data in the next available row. After appending the data, the workbook is saved to persist the changes.

        *Examples*
        | ***** Settings *****
        | Library    ExcelSage
        |
        | ***** Variables ******
        | @{data}     John     Doe     Maths   100
        |
        | ***** Test Cases *****
        | Example
        |   Open Workbook     workbook_name=\\path\\to\\excel\\file.xlsx
        |   Append Row     row_data=${data}
        """
        sheet_name = self.__get_active_sheet_name(sheet_name)
        self.__argument_type_checker({"row_data": [row_data, list]})
        active_workbook = self.__get_active_workbook()
        sheet = active_workbook[sheet_name]
        sheet.append(row_data)
        active_workbook.save(self.__get_active_workbook_name())
        logger.info(f"Row append to sheet {sheet_name}.")

    @keyword
    def insert_row(
        self, row_data: List[Any], row_index: int, sheet_name: Optional[str] = None
    ) -> None:
        """
        The ``Insert Row`` keyword inserts a new row at a specified index in an Excel sheet and populates that row with the provided data.
        The keyword validates the input and ensures that the row index is within Excel's allowable limits (1 to 1,048,576).

        *Examples*
        | ***** Settings *****
        | Library    ExcelSage
        |
        | ***** Variables ******
        | @{data}     John     Doe     Maths   100
        |
        | ***** Test Cases *****
        | Example
        |   Open Workbook     workbook_name=\\path\\to\\excel\\file.xlsx
        |   Insert Row     row_data=${data}     row_index=2
        """
        sheet_name = self.__get_active_sheet_name(sheet_name)
        self.__argument_type_checker(
            {"row_data": [row_data, list], "row_index": [row_index, int]}
        )

        if row_index < 1 or row_index > 1048576:
            raise InvalidRowIndexError(row_index)

        active_workbook = self.__get_active_workbook()
        sheet = active_workbook[sheet_name]
        sheet.insert_rows(row_index)
        for col_index, value in enumerate(row_data, start=1):
            sheet.cell(row=row_index, column=col_index, value=value)

        active_workbook.save(self.__get_active_workbook_name())
        logger.info(f"Inserted row at index {row_index} in sheet '{sheet_name}'.")

    @keyword
    def delete_row(self, row_index: int, sheet_name: str = None) -> None:
        """
        The ``Delete Row`` keyword deletes a specified row from the active workbook's sheet.
        You can optionally specify the sheet name; if not provided, the currently active sheet will be used.

        The row index must be within Excel's allowable range (1 to 1,048,576). If the index is out of bounds,
        the keyword raises an ``InvalidRowIndexError``. After deleting the row, the workbook is saved to persist the changes.

        *Examples*
        | ***** Settings *****
        | Library    ExcelSage
        |
        | ***** Test Cases *****
        | Example
        |   Open Workbook     workbook_name=\\path\\to\\excel\\file.xlsx
        |   Delete Row        row_index=3
        """
        sheet_name = self.__get_active_sheet_name(sheet_name)
        self.__argument_type_checker({"row_index": [row_index, int]})

        if row_index < 1 or row_index > 1048576:
            raise InvalidRowIndexError(row_index)

        active_workbook = self.__get_active_workbook()
        sheet = active_workbook[sheet_name]
        sheet.delete_rows(row_index)
        active_workbook.save(self.__get_active_workbook_name())
        logger.info(f"Deleted row at index {row_index}.")

    @keyword
    def append_column(
        self, col_data: Union[List[Any], Tuple[Any]], sheet_name: Optional[str] = None
    ) -> None:
        """
        The ``Append Column`` keyword appends a new column of data to the specified sheet in the active workbook.
        If no ``sheet_name`` is provided, it defaults to the currently active sheet.

        The column data will be appended in the next available column (after the last used column). After appending the data, the workbook is saved to persist the changes.

        *Examples*
        | ***** Settings *****
        | Library    ExcelSage
        |
        | ***** Variables ******
        | @{data}     Name     John     Mark    Dee
        |
        | ***** Test Cases *****
        | Example
        |   Open Workbook     workbook_name=\\path\\to\\excel\\file.xlsx
        |   Append Column     col_data=${data}
        """
        sheet_name = self.__get_active_sheet_name(sheet_name)
        self.__argument_type_checker({"col_data": [col_data, (list, tuple)]})

        active_workbook = self.__get_active_workbook()
        sheet = active_workbook[sheet_name]

        if sheet.max_row == 1 and sheet.max_column == 1 and sheet["A1"].value is None:
            next_column = 1
        else:
            next_column = sheet.max_column + 1

        col_letter = get_column_letter(next_column)

        for row_index, value in enumerate(col_data, start=1):
            sheet[f"{col_letter}{row_index}"] = value

        active_workbook.save(self.__get_active_workbook_name())
        logger.info(f"Column appended to sheet {sheet_name}.")

    @keyword
    def insert_column(
        self,
        col_data: Union[List[Any], Tuple[Any]],
        col_index: int,
        sheet_name: Optional[str] = None,
    ) -> None:
        """
        The ``Insert Column`` keyword inserts a new column at a specified index in an Excel sheet and populates that column with the provided data.
        The keyword validates the input and ensures that the column index is within Excel's allowable limits (1 to 16,384).

        *Examples*
        | ***** Settings *****
        | Library    ExcelSage
        |
        | ***** Variables ******
        | @{data}     John     Doe     Maths   100
        |
        | ***** Test Cases *****
        | Example
        |   Open Workbook     workbook_name=\\path\\to\\excel\\file.xlsx
        |   Insert Column     col_data=${data}     col_index=2
        """
        sheet_name = self.__get_active_sheet_name(sheet_name)
        self.__argument_type_checker(
            {"col_data": [col_data, (list, tuple)], "col_index": [col_index, int]}
        )

        if col_index < 1 or col_index > 16384:
            raise InvalidColumnIndexError(col_index)

        active_workbook = self.__get_active_workbook()
        sheet = active_workbook[sheet_name]
        sheet.insert_cols(col_index)

        for row_index, value in enumerate(col_data, start=1):
            sheet.cell(row=row_index, column=col_index, value=value)

        active_workbook.save(self.__get_active_workbook_name())
        logger.info(f"Inserted column at index {col_index}.")

    @keyword
    def delete_column(self, col_index: int, sheet_name: Optional[str] = None) -> None:
        """
        The ``Delete Column`` keyword deletes a specified column from the active workbook's sheet.
        You can optionally specify the sheet name; if not provided, the currently active sheet will be used.

        The column index must be within Excel's allowable range (1 to 16,384). If the index is out of bounds,
        the keyword raises an ``InvalidColumnIndexError``. After deleting the column, the workbook is saved to persist the changes.

        *Examples*
        | ***** Settings *****
        | Library    ExcelSage
        |
        | ***** Test Cases *****
        | Example
        |   Open Workbook     workbook_name=\\path\\to\\excel\\file.xlsx
        |   Delete Column     col_index=2
        """
        sheet_name = self.__get_active_sheet_name(sheet_name)
        self.__argument_type_checker({"col_index": [col_index, int]})

        if col_index < 1 or col_index > 16384:
            raise InvalidColumnIndexError(col_index)

        active_workbook = self.__get_active_workbook()
        sheet = active_workbook[sheet_name]
        sheet.delete_cols(col_index)
        active_workbook.save(self.__get_active_workbook_name())
        logger.info(f"Deleted column at index {col_index}.")

    @keyword
    def get_column_values(
        self,
        column_names_or_letters: Union[str, List[str]],
        output_format: str = "list",
        sheet_name: Optional[str] = None,
        starting_cell: str = "A1",
    ) -> Union[List[Any], dict, DataFrame]:
        """
        The ``Get Column Values`` keyword retrieves all values from the specified column(s) in an Excel sheet.
        The column(s) can be specified by header names (e.g., 'name', 'age') or by column letters (e.g., 'A', 'B', 'C').
        You can also specify a starting cell (e.g., 'A3') from which the header and data start.
        The output format can be a list, dictionary, or pandas DataFrame.

        *Examples*
        | ***** Settings *****
        | Library    ExcelSage
        |
        | ***** Variables *****
        | @{column_header}      Name    Age
        | @{column_letters}     A    C
        |
        | ***** Test Cases *****
        | Example
        |   Open Workbook     workbook_name=\\path\\to\\excel\\file.xlsx
        |   ${column_values}    Get Column Values    column_names_or_letters=name     output_format=list    sheet_name=Sheet1
        |   ${column_values}    Get Column Values    column_names_or_letters=A     starting_cell=B5
        |   ${multiple_columns} Get Column Values    column_names_or_letters=${column_header}     output_format=dict     starting_cell=B3
        |   ${multiple_columns} Get Column Values    column_names_or_letters=${column_letters}    output_format=dataframe     starting_cell=C4
        """
        sheet_name = self.__get_active_sheet_name(sheet_name)
        self.__argument_type_checker(
            {
                "column_names_or_letters": [
                    column_names_or_letters,
                    (str, list, tuple),
                ],
                "output_format": [output_format, str],
                "starting_cell": [starting_cell, str],
            }
        )

        if output_format.lower().strip() not in ["list", "dict", "dataframe"]:
            raise ValueError(
                "Invalid output format. Use 'list', 'dict', or 'dataframe'."
            )

        try:
            range_boundaries(starting_cell)
        except ValueError:
            raise InvalidCellAddressError(starting_cell)

        if isinstance(column_names_or_letters, str):
            column_names_or_letters = [column_names_or_letters]

        start_col_letter = "".join(filter(str.isalpha, starting_cell))
        start_row = int("".join(filter(str.isdigit, starting_cell)))
        start_col_index = column_index_from_string(start_col_letter)

        active_workbook = self.__get_active_workbook()
        sheet = active_workbook[sheet_name]
        headers_range = sheet.iter_rows(
            min_row=start_row,
            max_row=start_row,
            min_col=start_col_index,
            values_only=True,
        )
        first_row = next(headers_range)

        headers_to_fetch = []
        for col in column_names_or_letters:
            if isinstance(col, str) and col in first_row:
                headers_to_fetch.append(col)
            elif col.isalpha() and len(col) < 4:
                col_index = column_index_from_string(col)
                if col_index - 1 < len(first_row):
                    header = first_row[col_index - 1]
                    if isinstance(header, str):
                        headers_to_fetch.append(header)
                    else:
                        raise ValueError(
                            f"Column letter '{col}' does not have a valid string header: '{header}' found."
                        )
                else:
                    raise ValueError(
                        f"Column letter '{col}' is out of bounds for the provided sheet."
                    )
            else:
                raise ValueError(f"Invalid column name or letter: '{col}'")

        df = pd.read_excel(
            self.__get_active_workbook_name(),
            sheet_name=sheet_name,
            usecols=headers_to_fetch,
            header=start_row - 1,
        )

        if output_format.lower().strip() == "list":
            if len(headers_to_fetch) == 1:
                return df.iloc[:, 0].tolist()
            return [df[col].tolist() for col in df.columns]
        elif output_format.lower().strip() == "dict":
            return df.to_dict(orient="list")
        elif output_format.lower().strip() == "dataframe":
            return df.reset_index(drop=True)

    @keyword
    def get_row_values(
        self,
        row_indices: Union[int, List[int], Tuple[int]],
        output_format: str = "list",
        sheet_name: Optional[str] = None,
    ) -> Union[List[Any], dict]:
        """
        Retrieves all values from the specified row(s) in an Excel sheet.
        The row(s) can be specified by their index (starting from 1 for the first row).
        The output format can be a list or a dictionary.

        *Examples*
        | ***** Settings *****
        | Library    ExcelSage
        |
        | ***** Variables *****
        | @{rows}      2     5     7
        |
        | ***** Test Cases *****
        | Example
        |   Open Workbook     workbook_name=\\path\\to\\excel\\file.xlsx
        |   ${row_values}     Get Row Values    row_indices=2     output_format=list
        |   ${row_values}     Get Row Values    row_indices=${rows}    output_format=dict
        """
        sheet_name = self.__get_active_sheet_name(sheet_name)
        self.__argument_type_checker({"row_indices": [row_indices, (int, list, tuple)]})

        if output_format.lower().strip() not in ["list", "dict"]:
            raise ValueError("Invalid output format. Use 'list' or 'dict'.")

        if isinstance(row_indices, int):
            row_indices = [row_indices]

        active_workbook = self.__get_active_workbook()
        sheet = active_workbook[sheet_name]

        row_data = {}
        for row_index in row_indices:
            if row_index < 1 or row_index > 1048576:
                raise InvalidRowIndexError(row_index)

            row_values = [
                cell.value
                for cell in next(sheet.iter_rows(min_row=row_index, max_row=row_index))
            ]
            row_data[row_index] = row_values

        if output_format.lower().strip() == "list":
            if len(row_indices) == 1:
                return row_data[row_indices[0]]
            return [row_data[row] for row in row_indices]
        elif output_format.lower().strip() == "dict":
            return row_data

    @keyword
    def protect_sheet(self, password: str, sheet_name: Optional[str] = None) -> None:
        """
        The ``Protect Sheet`` keyword protects a specified sheet in the active workbook with a password. If the sheet is already protected, it raises a ``SheetAlreadyProtected`` error. Otherwise, it applies the password protection and saves the workbook.

        *Examples*
        | ***** Settings *****
        | Library    ExcelSage
        |
        | ***** Test Cases *****
        | Example
        |   Open Workbook     workbook_name=\\path\\to\\excel\\file.xlsx
        |   Protect Sheet     password=YourPassword     sheet_name=Sheet1
        """
        sheet_name = self.__get_active_sheet_name(sheet_name)
        self.__argument_type_checker({"password": [password, str]})
        active_workbook = self.__get_active_workbook()
        sheet = active_workbook[sheet_name]

        if sheet.protection.sheet:
            raise SheetAlreadyProtectedError(sheet_name)

        sheet.protection.set_password(password)
        active_workbook.save(self.__get_active_workbook_name())
        logger.info(f"Sheet {sheet_name} is protected successfully.")

    @keyword
    def unprotect_sheet(self, password: str, sheet_name: Optional[str] = None) -> None:
        """
        The ``Unprotect Sheet`` keyword removes password protection from a specified sheet in the active workbook.
        If the sheet is not protected, it raises a ``SheetNotProtectedError`` error.

        *Examples*
        | ***** Settings *****
        | Library    ExcelSage
        |
        | ***** Test Cases *****
        | Example
        |   Open Workbook     workbook_name=\\path\\to\\excel\\file.xlsx
        |   Unprotect Sheet   password=YourPassword     sheet_name=Sheet1
        """
        sheet_name = self.__get_active_sheet_name(sheet_name)
        self.__argument_type_checker({"password": [password, str]})
        active_workbook = self.__get_active_workbook()
        sheet = active_workbook[sheet_name]

        if not sheet.protection.sheet:
            raise SheetNotProtectedError(sheet_name)

        sheet.protection.set_password(password)
        sheet.protection.sheet = False

        active_workbook.save(self.__get_active_workbook_name())
        logger.info(f"Sheet {sheet_name} has been unprotected successfully.")

    @keyword
    def protect_workbook(self, password: str, protect_sheets: bool = True) -> None:
        """
        ``Protect Workbook`` keyword Protect the entire workbook structure with a password.
        This will prevent actions like adding, deleting, renaming, hiding,
        or moving sheets. Additionally, it can optionally protect all individual
        sheets, restricting various editing capabilities like inserting or deleting
        rows, columns, and modifying cell formats.

        *Examples*
        | ***** Settings *****
        | Library    ExcelSage
        |
        | ***** Test Cases *****
        | Example
        |   Open Workbook     workbook_name=\\path\\to\\excel\\file.xlsx
        |   Protect Workbook    password=YourPassword
        |   Protect Workbook    password=YourPassword       protect_sheets=False

        """
        active_workbook = self.__get_active_workbook()

        self.__argument_type_checker(
            {"password": [password, str], "protect_sheets": [protect_sheets, bool]}
        )

        if active_workbook.security and active_workbook.security.lockStructure:
            raise WorkbookAlreadyProtectedError()

        protection = WorkbookProtection()
        protection.workbookPassword = password
        protection.lockStructure = True
        protection.lockWindows = True

        active_workbook.security = protection

        if protect_sheets:
            for sheet in active_workbook.worksheets:
                sheet.protection = SheetProtection(password=password)
                sheet.protection.sheet = True
                sheet.protection.formatCells = False
                sheet.protection.formatColumns = False
                sheet.protection.formatRows = False
                sheet.protection.insertColumns = False
                sheet.protection.insertRows = False
                sheet.protection.deleteColumns = False
                sheet.protection.deleteRows = False
                sheet.protection.sort = False
                sheet.protection.autoFilter = False
                sheet.protection.objects = False
                sheet.protection.scenarios = False

        active_workbook.save(self.__get_active_workbook_name())
        logger.info("Workbook have been successfully protected.")

    @keyword
    def unprotect_workbook(self, unprotect_sheets: bool = False) -> None:
        """
        ``Unprotect Workbook`` keyword Unprotect the entire workbook by removing its structure protection.
        This allows actions like adding, deleting, renaming, hiding, or moving sheets.

        Optionally, individual sheet protections can also be removed by specifying ``unprotect_sheets=True``.

        *Examples*
        | ***** Settings *****
        | Library    ExcelSage
        |
        | ***** Test Cases *****
        | Example
        |   Open Workbook      workbook_name=\\path\\to\\excel\\file.xlsx
        |   Unprotect Workbook
        |   Unprotect Workbook      unprotect_sheets=True

        """
        active_workbook = self.__get_active_workbook()

        self.__argument_type_checker({"unprotect_sheets": [unprotect_sheets, bool]})

        if not active_workbook.security.lockStructure:
            raise WorkbookNotProtectedError()

        active_workbook.security.lockStructure = False

        if unprotect_sheets:
            for sheet in active_workbook.worksheets:
                if sheet.protection.sheet:
                    sheet.protection.sheet = False
                    logger.info(f"Sheet {sheet.title} unprotected.")

        active_workbook.save(self.__get_active_workbook_name())
        logger.info("Workbook have been successfully unprotected.")

    @keyword
    def clear_sheet(self, sheet_name: Optional[str] = None) -> str:
        """
        The ``Clear Sheet`` keyword clears all cell values in the specified sheet of the active workbook. Once the sheet is cleared, the keyword saves the workbook.

        *Examples*
        | ***** Settings *****
        | Library    ExcelSage
        |
        | ***** Test Cases *****
        | Example
        |   Open Workbook     workbook_name=\\path\\to\\excel\\file.xlsx
        |   ${cleared_sheet_name}     Clear Sheet     sheet_name=Sheet1
        """
        sheet_name = self.__get_active_sheet_name(sheet_name)
        active_workbook = self.__get_active_workbook()
        sheet = active_workbook[sheet_name]
        for row in sheet.iter_rows(
            min_row=1, max_col=sheet.max_column, max_row=sheet.max_row
        ):
            for cell in row:
                cell.value = None
        active_workbook.save(self.__get_active_workbook_name())
        logger.info(f"Cleared sheet {sheet_name}.")
        return sheet_name

    @keyword
    def copy_sheet(self, source_sheet_name: str, new_sheet_name: str) -> str:
        """
        The ``Copy Sheet`` keyword creates a copy of an existing sheet in the workbook, giving it a new name. It first validates that the workbook is open and checks that the source sheet exists. If the new sheet name is invalid or already exists, it raises an appropriate error. The keyword copies the source sheet, renames the copy, and saves the workbook.

        *Examples*
        | ***** Settings *****
        | Library    ExcelSage
        |
        | ***** Test Cases *****
        | Example
        |   Open Workbook     workbook_name=\\path\\to\\excel\\file.xlsx
        |   ${copied_sheet_name}     Copy Sheet      source_sheet_name=Sheet1     new_sheet_name=CopiedSheet
        """
        active_workbook = self.__get_active_workbook()

        self.__argument_type_checker(
            {
                "source_sheet_name": [source_sheet_name, str],
                "new_sheet_name": [new_sheet_name, str],
            }
        )

        new_sheet_name = new_sheet_name.strip()
        if source_sheet_name not in active_workbook.sheetnames:
            raise SheetDoesntExistsError(source_sheet_name)

        if (
            not new_sheet_name
            or len(new_sheet_name) > 31
            or any(
                char in new_sheet_name for char in [":", "/", "\\", "?", "*", "[", "]"]
            )
        ):
            raise InvalidSheetNameError(new_sheet_name)

        source = active_workbook[source_sheet_name]
        target = active_workbook.copy_worksheet(source)
        target.title = new_sheet_name
        active_workbook.save(self.__get_active_workbook_name())
        logger.info(f"Coppied sheet {source_sheet_name} to {new_sheet_name}.")
        return new_sheet_name

    @keyword
    def find_value(
        self, value: Any, sheet_name: Optional[str] = None, occurence: str = "first"
    ) -> Union[str, List[str]]:
        """
        The ``Find Value`` keyword searches for a specified value in a sheet and can return either the first occurrence or all occurrences, depending on the occurence parameter.
        If the value is found, the keyword returns the cell coordinate(s). If no match is found, it returns ``None``.

        *Examples*
        | ***** Settings *****
        | Library    ExcelSage
        |
        | ***** Test Cases *****
        | Example
        |   Open Workbook     workbook_name=\\path\\to\\excel\\file.xlsx
        |   ${cell_cordinate}     Find Value     value=John Doe     sheet_name=Sheet1     occurence=first
        |   ${cell_cordinates}     Find Value     value=John Doe     sheet_name=Sheet1     occurence=all
        |   ${cell_cordinate}     Find Value     value=John Doe

        """
        all_occurences = []
        sheet_name = self.__get_active_sheet_name(sheet_name)

        if occurence.lower().strip() not in ["first", "all"]:
            raise ValueError("Invalid occurence, use either 'first' or 'all'.")

        active_workbook = self.__get_active_workbook()
        sheet = active_workbook[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value == value:
                    if occurence.lower().strip() == "first":
                        logger.info(f"Value found in cell {cell.coordinate}.")
                        return cell.coordinate
                    elif occurence.lower().strip() == "all":
                        all_occurences.append(cell.coordinate)

        else:
            if all_occurences:
                logger.info(f"Value found in cell(s) {all_occurences}")
                return all_occurences
            logger.info("Value not found in any cell.")
            return None

    @keyword
    def find_and_replace(
        self,
        old_value: Any,
        new_value: Any,
        sheet_name: Optional[str] = None,
        occurence: str = "first",
    ) -> Union[str, List[str], None]:
        """
        The ``Find and Replace`` keyword searches for a specified value (``old_value``) in the specified sheet and replaces it with the ``new_value``.
        The ``occurence`` parameter determines whether only the first occurrence or all occurrences are replaced.
        If the value is found, the keyword returns the cell coordinate(s) where the replacement was made. If no match is found, it returns ``None``.

        *Examples*
        | ***** Settings *****
        | Library    ExcelSage
        |
        | ***** Test Cases *****
        | Example
        |   Open Workbook     workbook_name=\\path\\to\\excel\\file.xlsx
        |   ${cell_cordinate}     Find and Replace     old_value=John Doe     new_value=John Smith     sheet_name=Sheet1     occurence=first
        |   ${cell_cordinates}    Find and Replace     old_value=John Doe     new_value=John Smith     sheet_name=Sheet1     occurence=all
        |   ${cell_cordinate}     Find and Replace     old_value=John Doe     new_value=John Smith
        """
        replaced_cells = []
        sheet_name = self.__get_active_sheet_name(sheet_name)

        if occurence.lower().strip() not in ["first", "all"]:
            raise ValueError("Invalid occurence, use either 'first' or 'all'.")

        active_workbook = self.__get_active_workbook()
        sheet = active_workbook[sheet_name]

        for row in sheet.iter_rows():
            for cell in row:
                if cell.value == old_value:
                    if occurence.lower().strip() == "first":
                        cell.value = new_value
                        active_workbook.save(self.__get_active_workbook_name())
                        logger.info(
                            f"Replaced '{old_value}' with '{new_value}' in cell {cell.coordinate}."
                        )
                        return cell.coordinate
                    elif occurence.lower().strip() == "all":
                        cell.value = new_value
                        replaced_cells.append(cell.coordinate)

        else:
            if replaced_cells:
                active_workbook.save(self.__get_active_workbook_name())
                logger.info(
                    f"Replaced '{old_value}' with '{new_value}' in cells {replaced_cells}."
                )
                return replaced_cells
            logger.info(f"Value '{old_value}' not found in any cell.")
            return None

    @keyword
    def format_cell(
        self,
        cell_name: str,
        font_size: Optional[int] = None,
        font_color: Optional[str] = None,
        sheet_name: Optional[str] = None,
        alignment: Optional[dict] = None,
        wrap_text: Optional[bool] = None,
        bg_color: Optional[str] = None,
        cell_width: Optional[Union[int, float]] = None,
        cell_height: Optional[Union[int, float]] = None,
        font_name: Optional[str] = None,
        bold: Optional[bool] = None,
        italic: Optional[bool] = None,
        underline: Optional[bool] = None,
        strike_through: Optional[bool] = None,
        border: Optional[dict] = None,
        auto_fit_height: Optional[bool] = None,
        auto_fit_width: Optional[bool] = None,
    ) -> None:
        """
        The ``Format Cell`` keyword allows formatting of a specified cell in the active workbook. It accepts a variety of formatting options, including font size, color, alignment, background color, borders, and text formatting (bold, italic, underline, etc.). Invalid parameters trigger specific exceptions such as ``InvalidColorError``, ``InvalidAlignmentError``, or ``InvalidBorderStyleError``.

        *Examples*
        | ***** Settings *****
        | Library    ExcelSage
        |
        | ***** Variables *****
        | &{alignments}        vertical=center     horizontal=left
        | &{border}         left=True     right=True     top=False     bottom=True     style=thin     color=#FF0000
        |
        | ***** Test Cases *****
        | Example
        |   Open Workbook     workbook_name=\\path\\to\\excel\\file.xlsx
        |   Format Cell     cell_name=A1     font_size=15    font_color=#FFFF00     alignment=${alignments}     border=${border}
        |   Format Cell     cell_name=A1     wrap_text=True     font_name=Arial     cell_width=25
        """
        sheet_name = self.__get_active_sheet_name(sheet_name)
        self.__argument_type_checker(
            {
                "cell_name": [cell_name, str],
                "font_size": [font_size, int, None],
                "font_color": [font_color, str, None],
                "alignment": [alignment, dict, None],
                "wrap_text": [wrap_text, bool, None],
                "bg_color": [bg_color, str, None],
                "cell_width": [cell_width, (int, float), None],
                "cell_height": [cell_height, (int, float), None],
                "font_name": [font_name, str, None],
                "bold": [bold, bool, None],
                "italic": [italic, bool, None],
                "underline": [underline, bool, None],
                "strike_through": [strike_through, bool, None],
                "border": [border, dict, None],
                "auto_fit_height": [auto_fit_height, bool, None],
                "auto_fit_width": [auto_fit_width, bool, None],
            }
        )

        active_workbook = self.__get_active_workbook()
        sheet = active_workbook[sheet_name]

        try:
            cell = sheet[cell_name]
            current_font = cell.font
            cell_value = str(cell.value) if cell.value else ""
            col_letter = get_column_letter(cell.column)
            row_num = "".join(filter(str.isdigit, cell_name))

            font_args = {
                "name": font_name if font_name else current_font.name,
                "bold": bold if bold is not None else current_font.bold,
                "italic": italic if italic is not None else current_font.italic,
                "underline": "single" if underline else current_font.underline,
                "size": font_size if font_size else current_font.size,
                "color": current_font.color,
                "strike": strike_through
                if strike_through is not None
                else current_font.strike,
            }

            if font_color is not None:
                if not re.match(r"^#[0-9A-Fa-f]{6}$", font_color):
                    raise InvalidColorError(color_type="font", color=font_color)

                if font_color.startswith("#"):
                    font_color = "FF" + font_color[1:]
                font_args["color"] = font_color

            cell.font = Font(**font_args)

            # Apply background color
            if bg_color is not None:
                if not re.match(r"^#[0-9A-Fa-f]{6}$", bg_color):
                    raise InvalidColorError(color_type="background", color=bg_color)

                if bg_color.startswith("#"):
                    bg_color = "FF" + bg_color[1:]
                cell.fill = PatternFill(
                    start_color=bg_color, end_color=bg_color, fill_type="solid"
                )

            # Apply alignment and wrap text
            if alignment or wrap_text is not None:
                if alignment:
                    vertical_align = alignment.get("vertical")
                    horizontal_align = alignment.get("horizontal")

                    if (
                        vertical_align
                        and vertical_align not in self.VALID_VERTICAL_ALIGNMENTS
                    ):
                        raise InvalidAlignmentError(
                            alignment_type="vertical",
                            alignment_value=vertical_align,
                            allowed_values=self.VALID_VERTICAL_ALIGNMENTS,
                        )
                    if (
                        horizontal_align
                        and horizontal_align not in self.VALID_HORIZONTAL_ALIGNMENTS
                    ):
                        raise InvalidAlignmentError(
                            alignment_type="horizontal",
                            alignment_value=horizontal_align,
                            allowed_values=self.VALID_HORIZONTAL_ALIGNMENTS,
                        )

                align_args = {
                    "horizontal": horizontal_align if horizontal_align else None,
                    "vertical": vertical_align if vertical_align else None,
                    "wrap_text": wrap_text
                    if wrap_text is not None
                    else cell.alignment.wrap_text,
                }
                cell.alignment = Alignment(**align_args)

            # Set column width and row height
            if cell_width is not None:
                sheet.column_dimensions[col_letter].width = cell_width
            if cell_height is not None:
                sheet.row_dimensions[cell.row].height = cell_height

            # Apply borders
            if border is not None:
                border_sides = {}
                border_style = border.get("style", "thin")
                border_color = border.get("color", "#000000")

                if border_style not in self.VALID_BORDER_STYLES:
                    raise InvalidBorderStyleError(
                        border_style=border_style,
                        allowed_styles=self.VALID_BORDER_STYLES,
                    )

                if not re.match(r"^#[0-9A-Fa-f]{6}$", border_color):
                    raise InvalidColorError(color_type="border", color=border_color)

                # Convert color to ARGB
                if border_color.startswith("#"):
                    border_color = "FF" + border_color[1:]

                # Define sides based on the border dictionary
                if border.get("left"):
                    border_sides["left"] = Side(
                        border_style=border_style, color=border_color
                    )
                if border.get("right"):
                    border_sides["right"] = Side(
                        border_style=border_style, color=border_color
                    )
                if border.get("top"):
                    border_sides["top"] = Side(
                        border_style=border_style, color=border_color
                    )
                if border.get("bottom"):
                    border_sides["bottom"] = Side(
                        border_style=border_style, color=border_color
                    )

                cell.border = Border(**border_sides)

            if auto_fit_width:
                max_length = max(len(cell_value), len(col_letter))
                column_width = max_length + 2
                sheet.column_dimensions[col_letter].width = column_width

            if auto_fit_height:
                max_line_count = cell_value.count("\n") + 1
                row_height = max(15, max_line_count * 15)
                sheet.row_dimensions[int(row_num)].height = row_height

            active_workbook.save(self.__get_active_workbook_name())
            logger.info(f"Formatted cell {cell_name}.")

        except ValueError:
            raise InvalidCellAddressError(cell_name)

    @keyword
    def merge_excels(
        self,
        file_list: list,
        output_filename: str,
        merge_type: str = "multiple_sheets",
        skip_bad_rows: bool = False,
    ) -> None:
        """
        The ``Merge Excels`` keyword provides functionality to merge multiple Excel files into a single output file, supporting three different merge strategies and handling potential row issues with an optional flag.

        - *Case 1:* ``multiple_sheets``:
            In this mode, all sheets from all input files are copied to the output file.
            Each sheet is given a unique name by combining the sheet name with the base name of the file (sanitized to meet Excel's character restrictions).
            It ensures sheet names do not exceed 31 characters and replaces invalid characters with underscores.

        - *Case 2:* ``single_sheet``:
            All sheets from all files are merged into a single DataFrame, appending rows from all files.
            If the ``skip_bad_rows`` flag is set to ``True``, rows with column mismatches are skipped; otherwise, an exception is raised.

        - *Case 3:* ``sheet_wise``:
            Sheets from all files are merged sheet by sheet. The first sheets of all files are merged into one, the second sheets into another, and so on.
            The function handles cases where some files may not have as many sheets, logging a warning and skipping those sheets.
            The optional ``skip_bad_rows`` flag allows the keyword to skip problematic rows if set to ``True``.

        *Examples*
        | ***** Settings *****
        | Library    ExcelSage
        |
        | ***** Variables *****
        | @{files}        path\\to\\excel\\file1    path\\to\\excel\\file2
        | ${output_file}    path\\to\\output\\excel\\file.xlsx
        |
        | ***** Test Cases *****
        | Example
        |   Merge Excels     file_list=${files}     output_filename=${output_file}    merge_type=sheet_wise     skip_bad_rows=True
        |   Merge Excels     file_list=${files}     output_filename=${output_file}    merge_type=multiple_sheets
        """
        self.__argument_type_checker(
            {
                "file_list": [file_list, list],
                "output_filename": [output_filename, str],
                "merge_type": [merge_type, str],
                "skip_bad_rows": [skip_bad_rows, bool],
            }
        )
        if not file_list:
            raise ValueError(
                "The file list is empty. Provide at least one file to merge."
            )
        if merge_type not in ["multiple_sheets", "single_sheet", "sheet_wise"]:
            raise ValueError(
                "Invalid merge type. Use 'multiple_sheets', 'single_sheet', or 'sheet_wise'."
            )

        writer = pd.ExcelWriter(output_filename, engine="openpyxl")

        if merge_type == "multiple_sheets":
            # Case 1: Multiple Excel files with multiple sheets merged into a single Excel with all those sheets
            for file_path in file_list:
                if not os.path.exists(file_path):
                    raise ExcelFileNotFoundError(file_path)

                with pd.ExcelFile(file_path) as excel_file:
                    for sheet_name in excel_file.sheet_names:
                        df = pd.read_excel(file_path, sheet_name=sheet_name)
                        unique_sheet_name = (
                            f"{sheet_name}_{Path(os.path.basename(file_path)).stem}"
                        )
                        unique_sheet_name = re.sub(
                            r"[:/\\?*\[\]]", "_", unique_sheet_name
                        )[:31]
                        df.to_excel(writer, sheet_name=unique_sheet_name, index=False)

        elif merge_type == "single_sheet":
            # Case 2: Multiple Excel files with multiple sheets merged into a single sheet
            merged_df = pd.DataFrame()
            for file_path in file_list:
                if not os.path.exists(file_path):
                    raise ExcelFileNotFoundError(file_path)

                with pd.ExcelFile(file_path) as excel_file:
                    for sheet_name in excel_file.sheet_names:
                        try:
                            df = pd.read_excel(file_path, sheet_name=sheet_name)
                            merged_df = pd.concat(
                                [merged_df, df], ignore_index=True, sort=False
                            )
                        except Exception as e:
                            if skip_bad_rows:
                                print(
                                    f"Skipping rows with issues in {sheet_name} from {file_path}"
                                )
                            else:
                                raise e

            merged_df.to_excel(writer, sheet_name="Merged_Sheet", index=False)

        elif merge_type == "sheet_wise":
            # Case 3: Merging Excel files sheet-wise
            for file_path in file_list:
                if not os.path.exists(file_path):
                    raise ExcelFileNotFoundError(file_path)

            max_sheets = max([len(pd.ExcelFile(f).sheet_names) for f in file_list])

            for i in range(max_sheets):
                merged_df = pd.DataFrame()
                for file_path in file_list:
                    with pd.ExcelFile(file_path) as excel_file:
                        try:
                            sheet_name = excel_file.sheet_names[i]
                            df = pd.read_excel(file_path, sheet_name=sheet_name)
                            merged_df = pd.concat(
                                [merged_df, df], ignore_index=True, sort=False
                            )
                        except IndexError:
                            warnings.warn(
                                f"File {file_path} does not have sheet {i + 1}. Skipping.",
                                category=UserWarning,
                            )
                        except Exception as e:
                            if skip_bad_rows:
                                logger.warn(
                                    f"Skipping rows with issues in sheet {i + 1} from {file_path}"
                                )
                            else:
                                raise e

                merged_df.to_excel(writer, sheet_name=f"Sheet_{i + 1}", index=False)

        writer.close()
        logger.info(f"Merged Excel created: {output_filename}")

    @keyword
    def merge_cells(self, cell_range: str, sheet_name: Optional[str] = None) -> None:
        """
        The ``Merge Cells`` keyword merges a specified range of cells into one in the active workbook.
        The merged cells will span across the range provided by the ``cell_range`` argument.

        *Examples*
        | ***** Test Cases *****
        | Example
        |   Open Workbook     workbook_name=\\path\\to\\excel\\file.xlsx
        |   Merge Cells       cell_range=A1:D4     sheet_name=Sheet1
        |   Merge Cells       cell_range=B2:C2     sheet_name=Sheet2
        """
        sheet_name = self.__get_active_sheet_name(sheet_name)
        self.__argument_type_checker({"cell_range": [cell_range, str]})
        active_workbook = self.__get_active_workbook()
        sheet = active_workbook[sheet_name]

        try:
            min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        except ValueError as e:
            raise InvalidCellRangeError(e)

        if min_row > max_row or min_col > max_col:
            raise InvalidCellRangeError(
                f"Invalid cell range: {cell_range}. The start cell must be smaller than the end cell."
            )

        sheet.merge_cells(cell_range)
        active_workbook.save(self.__get_active_workbook_name())
        logger.info(f"Merged cells in range {cell_range}.")

    @keyword
    def unmerge_cells(self, cell_range: str, sheet_name: Optional[str] = None) -> None:
        """
        The ``Unmerge Cells`` keyword unmerges a specified range of cells into one in the active workbook.
        The unmerged cells will span across the range provided by the ``cell_range`` argument.

        *Examples*
        | ***** Test Cases *****
        | Example
        |   Open Workbook     workbook_name=\\path\\to\\excel\\file.xlsx
        |   Unmerge Cells       cell_range=A1:D4     sheet_name=Sheet1
        |   Unmerge Cells       cell_range=B2:C2     sheet_name=Sheet2
        """
        sheet_name = self.__get_active_sheet_name(sheet_name)
        self.__argument_type_checker({"cell_range": [cell_range, str]})
        active_workbook = self.__get_active_workbook()
        sheet = active_workbook[sheet_name]

        try:
            min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        except ValueError as e:
            raise InvalidCellRangeError(e)

        if min_row > max_row or min_col > max_col:
            raise InvalidCellRangeError(
                f"Invalid cell range: {cell_range}. The start cell must be smaller than the end cell."
            )

        sheet.unmerge_cells(cell_range)
        active_workbook.save(self.__get_active_workbook_name())
        logger.info(f"Unmerged cells in range {cell_range}.")

    @keyword
    def sort_column(
        self,
        column_name_or_letter: str,
        asc: bool = True,
        starting_cell: str = "A1",
        output_format: str = "list",
        sheet_name: Optional[str] = None,
    ) -> Union[List[Any], dict, DataFrame]:
        """
        The ``Sort Column`` keyword sorts the specified column in the sheet starting from a specific cell.
        The column can be specified by name (e.g., 'Salary') or by letter (e.g., 'A'), and the sorted values can be returned as a list, dictionary, or DataFrame.

        *Examples*
            | ***** Settings *****
            | Library    ExcelSage
            |
            | ***** Test Cases *****
            | Example
            |   Open Workbook     workbook_name=\\path\\to\\excel\\file.xlsx
            |   ${sorted_values}   Sort Column     column_name_or_letter=Age     output_format=list     sheet_name=Sheet1   asc=False
            |   ${sorted_values}   Sort Column     column_name_or_letter=Salary     output_format=dict     starting_cell=D6
        """
        sheet_name = self.__get_active_sheet_name(sheet_name)
        self.__argument_type_checker(
            {"column_name_or_letter": [column_name_or_letter, str]}
        )

        if output_format.lower().strip() not in ["list", "dict", "dataframe"]:
            raise ValueError(
                "Invalid output format. Use 'list', 'dict', or 'dataframe'."
            )

        try:
            range_boundaries(starting_cell)
        except ValueError:
            raise InvalidCellAddressError(starting_cell)

        start_col_letter = "".join(filter(str.isalpha, starting_cell))
        start_row = int("".join(filter(str.isdigit, starting_cell)))
        start_col_index = column_index_from_string(start_col_letter)

        active_workbook = self.__get_active_workbook()
        sheet = active_workbook[sheet_name]
        headers_range = sheet.iter_rows(
            min_row=start_row,
            max_row=start_row,
            min_col=start_col_index,
            values_only=True,
        )
        first_row = next(headers_range)

        if column_name_or_letter in first_row:
            header_to_fetch = column_name_or_letter
        elif column_name_or_letter.isalpha() and len(column_name_or_letter) < 4:
            col_index = column_index_from_string(column_name_or_letter)
            if col_index - 1 < len(first_row):
                header = first_row[col_index - 1]
                for col in first_row:
                    if not isinstance(col, str):
                        raise ValueError(
                            f"{sheet_name} does not have a valid string header: '{col}' found."
                        )
                header_to_fetch = header
            else:
                raise ValueError(
                    f"Column letter '{column_name_or_letter}' is out of bounds for the provided sheet."
                )
        else:
            raise ValueError(
                f"Invalid column name or letter: '{column_name_or_letter}'"
            )

        df = pd.read_excel(
            self.__get_active_workbook_name(),
            sheet_name=sheet_name,
            usecols=first_row,
            header=start_row - 1,
        )

        df_sorted = df.sort_values(by=header_to_fetch, ascending=asc)
        for row_idx, row in enumerate(
            df_sorted.itertuples(index=False), start=start_row + 1
        ):
            for col_idx, value in enumerate(row, start=start_col_index):
                sheet.cell(row=row_idx, column=col_idx, value=value)

        active_workbook.save(self.__get_active_workbook_name())
        logger.info(
            f"Sorted column '{column_name_or_letter}' and saved changes to '{sheet_name}'."
        )

        if output_format.lower() == "list":
            return df_sorted.values.tolist()
        elif output_format.lower() == "dict":
            return df_sorted.to_dict(orient="list")
        elif output_format.lower() == "dataframe":
            return df_sorted.reset_index(drop=True)

    @keyword
    def find_duplicates(
        self,
        column_names_or_letters: Optional[Union[str, List[str], Tuple[str]]] = None,
        output_format: str = "list",
        starting_cell: str = "A1",
        sheet_name: Optional[str] = None,
        delete: bool = False,
        output_filename: Optional[str] = None,
        overwrite_if_exists: bool = False,
    ) -> Union[List[Any], dict, DataFrame, int]:
        """
        The `Find Duplicates`` keyword identifies and retrieves duplicate rows from the specified column(s) in the Excel sheet.
        It can check for duplicates based on either column names or column letters, and the results can be returned in different formats such as a list, dictionary, or pandas DataFrame.
        Additionally, you can specify a starting cell from which the headers begin, and filter duplicates from that point onward.

        When `delete=True`, the duplicate rows are removed from the sheet (keeping the first occurrence) and saved to a file.
        In this case, `output_filename` is mandatory to avoid modifying the source file directly.
        The function returns the number of rows deleted (int) instead of the duplicate data.

        The `overwrite_if_exists` parameter controls whether an existing output file can be overwritten.
        If `overwrite_if_exists=False` (default) and the output file already exists, a `FileAlreadyExistsError` will be raised.
        Set `overwrite_if_exists=True` to allow overwriting existing files.

        *Examples*
        | ***** Settings *****
        | Library    ExcelSage
        |
        | ***** Variables *****
        | @{columns}    Age    Gender
        |
        | ***** Test Cases *****
        | Example
        |   Open Workbook     workbook_name=\\path\\to\\excel\\file.xlsx
        |   ${duplicates}    Find Duplicates    column_names_or_letters=Age    output_format=list    sheet_name=Sheet1
        |   ${duplicates}    Find Duplicates    column_names_or_letters=${columns}    output_format=dict    sheet_name=Sheet1
        |   ${duplicates}    Find Duplicates    column_names_or_letters=B
        |   ${duplicates}    Find Duplicates     output_format=dataframe     starting_cell=D6
        |   ${deleted_count}    Find Duplicates    column_names_or_letters=Age    sheet_name=Sheet1    delete=True    output_filename=\\path\\to\\output.xlsx
        |   ${deleted_count}    Find Duplicates    column_names_or_letters=Age    sheet_name=Sheet1    delete=True    output_filename=\\path\\to\\output.xlsx    overwrite_if_exists=True
        """

        sheet_name = self.__get_active_sheet_name(sheet_name)

        # Validate that output_filename is provided when delete=True
        if delete and not output_filename:
            raise ValueError(
                "When delete=True, output_filename is mandatory to avoid modifying the source file. "
                "Please provide an output_filename parameter."
            )

        # Validate overwrite_if_exists when output_filename is provided
        if (
            delete
            and output_filename
            and os.path.exists(output_filename)
            and not overwrite_if_exists
        ):
            raise FileAlreadyExistsError(output_filename)

        self.__argument_type_checker(
            {
                "column_names_or_letters": [
                    column_names_or_letters,
                    (str, list, tuple),
                    None,
                ],
                "output_format": [output_format, str],
                "starting_cell": [starting_cell, str],
                "overwrite_if_exists": [overwrite_if_exists, bool],
                "delete": [delete, bool],
                "output_filename": [output_filename, str, None],
            }
        )

        if not delete and output_format.lower().strip() not in [
            "list",
            "dict",
            "dataframe",
        ]:
            raise ValueError(
                "Invalid output format. Use 'list', 'dict', or 'dataframe'."
            )

        try:
            range_boundaries(starting_cell)
        except ValueError:
            raise InvalidCellAddressError(starting_cell)

        start_row = int("".join(filter(str.isdigit, starting_cell)))
        header_row = start_row - 1

        if column_names_or_letters:
            if isinstance(column_names_or_letters, str):
                column_names_or_letters = [column_names_or_letters]

            start_col_letter = "".join(filter(str.isalpha, starting_cell))
            start_col_index = column_index_from_string(start_col_letter)

            active_workbook = self.__get_active_workbook()
            sheet = active_workbook[sheet_name]
            headers_range = sheet.iter_rows(
                min_row=start_row,
                max_row=start_row,
                min_col=start_col_index,
                values_only=True,
            )
            first_row = next(headers_range)

            headers_to_fetch = []
            for col in column_names_or_letters:
                if isinstance(col, str) and col in first_row:
                    headers_to_fetch.append(col)
                elif col.isalpha() and len(col) < 4:
                    col_index = column_index_from_string(col)
                    if col_index - 1 < len(first_row):
                        header = first_row[col_index - 1]
                        for col in first_row:
                            if not isinstance(col, str):
                                raise ValueError(
                                    f"{sheet_name} does not have a valid string header: '{col}' found."
                                )
                            headers_to_fetch.append(header)
                    else:
                        raise ValueError(
                            f"Column letter '{col}' is out of bounds for the provided sheet."
                        )
                else:
                    raise ValueError(f"Invalid column name or letter: '{col}'")

            df = pd.read_excel(
                self.__get_active_workbook_name(),
                sheet_name=sheet_name,
                usecols=first_row,
                header=header_row,
            )
            duplicates = df[df.duplicated(subset=headers_to_fetch, keep=False)]
        else:
            start_col_letter = "".join(filter(str.isalpha, starting_cell))
            start_col_index = column_index_from_string(start_col_letter)

            active_workbook = self.__get_active_workbook()
            sheet = active_workbook[sheet_name]
            headers_range = sheet.iter_rows(
                min_row=start_row,
                max_row=start_row,
                min_col=start_col_index,
                values_only=True,
            )
            first_row = next(headers_range)

            # Read only columns from starting_cell onwards
            df = pd.read_excel(
                self.__get_active_workbook_name(),
                sheet_name=sheet_name,
                usecols=first_row,
                header=header_row,
            )
            duplicates = df[df.duplicated(keep=False)]

        if delete:
            start_col_letter = "".join(filter(str.isalpha, starting_cell))
            start_row_num = int("".join(filter(str.isdigit, starting_cell)))
            start_col_index = column_index_from_string(start_col_letter)

            if column_names_or_letters:
                df_full = df.copy()
                df_unique = df_full.drop_duplicates(
                    subset=headers_to_fetch, keep="first"
                )
                original_columns = df_full.columns.tolist()
            else:
                df_full = df.copy()
                df_unique = df_full.drop_duplicates(keep="first")
                original_columns = df_full.columns.tolist()

            rows_deleted = len(df_full) - len(df_unique)

            workbook_path = (
                output_filename
                if output_filename
                else self.__get_active_workbook_name()
            )
            source_path = self.__get_active_workbook_name()

            if (
                self.active_workbook_alias
                and self.active_workbook_alias in self.workbooks
            ):
                self.workbooks[self.active_workbook_alias]["workbook"].close()

            source_wb = excel.load_workbook(source_path)
            source_ws = source_wb[sheet_name]

            max_row = source_ws.max_row
            max_col = source_ws.max_column

            original_num_cols = len(original_columns)
            end_col_index = start_col_index + original_num_cols - 1

            data_start_row = start_row_num + 1

            original_max_data_row = data_start_row + len(df_full) - 1
            if max_row >= data_start_row:
                clear_to_row = max(original_max_data_row, max_row)
                for row in range(data_start_row, clear_to_row + 1):
                    for col in range(
                        start_col_index, min(end_col_index + 1, max_col + 1)
                    ):
                        cell = source_ws.cell(row=row, column=col)
                        cell.value = None

            if original_num_cols > 0:
                for col_idx, header_value in enumerate(original_columns):
                    target_col = start_col_index + col_idx
                    source_ws.cell(
                        row=start_row_num, column=target_col, value=header_value
                    )

            data_to_write = df_unique.values.tolist()
            for row_idx, row_data in enumerate(data_to_write, start=data_start_row):
                for col_idx, value in enumerate(row_data):
                    target_col = start_col_index + col_idx
                    source_ws.cell(row=row_idx, column=target_col, value=value)

            last_written_row = data_start_row + len(data_to_write) - 1
            original_last_data_row = data_start_row + len(df_full) - 1

            if original_last_data_row > last_written_row:
                rows_to_delete = original_last_data_row - last_written_row
                source_ws.delete_rows(last_written_row + 1, rows_to_delete)

            source_wb.save(workbook_path)
            source_wb.close()

            if output_filename:
                loaded_workbook = excel.load_workbook(workbook_path)
                if workbook_path not in self.workbooks:
                    self.workbooks[workbook_path] = {
                        "workbook": loaded_workbook,
                        "name": workbook_path,
                    }
                    self.active_workbook_alias = workbook_path
            else:
                loaded_workbook = excel.load_workbook(workbook_path)
                if (
                    self.active_workbook_alias
                    and self.active_workbook_alias in self.workbooks
                ):
                    self.workbooks[self.active_workbook_alias]["workbook"] = (
                        loaded_workbook
                    )
                    self.workbooks[self.active_workbook_alias]["name"] = workbook_path

            return rows_deleted

        if output_format.lower().strip() == "list":
            return duplicates.values.tolist()
        elif output_format.lower().strip() == "dict":
            return duplicates.to_dict(orient="list")
        elif output_format.lower().strip() == "dataframe":
            return duplicates.reset_index(drop=True)

    @keyword
    def remove_empty_rows(
        self,
        output_filename: str,
        sheet_name: Optional[str] = None,
        column_names_or_letters: Optional[Union[str, List[str], Tuple[str]]] = None,
        overwrite_if_exists: bool = False,
        starting_cell: str = "A1",
    ) -> int:
        """
        The ``Remove Empty Rows`` keyword removes empty rows from the specified sheet in the active workbook.
        It can check for empty rows based on either all cells in a row or only specified columns.

        If `column_names_or_letters=None`, a row is considered empty if ALL cells in that row are empty.
        If `column_names_or_letters` is specified, only those columns are checked for emptiness.

        The function returns the number of rows removed.

        The `output_filename` parameter is mandatory to avoid modifying the source file directly.
        The `overwrite_if_exists` parameter controls whether an existing output file can be overwritten.
        If `overwrite_if_exists=False` (default) and the output file already exists, a `FileAlreadyExistsError` will be raised.
        Set `overwrite_if_exists=True` to allow overwriting existing files.

        The `starting_cell` parameter specifies where the data begins (including headers), allowing you to preserve
        the structure of the Excel file (e.g., headers, formatting, or data that appears before the starting cell).

        *Examples*
        | ***** Settings *****
        | Library    ExcelSage
        |
        | ***** Test Cases *****
        | Example
        |   Open Workbook     workbook_name=\\path\\to\\excel\\file.xlsx
        |   ${removed}    Remove Empty Rows    sheet_name=Sheet1    output_filename=\\path\\to\\output.xlsx
        |   ${removed}    Remove Empty Rows    column_names_or_letters=Age    sheet_name=Sheet1    output_filename=\\path\\to\\output.xlsx
        |   ${removed}    Remove Empty Rows    column_names_or_letters=${columns}    starting_cell=D6    output_filename=\\path\\to\\output.xlsx    overwrite_if_exists=True
        """
        sheet_name = self.__get_active_sheet_name(sheet_name)

        if os.path.exists(output_filename) and not overwrite_if_exists:
            raise FileAlreadyExistsError(output_filename)

        self.__argument_type_checker(
            {
                "column_names_or_letters": [
                    column_names_or_letters,
                    (str, list, tuple),
                    None,
                ],
                "starting_cell": [starting_cell, str],
                "overwrite_if_exists": [overwrite_if_exists, bool],
                "output_filename": [output_filename, str],
            }
        )

        try:
            range_boundaries(starting_cell)
        except ValueError:
            raise InvalidCellAddressError(starting_cell)

        start_row = int("".join(filter(str.isdigit, starting_cell)))
        header_row = start_row - 1

        if column_names_or_letters:
            if isinstance(column_names_or_letters, str):
                column_names_or_letters = [column_names_or_letters]

            start_col_letter = "".join(filter(str.isalpha, starting_cell))
            start_col_index = column_index_from_string(start_col_letter)

            active_workbook = self.__get_active_workbook()
            sheet = active_workbook[sheet_name]
            headers_range = sheet.iter_rows(
                min_row=start_row,
                max_row=start_row,
                min_col=start_col_index,
                values_only=True,
            )
            first_row = next(headers_range)

            headers_to_fetch = []
            for col in column_names_or_letters:
                if isinstance(col, str) and col in first_row:
                    headers_to_fetch.append(col)
                elif col.isalpha() and len(col) < 4:
                    col_index = column_index_from_string(col)
                    if col_index - 1 < len(first_row):
                        header = first_row[col_index - 1]
                        for col in first_row:
                            if not isinstance(col, str):
                                raise ValueError(
                                    f"{sheet_name} does not have a valid string header: '{col}' found."
                                )
                            headers_to_fetch.append(header)
                    else:
                        raise ValueError(
                            f"Column letter '{col}' is out of bounds for the provided sheet."
                        )
                else:
                    raise ValueError(f"Invalid column name or letter: '{col}'")

            df = pd.read_excel(
                self.__get_active_workbook_name(),
                sheet_name=sheet_name,
                usecols=first_row,
                header=header_row,
            )
            df = df.replace("", pd.NA)
            df_filtered = df.dropna(subset=headers_to_fetch, how="all")
            original_columns = df.columns.tolist()
        else:
            start_col_letter = "".join(filter(str.isalpha, starting_cell))
            start_col_index = column_index_from_string(start_col_letter)

            active_workbook = self.__get_active_workbook()
            sheet = active_workbook[sheet_name]
            max_col = sheet.max_column
            headers_range = sheet.iter_rows(
                min_row=start_row,
                max_row=start_row,
                min_col=start_col_index,
                max_col=min(max_col, start_col_index + 100),
                values_only=True,
            )
            first_row = list(next(headers_range))
            while first_row and first_row[-1] is None:
                first_row.pop()

            first_row_filtered = [col for col in first_row if col is not None]

            max_row = sheet.max_row
            actual_last_data_row = start_row
            for row_idx in range(
                start_row + 1, min(max_row + 1, start_row + 10000)
            ):  # Limit search
                for col_idx in range(start_col_index, start_col_index + len(first_row)):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    if cell.value is not None and str(cell.value).strip() != "":
                        actual_last_data_row = row_idx
                        break

            read_up_to = actual_last_data_row
            if actual_last_data_row < max_row:
                next_row = actual_last_data_row + 1
                has_any_data_anywhere = False
                for col_idx in range(1, min(sheet.max_column + 1, 100)):
                    cell = sheet.cell(row=next_row, column=col_idx)
                    if cell.value is not None and str(cell.value).strip() != "":
                        has_any_data_anywhere = True
                        break

                if has_any_data_anywhere or (
                    next_row == actual_last_data_row + 1
                    and max_row == actual_last_data_row + 1
                ):
                    read_up_to = next_row

            max_row = read_up_to
            if not first_row_filtered:
                num_cols = len(first_row)
                data_rows = []
                for row in sheet.iter_rows(
                    min_row=start_row + 1,
                    max_row=max_row,
                    min_col=start_col_index,
                    max_col=start_col_index + num_cols - 1,
                    values_only=True,
                ):
                    data_rows.append(list(row))

                df = pd.DataFrame(data_rows, columns=first_row[:num_cols])
            else:
                num_cols = len(first_row_filtered)
                col_name_to_idx = {}
                for idx, val in enumerate(first_row):
                    if val is not None and val in first_row_filtered:
                        col_name_to_idx[val] = start_col_index + idx

                data_rows = []
                for row in sheet.iter_rows(
                    min_row=start_row + 1,
                    max_row=max_row,
                    min_col=start_col_index,
                    max_col=start_col_index + len(first_row) - 1,
                    values_only=True,
                ):
                    row_data = list(row)
                    filtered_row = []
                    for col_name in first_row_filtered:
                        col_idx = col_name_to_idx[col_name] - start_col_index
                        if col_idx < len(row_data):
                            filtered_row.append(row_data[col_idx])
                        else:
                            filtered_row.append(None)
                    data_rows.append(filtered_row)

                df = pd.DataFrame(data_rows, columns=first_row_filtered)

            df = df.replace("", pd.NA)
            df_filtered = df.dropna(axis=0, how="all")
            original_columns = df.columns.tolist()

        rows_removed = len(df) - len(df_filtered)

        workbook_path = output_filename
        source_path = self.__get_active_workbook_name()

        if rows_removed == 0:
            if (
                self.active_workbook_alias
                and self.active_workbook_alias in self.workbooks
            ):
                self.workbooks[self.active_workbook_alias]["workbook"].close()

            source_wb = excel.load_workbook(source_path)
            source_wb.save(workbook_path)
            source_wb.close()

            loaded_workbook = excel.load_workbook(workbook_path)
            if workbook_path not in self.workbooks:
                self.workbooks[workbook_path] = {
                    "workbook": loaded_workbook,
                    "name": workbook_path,
                }
                self.active_workbook_alias = workbook_path

            logger.info(f"No empty rows found in sheet '{sheet_name}'.")
            return 0

        if self.active_workbook_alias and self.active_workbook_alias in self.workbooks:
            self.workbooks[self.active_workbook_alias]["workbook"].close()

        source_wb = excel.load_workbook(source_path)
        source_ws = source_wb[sheet_name]

        max_row = source_ws.max_row
        max_col = source_ws.max_column

        start_col_letter = "".join(filter(str.isalpha, starting_cell))
        start_row_num = int("".join(filter(str.isdigit, starting_cell)))
        start_col_index = column_index_from_string(start_col_letter)

        original_num_cols = len(original_columns)
        end_col_index = start_col_index + original_num_cols - 1

        data_start_row = start_row_num + 1

        original_max_data_row = data_start_row + len(df) - 1
        if max_row >= data_start_row:
            clear_to_row = max(original_max_data_row, max_row)
            for row in range(data_start_row, clear_to_row + 1):
                for col in range(start_col_index, min(end_col_index + 1, max_col + 1)):
                    cell = source_ws.cell(row=row, column=col)
                    cell.value = None

        if original_num_cols > 0:
            for col_idx, header_value in enumerate(original_columns):
                target_col = start_col_index + col_idx
                source_ws.cell(row=start_row_num, column=target_col, value=header_value)

        data_to_write = df_filtered.values.tolist()
        for row_idx, row_data in enumerate(data_to_write, start=data_start_row):
            for col_idx, value in enumerate(row_data):
                target_col = start_col_index + col_idx
                source_ws.cell(row=row_idx, column=target_col, value=value)

        last_written_row = data_start_row + len(data_to_write) - 1
        original_last_data_row = data_start_row + len(df) - 1

        if original_last_data_row > last_written_row:
            rows_to_delete = original_last_data_row - last_written_row
            source_ws.delete_rows(last_written_row + 1, rows_to_delete)

        source_wb.save(workbook_path)
        source_wb.close()

        loaded_workbook = excel.load_workbook(workbook_path)
        if workbook_path not in self.workbooks:
            self.workbooks[workbook_path] = {
                "workbook": loaded_workbook,
                "name": workbook_path,
            }
            self.active_workbook_alias = workbook_path

        logger.info(f"Removed {rows_removed} empty row(s) from sheet '{sheet_name}'.")
        return rows_removed

    @keyword
    def compare_excels(
        self,
        source_excel: str,
        target_excel: str,
        source_excel_config: Optional[dict] = None,
        target_excel_config: Optional[dict] = None,
    ) -> DataFrame:
        """
        The ``Compare Excels`` keyword compares two Excel sheets and identifies differences in the data.
        The comparison is based on the values of the specified columns, and the output includes rows that are unique to either of the two sheets. It handles the comparison intelligently, providing options to configure which sheet, starting cell, and columns should be compared for each Excel file.

        *Examples*
        | ***** Settings *****
        | Library    ExcelSage
        |
        | ***** Variables *****
        | &{source_config}    Sheet=Sheet1      Starting_cell=C1
        | &{target_config}    Sheet=Sheet1
        |
        | ***** Test Cases *****
        | Example
        |   ${differences}    Compare Excels    source_excel=\\path\\to\\excel\\source\\file.xlsx    target_excel=\\path\\to\\excel\\target\\file.xlsx    source_excel_config=${source_config}     target_excel_config=${target_config}
        |   ${differences}    Compare Excels    source_excel=\\path\\to\\excel\\source\\file.xlsx    target_excel=\\path\\to\\excel\\target\\file.xlsx
        """
        self.__argument_type_checker(
            {
                "source_excel": [source_excel, str],
                "target_excel": [target_excel, str],
                "source_excel_config": [source_excel_config, dict, None],
                "target_excel_config": [source_excel_config, dict, None],
            }
        )

        def load_excel(file_name: str, config: Optional[dict] = None) -> DataFrame:
            if not os.path.exists(file_name):
                raise ExcelFileNotFoundError(file_name)

            sheet_name = config.get("sheet_name", 0) if config else 0
            starting_cell = config.get("starting_cell", "A1") if config else "A1"
            columns = config.get("columns", None) if config else None

            try:
                range_boundaries(starting_cell)
            except ValueError:
                raise InvalidCellAddressError(starting_cell)

            start_row = int("".join(filter(str.isdigit, starting_cell)))
            df = pd.read_excel(file_name, sheet_name=sheet_name, header=start_row - 1)

            if columns is not None:
                missing_columns = [col for col in columns if col not in df.columns]
                if missing_columns:
                    raise InvalidColumnNameError(sheet_name, missing_columns)
                df = df[columns]

            return df

        source_df = load_excel(source_excel, source_excel_config)
        target_df = load_excel(target_excel, target_excel_config)

        source_columns = set(source_df.columns.tolist())
        target_columns = set(target_df.columns.tolist())

        if source_columns != target_columns:
            missing_in_source = (
                target_columns - source_columns
                if len(target_columns - source_columns) != 0
                else None
            )
            missing_in_target = (
                source_columns - target_columns
                if len(source_columns - target_columns) != 0
                else None
            )
            error_message = f"Column mismatch found in excel files.\nMissing in source: {missing_in_source}\nMissing in target: {missing_in_target}"
            raise ColumnMismatchError(error_message)

        excel_column_name = "Excel_Source"
        if "Excel_Source" in source_df.columns or "Excel_Source" in target_df.columns:
            excel_column_name = "__Excel_Source__"

        source_df[excel_column_name] = "Source"
        target_df[excel_column_name] = "Target"

        diff_df = pd.concat([source_df, target_df]).drop_duplicates(
            subset=source_df.columns.difference([excel_column_name]), keep=False
        )

        if diff_df.empty:
            logger.info("No differences found between the two Excel sheets.")
        else:
            logger.info(f"Differences found between the two Excel sheets.\n{diff_df}")

        return diff_df

    @keyword
    def export_to_csv(
        self,
        filename: str,
        sheet_name: str,
        output_filename: str,
        separator: str = ",",
        overwrite_if_exists: bool = False,
    ) -> str:
        """
        The `Export To CSV` keyword reads the data from a specified sheet in an Excel file and exports it to a CSV file.

        *Note*
        Separator must be a string of length 1

        *Examples*
        | ***** Settings *****
        | Library    ExcelSage
        |
        | ***** Test Cases *****
        | Example
        |   Export To CSV     filename=\\path\\to\\excel\\file.xlsx    sheet_name=Sheet1    output_filename=\\path\\to\\csv\\file.csv    separator=;
        """
        self.__argument_type_checker(
            {
                "filename": [filename, str],
                "output_filename": [output_filename, str],
                "overwrite_if_exists": [overwrite_if_exists, bool],
                "separator": [separator, str],
            }
        )
        if not os.path.exists(filename):
            raise ExcelFileNotFoundError(filename)

        if os.path.exists(output_filename) and not overwrite_if_exists:
            raise FileAlreadyExistsError(output_filename)

        df = pd.read_excel(filename, sheet_name=sheet_name)
        df.to_csv(output_filename, index=False, sep=separator)
        return output_filename

    @keyword
    def get_column_headers(
        self, starting_cell: str = "A1", sheet_name: Optional[str] = None
    ) -> List[str]:
        """
        The ``Get Column Headers`` keyword retrieves all column headers (i.e., the first row of data) starting from the specified cell in a sheet.
        By default, it starts from cell A1, but the starting cell can be customized.
        It returns a list of column headers.

        *Examples*
        | ***** Test Cases *****
        | Example
        |   Open Workbook     workbook_name=\\path\\to\\excel\\file.xlsx
        |   ${headers}    Get Column Headers    starting_cell=B5    sheet_name=Sheet1
        |   Log    ${headers}
        """
        sheet_name = self.__get_active_sheet_name(sheet_name)
        self.__argument_type_checker({"starting_cell": [starting_cell, str]})

        try:
            range_boundaries(starting_cell)
        except ValueError:
            raise InvalidCellAddressError(starting_cell)

        start_col_letter = "".join(filter(str.isalpha, starting_cell))
        start_row = int("".join(filter(str.isdigit, starting_cell)))
        start_col_index = column_index_from_string(start_col_letter)

        active_workbook = self.__get_active_workbook()
        sheet = active_workbook[sheet_name]
        headers_range = sheet.iter_rows(
            min_row=start_row,
            max_row=start_row,
            min_col=start_col_index,
            values_only=True,
        )
        column_headers = next(headers_range)

        return column_headers
