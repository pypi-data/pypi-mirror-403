import shutil
import unittest
from pathlib import Path

import numpy as np
import pandas as pd
import pytest
from openpyxl import load_workbook

import unac.util.data as d
import unac.util.imports as imports
from unac.correction.ict_bindings import IctCorrection
from unac.correction.isocor_bindings import IsocorCorrection
from unac.correction.isocorrector_bindings import IsoCorrectoRCorrection
from unac.correction.naive import NaiveCorrection


class TestChemFormula(unittest.TestCase):
    def test_is_chem_formula(self):
        # an empty string is a valid formula
        self.assertTrue(d.ChemFormula.is_chem_formula(""))
        # this is glucose
        self.assertTrue(d.ChemFormula.is_chem_formula("C6H12O6"))
        # wrong capitalization
        self.assertFalse(d.ChemFormula.is_chem_formula("c6h12o6")[0])
        # Non alphanumeric
        self.assertFalse(d.ChemFormula.is_chem_formula("_C6H12O6")[0])
        self.assertFalse(d.ChemFormula.is_chem_formula("lorem ipsum")[0])
        # duplicate entries
        self.assertFalse(d.ChemFormula.is_chem_formula("C6H12O6C2")[0])
        # Non existing elements
        self.assertFalse(d.ChemFormula.is_chem_formula("C6H12O6Yt3")[0])


class TestMeasurement(unittest.TestCase):
    def test_possible_masses_ms(self):
        masses = []
        for pre in range(6):
            masses.append(d.MassLane(pre, pre, False))

        test_meas = d.Measurement(
            mid="Ser60",
            masses=masses,
            pre="C3H8NO3",
            pre_c=3,
            pre_mass=0,
            pro="C3H8NO3",
            pro_c=3,
            pro_mass=0,
        )
        pos = test_meas.generate_possible_masses()
        expected = []
        for pre in range(4):
            expected.append(d.MassLane(pre, pre, False))
        self.assertEqual(pos, expected)

    def test_possible_masses_msms(self):
        masses = []
        for (
            pre,
            pro,
        ) in [
            (0, 0),
            (1, 0),
            (1, 1),
            (2, 1),
            (2, 2),
            (3, 2),
            (3, 3),
            (4, 3),
            (4, 4),
        ]:
            masses.append(d.MassLane(pre, pro, True))

        test_meas = d.Measurement(
            mid="Ser60",
            masses=masses,
            pre="C3H8NO3",
            pre_c=3,
            pre_mass=0,
            pro="C2H6NO2",
            pro_c=2,
            pro_mass=0,
        )
        pos = test_meas.generate_possible_masses()
        expected = []
        for pre, pro in [
            (0, 0),
            (1, 0),
            (1, 1),
            (2, 1),
            (2, 2),
            (3, 2),
        ]:
            expected.append(d.MassLane(pre, pro, True))
        self.assertEqual(pos, expected)


class TestDataImport(unittest.TestCase):
    RESSOURCES = Path(__file__).parent / "ressources"

    # test whether integer (non string type) as header causes error
    def test_verfiy_msms_structure_string(self):
        path = self.RESSOURCES / "verify_msms_structure_string_error.xlsx"
        wb = load_workbook(path)
        ws = wb["Tabelle1"]
        self.assertRaises(d.DataError, imports.verify_msms_structure, ws)

    # test whether misspelled header causes error
    def test_verfiy_msms_structure_header(self):
        path = self.RESSOURCES / "verify_msms_structure_headers_typo_error.xlsx"
        wb = load_workbook(path)
        ws = wb["Tabelle1"]
        self.assertRaises(d.DataError, imports.verify_msms_structure, ws)

    # test whether correct template is read
    def test_import_raw_data(self):
        path = self.RESSOURCES / "import_raw_data_correct_template.xlsx"
        wb = load_workbook(path)
        has_data, collected_m_ids = imports.import_raw_data(wb, "Tabelle1")
        assert isinstance(collected_m_ids, dict)
        assert "Ser60_C3H8NO3_C2H6NO" in collected_m_ids.keys()
        assert "data.Measurement" in str(collected_m_ids.values())
        assert has_data


class TestMsCorrection(unittest.TestCase):
    # test raw correction of MS data
    def test_naive_correction(self):
        # define a d.Measurement object with known data for correction
        masses = []
        for pre in range(4):
            masses.append(d.MassLane(pre, pre, False))
        data2_value = d.Measurement(
            mid="Ser60",
            masses=masses,
            pre="C3H8NO3",
            pre_c=3,
            pre_mass=0,
            pro="C3H8NO3",
            pro_c=3,
            pro_mass=0,
        )
        # add_data: rep, times, data, masses; it's row-wise from the template so one isotopomer per list, i.e. array row
        data = np.transpose(np.asarray([[0.9, 0.1, 0.0, 0.0], [0.8, 0.15, 0.05, 0.0], [0.6, 0.25, 0.1, 0.05]]))
        times = [0, 5, 10]
        data2_value.add_data(rep="R1", times=times, data=data, masses=masses)
        data2 = {"Ser60_C3H8NO3_C3H8NO3": data2_value}
        # perform correction
        cor = NaiveCorrection()
        data2_cor = cor.perform_correction_ui(data2, silent=True)
        # test correction result
        df = data2_cor["Ser60_C3H8NO3_C3H8NO3"].data
        solution = np.transpose(
            np.asarray(
                [
                    [0.91077, 0.09599, -0.00617, -0.00059],
                    [0.80932, 0.14712, 0.04474, -0.00119],
                    [0.60645, 0.24922, 0.09591, 0.04843],
                ]
            )
        )
        for a in range(df.shape[1]):
            rounded_values = np.asarray(df.iloc[:, a]).round(decimals=5)
            self.assertTrue(
                (rounded_values == solution[:, a]).all(),
                f"In column {a} there is an error. {rounded_values} - {solution[:, a]}",
            )

    # test raw correction of MS data
    @pytest.mark.isocorrector
    @pytest.mark.ict
    @pytest.mark.isocor
    def test_ms_correction(self):
        # define a d.Measurement object with known data for correction
        masses = []
        for pre in range(4):
            masses.append(d.MassLane(pre, pre, False))
        data2_value = d.Measurement(
            mid="Ser60",
            masses=masses,
            pre="C3H8NO3",
            pre_c=3,
            pre_mass=0,
            pro="C3H8NO3",
            pro_c=3,
            pro_mass=0,
        )
        # add_data: rep, times, data, masses; it's row-wise from the template so one isotopomer per list, i.e. array row
        data = np.transpose(np.asarray([[0.7, 0.1, 0.15, 0.05], [0.2, 0.15, 0.35, 0.3], [0.6, 0.25, 0.1, 0.05]]))
        times = [0, 5, 10]
        data2_value.add_data(rep="R1", times=times, data=data, masses=masses)
        data2 = {"Ser60_C3H8NO3_C3H8NO3": data2_value}
        # perform correction
        # part 2: compare with other tools
        correction_tools = {}
        for c in [
            IctCorrection(),
            IsoCorrectoRCorrection(),
            IsocorCorrection(),
            NaiveCorrection(),
        ]:
            correction_tools[c.name] = c

        for c in correction_tools:
            data2_cor = correction_tools[c].perform_correction_ui(data2, "./delme/", True)
            # test correction result
            df = data2_cor["Ser60_C3H8NO3_C3H8NO3"].data
            solution = np.transpose(
                np.asarray(
                    [
                        [0.70730, 0.09700, 0.14664, 0.04906],
                        [0.20124, 0.14978, 0.35006, 0.29892],
                        [0.60644, 0.24922, 0.09591, 0.04843],
                    ]
                )
            )
            for a in range(df.shape[1]):
                calculated_values = np.asarray(df.iloc[:, a])
                self.assertTrue(
                    (np.abs(calculated_values - solution[:, a]) < 1e-5).all(),
                    f"With {c} in column {a} there is an error. {calculated_values} - {solution[:, a]}",
                )
        # remove temporary files
        shutil.rmtree("./delme/")

    # test raw correction of MS data with superfluous higher traces
    @pytest.mark.isocorrector
    @pytest.mark.ict
    @pytest.mark.isocor
    def test_ms_correction_super(self):
        # define a d.Measurement object with known data for correction
        masses = []
        for pre in range(6):
            masses.append(d.MassLane(pre, pre, False))
        data2_value = d.Measurement(
            mid="Ala260",
            masses=masses,
            pre="C11H26N1O2Si2",
            pre_c=3,
            pre_mass=260,
            pro="C11H26N1O2Si2",
            pro_c=3,
            pro_mass=260,
        )
        data = np.transpose(
            np.asarray(
                [
                    [
                        0.23143457875606332,
                        0.19954278710275314,
                        0.35893520878792884,
                        0.15354311660302722,
                        0.044215882325702185,
                        0.010558816869606295,
                    ]
                ]
            )
        )
        data2_value.add_data(rep="R1", times=[0], data=data, masses=masses)
        data2 = {"Ala": data2_value}

        correction_tools = {}
        for c in [
            IctCorrection(),
            IsoCorrectoRCorrection(),
            IsocorCorrection(),
            NaiveCorrection(),
        ]:
            correction_tools[c.name] = c
        for c in correction_tools:
            data2_cor = correction_tools[c].perform_correction_ui(data2, "./delme/", True)
            # test correction result
            df: pd.DataFrame = data2_cor["Ala"].data
            solution = np.transpose(np.array([[0.3, 0.2, 0.4, 0.1, 0, 0]]))

            self.assertTrue(
                (np.abs(solution - df.to_numpy()) < 1e-8).all(),
                f"difference to large between {np.transpose(solution)} and {np.transpose(df.to_numpy())} when using {c}",
            )
        shutil.rmtree("./delme/")


if __name__ == "__main__":
    unittest.main()
