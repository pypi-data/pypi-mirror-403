import openpyxl

import unac.util.data as d


def get_meas_times(data: dict[d.Measurement]):
    """
    collect measurement time-points
    :param data: a dict of d.Measurement
    :return: a list of all time-points for which measurements are available
    :rtype: list
    """
    # collect all time-points
    # (time-points may differ for measurements)
    times = set()
    for m_id in data:
        times.update(data[m_id].get_times())
    times = [x for x in times]
    times.sort()
    return times


def export_to_excel(data: dict[d.Measurement], workbook: openpyxl.workbook, sheet: str):
    means_sheet = workbook.create_sheet(f"{sheet}_corr_means")
    # collect all time-points to build column headers
    times = get_meas_times(data)

    # write header
    means_sheet.merge_cells(range_string="A1:A2")
    means_sheet.merge_cells(range_string="B1:B2")
    means_sheet.cell(row=1, column=1, value="Measurement ID")
    means_sheet.cell(row=1, column=2, value="shift")
    for t_idx, time in enumerate(times):
        means_sheet.merge_cells(
            start_row=1,
            start_column=3 + 2 * t_idx,
            end_row=1,
            end_column=3 + 2 * t_idx + 1,
        )
        means_sheet.cell(row=1, column=3 + 2 * t_idx, value=time)
        means_sheet.cell(row=2, column=3 + 2 * t_idx, value="value")
        means_sheet.cell(row=2, column=3 + 2 * t_idx + 1, value="stddev")

    current_row = 3
    for m_id in data:
        measurement: d.Measurement
        measurement = data[m_id]

        (means, stddevs) = measurement.calc_means_and_stddevs()
        masses = [str(x) for x in measurement.get_masses()]
        for m_idx, mass in enumerate(masses):
            means_sheet.cell(row=current_row, column=1, value=measurement.mid)
            means_sheet.cell(row=current_row, column=2, value=mass)
            for t_idx, time in enumerate(times):
                if time in means:
                    means_t = means[time]
                    stddevs_t = stddevs[time]
                    means_sheet.cell(row=current_row, column=3 + 2 * t_idx, value=means_t.iloc[m_idx])
                    means_sheet.cell(
                        row=current_row,
                        column=3 + 2 * t_idx + 1,
                        value=stddevs_t.iloc[m_idx],
                    )
            current_row = current_row + 1
    pass


def export_to_fluxml(data):
    # TODO implement
    pass


def export_to_omix(data: dict[d.Measurement], workbook: openpyxl.workbook, sheet_basename: str):
    """
    write the corrected values to a sheet from where omix can import it directly

    :param data: a dict of MGroup with the corrected data
    :param workbook: excel workbook for data
    :param sheet_basename: The name of the sheet where the values go

    """
    omix_sheet = workbook.create_sheet(f"{sheet_basename}_corr_omix")
    # Now everything is in the m_data list
    current_row = 1
    measurement: d.Measurement
    for measurement in data:
        datum: d.Measurement = data[measurement]
        msms = datum.is_msms()
        start_col = 1
        # write header
        if datum.is_inst():
            omix_sheet.cell(
                row=current_row,
                column=1,
                value="; ".join([f"{x}" for x in datum.get_times()]),
            )
            start_col += 1
        omix_sheet.cell(row=current_row, column=start_col + 0, value=datum.build_fluxml_name())
        current_row += 1

        (means, stddevs) = datum.calc_means_and_stddevs_no_neg()
        # write data
        for time in datum.get_times():
            for mlane in datum.generate_possible_masses():
                if datum.is_inst():
                    omix_sheet.cell(row=current_row, column=1, value=time)

                # m_lane
                m_lane = "M"
                if msms:
                    # omix needs precursor - product (mother - daughter) masses separated by a comma
                    m_lane += f"({mlane.pre_shift},{mlane.pro_shift})"
                else:
                    m_lane += f"{mlane.pre_shift}"
                omix_sheet.cell(row=current_row, column=start_col + 0, value=m_lane)
                # value
                omix_sheet.cell(row=current_row, column=start_col + 1, value=means[time][mlane])
                # stddev
                omix_sheet.cell(row=current_row, column=start_col + 2, value=stddevs[time][mlane])
                current_row += 1
    pass
