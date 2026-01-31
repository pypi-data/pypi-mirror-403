#!/usr/bin/env python3

import argparse
import os

import numpy as np
from openpyxl import load_workbook

import unac.correction.registry as registry
import unac.util.compare as compare
import unac.util.data as d
import unac.util.exports as exports
import unac.util.imports as imports
import unac.util.plots as p
from unac.util.config import Config


def self_test():
    tools = registry.available_backends()
    print("Available correction tools:")
    for t in tools:
        print(f"  - {t.name}")


def do_correction(
    f_name,
    d_stddev,
    tools_in=None,
):
    available_tools = registry.available_backends()
    if tools_in is None:
        tools = available_tools
    else:
        tools = [t for t in available_tools if t in tools_in]

    d.Measurement.default_stddev = d_stddev
    file = f_name
    correction_tools = collect_correction_tools(tools)

    wb = load_workbook(filename=file)
    name_spec = os.path.splitext(file)[0]

    input_data = prepare_data(wb)

    for sheet_name in input_data:
        print(f"\ncorrecting data from sheet: {sheet_name}")
        out_path = name_spec + "/" + sheet_name
        os.makedirs(out_path, exist_ok=True)

        cor_data = {}
        for ct in correction_tools:
            cor_data[ct] = correction_tools[ct].perform_correction_ui(input_data[sheet_name], out_path)

        problematic = compare.compare_results(cor_data)
        if problematic:
            out = "".join([f"- {x} has {problematic[x][1]}: {problematic[x][0]:.2e}\n" for x in problematic])
            print(f"\nWARNING: There were issues with the following measurements \n{out}")
            print("!!! These measurements will not be exported to the tabular output !!!")
            print(f"Review the raw data and the plots in {out_path}/problematic/")
            print(
                f'You can also adjust the values for "{Config.TOL_NEG_KEY}" and "{Config.TOL_DIFF_KEY}" in the '
                f"configuration section [{Config.TOL_KEY}]"
            )

        p.plot_problematic(cor_data, problematic, out_path + "/problematic")
        consolidated_data = compare.filter_problematic(cor_data, problematic)

        cor_data["uncorrected"] = input_data[sheet_name]
        p.plot_data(
            cor_data,
            out_path + "/compare",
        )

        if np.any([x.is_inst() for x in input_data[sheet_name].values()]):
            p.plot_timeseries(consolidated_data, out_path + "/time_series")
            p.plot_timeseries(input_data[sheet_name], out_path + "/time_series/uncorrected")

        exports.export_to_omix(consolidated_data, wb, sheet_name)
        exports.export_to_excel(consolidated_data, wb, sheet_name)

    output_name = f"{name_spec}_cor.xlsx"
    wb.save(filename=output_name)


def collect_correction_tools(tools):
    correction_tools = {}
    for c in tools:
        correction_tools[c.name] = c
    return correction_tools


def prepare_data(wb):
    out = {}

    # look for sheets that have a good name (contain MID and MS, but not corr)
    for sheet_name in wb.sheetnames:
        if ("MID" in sheet_name) and ("MS" in sheet_name) and ("corr" not in sheet_name):
            # parse the raw data into files for ict
            (has_data, data) = imports.import_raw_data(wb, sheet_name)
            # were there any data in the sheet? has_data is True if MS(MS) measurement data was found in the spreadsheet
            if has_data:
                out[sheet_name] = data

    return out


def main():
    parser = argparse.ArgumentParser(
        prog="uNAC",
        description="Perform natural abundance correction on standardized spreadsheet",
    )
    parser.add_argument("filename", metavar="file.xlsx", nargs="?", help="spreadsheet file")
    parser.add_argument(
        "-s",
        action="store",
        dest="stddev",
        type=float,
        help="The minimal standard deviation of a mass trace",
        default=0.01,
        required=False,
    )
    parser.add_argument(
        "-c",
        metavar="config.toml",
        action="store",
        dest="config",
        type=str,
        help="config file",
        default=None,
        required=False,
    )
    parser.add_argument(
        "-t",
        "--test",
        action="store_true",
        help="test available backends and exit",
    )
    args = parser.parse_args()

    if args.test:
        self_test()
        exit(0)
    elif not args.filename:
        run_gui()
    else:
        if args.config:
            try:
                Config.parse_config(args.config)
            except ValueError as e:
                print(f"Failed to read the config file {args.config}. The following reason was given:\n{e}")
                exit(-1)
        do_correction(args.filename, args.stddev)


def run_gui():
    pass


if __name__ == "__main__":
    main()
