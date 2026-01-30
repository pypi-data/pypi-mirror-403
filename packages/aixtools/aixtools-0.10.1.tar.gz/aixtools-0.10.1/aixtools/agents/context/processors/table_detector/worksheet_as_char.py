from datetime import date, datetime

from openpyxl.worksheet.worksheet import Worksheet


class WorksheetAsChar:
    """Character grid representation of a worksheet.

    Each cell is represented as a single character:
    - ' ' for empty cells
    - 's' for string cells
    - 'f' for float cells
    - 'i' for integer cells
    - 'b' for boolean cells
    - 'd' for date/datetime cells
    - '?' for other types
    """

    def __init__(
        self,
        worksheet: Worksheet,
        collapse_whitespace_lines: bool = True,
        remove_trailing_empty: bool = True,
        include_headers: bool = True,
        collapse_repeated_lines: bool = True,
        min_repeat_threshold: int = 5,
    ):
        """Initialize character grid from worksheet.

        Args:
            worksheet: The worksheet to convert
            collapse_whitespace_lines: Replace lines with only whitespace by empty line
                (affects __str__ only)
            remove_trailing_empty: Remove all trailing empty lines (affects __str__ only)
            collapse_repeated_lines: Collapse sequences of repeated lines when count
                > min_repeat_threshold (affects __str__ only)
            min_repeat_threshold: Minimum number of consecutive repeated lines required
                for collapsing
        """
        self.worksheet = worksheet
        self.collapse_whitespace_lines = collapse_whitespace_lines
        self.remove_trailing_empty = remove_trailing_empty
        self.include_headers = include_headers
        self.collapse_repeated_lines = collapse_repeated_lines
        self.min_repeat_threshold = min_repeat_threshold
        self._lines = self._build_lines(worksheet)

    def __getitem__(self, key: int | tuple[int, int]) -> str:
        """Get line or character at position using 0-based indexing.

        Args:
            key: Either row_idx (int) to get entire line, or (row_idx, col_idx) tuple to get specific character

        Returns:
            Line string if key is int, character if key is tuple, or empty string/' ' if out of bounds
        """
        if isinstance(key, int):
            return self.get_line(key)

        row_idx, col_idx = key
        if row_idx < 0 or row_idx >= len(self._lines):
            return " "
        line = self._lines[row_idx]
        if col_idx < 0 or col_idx >= len(line):
            return " "
        return line[col_idx]

    def __len__(self) -> int:
        """Return number of lines in character grid."""
        return len(self._lines)

    def __str__(self) -> str:
        """Return string representation of character grid with line numbers and column headers."""
        return self.to_str(
            include_headers=self.include_headers,
            collapse_whitespace_lines=self.collapse_whitespace_lines,
            remove_trailing_empty=self.remove_trailing_empty,
            collapse_repeated_lines=self.collapse_repeated_lines,
        )

    def __repr__(self) -> str:
        return self.__str__()

    def __iter__(self):
        """Iterate over rows (lines)."""
        return iter(self._lines)

    def get_line(self, row_idx: int) -> str:
        """Get line at row index (0-based).

        Args:
            row_idx: Row index (0-based)

        Returns:
            Line string or empty string if out of bounds
        """
        if row_idx < 0 or row_idx >= len(self._lines):
            return ""
        return self._lines[row_idx]

    def get_lines(self) -> list[str]:
        """Return all lines as a list."""
        return self._lines.copy()

    def get_max_cols(self) -> int:
        """Return maximum number of columns across all lines."""
        return max(len(line) for line in self._lines) if self._lines else 0

    def iter_cols(self):
        """Iterate over columns (yields column strings from top to bottom)."""
        max_cols = self.get_max_cols()
        for col_idx in range(max_cols):
            col_chars = []
            for line in self._lines:
                if col_idx < len(line):
                    col_chars.append(line[col_idx])
                else:
                    col_chars.append(" ")
            yield "".join(col_chars)

    def size(self) -> tuple[int, int]:
        """Return (num_rows, num_cols)."""
        return (len(self._lines), self.get_max_cols())

    def get_slice(self, start_row: int, end_row: int, start_col: int, end_col: int) -> str:
        """Get substring from character grid region (0-based, inclusive).

        Args:
            start_row: Starting row index (0-based)
            end_row: Ending row index (0-based, inclusive)
            start_col: Starting column index (0-based)
            end_col: Ending column index (0-based, inclusive)

        Returns:
            String containing characters from specified region
        """
        result = []
        for row_idx in range(start_row, end_row + 1):
            if row_idx >= len(self._lines):
                break
            line = self._lines[row_idx]
            row_chars = line[start_col : end_col + 1] if start_col < len(line) else ""
            result.append(row_chars)
        return "\n".join(result)

    def to_str(
        self,
        include_headers: bool = False,
        collapse_whitespace_lines: bool | None = False,
        remove_trailing_empty: bool | None = False,
        collapse_repeated_lines: bool | None = False,
    ) -> str:
        """Return string representation of character grid with optional formatting.

        Args:
            include_headers: Include column headers and row numbers
            collapse_whitespace_lines: Replace lines with only whitespace by empty line (None uses instance setting)
            remove_trailing_empty: Remove all trailing empty lines (None uses instance setting)
            collapse_repeated_lines: Collapse sequences of repeated lines (None uses instance setting)

        Returns:
            String representation of character grid
        """
        lines = [(i, line) for i, line in enumerate(self._lines)]

        if collapse_whitespace_lines:
            lines = [(i, "" if line.strip() == "" else line) for i, line in lines]

        if collapse_repeated_lines:
            lines = self._collapse_repeated_lines(lines)

        if remove_trailing_empty:
            while lines and lines[-1][1] == "":
                lines.pop()

        if include_headers:
            return self._format_with_headers(lines)
        else:
            return "\n".join(line for _, line in lines)

    def _col_index_to_letter(self, col_idx: int) -> str:
        """Convert 0-based column index to Excel-style letter (A, B, ..., Z, AA, AB, etc.)."""
        result = ""
        col_num = col_idx + 1
        while col_num > 0:
            col_num -= 1
            result = chr(65 + (col_num % 26)) + result
            col_num //= 26
        return result

    def _format_column_header(self, max_cols: int) -> str:
        """Generate column header lines (two lines for columns beyond Z)."""
        letters = [self._col_index_to_letter(i) for i in range(max_cols)]
        max_len = max(len(letter) for letter in letters)

        if max_len == 1:
            return "   |" + "".join(letters)

        lines = []
        for line_idx in range(max_len):
            header = "   |"
            for letter in letters:
                char_idx = line_idx - (max_len - len(letter))
                header += letter[char_idx] if char_idx >= 0 else " "
            lines.append(header)
        return "\n".join(lines)

    def _format_with_headers(self, lines: list[tuple[int | None, str]]) -> str:
        """Format lines with row numbers and column headers.

        Args:
            lines: List of (row_idx, line) tuples where row_idx can be None for summary lines

        Returns:
            Formatted string with headers and line numbers
        """
        if not lines:
            return ""

        max_cols = max(len(line) for _, line in lines if _ is not None) if lines else 0

        if max_cols == 0:
            return ""

        result = []
        result.append(self._format_column_header(max_cols))
        result.append("   |" + "-" * max_cols)

        for row_idx, line in lines:
            if row_idx is None:
                result.append(f"   |{line}")
            else:
                padded_line = line.ljust(max_cols)
                result.append(f"{row_idx:03d}|{padded_line}")

        return "\n".join(result)

    def _build_lines(self, worksheet: Worksheet) -> list[str]:
        """Build character grid lines from worksheet (unaltered)."""
        if worksheet.max_row == 0 or worksheet.max_column == 0:
            return []

        lines = []
        for row_idx, row in enumerate(
            worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column), start=1
        ):
            row_chars = []
            for cell in row:
                value = cell.value

                if value is None or (isinstance(value, str) and value.strip() == ""):
                    char = " "
                elif isinstance(value, bool):
                    char = "b"
                elif isinstance(value, int):
                    char = "i"
                elif isinstance(value, float):
                    char = "f"
                elif isinstance(value, str):
                    char = "s"
                elif isinstance(value, (datetime, date)):
                    char = "d"
                else:
                    char = "?"

                row_chars.append(char)

            line = "".join(row_chars)
            lines.append(line)

        return lines

    def _collapse_repeated_lines(self, lines: list[tuple[int, str]]) -> list[tuple[int | None, str]]:
        """Collapse sequences of repeated lines into summary format when count > min_repeat_threshold.

        Args:
            lines: List of (row_idx, line) tuples to process

        Returns:
            List of tuples with repeated sequences collapsed
        """
        if not lines:
            return lines

        result = []
        i = 0

        while i < len(lines):
            current_idx, current_line = lines[i]
            repeat_count = 1

            while i + repeat_count < len(lines) and lines[i + repeat_count][1] == current_line:
                repeat_count += 1

            if repeat_count > self.min_repeat_threshold:
                context_lines = 2
                for j in range(context_lines):
                    result.append(lines[i + j])

                result.append((None, f"... (repeated {repeat_count - 2 * context_lines} times)"))

                for j in range(context_lines):
                    result.append(lines[i + repeat_count - context_lines + j])
            else:
                for j in range(repeat_count):
                    result.append(lines[i + j])

            i += repeat_count

        return result
