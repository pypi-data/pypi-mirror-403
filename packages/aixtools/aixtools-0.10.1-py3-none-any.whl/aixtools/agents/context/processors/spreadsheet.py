import io
from collections.abc import Callable
from pathlib import Path

from aixtools.agents.context.data_models import FileExtractionResult, FileType, TruncationInfo
from aixtools.agents.context.processors.common import (
    check_and_apply_output_limit,
    create_error_result,
    create_file_metadata,
    format_section_header,
    truncate_string,
)
from aixtools.agents.context.processors.table_detector.table_detector_char import TableDetectorCharGrid
from aixtools.utils import config


def _format_sheet_as_csv(rows: list[list], max_cell_length: int, max_line_length: int) -> tuple[str, int]:
    """Format sheet rows as CSV with cell and line truncation."""
    output = io.StringIO()
    cells_truncated = 0

    for row in rows:
        # Truncate each cell
        truncated_cells = []
        for cell in row:
            truncated_cell, was_truncated = truncate_string(cell, max_cell_length)
            truncated_cells.append(truncated_cell)
            if was_truncated:
                cells_truncated += 1

        # Create CSV line
        line = ",".join(f'"{cell}"' if "," in cell or '"' in cell else cell for cell in truncated_cells)

        # Truncate line if needed
        if len(line) > max_line_length:
            line = line[:max_line_length] + "..."

        output.write(line + "\n")

    return output.getvalue(), cells_truncated


def _process_excel_file(
    file_path: Path,
    max_sheets: int,
    max_rows_head: int,
    max_rows_tail: int,
    max_columns: int,
    max_cell_length: int,
    max_line_length: int,
    max_total_output: int,
) -> tuple[str, TruncationInfo]:
    """Process Excel file (.xlsx) using openpyxl."""
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required for Excel file processing. Install with: uv add openpyxl")

    # Load workbook and detect tables
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet_names = workbook.sheetnames
    total_sheets = len(sheet_names)
    sheets_to_process = min(total_sheets, max_sheets)

    output = io.StringIO()
    output.write(f"Spreadsheet: {file_path.name}\n")
    output.write(f"Sheets: {sheets_to_process} of {total_sheets}\n\n")

    total_cells_truncated = 0
    total_columns_truncated = 0
    total_rows_shown = 0
    total_rows_available = 0

    # Detect tables using character grid analysis
    detector = TableDetectorCharGrid(workbook)
    detected_tables = detector.detect()

    for i in range(sheets_to_process):
        sheet_name = sheet_names[i]
        sheet = workbook[sheet_name]

        # Add sheet header
        output.write(format_section_header(f"{sheet_name}", i, total_sheets))

        # Process each detected table in this sheet
        sheet_tables = [t for t in detected_tables if t.sheet_name == sheet_name]
        if not sheet_tables:
            output.write("(no tables detected)\n")
            continue

        for table_idx, table in enumerate(sheet_tables):
            if len(sheet_tables) > 1:
                output.write(f"\n--- Table {table_idx + 1} of {len(sheet_tables)} ---\n")

            # Extract table data
            table_rows = []
            for row_idx in range(table.start_row, table.end_row + 1):
                row_data = []
                for col_idx in range(table.start_col, table.end_col + 1):
                    cell = sheet.cell(row_idx, col_idx)
                    row_data.append(cell.value)
                table_rows.append(row_data)

            total_rows = len(table_rows) - 1 if table.header_row else len(table_rows)
            total_columns = len(table_rows[0]) if table_rows else 0

            # Apply row truncation (head + tail)
            if table.header_row:
                header_offset = table.header_row - table.start_row
                header = table_rows[header_offset] if header_offset < len(table_rows) else []
                data_rows = table_rows[header_offset + 1 :] if header_offset + 1 < len(table_rows) else []

                if len(data_rows) <= (max_rows_head + max_rows_tail):
                    selected_rows = [header] + data_rows
                else:
                    head_rows = data_rows[:max_rows_head]
                    tail_rows = data_rows[-max_rows_tail:] if max_rows_tail > 0 else []
                    selected_rows = [header] + head_rows + tail_rows
            elif len(table_rows) <= (max_rows_head + max_rows_tail):
                selected_rows = table_rows
            else:
                head_rows = table_rows[:max_rows_head]
                tail_rows = table_rows[-max_rows_tail:] if max_rows_tail > 0 else []
                selected_rows = head_rows + tail_rows

            # Apply column truncation
            columns_to_show = min(total_columns, max_columns)
            truncated_rows = [row[:columns_to_show] for row in selected_rows]

            # Format as CSV
            csv_content, cells_truncated = _format_sheet_as_csv(truncated_rows, max_cell_length, max_line_length)
            total_cells_truncated += cells_truncated

            # Write metadata
            rows_shown = len(truncated_rows) - 1 if table.header_row else len(truncated_rows)
            output.write(f"Columns: {columns_to_show} (of {total_columns} total)\n")
            output.write(f"Rows: {rows_shown} (of {total_rows} total)\n")
            if table.header_row:
                output.write(f"Header row: {table.header_row}\n")
            output.write("\n")

            # Write data
            output.write(csv_content)

            total_columns_truncated += (total_columns - columns_to_show) if total_columns > columns_to_show else 0
            total_rows_shown += rows_shown
            total_rows_available += total_rows

            # Check output limit
            if output.tell() > max_total_output:
                output.write("\n...\n")
                break

        if output.tell() > max_total_output:
            break

    workbook.close()

    # Build truncation info
    truncation_info = TruncationInfo(
        cells_truncated=total_cells_truncated, total_output_limit_reached=output.tell() > max_total_output
    )

    if sheets_to_process < total_sheets:
        truncation_info.rows_shown = f"{sheets_to_process} sheets of {total_sheets} total"

    return output.getvalue(), truncation_info


def _process_xls_file(
    file_path: Path,
    max_sheets: int,
    max_rows_head: int,
    max_rows_tail: int,
    max_columns: int,
    max_cell_length: int,
    max_line_length: int,
    max_total_output: int,
) -> tuple[str, TruncationInfo]:
    """Process legacy Excel file (.xls) using xlrd."""
    try:
        import xlrd
    except ImportError:
        raise ImportError("xlrd is required for .xls file processing. Install with: uv add xlrd")

    workbook = xlrd.open_workbook(file_path)
    sheet_names = workbook.sheet_names()
    total_sheets = len(sheet_names)
    sheets_to_process = min(total_sheets, max_sheets)

    output = io.StringIO()
    output.write(f"Spreadsheet: {file_path.name}\n")
    output.write(f"Sheets: {sheets_to_process} of {total_sheets}\n\n")

    total_cells_truncated = 0
    total_columns_truncated = 0
    total_rows_shown = 0
    total_rows_available = 0

    for i in range(sheets_to_process):
        sheet_name = sheet_names[i]
        sheet = workbook.sheet_by_name(sheet_name)

        # Add sheet header
        output.write(format_section_header(f"{sheet_name}", i, total_sheets))

        if sheet.nrows == 0:
            output.write("(empty sheet)\n")
            continue

        # Read rows
        total_rows = sheet.nrows - 1  # Exclude header
        total_columns = sheet.ncols
        columns_to_show = min(total_columns, max_columns)

        # Build header
        header = [str(sheet.cell_value(0, col)) for col in range(columns_to_show)]

        # Select rows (head + tail pattern)
        if total_rows <= (max_rows_head + max_rows_tail):
            row_indices = list(range(1, sheet.nrows))
        else:
            head_indices = list(range(1, max_rows_head + 1))
            tail_indices = list(range(sheet.nrows - max_rows_tail, sheet.nrows)) if max_rows_tail > 0 else []
            row_indices = head_indices + tail_indices

        # Build rows
        rows = [header]
        for row_idx in row_indices:
            row = [str(sheet.cell_value(row_idx, col)) for col in range(columns_to_show)]
            rows.append(row)

        # Format as CSV
        csv_content, cells_truncated = _format_sheet_as_csv(rows, max_cell_length, max_line_length)
        total_cells_truncated += cells_truncated

        # Write metadata
        rows_shown = len(row_indices)
        output.write(f"Columns: {columns_to_show} (of {total_columns} total)\n")
        output.write(f"Rows: {rows_shown} (of {total_rows} total)\n\n")

        # Write data
        output.write(csv_content)

        total_columns_truncated += (total_columns - columns_to_show) if total_columns > columns_to_show else 0
        total_rows_shown += rows_shown
        total_rows_available += total_rows

        # Check output limit
        if output.tell() > max_total_output:
            output.write("\n...\n")
            break

    # Build truncation info
    truncation_info = TruncationInfo(
        cells_truncated=total_cells_truncated, total_output_limit_reached=output.tell() > max_total_output
    )

    if sheets_to_process < total_sheets:
        truncation_info.rows_shown = f"{sheets_to_process} sheets of {total_sheets} total"

    return output.getvalue(), truncation_info


def _process_ods_file(
    file_path: Path,
    max_sheets: int,
    max_rows_head: int,
    max_rows_tail: int,
    max_columns: int,
    max_cell_length: int,
    max_line_length: int,
    max_total_output: int,
) -> tuple[str, TruncationInfo]:
    """Process OpenDocument Spreadsheet (.ods) using odfpy or pandas."""
    try:
        import pandas as pd
    except ImportError:
        raise ImportError("pandas is required for ODS file processing. Install with: uv add pandas odfpy")

    # Read all sheets
    all_sheets = pd.read_excel(file_path, sheet_name=None, engine="odf")
    sheet_names = list(all_sheets.keys())
    total_sheets = len(sheet_names)
    sheets_to_process = min(total_sheets, max_sheets)

    output = io.StringIO()
    output.write(f"Spreadsheet: {file_path.name}\n")
    output.write(f"Sheets: {sheets_to_process} of {total_sheets}\n\n")

    total_cells_truncated = 0

    for i, sheet_name in enumerate(sheet_names[:sheets_to_process]):
        df = all_sheets[sheet_name]

        # Add sheet header
        output.write(format_section_header(f"{sheet_name}", i, total_sheets))

        if df.empty:
            output.write("(empty sheet)\n")
            continue

        # Truncate columns
        total_columns = len(df.columns)
        columns_to_show = min(total_columns, max_columns)
        df_truncated = df.iloc[:, :columns_to_show]

        # Truncate rows (head + tail)
        total_rows = len(df)
        if total_rows <= (max_rows_head + max_rows_tail):
            df_selected = df_truncated
        else:
            df_head = df_truncated.head(max_rows_head)
            df_tail = df_truncated.tail(max_rows_tail) if max_rows_tail > 0 else pd.DataFrame()
            df_selected = pd.concat([df_head, df_tail])

        rows_shown = len(df_selected)

        # Write metadata
        output.write(f"Columns: {columns_to_show} (of {total_columns} total)\n")
        output.write(f"Rows: {rows_shown} (of {total_rows} total)\n\n")

        # Convert to list of lists for CSV formatting
        header = list(df_selected.columns)
        data_rows = df_selected.values.tolist()
        all_rows = [header] + data_rows

        # Format as CSV with truncation
        csv_content, cells_truncated = _format_sheet_as_csv(all_rows, max_cell_length, max_line_length)
        total_cells_truncated += cells_truncated

        output.write(csv_content)

        # Check output limit
        if output.tell() > max_total_output:
            output.write("\n...\n")
            break

    # Build truncation info
    truncation_info = TruncationInfo(
        cells_truncated=total_cells_truncated, total_output_limit_reached=output.tell() > max_total_output
    )

    if sheets_to_process < total_sheets:
        truncation_info.rows_shown = f"{sheets_to_process} sheets of {total_sheets} total"

    return output.getvalue(), truncation_info


def process_spreadsheet(
    file_path: Path,
    max_sheets: int = 3,
    max_rows_head: int = config.DEFAULT_ROWS_HEAD,
    max_rows_tail: int = config.DEFAULT_ROWS_TAIL,
    max_columns: int = config.MAX_COLUMNS,
    max_cell_length: int = config.MAX_CELL_LENGTH,
    max_line_length: int = config.MAX_LINE_LENGTH,
    max_total_output: int = config.MAX_TOTAL_OUTPUT,
    tokenizer: Callable | None = None,
) -> FileExtractionResult:
    """Process spreadsheet files (.xlsx, .xls, .ods).

    .xlsx files are processed with intelligent table detection to identify and separate
    multiple tables within sheets. .xls and .ods files use simpler row/column processing.

    Args:
        file_path: Path to spreadsheet file
        max_sheets: Maximum number of sheets to process
        max_rows_head: Maximum rows from start of each sheet
        max_rows_tail: Maximum rows from end of each sheet
        max_columns: Maximum columns per sheet
        max_cell_length: Maximum characters per cell
        max_line_length: Maximum characters per line
        max_total_output: Maximum total output length
        tokenizer: Optional tokenizer function
    """
    try:
        metadata = create_file_metadata(file_path, mime_type=f"spreadsheet/{file_path.suffix[1:]}")

        # Process based on file type
        suffix = file_path.suffix.lower()
        if suffix == ".xlsx":
            content, truncation_info = _process_excel_file(
                file_path,
                max_sheets,
                max_rows_head,
                max_rows_tail,
                max_columns,
                max_cell_length,
                max_line_length,
                max_total_output,
            )
        elif suffix == ".xls":
            content, truncation_info = _process_xls_file(
                file_path,
                max_sheets,
                max_rows_head,
                max_rows_tail,
                max_columns,
                max_cell_length,
                max_line_length,
                max_total_output,
            )
        elif suffix == ".ods":
            content, truncation_info = _process_ods_file(
                file_path,
                max_sheets,
                max_rows_head,
                max_rows_tail,
                max_columns,
                max_cell_length,
                max_line_length,
                max_total_output,
            )
        else:
            return FileExtractionResult(
                content=None,
                success=False,
                error_message=f"Unsupported spreadsheet format: {suffix}",
                file_type=FileType.SPREADSHEET,
                metadata=metadata,
            )

        # Apply tokenizer if provided
        if tokenizer:
            truncation_info.tokens_shown = tokenizer(content)

        # Final length check
        content = check_and_apply_output_limit(content, max_total_output, truncation_info)

        return FileExtractionResult(
            content=content,
            success=True,
            was_extracted=True,
            file_type=FileType.SPREADSHEET,
            truncation_info=truncation_info,
            metadata=metadata,
        )

    except ImportError as e:
        return create_error_result(e, FileType.SPREADSHEET, file_path, "spreadsheet (missing dependencies)")
    except Exception as e:
        return create_error_result(e, FileType.SPREADSHEET, file_path, "spreadsheet")
