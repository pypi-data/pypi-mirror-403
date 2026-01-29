"""Excel Exporter - Clean Architecture з DTO.

BREAKING CHANGE (Фаза 6): Тепер використовує GraphDTO замість Graph.

Exports graph to Excel format with formatting and multiple sheets.
"""

import logging
import time
from pathlib import Path
from typing import Optional

from graph_crawler.application.dto import GraphDTO, NodeDTO, EdgeDTO
from graph_crawler.application.services.exporters.base_exporter import BaseExporter
from graph_crawler.domain.events.event_bus import EventBus
from graph_crawler.domain.events.events import EventType

try:
    import openpyxl
    import pandas as pd
    from openpyxl.styles import Alignment, Font, PatternFill

    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

logger = logging.getLogger(__name__)


class ExcelExporter(BaseExporter):
    """
    Export graph to Excel format з форматуванням через DTO.

    BREAKING CHANGE (Фаза 6): Тепер працює ТІЛЬКИ з GraphDTO.

    Features:
    - Multiple sheets (Nodes, Edges, Summary)
    - Header formatting (bold, colored background)
    - Auto column width
    - Data types preserved
    - Filters на headers

    Requirements:
        pip install pandas openpyxl

    Example:
        >>> from graph_crawler.application.services.exporters import ExcelExporter
        >>> from graph_crawler.application.dto import GraphDTO
        >>>
        >>> exporter = ExcelExporter()
        >>> exporter.export(graph_dto, './output/graph.xlsx')
    """

    def __init__(self, event_bus: Optional["EventBus"] = None, **kwargs):
        """
        Initialize Excel exporter.

        Args:
            event_bus: EventBus для публікації подій (опціонально)
            **kwargs: Additional options

        Raises:
            ImportError: Якщо pandas або openpyxl не встановлено
        """
        super().__init__(event_bus=event_bus, **kwargs)
        if not PANDAS_AVAILABLE:
            raise ImportError(
                "pandas and openpyxl are required for ExcelExporter. "
                "Install with: pip install pandas openpyxl"
            )

    def export(
        self, graph_dto: GraphDTO, output_path: str, include_summary: bool = True, **options
    ) -> bool:
        """
        Export graph to Excel file through DTO.

        BREAKING CHANGE (Фаза 6): Тепер приймає GraphDTO замість Graph.

        Args:
            graph_dto: GraphDTO для експорту
            output_path: Path до output Excel file
            include_summary: Включити summary sheet
            **options: Additional options

        Returns:
            bool: True якщо успішно

        Raises:
            Exception: При помилках запису
        """
        start_time = time.time()

        self.publish_event(
            EventType.EXPORT_STARTED,
            data={
                "exporter": "excel",
                "output_path": output_path,
                "nodes_count": len(graph_dto.nodes),
                "edges_count": len(graph_dto.edges),
            },
        )

        try:
            if not self.validate_graph(graph_dto):
                logger.warning("GraphDTO validation failed, exporting empty graph")

            # Створити output directory
            output_dir = Path(output_path).parent
            output_dir.mkdir(parents=True, exist_ok=True)

            # Створити Excel writer
            with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
                # Export nodes
                nodes_df = self._create_nodes_dataframe(graph_dto)
                nodes_df.to_excel(writer, sheet_name="Nodes", index=False)

                # Export edges
                edges_df = self._create_edges_dataframe(graph_dto)
                edges_df.to_excel(writer, sheet_name="Edges", index=False)

                # Export summary
                if include_summary:
                    summary_df = self._create_summary_dataframe(graph_dto)
                    summary_df.to_excel(writer, sheet_name="Summary", index=False)

                # Apply formatting
                self._apply_formatting(writer.book)

            duration = time.time() - start_time

            self.publish_event(
                EventType.EXPORT_SUCCESS,
                data={
                    "exporter": "excel",
                    "output_path": output_path,
                    "nodes_count": len(graph_dto.nodes),
                    "edges_count": len(graph_dto.edges),
                    "duration": round(duration, 3),
                },
            )

            logger.info(f"Exported graph to {output_path}")
            return True

        except Exception as e:
            error_msg = f"Failed to export graph to Excel: {e}"
            logger.error(error_msg)

            self.publish_event(
                EventType.EXPORT_ERROR,
                data={
                    "exporter": "excel",
                    "error": error_msg,
                    "error_type": type(e).__name__,
                    "output_path": output_path,
                },
            )

            raise

    def _create_nodes_dataframe(self, graph_dto: GraphDTO) -> "pd.DataFrame":
        """
        Create pandas DataFrame для nodes from GraphDTO.

        Args:
            graph_dto: GraphDTO instance

        Returns:
            pd.DataFrame: Nodes data
        """
        nodes_data = []
        for node_dto in graph_dto.nodes:
            node_dict = {
                "URL": node_dto.url,
                "Depth": node_dto.depth,
                "Title": node_dto.metadata.get("title", ""),
                "Status Code": node_dto.response_status or "",
                "Content Type": node_dto.metadata.get("content_type", ""),
                "Scanned": node_dto.scanned,
                "Node ID": node_dto.node_id,
            }

            # Додати додаткові metadata
            for key, value in node_dto.metadata.items():
                if key not in ["title", "status_code", "content_type"]:
                    node_dict[key.replace("_", " ").title()] = value

            nodes_data.append(node_dict)

        return pd.DataFrame(nodes_data)

    def _create_edges_dataframe(self, graph_dto: GraphDTO) -> "pd.DataFrame":
        """
        Create pandas DataFrame для edges from GraphDTO.

        Args:
            graph_dto: GraphDTO instance

        Returns:
            pd.DataFrame: Edges data
        """
        nodes_by_id = {node.node_id: node for node in graph_dto.nodes}

        edges_data = []
        for edge_dto in graph_dto.edges:
            source_node = nodes_by_id.get(edge_dto.source_node_id)
            target_node = nodes_by_id.get(edge_dto.target_node_id)

            edges_data.append(
                {
                    "Source URL": source_node.url if source_node else "",
                    "Target URL": target_node.url if target_node else "",
                    "Edge ID": edge_dto.edge_id,
                    "Link Type": str(edge_dto.metadata.get("link_type", [])),
                }
            )

        return pd.DataFrame(edges_data)

    def _create_summary_dataframe(self, graph_dto: GraphDTO) -> "pd.DataFrame":
        """
        Create summary DataFrame з статистикою from GraphDTO.

        Args:
            graph_dto: GraphDTO instance

        Returns:
            pd.DataFrame: Summary data
        """
        stats = graph_dto.stats

        summary_data = [
            {"Metric": "Total Nodes", "Value": stats.total_nodes},
            {"Metric": "Scanned Nodes", "Value": stats.scanned_nodes},
            {"Metric": "Unscanned Nodes", "Value": stats.unscanned_nodes},
            {"Metric": "Total Edges", "Value": stats.total_edges},
            {"Metric": "Average Depth", "Value": round(stats.avg_depth, 2)},
            {"Metric": "Max Depth", "Value": stats.max_depth},
        ]

        return pd.DataFrame(summary_data)

    def _apply_formatting(self, workbook: "openpyxl.Workbook"):
        """
        Apply formatting до Excel workbook.

        Args:
            workbook: openpyxl Workbook instance
        """
        # Header formatting
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Format header row
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

            # Auto-adjust column width
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter

                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)  # Max 50 chars
                sheet.column_dimensions[column_letter].width = adjusted_width

            # Add filters
            if sheet.max_row > 1:
                sheet.auto_filter.ref = sheet.dimensions
