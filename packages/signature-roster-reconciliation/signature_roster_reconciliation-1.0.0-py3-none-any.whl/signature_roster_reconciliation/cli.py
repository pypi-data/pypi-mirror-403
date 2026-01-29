#!/usr/bin/env python3
"""
Signature Locker - Roster vs Orders Reconciliation Script

This script compares roster data from Program Directors against orders received
to identify common issues:
1. Players who have not ordered yet
2. Players who have ordered more than once (duplicates)
3. Players who ordered for the wrong team
4. Potential misspellings of player last names

Usage:
    python roster_order_reconciliation.py <roster_file> <orders_file> [--roster-sheet SHEET] [--output OUTPUT]

Example:
    python roster_order_reconciliation.py roster.xlsx orders.xlsx --output reconciliation_report.xlsx
"""

import pandas as pd
import numpy as np
import argparse
import sys
from pathlib import Path
from difflib import SequenceMatcher
from collections import defaultdict
from datetime import datetime
import re
import unicodedata
import unicodedata


# ============================================================================
# CONFIGURATION - Column name mappings for flexible roster formats
# ============================================================================

# Possible column names for each key field in roster data
ROSTER_COLUMN_MAPPINGS = {
    'player_name': [
        'Player Name', 'player_name', 'Player', 'Name', 'Athlete Name',
        'Child Name', 'Student Name', 'Participant Name'
    ],
    'player_first': [
        'First', 'First Name', 'Player First Name', 'Athlete First Name',
        'FirstName', 'Given Name'
    ],
    'player_last': [
        'Last', 'Last Name', 'Player Last Name', 'Athlete Last Name',
        'LastName', 'Surname', 'Family Name'
    ],
    'team_name': [
        'Team Name', 'team_name', 'Team', 'Squad', 'Division', 'Group'
    ],
    'parent_email_1': [
        'Parent 1 Email', 'Parent Email 1', 'parent_email_1', 'Email 1',
        'Primary Email', 'Email', 'Parent Email', 'Guardian Email',
        'Mother Email', 'Contact Email'
    ],
    'parent_email_2': [
        'Parent 2 Email', 'Parent Email 2', 'parent_email_2', 'Email 2',
        'Secondary Email', 'Father Email', 'Alt Email', 'Alternate Email'
    ],
    'parent_name_1': [
        'Parent Name 1 Name', 'Parent 1 Name', 'Parent Name 1', 'parent_name_1',
        'Primary Parent', 'Parent', 'Guardian', 'Mother Name'
    ],
    'parent_name_2': [
        'Parent Name 2', 'Parent 2 Name', 'parent_name_2',
        'Secondary Parent', 'Father Name'
    ],
    'uniform_number': [
        'Uniform Number', 'Jersey Number', 'Number', 'Assigned Number',
        'Player Number', 'Jersey #', 'Jersey', 'Uniform #', 'Uni Number',
        'Kit Number', 'Squad Number', '#'
    ]
}

# Standard column names in Program Director Report (orders)
ORDERS_COLUMNS = {
    'email': 'Order Email',
    'player_first': "Property 'Player First Name'",
    'player_last': "Property 'Player Last Name'",
    'team': "Property 'Selected Team'",
    'order_number': 'Order number',
    'order_date': 'DAY Order Date',
    'billing_first': 'Order Billing First Name',
    'billing_last': 'Order Billing Last Name',
    'piece_1_size': 'Piece 1 Size',
    'piece_2_size': 'Piece 2 Size',
    'piece_3_size': 'Piece 3 Size',
    'uniform_number': "Property 'Uniform Number'",
    'product_name': 'Product name',
    'category_tag': 'Order Category Tag'
}

# Patterns that indicate an actual uniform order (vs apparel/accessories)
UNIFORM_PRODUCT_PATTERNS = [
    'player package',
    'pinnie',
    'uniform package',
    'jersey package',
    'kit package'
]

# Standard size ordering for comparison
SIZE_ORDER = {
    'yxs': 1, 'ys': 2, 'ym': 3, 'yl': 4, 'yxl': 5, 'yxxl': 6,
    'xs': 10, 's': 11, 'm': 12, 'l': 13, 'xl': 14, 'xxl': 15, '2xl': 15, 'xxxl': 16, '3xl': 16,
    # Additional variations
    'youth xs': 1, 'youth s': 2, 'youth m': 3, 'youth l': 4, 'youth xl': 5,
    'adult xs': 10, 'adult s': 11, 'adult m': 12, 'adult l': 13, 'adult xl': 14,
    'axs': 10, 'as': 11, 'am': 12, 'al': 13, 'axl': 14, 'axxl': 15,
    # Hat sizes
    's/m': 20, 'l/xl': 21, 'sm': 20, 'lxl': 21
}


# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================

def normalize_email(email):
    """Normalize email for comparison - lowercase and strip whitespace."""
    if pd.isna(email) or email in ['(blank)', '']:
        return None
    return str(email).lower().strip()


def is_valid_email(email):
    """Check if email has a valid format."""
    if not email:
        return False
    # Basic email regex pattern
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return bool(re.match(pattern, str(email).strip()))


def normalize_unicode(text):
    """
    Normalize unicode characters to their closest ASCII equivalents.
    Handles curly quotes, special apostrophes, accented characters, etc.
    
    Examples:
        O'Brien -> O'Brien (curly apostrophe to straight)
        Müller -> Muller
        café -> cafe
    """
    if not text or pd.isna(text):
        return text
    
    text = str(text)
    
    # First, normalize unicode to decomposed form (NFD)
    # This separates base characters from combining marks
    text = unicodedata.normalize('NFD', text)
    
    # Remove combining marks (accents, etc.) but keep base characters
    text = ''.join(char for char in text if unicodedata.category(char) != 'Mn')
    
    # Normalize back to composed form
    text = unicodedata.normalize('NFC', text)
    
    # Replace common unicode variants with ASCII equivalents
    unicode_replacements = {
        ''': "'",   # Right single quotation mark
        ''': "'",   # Left single quotation mark
        '"': '"',   # Left double quotation mark
        '"': '"',   # Right double quotation mark
        '–': '-',   # En dash
        '—': '-',   # Em dash
        '…': '...',  # Ellipsis
        '\u00A0': ' ',  # Non-breaking space
        '\u2019': "'",  # Another right single quotation mark
        '\u2018': "'",  # Another left single quotation mark
    }
    
    for unicode_char, ascii_char in unicode_replacements.items():
        text = text.replace(unicode_char, ascii_char)
    
    return text


# Common name suffixes to remove for comparison
NAME_SUFFIXES = [
    r'\s+jr\.?$', r'\s+sr\.?$', r'\s+ii$', r'\s+iii$', r'\s+iv$', r'\s+v$',
    r'\s+2nd$', r'\s+3rd$', r'\s+4th$', r'\s+5th$',
    r'\s+junior$', r'\s+senior$'
]


def remove_name_suffix(name):
    """Remove common name suffixes (Jr., III, IV, etc.) for comparison."""
    if not name:
        return name
    name_lower = name.lower().strip()
    for suffix_pattern in NAME_SUFFIXES:
        name_lower = re.sub(suffix_pattern, '', name_lower, flags=re.IGNORECASE)
    return name_lower.strip()


def normalize_name(name):
    """Normalize name for comparison - lowercase, strip, remove extra spaces, normalize unicode."""
    if pd.isna(name) or name in ['(blank)', '']:
        return None
    name = str(name)
    # Normalize unicode first
    name = normalize_unicode(name)
    name = name.lower().strip()
    name = re.sub(r'\s+', ' ', name)  # Collapse multiple spaces
    return name


def normalize_last_name(last_name):
    """Normalize last name - lowercase, remove suffixes, strip, normalize unicode."""
    if pd.isna(last_name) or last_name in ['(blank)', '']:
        return None
    name = str(last_name)
    # Normalize unicode first
    name = normalize_unicode(name)
    name = name.lower().strip()
    name = remove_name_suffix(name)
    name = re.sub(r'\s+', ' ', name)  # Collapse multiple spaces
    # Also remove hyphens for comparison (treat "Haines-Connor" same as "Haines Connor")
    name = name.replace('-', ' ').strip()
    return name


def normalize_team_name(team_name):
    """
    Normalize team name for comparison.
    - Lowercase
    - Remove extra whitespace
    - Normalize unicode
    - Remove common prefixes/suffixes that don't affect team identity
    """
    if pd.isna(team_name) or not team_name:
        return None
    
    name = str(team_name).strip()
    name = normalize_unicode(name)
    name = name.lower()
    name = re.sub(r'\s+', ' ', name)  # Collapse multiple spaces
    
    return name


def get_team_words(team_name):
    """
    Extract meaningful words from a team name for comparison.
    Removes common filler words and normalizes the result.
    """
    if not team_name:
        return set()
    
    normalized = normalize_team_name(team_name)
    if not normalized:
        return set()
    
    # Replace common separators with spaces
    normalized = normalized.replace('/', ' ').replace('-', ' ').replace('_', ' ')
    
    # Split into words
    words = normalized.split()
    
    # Remove very common filler words that don't identify the team
    filler_words = {'team', 'the', 'of', 'and', '&', 'a', 'an'}
    
    # Keep words that are meaningful
    meaningful_words = set()
    for word in words:
        # Clean the word
        word = re.sub(r'[^\w\d]', '', word)  # Remove punctuation
        if word and word not in filler_words and len(word) > 1:
            meaningful_words.add(word)
    
    return meaningful_words


def calculate_team_similarity(team1, team2):
    """
    Calculate similarity between two team names using multiple methods.
    Returns a score from 0 to 1.
    
    Uses:
    1. Word overlap (Jaccard similarity)
    2. String similarity (for catching typos)
    3. Containment check (one name contains the other)
    """
    if not team1 or not team2:
        return 0.0
    
    norm1 = normalize_team_name(team1)
    norm2 = normalize_team_name(team2)
    
    if not norm1 or not norm2:
        return 0.0
    
    # Exact match after normalization
    if norm1 == norm2:
        return 1.0
    
    # Get word sets
    words1 = get_team_words(team1)
    words2 = get_team_words(team2)
    
    if not words1 or not words2:
        # Fall back to string similarity if we can't extract words
        return similarity_score(norm1, norm2)
    
    # Calculate Jaccard similarity (intersection over union)
    intersection = words1 & words2
    union = words1 | words2
    jaccard = len(intersection) / len(union) if union else 0
    
    # Calculate containment (what fraction of smaller set is in larger)
    smaller = words1 if len(words1) <= len(words2) else words2
    containment = len(intersection) / len(smaller) if smaller else 0
    
    # Calculate string similarity on normalized names
    string_sim = similarity_score(norm1, norm2)
    
    # Use the highest score, weighted towards word-based matching
    # This handles cases like "2027 Elite" vs "2027 Elite CLT" well
    # (high containment even if Jaccard is lower)
    score = max(
        jaccard,
        containment * 0.95,  # Slight penalty for containment-only matches
        string_sim * 0.9     # Slight penalty for string-only matches
    )
    
    return score


def teams_are_equivalent(roster_team, order_team, threshold=0.90):
    """
    Determine if two team names refer to the same team.
    
    Args:
        roster_team: Team name from roster
        order_team: Team name from order
        threshold: Minimum similarity score to consider teams equivalent (default 0.90)
    
    Returns:
        bool: True if teams are likely the same
    """
    similarity = calculate_team_similarity(roster_team, order_team)
    return similarity >= threshold


def split_player_name(full_name):
    """Split a full player name into first and last name."""
    if not full_name:
        return None, None
    parts = full_name.strip().split()
    if len(parts) == 1:
        return parts[0], ''
    elif len(parts) == 2:
        return parts[0], parts[1]
    else:
        # Assume first word is first name, rest is last name
        return parts[0], ' '.join(parts[1:])


def similarity_score(str1, str2):
    """Calculate similarity score between two strings (0-1)."""
    if not str1 or not str2:
        return 0.0
    return SequenceMatcher(None, str1.lower(), str2.lower()).ratio()


def find_column(df, possible_names):
    """Find a column in dataframe matching any of the possible names."""
    for col_name in possible_names:
        if col_name in df.columns:
            return col_name
        # Also try case-insensitive match
        for actual_col in df.columns:
            if str(actual_col).lower().strip() == col_name.lower().strip():
                return actual_col
    return None


def find_roster_sheet(excel_file):
    """
    Automatically detect the roster sheet in an Excel file.
    Looks for sheets with key roster columns.
    """
    xls = pd.ExcelFile(excel_file)
    
    best_sheet = None
    best_score = 0
    
    for sheet_name in xls.sheet_names:
        try:
            df = pd.read_excel(xls, sheet_name=sheet_name, nrows=5)
            
            # Score based on how many roster columns are found
            score = 0
            for field, possible_names in ROSTER_COLUMN_MAPPINGS.items():
                if find_column(df, possible_names):
                    score += 1
            
            # Prefer sheets with more rows (actual data)
            if score > 0:
                full_df = pd.read_excel(xls, sheet_name=sheet_name)
                row_count = len(full_df.dropna(how='all'))
                if row_count > 10:  # Has meaningful data
                    score += 1
            
            if score > best_score:
                best_score = score
                best_sheet = sheet_name
                
        except Exception:
            continue
    
    return best_sheet


# ============================================================================
# DATA LOADING FUNCTIONS
# ============================================================================

def load_roster_data(filepath, sheet_name=None):
    """
    Load roster data from Excel file, automatically detecting column names.
    
    Returns a tuple: (DataFrame, has_uniform_numbers)
    DataFrame has columns:
    - player_name, player_first, player_last
    - team_name
    - parent_email_1, parent_email_2
    - parent_name_1, parent_name_2
    - uniform_number (if present in roster)
    """
    print(f"\n{'='*60}")
    print("LOADING ROSTER DATA")
    print(f"{'='*60}")
    
    # Auto-detect sheet if not specified
    if sheet_name is None:
        sheet_name = find_roster_sheet(filepath)
        if sheet_name:
            print(f"Auto-detected roster sheet: '{sheet_name}'")
        else:
            # Try first sheet as fallback
            xls = pd.ExcelFile(filepath)
            sheet_name = xls.sheet_names[0]
            print(f"Using first sheet: '{sheet_name}'")
    
    df = pd.read_excel(filepath, sheet_name=sheet_name)
    print(f"Loaded {len(df)} rows from '{sheet_name}'")
    
    # Map columns to standard names
    column_mapping = {}
    for field, possible_names in ROSTER_COLUMN_MAPPINGS.items():
        found_col = find_column(df, possible_names)
        if found_col:
            column_mapping[found_col] = field
            print(f"  Mapped '{found_col}' -> '{field}'")
    
    # Create standardized dataframe
    roster = pd.DataFrame()
    
    # Player name - handle both combined name and separate first/last columns
    player_col = find_column(df, ROSTER_COLUMN_MAPPINGS['player_name'])
    first_col = find_column(df, ROSTER_COLUMN_MAPPINGS['player_first'])
    last_col = find_column(df, ROSTER_COLUMN_MAPPINGS['player_last'])
    
    if player_col:
        # Combined player name column exists
        roster['player_name'] = df[player_col].apply(lambda x: normalize_name(x) if pd.notna(x) else None)
        roster['player_name_original'] = df[player_col]
        # Split into first/last
        names_split = roster['player_name'].apply(lambda x: split_player_name(x) if x else (None, None))
        roster['player_first'] = names_split.apply(lambda x: x[0])
        roster['player_last'] = names_split.apply(lambda x: x[1])
    elif first_col and last_col:
        # Separate first/last columns
        roster['player_first'] = df[first_col].apply(lambda x: normalize_name(x) if pd.notna(x) else None)
        roster['player_last'] = df[last_col].apply(lambda x: normalize_name(x) if pd.notna(x) else None)
        roster['player_first_original'] = df[first_col]
        roster['player_last_original'] = df[last_col]
        # Combine into full name
        roster['player_name'] = roster.apply(
            lambda row: f"{row['player_first']} {row['player_last']}".strip() 
            if row['player_first'] and row['player_last'] else (row['player_first'] or row['player_last']),
            axis=1
        )
        roster['player_name_original'] = roster.apply(
            lambda row: f"{row.get('player_first_original', '')} {row.get('player_last_original', '')}".strip(),
            axis=1
        )
    
    # Team name
    team_col = find_column(df, ROSTER_COLUMN_MAPPINGS['team_name'])
    if team_col:
        roster['team_name'] = df[team_col].apply(lambda x: str(x).strip() if pd.notna(x) else None)
        roster['team_name_normalized'] = roster['team_name'].apply(lambda x: normalize_name(x) if x else None)
    
    # Parent emails - handle semicolon-separated emails
    def extract_emails(email_str):
        """Extract and normalize emails, handling semicolon-separated values."""
        if pd.isna(email_str) or email_str in ['(blank)', '']:
            return []
        emails = str(email_str).replace(';', ',').split(',')
        return [normalize_email(e.strip()) for e in emails if normalize_email(e.strip())]
    
    email1_col = find_column(df, ROSTER_COLUMN_MAPPINGS['parent_email_1'])
    if email1_col:
        # Handle multiple emails in one cell
        email_lists = df[email1_col].apply(extract_emails)
        roster['parent_email_1'] = email_lists.apply(lambda x: x[0] if x else None)
        roster['parent_email_1_original'] = df[email1_col]
        # If multiple emails in first column, use second as email_2
        roster['parent_email_1b'] = email_lists.apply(lambda x: x[1] if len(x) > 1 else None)
    
    email2_col = find_column(df, ROSTER_COLUMN_MAPPINGS['parent_email_2'])
    if email2_col:
        roster['parent_email_2'] = df[email2_col].apply(normalize_email)
        roster['parent_email_2_original'] = df[email2_col]
    elif 'parent_email_1b' in roster.columns:
        # Use secondary email from first column if no dedicated email2 column
        roster['parent_email_2'] = roster['parent_email_1b']
    
    # Parent names
    pname1_col = find_column(df, ROSTER_COLUMN_MAPPINGS['parent_name_1'])
    if pname1_col:
        roster['parent_name_1'] = df[pname1_col].apply(lambda x: str(x).strip() if pd.notna(x) else None)
    
    pname2_col = find_column(df, ROSTER_COLUMN_MAPPINGS['parent_name_2'])
    if pname2_col:
        roster['parent_name_2'] = df[pname2_col].apply(lambda x: str(x).strip() if pd.notna(x) else None)
    
    # Uniform number
    uniform_col = find_column(df, ROSTER_COLUMN_MAPPINGS['uniform_number'])
    has_uniform_numbers = uniform_col is not None
    if uniform_col:
        roster['uniform_number'] = df[uniform_col].apply(
            lambda x: str(int(x)) if pd.notna(x) and str(x).replace('.0','').replace('-','').isdigit() 
            else (str(x).strip() if pd.notna(x) else None)
        )
        roster['uniform_number_original'] = df[uniform_col]
        print(f"  Found uniform numbers in column '{uniform_col}'")
    else:
        roster['uniform_number'] = None
        print(f"  No uniform number column found in roster")
    
    # Remove rows without player name or without any email
    roster = roster[roster['player_name'].notna()].copy()
    
    # Check for email columns
    has_email_1 = 'parent_email_1' in roster.columns and roster['parent_email_1'].notna().any()
    has_email_2 = 'parent_email_2' in roster.columns and roster['parent_email_2'].notna().any()
    
    if has_email_1 or has_email_2:
        roster = roster[
            roster.get('parent_email_1', pd.Series([None]*len(roster))).notna() | 
            roster.get('parent_email_2', pd.Series([None]*len(roster))).notna()
        ].copy()
    
    # Create composite keys for matching
    roster['email_key'] = roster.apply(
        lambda row: frozenset(filter(None, [row.get('parent_email_1'), row.get('parent_email_2')])),
        axis=1
    )
    
    roster = roster.reset_index(drop=True)
    print(f"\nProcessed {len(roster)} valid roster entries")
    print(f"Unique teams: {roster['team_name'].nunique()}")
    if has_uniform_numbers:
        print(f"Roster contains uniform numbers: YES")
    else:
        print(f"Roster contains uniform numbers: NO")
    
    return roster, has_uniform_numbers


def load_orders_data(filepath):
    """
    Load orders data from Program Director Report.
    
    Returns a standardized DataFrame with order details including sizes.
    """
    print(f"\n{'='*60}")
    print("LOADING ORDERS DATA")
    print(f"{'='*60}")
    
    xls = pd.ExcelFile(filepath)
    df = pd.read_excel(xls, sheet_name=xls.sheet_names[0])
    print(f"Loaded {len(df)} orders from '{xls.sheet_names[0]}'")
    
    orders = pd.DataFrame()
    
    # Map standard order columns
    orders['order_number'] = df.get(ORDERS_COLUMNS['order_number'])
    orders['order_date'] = df.get(ORDERS_COLUMNS['order_date'])
    orders['order_email'] = df.get(ORDERS_COLUMNS['email']).apply(normalize_email)
    orders['order_email_original'] = df.get(ORDERS_COLUMNS['email'])
    
    # Player name from order
    orders['player_first'] = df.get(ORDERS_COLUMNS['player_first']).apply(
        lambda x: normalize_name(x) if pd.notna(x) else None
    )
    orders['player_last'] = df.get(ORDERS_COLUMNS['player_last']).apply(
        lambda x: normalize_name(x) if pd.notna(x) else None
    )
    orders['player_first_original'] = df.get(ORDERS_COLUMNS['player_first'])
    orders['player_last_original'] = df.get(ORDERS_COLUMNS['player_last'])
    orders['player_name'] = orders.apply(
        lambda row: f"{row['player_first']} {row['player_last']}".strip() 
        if row['player_first'] and row['player_last'] else None,
        axis=1
    )
    
    # Team from order
    orders['order_team'] = df.get(ORDERS_COLUMNS['team'])
    orders['order_team_normalized'] = orders['order_team'].apply(
        lambda x: normalize_name(x) if pd.notna(x) else None
    )
    
    # Billing info
    orders['billing_first'] = df.get(ORDERS_COLUMNS['billing_first'])
    orders['billing_last'] = df.get(ORDERS_COLUMNS['billing_last'])
    
    # Size columns
    orders['piece_1_size'] = df.get(ORDERS_COLUMNS['piece_1_size'])
    orders['piece_2_size'] = df.get(ORDERS_COLUMNS['piece_2_size'])
    orders['piece_3_size'] = df.get(ORDERS_COLUMNS['piece_3_size'])
    
    # Normalized sizes for comparison
    orders['piece_1_size_normalized'] = orders['piece_1_size'].apply(
        lambda x: str(x).lower().strip() if pd.notna(x) else None
    )
    orders['piece_2_size_normalized'] = orders['piece_2_size'].apply(
        lambda x: str(x).lower().strip() if pd.notna(x) else None
    )
    orders['piece_3_size_normalized'] = orders['piece_3_size'].apply(
        lambda x: str(x).lower().strip() if pd.notna(x) else None
    )
    
    # Uniform number from order
    uniform_num_col = df.get(ORDERS_COLUMNS['uniform_number'])
    if uniform_num_col is not None:
        orders['order_uniform_number'] = uniform_num_col.apply(
            lambda x: str(int(x)) if pd.notna(x) and str(x).replace('.0','').replace('-','').isdigit() 
            else (str(x).strip() if pd.notna(x) else None)
        )
        orders['order_uniform_number_original'] = uniform_num_col
    else:
        orders['order_uniform_number'] = None
        orders['order_uniform_number_original'] = None
    
    # Product name and uniform order detection
    orders['product_name'] = df.get(ORDERS_COLUMNS['product_name'])
    orders['category_tag'] = df.get(ORDERS_COLUMNS['category_tag'])
    
    # Determine if this is an actual uniform order (vs apparel/accessories)
    # Check if product name contains any of the uniform patterns
    def is_uniform_order(product_name):
        if pd.isna(product_name):
            return False
        product_lower = str(product_name).lower()
        return any(pattern in product_lower for pattern in UNIFORM_PRODUCT_PATTERNS)
    
    orders['is_uniform_order'] = orders['product_name'].apply(is_uniform_order)
    
    orders = orders.dropna(subset=['order_email', 'player_name']).reset_index(drop=True)
    
    uniform_count = orders['is_uniform_order'].sum()
    apparel_count = len(orders) - uniform_count
    print(f"Processed {len(orders)} valid orders ({uniform_count} uniforms, {apparel_count} apparel/accessories)")
    print(f"Unique teams: {orders['order_team'].nunique()}")
    
    return orders


# ============================================================================
# RECONCILIATION FUNCTIONS
# ============================================================================

def match_by_email(roster, orders):
    """
    Match roster entries to orders by parent email address.
    Returns a dict mapping roster index to list of matching order indices.
    """
    # Build email to order indices map
    email_to_orders = defaultdict(list)
    for idx, row in orders.iterrows():
        if row['order_email']:
            email_to_orders[row['order_email']].append(idx)
    
    # Match roster to orders
    roster_matches = {}
    for idx, row in roster.iterrows():
        matching_orders = set()
        if row.get('parent_email_1'):
            matching_orders.update(email_to_orders.get(row['parent_email_1'], []))
        if row.get('parent_email_2'):
            matching_orders.update(email_to_orders.get(row['parent_email_2'], []))
        roster_matches[idx] = list(matching_orders)
    
    return roster_matches


def match_by_name(roster, orders, similarity_threshold=0.85):
    """
    Match roster entries to orders by player name similarity.
    Returns a dict mapping roster index to list of (order_idx, similarity_score) tuples.
    """
    roster_matches = {}
    
    for r_idx, r_row in roster.iterrows():
        roster_name = r_row['player_name']
        if not roster_name:
            continue
        
        matches = []
        for o_idx, o_row in orders.iterrows():
            order_name = o_row['player_name']
            if not order_name:
                continue
            
            score = similarity_score(roster_name, order_name)
            if score >= similarity_threshold:
                matches.append((o_idx, score))
        
        roster_matches[r_idx] = sorted(matches, key=lambda x: -x[1])
    
    return roster_matches


def find_not_ordered(roster, email_matches, name_matches):
    """Identify roster entries with no matching orders."""
    not_ordered = []
    
    for idx, row in roster.iterrows():
        email_order_matches = email_matches.get(idx, [])
        name_order_matches = [m[0] for m in name_matches.get(idx, [])]
        
        if not email_order_matches and not name_order_matches:
            not_ordered.append({
                'player_name': row.get('player_name_original', row.get('player_name')),
                'team_name': row.get('team_name'),
                'parent_email_1': row.get('parent_email_1_original', row.get('parent_email_1')),
                'parent_email_2': row.get('parent_email_2_original', row.get('parent_email_2')),
                'parent_name_1': row.get('parent_name_1'),
                'parent_name_2': row.get('parent_name_2'),
                'roster_index': idx
            })
    
    return not_ordered


def find_duplicate_orders(roster, orders, email_matches, name_matches):
    """
    Identify players who have ordered more than once.
    Now distinguishes between:
    - True duplicates: Same player (matching first + last name) has multiple UNIFORM orders
    - Sibling orders: Different players from same family (different first names)
    
    Note: Only UNIFORM orders count as duplicates. If a player orders a uniform
    plus apparel items (hoodies, etc.), that is NOT flagged as a duplicate.
    """
    true_duplicates = []
    sibling_orders = []
    
    for idx, row in roster.iterrows():
        email_order_matches = email_matches.get(idx, [])
        name_order_matches = name_matches.get(idx, [])
        
        # Get all unique order indices from both email and name matches
        all_order_indices = set(email_order_matches)
        all_order_indices.update([m[0] for m in name_order_matches])
        
        if len(all_order_indices) > 1:
            # This roster entry has multiple orders - need to categorize
            roster_first = normalize_name(row.get('player_first'))
            roster_last = normalize_last_name(row.get('player_last'))
            
            # Group orders by whether they match this player's name
            matching_player_orders = []
            matching_player_uniform_orders = []  # Only uniform orders for duplicate detection
            different_player_orders = []
            
            for o_idx in all_order_indices:
                o_row = orders.loc[o_idx]
                order_first = normalize_name(o_row.get('player_first'))
                order_last = normalize_last_name(o_row.get('player_last'))
                is_uniform = o_row.get('is_uniform_order', True)  # Default to True for backward compatibility
                
                order_detail = {
                    'order_number': o_row['order_number'],
                    'order_date': o_row['order_date'],
                    'order_email': o_row['order_email_original'],
                    'player_first': o_row['player_first_original'],
                    'player_last': o_row['player_last_original'],
                    'player_name_on_order': f"{o_row['player_first_original']} {o_row['player_last_original']}",
                    'team_on_order': o_row['order_team'],
                    'product_name': o_row.get('product_name', ''),
                    'is_uniform_order': is_uniform,
                    'sizes': {
                        'piece_1': o_row.get('piece_1_size'),
                        'piece_2': o_row.get('piece_2_size'),
                        'piece_3': o_row.get('piece_3_size')
                    }
                }
                
                # Check if first name matches (allowing for similarity)
                first_name_match = (
                    roster_first == order_first or 
                    (roster_first and order_first and similarity_score(roster_first, order_first) >= 0.85)
                )
                # Check if last name matches (already normalized, removes suffixes)
                last_name_match = roster_last == order_last
                
                if first_name_match and last_name_match:
                    matching_player_orders.append(order_detail)
                    # Only add to uniform list if it's an actual uniform order
                    if is_uniform:
                        matching_player_uniform_orders.append(order_detail)
                else:
                    different_player_orders.append(order_detail)
            
            # True duplicate: Same player has multiple UNIFORM orders
            # (We only flag if there are 2+ uniform orders, not if they have 1 uniform + apparel)
            if len(matching_player_uniform_orders) > 1:
                true_duplicates.append({
                    'player_name': row.get('player_name_original', row.get('player_name')),
                    'team_name': row.get('team_name'),
                    'parent_email_1': row.get('parent_email_1_original', row.get('parent_email_1')),
                    'parent_email_2': row.get('parent_email_2_original', row.get('parent_email_2')),
                    'order_count': len(matching_player_uniform_orders),
                    'orders': matching_player_uniform_orders,  # Only include the uniform orders
                    'total_orders': len(matching_player_orders),  # Total including apparel
                    'roster_index': idx,
                    'issue_type': 'TRUE_DUPLICATE'
                })
            
            # Sibling orders: Different players from same family
            if different_player_orders:
                sibling_orders.append({
                    'roster_player_name': row.get('player_name_original', row.get('player_name')),
                    'team_name': row.get('team_name'),
                    'parent_email_1': row.get('parent_email_1_original', row.get('parent_email_1')),
                    'parent_email_2': row.get('parent_email_2_original', row.get('parent_email_2')),
                    'sibling_orders': different_player_orders,
                    'sibling_names': [o['player_name_on_order'] for o in different_player_orders],
                    'roster_index': idx,
                    'issue_type': 'SIBLING_ORDER'
                })
    
    return true_duplicates, sibling_orders


def find_wrong_team(roster, orders, email_matches, name_matches):
    """
    Identify players who ordered for a different team than their roster team.
    
    Uses intelligent team name matching that handles common variations:
    - "2027 Elite" vs "2027 Elite CLT" -> treated as same team
    - Different teams with similar names are properly distinguished
    """
    wrong_team = []
    
    for idx, row in roster.iterrows():
        roster_team = row.get('team_name')
        if not roster_team:
            continue
        
        # Get all order indices for this roster entry
        email_order_matches = email_matches.get(idx, [])
        name_order_matches = [m[0] for m in name_matches.get(idx, []) if m[1] >= 0.9]  # High confidence name match
        
        all_order_indices = set(email_order_matches) | set(name_order_matches)
        
        for o_idx in all_order_indices:
            o_row = orders.loc[o_idx]
            order_team = o_row.get('order_team')
            
            if order_team and roster_team:
                # Use the intelligent team comparison
                team_similarity = calculate_team_similarity(roster_team, order_team)
                
                # Only flag as wrong team if similarity is below 90%
                if not teams_are_equivalent(roster_team, order_team, threshold=0.90):
                    wrong_team.append({
                        'player_name': row.get('player_name_original', row.get('player_name')),
                        'roster_team': roster_team,
                        'order_team': order_team,
                        'team_similarity': round(team_similarity, 2),
                        'order_number': o_row['order_number'],
                        'order_email': o_row['order_email_original'],
                        'parent_email_1': row.get('parent_email_1_original'),
                        'parent_email_2': row.get('parent_email_2_original'),
                        'roster_index': idx
                    })
    
    return wrong_team


def find_misspellings(roster, orders, email_matches, similarity_threshold=0.7):
    """
    Identify potential misspellings of player last names.
    Compares orders matched by email where names don't quite match.
    Now ignores case and common suffixes (Jr., III, IV, etc.)
    """
    misspellings = []
    
    for r_idx, row in roster.iterrows():
        order_indices = email_matches.get(r_idx, [])
        if not order_indices:
            continue
        
        roster_last = row.get('player_last')
        if not roster_last:
            continue
        
        # Normalize roster last name (remove suffixes, lowercase)
        roster_last_normalized = normalize_last_name(roster_last)
        
        for o_idx in order_indices:
            o_row = orders.loc[o_idx]
            order_last = o_row.get('player_last')
            
            if not order_last:
                continue
            
            # Normalize order last name (remove suffixes, lowercase)
            order_last_normalized = normalize_last_name(order_last)
            
            # Skip if normalized versions match exactly (not a misspelling)
            if roster_last_normalized == order_last_normalized:
                continue
            
            # Calculate similarity on normalized versions
            score = similarity_score(roster_last_normalized, order_last_normalized)
            
            # Flag if similar but not exact (potential typo)
            if similarity_threshold <= score < 1.0:
                misspellings.append({
                    'roster_player_name': row.get('player_name_original'),
                    'roster_last_name': row.get('player_last'),
                    'roster_last_normalized': roster_last_normalized,
                    'order_player_name': f"{o_row['player_first_original']} {o_row['player_last_original']}",
                    'order_last_name': o_row['player_last'],
                    'order_last_normalized': order_last_normalized,
                    'similarity': round(score, 2),
                    'team_name': row.get('team_name'),
                    'order_number': o_row['order_number'],
                    'order_email': o_row['order_email_original'],
                    'roster_index': r_idx
                })
    
    return misspellings


def get_size_numeric(size_str):
    """Convert size string to numeric value for comparison."""
    if pd.isna(size_str) or not size_str:
        return None
    size_lower = str(size_str).lower().strip()
    return SIZE_ORDER.get(size_lower, None)


def get_size_category(size_str):
    """Determine if a size is Youth, Adult, or Hat."""
    if pd.isna(size_str) or not size_str:
        return None
    size_lower = str(size_str).lower().strip()
    
    # Youth sizes
    if size_lower.startswith('y') or 'youth' in size_lower:
        return 'Youth'
    # Hat sizes
    if '/' in size_lower or size_lower in ['s/m', 'l/xl', 'sm', 'lxl']:
        return 'Hat'
    # Adult sizes
    return 'Adult'


def find_size_outliers(orders):
    """
    Identify potential sizing issues:
    1. Inconsistent sizes within an order (e.g., YS top with XL pants)
    2. Sizes that are outliers compared to teammates
    3. Mixed Youth/Adult sizing within an order
    """
    size_outliers = []
    
    # First, calculate team size statistics
    team_size_stats = {}
    for team in orders['order_team'].unique():
        if pd.isna(team):
            continue
        
        team_orders = orders[orders['order_team'] == team]
        
        # Get numeric sizes for each piece
        piece_1_sizes = team_orders['piece_1_size_normalized'].apply(get_size_numeric).dropna()
        piece_2_sizes = team_orders['piece_2_size_normalized'].apply(get_size_numeric).dropna()
        piece_3_sizes = team_orders['piece_3_size_normalized'].apply(get_size_numeric).dropna()
        
        team_size_stats[team] = {
            'piece_1': {
                'mean': piece_1_sizes.mean() if len(piece_1_sizes) > 0 else None,
                'std': piece_1_sizes.std() if len(piece_1_sizes) > 1 else None,
                'min': piece_1_sizes.min() if len(piece_1_sizes) > 0 else None,
                'max': piece_1_sizes.max() if len(piece_1_sizes) > 0 else None,
                'count': len(piece_1_sizes)
            },
            'piece_2': {
                'mean': piece_2_sizes.mean() if len(piece_2_sizes) > 0 else None,
                'std': piece_2_sizes.std() if len(piece_2_sizes) > 1 else None,
                'min': piece_2_sizes.min() if len(piece_2_sizes) > 0 else None,
                'max': piece_2_sizes.max() if len(piece_2_sizes) > 0 else None,
                'count': len(piece_2_sizes)
            },
            'piece_3': {
                'mean': piece_3_sizes.mean() if len(piece_3_sizes) > 0 else None,
                'std': piece_3_sizes.std() if len(piece_3_sizes) > 1 else None,
                'min': piece_3_sizes.min() if len(piece_3_sizes) > 0 else None,
                'max': piece_3_sizes.max() if len(piece_3_sizes) > 0 else None,
                'count': len(piece_3_sizes)
            }
        }
    
    # Analyze each order
    for idx, row in orders.iterrows():
        issues = []
        
        # Get sizes
        size_1 = row.get('piece_1_size')
        size_2 = row.get('piece_2_size')
        size_3 = row.get('piece_3_size')
        
        size_1_num = get_size_numeric(size_1)
        size_2_num = get_size_numeric(size_2)
        size_3_num = get_size_numeric(size_3)
        
        size_1_cat = get_size_category(size_1)
        size_2_cat = get_size_category(size_2)
        size_3_cat = get_size_category(size_3)
        
        # Check 1: Inconsistent sizes within order (more than 2 size difference)
        uniform_sizes = [s for s in [size_1_num, size_2_num, size_3_num] if s is not None and s < 20]  # Exclude hat sizes
        if len(uniform_sizes) >= 2:
            size_spread = max(uniform_sizes) - min(uniform_sizes)
            if size_spread > 2:
                issues.append(f"Size spread of {size_spread} within order (Pieces: {size_1}, {size_2}, {size_3})")
        
        # Check 2: Mixed Youth/Adult sizing (excluding hats)
        uniform_cats = [c for c in [size_1_cat, size_2_cat, size_3_cat] if c and c != 'Hat']
        if len(set(uniform_cats)) > 1:
            issues.append(f"Mixed Youth/Adult sizing: {size_1_cat} + {size_2_cat} + {size_3_cat}")
        
        # Check 3: Outlier compared to team (more than 2 std devs from mean)
        team = row.get('order_team')
        if team and team in team_size_stats:
            stats = team_size_stats[team]
            
            for piece_num, piece_size, piece_stats in [
                (1, size_1_num, stats['piece_1']),
                (2, size_2_num, stats['piece_2']),
                (3, size_3_num, stats['piece_3'])
            ]:
                if piece_size is not None and piece_stats['mean'] is not None and piece_stats['std'] is not None:
                    if piece_stats['std'] > 0:
                        z_score = abs(piece_size - piece_stats['mean']) / piece_stats['std']
                        if z_score > 2.0:
                            actual_size = row.get(f'piece_{piece_num}_size')
                            issues.append(f"Piece {piece_num} size '{actual_size}' is outlier for team (z-score: {z_score:.1f})")
        
        # If any issues found, add to outliers list
        if issues:
            size_outliers.append({
                'player_name': f"{row['player_first_original']} {row['player_last_original']}",
                'team_name': row['order_team'],
                'order_number': row['order_number'],
                'order_email': row['order_email_original'],
                'piece_1_size': size_1,
                'piece_2_size': size_2,
                'piece_3_size': size_3,
                'issues': issues,
                'issue_summary': '; '.join(issues)
            })
    
    return size_outliers, team_size_stats


def find_uniform_number_mismatches(roster, orders, email_matches, name_matches, has_uniform_numbers):
    """
    Identify orders where the uniform number doesn't match the roster's assigned number.
    
    Only performs this check if the roster contains uniform numbers.
    Returns tuple: (mismatches_list, check_performed_bool)
    """
    if not has_uniform_numbers:
        return [], False
    
    # Check if roster actually has uniform number data
    roster_has_numbers = roster['uniform_number'].notna().any()
    if not roster_has_numbers:
        return [], False
    
    mismatches = []
    
    for r_idx, row in roster.iterrows():
        roster_uniform = row.get('uniform_number')
        if not roster_uniform:
            continue
        
        # Get matching orders
        email_order_matches = email_matches.get(r_idx, [])
        name_order_matches = [m[0] for m in name_matches.get(r_idx, []) if m[1] >= 0.9]
        all_order_indices = set(email_order_matches) | set(name_order_matches)
        
        for o_idx in all_order_indices:
            o_row = orders.loc[o_idx]
            order_uniform = o_row.get('order_uniform_number')
            
            # Skip if order has no uniform number
            if not order_uniform or pd.isna(order_uniform):
                continue
            
            # Compare uniform numbers (normalize for comparison)
            roster_num_normalized = str(roster_uniform).strip().lstrip('0') or '0'
            order_num_normalized = str(order_uniform).strip().lstrip('0') or '0'
            
            if roster_num_normalized != order_num_normalized:
                mismatches.append({
                    'player_name': row.get('player_name_original', row.get('player_name')),
                    'team_name': row.get('team_name'),
                    'roster_uniform_number': roster_uniform,
                    'order_uniform_number': order_uniform,
                    'order_number': o_row['order_number'],
                    'order_email': o_row['order_email_original'],
                    'parent_email_1': row.get('parent_email_1_original', row.get('parent_email_1')),
                    'parent_email_2': row.get('parent_email_2_original', row.get('parent_email_2')),
                    'roster_index': r_idx
                })
    
    return mismatches, True


def generate_data_quality_warnings(roster, orders, has_uniform_numbers):
    """
    Generate data quality warnings to help identify potential issues with source data.
    
    Returns a list of warning dictionaries with category, severity, and details.
    """
    warnings = []
    
    # -------------------------------------------------------------------------
    # 1. Missing Email Warnings
    # -------------------------------------------------------------------------
    missing_email_players = []
    for idx, row in roster.iterrows():
        email1 = row.get('parent_email_1')
        email2 = row.get('parent_email_2')
        if not email1 and not email2:
            missing_email_players.append({
                'player_name': row.get('player_name_original', row.get('player_name')),
                'team_name': row.get('team_name')
            })
    
    if missing_email_players:
        warnings.append({
            'category': 'Missing Emails',
            'severity': 'High',
            'count': len(missing_email_players),
            'description': 'Roster entries with no parent email addresses. These players cannot be matched to orders by email.',
            'details': missing_email_players[:20],  # Limit to first 20
            'has_more': len(missing_email_players) > 20
        })
    
    # -------------------------------------------------------------------------
    # 2. Invalid Email Format Warnings
    # -------------------------------------------------------------------------
    invalid_email_entries = []
    multi_email_entries = []
    
    # Common placeholder values that should be treated as empty
    blank_placeholders = {'(blank)', 'blank', 'n/a', 'na', 'none', '-', ''}
    
    for idx, row in roster.iterrows():
        email1 = row.get('parent_email_1_original', row.get('parent_email_1'))
        email2 = row.get('parent_email_2_original', row.get('parent_email_2'))
        
        invalid_emails = []
        multi_emails = []
        
        for email in [email1, email2]:
            if email and not pd.isna(email):
                email_str = str(email).strip()
                # Skip placeholder values
                if email_str.lower() in blank_placeholders:
                    continue
                # Check for multiple emails in one cell
                if ';' in email_str or ',' in email_str:
                    multi_emails.append(email_str)
                elif not is_valid_email(email_str):
                    invalid_emails.append(email_str)
        
        if invalid_emails:
            invalid_email_entries.append({
                'player_name': row.get('player_name_original', row.get('player_name')),
                'team_name': row.get('team_name'),
                'invalid_emails': ', '.join(invalid_emails)
            })
        
        if multi_emails:
            multi_email_entries.append({
                'player_name': row.get('player_name_original', row.get('player_name')),
                'team_name': row.get('team_name'),
                'multi_emails': ', '.join(multi_emails)
            })
    
    if invalid_email_entries:
        warnings.append({
            'category': 'Invalid Email Format',
            'severity': 'Medium',
            'count': len(invalid_email_entries),
            'description': 'Roster entries with email addresses that appear to be invalid format.',
            'details': invalid_email_entries[:20],
            'has_more': len(invalid_email_entries) > 20
        })
    
    if multi_email_entries:
        warnings.append({
            'category': 'Multiple Emails in One Cell',
            'severity': 'Low',
            'count': len(multi_email_entries),
            'description': 'Roster entries with multiple email addresses in one cell. The first email will be used for matching.',
            'details': multi_email_entries[:20],
            'has_more': len(multi_email_entries) > 20
        })
    
    # -------------------------------------------------------------------------
    # 3. Duplicate Emails (Different Players)
    # -------------------------------------------------------------------------
    email_to_players = defaultdict(list)
    for idx, row in roster.iterrows():
        email1 = row.get('parent_email_1')
        email2 = row.get('parent_email_2')
        player_name = row.get('player_name_original', row.get('player_name'))
        player_last = row.get('player_last', '')
        
        if email1:
            email_to_players[email1].append((player_name, player_last, row.get('team_name')))
        if email2:
            email_to_players[email2].append((player_name, player_last, row.get('team_name')))
    
    # Find emails with multiple different last names (likely data issue, not siblings)
    suspicious_duplicates = []
    for email, players in email_to_players.items():
        if len(players) > 1:
            # Check if last names are different (not siblings)
            last_names = set(normalize_last_name(p[1]) for p in players if p[1])
            if len(last_names) > 1:
                suspicious_duplicates.append({
                    'email': email,
                    'players': [f"{p[0]} ({p[2]})" for p in players],
                    'player_count': len(players)
                })
    
    if suspicious_duplicates:
        warnings.append({
            'category': 'Duplicate Emails (Different Families)',
            'severity': 'Medium',
            'count': len(suspicious_duplicates),
            'description': 'Same email address used for players with different last names. This may indicate data entry errors.',
            'details': suspicious_duplicates[:20],
            'has_more': len(suspicious_duplicates) > 20
        })
    
    # -------------------------------------------------------------------------
    # 4. Teams in Roster with No Orders
    # -------------------------------------------------------------------------
    roster_teams = set(roster['team_name'].dropna().unique())
    order_teams = set(orders['order_team'].dropna().unique())
    
    # Normalize for comparison
    roster_teams_normalized = {normalize_team_name(t): t for t in roster_teams if t}
    order_teams_normalized = {normalize_team_name(t): t for t in order_teams if t}
    
    teams_no_orders = []
    for norm_team, orig_team in roster_teams_normalized.items():
        # Check if this team has any matching orders using the smart matching
        has_match = False
        for order_norm_team in order_teams_normalized.keys():
            if teams_are_equivalent(norm_team, order_norm_team, threshold=0.90):
                has_match = True
                break
        
        if not has_match:
            # Count players on this team
            player_count = len(roster[roster['team_name'] == orig_team])
            teams_no_orders.append({
                'team_name': orig_team,
                'player_count': player_count
            })
    
    if teams_no_orders:
        warnings.append({
            'category': 'Roster Teams with No Orders',
            'severity': 'High',
            'count': len(teams_no_orders),
            'description': 'Teams in roster that have no matching orders. May indicate team name mismatch or all players have not ordered.',
            'details': teams_no_orders,
            'has_more': False
        })
    
    # -------------------------------------------------------------------------
    # 5. Order Teams Not in Roster
    # -------------------------------------------------------------------------
    teams_not_in_roster = []
    for order_norm_team, order_orig_team in order_teams_normalized.items():
        # Check if this order team matches any roster team
        has_match = False
        for roster_norm_team in roster_teams_normalized.keys():
            if teams_are_equivalent(roster_norm_team, order_norm_team, threshold=0.90):
                has_match = True
                break
        
        if not has_match:
            # Count orders for this team
            order_count = len(orders[orders['order_team'] == order_orig_team])
            teams_not_in_roster.append({
                'team_name': order_orig_team,
                'order_count': order_count
            })
    
    if teams_not_in_roster:
        warnings.append({
            'category': 'Order Teams Not in Roster',
            'severity': 'Medium',
            'count': len(teams_not_in_roster),
            'description': 'Teams in orders that do not match any roster team. May indicate team name mismatch or missing roster data.',
            'details': teams_not_in_roster,
            'has_more': False
        })
    
    # -------------------------------------------------------------------------
    # 6. Non-Numeric Uniform Numbers
    # -------------------------------------------------------------------------
    if has_uniform_numbers:
        non_numeric_uniforms = []
        for idx, row in roster.iterrows():
            uniform_num = row.get('uniform_number')
            if uniform_num and not pd.isna(uniform_num):
                # Check if it's numeric (allowing for leading zeros)
                uniform_str = str(uniform_num).strip()
                if not uniform_str.replace('.', '').replace('-', '').isdigit():
                    non_numeric_uniforms.append({
                        'player_name': row.get('player_name_original', row.get('player_name')),
                        'team_name': row.get('team_name'),
                        'uniform_value': uniform_str
                    })
        
        if non_numeric_uniforms:
            warnings.append({
                'category': 'Non-Numeric Uniform Numbers',
                'severity': 'High',
                'count': len(non_numeric_uniforms),
                'description': 'Roster entries with non-numeric uniform numbers. This may indicate the wrong column was mapped as uniform number.',
                'details': non_numeric_uniforms[:20],
                'has_more': len(non_numeric_uniforms) > 20
            })
        
        # Also check for unusual uniform numbers (outside typical range)
        unusual_uniforms = []
        for idx, row in roster.iterrows():
            uniform_num = row.get('uniform_number')
            if uniform_num and not pd.isna(uniform_num):
                try:
                    num_value = int(float(str(uniform_num).strip()))
                    if num_value < 0 or num_value > 99:
                        unusual_uniforms.append({
                            'player_name': row.get('player_name_original', row.get('player_name')),
                            'team_name': row.get('team_name'),
                            'uniform_value': str(uniform_num)
                        })
                except (ValueError, TypeError):
                    pass  # Already caught by non-numeric check
        
        if unusual_uniforms:
            warnings.append({
                'category': 'Unusual Uniform Numbers',
                'severity': 'Low',
                'count': len(unusual_uniforms),
                'description': 'Uniform numbers outside typical range (0-99). May be valid but worth verifying.',
                'details': unusual_uniforms[:20],
                'has_more': len(unusual_uniforms) > 20
            })
    
    # -------------------------------------------------------------------------
    # 7. Players with Unusual Characters in Names
    # -------------------------------------------------------------------------
    unusual_names = []
    unusual_char_pattern = re.compile(r'[^\w\s\'-]', re.UNICODE)
    
    for idx, row in roster.iterrows():
        player_name = row.get('player_name_original', row.get('player_name', ''))
        if player_name:
            unusual_chars = unusual_char_pattern.findall(str(player_name))
            if unusual_chars:
                unusual_names.append({
                    'player_name': player_name,
                    'team_name': row.get('team_name'),
                    'unusual_characters': ''.join(set(unusual_chars))
                })
    
    if unusual_names:
        warnings.append({
            'category': 'Unusual Characters in Names',
            'severity': 'Low',
            'count': len(unusual_names),
            'description': 'Player names containing unusual characters. May affect matching.',
            'details': unusual_names[:20],
            'has_more': len(unusual_names) > 20
        })
    
    return warnings


# ============================================================================
# REPORTING FUNCTIONS
# ============================================================================

def generate_summary(roster, orders, not_ordered, true_duplicates, sibling_orders, wrong_team, misspellings, size_outliers, uniform_mismatches, uniform_check_performed, data_quality_warnings):
    """Generate summary statistics."""
    summary = {
        'Total Roster Entries': len(roster),
        'Total Orders Received': len(orders),
        'Unique Teams (Roster)': roster['team_name'].nunique(),
        'Unique Teams (Orders)': orders['order_team'].nunique(),
        'Players Not Ordered': len(not_ordered),
        'True Duplicate Orders': len(true_duplicates),
        'Sibling/Family Orders': len(sibling_orders),
        'Wrong Team Orders': len(wrong_team),
        'Potential Misspellings': len(misspellings),
        'Size Outliers': len(size_outliers),
        'Order Rate': f"{(1 - len(not_ordered)/len(roster))*100:.1f}%" if len(roster) > 0 else "N/A"
    }
    
    if uniform_check_performed:
        summary['Uniform Number Mismatches'] = len(uniform_mismatches)
    else:
        summary['Uniform Number Check'] = 'Not performed (no uniform numbers in roster)'
    
    # Count high severity warnings
    high_severity_count = sum(1 for w in data_quality_warnings if w['severity'] == 'High')
    if high_severity_count > 0:
        summary['Data Quality Warnings (High)'] = high_severity_count
    
    return summary


def save_report(output_path, summary, not_ordered, true_duplicates, sibling_orders, wrong_team, misspellings, size_outliers, uniform_mismatches, uniform_check_performed, data_quality_warnings, team_size_stats, roster, orders):
    """Save reconciliation report to Excel file."""
    print(f"\n{'='*60}")
    print("SAVING REPORT")
    print(f"{'='*60}")
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    wb = Workbook()
    
    # Styling
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='4472C4')
    warning_fill = PatternFill('solid', fgColor='FFEB9C')
    error_fill = PatternFill('solid', fgColor='FFC7CE')
    success_fill = PatternFill('solid', fgColor='C6EFCE')
    info_fill = PatternFill('solid', fgColor='BDD7EE')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # -------------------------------------------------------------------------
    # Summary Sheet
    # -------------------------------------------------------------------------
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    ws_summary['A1'] = "Roster vs Orders Reconciliation Report"
    ws_summary['A1'].font = Font(bold=True, size=16)
    ws_summary['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    
    row = 4
    for key, value in summary.items():
        ws_summary[f'A{row}'] = key
        ws_summary[f'B{row}'] = value
        ws_summary[f'A{row}'].font = Font(bold=True)
        row += 1
    
    # Issue counts with color coding
    row += 1
    ws_summary[f'A{row}'] = "Issue Summary"
    ws_summary[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    
    issues = [
        ('Players Not Ordered', len(not_ordered), error_fill if len(not_ordered) > 0 else success_fill),
        ('True Duplicate Orders', len(true_duplicates), error_fill if len(true_duplicates) > 0 else success_fill),
        ('Sibling/Family Orders (Info Only)', len(sibling_orders), info_fill),
        ('Wrong Team Orders', len(wrong_team), warning_fill if len(wrong_team) > 0 else success_fill),
        ('Potential Misspellings', len(misspellings), warning_fill if len(misspellings) > 0 else success_fill),
        ('Size Outliers', len(size_outliers), warning_fill if len(size_outliers) > 0 else success_fill),
    ]
    
    # Add uniform number info
    if uniform_check_performed:
        issues.append(('Uniform Number Mismatches', len(uniform_mismatches), error_fill if len(uniform_mismatches) > 0 else success_fill))
    else:
        issues.append(('Uniform Number Check', 'N/A - No uniform numbers in roster', info_fill))
    
    for issue_name, count, fill in issues:
        ws_summary[f'A{row}'] = issue_name
        ws_summary[f'B{row}'] = count
        ws_summary[f'B{row}'].fill = fill
        row += 1
    
    ws_summary.column_dimensions['A'].width = 35
    ws_summary.column_dimensions['B'].width = 15
    
    # -------------------------------------------------------------------------
    # Not Ordered Sheet
    # -------------------------------------------------------------------------
    ws_not_ordered = wb.create_sheet("Not Ordered")
    
    headers = ['Player Name', 'Team Name', 'Parent Email 1', 'Parent Email 2', 'Parent Name 1', 'Parent Name 2']
    for col, header in enumerate(headers, 1):
        cell = ws_not_ordered.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    for row_idx, item in enumerate(not_ordered, 2):
        ws_not_ordered.cell(row=row_idx, column=1, value=item.get('player_name'))
        ws_not_ordered.cell(row=row_idx, column=2, value=item.get('team_name'))
        ws_not_ordered.cell(row=row_idx, column=3, value=item.get('parent_email_1'))
        ws_not_ordered.cell(row=row_idx, column=4, value=item.get('parent_email_2'))
        ws_not_ordered.cell(row=row_idx, column=5, value=item.get('parent_name_1'))
        ws_not_ordered.cell(row=row_idx, column=6, value=item.get('parent_name_2'))
    
    for col in range(1, 7):
        ws_not_ordered.column_dimensions[chr(64+col)].width = 25
    
    # -------------------------------------------------------------------------
    # True Duplicate Orders Sheet
    # -------------------------------------------------------------------------
    ws_duplicates = wb.create_sheet("True Duplicates")
    
    headers = ['Player Name', 'Team Name', 'Parent Email 1', 'Parent Email 2', 'Uniform Orders', 'Order Numbers', 'Order Dates', 'Teams on Orders', 'Products Ordered']
    for col, header in enumerate(headers, 1):
        cell = ws_duplicates.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    for row_idx, item in enumerate(true_duplicates, 2):
        ws_duplicates.cell(row=row_idx, column=1, value=item.get('player_name'))
        ws_duplicates.cell(row=row_idx, column=2, value=item.get('team_name'))
        ws_duplicates.cell(row=row_idx, column=3, value=item.get('parent_email_1'))
        ws_duplicates.cell(row=row_idx, column=4, value=item.get('parent_email_2'))
        ws_duplicates.cell(row=row_idx, column=5, value=item.get('order_count'))
        
        order_nums = ', '.join([str(o['order_number']) for o in item.get('orders', [])])
        order_dates = ', '.join([str(o['order_date'])[:10] for o in item.get('orders', [])])
        order_teams = ' | '.join([str(o.get('team_on_order', ''))[:40] for o in item.get('orders', [])])
        order_products = ' | '.join([str(o.get('product_name', ''))[:50] for o in item.get('orders', [])])
        
        ws_duplicates.cell(row=row_idx, column=6, value=order_nums)
        ws_duplicates.cell(row=row_idx, column=7, value=order_dates)
        ws_duplicates.cell(row=row_idx, column=8, value=order_teams)
        ws_duplicates.cell(row=row_idx, column=9, value=order_products)
    
    ws_duplicates.column_dimensions['A'].width = 20
    ws_duplicates.column_dimensions['B'].width = 25
    ws_duplicates.column_dimensions['C'].width = 25
    ws_duplicates.column_dimensions['D'].width = 25
    ws_duplicates.column_dimensions['E'].width = 12
    ws_duplicates.column_dimensions['F'].width = 15
    ws_duplicates.column_dimensions['G'].width = 20
    ws_duplicates.column_dimensions['H'].width = 45
    ws_duplicates.column_dimensions['I'].width = 55
    
    # -------------------------------------------------------------------------
    # Sibling/Family Orders Sheet (Informational)
    # -------------------------------------------------------------------------
    ws_siblings = wb.create_sheet("Sibling Orders (Info)")
    
    headers = ['Roster Player', 'Team Name', 'Parent Email 1', 'Parent Email 2', 'Sibling Names on Orders', 'Sibling Teams']
    for col, header in enumerate(headers, 1):
        cell = ws_siblings.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = PatternFill('solid', fgColor='5B9BD5')  # Different blue for info
        cell.border = thin_border
    
    for row_idx, item in enumerate(sibling_orders, 2):
        ws_siblings.cell(row=row_idx, column=1, value=item.get('roster_player_name'))
        ws_siblings.cell(row=row_idx, column=2, value=item.get('team_name'))
        ws_siblings.cell(row=row_idx, column=3, value=item.get('parent_email_1'))
        ws_siblings.cell(row=row_idx, column=4, value=item.get('parent_email_2'))
        
        sibling_names = ', '.join(item.get('sibling_names', []))
        sibling_teams = ', '.join(set([o.get('team_on_order', '') for o in item.get('sibling_orders', [])]))
        
        ws_siblings.cell(row=row_idx, column=5, value=sibling_names)
        ws_siblings.cell(row=row_idx, column=6, value=sibling_teams)
    
    for col in range(1, 7):
        ws_siblings.column_dimensions[chr(64+col)].width = 25
    
    # -------------------------------------------------------------------------
    # Wrong Team Sheet
    # -------------------------------------------------------------------------
    ws_wrong_team = wb.create_sheet("Wrong Team")
    
    headers = ['Player Name', 'Roster Team', 'Order Team', 'Team Similarity', 'Order Number', 'Order Email', 'Parent Email 1', 'Parent Email 2']
    for col, header in enumerate(headers, 1):
        cell = ws_wrong_team.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    for row_idx, item in enumerate(wrong_team, 2):
        ws_wrong_team.cell(row=row_idx, column=1, value=item.get('player_name'))
        ws_wrong_team.cell(row=row_idx, column=2, value=item.get('roster_team'))
        ws_wrong_team.cell(row=row_idx, column=3, value=item.get('order_team'))
        ws_wrong_team.cell(row=row_idx, column=4, value=item.get('team_similarity'))
        ws_wrong_team.cell(row=row_idx, column=5, value=item.get('order_number'))
        ws_wrong_team.cell(row=row_idx, column=6, value=item.get('order_email'))
        ws_wrong_team.cell(row=row_idx, column=7, value=item.get('parent_email_1'))
        ws_wrong_team.cell(row=row_idx, column=8, value=item.get('parent_email_2'))
    
    for col in range(1, 9):
        ws_wrong_team.column_dimensions[chr(64+col)].width = 30
    
    # -------------------------------------------------------------------------
    # Misspellings Sheet
    # -------------------------------------------------------------------------
    ws_misspellings = wb.create_sheet("Potential Misspellings")
    
    headers = ['Roster Player Name', 'Roster Last Name', 'Order Player Name', 'Order Last Name', 'Similarity', 'Team Name', 'Order Number', 'Order Email']
    for col, header in enumerate(headers, 1):
        cell = ws_misspellings.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    for row_idx, item in enumerate(misspellings, 2):
        ws_misspellings.cell(row=row_idx, column=1, value=item.get('roster_player_name'))
        ws_misspellings.cell(row=row_idx, column=2, value=item.get('roster_last_name'))
        ws_misspellings.cell(row=row_idx, column=3, value=item.get('order_player_name'))
        ws_misspellings.cell(row=row_idx, column=4, value=item.get('order_last_name'))
        ws_misspellings.cell(row=row_idx, column=5, value=item.get('similarity'))
        ws_misspellings.cell(row=row_idx, column=6, value=item.get('team_name'))
        ws_misspellings.cell(row=row_idx, column=7, value=item.get('order_number'))
        ws_misspellings.cell(row=row_idx, column=8, value=item.get('order_email'))
    
    for col in range(1, 9):
        ws_misspellings.column_dimensions[chr(64+col)].width = 22
    
    # -------------------------------------------------------------------------
    # Size Outliers Sheet
    # -------------------------------------------------------------------------
    ws_sizes = wb.create_sheet("Size Outliers")
    
    headers = ['Player Name', 'Team Name', 'Order Number', 'Order Email', 'Piece 1 Size', 'Piece 2 Size', 'Piece 3 Size', 'Issues']
    for col, header in enumerate(headers, 1):
        cell = ws_sizes.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    for row_idx, item in enumerate(size_outliers, 2):
        ws_sizes.cell(row=row_idx, column=1, value=item.get('player_name'))
        ws_sizes.cell(row=row_idx, column=2, value=item.get('team_name'))
        ws_sizes.cell(row=row_idx, column=3, value=item.get('order_number'))
        ws_sizes.cell(row=row_idx, column=4, value=item.get('order_email'))
        ws_sizes.cell(row=row_idx, column=5, value=item.get('piece_1_size'))
        ws_sizes.cell(row=row_idx, column=6, value=item.get('piece_2_size'))
        ws_sizes.cell(row=row_idx, column=7, value=item.get('piece_3_size'))
        ws_sizes.cell(row=row_idx, column=8, value=item.get('issue_summary'))
        
        # Highlight the issues cell
        ws_sizes.cell(row=row_idx, column=8).fill = warning_fill
    
    ws_sizes.column_dimensions['A'].width = 25
    ws_sizes.column_dimensions['B'].width = 35
    ws_sizes.column_dimensions['C'].width = 12
    ws_sizes.column_dimensions['D'].width = 30
    ws_sizes.column_dimensions['E'].width = 12
    ws_sizes.column_dimensions['F'].width = 12
    ws_sizes.column_dimensions['G'].width = 12
    ws_sizes.column_dimensions['H'].width = 60
    
    # -------------------------------------------------------------------------
    # Uniform Number Mismatches Sheet
    # -------------------------------------------------------------------------
    ws_uniform = wb.create_sheet("Uniform Number Mismatches")
    
    if uniform_check_performed:
        headers = ['Player Name', 'Team Name', 'Roster Number', 'Order Number', 'Order #', 'Order Email', 'Parent Email 1', 'Parent Email 2']
        for col, header in enumerate(headers, 1):
            cell = ws_uniform.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        for row_idx, item in enumerate(uniform_mismatches, 2):
            ws_uniform.cell(row=row_idx, column=1, value=item.get('player_name'))
            ws_uniform.cell(row=row_idx, column=2, value=item.get('team_name'))
            ws_uniform.cell(row=row_idx, column=3, value=item.get('roster_uniform_number'))
            ws_uniform.cell(row=row_idx, column=4, value=item.get('order_uniform_number'))
            ws_uniform.cell(row=row_idx, column=5, value=item.get('order_number'))
            ws_uniform.cell(row=row_idx, column=6, value=item.get('order_email'))
            ws_uniform.cell(row=row_idx, column=7, value=item.get('parent_email_1'))
            ws_uniform.cell(row=row_idx, column=8, value=item.get('parent_email_2'))
            
            # Highlight the mismatched numbers
            ws_uniform.cell(row=row_idx, column=3).fill = warning_fill
            ws_uniform.cell(row=row_idx, column=4).fill = error_fill
        
        for col in range(1, 9):
            ws_uniform.column_dimensions[chr(64+col)].width = 20
        ws_uniform.column_dimensions['B'].width = 30
    else:
        # No uniform numbers in roster - add informational message
        ws_uniform.cell(row=1, column=1, value="Uniform Number Check Not Performed")
        ws_uniform.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws_uniform.cell(row=3, column=1, value="Reason: The roster data does not contain a column for uniform numbers.")
        ws_uniform.cell(row=4, column=1, value="To enable this check, ensure the roster includes a column named:")
        ws_uniform.cell(row=5, column=1, value="  - 'Uniform Number', 'Jersey Number', 'Number', 'Assigned Number', or similar")
        ws_uniform.column_dimensions['A'].width = 70
    
    # -------------------------------------------------------------------------
    # Team Summary Sheet
    # -------------------------------------------------------------------------
    ws_teams = wb.create_sheet("Team Summary")
    
    # Calculate per-team stats
    team_stats = []
    for team in roster['team_name'].unique():
        if pd.isna(team):
            continue
        
        team_roster = roster[roster['team_name'] == team]
        team_roster_indices = set(team_roster.index)
        
        not_ordered_indices = set([item['roster_index'] for item in not_ordered])
        team_not_ordered = len(team_roster_indices & not_ordered_indices)
        
        team_ordered = len(team_roster) - team_not_ordered
        
        team_stats.append({
            'Team Name': team,
            'Roster Count': len(team_roster),
            'Orders Placed': team_ordered,
            'Not Ordered': team_not_ordered,
            'Order Rate': f"{(team_ordered/len(team_roster))*100:.1f}%" if len(team_roster) > 0 else "N/A"
        })
    
    headers = ['Team Name', 'Roster Count', 'Orders Placed', 'Not Ordered', 'Order Rate']
    for col, header in enumerate(headers, 1):
        cell = ws_teams.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    for row_idx, stats in enumerate(sorted(team_stats, key=lambda x: x['Team Name']), 2):
        ws_teams.cell(row=row_idx, column=1, value=stats['Team Name'])
        ws_teams.cell(row=row_idx, column=2, value=stats['Roster Count'])
        ws_teams.cell(row=row_idx, column=3, value=stats['Orders Placed'])
        ws_teams.cell(row=row_idx, column=4, value=stats['Not Ordered'])
        ws_teams.cell(row=row_idx, column=5, value=stats['Order Rate'])
        
        # Color code low order rates
        if stats['Not Ordered'] > 0:
            ws_teams.cell(row=row_idx, column=4).fill = warning_fill
    
    ws_teams.column_dimensions['A'].width = 40
    for col in ['B', 'C', 'D', 'E']:
        ws_teams.column_dimensions[col].width = 15
    
    # -------------------------------------------------------------------------
    # Data Quality Warnings (for verification)
    # -------------------------------------------------------------------------
    ws_warnings = wb.create_sheet("Data Quality Warnings")
    
    # Add explanation row
    ws_warnings.cell(row=1, column=1, value="This tab highlights potential issues with the source data that may affect reconciliation accuracy.")
    ws_warnings.cell(row=1, column=1).font = Font(italic=True, color='666666')
    ws_warnings.merge_cells('A1:E1')
    
    if data_quality_warnings:
        # Header row
        warning_headers = ['Category', 'Severity', 'Count', 'Description', 'Details']
        for col, header in enumerate(warning_headers, 1):
            cell = ws_warnings.cell(row=2, column=col, value=header)
            cell.font = header_font
            cell.fill = PatternFill('solid', fgColor='FF6B6B')  # Red for warnings
            cell.border = thin_border
        
        current_row = 3
        for warning in data_quality_warnings:
            # Main warning row
            ws_warnings.cell(row=current_row, column=1, value=warning['category'])
            ws_warnings.cell(row=current_row, column=2, value=warning['severity'])
            ws_warnings.cell(row=current_row, column=3, value=warning['count'])
            ws_warnings.cell(row=current_row, column=4, value=warning['description'])
            
            # Color code by severity
            if warning['severity'] == 'High':
                severity_fill = error_fill
            elif warning['severity'] == 'Medium':
                severity_fill = warning_fill
            else:
                severity_fill = info_fill
            
            ws_warnings.cell(row=current_row, column=2).fill = severity_fill
            
            # Format details based on category
            details = warning.get('details', [])
            if details:
                if warning['category'] in ['Missing Emails', 'Invalid Email Format', 'Non-Numeric Uniform Numbers', 'Unusual Uniform Numbers', 'Unusual Characters in Names']:
                    # Player-based details
                    detail_str = '; '.join([f"{d.get('player_name', 'Unknown')} ({d.get('team_name', 'Unknown team')})" for d in details[:10]])
                elif warning['category'] == 'Duplicate Emails (Different Families)':
                    # Email-based details
                    detail_str = '; '.join([f"{d['email']}: {', '.join(d['players'][:3])}" for d in details[:5]])
                elif warning['category'] in ['Roster Teams with No Orders', 'Order Teams Not in Roster']:
                    # Team-based details
                    detail_str = '; '.join([f"{d['team_name']}" for d in details[:10]])
                else:
                    detail_str = str(details)[:200]
                
                if warning.get('has_more'):
                    detail_str += f" ... and {warning['count'] - len(details)} more"
                
                ws_warnings.cell(row=current_row, column=5, value=detail_str)
            
            current_row += 1
        
        ws_warnings.column_dimensions['A'].width = 35
        ws_warnings.column_dimensions['B'].width = 12
        ws_warnings.column_dimensions['C'].width = 10
        ws_warnings.column_dimensions['D'].width = 60
        ws_warnings.column_dimensions['E'].width = 80
    else:
        ws_warnings.cell(row=3, column=1, value="No data quality warnings detected.")
        ws_warnings.cell(row=3, column=1).font = Font(color='006400')  # Dark green
        ws_warnings.column_dimensions['A'].width = 40
    
    # -------------------------------------------------------------------------
    # Processed Roster Data (for verification)
    # -------------------------------------------------------------------------
    ws_roster_data = wb.create_sheet("Processed Roster Data")
    
    # Add explanation row
    ws_roster_data.cell(row=1, column=1, value="This tab shows how the roster data was interpreted. Review to verify correct column mapping.")
    ws_roster_data.cell(row=1, column=1).font = Font(italic=True, color='666666')
    ws_roster_data.merge_cells('A1:H1')
    
    roster_headers = ['Player First Name', 'Player Last Name', 'Full Name', 'Team Name', 'Parent Email 1', 'Parent Email 2', 'Uniform Number', 'Original Row']
    for col, header in enumerate(roster_headers, 1):
        cell = ws_roster_data.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = PatternFill('solid', fgColor='70AD47')  # Green for roster
        cell.border = thin_border
    
    for row_idx, (idx, row) in enumerate(roster.iterrows(), 3):
        ws_roster_data.cell(row=row_idx, column=1, value=row.get('player_first'))
        ws_roster_data.cell(row=row_idx, column=2, value=row.get('player_last'))
        ws_roster_data.cell(row=row_idx, column=3, value=row.get('player_name_original', row.get('player_name')))
        ws_roster_data.cell(row=row_idx, column=4, value=row.get('team_name'))
        ws_roster_data.cell(row=row_idx, column=5, value=row.get('parent_email_1_original', row.get('parent_email_1')))
        ws_roster_data.cell(row=row_idx, column=6, value=row.get('parent_email_2_original', row.get('parent_email_2')))
        ws_roster_data.cell(row=row_idx, column=7, value=row.get('uniform_number'))
        ws_roster_data.cell(row=row_idx, column=8, value=idx + 2)  # Original row number (accounting for header)
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws_roster_data.column_dimensions[col].width = 25
    
    # -------------------------------------------------------------------------
    # Processed Orders Data (for verification)
    # -------------------------------------------------------------------------
    ws_orders_data = wb.create_sheet("Processed Orders Data")
    
    # Add explanation row
    ws_orders_data.cell(row=1, column=1, value="This tab shows the Shopify order data used for reconciliation. Only uniform orders (Player Package/Pinnie products) are included.")
    ws_orders_data.cell(row=1, column=1).font = Font(italic=True, color='666666')
    ws_orders_data.merge_cells('A1:J1')
    
    orders_headers = ['Order Number', 'Order Date', 'Order Email', 'Player First Name', 'Player Last Name', 'Team Name', 'Uniform Number', 'Product Name', 'Is Uniform Order']
    for col, header in enumerate(orders_headers, 1):
        cell = ws_orders_data.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = PatternFill('solid', fgColor='5B9BD5')  # Blue for orders
        cell.border = thin_border
    
    for row_idx, (idx, row) in enumerate(orders.iterrows(), 3):
        ws_orders_data.cell(row=row_idx, column=1, value=row.get('order_number'))
        ws_orders_data.cell(row=row_idx, column=2, value=str(row.get('order_date'))[:10] if pd.notna(row.get('order_date')) else '')
        ws_orders_data.cell(row=row_idx, column=3, value=row.get('order_email_original'))
        ws_orders_data.cell(row=row_idx, column=4, value=row.get('player_first_original'))
        ws_orders_data.cell(row=row_idx, column=5, value=row.get('player_last_original'))
        ws_orders_data.cell(row=row_idx, column=6, value=row.get('order_team'))
        ws_orders_data.cell(row=row_idx, column=7, value=row.get('order_uniform_number'))
        ws_orders_data.cell(row=row_idx, column=8, value=row.get('product_name'))
        ws_orders_data.cell(row=row_idx, column=9, value='Yes' if row.get('is_uniform_order') else 'No')
        
        # Highlight non-uniform orders
        if not row.get('is_uniform_order'):
            for c in range(1, 10):
                ws_orders_data.cell(row=row_idx, column=c).fill = PatternFill('solid', fgColor='F2F2F2')
    
    ws_orders_data.column_dimensions['A'].width = 12
    ws_orders_data.column_dimensions['B'].width = 12
    ws_orders_data.column_dimensions['C'].width = 30
    ws_orders_data.column_dimensions['D'].width = 18
    ws_orders_data.column_dimensions['E'].width = 18
    ws_orders_data.column_dimensions['F'].width = 35
    ws_orders_data.column_dimensions['G'].width = 15
    ws_orders_data.column_dimensions['H'].width = 55
    ws_orders_data.column_dimensions['I'].width = 15
    
    # Save
    wb.save(output_path)
    print(f"Report saved to: {output_path}")


def print_console_summary(summary, not_ordered, true_duplicates, sibling_orders, wrong_team, misspellings, size_outliers, uniform_mismatches, uniform_check_performed, data_quality_warnings):
    """Print summary to console."""
    print(f"\n{'='*60}")
    print("RECONCILIATION SUMMARY")
    print(f"{'='*60}")
    
    for key, value in summary.items():
        print(f"  {key}: {value}")
    
    if not_ordered:
        print(f"\n--- Players Not Ordered ({len(not_ordered)}) ---")
        for item in not_ordered[:10]:
            print(f"  - {item['player_name']} ({item['team_name']})")
        if len(not_ordered) > 10:
            print(f"  ... and {len(not_ordered) - 10} more")
    
    if true_duplicates:
        print(f"\n--- True Duplicate Orders ({len(true_duplicates)}) ---")
        for item in true_duplicates[:10]:
            print(f"  - {item['player_name']}: {item['order_count']} orders (SAME PLAYER)")
        if len(true_duplicates) > 10:
            print(f"  ... and {len(true_duplicates) - 10} more")
    
    if sibling_orders:
        print(f"\n--- Sibling/Family Orders ({len(sibling_orders)}) - Info Only ---")
        for item in sibling_orders[:10]:
            siblings = ', '.join(item.get('sibling_names', []))
            print(f"  - {item['roster_player_name']} family also ordered: {siblings}")
        if len(sibling_orders) > 10:
            print(f"  ... and {len(sibling_orders) - 10} more")
    
    if wrong_team:
        print(f"\n--- Wrong Team ({len(wrong_team)}) ---")
        for item in wrong_team[:10]:
            roster_team = item['roster_team'][:35] if len(item['roster_team']) > 35 else item['roster_team']
            order_team = item['order_team'][:35] if len(item['order_team']) > 35 else item['order_team']
            print(f"  - {item['player_name']}: Roster={roster_team}... Order={order_team}...")
        if len(wrong_team) > 10:
            print(f"  ... and {len(wrong_team) - 10} more")
    
    if misspellings:
        print(f"\n--- Potential Misspellings ({len(misspellings)}) ---")
        for item in misspellings[:10]:
            print(f"  - Roster: '{item['roster_last_name']}' vs Order: '{item['order_last_name']}' ({item['similarity']*100:.0f}% similar)")
        if len(misspellings) > 10:
            print(f"  ... and {len(misspellings) - 10} more")
    
    if size_outliers:
        print(f"\n--- Size Outliers ({len(size_outliers)}) ---")
        for item in size_outliers[:10]:
            print(f"  - {item['player_name']} ({item['team_name']}): {item['issue_summary'][:60]}...")
        if len(size_outliers) > 10:
            print(f"  ... and {len(size_outliers) - 10} more")
    
    if uniform_check_performed:
        if uniform_mismatches:
            print(f"\n--- Uniform Number Mismatches ({len(uniform_mismatches)}) ---")
            for item in uniform_mismatches[:10]:
                print(f"  - {item['player_name']}: Roster #{item['roster_uniform_number']} vs Order #{item['order_uniform_number']}")
            if len(uniform_mismatches) > 10:
                print(f"  ... and {len(uniform_mismatches) - 10} more")
        else:
            print(f"\n--- Uniform Number Mismatches (0) ---")
            print(f"  No mismatches found!")
    else:
        print(f"\n--- Uniform Number Check ---")
        print(f"  Not performed: Roster data does not contain uniform numbers")
    
    # Data Quality Warnings
    if data_quality_warnings:
        high_warnings = [w for w in data_quality_warnings if w['severity'] == 'High']
        med_warnings = [w for w in data_quality_warnings if w['severity'] == 'Medium']
        low_warnings = [w for w in data_quality_warnings if w['severity'] == 'Low']
        
        print(f"\n--- Data Quality Warnings ({len(data_quality_warnings)} total) ---")
        if high_warnings:
            print(f"  HIGH SEVERITY ({len(high_warnings)}):")
            for w in high_warnings:
                print(f"    - {w['category']}: {w['count']} issues")
        if med_warnings:
            print(f"  MEDIUM SEVERITY ({len(med_warnings)}):")
            for w in med_warnings:
                print(f"    - {w['category']}: {w['count']} issues")
        if low_warnings:
            print(f"  LOW SEVERITY ({len(low_warnings)}):")
            for w in low_warnings:
                print(f"    - {w['category']}: {w['count']} issues")
    else:
        print(f"\n--- Data Quality Warnings ---")
        print(f"  No warnings detected!")


# ============================================================================
# MAIN FUNCTION
# ============================================================================

def main():
    parser = argparse.ArgumentParser(
        description='Signature Locker - Roster vs Orders Reconciliation',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python roster_order_reconciliation.py roster.xlsx orders.xlsx
  python roster_order_reconciliation.py roster.xlsx orders.xlsx --output report.xlsx
  python roster_order_reconciliation.py roster.xlsx orders.xlsx --roster-sheet "Players"
        """
    )
    
    parser.add_argument('roster_file', help='Path to roster Excel file')
    parser.add_argument('orders_file', help='Path to Program Director Report (orders) Excel file')
    parser.add_argument('--roster-sheet', '-s', help='Specific sheet name in roster file (auto-detected if not specified)')
    parser.add_argument('--output', '-o', default='reconciliation_report.xlsx', help='Output report filename')
    parser.add_argument('--misspelling-threshold', '-m', type=float, default=0.7, 
                        help='Similarity threshold for misspelling detection (0-1, default: 0.7)')
    
    args = parser.parse_args()
    
    # Validate input files
    if not Path(args.roster_file).exists():
        print(f"Error: Roster file not found: {args.roster_file}")
        sys.exit(1)
    
    if not Path(args.orders_file).exists():
        print(f"Error: Orders file not found: {args.orders_file}")
        sys.exit(1)
    
    print("\n" + "="*60)
    print("SIGNATURE LOCKER - ROSTER VS ORDERS RECONCILIATION")
    print("="*60)
    
    # Load data
    roster, has_uniform_numbers = load_roster_data(args.roster_file, args.roster_sheet)
    orders = load_orders_data(args.orders_file)
    
    # Perform matching
    print(f"\n{'='*60}")
    print("PERFORMING RECONCILIATION")
    print(f"{'='*60}")
    
    print("Matching by email...")
    email_matches = match_by_email(roster, orders)
    
    print("Matching by player name...")
    name_matches = match_by_name(roster, orders)
    
    # Find issues
    print("Finding players who have not ordered...")
    not_ordered = find_not_ordered(roster, email_matches, name_matches)
    
    print("Finding duplicate orders and sibling orders...")
    true_duplicates, sibling_orders = find_duplicate_orders(roster, orders, email_matches, name_matches)
    
    print("Finding wrong team orders...")
    wrong_team = find_wrong_team(roster, orders, email_matches, name_matches)
    
    print("Finding potential misspellings...")
    misspellings = find_misspellings(roster, orders, email_matches, args.misspelling_threshold)
    
    print("Analyzing size outliers...")
    size_outliers, team_size_stats = find_size_outliers(orders)
    
    print("Checking uniform number mismatches...")
    uniform_mismatches, uniform_check_performed = find_uniform_number_mismatches(
        roster, orders, email_matches, name_matches, has_uniform_numbers
    )
    
    print("Generating data quality warnings...")
    data_quality_warnings = generate_data_quality_warnings(roster, orders, has_uniform_numbers)
    
    # Generate summary
    summary = generate_summary(roster, orders, not_ordered, true_duplicates, sibling_orders, wrong_team, misspellings, size_outliers, uniform_mismatches, uniform_check_performed, data_quality_warnings)
    
    # Output
    print_console_summary(summary, not_ordered, true_duplicates, sibling_orders, wrong_team, misspellings, size_outliers, uniform_mismatches, uniform_check_performed, data_quality_warnings)
    save_report(args.output, summary, not_ordered, true_duplicates, sibling_orders, wrong_team, misspellings, size_outliers, uniform_mismatches, uniform_check_performed, data_quality_warnings, team_size_stats, roster, orders)
    
    print(f"\n{'='*60}")
    print("RECONCILIATION COMPLETE")
    print(f"{'='*60}")
    
    return 0


if __name__ == '__main__':
    sys.exit(main())
