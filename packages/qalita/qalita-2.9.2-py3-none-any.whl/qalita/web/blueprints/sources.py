"""
# QALITA (c) COPYRIGHT 2025 - ALL RIGHTS RESERVED -
"""

from typing import Any, Dict, Callable, Tuple, List
import os
import io
import csv
import yaml
from math import inf
from qalita.internal.utils import logger

from flask import (
    Blueprint,
    current_app,
    request,
    jsonify,
)
from qalita.commands.source import (
    validate_source_object,
    validate_source as _validate_all,
    push_single_programmatic,
)
from .helpers import get_platform_url


bp = Blueprint("sources", __name__)


@bp.get("/")
def list_sources():
    cfg = current_app.config["QALITA_CONFIG_OBJ"]
    cfg.load_source_config()
    sources = cfg.config.get("sources", [])

    # Resolve public platform URL using centralized helper
    platform_url = get_platform_url()

    return jsonify({
        "sources": sources,
        "platform_url": platform_url,
    })


@bp.post("/add")
def add_source_post():
    cfg = current_app.config["QALITA_CONFIG_OBJ"]
    cfg.load_source_config()

    name = request.form.get("name", "").strip()
    s_type = request.form.get("type", "").strip()

    # Build config from form according to type
    config_section = {}
    if s_type == "file":
        p = request.form.get("file_path", "").strip()
        if p:
            config_section["path"] = p
    elif s_type == "folder":
        p = request.form.get("folder_path", "").strip()
        if p:
            config_section["path"] = p
    elif s_type == "sqlite":
        fpath = request.form.get("sqlite_file_path", "").strip()
        if fpath:
            config_section["type"] = "sqlite"
            config_section["file_path"] = fpath
    elif s_type == "csv":
        p = request.form.get("csv_path", "").strip()
        if p:
            config_section["path"] = p
        d = request.form.get("csv_delimiter", ",").strip() or ","
        e = request.form.get("csv_encoding", "utf-8").strip() or "utf-8"
        h = request.form.get("csv_header") == "on"
        config_section.update(
            {
                "delimiter": d,
                "encoding": e,
                "has_header": h,
            }
        )
    elif s_type == "excel":
        p = request.form.get("excel_path", "").strip()
        if p:
            config_section["path"] = p
        sheet = request.form.get("excel_sheet", "").strip()
        header_row = request.form.get("excel_header_row", "1").strip()
        config_section.update(
            {
                "sheet": sheet,
                "header_row": header_row,
            }
        )
    elif s_type in ["mysql", "postgresql", "oracle", "mssql"]:
        config_section.update(
            {
                "type": s_type,
                "host": request.form.get("db_host", "").strip(),
                "port": request.form.get("db_port", "").strip(),
                "username": request.form.get("db_username", "").strip(),
                "password": request.form.get("db_password", "").strip(),
                "database": request.form.get("db_database", "").strip(),
                "table_or_query": request.form.get("db_table_or_query", "*").strip()
                or "*",
            }
        )
        # Optional schema for PostgreSQL and Oracle
        if s_type in ("postgresql", "oracle"):
            schema = request.form.get("db_schema", "").strip()
            if schema:
                config_section["schema"] = schema
    elif s_type == "mongodb":
        config_section.update(
            {
                "host": request.form.get("mongo_host", "").strip(),
                "port": request.form.get("mongo_port", "").strip(),
                "username": request.form.get("mongo_username", "").strip(),
                "password": request.form.get("mongo_password", "").strip(),
                "database": request.form.get("mongo_database", "").strip(),
            }
        )
    elif s_type == "s3":
        config_section.update(
            {
                "bucket": request.form.get("s3_bucket", "").strip(),
                "prefix": request.form.get("s3_prefix", "").strip(),
                "access_key": request.form.get("s3_access_key", "").strip(),
                "secret_key": request.form.get("s3_secret_key", "").strip(),
                "region": request.form.get("s3_region", "").strip(),
            }
        )
    elif s_type == "gcs":
        config_section.update(
            {
                "bucket": request.form.get("gcs_bucket", "").strip(),
                "prefix": request.form.get("gcs_prefix", "").strip(),
                "credentials_json": request.form.get("gcs_credentials", "").strip(),
            }
        )
    elif s_type == "azure_blob":
        config_section.update(
            {
                "container": request.form.get("az_container", "").strip(),
                "prefix": request.form.get("az_prefix", "").strip(),
                "connection_string": request.form.get("az_connection", "").strip(),
            }
        )
    elif s_type == "hdfs":
        config_section.update(
            {
                "namenode_host": request.form.get("hdfs_namenode", "").strip(),
                "port": request.form.get("hdfs_port", "").strip(),
                "user": request.form.get("hdfs_user", "").strip(),
                "path": request.form.get("hdfs_path", "").strip(),
            }
        )

    new_source = {
        "name": name,
        "type": s_type,
        "description": request.form.get("description", "").strip(),
        "reference": request.form.get("reference") == "on",
        "sensitive": request.form.get("sensitive") == "on",
        "visibility": request.form.get("visibility", "private"),
        "config": config_section,
    }

    if not validate_source_object(cfg, new_source, skip_connection=False):
        return jsonify({
            "ok": False,
            "error": "Validation failed. Check fields and connectivity.",
        }), 400

    cfg.config.setdefault("sources", []).append(new_source)
    cfg.save_source_config()
    # Validate all to compute status fields, then push only this source
    try:
        try:
            _validate_all.__wrapped__(cfg)  # type: ignore[attr-defined]
        except Exception:
            _validate_all(cfg)  # type: ignore[misc]
    except (SystemExit, Exception):
        # Gracefully skip full validation if agent context is not initialized
        pass
    try:
        push_single_programmatic(cfg, name, approve_public=False)
    except (SystemExit, Exception):
        # Skip push when not logged in / no .agent
        pass

    # Resolve platform_url using centralized helper
    platform_url = get_platform_url()

    # Try to get the created source id (if present after validate/push)
    src_id = None
    try:
        # Reload config to get potential IDs added during push/validate
        cfg.load_source_config()
        created = next(
            (s for s in cfg.config.get("sources", []) if s.get("name") == name), None
        )
        src_id = created.get("id") if isinstance(created, dict) else None
    except Exception:
        src_id = None

    return jsonify({
        "ok": True,
        "name": name,
        "platform_url": platform_url,
        "source_id": src_id,
        "message": f"Source '{name}' added successfully.",
    })


@bp.post("/edit/<name>")
def edit_source_post(name):
    cfg = current_app.config["QALITA_CONFIG_OBJ"]
    cfg.load_source_config()
    sources = cfg.config.get("sources", [])
    for i, src in enumerate(sources):
        if src.get("name") == name:
            new_name = request.form.get("name", src.get("name", "")).strip()
            new_type = request.form.get("type", src.get("type", "")).strip()
            new_desc = request.form.get(
                "description", src.get("description", "")
            ).strip()
            new_vis = request.form.get("visibility", src.get("visibility", "private"))
            new_ref = request.form.get("reference") == "on"
            new_sens = request.form.get("sensitive") == "on"
            # Build config
            config_section: Dict[str, Any] = {}
            if new_type == "file":
                p = request.form.get("file_path", "").strip()
                if p:
                    config_section["path"] = p
            elif new_type == "folder":
                p = request.form.get("folder_path", "").strip()
                if p:
                    config_section["path"] = p
            elif new_type == "sqlite":
                fpath = request.form.get("sqlite_file_path", "").strip()
                if fpath:
                    config_section["type"] = "sqlite"
                    config_section["file_path"] = fpath
            elif new_type in ["mysql", "postgresql", "oracle", "mssql"]:
                config_section.update(
                    {
                        "type": new_type,
                        "host": request.form.get("db_host", "").strip(),
                        "port": request.form.get("db_port", "").strip(),
                        "username": request.form.get("db_username", "").strip(),
                        "password": request.form.get("db_password", "").strip(),
                        "database": request.form.get("db_database", "").strip(),
                        "table_or_query": request.form.get(
                            "db_table_or_query", "*"
                        ).strip()
                        or "*",
                    }
                )
                if new_type in ("postgresql", "oracle"):
                    schema = request.form.get("db_schema", "").strip()
                    if schema:
                        config_section["schema"] = schema
            elif new_type == "mongodb":
                config_section.update(
                    {
                        "host": request.form.get("mongo_host", "").strip(),
                        "port": request.form.get("mongo_port", "").strip(),
                        "username": request.form.get("mongo_username", "").strip(),
                        "password": request.form.get("mongo_password", "").strip(),
                        "database": request.form.get("mongo_database", "").strip(),
                    }
                )
            elif new_type == "csv":
                p = request.form.get("csv_path", "").strip()
                if p:
                    config_section["path"] = p
                d = request.form.get("csv_delimiter", ",").strip() or ","
                e = request.form.get("csv_encoding", "utf-8").strip() or "utf-8"
                h = request.form.get("csv_header") == "on"
                config_section.update(
                    {
                        "delimiter": d,
                        "encoding": e,
                        "has_header": h,
                    }
                )
            elif new_type == "excel":
                p = request.form.get("excel_path", "").strip()
                if p:
                    config_section["path"] = p
                sheet = request.form.get("excel_sheet", "").strip()
                header_row = request.form.get("excel_header_row", "1").strip()
                config_section.update(
                    {
                        "sheet": sheet,
                        "header_row": header_row,
                    }
                )
            elif new_type == "s3":
                config_section.update(
                    {
                        "bucket": request.form.get("s3_bucket", "").strip(),
                        "prefix": request.form.get("s3_prefix", "").strip(),
                        "access_key": request.form.get("s3_access_key", "").strip(),
                        "secret_key": request.form.get("s3_secret_key", "").strip(),
                        "region": request.form.get("s3_region", "").strip(),
                    }
                )
            elif new_type == "gcs":
                config_section.update(
                    {
                        "bucket": request.form.get("gcs_bucket", "").strip(),
                        "prefix": request.form.get("gcs_prefix", "").strip(),
                        "credentials_json": request.form.get(
                            "gcs_credentials", ""
                        ).strip(),
                    }
                )
            elif new_type == "azure_blob":
                config_section.update(
                    {
                        "container": request.form.get("az_container", "").strip(),
                        "prefix": request.form.get("az_prefix", "").strip(),
                        "connection_string": request.form.get(
                            "az_connection", ""
                        ).strip(),
                    }
                )
            elif new_type == "hdfs":
                config_section.update(
                    {
                        "namenode_host": request.form.get("hdfs_namenode", "").strip(),
                        "port": request.form.get("hdfs_port", "").strip(),
                        "user": request.form.get("hdfs_user", "").strip(),
                        "path": request.form.get("hdfs_path", "").strip(),
                    }
                )

            updated = {
                "name": new_name,
                "type": new_type,
                "description": new_desc,
                "visibility": new_vis,
                "reference": new_ref,
                "sensitive": new_sens,
                "config": config_section if config_section else src.get("config", {}),
            }

            if not validate_source_object(
                cfg, updated, skip_connection=False, exclude_name=name
            ):
                return jsonify({
                    "ok": False,
                    "error": "Validation failed. Check fields and connectivity.",
                }), 400

            sources[i].update(updated)
            break
    cfg.save_source_config()
    # Re-validate and push only the edited source
    try:
        try:
            _validate_all.__wrapped__(cfg)  # type: ignore[attr-defined]
        except Exception:
            _validate_all(cfg)  # type: ignore[misc]
    except (SystemExit, Exception):
        pass
    try:
        push_single_programmatic(cfg, new_name, approve_public=False)
    except (SystemExit, Exception):
        pass
    return jsonify({
        "ok": True,
        "name": new_name,
        "message": f"Source '{new_name}' updated successfully.",
    })


@bp.post("/delete/<name>")
def delete_source_post(name):
    cfg = current_app.config["QALITA_CONFIG_OBJ"]
    cfg.load_source_config()
    cfg.config["sources"] = [
        s for s in cfg.config.get("sources", []) if s.get("name") != name
    ]
    cfg.save_source_config()
    return jsonify({
        "ok": True,
        "name": name,
        "message": f"Source '{name}' deleted successfully.",
    })


@bp.get("/pick-file")
def pick_file():
    try:
        import tkinter as tk  # type: ignore
        from tkinter import filedialog  # type: ignore

        root = tk.Tk()
        root.withdraw()
        try:
            root.wm_attributes("-topmost", 1)
        except Exception:
            pass
        path = filedialog.askopenfilename()
        root.update()
        root.destroy()
        return jsonify({"path": path})
    except Exception as exc:  # pragma: no cover
        logger.error(f"File picker unavailable: {exc}")
        return jsonify({"error": "File picker unavailable"}), 500


@bp.get("/pick-folder")
def pick_folder():
    try:
        import tkinter as tk  # type: ignore
        from tkinter import filedialog  # type: ignore

        root = tk.Tk()
        root.withdraw()
        try:
            root.wm_attributes("-topmost", 1)
        except Exception:
            pass
        path = filedialog.askdirectory()
        root.update()
        root.destroy()
        return jsonify({"path": path})
    except Exception as exc:  # pragma: no cover
        logger.error(f"Folder picker unavailable: {exc}")
        return jsonify({"error": "Folder picker unavailable"}), 500


# Lightweight validation endpoint for the edit form (no save/push)
@bp.post("/validate")
def validate_source_form():
    cfg = current_app.config["QALITA_CONFIG_OBJ"]
    cfg.load_source_config()

    # Accept either JSON or form-encoded payload
    data = request.get_json(silent=True) or request.form or {}

    def _get(name: str) -> str:
        v = data.get(name)
        if isinstance(v, list):
            return (v[0] or "").strip()
        return (v or "").strip()

    name = _get("name")
    s_type = _get("type")
    original_name = _get("original_name") or None

    # Build config from payload mirroring add/edit logic
    config_section: Dict[str, Any] = {}
    if s_type == "file":
        p = _get("file_path")
        if p:
            config_section["path"] = p
    elif s_type == "folder":
        p = _get("folder_path")
        if p:
            config_section["path"] = p
    elif s_type == "sqlite":
        fpath = _get("sqlite_file_path")
        if fpath:
            config_section["type"] = "sqlite"
            config_section["file_path"] = fpath
    elif s_type in ["mysql", "postgresql", "oracle", "mssql"]:
        config_section.update(
            {
                "type": s_type,
                "host": _get("db_host"),
                "port": _get("db_port"),
                "username": _get("db_username"),
                "password": _get("db_password"),
                "database": _get("db_database"),
                "table_or_query": _get("db_table_or_query") or "*",
            }
        )
        if s_type in ("postgresql", "oracle"):
            schema = _get("db_schema")
            if schema:
                config_section["schema"] = schema
    elif s_type == "mongodb":
        config_section.update(
            {
                "host": _get("mongo_host"),
                "port": _get("mongo_port"),
                "username": _get("mongo_username"),
                "password": _get("mongo_password"),
                "database": _get("mongo_database"),
            }
        )
    elif s_type == "csv":
        p = _get("csv_path")
        if p:
            config_section["path"] = p
        d = _get("csv_delimiter") or ","
        e = _get("csv_encoding") or "utf-8"
        h = (_get("csv_header").lower() == "on") if isinstance(data, dict) else False
        config_section.update(
            {
                "delimiter": d,
                "encoding": e,
                "has_header": h,
            }
        )
    elif s_type == "excel":
        p = _get("excel_path")
        if p:
            config_section["path"] = p
        sheet = _get("excel_sheet")
        header_row = _get("excel_header_row") or "1"
        config_section.update(
            {
                "sheet": sheet,
                "header_row": header_row,
            }
        )
    elif s_type == "s3":
        config_section.update(
            {
                "bucket": _get("s3_bucket"),
                "prefix": _get("s3_prefix"),
                "access_key": _get("s3_access_key"),
                "secret_key": _get("s3_secret_key"),
                "region": _get("s3_region"),
            }
        )
    elif s_type == "gcs":
        config_section.update(
            {
                "bucket": _get("gcs_bucket"),
                "prefix": _get("gcs_prefix"),
                "credentials_json": _get("gcs_credentials"),
            }
        )
    elif s_type == "azure_blob":
        config_section.update(
            {
                "container": _get("az_container"),
                "prefix": _get("az_prefix"),
                "connection_string": _get("az_connection"),
            }
        )
    elif s_type == "hdfs":
        config_section.update(
            {
                "namenode_host": _get("hdfs_namenode"),
                "port": _get("hdfs_port"),
                "user": _get("hdfs_user"),
                "path": _get("hdfs_path"),
            }
        )

    candidate = {
        "name": name,
        "type": s_type,
        "description": _get("description"),
        "reference": (
            (_get("reference").lower() == "on") if isinstance(data, dict) else False
        ),
        "sensitive": (
            (_get("sensitive").lower() == "on") if isinstance(data, dict) else False
        ),
        "visibility": _get("visibility") or "private",
        "config": config_section,
    }

    ok = validate_source_object(
        cfg, candidate, skip_connection=False, exclude_name=original_name
    )
    if ok:
        return jsonify({"ok": True, "message": "Source is valid."})
    return (
        jsonify(
            {
                "ok": False,
                "message": "Validation failed. Check fields and connectivity.",
            }
        ),
        400,
    )


# ---- Data preview (modular handlers) ----


def _qalita_home() -> str:
    try:
        cfg = current_app.config.get("QALITA_CONFIG_OBJ")
        return getattr(cfg, "qalita_home", os.path.expanduser("~/.qalita"))
    except Exception:
        return os.path.expanduser("~/.qalita")


def _read_qalita_conf() -> Dict[str, Any]:
    path = os.path.join(_qalita_home(), "sources-conf.yaml")
    try:
        if not os.path.isfile(path):
            return {}
        with open(path, "r", encoding="utf-8") as f:
            data = yaml.safe_load(f) or {}
        if not isinstance(data, dict):
            return {}
        return data
    except Exception:
        return {}


def _find_source_by_id(conf: Dict[str, Any], src_id: str) -> Dict[str, Any] | None:
    try:
        sources = conf.get("sources") if isinstance(conf.get("sources"), list) else []
        for s in sources:
            if not isinstance(s, dict):
                continue
            if str(s.get("id", "")) == str(src_id):
                return s
    except Exception:
        return None
    return None


def _csv_preview(source: Dict[str, Any], options: Dict[str, Any] | None = None) -> Tuple[Dict[str, Any], int]:
    from qalita.internal.utils import validate_file_path
    
    cfg = source.get("config") if isinstance(source.get("config"), dict) else {}
    raw_path = cfg.get("path") or source.get("path")
    if not raw_path:
        return {"ok": False, "message": "CSV file path not specified"}, 400
    
    # Validate and sanitize the file path to prevent path injection
    try:
        path = validate_file_path(raw_path)
    except (ValueError, FileNotFoundError) as e:
        logger.warning(f"CSV preview path validation failed: {e}")
        return {"ok": False, "message": "CSV file not found or invalid path"}, 404
    delimiter = cfg.get("delimiter") or ","
    encoding = cfg.get("encoding") or "utf-8"
    has_header = bool(cfg.get("has_header", True))
    max_rows = 200
    try:
        if isinstance(options, dict) and options.get("limit") is not None:
            limit_raw = options.get("limit")
            if isinstance(limit_raw, str) and limit_raw.strip().lower() == "all":
                max_rows = 10_000  # hard cap for safety
            else:
                v = int(limit_raw)
                if v > 0:
                    max_rows = min(v, 10_000)
    except Exception:
        pass
    verbose = bool(isinstance(options, dict) and options.get("verbose"))
    if verbose:
        logger.info(f"[preview.csv] path={path}, delimiter='{delimiter}', encoding='{encoding}', has_header={has_header}, limit={max_rows}")
    # Read a small sample for preview
    headers: List[str] = []
    rows: List[List[str]] = []
    try:
        with open(path, "r", encoding=encoding, newline="") as f:
            reader = csv.reader(f, delimiter=delimiter)
            first_row: List[str] | None = None
            for i, row in enumerate(reader):
                if i == 0:
                    first_row = [str(c) for c in row]
                    if has_header:
                        headers = first_row
                        continue
                if not headers:
                    # Generate default headers
                    headers = [f"col_{idx+1}" for idx in range(len(first_row or row))]
                rows.append([str(c) for c in row])
                if len(rows) >= max_rows:
                    break
    except Exception as exc:
        return {"ok": False, "message": f"Failed to read CSV: {exc}"}, 500
    if verbose:
        logger.info(f"[preview.csv] produced headers={len(headers)}, rows={len(rows)}")
    return {"ok": True, "view": {"type": "table", "headers": headers, "rows": rows}}, 200


PreviewHandler = Callable[[Dict[str, Any], Dict[str, Any] | None], Tuple[Dict[str, Any], int]]


_PREVIEW_HANDLERS: Dict[str, PreviewHandler] = {
    "csv": _csv_preview,
}


def _excel_preview(source: Dict[str, Any], options: Dict[str, Any] | None = None) -> Tuple[Dict[str, Any], int]:
    from qalita.internal.utils import validate_file_path
    
    try:
        import openpyxl  # type: ignore
    except Exception:
        return {"ok": False, "message": "Excel preview requires 'openpyxl' to be installed"}, 500

    cfg = source.get("config") if isinstance(source.get("config"), dict) else {}
    raw_path = cfg.get("path") or source.get("path")
    if not raw_path:
        return {"ok": False, "message": "Excel file path not specified"}, 400
    
    # Validate and sanitize the file path to prevent path injection
    try:
        path = validate_file_path(raw_path)
    except (ValueError, FileNotFoundError) as e:
        logger.warning(f"Excel preview path validation failed: {e}")
        return {"ok": False, "message": "Excel file not found or invalid path"}, 404

    sheet_name = (cfg.get("sheet") or "").strip()
    header_row_raw = cfg.get("header_row")
    try:
        header_row = int(header_row_raw) if header_row_raw not in (None, "") else 1
    except Exception:
        header_row = 1

    max_rows = 200
    try:
        if isinstance(options, dict) and options.get("limit") is not None:
            limit_raw = options.get("limit")
            if isinstance(limit_raw, str) and limit_raw.strip().lower() == "all":
                max_rows = 10_000
            else:
                v = int(limit_raw)
                if v > 0:
                    max_rows = min(v, 10_000)
    except Exception:
        pass
    verbose = bool(isinstance(options, dict) and options.get("verbose"))
    if verbose:
        logger.info(f"[preview.excel] path={path}, sheet='{sheet_name or '(active)'}', header_row={header_row}, limit={max_rows}")

    headers: List[str] = []
    rows: List[List[str]] = []
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active
        # Iterate rows; openpyxl is 1-based index
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            values = list(row)
            if header_row and i == header_row:
                headers = ["" if v is None else str(v) for v in values]
                # replace empty header names
                headers = [h if h else f"col_{idx+1}" for idx, h in enumerate(headers)]
                continue
            if not headers:
                # No header row: synthesize from width of first row
                headers = [f"col_{idx+1}" for idx in range(len(values))]
            rows.append(["" if v is None else str(v) for v in values])
            if len(rows) >= max_rows:
                break
        try:
            wb.close()
        except Exception:
            pass
    except Exception as exc:
        return {"ok": False, "message": f"Failed to read Excel: {exc}"}, 500

    if verbose:
        logger.info(f"[preview.excel] produced headers={len(headers)}, rows={len(rows)}")
    return {"ok": True, "view": {"type": "table", "headers": headers, "rows": rows}}, 200


_PREVIEW_HANDLERS["excel"] = _excel_preview


@bp.get("/preview")
def preview_source():
    """Preview a source by id from ~/.qalita/sources-conf.yaml

    For now supports CSV. Returns a generic 'view' payload consumable by the View Panel.
    """
    src_id = (request.args.get("source_id") or request.args.get("id") or "").strip()
    if not src_id:
        return jsonify({"ok": False, "message": "Missing source_id"}), 400
    conf = _read_qalita_conf()
    src = _find_source_by_id(conf, src_id)
    if not src:
        return jsonify({"ok": False, "message": "Source not found"}), 404
    # validate can be boolean or string: 'valid' | 'invalid'
    val = src.get("validate")
    is_valid = False
    if isinstance(val, bool):
        is_valid = val
    elif isinstance(val, str):
        is_valid = val.strip().lower() == "valid"
    if not is_valid:
        return jsonify({"ok": False, "message": "Source not validated"}), 400
    s_type = (src.get("type") or "").strip().lower()
    handler = _PREVIEW_HANDLERS.get(s_type)
    if not handler:
        return jsonify({"ok": False, "message": f"Preview not supported for type '{s_type}'"}), 400
    # Parse optional preview options
    options: Dict[str, Any] = {}
    try:
        limit = request.args.get("limit")
        if limit is not None:
            options["limit"] = limit
        # Verbose/debug mode
        verbose_flag = request.args.get("verbose") or request.args.get("debug")
        if isinstance(verbose_flag, str) and verbose_flag.strip().lower() in ("1", "true", "yes", "on"):  # noqa: E501
            options["verbose"] = True
    except Exception:
        pass
    if options.get("verbose"):
        logger.info(f"[preview] source_id={src_id}, type={s_type}, options={options}")
    body, code = handler(src, options)
    # Attach debug info when verbose
    try:
        if options.get("verbose") and isinstance(body, dict):
            cfg = src.get("config") if isinstance(src.get("config"), dict) else {}
            debug = {
                "source_id": src_id,
                "type": s_type,
                "options": options,
                "config_keys": list(cfg.keys()) if isinstance(cfg, dict) else [],
            }
            # add quick counts when table
            if isinstance(body.get("view"), dict):
                v = body["view"]
                if v.get("type") == "table":
                    debug["headers_count"] = len(v.get("headers") or [])
                    debug["rows_count"] = len(v.get("rows") or [])
            body["debug"] = debug
    except Exception:
        pass
    return jsonify(body), code
