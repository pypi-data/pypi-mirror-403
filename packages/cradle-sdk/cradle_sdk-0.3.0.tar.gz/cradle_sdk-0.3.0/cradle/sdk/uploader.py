import csv as csv_std
import tempfile
import time
from abc import ABC, abstractmethod
from collections.abc import Callable
from datetime import UTC, datetime, timedelta
from pathlib import Path
from typing import Self

import openpyxl
import pyarrow as pa
import pyarrow.compute as pc
import pyarrow.csv as csv
import pyarrow.parquet as pq
from typing_extensions import TypeVar

from cradle.sdk.client import DataLoadClient
from cradle.sdk.types.common import ContextProject, ContextRound
from cradle.sdk.types.data import (
    ArrayColumn,
    ArrayType,
    Column,
    ColumnType,
    DataLoadCreate,
    DataLoadResponse,
    DataLoadState,
    FileFormat,
    PrimitiveColumn,
    PrimitiveType,
    StructColumn,
    StructType,
    TableLoadConfig,
    TypeNames,
)


class Transform(ABC):
    """Base class for all transformations applied to UploadFile objects."""

    @abstractmethod
    def apply(self, data: pa.Table) -> pa.Table:
        """Apply the transformation to an UploadFile and return a new UploadFile."""
        ...

    def __call__(self, data: pa.Table) -> pa.Table:
        return self.apply(data)


class File:
    def __init__(self, data: pa.Table, source: Path | None, relative_path: Path | None = None):
        self.data = data
        self.source = source
        self.relative_path = relative_path
        self.columns: list[Column] = []

        try:
            for field in data.schema:
                self.columns.append(_convert_column(field.name, field.type))
        except _NullTypeError as e:
            raise ValueError(f"""
                {e}

                For CSV input files, provide an explicit column type (e.g. `convert_options=pyarrow.csv.ConvertOptions(column_types={{"c1": pa.string()}})`)
                or omit the affected column (via `ConvertOptions(include_columns=[...])`).
            """) from e


class FileSet:
    def __init__(self, files: list[File]):
        if len(files) == 0:
            raise ValueError("file list must not be empty")

        for f in files[1:]:
            if f.columns != files[0].columns:
                raise ValueError(
                    f"all files must have the same schema, found {f.columns} but expected {files[0].columns}"
                )

        self.files = files

    @property
    def columns(self) -> list[Column]:
        return self.files[0].columns

    def merge(self, other: "FileSet") -> "FileSet":
        if self.columns != other.columns:
            raise ValueError(f"FileSets have incompatible columns {self.columns} and {other.columns}")
        return FileSet(self.files + other.files)

    def transform(self, *transforms: Transform) -> "FileSet":
        """Apply all transformations to all files.

        Use this primarily to make the data in the files fit the schema of the target table.
        Avoid transformations that prepare data for a specific task. Those use cases should
        instead be deferred to views over base tables in the platform.
        """

        # Iterate over transforms in the outer loop so we know for each transform
        # whether it succeeds for all files before proceeding. This should make
        # debugging overall easier.
        files = self.files[:]
        for t in transforms:
            files = [File(t(f.data), f.source, f.relative_path) for f in files]
        return FileSet(files)

    @classmethod
    def _from_files(
        cls,
        open_fn: Callable[[Path], pa.Table],
        directory: str | Path,
        pattern: str | list[str],
    ) -> Self:
        if isinstance(directory, str):
            directory = Path(directory)
        if isinstance(pattern, str):
            pattern = [pattern]

        files = []
        for p in pattern:
            for f in directory.glob(p):
                relative_path = f.relative_to(directory)
                files.append(File(data=open_fn(f), source=f, relative_path=relative_path))
        return cls(files)

    @classmethod
    def from_parquet(
        cls,
        directory: Path | str,
        pattern: str | list[str],
        **parquet_options,
    ) -> Self:
        """Create a `FileSet` from Parquet source files.

        Args:
            directory: The base directory relative to which the file path pattern is matched.
            pattern: The glob pattern or list of patterns to match the Parquet files. The resulting paths
                relative to `directory` will be submitted as metadata with the uploaded files and be visible
                in the platform.
            **parquet_options: Additional options to pass to `pyarrow.parquet.read_table` to control how the
                Parquet files are read.
        """

        def _open(file: Path) -> pa.Table:
            return pq.read_table(file, **parquet_options)

        return cls._from_files(_open, directory, pattern)

    @classmethod
    def from_csv(
        cls,
        directory: Path | str,
        pattern: str | list[str],
        **csv_options,
    ) -> Self:
        """Create a `FileSet` from CSV source files.

        Args:
            directory: The base directory relative to which the file path pattern is matched.
            pattern: The glob pattern or list of patterns to match the CSV files. The resulting paths
                relative to `directory` will be submitted as metadata with the uploaded files and be visible
                in the platform.
            **csv_options: Additional options to pass to `pyarrow.csv.read_csv` to control how the
                CSV files are read.
        """

        def _open(file: Path) -> pa.Table:
            return csv.read_csv(file, **csv_options)

        return cls._from_files(_open, directory, pattern)

    @classmethod
    def from_excel(
        cls,
        directory: Path | str,
        pattern: str | list[str],
        sheet: str | int,
        **csv_options,
    ) -> Self:
        """Create a `FileSet` from Excel files (xlsx or xls).

        Excel files are first converted to CSV as-is and then read as CSV files like in `from_csv`.

        Args:
            directory: The base directory relative to which the file path pattern is matched.
            pattern: The glob pattern or list of patterns to match the Excel files. The resulting paths
                relative to `directory` will be submitted as metadata with the uploaded files and be visible
                in the platform.
            sheet: The sheet name or index to read from the Excel file. If multiple sheets should be read,
                call `from_excel` individually for each sheet.
            **csv_options: Additional options to pass to `pyarrow.csv.read_csv` to control how the
                converted CSV files are read.
        """

        def _open(file: Path) -> pa.Table:
            with tempfile.NamedTemporaryFile(suffix=".csv") as temp_csv:
                tmp_csv = Path(temp_csv.name)

                workbook = openpyxl.load_workbook(file, read_only=True, data_only=True)
                if isinstance(sheet, str):
                    if sheet not in workbook.sheetnames:
                        raise ValueError(f"Sheet '{sheet}' not found in {file}")
                    data = workbook[sheet]
                else:
                    if sheet >= len(workbook.sheetnames):
                        raise ValueError(f"Sheet index {sheet} is out of range for {file}")
                    data = workbook[workbook.sheetnames[sheet]]

                with tmp_csv.open("w", encoding="utf-8", newline="") as f:
                    writer = csv_std.writer(f)
                    for row in data.rows:
                        writer.writerow([cell.value for cell in row])

                return csv.read_csv(tmp_csv, **csv_options)

        return cls._from_files(_open, directory, pattern)


class Uploader:
    def __init__(self, client: DataLoadClient, context: ContextProject | ContextRound):
        """Initialize the Uploader with a Client instance.

        Args:
            client: An authenticated Client instance
            context: The context of the upload
        """
        self._load_id: int | None = None
        self._client = client
        self._context = context
        self._tables: dict[str, FileSet] = {}

    def add_files(self, table_name: str, files: FileSet) -> None:
        """Add files to the upload for the specified table.

        All files for a given table must have the same schema, which must be compatible
        with the schema of the target table. It is compatible if all required columns exist
        and have the right type.

        Args:
            table_name: The name of the table the files will be ingested into.
            files: The files to upload.
        """
        existing = self._tables.get(table_name)
        if existing is not None:
            files = existing.merge(files)
        self._tables[table_name] = files

    def load(self) -> DataLoadResponse:
        """Upload all files for this upload.

        For each added `UploadFile` this will upload the data in Parquet format to be ingested
        into the target table. Additionally, the source, if set, will be uploaded for archival purposes.

        The upload will be left in pending state until finalize is called.

        Returns:
            UploadResponse: The upload state retrieved from the server after all uploads completed.
        """
        if self._load_id is not None:
            raise ValueError("Upload already started")

        load = self._client.create(
            DataLoadCreate(
                context=self._context,
                tables={
                    name: TableLoadConfig(columns=files.columns, format=FileFormat.PARQUET)
                    for name, files in self._tables.items()
                },
            )
        )
        self._load_id = load.id

        with tempfile.TemporaryDirectory(prefix="cradle-upload-") as temp_dir:
            for ref, t in self._tables.items():
                base_dir = Path(temp_dir) / ref
                base_dir.mkdir(exist_ok=False, parents=True)

                for i, file in enumerate(t.files):
                    description = None
                    source_file_id = None
                    if file.source is not None:
                        print(f"Uploading source file {file.source}...")
                        resp = self._client.upload_file(
                            load_id=self._load_id,
                            file=file.source,
                            filepath=file.relative_path,
                            table_reference=None,
                        )
                        description = f"Source file: {file.source} ({resp.id})"
                        source_file_id = resp.id

                    path = base_dir / f"{i}.parquet"
                    print(f"Uploading file {path}...")
                    pq.write_table(file.data, path)
                    self._client.upload_file(
                        load_id=self._load_id,
                        file=path,
                        filepath=(
                            file.relative_path.with_suffix(".parquet") if file.relative_path is not None else None
                        ),
                        table_reference=ref,
                        description=description,
                        source_file_id=source_file_id,
                    )

        return self._client.get(load_id=self._load_id)

    def finalize(self, wait: bool = True, timeout: float = 60) -> DataLoadResponse:
        """Finalize the upload. Calls `upload()` if it has not already been called.

        Args:
            wait: Whether to wait for the upload to complete.
            timeout: Timeout in seconds for waiting for the upload to complete.

        Returns:
            UploadResponse: The most recent upload state retrieved from the server.
        """
        if self._load_id is None:
            load_id = self.load().id
        else:
            load_id = self._load_id

        u = self._client.finalize(load_id=load_id)
        if not wait:
            return u

        def _fn():
            u = self._client.get(load_id=load_id)
            if u.state == DataLoadState.FAILED:
                raise RuntimeError(f"Upload failed: {', '.join(u.errors)}")
            return u.state == DataLoadState.COMPLETED, u

        return _wait_for_condition(_fn, datetime.now(tz=UTC) + timedelta(seconds=timeout))


class RenameColumns(Transform):
    """Transform that renames columns in the dataframe."""

    def __init__(self, column_map: dict[str, str]):
        """Args:
        column_map: A dictionary mapping old column names to new column names
        """
        self.column_map = column_map

    def apply(self, data: pa.Table) -> pa.Table:
        names = [self.column_map.get(name, name) for name in data.column_names]
        return data.rename_columns(names)


class FilterRows(Transform):
    """Filter rows in the dataframe based on a provided condition.

    Example usage: FilterRows(lambda t: pa.compute.greater(t.column("a"), 42))
    """

    def __init__(self, condition: Callable[[pa.Table], pa.Array]):
        """Args:
        condition: A function that computes a boolean array over all rows in the provided tables.
            Rows for which the result index is False, will be dropped from the resulting table.
        """
        self._condition = condition

    def apply(self, data: pa.Table) -> pa.Table:
        return data.filter(self._condition(data))


class DropNullRows(FilterRows):
    """Transform that removes rows with null values in the specified columns."""

    def __init__(self, columns: str | list[str] | None = None):
        """Args:
        columns: Column or list of column names to check for nulls. If None, checks all columns.
        """
        if isinstance(columns, str):
            columns = [columns]
        self._columns = columns

    def apply(self, data: pa.Table) -> pa.Table:
        columns = self._columns
        if columns is None:
            columns = data.column_names
        mask = pa.array([True] * len(data))

        for col in columns:
            if col not in data.column_names:
                raise ValueError(f"Column '{col}' does not exist in table")
            mask = pc.and_(mask, pc.is_valid(data[col]))  # type: ignore[reportAttributeAccessIssue] seem to actually be missing in venv, but invoking them works ???

        return data.filter(mask)


T = TypeVar("T")


def _wait_for_condition(cond_fn: Callable[[], tuple[bool, T]], deadline: datetime, interval: float = 1) -> T:
    """Wait for the condition function to return True or the deadline is exceeded. It will be executed at least once."""

    while True:
        done, result = cond_fn()
        if done:
            return result
        if datetime.now(tz=UTC) > deadline:
            raise TimeoutError("Deadline exceeded")
        time.sleep(interval)


class _NullTypeError(ValueError): ...


def _convert_type(type_: pa.DataType) -> ColumnType:
    if pa.types.is_integer(type_):
        return PrimitiveType(type=TypeNames.INT64)
    if pa.types.is_floating(type_):
        return PrimitiveType(type=TypeNames.FLOAT64)
    if pa.types.is_boolean(type_):
        return PrimitiveType(type=TypeNames.BOOL)
    if pa.types.is_string(type_):
        return PrimitiveType(type=TypeNames.STRING)
    if pa.types.is_struct(type_):
        return StructType(columns=[_convert_column(f.name, f.type) for f in type_.fields])
    if pa.types.is_list(type_):
        return ArrayType(item_type=_convert_type(type_.value_type))
    if pa.types.is_null(type_):
        raise _NullTypeError(f"Unsupported PyArrow type {type_}")
    else:
        raise ValueError(f"Unsupported PyArrow type {type_}")


def _convert_column(name: str, type_: pa.DataType) -> Column:
    if pa.types.is_integer(type_):
        return PrimitiveColumn(name=name, type=TypeNames.INT64, nullable=True)
    if pa.types.is_floating(type_):
        return PrimitiveColumn(name=name, type=TypeNames.FLOAT64, nullable=True)
    if pa.types.is_boolean(type_):
        return PrimitiveColumn(name=name, type=TypeNames.BOOL, nullable=True)
    if pa.types.is_string(type_):
        return PrimitiveColumn(name=name, type=TypeNames.STRING, nullable=True)
    if pa.types.is_struct(type_):
        columns = [_convert_column(f.name, f.type) for f in type_.fields]
        return StructColumn(name=name, columns=columns, nullable=True)
    if pa.types.is_list(type_):
        return ArrayColumn(name=name, item_type=_convert_type(type_.value_type))
    if pa.types.is_null(type_):
        raise _NullTypeError(f"Unsupported PyArrow type {type_} for column {name}")
    else:
        raise ValueError(f"Unsupported PyArrow type {type_} for column {name}")
