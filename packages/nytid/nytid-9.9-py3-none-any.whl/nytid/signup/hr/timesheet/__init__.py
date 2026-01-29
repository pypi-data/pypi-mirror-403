"""Generate timesheets for TAs"""

import datetime
import io
import pkgutil
import os.path
import PIL.Image
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment

def test():
    """Tests the module"""
    personnr = "123456-7890"
    name = "Alexander Baltatzis"
    email = "alba@kth.se"
    hourly_salary = 150
    course_leader = ('Daniel Bosk', 'dbosk@kth.se')
    HoD = "Karl Meinke"
    events = []
    events.append({"datum":"2022-12-03",
                  'tid':"8-10",
                  'kurskod':'DD1321',
                  'typ':"handl",
                  'timmar':2,
                  'koeff':1.33,
                  'omr_tid':2*1.33,
                  'belopp':hourly_salary*2*1.33})
    events.append({"datum":"2022-12-04",
                  'tid':"8-10",
                  'kurskod':'DD1321',
                  'typ':"övning",
                  'timmar':2,
                  'koeff':3,
                  'omr_tid':2*3,
                  'belopp':hourly_salary*2*3})
    events.append({"datum":"2022-12-05",
                  'tid':"8-10",
                  'kurskod':'DD1321',
                  'typ':"handl",
                  'timmar':2.1,
                  'koeff':1.33,
                  'omr_tid':2.1*1.33,
                  'belopp':hourly_salary*2.1*1.33})
    events.append({"datum":"2022-12-05",
                  'tid':"10-12",
                  'kurskod':'DD1310',
                  'typ':"handl",
                  'timmar':2,
                  'koeff':1.33,
                  'omr_tid':2*1.33,
                  'belopp':hourly_salary*2*1.33})
    removed_events = []
    removed_events.append({"datum":"2022-11-06",
                           'tid':"8-10",
                           'kurskod':'DD1321',
                           'typ':"handl",
                           'timmar':2,
                           'koeff':1.33,
                           'omr_tid':2*1.33,
                           'belopp':hourly_salary*2*1.33})

    make_xlsx(personnr, name, email, events,
              course_leader, HoD,
              course_leader_signature="signature.png",
              removed_events=removed_events)

def fit_image(img, height=None, width=None):
    """
    Rescales `img` to fit either `height` or `width` pixels --- not both, we 
    keep aspect ratio.

    `img` can be a string (path to file) or an image (PIL.Image).
    """
    img = Image(img)

    if height:
        scale = height/img.height
    elif width:
        scale = width/img.width
    else:
        raise KeyError("fit_image: neither height, nor width given.")

    img.height *= scale
    img.width *= scale
    return img

def make_xlsx(personnummer, name, email, events,
              course_leader, HoD,
              org = "JH", project = "1102",
              hourly_salary = 165,
              output = None,
              course_leader_signature = None,
              HoD_signature = None,
              logo = "kth.png",
              removed_events=None):
    """
    Generates a time report for a TA:
    - `name` and `email` are name and email for the TA.
    - `events` is a list of dictionaries, each dictionary containing:
        - datum: str
        - tid: str
        - kurskod: str
        - typ: str
        - timmar: float, koeff: float, omr_tid: float
        - belopp: float
    - `course_leader` is a tuple (name: str, email: str)
    - `HoD` the name of the Head of Department
    - `org`, `project` is the organization and project, both strings.
    - `hourly_salary` is the hourly salary, float.
    - `output` is the desired output (xlsx) filename.
    - `course_leader_signature` is an image (or path to one) containing the 
      course responsible's signature.
    - `HoD_signature` is an image (or path to one) containing the Head of 
      Department's signature.
    - `logo` is an image (or path to one) containing the logotype of the 
      university.
    """
    login = email.replace("@kth.se", "")
    if not output:
        output = login + "_tid_" + datetime.date.today().strftime("%Y-%m-%d.xlsx")

    wb = Workbook()
    ark = wb.active


    #############################################################
    # Logo
    ark.title = login + " " + datetime.date.today().strftime("%Y-%b")
    if isinstance(logo, str):
        logo = os.path.expanduser(logo)
    try:
        logo = fit_image(logo, height=80)
    except FileNotFoundError:
        img_data = PIL.Image.open(
                io.BytesIO(pkgutil.get_data(__name__, "kth.png")))
        logo = fit_image(img_data, height=80)
        ark.add_image(logo, "A1")
    else:
        ark.add_image(logo, "A1")

    #############################################################
    # kolumnstorlekar
    ark.column_dimensions['A'].width = 12  # 'Schemalagd tid'
    ark.column_dimensions['B'].width = 11  # 'Typ'           
    ark.column_dimensions['C'].width = 7   # 'timmar'      
    ark.column_dimensions['D'].width = 6   # 'koeff'         
    ark.column_dimensions['E'].width = 12  # 'omräknad tid'        
    ark.column_dimensions['F'].width = 6   # 'Timlön'     
    ark.column_dimensions['G'].width = 9   # 'Belopp'        
    ark.column_dimensions['H'].width = 9

    #############################################################
    # Börja på rad 6
    rad = "6"
    ark['A' + rad] = "Redovisning av arbetade timmar, time sheet"

    rad = incr(rad, 2)
    ark['A' + rad] = "Namn"
    ark['B' + rad] = name
    ark['B' + rad].fill = PatternFill(start_color="00EEECE1", end_color="00EEECE1", fill_type="solid")
    ark['C' + rad].fill = PatternFill(start_color="00EEECE1", 
                                      end_color="00EEECE1", fill_type="solid")

    ark['E' + rad] = 'Personnr'
    ark['F' + rad] = personnummer
    ark['F' + rad].fill = PatternFill(start_color="00EEECE1", end_color="00EEECE1", fill_type="solid")
    ark['G' + rad].fill = PatternFill(start_color="00EEECE1", 
                                      end_color="00EEECE1", fill_type="solid")

    rad = incr(rad)
    ark['A' + rad] = 'E-post'
    ark['B' + rad] = email
    ark['B' + rad].fill = PatternFill(start_color="00EEECE1", end_color="00EEECE1", fill_type="solid")
    ark['C' + rad].fill = PatternFill(start_color="00EEECE1", 
                                      end_color="00EEECE1", fill_type="solid")

    rad = incr(rad, 2)
    ark['A' + rad] = 'Kurskod'
    ark['B' + rad] = ''
    kurskoder = []
    for kol in events:
        if 'kurskod' in kol and kol['kurskod'] not in kurskoder:
            ark['B' + rad].value += kol['kurskod'] + " "
            kurskoder.append( kol['kurskod'] )
    ark['B' + rad].fill = PatternFill(start_color="00EEECE1", 
                                      end_color="00EEECE1", fill_type="solid")
    ark['C' + rad].fill = PatternFill(start_color="00EEECE1", 
                                      end_color="00EEECE1", fill_type="solid")

    rad = incr(rad, 2)
    ark['A' + rad] = 'Timmar ska anges inklusive förberedelsetid enligt schablon'
    rad = incr(rad)
    ark['A' + rad] = 'Ange typ av undervisning övning, handledning, möte, etc'

    rad = incr(rad, 2)
    ark['A' + rad] = 'Datum'
    ark['B' + rad] = 'Typ'
    ark['C' + rad] = 'Klocktimmar'
    ark['D' + rad] = 'koeff'
    ark['E' + rad] = 'Lönetimmar'
    ark['F' + rad] = 'Timlön'
    ark['G' + rad] = 'Belopp'

    for kol in ['C', 'D', 'E', 'F', 'G']:
        ark[kol+rad].alignment = Alignment(horizontal="right")


    #############################################################
    # Summering på sista raden 
    rad = incr(rad)   
    sist = incr(rad, len(events))
    ark['E'+sist].font = Font(bold=True)  
    ark['G'+sist].font = Font(bold=True)  
    ark['E'+sist].fill = PatternFill(start_color="00EEECE1", end_color="00EEECE1", fill_type="solid")
    #ark['G'+sist].fill = PatternFill(start_color="00C0C0C0", end_color="00C0C0C0", fill_type="solid")

    #############################################################
    # Matris med timredovisningen
    for i, kol in enumerate(events):
        ark['A'+rad].value = kol['datum']
        ark['B'+rad] = kol['typ']
        ark['C'+rad] = float(kol['timmar'])
        ark['D'+rad] = float(kol['koeff'])
        ark['E'+rad].value = float(kol['omr_tid'])
        ark['F'+rad].value = float(hourly_salary)
        ark['G'+rad].value = float(kol['belopp'])

        if i % 2 == 0:
            for kol in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
                ark[kol+rad].fill = PatternFill(start_color="00E0E0E0", end_color="00E0E0E0", fill_type="solid")
        
        if i == 0:
            tidsumma = "=SUM(E"+rad
            ark['G'+sist].value = '=G'+rad
        else:
            tidsumma += ',E'+rad 
            ark['G'+sist].value += '+G'+rad

        rad = incr(rad)

    if events:
        ark['E'+sist].value = tidsumma + ')'

    if removed_events:
        #############################################################
        # Borttagna tillfällen
        # Summering på sista raden 
        rad = incr(sist, 2)
        ark['A'+rad].value = "Felaktigt rapporterade tillfällen"
        ark['A'+rad].font = Font(bold=True)
        rad = incr(rad)
        ark['A'+rad].value = "Dessa tillfällen har rapporterats felaktigt " \
                             "på tidigare tidrapporter"
        rad = incr(rad)
        ark['A'+rad].value = "och ska därför tas bort."
        rad = incr(rad, 2)
        sist = incr(rad, len(removed_events))
        ark['E'+sist].font = Font(bold=True)  
        ark['G'+sist].font = Font(bold=True)  
        ark['E'+sist].fill = PatternFill(start_color="00EEECE1", end_color="00EEECE1", fill_type="solid")
        #ark['G'+sist].fill = PatternFill(start_color="00C0C0C0", end_color="00C0C0C0", fill_type="solid")

        #############################################################
        # Matris med timredovisningen
        for i, kol in enumerate(removed_events):
            ark['A'+rad].value = kol['datum']
            ark['B'+rad] = kol['typ']
            ark['C'+rad] = -float(kol['timmar'])
            ark['D'+rad] = float(kol['koeff'])
            ark['E'+rad].value = -float(kol['omr_tid'])
            ark['F'+rad].value = float(hourly_salary)
            ark['G'+rad].value = -float(kol['belopp'])

            if i % 2 == 0:
                for kol in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
                    ark[kol+rad].fill = PatternFill(start_color="00E0E0E0", end_color="00E0E0E0", fill_type="solid")
            
            if i == 0:
                tidsumma = "=SUM(E"+rad
                ark['G'+sist].value = '=G'+rad
            else:
                tidsumma += ',E'+rad 
                ark['G'+sist].value += '+G'+rad

            rad = incr(rad)

        if events:
            ark['E'+sist].value = tidsumma + ')'


            
    #############################################################
    # Kontering
    rad = incr(sist, 3)
    ark['A'+rad].value = "Kontering"
    rad = incr(rad)
    ark['A'+rad].value = "Org.enhet"
    ark['A'+rad].fill  = PatternFill(start_color="00EEECE1", end_color="00EEECE1", fill_type="solid")
    ark['B'+rad].value = "Projekt"
    ark['B'+rad].fill = PatternFill(start_color="00EEECE1", end_color="00EEECE1", fill_type="solid")
    rad = incr(rad)
    ark['A'+rad].value = org
    ark['A'+rad].fill  = PatternFill(start_color="00EEECE1", end_color="00EEECE1", fill_type="solid")
    ark['B'+rad].value = project
    ark['B'+rad].fill  = PatternFill(start_color="00EEECE1", end_color="00EEECE1", fill_type="solid")

    #############################################################
    # Underskrift
    rad = incr(rad, 4)

    ark['A'+rad].value = "__________________________________"
    ark['E'+rad].value = "__________________________________"

    if isinstance(HoD_signature, str):
        HoD_signature = os.path.expanduser(HoD_signature)

    if HoD_signature:
        HoD_signature = fit_image(HoD_signature, height=60)
        ark.add_image(HoD_signature, "A"+incr(rad, -2))

    if isinstance(course_leader_signature, str):
        course_leader_signature = os.path.expanduser(course_leader_signature)

    if course_leader_signature:
        course_leader_signature = fit_image(course_leader_signature, height=60)
        ark.add_image(course_leader_signature, "E"+incr(rad, -2))

    rad = incr(rad)

    ark['A'+rad].value = "Ekonomisk attest " + HoD
    ark['E'+rad].value = "Kursansvarig"

    rad = incr(rad)
    ark['E'+rad].value = course_leader[0]
    rad = incr(rad)
    ark['E'+rad].value = course_leader[1]

    rad = incr(rad, 2)
    ark['A'+rad].value = "Underskriven blankett lämnas till HR"

    wb.save(output)

def incr(rad, i=1):
    """
    Bumps the row of a cell. Takes a string containing a number,
    increase by `i`.
    """
    return str(int(rad)+i)


#################################################################
#
# main - test
#
if __name__ == "__main__":
    test()

