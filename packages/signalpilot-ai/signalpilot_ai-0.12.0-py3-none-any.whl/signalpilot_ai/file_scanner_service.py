"""
File Scanner Service for SignalPilot AI.
Handles file scanning, schema extraction, and directory tracking.
"""

import os
import json
import hashlib
import threading
import asyncio
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Any, Tuple
import pandas as pd
import numpy as np
import concurrent.futures
from concurrent.futures import ThreadPoolExecutor
import pyarrow.dataset as ds
from openpyxl import load_workbook

from .cache_service import get_cache_service, get_file_scan_cache_manager
from .log_utils import print


class FileScannerService:
    """Service for scanning directories and extracting file schemas"""
    
    def __init__(self):
        self.cache_service = get_cache_service()
        self.file_scan_cache = get_file_scan_cache_manager()
        self._lock = threading.RLock()
        
        # Data file extensions
        self.DATA_EXTENSIONS = {'.csv', '.json', '.xlsx', '.xls', '.parquet',
                               '.feather', '.hdf5', '.h5', '.sql', '.db', '.sqlite', '.tsv', '.txt', '.ipynb'}
        
        # Directories to exclude from search
        self.EXCLUDE_DIRS = {'.git', '.ipynb_checkpoints', 'node_modules', '__pycache__', 
                            '.venv', 'venv', 'env', '.pytest_cache', '.mypy_cache', 
                            'dist', 'build', '.tox', 'logs', '.vscode'}
        
        # Thread pool for async schema extraction
        self._executor = ThreadPoolExecutor(max_workers=4, thread_name_prefix="schema_extractor")
        
        # Dedicated thread pool for directory scanning operations (I/O blocking)
        self._dir_executor = ThreadPoolExecutor(max_workers=2, thread_name_prefix="dir_scanner")
    
    async def _list_directory_async(self, directory_path: Path) -> List[Path]:
        """List directory contents using dedicated thread pool to avoid blocking."""
        loop = asyncio.get_running_loop()
        return await loop.run_in_executor(self._dir_executor, lambda: list(directory_path.iterdir()))
    
    def __del__(self):
        """Cleanup thread pools when service is destroyed."""
        if hasattr(self, '_dir_executor'):
            self._dir_executor.shutdown(wait=False)
        if hasattr(self, '_executor'):
            self._executor.shutdown(wait=False)
    
    def _get_directory_hash(self, directory: str) -> str:
        """Generate a hash for a directory path"""
        return hashlib.md5(directory.encode()).hexdigest()[:16]
    
    
    def _is_binary_file(self, filepath: str, chunk_size: int = 512) -> bool:
        """Ultra-fast binary file detection with minimal I/O"""
        try:
            with open(filepath, 'rb') as f:
                chunk = f.read(chunk_size)
                if not chunk:
                    return False
                # Fast null byte check - if any null bytes, it's binary
                if b'\x00' in chunk:
                    return True
                # Quick printable ratio check using bytes directly
                printable = sum(1 for b in chunk if 32 <= b <= 126 or b in (9, 10, 13))
                return (printable / len(chunk)) < 0.7
        except (IOError, OSError):
            return True

    def _parse_json_array_simple(self, filepath: str, max_items: int = 5) -> Tuple[List[Any], bool]:
        """
        Simple JSON array parsing that reads first chunk and extracts items.
        More robust for very large files.
        Returns (items_list, is_truncated)
        """
        import json
        
        try:
            # Read first 50KB of the file
            with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                chunk = f.read(50000)  # 50KB chunk
                
                # Try to find the opening bracket
                bracket_pos = chunk.find('[')
                if bracket_pos == -1:
                    # Not an array, try as single object
                    f.seek(0)
                    try:
                        single_obj = json.load(f)
                        return [single_obj], False
                    except:
                        return [], False
                
                # Find the first few complete JSON objects in the chunk
                items = []
                current_pos = bracket_pos + 1
                
                while len(items) < max_items and current_pos < len(chunk):
                    # Find the next complete JSON object
                    brace_count = 0
                    start_pos = current_pos
                    in_string = False
                    escape_next = False
                    
                    for i in range(current_pos, len(chunk)):
                        char = chunk[i]
                        
                        if escape_next:
                            escape_next = False
                            continue
                        elif char == '\\':
                            escape_next = True
                            continue
                        elif char == '"' and not escape_next:
                            in_string = not in_string
                        elif not in_string:
                            if char == '{':
                                brace_count += 1
                            elif char == '}':
                                brace_count -= 1
                                if brace_count == 0:
                                    # Found complete object
                                    try:
                                        obj_str = chunk[start_pos:i+1].strip()
                                        if obj_str.startswith(','):
                                            obj_str = obj_str[1:].strip()
                                        if obj_str:
                                            obj = json.loads(obj_str)
                                            items.append(obj)
                                    except:
                                        pass
                                    current_pos = i + 1
                                    break
                        elif char == ',' and brace_count == 0 and not in_string:
                            # End of current item
                            try:
                                obj_str = chunk[start_pos:i].strip()
                                if obj_str.startswith(','):
                                    obj_str = obj_str[1:].strip()
                                if obj_str:
                                    obj = json.loads(obj_str)
                                    items.append(obj)
                            except:
                                pass
                            current_pos = i + 1
                            break
                    else:
                        break
                
                # Check if there's more content
                remaining = f.read(1000)
                is_truncated = len(remaining) > 0 or len(items) == max_items
                
                return items, is_truncated
                
        except Exception:
            return [], False

    def _read_json_file(self, filepath: str, max_items: int = 5) -> Tuple[List[Any], str, bool]:
        """
        Read JSON file with smart loading strategy.
        Returns (data_list, file_format, is_truncated)
        file_format: 'object', 'array', or 'jsonl'
        """
        import json
        
        try:
            # Check file size
            file_size = os.path.getsize(filepath)
            size_mb = file_size / (1024 * 1024)
            
            # For small files (< 10MB), use simple json.load()
            if size_mb < 10:
                try:
                    with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                        data = json.load(f)
                    
                    if isinstance(data, list):
                        return data[:max_items], 'array', len(data) > max_items
                    elif isinstance(data, dict):
                        return [data], 'object', False
                    else:
                        return [data], 'primitive', False
                except json.JSONDecodeError:
                    # Try as JSONL if JSON parsing fails
                    pass
            
            # For large files or if JSON parsing failed, try incremental parsing
            try:
                # First check if it's JSONL (line-delimited JSON)
                with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                    first_line = f.readline().strip()
                    if first_line and first_line.startswith('{') and first_line.endswith('}'):
                        # Likely JSONL format
                        f.seek(0)
                        items = []
                        for i, line in enumerate(f):
                            if i >= max_items:
                                break
                            line = line.strip()
                            if line:
                                try:
                                    item = json.loads(line)
                                    items.append(item)
                                except:
                                    continue
                        return items, 'jsonl', True  # Assume truncated for large files
                
                # Try simple array parsing
                items, is_truncated = self._parse_json_array_simple(filepath, max_items)
                if items:
                    return items, 'array', is_truncated
                
                # Fallback: try to read first few lines as individual JSON objects
                try:
                    with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                        items = []
                        for i, line in enumerate(f):
                            if i >= max_items:
                                break
                            line = line.strip()
                            if line and (line.startswith('{') or line.startswith('[')):
                                try:
                                    obj = json.loads(line)
                                    items.append(obj)
                                except:
                                    continue
                        if items:
                            return items, 'jsonl', True
                except:
                    pass
                
                # Final fallback: try to read as single object
                with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                    content = f.read(10000)  # Read first 10KB
                    try:
                        data = json.loads(content)
                        if isinstance(data, dict):
                            return [data], 'object', True
                    except:
                        pass
                
                return [], 'unknown', False
                
            except Exception:
                return [], 'unknown', False
                
        except Exception:
            return [], 'unknown', False

    def _infer_json_type(self, value: Any) -> str:
        """Infer data type from JSON value"""
        if value is None:
            return 'null'
        elif isinstance(value, bool):
            return 'boolean'
        elif isinstance(value, int):
            return 'integer'
        elif isinstance(value, float):
            return 'float'
        elif isinstance(value, str):
            # Try to detect if it's a numeric string
            try:
                float(value)
                return 'numeric_string'
            except:
                return 'string'
        elif isinstance(value, list):
            return 'array'
        elif isinstance(value, dict):
            return 'object'
        else:
            return 'unknown'

    def _extract_keys_from_object(self, obj: dict, prefix: str = "", max_depth: int = 3, max_fields: int = 50) -> Dict[str, Dict[str, Any]]:
        """Recursively extract keys from nested objects with truncation limits"""
        keys = {}
        
        if max_depth <= 0 or len(keys) >= max_fields:
            return keys
            
        for i, (key, value) in enumerate(obj.items()):
            if len(keys) >= max_fields:
                break
                
            full_key = f"{prefix}.{key}" if prefix else key
            value_type = self._infer_json_type(value)
            
            # Truncate sample values to 50 characters max
            sample_value = None
            if not isinstance(value, (dict, list)):
                sample_str = str(value)
                sample_value = sample_str[:50] + "..." if len(sample_str) > 50 else sample_str
            
            keys[full_key] = {
                'type': value_type,
                'sample_value': sample_value,
                'is_nested': isinstance(value, (dict, list)),
                'depth': len(prefix.split('.')) if prefix else 0
            }
            
            # Recursively extract nested keys (reduced depth)
            if isinstance(value, dict) and max_depth > 1 and len(keys) < max_fields:
                nested_keys = self._extract_keys_from_object(value, full_key, max_depth - 1, max_fields - len(keys))
                keys.update(nested_keys)
            elif isinstance(value, list) and value and isinstance(value[0], dict) and len(keys) < max_fields:
                # For arrays of objects, analyze the first object only
                nested_keys = self._extract_keys_from_object(value[0], f"{full_key}[0]", max_depth - 1, max_fields - len(keys))
                keys.update(nested_keys)
        
        return keys

    def _analyze_json_structure(self, data_list: List[Any], filepath: str, filename: str, current_mtime: float) -> Dict[str, Any]:
        """
        Analyze JSON structure and extract comprehensive schema information.
        """
        if not data_list:
            return {
                'success': False,
                'error': f'No data found in JSON file. File may be too large ({os.path.getsize(filepath) / (1024*1024):.1f}MB) or malformed. Try using a smaller sample or check file format.'
            }
        
        # Analyze structure
        all_keys = {}
        sample_data = []
        total_records = len(data_list)
        max_depth = 0
        structure_types = set()
        
        # Process each record
        for i, record in enumerate(data_list):
            if isinstance(record, dict):
                structure_types.add('object')
                # Extract keys from this record
                record_keys = self._extract_keys_from_object(record)
                
                # Merge with all_keys, keeping track of types and sample values
                for key, key_info in record_keys.items():
                    if key not in all_keys:
                        all_keys[key] = {
                            'type': key_info['type'],
                            'sample_values': [],
                            'is_consistent': True,
                            'depth': key_info['depth']
                        }
                    
                    # Add sample value
                    if key_info['sample_value']:
                        all_keys[key]['sample_values'].append(key_info['sample_value'])
                    
                    # Check type consistency
                    if all_keys[key]['type'] != key_info['type']:
                        all_keys[key]['is_consistent'] = False
                        # Update to most common type or 'mixed'
                        all_keys[key]['type'] = 'mixed'
                    
                    # Track max depth
                    max_depth = max(max_depth, key_info['depth'])
                
                # Add to sample data (first 5 records, truncated)
                if i < 5:
                    # Truncate sample data to essential fields only
                    truncated_record = {}
                    for j, (k, v) in enumerate(record.items()):
                        if j >= 5:  # Only first 5 fields
                            break
                        if isinstance(v, (dict, list)):
                            truncated_record[k] = f"<{type(v).__name__}>"
                        else:
                            val_str = str(v)
                            truncated_record[k] = val_str[:50] + "..." if len(val_str) > 50 else val_str
                    sample_data.append(truncated_record)
                    
            elif isinstance(record, list):
                structure_types.add('array')
                if i < 3:
                    # Truncate arrays to first 5 items
                    truncated_array = record[:5]
                    sample_data.append(truncated_array)
            else:
                structure_types.add('primitive')
                if i < 3:
                    # Truncate primitive values
                    val_str = str(record)
                    truncated_val = val_str[:50] + "..." if len(val_str) > 50 else val_str
                    sample_data.append(truncated_val)
        
        # Convert keys to columns format with truncation
        columns = []
        for key, key_info in all_keys.items():
            # Determine final type
            final_type = key_info['type']
            if not key_info['is_consistent']:
                final_type = 'mixed'
            
            # Get unique sample values (limit to 2, truncate to 30 chars each)
            unique_samples = []
            for sample in list(set(key_info['sample_values']))[:2]:
                sample_str = str(sample)
                truncated_sample = sample_str[:30] + "..." if len(sample_str) > 30 else sample_str
                unique_samples.append(truncated_sample)
            
            columns.append({
                'name': key,
                'dataType': final_type,
                'description': f'Field {key} of type {final_type}',
                'sample_values': unique_samples,
                'is_consistent': key_info['is_consistent'],
                'depth': key_info['depth']
            })
        
        # Determine file format description
        if 'object' in structure_types and len(structure_types) == 1:
            format_desc = 'JSON object'
        elif 'array' in structure_types and len(structure_types) == 1:
            format_desc = 'JSON array'
        elif 'jsonl' in str(filepath).lower():
            format_desc = 'Line-delimited JSON (JSONL)'
        else:
            format_desc = 'Mixed JSON structure'
        
        # Calculate statistics
        unique_keys_count = len(all_keys)
        consistent_keys = sum(1 for k in all_keys.values() if k['is_consistent'])
        
        return {
            'success': True,
            'fileId': filepath,
            'fileName': filename,
            'filePath': filepath,
            'fileType': 'json',
            'extractedAt': datetime.now().isoformat(),
            'summary': f'{format_desc} with {total_records} record{"s" if total_records != 1 else ""}, {unique_keys_count} unique field{"s" if unique_keys_count != 1 else ""}',
            'totalRows': total_records,
            'totalColumns': unique_keys_count,
            'columns': columns,
            'sampleData': sample_data,
            'fileMtime': current_mtime,
            'structure_info': {
                'format_type': format_desc,
                'max_depth': max_depth,
                'consistent_fields': consistent_keys,
                'total_fields': unique_keys_count,
                'structure_types': list(structure_types)
            }
        }

    def _read_file_preview_optimized(self, filepath: str, max_chars: int = 2000, max_newlines: int = 5) -> Tuple[str, bool]:
        """
        Ultra-fast file preview reader using efficient buffered reading.
        Reads first 2000 characters OR first 3 newlines, whichever comes first.
        Returns (content, is_truncated)
        """
        try:
            file_size = os.path.getsize(filepath)
            
            # For very small files, read directly
            if file_size <= max_chars:
                with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                    content = f.read()
                    return content, False
            
            # For larger files, read in chunks and stop at limits
            with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read(max_chars)
                
                # Count newlines in the content we read
                newline_count = content.count('\n')
                
                # If we have 3 or fewer newlines, we're good
                if newline_count <= max_newlines:
                    # Check if there's more content (for truncation flag)
                    next_chunk = f.read(1)
                    is_truncated = bool(next_chunk)
                    return content, is_truncated
                
                # If we have more than 3 newlines, truncate to first 3
                lines = content.split('\n', max_newlines + 1)
                if len(lines) > max_newlines:
                    # We have more than max_newlines, so truncate
                    truncated_content = '\n'.join(lines[:max_newlines])
                    return truncated_content, True
                else:
                    # Exactly max_newlines, check if there's more content
                    next_chunk = f.read(1)
                    is_truncated = bool(next_chunk)
                    return content, is_truncated
                    
        except (UnicodeDecodeError, IOError, OSError):
            try:
                # Fallback for problematic files
                with open(filepath, 'rb') as f:
                    raw_bytes = f.read(max_chars)
                    content = raw_bytes.decode('utf-8', errors='replace')
                    # Apply newline limit to fallback content too
                    lines = content.split('\n')
                    if len(lines) > max_newlines:
                        content = '\n'.join(lines[:max_newlines])
                        return content, True
                    return content, len(raw_bytes) == max_chars
            except Exception:
                return f"<Error reading file: {filepath}>", False
    
    def _get_file_type_info(self, filepath: str, extension: str) -> Dict[str, Any]:
        """Get optimized metadata about file type"""
        file_info = {
            'extension': extension,
            'is_csv': extension == '.csv',
            'is_tsv': extension == '.tsv',
            'is_json': extension == '.json',
            'is_parquet': extension == '.parquet',
            'is_xlsx': extension == '.xlsx',
            'is_ipynb': extension == '.ipynb',
            'is_text': extension in ['.txt', '.md', '.py', '.js', '.ts', '.html', '.xml', '.ipynb'],
            'is_data': extension in ['.csv', '.tsv', '.json', '.jsonl', '.parquet', '.xlsx'],
            'is_binary': extension in ['.parquet', '.xlsx']  # Will be set later based on actual binary detection
        }
        
        try:
            file_info['size_bytes'] = os.path.getsize(filepath)
            file_info['last_modified'] = datetime.fromtimestamp(os.path.getmtime(filepath)).isoformat()
        except (IOError, OSError):
            file_info['size_bytes'] = 0
            file_info['last_modified'] = None
            
        return file_info
    
    def _process_csv_preview(self, content: str, filepath: str) -> Dict[str, Any]:
        """Fast CSV preview processing"""
        # Split into lines efficiently, limit to what we need
        newline_pos = content.find('\n')
        if newline_pos == -1:
            # Single line file
            header = content.strip()
            sample_rows = []
        else:
            # Multi-line file - get header and up to 5 sample rows
            lines = content.split('\n', 6)  # Get at most 6 lines (header + 5 samples)
            header = lines[0] if lines[0] else None
            sample_rows = [line for line in lines[1:6] if line.strip()]
        
        result = {
            'type': 'csv',
            'preview_type': 'header_and_sample',
            'header': header,
            'sample_rows': sample_rows
        }
        
        if header:
            result['estimated_columns'] = header.count(',') + 1
        
        return result
    
    def _process_json_preview(self, content: str, filepath: str) -> Dict[str, Any]:
        """Enhanced JSON structure analysis with actual parsing"""
        import json
        
        result = {
            'type': 'json',
            'preview_type': 'parsed_structure'
        }
        
        try:
            # Try to parse the content as JSON
            data = json.loads(content)
            
            if isinstance(data, dict):
                result['structure'] = 'object'
                result['keys'] = list(data.keys())[:5]  # First 5 keys only
                result['key_count'] = len(data.keys())
                
                # Analyze key types (truncated)
                key_types = {}
                for i, (key, value) in enumerate(data.items()):
                    if i >= 5:  # Limit to first 5 keys
                        break
                    key_types[key] = self._infer_json_type(value)
                result['key_types'] = key_types
                
            elif isinstance(data, list):
                result['structure'] = 'array'
                result['item_count'] = len(data)
                result['sample_items'] = data[:3]  # First 3 items
                
                # Analyze item types
                if data:
                    item_types = [self._infer_json_type(item) for item in data[:5]]
                    result['item_types'] = item_types
                    
                    # If items are objects, get their keys
                    if isinstance(data[0], dict):
                        result['sample_keys'] = list(data[0].keys())[:5]
                        
            else:
                result['structure'] = 'primitive'
                result['value'] = str(data)[:50]
                result['value_type'] = self._infer_json_type(data)
                
        except json.JSONDecodeError:
            # Fallback to line-by-line parsing for JSONL
            lines = content.strip().split('\n')
            if lines and lines[0].strip().startswith('{'):
                result['structure'] = 'jsonl'
                result['line_count'] = len(lines)
                
                # Try to parse first few lines
                sample_objects = []
                for line in lines[:2]:  # Only first 2 lines
                    try:
                        obj = json.loads(line.strip())
                        sample_objects.append(obj)
                    except:
                        continue
                
                if sample_objects:
                    result['sample_objects'] = sample_objects
                    # Get keys from first object
                    if isinstance(sample_objects[0], dict):
                        result['sample_keys'] = list(sample_objects[0].keys())[:5]
            else:
                # Fallback to basic analysis
                content_stripped = content.lstrip()
                if not content_stripped:
                    result['structure'] = 'empty'
                else:
                    first_char = content_stripped[0]
                    if first_char == '{':
                        result['structure'] = 'object'
                        result['estimated_keys'] = content_stripped.count('":')
                    elif first_char == '[':
                        result['structure'] = 'array'
                        result['estimated_items'] = content_stripped.count(',') + 1
                    else:
                        result['structure'] = 'unknown'
                        
        except Exception as e:
            result['structure'] = 'error'
            result['error'] = str(e)
        
        return result
    
    def _get_data_type(self, dtype_str: str) -> str:
        """Map pandas dtypes to readable types"""
        if dtype_str.startswith('int'):
            return 'integer'
        elif dtype_str.startswith('float'):
            return 'float'
        elif dtype_str == 'bool':
            return 'boolean'
        elif dtype_str.startswith('datetime'):
            return 'datetime'
        elif dtype_str == 'object':
            return 'string'
        else:
            return 'string'
    
    def _analyze_dataframe(self, df: pd.DataFrame, file_type: str) -> Dict[str, Any]:
        """Analyze DataFrame and return schema information"""
        # Get basic info
        total_rows_sample = len(df)
        total_columns = len(df.columns)
        
        # Get column information
        columns = []
        sample_data = []
        
        for col in df.columns:
            dtype = str(df[col].dtype)
            data_type = self._get_data_type(dtype)
            
            # For object columns, try to infer if it's a date
            if dtype == 'object' and not df[col].dropna().empty:
                sample_val = df[col].dropna().iloc[0]
                try:
                    pd.to_datetime(sample_val)
                    data_type = 'date'
                except:
                    pass
            
            columns.append({
                'name': str(col),
                'dataType': data_type,
                'description': f'Column {col} of type {data_type}'
            })
        
        # Get sample data (first 5 rows)
        for _, row in df.head(5).iterrows():
            sample_data.append(row.fillna('').astype(str).tolist())
        
        return {
            'success': True,
            'totalRows': total_rows_sample,
            'totalColumns': total_columns,
            'columns': columns,
            'sampleData': sample_data,
            'summary': f'{file_type.upper()} file with {total_columns} columns'
        }
    
    async def scan_directories(self, paths: List[str], force_refresh: bool = False, workspace_root: str = None) -> Dict[str, Any]:
        """Scan multiple directories and return file information without reprocessing unchanged files.

        Strategy:
        - Quickly enumerate directory contents (recursively)
        - For files, return cached rich entries if mtime unchanged
        - Otherwise, return a lightweight placeholder and process the file asynchronously to populate cache
        - Schema extraction reuses existing caching/async logic
        """
        all_files: List[Dict[str, Any]] = []
        scanned_directories: List[Dict[str, Any]] = []
        cached_count = 0

        if workspace_root is None:
            workspace_root = os.getcwd()
        original_root_path = Path(workspace_root)

        scanned_directories_cache = self.file_scan_cache.get_scanned_directories()

        for path in paths:
            scanned_directory_cache = next((dir for dir in scanned_directories_cache if dir.get('path') == path), None)
            dir_cached = True
            abs_path = os.path.abspath(os.getcwd() if path in ('.', './') else path)

            files_for_dir: List[Dict[str, Any]] = []
            base_dir = Path(abs_path)
            if not base_dir.exists() or not base_dir.is_dir():
                scanned_directories.append({
                    'path': path,
                    'file_count': 0,
                    'scanned_at': datetime.now().isoformat(),
                })
                continue

            # Walk directory tree with shallow Path.iterdir recursion to keep control and avoid reading file contents
            stack: List[Tuple[Path, int]] = [(base_dir, 0)]
            max_depth = 10
            while stack:
                current_dir, depth = stack.pop()
                if depth > max_depth:
                    continue
                try:
                    items = await self._list_directory_async(current_dir)
                    for item in items:
                        # Skip hidden and excluded directories
                        if item.is_dir():
                            if item.name.startswith('.') or item.name in self.EXCLUDE_DIRS:
                                continue
                            # Create directory entry (simple)
                            try:
                                relative_path = str(item.relative_to(original_root_path))
                            except ValueError:
                                relative_path = str(item.name)
                            try:
                                normalized_path = str(item.resolve().relative_to(original_root_path)) if item.resolve().is_relative_to(original_root_path) else str(item.name)
                            except Exception:
                                normalized_path = str(item)
                            dir_entry = {
                                'id': str(item),
                                'name': item.name,
                                'absolute_path': str(item.absolute()),
                                'path': str(item),
                                'normalized_path': normalized_path,
                                'relative_path': relative_path,
                                'is_directory': True,
                                'file_info': {'is_directory': True}
                            }
                            files_for_dir.append(dir_entry)
                            # Recurse
                            stack.append((item, depth + 1))
                            continue

                        # File handling
                        if item.name.startswith('.'):
                            continue

                        abs_file_path = str(item.absolute())
                        cached_entry = self.file_scan_cache.get_file_entry(abs_file_path)

                        current_mtime = None
                        try:
                            current_mtime = os.path.getmtime(abs_file_path)
                        except Exception:
                            pass

                        # Check if schema extraction has timed out (older than 60 seconds)
                        if (cached_entry and isinstance(cached_entry, dict) and 
                            cached_entry.get('schema') and cached_entry.get('schema', {}).get('started_at')):
                            schema_started_at = cached_entry.get('schema', {}).get('started_at')
                            try:
                                started_time = datetime.fromisoformat(schema_started_at)
                                time_diff = datetime.now() - started_time
                                if time_diff > timedelta(seconds=60):
                                    cached_entry['schema'] = {
                                        'loading': False,
                                        'error': 'Schema extraction timed out',
                                    }
                                    self.file_scan_cache.set_file_entry(abs_file_path, cached_entry)
                                    continue
                            except (ValueError, TypeError):
                                # If we can't parse the datetime, continue with normal processing
                                pass

                        use_cached = False
                        if cached_entry and isinstance(cached_entry, dict) and not force_refresh:
                            cached_mtime = cached_entry.get('file_mtime') or cached_entry.get('fileMtime')
                            schema_info = cached_entry.get('schema')
                            if schema_info and isinstance(schema_info, dict) and schema_info.get('loading') is True:
                                use_cached = True
                            if current_mtime is not None and cached_mtime is not None and abs(float(cached_mtime)) == abs(float(current_mtime)):
                                use_cached = True

                        if use_cached:
                            entry = dict(cached_entry)
                            cached_count += 1
                        else:
                            dir_cached = False
                            # Lightweight placeholder while we process in background
                            try:
                                relative_path = str(item.relative_to(original_root_path))
                            except ValueError:
                                relative_path = str(item.name)
                            try:
                                normalized_path = str(item.resolve().relative_to(original_root_path)) if item.resolve().is_relative_to(original_root_path) else str(item.name)
                            except Exception:
                                normalized_path = str(item)
                            entry = {
                                'id': str(item),
                                'name': item.stem,
                                'absolute_path': abs_file_path,
                                'path': str(item),
                                'normalized_path': normalized_path,
                                'relative_path': relative_path,
                                'is_directory': False,
                                'file_mtime': current_mtime,
                                'schema': {
                                    'loading': True,
                                    'started_at': datetime.now().isoformat(),
                                }
                            }

                            self.file_scan_cache.set_file_entry(abs_file_path, entry)
                            # Schedule background processing to populate cache
                            print(f"[FileScannerService] New file detected, scheduling schema extraction: {abs_file_path}")
                            self._executor.submit(self.extract_schema, abs_file_path)

                        files_for_dir.append(entry)
                except (IOError, OSError, PermissionError):
                    continue

            all_files.extend(files_for_dir)
            if dir_cached and scanned_directory_cache:
                scanned_directories.append(scanned_directory_cache)
            else:
                scanned_directories.append({
                    'path': path,
                    'file_count': len(files_for_dir),
                    'scanned_at': datetime.now().isoformat(),
                })

        # De-duplicate by absolute path and directory flag
        unique_seen = set()
        deduped_files: List[Dict[str, Any]] = []
        for entry in all_files:
            abs_path_val = entry.get('absolute_path')
            is_dir = bool(entry.get('is_directory'))
            key = (abs_path_val, is_dir)
            if not abs_path_val:
                deduped_files.append(entry)
                continue
            if key in unique_seen:
                continue
            unique_seen.add(key)
            deduped_files.append(entry)

        # Log summary of schema states
        loading_count = sum(1 for f in deduped_files if isinstance(f.get('schema'), dict) and f['schema'].get('loading') is True)
        completed_count = sum(1 for f in deduped_files if isinstance(f.get('schema'), dict) and 'success' in f['schema'])
        no_schema_count = sum(1 for f in deduped_files if not f.get('schema') and not f.get('is_directory'))
        print(f"[FileScannerService] scan_directories response: {len(deduped_files)} files, {loading_count} loading, {completed_count} completed, {no_schema_count} no-schema")

        return {
            'files': deduped_files,
            'scanned_directories': scanned_directories,
            'cached': cached_count > 0,
            'total_files': len(deduped_files)
        }
    
    def _analyze_notebook_structure(self, cells: List[Dict[str, Any]], metadata: Dict[str, Any],
                                    filepath: str, filename: str, current_mtime: float,
                                    total_cells: int) -> Dict[str, Any]:
        """
        Analyze notebook structure and extract schema information.
        """
        code_cells = []
        markdown_cells = []

        for idx, cell in enumerate(cells):
            cell_type = cell.get('cell_type', 'unknown')
            source = cell.get('source', [])

            # Join source lines if it's a list
            if isinstance(source, list):
                source_text = ''.join(source)
            else:
                source_text = source

            cell_info = {
                'index': idx,
                'cell_type': cell_type,
                'source': source_text[:500],  # Truncate to 500 chars
                'execution_count': cell.get('execution_count'),
            }

            # Add outputs for code cells (truncated)
            if cell_type == 'code':
                outputs = cell.get('outputs', [])
                truncated_outputs = []
                for output in outputs[:3]:  # Max 3 outputs
                    output_type = output.get('output_type', 'unknown')
                    output_info = {'type': output_type}

                    if output_type == 'stream':
                        text = output.get('text', [])
                        if isinstance(text, list):
                            text = ''.join(text)
                        output_info['text'] = text[:200]  # Truncate
                    elif output_type in ('execute_result', 'display_data'):
                        data = output.get('data', {})
                        if 'text/plain' in data:
                            plain_text = data['text/plain']
                            if isinstance(plain_text, list):
                                plain_text = ''.join(plain_text)
                            output_info['text'] = plain_text[:200]
                    elif output_type == 'error':
                        output_info['ename'] = output.get('ename')
                        output_info['evalue'] = output.get('evalue', '')[:200]

                    truncated_outputs.append(output_info)

                cell_info['outputs'] = truncated_outputs
                code_cells.append(cell_info)
            elif cell_type == 'markdown':
                markdown_cells.append(cell_info)

        # Extract kernel info
        kernel_info = metadata.get('kernelspec', {})
        kernel_name = kernel_info.get('name', 'unknown')
        kernel_language = kernel_info.get('language', 'unknown')

        # Create columns format for consistency with other file types
        columns = [
            {
                'name': 'cell_index',
                'dataType': 'integer',
                'description': 'Cell index in notebook'
            },
            {
                'name': 'cell_type',
                'dataType': 'string',
                'description': 'Type of cell (code, markdown, raw)'
            },
            {
                'name': 'source',
                'dataType': 'string',
                'description': 'Cell source code or markdown content'
            },
            {
                'name': 'execution_count',
                'dataType': 'integer',
                'description': 'Execution count for code cells'
            },
            {
                'name': 'outputs',
                'dataType': 'array',
                'description': 'Cell outputs (for code cells)'
            }
        ]

        summary = f'Jupyter Notebook with {total_cells} total cells ({len(code_cells)} code, {len(markdown_cells)} markdown), kernel: {kernel_name}'

        return {
            'success': True,
            'fileId': filepath,
            'fileName': filename,
            'filePath': filepath,
            'fileType': 'ipynb',
            'extractedAt': datetime.now().isoformat(),
            'summary': summary,
            'totalRows': total_cells,
            'totalColumns': len(columns),
            'columns': columns,
            'sampleData': cells,
            'fileMtime': current_mtime,
            'notebook_info': {
                'total_cells': total_cells,
                'code_cells': len(code_cells),
                'markdown_cells': len(markdown_cells),
                'kernel_name': kernel_name,
                'kernel_language': kernel_language
            }
        }

    def extract_schema(self, file_path: str, force_refresh: bool = False, start_cell: int = 0, end_cell: int = 5) -> Dict[str, Any]:
        """Extract schema from a file using pandas"""
        print(f"[FileScannerService] extract_schema START: {file_path}")
        try:
            # Convert to absolute path
            abs_path = file_path

            # Get current mtime for caching
            current_mtime = None
            try:
                current_mtime = os.path.getmtime(abs_path)
            except Exception:
                pass

            item = Path(abs_path)
            
            # Determine file type first
            extension = Path(abs_path).suffix.lower()
            file_type = None
            if extension == '.csv':
                file_type = 'csv'
            elif extension == '.tsv':
                file_type = 'tsv'
            elif extension == '.parquet':
                file_type = 'parquet'
            elif extension == '.xlsx':
                file_type = 'xlsx'
            elif extension == '.json' or extension == '.jsonl':
                file_type = 'json'
            elif extension == '.ipynb':
                file_type = 'ipynb'
            
            entry = self.file_scan_cache.get_file_entry(abs_path) or {
                'fileId': abs_path,
                'fileName': item.name,
                'filePath': abs_path,
                'fileType': file_type,
                'fileMtime': current_mtime
            }
            
            # Check if file type is supported
            if file_type is None:
                if entry:
                    entry['schema'] = {
                        'success': False,
                        'error': f'Unsupported file type: {extension}'
                    }
                    self.file_scan_cache.set_file_entry(abs_path, entry)
                    return entry
                else:
                    return {
                        'success': False,
                        'error': f'Unsupported file type: {extension}'
                    }

            # Extract schema
            try:
                if file_type in ['csv', 'tsv']:
                    separator = '\t' if file_type == 'tsv' else ','
                    df = pd.read_csv(abs_path, sep=separator, nrows=5)
                elif file_type == 'parquet':
                    df = ds.dataset(abs_path).scanner().head(5).to_pandas()
                    df = df.head(5)  # Limit to first 5 rows
                elif file_type == 'xlsx':
                    # Read .xlsx files using openpyxl engine
                    df = pd.read_excel(abs_path, engine='openpyxl', nrows=5)
                    
                    # Get sheet count and names using openpyxl
                    try:
                        workbook = load_workbook(abs_path, read_only=True)
                        sheet_names = workbook.sheetnames
                        total_sheets = len(sheet_names)
                        workbook.close()
                    except Exception:
                        sheet_names = ['Sheet1']  # Default sheet name
                        total_sheets = 1  # Default to 1 if we can't determine
                elif file_type == 'json':
                    # Read and analyze JSON file
                    json_data, file_format, is_truncated = self._read_json_file(abs_path)
                    schema = self._analyze_json_structure(json_data, abs_path, item.name, current_mtime)

                    # Get file info
                    file_info = self._get_file_type_info(str(item), extension)
                    entry['file_info'] = file_info

                    # JSON files are text files
                    file_info['is_binary'] = False

                    # Read file preview for JSON files
                    preview = self._read_file_preview_optimized(str(item))
                    entry['content_preview'] = preview[0]
                    entry['is_truncated'] = preview[1]

                    # Process JSON preview
                    if preview[0]:
                        entry['json_info'] = self._process_json_preview(preview[0], str(item))

                    # Cache the entry
                    if entry:
                        entry['schema'] = schema
                        self.file_scan_cache.set_file_entry(abs_path, entry)
                        print(f"[FileScannerService] extract_schema SUCCESS (json): {abs_path}")

                    return schema
                elif file_type == 'ipynb':
                    # Read notebook file once
                    with open(abs_path, 'r', encoding='utf-8') as f:
                        notebook = json.load(f)

                    cells_all = notebook.get('cells', [])
                    metadata = notebook.get('metadata', {})
                    total_cells = len(cells_all)

                    # Clamp indices and extract range
                    start_idx = max(0, start_cell)
                    end_idx = min(total_cells, end_cell)
                    cells = cells_all[start_idx:end_idx]

                    schema = self._analyze_notebook_structure(cells, metadata, abs_path, item.name, current_mtime, total_cells)

                    # Get file info
                    file_info = self._get_file_type_info(str(item), extension)
                    entry['file_info'] = file_info
                    file_info['is_binary'] = False

                    # Cache the entry
                    entry['schema'] = schema
                    self.file_scan_cache.set_file_entry(abs_path, entry)

                    return schema

                # Get file info for other file types
                file_info = self._get_file_type_info(str(item), extension)
                entry['file_info'] = file_info

                # Check if file is binary
                is_binary = self._is_binary_file(str(item))
                
                if is_binary:
                    # For binary files, just provide basic info and file path
                    entry['content_preview'] = f"Binary file: {str(item)}"
                    entry['is_truncated'] = False
                    # Mark as binary in file_info
                    file_info['is_binary'] = True
                else:
                    # Read file preview with limits for text files
                    preview = self._read_file_preview_optimized(str(item))
                    entry['content_preview'] = preview[0]
                    entry['is_truncated'] = preview[1]
                    file_info['is_binary'] = False
                
                content = entry['content_preview']

                if (file_type == 'csv' or file_type == 'tsv') and content:
                    entry['csv_info'] = self._process_csv_preview(content, str(item))
                elif file_type == 'json' and content:
                    entry['json_info'] = self._process_json_preview(content, str(item))

                if df is not None:
                    result = self._analyze_dataframe(df, file_type)
                    
                    if result['success']:
                        schema = {
                            'success': True,
                            'fileId': abs_path,
                            'fileName': Path(abs_path).name,
                            'filePath': abs_path,
                            'fileType': file_type,
                            'extractedAt': datetime.now().isoformat(),
                            'summary': result['summary'],
                            'totalRows': result['totalRows'],
                            'totalColumns': result['totalColumns'],
                            'columns': result['columns'],
                            'sampleData': result['sampleData'],
                            'fileMtime': current_mtime
                        }
                        
                        # Add sheet count and names for Excel files
                        if file_type == 'xlsx' and 'total_sheets' in locals():
                            schema['totalSheets'] = total_sheets
                            schema['sheetNames'] = sheet_names
                            schema['summary'] = f'Excel file with {total_sheets} sheet{"s" if total_sheets > 1 else ""} ({", ".join(sheet_names)}), {result["totalColumns"]} columns'

                        if entry:
                            entry['schema'] = schema
                            self.file_scan_cache.set_file_entry(abs_path, entry)
                            print(f"[FileScannerService] extract_schema SUCCESS: {abs_path} - {schema.get('totalColumns', '?')} columns, {schema.get('totalRows', '?')} rows")

                        return schema
                    else:
                        print(f"[FileScannerService] extract_schema FAILED (analyze): {abs_path} - {result.get('error', 'unknown')}")
                        if entry:
                            entry['schema'] = {
                                'success': False,
                                'error': f'Failed to extract schema: {result["error"]}'
                            }
                            self.file_scan_cache.set_file_entry(abs_path, entry)
                        return entry

            except Exception as e:
                print(f"[FileScannerService] extract_schema EXCEPTION (inner): {abs_path} - {str(e)}")
                if entry:
                    entry['schema'] = {
                        'success': False,
                        'error': f'Failed to extract schema: {str(e)}'
                    }
                    self.file_scan_cache.set_file_entry(abs_path, entry)
                    return entry
                return {
                    'success': False,
                    'error': f'Failed to extract schema: {str(e)}'
                }

        except Exception as e:
            print(f"[FileScannerService] extract_schema EXCEPTION (outer): {file_path} - {str(e)}")
            entry = self.file_scan_cache.get_file_entry(abs_path)
            if entry:
                entry['schema'] = {
                    'success': False,
                    'error': f'Error extracting schema: {str(e)}'
                }
                self.file_scan_cache.set_file_entry(abs_path, entry)
                return entry
            return {
                'success': False,
                'error': f'Error extracting schema: {str(e)}'
            }

    def get_scanned_directories(self) -> Dict[str, Any]:
        """Get list of currently scanned directories"""
        directories = self.file_scan_cache.get_scanned_directories()
        
        return {
            'directories': directories
        }
    
    def update_scanned_directories(self, directories: List[Dict[str, Any]]) -> bool:
        """Update the list of scanned directories"""
        return self.file_scan_cache.set_scanned_directories(directories)
    
    def shutdown(self):
        """Shutdown the service and cleanup resources"""
        if hasattr(self, '_executor'):
            self._executor.shutdown(wait=True)


# Global instance
_file_scanner_service = None


def get_file_scanner_service() -> FileScannerService:
    """Get the global file scanner service instance"""
    global _file_scanner_service
    if _file_scanner_service is None:
        _file_scanner_service = FileScannerService()
    return _file_scanner_service
