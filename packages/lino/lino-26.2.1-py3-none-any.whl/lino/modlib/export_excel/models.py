# -*- coding: UTF-8 -*-
# Copyright 2014-2025 Rumma & Ko Ltd
# License: GNU Affero General Public License v3 (see file COPYING for details)

from django.conf import settings
from django.db.models import Model
from django.utils.functional import Promise
from django.utils.html import SafeString
from django.utils.timezone import make_naive, is_aware
from lino.utils.html import iselement, to_rst

from lino.core import actions
from lino.core.choicelists import Choice
from lino.core.tables import AbstractTable
from lino.core.roles import DataExporter
from lino.utils import IncompleteDate
from lino.utils.media import TmpMediaFile
from lino.utils.quantities import Duration
from lino.api import dd, _

import datetime


def sheet_name(s):
    s = s[:31]
    for c in "[]:\\?/*\x00":
        s = s.replace(c, "_")
    return s


def ar2workbook(ar, column_names=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.styles import NamedStyle
    # local import to avoid the following traceback:
    # Error in sys.exitfunc:
    # Traceback (most recent call last):
    #   File "/usr/lib/python2.7/atexit.py", line 24, in _run_exitfuncs
    #     func(*targs, **kargs)
    #   File "/openpyxl/writer/write_only.py", line 38, in _openpyxl_shutdown
    #     for path in ALL_TEMP_FILES:
    # TypeError: 'NoneType' object is not iterable

    # workbook = Workbook(guess_types=True)

    # removed `guess_types=True` because it caused trouble in openpyxl
    # 3.4.0 and because I don't know whether it is needed.

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name(ar.get_title())

    duration_style = NamedStyle(name="duration", number_format="[hh]:mm;@")

    bold_font = Font(
        name="Calibri",
        size=11,
        bold=True,
    )

    fields, headers, widths = ar.get_field_info(column_names)

    for c, column in enumerate(fields):
        sheet.cell(row=1, column=c + 1).value = str(headers[c])
        sheet.cell(row=1, column=c + 1).font = bold_font
        # sheet.col(c).width = min(256 * widths[c] / 7, 65535)
        # 256 == 1 character width, max width=65535

    for c, column in enumerate(fields):
        for r, row in enumerate(ar.data_iterator, start=1):
            sf = column.field._lino_atomizer
            value = sf.full_value_from_object(row, ar)
            style = None
            if type(value) == bool:
                value = value and 1 or 0
            elif isinstance(value, Choice):
                value = str(value)
            elif isinstance(value, Duration):
                style = duration_style
                negative = False
                time = str(value)
                if time.startswith("-"):
                    time = time.strip("-")
                    negative = True
                time = time.split(":")
                value = datetime.timedelta(
                    hours=int(time[0]), minutes=int(time[1]))
                if negative:  # Make negative.
                    value = value - value - value
            elif iselement(value) or isinstance(value, SafeString):
                value = to_rst(value)
                # dd.logger.info("20160716 %s", value)
            elif isinstance(value, Promise):
                value = str(value)
            elif isinstance(value, IncompleteDate):
                if value.is_complete():
                    value = value.as_date()
                else:
                    value = str(value)
            elif isinstance(value, Model):
                value = str(value)
            elif isinstance(value, datetime.datetime):
                # Excel does not support timezones in datetimes. The tzinfo in
                # the datetime/time object must be set to None.
                if is_aware(value):
                    value = make_naive()
            elif isinstance(value, str):
                # if it is a future.newstr, change it to a real string to avoid
                # ValueError: Cannot convert 'Hans Altenberg' to Excel
                value = str(value)
            try:
                cell = sheet.cell(row=r + 1, column=c + 1)
                if style is not None:
                    cell.style = style
                cell.value = value
            except ValueError:
                raise Exception("20190222 {} {}".format(
                    value.__class__, value))

    return workbook


class ExportExcelAction(actions.Action):
    label = _("Export to .xls")
    help_text = _("Export this table as an .xls document")
    icon_name = "page_excel"
    ui5_icon_name = "sap-icon://excel-attachment"
    sort_index = -5
    select_rows = False
    default_format = "ajax"
    preprocessor = "Lino.get_current_grid_config"
    callable_from = "t"
    required_roles = dd.login_required(DataExporter)

    def run_from_ui(self, ar, **kw):
        # Prepare tmp file
        mf = TmpMediaFile(ar, "xlsx")
        settings.SITE.makedirs_if_missing(mf.path.parent)

        # Render
        self.render(ar, mf.path)

        # Tell client that the action was successful and that it
        # should open a new browser window on the generated file.
        ar.success(open_url=mf.url)

    def render(self, ar, file):
        workbook = ar2workbook(ar)
        workbook.save(file)


AbstractTable.export_excel = ExportExcelAction()
