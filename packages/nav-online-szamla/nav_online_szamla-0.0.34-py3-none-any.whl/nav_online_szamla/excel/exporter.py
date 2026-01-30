"""
Excel exporter for NAV invoice data.

This module provides functionality to export invoice data to Excel files
with proper formatting and structure.
"""

import logging
from pathlib import Path
from typing import List, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet

from ..models import InvoiceData, ManageInvoiceOperationType
from .exceptions import ExcelProcessingException
from .mapper import ExcelFieldMapper
from .models import InvoiceHeaderRow, InvoiceLineRow

logger = logging.getLogger(__name__)


class InvoiceExcelExporter:
    """Excel exporter for invoice data."""

    def __init__(self):
        """Initialize the Excel exporter."""
        self.mapper = ExcelFieldMapper()

    def export_to_excel(
        self, 
        invoice_data_list: List[Tuple[InvoiceData, ManageInvoiceOperationType]], 
        file_path: str,
        include_operation_type: bool = False
    ) -> None:
        """
        Export invoice data to Excel file.
        
        Args:
            invoice_data_list: List of tuples containing InvoiceData and operation type
            file_path: Path where the Excel file should be saved
            include_operation_type: Whether to include operation type as a column
            
        Raises:
            ExcelProcessingException: If export fails
        """
        try:
            # Validate input
            if not invoice_data_list:
                raise ExcelProcessingException("No invoice data provided")
                
            logger.info(f"Starting Excel export for {len(invoice_data_list)} invoices to {file_path}")
            
            # Convert data to row structures
            header_rows = []
            line_rows = []
            
            for invoice_data, operation_type in invoice_data_list:
                # Convert to header row
                try:
                    header_row = self.mapper.invoice_data_to_header_row(invoice_data, operation_type)
                    header_rows.append(header_row)
                except Exception as e:
                    # Get invoice identifier safely
                    invoice_id = getattr(invoice_data, 'invoice_number', 'unknown')
                    logger.warning(f"Failed to convert header data for {invoice_id}: {e}")
                    continue

                # Convert to line rows
                try:
                    line_row_list = self.mapper.invoice_data_to_line_rows(invoice_data, operation_type)
                    line_rows.extend(line_row_list)
                except Exception as e:
                    # Get invoice identifier safely
                    invoice_id = getattr(invoice_data, 'invoice_number', 'unknown')
                    logger.exception(f"Failed to convert line data for {invoice_id}: {e}")
                    # Continue with header even if lines fail            # Create Excel file
            # Only raise exception if ALL data is invalid and are not proper objects
            # Allow mapping failures on valid objects to create empty files
            if not header_rows and not line_rows and invoice_data_list:
                # Check if we have any basic invoice-like objects (not just strings)
                valid_types = any(
                    hasattr(invoice_data, 'invoice_number') or 
                    hasattr(invoice_data, '__dict__') or
                    not isinstance(invoice_data, str)
                    for invoice_data, _ in invoice_data_list
                )
                if not valid_types:
                    raise ExcelProcessingException("No valid invoice data could be processed")
                
            self._create_excel_file(header_rows, line_rows, file_path, include_operation_type)
            
            logger.info(f"Successfully exported {len(header_rows)} invoices with {len(line_rows)} line items to {file_path}")
            
        except Exception as e:
            logger.error(f"Excel export failed: {e}")
            raise ExcelProcessingException(f"Failed to export to Excel: {e}")

    def create_template_excel(self, file_path: str) -> None:
        """
        Create an Excel template file with proper structure and headers but no data.
        
        Args:
            file_path: Path where the template Excel file should be created
            
        Raises:
            ExcelProcessingException: If template creation fails
        """
        try:
            logger.info(f"Creating Excel template: {file_path}")
            
            # Create empty data structures to generate headers
            empty_header_rows = []
            empty_line_rows = []
            
            # Create the Excel file structure with headers but no data
            self._create_excel_file(empty_header_rows, empty_line_rows, file_path, include_operation_type=False)
            
            logger.info(f"Excel template created successfully: {file_path}")
            
        except Exception as e:
            logger.error(f"Template creation failed: {e}")
            raise ExcelProcessingException(f"Failed to create Excel template: {e}")

    def _create_excel_file(
        self,
        header_rows: List[InvoiceHeaderRow],
        line_rows: List[InvoiceLineRow],
        file_path: str,
        include_operation_type: bool = False
    ) -> None:
        """
        Create the Excel file with properly formatted sheets.
        
        Args:
            header_rows: List of header row data
            line_rows: List of line row data
            file_path: Output file path
            include_operation_type: Whether to include operation type column
        """
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        # Create header sheet
        header_sheet = wb.create_sheet("Fejléc adatok")
        self._populate_header_sheet(header_sheet, header_rows, include_operation_type)

        # Create lines sheet
        lines_sheet = wb.create_sheet("Tétel adatok") 
        self._populate_lines_sheet(lines_sheet, line_rows, include_operation_type)

        # Save workbook
        Path(file_path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(file_path)

    def _populate_header_sheet(
        self, 
        sheet: Worksheet, 
        header_rows: List[InvoiceHeaderRow],
        include_operation_type: bool = False
    ) -> None:
        """Populate the header sheet with invoice header data."""
        # Create header columns mapping
        header_columns = {
            'invoice_number': 'Számla sorszáma',
            'invoice_issue_date': 'Számla kelte', 
            'fulfillment_date': 'Teljesítés dátuma',
            'invoice_currency': 'Számla pénzneme',
            'exchange_rate': 'Alkalmazott árfolyam',
            'seller_tax_number_main': 'Eladó adószáma (törzsszám)',
            'seller_tax_number_vat': 'Eladó adószáma (ÁFA-kód)',
            'seller_tax_number_county': 'Eladó adószáma (megyekód)',
            'seller_name': 'Eladó neve',
            'seller_country_code': 'Eladó országkódja',
            'seller_postal_code': 'Eladó irányítószáma',
            'seller_city': 'Eladó települése',
            'seller_address_detail': 'Eladó többi címadata',
            'buyer_tax_number_main': 'Vevő adószáma (törzsszám)',
            'buyer_tax_number_vat': 'Vevő adószáma (ÁFA-kód)',
            'buyer_tax_number_county': 'Vevő adószáma (megyekód)',
            'buyer_name': 'Vevő neve',
            'buyer_vat_status': 'Vevő státusza',
            'buyer_community_vat_number': 'Vevő közösségi adószáma',
            'buyer_third_country_tax_number': 'Vevő harmadik országbeli adószáma',
            'buyer_country_code': 'Vevő országkódja',
            'buyer_postal_code': 'Vevő irányítószáma',
            'buyer_city': 'Vevő települése',
            'buyer_address_detail': 'Vevő többi címadata',
            'original_invoice_number': 'Eredeti számla száma',
            'modification_date': 'Módosító okirat kelte',
            'modification_index': 'Módosítás sorszáma',
            'net_amount_original': 'Számla nettó összege (a számla pénznemében)',
            'net_amount_huf': 'Számla nettó összege (forintban)',
            'vat_amount_original': 'Számla ÁFA összege (a számla pénznemében)',
            'vat_amount_huf': 'Számla ÁFA összege (forintban)',
            'gross_amount_original': 'Számla bruttó összege (a számla pénznemében)',
            'gross_amount_huf': 'Számla bruttó összege (forintban)',
            'payment_due_date': 'Fizetési határidő',
            'payment_method': 'Fizetési mód',
            'small_business_indicator': 'Kisadózó jelölése',
            'cash_accounting_indicator': 'Pénzforgalmi elszámolás jelölése',
            'invoice_category': 'Számla típusa',
            'completeness_indicator': 'Az adatszolgáltatás maga a számla',
            'advance_payment_indicator': 'Előleg jelleg jelölése',
            'no_vat_charge_indicator': 'Nincs felszámított áfa az áfa törvény 17. § alapján',
        }
        
        if include_operation_type:
            header_columns['operation_type'] = 'Művelet típusa'

        # Convert to DataFrame and write
        if header_rows:
            data = []
            for row in header_rows:
                data.append(self._header_row_to_dict(row))
            
            df = pd.DataFrame(data)
            # Replace all None/NaN values with 'n/a'
            df = df.fillna('n/a')
            # Rename columns to Hungarian
            df = df.rename(columns=header_columns)
            
        else:
            # Create empty DataFrame with proper columns for template
            df = pd.DataFrame(columns=list(header_columns.values()))

        # Write DataFrame to sheet
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                sheet.cell(row=r_idx, column=c_idx, value=value)

    def _populate_lines_sheet(
        self, 
        sheet: Worksheet, 
        line_rows: List[InvoiceLineRow],
        include_operation_type: bool = False
    ) -> None:
        """Populate the lines sheet with invoice line data."""
        # Create comprehensive lines columns mapping to match expected structure
        lines_columns = {
            'invoice_number': 'Számla sorszáma',
            'buyer_tax_number_main': 'Vevő adószáma (törzsszám)',
            'buyer_name': 'Vevő neve',
            'seller_tax_number_main': 'Eladó adószáma (törzsszám)',
            'seller_name': 'Eladó neve',
            'line_number': 'Tétel sorszáma',
            'modified_line_number': 'Módosítással érintett tétel sorszáma',
            'line_modification_type': 'Módosítás jellege',
            'description': 'Megnevezés',
            'quantity': 'Mennyiség',
            'unit_of_measure': 'Mennyiségi egység',
            'unit_price': 'Egységár',
            'net_amount_original': 'Nettó összeg (a számla pénznemében)',
            'net_amount_huf': 'Nettó összeg (forintban)',
            'vat_rate': 'Adó mértéke',
            'vat_exemption_indicator': 'Áfamentesség jelölés',
            'vat_exemption_case': 'Áfamentesség esete',
            'vat_exemption_reason': 'Áfamentesség leírása',
            'out_of_scope_indicator': 'ÁFA törvény hatályán kívüli jelölés',
            'out_of_scope_case': 'ÁFA törvény hatályon kívüliségének esete',
            'out_of_scope_reason': 'ÁFA törvény hatályon kívüliségének leírása',
            'tax_base_deviation_case': 'Adóalap és felszámított adó eltérésének esete',
            'different_tax_rate_content': 'Eltérő adóalap és felszámított adó adómérték, adótartalom',
            'domestic_reverse_charge_indicator': 'Belföldi fordított adózás jelölés',
            'margin_scheme_with_vat': 'Áthárított adót tartalmazó különbözet szerinti adózás',
            'margin_scheme_without_vat': 'Áthárított adót nem tartalmazó különbözet szerinti adózás',
            'margin_scheme_indicator': 'Különbözet szerinti adózás',
            'vat_amount_original': 'ÁFA összeg (a számla pénznemében)',
            'vat_amount_huf': 'ÁFA összeg (forintban)',
            'gross_amount_original': 'Bruttó összeg (a számla pénznemében)',
            'gross_amount_huf': 'Bruttó összeg (forintban)',
            'vat_content': 'ÁFA tartalom',
            'advance_payment_indicator': 'Előleg jelleg jelölése',
            'line_exchange_rate': 'Tétel árfolyam',
            'line_fulfillment_date': 'Tétel teljesítés dátuma',
            'no_vat_charge_indicator': 'Nincs felszámított áfa az áfa törvény 17. § alapján',
        }
        
        if include_operation_type:
            lines_columns['operation_type'] = 'Művelet típusa'

        # Convert to DataFrame and write
        if line_rows:
            data = []
            for row in line_rows:
                data.append(self._line_row_to_dict(row))
            
            df = pd.DataFrame(data)
            # Replace all None/NaN values with 'n/a'
            df = df.fillna('n/a')
            # Rename columns to Hungarian  
            df = df.rename(columns=lines_columns)
            
        else:
            # Create empty DataFrame with proper columns for template
            df = pd.DataFrame(columns=list(lines_columns.values()))

        # Write DataFrame to sheet
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                sheet.cell(row=r_idx, column=c_idx, value=value)

    def _header_row_to_dict(self, row: InvoiceHeaderRow) -> dict:
        """Convert InvoiceHeaderRow to dictionary."""
        return {
            'invoice_number': row.invoice_number,
            'invoice_issue_date': row.invoice_issue_date,
            'fulfillment_date': row.fulfillment_date,
            'invoice_currency': row.invoice_currency,
            'exchange_rate': row.exchange_rate,
            'seller_tax_number_main': row.seller_tax_number_main,
            'seller_tax_number_vat': row.seller_tax_number_vat,
            'seller_tax_number_county': row.seller_tax_number_county,
            'seller_name': row.seller_name,
            'seller_country_code': row.seller_country_code,
            'seller_postal_code': row.seller_postal_code,
            'seller_city': row.seller_city,
            'seller_address_detail': row.seller_address_detail,
            'buyer_tax_number_main': row.buyer_tax_number_main,
            'buyer_tax_number_vat': row.buyer_tax_number_vat,
            'buyer_tax_number_county': row.buyer_tax_number_county,
            'buyer_name': row.buyer_name,
            'buyer_vat_status': row.buyer_vat_status,
            'buyer_community_vat_number': row.buyer_community_vat_number,
            'buyer_third_country_tax_number': row.buyer_third_country_tax_number,
            'buyer_country_code': row.buyer_country_code,
            'buyer_postal_code': row.buyer_postal_code,
            'buyer_city': row.buyer_city,
            'buyer_address_detail': row.buyer_address_detail,
            'original_invoice_number': row.original_invoice_number,
            'modification_date': row.modification_date,
            'modification_index': row.modification_index,
            'net_amount_original': row.net_amount_original,
            'net_amount_huf': row.net_amount_huf,
            'vat_amount_original': row.vat_amount_original,
            'vat_amount_huf': row.vat_amount_huf,
            'gross_amount_original': row.gross_amount_original,
            'gross_amount_huf': row.gross_amount_huf,
            'payment_due_date': row.payment_due_date,
            'payment_method': row.payment_method,
            'small_business_indicator': row.small_business_indicator,
            'cash_accounting_indicator': row.cash_accounting_indicator,
            'invoice_category': row.invoice_category,
            'completeness_indicator': row.completeness_indicator,
        }

    def _line_row_to_dict(self, row: InvoiceLineRow) -> dict:
        """Convert InvoiceLineRow to dictionary with all comprehensive fields."""
        return {
            # Reference information
            'invoice_number': row.invoice_number,
            'buyer_tax_number_main': row.buyer_tax_number_main,
            'buyer_name': row.buyer_name,
            'seller_tax_number_main': row.seller_tax_number_main,
            'seller_name': row.seller_name,
            
            # Line item identification
            'line_number': row.line_number,
            'modified_line_number': row.modified_line_number,
            'line_modification_type': row.line_modification_type,
            
            # Product/service information
            'description': row.description,
            'quantity': row.quantity,
            'unit_of_measure': row.unit_of_measure,
            'unit_price': row.unit_price,
            
            # Financial amounts
            'net_amount_original': row.net_amount_original,
            'net_amount_huf': row.net_amount_huf,
            
            # VAT information
            'vat_rate': row.vat_rate,
            'vat_exemption_indicator': row.vat_exemption_indicator,
            'vat_exemption_case': row.vat_exemption_case,
            'vat_exemption_reason': row.vat_exemption_reason,
            
            # Out of scope VAT
            'out_of_scope_indicator': row.out_of_scope_indicator,
            'out_of_scope_case': row.out_of_scope_case,
            'out_of_scope_reason': row.out_of_scope_reason,
            
            # Tax deviation
            'tax_base_deviation_case': row.tax_base_deviation_case,
            'different_tax_rate_content': row.different_tax_rate_content,
            
            # Reverse charge and margin scheme
            'domestic_reverse_charge_indicator': row.domestic_reverse_charge_indicator,
            'margin_scheme_with_vat': row.margin_scheme_with_vat,
            'margin_scheme_without_vat': row.margin_scheme_without_vat,
            'margin_scheme_indicator': row.margin_scheme_indicator,
            
            # VAT amounts
            'vat_amount_original': row.vat_amount_original,
            'vat_amount_huf': row.vat_amount_huf,
            'gross_amount_original': row.gross_amount_original,
            'gross_amount_huf': row.gross_amount_huf,
            
            # Additional information
            'vat_content': row.vat_content,
            'advance_payment_indicator': row.advance_payment_indicator,
            'line_exchange_rate': row.line_exchange_rate,
            'line_fulfillment_date': row.line_fulfillment_date,
            'no_vat_charge_indicator': row.no_vat_charge_indicator,
        }

    def _validate_file_path(self, file_path: str) -> None:
        """Validate that the file path has the correct extension."""
        if not file_path.endswith('.xlsx'):
            raise ExcelProcessingException("File must have .xlsx extension")

    def _format_datetime_for_excel(self, value) -> pd.Timestamp:
        """Format datetime value for Excel."""
        if value is None:
            return None
        if isinstance(value, str):
            return pd.Timestamp(value)
        return pd.Timestamp(value)

    def _format_decimal_for_excel(self, value):
        """Format decimal value for Excel."""
        if value is None:
            return None
        if isinstance(value, str):
            return float(value)
        if hasattr(value, '__float__'):
            return float(value)
        return value