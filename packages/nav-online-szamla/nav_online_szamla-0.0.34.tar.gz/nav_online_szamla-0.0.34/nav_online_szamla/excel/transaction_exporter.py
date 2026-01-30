"""
Excel exporter for NAV transaction data.

This module provides functionality to export transaction data to Excel files
with proper formatting and structure including transaction status information.
"""

import logging
from pathlib import Path
from typing import List

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet

from ..models import QueryTransactionStatusResponse
from .exceptions import ExcelProcessingException
from .transaction_mapper import TransactionFieldMapper
from .models import InvoiceHeaderRow, InvoiceLineRow, TransactionStatusRow

logger = logging.getLogger(__name__)


class TransactionExcelExporter:
    """Excel exporter for transaction data."""

    def __init__(self):
        """Initialize the transaction Excel exporter."""
        self.mapper = TransactionFieldMapper()

    def export_to_excel(
        self, 
        transaction_responses: List[QueryTransactionStatusResponse], 
        file_path: str,
        transaction_ids: List[str] = None,
        request_statuses: List[str] = None,
        technical_annulments: List[bool] = None,
        ins_dates: List[str] = None
    ) -> None:
        """
        Export transaction data to Excel file with 3 sheets.
        
        Args:
            transaction_responses: List of transaction status responses
            file_path: Path where the Excel file should be saved
            transaction_ids: Optional list of transaction IDs corresponding to responses
            request_statuses: Optional list of request statuses corresponding to responses
            technical_annulments: Optional list of technical annulment flags corresponding to responses
            ins_dates: Optional list of submission timestamps (ins_date) corresponding to responses
            
        Raises:
            ExcelProcessingException: If export fails
        """
        try:
            # Handle empty data gracefully by creating template file (similar to invoice exporter)
            if not transaction_responses:
                logger.info(f"No transaction data provided, creating empty Excel template: {file_path}")
                self._create_excel_file([], [], [], file_path)
                logger.info(f"Empty Excel template created successfully: {file_path}")
                return
                
            logger.info(f"Starting Excel export for {len(transaction_responses)} transactions to {file_path}")
            
            # Convert data to row structures
            header_rows = []
            line_rows = []
            status_rows = []
            
            for i, transaction_response in enumerate(transaction_responses):
                try:
                    # Get the transaction ID for this response
                    transaction_id = None
                    if transaction_ids and i < len(transaction_ids):
                        transaction_id = transaction_ids[i]
                    
                    # Get request status from list if provided
                    request_status = None
                    if request_statuses and i < len(request_statuses):
                        request_status = request_statuses[i]
                    
                    # Get technical annulment from list if provided
                    technical_annulment = None
                    if technical_annulments and i < len(technical_annulments):
                        technical_annulment = technical_annulments[i]
                    
                    # Get ins_date from list if provided
                    ins_date = None
                    if ins_dates and i < len(ins_dates):
                        ins_date = ins_dates[i]
                    
                    # Extract invoice data and status information
                    transaction_header_rows, transaction_line_rows, transaction_status_rows = self.mapper.transaction_response_to_rows(
                        transaction_response, 
                        transaction_id,
                        request_status,
                        technical_annulment,
                        ins_date
                    )
                    
                    header_rows.extend(transaction_header_rows)
                    line_rows.extend(transaction_line_rows)
                    status_rows.extend(transaction_status_rows)
                    
                except Exception as e:
                    display_id = transaction_id if transaction_id else getattr(transaction_response, 'transaction_id', 'unknown')
                    logger.warning(f"Failed to convert transaction data for {display_id}: {e}")
                    continue
            
            # Create Excel file with 3 sheets
            self._create_excel_file(header_rows, line_rows, status_rows, file_path)
            
            logger.info(f"Successfully exported {len(header_rows)} invoice headers, {len(line_rows)} line items, and {len(status_rows)} status records to {file_path}")
            
        except Exception as e:
            logger.error(f"Transaction Excel export failed: {e}")
            raise ExcelProcessingException(f"Failed to export transactions to Excel: {e}")

    def _create_excel_file(
        self,
        header_rows: List[InvoiceHeaderRow],
        line_rows: List[InvoiceLineRow],
        status_rows: List[TransactionStatusRow],
        file_path: str
    ) -> None:
        """
        Create the Excel file with properly formatted sheets.
        
        Args:
            header_rows: List of header row data
            line_rows: List of line row data
            status_rows: List of transaction status row data
            file_path: Output file path
        """
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        # Create header sheet
        header_sheet = wb.create_sheet("Fejléc adatok")
        self._populate_header_sheet(header_sheet, header_rows)

        # Create lines sheet
        lines_sheet = wb.create_sheet("Tétel adatok") 
        self._populate_lines_sheet(lines_sheet, line_rows)

        # Create transaction status sheet
        status_sheet = wb.create_sheet("Tranzakció Státusz")
        self._populate_status_sheet(status_sheet, status_rows)

        # Save workbook
        Path(file_path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(file_path)

    def _populate_header_sheet(
        self, 
        sheet: Worksheet, 
        header_rows: List[InvoiceHeaderRow]
    ) -> None:
        """Populate the header sheet with invoice header data."""
        # Create header columns mapping
        header_columns = {
            'invoice_number': 'Számla sorszáma',
            'invoice_issue_date': 'Számla kelte', 
            'fulfillment_date': 'Teljesítés dátuma',
            'invoice_currency': 'Számla pénzneme',
            'exchange_rate': 'Alkalmazott árfolyam',
            'seller_tax_number_main': 'Eladó adószáma (törzsszám)',
            'seller_tax_number_vat': 'Eladó adószáma (ÁFA-kód)',
            'seller_tax_number_county': 'Eladó adószáma (megyekód)',
            'seller_name': 'Eladó neve',
            'seller_country_code': 'Eladó országkódja',
            'seller_postal_code': 'Eladó irányítószáma',
            'seller_city': 'Eladó települése',
            'seller_address_detail': 'Eladó többi címadata',
            'buyer_tax_number_main': 'Vevő adószáma (törzsszám)',
            'buyer_tax_number_vat': 'Vevő adószáma (ÁFA-kód)',
            'buyer_tax_number_county': 'Vevő adószáma (megyekód)',
            'buyer_name': 'Vevő neve',
            'buyer_vat_status': 'Vevő státusza',
            'buyer_community_vat_number': 'Vevő közösségi adószáma',
            'buyer_third_country_tax_number': 'Vevő harmadik országbeli adószáma',
            'buyer_country_code': 'Vevő országkódja',
            'buyer_postal_code': 'Vevő irányítószáma',
            'buyer_city': 'Vevő települése',
            'buyer_address_detail': 'Vevő többi címadata',
            'original_invoice_number': 'Eredeti számla száma',
            'modification_date': 'Módosító okirat kelte',
            'modification_index': 'Módosítás sorszáma',
            'net_amount_original': 'Számla nettó összege (a számla pénznemében)',
            'net_amount_huf': 'Számla nettó összege (forintban)',
            'vat_amount_original': 'Számla ÁFA összege (a számla pénznemében)',
            'vat_amount_huf': 'Számla ÁFA összege (forintban)',
            'gross_amount_original': 'Számla bruttó összege (a számla pénznemében)',
            'gross_amount_huf': 'Számla bruttó összege (forintban)',
            'payment_due_date': 'Fizetési határidő',
            'payment_method': 'Fizetési mód',
            'small_business_indicator': 'Kisadózó jelölése',
            'cash_accounting_indicator': 'Pénzforgalmi elszámolás jelölése',
            'invoice_category': 'Számla típusa',
            'completeness_indicator': 'Az adatszolgáltatás maga a számla',
        }

        # Convert to DataFrame and write
        if header_rows:
            data = []
            for row in header_rows:
                data.append(self._header_row_to_dict(row))
            
            df = pd.DataFrame(data)
            # Replace all None/NaN values with 'n/a'
            df = df.fillna('n/a')
            # Rename columns to Hungarian
            df = df.rename(columns=header_columns)
            
        else:
            # Create empty DataFrame with proper columns for template
            df = pd.DataFrame(columns=list(header_columns.values()))

        # Write DataFrame to sheet
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                sheet.cell(row=r_idx, column=c_idx, value=value)

    def _populate_lines_sheet(
        self, 
        sheet: Worksheet, 
        line_rows: List[InvoiceLineRow]
    ) -> None:
        """Populate the lines sheet with invoice line data."""
        # Create comprehensive lines columns mapping
        lines_columns = {
            'invoice_number': 'Számla sorszáma',
            'buyer_tax_number_main': 'Vevő adószáma (törzsszám)',
            'buyer_name': 'Vevő neve',
            'seller_tax_number_main': 'Eladó adószáma (törzsszám)',
            'seller_name': 'Eladó neve',
            'line_number': 'Tétel sorszáma',
            'modified_line_number': 'Módosítással érintett tétel sorszáma',
            'line_modification_type': 'Módosítás jellege',
            'description': 'Megnevezés',
            'quantity': 'Mennyiség',
            'unit_of_measure': 'Mennyiségi egység',
            'unit_price': 'Egységár',
            'net_amount_original': 'Nettó összeg (a számla pénznemében)',
            'net_amount_huf': 'Nettó összeg (forintban)',
            'vat_rate': 'Adó mértéke',
            'vat_exemption_indicator': 'Áfamentesség jelölés',
            'vat_exemption_case': 'Áfamentesség esete',
            'vat_exemption_reason': 'Áfamentesség leírása',
            'out_of_scope_indicator': 'ÁFA törvény hatályán kívüli jelölés',
            'out_of_scope_case': 'ÁFA törvény hatályon kívüliségének esete',
            'out_of_scope_reason': 'ÁFA törvény hatályon kívüliségének leírása',
            'tax_base_deviation_case': 'Adóalap és felszámított adó eltérésének esete',
            'different_tax_rate_content': 'Eltérő adóalap és felszámított adó adómérték, adótartalom',
            'domestic_reverse_charge_indicator': 'Belföldi fordított adózás jelölés',
            'margin_scheme_with_vat': 'Áthárított adót tartalmazó különbözet szerinti adózás',
            'margin_scheme_without_vat': 'Áthárított adót nem tartalmazó különbözet szerinti adózás',
            'margin_scheme_indicator': 'Különbözet szerinti adózás',
            'vat_amount_original': 'ÁFA összeg (a számla pénznemében)',
            'vat_amount_huf': 'ÁFA összeg (forintban)',
            'gross_amount_original': 'Bruttó összeg (a számla pénznemében)',
            'gross_amount_huf': 'Bruttó összeg (forintban)',
            'vat_content': 'ÁFA tartalom',
            'advance_payment_indicator': 'Előleg jelleg jelölése',
            'line_exchange_rate': 'Tétel árfolyam',
            'line_fulfillment_date': 'Tétel teljesítés dátuma',
            'no_vat_charge_indicator': 'Nincs felszámított áfa az áfa törvény 17. § alapján',
        }

        # Convert to DataFrame and write
        if line_rows:
            data = []
            for row in line_rows:
                data.append(self._line_row_to_dict(row))
            
            df = pd.DataFrame(data)
            # Replace all None/NaN values with 'n/a'
            df = df.fillna('n/a')
            # Rename columns to Hungarian  
            df = df.rename(columns=lines_columns)
            
        else:
            # Create empty DataFrame with proper columns for template
            df = pd.DataFrame(columns=list(lines_columns.values()))

        # Write DataFrame to sheet
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                sheet.cell(row=r_idx, column=c_idx, value=value)

    def _populate_status_sheet(
        self, 
        sheet: Worksheet, 
        status_rows: List[TransactionStatusRow]
    ) -> None:
        """Populate the status sheet with transaction status data."""
        # Create transaction status columns mapping
        status_columns = {
            'transaction_id': 'Tranzakció azonosító',
            'submission_timestamp': 'Beküldés időpontja',
            'invoice_number': 'Számla sorszáma',
            'invoice_status': 'Számla státusz',
            'operation_type': 'Művelet típusa',
            'request_status': 'Feldolgozási státusza',
            'technical_annulment': 'Technikai érvénytelenítés',
            'business_validation_messages': 'Üzleti validációs üzenetek',
            'technical_validation_messages': 'Technikai validációs üzenetek',
        }

        # Convert to DataFrame and write
        if status_rows:
            data = []
            for row in status_rows:
                data.append(self._status_row_to_dict(row))
            
            df = pd.DataFrame(data)
            # Replace all None/NaN values with 'n/a'
            df = df.fillna('n/a')
            # Rename columns to Hungarian  
            df = df.rename(columns=status_columns)
            
        else:
            # Create empty DataFrame with proper columns for template
            df = pd.DataFrame(columns=list(status_columns.values()))

        # Write DataFrame to sheet
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                sheet.cell(row=r_idx, column=c_idx, value=value)

    def _header_row_to_dict(self, row: InvoiceHeaderRow) -> dict:
        """Convert InvoiceHeaderRow to dictionary."""
        return {
            'invoice_number': row.invoice_number,
            'invoice_issue_date': row.invoice_issue_date,
            'fulfillment_date': row.fulfillment_date,
            'invoice_currency': row.invoice_currency,
            'exchange_rate': row.exchange_rate,
            'seller_tax_number_main': row.seller_tax_number_main,
            'seller_tax_number_vat': row.seller_tax_number_vat,
            'seller_tax_number_county': row.seller_tax_number_county,
            'seller_name': row.seller_name,
            'seller_country_code': row.seller_country_code,
            'seller_postal_code': row.seller_postal_code,
            'seller_city': row.seller_city,
            'seller_address_detail': row.seller_address_detail,
            'buyer_tax_number_main': row.buyer_tax_number_main,
            'buyer_tax_number_vat': row.buyer_tax_number_vat,
            'buyer_tax_number_county': row.buyer_tax_number_county,
            'buyer_name': row.buyer_name,
            'buyer_vat_status': row.buyer_vat_status,
            'buyer_community_vat_number': row.buyer_community_vat_number,
            'buyer_third_country_tax_number': row.buyer_third_country_tax_number,
            'buyer_country_code': row.buyer_country_code,
            'buyer_postal_code': row.buyer_postal_code,
            'buyer_city': row.buyer_city,
            'buyer_address_detail': row.buyer_address_detail,
            'original_invoice_number': row.original_invoice_number,
            'modification_date': row.modification_date,
            'modification_index': row.modification_index,
            'net_amount_original': row.net_amount_original,
            'net_amount_huf': row.net_amount_huf,
            'vat_amount_original': row.vat_amount_original,
            'vat_amount_huf': row.vat_amount_huf,
            'gross_amount_original': row.gross_amount_original,
            'gross_amount_huf': row.gross_amount_huf,
            'payment_due_date': row.payment_due_date,
            'payment_method': row.payment_method,
            'small_business_indicator': row.small_business_indicator,
            'cash_accounting_indicator': row.cash_accounting_indicator,
            'invoice_category': row.invoice_category,
            'completeness_indicator': row.completeness_indicator,
        }

    def _line_row_to_dict(self, row: InvoiceLineRow) -> dict:
        """Convert InvoiceLineRow to dictionary with all comprehensive fields."""
        return {
            # Reference information
            'invoice_number': row.invoice_number,
            'buyer_tax_number_main': row.buyer_tax_number_main,
            'buyer_name': row.buyer_name,
            'seller_tax_number_main': row.seller_tax_number_main,
            'seller_name': row.seller_name,
            
            # Line item identification
            'line_number': row.line_number,
            'modified_line_number': row.modified_line_number,
            'line_modification_type': row.line_modification_type,
            
            # Product/service information
            'description': row.description,
            'quantity': row.quantity,
            'unit_of_measure': row.unit_of_measure,
            'unit_price': row.unit_price,
            
            # Financial amounts
            'net_amount_original': row.net_amount_original,
            'net_amount_huf': row.net_amount_huf,
            
            # VAT information
            'vat_rate': row.vat_rate,
            'vat_exemption_indicator': row.vat_exemption_indicator,
            'vat_exemption_case': row.vat_exemption_case,
            'vat_exemption_reason': row.vat_exemption_reason,
            
            # Out of scope VAT
            'out_of_scope_indicator': row.out_of_scope_indicator,
            'out_of_scope_case': row.out_of_scope_case,
            'out_of_scope_reason': row.out_of_scope_reason,
            
            # Tax deviation
            'tax_base_deviation_case': row.tax_base_deviation_case,
            'different_tax_rate_content': row.different_tax_rate_content,
            
            # Reverse charge and margin scheme
            'domestic_reverse_charge_indicator': row.domestic_reverse_charge_indicator,
            'margin_scheme_with_vat': row.margin_scheme_with_vat,
            'margin_scheme_without_vat': row.margin_scheme_without_vat,
            'margin_scheme_indicator': row.margin_scheme_indicator,
            
            # VAT amounts
            'vat_amount_original': row.vat_amount_original,
            'vat_amount_huf': row.vat_amount_huf,
            'gross_amount_original': row.gross_amount_original,
            'gross_amount_huf': row.gross_amount_huf,
            
            # Additional information
            'vat_content': row.vat_content,
            'advance_payment_indicator': row.advance_payment_indicator,
            'line_exchange_rate': row.line_exchange_rate,
            'line_fulfillment_date': row.line_fulfillment_date,
            'no_vat_charge_indicator': row.no_vat_charge_indicator,
        }

    def _status_row_to_dict(self, row: TransactionStatusRow) -> dict:
        """Convert TransactionStatusRow to dictionary."""
        return {
            'transaction_id': row.transaction_id,
            'submission_timestamp': row.submission_timestamp,
            'invoice_number': row.invoice_number,
            'invoice_status': row.invoice_status,
            'operation_type': row.operation_type,
            'request_status': row.request_status,
            'technical_annulment': row.technical_annulment,
            'business_validation_messages': row.business_validation_messages,
            'technical_validation_messages': row.technical_validation_messages,
        }