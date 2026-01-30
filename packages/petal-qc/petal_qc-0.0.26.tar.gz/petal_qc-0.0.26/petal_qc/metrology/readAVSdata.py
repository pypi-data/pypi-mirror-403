#!/usr/bin/env pythoon3
"""Read AVS dats file."""
import sys
import re
import numpy as np
from argparse import ArgumentParser
from pathlib import Path

import dateutil.parser
import openpyxl as XL
from openpyxl.cell.cell import MergedCell
from openpyxl.utils.exceptions import InvalidFileException

try:
    import itkdb_gtk

except ImportError:
    cwd = Path(__file__).parent.parent
    sys.path.append(cwd.as_posix())

from itkdb_gtk import ITkDBlogin, ITkDButils

sk_defaults = {
    "institution": "AVS",
    "runNumber": "1",
}


class AVSDataException(Exception):
    """AVSData exception class."""

    def __init__(self, message):
        """Call the base class constructor with the parameters it needs."""
        super().__init__(message)


def create_weight(session, SN, the_date=None, manager="", passed=True, problems=False, comments=None):
    """Creates the dictionary for a WEIGHT test.

    Args:
        session: the DB session
        SN: Serial Number
        the_date: the date of the test
        manager: manager name
        passed: if test passed or not
        problems: if problems were found during test
        comments: list of comments to append to the test

    """
    if comments is None:
        comments = []

    the_date = ITkDButils.get_db_date(the_date)
    out = ITkDButils.get_test_skeleton(session, "CORE_PETAL", "WEIGHING", sk_defaults)
    out['component'] = SN
    out['date'] = the_date
    out["institution"] = "AVS"
    out["runNumber"] = "1"
    out['passed'] = passed
    out['problems'] = problems
    out['properties']['PRODUCTION_MANAGER'] = manager
    out['comments'] = comments
    return out


def create_manufacturing(session, SN, the_date=None, manager="", passed=True, problems=False, comments=None):
    """Create the dictionary or the MANUFACTURING test.

    Args:
        session: the DB session
        SN: Serial Number
        the_date: the date of the test
        manager: manager name
        passed: if test passed or not
        problems: if problems were found during test
        comments: list of comments to append to the test

    """
    if comments is None:
        comments = []
    the_date = ITkDButils.get_db_date(the_date)
    out = ITkDButils.get_test_skeleton(session, "CORE_PETAL", "MANUFACTURING", sk_defaults)
    out['component'] = SN
    out['date'] = the_date
    out["institution"] = "AVS"
    out["runNumber"] = "1"
    out['passed'] = passed
    out['problems'] = problems
    out['properties']['PRODUCTION_MANAGER'] = manager
    out['comments'] = comments

    return out


def create_visual_inpection(session, SN, the_date=None, operator="", passed=True, problems=False, comments=None):
    """Create Visual Inspection test skeleton."""
    if comments is None:
        comments = []
    the_date = ITkDButils.get_db_date(the_date)
    out = ITkDButils.get_test_skeleton(session, "CORE_PETAL", "VISUAL_INSPECTION", sk_defaults)
    out['component'] = SN
    out["institution"] = "AVS"
    out["runNumber"] = "1"
    out['date'] = the_date
    out['passed'] = passed
    out['problems'] = problems
    out['properties']['OPERATOR'] = operator
    out['comments'] = comments

    return out


def create_delamination_test(session, SN, the_date=None, operator="", passed=True, problems=False, comments=None):
    """Create the delamination test JSON.

    Args:
        session: the DB session
        SN: Serial Number
        the_date: the date of the test
        operator: operator name
        passed: if test passed or not
        problems: if problems were found during test
        comments: list of comments to append to the test

    """
    if comments is None:
        comments = []

    the_date = ITkDButils.get_db_date(the_date)
    out = ITkDButils.get_test_skeleton(session, "CORE_PETAL", "DELAMINATION", sk_defaults, {"boolean": False})
    out['component'] = SN
    out['date'] = the_date
    out["institution"] = "AVS"
    out["runNumber"] = "1"
    out['passed'] = passed
    out['problems'] = problems
    out['properties']['OPERATOR'] = operator
    out['comments'] = comments

    return out


def create_grounding_test(session, SN, the_date=None, operator="", passed=True, problems=False, comments=None):
    """Create grounding test.

    Args:
        session: the DB session
        SN: Serial Number
        the_date: the date of the test
        operator: operator name
        passed: if test passed or not
        problems: if problems were found during test
        comments: list of comments to append to the test

    """
    if comments is None:
        comments = []
    the_date = ITkDButils.get_db_date(the_date)
    out = ITkDButils.get_test_skeleton(session, "CORE_PETAL", "GROUNDING_CHECK", sk_defaults, {"boolean": False})
    out['component'] = SN
    out['date'] = the_date
    out["institution"] = "AVS"
    out["runNumber"] = "1"
    out['passed'] = passed
    out['problems'] = problems
    out['properties']['OPERATOR'] = operator
    out['comments'] = comments

    return out


def create_metrology_test(session, SN, the_date=None, operator="", passed=True, problems=False, comments=None):
    """Metrology test.

    Args:
        session: the DB session
        SN: Serial Number
        the_date: the date of the test
        operator: operator name
        passed: if test passed or not
        problems: if problems were found during test
        comments: list of comments to append to the test

    """
    if comments is None:
        comments = []

    the_date = ITkDButils.get_db_date(the_date)
    out = ITkDButils.get_test_skeleton(session, "CORE_PETAL", "METROLOGY_AVS",
                                       sk_defaults, {"integer": -1, "float": -1.0})
    out['component'] = SN
    out['date'] = the_date
    out['passed'] = passed
    out["institution"] = "AVS"
    out["runNumber"] = "1"
    out['problems'] = problems
    out['properties']['OPERATOR'] = operator
    out['comments'] = comments

    return out


def split_comp_list(lst):
    """Split a  list of components separated by  various possible characters."""
    if lst is None:
        return []

    if isinstance(lst, float):
        return [lst]

    if isinstance(lst, int):
        return [lst]

    out = [lst]
    for sep in ['/', '\\', '\n']:
        if lst.find(sep) >= 0:
            out = [x.strip() for x in lst.split(sep)]
            break

    return out


def get_comments(txt):
    """Return test DB comment."""
    return split_comp_list(txt)


def get_float(cell, separator=None, default=0.0):
    """Return float from string."""
    txt = cell.value
    if txt is None:
        return default

    if separator is None:
        if isinstance(txt, float):
            return txt

        if isinstance(txt, int):
            return float(txt)

        if isinstance(txt, str):
            try:
                txt = txt.replace(',', '.')
                rr = re.findall(r"[-+]?[.]?[\d]+(?:,\d\d\d)*[\.]?\d*(?:[eE][-+]?\d+)?", txt)
                if len(rr) == 0:
                    return default

                val = float(rr[0])
                return val
            except ValueError:
                return default

        return default

    else:
        values = []
        for val in split_comp_list(txt):
            if isinstance(val, float) or isinstance(val, int):
                values.append(val)
            else:
                try:
                    v = float(val.strip().replace(',', '.'))
                except ValueError:
                    print("get_float: Cannot convert {} in {}".format(val, cell.coordinate))
                    v = default

                values.append(v)

        return values

def get_int(cell, default=None):
    """Get an int from a cell."""
    value = cell.value
    if value is None:
        return default

    return int(value)

def get_boolean(cell):
    """Get a boolean from a cell."""
    value = cell.value
    if value is None:
        return False

    else:
        txt = value.strip().lower()
        return txt == "pass"


def get_text(cell):
    """Get a string from a cell."""
    value = cell.value
    if value:
        value = value.strip()

    return value


def get_res_and_accep(sheet, indx):
    """Return result and acceptancee."""
    sval = sheet["g{}".format(indx)].value
    if isinstance(sval, float) or isinstance(sval, int):
        val = sval

    else:
        sval = ' '.join(sval.strip().split()).split()

        scale = 1.0
        try:
            rr = re.findall(r"[-+]?[.]?[\d]+(?:,\d\d\d)*[\.]?\d*(?:[eE][-+]?\d+)?", sval[0])
            if len(rr) == 0:
                val = 0.0
            else:
                val = float(rr[0])
                if len(sval)>1:
                    U = sval[1].upper()[0]
                    if U=='G':
                        scale = 1000

        except ValueError:
            val = 0

        val = val * scale
    #val = get_float(sheet["g{}".format(indx)])
    pass_val = sheet["h{}".format(indx)]
    if pass_val.value is None:
        # Operator did not set the PASS/FAIL thing
        accept = True
    else:
        accept = get_boolean(pass_val)

    return val, accept


def find_idx(lst, val):
    """Return index of occurence of val in lst."""
    R = re.compile(val, re.DOTALL|re.IGNORECASE)
    idx = [i for i, x in enumerate(lst) if (x and R.search(x))]
    return idx


def cell_value(sheet, coord):
    """Return the cell value."""
    cell = sheet[coord]
    if not isinstance(cell, MergedCell):
        return cell.value

    # "Oh no, the cell is merged!"
    for cell_range in sheet.merged_cells.ranges:
        if coord in cell_range:
            return cell_range.start_cell.value

    return None

def distance(P1, P2):
    """Distance between 2 points."""
    P = (P2-P1)
    S = np.sum(np.square(P))
    D = np.sqrt(S)
    return D

def find_test_indx(tests, val):
    """Apply brute force."""
    try:
        indx = tests.index(val)

    except ValueError:
        for ix, key in enumerate(tests):
            if key is None:
                continue

            if val in key:
                indx = ix
                break

    return indx

def fix_commas(P):
    """Fix commas in cells."""
    for i in range(2):
        if P[i] > 1000:
            P[i] /= 1000


def get_locator_distance(sheet, results, tests, nom, result):
    """Compute distance to nominal."""
    Pnom = np.array([
        get_float(sheet['E{}'.format(find_test_indx(tests, "{}_X".format(nom)))]),
        get_float(sheet['E{}'.format(find_test_indx(tests, "{}_Y".format(nom)))])
        ])
    fix_commas(Pnom)

    P = np.array([results["{}_X".format(result)], results["{}_Y".format(result)]])
    fix_commas(P)

    D = distance(Pnom, P)
    return D


def check_locator_positions(sheet, the_test, tests):
    """Check position of locators."""

    defects = []
    locators = [("PL01", "LOCATOR1"), ("PL02", "LOCATOR2"), ("PL03", "LOCATOR3"),
                ("FD01", "FIDUCIAL1"), ("FD02", "FIDUCIAL2")]

    for nom, val in locators:
        D = get_locator_distance(sheet, the_test["results"], tests,  nom, val)
        if D > 0.075:
            defects.append({
                "name": nom,
                "description": "{:.3f} mm out".format(D)
            })

    if len(defects)>0:
        the_test["passed"] = False
        the_test["defects"].extend(defects)


def check_for_problems(sheet, the_test, row_range):
    """Finds FAIL massages in the Acceptance column."""
    nfail = 0
    for row in range(row_range[0], row_range[1]):
        txt = get_text(sheet["h{}".format(row)])
        if txt is None:
            continue

        txt = txt.lower()
        if txt[0] == 'f':
            nfail += 1
            hdr = get_text(sheet["d{}".format(row)])
            result = get_float(sheet["g{}".format(row)])
            reason = "{} [{}]".format(result, cell_value(sheet, "i{}".format(row)))

            if reason:
                if len(reason) < 1:
                    msg = "{}: {}".format(hdr, reason)
                    the_test["defects"].append({"name": msg})
                else:
                    the_test["defects"].append({"name": hdr,
                                                "description": reason})

    #if nfail:
    #    the_test["passed"] = False


def get_coreID(bustapeID):
    """Build core SN from bus tape SN."""
    SN = "20USEBC" + bustapeID[-7:]
    return SN


def readFATfile(session, file_path, SN=None):
    """Read data from FAT excel file.

    Args:
        session: the DB session
        file_path: File path
        SN: COre serial number

    """
    # Open spreadsheet
    try:
        wb = XL.load_workbook(file_path, data_only=True)
    except InvalidFileException as ee:
        print("Could not open input file: ", file_path)
        print(ee)
        return None

    # Assume active sheet is the good one, otherwise will have to find in wb.sheetnames
    sheet = wb.active
    if sheet.max_row < 50 or sheet.max_column < 9:
        raise AVSDataException("Wrong FAT file")

    # Check the SN of the petal core
    if SN is None or len(SN) == 0:
        coreID = split_comp_list(sheet['C6'].value)
        if len(coreID) == 0:
            raise AVSDataException("Cannot figure out core SN in FAT file.")

        for cID in coreID:
            cmp = ITkDButils.get_DB_component(session, cID)
            if cmp["type"]["code"] == "BT_PETAL_FRONT":
                SN = get_coreID(cID)
                break

    batch = sheet['E6'].value
    operator = sheet['G6'].value

    txt = list(map(str.strip, sheet['i4'].value.split(':')))[1]
    test_date = dateutil.parser.parse(txt)

    # Get the test index
    test_name = [str(sheet[x][1].value) for x in range(1, sheet.max_row)]
    tests = [str(sheet[x][3].value) for x in range(1, sheet.max_row)]

    # This is to avoid adding 1 for cell names...
    tests.insert(0, None)
    test_name.insert(0, None)

    #
    # Visual inspection
    vi_text = get_text(sheet['i9'])
    vi_result = get_text(sheet['g9'])
    vi_pass = sheet['h9'].value.strip().lower() == "pass"
    vi_defects = []
    if vi_pass:
        if vi_result and len(vi_result):
            if vi_text and len(vi_text):
                vi_text = vi_result + '\n' + vi_text
            else:
                vi_text = vi_result

    else:
        vi_defects.append({"name": "PETAL_VI_DEFECT", "description": vi_result})

    vi_test = create_visual_inpection(session, SN, test_date, operator, vi_pass,
                                      comments=get_comments(vi_text))
    for df in vi_defects:
        vi_test["defects"].append(df)

    #
    # Delamination test
    dl_text = get_text(sheet['i10'])
    dl_result = get_text(sheet['g10'])
    dl_pass = sheet['h10'].value.strip().lower() == "pass"
    dl_defects = []
    if dl_pass:
        if dl_result and len(dl_result):
            if dl_text and len(dl_text):
                dl_text = dl_result + '\n' + dl_text
            else:
                dl_text = dl_result

    else:
        dl_defects.append({"name": "PETAL_DL_DEFECT",
                           "description": dl_result})

    delamination_test = create_delamination_test(session, SN, test_date, operator, dl_pass,
                                                 comments=get_comments(dl_text))
    for df in dl_defects:
        delamination_test["defects"].append(df)

    #
    # Conductivity
    # TODO: read proper rows
    grounding_test = create_grounding_test(session, SN, test_date, operator)
    cond_val, cond_pass = get_res_and_accep(sheet, tests.index("COND"))
    if "INS_LOOP" in tests:
        loop_val, loop_pass = get_res_and_accep(sheet, tests.index("INS_LOOP"))
    else:
        loop_val, loop_pass = get_res_and_accep(sheet, tests.index("INS"))

    if "INS_LOOP_GND" in tests:
        loop_gnd_val, loop_gnd_pass = get_res_and_accep(sheet, tests.index("INS_LOOP_GND"))
    else:
        loop_gnd_val, loop_gnd_pass = get_res_and_accep(sheet, tests.index("INS_FACE"))

    passed = cond_pass and loop_pass and loop_gnd_pass
    grounding_test["passed"] = passed
    grounding_test["results"]["RESISTANCE_FB"] = cond_val
    grounding_test["results"]["RESISTANCE_PIPES"] = loop_val
    grounding_test["results"]["RESISTANCE_PIPE_GND"] = loop_gnd_val
    #check_for_problems(sheet, grounding_test, [tests.index('COND'), tests.index("WEIGH")])

    #
    # Weight
    petal_weight, weight_pass = get_res_and_accep(sheet, tests.index("WEIGH"))

    #
    # Metrology AVS
    metrology_test = create_metrology_test(session, SN, test_date, operator)
    metrology_test["results"]["LOCATOR1_DIAMETER"] = get_float(sheet['g{}'.format(tests.index("PL01_DIAM"))])
    metrology_test["results"]["LOCATOR2_DIAMETER"] = get_float(sheet['g{}'.format(tests.index("PL02_DIAM"))])
    metrology_test["results"]["LOCATOR3_DIAMETER"] = get_float(sheet['g{}'.format(tests.index("PL03_DIAM"))])
    metrology_test["results"]["LOCATOR1_X"] = get_float(sheet['g{}'.format(find_idx(tests, "PL01_X")[0])])
    metrology_test["results"]["LOCATOR1_Y"] = get_float(sheet['g{}'.format(find_idx(tests, "PL01_Y")[0])])
    metrology_test["results"]["LOCATOR2_X"] = get_float(sheet['g{}'.format(find_idx(tests, "PL02_X")[0])])
    metrology_test["results"]["LOCATOR2_Y"] = get_float(sheet['g{}'.format(find_idx(tests, "PL02_Y")[0])])
    metrology_test["results"]["LOCATOR3_X"] = get_float(sheet['g{}'.format(find_idx(tests, "PL03_X")[0])])
    metrology_test["results"]["LOCATOR3_Y"] = get_float(sheet['g{}'.format(find_idx(tests, "PL03_Y")[0])])
    metrology_test["results"]["FIDUCIAL1_DIAMETER"] = get_float(sheet["g{}".format(find_idx(tests, "FD01_DIAM")[0])])
    metrology_test["results"]["FIDUCIAL1_X"] = get_float(sheet["g{}".format(find_idx(tests, "FD01_X")[0])])
    metrology_test["results"]["FIDUCIAL1_Y"] = get_float(sheet["g{}".format(find_idx(tests, "FD01_Y")[0])])
    metrology_test["results"]["FIDUCIAL2_DIAMETER"] = get_float(sheet["g{}".format(find_idx(tests, "FD02_DIAM")[0])])
    metrology_test["results"]["FIDUCIAL2_X"] = get_float(sheet["g{}".format(find_idx(tests, "FD02_X")[0])])
    metrology_test["results"]["FIDUCIAL2_Y"] = get_float(sheet["g{}".format(find_idx(tests, "FD02_Y")[0])])
    metrology_test["results"]["ANGLE_VCHANNEL"] = get_float(sheet["g{}".format(find_idx(tests, "VANGL")[0])])
    metrology_test["results"]["ENVELOPE"] = get_float(sheet["g{}".format(find_idx(tests, "ENVEL")[0])])
    metrology_test["results"]["COPLANARITY_FRONT"] = get_float(sheet["g{}".format(find_idx(tests, "F.PL_PLAN")[0])])
    metrology_test["results"]["LOCAL_FLATNESS_FRONT"] = get_float(sheet["g{}".format(find_idx(tests, "F.FS_PLAN")[0])], '/')
    metrology_test["results"]["PARALLELISM_FRONT"] = get_float(sheet["g{}".format(find_idx(tests, "F.PARAL")[0])])
    metrology_test["results"]["COPLANARITY_BACK"] = get_float(sheet["g{}".format(find_idx(tests, "B.PL_PLAN")[0])])
    metrology_test["results"]["LOCAL_FLATNESS_BACK"] = get_float(sheet["g{}".format(find_idx(tests, "B.FS_PLAN")[0])], '/')
    metrology_test["results"]["PARALLELISM_BACK"] = get_float(sheet["g{}".format(find_idx(tests, "B.PARAL")[0])])

    # Get defects
    check_locator_positions(sheet, metrology_test, tests)
    check_for_problems(sheet, metrology_test, [tests.index("VANGL"), sheet.max_row])

    return vi_test, delamination_test, grounding_test, metrology_test, batch, petal_weight


def find_label(sheet, label, column, max_row=20):
    """Find label in given column

    Args:
        sheet (): The spread sheet
        label (): The label to search for
        column (): The column to scan.

    Return:
        indx (int) - the index. <0 means not found.
    """
    indx = -1
    for i in range(1, 15):
        val = sheet["{}{}".format(column, i)].value
        if val is None:
            continue

        if val.find(label)>=0:
            indx = i
            break

    return indx

def readProductionSheet(session, file_path, SN):
    """Read data fro AVS PS.

    Args:
        session: the DB session
        file_path: path of input file
        SN: The serial number
        write_json: if true, test json is writen to file.

    """
    try:
        wb = XL.load_workbook(file_path, data_only=True)
    except InvalidFileException as ee:
        print("Could not open input file: ", file_path)
        print(ee.message)
        return None

    # Sometimes it comes with an empty sheet in front
    sheet = None
    for name in wb.sheetnames:
        value = wb[name]["C7"].value
        if value is not None:
            sheet = wb[name]
            break
        
    #sheet = wb.active
    if sheet.max_row > 30 or sheet.max_column > 9:
        raise AVSDataException("Wrong PS file:\nmx row {} mx_col {}".format(sheet.max_row, sheet.max_column))

    # Find the start
    indx = find_label(sheet, "PETAL", "A")
    if indx < 0:
        # Try with second column
        indx = find_label(sheet, "PETAL", "B")
        if indx < 0:
            print("Wrong Production Sheet.")
            return None
        else:
            icol = ord('B')
            ccol = 'B'
    else:
        icol = ord('A')
        ccol = 'A'

    i_items = find_label(sheet, "DRAWING", ccol)
    if i_items < 0:
        print("Wrong Production Sheet.")
        return None

    i_items += 1

    # Get the SN of the fron facesheet and create the Petal SN
    ID = sheet["{}{}".format(chr(icol+2), i_items)].value.strip()
    SN = get_coreID(ID)
    nn = get_int(sheet["{}{}".format(chr(icol+1), indx)])
    petal_id = "PPC.{:03d}".format(nn)
    mould_id = sheet["{}{}".format(chr(icol+1), indx+3)].value

    # find the date (use the end date)
    start_date = sheet["{}{}".format(chr(icol+4), indx+2)].value
    end_date = sheet["{}{}".format(chr(icol+4), indx+3)].value
    test_date = start_date

    manager = sheet["{}{}".format(chr(icol+4), indx+1)].value

    # Manufacturing
    id_col = chr(icol+2)
    w_col = chr(icol+3)
    n_col = chr(icol+4)
    comments = get_comments(sheet['a25'].value)
    manufacturing = create_manufacturing(session, SN, test_date, manager, comments=comments)
    manufacturing['properties']['START_DATE'] = ITkDButils.get_db_date(start_date)
    manufacturing['properties']['FINISH_DATE'] = ITkDButils.get_db_date(end_date)
    manufacturing["properties"]["MOULD_ID"] = mould_id
    manufacturing["properties"]["PROCESS_DOCUMENT"] = sheet["{}{}".format(chr(icol+4), indx)].value
    manufacturing["results"]["LOCATOR_A"] = sheet["{}{}".format(id_col, i_items+2)].value
    manufacturing["results"]["LOCATOR_B"] = sheet["{}{}".format(id_col, i_items+3)].value
    manufacturing["results"]["LOCATOR_C"] = sheet["{}{}".format(id_col, i_items+4)].value
    manufacturing["results"]["HONEYCOMBSET"] = split_comp_list( sheet["{}{}".format(id_col, i_items+5)].value)
    manufacturing["results"]["HONEYCOMBSET"] = [ str(x) for x in manufacturing["results"]["HONEYCOMBSET"]]
    
    manufacturing["results"]["EPOXY_ADHESIVE"] = split_comp_list(sheet["{}{}".format(id_col, i_items+8)].value)
    manufacturing["results"]["EPOXY_PUTTY"] = split_comp_list( sheet["{}{}".format(id_col, i_items+9)].value)
    manufacturing["results"]["EPOXY_CONDUCTIVE"] = split_comp_list( sheet["{}{}".format(id_col, i_items+10)].value)

    # Weighing
    weighing = create_weight(session, SN, test_date, manager)
    scol = "{}{}:{}{}".format(w_col, i_items, w_col, i_items+10)
    comp_weight = [get_float(x[0]) for x in sheet[scol]]
    petal_weight = sum([float(x) for x in comp_weight])
    weighing["results"]["WEIGHT_FACING_FRONT"] = comp_weight[0]
    weighing["results"]["WEIGHT_FACING_BACK"] = comp_weight[1]
    weighing["results"]["WEIGHT_LOCATOR_A"] = comp_weight[2]
    weighing["results"]["WEIGHT_LOCATOR_B"] = comp_weight[3]
    weighing["results"]["WEIGHT_LOCATOR_C"] = comp_weight[4]
    weighing["results"]["WEIGHT_COOLINGLOOPASSEMBLY"] = comp_weight[6]
    weighing["results"]["WEIGHT_HONEYCOMBSET"] = comp_weight[5]
    weighing["results"]["WEIGHT_EPOXYADHESIVE"] = comp_weight[8]
    weighing["results"]["WEIGHT_EPOXYPUTTY"] = comp_weight[9]
    weighing["results"]["WEIGHT_EPOXYCONDUCTIVE"] = comp_weight[10]
    weighing["results"]["WEIGHT_CORE"] = petal_weight

    # Comments
    for i in range(i_items, i_items+11):
        cell_id = sheet['{}{}'.format(ccol, i)].value
        comment = sheet['{}{}'.format(n_col, i)].value
        if comment is not None:
            comment = comment.strip()
            if len(comment):
                msg = "{}: {}".format(cell_id, comment)
                weighing["comments"].append(msg)

    DESY = {
        "FacingFront": sheet["{}{}".format(id_col, i_items)].value.strip(),
        "FacingBack": sheet["{}{}".format(id_col, i_items+1)].value.strip(),
        "CoolingLoop": sheet["{}{}".format(id_col, i_items+6)].value.strip(),
        "AllcompSet": sheet["{}{}".format(id_col, i_items+7)].value.strip(),
        "HoneyCombSet": manufacturing["results"]["HONEYCOMBSET"]
    }

    return manufacturing, weighing, DESY, petal_id


if __name__ == "__main__":
    parser = ArgumentParser()
    parser.add_argument('files', nargs='*', help="Input files")
    parser.add_argument("--SN", dest="SN", type=str, default="SNnnn",
                        help="Module serial number")

    options = parser.parse_args()
    if len(options.files) == 0:
        print("I need an input file")
        sys.exit()

    dlg = ITkDBlogin.ITkDBlogin()
    client = dlg.get_client()
    if client is None:
        print("Could not connect to DB with provided credentials.")
        dlg.die()
        sys.exit()

    fnam = Path(options.files[0]).expanduser().resolve()
    readProductionSheet(client, fnam, options.SN)
#    readFATfile(client, fnam, options.SN)
    dlg.die()
