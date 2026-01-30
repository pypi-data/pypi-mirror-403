#!/usr/bin/env python3
"""Create a table from a set of PDB JSon files."""

import sys
import json
import openpyxl as xl
from openpyxl.worksheet.cell_range import CellRange
from pathlib import Path


def comparisonTable(files, options):
    """Creqt3 the table."""
    wb = xl.Workbook()
    ws = wb.active

    dir(ws)

    firstC = ["All", "R0", "R1", "R2", "R3_0", "R3_1", "R4_0", "R4_1", "R5_0", "R5_1"]
    for i, lbl in enumerate(firstC):
        ws.cell(row=3+i, column=1).value = lbl

    for i, ifile in enumerate(files):
        P = Path(ifile).expanduser().resolve()
        print(P.stem)
        petal, side = P.stem.split('_')
        with open(P, "r") as inp:
            Js = json.load(inp)
            if len(Js["results"]["METROLOGY_FRONT"]):
                Js = Js["results"]["METROLOGY_FRONT"]
            else:
                Js = Js["results"]["METROLOGY_BACK"]

        if i % 2 == 0:
            ws.cell(row=1, column=i+2).value = petal

        ws.cell(row=2, column=i+2).value = side
        ws.cell(row=3, column=i+2).value = Js["FLATNESS_GLOBAL"]
        ws.cell(row=3, column=i+2).number_format = '0.000'
        for j, v in enumerate(Js["FLATNESS_LOCAL"]):
            cell = ws.cell(row=4+j, column=i+2)
            cell.value = v
            cell.number_format = '0.000'

    wb.save(options.out)


if __name__ == "__main__":
    from argparse import ArgumentParser
    parser = ArgumentParser()
    parser.add_argument('files', nargs='*', help="Input files")
    parser.add_argument("--out", default="table.xlsx")

    options = parser.parse_args()
    if len(options.files) == 0:
        print(sys.argv[0])
        print("I need an input file")
        sys.exit()

    comparisonTable(options.files, options)
