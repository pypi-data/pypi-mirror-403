"""
Export utilities for Overwatch admin panel.
"""

import csv
import json
from pathlib import Path
from typing import Any


async def export_data(items: list[Any], format: str, model_info: Any) -> str:
    """
    Export data to various formats.

    Args:
        items: List of items to export
        format: Export format (csv, json, xlsx)
        model_info: Model information (can be ModelInfo or any object with get_simple_fields method)

    Returns:
        Path to exported file
    """
    if not items:
        raise ValueError("No items to export")

    # Create temporary directory for exports
    export_dir = Path("tmp/exports")
    export_dir.mkdir(parents=True, exist_ok=True)

    # Generate filename - handle None model_info
    model_name = getattr(model_info, 'name', 'data') if model_info else 'data'
    filename = f"{model_name}_export.{format}"
    file_path = export_dir / filename

    # Convert items to dictionaries
    data = []
    for item in items:
        if hasattr(item, "__dict__"):
            item_dict = {}
            # Handle None model_info by using all attributes
            if model_info and hasattr(model_info, 'get_simple_fields'):
                field_names = model_info.get_simple_fields()
            else:
                # Use all attributes that aren't private or methods
                field_names = [attr for attr in dir(item) if not attr.startswith('_') and not callable(getattr(item, attr))]

            for field_name in field_names:
                if hasattr(item, field_name):
                    value = getattr(item, field_name)
                    # Convert complex objects to strings
                    if hasattr(value, "__dict__"):
                        item_dict[field_name] = str(value)
                    else:
                        item_dict[field_name] = value
            data.append(item_dict)
        else:
            data.append(item)

    # Export based on format
    if format.lower() == "csv":
        await _export_csv(data, file_path, model_info)
    elif format.lower() == "json":
        await _export_json(data, file_path)
    elif format.lower() == "xlsx":
        await _export_xlsx(data, file_path, model_info)
    else:
        raise ValueError(f"Unsupported export format: {format}")

    return str(file_path)


async def _export_csv(
    data: list[dict[str, Any]], file_path: Path, model_info: Any
) -> None:
    """Export data to CSV format."""
    if not data:
        return

    # Get field names from model info
    if model_info and hasattr(model_info, 'get_simple_fields'):
        fieldnames = model_info.get_simple_fields()
    else:
        # Use all keys from the first data item
        fieldnames = list(data[0].keys()) if data else []

    with open(file_path, "w", newline="", encoding="utf-8") as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()

        for row in data:
            # Filter row to only include fieldnames
            filtered_row = {k: v for k, v in row.items() if k in fieldnames}
            writer.writerow(filtered_row)


async def _export_json(data: list[dict[str, Any]], file_path: Path) -> None:
    """Export data to JSON format."""
    with open(file_path, "w", encoding="utf-8") as jsonfile:
        json.dump(data, jsonfile, indent=2, ensure_ascii=False, default=str)


async def _export_xlsx(
    data: list[dict[str, Any]], file_path: Path, model_info: Any
) -> None:
    """Export data to XLSX format."""
    try:
        from openpyxl import Workbook
    except ImportError as e:
        raise ImportError(
            "openpyxl is required for XLSX export. Install with: pip install openpyxl"
        ) from e

    if not data:
        return

    # Create workbook
    wb = Workbook()
    ws = wb.active

    if ws is None:
        raise RuntimeError("Failed to create worksheet in workbook")

    # Handle None model_info for worksheet title
    sheet_title = getattr(model_info, 'name', 'Data') if model_info else 'Data'
    ws.title = sheet_title

    # Get field names from model info
    if model_info and hasattr(model_info, 'get_simple_fields'):
        fieldnames = model_info.get_simple_fields()
    else:
        # Use all keys from first data item
        fieldnames = list(data[0].keys()) if data else []

    # Write header
    for col_num, fieldname in enumerate(fieldnames, 1):
        ws.cell(row=1, column=col_num, value=fieldname)

    # Write data
    for row_num, row_data in enumerate(data, 2):
        for col_num, fieldname in enumerate(fieldnames, 1):
            value = row_data.get(fieldname, "")
            ws.cell(row=row_num, column=col_num, value=value)

    # Save workbook
    wb.save(file_path)
