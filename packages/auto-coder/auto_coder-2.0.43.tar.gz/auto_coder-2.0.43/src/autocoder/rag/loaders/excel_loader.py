from typing import Dict, List, Tuple
from pathlib import Path
from zipfile import ZipFile
from xml.etree import ElementTree as ET
from openpyxl import load_workbook
from jinja2 import Template

try:
    import pandas as pd

    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False


def extract_text_from_excel(excel_path) -> List[Tuple[str, str]]:
    """
    从 Excel 文件中提取每个 Sheet 的文本内容。
    首选使用 pandas 读取；若失败，则使用 openpyxl 读取；若仍失败，则降级使用 zip+xml 解析仅提取单元格文本。
    """
    tmpl = Template(
        """{% for row in rows %}
{% for cell in row %}"{{ cell }}"{% if not loop.last %},{% endif %}{% endfor %}{% if not loop.last %}{% endif %}{% endfor %}
        """
    )

    # 首选 pandas 读取
    if HAS_PANDAS:
        try:
            return _extract_text_from_excel_pandas(excel_path, tmpl)
        except Exception:
            pass

    # 降级到 openpyxl 读取
    try:
        return _extract_text_from_excel_openpyxl(excel_path, tmpl)
    except Exception:
        # 最后降级：不解析样式，直接解析 sharedStrings 和 sheetData
        return _extract_text_from_excel_zip(Path(excel_path), tmpl)


def _extract_text_from_excel_pandas(
    excel_path, tmpl: Template
) -> List[Tuple[str, str]]:
    """
    使用 pandas 读取 Excel 文件。
    """
    sheet_list: List[Tuple[str, str]] = []
    excel_file = pd.ExcelFile(excel_path)
    for sheet_name in excel_file.sheet_names:
        df = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)
        if df.empty:
            continue
        # 删除全为空的行
        df = df.dropna(how="all")
        if df.empty:
            continue
        # 将 NaN 转换为空字符串
        df = df.fillna("")
        # 转换为嵌套列表
        rows = df.values.tolist()
        content = tmpl.render(rows=rows)
        sheet_list.append([str(excel_path) + f"#{sheet_name}", content])
    excel_file.close()
    return sheet_list


def _extract_text_from_excel_openpyxl(
    excel_path, tmpl: Template
) -> List[Tuple[str, str]]:
    """
    使用 openpyxl 读取 Excel 文件。
    """
    sheet_list: List[Tuple[str, str]] = []
    wb = load_workbook(excel_path, data_only=True, read_only=True, keep_links=False)
    for ws in wb:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        # 过滤掉rows中全是null的行
        rows = [row for row in rows if any(row)]
        # 所有的None都转换成空字符串
        rows = [[cell if cell is not None else "" for cell in row] for row in rows]
        content = tmpl.render(rows=rows)
        sheet_list.append([str(excel_path) + f"#{ws.title}", content])
    wb.close()
    return sheet_list


def _extract_text_from_excel_zip(
    excel_path: Path, tmpl: Template
) -> List[Tuple[str, str]]:
    """
    直接解析 xlsx(zip) 的 XML 文件，忽略样式，仅还原单元格文本。
    """
    sheet_list: List[Tuple[str, str]] = []
    with ZipFile(excel_path) as zf:
        # 解析 sharedStrings
        shared_strings: List[str] = []
        if "xl/sharedStrings.xml" in zf.namelist():
            with zf.open("xl/sharedStrings.xml") as f:
                tree = ET.parse(f)
                root = tree.getroot()
                # {namespace} 标签处理
                ns = _ns(root.tag)
                for si in root.findall(f".//{ns}si"):
                    # si 内可能包含多个 t，需要拼接
                    text_fragments: List[str] = []
                    for t in si.findall(f".//{ns}t"):
                        text_fragments.append(t.text or "")
                    shared_strings.append("".join(text_fragments))

        # 解析 workbook.xml 获取 sheet 列表及 r:id
        with zf.open("xl/workbook.xml") as f:
            wb_tree = ET.parse(f)
            wb_root = wb_tree.getroot()
            wb_ns = _ns(wb_root.tag)
            sheets = wb_root.find(f"{wb_ns}sheets")
            if sheets is None:
                return sheet_list
            sheet_name_to_rid: List[Tuple[str, str]] = []
            for sheet in sheets.findall(f"{wb_ns}sheet"):
                name = sheet.attrib.get("name", "Sheet")
                rid = sheet.attrib.get(
                    "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id",
                    "",
                )
                if rid:
                    sheet_name_to_rid.append((name, rid))

        # 解析 workbook relationships: rId -> target (worksheets/sheetN.xml)
        rid_to_target: Dict[str, str] = {}
        with zf.open("xl/_rels/workbook.xml.rels") as f:
            rels_tree = ET.parse(f)
            rels_root = rels_tree.getroot()
            for rel in rels_root:
                if rel.tag.endswith("Relationship"):
                    rid = rel.attrib.get("Id", "")
                    target = rel.attrib.get("Target", "")
                    if rid and target:
                        rid_to_target[rid] = target

        # 逐个 sheet 解析
        for sheet_name, rid in sheet_name_to_rid:
            target = rid_to_target.get(rid)
            if not target:
                continue
            # 兼容前缀
            sheet_path = f"xl/{target}" if not target.startswith("xl/") else target
            if sheet_path not in zf.namelist():
                continue
            with zf.open(sheet_path) as f:
                sheet_tree = ET.parse(f)
                sheet_root = sheet_tree.getroot()
                sheet_ns = _ns(sheet_root.tag)
                rows_data: List[List[str]] = []
                for row in sheet_root.findall(f".//{sheet_ns}row"):
                    # 收集该行的 (col_index -> value)
                    col_to_val: Dict[int, str] = {}
                    max_col_idx = 0
                    for cell in row.findall(f"{sheet_ns}c"):
                        cell_type = cell.attrib.get("t")
                        cell_ref = cell.attrib.get("r", "")
                        col_idx = _col_index_from_ref(cell_ref)
                        if col_idx > max_col_idx:
                            max_col_idx = col_idx
                        value = ""
                        if cell_type == "s":
                            v = cell.find(f"{sheet_ns}v")
                            if v is not None and v.text is not None:
                                try:
                                    s_idx = int(v.text)
                                    if 0 <= s_idx < len(shared_strings):
                                        value = shared_strings[s_idx]
                                except Exception:
                                    value = v.text or ""
                        elif cell_type == "inlineStr":
                            is_el = cell.find(f"{sheet_ns}is")
                            if is_el is not None:
                                # 兼容多段 t
                                text_fragments: List[str] = []
                                for t in is_el.findall(f".//{sheet_ns}t"):
                                    text_fragments.append(t.text or "")
                                value = "".join(text_fragments)
                        else:
                            v = cell.find(f"{sheet_ns}v")
                            if v is not None and v.text is not None:
                                value = v.text
                        col_to_val[col_idx] = value
                    # 构造该行的顺序列表（从 1 到 max_col_idx）
                    if max_col_idx > 0:
                        row_vals = [
                            col_to_val.get(i, "") for i in range(1, max_col_idx + 1)
                        ]
                        # 过滤整行空
                        if any(val for val in row_vals):
                            rows_data.append(row_vals)
                if rows_data:
                    content = tmpl.render(rows=rows_data)
                    sheet_list.append([str(excel_path) + f"#{sheet_name}", content])
        return sheet_list


def _ns(tag: str) -> str:
    """
    从带命名空间的标签提取命名空间前缀形式。
    如: '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}worksheet'
    返回: '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}'
    """
    if tag.startswith("{"):
        return tag[: tag.index("}") + 1]
    return ""


def _col_index_from_ref(cell_ref: str) -> int:
    """
    从单元格引用（例如 'C5'）提取列索引（从 1 开始）。
    """
    col_letters = []
    for ch in cell_ref:
        if "A" <= ch <= "Z":
            col_letters.append(ch)
        else:
            break
    if not col_letters:
        return 1
    # 将列字母转换为索引（A->1, Z->26, AA->27 ...）
    idx = 0
    for ch in col_letters:
        idx = idx * 26 + (ord(ch) - ord("A") + 1)
    return idx
