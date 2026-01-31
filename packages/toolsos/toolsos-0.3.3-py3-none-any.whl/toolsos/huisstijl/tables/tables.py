import re
from itertools import groupby
from typing import Any, Callable, Dict

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter

Fmt = list[list[dict[str, Any]]]
Mapping = Dict[str, Dict[str, str | int | bool]]

LOOKUP: dict[str, Callable] = {
    "font": Font,
    "fill": PatternFill,
    "alignment": Alignment,
    "border": Border,
    "protection": Protection,
    "side": Side,
}

LETTERS = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"


def set_global_style(style: str) -> None:
    """Table styling is based on global style defintion, because constanly passing
    around style dictionary as an argument would be tedious. In the future this should
    however be implemented because setting a global parameter like this is probably
    an anti-pattern.S

    Args:
        style (str): Input either `old` or `new` to set global style dictionary
    """
    global STYLES
    if style == "old":
        from .table_styles import STYLE_OLD

        STYLES = STYLE_OLD
    elif style == "new":
        from .table_styles import STYLE_NEW

        STYLES = STYLE_NEW


def cols_to_str(df: pd.DataFrame) -> pd.DataFrame:
    """Change column names in to string. Multiindex column names are nog changed because
    these are always strings

    Args:
        df (pd.DataFrame): Dataframe

    Returns:
        pd.DataFrame: Dataframe with column names as strings
    """

    # Multiindex columns are always strings and therefore can't be casted as string
    if df.columns.nlevels == 1:
        df.columns = df.columns.astype(str)

    return df


def get_max_col_widths(data: pd.DataFrame | np.ndarray) -> list[float]:
    col_widths = []
    if isinstance(data, pd.DataFrame):
        for col in zip(*flatten_multiindex_columns(data)):
            col_widths.append(max(len(e) for e in col))
    else:
        for col in zip(*data):
            col_widths.append(max(len(str(e)) for e in col))

    col_widths = [col_width * 1.13 for col_width in col_widths]

    return col_widths


def flatten_multiindex_columns(df: pd.DataFrame) -> list:
    column_multi = []
    for level in range(df.columns.nlevels):
        column_multi.append(df.columns.get_level_values(level))
    return column_multi


def df_to_array(df: pd.DataFrame) -> np.ndarray:
    """Turn dataframe into array that includes column names as the first row

    Args:
        df (pd.DataFrame): Dataframe to be turned into an array

    Returns:
        np.array: Array that includes the data and the column names as the first row
    """
    column_names = flatten_multiindex_columns(df)

    return np.vstack([column_names, df.to_numpy()])


def get_cells_to_merge(df: pd.DataFrame) -> dict[int : list[int, int]]:
    """Pandas dataframes sometimes have mutliindex columns. For all but the last level
    a dictionary is created to merge the cells. The last level isn't merged because these
    are these contain unique column names

    Args:
        df (pd.DataFrame): Pandas dataframe. If the dataframe has multicolumn indices
        a dictionary containg the cells to merge is returned

    Returns:
        dict[int: list[int, int]]: Dictionary containg the cells to merge
    """
    levels = flatten_multiindex_columns(df)[:-1]

    cells_to_merge = {}
    for level_idx, level in enumerate(levels):
        start_col = 1
        start_col = 1
        merge_cells = []
        for _, val in groupby(level):
            val = list(val)
            merge_cells.append([start_col, start_col + len(val) - 1])
            start_col += len(val)
        cells_to_merge[level_idx] = merge_cells

    return cells_to_merge


def get_fmt_table(arr: np.ndarray) -> Fmt:
    """Create nested list with dictionary inside that is the same size as the original
    dataframe including column names

    Args:
        df (pd.DataFrame): Dataframe to be written to excel

    Returns:
        Fmt: Return empty nest list that will later be used to store formatting info
    """
    fmt = []
    for _ in range(arr.shape[0] + 1):
        row: list = []
        for _ in range(arr.shape[1]):
            row.append({})
        fmt.append(row)

    return fmt


def update_format(fmt: Fmt, row_idx: int, col_idx: int, mapping: Mapping) -> Fmt:
    """Update the cell containing the formatting info

    Args:
        fmt (Fmt): nested list containing the formatting info
        row_idx (int): row index
        col_idx (int): column index
        mapping (Mapping): formatting info

    Returns:
        Fmt: nested list containing the updated formatting info
    """
    for fmt_type, args in mapping.items():
        cell = fmt[row_idx][col_idx].get(fmt_type)

        if not cell:
            fmt[row_idx][col_idx][fmt_type] = args
        else:
            fmt[row_idx][col_idx][fmt_type] = cell | args

    return fmt


def set_style_all(fmt: Fmt, mapping: Mapping) -> Fmt:
    """Set the formatting for all cells

    Args:
        fmt (Fmt): nested list containing the formatting info
        mapping (Mapping): formatting info

    Returns:
        Fmt: nested list containing the formatting info
    """
    for row_idx, row in enumerate(fmt):
        for col_idx, _ in enumerate(row):
            update_format(fmt, row_idx, col_idx, mapping)

    return fmt


def set_style_row(
    fmt: Fmt,
    row_idxs: int | list,
    mapping: Mapping,
    exlude_col_ids: int | list | None = None,
) -> Fmt:
    """Set the formatting on a row

    Args:
        fmt (Fmt): nested list containing the formatting info
        row_idxs (int | list): The indices of the rows to be updated
        mapping (Mapping): formatting info
        exlude_col_ids (int | list | None, optional): Indices of the cols to be excluded
        when setting formatting for a row. Defaults to None.

    Returns:
        Fmt: nested list containing the formatting info
    """
    if isinstance(row_idxs, int):
        row_idxs = [row_idxs]

    if exlude_col_ids is not None:
        if isinstance(exlude_col_ids, int):
            exlude_col_ids = [exlude_col_ids]
    else:
        exlude_col_ids = []

    for row_idx in row_idxs:
        for col_idx, _ in enumerate(fmt[row_idx]):
            if col_idx not in exlude_col_ids:
                update_format(fmt, row_idx, col_idx, mapping)

    return fmt


def set_style_col(
    fmt: Fmt,
    col_idxs: int | list,
    mapping: Mapping,
    exlude_row_ids: int | list | None = None,
) -> Fmt:
    """Set the formatting on a row

    Args:
        fmt (Fmt): nested list containing the formatting info
        row_idxs (int | list): The indices of the rows to be updated
        mapping (Mapping): formatting info
        exlude_row_ids (int | list | None, optional): Indices of the rows to be excluded
        when setting formatting for a col. Defaults to None.

    Returns:
        Fmt: nested list containing the formatting info
    """
    if isinstance(col_idxs, int):
        col_idxs = [col_idxs]

    if exlude_row_ids is not None:
        if isinstance(exlude_row_ids, int):
            exlude_row_ids = [exlude_row_ids]
    else:
        exlude_row_ids = []

    for col_idx in col_idxs:
        for row_idx, _ in enumerate(fmt):
            if row_idx not in exlude_row_ids:
                update_format(fmt, row_idx, col_idx, mapping)

    return fmt


def excel_style(row: int, col: int) -> str:
    """Convert given row and column number to an Excel-style cell name."""
    result: list = []
    while col:
        col, rem = divmod(col - 1, 26)
        result[:0] = LETTERS[rem]
    return "".join(result) + str(row)


def get_cols_id_with_pattern(df: pd.DataFrame, pattern: str) -> list[int]:
    """Get columns indices from columns matching a regex pattern

    Args:
        df (pd.DataFrame): Input dataframe
        pattern (str): regex pattern to get columns indices when matching

    Returns:
        list[int]: list with column indices matching pattern
    """

    if isinstance(df.columns, pd.MultiIndex):
        # Use the lowest level in the multi-index column
        return [
            idx for idx, col in enumerate(df.columns) if re.findall(pattern, col[-1])
        ]
    else:
        return [idx for idx, col in enumerate(df.columns) if re.findall(pattern, col)]


def get_string_cols_ids(df: pd.DataFrame) -> list[int]:
    """Get column indices of string columns

    Args:
        df (pd.DataFrame): Input dataframe

    Returns:
        list[int]: list with column indices of string columns
    """
    return [i for i, dtype in enumerate(df.dtypes) if dtype == "O"]


def get_numeric_col_ids(df: pd.DataFrame) -> list[int]:
    """Get column indices of numeric columns

    Args:
        df (pd.DataFrame): Input dataframe

    Returns:
        list[int]: list with column indices of numeric columns
    """
    num_cols = df.select_dtypes("number").columns
    return [i for i, col in enumerate(df.columns) if col in num_cols]


def cell_formatting(
    arr: np.ndarray,
    default_format: Mapping | None = None,
    blue_row_ids: int | list | None = None,
    light_blue_row_ids: int | list | None = None,
    light_blue_col_ids: int | list | None = None,
    left_align_ids: int | list | None = None,
    right_align_ids: int | list | None = None,
    perc_col_ids: int | list | None = None,
    perc_col_format: str | None = None,
    float_col_ids: int | list | None = None,
    float_col_format: str | None = None,
    blue_border_ids: bool | None = None,
    number_format: str | None = None,
):
    """Function to create the nested list with the shape of the input data (including columns)
    containing dictionaries with the formatting

    Args:
        arr (np.ndarray): array representing the data
        default_format (Mapping | None, optional): Default format applied to all cells. Defaults to None.
        blue_row_ids (int | list | None, optional): The ids of the rows to be colored blue. Defaults to None.
        light_blue_row_ids (int | list | None, optional): _description_. Defaults to None.
        light_blue_col_ids (int | list | None, optional): _description_. Defaults to None.
        left_align_ids (int | list | None, optional): _description_. Defaults to None.
        right_align_ids (int | list | None, optional): _description_. Defaults to None.
        perc_col_ids (int | list | None, optional): _description_. Defaults to None.
        perc_col_format (str | None, optional): _description_. Defaults to None.
        blue_border (bool | None, optional): _description_. Defaults to None.
        number_format (str | None, optional): _description_. Defaults to None.

    Returns:
        _type_: _description_
    """
    fmt = get_fmt_table(arr)

    if default_format:
        fmt = set_style_all(fmt, default_format)

    if number_format:
        fmt = set_style_all(fmt, {"number_format": {"format": number_format}})

    if blue_row_ids:
        fmt = set_style_row(fmt, blue_row_ids, STYLES["blue_white"])

    if light_blue_row_ids:
        fmt = set_style_row(fmt, light_blue_row_ids, STYLES["light_blue"])

    if light_blue_col_ids:
        fmt = set_style_col(fmt, light_blue_col_ids, STYLES["light_blue"], blue_row_ids)

    if left_align_ids:
        fmt = set_style_col(fmt, left_align_ids, STYLES["left_align"])

    if right_align_ids:
        fmt = set_style_col(fmt, right_align_ids, STYLES["right_align"])

    if perc_col_ids:
        if not perc_col_format:
            perc_col_format = "0.0%"
        fmt = set_style_col(
            fmt, perc_col_ids, {"number_format": {"format": perc_col_format}}
        )

    if float_col_ids:
        if not float_col_format:
            float_col_format = "0.00"
        fmt = set_style_col(
            fmt, float_col_ids, {"number_format": {"format": float_col_format}}
        )

    if blue_border_ids:
        fmt = set_style_row(fmt, blue_border_ids, STYLES["blue_border_bottom"])

    return fmt


def write_to_worksheet(
    ws: Any,
    arr: np.ndarray,
    fmt: Fmt,
    title: str | None = None,
    source: str | None = None,
    col_filter: bool | None = None,
    col_widths: list | None = None,
    min_column_width: int | None = None,
    cells_to_merge: list[list[int]] | None = None,
) -> None:
    """Writing data to worksheet. Used for writing values to cells and formatting the cells
    and

    Args:
        ws (Any): openpyxl worksheet
        arr (np.ndarray): array containing the input data
        fmt (Fmt): nested list containing dictionaries with the formatting info per cell
        title (str | None, optional): Title to be inserted above the table. Defaults to None.
        col_filter (bool | None, optional): Set column filter in excel. Defaults to None.
        autofit_columns (bool | None, optional): Automatically fit column width. Defaults to None.
    """
    for row_idx, row in enumerate(arr):
        for col_idx, _ in enumerate(row):
            value = arr[row_idx][col_idx]
            # Cell indices are not zero-indexed but one-indexed
            cell = ws.cell(row_idx + 1, col_idx + 1, value)
            # Get formatting for specific cell
            cell_fmt = fmt[row_idx][col_idx]
            for t, kwa in cell_fmt.items():
                # The api for setting different kind of formatting options is not
                # consistent therefore depeding on the formatting we want to set we have
                # to use a different strategy
                if t == "number_format":
                    # cell.number_format = "0.0"
                    setattr(cell, t, kwa["format"])
                elif t.startswith("border"):
                    # cell.border = Border(bottom=Side(color="00a0e6"))
                    type_, side = t.split("_")
                    side_spec = Side(**kwa)
                    setattr(cell, type_, LOOKUP[type_](**{side: side_spec}))
                else:
                    # cell.font = Font(color="B1D9F5", bold=True)
                    setattr(cell, t, LOOKUP[t](**kwa))

    if col_filter:
        filters = ws.auto_filter
        filters.ref = f"A1:{excel_style(len(fmt), len(fmt[0]))}"

    if source:
        _insert_source(ws, source, arr)

    if col_widths:
        _set_column_widths(ws, col_widths, min_column_width)

    if title:
        _insert_title(ws, title)

    if cells_to_merge:
        _merge_cells(ws, cells_to_merge, title)


def _set_column_widths(
    ws: Any, col_widths: list[int], min_column_width: int | None
) -> None:
    for idx, col_width in enumerate(col_widths):
        col_letter = get_column_letter(idx + 1)

        if min_column_width:
            if col_width < min_column_width:
                col_width = min_column_width

        ws.column_dimensions[col_letter].width = col_width


def _merge_cells(ws, cells_to_merge, title: str | None = None) -> None:
    add = 0
    if title:
        add = 1

    for row_idx, merge in cells_to_merge.items():
        row_idx = row_idx + add
        for start, stop in merge:
            cell = ws.cell(row_idx + 1, start)
            cell.alignment = Alignment(horizontal="center")
            ws.merge_cells(
                start_row=row_idx + 1,
                end_row=row_idx + 1,
                start_column=start,
                end_column=stop,
            )


def _insert_source(ws, source, arr):
    height, width = arr.shape
    cell = ws.cell(height + 1, width, source)
    cell.alignment = Alignment(horizontal="right")
    cell.font = Font(**STYLES["calibri"]["font"])


def _insert_title(ws: Any, title: str) -> None:
    ws.insert_rows(0)
    cell = ws.cell(1, 1, title)
    cell.alignment = Alignment(horizontal="left")
    for t, kwa in STYLES["title_bold"].items():
        setattr(cell, t, LOOKUP[t](**kwa))


def write_table(
    data: pd.DataFrame | dict[str, pd.DataFrame],
    file: str,
    header_row: int = 0,
    title: str | dict[str, str] | None = None,
    source: str | None = None,
    total_row: bool | None = None,
    light_blue_row_ids: int | list[int] | None = None,
    total_col: bool | None = None,
    right_align_ids: list | None = None,
    right_align_pattern: str | None = None,
    right_align_numeric: bool | None = True,
    left_align_ids: list | None = None,
    left_align_pattern: str | None = None,
    left_align_string: bool | None = True,
    perc_ids: list | None = None,
    perc_pattern: str | None = None,
    perc_col_format: str | None = None,
    float_ids: list | None = None,
    float_pattern: str | None = None,
    float_col_format: str | None = None,
    blue_border: bool | None = True,
    blue_border_row_ids: int | list[int] | None = None,
    number_format: str = "0.0",
    autofit_columns: str | None = None,
    column_widths: list[int] | None = None,
    min_column_width: int | None = None,
    col_filter: bool | None = False,
    style: str = "old",
    combine_multiindex: bool | int = False,
    column_names_to_string: bool = True,
):
    """_summary_

    Args:
        data (pd.DataFrame | dict[pd.DataFrame]): dataframe or dicts with dataframes
        file (str): destination file path
        header_row (int, optional): Id of header row. Defaults to 0.
        header_row (int): Set the number of rows to be dark blue (zero-indexed). Defaults to 0 (top row)
        title (str): Set the title above the table. In the case of multiple tables provide a dict in
        which te keys correspond to the sheet name. Defaults to none
        source (str | None, optional): Descriptiopn to be added underneath the table. Defaults to None.
        total_row (bool, optional): Color bottom row blue
        light_blue_row_ids (int | list[int] | None, optional): Row to be formatted light_blue. Defaults to None.
        total_col (bool, optional): Color last column blue.
        right_align_ids (list, optional): The ids of the columns to right align. Defaults to None
        right_align_pattern (str, optional): Pattern of columns to right align. Defaults to None.
        right_align_numeric (bool, optional): Right align numeric columns. Defaults to True.
        left_align_ids (list, optional): The ids of the columns to left align. Defaults to None.
        left_align_pattern (str, optional): Pattern of columns to left align. Defaults to None.
        left_align_string (bool, optional): Left align string columns. Defaults to True.
        perc_ids (list, optional): The ids of the columns to format as percentage. Defaults to None.
        perc_pattern (str, optional): The pattern of columns to format as percentage. Defaults to None.
        perc_col_format (str, optional): The formatting string of percentage columns. Defaults to None.
        blue_border (bool | None, optional): _description_. Defaults to True.
        blue_border_row_ids (int | list[int] | None, optional): _description_. Defaults to None.
        number_format (str, optional): _description_. Defaults to "0.0".
        autofit_columns (str | None, optional): _description_. Defaults to None.
        column_widths (list[int] | None, optional): _description_. Defaults to None.
        min_column_width (int | None, optional): _description_. Defaults to None.
        col_filter (bool, optional): Set filter on columns. Defaults to False.
        style (str, optional): _description_. Defaults to "old".
        combine_multiindex (bool | int, optional): _description_. Defaults to False.
        column_names_to_string (bool, optional): _description_. Defaults to True.
    """
    wb = Workbook()
    # Empty sheet is created on Workbook creation
    del wb["Sheet"]

    set_global_style(style)

    if not isinstance(data, dict):
        data = {"Sheet1": data}

    for sheet_name, df in data.items():
        format_worksheet(
            wb=wb,
            df=df,
            sheet_name=sheet_name,
            header_row=header_row,
            title=title,
            source=source,
            total_row=total_row,
            light_blue_row_ids=light_blue_row_ids,
            total_col=total_col,
            right_align_ids=right_align_ids,
            right_align_pattern=right_align_pattern,
            right_align_numeric=right_align_numeric,
            left_align_ids=left_align_ids,
            left_align_pattern=left_align_pattern,
            left_align_string=left_align_string,
            perc_ids=perc_ids,
            perc_pattern=perc_pattern,
            perc_col_format=perc_col_format,
            float_ids=float_ids,
            float_pattern=float_pattern,
            float_col_format=float_col_format,
            blue_border=blue_border,
            blue_border_row_ids=blue_border_row_ids,
            number_format=number_format,
            autofit_columns=autofit_columns,
            column_widths=column_widths,
            min_column_width=min_column_width,
            col_filter=col_filter,
            combine_multiindex=combine_multiindex,
            column_names_to_string=column_names_to_string,
        )

    wb.save(file)


def write_table_from_dict(
    file,
    write_info,
    style: str = "old",
):
    wb = Workbook()
    # Empty sheet is created on Workbook creation
    del wb["Sheet"]

    set_global_style(style)

    for sheet in write_info:
        format_worksheet(wb=wb, **sheet)

    wb.save(file)


def format_worksheet(
    wb: Any,
    df: pd.DataFrame,
    sheet_name: str,
    header_row: int = 0,
    title: str | dict[str, str] | None = None,
    source: str | None = None,
    total_row: bool | None = None,
    light_blue_row_ids: int | list[int] | None = None,
    total_col: bool | None = None,
    right_align_ids: list | None = None,
    right_align_pattern: str | None = None,
    right_align_numeric: bool | None = True,
    left_align_ids: list | None = None,
    left_align_pattern: str | None = None,
    left_align_string: bool | None = True,
    perc_ids: list | None = None,
    perc_pattern: str | None = None,
    perc_col_format: str | None = None,
    float_ids: list | None = None,
    float_pattern: str | None = None,
    float_col_format: str | None = None,
    blue_border: bool | None = True,
    blue_border_row_ids: int | list[int] | None = None,
    number_format: str = "0.0",
    autofit_columns: str | None = None,
    column_widths: list[int] | None = None,
    min_column_width: int | None = None,
    col_filter: bool | None = False,
    combine_multiindex: bool | int = False,
    column_names_to_string: bool = True,
):
    """_summary_

    Args:
        data (pd.DataFrame | dict[pd.DataFrame]): dataframe or dicts with dataframes
        name (str): name of excel file
        header_row (int): Set the number of rows to be dark blue (zero-indexed). Defaults to 0 (top row)
        title (str): Set the title above the table. In the case of multiple tables provide a dict in
        which te keys correspond to the sheet name. Defaults to none
        total_row (bool, optional): Color bottom row blue
        total_col (bool, optional): Color last column blue.
        right_align_ids (list, optional): The ids of the columns to right align. Defaults to None
        right_align_pattern (str, optional): Pattern of columns to right align. Defaults to None.
        right_align_numeric (bool, optional): Right align numeric columns. Defaults to True.
        left_align_ids (list, optional): The ids of the columns to left align. Defaults to None.
        left_align_pattern (str, optional): Pattern of columns to left align. Defaults to None.
        left_align_string (bool, optional): Left align string columns. Defaults to True.
        perc_ids (list, optional): The ids of the columns to format as percentage. Defaults to None.
        perc_pattern (str, optional): The pattern of columns to format as percentage. Defaults to None.
        perc_col_format (str, optional): The formatting string of percentage columns. Defaults to None.
        col_filter (bool, optional): Set filter on columns. Defaults to False.
    """
    if column_names_to_string:
        df = cols_to_str(df)

    arr = df_to_array(df)

    blue_rows = []
    light_blue_rows = []
    light_blue_cols = []
    blue_border_ids = []
    r_align_ids = []
    l_align_ids = []
    p_ids = []
    f_ids = []
    cells_to_merge = []
    title_tbl = None
    title_src = None

    if isinstance(header_row, int):
        blue_rows.extend(list(range(0, header_row + 1)))

    if title:
        if isinstance(title, str):
            title_tbl = title
        elif isinstance(title, dict):
            title_tbl = title.get(sheet_name)

    if source:
        if isinstance(source, str):
            title_src = source
        elif isinstance(title, dict):
            title_src = source.get(sheet_name)

    if right_align_ids:
        r_align_ids.extend(right_align_ids)

    if right_align_pattern:
        r_align_ids.extend(get_cols_id_with_pattern(df, right_align_pattern))

    if right_align_numeric:
        r_align_ids.extend(get_numeric_col_ids(df))

    if left_align_ids:
        r_align_ids.extend(left_align_ids)

    if left_align_pattern:
        l_align_ids.extend(get_cols_id_with_pattern(df, left_align_pattern))

    if left_align_string:
        l_align_ids.extend(get_string_cols_ids(df))

    if perc_ids:
        p_ids.extend(perc_ids)

    if perc_pattern:
        pr_id = get_cols_id_with_pattern(df, perc_pattern)
        p_ids.extend(pr_id)
        r_align_ids.extend(pr_id)

    if float_ids:
        f_ids.extend(float_ids)

    if float_pattern:
        fr_id = get_cols_id_with_pattern(df, float_pattern)
        f_ids.extend(fr_id)
        r_align_ids.extend(fr_id)

    if total_row:
        light_blue_rows.append(arr.shape[0] - 1)

    if light_blue_row_ids:
        light_blue_rows.extend(light_blue_row_ids)

    if total_col:
        light_blue_cols.append(arr.shape[1] - 1)

    if blue_border:
        blue_border_ids.append(arr.shape[0] - 1)

    if blue_border_row_ids:
        blue_border_ids.extend(blue_border_row_ids)

    if combine_multiindex:
        cells_to_merge = get_cells_to_merge(df)

    if column_widths:
        if len(arr[0]) != len(column_widths):
            raise Warning(
                "The number of widths defined in column_widths should be equal to the number of columsn in the dataframe"
            )
        col_widths = column_widths

    elif (autofit_columns == "column_names") or (autofit_columns is None):
        col_widths = get_max_col_widths(df)
    elif autofit_columns == "all_data":
        col_widths = get_max_col_widths(arr)
    else:
        col_widths = None

    ws = wb.create_sheet(sheet_name)

    fmt = cell_formatting(
        arr=arr,
        default_format=STYLES["calibri"],
        blue_row_ids=blue_rows,
        light_blue_row_ids=light_blue_rows,
        light_blue_col_ids=light_blue_cols,
        left_align_ids=l_align_ids,
        right_align_ids=r_align_ids,
        perc_col_ids=p_ids,
        perc_col_format=perc_col_format,
        float_col_ids=f_ids,
        float_col_format=float_col_format,
        number_format=number_format,
        blue_border_ids=blue_border_ids,
    )

    write_to_worksheet(
        ws=ws,
        arr=arr,
        fmt=fmt,
        title=title_tbl,
        source=title_src,
        col_filter=col_filter,
        col_widths=col_widths,
        cells_to_merge=cells_to_merge,
        min_column_width=min_column_width,
    )
