#!/usr/bin/env python3
"""
MarkShark
score_core.py  —  Axis-based MarkShark grading engine

Features:
 - Multi-version exam support (NEW)
 - CSV includes: correct, incorrect, blank, multi, percent
 - KEY row(s) written below header (when key(s) provided)
 - Annotated PNGs:
     * Names/ID: blue circles, optional white % text
     * Answers: green=correct, red=incorrect, grey=blank, orange=multi
 - Optional % fill text via --label-density
 - Columns limited to len(key) when a key is provided
"""

from __future__ import annotations
import os
import csv
import sys
from typing import Optional, List, Tuple, Dict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

import numpy as np
import cv2
from .defaults import (
    ANNOTATION_DEFAULTS,
    AnnotationDefaults,
    SCORING_DEFAULTS,
    RENDER_DEFAULTS,
    resolve_scored_pdf_path,
)

from .tools.bubblemap_io import load_bublmap, Bubblemap, GridLayout, PageLayout
from .tools import io_pages as IO
from .tools.score_tools import (
    process_page_all,
    load_key_txt,
    load_multi_version_keys,
    score_against_multi_keys,
    grid_centers_axis_mode,
    centers_to_circle_rois,
    roi_fill_scores,
    calibrate_fixed_thresh_for_page,
)


# ----------------------------
# Basic Stats Functions (for inline stats during scoring)
# ----------------------------

def _compute_basic_stats(
    all_student_data: List[Dict],  # List of {answers, correct, total, version}
    keys_dict: Optional[Dict[str, List[str]]],
    q_out: int,
) -> Dict:
    """
    Compute basic exam and item statistics from scoring results.
    Handles multi-version exams properly by computing per-version item stats.
    
    Args:
        all_student_data: List of dicts with keys: answers, correct, total, version
        keys_dict: Dict mapping version -> answer key list
        q_out: Number of questions to analyze
    
    Returns dict with:
        - n_students: total number of students
        - mean_score: mean number correct (pooled)
        - mean_percent: mean percentage correct (pooled)
        - std_dev: standard deviation of scores (pooled)
        - high_score: highest score (pooled)
        - low_score: lowest score (pooled)
        - kr20_overall: KR-20 computed across all students (pooled)
        - versions: dict of per-version stats:
            {version: {n, mean, kr20, per_item_pct, per_item_pb}}
    """
    n_students = len(all_student_data)
    
    empty_result = {
        "n_students": 0,
        "mean_score": 0,
        "mean_percent": 0,
        "std_dev": 0,
        "high_score": 0,
        "low_score": 0,
        "kr20_overall": float("nan"),
        "versions": {},
    }
    
    if n_students == 0:
        return empty_result
    
    # Extract scores for pooled stats
    scores = [d["correct"] for d in all_student_data if d["total"] > 0]
    totals = [d["total"] for d in all_student_data if d["total"] > 0]
    
    if not scores:
        return empty_result
    
    # Pooled exam-level stats
    mean_score = np.mean(scores)
    mean_total = np.mean(totals) if totals else q_out
    mean_percent = (mean_score / mean_total * 100) if mean_total > 0 else 0
    std_dev = np.std(scores, ddof=1) if len(scores) > 1 else 0
    high_score = max(scores)
    low_score = min(scores)
    
    # Group students by version
    students_by_version: Dict[str, List[Dict]] = {}
    for d in all_student_data:
        ver = d.get("version", "").rstrip("*") or "A"  # Default to A if no version
        if ver not in students_by_version:
            students_by_version[ver] = []
        students_by_version[ver].append(d)
    
    # Determine if this is a multi-version exam
    is_multi_version = len(students_by_version) > 1 and keys_dict and len(keys_dict) > 1
    
    # Compute per-version stats
    version_stats = {}
    all_correctness_matrices = []  # For pooled KR-20
    
    for ver, students in sorted(students_by_version.items()):
        ver_n = len(students)
        ver_scores = [d["correct"] for d in students if d["total"] > 0]
        
        if not ver_scores:
            continue
        
        ver_mean = np.mean(ver_scores)
        
        # Get the key for this version
        key = None
        if keys_dict:
            key = keys_dict.get(ver, keys_dict.get(list(keys_dict.keys())[0]))
            key = key[:q_out] if key else None
        
        # Build correctness matrix for this version
        correctness = np.zeros((ver_n, q_out))
        for i, d in enumerate(students):
            answers = d.get("answers", [])
            for j, ans in enumerate(answers[:q_out]):
                if key and j < len(key):
                    if ans and ans == key[j]:
                        correctness[i, j] = 1
        
        all_correctness_matrices.append(correctness)
        
        # Per-item stats for this version
        per_item_pct = (correctness.mean(axis=0) * 100).tolist() if ver_n > 0 else []
        
        # Point-biserial for this version
        total_scores = correctness.sum(axis=1)
        per_item_pb = []
        for j in range(q_out):
            item_col = correctness[:, j]
            total_minus_item = total_scores - item_col
            pb = _point_biserial(item_col, total_minus_item)
            per_item_pb.append(pb)
        
        # KR-20 for this version
        ver_kr20 = _kr20(correctness, total_scores)
        
        version_stats[ver] = {
            "n": ver_n,
            "mean": round(ver_mean, 2),
            "kr20": round(ver_kr20, 3) if not np.isnan(ver_kr20) else float("nan"),
            "per_item_pct": [round(p, 1) for p in per_item_pct],
            "per_item_pb": [round(p, 3) if not np.isnan(p) else float("nan") for p in per_item_pb],
        }
    
    # Compute overall KR-20 (pooled across all versions)
    # This treats each student's correctness pattern as a row, regardless of version
    kr20_overall = float("nan")
    if all_correctness_matrices:
        pooled_correctness = np.vstack(all_correctness_matrices)
        pooled_totals = pooled_correctness.sum(axis=1)
        kr20_overall = _kr20(pooled_correctness, pooled_totals)
    
    return {
        "n_students": n_students,
        "mean_score": round(mean_score, 2),
        "mean_percent": round(mean_percent, 1),
        "std_dev": round(std_dev, 2),
        "high_score": high_score,
        "low_score": low_score,
        "kr20_overall": round(kr20_overall, 3) if not np.isnan(kr20_overall) else float("nan"),
        "versions": version_stats,
        "is_multi_version": is_multi_version,
    }


def _point_biserial(item: np.ndarray, total_minus_item: np.ndarray) -> float:
    """Compute point-biserial correlation for an item."""
    p = item.mean()
    q = 1 - p
    if p == 0 or p == 1:
        return float("nan")
    
    mask_1 = item == 1
    mask_0 = item == 0
    
    if mask_1.sum() == 0 or mask_0.sum() == 0:
        return float("nan")
    
    M1 = total_minus_item[mask_1].mean()
    M0 = total_minus_item[mask_0].mean()
    s = total_minus_item.std(ddof=1)
    
    if s == 0 or np.isnan(s):
        return float("nan")
    
    return float(((M1 - M0) / s) * np.sqrt(p * q))


def _kr20(correctness: np.ndarray, total_scores: np.ndarray) -> float:
    """Compute KR-20 reliability coefficient."""
    k = correctness.shape[1]
    if k < 2:
        return float("nan")
    
    p = correctness.mean(axis=0)
    pq_sum = (p * (1 - p)).sum()
    var_total = np.var(total_scores, ddof=1)
    
    if var_total <= 0 or np.isnan(var_total):
        return float("nan")
    
    return float((k / (k - 1.0)) * (1.0 - (pq_sum / var_total)))


# ----------------------------
# Utilities
# ----------------------------

def _ensure_dir(path: str) -> None:
    if path and not os.path.isdir(path):
        os.makedirs(path, exist_ok=True)


def _rowwise_scores(
    gray: np.ndarray,
    rois: List[Tuple[int, int, int, int]],
    rows: int,
    cols: int,
    fixed_thresh: Optional[int] = None,
) -> List[List[float]]:
    """Return a rows x cols matrix of fill scores (0-1)."""
    flat = roi_fill_scores(
        gray,
        rois,
        inner_radius_ratio=0.70,
        blur_ksize=5,
        fixed_thresh=fixed_thresh,
    )
    return [flat[r * cols:(r + 1) * cols] for r in range(rows)]

# ----------------------------
# Multi-page support helper
# ----------------------------

def _process_single_page(
    img_bgr: np.ndarray,
    page_layout: 'PageLayout',
    *,
    min_fill: float,
    top2_ratio: float,
    min_top2_diff: float,
    fixed_thresh: Optional[int] = None,
) -> Tuple[dict, List[Optional[str]]]:
    """
    Process a single page using a specific PageLayout from a multi-page bubblemap.
    Returns (info_dict, answers_list) for this page only.
    """
    from .tools.score_tools import decode_layout, indices_to_text_col
    
    gray = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)
    
    info = {"last_name": "", "first_name": "", "student_id": "", "version": ""}
    
    # Decode name/ID from this page (if layouts are present)
    if page_layout.last_name_layout:
        picked, _, _ = decode_layout(
            gray,
            page_layout.last_name_layout,
            min_fill=min_fill,
            top2_ratio=top2_ratio,
            min_top2_diff=min_top2_diff,
            fixed_thresh=fixed_thresh,
        )
        info["last_name"] = indices_to_text_col(
            picked, page_layout.last_name_layout.labels or " ABCDEFGHIJKLMNOPQRSTUVWXYZ"
        ).strip()
    
    if page_layout.first_name_layout:
        picked, _, _ = decode_layout(
            gray,
            page_layout.first_name_layout,
            min_fill=min_fill,
            top2_ratio=top2_ratio,
            min_top2_diff=min_top2_diff,
            fixed_thresh=fixed_thresh,
        )
        info["first_name"] = indices_to_text_col(
            picked, page_layout.first_name_layout.labels or " ABCDEFGHIJKLMNOPQRSTUVWXYZ"
        ).strip()
    
    if page_layout.id_layout:
        picked, _, _ = decode_layout(
            gray,
            page_layout.id_layout,
            min_fill=min_fill,
            top2_ratio=top2_ratio,
            min_top2_diff=min_top2_diff,
            fixed_thresh=fixed_thresh,
        )
        info["student_id"] = indices_to_text_col(
            picked, page_layout.id_layout.labels or "0123456789"
        ).strip()
    
    if page_layout.version_layout:
        # Version detection (usually only on page 1)
        picked, _, _ = decode_layout(
            gray,
            page_layout.version_layout,
            min_fill=min_fill,
            top2_ratio=top2_ratio,
            min_top2_diff=min_top2_diff,
            fixed_thresh=fixed_thresh,
        )
        if picked and len(picked) > 0 and picked[0] is not None:
            labels = page_layout.version_layout.labels or "ABCD"
            info["version"] = labels[picked[0]] if picked[0] < len(labels) else ""
    
    # Decode answers from this page
    answers: List[Optional[str]] = []
    for layout in page_layout.answer_layouts:
        picked, _, scores = decode_layout(
            gray,
            layout,
            min_fill=min_fill,
            top2_ratio=top2_ratio,
            min_top2_diff=min_top2_diff,
            fixed_thresh=fixed_thresh,
        )
        choice_labels = list(layout.labels) if layout.labels else [chr(ord("A") + k) for k in range(layout.numcols)]
        if layout.selection_axis == "row":
            from .tools.score_tools import scores_to_labels_row
            answers.extend(
                scores_to_labels_row(
                    scores,
                    layout.numrows,
                    layout.numcols,
                    choice_labels,
                    min_fill=min_fill,
                    top2_ratio=top2_ratio,
                    min_top2_diff=min_top2_diff,
                )
            )
        else:
            from .tools.score_tools import indices_to_labels_row
            answers.extend(indices_to_labels_row(picked, layout.numcols, choice_labels))
    
    return info, answers


# ----------------------------
# Annotation helpers
# ----------------------------

def _annotate_names_ids(
    img_bgr: np.ndarray,
    bmap: Bubblemap,
    label_density: bool,
    color_zone=None,
    text_color=None,
    thickness: Optional[int] = None,
    font_scale: Optional[float] = None,
    label_thickness: Optional[int] = None,
    annotation_defaults: Optional[AnnotationDefaults] = None,
) -> np.ndarray:
    """
    Draw blue circles for Last/First Name and Student ID grids.
    If label_density=True, write white % fill text in each bubble.
    Returns a new image with drawings (does not modify input in place).
    """
    out = img_bgr.copy()
    gray = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)
    H, W = gray.shape[:2]

    ad = annotation_defaults or ANNOTATION_DEFAULTS
    if color_zone is None:
        color_zone = ad.color_zone
    if text_color is None:
        text_color = ad.percent_text_color
    if thickness is None:
        thickness = ad.thickness_names
    if font_scale is None:
        font_scale = ad.label_font_scale
    if label_thickness is None:
        label_thickness = ad.label_thickness

    def draw_layout(layout: GridLayout) -> None:
        centers = grid_centers_axis_mode(
            layout.x_topleft, layout.y_topleft,
            layout.x_bottomright, layout.y_bottomright,
            layout.numrows, layout.numcols
        )
        rois = centers_to_circle_rois(centers, W, H, layout.radius_pct)
        scores = roi_fill_scores(gray, rois, inner_radius_ratio=0.70, blur_ksize=5)
        for idx, (x, y, w, h) in enumerate(rois):
            cx, cy = x + w // 2, y + h // 2
            radius = min(w, h) // 2
            cv2.circle(out, (cx, cy), radius, color_zone, thickness, lineType=cv2.LINE_AA)
            if label_density:
                pct = int(round(100 * scores[idx]))
                cv2.putText(out, f"{pct}", (cx - 8, cy + 5),
                            cv2.FONT_HERSHEY_SIMPLEX, float(font_scale), text_color, int(label_thickness), cv2.LINE_AA)

    for attr in ("last_name_layout", "first_name_layout", "id_layout"):
        lay = getattr(bmap, attr, None)
        if isinstance(lay, GridLayout):
            draw_layout(lay)

    return out


def _annotate_answers(
img_bgr: np.ndarray,
bmap: Bubblemap,
key_letters: Optional[List[str]],
label_density: bool,
annotate_all_cells: bool,
min_fill: float,
top2_ratio: float,
min_top2_diff: float,
answers_for_annotation: Optional[List[Optional[str]]] = None,
fixed_thresh: Optional[int] = None,
calibrate_background: bool = False,
background_percentile: float = SCORING_DEFAULTS.background_percentile,
color_correct=None,
color_incorrect=None,
color_blank=None,
color_blank_answer_row=None,
color_multi=None,
thickness: Optional[int] = None,
pct_fill_font_color=None,
pct_fill_font_scale=None,
pct_fill_font_thickness=None,
pct_fill_font_position=None,
annotation_defaults: Optional[AnnotationDefaults] = None,
box_multi: Optional[bool] = None,
box_blank_answer_row: Optional[bool] = None,
box_color_multi=None,
box_color_blank_answer_row=None,
box_thickness: Optional[int] = None,
box_pad: Optional[int] = None,
box_top_extra: Optional[int] = None,
) -> np.ndarray:
    """
    Draw per-bubble overlays for answer blocks:
      - green circle for correct,
      - red for incorrect,
      - grey for blank,
      - orange for multi.
    Optionally put % fill text in each bubble (label_density=True).
    Returns a new image with drawings (does not modify input in place).
    """
    out = img_bgr.copy()
    H, W = out.shape[:2]
    key_seq = [k.upper() for k in key_letters] if key_letters else None

    # Resolve annotation defaults
    ad = annotation_defaults or ANNOTATION_DEFAULTS
    if color_correct is None:
        color_correct = ad.color_correct
    if color_incorrect is None:
        color_incorrect = ad.color_incorrect
    if color_blank is None:
        color_blank = ad.color_blank
    if color_blank_answer_row is None:
        color_blank_answer_row = getattr(ad, "color_blank_answer_row", (255, 0, 255))
    if color_multi is None:
        color_multi = ad.color_multi
    if thickness is None:
        thickness = ad.thickness_answers

    # Resolve annotation defaults
    if pct_fill_font_color is None:
        pct_fill_font_color  = getattr(ad, "pct_fill_font_color", (255, 0, 255))
    if pct_fill_font_scale is None:
        pct_fill_font_scale = getattr(ad, "pct_fill_font_scale", 0.5)
    if pct_fill_font_thickness is None:
        pct_fill_font_thickness = getattr(ad, "pct_fill_font_thickness", 1)
    if pct_fill_font_position is None:
        pct_fill_font_position = getattr(ad, "pct_fill_font_position", 5)

    # Row box defaults
    if box_multi is None:
        box_multi = getattr(ad, "box_multi", False)
    if box_blank_answer_row is None:
        box_blank_answer_row = getattr(ad, "box_blank_answer_row", False)
    if box_color_multi is None:
        box_color_multi = getattr(ad, "box_color_multi", color_multi)
    if box_color_blank_answer_row is None:
        box_color_blank_answer_row = getattr(ad, "box_color_blank_answer_row", color_blank_answer_row)
    if box_thickness is None:
        box_thickness = getattr(ad, "box_thickness", thickness)
    if box_pad is None:
        box_pad = getattr(ad, "box_pad", 6)
    if box_top_extra is None:
        box_top_extra = getattr(ad, "box_top_extra", 0)

    q_global = 0
    gray = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)

    for layout in bmap.answer_layouts:
        centers = grid_centers_axis_mode(
            layout.x_topleft, layout.y_topleft,
            layout.x_bottomright, layout.y_bottomright,
            layout.numrows, layout.numcols
        )
        rois = centers_to_circle_rois(centers, W, H, layout.radius_pct)
        M = _rowwise_scores(gray, rois, layout.numrows, layout.numcols, fixed_thresh=fixed_thresh,)

        # Apply background calibration if enabled
        if calibrate_background and layout.numrows > 1:
            from .tools.score_tools import calibrate_column_backgrounds, subtract_column_backgrounds
            flat_scores = [score for row in M for score in row]  # Flatten to list
            backgrounds = calibrate_column_backgrounds(flat_scores, layout.numrows, layout.numcols, background_percentile)
            calibrated_flat = subtract_column_backgrounds(flat_scores, layout.numrows, layout.numcols, backgrounds)
            # Reshape back to rows x cols
            M = [calibrated_flat[r * layout.numcols:(r + 1) * layout.numcols] for r in range(layout.numrows)]

        choice_labels = list(layout.labels) if layout.labels else [chr(ord("A") + i) for i in range(layout.numcols)]
        label_to_idx = {str(lab).strip().upper(): i for i, lab in enumerate(choice_labels)}

        for r in range(layout.numrows):
            # Skip annotation for questions beyond the answer key length
            if answers_for_annotation is not None and q_global >= len(answers_for_annotation):
                q_global += 1
                continue
            
            row_scores = M[r]
            order = np.argsort(row_scores)[::-1]
            best = int(order[0])
            best_val = float(row_scores[best])
            second_val = float(row_scores[order[1]]) if layout.numcols > 1 else 0.0
            is_blank = best_val < min_fill
            is_multi = (not is_blank) and (layout.numcols > 1) and (second_val > top2_ratio * best_val)

            # If available, drive annotation from the same scoring decisions used for the CSV.
            # This prevents borderline cases being labeled blank in the CSV but shown as selected in the PDF overlay.
            selected_labels = []  # type: List[str]
            selected_idx = set()

            if answers_for_annotation is not None and q_global < len(answers_for_annotation):
                ans = answers_for_annotation[q_global]
                if ans is None or ans == "":
                    is_blank, is_multi = True, False
                else:
                    ans_str = str(ans).strip().upper()
                    if "," in ans_str:
                        is_blank, is_multi = False, True
                        selected_labels = [p.strip() for p in ans_str.split(",") if p.strip()]
                    else:
                        is_blank, is_multi = False, False
                        selected_labels = [ans_str]

                for lab in selected_labels:
                    if lab in label_to_idx:
                        selected_idx.add(label_to_idx[lab])

            # Fallback: if we did not map any selected labels, use local score winners.
            if not selected_idx:
                if is_multi and layout.numcols > 1:
                    selected_idx = {best, int(order[1])}
                elif not is_blank:
                    selected_idx = {best}


            key_char = key_seq[q_global] if key_seq and q_global < len(key_seq) else None
            answer_row_blank = bool(is_blank and key_char and (key_char in choice_labels))

            # Optional row-level boxes, draw these before circles and text
            if (is_multi and box_multi) or (answer_row_blank and box_blank_answer_row):
                row_rois = rois[r * layout.numcols:(r + 1) * layout.numcols]
                x0 = min(x for x, y, w, h in row_rois)
                y0 = min(y for x, y, w, h in row_rois)
                x1 = max(x + w for x, y, w, h in row_rois)
                y1 = max(y + h for x, y, w, h in row_rois)

                x0 = max(0, x0 - int(box_pad))
                y0 = max(0, y0 - int(box_pad) - int(box_top_extra))
                x1 = min(W - 1, x1 + int(box_pad))
                y1 = min(H - 1, y1 + int(box_pad))

                if is_multi and box_multi:
                    cv2.rectangle(out, (x0, y0), (x1, y1), box_color_multi, int(box_thickness), lineType=cv2.LINE_AA)
                if answer_row_blank and box_blank_answer_row:
                    cv2.rectangle(out, (x0, y0), (x1, y1), box_color_blank_answer_row, int(box_thickness), lineType=cv2.LINE_AA)


            for c in range(layout.numcols):
                x, y, w, h = rois[r * layout.numcols + c]
                cx, cy = x + w // 2, y + h // 2
                radius = min(w, h) // 2

                draw_this = annotate_all_cells or (c in selected_idx) or is_blank or is_multi
                if not draw_this:
                    continue

                if is_blank:
                    col = (color_blank_answer_row if answer_row_blank else color_blank)
                elif is_multi:
                    col = color_multi
                else:
                    if key_char:
                        is_selected = c in selected_idx

                        selected_single = None
                        if selected_labels and len(selected_labels) == 1:
                            selected_single = selected_labels[0]
                        elif (not selected_labels) and (not is_blank) and (not is_multi) and len(selected_idx) == 1:
                            only_idx = next(iter(selected_idx))
                            selected_single = str(choice_labels[only_idx]).strip().upper()

                        col = color_correct if (is_selected and selected_single == key_char) else (
                            color_incorrect if is_selected else (200, 200, 200)
                        )
                    else:
                        col = (0, 200, 200) if (c in selected_idx) else (200, 200, 200)

                cv2.circle(out, (cx, cy), radius, col, thickness, lineType=cv2.LINE_AA)

                if label_density:
                    pct = int(round(100 * row_scores[c]))
                
                    # Put text above the circle
                    text = f"{pct}"
                    font = cv2.FONT_HERSHEY_SIMPLEX
                
                    (tw, th), baseline = cv2.getTextSize(text, font, pct_fill_font_scale, 
                    pct_fill_font_thickness)
                
                    # this formula sets text baseline position above or below circle
                    # pct_fill_font_position values correspond to pixels.
                    # these values should be set in the defaults.py script
                    # a value of zero puts the text right on top of the circle
                    # negative values pull the text into the circle.
                    # positive values push the text farther above the circle
                    tx = cx - tw // 2
                    ty = cy - radius - pct_fill_font_position  
                
                    # Keep text inside the image bounds
                    tx = max(0, min(tx, W - tw - 1))
                    ty = max(th + 1, ty)  # avoid going off the top
                
                    cv2.putText(out, text, (tx, ty),
                                font, pct_fill_font_scale, pct_fill_font_color,
                                pct_fill_font_thickness, cv2.LINE_AA)

            q_global += 1

    return out


# ----------------------------
# Main grading entry point
# ----------------------------

def score_pdf(
    input_path: str,
    bublmap_path: str,
    out_csv: str,
    min_fill: float,
    top2_ratio: float,
    min_top2_diff: float,
    fixed_thresh: Optional[int] = None,
    key_txt: Optional[str] = None,
    out_annotated_dir: Optional[str] = None,
    out_pdf: Optional[str] = None,
    dpi: int = RENDER_DEFAULTS.dpi,
    pdf_quality: int = RENDER_DEFAULTS.pdf_quality,
    annotate_all_cells: bool = False,
    label_density: bool = False,
    pdf_renderer: str = "auto",
    auto_calibrate_thresh: bool = True,
    verbose_calibration: bool = False,
    # Background calibration
    calibrate_background: bool = SCORING_DEFAULTS.calibrate_background,
    background_percentile: float = SCORING_DEFAULTS.background_percentile,
    # Adaptive rescoring
    adaptive_rescoring: bool = SCORING_DEFAULTS.adaptive_rescoring,
    adaptive_max_adjustment: int = SCORING_DEFAULTS.adaptive_max_adjustment,
    adaptive_min_above_floor: float = SCORING_DEFAULTS.adaptive_min_above_floor,
    # NEW: Review/flagging options
    review_pdf: Optional[str] = None,  # Output PDF containing only flagged pages
    flagged_xlsx: Optional[str] = None,  # Output XLSX listing flagged items for review
    low_confidence_threshold: float = 0.15,  # Flag answers with separation below this
    # NEW: Inline stats option
    include_stats: bool = True,  # Append summary stats rows to CSV
) -> str:
    
    """
    Grade a PDF or image stack using axis-based geometry.
    
    NEW: Multi-page bubble sheet support!
    NEW: Multi-version exam support!
    NEW: Flagging and review PDF support!
    NEW: Inline basic statistics!

    Behavior:
      - Supports single-page or multi-page bubble sheets
      - For multi-page: processes pages in groups (page 1+2 = student 1, page 3+4 = student 2, etc.)
      - Name/ID taken from page 1, answers combined from all pages
      - Detects version from version_layout bubbles
      - Loads multi-version keys if key file contains #Version headers
      - Scores each student against their version's key
      - If key is provided: limit output columns and scoring to first len(key) questions.
      - CSV includes metrics: correct, incorrect, blank, multi, percent.
      - KEY row(s) written under header (one per version).
      - Annotated images combine blue name/ID overlays and answer overlays.
      - Page column shows "1-2" for 2-page sheets, "1" for single-page
      
    Flagging (NEW):
      - Tracks answers that are blank, multi, or low-confidence
      - If review_pdf is provided: generates a PDF with just the flagged pages
      - If flagged_xlsx is provided: writes XLSX with flagged items + Corrected_Answer column
      - Flagged pages get special annotations highlighting problem areas
      
    Inline Stats (NEW):
      - If include_stats=True and key is provided: appends summary rows to CSV
      - Exam stats: N, Mean, StdDev, High, Low, KR-20
      - Item stats: % Correct, Point-Biserial for each question
    """
    bmap: Bubblemap = load_bublmap(bublmap_path)
    pages = IO.load_pages(input_path, dpi=dpi, renderer=pdf_renderer)
    
    # Try loading multi-version keys, fall back to single version
    keys_dict: Optional[Dict[str, List[str]]] = None
    single_key: Optional[List[str]] = None
    
    if key_txt:
        try:
            keys_dict = load_multi_version_keys(key_txt)
        except Exception:
            # Fall back to single-version key
            single_key = load_key_txt(key_txt)
            if single_key:
                keys_dict = {"A": single_key}  # Default to version A

    # Calculate total questions across all pages, and per-page question counts
    questions_per_page = []
    for page in bmap.pages:
        page_q = sum(layout.numrows for layout in page.answer_layouts)
        questions_per_page.append(page_q)
    total_q = sum(questions_per_page)

    if keys_dict:
        # Use length of first version key
        first_version = sorted(keys_dict.keys())[0]
        q_out = len(keys_dict[first_version])
    else:
        q_out = total_q

    q_out = max(0, min(q_out, total_q))

    # Calculate how many active questions are on each page (for calibration)
    # e.g. 57 questions on a 2-page template with 32+32 rows -> page 1 uses 32, page 2 uses 25
    active_rows_per_page = []
    remaining = q_out
    for page_q in questions_per_page:
        active = min(remaining, page_q)
        active_rows_per_page.append(active)
        remaining -= active

    # Make the CSV header
    header = ["Version", "Page", "LastName", "FirstName", "StudentID"]

    if keys_dict:
        header += ["Correct", "Incorrect", "Blank", "Multi", "Percent"]
    else:
        header += ["Blank"]
        
    header += [f"Q{i+1}" for i in range(q_out)]

    _ensure_dir(os.path.dirname(out_csv) or ".")
    if out_annotated_dir:
        _ensure_dir(out_annotated_dir)


    # If requested, write an annotated PDF (default) and optionally a PNG directory.
    out_pdf_path: Optional[str] = None
    pdf_writer = None
    annotated_pages: List[np.ndarray] = []  # fallback only (used if streaming writer is unavailable)


    # out_pdf None means use the default name from defaults.py
    if out_pdf is None:
        out_pdf = SCORING_DEFAULTS.out_pdf

    # out_pdf "" (empty string) means disable PDF output
    if out_pdf:
        out_pdf_path = resolve_scored_pdf_path(out_pdf, out_csv=out_csv, out_pdf_dir=SCORING_DEFAULTS.out_pdf_dir)
        _ensure_dir(os.path.dirname(out_pdf_path) or ".")
        # Prefer a streaming PDF writer (small PDFs via JPEG embedding). Falls back to PIL at end if needed.
        try:
            pdf_writer = IO.PdfPageWriter(out_pdf_path, dpi=dpi)
        except Exception:
            pdf_writer = None
    
    # ==================== FLAGGING INFRASTRUCTURE ====================
    # Track flagged items for review PDF generation
    flagged_items: List[Dict] = []  # List of {student_id, page, question, issue, answer, fill_pct, ...}
    flagged_page_images: List[Tuple[int, np.ndarray]] = []  # (page_idx, annotated_image) for review PDF
    
    # Helper to check if a student has flagged answers
    def _has_flags(answers: List[Optional[str]]) -> bool:
        for a in answers:
            if a is None or a == "":  # blank
                return True
            if isinstance(a, str) and "," in a:  # multi
                return True
        return False
    
    # Helper to record flagged items for a student
    def _record_flags(
        student_id: str,
        page_num: int,
        answers: List[Optional[str]],
        student_name: str = "",
    ):
        for q_idx, ans in enumerate(answers):
            issue = None
            if ans is None or ans == "":
                issue = "blank"
            elif isinstance(ans, str) and "," in ans:
                issue = "multi"
            
            if issue:
                flagged_items.append({
                    "student_id": student_id,
                    "student_name": student_name,
                    "page": page_num,
                    "question": q_idx + 1,
                    "issue": issue,
                    "current_answer": ans if ans else "",
                })
    # ==================================================================
    
    # ==================== STATS COLLECTION ============================
    # Collect data for computing basic statistics at the end
    # Now includes version info for multi-version support
    # Also stores full CSV row for grouped output by version
    all_student_data: List[Dict] = []  # {answers, correct, total, version, csv_row}
    # ==================================================================
    
    # Store header for later use when writing grouped output
    csv_header = header
            
    with open(out_csv, "w", newline="", encoding="utf-8") as f:
        # We no longer write header/keys here - they'll be written grouped by version at the end
        # Just process all students and collect their data

        # Determine if multi-page mode
        pages_per_student = bmap.num_pages
        is_multipage = pages_per_student > 1
        
        if is_multipage:
            # MULTI-PAGE MODE: Process pages in groups
            num_students = len(pages) // pages_per_student
            
            # Check for incomplete students (odd number of pages)
            if len(pages) % pages_per_student != 0:
                raise ValueError(
                    f"ERROR: Input PDF has {len(pages)} pages, but template expects {pages_per_student} pages per student. "
                    f"Expected a multiple of {pages_per_student} pages (got {len(pages) % pages_per_student} extra pages). "
                    f"Please check your scans - each student should have exactly {pages_per_student} pages."
                )
            
            for student_idx in range(num_students):
                # Get all pages for this student
                start_page_idx = student_idx * pages_per_student
                student_pages = pages[start_page_idx:start_page_idx + pages_per_student]
                
                # Process each page and collect data
                all_answers = []
                student_info = {}
                page_thresholds = []
                
                for page_num, img_bgr in enumerate(student_pages, start=1):
                    page_layout = bmap.get_page(page_num)

                    # Per-page calibration
                    page_fixed_thresh = fixed_thresh
                    if fixed_thresh is None and auto_calibrate_thresh:
                        # Use page 1 layout for calibration reference
                        calib_page = bmap.get_page(1)
                        if calib_page:
                            # Create temp bmap-like object for calibration
                            temp_bmap = type('TempBmap', (object,), {
                                'answer_layouts': calib_page.answer_layouts,
                                'last_name_layout': calib_page.last_name_layout,
                                'first_name_layout': calib_page.first_name_layout,
                                'id_layout': calib_page.id_layout,
                            })()
                            # Only calibrate on active rows (skip unused questions)
                            calib_max_rows = active_rows_per_page[page_num - 1] if page_num <= len(active_rows_per_page) else None
                            page_fixed_thresh, _calib_stats = calibrate_fixed_thresh_for_page(
                                img_bgr,
                                temp_bmap,
                                max_rows=calib_max_rows,
                                verbose=verbose_calibration,
                            )
                    
                    page_thresholds.append(page_fixed_thresh)
                    
                    # Process this page
                    info, answers = _process_single_page(
                        img_bgr, page_layout,
                        min_fill=min_fill,
                        top2_ratio=top2_ratio,
                        min_top2_diff=min_top2_diff,
                        fixed_thresh=page_fixed_thresh,
                    )
                    
                    # Take student info from page 1
                    if page_num == 1:
                        student_info = info
                    
                    # Accumulate all answers
                    all_answers.extend(answers)
                
                # Limit answers to key length
                answers_out = all_answers[:q_out]
                answers_csv = [a if a is not None else "" for a in answers_out]
                
                # Get detected version
                student_version = student_info.get("version", "")
                
                # Metrics
                blanks = sum(1 for a in answers_out if (a is None or a == ""))
                multi = sum(1 for a in answers_out if (isinstance(a, str) and "," in a))
                correct = incorrect = 0
                percent = 0.0
                version_used = student_version
                key_out = None
                
                if keys_dict:
                    # Multi-version scoring
                    correct, total_scored, version_used = score_against_multi_keys(
                        answers_out,
                        student_version,
                        keys_dict,
                    )
                    
                    # Count incorrect
                    answered_single = sum(1 for a in answers_out if a and "," not in a)
                    incorrect = answered_single - correct
                    
                    percent = (100.0 * correct / max(1, total_scored))
                    
                    # Get the key for annotation
                    key_version = version_used.rstrip("*")
                    key_out = keys_dict.get(key_version, [])[:q_out]
                
                # Compose CSV row - Page column shows "1-2" for multi-page
                page_range = f"1-{pages_per_student}" if pages_per_student > 1 else "1"
                row = [
                    version_used,
                    page_range,
                    student_info.get("last_name", ""),
                    student_info.get("first_name", ""),
                    student_info.get("student_id", ""),
                ]
                
                if keys_dict:
                    row += [str(correct), str(incorrect), str(blanks), str(multi), f"{percent:.1f}"]
                else:
                    row += [str(blanks)]
                
                row += answers_csv

                # ==================== FLAGGING ====================
                # Record flagged items for this student (multi-page)
                if _has_flags(answers_out):
                    student_name = f"{student_info.get('last_name', '')} {student_info.get('first_name', '')}".strip()
                    _record_flags(
                        student_id=student_info.get("student_id", f"student_{student_idx+1}"),
                        page_num=start_page_idx + 1,  # First page of this student's set
                        answers=answers_out,
                        student_name=student_name,
                    )
                # ==================================================
                
                # ==================== STATS COLLECTION ====================
                # Collect data for stats computation (with version for multi-version support)
                all_student_data.append({
                    "answers": answers_out,
                    "correct": correct if keys_dict else 0,
                    "total": total_scored if keys_dict else q_out,
                    "version": version_used.rstrip("*") if version_used else "",
                    "csv_row": row,  # Store for grouped output
                })
                # ==========================================================
                
                # Annotate all pages for this student
                if out_annotated_dir or out_pdf_path:
                    for page_num, img_bgr in enumerate(student_pages, start=1):
                        page_layout = bmap.get_page(page_num)
                        
                        if page_layout is None:
                            raise ValueError(f"Could not get layout for page {page_num}")
                        
                        # DEBUG: Verify we have the right page
                        num_answer_layouts_this_page = len(page_layout.answer_layouts)
                        if verbose_calibration:
                            print(f"  Annotating page {page_num} with {num_answer_layouts_this_page} answer layout(s)")
                            if num_answer_layouts_this_page > 0:
                                first_layout = page_layout.answer_layouts[0]
                                print(f"    First layout coords: ({first_layout.x_topleft:.4f}, {first_layout.y_topleft:.4f}) to ({first_layout.x_bottomright:.4f}, {first_layout.y_bottomright:.4f})")
                                print(f"    First layout: {first_layout.numrows} rows × {first_layout.numcols} cols")
                        
                        page_fixed_thresh = page_thresholds[page_num - 1]
                        
                        # Get answers for this page only
                        # Calculate which answers belong to this page
                        answers_before_page = sum(
                            sum(layout.numrows for layout in bmap.get_page(p).answer_layouts)
                            for p in range(1, page_num)
                        )
                        answers_in_page = sum(layout.numrows for layout in page_layout.answer_layouts)
                        
                        # For annotation, we need the answers specific to THIS page
                        # Slice from all_answers to get just this page's answers
                        page_answers = all_answers[answers_before_page:answers_before_page + answers_in_page]
                        
                        # Limit to key length - only annotate answers that are within the key
                        # But we need to account for the offset
                        if q_out > answers_before_page:
                            # Some or all of this page's questions are within the key
                            page_answers_limit = min(answers_in_page, q_out - answers_before_page)
                            page_answers_out = page_answers[:page_answers_limit]
                        else:
                            # This entire page is beyond the key length
                            page_answers_out = []
                        
                        # Annotate with name/ID zones
                        # Create a proper Bubblemap-like object from PageLayout
                        temp_bmap_names = type('TempBmap', (object,), {
                            'last_name_layout': page_layout.last_name_layout,
                            'first_name_layout': page_layout.first_name_layout,
                            'id_layout': page_layout.id_layout,
                        })()
                        
                        vis = _annotate_names_ids(
                            img_bgr,
                            temp_bmap_names,
                            label_density=label_density,
                            annotation_defaults=ANNOTATION_DEFAULTS,
                        )
                        
                        # Annotate with answer zones
                        # Create a proper Bubblemap-like object from PageLayout
                        temp_bmap_answers = type('TempBmap', (object,), {
                            'answer_layouts': page_layout.answer_layouts,
                        })()
                        
                        vis = _annotate_answers(
                            vis,
                            temp_bmap_answers,
                            key_out[answers_before_page:answers_before_page + answers_in_page] if key_out else None,
                            answers_for_annotation=page_answers_out,
                            label_density=label_density,
                            annotate_all_cells=annotate_all_cells,
                            min_fill=min_fill,
                            top2_ratio=top2_ratio,
                            min_top2_diff=min_top2_diff,
                            fixed_thresh=page_fixed_thresh,
                            calibrate_background=calibrate_background,
                            background_percentile=background_percentile,
                            annotation_defaults=ANNOTATION_DEFAULTS,
                        )
                        
                        # Save annotated page
                        if out_annotated_dir:
                            actual_page_num = start_page_idx + page_num
                            out_png = os.path.join(out_annotated_dir, f"page_{actual_page_num:03d}_overlay.png")
                            cv2.imwrite(out_png, vis)
                        
                        if out_pdf_path:
                            if pdf_writer is not None:
                                pdf_writer.add_page(vis)
                            else:
                                annotated_pages.append(vis)
                        
                        # Track flagged pages for review PDF (multi-page mode)
                        if (review_pdf or flagged_xlsx) and _has_flags(answers_out):
                            actual_page_num = start_page_idx + page_num
                            flagged_page_images.append((actual_page_num, vis.copy()))
        
        else:
            # SINGLE-PAGE MODE: Original logic (unchanged)
            for page_idx, img_bgr in enumerate(pages, start=1):
                # Per-page calibration for lightly marked sheets
                page_fixed_thresh = fixed_thresh
                if fixed_thresh is None and auto_calibrate_thresh:
                    page_fixed_thresh, _calib_stats = calibrate_fixed_thresh_for_page(
                        img_bgr,
                        bmap,
                        max_rows=q_out if q_out < total_q else None,
                        verbose=verbose_calibration,
                    )
                # Decode all fields using the shared axis-mode pipeline
                info, answers, backgrounds = process_page_all(
                    img_bgr, bmap,
                    min_fill=min_fill,
                    top2_ratio=top2_ratio,
                    min_top2_diff=min_top2_diff,
                    fixed_thresh=page_fixed_thresh,
                    calibrate_background=calibrate_background,
                    background_percentile=background_percentile,
                    adaptive_min_above_floor=adaptive_min_above_floor,
                    verbose_calibration=verbose_calibration,
                )

                # Try adaptive rescoring if enabled and there are blanks
                if adaptive_rescoring:
                    from .tools.score_tools import adaptive_rescore_page
                    answers, page_fixed_thresh, adapted = adaptive_rescore_page(
                        img_bgr, bmap,
                        initial_threshold=page_fixed_thresh,
                        initial_answers=answers,
                        min_fill=min_fill,
                        top2_ratio=top2_ratio,
                        min_top2_diff=min_top2_diff,
                        calibrate_background=calibrate_background,
                        background_percentile=background_percentile,
                        adaptive_min_above_floor=adaptive_min_above_floor,
                        adaptive_max_adjustment=adaptive_max_adjustment,
                        max_questions=q_out if q_out < total_q else None,
                        verbose=verbose_calibration,
                    )

                # Limit answers to the Qs we output (based on key length if present)
                answers_out = answers[:q_out]
                answers_csv = [a if a is not None else "" for a in answers_out]

                # Get detected version
                student_version = info.get("version", "")
                
                # Metrics
                blanks = sum(1 for a in answers_out if (a is None or a == ""))
                multi = sum(1 for a in answers_out if (isinstance(a, str) and "," in a))
                correct = incorrect = 0
                percent = 0.0
                version_used = student_version
                key_out = None

                if keys_dict:
                    # Multi-version scoring
                    correct, total_scored, version_used = score_against_multi_keys(
                        answers_out,
                        student_version,
                        keys_dict,
                    )
                    
                    # Count incorrect (total scored - blanks - multi - correct)
                    answered_single = sum(1 for a in answers_out if a and "," not in a)
                    incorrect = answered_single - correct
                    
                    percent = (100.0 * correct / max(1, total_scored))
                    
                    # Get the key for annotation
                    key_version = version_used.rstrip("*")
                    key_out = keys_dict.get(key_version, [])[:q_out]
                
                # Compose CSV row
                row = [
                    version_used,
                    str(page_idx),
                    info.get("last_name", ""),
                    info.get("first_name", ""),
                    info.get("student_id", ""),
                ] 

                if keys_dict:
                    row += [str(correct), str(incorrect), str(blanks), str(multi), f"{percent:.1f}"]
                else:
                    row += [str(blanks)]

                row += answers_csv

                # ==================== FLAGGING ====================
                # Record flagged items for this student
                if _has_flags(answers_out):
                    student_name = f"{info.get('last_name', '')} {info.get('first_name', '')}".strip()
                    _record_flags(
                        student_id=info.get("student_id", f"page_{page_idx}"),
                        page_num=page_idx,
                        answers=answers_out,
                        student_name=student_name,
                    )
                # ==================================================
                
                # ==================== STATS COLLECTION ====================
                # Collect data for stats computation (with version for multi-version support)
                all_student_data.append({
                    "answers": answers_out,
                    "correct": correct if keys_dict else 0,
                    "total": total_scored if keys_dict else q_out,
                    "version": version_used.rstrip("*") if version_used else "",
                    "csv_row": row,  # Store for grouped output
                })
                # ==========================================================

                # Annotated image: names/IDs in blue (with optional %), then answers overlay
                if out_annotated_dir or out_pdf_path:
                    vis = _annotate_names_ids(
                        img_bgr,
                        bmap,
                        label_density=label_density,
                        annotation_defaults=ANNOTATION_DEFAULTS,
                    )
                    vis = _annotate_answers(
                        vis,
                        bmap,
                        key_out,
                        answers_for_annotation=answers_out,  # Only annotate questions in the key
                        label_density=label_density,
                        annotate_all_cells=annotate_all_cells,
                        min_fill=min_fill,
                        top2_ratio=top2_ratio,
                        min_top2_diff=min_top2_diff,
                        fixed_thresh=page_fixed_thresh,
                        calibrate_background=calibrate_background,
                        background_percentile=background_percentile,
                        annotation_defaults=ANNOTATION_DEFAULTS,
                    )
                    if out_annotated_dir:
                        out_png = os.path.join(out_annotated_dir, f"page_{page_idx:03d}_overlay.png")
                        cv2.imwrite(out_png, vis)
                    if out_pdf_path:
                        if pdf_writer is not None:
                            pdf_writer.add_page(vis)
                        else:
                            annotated_pages.append(vis)
                    
                    # Track flagged pages for review PDF (single-page mode)
                    if (review_pdf or flagged_xlsx) and _has_flags(answers_out):
                        flagged_page_images.append((page_idx, vis.copy()))
                        
    if out_pdf_path:
        if pdf_writer is not None:
            pdf_writer.close(save=True)
        elif annotated_pages:
            IO.save_images_as_pdf(annotated_pages, out_pdf_path, dpi=dpi, quality=pdf_quality)

    # ==================== WRITE GROUPED CSV + STATS ====================
    # Write CSV grouped by version:
    # - Version A: header, key, students
    # - Version B: header, key, students  
    # - Summary statistics at the end
    
    # Group students by version
    students_by_version: Dict[str, List[Dict]] = {}
    for d in all_student_data:
        ver = d.get("version", "") or "A"  # Default to A if no version
        if ver not in students_by_version:
            students_by_version[ver] = []
        students_by_version[ver].append(d)
    
    # Check if multi-version
    is_multi_version_output = len(students_by_version) > 1 and keys_dict and len(keys_dict) > 1
    
    with open(out_csv, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        
        # Compute stats FIRST so we have per-version stats available
        stats = None
        if include_stats and keys_dict and all_student_data:
            stats = _compute_basic_stats(
                all_student_data=all_student_data,
                keys_dict=keys_dict,
                q_out=q_out,
            )
        
        # Helper for stats rows
        n_cols_before_q = 5 + 5 if keys_dict else 5 + 1
        
        def make_stats_row(label: str, value, q_values=None):
            """Helper to create a stats row with label and optional per-Q values."""
            row = [""] * n_cols_before_q
            row[0] = label
            if q_values:
                row.extend([str(v) if v == v else "" for v in q_values])
            else:
                row.extend([""] * q_out)
            if value is not None:
                row[1] = str(value)
            return row
        
        # Write each version's section (with its item stats immediately after)
        for ver_idx, ver in enumerate(sorted(students_by_version.keys())):
            students = students_by_version[ver]
            
            # Add blank line between versions (except before first)
            if ver_idx > 0:
                writer.writerow([])
            
            # Write section header for multi-version exams
            if is_multi_version_output:
                writer.writerow([f"=== VERSION {ver} ({len(students)} students) ==="])
            
            # Write column header
            writer.writerow(csv_header)
            
            # Write KEY row for this version
            if keys_dict and ver in keys_dict:
                key_row = [ver, "0", "KEY", "KEY", "KEY"]
                if keys_dict:
                    key_row += ["", "", "", "", ""]
                else:
                    key_row += [""]
                key_row += keys_dict[ver][:q_out]
                writer.writerow(key_row)
            
            # Write student rows for this version
            for student in students:
                writer.writerow(student["csv_row"])
            
            # Write item stats for THIS version immediately after its students
            if stats and stats.get("versions") and ver in stats["versions"]:
                ver_stats = stats["versions"][ver]
                if ver_stats.get("per_item_pct"):
                    writer.writerow([])
                    if is_multi_version_output:
                        writer.writerow(make_stats_row(f"--- ITEM STATISTICS: VERSION {ver} (N={ver_stats['n']}) ---", ""))
                    else:
                        writer.writerow(make_stats_row("--- ITEM STATISTICS ---", ""))
                    writer.writerow(make_stats_row(f"PCT_CORRECT{'_'+ver if is_multi_version_output else ''}", "", ver_stats["per_item_pct"]))
                    writer.writerow(make_stats_row(f"POINT_BISERIAL{'_'+ver if is_multi_version_output else ''}", "", ver_stats["per_item_pb"]))
        
        # ==================== WRITE OVERALL STATS AT END ====================
        if stats:
            # Separator
            writer.writerow([])
            writer.writerow([])
            
            # Overall exam statistics header
            writer.writerow(make_stats_row("=== EXAM STATISTICS (ALL VERSIONS) ===", ""))
            writer.writerow(make_stats_row("N_STUDENTS", stats["n_students"]))
            writer.writerow(make_stats_row("MEAN_SCORE", stats["mean_score"]))
            writer.writerow(make_stats_row("MEAN_PERCENT", f"{stats['mean_percent']:.1f}%"))
            writer.writerow(make_stats_row("STD_DEV", stats["std_dev"]))
            writer.writerow(make_stats_row("HIGH_SCORE", stats["high_score"]))
            writer.writerow(make_stats_row("LOW_SCORE", stats["low_score"]))
            
            # KR-20: Overall
            if not np.isnan(stats["kr20_overall"]):
                writer.writerow(make_stats_row("KR20_OVERALL", stats["kr20_overall"]))
            
            # KR-20: Per version (if multi-version)
            if stats.get("is_multi_version") and stats.get("versions"):
                for ver, ver_stats in sorted(stats["versions"].items()):
                    if not np.isnan(ver_stats["kr20"]):
                        writer.writerow(make_stats_row(f"KR20_VERSION_{ver}", f"{ver_stats['kr20']} (N={ver_stats['n']})"))
            
            # Note: Per-version item stats are now written immediately after each version's students
            
            # Print stats summary to stderr
            print(f"[info] Exam stats: N={stats['n_students']}, Mean={stats['mean_score']:.1f} ({stats['mean_percent']:.1f}%), "
                  f"SD={stats['std_dev']:.2f}, Range={stats['low_score']}-{stats['high_score']}", file=sys.stderr)
            
            if not np.isnan(stats["kr20_overall"]):
                print(f"[info] KR-20 overall: {stats['kr20_overall']:.3f}", file=sys.stderr)
            
            if stats.get("is_multi_version") and stats.get("versions"):
                for ver, ver_stats in sorted(stats["versions"].items()):
                    if not np.isnan(ver_stats["kr20"]):
                        print(f"[info] KR-20 version {ver}: {ver_stats['kr20']:.3f} (N={ver_stats['n']})", file=sys.stderr)
    # =================================================================

    # ==================== GENERATE REVIEW OUTPUTS ====================
    # Write flagged XLSX if requested
    if flagged_xlsx and flagged_items:
        _ensure_dir(os.path.dirname(flagged_xlsx) or ".")
        wb = Workbook()
        ws = wb.active
        ws.title = "Flagged Items"

        # Header row
        headers = ["Student ID", "Student Name", "Page", "Question", "Issue", "Current Answer", "Corrected Answer"]
        ws.append(headers)

        # Style header row
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

        # Data rows
        for item in flagged_items:
            ws.append([
                item["student_id"],
                item["student_name"],
                item["page"],
                item["question"],
                item["issue"],
                item["current_answer"],
                "",  # Empty Corrected Answer column for teacher input
            ])

        # Style data rows and highlight Corrected Answer column
        yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        for row_num in range(2, len(flagged_items) + 2):
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.border = thin_border
                # Highlight the Corrected Answer column
                if col_num == 7:  # Corrected Answer column
                    cell.fill = yellow_fill

        # Set column widths
        ws.column_dimensions['A'].width = 12  # Student ID
        ws.column_dimensions['B'].width = 20  # Student Name
        ws.column_dimensions['C'].width = 8   # Page
        ws.column_dimensions['D'].width = 10  # Question
        ws.column_dimensions['E'].width = 10  # Issue
        ws.column_dimensions['F'].width = 15  # Current Answer
        ws.column_dimensions['G'].width = 18  # Corrected Answer

        # Freeze header row
        ws.freeze_panes = 'A2'

        wb.save(flagged_xlsx)
        print(f"[info] Wrote {len(flagged_items)} flagged items to {flagged_xlsx}", file=sys.stderr)
    
    # Generate review PDF with flagged pages
    if review_pdf and flagged_page_images:
        _ensure_dir(os.path.dirname(review_pdf) or ".")
        review_pages = [img for _, img in sorted(flagged_page_images, key=lambda x: x[0])]
        
        try:
            review_writer = IO.PdfPageWriter(review_pdf, dpi=dpi)
            for img in review_pages:
                review_writer.add_page(img)
            review_writer.close(save=True)
        except Exception:
            # Fallback to PyMuPDF-based saving
            IO.save_images_as_pdf(review_pages, review_pdf, dpi=dpi, quality=pdf_quality)
        
        print(f"[info] Wrote review PDF with {len(review_pages)} flagged pages to {review_pdf}", file=__import__('sys').stderr)
    elif review_pdf and not flagged_page_images:
        print("[info] No flagged pages - review PDF not created", file=__import__('sys').stderr)
    
    # Print summary
    if flagged_items:
        blank_count = sum(1 for item in flagged_items if item["issue"] == "blank")
        multi_count = sum(1 for item in flagged_items if item["issue"] == "multi")
        print(f"[info] Flagging summary: {blank_count} blank, {multi_count} multi across {len(set(item['student_id'] for item in flagged_items))} students", file=__import__('sys').stderr)
    # =================================================================

    return out_csv
