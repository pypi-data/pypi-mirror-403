#!/usr/bin/env python3
"""
MarkShark
report_tools.py
Generate teacher-friendly Excel reports from scored CSV results

Features:
- Multi-version support: separate tab per version
- Roster matching: flags absent students and orphan scans
- Color-coded item quality indicators
- Summary statistics and item analysis
"""

from __future__ import annotations
from typing import Optional, List, Dict, Tuple

import pandas as pd
import numpy as np

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    raise ImportError(
        "openpyxl is required for Excel report generation. "
        "Install it with: pip install openpyxl"
    )

try:
    from rapidfuzz import fuzz
except ImportError:
    raise ImportError(
        "rapidfuzz is required for student roster matching. "
        "Install it with: pip install rapidfuzz"
    )

from .stats_tools import (
    detect_item_columns,
    detect_key_row_index,
    prepare_correctness_matrix,
    point_biserial,
    kr20,
    kr21,
)


# ==================== CORRECTIONS HANDLING ====================

def _normalize_id(val) -> str:
    """Normalize a student ID for comparison: stringify, strip, remove .0 suffix."""
    s = str(val).strip()
    if s.endswith('.0'):
        s = s[:-2]
    return s


def _find_col(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
    """Find the first column name from candidates that exists in df."""
    for col in candidates:
        if col in df.columns:
            return col
    return None


def load_corrections(corrections_xlsx: str) -> pd.DataFrame:
    """
    Load corrections from a filled-in flagged.xlsx file.

    The file should have columns:
    - Student ID
    - Question
    - Corrected Answer (non-empty means a correction)

    Returns:
        DataFrame with columns: student_id, question, corrected_answer
        Only rows with non-empty Corrected Answer are returned.
    """
    import sys

    # Try 'Flagged Items' sheet first, then fall back to first sheet
    try:
        df = pd.read_excel(corrections_xlsx, sheet_name="Flagged Items")
    except ValueError:
        df = pd.read_excel(corrections_xlsx, sheet_name=0)
    print(f"[corrections] Loaded XLSX with columns: {list(df.columns)}", file=sys.stderr)

    # Normalize column names to handle variations
    col_map = {}
    for col in df.columns:
        col_lower = col.lower().replace(' ', '_').strip()
        if 'student' in col_lower and 'id' in col_lower:
            col_map[col] = 'student_id'
        elif col_lower == 'question':
            col_map[col] = 'question'
        elif 'corrected' in col_lower and 'answer' in col_lower:
            col_map[col] = 'corrected_answer'

    df = df.rename(columns=col_map)
    print(f"[corrections] After rename, columns: {list(df.columns)}", file=sys.stderr)

    # Filter to only rows with corrections
    if 'corrected_answer' not in df.columns:
        print("[corrections] Warning: No 'corrected_answer' column found!", file=sys.stderr)
        return pd.DataFrame(columns=['student_id', 'question', 'corrected_answer'])

    # Convert to string, then filter out empties and NaN placeholders
    df['corrected_answer'] = df['corrected_answer'].astype(str).str.strip()
    mask = (df['corrected_answer'] != '') & (df['corrected_answer'].str.lower() != 'nan') & (df['corrected_answer'].str.lower() != 'none')
    corrections = df[mask]

    print(f"[corrections] Found {len(corrections)} rows with corrections", file=sys.stderr)
    if not corrections.empty:
        print(f"[corrections] First correction: {corrections.iloc[0].to_dict()}", file=sys.stderr)

    if corrections.empty:
        return pd.DataFrame(columns=['student_id', 'question', 'corrected_answer'])

    return corrections[['student_id', 'question', 'corrected_answer']].copy()


def _question_to_col(question_val, item_cols: List[str]) -> Optional[str]:
    """
    Map a question identifier (from the flagged XLSX) to a CSV column name.

    Handles: bare integer 22 -> 'Q22', string 'Q22' -> 'Q22', string '22' -> 'Q22'.
    """
    q_str = str(question_val).strip()

    # Direct match (e.g., 'Q22' in item_cols)
    if q_str in item_cols:
        return q_str
    if q_str.upper() in [c.upper() for c in item_cols]:
        for c in item_cols:
            if c.upper() == q_str.upper():
                return c

    # Bare number -> Q-prefixed (e.g., 22 -> 'Q22')
    q_str_digits = q_str.lstrip('Qq')
    if q_str_digits.isdigit():
        candidate = f"Q{int(q_str_digits)}"
        if candidate in item_cols:
            return candidate

    return None


def merge_corrections(
    df: pd.DataFrame,
    corrections: pd.DataFrame,
    item_cols: List[str],
    key_row_idx: int,
) -> Tuple[pd.DataFrame, int]:
    """
    Apply corrections to the results DataFrame.

    For each correction, updates the corresponding question answer
    in the student's row and recalculates the scoring columns.

    Args:
        df: Results DataFrame from scored CSV
        corrections: DataFrame with columns: student_id, question, corrected_answer
        item_cols: List of question column names (Q1, Q2, etc.)
        key_row_idx: Index of the KEY row

    Returns:
        Tuple of (modified DataFrame, number of corrections applied)
    """
    import sys

    if corrections.empty:
        return df, 0

    df = df.copy()
    corrections_applied = 0

    # Get the key values for recalculating scores
    key_row = df.iloc[key_row_idx]

    # Find the student ID column
    id_col = _find_col(df, ['studentid', 'StudentID', 'Student_ID', 'student_id', 'ID', 'id'])
    if id_col is None:
        print(f"[corrections] Warning: No student ID column found in {list(df.columns)}", file=sys.stderr)
        return df, 0

    # Pre-normalize all IDs in the DataFrame for matching
    df_ids_normalized = df[id_col].apply(_normalize_id)

    for _, correction in corrections.iterrows():
        student_id_normalized = _normalize_id(correction['student_id'])
        new_answer = str(correction['corrected_answer']).upper().strip()

        # Map question to column name
        q_col = _question_to_col(correction['question'], item_cols)
        if q_col is None:
            print(f"[corrections] Warning: Could not find column for question '{correction['question']}' "
                  f"(available: {item_cols[:5]}...)", file=sys.stderr)
            continue

        # Find matching student rows (excluding KEY row)
        student_mask = df_ids_normalized == student_id_normalized
        student_indices = [i for i in df[student_mask].index.tolist() if i != key_row_idx]

        if not student_indices:
            sample_ids = df_ids_normalized[df_ids_normalized != 'KEY'].head(5).tolist()
            print(f"[corrections] Warning: Could not find student '{student_id_normalized}' "
                  f"(sample IDs: {sample_ids})", file=sys.stderr)
            continue

        # Apply correction
        for idx in student_indices:
            old_answer = str(df.at[idx, q_col]).upper().strip()
            df.at[idx, q_col] = new_answer
            corrections_applied += 1
            print(f"[corrections] Applied: Student {student_id_normalized}, "
                  f"{q_col}: '{old_answer}' -> '{new_answer}'", file=sys.stderr)

            _recalculate_student_scores(df, idx, item_cols, key_row)

    return df, corrections_applied


def _recalculate_student_scores(
    df: pd.DataFrame,
    student_idx: int,
    item_cols: List[str],
    key_row: pd.Series,
):
    """
    Recalculate correct/incorrect/blank/multi/percent for a student after corrections.
    """
    correct = 0
    incorrect = 0
    blank = 0
    multi = 0

    for col in item_cols:
        answer = str(df.at[student_idx, col]).upper().strip()
        key_answer = str(key_row[col]).upper().strip()

        if not answer or answer in ('', 'NAN', 'NONE', 'BLANK', '-'):
            blank += 1
        elif ',' in answer:
            multi += 1
        elif answer == key_answer:
            correct += 1
        else:
            incorrect += 1

    total = len(item_cols)
    percent = (correct / total * 100) if total > 0 else 0

    # Update the score columns (handle all case variants from CSV normalization)
    for name, value in [('correct', correct), ('incorrect', incorrect),
                        ('blank', blank), ('multi', multi), ('percent', round(percent, 2))]:
        col = _find_col(df, [name, name.capitalize(), name.upper()])
        if col is not None:
            df.at[student_idx, col] = value


def apply_corrections_to_csv(
    input_csv: str,
    corrections_xlsx: str,
    output_csv: str,
) -> int:
    """
    Apply teacher corrections to a scored CSV and write a new corrected CSV.

    Reads the scored CSV, loads corrections from the filled flagged.xlsx,
    updates the answer columns, recalculates scores, and writes a new CSV.

    Args:
        input_csv: Path to original scored CSV
        corrections_xlsx: Path to filled flagged.xlsx with corrections
        output_csv: Path to write the corrected CSV

    Returns:
        Number of corrections applied
    """
    import sys

    # Load the CSV
    df = _load_score_csv_robust(input_csv)

    # Detect item columns and key row
    item_cols = detect_item_columns(df, r"Q\d+")
    if not item_cols:
        raise ValueError(f"No item columns (Q1, Q2...) found in CSV. Columns: {list(df.columns)}")

    key_row_idx = detect_key_row_index(df, item_cols, key_label="KEY")

    # Load and apply corrections
    corrections = load_corrections(corrections_xlsx)
    if corrections.empty:
        print("[corrections] No corrections found in XLSX", file=sys.stderr)
        # Still write the output (a clean copy)
        df.to_csv(output_csv, index=False)
        return 0

    df, corrections_applied = merge_corrections(df, corrections, item_cols, key_row_idx)

    # Write corrected CSV
    df.to_csv(output_csv, index=False)
    print(f"[corrections] Wrote corrected CSV with {corrections_applied} corrections to {output_csv}", file=sys.stderr)

    return corrections_applied


# ==================== ROSTER MATCHING ====================

def load_roster(roster_path: str) -> pd.DataFrame:
    """
    Load and normalize a class roster CSV.

    Expected columns (case-insensitive, auto-detected):
    - StudentID / ID / Student_ID
    - LastName / Last / Surname
    - FirstName / First (optional)

    Returns DataFrame with standardized columns: StudentID, LastName, FirstName
    """
    df = pd.read_csv(roster_path)

    # Normalize column names
    col_map = {}
    for col in df.columns:
        col_lower = col.lower().strip()
        if col_lower in ('studentid', 'id', 'student_id', 'sid'):
            col_map[col] = 'StudentID'
        elif col_lower in ('lastname', 'last', 'surname', 'last_name'):
            col_map[col] = 'LastName'
        elif col_lower in ('firstname', 'first', 'first_name'):
            col_map[col] = 'FirstName'

    if 'StudentID' not in col_map.values():
        raise ValueError(
            f"Roster CSV must have a student ID column. "
            f"Expected: StudentID/ID/Student_ID. Found: {list(df.columns)}"
        )

    if 'LastName' not in col_map.values():
        raise ValueError(
            f"Roster CSV must have a last name column. "
            f"Expected: LastName/Last/Surname. Found: {list(df.columns)}"
        )

    df = df.rename(columns=col_map)

    # Fill missing FirstName with empty string
    if 'FirstName' not in df.columns:
        df['FirstName'] = ''

    # Convert StudentID to string, strip whitespace, and remove float artifacts
    # (pandas reads numeric IDs as float when NaNs are present, producing "1234.0")
    df['StudentID'] = df['StudentID'].apply(_normalize_id)
    df['LastName'] = df['LastName'].astype(str).str.strip()
    df['FirstName'] = df['FirstName'].astype(str).str.strip()

    return df[['StudentID', 'LastName', 'FirstName']]


def fuzzy_match_student(
    scanned_id: str,
    scanned_last: str,
    scanned_first: str,
    roster: pd.DataFrame,
    id_threshold: float = 85.0,
    name_threshold: float = 85.0,
) -> Tuple[Optional[str], float, str]:
    """
    Attempt to match a scanned student to the roster using fuzzy matching.

    Returns:
        (matched_roster_id, confidence, match_type)

        match_type can be:
        - "exact": Exact StudentID match
        - "high_confidence": High ID similarity or ID + name match
        - "probable": Moderate confidence match
        - "no_match": No good match found
    """
    if roster.empty:
        return None, 0.0, "no_match"

    # Clean inputs
    scanned_id = str(scanned_id).strip()
    scanned_last = str(scanned_last).strip().upper()
    scanned_first = str(scanned_first).strip().upper()

    best_match = None
    best_score = 0.0
    match_type = "no_match"

    for _, row in roster.iterrows():
        roster_id = str(row['StudentID']).strip()
        roster_last = str(row['LastName']).strip().upper()
        roster_first = str(row['FirstName']).strip().upper()

        # Exact ID match
        if scanned_id == roster_id:
            return roster_id, 100.0, "exact"

        # Fuzzy ID match
        id_score = fuzz.ratio(scanned_id, roster_id)

        # Name matching
        last_score = fuzz.ratio(scanned_last, roster_last) if scanned_last else 0
        first_score = fuzz.ratio(scanned_first, roster_first) if scanned_first and roster_first else 0

        # Combined scoring strategies
        # Strategy 1: Very high ID match (typo in one digit)
        if id_score >= 95:
            confidence = id_score
            if confidence > best_score:
                best_match = roster_id
                best_score = confidence
                match_type = "high_confidence"

        # Strategy 2: Good ID match + exact last name
        elif id_score >= id_threshold and last_score == 100:
            confidence = (id_score + last_score) / 2
            if confidence > best_score:
                best_match = roster_id
                best_score = confidence
                match_type = "high_confidence"

        # Strategy 3: Perfect name match (both first and last)
        elif last_score == 100 and first_score == 100 and scanned_first:
            confidence = 100.0
            if confidence > best_score:
                best_match = roster_id
                best_score = confidence
                match_type = "probable"

        # Strategy 4: Good overall match
        elif id_score >= id_threshold or (last_score >= name_threshold and id_score >= 70):
            confidence = max(id_score, (id_score + last_score) / 2)
            if confidence > best_score:
                best_match = roster_id
                best_score = confidence
                match_type = "probable"

    return best_match, best_score, match_type


def match_students_to_roster(
    students_df: pd.DataFrame,
    roster_df: pd.DataFrame,
) -> Tuple[pd.DataFrame, List[Dict], List[Dict]]:
    """
    Match scanned students to roster.

    Returns:
        (students_df_with_matches, orphan_scans, absent_students)

        students_df_with_matches: Original DataFrame with added columns:
            - RosterID: Matched roster ID (or None)
            - MatchConfidence: 0-100
            - MatchType: exact/high_confidence/probable/no_match

        orphan_scans: List of dicts for students who couldn't be matched confidently
        absent_students: List of dicts for roster students not found in scans
    """
    # Add matching columns
    students_df['RosterID'] = None
    students_df['MatchConfidence'] = 0.0
    students_df['MatchType'] = 'no_match'

    matched_roster_ids = set()

    # Resolve column names (scored CSV may use lowercase 'studentid', 'lastname', etc.)
    _sid_col = _find_col(students_df, ['StudentID', 'studentid', 'Student_ID', 'student_id', 'ID', 'id'])
    _last_col = _find_col(students_df, ['LastName', 'lastname', 'Last', 'last', 'Surname', 'last_name'])
    _first_col = _find_col(students_df, ['FirstName', 'firstname', 'First', 'first', 'first_name'])

    for idx, row in students_df.iterrows():
        scanned_id = _normalize_id(row.get(_sid_col, '')) if _sid_col else ''
        scanned_last = str(row.get(_last_col, '')).strip() if _last_col else ''
        scanned_first = str(row.get(_first_col, '')).strip() if _first_col else ''

        if not scanned_id and not scanned_last:
            # Can't match without any identifying info
            continue

        matched_id, confidence, match_type = fuzzy_match_student(
            scanned_id, scanned_last, scanned_first, roster_df
        )

        if matched_id:
            students_df.at[idx, 'RosterID'] = matched_id
            students_df.at[idx, 'MatchConfidence'] = confidence
            students_df.at[idx, 'MatchType'] = match_type
            matched_roster_ids.add(matched_id)

    # Find orphan scans (low confidence or no match)
    orphan_scans = []
    for idx, row in students_df.iterrows():
        if row['MatchType'] in ('no_match', 'probable'):
            orphan_scans.append({
                'ScannedID': _normalize_id(row.get(_sid_col, '')) if _sid_col else '',
                'LastName': str(row.get(_last_col, '')).strip() if _last_col else '',
                'FirstName': str(row.get(_first_col, '')).strip() if _first_col else '',
                'MatchType': row['MatchType'],
                'PossibleMatch': row['RosterID'] if row['MatchType'] == 'probable' else None,
                'Confidence': row['MatchConfidence'],
            })

    # Find absent students (matched_roster_ids contains normalized strings)
    absent_students = []
    for _, row in roster_df.iterrows():
        roster_id = _normalize_id(row['StudentID'])
        if roster_id not in matched_roster_ids:
            absent_students.append({
                'StudentID': roster_id,
                'LastName': row['LastName'],
                'FirstName': row['FirstName'],
            })

    return students_df, orphan_scans, absent_students


# ==================== EXCEL FORMATTING ====================

# Color scheme
COLOR_HEADER = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
COLOR_KEY_ROW = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
COLOR_GOOD = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
COLOR_WARNING = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
COLOR_PROBLEM = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
COLOR_ORPHAN = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
COLOR_BLANK = PatternFill(start_color="F056E3", end_color="F056E3", fill_type="solid")  # Pink-purple for blank answers
COLOR_MULTI = PatternFill(start_color="FFB366", end_color="FFB366", fill_type="solid")  # Orange for multi-answer

FONT_HEADER = Font(bold=True, color="FFFFFF")
FONT_BOLD = Font(bold=True)
BORDER_THIN = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def format_header_row(ws, row_num: int):
    """Apply header formatting to a row."""
    for cell in ws[row_num]:
        cell.fill = COLOR_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER_THIN


def format_key_row(ws, row_num: int):
    """Apply KEY row formatting."""
    for cell in ws[row_num]:
        cell.fill = COLOR_KEY_ROW
        cell.font = FONT_BOLD
        cell.border = BORDER_THIN


def auto_size_columns(ws):
    """Auto-size all columns based on content."""
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


# ==================== CSV LOADING ====================

def _load_score_csv_robust(input_csv: str) -> pd.DataFrame:
    """
    Load CSV from score command, handling messy format with include_stats.

    The score command with include_stats can create CSVs with:
    - Section headers like "=== VERSION A (112 students) ==="
    - Multiple header rows (one per version)
    - Stats rows at the bottom

    This function:
    1. Finds the first valid header row
    2. Reads only the student data rows (skipping section headers and stats)
    3. Returns a clean DataFrame
    """
    import csv

    # Read the raw CSV to find structure
    with open(input_csv, 'r', encoding='utf-8-sig') as f:
        lines = f.readlines()

    # Find the first valid header row (has question columns like Q1, Q2, etc.)
    header_idx = None
    header = None

    for idx, line in enumerate(lines):
        row = next(csv.reader([line]))
        # Valid header should have StudentID and Q1
        if 'StudentID' in row or 'StudentID' in str(row):
            # Check if this looks like a real header with questions
            if any(col.startswith('Q') and col[1:].isdigit() for col in row if isinstance(col, str)):
                header_idx = idx
                header = row
                break

    if header_idx is None:
        # Fallback: just try reading normally
        return pd.read_csv(input_csv)

    # Collect all data rows (skip section headers and stats rows)
    data_rows = []
    reader = csv.reader(lines[header_idx + 1:])

    for row in reader:
        # Skip empty rows
        if not row or all(cell.strip() == '' for cell in row):
            continue

        # Skip section headers (=== VERSION A === etc.)
        if row and row[0].strip().startswith('==='):
            continue

        # Skip stats rows (start with specific labels)
        if row and row[0].strip() in ['PCT_CORRECT', 'POINT_BISERIAL', 'N_STUDENTS',
                                        'MEAN_SCORE', 'MEAN_PERCENT', 'STD_DEV',
                                        'HIGH_SCORE', 'LOW_SCORE', 'KR20_OVERALL'] or \
           (row and 'PCT_CORRECT' in row[0]) or \
           (row and 'POINT_BISERIAL' in row[0]) or \
           (row and 'KR20_VERSION' in row[0]) or \
           (row and '--- ITEM STATISTICS' in row[0]):
            continue

        # Skip duplicate header rows (multi-version CSVs repeat the header)
        if row == header:
            continue
        # Also catch headers that match loosely (e.g., same first few columns)
        if len(row) >= 3 and row[0].strip() == header[0].strip() and row[1].strip() == header[1].strip() and row[2].strip() == header[2].strip():
            continue

        # This looks like a valid data row
        # Ensure row has same length as header
        if len(row) < len(header):
            row.extend([''] * (len(header) - len(row)))
        elif len(row) > len(header):
            row = row[:len(header)]

        data_rows.append(row)

    # Create DataFrame
    df = pd.DataFrame(data_rows, columns=header)

    # Normalize column names to lowercase for consistency
    # This handles both 'Correct' and 'correct', 'Version' and 'version', etc.
    column_mapping = {}
    for col in df.columns:
        col_lower = col.lower().strip()
        # Map common variations to standard names
        if col_lower in ('correct', 'incorrect', 'blank', 'multi', 'percent',
                         'version', 'studentid', 'lastname', 'firstname',
                         'page', 'page_index'):
            column_mapping[col] = col_lower
        # Keep question columns as-is (Q1, Q2, etc.)
        elif col.startswith('Q') and col[1:].isdigit():
            column_mapping[col] = col
        else:
            column_mapping[col] = col

    df = df.rename(columns=column_mapping)

    # Convert score columns from strings to proper numeric types so Excel
    # formats them as numbers instead of left-aligned text.
    for col in ('correct', 'incorrect', 'blank', 'multi'):
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')
    if 'percent' in df.columns:
        df['percent'] = pd.to_numeric(df['percent'], errors='coerce')

    return df


# ==================== REPORT GENERATION ====================

def generate_report(
    input_csv: str,
    out_xlsx: str,
    roster_csv: Optional[str] = None,
    item_pattern: str = r"Q\d+",
    project_name: Optional[str] = None,
    run_label: Optional[str] = None,
    corrections_applied: int = 0,
    corrections_xlsx: Optional[str] = None,
):
    r"""
    Generate comprehensive Excel report from scored CSV.

    Args:
        input_csv: Path to scored CSV from markshark score
        out_xlsx: Path to output Excel file
        roster_csv: Optional path to class roster CSV
        item_pattern: Regex pattern for item columns (default: Q\d+)
        project_name: Optional project name for report header
        run_label: Optional run label (e.g., "run_001_2025-01-21_1430")
        corrections_applied: Number of corrections applied (for display on Summary tab)
        corrections_xlsx: Optional path to the corrections XLSX (for listing details on Summary tab)
    """
    # Load scored results - handle messy CSV from score with include_stats
    df = _load_score_csv_robust(input_csv)

    # Detect item columns and key row
    item_cols = detect_item_columns(df, item_pattern)
    if not item_cols:
        raise ValueError(
            f"No item columns found matching pattern '{item_pattern}'. "
            f"Available columns: {list(df.columns)}"
        )

    k = len(item_cols)

    # Load roster if provided
    roster_df = None
    orphan_scans = []
    absent_students = []

    if roster_csv:
        roster_df = load_roster(roster_csv)
        # Remove all KEY rows before matching (multi-version CSVs have one per version)
        key_mask = df.apply(
            lambda row: any(str(cell).strip().upper() == 'KEY' for cell in row), axis=1
        )
        students_only = df[~key_mask].reset_index(drop=True)
        students_only, orphan_scans, absent_students = match_students_to_roster(
            students_only, roster_df
        )

    # Group by version (handle both 'Version' and 'version' column names)
    version_col = None
    if 'version' in df.columns:
        version_col = 'version'
    elif 'Version' in df.columns:
        version_col = 'Version'

    if version_col:
        # Get unique versions, filtering out invalid values
        all_versions = df[version_col].dropna().astype(str).str.strip().unique()
        # Filter out "VERSION" and other non-single-letter values
        versions = sorted([v for v in all_versions
                          if v and len(v) <= 2 and v.upper() != 'VERSION' and v != 'KEY'])
    else:
        versions = ['A']  # Default single version

    # Compute per-version statistics
    # Each version has its own answer key, so correctness must be computed per-version.
    # Pooled (amalgamated) stats are then derived by combining per-version results.
    version_stats = {}
    all_items_num = []      # Collect correctness matrices for pooled stats
    all_total_scores = []   # Collect total scores for pooled stats
    total_n_students = 0

    for version in versions:
        # Filter to only students who took this version
        if version_col:
            version_mask = df[version_col].astype(str).str.strip() == str(version).strip()
            df_version = df[version_mask].copy()
        else:
            df_version = df.copy()

        # Find KEY row for this version
        key_row_idx_version = detect_key_row_index(df_version, item_cols, key_label="KEY")

        # Prepare correctness matrix for this version only
        items_num_v, total_scores_v, students_df_v, key_series_v = prepare_correctness_matrix(
            df_version, item_cols, key_row_idx_version, answers_mode="auto"
        )

        n_students_v = len(students_df_v)
        total_n_students += n_students_v

        # Compute difficulty (% correct) for this version
        difficulty_v = items_num_v.mean(axis=0)

        # Compute point-biserial for this version
        pb_vals_v = {}
        for col in item_cols:
            item_series = items_num_v[col]
            total_minus = total_scores_v - item_series.fillna(0)
            pb_vals_v[col] = point_biserial(item_series, total_minus)

        # Version-level exam stats
        mean_v = float(total_scores_v.mean()) if n_students_v > 0 else 0.0
        std_v = float(total_scores_v.std(ddof=1)) if n_students_v > 1 else 0.0
        kr20_v = kr20(items_num_v, total_scores_v)
        kr21_v = kr21(items_num_v, total_scores_v)

        # Store version-specific stats
        version_stats[version] = {
            'difficulty': difficulty_v,
            'pb_vals': pb_vals_v,
            'key_series': key_series_v,
            'n_students': n_students_v,
            'mean': mean_v,
            'std': std_v,
            'kr20': kr20_v,
            'kr21': kr21_v,
        }

        # Accumulate for pooled stats
        all_items_num.append(items_num_v)
        all_total_scores.append(total_scores_v)

    # Compute pooled (amalgamated) stats from per-version correctness matrices.
    # Each student was scored against their own version's key, so these are correct.
    if all_items_num:
        pooled_items = pd.concat(all_items_num, ignore_index=True)
        pooled_totals = pd.concat(all_total_scores, ignore_index=True)
    else:
        pooled_items = pd.DataFrame(columns=item_cols)
        pooled_totals = pd.Series(dtype=float)

    mean_total = float(pooled_totals.mean()) if total_n_students > 0 else 0.0
    std_total = float(pooled_totals.std(ddof=1)) if total_n_students > 1 else 0.0
    kr20_val = kr20(pooled_items, pooled_totals)
    kr21_val = kr21(pooled_items, pooled_totals)

    # Load full flagged xlsx for display on summary tab (all rows, all columns)
    corrections_detail = None
    if corrections_xlsx:
        try:
            try:
                corrections_detail = pd.read_excel(corrections_xlsx, sheet_name="Flagged Items")
            except ValueError:
                corrections_detail = pd.read_excel(corrections_xlsx, sheet_name=0)
        except Exception:
            corrections_detail = None

    # Create Excel workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # ========== SUMMARY TAB ==========
    create_summary_tab(
        wb, k, total_n_students, mean_total, std_total, kr20_val, kr21_val,
        versions, version_stats,
        orphan_scans, absent_students, project_name, run_label,
        corrections_applied, corrections_detail
    )

    # ========== PER-VERSION TABS ==========
    for version in versions:
        # Get version-specific stats
        vstats = version_stats[version]
        create_version_tab(
            wb, df, None, version, item_cols, vstats['key_series'],
            vstats['difficulty'], vstats['pb_vals'], roster_df, orphan_scans if roster_csv else None
        )

    # ========== CLASS SCORES TAB ==========
    create_class_scores_tab(wb, df, item_cols, k)

    # Save workbook
    wb.save(out_xlsx)
    print(f"Excel report generated: {out_xlsx}")


def create_summary_tab(
    wb, k, n_students, mean_total, std_total, kr20_val, kr21_val,
    versions, version_stats,
    orphan_scans, absent_students, project_name=None, run_label=None,
    corrections_applied=0, corrections_detail=None
):
    """Create summary tab with per-version and amalgamated exam statistics."""
    from datetime import datetime

    ws = wb.create_sheet("Summary", 0)
    n_versions = len(versions)
    is_multi_version = n_versions > 1

    # Title
    ws['A1'] = "MarkShark Exam Report"
    ws['A1'].font = Font(size=16, bold=True)

    # Project metadata (if provided)
    row = 2
    if project_name:
        ws[f'A{row}'] = "Project:"
        ws[f'A{row}'].font = FONT_BOLD
        ws[f'B{row}'] = project_name
        row += 1

    if run_label:
        ws[f'A{row}'] = "Run:"
        ws[f'A{row}'].font = FONT_BOLD
        ws[f'B{row}'] = run_label
        row += 1

    # Always show generation timestamp
    ws[f'A{row}'] = "Generated:"
    ws[f'A{row}'].font = FONT_BOLD
    ws[f'B{row}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    row += 1

    # Show corrections applied (if any)
    if corrections_applied > 0:
        ws[f'A{row}'] = "Corrections Applied:"
        ws[f'A{row}'].font = FONT_BOLD
        ws[f'B{row}'] = f"{corrections_applied} manual corrections from review"
        ws[f'B{row}'].font = Font(color="0000FF", italic=True)  # Blue text
        row += 1

    row += 1  # Extra space before stats

    # ---- Per-version stats (shown first so teachers can spot version-level issues) ----
    if is_multi_version:
        ws[f'A{row}'] = "Per-Version Statistics"
        ws[f'A{row}'].font = Font(size=13, bold=True)
        row += 1

        # Table header
        ver_headers = ["", "N Students", "Mean Score", "Mean %", "Std Dev",
                        "KR-20", "KR-21"]
        for col_idx, hdr in enumerate(ver_headers, start=1):
            ws.cell(row=row, column=col_idx, value=hdr)
        format_header_row(ws, row)
        row += 1

        for ver in versions:
            vs = version_stats[ver]
            ver_mean = vs['mean']
            ver_pct = (ver_mean / k * 100) if k > 0 else 0.0
            ver_kr20 = vs['kr20']
            ver_kr21 = vs['kr21']

            ws.cell(row=row, column=1, value=f"Version {ver}")
            ws.cell(row=row, column=1).font = FONT_BOLD
            ws.cell(row=row, column=2, value=vs['n_students'])
            ws.cell(row=row, column=3, value=f"{ver_mean:.2f}")
            ws.cell(row=row, column=4, value=f"{ver_pct:.1f}%")
            ws.cell(row=row, column=5, value=f"{vs['std']:.2f}")
            ws.cell(row=row, column=6, value=f"{ver_kr20:.3f}" if not np.isnan(ver_kr20) else "N/A")
            ws.cell(row=row, column=7, value=f"{ver_kr21:.3f}" if not np.isnan(ver_kr21) else "N/A")

            # Apply border to data cells
            for col_idx in range(1, len(ver_headers) + 1):
                ws.cell(row=row, column=col_idx).border = BORDER_THIN

            row += 1

        row += 1  # Space before amalgamated stats

    # ---- Amalgamated (overall) stats ----
    if is_multi_version:
        ws[f'A{row}'] = "Amalgamated Exam Statistics (All Versions)"
    else:
        ws[f'A{row}'] = "Overall Exam Statistics"
    ws[f'A{row}'].font = Font(size=13, bold=True)
    row += 1

    stats = [
        ("Number of Students", n_students),
        ("Number of Questions", k),
        ("Number of Versions", n_versions),
        ("Mean Score", f"{mean_total:.2f}"),
        ("Mean Percentage", f"{mean_total/k*100:.1f}%" if k > 0 else "N/A"),
        ("Standard Deviation", f"{std_total:.2f}"),
        ("KR-20 Reliability", f"{kr20_val:.3f}" if not np.isnan(kr20_val) else "N/A"),
        ("KR-21 Reliability", f"{kr21_val:.3f}" if not np.isnan(kr21_val) else "N/A"),
    ]

    for stat_name, stat_value in stats:
        ws[f'A{row}'] = stat_name
        ws[f'B{row}'] = stat_value
        row += 1

    # Reliability interpretation
    row += 1
    ws[f'A{row}'] = "Reliability Interpretation"
    ws[f'A{row}'].font = FONT_BOLD
    row += 1

    if not np.isnan(kr20_val):
        if kr20_val >= 0.80:
            ws[f'A{row}'] = "Excellent reliability (≥0.80)"
            ws[f'A{row}'].fill = COLOR_GOOD
        elif kr20_val >= 0.70:
            ws[f'A{row}'] = "Good reliability (0.70-0.80)"
            ws[f'A{row}'].fill = COLOR_GOOD
        elif kr20_val >= 0.60:
            ws[f'A{row}'] = "Acceptable reliability (0.60-0.70)"
            ws[f'A{row}'].fill = COLOR_WARNING
        else:
            ws[f'A{row}'] = "Poor reliability (<0.60) - exam needs work"
            ws[f'A{row}'].fill = COLOR_PROBLEM

    # Roster issues
    if orphan_scans or absent_students:
        row += 2
        ws[f'A{row}'] = "Roster Issues"
        ws[f'A{row}'].font = Font(size=14, bold=True, color="FF0000")
        row += 1

        if orphan_scans:
            ws[f'A{row}'] = f"⚠ {len(orphan_scans)} orphan scan(s)"
            ws[f'A{row}'].fill = COLOR_ORPHAN
            row += 1

            # List orphan scans
            ws[f'A{row}'] = "Orphan Scans (ID mismatch or fuzzy match):"
            ws[f'A{row}'].font = FONT_BOLD
            row += 1
            ws[f'A{row}'] = "Scanned ID"
            ws[f'B{row}'] = "Last Name"
            ws[f'C{row}'] = "First Name"
            ws[f'D{row}'] = "Match Type"
            ws[f'E{row}'] = "Possible Match"
            format_header_row(ws, row)
            row += 1

            for orphan in orphan_scans:
                ws[f'A{row}'] = orphan.get('ScannedID', '')
                ws[f'B{row}'] = orphan.get('LastName', '')
                ws[f'C{row}'] = orphan.get('FirstName', '')
                ws[f'D{row}'] = orphan.get('MatchType', 'no_match')
                ws[f'E{row}'] = orphan.get('PossibleMatch', '')
                row += 1
            row += 1

        if absent_students:
            ws[f'A{row}'] = f"⚠ {len(absent_students)} absent student(s)"
            ws[f'A{row}'].fill = COLOR_WARNING
            row += 1

            # List absent students
            ws[f'A{row}'] = "Absent Students:"
            ws[f'A{row}'].font = FONT_BOLD
            row += 1
            ws[f'A{row}'] = "Student ID"
            ws[f'B{row}'] = "Last Name"
            ws[f'C{row}'] = "First Name"
            format_header_row(ws, row)
            row += 1

            for student in absent_students:
                ws[f'A{row}'] = student['StudentID']
                ws[f'B{row}'] = student['LastName']
                ws[f'C{row}'] = student['FirstName']
                row += 1

    # Corrections detail (copy of the flagged XLSX rows)
    if corrections_detail is not None and not corrections_detail.empty:
        row += 2
        ws[f'A{row}'] = "Corrections Applied"
        ws[f'A{row}'].font = Font(size=14, bold=True, color="FF0000")
        row += 1

        # Write all columns from the flagged xlsx as-is
        flagged_cols = list(corrections_detail.columns)
        for col_idx, col_name in enumerate(flagged_cols, start=1):
            ws.cell(row=row, column=col_idx, value=str(col_name))
        format_header_row(ws, row)
        row += 1

        for _, corr_row in corrections_detail.iterrows():
            for col_idx, col_name in enumerate(flagged_cols, start=1):
                val = corr_row.get(col_name, '')
                # Clean up NaN display
                if pd.isna(val):
                    val = ''
                ws.cell(row=row, column=col_idx, value=val)
            row += 1

    auto_size_columns(ws)


def create_version_tab(
    wb, df_full, students_df, version, item_cols, key_series,
    difficulty, pb_vals, roster_df, orphan_scans
):
    """Create a tab for a specific exam version."""
    ws = wb.create_sheet(f"Version {version}")

    # Filter for this version (handle both 'Version' and 'version')
    version_col = 'version' if 'version' in df_full.columns else 'Version' if 'Version' in df_full.columns else None

    if version_col:
        version_mask = df_full[version_col].astype(str).str.strip() == str(version).strip()
        df_version = df_full[version_mask].copy()
    else:
        df_version = df_full.copy()

    # Get KEY row for this version
    key_row_data = df_version[df_version.apply(
        lambda row: any(str(cell).strip().upper() == 'KEY' for cell in row), axis=1
    )]

    # Get student rows
    student_rows = df_version[~df_version.apply(
        lambda row: any(str(cell).strip().upper() == 'KEY' for cell in row), axis=1
    )]

    # Determine columns to display in the desired order:
    # LastName, FirstName, StudentID, Issue, correct, incorrect, blank, multi, percent, Version, Q1, Q2, ...
    display_cols = []

    # Identity columns first (check both cases)
    for col_variants in [('lastname', 'LastName'), ('firstname', 'FirstName'), ('studentid', 'StudentID')]:
        for variant in col_variants:
            if variant in df_version.columns:
                display_cols.append(variant)
                break

    # Add Issue column (will be computed)
    display_cols.append('Issue')

    # Score columns (check both cases)
    for col in ['correct', 'incorrect', 'blank', 'multi', 'percent']:
        if col in df_version.columns:
            display_cols.append(col)
        elif col.capitalize() in df_version.columns:
            display_cols.append(col.capitalize())

    # Version column (check both cases)
    if 'version' in df_version.columns:
        display_cols.append('version')
    elif 'Version' in df_version.columns:
        display_cols.append('Version')

    # Question columns
    display_cols.extend(item_cols)

    # Build column index map for quick lookup
    col_idx_map = {col: idx + 1 for idx, col in enumerate(display_cols)}

    # Write header
    for col_idx, col_name in enumerate(display_cols, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)
    format_header_row(ws, 1)

    # Get KEY answers for this version
    key_answers = {}
    if not key_row_data.empty:
        key_row = key_row_data.iloc[0]
        for col in item_cols:
            key_answers[col] = str(key_row.get(col, '')).strip().upper()

    # Write student rows
    row_num = 2
    for _, student_row in student_rows.iterrows():
        # Determine issues for this student
        issues = []

        # Check for blank/multi answers
        blank_count = int(student_row.get('blank', 0))
        multi_count = int(student_row.get('multi', 0))
        if blank_count > 0:
            issues.append(f"{blank_count} blank")
        if multi_count > 0:
            issues.append(f"{multi_count} multi")

        # Check roster matching (if available)
        if roster_df is not None:
            student_id = str(student_row.get('StudentID', '')).strip()
            if orphan_scans:
                # Check if this student is an orphan
                for orphan in orphan_scans:
                    if str(orphan.get('ScannedID', '')).strip() == student_id:
                        match_type = orphan.get('MatchType', 'no_match')
                        if match_type == 'no_match':
                            issues.append("ID mismatch")
                        elif match_type == 'probable':
                            issues.append("Fuzzy match")
                        break

        issue_text = "; ".join(issues) if issues else ""

        # Write all column values
        for col_idx, col_name in enumerate(display_cols, start=1):
            if col_name == 'Issue':
                cell = ws.cell(row=row_num, column=col_idx, value=issue_text)
                if issue_text:
                    cell.fill = COLOR_WARNING
            else:
                value = student_row.get(col_name, '')
                # Handle NaN/None from pandas (blank CSV cells become NaN)
                if pd.isna(value):
                    student_answer = ''
                else:
                    student_answer = str(value).strip().upper()

                # Check for blank or multi-answer cells in question columns
                if col_name in item_cols:
                    # Detect blank answers (empty, NaN, whitespace, or common blank indicators)
                    is_blank = (
                        not student_answer
                        or student_answer in ('', 'NAN', 'BLANK', 'NONE', '?')
                    )

                    # Detect multi-answer (contains comma, multiple letters, or "MULTI" indicator)
                    is_multi = (
                        not is_blank and (
                            ',' in student_answer or
                            student_answer == 'MULTI' or
                            (len(student_answer) > 1 and student_answer not in ('BLANK', 'NONE', 'MULTI', 'NAN'))
                        )
                    )

                    # Apply formatting based on answer type
                    if is_blank:
                        cell = ws.cell(row=row_num, column=col_idx, value="BLANK")
                        cell.fill = COLOR_BLANK
                        cell.alignment = Alignment(horizontal='center')
                    elif is_multi:
                        cell = ws.cell(row=row_num, column=col_idx, value=value)
                        cell.fill = COLOR_MULTI
                        cell.alignment = Alignment(horizontal='center')
                    else:
                        cell = ws.cell(row=row_num, column=col_idx, value=value)
                        cell.alignment = Alignment(horizontal='center')
                        # Highlight incorrect answers in light red (only if not blank/multi)
                        if key_answers.get(col_name):
                            correct_answer = key_answers[col_name]
                            if student_answer and student_answer != correct_answer:
                                cell.fill = PatternFill(start_color="FFD7D7", end_color="FFD7D7", fill_type="solid")
                else:
                    # Non-question columns - just write the value
                    ws.cell(row=row_num, column=col_idx, value=value)

        row_num += 1

    # Add KEY answer row before statistics
    row_num += 1
    ws.cell(row=row_num, column=1, value="KEY")
    ws.cell(row=row_num, column=1).font = FONT_BOLD
    for col_name in item_cols:
        col_idx = col_idx_map.get(col_name)
        if col_idx and col_name in key_answers:
            ws.cell(row=row_num, column=col_idx, value=key_answers[col_name])
    format_key_row(ws, row_num)
    row_num += 1

    # Add item statistics rows
    ws.cell(row=row_num, column=1, value="% Correct")
    ws.cell(row=row_num, column=1).font = FONT_BOLD
    for col_name in item_cols:
        col_idx = col_idx_map.get(col_name)
        if col_idx:
            pct = difficulty[col_name] * 100 if not np.isnan(difficulty[col_name]) else 0
            ws.cell(row=row_num, column=col_idx, value=f"{pct:.1f}%")
    row_num += 1

    ws.cell(row=row_num, column=1, value="Point-Biserial")
    ws.cell(row=row_num, column=1).font = FONT_BOLD
    for col_name in item_cols:
        col_idx = col_idx_map.get(col_name)
        if col_idx:
            pb = pb_vals[col_name]
            if not np.isnan(pb):
                cell = ws.cell(row=row_num, column=col_idx, value=f"{pb:.3f}")
                # Color code based on quality
                if pb >= 0.20:
                    cell.fill = COLOR_GOOD
                elif pb >= 0.10:
                    cell.fill = COLOR_WARNING
                else:
                    cell.fill = COLOR_PROBLEM
    row_num += 1

    ws.cell(row=row_num, column=1, value="Item Quality")
    ws.cell(row=row_num, column=1).font = FONT_BOLD
    for col_name in item_cols:
        col_idx = col_idx_map.get(col_name)
        if col_idx:
            pb = pb_vals[col_name]
            if not np.isnan(pb):
                if pb >= 0.20:
                    quality = "✓ Good"
                    fill = COLOR_GOOD
                elif pb >= 0.10:
                    quality = "⚠ Review"
                    fill = COLOR_WARNING
                else:
                    quality = "✗ Problem"
                    fill = COLOR_PROBLEM
                cell = ws.cell(row=row_num, column=col_idx, value=quality)
                cell.fill = fill

    auto_size_columns(ws)


def create_class_scores_tab(wb, df_full, item_cols, k):
    """
    Create a "Class Scores" tab with all students sorted alphabetically.

    This provides a simple roster view suitable for pasting into gradebooks:
    - LastName, FirstName, StudentID, RawScore, Percent, Version
    - Sorted by LastName first, then FirstName
    """
    ws = wb.create_sheet("Class Scores")

    # Determine column names (handle case variations) - do this first for filtering
    lastname_col = 'lastname' if 'lastname' in df_full.columns else 'LastName' if 'LastName' in df_full.columns else None
    firstname_col = 'firstname' if 'firstname' in df_full.columns else 'FirstName' if 'FirstName' in df_full.columns else None
    studentid_col = 'studentid' if 'studentid' in df_full.columns else 'StudentID' if 'StudentID' in df_full.columns else None
    correct_col = 'correct' if 'correct' in df_full.columns else 'Correct' if 'Correct' in df_full.columns else None
    percent_col = 'percent' if 'percent' in df_full.columns else 'Percent' if 'Percent' in df_full.columns else None
    version_col = 'version' if 'version' in df_full.columns else 'Version' if 'Version' in df_full.columns else None

    # Filter out KEY rows and header rows that got mixed in
    def is_non_student_row(row):
        # Check if any identity field contains "KEY" or is a header value
        header_values = {'KEY', 'LASTNAME', 'FIRSTNAME', 'STUDENTID', 'STUDENT_ID', 'LAST_NAME', 'FIRST_NAME'}
        for col in [lastname_col, firstname_col, studentid_col]:
            if col and col in row.index:
                val = str(row[col]).strip().upper()
                if val in header_values:
                    return True
        return False

    df_students = df_full[~df_full.apply(is_non_student_row, axis=1)].copy()

    # Sort by last name first, then first name (case-insensitive)
    sort_cols = []
    if lastname_col:
        df_students['_sort_last'] = df_students[lastname_col].astype(str).str.upper()
        sort_cols.append('_sort_last')
    if firstname_col:
        df_students['_sort_first'] = df_students[firstname_col].astype(str).str.upper()
        sort_cols.append('_sort_first')

    if sort_cols:
        df_students = df_students.sort_values(sort_cols)
        # Drop sort columns
        df_students = df_students.drop(columns=[c for c in sort_cols if c in df_students.columns])

    # Write header
    headers = ['Last Name', 'First Name', 'Student ID', 'Raw Score', 'Percent', 'Version']
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    format_header_row(ws, 1)

    # Write student rows
    row_num = 2
    for _, row in df_students.iterrows():
        # Last Name
        ws.cell(row=row_num, column=1, value=row.get(lastname_col, '') if lastname_col else '')

        # First Name
        ws.cell(row=row_num, column=2, value=row.get(firstname_col, '') if firstname_col else '')

        # Student ID
        ws.cell(row=row_num, column=3, value=row.get(studentid_col, '') if studentid_col else '')

        # Raw Score (correct answers)
        raw_score = row.get(correct_col, '') if correct_col else ''
        ws.cell(row=row_num, column=4, value=raw_score)

        # Percent
        percent = row.get(percent_col, '') if percent_col else ''
        if percent and str(percent).strip():
            try:
                pct_val = float(percent)
                ws.cell(row=row_num, column=5, value=f"{pct_val:.1f}%")
            except (ValueError, TypeError):
                ws.cell(row=row_num, column=5, value=percent)
        else:
            ws.cell(row=row_num, column=5, value='')

        # Version
        ws.cell(row=row_num, column=6, value=row.get(version_col, '') if version_col else '')

        row_num += 1

    # Add summary at bottom
    row_num += 1
    ws.cell(row=row_num, column=1, value="Total Students:")
    ws.cell(row=row_num, column=1).font = FONT_BOLD
    ws.cell(row=row_num, column=2, value=len(df_students))

    auto_size_columns(ws)
