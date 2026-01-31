"""Excel file processing utilities with security validation.

Provides utilities for safely processing Excel files with configurable
security limits to prevent resource exhaustion attacks.
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import TYPE_CHECKING, Any

from pydantic import BaseModel, Field, model_validator

if TYPE_CHECKING:
    from collections.abc import Callable


# ============================================================================
# Exceptions (defined here to avoid circular imports)
# ============================================================================


class ExcelSecurityError(Exception):
    """Exception raised when Excel file security validation fails."""

    def __init__(
        self,
        message: str,
        violation_type: str | None = None,
        limit: int | None = None,
        actual: int | None = None,
    ):
        super().__init__(message)
        self.message = message
        self.violation_type = violation_type
        self.limit = limit
        self.actual = actual


class ExcelParsingError(Exception):
    """Exception raised when Excel file parsing encounters errors."""

    def __init__(
        self,
        message: str,
        file_path: str | None = None,
        original_error: Exception | None = None,
    ):
        super().__init__(message)
        self.message = message
        self.file_path = file_path
        self.original_error = original_error


# ============================================================================
# Configuration Model
# ============================================================================


class ExcelSecurityConfig(BaseModel):
    """Security configuration for Excel file processing.

    Defines essential security limits for Excel file processing to prevent
    resource exhaustion attacks.
    """

    max_file_size_mb: int = Field(default=10, ge=1, le=1000)
    max_rows: int = Field(default=100000, ge=1, le=100000)
    max_columns: int = Field(default=50, ge=1, le=16384)
    max_memory_usage_mb: int = Field(default=30, ge=1, le=1000)
    max_filename_length: int = Field(default=255, ge=1, le=1000)
    max_column_name_length: int = Field(default=100, ge=1, le=500)
    max_metadata_value_length: int = Field(default=1000, ge=1, le=10000)
    validation_check_interval: int = Field(default=1000, ge=100, le=10000)

    model_config = {'validate_assignment': True, 'extra': 'forbid'}

    @model_validator(mode='after')
    def validate_resource_limits(self) -> 'ExcelSecurityConfig':
        """Validate that resource limits are reasonable."""
        estimated_cells = self.max_rows * self.max_columns
        if estimated_cells > 50_000_000:
            raise ValueError(
                f'Combination of max_rows ({self.max_rows}) and max_columns ({self.max_columns}) '
                f'would allow too many cells ({estimated_cells:,})'
            )
        return self

    @property
    def max_file_size_bytes(self) -> int:
        """Get maximum file size in bytes."""
        return self.max_file_size_mb * 1024 * 1024

    @property
    def max_memory_usage_bytes(self) -> int:
        """Get maximum memory usage in bytes."""
        return self.max_memory_usage_mb * 1024 * 1024

    @classmethod
    def from_action_config(cls, action_config: dict[str, Any] | None) -> 'ExcelSecurityConfig':
        """Create ExcelSecurityConfig from plugin action configuration."""
        if not action_config or 'excel_config' not in action_config:
            return cls()
        excel_config = action_config['excel_config']
        return cls(**{k: v for k, v in excel_config.items() if k in cls.model_fields})


# ============================================================================
# Utility Classes
# ============================================================================


class PathAwareJSONEncoder(json.JSONEncoder):
    """Custom JSON encoder that handles Path objects and datetime objects.

    Extends the default JSON encoder to properly serialize Path objects
    and datetime objects that are commonly used in upload operations.

    Supported object types:
    - Path objects (converts to string using __fspath__ or as_posix)
    - Datetime objects (converts using isoformat)
    - All other standard JSON-serializable types

    Example:
        >>> from pathlib import Path
        >>> from datetime import datetime
        >>> data = {"path": Path("/tmp/file.txt"), "timestamp": datetime.now()}
        >>> json.dumps(data, cls=PathAwareJSONEncoder)
        '{"path": "/tmp/file.txt", "timestamp": "2023-01-01T12:00:00"}'
    """

    def default(self, obj: Any) -> Any:
        """Convert non-serializable objects to JSON-compatible types."""
        if hasattr(obj, '__fspath__'):
            return obj.__fspath__()
        elif hasattr(obj, 'as_posix'):
            return obj.as_posix()
        elif hasattr(obj, 'isoformat'):
            return obj.isoformat()
        return super().default(obj)


class ExcelMetadataUtils:
    """Utility class for Excel metadata processing with security validation.

    Provides methods for validating Excel file constraints and processing
    metadata values with length limits.

    Attributes:
        config: ExcelSecurityConfig instance defining security limits.

    Example:
        >>> config = ExcelSecurityConfig(max_filename_length=100)
        >>> utils = ExcelMetadataUtils(config)
        >>> utils.is_valid_filename_length("short_name.xlsx")
        True
        >>> utils.is_valid_filename_length("a" * 200 + ".xlsx")
        False
    """

    def __init__(self, config: ExcelSecurityConfig):
        """Initialize with Excel security configuration.

        Args:
            config: ExcelSecurityConfig instance defining security limits.
        """
        self.config = config

    def is_valid_filename_length(self, filename: str) -> bool:
        """Check if filename length is within limits.

        Args:
            filename: The filename to validate.

        Returns:
            True if filename length is within the configured limit.
        """
        return len(filename) <= self.config.max_filename_length

    def validate_and_truncate_string(self, value: str, max_length: int) -> str:
        """Validate and truncate string to maximum length.

        Args:
            value: The string value to process.
            max_length: Maximum allowed length.

        Returns:
            The processed string, stripped and truncated if necessary.
        """
        if not isinstance(value, str):
            value = str(value)

        # Strip whitespace
        value = value.strip()

        # Truncate if too long
        if len(value) > max_length:
            value = value[:max_length]

        return value

    def is_valid_column_name(self, column_name: str) -> bool:
        """Check if column name is valid.

        Args:
            column_name: The column name to validate.

        Returns:
            True if column name is non-empty and within length limit.
        """
        if not column_name or not isinstance(column_name, str):
            return False
        return len(column_name.strip()) <= self.config.max_column_name_length

    def is_valid_metadata_value(self, value: str | None) -> bool:
        """Check if metadata value is valid.

        Args:
            value: The metadata value to validate.

        Returns:
            True if value is None or within length limit.
        """
        if value is None:
            return True
        if not isinstance(value, str):
            value = str(value)
        return len(value) <= self.config.max_metadata_value_length


def validate_excel_file_security(
    file_path: Path,
    config: ExcelSecurityConfig,
    *,
    on_log: Callable[[str, Any], None] | None = None,
) -> None:
    """Validate Excel file against security constraints.

    Checks file size and estimates memory usage before allowing processing.

    Args:
        file_path: Path to the Excel file.
        config: Security configuration with limits.
        on_log: Optional callback for logging events.

    Raises:
        ExcelSecurityError: If any security constraint is violated.
        FileNotFoundError: If the file doesn't exist.

    Example:
        >>> config = ExcelSecurityConfig(max_file_size_mb=10)
        >>> validate_excel_file_security(Path("data.xlsx"), config)
    """
    if not file_path.exists():
        raise FileNotFoundError(f'Excel file not found: {file_path}')

    file_size = file_path.stat().st_size

    if on_log:
        on_log('security_validation_started', {'file_size': file_size})

    # Check file size
    if file_size > config.max_file_size_bytes:
        raise ExcelSecurityError(
            f'Excel file size ({file_size:,} bytes) exceeds limit ({config.max_file_size_bytes:,} bytes)',
            violation_type='file_size',
            limit=config.max_file_size_bytes,
            actual=file_size,
        )

    # Estimate memory usage (file size * 3 is a conservative estimate)
    estimated_memory = file_size * 3

    if on_log:
        on_log('memory_estimation', {'file_size': file_size, 'estimated': estimated_memory})

    if estimated_memory > config.max_memory_usage_bytes:
        raise ExcelSecurityError(
            f'Estimated memory usage ({estimated_memory:,} bytes) exceeds limit '
            f'({config.max_memory_usage_bytes:,} bytes)',
            violation_type='memory',
            limit=config.max_memory_usage_bytes,
            actual=estimated_memory,
        )


def load_excel_metadata(
    file_path: Path,
    config: ExcelSecurityConfig,
    *,
    sheet_name: str | int = 0,
    filename_column: str = 'filename',
    on_log: Callable[[str, Any], None] | None = None,
) -> dict[str, dict[str, Any]]:
    """Load metadata from an Excel file with security validation.

    Reads an Excel file and extracts metadata as a dictionary keyed by filename.
    Performs security validation before processing.

    Args:
        file_path: Path to the Excel file.
        config: Security configuration with limits.
        sheet_name: Sheet name or index to read. Default is first sheet.
        filename_column: Column name containing filenames. Default is 'filename'.
        on_log: Optional callback for logging events.

    Returns:
        Dictionary mapping filenames to their metadata dictionaries.

    Raises:
        ExcelSecurityError: If security constraints are violated.
        ExcelParsingError: If the file cannot be parsed.
        ImportError: If openpyxl is not installed.

    Example:
        >>> config = ExcelSecurityConfig()
        >>> metadata = load_excel_metadata(Path("metadata.xlsx"), config)
        >>> metadata["image_001.jpg"]
        {'label': 'cat', 'confidence': 0.95}
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.utils.exceptions import InvalidFileException
    except ImportError as e:
        raise ImportError(
            'openpyxl is required for Excel metadata processing. Install it with: pip install openpyxl'
        ) from e

    # Validate security constraints
    validate_excel_file_security(file_path, config, on_log=on_log)

    utils = ExcelMetadataUtils(config)
    metadata: dict[str, dict[str, Any]] = {}

    try:
        # Load workbook in read-only mode for better memory efficiency
        workbook = load_workbook(file_path, read_only=True, data_only=True)

        if on_log:
            on_log('workbook_loaded', {'file': str(file_path)})

        # Get the specified sheet
        if isinstance(sheet_name, int):
            if sheet_name >= len(workbook.sheetnames):
                raise ExcelParsingError(
                    f'Sheet index {sheet_name} out of range. Available sheets: {len(workbook.sheetnames)}',
                    file_path=str(file_path),
                )
            sheet = workbook[workbook.sheetnames[sheet_name]]
        else:
            if sheet_name not in workbook.sheetnames:
                raise ExcelParsingError(
                    f"Sheet '{sheet_name}' not found. Available sheets: {workbook.sheetnames}",
                    file_path=str(file_path),
                )
            sheet = workbook[sheet_name]

        # Read header row
        rows = sheet.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            raise ExcelParsingError(
                'Excel file is empty or has no header row',
                file_path=str(file_path),
            )

        # Validate header
        if header is None or all(h is None for h in header):
            raise ExcelParsingError(
                'Excel file has empty header row',
                file_path=str(file_path),
            )

        # Find filename column index
        header_list = [str(h).strip() if h is not None else '' for h in header]

        if filename_column not in header_list:
            raise ExcelParsingError(
                f"Filename column '{filename_column}' not found in header. Available columns: {header_list}",
                file_path=str(file_path),
            )

        filename_idx = header_list.index(filename_column)

        # Validate column names
        valid_columns = []
        for i, col_name in enumerate(header_list):
            if col_name and utils.is_valid_column_name(col_name):
                valid_columns.append((i, col_name))

        # Check column count
        if len(valid_columns) > config.max_columns:
            raise ExcelSecurityError(
                f'Too many columns ({len(valid_columns)}) exceeds limit ({config.max_columns})',
                violation_type='columns',
                limit=config.max_columns,
                actual=len(valid_columns),
            )

        # Process rows
        row_count = 0
        for row in rows:
            row_count += 1

            # Check row limit
            if row_count > config.max_rows:
                raise ExcelSecurityError(
                    f'Too many rows ({row_count}) exceeds limit ({config.max_rows})',
                    violation_type='rows',
                    limit=config.max_rows,
                    actual=row_count,
                )

            # Skip empty rows
            if row is None or all(cell is None for cell in row):
                continue

            # Get filename
            filename_value = row[filename_idx] if filename_idx < len(row) else None
            if filename_value is None:
                continue

            filename = str(filename_value).strip()
            if not filename:
                continue

            # Check filename length
            if not utils.is_valid_filename_length(filename):
                if on_log:
                    on_log('filename_too_long', {'filename': filename[:50]})
                continue

            # Extract metadata from other columns
            row_metadata: dict[str, Any] = {}
            for col_idx, col_name in valid_columns:
                if col_idx == filename_idx:
                    continue

                if col_idx < len(row):
                    value = row[col_idx]
                    if value is not None:
                        # Validate and convert value
                        str_value = str(value)
                        if utils.is_valid_metadata_value(str_value):
                            row_metadata[col_name] = value

            metadata[filename] = row_metadata

            # Periodic validation check
            if row_count % config.validation_check_interval == 0:
                if on_log:
                    on_log('processing_progress', {'rows': row_count})

        workbook.close()

        if on_log:
            on_log('metadata_loaded', {'file_count': len(metadata)})

        return metadata

    except InvalidFileException as e:
        raise ExcelParsingError(
            f'Invalid Excel file format: {e}',
            file_path=str(file_path),
            original_error=e,
        )
    except ExcelSecurityError:
        raise
    except ExcelParsingError:
        raise
    except Exception as e:
        raise ExcelParsingError(
            f'Unexpected error reading Excel file: {e}',
            file_path=str(file_path),
            original_error=e,
        )


__all__ = [
    # Exceptions
    'ExcelSecurityError',
    'ExcelParsingError',
    # Configuration
    'ExcelSecurityConfig',
    # Utilities
    'PathAwareJSONEncoder',
    'ExcelMetadataUtils',
    'validate_excel_file_security',
    'load_excel_metadata',
]
