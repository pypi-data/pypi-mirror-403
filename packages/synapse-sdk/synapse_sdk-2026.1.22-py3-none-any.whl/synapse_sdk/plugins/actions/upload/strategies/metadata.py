"""Metadata extraction strategies for upload operations.

Provides strategies for extracting and validating metadata:
    - ExcelMetadataStrategy: Extract metadata from Excel files
    - NoneMetadataStrategy: No-op strategy for uploads without metadata
"""

from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Any

from synapse_sdk.plugins.actions.upload.strategies.base import (
    MetadataStrategy,
    ValidationResult,
)
from synapse_sdk.utils.excel import (
    ExcelMetadataUtils,
    ExcelParsingError,
    ExcelSecurityConfig,
    ExcelSecurityError,
)


class ExcelMetadataStrategy(MetadataStrategy):
    """Excel metadata extraction strategy.

    Extracts metadata from Excel files with security validation.
    Expects first column to be 'filename' or 'file_name'.

    Example:
        >>> strategy = ExcelMetadataStrategy()
        >>> metadata = strategy.extract(Path("metadata.xlsx"))
        >>> metadata
        {'image_001.jpg': {'label': 'cat', 'confidence': '0.95'}}
    """

    def __init__(self, config: ExcelSecurityConfig | None = None):
        """Initialize with optional security configuration.

        Args:
            config: ExcelSecurityConfig instance. Uses defaults if not provided.
        """
        self.config = config or ExcelSecurityConfig()
        self.utils = ExcelMetadataUtils(self.config)

    def extract(self, source_path: Path) -> dict[str, dict[str, Any]]:
        """Extract metadata from Excel file.

        Args:
            source_path: Path to Excel file.

        Returns:
            Dictionary mapping filenames to their metadata.

        Raises:
            ExcelSecurityError: If file violates security constraints.
            ExcelParsingError: If file cannot be parsed.
        """
        try:
            # Import openpyxl here to allow graceful failure
            from openpyxl import load_workbook
            from openpyxl.utils.exceptions import InvalidFileException
        except ImportError as e:
            raise ImportError(
                'openpyxl is required for Excel metadata processing. Install it with: pip install openpyxl'
            ) from e

        try:
            excel_stream = self._prepare_excel_file(source_path)
            workbook = load_workbook(excel_stream, read_only=True, data_only=True)

            try:
                return self._process_worksheet(workbook.active)
            finally:
                workbook.close()

        except ExcelSecurityError:
            raise
        except ExcelParsingError:
            raise
        except InvalidFileException as e:
            raise ExcelParsingError(
                f'Invalid Excel file format: {e}',
                file_path=str(source_path),
                original_error=e,
            )
        except MemoryError:
            raise ExcelSecurityError(
                'Excel file exceeds memory limits',
                violation_type='memory',
            )
        except (OSError, IOError) as e:
            raise ExcelParsingError(
                f'File access error: {e}',
                file_path=str(source_path),
                original_error=e,
            )
        except Exception as e:
            # Handle ZIP errors (Excel files are ZIP archives)
            if 'zip file' in str(e).lower():
                raise ExcelParsingError(
                    f'Invalid Excel file format: {e}',
                    file_path=str(source_path),
                    original_error=e,
                )
            raise ExcelParsingError(
                f'Unexpected error: {e}',
                file_path=str(source_path),
                original_error=e,
            )

    def validate(self, metadata: dict[str, dict[str, Any]]) -> ValidationResult:
        """Validate extracted metadata.

        Args:
            metadata: Metadata dictionary to validate.

        Returns:
            ValidationResult with validation status and any errors.
        """
        errors: list[str] = []

        if not isinstance(metadata, dict):
            errors.append('Metadata must be a dictionary')
            return ValidationResult(valid=False, errors=errors)

        for file_name, file_metadata in metadata.items():
            if not isinstance(file_metadata, dict):
                errors.append(f"Metadata for file '{file_name}' must be a dictionary")
                continue

            if not self.utils.is_valid_filename_length(file_name):
                errors.append(f"Filename '{file_name}' exceeds maximum length")

        return ValidationResult(valid=len(errors) == 0, errors=errors)

    def _validate_security(self, excel_path: Path) -> None:
        """Validate Excel file against security constraints."""
        file_size = excel_path.stat().st_size

        if file_size > self.config.max_file_size_bytes:
            raise ExcelSecurityError(
                f'Excel file too large: {file_size:,} bytes (max: {self.config.max_file_size_bytes:,})',
                violation_type='file_size',
                limit=self.config.max_file_size_bytes,
                actual=file_size,
            )

        # Estimate memory usage (file size * 3)
        estimated_memory = file_size * 3
        if estimated_memory > self.config.max_memory_usage_bytes:
            raise ExcelSecurityError(
                f'Excel file may consume too much memory: ~{estimated_memory:,} bytes '
                f'(max: {self.config.max_memory_usage_bytes:,})',
                violation_type='memory',
                limit=self.config.max_memory_usage_bytes,
                actual=estimated_memory,
            )

    def _prepare_excel_file(self, excel_path: Path) -> BytesIO:
        """Prepare Excel file for processing."""
        self._validate_security(excel_path)
        excel_bytes = excel_path.read_bytes()
        return BytesIO(excel_bytes)

    def _process_worksheet(self, worksheet) -> dict[str, dict[str, Any]]:
        """Process Excel worksheet and extract metadata."""
        if worksheet is None:
            raise ExcelParsingError('Excel file has no active worksheet')

        metadata: dict[str, dict[str, Any]] = {}
        headers: tuple | None = None
        data_row_count = 0

        for row_idx, row in enumerate(worksheet.iter_rows(values_only=True)):
            # Skip empty rows
            if not row or not row[0] or str(row[0]).strip() == '':
                continue

            # Process header row
            if row_idx == 0:
                headers = self._process_headers(row)
                continue

            if headers is None:
                raise ExcelParsingError('Excel file missing header row')

            data_row_count += 1

            # Check row limit
            if data_row_count > self.config.max_rows:
                raise ExcelSecurityError(
                    f'Too many rows: {data_row_count} (max: {self.config.max_rows})',
                    violation_type='rows',
                    limit=self.config.max_rows,
                    actual=data_row_count,
                )

            # Process data row
            row_result = self._process_data_row(row, headers)
            if row_result:
                metadata.update(row_result)

        return metadata

    def _process_headers(self, headers: tuple) -> tuple:
        """Process and validate header row."""
        if len(headers) < 2:
            raise ExcelParsingError('Excel file must have at least 2 columns (filename and metadata)')

        # Validate first column is filename
        first_header = str(headers[0]).strip().lower() if headers[0] else ''
        valid_headers = {'filename', 'file_name'}

        if first_header not in valid_headers:
            raise ExcelParsingError(
                f'First column header must be "filename" or "file_name", '
                f'got: "{headers[0]}". Valid options: {", ".join(valid_headers)}'
            )

        # Check column count
        if len(headers) > self.config.max_columns:
            raise ExcelSecurityError(
                f'Too many columns: {len(headers)} (max: {self.config.max_columns})',
                violation_type='columns',
                limit=self.config.max_columns,
                actual=len(headers),
            )

        return headers

    def _process_data_row(
        self,
        row: tuple,
        headers: tuple,
    ) -> dict[str, dict[str, Any]] | None:
        """Process a single data row."""
        if not row[0] or str(row[0]).strip() == '':
            return None

        file_name = str(row[0]).strip()

        # Validate filename length
        if not self.utils.is_valid_filename_length(file_name):
            return None

        # Extract metadata from remaining columns
        file_metadata: dict[str, Any] = {}

        for i, value in enumerate(row[1:], start=1):
            if i >= len(headers):
                continue

            header_value = headers[i]
            column_name = str(header_value).strip() if header_value else f'column_{i}'

            # Truncate column name if needed
            column_name = self.utils.validate_and_truncate_string(
                column_name,
                self.config.max_column_name_length,
            )

            # Convert value to string
            value_str = '' if value is None else str(value)
            value_str = self.utils.validate_and_truncate_string(
                value_str,
                self.config.max_metadata_value_length,
            )

            file_metadata[column_name] = value_str

        return {file_name: file_metadata} if file_metadata else None


class NoneMetadataStrategy(MetadataStrategy):
    """No-op metadata strategy for uploads without metadata.

    Use this strategy when no metadata extraction is needed.

    Example:
        >>> strategy = NoneMetadataStrategy()
        >>> metadata = strategy.extract(Path("anything"))  # Returns empty dict
        >>> strategy.validate({})  # Always valid
    """

    def extract(self, source_path: Path) -> dict[str, dict[str, Any]]:
        """Return empty metadata.

        Args:
            source_path: Ignored.

        Returns:
            Empty dictionary.
        """
        return {}

    def validate(self, metadata: dict[str, dict[str, Any]]) -> ValidationResult:
        """Always return valid result.

        Args:
            metadata: Ignored.

        Returns:
            ValidationResult with valid=True.
        """
        return ValidationResult(valid=True)


__all__ = [
    'ExcelMetadataStrategy',
    'NoneMetadataStrategy',
]
