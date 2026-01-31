#!/usr/bin/env python3

from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles import Font


class ExcelFormat:
    def __init__(self, ws):
        self.ws = ws
        
    def set_column(self):
        for column in self.ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 10)
            self.ws.column_dimensions[column[0].column_letter].width = adjusted_width
            
    def set_rows_center(self):
        # 将所有单元格的文字居中
        for row in self.ws.iter_rows():
            
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        for row in self.ws.iter_rows():
            self.ws.row_dimensions[row[0].row].height = 24
    
    def set_freeze_first_row(self):
        """设置首行锁定/冻结首行"""
        # 冻结首行，从第二行开始滚动
        self.ws.freeze_panes = 'A2'
    
    def set_freeze_first_column(self):
        """设置首列锁定/冻结首列"""
        # 冻结首列，从第二列开始滚动
        self.ws.freeze_panes = 'B1'
    
    def set_first_row_bold_color(self, font_color='FF0000FF'):
        """设置首行字体加粗并改变颜色
        
        Args:
            font_color (str): 字体颜色的十六进制代码，默认为蓝色(FF0000FF)
                            格式：AARRGGBB 或 RRGGBB
                            例如：'FF0000FF'(蓝色), 'FFFF0000'(红色), 'FF008000'(绿色)
        """
        # 遍历首行的所有单元格
        for cell in self.ws[1]:
            if cell.value is not None:  # 只对有内容的单元格设置样式
                # 设置字体为粗体并改变颜色
                cell.font = Font(bold=True, color=font_color)
    
    def set_freeze_first_row_and_column(self):
        """同时冻结首行和首列"""
        # 冻结首行首列，从第二行第二列开始滚动
        self.ws.freeze_panes = 'B2'
    
    