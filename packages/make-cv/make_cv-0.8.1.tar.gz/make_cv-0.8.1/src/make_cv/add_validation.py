#!/usr/bin/env python3

from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import quote_sheetname
from openpyxl.worksheet.datavalidation import DataValidation

def pa(): 
	personal_awards = "Awards/personal awards data.xlsx"
	wb = load_workbook(personal_awards)
	ws = wb['Data']
	
	dvList = DataValidation(type="list",formula1="Notes!$A$10:$A$13",allow_blank=False)
	ws.add_data_validation(dvList)
	dvList.add('B2:B1048576') # This is the same as for the whole of column B
	
	dvYear = DataValidation(type="whole",
						operator="between",
						formula1=1900,formula2=2100)
	ws.add_data_validation(dvYear)
	dvYear.add('C2:C1048576') # This is the same as for the whole of column B
	
	wb.save(personal_awards)
	
def sa(): 
	student_awards = "Awards/student awards data.xlsx"
	wb = load_workbook(student_awards)
	ws = wb['Data']
	
	dvList1 = DataValidation(type="list",formula1="Notes!$A$7:$A$9",allow_blank=False)
	ws.add_data_validation(dvList1)
	dvList1.add('D2:D1048576') # This is the same as for the whole of column D

	dvList2 = DataValidation(type="list",formula1="Notes!$B$7:$B$8",allow_blank=False)
	ws.add_data_validation(dvList2)
	dvList2.add('E2:E1048576') # This is the same as for the whole of column D

	dvYear = DataValidation(type="whole",
						operator="between",
						formula1=1900,formula2=2100,allow_blank=False)
	ws.add_data_validation(dvYear)
	dvYear.add('F2:F1048576') # This is the same as for the whole of column B
	
	dvD = DataValidation(type="decimal",
                    operator="greaterThan",
                    formula1=0,allow_blank=True)
	ws.add_data_validation(dvD)
	dvD.add('C2:C1048576') # This is the same as for the whole of column B

	wb.save(student_awards)
	

def pg():
	PandG = "Proposals & Grants/proposals & grants.xlsx"
	wb = load_workbook(PandG)
	ws = wb['Data']

	dvYN = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
	ws.add_data_validation(dvYN)
	dvYN.add('F2:F1048576') # This is the same as for the whole of column B
	
	dvD = DataValidation(type="decimal",
                    operator="greaterThan",
                    formula1=0,allow_blank=True)
	ws.add_data_validation(dvD)
	dvD.add('D2:E1048576') # This is the same as for the whole of column B
	
	start_date = "1930-01-01"
	end_date = "2100-12-31"
	dvS = DataValidation(type="date",operator="between",formula1=f'"{start_date}"',formula2=f'"{end_date}"',allow_blank=False)
	ws.add_data_validation(dvS)
	dvS.add('H2:H1048576') # This is the same as for the whole of column B
	
	dvS1 = DataValidation(type="date",operator="between",formula1=f'"{start_date}"',formula2=f'"{end_date}"',allow_blank=True)
	ws.add_data_validation(dvS1)
	dvS1.add('I2:J1048576') # This is the same as for the whole of column B
	
	wb.save(PandG)
	
def cu(): 
	current = "Scholarship/current student data.xlsx"
	wb = load_workbook(current)
	ws = wb['Data']
	
	dvList1 = DataValidation(type="list",formula1="Notes!$B$5:$D$5",allow_blank=False)
	ws.add_data_validation(dvList1)
	dvList1.add('B2:B1048576') # This is the same as for the whole of column D
	
	start_date = "1930-01-01"
	end_date = "2100-12-31"
	dvS = DataValidation(type="date",operator="between",formula1=f'"{start_date}"',formula2=f'"{end_date}"',allow_blank=False)
	ws.add_data_validation(dvS)
	dvS.add('C2:C1048576') # This is the same as for the whole of column B
	
	wb.save(current)
	
def th(): 
	thesis = "Scholarship/thesis data.xlsx"
	wb = load_workbook(thesis)
	ws = wb['Data']
	
	dvList1 = DataValidation(type="list",formula1="Notes!$B$7:$D$7",allow_blank=False)
	ws.add_data_validation(dvList1)
	dvList1.add('D2:D1048576') # This is the same as for the whole of column D
	
	start_date = "1930-01-01"
	end_date = "2100-12-31"
	dvS = DataValidation(type="date",operator="between",formula1=f'"{start_date}"',formula2=f'"{end_date}"',allow_blank=True)
	ws.add_data_validation(dvS)
	dvS.add('B2:B1048576') # This is the same as for the whole of column B
	
	dvYear = DataValidation(type="whole",
						operator="between",
						formula1=1900,formula2=2100,allow_blank=False)
	ws.add_data_validation(dvYear)
	dvYear.add('C2:C1048576') # This is the same as for the whole of column B

	wb.save(thesis)

def pd(): 
	profd = "Service/professional development data.xlsx"
	wb = load_workbook(profd)
	ws = wb['Data']
	
	dvList1 = DataValidation(type="list",formula1="Notes!$A$5:$A$7",allow_blank=False)
	ws.add_data_validation(dvList1)
	dvList1.add('B2:B1048576') # This is the same as for the whole of column D

	dvList2 = DataValidation(type="list",formula1="Notes!$A$10:$A$13",allow_blank=False)
	ws.add_data_validation(dvList2)
	dvList2.add('C2:C1048576') # This is the same as for the whole of column D

	dvYear = DataValidation(type="whole",
						operator="between",
						formula1=1900,formula2=2100,allow_blank=False)
	ws.add_data_validation(dvYear)
	dvYear.add('D2:D1048576') # This is the same as for the whole of column B
	
	dvD = DataValidation(type="decimal",
                    operator="greaterThan",
                    formula1=0,allow_blank=True)
	ws.add_data_validation(dvD)
	dvD.add('E2:E1048576') # This is the same as for the whole of column B

	wb.save(profd)

def se(): 
	service = "Service/service data.xlsx"
	wb = load_workbook(service)
	ws = wb['Data']
	
	dvList1 = DataValidation(type="list",formula1="Notes!$B$11:$E$11",allow_blank=False)
	ws.add_data_validation(dvList1)
	dvList1.add('B2:B1048576') # This is the same as for the whole of column D

	dvList2 = DataValidation(type="list",formula1="Notes!$B$12:$C$12",allow_blank=False)
	ws.add_data_validation(dvList2)
	dvList2.add('C2:C1048576') # This is the same as for the whole of column D

	dvList3 = DataValidation(type="list",formula1="Notes!$B$13:$E$13",allow_blank=False)
	ws.add_data_validation(dvList3)
	dvList3.add('D2:D1048576') # This is the same as for the whole of column D


	dvYear = DataValidation(type="whole",
						operator="between",
						formula1=1900,formula2=2100,allow_blank=False)
	ws.add_data_validation(dvYear)
	dvYear.add('E2:E1048576') # This is the same as for the whole of column B
	
	dvD = DataValidation(type="decimal",
                    operator="greaterThan",
                    formula1=0,allow_blank=True)
	ws.add_data_validation(dvD)
	dvD.add('F2:F1048576') # This is the same as for the whole of column B

	wb.save(service)

def ur(): 
	undergraduate = "Service/undergraduate research data.xlsx"
	wb = load_workbook(undergraduate)
	ws = wb['Data']
	
	dvList1 = DataValidation(type="list",formula1="Notes!$A$7:$A$11",allow_blank=False)
	ws.add_data_validation(dvList1)
	dvList1.add('C2:C1048576') # This is the same as for the whole of column D
	
	dvList2 = DataValidation(type="list",formula1="Notes!$A$14:$A$17",allow_blank=False)
	ws.add_data_validation(dvList2)
	dvList2.add('D2:D1048576') # This is the same as for the whole of column D
	
	dvYear = DataValidation(type="whole",
						operator="between",
						formula1=1900,formula2=2100,allow_blank=False)
	ws.add_data_validation(dvYear)
	dvYear.add('E2:E1048576') # This is the same as for the whole of column B

	wb.save(undergraduate)



pa()
sa()
pg()
cu()
th()
pd()
se()	
ur()