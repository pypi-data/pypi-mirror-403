"""
author: 馒头
email: neihanshenshou@163.com
"""
import os

from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from openpyxl.styles import PatternFill

from SteamedBun import FileOperate


def dye_color(covered_api_filename,
              all_api_filename,
              delete_covered=True,
              url_title="urls",
              method_title="methods",
              time_title="times"):
    """
    :param covered_api_filename: api接口覆盖文件
        example - urls + 请求方法 + times
                /api/api    get     3

    :param all_api_filename: 总的api接口文件
        example - urls + methods + times
                /api/api    get     0
                /api/api    post    0

    :param delete_covered: 是否删除第一个文件
    :param url_title: 第一列表头 正常是指请求路径
    :param method_title: 第二列表头 正常是指请求方法
    :param time_title: 第三列表头 正常是指请求次数

    :return 重新渲染 总的api接口文件
        example - urls + methods + times
                /api/api    get     3(绿色)
                /api/api    post    0(红色)
    """
    color_green = PatternFill(start_color='a4e079', end_color='a4e079', fill_type='solid')
    color_red = PatternFill(start_color='e02d09', end_color='e02d09', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'))

    workbook1 = load_workbook(filename=covered_api_filename)
    workbook2 = load_workbook(filename=all_api_filename)
    sheet1 = workbook1.active
    sheet2 = workbook2.active
    sheet_data1 = FileOperate.read_excel(filename=covered_api_filename, sheet_name=sheet1.title)
    sheet_data2 = FileOperate.read_excel(filename=all_api_filename, sheet_name=sheet2.title)

    dict_data1 = sheet_data1.to_dict()
    dict_data2 = sheet_data2.to_dict()

    covered_count, all_api_count = 0, len(sheet_data2.values) or 0
    for index1, value1 in dict_data1.get(url_title).items():
        for index2, value2 in dict_data2.get(url_title).items():
            if value2 == value1:
                if dict_data1.get(method_title).get(index1) == dict_data2.get(method_title).get(index2):
                    for _ in ("A", "B", "C"):
                        sheet2[f"{_}{index2 + 2}"].fill = color_green
                        sheet2[f"{_}{index2 + 2}"].border = thin_border

                    sheet2[f"C{index2 + 2}"].value = sheet1[f"C{index1 + 2}"].value
                    covered_count += 1

    if all_api_count != 0:
        from SteamedBun import NumberTools
        coverage = NumberTools.percent(
            number=covered_count * 100 / all_api_count,
            rate=2
        )
    else:
        coverage = "0 %"
    sheet2["D1"].value = "接口覆盖率"
    sheet2["D2"].value = coverage
    workbook2.save(filename=all_api_filename)
    workbook1.close()

    workbook = load_workbook(filename=all_api_filename)
    sheet = workbook.active
    sheet_data = FileOperate.read_excel(filename=all_api_filename, sheet_name=sheet.title)
    dict_data = sheet_data.to_dict()
    for each, value in dict_data.get(time_title).items():
        if str(value) == "0":
            for _ in ("A", "B", "C"):
                sheet[f"{_}{each + 2}"].fill = color_red
                sheet[f"{_}{each + 2}"].border = thin_border

    workbook.save(filename=all_api_filename)

    if delete_covered:
        os.remove(covered_api_filename)
