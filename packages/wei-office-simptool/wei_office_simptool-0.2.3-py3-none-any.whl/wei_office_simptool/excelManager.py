from pathlib import Path
from typing import List, Optional, Sequence, Tuple
import pandas as pd
import xlwings as xw
import openpyxl
from openpyxl import load_workbook
from contextlib import contextmanager
from .stringManager import StringBaba

def create_workbook(file_name: str, default_sheet: str = "sheet1") -> None:
    """
    创建一个新的 Excel 工作簿并指定默认工作表名称
    """
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = default_sheet
    wb.save(file_name)

def _auto_range(sr: int, sc: int, results: Sequence[Sequence], re: int, er: int, ec: int) -> Tuple[int, int]:
    """
    根据传入数据自动计算写入的结束行列（re=0 自动计算；re=1 使用传入的 er/ec）
    """
    if re == 0 and results:
        er = len(results) + sr - 1
        ec = len(results[0]) + sc - 1
    return er, ec


class ExcelHandler:
    """
    ExcelHandler：面向已有文件的读取/写入工具
    - 支持自动创建缺失的工作表
    - 提供快速写入（fast_write）以简化范围计算
    """
    def __init__(self, file_name: str):
        self.file_name = file_name
        if not Path(file_name).exists():
            create_workbook(file_name)
        self.wb = load_workbook(self.file_name)

    def _ensure_sheet(self, sheet_name: str):
        if sheet_name not in self.wb.sheetnames:
            self.wb.create_sheet(title=sheet_name)
        return self.wb[sheet_name]

    def excel_write(self, sheet_name: str, results: Sequence[Sequence], start_row: int, start_col: int, end_row: int, end_col: int):
        try:
            if not results:
                return
            sheet = self._ensure_sheet(sheet_name)
            for i, row in enumerate(range(start_row, end_row + 1)):
                for j, value in enumerate(range(start_col, end_col + 1)):
                    sheet.cell(row=row, column=value, value=results[i][j])
            print("Results have been written!")
            self.wb.save(self.file_name)
        except Exception as e:
            print(e)

    def excel_read(self, sheet_name: str, start_row: int, start_col: int, end_row: int, end_col: int):
        try:
            sheet = self._ensure_sheet(sheet_name)
            values = [
                [sheet.cell(row=row, column=col).value for col in range(start_col, end_col + 1)]
                for row in range(start_row, end_row + 1)
            ]
            print("Results have been read!")
            return values
        except Exception as e:
            print(e)

    def excel_save_as(self, file_name2: Optional[str]):
        try:
            self.wb.save(file_name2 or self.file_name)
            print("The file has been saved as " + str(file_name2))
        except Exception as e:
            print(e)

    def excel_quit(self):
        try:
            self.wb.close()
        except Exception as e:
            print(e)

    @staticmethod
    def fast_write(sheet_name: str, results: Sequence[Sequence], start_row: int, start_col: int, end_row: int = 0, end_col: int = 0, re: int = 0, xl_book: Optional["ExcelHandler"] = None):
        end_row, end_col = _auto_range(start_row, start_col, results, re, end_row, end_col)
        xl_book.excel_write(sheet_name, results, start_row=start_row, start_col=start_col, end_row=end_row, end_col=end_col)


class OpenExcel:
    """
    OpenExcel：通过 Excel 应用打开工作簿，适合需要 RefreshAll 的场景
    - my_open 上下文：返回 eExcel 对象，退出时保存
    - open_save_Excel 上下文：返回 xlwings 的 Workbook，退出时刷新数据并保存
    """
    def __init__(self, openfile: str, savefile: Optional[str] = None):
        self.openfile = openfile
        self.savefile = savefile

    @contextmanager
    def my_open(self):
        print(f"Opening Excel file: {self.openfile}")
        wb = eExcel(file_name=self.openfile)
        yield wb
        wb.excel_save_as(self.savefile or self.openfile)

    @contextmanager
    def open_save_Excel(self):
        app = None
        wb = None
        try:
            app = xw.App(visible=False)
            wb = app.books.open(self.openfile)
        except Exception as e:
            if app:
                app.quit()
            raise e
        try:
            yield wb
        finally:
            try:
                wb.api.RefreshAll()
                wb.save(self.savefile or self.openfile)
            finally:
                if app:
                    app.quit()

    def file_show(self, filter: Optional[Sequence[str]] = None):
        app = xw.App(visible=False)
        wb = app.books.open(self.openfile)
        wbsn = wb.sheet_names
        app.quit()
        if filter is not None:
            f = filter if isinstance(filter, (list, tuple)) else [filter]
            wbsn = StringBaba(wbsn).filter_string_list(f)
        return wbsn


class ExcelOperation:
    """
    ExcelOperation：数据处理类
    - split_table：按工作表拆分为多个文件保存至输出目录
    """
    def __init__(self, input_file: str, output_folder: str):
        self.input_file = input_file
        self.output_folder = output_folder

    def split_table(self):
        excel_file = pd.ExcelFile(self.input_file)
        out_dir = Path(self.output_folder)
        out_dir.mkdir(parents=True, exist_ok=True)
        for sheet_name in excel_file.sheet_names:
            df = pd.read_excel(self.input_file, sheet_name=sheet_name)
            output_file = f'{sheet_name}.xlsx'
            df.to_excel(str(out_dir / output_file), index=False)


class eExcel:
    """
    eExcel：轻量级工作簿操作（仅依赖 openpyxl）
    - 支持不存在文件时自动创建
    - 提供快速写入（fast_write）
    """
    def __init__(self, file_name: Optional[str] = None):
        self.file_name = file_name
        if not Path(file_name).exists():
            create_workbook(file_name)
        self.wb = openpyxl.load_workbook(file_name)
        self.ws = self.wb.active

    def create_new_sheet(self, ws: str):
        self.wb.create_sheet(ws)

    def excel_write(self, ws: str, results: Sequence[Sequence], start_row: int, start_col: int, end_row: int, end_col: int):
        ws_obj = self.wb[ws]
        for i, row in enumerate(range(start_row, end_row + 1)):
            for j, value in enumerate(range(start_col, end_col + 1)):
                ws_obj.cell(row=row, column=value, value=results[i][j])

    def excel_read(self, start_row: int, start_col: int, end_row: int, end_col: int):
        valueA = [
            [self.ws.cell(row=row, column=col).value for col in range(start_col, end_col + 1)]
            for row in range(start_row, end_row + 1)
        ]
        return valueA

    def excel_save_as(self, file_name2: str):
        self.wb.save(file_name2)

    def fast_write(self, ws: str, results: Sequence[Sequence], sr: int, sc: int, er: int = 0, ec: int = 0, re: int = 0, wb: Optional["eExcel"] = None):
        er, ec = _auto_range(sr, sc, results, re, er, ec)
        target = wb if wb else self
        target.excel_write(ws, results, start_row=sr, start_col=sc, end_row=er, end_col=ec)

    @classmethod
    def quick(cls, file_name: str, default_sheet: str = "sheet1") -> "eExcel":
        """
        快速创建并返回 eExcel 对象（如果文件不存在则创建）
        """
        if not Path(file_name).exists():
            create_workbook(file_name, default_sheet)
        return cls(file_name=file_name)
