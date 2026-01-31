#
# -*- coding: utf-8  -*-
#=================================================================
# Created by: Jacky Lai
# Created on: June 2024
# Last Modified: June 2024
#=================================================================
# Copyright (c) 2024 [Jieming Ye]
#
# This Python source code is licensed under the 
# Open Source Non-Commercial License (OSNCL) v1.0
# See LICENSE for details.
#=================================================================
"""
Pre-requisite: 
SimulationName.oof: default oslo output file after successfully run a simulation.
DC Single End Feeding Input.xlsx summary spreadsheet with user configuration settings for DC single end feeding assessment
Used Input:
simname: to locate the result file or file rename
time_start: to define the analysis start time
time_end: to define the analysis end time
option_select: to define the option selected in subfunction
text_input: to locate the excel file name
other input for oslo_extraction.py only.
Expected Output:
Various .csv file containing individual detailed output or summary output.
Writes to DC Single End Feeding Input.xlsx with results and results summary (color coded)
Description:
This script defines the process of DC data calculation for single end feeding.

In branch_data_process(), it generates the list file first and read list file line by line, saving the result is a data frame during the reading process. It then doing individual calculation and output the result to several .csv files. The process is relative easy to follow via reading the code.

"""
#=================================================================
# VERSION CONTROL
# V0.2 (Jacky Lai) - Beta Test Version updated June 2024
# DC processing tool for single end feeding 1st stage assessment 
#=================================================================
# Set Information Variable
# N/A
#=================================================================

import os
import math
import csv

import pandas as pd
import numpy as np

from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, NamedStyle
# from openpyxl.formatting.rule import CellIsRule,FormulaRule
from vision_oslo_extension.shared_contents import SharedMethods

def main(simname, main_option, time_start, time_end, option_select, text_input, low_v, high_v, time_step):
    
    print("")
    print("DC Single End Feeding Assessment - - - > ")
    print("")
    
    #simname = simname

    option = option_select

    start = 7 # result start from row 7
    space = 5
    # time_increment = 5
    time_windows_total= ['15min','30min']

    # Option:
    # 1: DC SEF for Track Circuit Breakers TCBs 15min
    # 2: DC SEF for +ve ETE                     30min
    # 3: DC SEF for Track Circuit Breakers TCBs 15min with CSV step outputs
    # 4: DC SEF for +ve ETE                     30min with CSV step outputs

    #check option for TCB or ETE is selected
    if option not in ["0","1","2","3","4"]:
        SharedMethods.print_message("ERROR: Error in DC single end feeding. Please contact Support...","31")
        return False

    if option == "0":
        SharedMethods.print_message("ERROR: Please select an option to proceed.","31")
        return False
    elif option == "1": #option for TCB use 15 min timewindow
        time_windows= ['15min']
        time_step = 900
    elif option == "2": #option for ETE use 30 min timewindow
        time_windows= ['30min']
        time_step = 1800

    elif option == "3": #option for TCB use 15 min timewindow, extract CSV
        time_windows= ['15min']
        time_step = 900

    elif option == "4": #option for ETE use 30 min timewindow, extract CSV
        time_windows= ['30min']
        time_step = 1800

    # Specify Excel file name
    excel_file = text_input + ".xlsx"

    if not SharedMethods.check_existing_file(excel_file):
        return False

    time_increment = 5
    
    #read excel and get branch list
    result = start_reading_process(simname,excel_file,option)
    if result == False:
        return False
    else:
        start_df,oslo_total = result

    # read branch names from excel and create modified branch list to do extraction (RMS)
    branch_list_1 = read_all_branches(start_df)

    #create df for results
    results_df = start_df

    #run modified branch_data_process using xlsx as input        
    results_df = sef_branch_data_process(simname,option_select,time_start,time_end,time_step,text_input,time_increment,branch_list_1,results_df)

    if not SEF_ratings_excel(simname,excel_file,option_select,results_df):
        return False

    return True

# read the start tab and collect informaiton
def start_reading_process(simname, excel_file,option):
    
    # File check
    print("Check Excel and Deleting Existing Contents ...")
    if option == "1" or option == "3":
        columns = ['Substation Name', 'OSLO', 'Feeder ID', 'TCB Type', 'Ref Rating (kA)','Rating (kA)']
        col_num = 6

    elif option == "2" or option == "4":
        columns = ['Substation Name', 'OSLO', 'Feeder ID', 'TCB Type', 'Ref Rating (kA)','Rating (kA)']
        col_num = 6

    try:
        wb = load_workbook(excel_file)
        # Delete all sheets except 'Start'
        for sheet_name in wb.sheetnames:
            if sheet_name == 'Start':
                result = check_reading_frame(wb[sheet_name],option) # list, dictionary type
                if result == False:
                    return False
                else:
                    #data_start_row,data_end_row,oslo_total,scenariolist = result
                    data_start_row,data_end_row,oslo_total = result
                
                start_df = pd.read_excel(excel_file,sheet_name = 'Start',header = 0,usecols=range(col_num), skiprows=data_start_row-1, nrows=oslo_total, names = columns)
                oslolist = start_df.iloc[:, 1].tolist()
                
                # check duplication in OSLO id
                if not SharedMethods.find_duplicates(oslolist): return False

                flag = False
                # check oslo list 
                if not option == "8": # do not check option 8 as the format is very different
                    for index, oslo in enumerate(oslolist):
                        if pd.isna(oslo):
                            SharedMethods.print_message(f"ERROR: Items {start_df.iloc[index,0]} is not assigned with OSLO ID","31")
                            flag = True

                # decision point
                if flag:
                    return False

            else:
                #delete other workbooks
                wb.remove(wb[sheet_name])
        
        # Save changes
        wb.save(excel_file)
        
    except Exception as e:
        SharedMethods.print_message(f"ERROR: (Close the Excel file and Start Again) Error: {e}","31")
        return False
    
    return start_df,oslo_total

# rating data frame
def check_reading_frame(sheet,option):
    table_start_row = 11

    # define the outage list section to be read # Compare against function above
    print("Check Data Entry & Scenario List...")
    if option == "1" or option == "3":
        table_row = 12
        table_start_column = 2

    elif option == "2" or option == "4":
        table_row = 12
        table_start_column = 2
 
    # create ole list and rating data
    # find total number of substations
    index = table_row
    column = table_start_column
    if sheet.cell(row=index, column=column).value is not None:
        while True:          
            index += 1
            check = sheet.cell(row=index, column=column).value
            if check is None:
                table_row_end = index
                oslo_total = index - table_row
                #oslo_total is number of substaions
                break
    else: 
        #if no substation is found
        SharedMethods.print_message("ERROR: Wrong data format. No information at B12","31")
        return False
    
    return table_start_row,table_row_end,oslo_total

#read all branch list and create a list for extracting all branch step RMS outputs
def read_all_branches(start_df):

    # columns = ['Substation Name', 'OSLO', 'Feeder ID', 'TCB Type', 'Ref Rating (kA)','Rating (kA)']
    bl_1 = start_df['OSLO'].tolist()
        
    #for each item in bl_1 check if it end with /E or /S add the opposite node if it not already exist in bl_1
    bl_1 = bl_check_and_append(bl_1)
 
    return bl_1

#function for read_all_branches to add /E and /S to branch list
def bl_check_and_append(bl_append):
    """
    Check if an item ends with '/E' or '/S' and append the missing counterpart to the list.
    
    Args:
        row (pd.Series): The row to check and append.
        
    Returns:
        pd.Series: The updated row.
    """

    updated_list = bl_append.copy()  # Make a copy to avoid modifying the original list
    
    for item in bl_append:
        if item.endswith('/E'):
            start_item = item[:-2] + '/S'
            if start_item not in updated_list:
                updated_list.append(start_item)
        elif item.endswith('/S'):
            end_item = item[:-2] + '/E'
            if end_item not in updated_list:
                updated_list.append(end_item)
    
    return updated_list


# Updated from batch_processing.py --> Option 10 for DC assessment
# option 4 main sub function for data extraction
#changed to extract RMS step, RMS Max and add step value together
def sef_branch_data_process(simname, option_select,time_start, time_end, time_step, text_input, time_increment,branch_list,results_df):

    # print("5: Customised branch rolling RMS current (BranchNodeList.txt is required)")
    # will always run selected branch rolling RMS current step output extraction aka opiton 5

    simname_sef = simname + '_SEF'

    #create branch list for extractions
    branch_list_0 = results_df['OSLO'].tolist()

    second = int(time_step)
    
    #read branches using list from excel and previous loops
    branch_input = [node for node in branch_list if node]  # Remove empty strings
    for node in branch_input:
        if node[-2:] == '/S' or node[-2:] == '/E':
            pass
        else:
            SharedMethods.print_message(f"ERROR: Branch node list in xlsx has information {node} which is not the supported format.","31")
            return False

    filename = simname + ".osop.lst"

    if not SharedMethods.check_and_extract_lst_file(simname,time_start=time_start, time_end=time_end):
        return False

    sim_time, branch_list, step_branch_output, step_df = lst_file_read_branch(filename)

    #create a filtered branch_df for results
    # filte the table based on user input
    #add a column of the time
    selected_columns = ['Time']
    for item in branch_input:
        if item in step_df.columns:
            #add column for each branch
            selected_columns.append(item)
        else:
            SharedMethods.print_message(f"WARNING: Input File Node {item} does NOT exist in the simulation. Ignored....","33")
    branch_df = step_df[selected_columns]


    #for each branch in results_df create a sum of the step outputs on branches of both ends and create RMS output
    # SEF_results_df is step results of the sum of loads on both ends of the elctrical section
    SEF_results_df = pd.DataFrame()
    SEF_results_df['Time'] = branch_df['Time']

    for item in branch_list_0:
        node_s = item[:-2] + '/S'
        node_e = item[:-2] + '/E'
        SEF_results_df.loc[:,item]=branch_df[node_s] +branch_df[node_e]
        #performance issue errors here

    print('Finished processing branch SEF step outputs...')
    #simname_sef = simname + '_SEF'
    #write SEF step output to CSV file only do if option 3 or 4
    if option_select == "3" or option_select == "4":
        branch_step_csv_out(simname_sef,SEF_results_df)
        print('Finished saving step output CSV for SEF branch... option 3 or 4 selected')

    #write original step output to CSV file only do if option 3 or 4
    if option_select == "3" or option_select == "4":
        branch_step_csv_out(simname,branch_df)
        print('Finished saving step output CSV for (not SEF) branch ... option 3 or 4 selected')


    #process RMS step and RMS SUM for SEF loads
    print("Processing SEF_RMS calculation...")
    SEF_rms_df,SEF_max_df = branch_list_RMS_process(simname,SEF_results_df,second,time_increment)
    print('Finished processing SEF_RMS calculation...')
    
    results_df[['Start Time','End Time','Max RMS SEF Current (A)']] = SEF_max_df[['Start Time','End Time','Maximum RMS Current (A)']]
    
    #write SEF RMS step output to CSV file, only do if option 3 or 4
    if option_select == "3" or option_select == "4":
        branch_RMS_csv_out(simname_sef,second,SEF_rms_df)
        print('Finished saving RMS step CSV for SEF branch nodes ... option 3 or 4 selected')  

    #process RMS for branch nodes (not single end feeding)
    print("Processing Branch Normal Feed RMS calculation...")
    rms_df,max_df = branch_list_RMS_process(simname,branch_df,second,time_increment)
    print('\nFinished processing Branch Normal Feed RMS output...')

    #write RMS step output to CSV file only do if option 3 or 4
    if option_select == "3" or option_select == "4":
        branch_RMS_csv_out(simname,second,rms_df)
        print('Finished saving RMS step CSV for (not SEF) branch nodes ... option 3 or 4 selected')

    #add Max RMS node loads for each end of the branch to results_df
    #working on
    #set node as index for RMS max data frame
    max_df.set_index('Node',inplace=True)
    RMS_opp = []
    RMS_node =[]
    Opp_OSLO = []

    for nodes in results_df['OSLO']:
        if nodes.endswith('/E'):
            opp_node = nodes[:-2] + '/S'
            Opp_OSLO.append(opp_node)
            Node_RMS = max_df.loc[nodes,'Maximum RMS Current (A)']
            Opp_RMS = max_df.loc[opp_node,'Maximum RMS Current (A)']
            #list showing the RMS value of all Opposite node (RMS max for other end of branch)
            RMS_opp.append(Opp_RMS)
            #list showing the MAX RMS value at the node during normal (not single end) feeding
            RMS_node.append(Node_RMS)

        else:
            opp_node = nodes[:-2] + '/E'
            Opp_OSLO.append(opp_node)
            Node_RMS = max_df.loc[nodes,'Maximum RMS Current (A)']
            Opp_RMS = max_df.loc[opp_node,'Maximum RMS Current (A)']
            RMS_opp.append(Opp_RMS)
            RMS_node.append(Node_RMS)

    #add the RMS_opp and RMS_node to results list
    results_df['RMS_MAX_node (A)'] = RMS_node
    results_df['RMS_MAX_opposite (A)'] = RMS_opp
    results_df['Opposite_Node_OSLO_Name'] = Opp_OSLO

    # write SEF RMS sum to CSV file only do if option 3 or 4
    if option_select == "3" or option_select == "4":
        branch_RMS_sum_csv_out(simname_sef,second,results_df)
        print('Finished saving RMS MAX SUM CSV for SEF branch nodes ... option 3 or 4 selected')

    return results_df

# Read the list file and get all branches information in list:
def lst_file_read_branch(filename):

    sim_time = [] # Time list for reading
    branch_list = [] # initialise branch list
    step_branch_output = [] # step branch output information (2D list)

    # get all branches in the model
    section_flag = False # judge if N-Link section is reached

    print('\nInitialising...')
    with open(filename) as fp:
        total_line = sum(1 for line in enumerate(fp)) # get the total line number

    with open(filename) as fp:
        print('\nReading Branch Information...')
        for index, line in enumerate(fp):            
            if line[:7].strip() == '':
                continue
            if line[:7].strip() == 'NLINK':
                section_flag = True
                continue

            if line[:7].strip() == 'NFIXC':
                section_flag = False
                break
            
            if section_flag == True:
                branch_temp = line[10:14].strip()
                branch_list.append(branch_temp) # Save branches in a list file
                step_branch_output.append([branch_temp])

    
    branch_read_flag = False
    time_count = 0 # time counter
    branch_count = 0 # branch counter
    # create columns for pandadataframe
    columns = ['Time']
    for item in branch_list:
        columns.append(item +'/S')
        columns.append(item +'/E')
    row_data = []
    row = []
    
    print('\nProcess branch information...')
    with open(filename) as fp:
        for index, line in enumerate(fp):
            # flag judgement
            if line[:7].strip() == 'INCSEC':
                if row:
                    row_data.append(row)

                time_count = time_count + 1
                temp_time = line[20:22].strip()+":"+line[23:25].strip()+":"+line[26:28].strip()
                sim_time.append(temp_time)
                branch_read_flag = True

                row = []
                row.append(temp_time)
                continue

            if line[:7].strip() == 'TRAIN':
                branch_count = 0
                branch_read_flag = False
            
            # add to consider if there is no train at specific time 
            if line[:8].strip() == 'NO OSLO':
                row = []
                branch_count = 0
                branch_read_flag = False
            
            # write down information
            if branch_read_flag == True:
                if line[76:82].strip() == '':
                    continue

                if line[:7].strip() == '':
                    branch_count = branch_count + 1
                    c_start_real = float(line[85:95].strip())
                    c_start_imag = float(line[95:105].strip())
                    c_end_real = float(line[110:120].strip())
                    c_end_imag = float(line[120:130].strip())

                    c_start_total = math.sqrt(math.pow(c_start_real,2)+math.pow(c_start_imag,2))
                    c_end_total = math.sqrt(math.pow(c_end_real,2)+math.pow(c_end_imag,2))

                    temp_line = [temp_time,c_start_real,c_start_imag,c_end_real,c_end_imag]
                    step_branch_output[branch_count-1].append(temp_line)

                    # temp_line = [c_start_total,c_end_total]
                    # step_branch_sum_c[branch_count-1].append(temp_line)
                    
                    row.append(c_start_total)
                    row.append(c_end_total)
                else:
                    continue

            SharedMethods.text_file_read_progress_bar(index, total_line)
    
    # create data frame straighaway:
    step_df = pd.DataFrame(row_data, columns = columns)

    return sim_time, branch_list, step_branch_output, step_df

# Write branch step output information to csv file
def branch_step_csv_out(simname,branch_df):

    # write summary info to CSV file
    with open(simname+'_branch_step.csv','w',newline='') as f:
        writer = csv.writer(f) # create the csv writer
        row = ['OSLO Branch Step Output']
        writer.writerow(row)
        row = ['Branch Node List']
        writer.writerow(row)

        # Write df1 to the file
        # df.insert(0, 'Blank', '')
        branch_df.to_csv(f, index=False)
        
    #print('\nBranch Step Output Summary Completed.')

    return True

# branch RMS output information to csv file
def branch_RMS_csv_out(simname,second,rms_df):

    # write summary info to CSV file
    with open(simname+'_branch_'+str(second)+'_rms_step.csv','w',newline='') as f:
        writer = csv.writer(f) # create the csv writer
        row = [f'OSLO Branch RMS Output over {second} seconds']
        writer.writerow(row)
        row = ['Branch Node List']
        writer.writerow(row)

        # Write df1 to the file
        # df.insert(0, 'Blank', '')
        rms_df.to_csv(f, index=False)
        
    #print('\nBranch RMS Output Summary Completed.')

    return

# branch RMS summary output information to csv file
def branch_RMS_sum_csv_out(simname,second,max_df):
    # write summary info to CSV file
    with open(simname+'_branch_'+str(second)+'_rms_sum_max.csv','w',newline='') as f:
        
        writer = csv.writer(f) # create the csv writer
        row = [f'OSLO Branch RMS Output over {second} seconds - Summary']
        writer.writerow(row)

        # Write df1 to the file
        # df.insert(0, 'Blank', '')
        max_df.to_csv(f, index=False)
        
    #print('\nBranch RMS Output Summary Completed.')

    return

# calculate rms
def calculate_rolling_rms(data, column, window_size, time_increment):
    return data[column].rolling(window=int(window_size / time_increment)).apply(lambda x: np.sqrt(np.mean(x**2)), raw=True)

# process branch list RMS calculation
def branch_list_RMS_process(simname,branch_df,second,time_increment):

    time_window = int(second/time_increment)
    # Calculate rolling RMS for each column except the first one
    print("Calculating in process.. This will take a bit of time...")
    print("Do not close this window or do other things...")

    total = branch_df.shape[1] - 1 # total number of columns
    node_list = branch_df.columns.tolist()
    node_list = node_list[1:] #exclude the first one

    rms_df = pd.DataFrame()

    rms_df['Time'] = branch_df['Time']

    # Create a list to store calculated RMS values for each node
    rms_values_list = []

    for index, node in enumerate(node_list):
        print(f"\rProcessing node: {node}...({index+1}/{total})", end='', flush=True)
        rms_values  = calculate_rolling_rms(branch_df, node, second, time_increment)
        rms_values_list.append(rms_values)
    print(f"\rProgress all node successfully completed.", flush=True)
    
    # Concatenate all calculated RMS values into a single DataFrame
    rms_values_df = pd.concat(rms_values_list, axis=1)

    # Concatenate rms_df with rms_values_df to combine 'Time' column and calculated RMS values
    rms_df = pd.concat([rms_df, rms_values_df], axis=1)

    print('Calculate Maximum Value...')

    # finding max summary
    row_data = []
    for index, node in enumerate(node_list):
        if rms_df[node].dtype in [int, float]:
            max_value = rms_df[node].max()
            if not np.isnan(max_value):
                time_of_max_e = rms_df.loc[rms_df[node].idxmax(), 'Time']
                time_of_max_s = rms_df.loc[rms_df[node].idxmax() - time_window + 1, 'Time']
            else:
                time_of_max_e = 'NA'
                time_of_max_s = 'NA'
        row_data.append([node[:-2],node,time_of_max_s,time_of_max_e,max_value])
    
    columns = ['Branch','Node','Start Time','End Time','Maximum RMS Current (A)']
    max_df = pd.DataFrame(row_data,columns = columns)

    return rms_df,max_df

# compare ratings and export to excel
# uses the updated results_df from branch_data_process
def SEF_ratings_excel(simname,excel_file,option_select,results_df):
    #create columns of results in kA instead of A
    results_df['RMS_MAX_node (kA)'] = results_df['RMS_MAX_node (A)']/1000
    results_df['RMS_MAX_opposite (kA)'] = results_df['RMS_MAX_opposite (A)']/1000
    results_df['Max RMS SEF Current (kA)'] = results_df['Max RMS SEF Current (A)'] /1000

    columns2 = ['Substation Name', 'OSLO', 'Opposite_Node_OSLO_Name', 'Feeder ID', 'TCB Type', 'Ref Rating (kA)','Rating (kA)','Start Time','End Time','Max RMS SEF Current (kA)','RMS_MAX_node (kA)','RMS_MAX_opposite (kA)']

    #write to results sheet
    excel_results_df = pd.DataFrame(columns=columns2)
    #write to results summary sheet
    excel_results_sum_df = pd.DataFrame(columns=columns2)

    results_list = []
    results_sum_list = []
    ratingc = float(0)
    ref_ratingc = float(0)

    #compare the ratings and add to the respective lists
    for index, row in results_df.iterrows():

        #before comparing the ratings check that ratings are valid
        # Check if 'Rating (kA)' is a valid number
        if is_number(row['Rating (kA)']):
            ratingc=float(row['Rating (kA)'])
        #if it is not a number set rating to zero     
        else:
            ratingc = float(0)

        # Check if 'Ref Rating (kA)' is a valid number
        if is_number(row['Ref Rating (kA)']):
            ref_ratingc=float(row['Ref Rating (kA)'])
        #if it is not a number set rating to zero     
        else:
            ref_ratingc = float(0)            

        # compare Max RMS SEF to rating (criteria for red faliure)
        if float(row['Max RMS SEF Current (kA)'])> ratingc:
            results_list.append(row[columns2].tolist())
            results_sum_list.append(row[columns2].tolist())
            #excel_results_sum_df = pd.concat(row[columns2], ignore_index=True)
            #excel_results_df = pd.concat(row[columns2], ignore_index=True)

        # compare Max RMS SEF to ref rating (criteria for amber faliure)
        elif float(row['Max RMS SEF Current (kA)'])> ref_ratingc:
            results_list.append(row[columns2].tolist())
            results_sum_list.append(row[columns2].tolist())
            #excel_results_sum_df = pd.concat(row[columns2], ignore_index=True)
            #excel_results_df = pd.concat(row[columns2], ignore_index=True)

        else:
            results_list.append(row[columns2].tolist())
            #excel_results_df = pd.concat(row[columns2], ignore_index=True)

    excel_results_sum_df = pd.DataFrame(results_sum_list, columns=columns2)
    excel_results_df = pd.DataFrame(results_list, columns=columns2)
    
    start = 0

    #write the lists to excel
    with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a',if_sheet_exists='overlay') as writer:

        # get currrent workbook to do process
        wb = writer.book
        print("Generate Failure Result Page in excel...")
        r_start = start + 10
        excel_results_sum_df.to_excel(writer, sheet_name="Result Summary", index=False, startrow = r_start, startcol = 1)
        
        # get currrent workbook to do process
        print("Generate Single End Feeding results page in excel...")
        excel_results_df.to_excel(writer, sheet_name="Full Single End Feeding Results", index=False, startrow = r_start, startcol = 1)

        sheets = wb['Start']
        sheet1 = wb["Result Summary"]
        sheet2 = wb["Full Single End Feeding Results"]
            
        project_name = sheets['B2'].value
        feeding_desp = sheets['B4'].value
        modeller = sheets['B5'].value
        #date = sheets['B6'].value

        # Add titles to results tabs
        sheet1['B2'].value = "Project Name:"
        sheet1['C2'].value = project_name
        sheet1['B3'].value = "Simulation Name:"
        sheet1['C3'].value = simname
        sheet1['B4'].value = "Feeding Arrangement:"
        sheet1['C4'].value = feeding_desp
        sheet1['B5'].value = "Result Created by:"
        sheet1['C5'].value = modeller
        sheet1['B6'].value = "Result Created at:"
        sheet1['C6'].value = datetime.now().strftime("%d-%m-%Y %H:%M")
        # Add titles to results tabs
        sheet2['B2'].value = "Project Name:"
        sheet2['C2'].value = project_name
        sheet2['B3'].value = "Simulation Name:"
        sheet2['C3'].value = simname
        sheet2['B4'].value = "Feeding Arrangement:"
        sheet2['C4'].value = feeding_desp
        sheet2['B5'].value = "Result Created by:"
        sheet2['C5'].value = modeller
        sheet2['B6'].value = "Result Created at:"
        sheet2['C6'].value = datetime.now().strftime("%d-%m-%Y %H:%M")

        sheet1['B9'].value = "Single End Feeding Faliures"
        cell = sheet1['B9']
        cell.font = Font(bold = True, size = 11)
        sheet2['B9'].value = "Single End Feeding Results"
        cell = sheet2['B9']
        cell.font = Font(bold = True, size = 11)
        #sheet1['C9'].value = "Failure " + name + " Summary"

        #make small changes to title rows
        sheet1['C11'] = 'Assessed OSLO Branch Node'
        sheet1['D11'] = 'Opposite OSLO Branch Node'

        sheet2['C11'] = 'Assessed OSLO Branch Node'
        sheet2['D11'] = 'Opposite OSLO Branch Node'

        if option_select == "1" or option_select == "3":
            sheet1['G11'] = 'TCB rating (kA)'
            sheet1['H11'] = 'Assessed rating(kA)'
            sheet2['G11'] = 'TCB rating (kA)'
            sheet2['H11'] = 'Assessed rating(kA)'

        if option_select == "2" or option_select == "4":
            sheet1['F11'] = 'ETE Type'
            sheet1['G11'] = '90 percent rating (kA)'
            sheet1['H11'] = 'Cable rating(kA)'
            sheet2['F11'] = 'ETE Type'
            sheet2['G11'] = '90 percent rating (kA)'
            sheet2['H11'] = 'Cable rating(kA)'

        sheet1['K10'] = 'Single End Feeding Current'
        sheet1['K11'] = 'Max RMS Combined Current (kA)'
        sheet2['K10'] = 'Single End Feeding Current'
        sheet2['K11'] = 'Max RMS Combined Current (kA)'

        #apply resize and reformatting
        resize_columns(sheet1)
        apply_conditional_formatting(sheet1)

        resize_columns(sheet2)
        apply_conditional_formatting(sheet2)

    return True

def resize_columns(sheet):
    for column in sheet.columns:
        max_length = 0
        column = column[0].column_letter
        for cell in sheet[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 1)
        sheet.column_dimensions[column].width = adjusted_width

def apply_conditional_formatting(sheet):
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red (0% S)
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 255,152,51 (80% S)
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green (00% S)

    Ref_rat = 0
    rat = 0 
    Cell_result = 0 

    center_alignment = Alignment(horizontal='center')

    #read Ratings reference ratings and results from excel values and apply formatting 
    for row in sheet.iter_rows(min_row=12, max_row=sheet.max_row, min_col=2, max_col=14):
        for cell in row:
            if cell.column == 7:
                #apply number check to reference rating
                if is_number(cell.value):
                    Ref_rat = cell.value
                #if it is not a number set ref rating to zero     
                else:
                    Ref_rat = float(0)
            
            if cell.column == 8:
                #apply number check to rating
                if is_number(cell.value):
                    rat = cell.value
                #if it is not a number set rating to zero     
                else:
                    rat = float(0)               

            #get results value from column 11
            if cell.column == 11:
                Cell_result = cell.value

        if Cell_result is not None and rat is not None:

            if Cell_result > rat:
                sheet.cell(row=row[0].row, column=11).fill = red_fill

            elif Cell_result > Ref_rat:
                sheet.cell(row=row[0].row, column=11).fill = yellow_fill

            elif Cell_result > 0:
                sheet.cell(row=row[0].row, column=11).fill = green_fill

    for row in sheet.iter_rows(min_row=12, max_row=sheet.max_row, min_col=2, max_col=14):
        for cell in row:
        #conditional formattng to show 2 decimal place only
            if cell.column in [11,12,13]:
                cell.number_format = '0.00'

        #Set alignment to center for columns 7 to 13
            if cell.column in range(7, 14):
                cell.alignment = center_alignment

#check if a value is a number
def is_number(variable):
    if variable is None:
        return False
    try:
        float_value = float(variable)  # Try to convert the variable to a float
        
        if math.isnan(float_value):
            return False

        return True
    except ValueError:
        return False
            
#============================================================================
# Check if the script is run as the main module
if __name__ == "__main__":
    # Add your debugging code here
    simname = "DCF312"  # Provide a simulation name or adjust as needed
    main_option = "12"  # Adjust as needed
    time_start = "0100000"  # Adjust as needed
    time_end = "0160000"  # Adjust as needed
    option_select = "3"  # Adjust as needed
    text_input = "DC Single End Feeding Input for TCB"  # Adjust as needed
    low_v = None  # Adjust as needed
    high_v = None  # Adjust as needed
    time_step = 900  # Adjust as needed

    # Call your main function with the provided parameters
    main(simname, main_option, time_start, time_end, option_select, text_input, low_v, high_v, time_step)