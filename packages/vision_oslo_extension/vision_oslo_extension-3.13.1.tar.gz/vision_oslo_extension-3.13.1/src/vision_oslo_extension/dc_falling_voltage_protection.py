#
# -*- coding: utf-8  -*-
#=================================================================
# Created by: Jacky Lai
# Created on: Jun 2024
# Last Modified: July 2024
#=================================================================
# Copyright (c) 2024 [Jieming Ye]
#
# This Python source code is licensed under the 
# Open Source Non-Commercial License (OSNCL) v1.0
# See LICENSE for details.
#=================================================================
"""
Pre-requisite: 
SimulationName.oof: default oslo output file after successfully run a simulation.
FVP input.xlsx
Used Input:
simname: to locate the result file or file rename
time_start: to define the analysis start time
time_end: to define the analysis end time
option_select: to define the option selected in subfunction
text_input: to locate the excel file name
other input for oslo_extraction.py only.
Expected Output:
Various .csv file containing individual detailed output or summary output.
Description:
This script defines the process of DC data calculation for falling voltage protection.

In branch_data_process(), it generates the list file first and read list file line by line, saving the result is a data frame during the reading process. It then doing individual calculation and output the result to several .csv files. The process is relative easy to follow via reading the code.

"""
#=================================================================
# VERSION CONTROL
# V0.1 (Jacky Lai) - Test Version
#=================================================================
# Set Information Variable
# N/A
#=================================================================

import os
import math
import csv

import pandas as pd
import numpy as np

from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, NamedStyle

from datetime import datetime
from functools import reduce

from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule,FormulaRule
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties
from openpyxl.drawing.colors import ColorChoice
from openpyxl.styles import NamedStyle, numbers, Border, Side
from openpyxl.chart.label import DataLabel, DataLabelList
from openpyxl.drawing.line import LineProperties
from openpyxl.chart.shapes import GraphicalProperties


from vision_oslo_extension import oslo_extraction
from vision_oslo_extension.shared_contents import SharedMethods

def main(simname, main_option, time_start, time_end, option_select, text_input, low_v, high_v, time_step):
    
    print("")
    print("DC Falling Voltage Protection Assessment - - - > ")
    print("")
    
    simname = simname

    option = option_select

    start = 7 # result start from row 7
    space = 5
    time_increment = 5
    time_windows_total= ['15min','30min']

    # Option:
    # 1: DC FVP protection assessment
    # 2: 
    #check option for TCB or ETE is selected
    if option not in ["0","1","2","3","4"]:
        SharedMethods.print_message("ERROR: Error in DC FVP processing . Please contact Support...","31")
        return False

    if option == "0":
        SharedMethods.print_message("ERROR: Please select an option to proceed.","31")
        return False

    elif option == "1": #option for simple extraction
        #time_windows= ['15min']
        time_step = 5
    elif option == "2": #option for extraction with step outputs and graphs
        #functionality in development
        #time_windows= ['30min']
        time_step = 5

    elif option == "3": #option for simple extraction, no osop extraction required
        #time_windows= ['15min']
        time_step = 5
    elif option == "4": #option for extraction with step outputs and graphs, no osop extraction required
        #functionality in development
        #time_windows= ['30min']
        time_step = 5

    # Specify Excel file name
    excel_file = text_input + ".xlsx"

    if not SharedMethods.check_existing_file(excel_file):
        return False

    #check if main option is 13
    if main_option not in ["0","13"]:
        SharedMethods.print_message("ERROR: Invalid Selection. Contact Support. Issue in batch_processing.py","31")
        return False
    
    if main_option == "0":
        SharedMethods.print_message("ERROR: Please select an option and run again","31")
        return False

    try:
        #time_windows = ['15min','30min']
        #time_increment = 5
        
        #read excel and get branch list
        result = start_reading_process(simname,excel_file,option)
        if result == False:
            return False
        else:
            start_df,oslo_total = result

        #create data frame for processing branches, FVP coefficients will be added to this dataframe
        branch_process_df = start_df

        # read branch names from excel and create branch list for OSOP extraction
        branch_list_1 = read_all_branches(start_df)

        # create branch list 1 new in different format
        branch_list_1new = check_branch_list(branch_list_1)
        if branch_list_1new == False:
            return False

        # branch_list_1: branch format in XXXX/X format
        # branch_list_1new: branch format in XXXX-X format this is because many application does not allow / charaters in naming

        #create df for results
        results_df = start_df
        
        #check oof file of simname exist
        if option == "1" or option =="2":              
            if not SharedMethods.check_oofresult_file(simname):
                return False

        #get ratings data table from input spreadsheet
        rating_data = check_FVP_list(excel_file)

        #check all specified FVP ratings are within FVP list

        # Create a set of valid (FVP Type, FVP setting (kA)) combinations from rating_data
        valid_combinations = set(zip(rating_data['FVP Type'], rating_data['FVP setting (kA)']))
        branch_process_df_t = branch_process_df

        # Check if each (FVP Type, FVP setting (kA)) in branch_process_df is in the valid_combinations set
        branch_process_df_t['is_valid'] = branch_process_df_t.apply(lambda row: (row['FVP Type'], row['FVP setting (kA)']) in valid_combinations, axis=1)

        # If there are any invalid combinations, raise an error
        if not branch_process_df_t['is_valid'].all():
            invalid_rows = branch_process_df_t[~branch_process_df_t['is_valid']]
            SharedMethods.print_message(f"Invalid FVP Type or FVP setting (kA) compination found in the following rows:\n{invalid_rows}","31")
            return False
        #if no errors continue with extraction

        # extract the branches
        #extract all branch step output using OSOP for branches in branch list  
        #only extract branches if in option 1 or 2
        if option == "1" or option =="2":
            for branch_id in branch_list_1:
                if not oslo_extraction.branch_step(simname, time_start, time_end, "1", branch_id):
                    return False

        #read branch step voltage and current into data_frame from d4.osop
        branchstep_d4_df = branch_reading_process(simname,branch_list_1new)



        #merge ratings data coefficient to branch_process_df_r which is the dataframe for processing branch results
        print("merging FVP ratings to each branch")
        branch_process_df_r = pd.merge(branch_process_df,rating_data, on=[ 'FVP Type','FVP setting (kA)'],how='left')

        #change branch format for comparison in next step
        branch_process_df_r['OSLO']=branch_list_1new

        #for each branch step see if the voltage at the recorded current is below the FVP curves
        # 0 is pass, 1 is below the -% rating curve, 2 is below the 0% nominal rating curve, 3 is below the +% rating curve
        branchstep_r = branch_FVP_rating(branchstep_d4_df,branch_process_df_r)

        #count number of 0 1 2 3 in branchstep_r
        results_df['OSLO']=branch_list_1new
        results_df = Branch_FVP_sum(results_df,branchstep_r)

        #create substation summary for each substation
        sub_sum_df = Summary_by_substation(results_df)

        #export results summary to excel
        if not SEF_ratings_excel(simname,excel_file,option_select,time_start,time_end,time_step,text_input,time_increment,results_df,sub_sum_df):
            return False

        #the following are only requried for detailed step results and graphs run only with option 2 or 4
        if option == "2" or option =="4":

            #create a dictionary for each substation row in the above for a list of Nodes
            #dict_substation_stock added, same format as a dictionary of list of rolling stocks
            dict_substation, dict_substation_name, dict_substation_stock = substations_node_list(sub_sum_df,results_df)
            
            #get dictionary of rolling stock details
            rolling_stock_df = check_rolling_stock(excel_file)

            #dict_substation_stock = substations_stock(dict_substation_name,results_df)

            #create a dictionary for each substation that contains each Nodes step resutls
            dict_sub_step = substations_step_list(branchstep_r,dict_substation)

            #run the main loop for extracting step resutls to step_excel and plot graphs
            #graph plotting loop within this loop
            if not FVP_step_excel(simname,excel_file, text_input,results_df,sub_sum_df,dict_sub_step,dict_substation_name,rating_data,dict_substation_stock,rolling_stock_df):
                return False
    
    except KeyError as e:
        SharedMethods.print_message(f"ERROR: Unexpected error occured: {e}. Possibly due to incompleted data.","31")
        return False
    
    except Exception as e:
        SharedMethods.print_message(f"ERROR: Unexpected error occured: {e}","31")
        return False

    return True 
        

# read the start tab and collect informaiton
def start_reading_process(simname, excel_file,option):
    
    # File check
    print("Check Excel and Deleting Existing Contents ...")
    if option == "1":
        columns = ['Substation Name', 'OSLO', 'TCB Type', 'FVP Type','FVP setting (kA)','Rolling Stock Types']
        col_num = 6

    elif option == "2":
        columns = ['Substation Name', 'OSLO', 'TCB Type', 'FVP Type','FVP setting (kA)','Rolling Stock Types']
        col_num = 6

    elif option == "3":
        columns = ['Substation Name', 'OSLO', 'TCB Type', 'FVP Type','FVP setting (kA)','Rolling Stock Types']
        col_num = 6

    elif option == "4":
        columns = ['Substation Name', 'OSLO', 'TCB Type', 'FVP Type','FVP setting (kA)','Rolling Stock Types']
        col_num = 6

    try:
        wb = load_workbook(excel_file)
        # Delete all sheets except 'Start'
        for sheet_name in wb.sheetnames:
            if sheet_name == 'Start':
                result = check_reading_frame(wb[sheet_name],option) # list, dictionary type
                if result == False:
                    return False
                else:
                    #data_start_row,data_end_row,oslo_total,scenariolist = result
                    data_start_row,data_end_row,oslo_total = result
                
                start_df = pd.read_excel(excel_file,sheet_name = 'Start',header = 0,usecols=range(col_num), skiprows=data_start_row-1, nrows=oslo_total, names = columns)
                oslolist = start_df.iloc[:, 1].tolist()

                # check duplication in OSLO id
                if not SharedMethods.find_duplicates(oslolist): return False

                start_df['FVP setting (kA)'] = start_df['FVP setting (kA)'].astype(float)

                flag = False
                # check oslo list 
                if not option == "8": # do not check option 8 as the format is very different
                    for index, oslo in enumerate(oslolist):
                        if pd.isna(oslo):
                            SharedMethods.print_message(f"ERROR: Items {start_df.iloc[index,0]} is not assigned with OSLO ID","31")
                            flag = True

                # decision point
                if flag:
                    return False

            else:
                #delete other workbooks
                wb.remove(wb[sheet_name])
        
        # Save changes
        wb.save(excel_file)
        

    except Exception as e:
        SharedMethods.print_message(f"ERROR: (Close the Excel file and Start Again) Error: {e}","31")
        return False
    
    return start_df,oslo_total

# rating data frame
def check_reading_frame(sheet,option):
    table_start_row = 11

    # define the outage list section to be read # Compare against function above
    print("Check Data Entry & Scenario List...")
    if option == "1" or option == "2" or option == "3" or option == "4":
        table_row = 12
        table_start_column = 2
        # table_end_column = 11
        #table_row_2 = 12
        #table_start_column_2 = 26
    #elif option == "2":
    #    table_row = 12
    #    table_start_column = 2
        #table_row_2 = 12
        #table_start_column_2 = 9

   
    # create ole list and rating data
    # find total number of substations
    index = table_row
    column = table_start_column

    if sheet.cell(row=index, column=column).value is not None:
        while True:          
            index += 1
            check = sheet.cell(row=index, column=column).value
            if check is None:
                table_row_end = index
                oslo_total = index - table_row
                #oslo_total is number of substaions
                break
    else: 
        #if no substation is found
        SharedMethods.print_message("ERROR: Wrong data check Falling Voltage Protection ratings table in Cell J12")
        return False
    
    
    return table_start_row,table_row_end,oslo_total

#read FVP protection curves list
# check FVP curve data
def check_FVP_list(excel_file):
    wb = load_workbook(excel_file)
    #sheet=wb[sheet_name]
    print("Reading FVP curves Info ...")
    table_row = 12
    table_start_column = 8
    table_end_column = 21
    # create ole list and rating data
    Type_List = []
    Settiing_list = []

    columns = ['FVP Type','FVP setting (kA)', '-% A', '-% B','-% C','-% D', '0% A', '0% B','0% C','0% D', '+% A', '+% B','+% C','+% D']
    col_num = 'H:U'

    # check table length
    index = table_row
    column = table_start_column+2
    
    for sheet_name in wb.sheetnames:
        if sheet_name == 'Start':
            if wb[sheet_name].cell(row=index, column=column).value is not None:
                while True:          
                    index += 1
                    check = wb[sheet_name].cell(row=index, column=column).value
                    if check is None:
                        table_row_end = index
                        rating_total = index - table_row
                        #rating_total is number of FVP types and settings combinations
                        break

            else: 
            #if no ratings is found
                SharedMethods.print_message("ERROR: Wrong data format. No information at B12","31")
                return False
    

    start_df = pd.read_excel(excel_file,sheet_name = 'Start',header = 0,usecols=col_num, skiprows=table_row-2, nrows=rating_total, names = columns)

    rating_data = start_df

    return rating_data

#read Rolling Stock maximum current curves list
def check_rolling_stock(excel_file):
    wb = load_workbook(excel_file)
    print("Reading rolling stock Info ...")
    table_row = 12
    table_start_column = 23
    table_end_row = 22
    table_end_column = 24
    total_rolling_stock = 1
    
    #find ending column
    row_no = 11
    index = table_start_column
    for sheet_name in wb.sheetnames:
        if sheet_name == 'Start':
            if wb[sheet_name].cell(row=row_no, column=index).value is not None:
                while True:          
                    index += 1
                    check = wb[sheet_name].cell(row=row_no, column=index).value
                    if check is None:
                        table_end_column = index
                        total_rolling_stock = (index - table_start_column)//2
                        #total_rolling_stock is number of rolling stock types
                        break

            else: 
            #if rolling stock is found
                SharedMethods.print_message("ERROR: Wrong data format for rolling stock table. No information at W11","31")
                return False

    #create dictionary of data frames for each rolling stock
    rolling_stock_df = {}
    columns = ['Voltage','Current']
    i=0
    sheet_name = 'Start'
    while i <total_rolling_stock:
        rolling_stock_name = wb[sheet_name].cell(row=10, column=24+i*2).value
        col_num = [table_start_column+i*2-1,table_start_column+i*2]
        stock_df = pd.read_excel(excel_file,sheet_name = sheet_name,header = 0,usecols=col_num, skiprows=table_row-2, nrows=table_end_row-table_row+1, names = columns)
        rolling_stock_df [rolling_stock_name] = stock_df
        i +=1


    return rolling_stock_df
    

#read all branch list and create a list for extracting branch step outputs
def read_all_branches(start_df):

    columns = ['Substation Name', 'OSLO', 'TCB Type', 'FVP Type','FVP setting (kA)']

    bl_1 = pd.DataFrame(columns=columns)
    # copy branch list from start_df to bl_1
    bl_1 = start_df['OSLO'].tolist()
        
    #for each item in bl_1 check if it end with /E or /S add the opposite node if it not already exist in bl_1
    #not used for FVP
    
 
    return bl_1


# create branch list 1 in different format
# check branch list
def check_branch_list(branch_list):
    newlst = []
    for branch_id in branch_list:
            if len(branch_id) <= 6:
                if branch_id[-2:-1] =="/":
                    branch = branch_id[:len(branch_id)-2].ljust(4)+branch_id[-2:]
                    branch = branch[:4]+"-"+ branch[-1:]
                    newlst.append(branch)
                else:
                    SharedMethods.print_message(f"ERROR: Branch OSLO ID {branch_id} does not contains proper symbol. Check Input","31")
                    return False
            else:
                SharedMethods.print_message(f"ERROR: Branch OSLO ID {branch_id} is NOT in correct format. Check Input'.format(branch_id)","31")       
                return False

    return newlst

# read individual d4 file
def branch_reading_process(simname,branch_list):

    total = len(branch_list)

    # create dataframe
    branchstep_d4_df = {}

    for index, branch in enumerate(branch_list):
        print(f"Processing {branch} ...")
        filename = simname + "_" + branch +".osop.d4"
        delimiter = '\\s+'
        columns = ["BranchID","Type","Time","P_inst","Q_inst","Voltage","V_angle","Current","I_angle"]
        dtype_mapping = {"Time": str,}
        df = pd.read_csv(filename, delimiter=delimiter, names = columns, skiprows = 11,dtype = dtype_mapping) 
        # Extracting parts from the string and formatting the 'Time' column
        df['Time'] = df['Time'].apply(lambda x: f"{int(x[1:3]):02d}:{int(x[3:5]):02d}:{int(x[5:]):02d}")
        # data frame process
        # save the data
        
        branchstep_d4_df[branch] = df

    return branchstep_d4_df
    #this is the dataframe containing step outputs for each branch

#compare the ratings in each step result of each branch
def branch_FVP_rating(branchstep_d4_df,branch_process_df_r):
    #add a column of current in kA for later calculations
    for key, df in branchstep_d4_df.items():
        df["Current kA"] = df["Current"]/1000
        df["Voltage V"] = df["Voltage"]*1000       

    for key, df in branchstep_d4_df.items():

        print(f"Comparing step values to FVP curve in branch {key} ...")

        #find the row of this branch in the input branch list and get the FVP coefficients
        if key in branch_process_df_r['OSLO'].values:

            #matches the OSLO name from each key in branchstep_d4_df.items to the branch process dataframe
            matching_row = branch_process_df_r.loc[branch_process_df_r['OSLO'] == key]

            #if a match is found for that OSLO name, extract the Falling Voltage Protection coefficients from the branch_process_df_r row
            if not matching_row.empty:
                # Access values directly from the matching row
                #'-% A', '-% B','-% C','-% D', '0% A', '0% B','0% C','0% D', '+% A', '+% B','+% C','+% D'
                M_A = matching_row['-% A'].values[0]
                M_B = matching_row['-% B'].values[0]
                M_C = matching_row['-% C'].values[0]
                M_D = matching_row['-% D'].values[0]

                N_A = matching_row['0% A'].values[0]
                N_B = matching_row['0% B'].values[0]
                N_C = matching_row['0% C'].values[0]
                N_D = matching_row['0% D'].values[0]

                P_A = matching_row['+% A'].values[0]
                P_B = matching_row['+% B'].values[0]
                P_C = matching_row['+% C'].values[0]
                P_D = matching_row['+% D'].values[0]

                #these variables are the coefficients for calculating the FVP curve
                #M is minus, N for Nominal and P for plus
            
            else:
                #return error if cannot get coefficients
                SharedMethods.print_message(f"ERROR: Could not match branch names when comparing ratings in step results, could not get FVP coefficients","31")
       
        else:
                SharedMethods.print_message(f"ERROR: Could not match branch names when comparing ratings in step results","31")
        #df["Current kA"] = df["Current"]/1000  

        #create list to contain the test results
        FVP_test_results =[]

        #for each row (step values) in branch step data frame, after doing the FVP test add the results in 0, 1 , 2 ,3 number to a new key of "FVP_test" in the branchstep_d4_df
        for index, row in df.iterrows():
            FVP_test_result = FVP_test(row["Current kA"], row["Voltage V"], M_A, M_B, M_C, M_D, N_A, N_B, N_C, N_D, P_A, P_B, P_C, P_D)
            FVP_test_results.append(FVP_test_result)

        df["FVP_test"] = FVP_test_results

    return branchstep_d4_df
    #return branchstep_d4_df with the "FVP_test" added

#used for comparing the voltage against FVP curves
def FVP_test(Current, Voltage, M_A, M_B, M_C, M_D, N_A, N_B, N_C, N_D, P_A, P_B, P_C, P_D):
    if Voltage < Current**3 * P_A + Current**2 * P_B + Current*P_C +P_D:
        return 3
    elif Voltage < Current**3 * N_A + Current**2 * N_B + Current*N_C +N_D:
        return 2
    elif Voltage < Current** 3 * M_A + Current** 2 * M_B + Current*M_C +M_D:
        return 1
    else:
        return 0       

#used for counting FVP test results in each branch 
def Branch_FVP_sum(results_df,branchstep_r):

    print("Counting FVP exceedances for each branch")

    for index, row in results_df.iterrows():
        B_name = row['OSLO']

        branch_steps = branchstep_r[B_name]
        num_data = len(branch_steps)

        T_1 = branch_steps["FVP_test"].value_counts().get(1,0)
        T_2 = branch_steps["FVP_test"].value_counts().get(2,0)
        T_3 = branch_steps["FVP_test"].value_counts().get(3,0)


        results_df.at[index, 'Total Data Point'] = num_data
        results_df.at[index, 'Total No. Exceed rating lower trip limit'] = T_1 + T_2 + T_3
        results_df.at[index, 'Total No. Exceed rating nominal trip limit'] = T_2 + T_3
        results_df.at[index, 'Total No. Exceed rating upper trip limit'] = T_3

    return results_df

#create substation summary for each substation
#assume as same substation if substation name, TCB type, FVP type and FVP setting match
def Summary_by_substation(results_df):

    print("Creating Substation Summary for each substation")

    # Group by the required columns and sum the remaining columns
    sub_sum_df = results_df.groupby(["Substation Name", 'TCB Type', "FVP Type", "FVP setting (kA)"]).sum(numeric_only=True).reset_index()

    # Add a column with the count of rows in each group
    sub_sum_df['Number of Nodes'] = results_df.groupby(["Substation Name", 'TCB Type', "FVP Type", "FVP setting (kA)"]).size().values

    return sub_sum_df

#create a dictionary for each substation row defined as the above for a list of Nodes and list of rolling stock that lies within that substation
def substations_node_list(sub_sum_df,results_df):
    key_columns = ["Substation Name", "FVP Type", "FVP setting (kA)"]
    key_columns2 = ["Substation Name", 'TCB Type', "FVP Type", "FVP setting (kA)"]
    dict_substation = {}

    dict_substation_name = {}

    # add substation stock extractions
    dict_substation_stock = {}
    rolling_stock_list = []
    
    for index, row in sub_sum_df.iterrows():
        #Join the values of the selected columns to form a key
        key = '_'.join(str(row[col]) for col in key_columns) # Using string to create a composite key

        sub_name = row[key_columns2].to_frame().T

        OSLO_nodes = [] #create empty list of nodes
        rolling_stock_list = [] #create empty list of rolling stock

        for index2, row2 in results_df.iterrows():
            if row["Substation Name"] == row2["Substation Name"] and row["FVP Type"] == row2["FVP Type"]and row["FVP setting (kA)"] == row2["FVP setting (kA)"]:
                OSLO_nodes.append(row2['OSLO'])

                rolling_stock_types = row2['Rolling Stock Types']

                #check if there is valid entry, only add trains to rolling_stock_list_a if entry exist. 
                if isinstance(rolling_stock_types, str):                    
                    if len(rolling_stock_types)>3:
                        #check if the divider is within the entry 
                        if ", " in rolling_stock_types:
                            rolling_stock_list_a = rolling_stock_types.split(", ")
                        else:
                            rolling_stock_list_a = [rolling_stock_types]
                    else:
                        rolling_stock_list_a = []
                else:
                    rolling_stock_list_a = []

                #append rolling stock if not already in the list for this substation
                for rs in rolling_stock_list_a:
                    if rs not in rolling_stock_list:
                        rolling_stock_list.append(rs)


        #key (substation name) gives all OSLO nodes in the substation
        dict_substation[key] = OSLO_nodes
        #key (substation name) gives all settings in the substation
        dict_substation_name[key] = sub_name
        #key (substation name) gives all rolling stock in this substation
        dict_substation_stock[key] = rolling_stock_list

    return dict_substation, dict_substation_name, dict_substation_stock

#create a dictionary for each substation, based on the branch nodes in the dictionary in the previous part
#create a dictionary to contain the step results of all branches in each substations
def substations_step_list(branchstep_r,dict_substation):

    #dictionary containing all step results in each substation    
    dict_sub_step = {}

    for key, value in dict_substation.items():

        temp_df = pd.DataFrame()

        for oslo in value:

            if oslo in branchstep_r:
                temp_df = pd.concat([temp_df, branchstep_r[oslo]], ignore_index=True)

        # Assign the concatenated DataFrame to the dictionary
        dict_sub_step[key] = temp_df
    
    return dict_sub_step

#write to CSV
def FVP_sum_csv_out(simname,results_df):
    # write summary info to CSV file
    with open(simname+'_FVP_sum.csv','w',newline='') as f:
        
        writer = csv.writer(f) # create the csv writer
        row = [f'OSLO Branch FVP Summary Output']
        writer.writerow(row)

        # Write df1 to the file
        # df.insert(0, 'Blank', '')
        results_df.to_csv(f, index=False)
        
    print('\nBranch FVP Output Summary Completed.')

    return

#export results summary to excel
def SEF_ratings_excel(simname,excel_file,option_select,time_start,time_end,time_step,text_input,time_increment,results_df,sub_sum_df):

    columns1 = ['Substation Name', 'OSLO', 'TCB Type', 'FVP Type','FVP setting (kA)','Total Data Point','Total No. Exceed rating lower trip limit', 'Total No. Exceed rating nominal trip limit', 'Total No. Exceed rating upper trip limit']
    columns2 = ['Substation Name', 'TCB Type', 'FVP Type','FVP setting (kA)', 'Number of Nodes', 'Total Data Point' ,'Total No. Exceed rating lower trip limit', 'Total No. Exceed rating nominal trip limit', 'Total No. Exceed rating upper trip limit']
    #write to results sheet
    excel_b_results_df = pd.DataFrame(columns=columns1)
    #write to results summary sheet
    excel_s_results_df = pd.DataFrame(columns=columns2)

    excel_b_results_df = pd.DataFrame(results_df, columns=columns1)
    excel_s_results_df = pd.DataFrame(sub_sum_df, columns=columns2)
    
    # Calculate the percentage columns
    excel_b_results_df['% Exceed lower trip limit'] = excel_b_results_df['Total No. Exceed rating lower trip limit'] / excel_b_results_df['Total Data Point']
    excel_b_results_df['% Exceed nominal trip limit'] = excel_b_results_df['Total No. Exceed rating nominal trip limit'] / excel_b_results_df['Total Data Point']
    excel_b_results_df['% Exceed upper trip limit'] = excel_b_results_df['Total No. Exceed rating upper trip limit'] / excel_b_results_df['Total Data Point']

    excel_s_results_df['% Exceed lower trip limit'] = excel_s_results_df['Total No. Exceed rating lower trip limit'] / excel_s_results_df['Total Data Point']
    excel_s_results_df['% Exceed nominal trip limit'] = excel_s_results_df['Total No. Exceed rating nominal trip limit'] / excel_s_results_df['Total Data Point']
    excel_s_results_df['% Exceed upper trip limit'] = excel_s_results_df['Total No. Exceed rating upper trip limit'] / excel_s_results_df['Total Data Point']

    #write the lists to excel
    with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a',if_sheet_exists='overlay') as writer:

        # get currrent workbook to do process
        wb = writer.book
        print("Generate Branches Result excel Page...")
        start = 0
        r_start = start + 10
        excel_b_results_df.to_excel(writer, sheet_name="Results in branches", index=False, startrow = r_start, startcol = 1)
        
        # get currrent workbook to do process
        wb = writer.book
        print("Generate Substations results excel Page...")
        r_start = start + 10
        excel_s_results_df.to_excel(writer, sheet_name="Summary results each substation", index=False, startrow = r_start, startcol = 1)

        workbook = writer.book
        sheets = workbook['Start']
        sheet1 = workbook["Results in branches"]
        sheet2 = workbook["Summary results each substation"]
            
        project_name = sheets['B2'].value
        feeding_desp = sheets['B4'].value
        modeller = sheets['B5'].value
        #date = sheets['B6'].value

        # Add titles to results tabs
        sheet1['B2'].value = "Project Name:"
        sheet1['C2'].value = project_name
        sheet1['B3'].value = "Simulation Name:"
        sheet1['C3'].value = simname
        sheet1['B4'].value = "Feeding Arrangement:"
        sheet1['C4'].value = feeding_desp
        sheet1['B5'].value = "Result Created by:"
        sheet1['C5'].value = modeller
        sheet1['B6'].value = "Result Created at:"
        sheet1['C6'].value = datetime.now().strftime("%d-%m-%Y %H:%M")
        # Add titles to results tabs
        sheet2['B2'].value = "Project Name:"
        sheet2['C2'].value = project_name
        sheet2['B3'].value = "Simulation Name:"
        sheet2['C3'].value = simname
        sheet2['B4'].value = "Feeding Arrangement:"
        sheet2['C4'].value = feeding_desp
        sheet2['B5'].value = "Result Created by:"
        sheet2['C5'].value = modeller
        sheet2['B6'].value = "Result Created at:"
        sheet2['C6'].value = datetime.now().strftime("%d-%m-%Y %H:%M")

        #make small changes to title rows


        #apply resize and reformatting
        resize_columns(sheet1)
        apply_conditional_formatting(sheet1,3)

        resize_columns(sheet2)
        apply_conditional_formatting(sheet2,2)


        #format the new columns as percentage
        #percentage_format = NamedStyle(name="percentage")
        #percentage_format.number_format = numbers.FORMAT_PERCENTAGE_00

        for col in ['K', 'L', 'M']:  # for column K L M
            for cell in sheet1[col][1:]:
                cell.number_format = numbers.FORMAT_PERCENTAGE_00  # Format cell as percentage

        for col in ['K', 'L', 'M']:  # for column K L M
            for cell in sheet2[col][1:]:
                cell.number_format = numbers.FORMAT_PERCENTAGE_00  # Format cell as percentage


    return True

#read data from existing excel_file
def read_data_from_main_input(excel_file):
    with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a',if_sheet_exists='overlay') as writer:

        workbook = writer.book
        sheets = workbook['Start']

            
        project_name = sheets['B2'].value
        feeding_desp = sheets['B4'].value
        modeller = sheets['B5'].value

        return project_name, feeding_desp, modeller

#export step results to excel
def FVP_step_excel(simname,excel_file,text_input,results_df,sub_sum_df,dict_sub_step,dict_substation_name,rating_data,dict_substation_stock,rolling_stock_df):

    project_name, feeding_desp, modeller = read_data_from_main_input(excel_file)

    excel_file2 = text_input + "_step_outputs_plots.xlsx"

    columns2 = ['Substation Name', 'TCB Type', 'FVP Type','FVP setting (kA)', 'Number of Nodes' ,'Total No. Exceed rating lower trip limit', 'Total No. Exceed rating nominal trip limit', 'Total No. Exceed rating upper trip limit']
    #write to results summary sheet
    #excel_s_results_df = pd.DataFrame(columns=columns2)
    excel_s_results_df = pd.DataFrame(sub_sum_df, columns=columns2)

    # Check if the file exists
    # and write the substation results summary first

    print("Generate Step results...")
    r_start = 10

    if os.path.exists(excel_file2):
        
        # Delete the file if it exist
        os.remove(excel_file2)

        # File does not exist, create a new file
        with pd.ExcelWriter(excel_file2, engine='openpyxl') as writer:          
            excel_s_results_df.to_excel(writer, sheet_name="Sum results each substation", index=False, startrow = r_start, startcol = 1)

            workbook = writer.book
            sheet1 = workbook["Sum results each substation"]
            # Add titles to results tabs
            sheet1['B2'].value = "Project Name:"
            sheet1['C2'].value = project_name
            sheet1['B3'].value = "Simulation Name:"
            sheet1['C3'].value = simname
            sheet1['B4'].value = "Feeding Arrangement:"
            sheet1['C4'].value = feeding_desp
            sheet1['B5'].value = "Result Created by:"
            sheet1['C5'].value = modeller
            sheet1['B6'].value = "Result Created at:"
            sheet1['C6'].value = datetime.now().strftime("%d-%m-%Y %H:%M")            

    # File does not exist, create a new file
    else:
        with pd.ExcelWriter(excel_file2, engine='openpyxl') as writer:          
            excel_s_results_df.to_excel(writer, sheet_name="Sum results each substation", index=False, startrow = r_start, startcol = 1)

            workbook = writer.book
            sheet1 = workbook["Sum results each substation"]
            # Add titles to results tabs
            sheet1['B2'].value = "Project Name:"
            sheet1['C2'].value = project_name
            sheet1['B3'].value = "Simulation Name:"
            sheet1['C3'].value = simname
            sheet1['B4'].value = "Feeding Arrangement:"
            sheet1['C4'].value = feeding_desp
            sheet1['B5'].value = "Result Created by:"
            sheet1['C5'].value = modeller
            sheet1['B6'].value = "Result Created at:"
            sheet1['C6'].value = datetime.now().strftime("%d-%m-%Y %H:%M")

    #write the step results to a new excel file
    with pd.ExcelWriter(excel_file2, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:

        dict_substation_rating_data = {}

        #for each substation dataframes exist in the dictionary of dict_sub_step
        for key, branch_step in dict_sub_step.items():

            print(f"Adding A sheet for substation {key} ...")
            #print("Adding A sheet for each substation...")

            #make sure the sheet_name is never longer than 30 characters
            if len(key) <31:
                sheet_name_variable = key
            else:
                sheet_name_variable = key[-30:]

            #write the branch step dataframe to excel with sheet name the same name as the key
            branch_step.to_excel(writer, sheet_name=sheet_name_variable,index=False, startrow = r_start, startcol = 1)

            workbook = writer.book
            sheet2 = workbook[sheet_name_variable]

            #write titles in the worksheet
            sheet2['B2'].value = "Project Name:"
            sheet2['C2'].value = project_name
            sheet2['B3'].value = "Simulation Name:"
            sheet2['C3'].value = simname
            sheet2['B4'].value = "Feeding Arrangement:"
            sheet2['C4'].value = feeding_desp
            sheet2['B5'].value = "Result Created by:"
            sheet2['C5'].value = modeller
            sheet2['B6'].value = "Result Created at:"
            sheet2['C6'].value = datetime.now().strftime("%d-%m-%Y %H:%M")

            #write 'Substation Name', 'TCB Type', 'FVP Type','FVP setting (kA)'
            sheet2['B7'].value = "Substation Name:"
            sheet2['C7'].value = dict_substation_name[key]['Substation Name'].iloc[0]
            sheet2['B8'].value = "TCB Type:"
            sheet2['C8'].value = dict_substation_name[key]['TCB Type'].iloc[0]
            sheet2['B9'].value = "FVP Type:"
            sheet2['C9'].value = dict_substation_name[key]['FVP Type'].iloc[0]
            sheet2['B10'].value = "FVP setting (kA):"
            sheet2['C10'].value = dict_substation_name[key]['FVP setting (kA)'].iloc[0]

            #write ratings coefficient in the worksheet
            #merge ratings data coefficient to branch_process_df_r which is the dataframe for processing branch results
            dict_substation_rating_data[key] = pd.merge(dict_substation_name[key],rating_data, on=['FVP Type','FVP setting (kA)'],how='left')

            sheet2['E6'].value = "FVP coefficients"
            sheet2['F6'].value = "for plotting curves"

            sheet2['E7'].value = "FVP Type:"
            sheet2['F7'].value = "FVP Setting (kA):"
            sheet2['G7'].value = "-% A (^3)"
            sheet2['H7'].value = "-% B (^2)"
            sheet2['I7'].value = "-% C (^1)"
            sheet2['J7'].value = "-% D "
            sheet2['K7'].value = "100% A (^3)"
            sheet2['L7'].value = "100% B (^2)"
            sheet2['M7'].value = "100% C (^1)"
            sheet2['N7'].value = "100% D "
            sheet2['O7'].value = "+% A (^3)"
            sheet2['P7'].value = "+% B (^2)"
            sheet2['Q7'].value = "+% C (^1)"
            sheet2['R7'].value = "+% D "

            #copy values for 'FVP Type','FVP setting (kA)', '-% A', '-% B','-% C','-% D', '0% A', '0% B','0% C','0% D', '+% A', '+% B','+% C','+% D' in
            sheet2['E8'].value = dict_substation_rating_data[key].loc[0,'FVP Type']
            sheet2['F8'].value = dict_substation_rating_data[key].loc[0,'FVP setting (kA)']
            sheet2['G8'].value = dict_substation_rating_data[key].loc[0,'-% A']
            sheet2['H8'].value = dict_substation_rating_data[key].loc[0,'-% B']
            sheet2['I8'].value = dict_substation_rating_data[key].loc[0,'-% C']
            sheet2['J8'].value = dict_substation_rating_data[key].loc[0,'-% D']
            sheet2['K8'].value = dict_substation_rating_data[key].loc[0,'0% A']
            sheet2['L8'].value = dict_substation_rating_data[key].loc[0,'0% B']
            sheet2['M8'].value = dict_substation_rating_data[key].loc[0,'0% C']
            sheet2['N8'].value = dict_substation_rating_data[key].loc[0,'0% D']
            sheet2['O8'].value = dict_substation_rating_data[key].loc[0,'+% A']
            sheet2['P8'].value = dict_substation_rating_data[key].loc[0,'+% B']
            sheet2['Q8'].value = dict_substation_rating_data[key].loc[0,'+% C']
            sheet2['R8'].value = dict_substation_rating_data[key].loc[0,'+% D']           

        #run plot graphs here for same with loop
        if not  plot_scatter(writer, dict_sub_step,dict_substation_stock,rolling_stock_df,dict_substation_name):
            return False


        print("Saving step results and in Excel, will take some time")

    return True

#plot the graphs in all sheets created in FVP_step_excel()
#for running within FVP_step_excel()
#added plotting rolling stock
def plot_scatter(writer, dict_sub_step,dict_substation_stock,rolling_stock_df,dict_substation_name):

    #excel_file2 = text_input + "_step_outputs.xlsx"

    # Current (kA) values in 0.5 increment from 0 to 10
    values = [i * 0.5 for i in range(21)]

    #get same sheet names as FVP_step_excel()
    # open the relevant excel sheets and add plot
    #with pd.ExcelWriter(excel_file2, engine='openpyxl', mode='a') as writer:

    for key, branch_step in dict_sub_step.items():

        print(f"Adding scatter plot for substaion {key} ...")

        #make sure the sheet_name is never longer than 30 characters
        if len(key) <31:
            sheet_name_variable = key
        else:
            sheet_name_variable = key[-30:]

        #copy ratings values in spreadsheet for plotting
        workbook = writer.book
        sheet1 = workbook[sheet_name_variable]


        #information for chart new title format
        Sub_name = dict_substation_name[key]['Substation Name'].iloc[0]
        TCB_type = dict_substation_name[key]['TCB Type'].iloc[0]
        FVP_type = dict_substation_name[key]['FVP Type'].iloc[0]
        FVP_setting = dict_substation_name[key]['FVP setting (kA)'].iloc[0]

        plus_minus = "\u00B1"  # ± symbol
        if FVP_type == "HSL" :
            tol = "\u00B1" + "10%"
        else:
            tol = "\u00B1" + "15%"               

        sheet1['G6'].value = tol

        a1 = sheet1['G8'].value  # Coefficient for x^3
        b1 = sheet1['H8'].value  # Coefficient for x^2
        c1 = sheet1['I8'].value  # Coefficient for x
        d1 = sheet1['J8'].value  # Constant term

        a2 = sheet1['K8'].value  # Coefficient for x^3
        b2 = sheet1['L8'].value  # Coefficient for x^2
        c2 = sheet1['M8'].value  # Coefficient for x
        d2 = sheet1['N8'].value  # Constant term

        a3 = sheet1['O8'].value  # Coefficient for x^3
        b3 = sheet1['P8'].value  # Coefficient for x^2
        c3 = sheet1['Q8'].value  # Coefficient for x
        d3 = sheet1['R8'].value  # Constant term

        sheet1['O11'].value = "Current (kA)"
        sheet1['P11'].value = "-% Voltage (V) lower trip limit"
        sheet1['Q11'].value = "100% Voltage (V) nominal trip limit"
        sheet1['R11'].value = "+% Voltage (V) upper trip limit"

        # Write the values in column O from O12 onwards
        for idx, value in enumerate(values, start=12):
            sheet1[f'O{idx}'].value = value

        # For -% lower trip limit Calculate the polynomial values and write them in column P from P12 onwards
        for idx, value in enumerate(values, start=12):
            x = value
            poly_value = a1 * x**3 + b1 * x**2 + c1 * x + d1
            sheet1[f'P{idx}'].value = poly_value

        # For 100% Voltage (V) nominal trip limit Calculate the polynomial values and write them in column Q from P12 onwards
        for idx, value in enumerate(values, start=12):
            x = value
            poly_value = a2 * x**3 + b2 * x**2 + c2 * x + d2
            sheet1[f'Q{idx}'].value = poly_value

        # For +% Voltage (V) upper trip limit Calculate the polynomial values and write them in column R from P12 onwards
        for idx, value in enumerate(values, start=12):
            x = value
            poly_value = a3 * x**3 + b3 * x**2 + c3 * x + d3
            sheet1[f'R{idx}'].value = poly_value

        # Find the last non-empty row in columns K and L
        max_row = sheet1.max_row
        for row in range(12, max_row + 1):
            if sheet1[f'K{row}'].value is None and sheet1[f'L{row}'].value is None:
                max_row = row - 1
                break

        # Under the coefficients in column O, enter the rolling stock details for plotting
        max_row2 = max_row

        #find the max row in Column O and P
        for row in range(12, max_row + 1):
            if sheet1[f'O{row}'].value is None and sheet1[f'P{row}'].value is None:
                max_row2 = row - 1
                break

        max_row2 = max_row2+2
        

        sheet1[f'O{max_row2}'].value = "All Rolling Stocks Used"
        row=max_row2+1
        column = 14
        
        # find number of rolling stock 
        no_rolling_stock = len(dict_substation_stock[key])
        
        rs_list= dict_substation_stock[key]

        # for each rolling stock in rs_list
        #check for all avalaible rolling stock details in rolling stock data frame

        print(f"Adding rolling stock information for substaion {key} ...")

        for key2, rs_details in rolling_stock_df.items():
            #if this key is in hte rs list then add it to the sheet
            if key2 in rs_list:
                sheet1.cell(row=row, column =column+1, value=key2)
                rs_details.to_excel(writer, sheet_name=sheet_name_variable, index=False, startrow = row, startcol = column)
                column = column+2

        # Create a scatter chart
        chart = ScatterChart()
        chart.title = f"FVP Plot, {Sub_name} ({FVP_type} {FVP_setting}kA, {tol})"
        chart.x_axis.title = "Current (kA)"
        chart.y_axis.title = "Voltage (V)"

        # Set the y-axis limits
        chart.y_axis.scaling.min = 300  # Replace with your desired min y value
        chart.y_axis.scaling.max = 900  # Maximum Voltage to Plot
        # Set the x-axis limits
        chart.x_axis.scaling.min = 0  # Replace with your desired min x value
        chart.x_axis.scaling.max = 8  # Maximum Current (kA) to Plot

        #set chart size
        chart.width = 25
        chart.height = 13

        # Define data for the scatter plot (from branch_step)
        x_values_scatter = Reference(sheet1, min_col=11, min_row=12, max_row=max_row)  # Column K
        y_values_scatter = Reference(sheet1, min_col=12, min_row=12, max_row=max_row)  # Column L
        series1 = Series(y_values_scatter, x_values_scatter,title_from_data=False, title="Voltage vs Current Scatter")
                    
        series1.graphicalProperties.line.noFill = True
        series1.marker.symbol = "circle"
        series1.marker.size = 2
        series1.marker.graphicalProperties.line.noFill = True
        chart.series.append(series1)

        # Define -% lower trip limit data for the curve (columns O and P)
        x_values_curve1 = Reference(sheet1, min_col=15, min_row=12, max_row=32)  # Column O
        y_values_curve1 = Reference(sheet1, min_col=16, min_row=12, max_row=32)  # Column P
        series2 = Series(y_values_curve1, x_values_curve1,title_from_data=False, title="lower trip limit")
        series2.graphicalProperties.line.solidFill = ColorChoice(prstClr="red")
        series2.graphicalProperties.line.width = 15000  # Width = 1.5 pt (Openpyxl uses EMUs, 12700 EMUs = 1pt)
        series2.graphicalProperties.line.dashStyle = "sysDash"  # Dashed line

        # Create data label list and add a label to a specific point (e.g., the last point)
        dLbls = DataLabelList()
        dLbl = DataLabel(idx=12)  # Set the index of the data point
        dLbl.showVal = False # Hide value
        dLbl.showSerName = True  # Show the series name for the label
        
        # Add a colored border to the data label
        dLbl.spPr = GraphicalProperties()
        dLbl.spPr.ln = LineProperties(solidFill="FF0000", w=12000)  # width in EMUs, 12000 EMUs = 1.2 point
        
        # Add a solid white background to the data label
        dLbl.spPr.solidFill = "FFFFFF"  # White background

        dLbls.dLbl.append(dLbl)
        # Assign data labels to the series
        series2.dLbls = dLbls

        chart.series.append(series2)

        # Define nominal trip limit data for the curve (columns O and Q)
        x_values_curve2 = Reference(sheet1, min_col=15, min_row=12, max_row=32)  # Column O
        y_values_curve2 = Reference(sheet1, min_col=17, min_row=12, max_row=32)  # Column Q
        series3 = Series(y_values_curve2, x_values_curve2,title_from_data=False, title="nominal trip limit")
        series3.graphicalProperties.line.solidFill = "000000"  # Black color
        series3.graphicalProperties.line.width = 15000  # Width = 1.5 pt

        # Create data label list and add a label to a specific point (e.g., the last point)
        dLbls = DataLabelList()
        dLbl = DataLabel(idx=13)  # Set the index of the data point
        dLbl.showVal = False # Hide value
        dLbl.showSerName = True  # Show the series name for the label

        # Add a colored border to the data label
        dLbl.spPr = GraphicalProperties()
        dLbl.spPr.ln = LineProperties(solidFill="000000", w=12000)  # width in EMUs, 12000 EMUs = 1.2 point
        
        # Add a solid white background to the data label
        dLbl.spPr.solidFill = "FFFFFF"  # White background

        dLbls.dLbl.append(dLbl)
        # Assign data labels to the series
        series3.dLbls = dLbls

        chart.series.append(series3)

        # Define nominal trip limit data for the curve (columns O and R)
        x_values_curve3 = Reference(sheet1, min_col=15, min_row=12, max_row=32)  # Column O
        y_values_curve3 = Reference(sheet1, min_col=18, min_row=12, max_row=32)  # Column R
        series4 = Series(y_values_curve3, x_values_curve3,title_from_data=False, title="upper trip limit")
        series4.graphicalProperties.line.solidFill = ColorChoice(prstClr="red")
        series4.graphicalProperties.line.width = 15000  # Width = 1.5 pt
        series4.graphicalProperties.line.dashStyle = "sysDash"  # Dashed line

        # Create data label list and add a label to a specific point (e.g., the last point)
        dLbls = DataLabelList()
        dLbl = DataLabel(idx=13)  # Set the index of the data point
        dLbl.showVal = False # Hide value
        dLbl.showSerName = True  # Show the series name for the label

        # Add a colored border to the data label
        dLbl.spPr = GraphicalProperties()
        dLbl.spPr.ln = LineProperties(solidFill="FF0000", w=12000)  # width in EMUs, 12000 EMUs = 1.2 point
        
        # Add a solid white background to the data label
        dLbl.spPr.solidFill = "FFFFFF"  # White background

        dLbls.dLbl.append(dLbl)
        # Assign data labels to the series
        series4.dLbls = dLbls

        chart.series.append(series4)


        row=max_row2+3
        column = 14
        max_row = sheet1.max_row

        #find the max row in Column O and P for rolling stock info
        for rows in range(max_row2, max_row + 1):
            if sheet1[f'O{rows}'].value is None and sheet1[f'P{rows}'].value is None:
                max_row = rows - 1
                break

        # Define a list of custom colors (avoiding blue, black, and red)
        custom_colors = [
            "00FF00",  # Green
            "FFA500",  # Orange
            "800080",  # Purple
            "FFFF00",  # Yellow
            "00CED1",  # DarkTurquoise
            "FF69B4",  # HotPink
            "8B4513",  # SaddleBrown
            "4B0082",  # Indigo
            "A52A2A",  # Brown
            "7FFF00",  # Chartreuse
        ]

        #add the rolling stocks to scatter graph
        for rs in range(no_rolling_stock):
            # Define the x and y values for the current rolling stock series
            x_values_rs = Reference(sheet1, min_col=16+rs*2, min_row=row, max_row=max_row)
            y_values_rs = Reference(sheet1, min_col=15+rs*2, min_row=row, max_row=max_row)
            # use rolling stock name as title of the series
            title = sheet1.cell(row=row-2, column =15+rs*2).value
            series5 = Series(y_values_rs, x_values_rs,title_from_data=False, title=title)

            # Assign a custom color from the list (loop over the list if more series than colors)
            color = custom_colors[rs % len(custom_colors)]

            series5.graphicalProperties.line.solidFill = color  # use the above color
            series5.graphicalProperties.line.width = 20000  # Width = 2 pt

            # Create data label list and add a label to a specific point (e.g., the last point)
            dLbls = DataLabelList()

            dLbl = DataLabel(idx=7)  # Set the index of the data point
            dLbl.showVal = False # Hide value
            dLbl.showSerName = True  # Show the series name for the label

            # Add a colored border to the data label
            dLbl.spPr = GraphicalProperties()
            dLbl.spPr.ln = LineProperties(solidFill=color, w=12000)  # width in EMUs, 12000 EMUs = 1.2 point
            
            # Add a solid white background to the data label
            dLbl.spPr.solidFill = "FFFFFF"  # White background

            dLbls.dLbl.append(dLbl)

            # Assign data labels to the series
            series5.dLbls = dLbls

            # Append the series to the chart
            chart.series.append(series5)

        chart.legend.position = 'b' # below legend position

        sheet1.add_chart(chart, "T3")

        #apply resize columns
        resize_columns(sheet1)

    return True




def resize_columns(sheet):
    for column in sheet.columns:
        max_length = 0
        column = column[0].column_letter
        for cell in sheet[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 1)
        sheet.column_dimensions[column].width = adjusted_width

def apply_conditional_formatting(sheet,column):
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red (0% S)
    amber_fill = PatternFill(start_color="FFBF00", end_color="FFBF00", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 255,152,51 (80% S)
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green (00% S)

    center_alignment = Alignment(horizontal='center')

    for row in sheet.iter_rows(min_row=12, max_row=sheet.max_row, min_col=2, max_col=14):
        for cell in row:
            if cell.column == 10:
                Red = cell.value
            
            if cell.column == 9:
                Amber = cell.value

            if cell.column == 8:
                Yellow = cell.value

            if cell.column == column:
                Name = cell.value

        if Name is not None:

            if Red > 0:
                sheet.cell(row=row[0].row, column=column).fill = red_fill

            elif Amber > 0:
                sheet.cell(row=row[0].row, column=column).fill = amber_fill

            elif Yellow > 0:
                sheet.cell(row=row[0].row, column=column).fill = yellow_fill

            else:
                sheet.cell(row=row[0].row, column=column).fill = green_fill

#============================================================================
# Check if the script is run as the main module
if __name__ == "__main__":
    # Add your debugging code here
    simname = "BEL013"  # Provide a simulation name or adjust as needed
    main_option = "13"  # Adjust as needed
    time_start = "0100000"  # Adjust as needed
    time_end = "0160000"  # Adjust as needed
    option_select = "4"  # Adjust as needed
    text_input = "FVP Wallington"  # Adjust as needed
    low_v = None  # Adjust as needed
    high_v = None  # Adjust as needed
    time_step = 900  # Adjust as needed

    # Call your main function with the provided parameters
    main(simname, main_option, time_start, time_end, option_select, text_input, low_v, high_v, time_step)