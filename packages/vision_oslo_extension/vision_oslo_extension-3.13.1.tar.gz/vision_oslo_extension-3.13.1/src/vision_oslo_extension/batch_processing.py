#
# -*- coding: utf-8  -*-
#=================================================================
# Created by: Jieming Ye
# Created on: Feb 2024
# Last Modified: Feb 2024
#=================================================================
# Copyright (c) 2024 [Jieming Ye]
#
# This Python source code is licensed under the 
# Open Source Non-Commercial License (OSNCL) v1.0
# See LICENSE for details.
#=================================================================
"""
Pre-requisite: 
SimulationName.oof: default oslo output file after successfully run a simulation.
FeederList.txt: feeder list config file for processing.
BranchNodeList.txt: node list config file for processing.
GridAllocation.csv: configuration file for processing.
Used Input:
simname: to locate the result file or file rename
time_start: to define the analysis start time
time_end: to define the analysis end time
option_select: to define the option selected in subfunction
text_input: to locate the excel file name
other input for oslo_extraction.py only.
Expected Output:
Various .csv file containing individual detailed output or summary output.
Description:
This script defines the process of DC data calculation.
In substation_process(), it extract all feeder step output first (due to the fact that the current list output does not contain feeder information directly). Then in d4_file_process, it reads the individual d4 files, doing proper calculation. And then output the result to .csv files in other sub functions.
In grid_power_processing(), this reads the csv information and generate a excel with all summary information included. The output selection could be further filtered or refined in the grid_summary_ouptut() function.
In branch_data_process(), it generates the list file first and read list file line by line, saving the result is a data frame during the reading process. It then doing individual calculation and output the result to several .csv files. The process is relative easy to follow via reading the code.

"""
#=================================================================
# VERSION CONTROL
# V1.0 (Jieming Ye) - Initial Version
# V1.1 (Jieming Ye) - 2024-13-04: Consider edge case of duplication and empty grid allocation info
# V2.0 (Jieming Ye) - Allow User Setting of Grid Calculation (Different Output Format)
#=================================================================
# Set Information Variable
# N/A
#=================================================================

import os
import math
import csv

import pandas as pd
import numpy as np

from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, NamedStyle

from vision_oslo_extension import oslo_extraction
from vision_oslo_extension.shared_contents import SharedMethods

def main(simname, main_option, time_start, time_end, option_select, text_input, low_v, high_v, time_step):

    #User Interface - Welcome message:
    print("")
    print("Data Batch Processing - - - > ")
    print("")
    print(f"\nOption Selected --> Option {main_option}")
    print(f"Sub-option Selected --> Option {option_select}")
    
    # get simulation name name from input
    print("Checking Result File...")

    if main_option not in ["0","9","10"]:
        SharedMethods.print_message("ERROR: Invalid Selection. Contact Support. Issue in batch_processing.py","31")
        return False
    
    if main_option == "0":
        SharedMethods.print_message("ERROR: Please select an option and run again","31")
        return False
    
    if not (main_option == "9" and option_select == "2"): # if for main =9 and option =2, no need to check file   
        check = SharedMethods.check_oofresult_file(simname)
        if check == False:
            return False

    if main_option == "9":

        time_windows = ["1min","4min","5min","15min","30min","60min","120min","180min","P30min"]
        time_increment = 5

        if not substation_process(simname, option_select,time_start, time_end,time_windows,time_increment):
            return False

    if main_option == "10":
        text_input = 'BranchNodeList.txt'
        time_increment = 5
        if not branch_data_process(simname, option_select,time_start, time_end, time_step,text_input,time_increment):
            return False

    return True

#########################
# calculate rms
def calculate_rolling_rms(data, column, window_size, time_increment):
    return data[column].rolling(window=int(window_size / time_increment)).apply(lambda x: np.sqrt(np.mean(x**2)), raw=True)

#==============Option 9 for DC process====================================(TOO DIFFICULT OMG)
def substation_process(simname, option_select,time_start, time_end,time_windows,time_increment):
    if option_select not in ["0","1","2"]:
        SharedMethods.print_message("ERROR: Contact Support. Issue in batch_processing.py --> substation_process","31")
        return False
    
    # User defined time windows extraction
    time_start = SharedMethods.time_input_process(time_start,1)
    time_end = SharedMethods.time_input_process(time_end,1)

    if time_start == False or time_end == False:
        return False
    
    if option_select == "1":
        if not suboption1_calculation_main_process(simname,time_start,time_end,time_windows,time_increment):
            return False
        
    if option_select == "2":

        # read the grid allocation csv
        print("Checking input files...")
        name = 'GridAllocation.csv'
        if not os.path.isfile(name):
            SharedMethods.print_message("ERROR: Please provide the GridAllocation.csv. Process terminated...","31")
            return False 
        
        columns = ['Substation','OSLO ID','Grid Feeder','Assessed Time']
        grid_df = pd.read_csv(name, usecols=[0, 1, 2, 3])  # Specify column indices
        grid_df.columns = columns # rename with programming

        # check duplication (JY:2024-12-04)
        duplicated_oslo_ids = grid_df[grid_df["OSLO ID"].duplicated(keep=False)]
        if not duplicated_oslo_ids.empty:
            duplication = duplicated_oslo_ids["OSLO ID"].unique()
            SharedMethods.print_message(f"ERROR: Substation ID: {duplication} was used at least twice. Check your Grid Allocation.","31")
            return False

        assessed_time = [] # in minutes can be either int or float
        # check if the list contents is a number
        for item in grid_df['Assessed Time'].unique().tolist():
            if str(item) == 'nan':
                continue
            try:
                minutes = float(item)
                seconds = minutes * 60
                # if int(item) == 0: # if user mannualy enter 0, it will be ignored.
                #     SharedMethods.print_message(f"WARNING: Time window {int(item)} does not fit time increment. Ignored...","33")
                #     continue
                if seconds % time_increment == 0: 
                    if minutes.is_integer():
                        assessed_time.append(int(item))
                    else:
                        assessed_time.append(float(item))
                else:
                    SharedMethods.print_message(f"WARNING: Time window {float(item)} min does not fit time increment. Ignored...","33")
            except ValueError:
                SharedMethods.print_message(f"WARNING: Assessed Time Input {item} is not valid seconds. Ignored...","33")
        
        # check if the list is empty
        if assessed_time == []:
            SharedMethods.print_message(f"WARNING: No time windows settings detected. The default 30 min will be used...","33")
            assessed_time.append(30)
            
        calculation_flag = False
        for item in assessed_time:
            file = simname + f'_{item}min_AVEpower_sum.csv'
            if not os.path.isfile(file):
                SharedMethods.print_message(f"WARNING: The summary file {file} does not exist. Will try to generate all files...","33")
                calculation_flag = True
            
        # prepare for the grid calculation using new time window
        if calculation_flag:
            print("Preparing for Grid Calculation...")
            # sort the assessed time from small to big
            assessed_time.sort()
            # remove the last item from the time window
            time_windows = time_windows[:-1]
            # add assessed time to time windows
            for time in assessed_time:
                if isinstance(time, int):
                    time_windows.append(f"P{time}min")
                else:
                    time_windows.append(f"P{time:.2f}min") # two decimal place
            # redo option1 with new time windows for the grid calculation
            suboption1_calculation_main_process(simname,time_start,time_end,time_windows,time_increment)
                                        
        # process all time windows one by one
        result = one_time_window_grid_calculation(simname, grid_df,assessed_time,time_increment)
        if not result:
            return False
        else:
            sum_df_1, sum_df_2 = result    

        # print result to excel
        result_to_excel(simname, assessed_time, sum_df_1,sum_df_2)
        
    return True

##==================suboption 1
def suboption1_calculation_main_process(simname,time_start,time_end,time_windows,time_increment):
    feeder_list = d4_file_check("FeederList.txt",simname)      
    if feeder_list == False:
        return False
    if feeder_list == True:
        if not oslo_extraction.feeder_step(simname,time_start,time_end,"2",""):
            SharedMethods.print_message(f"WARNING: Error will be ignored and process continued...","33")
    
    feeder_list = d4_file_check("FeederList.txt",simname) # incase version error that no file gets generated
    if feeder_list == False or feeder_list == True:
        return False
    
    print("Pre-checking completed. Ready to GO!")
    
    # Processing start
    print("\nProcessing Initiated...")

    df_list = [] #list to save individual dataframe d4 (including RMS calculation)
    sum_df_list = [] # list to calculate maximum

    for index, feeder in enumerate(feeder_list):

        print(f"Processing {feeder}...")

        sum_df,df = d4_file_process(simname,feeder,time_windows,time_increment)
        df_list.append(df)
        sum_df_list.append(sum_df)

    # writing summary to csv file    
    print('')
    print('Write summary to csv file...')
    # write indiviual time frame summary    
    rms_summary_output(simname,df_list,sum_df_list,time_windows,feeder_list)
    # write single summary to csv
    final_sum(simname,sum_df_list,time_windows,feeder_list)

    return True
# module to check d4 file exist or not  
def d4_file_check(filename,simname):
    print('')
    print("Pre-check: Have you provided the ""FeederList.txt"" file?")
    print("Checking...")

    if not os.path.isfile(filename):
        SharedMethods.print_message("ERROR: Please provide the FeederList.txt file. Process terminated...","31")
        return False
    else:
        print("FeederList input file exist. Continue...")

    print("Checking essential d4 files...")
    
    # reading the Feeder List file
    text_input = []
    with open(filename) as fbrlist:
        for index, line in enumerate(fbrlist):
            text_input.append(line[:50].strip())

    for items in text_input:
        name = simname + "_" + items + ".osop.d4"
    
        if not os.path.isfile(name): # if the d4 file does not exist
            SharedMethods.print_message(f"WARNING: Supply Point d4 file {name} does not exist. Will auto-generate all files...","33")
            return True

    return text_input

# process single d4 files
def d4_file_process(simname,feeder,time_windows,time_increment):

    name = simname + "_" + feeder + ".osop.d4"
    # New section included here
    # Reding files using panda
    delimiter = '\\s+'
    columns = ["SupplyID","Type","Time","P_inst","Q_inst","Voltage","V_angle","Current","I_angle"]
    dtype_mapping = {"Time": str,}
    df = pd.read_csv(name, delimiter=delimiter, names = columns, skiprows = 11,dtype = dtype_mapping)  # You might need to adjust the delimiter based on your file
    # Extracting parts from the string and formatting the 'Time' column
    df['Time'] = df['Time'].apply(lambda x: f"{int(x[1:3]):02d}:{int(x[3:5]):02d}:{int(x[5:]):02d}")

    df['I_Square'] = df['Current']**2

    time_interval = []
    p_time_interval = []
    for time in time_windows:
        if time[0] == "P":
            min = time[1:-3]
            p_time_interval.append(math.ceil(float(min)*60))
        else:
            min = time[:-3]
            time_interval.append(math.ceil(float(min)*60))
            
    # current process
    # column name list
    column_name_list = []
    for index, sec in enumerate(time_interval):
        column_name = f'I_{time_windows[index]}_RMS (A)'
        column_name_list.append(column_name)
        df[column_name] = calculate_rolling_rms(df, 'Current', sec, time_increment)
    
    # power process 
    for index, sec in enumerate(p_time_interval):
        minute = sec/60
        if minute.is_integer():
            column_name = f'P_{int(minute)}min_Ave (MW)'
        else:
            column_name = f'P_{minute:.2f}min_Ave (MW)'
        if index == 0:
            first_power_column = column_name
    # powersec = 1800
    # column_name = 'P_30min_Ave (MW)'
        column_name_list.append(column_name)
        df[column_name] = df['P_inst'].rolling(window=int(sec/time_increment)).mean()

    # insert empty columns
    df.insert(df.columns.get_loc('I_Square'),'New_C_1', np.nan)
    df.insert(df.columns.get_loc(first_power_column),'New_C_2', np.nan)

    # finding min max summary
    sum_df = pd.DataFrame(columns=df.columns,index=range(4))
    sum_df.iloc[0, 0] = "Maximum Value"
    sum_df.iloc[1, 0] = "Maximum Value at Time"
    sum_df.iloc[2, 0] = "Minimum Value"
    sum_df.iloc[3, 0] = "Minimum Value at Time"

    if df.empty:
        sum_df.iloc[0, 1] = "DATA FOR THIS FEEDER IS NOT AVAILABLE"

    else:
        for column in column_name_list:
            if df[column].dtype in [int, float]:
                max_value = df[column].max()
                if not np.isnan(max_value):
                    time_of_max = df.loc[df[column].idxmax(), 'Time']
                    sum_df.at[0, column] = max_value
                    sum_df.at[1, column] = time_of_max
                
                min_value = df[column].min()
                if not np.isnan(min_value):
                    time_of_min = df.loc[df[column].idxmin(), 'Time']
                    sum_df.at[2, column] = min_value
                    sum_df.at[3, column] = time_of_min

    # Writing both DataFrames to a CSV file
    filename = simname+"_"+feeder+'.csv'
    with open(filename, 'w', newline='') as f:
        # Write df1 to the file
        sum_df.to_csv(f, index=False)
        # Add a newline to separate the two DataFrames
        f.write('\n')
        # Write df2 to the file, ignoring the index
        df.to_csv(f, index=False)

    return sum_df,df

# output individual time summary to csv
def rms_summary_output(simname,df_list,sum_df_list,time_windows,feeder_list):
    print('')
    print('Write summary to csv file...')

    # summary of individual time windows
    for time in time_windows:
        print(f"Processing {time} Current / Power Summary...")
        # dataframe column name
        if time[0] == "P":
            time = time[1:]
            column = f'P_{time}_Ave (MW)'
            filename = simname+"_"+time+'_AVEpower_sum.csv'
        else:
            column = f'I_{time}_RMS (A)'
            filename = simname+"_"+time+'_RMScurrent_sum.csv'

        columns = ['OSLO ID'] + feeder_list
        max_df = pd.DataFrame(columns = columns)
        max_df['OSLO ID'] = ['Max Value','Time of Max']

        for index, sum_df in enumerate(sum_df_list):            
            max_df[feeder_list[index]] = sum_df.loc[0:1,column]
        

        rms_df = pd.DataFrame(columns = ['Time'])
        for index, df in enumerate(df_list):
            select = ['Time', column]

            df1 = rms_df
            df2 = df[select]
            # rename the df2 columns
            df2.columns = ['Time', feeder_list[index]]
            
            rms_df = pd.merge(df1, df2, left_on='Time', right_on='Time', how='outer') # merged two 
        
        # Writing both DataFrames to a CSV file
        with open(filename, 'w', newline='') as f:
            # Write df1 to the file
            # use this method to avoid insert warning
            blankc = pd.DataFrame({'Blank':[np.nan,np.nan]})
            max_df = pd.concat([blankc,max_df],axis = 1)
            max_df.to_csv(f, index=False)

            # Write df2 to the file, ignoring the index
            rms_df.insert(0, 'Blank', '')
            rms_df.to_csv(f, index=False, header=False)

    return

# output summary to be used by other programme
def final_sum(simname,sum_df_list,time_windows,feeder_list):
    print('')
    print('Write Final Summary to csv file...')

    # summary of individual time windows adding suffix _t
    time_t = [time + '_t' for time in time_windows]
    columns = ['OSLO ID'] + time_windows + time_t

    df = pd.DataFrame(columns = columns)

    # populate feeder list
    df['OSLO ID'] = feeder_list
    for index, feeder in enumerate(feeder_list):
        for time in time_windows:
            if time[0] == "P":
                column = f'P_{time[1:]}_Ave (MW)'
            else:
                column = f'I_{time}_RMS (A)'
            
            df.loc[index,time] = sum_df_list[index].loc[0,column]
            df.loc[index,time+'_t'] = sum_df_list[index].loc[1,column]


    filename = simname + '_RMSCurrent_Sum.csv'
    # Writing both DataFrames to a CSV file
    with open(filename, 'w', newline='') as f:
        writer = csv.writer(f) # create the csv writer
        row = [''] * 2 + ['Time Windows'] * len(time_windows) + ['Time at Max']
        writer.writerow(row)

        # Write df1 to the file
        df.insert(0, 'Blank', '')
        df.to_csv(f, index=False)

    return

##==========================subotpion 2
#process signle time window
def one_time_window_grid_calculation(simname, grid_df,assessed_time,time_increment):

    sum_df_1 = []
    sum_df_2 = []

    for time in assessed_time:

        print(f"Processing {time} min Grid Calculation...")

        if isinstance(time, int):
            file = simname + f'_{time}min_AVEpower_sum.csv'
        else:
            file = simname + f'_{time:.2f}min_AVEpower_sum.csv'
        
        sum_df = pd.read_csv(file,low_memory=False)
        # Check and drop the 'Blank' column if it exists
        if 'Blank' in sum_df.columns:
            sum_df = sum_df.drop(columns=['Blank'])
        # Check and drop the 'Unnamed: 0' column if it exists # This is added for backwards compatible.
        elif 'Unnamed: 0' in sum_df.columns:
            sum_df = sum_df.drop(columns=['Unnamed: 0'])

        # get  summary result from grid_power procesing
        result = grid_power_processing(simname, grid_df,time, time_increment)
        if result == False:
            return False
        else:
            feeder_list,sum_df_list = result

        # Ready data prepare:
        df_1, df_2 = grid_summary_output(grid_df,sum_df_list,feeder_list,sum_df,time)
        sum_df_1.append(df_1)
        sum_df_2.append(df_2)

    return sum_df_1, sum_df_2

# dataframe grid_calculation:
def grid_power_processing(simname, grid_df,time, time_increment):    # Group by Grid Feeder and list OSLO IDs for each feeder
    grouped = grid_df.groupby("Grid Feeder")["OSLO ID"].apply(list)
    feeder_list = []
    sum_df_list = []

    # Iterate over each group 
    for feeder, oslo_ids in grouped.items():
        print(f"Feeder: {feeder}")
        # Iterate over each OSLO ID within the group
        df_list = []
        # flag if at least one dataset exsit
        df_fail_flag = True
        for oslo in oslo_ids:
            print(f"- OSLO ID: {oslo}")                
            filename = simname + "_" + oslo + ".csv"
            if not os.path.isfile(filename):
                SharedMethods.print_message(f"WARNING: OSLO Step Output file {filename} does NOT exist. Skipping....","33")
                df = pd.DataFrame(columns = ['Time',oslo])
            else:
                df = pd.read_csv(filename, usecols=[2, 3], skiprows=range(1, 7))
                df.columns = ['Time',oslo]
                df_fail_flag = False                
            df_list.append(df)

        # check at least one data set is available
        if df_fail_flag:
            SharedMethods.print_message(f"ERROR: All OSLO data for {feeder} does NOT exist. Extract at least one to continue...", "31")
            return False

        # Merge all DataFrames based on 'Time'
        merged_df = df_list[0]  # Initialize with the first DataFrame
        for df in df_list[1:]:
            merged_df = pd.merge(merged_df, df, on='Time', how='outer')
        
        # Move 'Time' column to the first position
        columns = list(merged_df.columns)
        columns.remove('Time')
        columns = ['Time'] + columns
        merged_df = merged_df[columns]

        # Calculate sum of all values except 'Time' column
        merged_df['Total'] = pd.to_numeric(merged_df.iloc[:, 1:].sum(axis=1), errors='coerce')
        # Calculate rolling 300-row average
        time_window = time*60
        if isinstance(time, int):
            time_string = f"P{time}min"
        else:
            time_string = f"P{time:.2f}min"
        merged_df[time_string] = pd.to_numeric(merged_df['Total'].rolling(window=int(time_window/time_increment)).mean(), errors='coerce')

        # finding min max summary
        sum_df = pd.DataFrame(columns=merged_df.columns,index=range(2))
        sum_df.iloc[0, 0] = "Time of Max"
        sum_df.iloc[1, 0] = "Maximum"

        df = merged_df
        if df.empty:
            sum_df.iloc[0, 1] = "DATA FOR THIS GRID IS NOT AVAILABLE"

        else:
            for column in df.columns:
                if df[column].dtype in [int, float]:
                    max_value = df[column].max()
                    if not np.isnan(max_value):
                        time_of_max = df.loc[df[column].idxmax(), 'Time']
                        sum_df.at[0, column] = time_of_max
                        sum_df.at[1, column] = max_value
        
        # read to output:
        sum_df_list.append(sum_df)
        feeder_list.append(feeder)
                           
        # write to csv
        filename = simname+"_"+time_string[1:]+"_"+feeder+'.csv'
        with open(filename, 'w', newline='') as f:
            # Write df1 to the file
            sum_df.to_csv(f, index=False)
            # Write df2 to the file, ignoring the index
            df.to_csv(f, index=False,header = False)
            

    return feeder_list,sum_df_list

# calculate the final summary and make report for HV team
def grid_summary_output(grid_df,sum_df_list,feeder_list,sum_df,time):
# sum_df
#          OSLO ID                PRLY                COUL
# 0      Max Value  2.1092666666666666  1.5632416666666669
# 1    Time of Max            08:00:00            08:00:00
# 2       07:00:00                 NaN                 NaN
# 3       07:00:05                 NaN                 NaN
# 4       07:00:10                 NaN                 NaN

    if isinstance(time, int):
        time_string = f"P{time}min"
    else:
        time_string = f"P{time:.2f}min"
    
    # Create an empty dictionary to store the mapping
    substation_name = {}
    # Iterate over the rows of the DataFrame
    for index, row in grid_df.iterrows():
        oslo = row['OSLO ID']
        substation = row['Substation']
        if oslo not in substation_name:
            substation_name[oslo] = substation
    
    # summary oslo list
    sum_oslo_list = sum_df.columns.tolist()

    # create first summary dataframe:
    dataset1 = []
    for index, df in enumerate(sum_df_list):
        feeder = feeder_list[index]
        time = df.loc[0,time_string]
        power = df.loc[1,time_string]

        dataset1.append([feeder,time,power])
    
    columns = ['Grid Name','at Time (hh:mm:ss)','Maximum Power (MW)']

    sum_df_1 = pd.DataFrame(dataset1,columns = columns)

    # create second summary dataframe with abig more detailed
    dataset2 = []
    for index, df in enumerate(sum_df_list):
        columns = df.columns

        #remove columns 'Time','Total', 'P30min'
        df_oslo = df.drop(columns=['Time', 'Total', time_string])
        columns = df_oslo.columns
        for oslo in columns:
            feeder = feeder_list[index]
            time = df.loc[0,time_string]
            power = df.loc[1,time_string]
            substation = substation_name[oslo]
            oslo_id = oslo
            # find the maximum time row
            row_index = sum_df.loc[sum_df['OSLO ID'] == time].index[0]
            # assing the value
            if oslo in sum_oslo_list:
                oslo_power_at = pd.to_numeric(sum_df.loc[row_index,oslo], errors='coerce')
                oslo_time_abs  = sum_df.loc[1,oslo]
                oslo_power_abs = pd.to_numeric(sum_df.loc[0,oslo], errors='coerce')
                if power != 0:
                    oslo_contribute = oslo_power_at/power
                    # JY: if oslo_power_at is too small (considered as zero), the calculation will have runtime warning
                    oslo_diff = oslo_power_abs/oslo_power_at - 1
                else:
                    oslo_contribute = 1
                    oslo_diff = 1
            else:
                oslo_power_at = np.nan
                oslo_time_abs  = np.nan
                oslo_power_abs = np.nan
                if power != 0:
                    oslo_contribute = 0
                    oslo_diff = 0
                else:
                    oslo_contribute = np.nan
                    oslo_diff = np.nan
            
            dataset2.append([feeder,time,power,substation,oslo_id,oslo_power_at,oslo_contribute,oslo_time_abs,oslo_power_abs,oslo_diff])
    
    columns = ['Grid Name','Time at Grid Max (hh:mm:ss)','Max Power (MW)','Substation','OSLOID', \
               'Sub P at time of Grid Max (MW)','Sub Contribute %', 'Time at Sub Max (hh:mm:ss)', \
                'Max Sub Power (MW)','Power Diff %']
    
    sum_df_2 = pd.DataFrame(dataset2,columns = columns)
            
        
    return sum_df_1, sum_df_2

# print result to excel
def result_to_excel(simname,assessed_time,sum_df_1,sum_df_2):
    print('Writting Summary to Excel Ready Format....')

    filename = simname + '_Grid_Summary.xlsx'
    # Create a Pandas Excel writer using ExcelWriter
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:

        wb = writer.book
        # Write each DataFrame to a different worksheet 
        for index, time in enumerate(assessed_time):
            
            if isinstance(time, int):
                min_string = f"{time}"
            else:
                min_string = f"{time:.2f}"

            sheet_name = min_string + "min result"

            # Write sum_df_1 to cell B3
            sum_df_1[index].to_excel(writer, sheet_name=sheet_name, startrow=2, startcol=1, index=False)
            
            # Write sum_df_2 to cell G3
            sum_df_2[index].to_excel(writer, sheet_name=sheet_name, startrow=2, startcol=6, index=False)

            # formatting
            sheet = wb[sheet_name]

            sheet['B2'].value = f'Grid {min_string}min Average Power Summary'
            sheet['G2'].value = f'Detailed {min_string}min Maximum Average Power Summary'

            minrow = 2 + 1
            maxrow1 = len(sum_df_1[index]) + minrow
            maxrow2 = len(sum_df_2[index]) + minrow

            mincol1 = 2
            maxcol1 = 4

            mincol2 = 7
            maxcol2 = 16
            # Apply formatting: add borders and auto-adjust column width
            for row in sheet.iter_rows(min_row=minrow, min_col=mincol1, max_row=maxrow1, max_col=maxcol1):
                for cell in row:
                    cell.border = Border(left=Side(border_style='thin'), 
                                        right=Side(border_style='thin'), 
                                        top=Side(border_style='thin'), 
                                        bottom=Side(border_style='thin'))
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.number_format = '0.00'
            
            for row in sheet.iter_rows(min_row=minrow, min_col=mincol2, max_row=maxrow2, max_col=maxcol2):
                for cell in row:
                    cell.border = Border(left=Side(border_style='thin'), 
                                        right=Side(border_style='thin'), 
                                        top=Side(border_style='thin'), 
                                        bottom=Side(border_style='thin'))
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.number_format = '0.00'
            

            for row in sheet.iter_rows(min_row=minrow, min_col=13, max_row=maxrow2, max_col=13):
                for cell in row:
                    cell.number_format = '0.00%'
            
            for row in sheet.iter_rows(min_row=minrow, min_col=16, max_row=maxrow2, max_col=16):
                for cell in row:
                    cell.number_format = '0.00%'
            
            sheet.column_dimensions['B'].width = 20
            sheet.column_dimensions['G'].width = 20
            sheet.column_dimensions['J'].width = 15
    return

#===============Option 10 for DC assessment===========================================================
# option 4 main sub function for data extraction
def branch_data_process(simname, option_select,time_start, time_end, time_step, text_input, time_increment):

    # print("\nPlease select from the following options:(awaiting input)")
    # print("1: All branch step output extraction")
    # print("2: All branch rolling RMS current")
    # print("3: All branch Maximum rolling RMS current summary")
    # #print("4: All option 1-3")
    # print("4: Customised branch step output extraction (BranchNodeList.txt is required)")
    # print("5: Customised branch rolling RMS current (BranchNodeList.txt is required)")
    # print("6: Customised branch Maximum rolling RMS current summary (BranchNodeList.txt is required)")
    # print("Option 5-7 is under development and should NOT be selected!")

    option1 = option_select
    # option check and prepare
    if option1 not in ["0","1","2","3","4","5","6"]:
        SharedMethods.print_message("ERROR: Invalid Selection. Contact Support. Issue in batch_processing.py --> branch_data_process.","31")
        return False

    # if option1 in ["2","3","5","6"]:
    print(f"\nThe time window in seconds (0 - 86400): {time_step} seconds")
    second = int(time_step)
    if second <= 0 or second > 86400:
        SharedMethods.print_message("ERROR: Not valid time windows. Please reenter a valid assessment window","31")
        return False
    
    if option1 in ["4","5","6"]:
        branch_input = SharedMethods.file_read_import(text_input,simname)

        # check all item is unique
        if len(branch_input) != len(set(branch_input)):
            SharedMethods.print_message(f"ERROR: Duplication not allowed in this feature. Check your input...","31")
            return False

        for node in branch_input:
            if node[-2:] == '/S' or node[-2:] == '/E':
               pass
            else:
                SharedMethods.print_message(f"ERROR: BranchNodeList.txt has information {node} which is not the supported format.","31")
                SharedMethods.print_message("ERROR: Choose Option 1-3, or modify BranchNodeList.txt","31")
                return False

    filename = simname + ".osop.lst"
    if not SharedMethods.check_and_extract_lst_file(simname,time_start=time_start, time_end=time_end):
        return False

    sim_time, branch_list, step_branch_output, step_df = lst_file_read_branch(filename)

    if option1 in ["4","5","6"]:
        # filte the table based on user input
        selected_columns = ['Time']
        for item in branch_input:
            if item in step_df.columns:
                selected_columns.append(item)
            else:
                SharedMethods.print_message(f"WARNING: Input File Node {item} does NOT exist in the simulation. Ignored....","33")
        branch_df = step_df[selected_columns]
    else:
        branch_df = step_df

    if option1 in ["1","4"]:
        print('Saving branch step output summary...')
        branch_step_csv_out(simname,branch_df)
    
    if option1 in ["2","5"]:
        print("Processing RMS calculation...")
        rms_df,max_df = branch_list_RMS_process(simname,branch_df,second,time_increment)
        print('\nSaving branch RMS output summary...')
        branch_RMS_csv_out(simname,second,rms_df)
    
    if option1 in ["3","6"]:
        print("Processing RMS calculation...")
        rms_df,max_df = branch_list_RMS_process(simname,branch_df,second,time_increment)
        print('Saving branch RMS Summary output summary...')
        branch_RMS_sum_csv_out(simname,second,max_df)

    return True

# Read the list file and get all branches information in list:
def lst_file_read_branch(filename):

    sim_time = [] # Time list for reading
    branch_list = [] # initialise branch list
    step_branch_output = [] # step branch output information (2D list)
    step_branch_sum_c = [] # total current for each time step (2D list)

    # get all branches in the model
    section_flag = False # judge if N-Link section is reached

    print('\nInitialising...')
    with open(filename) as fp:
        total_line = sum(1 for line in enumerate(fp)) # get the total line number

    with open(filename) as fp:
        print('\nReading Branch Information...')
        for index, line in enumerate(fp):            
            if line[:7].strip() == '':
                continue
            if line[:7].strip() == 'NLINK':
                section_flag = True
                continue

            if line[:7].strip() == 'NFIXC':
                section_flag = False
                break
            
            if section_flag == True:
                branch_temp = line[10:14].strip()
                branch_list.append(branch_temp) # Save branches in a list file
                step_branch_output.append([branch_temp])
                # step_branch_sum_c.append([branch_temp])
    
    branch_read_flag = False
    time_count = 0 # time counter
    branch_count = 0 # branch counter
    # create columns for pandadataframe
    columns = ['Time']
    for item in branch_list:
        columns.append(item +'/S')
        columns.append(item +'/E')
    row_data = []
    row = []
    
    print('\nProcess branch information...')
    with open(filename) as fp:
        for index, line in enumerate(fp):
            # flag judgement
            if line[:7].strip() == 'INCSEC':
                if row:
                    row_data.append(row)

                time_count = time_count + 1
                temp_time = line[20:22].strip()+":"+line[23:25].strip()+":"+line[26:28].strip()
                sim_time.append(temp_time)
                branch_read_flag = True

                row = []
                row.append(temp_time)
                continue

            if line[:7].strip() == 'TRAIN':
                branch_count = 0
                branch_read_flag = False
            
            # add to consider if there is no train at specific time 
            if line[:8].strip() == 'NO OSLO':
                row = []
                branch_count = 0
                branch_read_flag = False
            
            # write down information
            if branch_read_flag == True:
                if line[76:82].strip() == '':
                    continue

                if line[:7].strip() == '':
                    branch_count = branch_count + 1
                    c_start_real = float(line[85:95].strip())
                    c_start_imag = float(line[95:105].strip())
                    c_end_real = float(line[110:120].strip())
                    c_end_imag = float(line[120:130].strip())

                    c_start_total = math.sqrt(math.pow(c_start_real,2)+math.pow(c_start_imag,2))
                    c_end_total = math.sqrt(math.pow(c_end_real,2)+math.pow(c_end_imag,2))

                    temp_line = [temp_time,c_start_real,c_start_imag,c_end_real,c_end_imag]
                    step_branch_output[branch_count-1].append(temp_line)

                    # temp_line = [c_start_total,c_end_total]
                    # step_branch_sum_c[branch_count-1].append(temp_line)
                    
                    row.append(c_start_total)
                    row.append(c_end_total)
                else:
                    continue
            
            SharedMethods.text_file_read_progress_bar(index, total_line)
    
    # create data frame straighaway:
    step_df = pd.DataFrame(row_data, columns = columns)

    return sim_time, branch_list, step_branch_output, step_df

# Write branch step output information to csv file
def branch_step_csv_out(simname,branch_df):

    # write summary info to CSV file
    with open(simname+'_branch_step_sum.csv','w',newline='') as f:
        writer = csv.writer(f) # create the csv writer
        row = ['OSLO Branch Step Output']
        writer.writerow(row)
        row = ['Branch Node List']
        writer.writerow(row)

        # Write df1 to the file
        # df.insert(0, 'Blank', '')
        branch_df.to_csv(f, index=False)
        
    print('\nBranch Step Output Summary Completed.')

    return True

# branch RMS output information to csv file
def branch_RMS_csv_out(simname,second,rms_df):

    # write summary info to CSV file
    with open(simname+'_branch_'+str(second)+'_rms_sum.csv','w',newline='') as f:
        writer = csv.writer(f) # create the csv writer
        row = [f'OSLO Branch RMS Output over {second} seconds']
        writer.writerow(row)
        row = ['Branch Node List']
        writer.writerow(row)

        # Write df1 to the file
        # df.insert(0, 'Blank', '')
        rms_df.to_csv(f, index=False)
        
    print('\nBranch RMS Output Summary Completed.')

    return

# branch RMS summary output information to csv file
def branch_RMS_sum_csv_out(simname,second,max_df):
    # write summary info to CSV file
    with open(simname+'_branch_'+str(second)+'_rms_sum_max.csv','w',newline='') as f:
        
        writer = csv.writer(f) # create the csv writer
        row = [f'OSLO Branch RMS Output over {second} seconds - Summary']
        writer.writerow(row)

        # Write df1 to the file
        # df.insert(0, 'Blank', '')
        max_df.to_csv(f, index=False)
        
    print('\nBranch RMS Output Summary Completed.')

    return

# process branch list RMS calculation
def branch_list_RMS_process(simname,branch_df,second,time_increment):

    time_window = int(second/time_increment)
    # Calculate rolling RMS for each column except the first one
    print("Calculating in process.. This will take a bit of time...")
    print("Do not close this window or do other things...")

    total = branch_df.shape[1] - 1 # total number of columns
    node_list = branch_df.columns.tolist()
    node_list = node_list[1:] #exclude the first one

    rms_df = pd.DataFrame()

    rms_df['Time'] = branch_df['Time']

    # Create a list to store calculated RMS values for each node
    rms_values_list = []

    for index, node in enumerate(node_list):
        print(f"\rProcessing node: {node}...({index+1}/{total})", end='', flush=True)
        rms_values  = calculate_rolling_rms(branch_df, node, second, time_increment)
        rms_values_list.append(rms_values)
    print(f"\rProgress all node successfully completed.", flush=True)
    
    # Concatenate all calculated RMS values into a single DataFrame
    rms_values_df = pd.concat(rms_values_list, axis=1)

    # Concatenate rms_df with rms_values_df to combine 'Time' column and calculated RMS values
    rms_df = pd.concat([rms_df, rms_values_df], axis=1)

    print('Calculate Maximum Value...')

    # finding max summary
    row_data = []
    for index, node in enumerate(node_list):
        if rms_df[node].dtype in [int, float]:
            max_value = rms_df[node].max()
            if not np.isnan(max_value):
                time_of_max_e = rms_df.loc[rms_df[node].idxmax(), 'Time']
                time_of_max_s = rms_df.loc[rms_df[node].idxmax() - time_window + 1, 'Time']
            else:
                time_of_max_e = 'NA'
                time_of_max_s = 'NA'
        row_data.append([node[:-2],node,time_of_max_s,time_of_max_e,max_value])
    
    columns = ['Branch','Node','Start Time','End Time','Maximum RMS Current (A)']
    max_df = pd.DataFrame(row_data,columns = columns)

    return rms_df,max_df

#============================================================================
# Check if the script is run as the main module
if __name__ == "__main__":
    # Add your debugging code here
    simname = "DCF321"  # Provide a simulation name or adjust as needed
    main_option = "9"  # Adjust as needed
    time_start = "0070000"  # Adjust as needed
    time_end = "0090000"  # Adjust as needed
    option_select = "2"  # Adjust as needed
    text_input = "FeederList"  # Adjust as needed
    low_v = None  # Adjust as needed
    high_v = None  # Adjust as needed
    time_step = 900  # Adjust as needed

    # Call your main function with the provided parameters
    main(simname, main_option, time_start, time_end, option_select, text_input, low_v, high_v, time_step)