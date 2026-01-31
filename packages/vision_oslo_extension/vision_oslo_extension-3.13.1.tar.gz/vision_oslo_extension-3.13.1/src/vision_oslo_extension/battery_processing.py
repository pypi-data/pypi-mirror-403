#
# -*- coding: utf-8  -*-
#=================================================================
# Created by: Jieming Ye
# Created on: Feb 2024
# Last Modified: Feb 2024
#=================================================================
# Copyright (c) 2024 [Jieming Ye]
#
# This Python source code is licensed under the 
# Open Source Non-Commercial License (OSNCL) v1.0
# See LICENSE for details.
#=================================================================
"""
Pre-requisite: 
SimulationName.oof: default oslo output file after successfully run a simulation.
xxxxx.xlsx: excel spreadsheet with proper user settings.
SimulationName.lst.txt: required for option 4
Used Input:
simname: to locate the result file or file rename
time_start: to define the analysis start time
time_end: to define the analysis end time
option_select: to define the option selected in subfunction
text_input: to locate the excel file name
other input for oslo_extraction.py only.
Expected Output:
Updated excel spreadsheet with analysis result.
Description:
This script defines the process of doing average load assessment.
First, it reads the excel file and save the user input in a data frame called start_df. Then it find the related .d4 files and read information one by one and saved in a data frame list called d4dataframe. The calculation result is saved in a similar format data frame list called sumdataframe. It then doing some analysis and save the final result in an updated data frame. Final step is to save all information in the excel in a proper format. The whole process is easy to follow via reading the code.
The key manual updating part is get_result_range() function which depends on the desired output format in excel, followed by table_formatting() function where some reading needs manual input. The other process should be auto adjustable as long as the excel input format is followed.

"""
#=================================================================
# VERSION CONTROL
# V1.0 (Jieming Ye) - Initial Version
#=================================================================
# Set Information Variable
# N/A
#=================================================================

# vision_oslo_extension/excel_processing.py
import math
import pandas as pd
import numpy as np
# import openpyxl
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, NamedStyle
from openpyxl.formatting.rule import CellIsRule,FormulaRule
from openpyxl.chart import ScatterChart,Reference,Series
from openpyxl.drawing.text import CharacterProperties
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.layout import Layout, ManualLayout
# import science for py
from scipy.interpolate import griddata
from scipy.spatial import Delaunay


# import vision_oslo
from vision_oslo_extension import oslo_extraction
from vision_oslo_extension.shared_contents import SharedMethods
from vision_oslo_extension import fs_bemu_capacity
from vision_oslo_extension import battery_processing_new


def main(simname, main_option, time_start, time_end, option_select, text_input, low_v, high_v, time_step):
    
    print("")
    print("Battery Train Assessment - - - > ")
    print("")
    
    simname = simname
    # Specify Excel file name
    excel_file = 'output.xlsx'
    excel_file = text_input + ".xlsx"

    if not SharedMethods.check_existing_file(excel_file):
        return False
    
    #option = "1" # Preliminary Assessment (simple thing)
    #option = "3" # Detailed spreadsheet, oof only
    #option = "4" # manual without extraction, exisiting d4

    option = option_select # 1:Fully Restart, require oof, and list

    time_increment = 5
    start = 10

    space  = 5

    if option not in ["0","1","2","3","4","5"]:
        SharedMethods.print_message("ERROR: Error in battery_processing.py. Please contact Support.","31")
        return False
    
    if option == "0":
        SharedMethods.print_message("ERROR: Please select an option to proceed.","31")
        return False
    
    if option in ["1","2"]:
        if not fs_bemu_capacity.main(simname, main_option, time_start, time_end, option_select, text_input, low_v, high_v, time_step):
            SharedMethods.print_message("ERROR: Error in fs_bemu_capacity.py. Please contact Support.","31")
            return False
        else:
            return True
    
    if option == "5":
        if not battery_processing_new.main(simname, main_option, time_start, time_end, option_select, text_input, low_v, high_v, time_step):
            SharedMethods.print_message("ERROR: Error in battery_processing_new.py. Please contact Support.","31")
            return False
        else:
            return True
    
    if option == "3":
        if not SharedMethods.check_oofresult_file(simname):
            return False

    # start_cell = 'B11'
    # read data from start tab
    result = start_reading_process(
        simname, time_start, time_end, text_input, low_v, high_v, time_step,excel_file,option
    )
    if result == False:
        return False
    else:
        route_list, train_list, traction_list, start_df = result

    # check bhtpbank existance:
    for bhtp in traction_list:
        if not SharedMethods.check_bhtpbank_from_root(bhtp):
            return False
        
    # check essential d4 files
    if option == "4":
        print("Checking essential ds1 files ...")
        for train in train_list:
            filename = get_train_ds1_filename(simname,train)
            if not SharedMethods.check_existing_file(filename):
                SharedMethods.print_message(f"ERROR: ds1 file for train {train} do not exist. Select option 3 to proceed.","31")
                return False
                            
    # extraction train step output list
    if option == "3":
        if not one_stop_train_extraction(simname, time_start, time_end,train_list):
            return False
    
    # check the file contents regardless
    for train in train_list:
        filename = get_train_ds1_filename(simname,train)
        if not SharedMethods.validate_extracted_result(filename, force_data = True):
            SharedMethods.print_message(f"ERROR: ds1 file for train {train} is empty. Not allowed in this function. Check your settings...","31")
            return False
        
    # number of dataframes to create
    total = len(train_list)

    # read bhtpbank files
    result = bhtpbank_analysis_process(traction_list)
    if result == False:
        return False
    else:
        bhtpbank_df_effi, bhtpbank_df_beffi, bhtpbank_df_be, aux_load = result

    # process ds1 file
    battery_level_change, ds1dataframe = train_reading_process(simname, train_list,time_increment,start_df,result)

    # update result
    result = result_table_prepare(start_df, battery_level_change)
    if result == False:
        return False
    else:
        result_df, route_list = result
    
    trainnumber = len(train_list)

    # prepare route list for plotting (plotting purely based on train jounery settings)
    train_list_for_plot = prepare_plotting_list(start_df)

    data_write_save(simname, excel_file,start,total,space,start_df,result_df,train_list,route_list,train_list_for_plot,ds1dataframe,trainnumber)

    return True

# get plotting list from the setting
def prepare_plotting_list(start_df):

    route_name_list = []
    from_list = []
    to_list = []
    head_list = []
    
    # convert analysis info into three formated list
    for _, row in start_df.iterrows():
        route_name = row['Train Journey']
        train_id = row['Train ID']
        previous_id = row['Link Previous']
        
        route_name_list.append(route_name)
        if str(previous_id).isdigit():
            from_list.append(previous_id)
        else:
            from_list.append(None)
            head_list.append([route_name,train_id])        
        to_list.append(train_id)
    
    # step1: arrange the linked train blindly
    linked_list = []
    linked_name_list = []

    for header in head_list:
        search = header[1] # train id
        jounery = []
        jounery_name = []
        
        route_finish = False # flag for whether this jounery is finished

        if to_list != []:
            while True:         
                for index, item in enumerate(to_list): # iterate all jounery end
                    update_flag = False # search updated or not

                    if item == search:
                        jounery.append(to_list[index])
                        jounery_name.append(route_name_list[index])
                        for index1, next in enumerate(from_list): # find next target
                            if next == search:
                                search = to_list[index1]
                                break
                            if index1 == len(from_list)-1: # until there is no link
                                search = 0
                        update_flag = True
                        break
                    
                    if index == len(to_list)-1 and update_flag == False: # if searched all link and did find the next one
                        route_finish = True
                
                if route_finish == True:
                    break
        
        linked_list.append(jounery)
        linked_name_list.append(jounery_name)
    
    # step2: further split the linked list based on the name list
    final_list = []
    count = -1
    for index, route in enumerate(linked_list):
        final_list.append([])
        count = count + 1
        target = linked_name_list[index][0]
        for index1, train in enumerate(route):
            if linked_name_list[index][index1] == target:
                final_list[count].append(train)
            else:
                final_list.append([])
                count = count + 1
                target = linked_name_list[index][index1]
                final_list[count].append(train)

    return final_list


# get ds1 file name from train no
def get_train_ds1_filename(simname, train):

    if 1 <= len(str(train)) <= 3:
        train_no = format(int(train),'03d')
    elif 4 <= len(str(train)) <= 5:
        train_no = format(int(train),'05d')
    filename = simname + "_train_" + train_no + ".osop.ds1"

    return filename

# extraction
def one_stop_train_extraction(simname, time_start, time_end,train_list):

    # User defined time windows extraction
    time_start = SharedMethods.time_input_process(time_start,1)
    time_end = SharedMethods.time_input_process(time_end,1)

    if time_start == False or time_end == False:
        return False

    # process 1: feeder step output
    print("Extrating train step output from Start Tab ... ")

    # processing the list
    for items in train_list:
        #print(items)
        if not oslo_extraction.train_step_one(simname,items,time_start,time_end):
            SharedMethods.print_message(f"WARNING: Error for {items} will be ignored and process continued...","33")

    return True

# write data to excel
def data_write_save(simname, excel_file,start,total,space,start_df,result_df,train_list,route_list,train_list_for_plot,ds1dataframe,trainnumber):
    # write data to excel
    with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a',if_sheet_exists='overlay') as writer:

        # get currrent workbook to do process
        wb = writer.book

        print("Generate Result Page...")
        result_df.to_excel(writer, sheet_name="Result", index=False, startrow = start, startcol = 1)
        
        # Write each DataFrame to a different sheet in the Excel file
        print("Writing individual Trains...")
        for index, train in enumerate(train_list):
            print(f"Saving {train}...")

            dflist = ds1dataframe[train]
            # emty columns
            dflist.insert(dflist.columns.get_loc('REF2')+1,'New_C_1', np.nan)
            dflist.insert(dflist.columns.get_loc('Effort for Traction (kN)')+1,'New_C_2', np.nan)
            dflist.insert(dflist.columns.get_loc('Battery Egy (dis)charging (kWh)')+1,'Excel_Time', np.nan)
            dflist.insert(dflist.columns.get_loc('Excel_Time')+1,'Excel_Position', np.nan)
            dflist.insert(dflist.columns.get_loc('Excel_Position')+1,'Excel_Battery', np.nan)
            dflist.insert(dflist.columns.get_loc('Excel_Battery')+1,'Excel_SoC', np.nan)


            dflist.to_excel(writer, sheet_name=str(train), index=False, startrow = 2)
        
        data_row = 4 # first data row number

        print("Add Information Supporting Plotting....")
        addtional_info(wb,route_list,ds1dataframe,result_df,data_row)

        # prepare sheet for plotting
        print("Plot the Battery Information....")
        empty_df = pd.DataFrame()
        for index, route in enumerate(train_list_for_plot):
            route_name = result_df.loc[result_df['Train ID'] == route[0], 'Train Journey'].values[0]
            empty_df.to_excel(writer, sheet_name=route_name, index=False)

        jounery_plot(wb,train_list_for_plot,ds1dataframe,result_df,data_row)
        jounery_plot_add(wb,train_list_for_plot,ds1dataframe,result_df,data_row)
        
        # # Calculate the Excel range for each DataFrame
        range_list = get_result_range(start,total,space,trainnumber) # range list as [power data range, current data range]
        
        # # table formatting
        table_formatting(simname,wb,range_list)

        print("Saving Data...")

    return

def addtional_info(wb,route_list,ds1dataframe,result_df,data_row):
    
    for route in route_list:
        pos = 0
        
        for train in route:

            sheet = wb[str(train)]
            dflist = ds1dataframe[train]

            # add formula in
            sheet["W2"].value = "Convert mph to kmh"
            sheet["X2"].value = "Convert mph to m/s"
            sheet["Y1"].value = "Interpolation, convert lbf to kN"
            sheet["Y2"].value = "If TE < 0 and 100%, do interpolation"
            sheet["AA2"].value = "TE * Ave Speed * 5 / 3600"
            sheet["AB2"].value = "Active Power * 5 / 3600"
            sheet["AC1"].value = "complex linear interpolation"
            sheet["AC2"].value = "mechanical power and efficiency"
            sheet["AD2"].value = "Auxiliary load * 5 / 3600"
            sheet["AE1"].value = "electrical - traction - aux On wire"
            sheet["AE2"].value = "0 - electrical off wire"
            sheet["AF1"].value = "Battery energy * charging effi On wire"
            sheet["AF2"].value = "Battery energy / discharging effi Off wire"
            sheet["AG2"].value = "Excel Time formula"
            sheet["AH2"].value = "Absolute Distance Gone for route"
            sheet["AI2"].value = "Absoulate Battery level for route"
            sheet["AJ2"].value = "SoC for route"

            # create excel time format
            index = data_row
            data_range = f"AG{data_row}:AG{data_row+len(dflist)-1}"
            for line in sheet[data_range]:
                for cell in line:
                    cell.value = f'=TIME(LEFT(C{index},2),MID(C{index},4,2),RIGHT(C{index},2))'
                    cell.number_format = 'hh:mm:ss'
                index = index + 1
            
            # create location info
            index = data_row
            data_range = f"AH{data_row}:AH{data_row+len(dflist)-1}"
            for line in sheet[data_range]:
                for cell in line:
                    #cell.value = f'=0.9144*F{index}+{pos}' # convert from yards to meters
                    cell.value = 0.9144 * sheet[f'F{index}'].value + pos
                    cell.number_format = '0.0'
                index = index + 1
            # update position information
            pos = pos + 0.9144 * dflist["Distance Gone"].iloc[-1] # get the last value

            # create battery info
            bat = result_df.loc[result_df['Train ID'] == train, 'Start Battery Level (kWh)']
            bat = bat.values[0]
            size = result_df.loc[result_df['Train ID'] == train, 'Battery Size (kWh)']
            size = size.values[0]

            sheet[f'AI{data_row}'].value = bat
            data_range = f"AI{data_row+1}:AI{data_row+len(dflist)-1}"

            index = data_row+1
            for line in sheet[data_range]:
                for cell in line:
                    #cell.value = f'=AF{index-1}+AI{index-1}' # battery info
                    cell.value = sheet[f'AF{index-1}'].value + sheet[f'AI{index-1}'].value
                index = index + 1
            
            # SoC Calculate
            index = data_row
            data_range = f"AJ{data_row}:AJ{data_row+len(dflist)-1}"
            for line in sheet[data_range]:
                for cell in line:
                    #cell.value = f'=AI{index}/{size}' # battery SoC
                    cell.value = sheet[f'AI{index}'].value /size  # battery SoC
                    cell.number_format = '0.00%'
                index = index + 1

    return

def jounery_plot(wb,route_list,ds1dataframe,result_df,data_row):
    # set default variable
    for index, route in enumerate(route_list):
        route_name = result_df.loc[result_df['Train ID'] == route[0], 'Train Journey'].values[0]
        # plot two plot, battery level vs time / location
        sheet = wb[route_name]

        chart = ScatterChart('smoothMarker')
        chart.title = route_name + " Battery SoC Profile"
        chart.x_axis.title = "Time (hh:mm:ss)"
        chart.y_axis.title = "State of Charge SoC (%)"

        # Set chart size
        chart.height = 10  # Adjust width as needed
        chart.width = 30  # Adjust height as needed
        xlimmin = 1
        xlimmax = 0

        for train in route:
            xvalues = Reference(wb[str(train)], min_col=33, min_row=data_row, max_row=data_row+ds1dataframe[train].shape[0]-1)
            yvalues = Reference(wb[str(train)], min_col=36, min_row=data_row, max_row=data_row+ds1dataframe[train].shape[0]-1)
            head = result_df.loc[result_df['Train ID'] == train, 'Section'].values[0]
            
            series = Series(yvalues,xvalues,title_from_data=False,title = head)
            chart.series.append(series)

            t_start = ds1dataframe[train]['Time'].iloc[0]
            t_start = t_start.split(":")
            ts = (float(t_start[0])+float(t_start[1])/60+float(t_start[1])/3600)/24

            t_end = ds1dataframe[train]['Time'].iloc[-1]
            t_end = t_end.split(":")
            te = (float(t_end[0])+float(t_end[1])/60+float(t_end[1])/3600)/24

            if ts < xlimmin:
                xlimmin = ts
            if te > xlimmax:
                xlimmax = te
                
        # find nearest 10 min timeline.
        for i in range(0,144):
            if i*1/144 > xlimmin:
                xlimmin = (i-1)*1/144
                for j in range(i,144):
                    if j*1/144 >= xlimmax:
                        xlimmax = j*1/144
                        break
                break
        
        tspace = (xlimmax-xlimmin)/8
        
        chart.legend.position = 'b' # top legend position
        chart.y_axis.scaling.min = 0      
        chart.y_axis.scaling.max = 1

        chart.x_axis.scaling.min = xlimmin
        chart.x_axis.scaling.max = xlimmax

        chart.x_axis.majorUnit = tspace # 
        chart.y_axis.majorUnit = 0.1 # 10 percent
        
        cp = CharacterProperties(sz=1400)
        chart.title.tx.rich.p[0].r[0].rPr = cp 

        cl = GraphicalProperties()
        cl.line.solidFill = "E0E0E0"
        chart.x_axis.majorGridlines.spPr = cl  # Light grey color
        chart.y_axis.majorGridlines.spPr = cl  #  # Light grey color

        # # # Adjust the layout to ensure there's enough space around the plot area
        # chart.layout = Layout(
        #     ManualLayout(
        #         x=0.15, y=0.25,  # Position from the left, top
        #         h=0.5, w=0.7,  # height, Width
        #         xMode="edge",
        #         yMode="edge",
        #     )
        # )
        sheet.add_chart(chart,"B2")


        chart = ScatterChart('smoothMarker')
        chart.title = route_name + " Battery SoC Profile"
        chart.x_axis.title = "Distance Gone (meter)"
        chart.y_axis.title = "State of Charge SoC (%)"

        # Set chart size
        chart.height = 10  # Adjust width as needed
        chart.width = 30  # Adjust height as needed
        xlimmin = 1000000 # 1000 km
        xlimmax = 0

        for train in route:
            xvalues = Reference(wb[str(train)], min_col=34, min_row=data_row, max_row=data_row+ds1dataframe[train].shape[0]-1)
            yvalues = Reference(wb[str(train)], min_col=36, min_row=data_row, max_row=data_row+ds1dataframe[train].shape[0]-1)
            head = result_df.loc[result_df['Train ID'] == train, 'Section'].values[0]
            
            series = Series(yvalues,xvalues,title_from_data=False,title = head)
            chart.series.append(series)

            sheet = wb[str(train)]
            t_start = sheet[f'AH{data_row}'].value
            t_end = sheet[f'AH{data_row+ds1dataframe[train].shape[0]-1}'].value

            if t_start < xlimmin:
                xlimmin = t_start
            if t_end > xlimmax:
                xlimmax = t_end
                
        sheet = wb[route_name]
        tspace = (xlimmax-xlimmin)/10
        
        chart.legend.position = 'b' # top legend position
        chart.y_axis.scaling.min = 0      
        chart.y_axis.scaling.max = 1

        chart.x_axis.scaling.min = xlimmin
        chart.x_axis.scaling.max = xlimmax

        chart.x_axis.majorUnit = tspace # 
        chart.y_axis.majorUnit = 0.1 # 10 percent
        
        cp = CharacterProperties(sz=1400)
        chart.title.tx.rich.p[0].r[0].rPr = cp 

        cl = GraphicalProperties()
        cl.line.solidFill = "E0E0E0"
        chart.x_axis.majorGridlines.spPr = cl  # Light grey color
        chart.y_axis.majorGridlines.spPr = cl  #  # Light grey color

        # # # Adjust the layout to ensure there's enough space around the plot area
        # chart.layout = Layout(
        #     ManualLayout(
        #         x=0.15, y=0.25,  # Position from the left, top
        #         h=0.5, w=0.7,  # height, Width
        #         xMode="edge",
        #         yMode="edge",
        #     )
        # )

        sheet.add_chart(chart,"B25")

    return

# As of DC work request, 2nd plot is needed with regard to actual level vs time
def jounery_plot_add(wb,route_list,ds1dataframe,result_df,data_row):
    # set default variable
    for index, route in enumerate(route_list):
        route_name = result_df.loc[result_df['Train ID'] == route[0], 'Train Journey'].values[0]
        # plot two plot, battery level vs time / location
        sheet = wb[route_name]

        chart = ScatterChart('smoothMarker')
        chart.title = route_name + " Battery Level Profile"
        chart.x_axis.title = "Time (hh:mm:ss)"
        chart.y_axis.title = "Battery Level (kWh)"

        # Set chart size
        chart.height = 10  # Adjust width as needed
        chart.width = 30  # Adjust height as needed
        xlimmin = 1
        xlimmax = 0

        for train in route:
            xvalues = Reference(wb[str(train)], min_col=33, min_row=data_row, max_row=data_row+ds1dataframe[train].shape[0]-1)
            yvalues = Reference(wb[str(train)], min_col=35, min_row=data_row, max_row=data_row+ds1dataframe[train].shape[0]-1)
            head = result_df.loc[result_df['Train ID'] == train, 'Section'].values[0]
            
            series = Series(yvalues,xvalues,title_from_data=False,title = head)
            chart.series.append(series)

            t_start = ds1dataframe[train]['Time'].iloc[0]
            t_start = t_start.split(":")
            ts = (float(t_start[0])+float(t_start[1])/60+float(t_start[1])/3600)/24

            t_end = ds1dataframe[train]['Time'].iloc[-1]
            t_end = t_end.split(":")
            te = (float(t_end[0])+float(t_end[1])/60+float(t_end[1])/3600)/24

            if ts < xlimmin:
                xlimmin = ts
            if te > xlimmax:
                xlimmax = te
                
        # find nearest 10 min timeline.
        for i in range(0,144):
            if i*1/144 > xlimmin:
                xlimmin = (i-1)*1/144
                for j in range(i,144):
                    if j*1/144 >= xlimmax:
                        xlimmax = j*1/144
                        break
                break
        
        tspace = (xlimmax-xlimmin)/8
        
        chart.legend.position = 'b' # top legend position
        # chart.y_axis.scaling.min = 0      
        # chart.y_axis.scaling.max = 1

        chart.x_axis.scaling.min = xlimmin
        chart.x_axis.scaling.max = xlimmax

        chart.x_axis.majorUnit = tspace # 
        # chart.y_axis.majorUnit = 0.1 # 10 percent
        
        cp = CharacterProperties(sz=1400)
        chart.title.tx.rich.p[0].r[0].rPr = cp 

        cl = GraphicalProperties()
        cl.line.solidFill = "E0E0E0"
        chart.x_axis.majorGridlines.spPr = cl  # Light grey color
        chart.y_axis.majorGridlines.spPr = cl  #  # Light grey color

        sheet.add_chart(chart,"T2")


        chart = ScatterChart('smoothMarker')
        chart.title = route_name + " Battery Level Profile"
        chart.x_axis.title = "Distance Gone (meter)"
        chart.y_axis.title = "Battery Level (kWh)"

        # Set chart size
        chart.height = 10  # Adjust width as needed
        chart.width = 30  # Adjust height as needed
        xlimmin = 1000000 # 1000 km
        xlimmax = 0

        for train in route:
            xvalues = Reference(wb[str(train)], min_col=34, min_row=data_row, max_row=data_row+ds1dataframe[train].shape[0]-1)
            yvalues = Reference(wb[str(train)], min_col=35, min_row=data_row, max_row=data_row+ds1dataframe[train].shape[0]-1)
            head = result_df.loc[result_df['Train ID'] == train, 'Section'].values[0]
            
            series = Series(yvalues,xvalues,title_from_data=False,title = head)
            chart.series.append(series)

            sheet = wb[str(train)]
            t_start = sheet[f'AH{data_row}'].value
            t_end = sheet[f'AH{data_row+ds1dataframe[train].shape[0]-1}'].value

            if t_start < xlimmin:
                xlimmin = t_start
            if t_end > xlimmax:
                xlimmax = t_end
                
        sheet = wb[route_name]
        tspace = (xlimmax-xlimmin)/10
        
        chart.legend.position = 'b' # top legend position
        # chart.y_axis.scaling.min = 0      
        # chart.y_axis.scaling.max = 1

        chart.x_axis.scaling.min = xlimmin
        chart.x_axis.scaling.max = xlimmax

        chart.x_axis.majorUnit = tspace # 
        # chart.y_axis.majorUnit = 0.1 # 10 percent
        
        cp = CharacterProperties(sz=1400)
        chart.title.tx.rich.p[0].r[0].rPr = cp 

        cl = GraphicalProperties()
        cl.line.solidFill = "E0E0E0"
        chart.x_axis.majorGridlines.spPr = cl  # Light grey color
        chart.y_axis.majorGridlines.spPr = cl  #  # Light grey color

        sheet.add_chart(chart,"T25")

    return

# read the start tab and collect informaiton
def start_reading_process(simname, time_start, time_end, text_input, low_v, high_v, time_step, excel_file,option):
    route_list = []
    train_list = []
    traction_list = []
    
    # File check
    print("Check Excel and Deleting Existing Contents ...")

    try:
        wb = load_workbook(excel_file)
        # Delete all sheets except 'Start'
        for sheet_name in wb.sheetnames:
            if sheet_name == 'Start':
                table_list = check_table_list(wb[sheet_name])
                
                if table_list is not False:
                    for line in table_list:
                        link = line[5]

                        if str(link).isdigit():  # This works for positive integers represented as strings
                            link = int(link)
                        else:
                            route_list.append(line[0])  # Append to route_list if conversion fail

                        bhtpbank = line[4]
                        if bhtpbank not in traction_list:
                            traction_list.append(bhtpbank) # unique bhtpbank list

                        train_list.append(line[1])

                else:
                    return False
            else:
                wb.remove(wb[sheet_name])
        
        # Save changes
        wb.save(excel_file)

    except Exception as e:
        SharedMethods.print_message(f"(Close the Excel file and Start Again) Error: {e}","31")
        return False
        
    # read start up information 
    columns = ["Train Journey","Train ID","Section","On/Off Wire","Traction Profile","Link Previous","Journey Start Battery Level (%)", \
               "Auxiliary Power (kW)","Charging Efficiency (%)","Discharging Efficiency (%)","Battery Size (kWh)"]
    start_df = pd.DataFrame(table_list,columns=columns)

    # check duplication in OSLO id
    if not SharedMethods.find_duplicates(train_list): return False

    # validate route
    if route_list == []:
        SharedMethods.print_message(f"ERROR: No routes detected. Adjust your settings. Ensure 'NA' set up for route start...","31")
        return False
    if not SharedMethods.find_duplicates(route_list): return False
    
    return route_list, train_list, traction_list, start_df

# check table list on start page
def check_table_list(sheet):
    print("Reading Configuration Setting ...")
    table_row = 11
    table_start_column = 1
    table_end_column = 11
    # create node and feeder list
    table_list = []

    # check feeder
    index = table_row
    column = table_start_column
    if sheet.cell(row=index, column=column+1).value is not None:
        while True:
            row_data = []
            for temp in range(table_start_column,table_end_column + 1):
                row_data.append(sheet.cell(row=index, column=temp).value)
            table_list.append(row_data)
            index += 1
            
            check = sheet.cell(row=index, column=column+1).value
            if check is None:
                break
    else:
        SharedMethods.print_message("ERROR: Wrong data format. No information at B11.","31")
        return False

    return table_list

# check the bhtpbank file and get required data in.
def bhtpbank_analysis_process(traction_list):
    bhtpbank_df_effi = {} # summary of bhtpbank data frame dictionary type {name: dataframe}
    bhtpbank_df_beffi = {} # braking efficiency
    bhtpbank_df_be ={} # braking effort
    aux_load = {} # auxiliary load for each rolling stock

    for bhtp in traction_list:
        print(f"Reading {bhtp} from VISION library. {bhtp} ...")
        bhtpname = SharedMethods.check_bhtpbank_from_root(bhtp)

        # Create Essential Data Set
        # Level 0: voltage entry [V1,V2,V3,...]
        # Level 1: data entry e.g. [d1,d2,d3...] under V1
        # Level 2: percetage entry e.g [0, 100] under d1
        vol_list = [] # data list saving voltage list

        #date format: list = [[d1,d2,d3..],[d1,d2,d3...],[...],...]
        speed_list = [] # nested speed for voltage
        TE_list = [] # nested tractive effort for voltage
        Re_I_list = [] # nested real current for voltage
        Im_I_list = [] # nested imag current for voltage
        
        TE_list_p1 = [] # nested list for first percentage
        Re_I_list_p1 = [] # nested real current for first percentage
        Im_I_list_p1 = [] # nested imag current for first percentage
        TE_list_p2 = [] # nested list for 2nd percentage
        Re_I_list_p2 = [] # nested real current for 2nd percentage
        Im_I_list_p2 = [] # nested imag current for 2nd percentage

        # define for braking part
        vol_list_b = [] # data list saving voltage list

        #date format: list = [[d1,d2,d3..],[d1,d2,d3...],[...],...]
        speed_list_b = [] # nested speed for voltage
        TE_list_b = [] # nested tractive effort for voltage
        Re_I_list_b = [] # nested real current for voltage
        Im_I_list_b = [] # nested imag current for voltage
        
        TE_list_p1_b = [] # nested list for first percentage
        Re_I_list_p1_b = [] # nested real current for first percentage
        Im_I_list_p1_b = [] # nested imag current for first percentage
        TE_list_p2_b = [] # nested list for 2nd percentage
        Re_I_list_p2_b = [] # nested real current for 2nd percentage
        Im_I_list_p2_b = [] # nested imag current for 2nd percentage
    
        vo_index = -1 # votlage index (level 0 index)

        vol = False # Flag if reading voltage line
        spe = False # Flag if reading speed line
        pec = 0 # pecentage line indicator

        brake_flag = False # Flag if reading/checking braking effort part
        BE_flag = False # confirm checking TE or BE part
    
        # Open the bhtpbank and read data
        with open(bhtpname) as fp:
            for index, line in enumerate(fp):
                # decide which section the code is looking            
                if line[:2].strip() == 'DB':
                    bhtpbank = line[2:10].strip()
                    bhtpcomment = line[10:].strip()
                if line[:3].strip() == '  M':
                    continue
                if line[:3].strip() == 'REG':
                    regen = True
                if line[:5].strip() == 'TRAIN':
                    aux_load[bhtp] = int(line[25:28].strip())
                if line[:1].strip() == '*':
                    continue
                if line[:3].strip() == 'VOL':
                    vo_index = vo_index + 1 # go to the next voltage entry    
                    pec = 0
                    vol = True
                if line[:3].strip() == 'SPE':
                    pec = 0
                    spe = True
                    vol = False
                if line[:3].strip() == '':
                    spe = False
                    pec = pec + 1
                if line[:15].strip() == 'END OF TRACTIVE':
                    brake_flag = True
                    vo_index = -1
                    continue
                if line[:11].strip() == 'END OF DATA':
                    break

                # excute data list
                if brake_flag == False:
                    list_data_action(line,vol_list,speed_list,TE_list,Re_I_list,Im_I_list, \
                        TE_list_p1,Re_I_list_p1,Im_I_list_p1,TE_list_p2,Re_I_list_p2,Im_I_list_p2, \
                            vo_index,vol,spe,pec)
                else:
                    list_data_action(line,vol_list_b,speed_list_b,TE_list_b,Re_I_list_b,Im_I_list_b, \
                        TE_list_p1_b,Re_I_list_p1_b,Im_I_list_p1_b,TE_list_p2_b,Re_I_list_p2_b,Im_I_list_p2_b, \
                            vo_index,vol,spe,pec)
                
        aux_list = aux_power_calculate(Re_I_list,Re_I_list_p1,vol_list)
        total_I_list = total_current(Re_I_list,Im_I_list)

        # use real current here for efficiency calculaion Re_I_list
        effi_list, mech_list, ele_list, ele_traction_list = effi_calculation(TE_list,Re_I_list,Re_I_list_p1,vol_list,speed_list,BE_flag)
        effi_check(effi_list,vol_list,BE_flag)

        # check if braking data is available
        if len(TE_list_b) == 0:
            brake_flag = False

        if brake_flag == True:
            BE_flag = True

            aux_list_b = aux_power_calculate(Re_I_list_b,Re_I_list_p1_b,vol_list_b)
            effi_list_b, mech_list_b, ele_list_b, ele_traction_list_b = effi_calculation(TE_list_b,Re_I_list_b,Re_I_list_p1_b,vol_list_b,speed_list_b,BE_flag)
            effi_check(effi_list_b,vol_list_b,BE_flag)

        # Traction Efficiency Datagrid
        x = []
        y = []
        z = []
        for index, volt in enumerate(vol_list):
            for index2, speed in enumerate(speed_list[index]):
                x.append(volt/1000)
                y.append(speed)
                z.append(effi_list[index][index2])

        data = {"x":x, "y": y, "z": z}
        df = pd.DataFrame(data)
        bhtpbank_df_effi[bhtp] = df

        # braking efficiency datagrid
        x = []
        y = []
        z = []
        for index, volt in enumerate(vol_list_b):
            for index2, speed in enumerate(speed_list_b[index]):
                x.append(volt/1000)
                y.append(speed)
                z.append(effi_list_b[index][index2])

        data = {"x":x, "y": y, "z": z}
        df = pd.DataFrame(data)
        bhtpbank_df_beffi[bhtp] = df

        # braking effort datagrid
        x = []
        y = []
        z = []
        for index, volt in enumerate(vol_list_b):
            for index2, speed in enumerate(speed_list_b[index]):
                x.append(volt/1000)
                y.append(speed)
                z.append(TE_list_b[index][index2])

        data = {"x":x, "y": y, "z": z}
        df = pd.DataFrame(data)
        bhtpbank_df_be[bhtp] = df

        if bhtp not in aux_load:
            SharedMethods.print_message(f"ERROR: BHTPBANK {bhtp} does not contain Auxiliary Load information. Update bhtpbank required...",31)
            return False

    return bhtpbank_df_effi, bhtpbank_df_beffi, bhtpbank_df_be, aux_load

# train information sort and output
# #speed_list,TE_list,Re_I_list,Im_I_list,vo_index,da_index,p_index,vol,spe    
def list_data_action(line,vol_list,speed_list,TE_list,Re_I_list,Im_I_list, \
    TE_list_p1,Re_I_list_p1,Im_I_list_p1,TE_list_p2,Re_I_list_p2,Im_I_list_p2, \
        vo_index,vol,spe,pec):
    if vol == True:
        data = int(line[8:].strip()) # get the value of voltage
        vol_list.append(data)
        speed_list.append([])
        TE_list.append([])
        Re_I_list.append([])
        Im_I_list.append([])
        TE_list_p1.append([])
        Re_I_list_p1.append([])
        Im_I_list_p1.append([])
        TE_list_p2.append([])
        Re_I_list_p2.append([])
        Im_I_list_p2.append([])
    
    if spe == True:
        data = float(line[5:15].strip()) # get the value of speed
        speed_list[vo_index].append(data)
        data = float(line[20:30].strip()) # get the value of TE
        TE_list[vo_index].append(data)
        data = float(line[40:50].strip()) # get the value of Real Current
        Re_I_list[vo_index].append(data)
        data = float(line[50:].strip()) # get the value of Imag Current
        Im_I_list[vo_index].append(data)

    if pec == 1:
        data = float(line[:10].strip()) # get the value of TE percentage
        TE_list_p1[vo_index].append(data)
        data = float(line[10:20].strip()) # get the value of Real perc
        Re_I_list_p1[vo_index].append(data)
        data = float(line[20:].strip()) # get the value of Imag perc
        Im_I_list_p1[vo_index].append(data)

    if pec == 2:
        data = float(line[:10].strip()) # get the value of TE percentage
        TE_list_p2[vo_index].append(data)
        data = float(line[10:20].strip()) # get the value of Real perc
        Re_I_list_p2[vo_index].append(data)
        data = float(line[20:].strip()) # get the value of Imag perc
        Im_I_list_p2[vo_index].append(data)

    return

# bhtpbank calculate total current
def total_current(Re_I_list,Im_I_list):
    total_I_list =[]
    for index, lst in enumerate(Re_I_list):
        total_I_list.append([])
        for index2, value in enumerate(Re_I_list[index]):
            data = math.sqrt(value*value + Im_I_list[index][index2]*Im_I_list[index][index2])
            total_I_list[index].append(data)
    return total_I_list

# bhtpbank auxiliary power check
def aux_power_calculate(Re_I_list,Re_I_list_p1,vol_list):
    aux_list = []
        
    for index, value in enumerate(vol_list):
        aux_list.append(value*Re_I_list[index][0]*Re_I_list_p1[index][0]/100/1000)
    
    return aux_list

# bhtpbank efficient calculate
def effi_calculation(TE_list,total_I_list,total_I_list_p1,vol_list,speed_list,BE_flag):

    effi_list = []
    mech_list = [] # mechnical power
    ele_list = [] # electric power
    ele_traction_list = [] # electrical power for traction
    
    for index, value in enumerate(vol_list):
        effi_list.append([])
        mech_list.append([])
        ele_list.append([])
        ele_traction_list.append([])
        for index2, value2 in enumerate(total_I_list[index]):
            data1 = TE_list[index][index2]*speed_list[index][index2]/3.6 # calculate mechnical power N*km/h/3.6
            data2 = value*value2/1000 # electrical power V * I/1000
            aux = data2*total_I_list_p1[index][index2]/100 # total power * percentage at 0 percent of mechanical           
            if aux >= 0:
                data3 = data2 - aux # power for traction, total electrical power - auxiliary power
            else:
                data3 = data2 # For comparision purpose only
            
            mech_list[index].append(data1)
            if BE_flag == False:
                ele_list[index].append(data2)
                ele_traction_list[index].append(data3)
                if data3 == 0:
                    effi_list[index].append(0.0)
                else:
                    effi_list[index].append(data1/data3) # mechanical / electrical
            else:
                ele_list[index].append(-data2)
                ele_traction_list[index].append(-data3)
                if data1 == 0:
                    effi_list[index].append(0.0)
                else:
                    effi_list[index].append(-data3/data1) # electrical / mechanical

    return effi_list, mech_list, ele_list, ele_traction_list

# efficiency check
def effi_check(effi_list,vol_list,BE_flag):
    if BE_flag == False:       
        # print("Traction Efficiency Checking Process...") 
        for index, value in enumerate(vol_list):
            if max(effi_list[index]) > 0.95:
                text = "WARNING: Maximum traction efficiency is {:.3f} at voltage level {} V!" \
                    .format(max(effi_list[index]),value)
                SharedMethods.print_message(text,"33")
    else:
        # print("Regen Efficiency Checking Process...")
        for index, value in enumerate(vol_list):
            if max(effi_list[index]) > 0.95:
                text = "WARNING: Maximum regen efficiency is {:.3f} at voltage level {} V!" \
                    .format(max(effi_list[index]),value)
                SharedMethods.print_message(text,"33")
    return True

# read individual ds1 file
def train_reading_process(simname, train_list,time_increment,start_df,result):
    # crate battery change summary
    battery_level_change = {}

    # create dataframe
    ds1dataframe = {}

    for train in train_list:
        df = None
        # find start_df column ("train") cororesponding value in column ("BHTP")
        # Get the row where the "train" column matches the current train value
        bhtp = start_df.loc[start_df['Train ID'] == train, 'Traction Profile']
        bhtp = bhtp.values[0]

        charging_effi = start_df.loc[start_df['Train ID'] == train, 'Charging Efficiency (%)']
        discharging_effi = start_df.loc[start_df['Train ID'] == train, 'Discharging Efficiency (%)']
        charging_effi = charging_effi.values[0]
        discharging_effi = discharging_effi.values[0]

        on_off_wire = start_df.loc[start_df['Train ID'] == train, 'On/Off Wire']
        on_off_wire = on_off_wire.values[0]
        
        
        print(f"Processing train No. {train} using {bhtp} ...")
        if 1 <= len(str(train)) <= 3:
            train_no = format(int(train),'03d')
        elif 4 <= len(str(train)) <= 5:
            train_no = format(int(train),'05d')
        filename = simname + "_train_" + train_no + ".osop.ds1"
        delimiter = '\\s+'
        columns = ["Train ID","Head Code","Time","Line Section (LS)","LS-Distance","Distance Gone","Branch","Branch Position","TE/BE Used", \
                   "TE/BE Used in %", "Pan Voltage (kV)", "Pan Angle (DEG)", "Current (A)", "Current Angle (DEG)", "Active Power (kW)", \
                    "Reactive Power (kVAr)", "Displacement Factor", "Inst Speed", "Average Speed", "REF1", "REF2"]
        
        dtype_mapping = {"Time": str,}

        df = pd.read_csv(filename, delimiter=delimiter, names = columns, skiprows = 16,dtype = dtype_mapping)  # You might need to adjust the delimiter based on your file
        #df["Time"] = pd.to_datetime(df["Time"],format='%H:%M:%S')
        # Extracting parts from the string and formatting the 'Time' column
        df['Time'] = df['Time'].apply(lambda x: f"{int(x[1:3]):02d}:{int(x[3:5]):02d}:{int(x[5:]):02d}")

        # data frame process
        df = ds1_file_process(df,time_increment,bhtp,charging_effi,discharging_effi,result,on_off_wire)

        # calcualte summary
        battery = df["Battery Egy (dis)charging (kWh)"].sum()

        #update result
        battery_level_change[train] = battery

        ds1dataframe[train] = df
    
    return battery_level_change, ds1dataframe

# ceate table 1 frame / power
def result_table_prepare(start_df, battery_level_change):    

    columns = start_df.columns.tolist()
    columns = columns + ['Battery Charged in Section (kWh)','Start Battery Level (kWh)','End Battery Level (kWh)', \
                         'Battery Charged in Section (%)', 'Start SoC (%)', 'End SoC (%)']
    
    # Iterate over rows and modify values to find out how many total jounerys
    start = []
    link = []
    for index, row in start_df.iterrows():
        # normally require use define NA, but change to not a number in case typo or different start
        link_previous = row["Link Previous"]
        if str(link_previous).isdigit():
            link_previous = int(link_previous)
            link.append([row["Link Previous"],row["Train ID"]])
        else:
            start.append(row["Train ID"])
        # if not isinstance(row["Link Previous"], (int, float)):
        #     start.append(row["Train ID"])
        # else:
        #     link.append([row["Link Previous"],row["Train ID"]])
    
    # interate to get jounery list
    route_list = []
    for header in start:
        search = header
        jounery = []
        jounery.append(header)
        
        route_finish = False # flag for whether this jounery is finished

        if link != []:
            while True:         
                for index, item in enumerate(link): # iterate all jounery link
                    update_flag = False # search updated or not

                    if item[0] == search:
                        jounery.append(item[1])
                        search = item[1]
                        update_flag = True
                        break
                    
                    if index == len(link)-1 and update_flag == False: # if searched all link and did find the next one
                        route_finish = True
                
                if route_finish == True:
                    break
        route_list.append(jounery)
    
    # preprae the result data in a sequence of train link
    new_row_list = []
    for jounery in route_list:
        level = 0 # ongoing battery level
        for train in jounery:
            row = start_df.loc[start_df['Train ID'] == train]
            battery = battery_level_change[train]
            size = row["Battery Size (kWh)"].iloc[0]
            ini = row["Journey Start Battery Level (%)"].iloc[0]
            if isinstance(ini, (int, float, np.int64, np.float64)):
                level = ini * size
            
            new_row = row.iloc[0].values.tolist()
            new_row.append(battery) # battery charged in section
            new_row.append(level) # section start energy
            new_row.append(level + battery) # section end energy
            if isinstance(size, (int, float, np.int64, np.float64)):
                new_row.append(battery / size) # in percentage
                new_row.append(level / size) # section start energy in per
                new_row.append((level + battery)/size) # section end energy
            else:
                new_row.append(1) # in percentage
                new_row.append(1) # section start energy
                new_row.append(1) # section end energy
            
            new_row_list.append(new_row)
            
            level = level + battery
    
    result_df = pd.DataFrame(data = new_row_list, columns = columns)

    return result_df, route_list

# doing invididual ds1 file process and calculatoin
def ds1_file_process(df, time_increment,bhtp,charging_effi,discharging_effi,result,on_off_wire):

    print("Calculating... Multiple data interpolation on-going. Please wait.......")
    bhtpbank_df_effi, bhtpbank_df_beffi, bhtpbank_df_be, aux_load = result

    # calculate instand speed
    df["Inst Speed (km/h)"] = df["Inst Speed"] * 1.609344 # Convert mph to kmh
    df["Average Speed (m/s)"] = df["Average Speed"] / 2.237 # Convert mph to m/s

    # calculate traction effort used for traction
    df["Effort for Traction (kN)"] = df.apply(calculate_BE_column, axis=1, args = (bhtpbank_df_be[bhtp],))

    # calculate mechanical energy
    df["Mechanical Egy Used (kWh)"] = df["Effort for Traction (kN)"] * df["Average Speed (m/s)"] * time_increment / 3600

    # calculate electricla energy consumped including battery power
    df["Electrical Egy Used (kWh)"] = df["Active Power (kW)"] * time_increment / 3600

    # calculate traction energy used (for traction only)
    df["Traction Egy Used (kWh)"] = df.apply(calculate_traction_column, axis=1, args = (bhtpbank_df_effi[bhtp],bhtpbank_df_beffi[bhtp]))

    # calculate auxiliary energu consumed
    df["Auxiliary Egy Used (kWh)"] = aux_load[bhtp] * time_increment / 3600
    
    if on_off_wire == "On":

        # calculate battery energy used (electrical - traction - auxilinary)
        df["Battery Egy Used (kWh)"] = df["Electrical Egy Used (kWh)"] - df["Traction Egy Used (kWh)"] - df["Auxiliary Egy Used (kWh)"]

        # calculat the battery energy changed
        df["Battery Egy (dis)charging (kWh)"] = df.apply(calculate_battery_column, axis=1, args = (charging_effi,discharging_effi,on_off_wire))
    
    else:

        # calculate battery energy used (electrical - traction - auxilinary)
        df["Battery Egy Used (kWh)"] = 0 - df["Electrical Egy Used (kWh)"]

        # calculat the battery energy changed
        df["Battery Egy (dis)charging (kWh)"] = df.apply(calculate_battery_column, axis=1, args = (charging_effi,discharging_effi,on_off_wire))

    return df

# calculate braking effort used
def calculate_BE_column(row,data_df):
    te_be_used = row['TE/BE Used']
    te_be_used_percent = row['TE/BE Used in %']
    voltage = row['Pan Voltage (kV)']
    speed = row['Inst Speed (km/h)']

    if te_be_used < 0 and te_be_used_percent == 100:
        # Perform data grid interpolation
        interpolated_value = interpolate_cal_value(voltage, speed, data_df)
        return -interpolated_value
    else:
        # Take TE/BE Used value and convert from lbf to kN
        return te_be_used * 0.0044482216

# calculate traction energy
def calculate_traction_column(row,data_df,data_df_b):
    traction_used = row['Effort for Traction (kN)']
    mech_energy = row['Mechanical Egy Used (kWh)']
    voltage = row['Pan Voltage (kV)']
    speed = row['Inst Speed (km/h)']

    if traction_used >= 0:
        if mech_energy == 0:
            return 0
        else:
            interpolated_value = interpolate_cal_value(voltage, speed, data_df)
            return (mech_energy / interpolated_value)
    else:
        interpolated_value = interpolate_cal_value(voltage, speed, data_df_b)

        return (mech_energy * interpolated_value)

# linear interploation formula
def interpolate_cal_value(voltage, speed, data_df):

    # Points for interpolation
    points = np.array(list(zip(data_df['x'], data_df['y'])))
    values = data_df['z']

    # Point to interpolate
    xi = np.array([voltage, speed])

    # Check if the point is within the convex hull of the input points
    if Delaunay(points).find_simplex(xi) >= 0:
        # Interpolate using linear method
        zi = griddata(points, values, xi, method='linear')
    else:
        # Interpolate using nearest method
        zi = griddata(points, values, xi, method='nearest')

    return (zi[0])

# calcualte battery charging based on efficiency
def calculate_battery_column(row, effi, dis_effi, on_off):
    traction_used = row['Effort for Traction (kN)']
    battery_used = row["Battery Egy Used (kWh)"]

    if on_off == "On":
        return (battery_used * effi)
    
    else:
        if battery_used < 0:
            return (battery_used / dis_effi)
        else:
            return (battery_used * effi)

def get_result_range(start,total,space,trainnumber):

    """
    Generates a list of range strings for different sections in a spreadsheet.

    Parameters:
    - start (int): Starting row index.
    - total (int): Total number of rows.
    - space (int): Space between sections.
    - fsnumber (int): Placeholder for fsnumber.

    Returns:
    List of range strings for different sections.
    """

    range_list = []
    # 0: table frame range 
    # 1: table data range 
    # 2: title range (2nd row)
    # 3: title range (1st row simulation info section1,2,3)
    # 4: title range (1st data)

    
    # 0
    result_range = [f"B{start + 1}:R{start + total + 1}"]
    range_list.append(result_range)
    # 1
    data_range = [f"M{start + 2}:R{start + total + 1}"]
    range_list.append(data_range)
   
    # 2
    title_range = [f"B{start+1}:R{start+1}"]
    range_list.append(title_range)
    
    #3 4
    add_range_1 = [f"B{start}:L{start}"]
    range_list.append(add_range_1)

    add_range_2 = [f"M{start}:R{start}"]
    range_list.append(add_range_2)

    # 5 percentage result
    add_range_3 = [f"H{start + 2}:H{start + total + 1}",f"J{start + 2}:K{start + total + 1}",f"P{start + 2}:R{start + total + 1}"]
    range_list.append(add_range_3)

    # 6 color coding range
    SoC_range = [f"Q{start + 2}:R{start + total + 1}"]
    range_list.append(SoC_range)

    return range_list

# Result table table formating
def table_formatting(simname,wb,range_list):
    # format the result table

    print("Formatting Process ...")
    # wb = load_workbook(excel_file)
    print("Information Collection ...")
    sheet = wb["Start"]
    project_name = sheet['B2'].value
    feeding_desp = sheet['B4'].value
    modeller = sheet['B5'].value
    date = sheet['B6'].value
    
    # Result Tab process
    sheet = wb["Result"]
    sheet['B2'].value = "Project Name:"
    sheet['C2'].value = project_name
    sheet['B3'].value = "Simulation Name:"
    sheet['C3'].value = simname
    sheet['B4'].value = "Feeding Arrangement:"
    sheet['C4'].value = feeding_desp
    sheet['B5'].value = "Result Created by:"
    sheet['C5'].value = modeller
    sheet['B6'].value = "Result Created at:"
    sheet['C6'].value = datetime.now().strftime("%d-%m-%Y %H:%M")

    apply_border(sheet,range_list)
    apply_title(sheet,range_list)
    
    print("Apply Numbering Format ...")
    # Define a custom style for formatting with two decimal places
    for range_name in range_list[1]:
        for row in sheet[range_name]:
            for cell in row:
                #cell.style = NamedStyle(name='decimal_style', number_format='0.00')
                cell.number_format = '0.00'
    
    #for range_name in range_list[7][0]:
    for range_name in range_list[5]:
        for row in sheet[range_name]:
            for cell in row:
                #cell.style = NamedStyle(name='decimal_style', number_format='0.00')
                cell.number_format = '0.00%'

    print("Apply Font ...")
    # Shade the range B11:Z11 with a light gray background color
    for range_name in range_list[2]:
        for row in sheet[range_name]:
            for cell in row:
                cell.font = Font(italic=True, size = 10)
    
    print("Apply Shading ...")
    # Shade the range B11:Z11 with a light gray background color
    for range_name in range_list[2]+range_list[3]+range_list[4]:
        for row in sheet[range_name]:
            for cell in row:
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    condtional_formating(sheet,range_list)

    print("Apply Column Length ...")
    total = int(range_list[0][0][-2:])
    # Auto-adjust the width of column B based on the content in B2 to B6
    max_length = max(len(str(sheet.cell(row=i, column=2).value)) for i in range(2, total+1))
    sheet.column_dimensions['B'].width = max_length + 2  # Add a little extra space
    max_length = max(len(str(sheet.cell(row=i, column=4).value)) for i in range(2, total+1))
    sheet.column_dimensions['D'].width = max_length + 3  # Add a little extra space

    # Auto-size columns after applying formatting
    for col_letter in ["C","E","F","G","H","I","J","K","L"]:
        sheet.column_dimensions[col_letter].auto_size = True

    # # Save changes
    # wb.save(excel_file)
    return

def apply_border(sheet,range_list):
    print("Apply Border ...")
    for range_name in range_list[0]+range_list[2]+range_list[3]+range_list[4]:
        for row in sheet[range_name]:
            for cell in row:
                # Apply border to all sides of the cell
                cell.border = Border(left=Side(border_style='thin'),
                                    right=Side(border_style='thin'),
                                    top=Side(border_style='thin'),
                                    bottom=Side(border_style='thin'))

                # Align cell content to the middle
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    return

def apply_title(sheet,range_list):

    print("Apply Title ...")
    # for range_name in range_list[3]:
    #     sheet.merge_cells(range_name)
    #     part = range_name.split(":")
    #     cell = sheet[part[0]]
    #     cell.value = "Site Information Summary"
    #     cell.font = Font(bold=True)

    sheet.merge_cells(range_list[3][0])
    part = range_list[3][0].split(":")
    cell = sheet[part[0]]
    cell.value = "Train Section Information"
    cell.font = Font(bold=True)

    sheet.merge_cells(range_list[4][0])
    part = range_list[4][0].split(":")
    cell = sheet[part[0]]
    cell.value = "Simulation Results"
    cell.font = Font(bold=True)

    return

def condtional_formating(sheet,range_list):
    print("Apply Conditional Formatting ...")
    # Compare values in columns I and J for each row and shade accordingly
    # set the pattern fill
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red (0% S)
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 255,152,51 (80% S)
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green (00% S)

    # power
    #yellow_line = str(float(sheet[range_list[6][0]])*0.9)
    # sheet.conditional_formatting.add(range_list[6][1],CellIsRule(operator = 'lessThan',formula=[range_list[6][0]+'*0.9'],fill=green_fill))
    # sheet.conditional_formatting.add(range_list[6][1],CellIsRule(operator = 'greaterThanOrEqual',formula=[range_list[6][0]],fill=red_fill))
    # sheet.conditional_formatting.add(range_list[6][1],CellIsRule(operator = 'between',formula=[range_list[6][0]+'*0.9',range_list[6][0]],fill=yellow_fill))
    # SoC
    sheet.conditional_formatting.add(range_list[6][0],CellIsRule(operator = 'between',formula=['0.15','0.85'],fill=green_fill))
    sheet.conditional_formatting.add(range_list[6][0],CellIsRule(operator = 'between',formula=['0','0.15'],fill=yellow_fill))
    sheet.conditional_formatting.add(range_list[6][0],CellIsRule(operator = 'between',formula=['0.85','1'],fill=yellow_fill))
    sheet.conditional_formatting.add(range_list[6][0],CellIsRule(operator = 'lessThan',formula=['0'],fill=red_fill))
    sheet.conditional_formatting.add(range_list[6][0],CellIsRule(operator = 'greaterThan',formula=['1'],fill=red_fill))

    return


if __name__ == "__main__":
    # Add your debugging code here
    simname = "test_model_3"  # Provide a simulation name or adjust as needed
    main_option = "1"  # Adjust as needed
    time_start = "0070000"  # Adjust as needed
    time_end = "0080000"  # Adjust as needed
    option_select = "4"  # Adjust as needed
    text_input = "AC_07_BEMU2"  # Adjust as needed
    low_v = None  # Adjust as needed
    high_v = None  # Adjust as needed
    time_step = None  # Adjust as needed

    # Call your main function with the provided parameters
    main(simname, main_option, time_start, time_end, option_select, text_input, low_v, high_v, time_step)

