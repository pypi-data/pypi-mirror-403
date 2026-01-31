#
# -*- coding: utf-8  -*-
#=================================================================
# Created by: Jieming Ye
# Created on: April 2025
# Last Modified: April 2025
#=================================================================
# Copyright (c) 2025 [Jieming Ye]
#
# This Python source code is licensed under the 
# Open Source Non-Commercial License (OSNCL) v1.0
# See LICENSE for details.
#=================================================================
"""
Pre-requisite: 
SimulationName.battery.txt: default oslo output file after successfully run a simulation from version RN29 with battery train running.
xxxxx.xlsx: excel spreadsheet with proper user settings.
Used Input:
simname: to locate the result file or file rename
time_start: to define the analysis start time
time_end: to define the analysis end time
option_select: to define the option selected in subfunction
text_input: to locate the excel file name
other input for oslo_extraction.py only.
Expected Output:
Updated excel spreadsheet with analysis result.
Description:
TBC
"""
#=================================================================
# VERSION CONTROL
# V1.0 (Jieming Ye) - Initial Version
#=================================================================
# Set Information Variable
# N/A
#=================================================================

# vision_oslo_extension/excel_processing.py
import math
import pandas as pd
import numpy as np
import string
# import openpyxl
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, NamedStyle
from openpyxl.formatting.rule import CellIsRule,FormulaRule
from openpyxl.chart import ScatterChart,Reference,Series
from openpyxl.drawing.text import CharacterProperties
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import SeriesLabel

# import vision_oslo
from vision_oslo_extension.shared_contents import SharedMethods


def main(simname, main_option, time_start, time_end, option_select, text_input, low_v, high_v, time_step):
    
    print("")
    print("New Battery Train Assessment From VISION RN29- - - > ")
    print("")

    simname = simname
    excel_file = text_input + ".xlsx"
    result_file = simname + ".battery.txt"
    result_file_2 = simname + ".traction.txt"

    option = option_select # 1:Fully Restart, require oof, and list

    time_increment = 5
    start = 10
    space  = 5

    if not SharedMethods.check_existing_file(result_file):
        SharedMethods.print_message(f"ERROR: Ensure you rerun the simulation with 'Run Battery Train' option selected !","31")
        return False
    
    if not SharedMethods.check_existing_file(result_file_2):
        SharedMethods.print_message(f"ERROR: Ensure you rerun the simulation with 'Run Battery Train' option selected !","31")
        return False
    
    if not SharedMethods.check_existing_file(excel_file):
        return False

    # start_cell = 'B11'
    # read data from start tab
    result = start_reading_process(
        simname, time_start, time_end, text_input, low_v, high_v, time_step,excel_file,option
    )
    if result == False:
        return False
    else:
        plot_type,marker_route_list, marker_location_list, marker_distance_list, start_df = result

    # process .battery.txt file
    result = train_reading_process(simname,time_increment,start_df,result_file)
    if result == False:
        return False
    else:
        result_data, batterydataframe = result

    # process .traction.txt file
    tractiondataframe = train_reading_process_traction(simname,time_increment,start_df,result_file_2)

    # append traction data to battery for bettery analysis
    result_data, ds1dataframe = additional_data_process(result_data,time_increment,batterydataframe,tractiondataframe)

    # update result
    result_df = result_table_prepare(start_df, result_data)

    # update the excel
    data_write_save(simname,excel_file,start,space,start_df,result_df,ds1dataframe,plot_type,marker_route_list, marker_location_list, marker_distance_list)

    return True


# write data to excel
def data_write_save(simname, excel_file,start,space,start_df,result_df,ds1dataframe,plot_type,marker_route_list, marker_location_list, marker_distance_list):
    # write data to excel
    with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a',if_sheet_exists='overlay') as writer:

        # get currrent workbook to do process
        wb = writer.book

        print("Generate Result Page...")
        result_df.to_excel(writer, sheet_name="Result", index=False, startrow = start, startcol = 1)
        
        # Write each DataFrame to a different sheet in the Excel file
        print("Writing individual Trains...")
        train_list = start_df['Train ID'].tolist()
        for index, train in enumerate(train_list):
            
            print(f"Saving {train}...")
            dflist = ds1dataframe[index]
            # emty columns
            dflist.insert(dflist.columns.get_loc('Control Mode No.')+1,'New_C_1', np.nan)
            dflist.insert(dflist.columns.get_loc('Track Energy (kWh)')+1,'New_C_2', np.nan)
            dflist.insert(dflist.columns.get_loc('New_C_2')+1,'Excel_Time', np.nan)
            dflist.insert(dflist.columns.get_loc('Excel_Time')+1,'Excel_SoC', np.nan)

            dflist.to_excel(writer, sheet_name=f'{index+1}_Train_{train}', index=False, startrow = 2)
        
        data_row = 4 # first data row number

        print("Add Information Supporting Plotting....")
        addtional_info(wb,train_list,ds1dataframe,result_df,data_row)
        
        # Calculate the Excel range for each DataFrame
        total = len(train_list)
        range_list = get_result_range(start,total,space) # range list as [power data range, current data range]
        
        # table formatting
        table_formatting(simname,wb,range_list)

        # prepare for additional marker
        datapoint_row,datapoint_location =  put_dummy_marker(wb,train_list,ds1dataframe,result_df,data_row,marker_route_list, marker_location_list, marker_distance_list)

        # jounery plotting
        jounery_plot(wb,train_list,ds1dataframe,result_df,data_row,datapoint_row,datapoint_location,plot_type)

        print("Saving Data...")

    return

def put_dummy_marker(wb,train_list,ds1dataframe,result_df,data_row,marker_route_list, marker_location_list, marker_distance_list):
    
    print("Analysing Additional Marker...")
    # this will return a final list as data point that could be shown as a new series and added to to the plot
    # datapoint_row = [] (excel row number)
    datapoint_row = []
    datapoint_location = []
    for index, train in enumerate(train_list):
        datapoint_row.append([])
        datapoint_location.append([])

        sheet = wb[f'{index+1}_Train_{train}']
        route_name = result_df.loc[index, 'Train Journey'] # get the route name
        train_df = ds1dataframe[index] # get the data frame

        for index2, marker_route in enumerate(marker_route_list):
            if marker_route == route_name:

                for index3,item in enumerate(marker_distance_list[index2]):
                    # find dataframe index with closest distance
                    closest_index = (train_df['Distance Gone (m)'] - item).abs().idxmin()
                    relative_index = train_df.index.get_loc(closest_index)
                    datapoint_row[index].append(int(data_row+relative_index))
                    datapoint_location[index].append(marker_location_list[index2][index3])
            
    return datapoint_row,datapoint_location

def addtional_info(wb,train_list,ds1dataframe,result_df,data_row):
       
    for index, train in enumerate(train_list):

        sheet = wb[f'{index+1}_Train_{train}']
        dflist = ds1dataframe[index]

        # add formula in
        # sheet["W2"].value = "Convert mph to kmh"

        # create excel time format
        index = data_row
        data_range = f"Y{data_row}:Y{data_row+len(dflist)-1}"
        for line in sheet[data_range]:
            for cell in line:
                cell.value = f'=TIME(MID(B{index},4,2),MID(B{index},7,2),RIGHT(B{index},2))'
                cell.number_format = 'hh:mm:ss'
            index = index + 1
        
        # create SoC info
        index = data_row
        data_range = f"Z{data_row}:Z{data_row+len(dflist)-1}"
        for line in sheet[data_range]:
            for cell in line:
                cell.value = f'=P{index}/100'
                cell.number_format = '0.00%'
            index = index + 1

    return

def jounery_plot(wb,train_list,ds1dataframe,result_df,data_row,datapoint_row,datapoint_location,plot_type):
    
    print("Plotting Jounery Profile...")
    
    # set default variable
    c1 = 25 # Time
    c2 = 6  # distance gone
    c3 = 15 # Energy Level
    c4 = 26 # SoC

    for plot_no in range(1,3): # iterate i == 1 (SOC) and i==2 (ENERGY)
        if plot_no == 1:
            cp1_start = c1
            cp1_end = c4
            cp2_start = c2
            cp2_end = c4
            cp1_loc = "B2"
            cp2_loc = "B25"
            title_suffix = " Battery SoC Profile"
            y_title = "State of Charge SoC (%)"

        else:
            cp1_start = c1
            cp1_end = c3
            cp2_start = c2
            cp2_end = c3
            cp1_loc = "T2"
            cp2_loc = "T25"
            title_suffix = " Battery Energy Level Profile"
            y_title = "Battery Energy Level (kWh)"

        for index, train in enumerate(train_list):
            # get essential row ID from dataframe
            train_df = ds1dataframe[index]
            total_rows, row_list,mode_changes = train_plot_prepare(train_df)

            route_name = result_df.loc[index, 'Train Journey']
            sheet_name = f'{index+1}_Train_{train}'

            # plot two plot, battery level vs time / location
            sheet = wb[sheet_name]        

            chart = ScatterChart('smoothMarker')
            chart.title = route_name + title_suffix
            chart.x_axis.title = "Time (hh:mm:ss)"
            chart.y_axis.title = y_title

            # Set chart size
            chart.height = 10  # Adjust width as needed
            chart.width = 30  # Adjust height as needed
            xlimmin = 1
            xlimmax = 0

            if plot_type == 2:
                for index2, mode in enumerate(mode_changes):
                    if index2 == len(mode_changes)-1: # last mode
                        xvalues = Reference(sheet, min_col=cp1_start, min_row=data_row+row_list[index2], max_row=data_row+total_rows-1)
                        yvalues = Reference(sheet, min_col=cp1_end, min_row=data_row+row_list[index2], max_row=data_row+total_rows-1)
                    else:
                        xvalues = Reference(sheet, min_col=cp1_start, min_row=data_row+row_list[index2], max_row=data_row+row_list[index2+1]-1)
                        yvalues = Reference(sheet, min_col=cp1_end, min_row=data_row+row_list[index2], max_row=data_row+row_list[index2+1]-1)
                    
                    series = Series(yvalues,xvalues,title_from_data=False,title = mode)
                    chart.series.append(series)
            else:
                xvalues = Reference(sheet, min_col=cp1_start, min_row=data_row, max_row=data_row+total_rows-1)
                yvalues = Reference(sheet, min_col=cp1_end, min_row=data_row, max_row=data_row+total_rows-1)
                series = Series(yvalues,xvalues,title_from_data=False,title = route_name)
                chart.series.append(series)

            t_start = ds1dataframe[index]['Time(D/HH:MM:SS)'].iloc[0]
            t_start = t_start.split(":")
            ts = (float(t_start[0][-2:])+float(t_start[1])/60+float(t_start[1])/3600)/24

            t_end = ds1dataframe[index]['Time(D/HH:MM:SS)'].iloc[-1]
            t_end = t_end.split(":")
            te = (float(t_end[0][-2:])+float(t_end[1])/60+float(t_end[1])/3600)/24

            if ts < xlimmin:
                xlimmin = ts
            if te > xlimmax:
                xlimmax = te
                    
            # find nearest 10 min timeline.
            for i in range(0,144):
                if i*1/144 > xlimmin:
                    xlimmin = (i-1)*1/144
                    for j in range(i,144):
                        if j*1/144 >= xlimmax:
                            xlimmax = j*1/144
                            break
                    break
            
            tspace = (xlimmax-xlimmin)/8
            
            chart.legend.position = 'b' # top legend position
            if plot_no == 1:
                chart.y_axis.scaling.min = 0      
                # chart.y_axis.scaling.max = 1
                chart.y_axis.majorUnit = 0.1 # 10 percent

            chart.x_axis.scaling.min = xlimmin
            chart.x_axis.scaling.max = xlimmax
            chart.x_axis.majorUnit = tspace # 
            
            cp = CharacterProperties(sz=1400)
            chart.title.tx.rich.p[0].r[0].rPr = cp 

            cl = GraphicalProperties()
            cl.line.solidFill = "E0E0E0"
            chart.x_axis.majorGridlines.spPr = cl  # Light grey color
            chart.y_axis.majorGridlines.spPr = cl  #  # Light grey color

            # Addtitional data label marker as data point series
            if datapoint_row[index] != []:
                for index2, location in enumerate(datapoint_location[index]):
                    xvalues = Reference(sheet, min_col=cp1_start, min_row=datapoint_row[index][index2], max_row=datapoint_row[index][index2])
                    yvalues = Reference(sheet, min_col=cp1_end, min_row=datapoint_row[index][index2], max_row=datapoint_row[index][index2])
                    label_series = Series(yvalues, xvalues)

                    label_series.dLbls = DataLabelList()
                    label_series.dLbls.showVal = False
                    label_series.dLbls.showSerName = True
                    label_series.dLbls.showLegendKey = False
                    label_series.dLbls.showCatName = False # altenative solution is to change this to name, not too good
                    label_series.dLbls.dLblPos = "t"  # position: right, other options: 't', 'l', 'b', 'ctr'
                    label_series.title = SeriesLabel(v=location)

                    # Make line invisible
                    label_series.graphicalProperties.line.noFill = True
                    label_series.marker.symbol = "circle"
                    label_series.marker.size = 5  # increase size to visible

                    # Add to chart
                    chart.series.append(label_series)
            
            sheet.add_chart(chart,cp1_loc)

            ###########################################################################################
            chart = ScatterChart('smoothMarker')
            chart.title = route_name + title_suffix
            chart.x_axis.title = "Distance Gone (meter)"
            chart.y_axis.title = y_title

            # Set chart size
            chart.height = 10  # Adjust width as needed
            chart.width = 30  # Adjust height as needed
            xlimmin = 1000000 # 1000 km
            xlimmax = 0

            if plot_type == 2:
                for index2, mode in enumerate(mode_changes):
                    if index2 == len(mode_changes)-1: # last mode
                        xvalues = Reference(sheet, min_col=cp2_start, min_row=data_row+row_list[index2], max_row=data_row+total_rows-1)
                        yvalues = Reference(sheet, min_col=cp2_end, min_row=data_row+row_list[index2], max_row=data_row+total_rows-1)
                    else:
                        xvalues = Reference(sheet, min_col=cp2_start, min_row=data_row+row_list[index2], max_row=data_row+row_list[index2+1]-1)
                        yvalues = Reference(sheet, min_col=cp2_end, min_row=data_row+row_list[index2], max_row=data_row+row_list[index2+1]-1)
                    
                    series = Series(yvalues,xvalues,title_from_data=False,title = mode)
                    chart.series.append(series)
            else:
                xvalues = Reference(sheet, min_col=cp2_start, min_row=data_row, max_row=data_row+total_rows-1)
                yvalues = Reference(sheet, min_col=cp2_end, min_row=data_row, max_row=data_row+total_rows-1)
                series = Series(yvalues,xvalues,title_from_data=False,title = route_name)
                chart.series.append(series)

            t_start = ds1dataframe[index]['Distance Gone (m)'].iloc[0]
            t_end = ds1dataframe[index]['Distance Gone (m)'].iloc[-1]

            if t_start < xlimmin:
                xlimmin = float(t_start)
            if t_end > xlimmax:
                xlimmax = float(t_end)
                    
            tspace = (xlimmax-xlimmin)/10

            chart.legend.position = 'b' # top legend position

            if plot_no == 1:
                chart.y_axis.scaling.min = 0      
                # chart.y_axis.scaling.max = 1
                chart.y_axis.majorUnit = 0.1 # 10 percent

            chart.x_axis.scaling.min = xlimmin
            chart.x_axis.scaling.max = xlimmax

            chart.x_axis.majorUnit = tspace # 

            
            cp = CharacterProperties(sz=1400)
            chart.title.tx.rich.p[0].r[0].rPr = cp 

            cl = GraphicalProperties()
            cl.line.solidFill = "E0E0E0"
            chart.x_axis.majorGridlines.spPr = cl  # Light grey color
            chart.y_axis.majorGridlines.spPr = cl  #  # Light grey color

            # Addtitional data label marker as data point series
            if datapoint_row[index] != []:
                for index2, location in enumerate(datapoint_location[index]):
                    xvalues = Reference(sheet, min_col=cp2_start, min_row=datapoint_row[index][index2], max_row=datapoint_row[index][index2])
                    yvalues = Reference(sheet, min_col=cp2_end, min_row=datapoint_row[index][index2], max_row=datapoint_row[index][index2])
                    label_series = Series(yvalues, xvalues)

                    label_series.dLbls = DataLabelList()
                    label_series.dLbls.showVal = False
                    label_series.dLbls.showSerName = True
                    label_series.dLbls.showLegendKey = False
                    label_series.dLbls.showCatName = False # altenative solution is to change this to name, not too good
                    label_series.dLbls.dLblPos = "t"  # position: right, other options: 't', 'l', 'b', 'ctr'
                    label_series.title = SeriesLabel(v=location)

                    # Make line invisible
                    label_series.graphicalProperties.line.noFill = True
                    label_series.marker.symbol = "circle"
                    label_series.marker.size = 5  # increase size to visible

                    # Add to chart
                    chart.series.append(label_series)

            sheet.add_chart(chart,cp2_loc)

    return

# find row number based on dataframe
def train_plot_prepare(train_df):
    # get the total rows

    total_rows = train_df.shape[0]

    mode_changes = []
    row_list = []
    previous_mode = None

    # check row number for each mode change
    # interate each row, when never there is a change in 'Operation Mode No.', record the row number
    # start a new count, and record the Operation Mode No.
    for i, current_mode in enumerate(train_df['Operation Mode No.']):
        if current_mode != previous_mode:
            match current_mode:
                case 0:
                    mode_changes.append("Battery Suspended")
                case 1:
                    mode_changes.append("On-Wire Charging")
                case 2:
                    mode_changes.append("On-Wire Discharging")
                case 3:
                    mode_changes.append("Off-Wire Charging")
                case 4:
                    mode_changes.append("Off-Wire Discharging")
            row_list.append(i)
            previous_mode = current_mode

    return total_rows,row_list,mode_changes

# read the start tab and collect informaiton
def start_reading_process(simname, time_start, time_end, text_input, low_v, high_v, time_step, excel_file, option):
    
    # File check
    print("Check Excel and Deleting Existing Contents ...")

    try:
        wb = load_workbook(excel_file)
        # Delete all sheets except 'Start'
        for sheet_name in wb.sheetnames:
            if sheet_name == 'Start':
                table_list = check_table_list(wb[sheet_name])
                marker_route_list, marker_location_list, marker_distance_list = check_location_list(wb[sheet_name])
                
                if table_list is not False:
                    for line in table_list:
                        route = line[0]
                        if route not in marker_route_list:
                            print(f"INFO: Route '{route}' does not have additional marker information for plotting.")
                        train = line[1]
                        if not str(train).isdigit():  # This works for positive integers represented as strings
                            SharedMethods.print_message(f"ERROR: Train ID '{train}' is not a valid number. Please check your input.","31")
                            return False
                else:
                    return False
                # read the plotting settings (K7)
                if wb[sheet_name].cell(row=7, column=11).value == 'Mode Split':
                    plot_type = 2
                elif wb[sheet_name].cell(row=7, column=11).value == 'Continues':
                    plot_type = 1
                else:
                    SharedMethods.print_message(f"WARNING: Plotting type (Mode Split/Continues) not defined at K7. Default to 'Continues'.","33")
                    plot_type = 1
            else:
                wb.remove(wb[sheet_name])
        
        # Save changes
        wb.save(excel_file)

    except Exception as e:
        SharedMethods.print_message(f"(Close the Excel file and Start Again) ERROR: {e}","31")
        return False
        
    # read start up information 
    columns = ["Train Journey","Train ID","Modelled Battery Size (kWh)","High SoC Warning (%)", \
               "Low SoC Warning (%)","Optional Battery Size (kWh)","Optional Battery SoC (%)"]
    start_df = pd.DataFrame(table_list,columns=columns)
    
    return plot_type, marker_route_list, marker_location_list, marker_distance_list, start_df

# check table list on start page
def check_table_list(sheet):
    print("Reading Configuration Setting ...")
    table_row = 11
    table_start_column = 1
    table_end_column = 7
    # create node and feeder list
    table_list = []

    # check feeder
    index = table_row
    column = table_start_column
    if sheet.cell(row=index, column=column+1).value is not None:
        while True:
            row_data = []
            for temp in range(table_start_column,table_end_column + 1):
                row_data.append(sheet.cell(row=index, column=temp).value)
            table_list.append(row_data)
            index += 1
            
            check = sheet.cell(row=index, column=column+1).value
            if check is None:
                break
    else:
        SharedMethods.print_message("ERROR: Wrong data format. No information at B11.","31")
        return False

    return table_list

# check location list on start page
def check_location_list(sheet):
    print("Reading Addtional Location Marker ...")
    table_row = 11
    table_start_column = 10
    table_end_column = 12
    # create location list
    marker_route_list = [] # format[route1, route2,...]
    marker_location_list = [] # format[[marker1, marker2 ...],[...],...]
    marker_distance_list = [] # format[[distance1, distance2 ...],[...],...]

    # check configuration list
    index = table_row
    column = table_start_column
    marker_flag = False # flag if read marker line
    index_location = 0
    if sheet.cell(row=index, column=column+1).value is not None:
        while True:
            route = sheet.cell(row=index, column=column).value
            distance = sheet.cell(row=index, column=column+1).value
            location = sheet.cell(row=index, column=column+2).value
            if route is not None:
                if route not in marker_route_list:
                    marker_route_list.append(route)
                    marker_location_list.append([])
                    marker_distance_list.append([])
                    index_location = index_location + 1
                    marker_flag = True
                else:
                    SharedMethods.print_message(f"WARNING: Duplicate route '{route}' detected. 2nd one will be ignored.","33")
                    marker_flag = False
            if marker_flag == True:
                if distance is not None:
                    marker_distance_list[index_location-1].append(distance)
                    marker_location_list[index_location-1].append(location)
            # check whether we should exit the loop
            if sheet.cell(row=index+1, column=column+1).value is None:
                marker_flag = False
                if sheet.cell(row=index+2, column=column+1).value is None:
                    break
            # read next line
            index = index + 1
    else:
        print(f"INFO: No additional marker defined as no information in 'K11'.")

    return marker_route_list, marker_location_list, marker_distance_list

# read battery file summary
def train_reading_process(simname,time_increment,start_df,result_file):
    print("Analysing Train Information...")

    # read the result file using comma spaced
    columns = ["TimeStep(S)","Time(D/HH:MM:SS)","Train ID","Rolling Stock No.","Battery Unit No.","Distance Gone (m)","Branch","Branch Position", \
                "Voltage Re (V)", "Voltage Im (V)", "Current Re (A)", "Current Im (A)", "Active Power (kW)", "Reactive Power (kVAR)", \
                "Battery Energy Level (kWh)", "Battery SoC Level (%)", "Operation Mode No.","Control Mode No."]

    df = pd.read_csv(result_file, delimiter=',', names = columns, skiprows = 8)  # You might need to adjust the delimiter based on your file

    # Extracting parts from the string and formatting the 'Time' column
    # df['Time'] = df['Time'].apply(lambda x: f"{int(x[1:3]):02d}:{int(x[3:5]):02d}:{int(x[5:]):02d}")

    train_list = start_df['Train ID'].tolist()
    # create dataframe
    ds1dataframe = []
    result_data = []
    for train in train_list:
        # Filter the DataFrame for the current Train ID
        train_df = df[df['Train ID'] == train].copy()

        if train_df.empty:
            SharedMethods.print_message(f"ERROR: Train {train} does NOT exist in the simulation result. Check your settings...","31")
            return False

        # Get start and end values
        start_energy = train_df['Battery Energy Level (kWh)'].iloc[0]
        end_energy = train_df['Battery Energy Level (kWh)'].iloc[-1]
        change_energy = end_energy - start_energy
        start_SoC = train_df['Battery SoC Level (%)'].iloc[0]
        end_SoC = train_df['Battery SoC Level (%)'].iloc[-1]
        change_SoC = end_SoC - start_SoC

        # Get max and min values
        max_energy = train_df['Battery Energy Level (kWh)'].max()
        min_energy = train_df['Battery Energy Level (kWh)'].min()
        max_SoC = train_df['Battery SoC Level (%)'].max()
        min_SoC = train_df['Battery SoC Level (%)'].min()

        # Count time spent in each operation mode
        mode_counts = train_df['Operation Mode No.'].value_counts()

        mode0_time = mode_counts.get(0, 0) * time_increment
        mode1_time = mode_counts.get(1, 0) * time_increment
        mode2_time = mode_counts.get(2, 0) * time_increment
        mode3_time = mode_counts.get(3, 0) * time_increment
        mode4_time = mode_counts.get(4, 0) * time_increment

        result_data.append([mode0_time,mode1_time,mode2_time,mode3_time,mode4_time,start_energy,end_energy,change_energy, \
                            max_energy,min_energy,start_SoC,end_SoC,change_SoC,max_SoC,min_SoC])
        ds1dataframe.append(train_df)
    
    return result_data, ds1dataframe

# read traction file summary
def train_reading_process_traction(simname,time_increment,start_df,result_file):
    # read the result file using comma spaced
    columns = ["TimeStep(S)","Time(D/HH:MM:SS)","Train ID","Rolling Stock No.","Battery Unit No.","Distance Gone (m)","Branch","Branch Position", \
                "Voltage Re (V)", "Voltage Im (V)", "Current Re (A)", "Current Im (A)", "Active Power (kW)", "Reactive Power (kVAR)", \
                "Reserve", "Operation Mode No.","Control Mode No."]

    df = pd.read_csv(result_file, delimiter=',', names = columns, skiprows = 8)  # You might need to adjust the delimiter based on your file

    # Extracting parts from the string and formatting the 'Time' column
    # df['Time'] = df['Time'].apply(lambda x: f"{int(x[1:3]):02d}:{int(x[3:5]):02d}:{int(x[5:]):02d}")

    train_list = start_df['Train ID'].tolist()
    # create dataframe
    ds1dataframe = []

    for train in train_list:
        # Filter the DataFrame for the current Train ID
        train_df = df[df['Train ID'] == train].copy()
        ds1dataframe.append(train_df)
    
    return ds1dataframe

# append addtional data from traction file to battery data list
def additional_data_process(result_data,time_increment,ds1dataframe,ds2dataframe):
    for index, train_df in enumerate(ds1dataframe):
        train_df["Track P (kW)"] = ds2dataframe[index]["Active Power (kW)"]
        train_df["Track Q (kVAR)"] = ds2dataframe[index]["Reactive Power (kVAR)"]
        train_df["Track S (kVA)"] = np.sqrt(train_df["Track P (kW)"]**2 +train_df["Track Q (kVAR)"]**2)
        train_df["Track Energy (kWh)"] = train_df["Track P (kW)"] * time_increment / 3600

        # adding column track energy when Operation Mode No. is zero and track P is postive.
        # Define mask for the condition
        mask = (train_df["Operation Mode No."] == 0) & (train_df["Battery SoC Level (%)"] == 0) # battery turned off but still require traction
        additional_energy = train_df.loc[mask,"Track Energy (kWh)"].sum()

        # insert the data to proper place
        result_row = result_data[index]
        result_data[index] = result_row[:-5] + [additional_energy] + result_row[-5:] # put it five column before

    return result_data,ds1dataframe

# ceate table 1 frame / power
def result_table_prepare(start_df, result_data):    

    columns = start_df.columns.tolist()
    columns = columns + ['Battery Idle','On-Wire Charging','On-wire Discharging','Off-wire Charging', \
                         'Off-wire Discharging', 'Start Energy (kWh)', 'End Energy (kWh)','Jounery Energy (kWh)', \
                            'Max Energy (kWh)','Min Energy (kWh)','Extra Required (kWh)', 'Start SoC (%)', 'End SoC (%)', \
                                'Jounery SoC (%)','Max SoC (%)','Min SoC (%)']
    
    # Build a list of combined rows
    result_rows = []

    for index, row in start_df.iterrows():
        result_row = result_data[index]
        # last 5 columns is SoC, convernt to percentage value
        result_row[-5:] = [x / 100 for x in result_row[-5:]]
        # Convert the row to a list, then add corresponding result_data
        combined_row = row.tolist() + result_row
        result_rows.append(combined_row)

    # Create the final DataFrame
    result_df = pd.DataFrame(data=result_rows, columns=columns)

    # drop two optional columns 'Optional Battery Size (kWh)","Optional Battery SoC (%)"
    result_df = result_df.drop(columns=["Optional Battery Size (kWh)", "Optional Battery SoC (%)"])
    
    return result_df

def get_result_range(start,total,space):

    """
    Generates a list of range strings for different sections in a spreadsheet.

    Parameters:
    - start (int): Starting row index.
    - total (int): Total number of rows.
    - space (int): Space between sections.

    Returns:
    List of range strings for different sections.
    """

    range_list = []
    # 0: table frame range 
    # 1: table data range 
    # 2: title range (2nd row)
    # 3: title range (1st row simulation info section1,2,3)
    # 4: title range (1st data)

    
    # 0
    result_range = [f"B{start + 1}:V{start + total + 1}"]
    range_list.append(result_range)
    # 1
    data_range = [f"G{start + 2}:V{start + total + 1}"]
    range_list.append(data_range)
   
    # 2
    title_range = [f"B{start+1}:V{start+1}"]
    range_list.append(title_range)
    
    #3 4
    add_range_1 = [f"B{start}:F{start}"]
    range_list.append(add_range_1)

    add_range_2 = [f"G{start}:K{start}",f"L{start}:Q{start}",f"R{start}:V{start}"]
    range_list.append(add_range_2)

    # 5 time result / integer
    time_range = [f"G{start + 2}:K{start + total + 1}"]
    range_list.append(time_range)

    # 6 Energy result / double decimal
    energy_range = [f"L{start + 2}:Q{start + total + 1}"]
    range_list.append(energy_range)

    # 7 SoC result / double decimal
    SoC_range = [f"R{start + 2}:V{start + total + 1}",f"E{start + 2}:F{start + total + 1}"]
    range_list.append(SoC_range)

    # 8 color coding SoC range (lower criteria, higher criteria, range)
    color_range = [f"$E{start + 2}",f"$F{start + 2}",f"$R${start + 2}:$S${start + total + 1}"]
    range_list.append(color_range)

    # 9 color coding SoC range (lower criteria, higher criteria, range)
    color_range1 = [f"$E{start + 2}",f"$F{start + 2}",f"$U${start + 2}:$V${start + total + 1}"]
    range_list.append(color_range1)   

    # 10 color coding range egy
    color_range2 = [f"Q{start + 2}:Q{start + total + 1}"]
    range_list.append(color_range2)

    return range_list

# Result table table formating
def table_formatting(simname,wb,range_list):
    # format the result table

    print("Formatting Process ...")
    # wb = load_workbook(excel_file)
    print("Information Collection ...")
    sheet = wb["Start"]
    project_name = sheet['B2'].value
    feeding_desp = sheet['B4'].value
    modeller = sheet['B5'].value
    date = sheet['B6'].value
    
    # Result Tab process
    sheet = wb["Result"]
    sheet['B2'].value = "Project Name:"
    sheet['C2'].value = project_name
    sheet['B3'].value = "Simulation Name:"
    sheet['C3'].value = simname
    sheet['B4'].value = "Feeding Arrangement:"
    sheet['C4'].value = feeding_desp
    sheet['B5'].value = "Result Created by:"
    sheet['C5'].value = modeller
    sheet['B6'].value = "Result Created at:"
    sheet['C6'].value = datetime.now().strftime("%d-%m-%Y %H:%M")

    apply_border(sheet,range_list)
    apply_title(sheet,range_list)
    
    print("Apply Numbering Format ...")
    # Define a custom style for formatting with two decimal places
    for range_name in range_list[6]:
        for row in sheet[range_name]:
            for cell in row:
                #cell.style = NamedStyle(name='decimal_style', number_format='0.00')
                cell.number_format = '0.00'
    
    #for range_name in range_list[7][0]:
    for range_name in range_list[7]:
        for row in sheet[range_name]:
            for cell in row:
                #cell.style = NamedStyle(name='decimal_style', number_format='0.00')
                cell.number_format = '0.00%'

    print("Apply Font ...")
    # Shade the range B11:Z11 with a light gray background color
    for range_name in range_list[2]:
        for row in sheet[range_name]:
            for cell in row:
                cell.font = Font(italic=True, size = 10)
    
    print("Apply Shading ...")
    # Shade the range B11:Z11 with a light gray background color
    for range_name in range_list[2]+range_list[3]+range_list[4]:
        for row in sheet[range_name]:
            for cell in row:
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    condtional_formating(sheet,range_list)

    print("Apply Column Length ...")
    total = int(range_list[0][0][-2:])
    # Auto-adjust the width of column B based on the content in B2 to B6
    max_length = max(len(str(sheet.cell(row=i, column=2).value)) for i in range(2, total+1))
    sheet.column_dimensions['B'].width = max_length + 2  # Add a little extra space

    # Auto-size columns after applying formatting
    for col_letter in string.ascii_uppercase[2:22]:  # "C" to "V"
        sheet.column_dimensions[col_letter].auto_size = True 

    # # Save changes
    # wb.save(excel_file)
    return

def apply_border(sheet,range_list):
    print("Apply Border ...")
    for range_name in range_list[0]+range_list[2]+range_list[3]+range_list[4]:
        for row in sheet[range_name]:
            for cell in row:
                # Apply border to all sides of the cell
                cell.border = Border(left=Side(border_style='thin'),
                                    right=Side(border_style='thin'),
                                    top=Side(border_style='thin'),
                                    bottom=Side(border_style='thin'))

                # Align cell content to the middle
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    return

def apply_title(sheet,range_list):

    print("Apply Title ...")
    # for range_name in range_list[3]:
    #     sheet.merge_cells(range_name)
    #     part = range_name.split(":")
    #     cell = sheet[part[0]]
    #     cell.value = "Site Information Summary"
    #     cell.font = Font(bold=True)

    sheet.merge_cells(range_list[3][0])
    part = range_list[3][0].split(":")
    cell = sheet[part[0]]
    cell.value = "Train Jounery Settings"
    cell.font = Font(bold=True)

    sheet.merge_cells(range_list[4][0])
    part = range_list[4][0].split(":")
    cell = sheet[part[0]]
    cell.value = "Total Time Spent on Each Mode (s)"
    cell.font = Font(bold=True)

    sheet.merge_cells(range_list[4][1])
    part = range_list[4][1].split(":")
    cell = sheet[part[0]]
    cell.value = "Train Battery Unit Energy Result"
    cell.font = Font(bold=True)

    sheet.merge_cells(range_list[4][2])
    part = range_list[4][2].split(":")
    cell = sheet[part[0]]
    cell.value = "Train Battery Unit SoC Result"
    cell.font = Font(bold=True)

    return

def condtional_formating(sheet,range_list):
    print("Apply Conditional Formatting ...")
    # Compare values in columns I and J for each row and shade accordingly
    # set the pattern fill
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red (0% S)
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 255,152,51 (80% S)
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green (00% S)

    # power
    #yellow_line = str(float(sheet[range_list[6][0]])*0.9)
    # sheet.conditional_formatting.add(range_list[6][1],CellIsRule(operator = 'lessThan',formula=[range_list[6][0]+'*0.9'],fill=green_fill))
    # sheet.conditional_formatting.add(range_list[6][1],CellIsRule(operator = 'greaterThanOrEqual',formula=[range_list[6][0]],fill=red_fill))
    # sheet.conditional_formatting.add(range_list[6][1],CellIsRule(operator = 'between',formula=[range_list[6][0]+'*0.9',range_list[6][0]],fill=yellow_fill))
    # SoC
    sheet.conditional_formatting.add(range_list[8][2],CellIsRule(operator = 'lessThanOrEqual',formula=['0'],fill=red_fill))
    sheet.conditional_formatting.add(range_list[8][2],CellIsRule(operator = 'greaterThanOrEqual',formula=[range_list[8][0]],fill=yellow_fill))
    sheet.conditional_formatting.add(range_list[8][2],CellIsRule(operator = 'lessThan',formula=[range_list[8][1]],fill=yellow_fill))
    sheet.conditional_formatting.add(range_list[8][2],CellIsRule(operator = 'between',formula=[range_list[8][0],range_list[8][1]],fill=green_fill))
    # SoC2
    sheet.conditional_formatting.add(range_list[9][2],CellIsRule(operator = 'lessThanOrEqual',formula=['0'],fill=red_fill))
    sheet.conditional_formatting.add(range_list[9][2],CellIsRule(operator = 'greaterThanOrEqual',formula=[range_list[9][0]],fill=yellow_fill))
    sheet.conditional_formatting.add(range_list[9][2],CellIsRule(operator = 'lessThan',formula=[range_list[9][1]],fill=yellow_fill))
    sheet.conditional_formatting.add(range_list[9][2],CellIsRule(operator = 'between',formula=[range_list[9][0],range_list[9][1]],fill=green_fill)) 

    # add energy
    sheet.conditional_formatting.add(range_list[10][0],CellIsRule(operator = 'greaterThan',formula=['0'],fill=red_fill))

    return


if __name__ == "__main__":
    # Add your debugging code here
    simname = "test_model_6"  # Provide a simulation name or adjust as needed
    main_option = "1"  # Adjust as needed
    time_start = "0070000"  # Adjust as needed
    time_end = "0080000"  # Adjust as needed
    option_select = "5"  # Adjust as needed
    text_input = "AC_07_BEMU3"  # Adjust as needed
    low_v = None  # Adjust as needed
    high_v = None  # Adjust as needed
    time_step = None  # Adjust as needed

    # Call your main function with the provided parameters
    main(simname, main_option, time_start, time_end, option_select, text_input, low_v, high_v, time_step)

