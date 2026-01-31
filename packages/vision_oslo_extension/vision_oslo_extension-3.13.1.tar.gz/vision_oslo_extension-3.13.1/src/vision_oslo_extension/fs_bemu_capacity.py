#
# -*- coding: utf-8  -*-
#=================================================================
# Created by: Jieming Ye
# Created on: Feb 2024
# Last Modified: Feb 2024
#=================================================================
# Copyright (c) 2024 [Jieming Ye]
#
# This Python source code is licensed under the 
# Open Source Non-Commercial License (OSNCL) v1.0
# See LICENSE for details.
#=================================================================
"""
Pre-requisite: 
SimulationName.oof: default oslo output file after successfully run a simulation.
xxxxx.xlsx: excel spreadsheet with proper user settings.
SimulationName.lst.txt: required for option 4
Used Input:
simname: to locate the result file or file rename
time_start: to define the analysis start time
time_end: to define the analysis end time
option_select: to define the option selected in subfunction
text_input: to locate the excel file name
other input for oslo_extraction.py only.
Expected Output:
Updated excel spreadsheet with analysis result.
Description:
This script defines the process of doing average load assessment.
First, it reads the excel file and save the user input in a data frame called start_df. Then it find the related .d4 files and read information one by one and saved in a data frame list called d4dataframe. The calculation result is saved in a similar format data frame list called sumdataframe. It then doing some analysis and save the final result in an updated data frame. Final step is to save all information in the excel in a proper format. The whole process is easy to follow via reading the code.
The key manual updating part is get_result_range() function which depends on the desired output format in excel, followed by table_formatting() function where some reading needs manual input. The other process should be auto adjustable as long as the excel input format is followed.

"""
#=================================================================
# VERSION CONTROL
# V1.0 (Jieming Ye) - Initial Version
# V1.1 (Jieming Ye) - Logic change for average calculation and counter for negative values
#=================================================================
# Set Information Variable
# N/A
#=================================================================

# vision_oslo_extension/excel_processing.py
import pandas as pd
import numpy as np
# import openpyxl
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, NamedStyle
from openpyxl.formatting.rule import CellIsRule,FormulaRule
from openpyxl.chart import ScatterChart,Reference,Series
from openpyxl.drawing.text import CharacterProperties
from openpyxl.chart.shapes import GraphicalProperties


# import vision_oslo
from vision_oslo_extension import oslo_extraction
from vision_oslo_extension.shared_contents import SharedMethods


def main(simname, main_option, time_start, time_end, option_select, text_input, low_v, high_v, time_step):
    
    print("")
    print("Supply Point Preliminary BEMU Support Assessment - - - > ")
    print("")
    
    simname = simname
    # Specify Excel file name
    excel_file = text_input + ".xlsx"

    #option = "1" # Full Auto
    #option = "2" # Update Spreadsheet only

    option = option_select # 1:Fully Restart, require oof, and list

    time_increment = 5
    start = 10

    space  = 5
    
    if option == "1":
        if not SharedMethods.check_oofresult_file(simname):
            return False

    # start_cell = 'B11'
    # read data from start tab
    # start_df = pd.read_excel(writer,sheet_name = "Start")
    result = start_reading_process(
        simname, time_start, time_end, text_input, low_v, high_v, time_step,excel_file,option
    )
    if result == False:
        return False
    else:
        supply_list, start_df = result

    # number of dataframes to create
    total = len(supply_list)

    # check if want to go throught the feeder check process
    if option == "1":

        if not one_stop_AC_extraction(simname, time_start, time_end,supply_list):
            return False

        if option == "4":
            return True
        
    if option == "2":
        print("Checking essential d4 files files...")
        for supply in supply_list:
            filename = simname + "_" + supply + ".osop.d4"
            if not SharedMethods.check_existing_file(filename):
                SharedMethods.print_message(f"ERROR: d4 file {supply} do not exist. Select option 1 to proceed.","31")
                return False
    
    # check if the d4 file is empty or not
    for feeder in supply_list:
        filename = simname + "_" + feeder +".osop.d4"
        if not SharedMethods.validate_extracted_result(filename,force_data = True):
            SharedMethods.print_message(f"ERROR: Empty d4 file from {feeder} is not allowed in the function. Check your settings.")
            return False
        
    # process d4 file
    d4dataframe = feeder_reading_process(simname, supply_list,time_increment)

    data_write_save(simname, excel_file,start,total,space,start_df,d4dataframe,supply_list,time_increment)

    return True

# extraction
def one_stop_AC_extraction(simname, time_start, time_end,supply_list):

    # User defined time windows extraction
    time_start = SharedMethods.time_input_process(time_start,1)
    time_end = SharedMethods.time_input_process(time_end,1)

    if time_start == False or time_end == False:
        return False

    # process 1: feeder step output
    print("Extrating feeder step output from Start Tab ... ")

    # processing the list
    for items in supply_list:
        #print(items)
        oslo_extraction.feeder_step_one(simname,items,time_start,time_end)

    # process 2: minmax value extraction
    print("Extrating min-max output...")
    
    # define the default osop command
        
    opcname = simname + ".opc"

    with open(opcname,"w") as fopc:
        fopc.writelines("MINMAX VALUES REQUIRED\n")

    if not SharedMethods.osop_running(simname):
        return False

    return True

# write data to excel
def data_write_save(simname, excel_file,start,total,space,start_df,d4dataframe,supply_list,time_increment):
    # write data to excel
    with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a',if_sheet_exists='overlay') as writer:

        # get currrent workbook to do process
        wb = writer.book
        data_row = 25

        print("Generate Result Page...")
        # empty_df = pd.DataFrame()
        # empty_df.to_excel(writer, sheet_name="Result", index=False)
        
        # Write each DataFrame to a different sheet in the Excel file
        print("Writing individual feeder...")
        for index, dflist in enumerate(d4dataframe):
            print(f"Saving {supply_list[index]}...")
            # emty columns
            new_columns = ['New_C_1', 'New_C_2', 'New_C_3', 'Excel_Time', 'Inst Current Headroom (A)', \
                           'Instat Max Train', 'New_C_4', 'MISL Headroom (MVA)', 'MISL Current (A)', \
                            'Charging Headroom (A)', 'Charging per OLE (A)', '30 min Average (A)', \
                                '30 min Current Headroom for charing (A)', '30 min Equvilent Power (MVA)', \
                                    '30 min Max Train', 'Max Train']
            
            new_df = pd.DataFrame(np.nan, index=dflist.index, columns=new_columns)
            # Concatenate the new columns to the original DataFrame
            dflist = pd.concat([dflist, new_df], axis=1)

            dflist.to_excel(writer, sheet_name=supply_list[index], index=False, startrow = data_row)
        

        for index, supply in enumerate(supply_list):
            print(f"Assessment for {supply}...")
            sheet = wb[supply]
            row = start_df.loc[start_df['OSLO ID'] == supply]

            apply_default_value(sheet,row)
            apply_excel_formula(sheet,len(d4dataframe[index]),data_row+2,time_increment)
            train_plot(sheet,row,len(d4dataframe[index]),data_row+2)
            train_plot_cont_only(sheet,row,len(d4dataframe[index]),data_row+2)

            # Calculate the Excel range for each DataFrame
            range_list = get_result_range(data_row)
            # table formatting
            table_formatting(wb,sheet,range_list)


        print("Saving Data...")

    return

# adding default value to each sheet from start setup
def apply_default_value(sheet,row):

    sheet["B2"].value = "Source Impedance"
    sheet["B3"].value = "Expected Usage"
    sheet["B4"].value = "MISL (MVA)"
    sheet["B5"].value = "MISL Limit (MVA)"
    sheet["B6"].value = "IF O/C (A)"
    sheet["B7"].value = "IF O/C Limit(A)"
    sheet["B8"].value = "OLE Section"
    sheet["B9"].value = "BEMU Limit (A)"
    sheet["B10"].value = "BEMU Limit (MW)"
    sheet["B11"].value = "OLE Rating (A)"
    
    sheet["C2"].value = f'=SQRT(D2^2+E2^2)'
    sheet["D2"].value = row["Source Resistance (Ω)"].iloc[0]
    sheet["E2"].value = row["Source Reactance (Ω)"].iloc[0]

    sheet["C3"].value = row["Expected Site Usage (%)"].iloc[0]
    sheet["C4"].value = row["MISL (MVA)"].iloc[0]
    sheet["C5"].value = f'=C4*C3'
    sheet["C6"].value = row["Incoming Feeder Overcurrent (A)"].iloc[0]
    sheet["C7"].value = f'=C6*C3'
    sheet["C8"].value = row["OLE Paralleled Section"].iloc[0]
    sheet["C9"].value = row["Instant Charging Limit (A)"].iloc[0]
    sheet["C10"].value = row["Continouse Charging Limit (MW)"].iloc[0]
    sheet["C11"].value = row["OLE Continous Rating (A)"].iloc[0]

    sheet["A17"].value = "NOTE: Calculation from Column W should is trackable from the formula"
    sheet["A18"].value = "NOTE: You can adjust the settings above to manipulate the plot"
    sheet["A19"].value = "Support for Column 'AC' Charging Headroom calculation as follows:"
    sheet["A20"].value = "Considering the source energy consumption --> VI + I^2 R = MISL --> RI^2+VI-MISL = 0"
    sheet["A21"].value = "I could be sorted using standard 2nd order equation solution"
    sheet["A22"].value = "I = [-V + sqrt(V^2 - 4 R MISL)] / 2 R"

    return

# adding formula to BEMU assessment (adjusting require testing)
def apply_excel_formula(sheet,total,start,time_increment):

    index = start
    data_range = f"W{start}:W{start+total-1}"
    for line in sheet[data_range]:
        for cell in line:
            cell.value = f'=TIME(LEFT(C{index},2),MID(C{index},4,2),RIGHT(C{index},2))'
            cell.number_format = 'hh:mm:ss'
        index = index + 1
    
    # instant current headroom
    index = start
    data_range = f"X{start}:X{start+total-1}"
    for line in sheet[data_range]:
        for cell in line:
            cell.value = f'=$C$7-H{index}'
        index = index + 1
    
    # instant max train
    index = start
    data_range = f"Y{start}:Y{start+total-1}"
    for line in sheet[data_range]:
        for cell in line:
            cell.value = f'=ROUNDDOWN(X{index}/$C$9,0)'
        index = index + 1
    
    # MISL headroom
    index = start
    data_range = f"AA{start}:AA{start+total-1}"
    for line in sheet[data_range]:
        for cell in line:
            cell.value = f'=$C$5-P{index}'
        index = index + 1
    
    # MISL headroom
    index = start
    data_range = f"AB{start}:AB{start+total-1}"
    for line in sheet[data_range]:
        for cell in line:
            cell.value = f'=AA{index}*1000/F{index}'
        index = index + 1
    
    # charging headroom
    index = start
    data_range = f"AC{start}:AC{start+total-1}"
    for line in sheet[data_range]:
        for cell in line:
            cell.value = f'=(-F{index}+SQRT(F{index}^2+4*$C$2*$C$5))/(2*$C$2)*1000'
        index = index + 1
    
    # charging per OLE
    index = start
    data_range = f"AD{start}:AD{start+total-1}"
    for line in sheet[data_range]:
        for cell in line:
            cell.value = f'=AC{index}/$C$8'
        index = index + 1
    
    # 30 min average current
    index = start
    data_range = f"AE{start}:AE{start+total-1}"
    for line in sheet[data_range]:
        # update logic so that for time below 30 min, doing average from the start anyway
        target_index = int(start+(1800/time_increment)-1)
        if index <= target_index:
            temp = start
        else:
            temp = int(index - (1800/time_increment) + 1)
        
        for cell in line:
            cell.value = f'=AVERAGE(AD{temp}:AD{index})'
        index = index + 1
    
    # 30 min charging
    index = start
    data_range = f"AF{start}:AF{start+total-1}"
    for line in sheet[data_range]:
        for cell in line:
            cell.value = f'=IF(AE{index}>$C$11,$C$11*$C$8,AC{index})'
        index = index + 1
    
    # 30 min charging
    index = start
    data_range = f"AG{start}:AG{start+total-1}"
    for line in sheet[data_range]:
        for cell in line:
            cell.value = f'=AF{index}*F{index}/1000'
        index = index + 1
    
    # 30 min max train
    index = start
    data_range = f"AH{start}:AH{start+total-1}"
    for line in sheet[data_range]:
        for cell in line:
            cell.value = f'=ROUNDDOWN(AG{index}/$C$10,0)'
        index = index + 1

    # 30 min max train
    index = start
    data_range = f"AI{start}:AI{start+total-1}"
    for line in sheet[data_range]:
        for cell in line:
            cell.value = f'=MAX(MIN(AH{index},Y{index}),0)' # maximum of 0 and min of instant and 30 min max values
        index = index + 1
    

    return

# plot the 30min average power:
def train_plot(sheet,row,total,start):
    name = row["Supply Point Name"].iloc[0]

    time = sheet[f'C{start}'].value
    time = time.split(":")
    ts = (float(time[0])+float(time[1])/60+float(time[1])/3600)/24

    time = sheet[f'C{start+total-1}'].value
    time = time.split(":")
    te = (float(time[0])+float(time[1])/60+float(time[1])/3600)/24

    xlim = 0
    ylim = 1
    
    # find nearest 10 min timeline.
    for i in range(0,144):
        if i*1/144 > ts:
            xlim = (i-1)*1/144
            for j in range(i,144):
                if j*1/144 >= te:
                    ylim = j*1/144
                    break
            break
    
    tspace = (ylim-xlim)/8

    # Create a scatter plot
    # chart = writer.book.add_chart({'type': 'scatter'})
    chart = ScatterChart()
    chart.title = f"Max Number of BEMU Chargable within {name}"
    chart.x_axis.title = "Time (hh:mm:ss)"
    chart.y_axis.title = "No. of Trains"

    #time_format = NamedStyle(name='time_format', number_format='HH:MM:SS')
    # Set chart size
    chart.height = 10  # Adjust width as needed
    chart.width = 25  # Adjust height as needed

    xvalues = Reference(sheet, min_col=23, min_row=start, max_row=start+total-1)
    yvalues = Reference(sheet, min_col=35, min_row=start, max_row=start+total-1)
    series = Series(yvalues,xvalues,title_from_data=False)
    chart.series.append(series)
    # Remove legend
    chart.legend = None
    # chart.legend.position = 't' # top legend position
    cp = CharacterProperties(sz=1400)
    chart.title.tx.rich.p[0].r[0].rPr = cp 
    # chart.title.tx.rich.paragraphs = cp 

    # Access major gridlines and set color to light grey
    cl = GraphicalProperties()
    cl.line.solidFill = "E0E0E0"
    chart.x_axis.majorGridlines.spPr = cl  # Light grey color
    chart.y_axis.majorGridlines.spPr = cl  #  # Light grey color
    chart.x_axis.scaling.min = xlim      
    chart.x_axis.scaling.max = ylim
    chart.x_axis.majorUnit = tspace

    sheet.add_chart(chart,"J2")


    return


# plot the 30min average power:
def train_plot_cont_only(sheet,row,total,start):
    name = row["Supply Point Name"].iloc[0]

    time = sheet[f'C{start}'].value
    time = time.split(":")
    ts = (float(time[0])+float(time[1])/60+float(time[1])/3600)/24

    time = sheet[f'C{start+total-1}'].value
    time = time.split(":")
    te = (float(time[0])+float(time[1])/60+float(time[1])/3600)/24

    xlim = 0
    ylim = 1
    
    # find nearest 10 min timeline.
    for i in range(0,144):
        if i*1/144 > ts:
            xlim = (i-1)*1/144
            for j in range(i,144):
                if j*1/144 >= te:
                    ylim = j*1/144
                    break
            break
    
    tspace = (ylim-xlim)/8

    # Create a scatter plot
    # chart = writer.book.add_chart({'type': 'scatter'})
    chart = ScatterChart()
    chart.title = f"Max Number of BEMU Chargable for {name} (Continous Loading Only)"
    chart.x_axis.title = "Time (hh:mm:ss)"
    chart.y_axis.title = "No. of Trains"

    #time_format = NamedStyle(name='time_format', number_format='HH:MM:SS')
    # Set chart size
    chart.height = 10  # Adjust width as needed
    chart.width = 25  # Adjust height as needed

    xvalues = Reference(sheet, min_col=23, min_row=start, max_row=start+total-1)
    yvalues = Reference(sheet, min_col=34, min_row=start, max_row=start+total-1)
    series = Series(yvalues,xvalues,title_from_data=False)
    chart.series.append(series)
    # Remove legend
    chart.legend = None
    # chart.legend.position = 't' # top legend position
    cp = CharacterProperties(sz=1400)
    chart.title.tx.rich.p[0].r[0].rPr = cp 
    # chart.title.tx.rich.paragraphs = cp 

    # Access major gridlines and set color to light grey
    cl = GraphicalProperties()
    cl.line.solidFill = "E0E0E0"
    chart.x_axis.majorGridlines.spPr = cl  # Light grey color
    chart.y_axis.majorGridlines.spPr = cl  #  # Light grey color
    chart.x_axis.scaling.min = xlim      
    chart.x_axis.scaling.max = ylim
    chart.x_axis.majorUnit = tspace

    sheet.add_chart(chart,"Y2")
    return

# read the start tab and collect informaiton
def start_reading_process(simname, time_start, time_end, text_input, low_v, high_v, time_step, excel_file,option):
    supply_list = []
    node_list = []
    name_list = []
    
    # File check
    print("Check Excel and Deleting Existing Contents ...")

    try:
        wb = load_workbook(excel_file)
        # Delete all sheets except 'Start'
        for sheet_name in wb.sheetnames:
            if sheet_name == 'Start':
                table_list = check_table_list(wb[sheet_name])
                
                if table_list is False:
                    return False
            else:
                wb.remove(wb[sheet_name])
        
        # Save changes
        wb.save(excel_file)

    except Exception as e:
        SharedMethods.print_message(f"(Close the Excel file and Start Again) Error: {e}","31")
        return False
        
    # read start up information 
    columns = ["Supply Point Name","OSLO ID","Source Resistance (Ω)","Source Reactance (Ω)","OLE Paralleled Section","Fault Level (MVA)", \
               "Incoming Feeder Overcurrent (A)","MISL (MVA)","Instant Charging Limit (A)","Continouse Charging Limit (MW)","Expected Site Usage (%)", \
                "OLE Continous Rating (A)"]
    start_df = pd.DataFrame(table_list,columns=columns)
    # get supply list oslo id
    for items in table_list:
        supply_list.append(items[1])
    
    # check duplication in OSLO id
    if not SharedMethods.find_duplicates(supply_list): return False
    
    return supply_list, start_df

# check table list on start page
def check_table_list(sheet):
    print("Reading Configuration Setting ...")
    table_row = 11
    table_start_column = 1
    table_end_column = 12
    # create node and feeder list
    table_list = []

    # check feeder
    index = table_row
    column = table_start_column
    if sheet.cell(row=index, column=column+1).value is not None:
        while True:
            row_data = []
            for temp in range(table_start_column,table_end_column + 1):
                row_data.append(sheet.cell(row=index, column=temp).value)
            table_list.append(row_data)
            index += 1
            
            check = sheet.cell(row=index, column=column+1).value
            if check is None:
                break
    else:
        SharedMethods.print_message("ERROR: Wrong data format. No information at B11.","31")
        return False

    return table_list

# read individual d4 file
def feeder_reading_process(simname, supply_list,time_increment):
    
    # create dataframe
    d4dataframe = []

    for feeder in supply_list:
        print(f"Processing {feeder} ...")
        filename = simname + "_" + feeder +".osop.d4"
        delimiter = '\\s+'
        columns = ["SupplyID","Type","Time","P_inst","Q_inst","Voltage","V_angle","Current","I_angle"]
        
        dtype_mapping = {"Time": str,}

        df = pd.read_csv(filename, delimiter=delimiter, names = columns, skiprows = 11,dtype = dtype_mapping)  # You might need to adjust the delimiter based on your file
        #df["Time"] = pd.to_datetime(df["Time"],format='%H:%M:%S')
        # Extracting parts from the string and formatting the 'Time' column
        df['Time'] = df['Time'].apply(lambda x: f"{int(x[1:3]):02d}:{int(x[3:5]):02d}:{int(x[5:]):02d}")

        # data frame process
        df = d4_file_process(df,time_increment)

        d4dataframe.append(df)

    return d4dataframe

# doing invididual d4 file process and calculatoin
def d4_file_process(df, time_increment):

    df['S_inst'] = np.sqrt(df['P_inst']**2 + df['Q_inst']**2)

    # window_sizes = {'1min': 60, '2min': 120, '10min': 600, '20min': 1200, '30min': 1800}
    window_sizes = {'20min': 1200, '30min': 1800}

    for time_interval, window_size in window_sizes.items():
        df[f'P_{time_interval}'] = df['P_inst'].rolling(window=int(window_size / time_increment)).mean()
        df[f'Q_{time_interval}'] = df['Q_inst'].rolling(window=int(window_size / time_increment)).mean()
        df[f'S_{time_interval}'] = np.sqrt(df[f'P_{time_interval}']**2 + df[f'Q_{time_interval}']**2)

    df['I_Inst'] = df['Current'] # Current Duplicate

    # for time_interval in [5, 10, 15, 20, 25, 30, 35, 40, 60, 120, 600, 1200, 1800]:
    for time_interval in [1200, 1800]:
        column_name = f'I_{time_interval}s_RMS'
        df[column_name] = calculate_rolling_rms(df, 'Current', time_interval, time_increment)

    # df['P_CosPhi'] = df['P_inst'] / df['S_inst']
    
    # for time_interval in ['inst', '1min', '2min', '10min', '20min', '30min']:
    #     df[f'PF_{time_interval}'] = df[f'P_{time_interval}'] / df[f'S_{time_interval}']

    return df

# calculate rms
def calculate_rolling_rms(data, column, window_size, time_increment):
    return data[column].rolling(window=int(window_size / time_increment)).apply(lambda x: np.sqrt(np.mean(x**2)), raw=True)

def get_result_range(data_row):

    range_list = []
    # 0
    title_range = [f"A{data_row+1}:AI{data_row+1}"]
    range_list.append(title_range)
    # 1
    perc_range = [f"C3:C3"]
    range_list.append(perc_range)

    return range_list

# Result table table formating
def table_formatting(wb,sheet,range_list):
    # format the result table

    print("Formatting Process ...")
    # # wb = load_workbook(excel_file)
    # print("Information Collection ...")
    # sheet = wb["Start"]
    # project_name = sheet['B2'].value
    # feeding_desp = sheet['B4'].value
    # modeller = sheet['B5'].value
    # date = sheet['B6'].value

    # # Result Tab process
    # sheet = wb["Result"]
    # sheet['B2'].value = "Project Name:"
    # sheet['C2'].value = project_name
    # sheet['B3'].value = "Simulation Name:"
    # sheet['C3'].value = simname
    # sheet['B4'].value = "Feeding Arrangement:"
    # sheet['C4'].value = feeding_desp
    # sheet['B5'].value = "Result Created by:"
    # sheet['C5'].value = modeller
    # sheet['B6'].value = "Result Created at:"
    # sheet['C6'].value = datetime.now().strftime("%d-%m-%Y %H:%M")
    
    for range_name in range_list[0]:
        for row in sheet[range_name]:
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for range_name in range_list[1]:
        for row in sheet[range_name]:
            for cell in row:
                cell.number_format = '0.0%'

    print("Apply Column Length ...")
    # Auto-adjust the width of column B based on the content in B2 to B6
    max_length = max(len(str(sheet.cell(row=i, column=2).value)) for i in range(2, 20))
    sheet.column_dimensions['B'].width = max_length + 2  # Add a little extra space

    # Auto-size columns after applying formatting
    for col_letter in ["C","D","E","F","G","H","I","J","K","L"]:
        sheet.column_dimensions[col_letter].auto_size = True

    # # Save changes
    # wb.save(excel_file)
    return

def condtional_formating(sheet,range_list):
    print("Apply Conditional Formatting ...")
    # Compare values in columns I and J for each row and shade accordingly
    # set the pattern fill
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red (0% S)
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 255,152,51 (80% S)
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green (00% S)

    # power
    #yellow_line = str(float(sheet[range_list[6][0]])*0.9)
    sheet.conditional_formatting.add(range_list[6][1],CellIsRule(operator = 'lessThan',formula=[range_list[6][0]+'*0.9'],fill=green_fill))
    sheet.conditional_formatting.add(range_list[6][1],CellIsRule(operator = 'greaterThanOrEqual',formula=[range_list[6][0]],fill=red_fill))
    sheet.conditional_formatting.add(range_list[6][1],CellIsRule(operator = 'between',formula=[range_list[6][0]+'*0.9',range_list[6][0]],fill=yellow_fill))
    # 1 min NPS
    sheet.conditional_formatting.add(range_list[7][1],CellIsRule(operator = 'lessThan',formula=['0.018'],fill=green_fill))
    sheet.conditional_formatting.add(range_list[7][1],CellIsRule(operator = 'greaterThanOrEqual',formula=['0.02'],fill=red_fill))
    sheet.conditional_formatting.add(range_list[7][1],CellIsRule(operator = 'between',formula=['0.018','0.02'],fill=yellow_fill))
    # 10 min NPS
    sheet.conditional_formatting.add(range_list[7][2],CellIsRule(operator = 'lessThan',formula=['0.0135'],fill=green_fill))
    sheet.conditional_formatting.add(range_list[7][2],CellIsRule(operator = 'greaterThanOrEqual',formula=['0.015'],fill=red_fill))
    sheet.conditional_formatting.add(range_list[7][2],CellIsRule(operator = 'between',formula=['0.0135','0.015'],fill=yellow_fill))
    # 30 min NPS
    sheet.conditional_formatting.add(range_list[7][3],CellIsRule(operator = 'lessThan',formula=['0.009'],fill=green_fill))
    sheet.conditional_formatting.add(range_list[7][3],CellIsRule(operator = 'greaterThanOrEqual',formula=['0.01'],fill=red_fill))
    sheet.conditional_formatting.add(range_list[7][3],CellIsRule(operator = 'between',formula=['0.009','0.01'],fill=yellow_fill))
    # Min V
    sheet.conditional_formatting.add(range_list[8][0],CellIsRule(operator = 'greaterThan',formula=['22.5'],fill=green_fill))
    sheet.conditional_formatting.add(range_list[8][0],CellIsRule(operator = 'between',formula=['19','22.5'],fill=yellow_fill))
    sheet.conditional_formatting.add(range_list[8][0],CellIsRule(operator = 'lessThanOrEqual',formula=['19'],fill=red_fill))
    # Max V
    sheet.conditional_formatting.add(range_list[8][1],CellIsRule(operator = 'lessThan',formula=['27.5'],fill=green_fill))
    sheet.conditional_formatting.add(range_list[8][1],CellIsRule(operator = 'greaterThanOrEqual',formula=['27.5'],fill=red_fill))

    # power
    #yellow_line = str(float(sheet[range_list[10][0]])*0.9)
    sheet.conditional_formatting.add(range_list[10][1],CellIsRule(operator = 'lessThan',formula=[range_list[10][0]+'*0.9'],fill=green_fill))
    sheet.conditional_formatting.add(range_list[10][1],CellIsRule(operator = 'between',formula=[range_list[10][0]+'*0.9',range_list[10][0]],fill=yellow_fill))
    sheet.conditional_formatting.add(range_list[10][1],CellIsRule(operator = 'greaterThanOrEqual',formula=[range_list[10][0]],fill=red_fill))
    # 1 min NPS
    sheet.conditional_formatting.add(range_list[11][1],CellIsRule(operator = 'lessThan',formula=['0.018'],fill=green_fill))
    sheet.conditional_formatting.add(range_list[11][1],CellIsRule(operator = 'between',formula=['0.018','0.02'],fill=yellow_fill))
    sheet.conditional_formatting.add(range_list[11][1],CellIsRule(operator = 'greaterThanOrEqual',formula=['0.02'],fill=red_fill))
    # 10 min NPS
    sheet.conditional_formatting.add(range_list[11][2],CellIsRule(operator = 'lessThan',formula=['0.0135'],fill=green_fill))
    sheet.conditional_formatting.add(range_list[11][2],CellIsRule(operator = 'between',formula=['0.0135','0.015'],fill=yellow_fill))
    sheet.conditional_formatting.add(range_list[11][2],CellIsRule(operator = 'greaterThanOrEqual',formula=['0.015'],fill=red_fill))
    # 30 min NPS
    sheet.conditional_formatting.add(range_list[11][3],CellIsRule(operator = 'lessThan',formula=['0.009'],fill=green_fill))
    sheet.conditional_formatting.add(range_list[11][3],CellIsRule(operator = 'between',formula=['0.009','0.01'],fill=yellow_fill))
    sheet.conditional_formatting.add(range_list[11][3],CellIsRule(operator = 'greaterThanOrEqual',formula=['0.01'],fill=red_fill))

    return


if __name__ == "__main__":
    # Add your debugging code here
    simname = "test_model_1"  # Provide a simulation name or adjust as needed
    main_option = "1"  # Adjust as needed
    time_start = "0070000"  # Adjust as needed
    time_end = "0080000"  # Adjust as needed
    option_select = "2"  # Adjust as needed
    text_input = "AC_06_BEMU1"  # Adjust as needed
    low_v = None  # Adjust as needed
    high_v = None  # Adjust as needed
    time_step = None  # Adjust as needed

    # Call your main function with the provided parameters
    main(simname, main_option, time_start, time_end, option_select, text_input, low_v, high_v, time_step)

