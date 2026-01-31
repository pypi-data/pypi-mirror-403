#
# -*- coding: utf-8  -*-
#=================================================================
# Created by: Jieming Ye
# Created on: Feb 2024
# Last Modified: Feb 2024
#=================================================================
# Copyright (c) 2024 [Jieming Ye]
#
# This Python source code is licensed under the 
# Open Source Non-Commercial License (OSNCL) v1.0
# See LICENSE for details.
#=================================================================
"""
Pre-requisite: 
SimulationName.oof: default oslo output file after successfully run a simulation.
xxxxx.xlsx: excel spreadsheet with proper user settings.
Used Input:
simname: to locate the result file or file rename
time_start: to define the analysis start time
time_end: to define the analysis end time
option_select: to define the option selected in subfunction
text_input: to locate the excel file name
other input for oslo_extraction.py only.
Expected Output:
Updated excel spreadsheet with analysis result.
Description:
This script defines the process of doing OLE rating assessment.
First, it reads the excel file and save the user input in a data frame called start_df. Then it calculated the rating_df as a comparison data frame. Then it processing related  .d4 files and read information one by one and saved in a data frame list called d4dataframe. The calculation result is saved in a similar format data frame list called sumdataframe. It then doing some analysis and save the final result in an updated data frame. Final step is to save all information in the excel in a proper format. The whole process is easy to follow via reading the code.
If additional time window to be assessed in the future, several places needs to updated, specially column definition in each sub functions.
The key manual updating part is get_result_range() function which depends on the desired output format in excel, followed by table_formatting() function where some reading needs manual input. The other process should be auto adjustable as long as the excel input format is followed.

"""
#=================================================================
# VERSION CONTROL
# V1.0 (Jieming Ye) - Initial Version
# V1.1 (Jieming Ye) - Including pre-checking of OLE type
#=================================================================
# Set Information Variable
# N/A
#=================================================================
# vision_oslo_extension/excel_processing.py
import pandas as pd
import numpy as np
# import openpyxl
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, NamedStyle
from openpyxl.formatting.rule import CellIsRule,FormulaRule

# import vision_oslo
from vision_oslo_extension import oslo_extraction
from vision_oslo_extension.shared_contents import SharedMethods
from vision_oslo_extension.shared_excel_helper import export_result_sheet_from_wb


def main(simname, main_option, time_start, time_end, option_select, text_input, low_v, high_v, time_step):

    print("")
    print("OLE Thermal Rating Assessment ----->")
    print("")
    
    simname = simname
    # Specify Excel file name
    excel_file = text_input + ".xlsx"

    if not SharedMethods.check_existing_file(excel_file):
        return False
    
    #option = "1" # Fully Restart, require oof,
    #option = "2" # Process only

    option = option_select # 1:Fully Restart, require oof, and list

    start = 10 # result start from row 10
    space = 5
    time_increment = 5

    if option not in ["0","1","2"]:
        SharedMethods.print_message("ERROR: Error in ole_processing.py. Please contact Support...","31")
        return False
    
    if option == "0":
        SharedMethods.print_message("ERROR: Please select an option to proceed.","31")
        return False

    # start_cell = 'B11'
    # read data from start tab
    # start_df = pd.read_excel(writer,sheet_name = "Start")
    result = start_reading_process(simname,excel_file)
    if result == False:
        return False
    else:
        start_df,rating_df,ole_list = result

    # get essential info
    branch_list = start_df['OSLO ID'].tolist()
    
    # check duplication in OSLO id
    if not SharedMethods.find_duplicates(branch_list): return False

    type_list = start_df['OLE Type'].tolist()
    # check branch_list
    branch_list_new = check_branch_list(branch_list)
    if branch_list_new == False:
        return False        
    # branch_list: branch format in XXXX/X format
    # branch_list_new: branch format in XXXX-X format this is because many application does not allow / charaters in naming

    # check type information
    ole_flag = False
    for index, ole in enumerate(type_list):
        if ole not in ole_list:
            SharedMethods.print_message(f"ERROR: OLE {start_df.loc[index,'OSLO ID']} Type '{ole}' not defined in the list. Check Input Data.","31")
            ole_flag = True
    if ole_flag: return False        

    # check if want to go throught the feeder check process
    if option == "1":
        if not SharedMethods.check_oofresult_file(simname):
            return False
        # extract the feeder
        for branch_id in branch_list:
            if not oslo_extraction.branch_step(simname, time_start, time_end, "1", branch_id):
                return False
        
    if option == "2":
        print("Checking essential d4 files...")
        for branch in branch_list_new:
            filename = simname + "_" + branch + ".osop.d4"
            if not SharedMethods.check_existing_file(filename):
                SharedMethods.print_message(f"ERROR: d4 file {branch} do not exist. Select option 1 or check OSLO ID to proceed.","31")
                return False
    try:
        # process d4 file
        d4dataframe, sumdataframe, result_df = branch_reading_process(simname,branch_list_new,time_increment)

        # Update the result table for Rating
        result_df = result_rating_update(simname,result_df,start_df,rating_df)

        # save the data
        data_write_save(simname,excel_file,start,space,d4dataframe,sumdataframe,start_df,result_df,branch_list_new)
    
    except KeyError as e:
        SharedMethods.print_message(f"ERROR: Unexpected error occured: {e}. Possibly due to incompleted data.","31")
        return False
    
    except Exception as e:
        SharedMethods.print_message(f"ERROR: Unexpected error occured: {e}","31")
        return False

    return True


# write data to excel
def data_write_save(simname,excel_file,start,space,d4dataframe,sumdataframe,start_df,result_df,branch_list_new):

    # write data to excel
    with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a',if_sheet_exists='overlay') as writer:

        # get currrent workbook to do process
        wb = writer.book

        print("Generate Result Page...")
        start_df.to_excel(writer, sheet_name="Result", index=False, startrow = start, startcol = 1)
        result_df.to_excel(writer, sheet_name="Result", index=False, startrow = start, startcol = start_df.shape[1] + 1)

        # saving individual feeder
        for index, dflist in enumerate(d4dataframe):

            print(f"Saving {branch_list_new[index]}...")
            # emty columns
            dflist.insert(dflist.columns.get_loc('I_angle')+1,'New_C_1', np.nan)

            sumdataframe[index].insert(sumdataframe[index].columns.get_loc('I_angle')+1,'New_C_1', np.nan)

            sumdataframe[index].to_excel(writer, sheet_name=branch_list_new[index], index=False, startrow = 0)

            dflist.to_excel(writer, sheet_name=branch_list_new[index], index=False, startrow = sumdataframe[index].shape[0]+2)

        
        # # Calculate the Excel range for post-processing
        range_list = get_result_range(branch_list_new,start,space,wb['Result'])
        
        # # table formatting
        table_formatting(simname,wb,range_list)

        # Output result dashboard only
        print("Exporting Result Dashboard...")
        dashboard_name = "0_"+simname+"_ole_loading.xlsx"
        export_result_sheet_from_wb(wb,"Result",dashboard_name)

        print("Saving Excel File...")

    return

# check branch list
def check_branch_list(branch_list):
    newlst = []
    for branch_id in branch_list:
            if len(branch_id) <= 6:
                if branch_id[-2:-1] =="/":
                    branch = branch_id[:len(branch_id)-2].ljust(4)+branch_id[-2:]
                    branch = branch[:4]+"-"+ branch[-1:]
                    newlst.append(branch)
                else:
                    SharedMethods.print_message(f"ERROR: Branch OSLO ID {branch_id} does not contains proper symbol. Check Input","31")
                    return False
            else:
                SharedMethods.print_message(f"ERROR: Branch OSLO ID {branch_id} is NOT in correct format. Check Input'.format(branch_id)","31")       
                return False

    return newlst

# read the start tab and collect informaiton
def start_reading_process(simname, excel_file):
    
    # File check
    print("Check Excel and Deleting Existing Contents ...")

    try:
        wb = load_workbook(excel_file)
        # Delete all sheets except 'Start'
        for sheet_name in wb.sheetnames:
            if sheet_name == 'Start':
                result = check_rating_list(wb[sheet_name]) # list, dictionary type
                if result == False:
                    return False
                else:
                    ole_list, rating_data = result
                    
                table_list = check_table_list(wb[sheet_name]) # list
                if table_list == False:
                    return False
            else:
                wb.remove(wb[sheet_name])
        
        # Save changes
        wb.save(excel_file)
        
    except Exception as e:
        SharedMethods.print_message(f"ERROR: (Close the Excel file and Start Again) Error: {e}","31")
        return False
        
    # creat dataframe
    columns = ['Location','OLE Section','OSLO ID','OLE Type']
    start_df = pd.DataFrame(table_list,columns=columns)
    rating_df = pd.DataFrame(rating_data, index = ole_list)
    
    
    return start_df,rating_df,ole_list

# check rating data
def check_rating_list(sheet):
    print("Reading Ratings Info ...")
    table_row = 12
    table_start_column = 7
    table_end_column = 11
    # create ole list and rating data
    ole_list = []
    cont_list = []
    r_20min = []
    r_10min = []
    r_5min = []

    # check table
    index = table_row
    column = table_start_column
    if sheet.cell(row=index, column=column).value is not None:
        while True:
            ole_list.append(sheet.cell(row=index, column=column).value)

            data = sheet.cell(row=index, column=column+1).value
            if isinstance(data,(int,float)):
                cont_list.append(data)
            else:
                cont_list.append('-')
            
            data = sheet.cell(row=index, column=column+2).value
            if isinstance(data,(int,float)):
                r_20min.append(data)
            else:
                r_20min.append('-')

            data = sheet.cell(row=index, column=column+3).value
            if isinstance(data,(int,float)):
                r_10min.append(data)
            else:
                r_10min.append('-')
            
            data = sheet.cell(row=index, column=column+4).value
            if isinstance(data,(int,float)):
                r_5min.append(data)
            else:
                r_5min.append('-')

            index += 1
            check = sheet.cell(row=index, column=column).value
            if check is None:
                break
    else:
        SharedMethods.print_message("ERROR: Wrong data format. No information at G12","31")
        return False
    
    rating_data = {'Cont.':cont_list,'20min': r_20min,'10min': r_10min,'5min': r_5min}

    return ole_list,rating_data

# check start table
def check_table_list(sheet):
    print("Reading Assessment Settings ...")
    table_row = 12
    table_start_column = 1
    table_end_column = 4
    # create node and feeder list
    table_list = []

    # check feeder
    index = table_row
    column = table_start_column
    if sheet.cell(row=index, column=column+2).value is not None:
        while True:
            row_data = []
            for temp in range(table_start_column,table_end_column+1):
                row_data.append(sheet.cell(row=index, column=temp).value)
            table_list.append(row_data)

            index += 1
            check = sheet.cell(row=index, column=column+2).value
            if check is None:
                break
    else:
        SharedMethods.print_message("ERROR: Wrong data format. No information at C12","31")
        return False

    return table_list

#update the result with rating information
def result_rating_update(simname,result_df,start_df,rating_df):

    type_list = start_df['OLE Type'].tolist()

    for index, ole_type in enumerate(type_list):
        result_df.loc[index,'Cont.'] = rating_df.loc[ole_type,'Cont.']
        result_df.loc[index,'20min'] = rating_df.loc[ole_type,'20min']
        result_df.loc[index,'10min'] = rating_df.loc[ole_type,'10min']
        result_df.loc[index,'5min'] = rating_df.loc[ole_type,'5min']

    return result_df

# read individual d4 file
def branch_reading_process(simname,branch_list,time_increment):

    total = len(branch_list)

    # create dataframe
    d4dataframe = []
    sumdataframe = []

    # creat result dataframe
    columns = ['Cont.','20min','10min','5min','r_30min','r_max_t1','r_20min','r_max_t2','r_10min','r_max_t3','r_5min','r_max_t4']
    result_df = pd.DataFrame(columns = columns, index=range(total))

    for index, branch in enumerate(branch_list):
        print(f"Processing {branch} ...")
        filename = simname + "_" + branch +".osop.d4"
        delimiter = '\\s+'
        columns = ["BranchID","Type","Time","P_inst","Q_inst","Voltage","V_angle","Current","I_angle"]
        dtype_mapping = {"Time": str,}
        df = pd.read_csv(filename, delimiter=delimiter, names = columns, skiprows = 11,dtype = dtype_mapping) 
        # Extracting parts from the string and formatting the 'Time' column
        df['Time'] = df['Time'].apply(lambda x: f"{int(x[1:3]):02d}:{int(x[3:5]):02d}:{int(x[5:]):02d}")
        # data frame process
        df = d4_file_br_process(df,time_increment)
        # calculate sum value
        df_sum = d4_find_max(df)
        #update summary result
        result_df = result_update(result_df, df_sum, index)

        # save the data
        sumdataframe.append(df_sum)
        d4dataframe.append(df)

    return d4dataframe, sumdataframe, result_df

# doing invididual d4 file process and calculatoin for branch step output
def d4_file_br_process(df, time_increment):

    window_sizes = {'30min': 1800,'20min': 1200, '10min': 600, '5min': 300}

    for time_interval, window_size in window_sizes.items():
        df[f'I_{time_interval}'] = calculate_rolling_rms(df, 'Current', window_size, time_increment)

    return df

# calculate rms
def calculate_rolling_rms(data, column, window_size, time_increment):
    return data[column].rolling(window=int(window_size / time_increment)).apply(lambda x: np.sqrt(np.mean(x**2)), raw=True)

# find maximum value and time of each d4 file
def d4_find_max(df):
    # # Insert two empty rows above the first row
    # df = pd.concat([pd.DataFrame(index=range(2)), df], ignore_index=True)
    sum_df = pd.DataFrame(columns=df.columns,index=range(4))
    sum_df.iloc[0, 0] = "Maximum Value"
    sum_df.iloc[1, 0] = "Maximum Value at Time"
    sum_df.iloc[2, 0] = "Minimum Value"
    sum_df.iloc[3, 0] = "Minimum Value at Time"

    start = 9

    if df.empty:
        sum_df.iloc[0, 1] = "DATA FOR THIS FEEDER IS NOT AVAILABLE"

    else:
        for column in df.columns[start:]:
            if df[column].dtype in [int, float]:
                max_value = df[column].max()
                time_of_max = df.loc[df[column].idxmax(), 'Time']
                sum_df.at[0, column] = max_value
                sum_df.at[1, column] = time_of_max
                min_value = df[column].min()
                time_of_min = df.loc[df[column].idxmin(), 'Time']
                sum_df.at[2, column] = min_value
                sum_df.at[3, column] = time_of_min

    return sum_df

# update result table based on output of d4 process
def result_update(target,df_sum,index):
    
    target.loc[index,'r_30min'] = df_sum.at[0,'I_30min']
    target.loc[index,'r_max_t1'] = df_sum.at[1,'I_30min']

    target.loc[index,'r_20min'] = df_sum.at[0,'I_20min']
    target.loc[index,'r_max_t2'] = df_sum.at[1,'I_20min']

    target.loc[index,'r_10min'] = df_sum.at[0,'I_10min']
    target.loc[index,'r_max_t3'] = df_sum.at[1,'I_10min']

    target.loc[index,'r_5min'] = df_sum.at[0,'I_5min']
    target.loc[index,'r_max_t4'] = df_sum.at[1,'I_5min']
    
    return target

# check the range
def get_result_range(branch_list_new,start,space,sheet):

    range_list = []
    total = len(branch_list_new)
    # 0: table frame range
    # 1: table data range (threshold, result)
    # 2: title range 
    # 3: title range (1st row info + threshold + result)
    # 4: Conditional formatting range (Cont,30min)

    # 5: Loop one by one to decide (threshold,result)

    
    # 0
    result_range = [f"B{start + 1}:Q{start + total + 1}"]
    range_list.append(result_range)
    # 1
    data_range = [f"F{start + 2}:I{start + total + 1}",f"J{start + 2}:Q{start + total + 1}"]
    range_list.append(data_range)
   
    # 2
    title_range = [f"B{start + 1}:Q{start + 1}"]
    range_list.append(title_range)
    
    #3 4
    add_range_1 = [f"B{start}:E{start}",f"F{start}:I{start}",f"J{start}:Q{start}"]
    range_list.append(add_range_1)

    add_range_2 = [f"F{start+2}",f"J{start+2}:J{start+total+1}"]
    range_list.append(add_range_2)

    # from 5 to the end
    for index in range(total):
        if sheet[f'G{12+index}'].value == '-':
            add = [f'F{12+index}',f'L{12+index}']
        else:
            add = [f'G{12+index}',f'L{12+index}']
        range_list.append(add)

        if sheet[f'H{12+index}'].value == '-':
            pass
        else:
            add = [f'H{12+index}',f'N{12+index}']
        
        range_list.append(add)
        
        if sheet[f'I{12+index}'].value == '-':
            pass
        else:
            add = [f'I{12+index}',f'P{12+index}']

        range_list.append(add)

    return range_list

# Result table table formating
def table_formatting(simname,wb,range_list):
    # format the result table

    print("Formatting Process ...")
    # wb = load_workbook(excel_file)
    print("Information Collection ...")
    sheet = wb["Start"]
    project_name = sheet['B2'].value
    feeding_desp = sheet['B4'].value
    modeller = sheet['B5'].value
    date = sheet['B6'].value

    # Result Tab process
    sheet = wb["Result"]
    sheet['B2'].value = "Project Name:"
    sheet['C2'].value = project_name
    sheet['B3'].value = "Simulation Name:"
    sheet['C3'].value = simname
    sheet['B4'].value = "Feeding Arrangement:"
    sheet['C4'].value = feeding_desp
    sheet['B5'].value = "Result Created by:"
    sheet['C5'].value = modeller
    sheet['B6'].value = "Result Created at:"
    sheet['C6'].value = datetime.now().strftime("%d-%m-%Y %H:%M")

    sheet['J11'].value = "30min"
    sheet['K11'].value = "time at"
    sheet['L11'].value = "20min"
    sheet['M11'].value = "time at"
    sheet['N11'].value = "10min"
    sheet['O11'].value = "time at"
    sheet['P11'].value = "5min"
    sheet['Q11'].value = "time at"

    # sheet['H4'].value = "Note:"
    # sheet['H5'].value = 
    # sheet['H6'].value = 


    print("Apply Border and Alignment...")
    # Define a custom style for formatting with two decimal places
    for range_name in range_list[0]+range_list[1]+range_list[2]+range_list[3]:
        for row in sheet[range_name]:
            for cell in row:
                # Apply border to all sides of the cell
                cell.border = Border(left=Side(border_style='thin'),
                                    right=Side(border_style='thin'),
                                    top=Side(border_style='thin'),
                                    bottom=Side(border_style='thin'))
                
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    apply_title(sheet,range_list)
    
    print("Apply Numbering Format ...")
    # Define a custom style for formatting with two decimal places
    for range_name in range_list[1]:
        for row in sheet[range_name]:
            for cell in row:
                cell.number_format = '0.00'


    print("Apply Font and Shading...")
    # Shade the range B11:Z11 with a light gray background color
    for range_name in range_list[2]+range_list[3]:
        for row in sheet[range_name]:
            for cell in row:
                cell.font = Font(bold = True, italic = True, size = 12)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    

    condtional_formating(sheet,range_list)

    print("Apply Column Length ...")
    # Auto-adjust the width of column B based on the content in B2 to B6
    max_length = max(len(str(sheet.cell(row=i, column=2).value)) for i in range(2, 8))
    sheet.column_dimensions['B'].width = max_length + 2  # Add a little extra space

    total = int(range_list[0][0][-2:])
    max_length = max(len(str(sheet.cell(row=i, column=3).value)) for i in range(2, total+1))
    sheet.column_dimensions['C'].width = max_length + 2  # Add a little extra space

    max_length = max(len(str(sheet.cell(row=i, column=5).value)) for i in range(2, total+1))
    sheet.column_dimensions['E'].width = max_length + 2  # Add a little extra space

    # # Auto-size columns after applying formatting
    # for col_letter in ["D","E"]:
    #     sheet.column_dimensions[col_letter].auto_size = True

    
    return

def apply_title(sheet,range_list):

    print("Apply Title ...")

    sheet.merge_cells(range_list[3][0])
    part = range_list[3][0].split(":")
    cell = sheet[part[0]]
    cell.value = "Contact System Information"
    cell.font = Font(bold=True)

    sheet.merge_cells(range_list[3][1])
    part = range_list[3][1].split(":")
    cell = sheet[part[0]]
    cell.value = "Equipment Rating (A)"
    cell.font = Font(bold=True)

    sheet.merge_cells(range_list[3][2])
    part = range_list[3][2].split(":")
    cell = sheet[part[0]]
    cell.value = "Simulation Result: Maximum RMS Current (A)"
    cell.font = Font(bold=True)

    return

# conditional formatting
def condtional_formating(sheet,range_list):
    print("Apply Conditional Formatting ...")
    # Compare values in columns I and J for each row and shade accordingly
    # set the pattern fill
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red (0% S)
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 255,152,51 (80% S)
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green (00% S)
    
    # general
    sheet.conditional_formatting.add(range_list[4][1],CellIsRule(operator = 'greaterThanOrEqual',formula=[range_list[4][0]],fill=red_fill))
    sheet.conditional_formatting.add(range_list[4][1],CellIsRule(operator = 'between',formula=[range_list[4][0]+'*0.9',range_list[4][0]],fill=yellow_fill))
    sheet.conditional_formatting.add(range_list[4][1],CellIsRule(operator = 'lessThan',formula=[range_list[4][0]],fill=green_fill))

    for index, lst in enumerate(range_list):
        if index > 4:
            sheet.conditional_formatting.add(lst[1],CellIsRule(operator = 'greaterThanOrEqual',formula=[lst[0]],fill=red_fill))
            sheet.conditional_formatting.add(lst[1],CellIsRule(operator = 'between',formula=[lst[0]+'*0.9',lst[0]],fill=yellow_fill))
            sheet.conditional_formatting.add(lst[1],CellIsRule(operator = 'lessThan',formula=[lst[0]],fill=green_fill))

    return


if __name__ == "__main__":
    # Add your debugging code here
    simname = "StraightLine1"  # Provide a simulation name or adjust as needed
    main_option = "1"  # Adjust as needed
    time_start = "0070000"  # Adjust as needed
    time_end = "0100000"  # Adjust as needed
    option_select = "2"  # Adjust as needed
    text_input = "output"  # Adjust as needed
    low_v = None  # Adjust as needed
    high_v = None  # Adjust as needed
    time_step = "1"  # Adjust as needed

    # Call your main function with the provided parameters
    main(simname, main_option, time_start, time_end, option_select, text_input, low_v, high_v, time_step)

