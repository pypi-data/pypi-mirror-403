#
# -*- coding: utf-8  -*-
#=================================================================
# Created by: Jieming Ye
# Created on: Feb 2024
# Last Modified: Feb 2024
#=================================================================
# Copyright (c) 2024 [Jieming Ye]
#
# This Python source code is licensed under the 
# Open Source Non-Commercial License (OSNCL) v1.0
# See LICENSE for details.
#=================================================================
"""
Pre-requisite: 
SimulationName.oof: default oslo output file after successfully run a simulation.
xxxxx.xlsx: excel spreadsheet with proper user settings.
Used Input:
simname: to locate the result file or file rename
time_start: to define the analysis start time
time_end: to define the analysis end time
option_select: to define the option selected in subfunction
text_input: to locate the excel file name
other input for oslo_extraction.py only.
Expected Output:
Updated excel spreadsheet with analysis result.
Description:
This script defines the process of doing new grid compliance assessment.
First, it reads the excel file and save the user input in a data frame called start_df. Then it reads the related .d4 file and saved in a data frame list called d4dataframe. This is for future proof as only one .df file is needed for this assessment. The calculation is then carried out. It then doing some analysis and save the final result in an updated data frame. Final step is to save all information in the excel in a proper format. The whole process is easy to follow via reading the code.
The calculation to be checked is saved in d4_file_sp_process() function. Some formulas are hard coded in this function, i,e, RVC calculation, memory time calculation.
The key manual updating part is get_result_range() function which depends on the desired output format in excel, followed by table_formatting() function where some reading needs manual input. The curve_plot() function is also need manual change if some new plots is desired in the future. The other process should be auto adjustable as long as the excel input format is followed.

"""
#=================================================================
# VERSION CONTROL
# V1.0 (Jieming Ye) - Initial Version
#=================================================================
# Set Information Variable
# N/A
#=================================================================

# vision_oslo_extension/excel_processing.py
import pandas as pd
import numpy as np
# import openpyxl
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, NamedStyle
from openpyxl.formatting.rule import CellIsRule,FormulaRule
from openpyxl.chart import ScatterChart,Reference,Series
from openpyxl.drawing.text import CharacterProperties
from openpyxl.chart.shapes import GraphicalProperties

# import vision_oslo
from vision_oslo_extension import oslo_extraction
from vision_oslo_extension.shared_contents import SharedMethods


def main(simname, main_option, time_start, time_end, option_select, text_input, low_v, high_v, time_step):

    print("")
    print("New Grid Connection Assessment: ------>")
    print("")
    
    simname = simname
    # Specify Excel file name
    excel_file = text_input + ".xlsx"

    if not SharedMethods.check_existing_file(excel_file):
        return False
    
    #option = "1" # Fully Restart, require oof,
    #option = "2" # Process only

    option = option_select # 1:Fully Restart, require oof, and list

    start = 10
    plot_start  = 10 # for feeder table, start from row 10
    total = 15

    if option not in ["0","1","2"]:
        SharedMethods.print_message("ERROR: Error in grid_connection.py. Please contact Support.","31")
        return False
    
    if option == "0":
        SharedMethods.print_message("ERROR: Please select an option to proceed.","31")
        return False

    # start_cell = 'B11'
    # read data from start tab
    # start_df = pd.read_excel(writer,sheet_name = "Start")
    setting_df = start_reading_process(simname,excel_file)
    if isinstance(setting_df, pd.DataFrame):
        pass
    elif setting_df is False:
        SharedMethods.print_message("ERROR: Data Reading Failed. Check your settings in Excel.","31")
        return False
        # Handle the case when the function returns False

    # get the feeder ID, ratio, time, etc, important info.    
    sec_v = setting_df.at[0,'Settings']  # modelled v
    time_increment = setting_df.at[1,'Settings'] # time increament

    feeder = str(setting_df.at[2,'Settings'])

    pri_v = setting_df.at[4,'Settings']  # grid voltage level
    ratio = pri_v / sec_v  # ratio between actual and modelled

    max_op_range = setting_df.at[5,'Settings']  # grid voltage max level
    min_op_range = setting_df.at[6,'Settings']  # grid voltage max level
    fault = setting_df.at[7,'Settings']  # grid voltage max level

    load_1_min = setting_df.at[9,'Settings']  # load from BA
    load_10_min = setting_df.at[10,'Settings']
    load_30_min = setting_df.at[11,'Settings']

    event_f = setting_df.at[13,'Settings']  # Assumed RVC current
    event_if = setting_df.at[14,'Settings']
    event_vif = setting_df.at[15,'Settings']

    # check input sensible or not:
    result = check_input_info(sec_v,time_increment,feeder,pri_v,max_op_range,min_op_range,fault)
    if result == False:
        return False
    
    # calculate and create result data table, calculate the criteria
    result_df = criteria_calculate(pri_v,max_op_range,min_op_range,load_1_min,load_10_min,load_30_min,time_increment)

    # check if want to go throught the feeder check process
    if option == "1":
        if not SharedMethods.check_oofresult_file(simname):
            return False
        # extract the feeder
        oslo_extraction.feeder_step(simname, time_start, time_end, "1", feeder)
        
    if option == "2":
        print("Checking essential d4 files...")
        filename = simname + "_" + feeder + ".osop.d4"
        if not SharedMethods.check_existing_file(filename):
            SharedMethods.print_message(f"ERROR: d4 file {feeder} do not exist. Select option 1 or check Supply Point ID to proceed.","31")
            return False
    
    # process d4 file
    feeder_list = [feeder]
    d4_df, d4_sum, result_df = feeder_reading_process(simname,feeder_list,pri_v,ratio,time_increment,result_df)
    
    # Update the result table for RVC
    result_df = result_rvc_update(simname, result_df,pri_v, fault, event_f,event_if,event_vif)

    # save the data
    data_write_save(simname, excel_file,start,plot_start,d4_df,d4_sum,result_df,feeder)

    return True


# write data to excel
def data_write_save(simname, excel_file,start,plot_start,d4_df,d4_sum,result_df,feeder):

    # write data to excel
    with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a',if_sheet_exists='overlay') as writer:

        # get currrent workbook to do process
        wb = writer.book

        print("Generate Result Page...")
        result_df.to_excel(writer, sheet_name="Result", index=False, startrow = start, startcol = 1)

        # print("Generate Plot Page...")
        print(f"Saving {feeder}...")

        # Add an empty row with NaN values
        d4_sum.loc[len(d4_sum)] = np.nan

        # write calculation instruction
        info_row = {'S_inst':'Refer Appendix E in 157897-NRD-REP-EDS-000003 and 000004, MML Braybrooke Assessment',
                    'SVC':'Step Voltage Change In Percentage',
                    'worst_case_d':'d = (step voltage change/nominal voltage)*(1+10%)',
                    'tn_step': 'tn = 18.4*(100*d*F)^3 with F=1',
                    't_memory': 't = rolling 10 min sum of tn_step'}
        d4_sum.loc[len(d4_sum)] = info_row


        # save information to spreadsheet
        d4_df.insert(d4_df.columns.get_loc('I_angle') + 1, 'New_C_1', np.nan)
        d4_sum.insert(d4_sum.columns.get_loc('I_angle') + 1, 'New_C_1', np.nan)

        d4_df.insert(d4_df.columns.get_loc('New_C_1') + 1, 'T_F_Plot', np.nan)
        d4_sum.insert(d4_sum.columns.get_loc('New_C_1') + 1, 'T_F_Plot', np.nan)

        d4_df.insert(d4_df.columns.get_loc('S_30min') + 1, 'New_C_2', np.nan)
        d4_sum.insert(d4_sum.columns.get_loc('S_30min') + 1, 'New_C_2', np.nan)

        d4_df.insert(d4_df.columns.get_loc('abs_SVC') + 1, 'New_C_3', np.nan)
        d4_sum.insert(d4_sum.columns.get_loc('abs_SVC') + 1, 'New_C_3', np.nan)

        d4_sum.to_excel(writer, sheet_name=feeder, index=False, startrow = 0)

        d4_df.to_excel(writer, sheet_name=feeder, index=False, startrow = d4_sum.shape[0]+2)
        
        # # Calculate the Excel range for each DataFrame
        range_list = get_result_range(feeder,d4_df)

        # do the plot
        total = len(d4_df)
        curve_plot(wb,feeder,range_list,result_df,plot_start,total)
        
        # # table formatting
        table_formatting(simname,wb,range_list,feeder)

        print("Saving Excel File...")

    return

# plot some essential graphs
def curve_plot(wb,feeder,range_list,result_df,plot_start,total):

    print("Plotting Graphs in Result... ")
    # define the time column
    sheet = wb[feeder]
    
    index = plot_start
    for line in sheet[range_list[5][0]]:
        for cell in line:
            cell.value = f'=TIME(LEFT(C{index},2),MID(C{index},4,2),RIGHT(C{index},2))'
            cell.number_format = 'hh:mm:ss'
        index = index + 1

    cell_list = range_list[5][0].split(":")
    time_start = sheet[cell_list[0]].value
    time_end = sheet[cell_list[1]].value
    v_max = result_df.iloc[0,2]
    v_min = result_df.iloc[1,2]

    time = sheet['C10'].value
    time = time.split(":")
    ts = (float(time[0])+float(time[1])/60+float(time[1])/3600)/24

    time = sheet[f'C{10+total-1}'].value
    time = time.split(":")
    te = (float(time[0])+float(time[1])/60+float(time[1])/3600)/24

    xlim = 0
    ylim = 1
    
    # find nearest 10 min timeline.
    for i in range(0,144):
        if i*1/144 > ts:
            xlim = (i-1)*1/144
            for j in range(i,144):
                if j*1/144 >= te:
                    ylim = j*1/144
                    break
            break
    
    tspace = (ylim-xlim)/6

    # start plot
    sheet = wb['Result']
    start = plot_start
    index = 11 # location on the result from row 11

    ## plot 1 - Votlage ###################################################################
    # Create a scatter plot
    chart = ScatterChart('smoothMarker')
    chart.title = "Point of Common Coupling - Voltage Profile"
    chart.x_axis.title = "Time (hh:mm:ss)"
    chart.y_axis.title = "Voltage (kV)"

    # Set chart size
    chart.height = 10  # Adjust width as needed
    chart.width = 25  # Adjust height as needed

    x = 11 # column index for time
    y = 23 # column index for supply point voltage
    xvalues = Reference(wb[feeder], min_col=x, min_row=start, max_row=start+total-1)
    yvalues = Reference(wb[feeder], min_col=y, min_row=start, max_row=start+total-1)
    series = Series(yvalues,xvalues,title_from_data=False,title = 'PCC Voltage')
    chart.series.append(series)

    # to be discussed if plotting threshold is nneded in the future
    
    # rows = [['Time','Threshold'],[time_start, v_max],[time_end, v_max],[time_start, v_min],[time_end, v_min]]
    # for row_index, row_data in enumerate(rows, start=1):
    #     for col_index, value in enumerate(row_data, start=32):  # Column AF is the 32nd column
    #         wb[feeder].cell(row=row_index, column=col_index, value=value)
    
    # xvalues = Reference(wb[feeder], min_col=32, min_row=2, max_row=3)
    # yvalues = Reference(wb[feeder], min_col=33, min_row=2, max_row=3)
    # series = Series(yvalues,xvalues,title_from_data=False,title = 'Maximum Voltage')
    # chart.series.append(series)

    # xvalues = Reference(wb[feeder], min_col=32, min_row=4, max_row=5)
    # yvalues = Reference(wb[feeder], min_col=33, min_row=4, max_row=5)
    # series = Series(yvalues,xvalues,title_from_data=False,title = 'Minimum Voltage')
    # chart.series.append(series)

    chart.legend.position = 't' # top legend position

    cp = CharacterProperties(sz=1400)
    chart.title.tx.rich.p[0].r[0].rPr = cp 
    # chart.title.tx.rich.paragraphs = cp 

    # Access major gridlines and set color to light grey
    cl = GraphicalProperties()
    cl.line.solidFill = "E0E0E0"
    chart.x_axis.majorGridlines.spPr = cl  # Light grey color
    chart.y_axis.majorGridlines.spPr = cl  #  # Light grey color
    chart.x_axis.scaling.min = xlim      
    chart.x_axis.scaling.max = ylim
    chart.x_axis.majorUnit = tspace

    sheet.add_chart(chart,f'H{index}')
    index = index + 20

    ## plot 2 - SVC ###################################################################
    # Create a scatter plot
    chart = ScatterChart('smoothMarker')
    chart.title = "Point of Common Coupling - Step Voltage Change in Percentage"
    chart.x_axis.title = "Time (hh:mm:ss)"
    chart.y_axis.title = "Percentage %"

    # Set chart size
    chart.height = 10  # Adjust width as needed
    chart.width = 25  # Adjust height as needed

    x = 11 # column index for time
    y = 24 # column index for SVC
    xvalues = Reference(wb[feeder], min_col=x, min_row=start, max_row=start+total-1)
    yvalues = Reference(wb[feeder], min_col=y, min_row=start, max_row=start+total-1)
    series = Series(yvalues,xvalues,title_from_data=False,title = 'SVC')
    chart.series.append(series)

    chart.legend.position = 't' # top legend position

    cp = CharacterProperties(sz=1400)
    chart.title.tx.rich.p[0].r[0].rPr = cp 
    # chart.title.tx.rich.paragraphs = cp 

    # Access major gridlines and set color to light grey
    cl = GraphicalProperties()
    cl.line.solidFill = "E0E0E0"
    chart.x_axis.majorGridlines.spPr = cl  # Light grey color
    chart.y_axis.majorGridlines.spPr = cl  #  # Light grey color
    chart.x_axis.scaling.min = xlim      
    chart.x_axis.scaling.max = ylim
    chart.x_axis.majorUnit = tspace

    sheet.add_chart(chart,f'H{index}')
    index = index + 20

    ## plot 3 - Memory time ###################################################################
    # Create a scatter plot
    chart = ScatterChart('smoothMarker')
    chart.title = "Flicker Severity Stage 2 Assessment - Memory Time Technique"
    chart.x_axis.title = "Time (hh:mm:ss)"
    chart.y_axis.title = "Second (s)"

    # Set chart size
    chart.height = 10  # Adjust width as needed
    chart.width = 25  # Adjust height as needed

    x = 11 # column index for time
    y = 28 # column index for tn
    xvalues = Reference(wb[feeder], min_col=x, min_row=start, max_row=start+total-1)
    yvalues = Reference(wb[feeder], min_col=y, min_row=start, max_row=start+total-1)
    series = Series(yvalues,xvalues,title_from_data=False,title = 'tn: memory time for each step')
    chart.series.append(series)

    y = 29 # column index for t
    yvalues = Reference(wb[feeder], min_col=y, min_row=start, max_row=start+total-1)
    series = Series(yvalues,xvalues,title_from_data=False,title = 't: memory time for 10 min sum')
    chart.series.append(series)

    chart.legend.position = 't' # top legend position

    cp = CharacterProperties(sz=1400)
    chart.title.tx.rich.p[0].r[0].rPr = cp 
    # chart.title.tx.rich.paragraphs = cp 

    # Access major gridlines and set color to light grey
    cl = GraphicalProperties()
    cl.line.solidFill = "E0E0E0"
    chart.x_axis.majorGridlines.spPr = cl  # Light grey color
    chart.y_axis.majorGridlines.spPr = cl  #  # Light grey color
    chart.x_axis.scaling.min = xlim      
    chart.x_axis.scaling.max = ylim
    chart.x_axis.majorUnit = tspace

    sheet.add_chart(chart,f'H{index}')
    index = index + 20

    ## plot 4/5 - 30 min ###################################################################
    # Create a scatter plot
    chart = ScatterChart('smoothMarker')
    chart.title = "30 min Average Reactive Power"
    chart.x_axis.title = "Time (hh:mm:ss)"
    chart.y_axis.title = "Reactive Power (MVAr)"

    # Set chart size
    chart.height = 10  # Adjust width as needed
    chart.width = 25  # Adjust height as needed

    x = 11 # column index for time
    y = 20 # column index for reactieve power
    xvalues = Reference(wb[feeder], min_col=x, min_row=start, max_row=start+total-1)
    yvalues = Reference(wb[feeder], min_col=y, min_row=start, max_row=start+total-1)
    series = Series(yvalues,xvalues,title_from_data=False,title = '30-min Average Reactive Power')
    chart.series.append(series)

    chart.legend.position = 't' # top legend position

    cp = CharacterProperties(sz=1400)
    chart.title.tx.rich.p[0].r[0].rPr = cp 
    # chart.title.tx.rich.paragraphs = cp 

    # Access major gridlines and set color to light grey
    cl = GraphicalProperties()
    cl.line.solidFill = "E0E0E0"
    chart.x_axis.majorGridlines.spPr = cl  # Light grey color
    chart.y_axis.majorGridlines.spPr = cl  #  # Light grey color
    chart.x_axis.scaling.min = xlim      
    chart.x_axis.scaling.max = ylim
    chart.x_axis.majorUnit = tspace

    if result_df.iloc[4,2] == "-":
        pass
    else:
        sheet.add_chart(chart,f'H{index}')
        index = index + 20

    
    # Create a scatter plot
    chart = ScatterChart('smoothMarker')
    chart.title = "30 min Average Apparent Power"
    chart.x_axis.title = "Time (hh:mm:ss)"
    chart.y_axis.title = "Power (MVA)"

    # Set chart size
    chart.height = 10  # Adjust width as needed
    chart.width = 25  # Adjust height as needed

    x = 11 # column index for time
    y = 21 # column index for reactieve power
    xvalues = Reference(wb[feeder], min_col=x, min_row=start, max_row=start+total-1)
    yvalues = Reference(wb[feeder], min_col=y, min_row=start, max_row=start+total-1)
    series = Series(yvalues,xvalues,title_from_data=False,title = '30-min Average Power')
    chart.series.append(series)

    chart.legend.position = 't' # top legend position

    cp = CharacterProperties(sz=1400)
    chart.title.tx.rich.p[0].r[0].rPr = cp 
    # chart.title.tx.rich.paragraphs = cp 

    # Access major gridlines and set color to light grey
    cl = GraphicalProperties()
    cl.line.solidFill = "E0E0E0"
    chart.x_axis.majorGridlines.spPr = cl  # Light grey color
    chart.y_axis.majorGridlines.spPr = cl  #  # Light grey color
    chart.x_axis.scaling.min = xlim      
    chart.x_axis.scaling.max = ylim
    chart.x_axis.majorUnit = tspace

    if result_df.iloc[4,2] == "-":
        pass
    else:
        sheet.add_chart(chart,f'H{index}')
        index = index + 20

    return

# read the start tab and collect informaiton
def start_reading_process(simname, excel_file):
    
    # File check
    print("Check Excel and Deleting Existing Contents ...")

    try:
        wb = load_workbook(excel_file)
        # Delete all sheets except 'Start'
        for sheet_name in wb.sheetnames:
            if sheet_name != 'Start':
                wb.remove(wb[sheet_name])
        
        # Save changes
        wb.save(excel_file)

    except Exception as e:
        SharedMethods.print_message(f"ERROR: (Close the Excel file and Start Again) Error: {e}","31")
        return False
        
    # # read start up information 
    # columns = ["Items","Description","Settings","Note"]

    print("Reading Setting Tables...")
    # Specify the range A11 to D25  attention to range
    start_row = 10  # 11th row in zero-based indexing
    end_row = 26    # 27th row in zero-based indexing
    setting_df = pd.read_excel(excel_file,sheet_name = 'Start',header = 0,usecols=range(4), skiprows=start_row, nrows=end_row - start_row)

    # print(setting_df)
    # print(criteria_df)
    
    return setting_df

# create criteria table and ready for result
def criteria_calculate(pri_v,max_op_range,min_op_range,load_1_min,load_10_min,load_30_min,time_increment):
    # find the NPS value
    if pri_v < 150:
        a = 0.02 # 1 min NPS
        b = '-' # 10 min NPS
        c = 0.01 # 30 min NPS
    else:
        a = '-'
        b = 0.015
        c = 0.01

    columns = ['Assessment Item','Unit','Assessment Criteria','Simulation Result']

    # Example data to populate the DataFrame
    data = [
        ['PCC Maximum Voltage', 'kV', pri_v + pri_v*max_op_range, np.nan],
        ['PCC Minimum Voltage', 'kV', pri_v + pri_v*min_op_range, np.nan],
        ['PCC 1-min Maximum Apparent Load', 'MVA', np.nan, np.nan],
        ['PCC 10-min Maximum Apparent Load', 'MVA', np.nan, np.nan],
        ['PCC 30-min Maximum Apparent Load', 'MVA', np.nan, np.nan],
        ['PCC 1-min Maximum Reactive Load', 'MVAr', np.nan, np.nan],
        ['PCC 10-min Maximum Reactive Load', 'MVAr', np.nan, np.nan],
        ['PCC 30-min Maximum Reactive Load', 'MVAr', np.nan, np.nan],
        ['PCC 1-min Negative Phase Sequence (NPS)', '%', a, np.nan],
        ['PCC 10-min Negative Phase Sequence (NPS)', '%', b, np.nan],
        ['PCC 30-min Negative Phase Sequence (NPS)', '%', c, np.nan],
        ['PCC Step Voltage Change (SVC)', '%', 0.03, np.nan],
        ['PCC Rapid Voltage Change (RVC) - Frequent Event', '%', 0.06, np.nan],
        ['PCC Rapid Voltage Change (RVC) - Infrequent Event', '%', 0.1, np.nan],
        ['PCC Rapid Voltage Change (RVC) - Very Infrequent Event', '%', 0.12, np.nan],
        ['Flicker Severity - Short Term - RVC with Pst <= 0.5', '%', np.nan, np.nan],
        ['Flicker Severity - Short Term - Memory Time Technique', 's', 600, np.nan],
        ['Flicker Severity - Long Term - Plt', 'N/A', 'Pass if Stage 2 Pst pass', np.nan],
        # Add more rows as needed
    ]

    # Create the DataFrame from the data
    df = pd.DataFrame(data, columns=columns)

    # if load_1_min is a number than assign the value to 1min load 
    coff = 0.48/np.sqrt(1+0.48*0.48)
    if isinstance(load_1_min, (int, float)) and not np.isnan(load_1_min):
        df.at[2, 'Assessment Criteria'] = load_1_min
        df.at[5, 'Assessment Criteria'] = load_1_min*coff
    else:
        df.at[2, 'Assessment Criteria'] = '-'
        df.at[5, 'Assessment Criteria'] = '-'
        
    if isinstance(load_10_min, (int, float)) and not np.isnan(load_10_min):
        df.at[3, 'Assessment Criteria'] = load_10_min
        df.at[6, 'Assessment Criteria'] = load_10_min*coff
    else:
        df.at[3, 'Assessment Criteria'] = '-'
        df.at[6, 'Assessment Criteria'] = '-'
        
    if isinstance(load_30_min, (int, float)) and not np.isnan(load_30_min):
        df.at[4, 'Assessment Criteria'] = load_30_min
        df.at[7, 'Assessment Criteria'] = load_30_min*coff
    else:
        df.at[4, 'Assessment Criteria'] = '-'
        df.at[7, 'Assessment Criteria'] = '-'
        
    # assign the Flicker criteria stage 2
    # based on 1 ,2,3,4,5 sec
    criteria_values = [0.004, 0.0045, 0.005, 0.0055, 0.006]
    df.at[15, 'Assessment Criteria'] = criteria_values[time_increment - 1]
    
    return df

#create protection limit dataframe
def result_rvc_update(simname, result_df, pri_v, fault, event_f,event_if,event_vif):

    vo_impedance = 2* (pri_v*pri_v) / fault  #vision oslo equivalent source impedance
    nominal_v = 25  # inrush current at 25 kV (nominla voltage setting)

    temp = result_df.at[2, 'Simulation Result'] / fault # 1 min NPS = 1 min Power / falue
    result_df.at[8, 'Simulation Result'] = temp

    temp = result_df.at[3, 'Simulation Result'] / fault # 10 min NPS = 10 min Power / falue
    result_df.at[9, 'Simulation Result'] = temp

    temp = result_df.at[4, 'Simulation Result'] / fault # 30 min NPS = 30 min Power / falue
    result_df.at[10, 'Simulation Result'] = temp

    if isinstance(event_f, (int, float)) and not np.isnan(event_f):
        temp = (event_f * nominal_v / pri_v) * vo_impedance / pri_v /1000 # current at primary * impedance / primary nominal voltag
        result_df.at[12, 'Simulation Result'] = temp
    else:
        result_df.at[12, 'Simulation Result'] = '-'

    if isinstance(event_if, (int, float)) and not np.isnan(event_if):
        temp = (event_if * nominal_v / pri_v) * vo_impedance / pri_v /1000 # current at primary * impedance / primary nominal voltag
        result_df.at[13, 'Simulation Result'] = temp
    else:
        result_df.at[13, 'Simulation Result'] = '-'

    if isinstance(event_vif, (int, float)) and not np.isnan(event_vif):
        temp = (event_vif * nominal_v / pri_v) * vo_impedance / pri_v /1000 # current at primary * impedance / primary nominal voltag
        result_df.at[14, 'Simulation Result'] = temp
    else:
        result_df.at[14, 'Simulation Result'] = '-'

    return result_df

# read individual d4 file
def feeder_reading_process(simname,feeder_list,pri_v,ratio,time_increment,result_df):

    # create dataframe
    d4dataframe = []
    #sumdataframe = []

    for index, feeder in enumerate(feeder_list):
        print(f"Processing {feeder} ...")
        filename = simname + "_" + feeder +".osop.d4"
        delimiter = '\\s+'
        columns = ["FeederID","Type","Time","P_inst","Q_inst","Voltage","V_angle","Current","I_angle"]
        dtype_mapping = {"Time": str,}
        df = pd.read_csv(filename, delimiter=delimiter, names = columns, skiprows = 11,dtype = dtype_mapping) 
        # Extracting parts from the string and formatting the 'Time' column
        df['Time'] = df['Time'].apply(lambda x: f"{int(x[1:3]):02d}:{int(x[3:5]):02d}:{int(x[5:]):02d}")
        # data frame process
        df = d4_file_sp_process(df,pri_v, ratio,time_increment)
        # calculate sum value
        df_sum = d4_find_max(df)
        #update summary result
        result_df = result_update(result_df, df_sum)

    return df, df_sum, result_df

# doing invididual d4 file process and calculatoin for supply points
def d4_file_sp_process(df, pri_v, ratio, time_increment):

    # calculate Power 
    df['S_inst'] = np.sqrt(df['P_inst']**2 + df['Q_inst']**2)

    window_sizes = {'1min': 60, '10min': 600, '30min': 1800}

    for time_interval, window_size in window_sizes.items():
        df[f'P_{time_interval}'] = df['P_inst'].rolling(window=int(window_size / time_increment)).mean()
        df[f'Q_{time_interval}'] = df['Q_inst'].rolling(window=int(window_size / time_increment)).mean()
        df[f'S_{time_interval}'] = np.sqrt(df[f'P_{time_interval}']**2 + df[f'Q_{time_interval}']**2)

    df['SP_Voltage'] = df['Voltage']*ratio  # actually supply point voltage
    df['SVC'] = df['SP_Voltage'].pct_change()  # default method to calculate percentage difference
    #df['SVC'] = df['SVC'].map('{:.2f}%'.format) # change the data shown in percentage format note this will hcange the data type to str

    df['abs_SVC'] = np.abs(df['SVC']) # calculate the absolute value of SVC
    #df['abs_SVC'] = df['abs_SVC'].map('{:.2f}%'.format) # change the data shown in percentage format

    df['worst_case_d'] = np.abs(df['SP_Voltage'].diff())/pri_v*1.1  # calcuate the worst case d for memory time relative votlage change
    df['tn_step'] = (df['worst_case_d']*100)**3 * 18.4  # calculate tn = 18.4 * (100*d*F)^3, with F=1 for step voltage change.
    
    df['t_memory'] = df['tn_step'].rolling(window=int(600 / time_increment)).sum() # rolling 10 min sum as required

    return df

# find maximum value and time of each d4 file
def d4_find_max(df):
    # # Insert two empty rows above the first row
    # df = pd.concat([pd.DataFrame(index=range(2)), df], ignore_index=True)
    sum_df = pd.DataFrame(columns=df.columns,index=range(4))
    sum_df.iloc[0, 0] = "Maximum Value"
    sum_df.iloc[1, 0] = "Maximum Value at Time"
    sum_df.iloc[2, 0] = "Minimum Value"
    sum_df.iloc[3, 0] = "Minimum Value at Time"

    start = 9

    if df.empty:
        sum_df.iloc[0, 1] = "DATA FOR THIS FEEDER IS NOT AVAILABLE"

    else:
        for column in df.columns[start:]:
            if df[column].dtype in [int, float]:
                max_value = df[column].max()
                time_of_max = df.loc[df[column].idxmax(), 'Time']
                sum_df.at[0, column] = max_value
                sum_df.at[1, column] = time_of_max
                min_value = df[column].min()
                time_of_min = df.loc[df[column].idxmin(), 'Time']
                sum_df.at[2, column] = min_value
                sum_df.at[3, column] = time_of_min

    return sum_df

# update result table based on output of d4 process
def result_update(target,df_sum):

#     Assessment Item  Unit       Assessment Criteria  Simulation Result
# 0                                 PCC Maximum Voltage    kV                     420.0                NaN
# 1                                 PCC Minimum Voltage    kV                     360.0                NaN
# 2                     PCC 1-min Maximum Apparent Load   MVA                       100                NaN
# 3                    PCC 10-min Maximum Apparent Load   MVA                         -                NaN
# 4                    PCC 30-min Maximum Apparent Load   MVA                        40                NaN
# 5                     PCC 1-min Maximum Reactive Load  MVAr                 43.273107                NaN
# 6                    PCC 10-min Maximum Reactive Load  MVAr                         -                NaN
# 7                    PCC 30-min Maximum Reactive Load  MVAr                 17.309243                NaN

# 8                     PCC 1-min NPS                       %                 43.273107                NaN
# 9                    PCC 10-min NPS                       %                         -                NaN
# 10                   PCC 30-min NPS                       %                 17.309243                NaN

# 11                      PCC Step Voltage Change (SVC)     %                      0.03                NaN
# 12    PCC Rapid Voltage Change (RVC) - Frequent Event     %                      0.06                NaN
# 13  PCC Rapid Voltage Change (RVC) - Infrequent Event     %                       0.1                NaN
# 14  PCC Rapid Voltage Change (RVC) - Very Infreque...     %                      0.12                NaN
# 15  Flicker Severity - Short Term - RVC with Pst <...     %                     0.006                NaN
# 16  Flicker Severity - Short Term - Memory Time Te...     s                       600                NaN
# 17                 Flicker Severity - Long Term - Plt   N/A  Pass if Stage 2 Pst pass                NaN
    
    target.loc[0,'Simulation Result'] = df_sum.at[0,'SP_Voltage']
    target.loc[1,'Simulation Result'] = df_sum.at[2,'SP_Voltage']

    target.loc[2,'Simulation Result'] = df_sum.at[0,'S_1min']
    target.loc[3,'Simulation Result'] = df_sum.at[0,'S_10min']
    target.loc[4,'Simulation Result'] = df_sum.at[0,'S_30min']
    target.loc[5,'Simulation Result'] = df_sum.at[0,'Q_1min']
    target.loc[6,'Simulation Result'] = df_sum.at[0,'Q_10min']
    target.loc[7,'Simulation Result'] = df_sum.at[0,'Q_30min']

    target.loc[11,'Simulation Result'] = df_sum.at[0,'abs_SVC'] # SVC
    target.loc[15,'Simulation Result'] = df_sum.at[0,'abs_SVC'] # Flicker Short Term stage 2
    target.loc[16,'Simulation Result'] = df_sum.at[0,'t_memory'] # Flicker memory time
    
    return target

# check input information:
def check_input_info(sec_v,time_increment,feeder,pri_v,max_op_range,min_op_range,fault):

    # check if time increament is sensible or not
    if 0 <= time_increment <= 5:
        print(f"Simulation Time Increment was Set to {time_increment} second.")
    else:
        SharedMethods.print_message("ERROR: Please set simulation time increment between 1-5 and restart the process....","31")
        return False
    
    if isinstance(sec_v, str) or np.isnan(sec_v):
        SharedMethods.print_message("ERROR: Please check input in Modelled Supply Voltage. Value Required...","31")
        return False
    else:
        if sec_v < 10 or sec_v > 450:
            SharedMethods.print_message("WARNING: Norminal Voltage in kV. Abnormal norminal votage detected.","33")
    
    if isinstance(pri_v, str) or np.isnan(pri_v):
        SharedMethods.print_message("ERROR: Please check input in Nominal Supply Voltage. Value Required...","31")
        return False
    else:
        if pri_v < 10 or pri_v > 450:
            SharedMethods.print_message("WARNING: Norminal Voltage in kV. Abnormal norminal votage detected.","33")
    
    if isinstance(max_op_range, str) or np.isnan(max_op_range):
        SharedMethods.print_message("ERROR: Check Maximum Operating Range. Value Required...","31")
        return False
    
    if isinstance(min_op_range, str) or np.isnan(min_op_range):
        SharedMethods.print_message("ERROR: Check Minimum Operating Range. Value Required...","31")
        return False
    
    if isinstance(fault, str) or np.isnan(fault) or fault == 0:
        SharedMethods.print_message("ERROR: Please check input in Fault Level. Value Required...","31")
        return False
    
    if len(feeder) > 4:
        SharedMethods.print_message("ERROR: Feeder ID larger than 4 digit. Extraction is not possible...","31")
        return False

    return True

# check the range
def get_result_range(feeder,d4_df):

    range_list = []
    total = len(d4_df)
    # 0: table frame range
    # 1: table data range (middle setting)
    # 2: title range 
    # 3: 2 digit range
    # 4: percentage range
    # 5: Feeder Time Range

    # 6: Conditional formatting range (critiera, range)
    # 7: speical range (check cell, 1,2,3,4. decision cell 5)

    # 8: range for percetage in result
    
    # 0
    result_range = [f"B11:E29"]
    range_list.append(result_range)
    # 1
    data_range = [f"C12:E29"]
    range_list.append(data_range)
   
    # 2
    title_range = [f"B11:E11"]
    range_list.append(title_range)
    
    #3 4 5
    add_range_1 = [f"D14:D19",f"E12:E19",f"E28:E28"]
    range_list.append(add_range_1)

    add_range_2 = [f"D20:E27"]
    range_list.append(add_range_2)

    time_range = [f"K10:K{10 + total - 1}"]
    range_list.append(time_range)

    #6,7
    condi = [f"D12",f"E12:E28",f"D13",f"E13"]
    range_list.append(condi)

    spec = ["D27","E27","D28","E28","E29"]
    range_list.append(spec)

    #8
    pec_range = [f"X10:Y{10 + total - 1}"]
    range_list.append(pec_range)

    return range_list

# Result table table formating
def table_formatting(simname,wb,range_list,feeder):
    # format the result table

    print("Formatting Process ...")
    # wb = load_workbook(excel_file)
    print("Information Collection ...")
    sheet = wb["Start"]
    project_name = sheet['B2'].value
    feeding_desp = sheet['B4'].value
    modeller = sheet['B5'].value
    date = sheet['B6'].value

    # Result Tab process
    sheet = wb["Result"]
    sheet['B2'].value = "Project Name:"
    sheet['C2'].value = project_name
    sheet['B3'].value = "Simulation Name:"
    sheet['C3'].value = simname
    sheet['B4'].value = "Feeding Arrangement:"
    sheet['C4'].value = feeding_desp
    sheet['B5'].value = "Result Created by:"
    sheet['C5'].value = modeller
    sheet['B6'].value = "Result Created at:"
    sheet['C6'].value = datetime.now().strftime("%d-%m-%Y %H:%M")

    sheet['H4'].value = "Note:"
    sheet['H5'].value = "Detailed calculation principle --> Refer to Braybrooke Flicker & Reactive Power Assessment Report "
    sheet['H6'].value = "Doc Ref: 157897-NRD-REP-EDS-000003 & 000004"


    print("Apply Border and Alignment...")
    # Define a custom style for formatting with two decimal places
    for range_name in range_list[0]:
        for row in sheet[range_name]:
            for cell in row:
                # Apply border to all sides of the cell
                cell.border = Border(left=Side(border_style='thin'),
                                    right=Side(border_style='thin'),
                                    top=Side(border_style='thin'),
                                    bottom=Side(border_style='thin'))
    
    for range_name in range_list[1]:
        for row in sheet[range_name]:
            for cell in row:
                # Align cell content to the middle
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    print("Apply Numbering Format ...")
    # Define a custom style for formatting with two decimal places
    for range_name in range_list[3]:
        for row in sheet[range_name]:
            for cell in row:
                cell.number_format = '0.00'

    for range_name in range_list[4]:
        for row in sheet[range_name]:
            for cell in row:
                cell.number_format = '0.00%'

    print("Apply Font and Shading...")
    # Shade the range B11:Z11 with a light gray background color
    for range_name in range_list[2]:
        for row in sheet[range_name]:
            for cell in row:
                cell.font = Font(bold = True, italic = True, size = 12)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    print("Check Special Condition...")
    if sheet[range_list[7][1]].value < sheet[range_list[7][0]].value or \
        sheet[range_list[7][3]].value < sheet[range_list[7][2]].value:
        sheet[range_list[7][4]].value = "PASS"
        sheet[range_list[7][4]].fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    else:
        sheet[range_list[7][4]].value = "TO BE CONFIRMED"
        sheet[range_list[7][4]].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    

    condtional_formating(sheet,range_list)

    print("Apply Column Length ...")
    # Auto-adjust the width of column B based on the content in B2 to B6
    max_length = max(len(str(sheet.cell(row=i, column=2).value)) for i in range(2, 26))
    sheet.column_dimensions['B'].width = max_length + 2  # Add a little extra space

    # Auto-size columns after applying formatting
    for col_letter in ["C","D","E","F"]:
        sheet.column_dimensions[col_letter].auto_size = True

    
    # detail page formatting
    sheet = wb[feeder]
    for range_name in range_list[8]:
        for row in sheet[range_name]:
            for cell in row:
                cell.number_format = '0.00%'
    return

# conditional formatting
def condtional_formating(sheet,range_list):
    print("Apply Conditional Formatting ...")
    # Compare values in columns I and J for each row and shade accordingly
    # set the pattern fill
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red (0% S)
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 255,152,51 (80% S)
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green (00% S)

    # speicial case for minimum voltage
    sheet.conditional_formatting.add(range_list[6][3],CellIsRule(operator = 'greaterThan',formula=[range_list[6][2]],fill=green_fill))
    sheet.conditional_formatting.add(range_list[6][3],CellIsRule(operator = 'lessThanOrEqual',formula=[range_list[6][2]+'*0.9'],fill=red_fill))
    sheet.conditional_formatting.add(range_list[6][3],CellIsRule(operator = 'between',formula=[range_list[6][2],range_list[6][2]+'+5'],fill=yellow_fill))
    
    # general
    #yellow_line = str(float(sheet[range_list[6][0]])*0.9)
    sheet.conditional_formatting.add(range_list[6][1],CellIsRule(operator = 'lessThan',formula=[range_list[6][0]],fill=green_fill))
    sheet.conditional_formatting.add(range_list[6][1],CellIsRule(operator = 'greaterThanOrEqual',formula=[range_list[6][0]+'*0.9'],fill=red_fill))
    sheet.conditional_formatting.add(range_list[6][1],CellIsRule(operator = 'between',formula=[range_list[6][0]+'*0.9',range_list[6][0]],fill=yellow_fill))

    return


if __name__ == "__main__":
    # Add your debugging code here
    simname = "MMLE3MOD014"  # Provide a simulation name or adjust as needed
    main_option = "1"  # Adjust as needed
    time_start = "0070000"  # Adjust as needed
    time_end = "0100000"  # Adjust as needed
    option_select = "1"  # Adjust as needed
    text_input = "output"  # Adjust as needed
    low_v = None  # Adjust as needed
    high_v = None  # Adjust as needed
    time_step = "1"  # Adjust as needed

    # Call your main function with the provided parameters
    main(simname, main_option, time_start, time_end, option_select, text_input, low_v, high_v, time_step)

