#
# -*- coding: utf-8  -*-
#=================================================================
# Created by: Jacky Lai
# Created on: Oct 2025
# Last Modified: Oct 2025
#=================================================================
# Copyright (c) 2025 [Jacky Lai]
#
# This Python source code is licensed under the 
# Open Source Non-Commercial License (OSNCL) v1.0
# See LICENSE for details.
#=================================================================

"""
Pre-requisite: 
xxxxx.xlsx: summary spreadsheet with user configuration settings
Result saved in a proper folder structure
train_list.csv for each simulation already extracted within each folder
Used Input:
simname: to locate the result file or file rename
option_select: to define the option selected in subfunction
text_input: to locate the excel file name
Expected Output:
Updated excel spreadsheet with low voltage analysis result.
Description:
This script summarised the low voltage results using the train_list.csv various simulation to one place.
It will highlight the low voltage events based on the user defined thresholds.
It will generate the following sheets:
- Filtered Branch Results
- Filtered Train Results
- Branch Results
- Train Results
The script reads the configuration from the "Start" sheet of the provided Excel file, 
processes the train_list.csv files from specified simulation folders, 
and outputs the results back into the same Excel file with appropriate formatting."""

#=================================================================
# VERSION CONTROL
# V0.1 (Jacky Lai) - Draft Version
# V1.0 (Jieming Ye) - Update following review
#=================================================================
# Set Information Variable
# N/A
#=================================================================

import pandas as pd
import os

from datetime import datetime, time

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

# import vision_oslo
from vision_oslo_extension.shared_contents import SharedMethods, SharedVariables

def main(simname, main_option, time_start, time_end, option_select, text_input, low_v, high_v, time_step):

    print("")
    print("Low Voltage Summary for AC or DC - - - > ")
    print("")
    
    # Specify Excel file name
    excel_file = text_input + ".xlsx"

    if not SharedMethods.check_existing_file(excel_file):
        return False
    
    # Option:
    # 1: Low Voltage Assessment

    # Read the start tab in the excel and collect information
    result = start_reading_process(excel_file)
    if not result:
        return False
    sims_df, filters, limits = result

    try:
        # Validate inputs
        if sims_df is None or limits is None:
            SharedMethods.print_message("ERROR: Missing simulations or limits from the 'Start' sheet.", "31")
            return False

        # check essential files
        if not check_essential_files(sims_df):
            return False

        # read all train_list.csv for each simulation
        sim_data = load_simulation_data(sims_df, filters)
        if not sim_data:
            SharedMethods.print_message("ERROR: No simulation data loaded. Ensure train_list.csv files exist and filters are correct.", "31")
            return False

        # get unique sorted branches and trains
        sorted_branches, sorted_trains = get_unique_sorted_lists(sim_data)
        if len(sorted_branches) == 0 or len(sorted_trains) == 0:
            SharedMethods.print_message("WARNING: No branches or trains found after loading data. Results may be empty.", "33")

    except ValueError as ve:
        SharedMethods.print_message(f"ERROR: Input validation failed - {ve}", "31")
        return False
    except Exception as e:
        SharedMethods.print_message(f"ERROR: Unexpected error during pre-processing: {e}", "31")
        return False
    
    
    sim_data_v = assess_voltage(sim_data, limits)

    time_interval = 5  # Assuming each record represents a 5-second interval
    
    # generate train results
    train_results_df = generate_train_results(sim_data_v, sorted_trains,time_interval)

    # generate branch results
    branch_results_df = generate_branch_results(sim_data_v, sorted_branches)

    # Filter train results
    filtered_train_results_df = filter_results_by_amber_threshold(train_results_df, limits['amber'])

    # Filter branch results
    filtered_branch_results_df = filter_results_by_amber_threshold(branch_results_df, limits['amber'])

    # Export results to Excel
    export_results_to_excel(excel_file, train_results_df, branch_results_df, filtered_train_results_df, filtered_branch_results_df, sims_df, limits)

    return True

# read the start tab and collect informaiton
def start_reading_process(excel_file):

    # File check
    print("Check Excel and Deleting Existing Contents ...")
    try:
        wb = load_workbook(excel_file)
        # Delete all sheets except 'Start'
        for sheet_name in wb.sheetnames:
            if sheet_name != 'Start':
                wb.remove(wb[sheet_name])
        wb.save(excel_file)
    except Exception as e:
        SharedMethods.print_message(f"ERROR: (Close the Excel file and Start Again) Error: {e}", "31")
        return False
    
    print("Reading simulations, filters, and limits from Excel...")

    # Define cell positions (1 based index)
    sim_start_row = 12      # Starting row for simulations name
    sim_name_col = 10       # Column J
    sim_folder_col = 9      # Column I
    filter_row = 12         # Row for filters
    branch_col = 4          # Column D
    start_col = 5           # Column E
    end_col = 6             # Column F
    limits_row = 12         # Row for limits
    amber_col = 2           # Column B
    red_col = 1             # Column A
    time_threshold_col = 3  # Column C

    result = parse_inputs_by_cells(
        excel_file, "Start",
        sim_start_row, sim_name_col, sim_folder_col,
        filter_row, branch_col, start_col, end_col,
        limits_row, amber_col, red_col, time_threshold_col
    )

    if not result:
        return False
    
    sims_df, filters, limits = result

    print(f"Simulations found ({len(sims_df)} rows):")
    print(str(sims_df))
    print("Filters:")
    print(str(filters))
    print("Limits:")
    print(str(limits))

    # Return for further processing
    return sims_df, filters, limits

# Read the Inputs sheet using fixed cell positions
def parse_inputs_by_cells(excel_path, sheet,
                          sim_start_row, sim_name_col, sim_folder_col,
                          filter_row, branch_col, start_col, end_col,
                          limits_row, amber_col, red_col, time_threshold_col):
    """
    Read simulations list, filters, and limits from the Inputs sheet using fixed cell positions.
    """
    wb = pd.ExcelFile(excel_path)
    df = pd.read_excel(wb, sheet_name=sheet, header=None)

    # --- Simulations table ---
    sims_data = []
    row_idx = sim_start_row - 1  # 0-based index
    while row_idx < len(df):
        sim_name = df.iat[row_idx, sim_name_col - 1]
        folder = df.iat[row_idx, sim_folder_col - 1]
        if pd.isna(sim_name) or pd.isna(folder):
            break
        sims_data.append({"SimName": str(sim_name).strip(), "Folder": str(folder).strip()})
        row_idx += 1
    sims_df = pd.DataFrame(sims_data)

    # --- Filters ---
    branches = []
    row_idx = filter_row - 1  # Start from the specified row (e.g., D12)
    while row_idx < len(df):
        branch = df.iat[row_idx, branch_col - 1]  # Read the value in the branch column
        if pd.isna(branch):  # Stop if the cell is empty
            break
        branches.append(str(branch).strip())  # Add the branch to the list
        row_idx += 1  # Move to the next row

    start_raw = df.iat[filter_row - 1, start_col - 1] 
    end_raw = df.iat[filter_row - 1, end_col - 1]
    start = None if pd.isna(start_raw) else start_raw
    end = None if pd.isna(end_raw) else end_raw

    # Validate start/end time filters if provided (H.MM.SS or HH.MM.SS)
    if start is not None:
        if not is_valid_time_string(start):
            SharedMethods.print_message(f"ERROR: Invalid Time (hh:mm:ss): '{start}'. Check Input...","31")
            return False
    if end is not None:
        if not is_valid_time_string(end):
            SharedMethods.print_message(f"ERROR: Invalid Time (hh:mm:ss): '{end}'. Check Input...","31")
            return False

    filters = {"branches": branches, "start": start, "end": end}

    # --- Limits ---
    amber = df.iat[limits_row - 1, amber_col - 1]
    red = df.iat[limits_row - 1, red_col - 1]
    time_threshold_s = df.iat[limits_row - 1, time_threshold_col - 1]

    if any(pd.isna(x) for x in [amber, red, time_threshold_s]):
        SharedMethods.print_message("ERROR: AmberLimit, RedLimit, and TimeThresholdSeconds are compulsory setting input.","31")
        return False

    limits = {"amber": float(amber), "red": float(red), "time_threshold_s": float(time_threshold_s)}

    return sims_df, filters, limits

# checking essential files existence
def check_essential_files(sims_df):
    # check essential files
    print('Checking essential files...')
    for _, row in sims_df.iterrows():
        folder = row['Folder']
        sim = row['SimName']
        if pd.isna(folder) or pd.isna(sim):
            SharedMethods.print_message("WARNING: Scenario List Information Not Complete. Please check.","33")
        else:
            # Adjust the require checking files below
            file_to_check = "train_list.csv"
            if not SharedMethods.folder_file_check(folder, file_to_check, True):
                SharedMethods.print_message("ERROR: Check info above. Adjust Setting or Do extraction first","31")
                return False
        
    return True

# load all train_list.csv for each simulation
def load_simulation_data(sims_df, filters):
    sim_data = {}
    for _, row in sims_df.iterrows():
        sim_name = row['SimName']
        folder = row['Folder']
        csv_path = os.path.join(folder, "train_list.csv")
        print(f"Loaded train_list.csv for {sim_name}")
        
        # Read only the columns of interest
        df = pd.read_csv(csv_path, usecols=[0, 2, 5, 12], 
                            names=['Train', 'Time', 'Voltage', 'Branch'], header=0)
        
        # Remove data points with zero voltage (neutral section)
        df = df[df['Voltage'] != 0]

        # Ensure Time is string
        df['Time'] = df['Time'].astype(str)

        # Convert h.mm.ss / hh.mm.ss  →  HH:MM:SS
        df['Time'] = (
            df['Time']
            .str.strip()
            .str.split('.', expand=True)
            .apply(lambda x: f"{int(x[0]):02d}:{int(x[1]):02d}:{int(x[2]):02d}", axis=1)
        )
        # convert to datatime format for date time
        df['Time'] = pd.to_datetime(df['Time'], format="%H:%M:%S",errors='coerce').dt.time
        
        # Apply filters if provided
        mask = pd.Series(True, index=df.index)

        # Branch filter (only if provided)
        if filters.get('branches'):
            mask &= df['Branch'].isin(filters['branches'])

        # Time filter (only if both provided)
        start = filters.get('start')
        end = filters.get('end')

        if start is not None and end is not None:
            mask &= (df['Time'] >= start) & (df['Time'] <= end)

        filtered_df = df.loc[mask]
        sim_data[sim_name] = filtered_df

    return sim_data
    
def get_unique_sorted_lists(sim_data):
    """
    Get unique and sorted lists of branches and trains from all simulation data frames.

    Args:
        sim_data (dict): Dictionary with simulation names as keys and data frames as values.

    Returns:
        tuple: Sorted list of unique branches, sorted list of unique trains.
    """
    all_branches = set()
    all_trains = set()

    for sim_name, df in sim_data.items():
        all_branches.update(df['Branch'].unique())
        all_trains.update(df['Train'].unique())

    # Convert sets to sorted lists
    sorted_branches = sorted(all_branches)
    sorted_trains = sorted(all_trains)

    return sorted_branches, sorted_trains

def assess_voltage(sim_data, limits):
    """
    Add a column to each data frame in sim_data to assess voltage levels.

    Args:
        sim_data (dict): Dictionary with simulation names as keys and data frames as values.
        limits (dict): Dictionary containing 'amber' and 'red' voltage thresholds.

    Returns:
        dict: Updated sim_data with an additional 'Assessment' column in each data frame.
    """
    print("Voltage Assessment In Progress...")
    amber_threshold = limits['amber']
    red_threshold = limits['red']

    for sim_name, df in sim_data.items():
        # Add a new column 'Assessment' based on voltage thresholds
        df['Assessment'] = df['Voltage'].apply(
            lambda v: 2 if v < red_threshold else (1 if v < amber_threshold else 0)
        )

        # SharedMethods.print_message(f"Voltage assessment added for {sim_name}", "32")

    return sim_data

def generate_train_results(sim_data, sorted_trains, time_interval):
    """
    Generate a results data frame for each train across all simulations.

    Args:
        sim_data (dict): Dictionary with simulation names as keys and data frames as values.
        sorted_trains (list): Sorted list of unique trains.

    Returns:
        pd.DataFrame: Results data frame with train statistics across simulations.
    """
    results = []

    for train in sorted_trains:
        train_row = {"Train": train}
        for sim_name, df in sim_data.items():
            # Filter data for the current train
            train_data = df[df['Train'] == train]

            if not train_data.empty:
                # Calculate statistics
                min_voltage = train_data['Voltage'].min()
                min_voltage_time = train_data.loc[train_data['Voltage'].idxmin(), 'Time']
                time_under_amber = len(train_data[train_data['Assessment'].isin([1, 2])]) * time_interval  # can be adjusted based on value set in main loop
                #amber_events = len(train_data[train_data['Assessment'] == 1])
                #red_events = len(train_data[train_data['Assessment'] == 2])

                # Update the row for this simulation
                train_row[f"{sim_name}_MinV"] = min_voltage
                train_row[f"{sim_name}_MinV_Time"] = min_voltage_time
                train_row[f"{sim_name}_TimeUnderAmber"] = time_under_amber
                #train_row[f"{sim_name}_AmberEvents"] = amber_events
                #train_row[f"{sim_name}_RedEvents"] = red_events
            else:
                # If no data for this train in the simulation, fill with NaN or default values
                train_row[f"{sim_name}_MinV"] = None
                train_row[f"{sim_name}_MinV_Time"] = None
                train_row[f"{sim_name}_TimeUnderAmber"] = 0
                #train_row[f"{sim_name}_AmberEvents"] = 0
                #train_row[f"{sim_name}_RedEvents"] = 0

        # Append the row to the results
        results.append(train_row)

    # Convert results to a DataFrame
    results_df = pd.DataFrame(results)

    return results_df

def generate_branch_results(sim_data, sorted_branches):
    """
    Generate a results data frame for each branch across all simulations.

    Args:
        sim_data (dict): Dictionary with simulation names as keys and data frames as values.
        sorted_branches (list): Sorted list of unique branches.

    Returns:
        pd.DataFrame: Results data frame with branch statistics across simulations.
    """
    results = []

    for branch in sorted_branches:
        branch_row = {"Branch": branch}
        for sim_name, df in sim_data.items():
            # Filter data for the current branch
            branch_data = df[df['Branch'] == branch]

            if not branch_data.empty:
                # Calculate statistics
                min_voltage = branch_data['Voltage'].min()
                min_voltage_train = branch_data.loc[branch_data['Voltage'].idxmin(), 'Train']
                min_voltage_time = branch_data.loc[branch_data['Voltage'].idxmin(), 'Time']
                
                # Update the row for this simulation
                branch_row[f"{sim_name}_MinV"] = min_voltage
                branch_row[f"{sim_name}_MinV_Train"] = min_voltage_train
                branch_row[f"{sim_name}_MinV_Time"] = min_voltage_time
            else:
                # If no data for this branch in the simulation, fill with NaN or default values
                branch_row[f"{sim_name}_MinV"] = None
                branch_row[f"{sim_name}_MinV_Train"] = None
                branch_row[f"{sim_name}_MinV_Time"] = None

        # Append the row to the results
        results.append(branch_row)

    # Convert results to a DataFrame
    results_df = pd.DataFrame(results)

    return results_df

def filter_results_by_amber_threshold(results_df, amber_threshold):
    """
    Filter the results data frame to include only rows where the minimum voltage is below the amber threshold.

    Args:
        results_df (pd.DataFrame): The results data frame (train or branch).
        amber_threshold (float): The amber voltage threshold.

    Returns:
        pd.DataFrame: Filtered results data frame.
    """
    # Identify columns that represent minimum voltage for simulations
    min_v_columns = [col for col in results_df.columns if col.endswith("_MinV")]

    # Filter rows where any of the MinV columns have a value below the amber threshold
    filtered_df = results_df[
        results_df[min_v_columns].apply(lambda row: any(row < amber_threshold), axis=1)
    ]

    return filtered_df

def export_results_to_excel(excel_file, train_results_df, branch_results_df, filtered_train_results_df, filtered_branch_results_df, sims_df, limits):
    """
    Export results data frames to the input Excel file with specified layout and formatting.

    Args:
        excel_file (str): Path to the Excel file.
        train_results_df (pd.DataFrame): Train results data frame.
        branch_results_df (pd.DataFrame): Branch results data frame.
        filtered_train_results_df (pd.DataFrame): Filtered train results data frame.
        filtered_branch_results_df (pd.DataFrame): Filtered branch results data frame.
        time_limit (float): Time limit for coloring TimeUnderAmber cells.
    """

    # write data to excel
    with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a',if_sheet_exists='overlay') as writer:
        # get currrent workbook to do process
        wb = writer.book
        print("Writing to Excel...")
        simulations = sims_df["SimName"].tolist()
        sheets = ["Branch Results (Highlighted)","Train Results (Highlighted)","Branch Results (All Lists)","Train Results (All Lists)"]

        # writing to each sheet
        # 1 Branch Sheet, 2 Train Sheet
        filtered_branch_results_df.to_excel(writer, sheet_name=sheets[0], index=False, startrow = 1)
        format_result_sheet(wb[sheets[0]],simulations,limits,1)
        filtered_train_results_df.to_excel(writer, sheet_name=sheets[1], index=False, startrow = 1)
        format_result_sheet(wb[sheets[1]],simulations,limits,2)
        branch_results_df.to_excel(writer, sheet_name=sheets[2], index=False, startrow = 1)
        format_result_sheet(wb[sheets[2]],simulations,limits,1)
        train_results_df.to_excel(writer, sheet_name=sheets[3], index=False, startrow = 1)
        format_result_sheet(wb[sheets[3]],simulations,limits,2)
        
        print("Saving Excel File...")

def format_result_sheet(ws,simulation_names,limits,sheet_type,header_row=2,start_data_row=3):
    """
    Docstring for format_result_sheet
    
    :param ws: Current worksheet
    :param simulation_names: Description
    :param limits: Description
    :param sheet_type: 1: Branch, 2: Train
    :param header_row: Description
    :param start_data_row: Description
    """
    max_row = ws.max_row
    max_col = ws.max_column
    table_align = Alignment(horizontal="center", vertical="center",wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    amber_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Amber color
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red color

    # center align + border
    for row in ws.iter_rows(
        min_row=header_row,
        max_row=max_row,
        min_col=1,
        max_col=max_col,
    ):
        for cell in row:
            cell.alignment = table_align
            cell.border = thin_border

    # format voltage be only intergel
    col = 2
    while col <= max_col:
        for row in range(start_data_row, max_row + 1):
            ws.cell(row=row, column=col).number_format = "0"
        col += 3
    
    # header data
    col = 2
    for sim in simulation_names:
        ws.merge_cells(start_row=header_row-1,start_column=col,end_row=header_row-1,end_column=col + 2)
        cell = ws.cell(row=header_row-1, column=col)
        cell.value = sim
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True)
        cell.border = thin_border
        col += 3
    
    # Branche Table Formatting
    if sheet_type == 1: # Branche Table Formatting
        # rename header
        col = 2  # column 1 is Branch / Train
        for sim in simulation_names:
            ws.cell(row=header_row, column=col).value = f"Mimimum Voltage (V)"
            ws.cell(row=header_row, column=col + 1).value = f"Train VO ID"
            ws.cell(row=header_row, column=col + 2).value = f"At Time (hh:mm:ss)"
            col += 3
        # conditional formatting
        col = 2
        for sim in simulation_names:
            value_col_letter = get_column_letter(col)
            # red critera
            formula = (f'{value_col_letter}{start_data_row}<{limits["red"]}')
            rule = FormulaRule(formula=[formula],fill=red_fill)
            ws.conditional_formatting.add(
                f"{value_col_letter}{start_data_row}:{value_col_letter}{max_row}",
                rule,
            )
            # yellow critera
            formula = (f'{value_col_letter}{start_data_row}<{limits["amber"]}')
            rule = FormulaRule(formula=[formula],fill=amber_fill)
            ws.conditional_formatting.add(
                f"{value_col_letter}{start_data_row}:{value_col_letter}{max_row}",
                rule,
            )
            col += 3
    # Train Sheet Formatting
    elif sheet_type == 2: # Train Table Formatting
        # rename header
        col = 2  # column 1 is Branch / Train
        for sim in simulation_names:
            ws.cell(row=header_row, column=col).value = f"Mimimum Voltage (V)"
            ws.cell(row=header_row, column=col + 1).value = f"At Time (hh:mm:ss)"
            ws.cell(row=header_row, column=col + 2).value = f"Duration < {limits['amber']}V (S)"
            col += 3
        # conditional formatting
        col = 2
        for sim in simulation_names:
            value_col_letter = get_column_letter(col)
            time_col_letter = get_column_letter(col + 2)
            # red critera
            formula = (
                f'OR({value_col_letter}{start_data_row}<{limits["red"]},'
                f'AND({value_col_letter}{start_data_row}<{limits["amber"]},'
                f'{time_col_letter}{start_data_row}>{limits["time_threshold_s"]}))'
            )
            rule = FormulaRule(formula=[formula],fill=red_fill)
            ws.conditional_formatting.add(
                f"{value_col_letter}{start_data_row}:{value_col_letter}{max_row}",
                rule,
            )
            # yellow critera
            formula = (f'{value_col_letter}{start_data_row}<{limits["amber"]}')
            rule = FormulaRule(formula=[formula],fill=amber_fill)
            ws.conditional_formatting.add(
                f"{value_col_letter}{start_data_row}:{value_col_letter}{max_row}",
                rule,
            )
            col += 3
    
    # Adjust column widths
    for col_idx in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].auto_size = True

def is_valid_time_string(val):
    """
    Validate string in HH:MM:SS format.
    And also a valid time.
    """
    if val is None or pd.isna(val):
        return False

    # Already a time object
    if isinstance(val, time):
        return True

    # datetime or pandas Timestamp
    if isinstance(val, (datetime, pd.Timestamp)):
        return True

    # String case
    if isinstance(val, str):
        try:
            datetime.strptime(val.strip(), "%H:%M:%S")
            return True
        except ValueError:
            return False

    return False


if __name__ == "__main__":
    # Add your debugging code here
    simname = "DC000"  # Provide a simulation name or adjust as needed
    main_option = "10"  # Adjust as needed
    time_start = "0070000"  # Adjust as needed
    time_end = "0080000"  # Adjust as needed
    option_select = "1"  # Adjust as needed
    text_input = "Low_V_Summary_template"  # Adjust as needed
    low_v = None  # Adjust as needed
    high_v = None  # Adjust as needed
    time_step = "1"  # Adjust as needed

    # Call your main function with the provided parameters
    main(simname, main_option, time_start, time_end, option_select, text_input, low_v, high_v, time_step)