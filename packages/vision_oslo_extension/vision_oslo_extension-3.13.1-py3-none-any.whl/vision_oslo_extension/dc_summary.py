#
# -*- coding: utf-8  -*-
#=================================================================
# Created by: Jieming Ye
# Created on: Feb 2024
# Last Modified: Feb 2024
#=================================================================
# Copyright (c) 2024 [Jieming Ye]
#
# This Python source code is licensed under the 
# Open Source Non-Commercial License (OSNCL) v1.0
# See LICENSE for details.
#=================================================================
"""
Pre-requisite: 
xxxxx.xlsx: summary spreadsheet with user configuration settings
Result saved in a proper folder structure
Used Input:
simname: to locate the result file or file rename
option_select: to define the option selected in subfunction
text_input: to locate the excel file name
Expected Output:
Updated excel spreadsheet with analysis result.
Description:
This script summarised the DC data from various simulation to one place.
First, it reads the excel file and save the user input in a data frame called start_df and understanding all related files required for summary. It then collecting data from dedicated files in dedicated folder and filter the information required for the final summary.
The format needs to work alongside with batch_processing.py
The key manual updating part is get_result_range() function which depends on the desired output format in excel, followed by table_formatting() function where some reading needs manual input. The other process should be auto adjustable as long as the excel input format is followed.

"""
#=================================================================
# VERSION CONTROL
# V1.0 (Jieming Ye) - Initial Version
#=================================================================
# Set Information Variable
# N/A
#=================================================================

# vision_oslo_extension/excel_processing.py
import pandas as pd
import numpy as np
import os
# import openpyxl
from datetime import datetime
from functools import reduce

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, NamedStyle
from openpyxl.formatting.rule import CellIsRule,FormulaRule

# import vision_oslo
from vision_oslo_extension.shared_contents import SharedMethods, SharedVariables


def main(simname, main_option, time_start, time_end, option_select, text_input, low_v, high_v, time_step):
    
    print("")
    print("DC Result Summary - - - > ")
    print("")
    
    simname = simname
    # Specify Excel file name
    excel_file = text_input + ".xlsx"

    if not SharedMethods.check_existing_file(excel_file):
        return False
    
    # Option:
    # 1: TRU Assessment
    # 2: Main DC CB (30 min RMS) - Sub
    # 3: DC Busbar (2h RMS) - Sub
    # 4: Negatvie ETE - Sub
    # 5: Impedance Bond - Sub
    # 6: Track CB (15 min RMS) - Branch
    # 7: ETE (30 min RMS) - Branch
    # 8: Min V

    option = option_select # 1:Fully Restart, require oof, and list

    start = 7 # result start from row 7
    space = 5
    # time_increment = 5
    time_windows_total= ['1min','4min','5min','15min','30min','60min','120min','180min','P30min']

    if option not in ["0","1","2","3","4","5","6","7","8"]:
        SharedMethods.print_message("ERROR: Error in dc_summary.py. Please contact Support...","31")
        return False
    
    if option == "0":
        SharedMethods.print_message("ERROR: Please select an option to proceed.","31")
        return False
    elif option == "1":
        time_windows= ['1min','4min','5min','15min','30min','60min','120min','180min','P30min']
    elif option == "2":
        time_windows= ['30min']
    elif option == "3":
        time_windows= ['120min']
    elif option == "4":
        time_windows= ['60min']
    elif option == "5":
        time_windows= ['60min','120min']
    elif option == "6":
        time_windows= ['15min']
    elif option == "7":
        time_windows= ['30min']
    elif option == "8":
        time_windows= ['MinV']
    
    # start_cell = 'B11'
    # read data from start tab
    # start_df = pd.read_excel(writer,sheet_name = "Start")
    result = start_reading_process(simname,excel_file,option)
    if result == False:
        return False
    else:
        start_df,oslo_total,scenariolist = result

    # check essential files:
    if not check_essential_files(scenariolist,option):
        return False
    
    try:
        # read all essential files: rating summary , result summary
        r_df_sum = read_all_files(start_df,oslo_total,scenariolist,time_windows,time_windows_total,option)
        t_df_sum = creat_summary_dataframe(start_df,oslo_total,scenariolist,time_windows,time_windows_total,option)

        # processing the information
        t_df_sum = result_update(start_df,oslo_total,scenariolist,time_windows,t_df_sum,r_df_sum,option)

        # check failure summary
        fail_df_sum = check_failure(t_df_sum,scenariolist,option)

        # substation group
        substation_df = substation_group(fail_df_sum,scenariolist,time_windows,option)

        # save the data
        data_write_save(simname,option,excel_file,start,space,oslo_total,scenariolist,time_windows,t_df_sum,fail_df_sum,substation_df)
    
    except KeyError as e:
        SharedMethods.print_message(f"ERROR: Unexpected error occured: {e}. Possibly due to incompleted data.","31")
        return False
    
    except Exception as e:
        SharedMethods.print_message(f"ERROR: Unexpected error occured: {e}","31")
        return False

    return True


# write data to excel
def data_write_save(simname,option,excel_file,start,space,oslo_total,scenariolist,time_windows,t_df_sum,fail_df_sum,substation_df):

    # write data to excel
    with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a',if_sheet_exists='overlay') as writer:

        # get currrent workbook to do process
        wb = writer.book
        print("Generate Failure Result Page...")
        r_start = start + 2
        substation_df.to_excel(writer, sheet_name="Result", index=False, startrow = r_start, startcol = 1)

        r_start = r_start + len(substation_df) + space
        for index, dflist in enumerate(fail_df_sum):
            dflist.to_excel(writer, sheet_name="Result", index=False, startrow = r_start, startcol = 1)
            r_start = r_start+ dflist.shape[0] + space

        # saving individual feeder
        for index, dflist in enumerate(t_df_sum):

            print(f"Saving {time_windows[index]} Result...")
            dflist.to_excel(writer, sheet_name=time_windows[index], index=False, startrow = start-1)

        # # Calculate the Excel range for post-processing
        range_list = get_result_range(start,space,oslo_total,scenariolist,fail_df_sum,substation_df,option)
        
        # # table formatting
        table_formatting(simname,wb,range_list,time_windows,scenariolist,option)

        print("Saving Excel File...")

    return

# read the start tab and collect informaiton
def start_reading_process(simname, excel_file,option):
    
    # File check
    print("Check Excel and Deleting Existing Contents ...")
    if option == "1":
        columns = ['Substation Name', 'OSLO', 'Outage Scenario', 'TRU Rating (MW)', 'TRU Type', \
                   '1min', '4min', '5min', '15min' ,'30min','60min','120min','180min', 'P30min', \
                    '1min_out', '4min_out', '5min_out', '15min_out', '30min_out', '60min_out','120min_out','180min_out','P30min_out']
        col_num = 23
    elif option == "2":
        columns = ['Substation Name', 'OSLO', 'Outage Scenario', 'DCCB Type', 'N-0 Rating (kA)','N-1 Rating (kA)']
        col_num = 6
    elif option == "3":
        columns = ['Substation Name', 'OSLO', 'Outage Scenario', 'DCBB Type', 'N-0 Rating (kA)','N-1 Rating (kA)']
        col_num = 6
    elif option == "4":
        columns = ['Substation Name', 'OSLO', 'Outage Scenario', 'Track Number', 'Rating (kA)']
        col_num = 5
    elif option == "5":
        columns = ['Substation Name', 'OSLO', 'Outage Scenario', 'Track Number', 'Bonding Type', '60min','120min']
        col_num = 7
    elif option == "6":
        columns = ['Substation Name', 'OSLO', 'Feeder ID', 'TCB Type', 'Ref Rating','Rating (kA)']
        col_num = 6
    elif option == "7":
        columns = ['Substation Name', 'OSLO', 'Cable No', 'Cable Type','Rating (kA)']
        col_num = 5
    elif option == "8":
        columns = ['OSLO', 'HeadCode', 'Origin', 'Destination','Departure Time','Rolling Stock','Voltage (V)']
        col_num = 7

    try:
        wb = load_workbook(excel_file)
        # Delete all sheets except 'Start'
        for sheet_name in wb.sheetnames:
            if sheet_name == 'Start':
                result = check_reading_frame(wb[sheet_name],option) # list, dictionary type
                if result == False:
                    return False
                else:
                    data_start_row,data_end_row,oslo_total,scenariolist = result
                
                if option == "8":
                    start_df = pd.read_excel(excel_file,sheet_name = 'Start',header = 0,usecols=range(col_num), skiprows=data_start_row-1, nrows=oslo_total, names = columns, dtype={columns[0]: str, columns[1]: str})
                else:
                    start_df = pd.read_excel(excel_file,sheet_name = 'Start',header = 0,usecols=range(col_num), skiprows=data_start_row-1, nrows=oslo_total, names = columns)
                    oslolist = start_df.iloc[:, 1].tolist()
                    outlist =  start_df.iloc[:, 2].tolist()

                    # check duplication in OSLO id
                    if not SharedMethods.find_duplicates(oslolist): return False

                flag = False
                # check oslo list 
                if not option == "8": # do not check option 8 as the format is very different
                    for index, oslo in enumerate(oslolist):
                        if pd.isna(oslo):
                            SharedMethods.print_message(f"ERROR: Items {start_df.iloc[index,0]} is not assigned with OSLO ID","31")
                            flag = True
                
                # Check if all numbers in outlist are in the keys of scearniolist only for 1-5 with regard to Sub
                if option in ["1","2","3","4","5"]:
                    for index, out in enumerate(outlist):
                        if pd.isna(out):
                            SharedMethods.print_message(f"WARNING: Item {start_df.iloc[index,1]} is assigned to Zero due to no outage mentioned.","33")
                            start_df.iloc[index,2] = 0.0
                        elif out not in scenariolist:
                            SharedMethods.print_message(f"ERROR: Outage Scenario {out} is not covered in scenario list. Check Data Input","31")
                            flag = True
                
                # check outgoing track feeder number for option 4:
                if option in ["4","5"]:
                    track_list = start_df.iloc[:, 3].tolist()
                    for trackno in track_list:
                        if pd.isna(trackno):
                            SharedMethods.print_message(f"WARNING: Item {start_df.iloc[index,1]} is assigned to Track Number 1 due to no Track Number mentioned.","33")
                            start_df.iloc[index,3] = 1
                        elif not isinstance(trackno, (int, float)) or trackno == 0:
                            SharedMethods.print_message(f"ERROR: Track Number at {start_df.iloc[index,1]} is NOT an Number or Zero. Check Data Input","31")
                            flag = True
                
                # decision point
                if flag:
                    return False

            else:
                wb.remove(wb[sheet_name])
        
        # Save changes
        wb.save(excel_file)
        

    except Exception as e:
        SharedMethods.print_message(f"ERROR: (Close the Excel file and Start Again) Error: {e}","31")
        return False
    
    return start_df,oslo_total,scenariolist

# rating data frame
def check_reading_frame(sheet,option):
    table_start_row = 11

    # define the outage list section to be read # Compare against function above
    print("Check Data Entry & Scenario List...")
    if option == "1":
        table_row = 12
        table_start_column = 2
        # table_end_column = 11
        table_row_2 = 12
        table_start_column_2 = 26
    elif option in ["2","3","6"]:
        table_row = 12
        table_start_column = 2
        table_row_2 = 12
        table_start_column_2 = 9
    elif option in ["4","7"]:
        table_row = 12
        table_start_column = 2
        table_row_2 = 12
        table_start_column_2 = 8
    elif option in ["5","8"]:
        table_row = 12
        table_start_column = 2
        table_row_2 = 12
        table_start_column_2 = 10
    
    # create ole list and rating data
    # find total number of substations
    index = table_row
    column = table_start_column
    if sheet.cell(row=index, column=column).value is not None:
        while True:          
            index += 1
            check = sheet.cell(row=index, column=column).value
            if check is None:
                table_row_end = index
                oslo_total = index - table_row
                break
    else:
        SharedMethods.print_message("ERROR: Wrong data format. No information at B12","31")
        return False
    
    # reading scearnio list
    scenariolist = {}
    index = table_row_2
    column = table_start_column_2
    if sheet.cell(row=index, column=column).value is not None:
        while True:
            s_temp = sheet.cell(row=index, column=column).value
            sim_temp = sheet.cell(row=index, column=column+1).value
            file_temp = sheet.cell(row=index, column=column+2).value
            scenariolist[s_temp] = [sim_temp, file_temp]
            
            index += 1
            check = sheet.cell(row=index, column=column).value
            if check is None:
                break
        # Check if key '0' is not in scenariolist and add it to the beginning
        if 0 not in scenariolist:
            scenariolist[0] = [None, None]

    else:
        SharedMethods.print_message("ERROR: Wrong data format. No information at Outage List Row 12","31")
        return False
    
    return table_start_row,table_row_end,oslo_total,scenariolist

# checking essential files existance
def check_essential_files(scenariolist,option):
    
    # check essential files
    print('Checking essential files...')
    for out,value in scenariolist.items():
        folder = value[0]
        sim = value[1]
        if folder == None or sim == None:
            SharedMethods.print_message("WARNING: Scenario List Information Not Complete. Please check.","33")
        else:
            # Adjust the require checking files below.
            if option in ["1","2","3","4","5"]:
                sim = sim + "_RMSCurrent_Sum.csv"
            elif option == "6":
                sim = sim + "_branch_900_rms_sum_max.csv"
            elif option == "7":
                sim = sim + "_branch_1800_rms_sum_max.csv"
            elif option == "8":
                sim = sim + ".osop.vlt"
            else:
                return False

            if not SharedMethods.folder_file_check(folder,sim,True):
                SharedMethods.print_message("ERROR: Check info above. Adjust Setting or Do extraction first","31")
                return False
        
    return True

# reading essential files and save
def read_all_files(start_df,oslo_total,scenariolist,time_windows,time_windows_total,option):
    if option == "1":
        result_header = ['OSLO'] + time_windows
        csv_columns = list(range(1,len(result_header)+1))
    elif option in ["2","3","4"]:
        result_header = ['OSLO'] + time_windows
        index_finder = time_windows_total.index(time_windows[0]) # assumed only 1 time in time slots
        csv_columns = [1,index_finder+2]
    elif option == "5":
        result_header = ['OSLO'] + time_windows
        temp = []
        for time in time_windows:
            index_finder = time_windows_total.index(time) # multiple time in time slots
            temp.append(index_finder+2)
        csv_columns = [1] + temp
    elif option in ["6","7"]:
        result_header = ['OSLO'] + time_windows # Assum only 1 time in time slots
        csv_columns = [1,4] # OSLO branch Node + maximum value only
    elif option == "8":
        result_header = ['OSLO','MinV'] # Assum only 1 time in time slots
        csv_columns = [0,6] # OSLO branch Node + maximum value only

    
    r_df_sum = [] # result summary
    
    # Reading result file in each folder
    for key,value in scenariolist.items():
        folder = value[0]
        sim = value[1]
        if folder == None or sim == None:
            SharedMethods.print_message("WARNING: Scenario List Information Not Complete. Skipping...","33")
        else:
            if option in ["1","2","3","4","5"]:
                sim = sim + "_RMSCurrent_Sum.csv"
                
            elif option == "6":
                sim = sim + "_branch_900_rms_sum_max.csv"
            
            elif option == "7":
                sim = sim + "_branch_1800_rms_sum_max.csv"
            
            if not option == "8":
                # if in double, move this under subtaiton section and do customised one for each option
                file_path = os.path.join(os.getcwd(),folder,sim)
                #result_df = pd.read_csv(file_path,usecols=csv_columns,nrows=oslo_total,header = None, skiprows=2, names= result_header)
                result_df = pd.read_csv(file_path,usecols=csv_columns,header = None, skiprows=2, names= result_header)
                if option == "1": # For TRU assessment, the last column is power in MW
                    # Divide all columns by 1000 except the first column and last column (updated to consider the P30min power)
                    result_df.iloc[:, 1:-1] = result_df.iloc[:, 1:-1].divide(1000)
                else:
                    # Divide all columns by 1000 except the first column
                    result_df.iloc[:, 1:] = result_df.iloc[:, 1:].divide(1000)
                r_df_sum.append(result_df)
            else:
                sim = sim + ".osop.vlt"
                file_path = os.path.join(os.getcwd(),folder,sim)

                # Define column widths based on your data
                # (TODO): I don't know the best way to validate which version quickly without reading the file in detail yet
                if SharedVariables.osop_version in [1,2]:
                    SharedMethods.print_message("ATTENTION: Result extracted from RN28 or earlier... Change configuration settings if not correct...","33")
                    col_widths = [4, 1, 4, 11, 8, 2, 9]  # Replace ... with the widths of your columns
                else: # from RN29 it uses 5 digit train number
                    SharedMethods.print_message("ATTENTION: Result extracted from RN29... Change configuration settings if not correct...","33")
                    col_widths = [5, 1, 4, 10, 8, 2, 9]
                # Read the file using fixed width format
                df = pd.read_fwf(file_path, widths=col_widths, header=None, skiprows=17)
                # Remove rows where the first element is not a number
                df = df[pd.to_numeric(df[0], errors='coerce').notnull()]
                # convert back to str
                df[0] = df[0].astype(str)
                # Convert the fourth column to numeric / voltage
                df[4] = pd.to_numeric(df[4], errors='coerce')
                # Multiply values in the fourth column by 1000
                df[4] *= 1000
                result_df = df[[0,4]]
                result_df.columns =  result_header
                # convert time in to time format
                # df['Time'] = df['Time'].apply(lambda x: f"{int(x[1:3]):02d}:{int(x[3:5]):02d}:{int(x[5:]):02d}")
                r_df_sum.append(result_df)
    
    return r_df_sum

# create summary dataframe for result
def creat_summary_dataframe(start_df,oslo_total,scenariolist,time_windows,time_windows_total,option):

    t_df_sum = [] # time summary

    # define final result tab format
    for time in time_windows:
        print(f'Processing {time} information in all folders...')
        if option == "1":
            columns = ['OSLO','Substation','Outage in Scenario No.','N-0 Rating (kA)','N-1 Rating (kA)']
            for key in scenariolist.keys():
                out = "outage_scenario_" + str(key)
                columns.append(out)
            t_df = pd.DataFrame(columns=columns) #             
            # copy essential info
            t_df['OSLO'] = start_df['OSLO']
            t_df['Substation'] = start_df['Substation Name']
            t_df['Outage in Scenario No.'] = start_df['Outage Scenario']
            # copy N-0 rating        
            t_df['N-0 Rating (kA)'] = start_df[time]
            time1 = time + '_out' # required due to the duplication of titile name
            t_df['N-1 Rating (kA)'] = start_df[time1]
        
        elif option in ["2","3"]:
            columns = ['OSLO','Substation','Outage in Scenario No.','N-0 Rating (kA)','N-1 Rating (kA)']
            for key in scenariolist.keys():
                out = "outage_scenario_" + str(key)
                columns.append(out)
            t_df = pd.DataFrame(columns=columns)
            # copy essential info
            t_df['OSLO'] = start_df['OSLO']
            t_df['Substation'] = start_df['Substation Name']
            t_df['Outage in Scenario No.'] = start_df['Outage Scenario']
            # copy N-0 rating        
            t_df['N-0 Rating (kA)'] = start_df['N-0 Rating (kA)']
            t_df['N-1 Rating (kA)'] = start_df['N-1 Rating (kA)']

        elif option == "4":
            columns = ['OSLO','Substation','Outage in Scenario No.','Track Number','Rating (kA)']
            for key in scenariolist.keys():
                out = "outage_scenario_" + str(key)
                columns.append(out)
            t_df = pd.DataFrame(columns=columns)
            # copy essential info
            t_df['OSLO'] = start_df['OSLO']
            t_df['Substation'] = start_df['Substation Name']
            t_df['Outage in Scenario No.'] = start_df['Outage Scenario']     
            t_df['Track Number'] = start_df['Track Number']
            t_df['Rating (kA)'] = start_df['Rating (kA)']
        
        elif option == "5":
            columns = ['OSLO','Substation','Outage in Scenario No.','Track Number','Bonding Type','Rating (kA)']
            for key in scenariolist.keys():
                out = "outage_scenario_" + str(key)
                columns.append(out)
            t_df = pd.DataFrame(columns=columns)
            # copy essential info
            t_df['OSLO'] = start_df['OSLO']
            t_df['Substation'] = start_df['Substation Name']
            t_df['Outage in Scenario No.'] = start_df['Outage Scenario']     
            t_df['Track Number'] = start_df['Track Number']
            t_df['Bonding Type'] = start_df['Bonding Type']
            t_df['Rating (kA)'] = start_df[time]
        
        elif option == "6":
            columns = ['OSLO','Substation','Track CB Type','Rating (kA)']
            for key in scenariolist.keys():
                out = "outage_scenario_" + str(key)
                columns.append(out)
            t_df = pd.DataFrame(columns=columns)
            # copy essential info
            t_df['OSLO'] = start_df['OSLO']
            t_df['Substation'] = start_df['Substation Name']
            # t_df['Outage in Scenario No.'] = start_df['Outage Scenario']     
            t_df['Track CB Type'] = start_df['TCB Type']
            t_df['Rating (kA)'] = start_df['Rating (kA)']
        
        elif option == "7":
            columns = ['OSLO','Location','Cable Number','Rating (kA)']
            for key in scenariolist.keys():
                out = "outage_scenario_" + str(key)
                columns.append(out)
            t_df = pd.DataFrame(columns=columns)
            # copy essential info
            t_df['OSLO'] = start_df['OSLO']
            t_df['Location'] = start_df['Substation Name']
            #t_df['Outage in Scenario No.'] = start_df['Outage Scenario']     
            t_df['Cable Number'] = start_df['Cable No']
            t_df['Rating (kA)'] = start_df['Rating (kA)']
        
        elif option == "8":
            columns = ['OSLO', 'HeadCode', 'Origin', 'Destination','Departure Time','Rolling Stock','Voltage Threshod (V)'] 
            for key in scenariolist.keys():
                out = "outage_scenario_" + str(key)
                columns.append(out)
            t_df = pd.DataFrame(columns=columns)
            # copy essential info
            t_df['OSLO'] = start_df['OSLO']
            t_df['HeadCode'] = start_df['HeadCode']  
            t_df['Origin'] = start_df['Origin']
            t_df['Destination'] = start_df['Destination']
            t_df['Departure Time'] = start_df['Departure Time']
            t_df['Rolling Stock'] = start_df['Rolling Stock']
            t_df['Voltage Threshod (V)'] = start_df['Voltage (V)']

        t_df_sum.append(t_df)

    return t_df_sum

# update result table based on result
def result_update(start_df,oslo_total,scenariolist,time_windows,t_df_sum,r_df_sum,option):
    # r_df
    #     OSLO         1min         4min         5min        30min        60min  120min  180min
    # 0    PRLY  6838.572292  4319.651351  4216.858674  3180.263272  3047.397089     NaN     NaN
    # 1    COUL  3957.541043  3088.845394  2958.794839  2281.910857  2199.222286     NaN     NaN
    # 2    SHEP  3151.558321  2340.795600  2306.909161  1668.882423  1581.344310     NaN     NaN
    # t_df
    # OSLO    Substation  Outage in Scenario No.  N-0 Rating (kA)  N-1 Rating (kA) outage_scenario_0 outage_scenario_1 outage_scenario_2
    # 0  COUL  Substation A           1        80.000000        70.000000      NaN      NaN      NaN
    # 1  SHEP  Substation A           1        61.264000        40.842667      NaN      NaN      NaN
    # 2  HOLE  Substation A           2        40.842667        20.421333      NaN      NaN      NaN
    # 3  REDA  Substation A           2        30.632000        15.316000      NaN      NaN      NaN
    print('Updating the Final Result...')
    
    for index, df in enumerate(t_df_sum):
        time = time_windows[index]
        rindex = 0 # result index
        for key,value in scenariolist.items():
            folder = value[0]
            sim = value[1]
            if folder == None or sim == None:
                pass
                #print("WARNING: Scenario List Information Not Complete. Skipping...")
            else:            
                # merge dataframe sbased on the OSLO columns
                merged_df = pd.merge(df, r_df_sum[rindex][['OSLO', time]], left_on='OSLO', right_on='OSLO', how='left') # joing to left, left join
                rindex = rindex + 1
                # update outage_scenario column
                out = 'outage_scenario_' + str(key)
                # update the column in df based on r_df
                df[out] = merged_df[time]
                
        # update the result summary
        t_df_sum[index] = df
    
    # update the result based on specific quesiton
    # Option 4 negative ETE
    if option in ["4","5"]:
        for df in t_df_sum:
            for out in scenariolist.keys():
                if out == 0:
                    df['outage_scenario_'+str(out)] = df['outage_scenario_'+str(out)] / (2* df['Track Number']-1)
                else:
                    df['outage_scenario_'+str(out)] = df['outage_scenario_'+str(out)] / (2* df['Track Number'])
    
    
    return t_df_sum

# check failure and return dataframe of failures:
def check_failure(t_df_sum,scenariolist,option):
    print('Checking Failure Summary...')
    fail_df_sum = []
    if option in ["1","2","3","4"]: # check how many columns used in section create summary. (info columns)
        remain = 5
    elif option in ["5"]:
        remain = 6
    elif option in ["6","7"]:
        remain = 4
    elif option in ["8"]:
        remain = 7

    for df in t_df_sum:
        new_rows = []
        # Iterate over rows using index
        for index in range(len(df)):
            row = df.iloc[index]
            condition = False

            for scn, (key, value) in enumerate(scenariolist.items()):
                if option in ["4","5","6","7"]: # single rating assessment
                    rating_column = 'Rating (kA)'
                    # if row.iloc[scn + remain] > row[rating_column]: # use > to avoid zero rating issue
                    if row.iloc[scn + remain] > 0.9 * row[rating_column]: # use > to avoid zero rating issue
                        condition = True
                        break  # No need to continue checking other scenarios
                    
                elif option == "8": # voltage speical
                    rating_column = 'Voltage Threshod (V)'
                    # if row.iloc[scn + remain] < row[rating_column]: # voltage
                    if row.iloc[scn + remain] < 488: # voltage get 488 out
                        condition = True
                        break  # No need to continue checking other scenarios
                
                else:
                    if key == row['Outage in Scenario No.'] and key != 0:
                        rating_column = 'N-1 Rating (kA)'
                    else:
                        rating_column = 'N-0 Rating (kA)'       
                    # if row.iloc[scn + remain] > row[rating_column]: # use > to avoid zero rating issue
                    if row.iloc[scn + remain] > 0.9 * row[rating_column]: # use > to avoid zero rating issue
                        condition = True
                        break  # No need to continue checking other scenarios

            if condition:
                new_rows.append(row)

        # Create a new DataFrame at the end
        new_df = pd.DataFrame(new_rows, columns=df.columns)

        # Append the new_df to fail_df_sum
        fail_df_sum.append(new_df)


    return fail_df_sum

# group by substations:
def substation_group(fail_df_sum,scenariolist,time_windows,option):
    if option in ["1","2","3"]:
        name = 'Substation'
    elif option in ["4","5","6","7","8"]:
        df = pd.DataFrame(columns=['See Below','This Table is left Blank.'])
        return df
    else:
        name = 'Substation'

    print(f'{name} Summary Analysis...')

    df_0 = [] # N-0 Summary
    df_1 = [] # N-1 Summary
    df_0_0 = [] # Individual time windows N-0
    df_0_1 = [] # Indivisual time windows N-1
    # create df for each scenario
    for key in scenariolist.keys():
        title = 'outage_scenario_' + str(key)

        if key == 0:
            for index, df in enumerate(fail_df_sum):
                time = time_windows[index]            
                lst = []
                for index1, value in df[title].items():
                    rating = df.loc[index1, 'N-0 Rating (kA)']
                    # if not pd.isna(rating) and value >= rating:
                    if not pd.isna(rating) and value >= 0.9 * rating:
                        oslo = df.loc[index1, 'OSLO']
                        sub = df.loc[index1, name]
                        if rating == 0:
                            perc = value
                        else:
                            perc = value / rating
                        lst.append([oslo,sub,perc])                
                new_df = pd.DataFrame(lst,columns = ['OSLO',name,time])
                df_0_0.append(new_df)
        else:
            for index, df in enumerate(fail_df_sum):
                time = time_windows[index]            
                lst_0 = []
                lst_1 = []
                for index1, value in df[title].items():
                    rating_0 = df.loc[index1, 'N-0 Rating (kA)']
                    rating_1 = df.loc[index1, 'N-1 Rating (kA)']
                    outage = df.loc[index1, 'Outage in Scenario No.']

                    # if not pd.isna(rating_0) and outage != key and value >= rating_0:
                    if not pd.isna(rating_0) and outage != key and value >= 0.9 * rating_0:
                        oslo = df.loc[index1, 'OSLO']
                        sub = df.loc[index1, name]
                        if rating_0 == 0:
                            perc = value
                        else:
                            perc = value / rating_0
                        
                        lst_0.append([oslo,sub,perc])

                    # if not pd.isna(rating_1) and outage == key and value >= rating_1:
                    if not pd.isna(rating_1) and outage == key and value >= 0.9 * rating_1:
                        oslo = df.loc[index1, 'OSLO']
                        sub = df.loc[index1, name]
                        if rating_1 == 0:
                            perc = value
                        else:
                            perc = value / rating_1
                        lst_1.append([oslo,sub,perc])
                
                # N-0 result  under N-1 outage
                new_df_0 = pd.DataFrame(lst_0,columns = ['OSLO',name,time])
                df_0_1.append(new_df_0)
                        
                # attached N-1
                new_df_1 = pd.DataFrame(lst_1,columns = ['OSLO',name,time])
                df_1.append(new_df_1)

    # dataframe merging
    # mergeing N-0 summary first and finding the maximum
    total = len(time_windows)
    for index1, time in enumerate(time_windows):               
        df_0.append(df_0_0[index1])
    t = 0
    
    for index, key in enumerate(scenariolist.keys()):
        if key != 0:
            for index1, time in enumerate(time_windows):
                df_0_temp = pd.merge(df_0[index1], df_0_1[total*t+index1], left_on='OSLO', right_on='OSLO', how='outer') # merged two 
                df_0_temp[name] = df_0_temp[name+'_x'].combine_first(df_0_temp[name+'_y']) # find the substation name
                df_0_temp[time] = df_0_temp[[time+'_x', time+'_y']].max(axis=1) # Find Maximum
                df_0_temp = df_0_temp.drop(columns=[time+'_x', time+'_y',name +'_x',name +'_y'])
                
                df_0[index1] = df_0_temp
            t = t + 1

    # Merge all DataFrames in df_0 based on 'OSLO' and keep each DataFrame's column 'time'
    merged_df_0 = reduce(lambda left, right: left.set_index('OSLO').combine_first(right.set_index('OSLO')).reset_index(), df_0)
    # Move 'Substation' column to the 2nd position
    substation_column = merged_df_0.pop(name)
    merged_df_0.insert(1, name, substation_column)

    if df_1 == []: # if there is no N-1 scearnio input
        merged_df_1 = pd.DataFrame(columns = ['OSLO',name] + time_windows)

    else:
        merged_df_1 = reduce(lambda left, right: left.set_index('OSLO').combine_first(right.set_index('OSLO')).reset_index(), df_1)
        substation_column = merged_df_1.pop(name)
        merged_df_1.insert(1, name, substation_column)

    # Merge N-0 and N-1 based on 'OSLO' and keep all time windows from both DataFrames
    merged_df = pd.merge(merged_df_0, merged_df_1, left_on='OSLO', right_on='OSLO', how='outer', suffixes=('_N-0', '_N-1')) # check how works

    # Combine 'Substation_N-0' and 'Substation_N-1' into a single 'Substation' column
    merged_df[name] = merged_df[name + '_N-0'].combine_first(merged_df[name + '_N-1'])

    # Drop the specified columns
    merged_df = merged_df.drop(columns=['Substation_N-0','Substation_N-1'])

    #Rearrange the sequence
    time_0 = [time + '_N-0' for time in time_windows]
    time_1 = [time + '_N-1' for time in time_windows]
    columns = ['OSLO',name] + time_0 + time_1
    merged_df = merged_df[columns]

    return merged_df

# check the range
def get_result_range(start,space,oslo_total,scenariolist,fail_df_sum,substation_df,option):

    if option in ["1","2","3","4"]:
        fix_width = 5 # 5 default columns [id, sub,out, n-0,n-1]
    elif option == "5":
        fix_width = 6
    elif option in ["6","7"]:
        fix_width = 4
    elif option in ["8"]:
        fix_width = 7

    
    range_list = []
    columns = fix_width + len(scenariolist)
    # sub_width = 2 + len(fail_df_sum) *2
    sub_total = len(substation_df)
    sub_column_start = 64 + 2
    sub_column_end = 64 + 3 + len(fail_df_sum) *2

    # 0: indivisual table frame range
    # 1: individual table data range (2 digit range)
    # 2: indivisual title range 


    # 3: result table frame range(multiple)
    # 4: result 2 digit range
    # 5: result title range

    # 6: Result  title location 
    # 7: Result  conditional formatting location [[start_row, end_row, start_column, end_column],[],[],...]

    # 8: individual  conditional formatting location [start_row, end_row, start_column, end_column]

    # 9: Percentage range

    # 0
    result_range = [f"{chr(64 + 1)}{start}:{chr(64 + columns)}{start+oslo_total}"]
    range_list.append(result_range)

    # 1
    data_range = [f"{chr(64 + 1 + fix_width - 2)}{start+1}:{chr(64 + columns)}{start+oslo_total}"]
    range_list.append(data_range)
   
    # 2
    title_range = [f"{chr(64 + 1)}{start}:{chr(64 + columns)}{start}"]
    range_list.append(title_range)

    # 3 4 5
    range_list.append([])
    range_list.append([])
    range_list.append([])
    r_start = start + 2 + sub_total + space
    for index, dflist in enumerate(fail_df_sum):
        length = dflist.shape[0]
        if not length == 0:
            range_list[3].append(f"{chr(64 + 2)}{r_start+1}:{chr(64 + 1 + columns)}{r_start+1+length}")
            range_list[4].append(f"{chr(64 + 2 + fix_width - 2)}{r_start+2}:{chr(64 + 1 + columns)}{r_start+1+length}")
        
        range_list[5].append(f"{chr(64 + 2)}{r_start+1}:{chr(64 + 1 + columns)}{r_start+1}")
        r_start = r_start + length + space
    
    range_list[3].append(f"{chr(sub_column_start)}{start + 3}:{chr(sub_column_end)}{start + 3 + sub_total}") # substation summary
    range_list[5].append(f"{chr(sub_column_start)}{start + 3}:{chr(sub_column_end)}{start + 3}") # substation summary
        
    #6
    # result title
    range_list.append([])
    r_start = start + 2 + sub_total + space
    for index, dflist in enumerate(fail_df_sum):
        length = dflist.shape[0]
        a = f"{chr(64 + 2)}{r_start-1}" # summary title
        b = f"{chr(64 + 2 + 1)}{r_start-1}" # xxx min rms curret
        c = f"{chr(64 + 2)}{r_start}" # result
        d = f"{chr(64 + 2 + 1)}{r_start}" # pass / fail

        if not length == 0:
            e = 'Failure Substation as Table Below'
        else:
            e = 'Assessment All Pass'

        r_start = r_start + length + space
        range_list[6].append([a,b,c,d,e])

    
    # 7 result table  location
    range_list.append([])
    r_start = start + 2 + sub_total + space
    start_row, end_row = r_start+2, r_start+1+length
    start_col, end_col = 2+fix_width, 1+columns
    
    for index, dflist in enumerate(fail_df_sum):
        length = dflist.shape[0]
        if not length == 0:
            start_row, end_row = r_start+2, r_start+1+length
            start_col, end_col = 2+fix_width, 1+columns
            range_list[7].append([start_row,end_row,start_col,end_col])
        
        r_start = r_start + length + space

    
    # 8 individual section
    start_row, end_row = start+1, start+oslo_total
    start_col, end_col = 1+fix_width, columns
    range_list.append([start_row,end_row,start_col,end_col])

    # 9 percetage section
    if sub_total == 0:
        sub_total = 1 # aovid emtpry range formatting
    data_range = [f"{chr(sub_column_start+2)}{start + 4}:{chr(sub_column_end)}{start + 3 + sub_total}"]
    range_list.append(data_range)


    return range_list

# Result table table formating
def table_formatting(simname,wb,range_list,time_windows,scenariolist,option):
    if option == "1":
        name = 'Substation'
        assess = 'TRU'
    elif option == "2":
        name = 'Substation'
        assess = 'Main DC Circuit Breaker'
    elif option == "3":
        name = 'Substation'
        assess = 'Main DC Busbar'
    elif option == "4":
        name = 'Substation'
        assess = 'Negative Electrical Track Equipment'
    elif option == "5":
        name = 'Substation'
        assess = 'Impedance Bonding'
    elif option == "6":
        name = 'Branch'
        assess = 'Track Feeder Circuit Breaker'
    elif option == "7":
        name = 'Branch'
        assess = 'Postive Electrical Track Equipment'
    elif option == "8":
        name = 'Train'
        assess = 'Minimum Voltage'


    print("Formatting Process ...")
    # wb = load_workbook(excel_file)
    for time in time_windows:
        print(f"Formatting {time} Result...")
        sheet = wb[time]

        sheet['A2'].value = "Result:"
        sheet['B2'].value = name + " " + time + " RMS Current Summary"
        sheet['A3'].value = "Pass:"
        sheet['A4'].value = "Failure:"
        
        
        # Define a custom style for formatting with two decimal places
        for range_name in range_list[0]+range_list[1]:
            for row in sheet[range_name]:
                for cell in row:
                    # Apply border to all sides of the cell
                    cell.border = Border(left=Side(border_style='thin'),
                                        right=Side(border_style='thin'),
                                        top=Side(border_style='thin'),
                                        bottom=Side(border_style='thin'))
                    
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        
        #print("Apply Numbering Format ...")
        # Define a custom style for formatting with two decimal places
        for range_name in range_list[1]:
            for row in sheet[range_name]:
                for cell in row:
                    cell.number_format = '0.00'

        #print("Apply Font and Shading...")
        # Shade the range B11:Z11 with a light gray background color
        for range_name in range_list[2]:
            for row in sheet[range_name]:
                for cell in row:
                    cell.font = Font(bold = True, italic = True, size = 11)
                    cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        condtional_formating(sheet,range_list,scenariolist,1,option)

        if time[0] == "P":
            sheet['B2'].value = name + " " + time + " Average Power Summary"
            sheet['D7'].value = "N-0 Rating (MW)"
            sheet['E7'].value = "N-1 Rating (MW)"

        # Auto-size columns after applying formatting
        for col_letter in ['A','B',"C","D"]:
            sheet.column_dimensions[col_letter].auto_size = True
        
        if option == "8":
            sheet.column_dimensions["E"].auto_size = True
            sheet.column_dimensions["F"].auto_size = True

    
    # format the result table
    sheet = wb['Start']
    project_name = sheet['B2'].value
    feeding_desp = sheet['B4'].value
    modeller = sheet['B5'].value
    date = sheet['B6'].value

    # Result Tab process
    sheet = wb["Result"]
    sheet['B2'].value = "Project Name:"
    sheet['C2'].value = project_name
    sheet['B3'].value = "Simulation Name:"
    sheet['C3'].value = simname
    sheet['B4'].value = "Feeding Arrangement:"
    sheet['C4'].value = feeding_desp
    sheet['B5'].value = "Result Created by:"
    sheet['C5'].value = modeller
    sheet['B6'].value = "Result Created at:"
    sheet['C6'].value = datetime.now().strftime("%d-%m-%Y %H:%M")

    sheet['B9'].value = "Summary Title:"
    cell = sheet['B9']
    cell.font = Font(bold = True, size = 11)
    sheet['C9'].value = "Failure " + name + " Summary"

    # sheet['H4'].value = "Note:"
    # sheet['H5'].value = 
    # sheet['H6'].value = 

    # Define a custom style for formatting with two decimal places
    for range_name in range_list[3]+range_list[4]:
        for row in sheet[range_name]:
            for cell in row:
                # Apply border to all sides of the cell
                cell.border = Border(left=Side(border_style='thin'),
                                    right=Side(border_style='thin'),
                                    top=Side(border_style='thin'),
                                    bottom=Side(border_style='thin'))
                
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # applying title
    print("Apply Title ...")

    for index, time in enumerate(time_windows):
        cell = sheet[range_list[6][index][0]]
        cell.value = "Summary Title"
        cell.font = Font(bold=True)

        cell = sheet[range_list[6][index][1]]
        cell.value = time + " " + assess + " Assessment (RMS Current)"

        cell = sheet[range_list[6][index][2]]
        cell.value = "Result"
        cell.font = Font(bold=True)

        cell = sheet[range_list[6][index][3]]
        cell.value = range_list[6][index][4]

        if time[0] == "P": # time
            cell = sheet[range_list[6][index][1]]
            cell.value = time + " " + assess + " Assessment (Average Power) - Note that the Rating Unit Below should be MW"
    
    print("Apply Numbering Format ...")
    # Define a custom style for formatting with two decimal places
    for range_name in range_list[4]:
        for row in sheet[range_name]:
            for cell in row:
                cell.number_format = '0.00'
    
    for range_name in range_list[9]:
        for row in sheet[range_name]:
            for cell in row:
                cell.number_format = '0.00%'


    print("Apply Font and Shading...")
    # Shade the range B11:Z11 with a light gray background color
    for range_name in range_list[5]:
        for row in sheet[range_name]:
            for cell in row:
                cell.font = Font(bold = True, italic = True, size = 11)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    print("Apply Conditional Formatting ...")
    condtional_formating(sheet,range_list,scenariolist,2,option)
    
    #print("Apply Column Length ...")
    # Auto-adjust the width of column B based on the content in B2 to B6
    max_length = max(len(str(sheet.cell(row=i, column=2).value)) for i in range(2, 8))
    sheet.column_dimensions['B'].width = max_length + 2  # Add a little extra space

    # Auto-size columns after applying formatting
    for index in range(2,5):
        col_letter = f'{chr(64 + index)}'
        sheet.column_dimensions[col_letter].auto_size = True
    
    if option == "8":
        sheet.column_dimensions["E"].auto_size = True
        sheet.column_dimensions["F"].auto_size = True
        sheet.column_dimensions["G"].auto_size = True


    return

# conditional formatting
def condtional_formating(sheet,range_list,scenariolist,compare_op,option):
    # Compare values in columns I and J for each row and shade accordingly
        # set the pattern fill
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red (0% S)
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 255,152,51 (80% S)
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green (00% S)

    key = list(scenariolist.keys())

    if option == "8": # train min voltage
        if compare_op == 2:
            checking = range_list[7]
        else:
            checking = [range_list[8]]
        
        for index, lst in enumerate(checking):
            for r in range(lst[0],lst[1]+1):
                for c in range(lst[2],lst[3]+1):
                    cell = sheet.cell(row = r, column = c) # result
                    compare = sheet.cell(row = r, column = lst[2]-1) # N-0 / Rating All the time
                    if cell.value != '' and compare.value != '':
                        if cell.value < compare.value:
                            cell.fill = red_fill
                        elif cell.value <= 488: # Yellow if the votlage is below 488
                            cell.fill = yellow_fill

    else: # below if for current related
        if compare_op == 2:
            checking = range_list[7]
            # general
            sheet.conditional_formatting.add(range_list[9][0],CellIsRule(operator = 'greaterThanOrEqual',formula=['1'],fill=red_fill))
            sheet.conditional_formatting.add(range_list[9][0],CellIsRule(operator = 'between',formula=['0.9','1'],fill=yellow_fill))
        else:
            checking = [range_list[8]]
            
        for index, lst in enumerate(checking):
            for r in range(lst[0],lst[1]+1):
                for c in range(lst[2],lst[3]+1):
                    cell_c = sheet.cell(row = r, column = lst[2]-3) # outage scearnio
                    cell = sheet.cell(row = r, column = c) # result

                    if option in ["1","2","3"]: # TRU, Main CB, DCBB                
                        if cell_c.value == key[c-lst[2]] and cell_c.value != 0:
                            compare = sheet.cell(row = r, column = lst[2]-1) # Compare with N-1 rating if outage = selected outage
                        else:
                            compare = sheet.cell(row = r, column = lst[2]-2) # Compare with N-0 rating
                    else:
                        compare = sheet.cell(row = r, column = lst[2]-1) # N-0 / Rating All the time
                    
                    if cell.value != '' and compare.value != '':
                        if cell.value > compare.value:
                            cell.fill = red_fill
                        elif cell.value >= 0.9*compare.value:
                            cell.fill = yellow_fill

    return


if __name__ == "__main__":
    # Add your debugging code here
    simname = "DC000"  # Provide a simulation name or adjust as needed
    main_option = "10"  # Adjust as needed
    time_start = "0070000"  # Adjust as needed
    time_end = "0080000"  # Adjust as needed
    option_select = "1"  # Adjust as needed
    text_input = "TRU_rating_template"  # Adjust as needed
    low_v = None  # Adjust as needed
    high_v = None  # Adjust as needed
    time_step = "1"  # Adjust as needed

    # Call your main function with the provided parameters
    main(simname, main_option, time_start, time_end, option_select, text_input, low_v, high_v, time_step)

