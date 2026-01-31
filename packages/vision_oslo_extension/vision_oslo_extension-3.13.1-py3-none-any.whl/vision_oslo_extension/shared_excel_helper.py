#=================================================================
# Created by: Jieming Ye
# Created on: Feb 2026
# Last Modified: Feb 2026
#=================================================================
# Copyright (c) 2026 [Jieming Ye]
#
# This Python source code is licensed under the 
# Open Source Non-Commercial License (OSNCL) v1.0
# See LICENSE for details.
#=================================================================
"""
Commonly used excel function helper

"""
#=================================================================
# VERSION CONTROL
# V1.0 (Jieming Ye) - Initial Version
#=================================================================
# Set Information Variable
# N/A
#=================================================================

from openpyxl import Workbook
from copy import copy

def export_result_sheet_from_wb(wb: Workbook, sheet_name: str, output_file: str):
    '''
    Docstring for export_result_sheet_from_wb

    :param wb: The source workbook object
    :param sheet_name: The name of the sheet to export
    :param output_file: The path to save the exported file
    '''
    src_ws = wb[sheet_name]

    dst_wb = Workbook()
    dst_ws = dst_wb.active
    dst_ws.title = sheet_name

    # Copy cells + styles
    for row in src_ws.iter_rows():
        for cell in row:
            new_cell = dst_ws[cell.coordinate]
            new_cell.value = cell.value

            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = cell.number_format
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

    # Column widths
    for col, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col].width = dim.width

    # Row heights
    for row, dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[row].height = dim.height

    # Merged cells
    for merged in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merged))

    # Freeze panes
    dst_ws.freeze_panes = src_ws.freeze_panes

    # Tables
    for table in src_ws.tables.values():
        dst_ws.add_table(copy(table))

    # Conditional formatting
    dst_ws.conditional_formatting = src_ws.conditional_formatting

    dst_wb.save(output_file)

    return
