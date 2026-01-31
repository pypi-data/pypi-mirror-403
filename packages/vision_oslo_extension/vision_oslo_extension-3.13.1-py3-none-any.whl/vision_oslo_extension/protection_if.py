#
# -*- coding: utf-8  -*-
#=================================================================
# Created by: Jieming Ye
# Created on: Feb 2024
# Last Modified: Feb 2024
#=================================================================
# Copyright (c) 2024 [Jieming Ye]
#
# This Python source code is licensed under the 
# Open Source Non-Commercial License (OSNCL) v1.0
# See LICENSE for details.
#=================================================================
"""
Pre-requisite: 
SimulationName.oof: default oslo output file after successfully run a simulation.
xxxxx.xlsx: excel spreadsheet with proper user settings.
Used Input:
simname: to locate the result file or file rename
time_start: to define the analysis start time
time_end: to define the analysis end time
option_select: to define the option selected in subfunction
text_input: to locate the excel file name
other input for oslo_extraction.py only.
Expected Output:
Updated excel spreadsheet with analysis result.
Description:
This script defines the process of doing incoming feeder protection assessment.
First, it reads the excel file and save the user input in a data frame called start_df. Then it find the related .d4 files and read information one by one and saved in a data frame list called d4dataframe. The calculation result is saved in a similar format data frame list called sumdataframe. It then doing some analysis and save the final result in an updated data frame. Final step is to save all information in the excel in a proper format. The whole process is easy to follow via reading the code.
The key manual updating part is get_result_range() function which depends on the desired output format in excel, followed by table_formatting() function where some reading needs manual input. The curve_plot() function is also need manual change if some new plots is desired in the future. The other process should be auto adjustable as long as the excel input format is followed.

"""
#=================================================================
# VERSION CONTROL
# V1.0 (Jieming Ye) - Initial Version
#=================================================================
# Set Information Variable
# N/A
#=================================================================


# vision_oslo_extension/excel_processing.py
import pandas as pd
import numpy as np
# import openpyxl
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, NamedStyle
from openpyxl.formatting.rule import CellIsRule,FormulaRule
from openpyxl.chart import ScatterChart,Reference,Series


# import vision_oslo
from vision_oslo_extension import oslo_extraction
from vision_oslo_extension.shared_contents import SharedMethods
from vision_oslo_extension.shared_excel_helper import export_result_sheet_from_wb


def main(simname, main_option, time_start, time_end, option_select, text_input, low_v, high_v, time_step):

    print("")
    print("Incoming Feeder Protection Assessment --->")
    print("")
    
    simname = simname
    excel_file = text_input + ".xlsx"

    if not SharedMethods.check_existing_file(excel_file):
        return False
    
    #option = "1" # Fully Restart, require oof,
    #option = "2" # Process only

    option = option_select # 1:Fully Restart, require oof, and list

    time_increment = 5
    start = 10

    space  = 5

    if option not in ["0","1","2"]:
        SharedMethods.print_message("ERROR: Error in protection_if.py. Please contact Support.","31")
        return False
    
    if option == "0":
        SharedMethods.print_message("ERROR: Please select an option to proceed.","31")
        return False

    result = start_reading_process(simname,excel_file)
    if result == False:
        return False
    else:
        feeder_list, type_list, p_type, p_setting, start_df = result

    # number of dataframes to create
    total = len(feeder_list)

    # check if want to go throught the feeder check process
    if option == "1":
        if not SharedMethods.check_oofresult_file(simname):
            return False
        if not one_stop_protection_extraction(simname, time_start, time_end, feeder_list, type_list):
            return False
        
    if option == "2":
        print("Checking essential d4 files and mxn files...")
        for feeder in feeder_list:
            filename = simname + "_" + feeder + ".osop.d4"
            if not SharedMethods.check_existing_file(filename):
                SharedMethods.print_message(f"ERROR: d4 file {feeder} do not exist. Select option 1 to proceed.","31")
                return False

    # process d4 file
    i_r_df,d4dataframe,sumdataframe = feeder_reading_process(simname,feeder_list,type_list,time_increment)
    
    # process assessment threshold table
    limit_r_df = protection_limit(simname, feeder_list,p_type,p_setting,time_increment)

    data_write_save(simname, excel_file,start,total,space,start_df,i_r_df,limit_r_df, \
                    d4dataframe,sumdataframe,feeder_list,type_list,time_increment)

    return True


def one_stop_protection_extraction(simname, time_start, time_end, feeder_list, type_list):

    # User defined time windows extraction
    time_start = SharedMethods.time_input_process(time_start,1)
    time_end = SharedMethods.time_input_process(time_end,1)

    if time_start == False or time_end == False:
        return False

    # process 1: feeder step output
    print("Extrating step output from Start Tab ... ")

    # processing the list
    for index, items in enumerate(feeder_list):
        #print(items)
        if type_list[index] == "SP":
            if not oslo_extraction.feeder_step_one(simname,items,time_start,time_end):
                SharedMethods.print_message(f"WARNING: Error for {items} will be ignored and process continued...","33")
        elif type_list[index] == "TR":
            if not oslo_extraction.tranx_step_one(simname,items,time_start,time_end):
                SharedMethods.print_message(f"WARNING: Error for {items} will be ignored and process continued...","33")
        else:
            SharedMethods.print_message(f"WARNING: {type_list[index]} is not a recongized type for {items}. Feeder applies by default.","33")
            if not oslo_extraction.feeder_step_one(simname,items,time_start,time_end):
                SharedMethods.print_message(f"WARNING: Error for {items} will be ignored and process continued...","33")
    
    return True

# write data to excel
def data_write_save(simname, excel_file,start,total,space,start_df,i_r_df,limit_r_df, \
                    d4dataframe,sumdataframe,feeder_list,type_list,time_increment):
    # write data to excel
    with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a',if_sheet_exists='overlay') as writer:

        # get currrent workbook to do process
        wb = writer.book

        print("Generate Result Page...")
        start_df.to_excel(writer, sheet_name="Result", index=False, startrow = start, startcol = 1)
        limit_r_df.to_excel(writer, sheet_name="Result", index=False, startrow = start, startcol = start_df.shape[1] + 1)
        #p_r_df.to_excel(writer, sheet_name="Result", index=False, startrow = start, startcol = start_df.shape[1] + 1, float_format="%.2f")

        start_df.to_excel(writer, sheet_name="Result", index=False, startrow = start + total + space, startcol = 1)
        i_r_df.to_excel(writer, sheet_name="Result", index=False, startrow = start + total + space, startcol = start_df.shape[1] + 1)

        # plot graph in new sheet
        #limit = limit_r_df.iloc[index].tolist()
        print("Generate Plot Page...")
        curve_plot(sumdataframe,limit_r_df,feeder_list,wb,time_increment)
        
        # Write each DataFrame to a different sheet in the Excel file
        print("Writing individual incoming feeder...")
        for index, dflist in enumerate(d4dataframe):
            print(f"Saving {feeder_list[index]}...")
            if type_list[index] == 'TR':
                dflist.insert(dflist.columns.get_loc('2nd_I_angle') + 1, 'New_C_1', np.nan)
                sumdataframe[index].insert(sumdataframe[index].columns.get_loc('2nd_I_angle')+1,'New_C_1', np.nan)
            
            else:
                dflist.insert(dflist.columns.get_loc('I_angle') + 1, 'New_C_1', np.nan)
                dflist.insert(dflist.columns.get_loc('New_C_1') + 1, 'New_C_2', np.nan)
                dflist.insert(dflist.columns.get_loc('New_C_2') + 1, 'New_C_3', np.nan)
                dflist.insert(dflist.columns.get_loc('New_C_3') + 1, 'New_C_4', np.nan)
                sumdataframe[index].insert(sumdataframe[index].columns.get_loc('I_angle')+1,'New_C_1', np.nan)
                sumdataframe[index].insert(sumdataframe[index].columns.get_loc('New_C_1')+1,'New_C_2', np.nan)
                sumdataframe[index].insert(sumdataframe[index].columns.get_loc('New_C_2')+1,'New_C_3', np.nan)
                sumdataframe[index].insert(sumdataframe[index].columns.get_loc('New_C_3')+1,'New_C_4', np.nan)

            sumdataframe[index].to_excel(writer, sheet_name=feeder_list[index], index=False, startrow = 0)

            dflist.to_excel(writer, sheet_name=feeder_list[index], index=False, startrow = sumdataframe[index].shape[0]+2)
            
        # Calculate the Excel range for each DataFrame
        range_list = get_result_range(start,total,space) # range list as [power data range, current data range]

        # Output result dashboard only
        print("Exporting Result Dashboard...")
        dashboard_name = "0_"+simname+"_result_if_current_loading.xlsx"
        export_result_sheet_from_wb(wb,"Result",dashboard_name)
        
        # # table formatting
        table_formatting(simname,wb,range_list)

    return

# plot the IDMT Protection Curve
def curve_plot(sumdataframe,limit_r_df,feeder_list,wb,time_increment):
    # limit as list

    sheet = wb.create_sheet(title="Plot")

    row = 1
    start = 2
    p_row = 1
    time = [time_increment, 5, 10, 15, 20, 25, 30, 35, 40, 60, 120, 600, 1200, 1800]
    

    sheet.cell(row,start).value = 'Time'

    for index, t in enumerate(time):
        sheet.cell(row,start+index+1).value = t


    for index, feeder in enumerate(feeder_list):
        row = row + 1
        sheet.cell(row,start).value = feeder

        row = row + 1
        sheet.cell(row,start).value = "Limit"
        limit = limit_r_df.iloc[index].tolist()
        for index1,lmt in enumerate(limit):
            sheet.cell(row,start+index1+1).value = lmt

        row = row + 1
        sheet.cell(row,start).value = "Result"
        df = sumdataframe[index]
        start_index = df.columns.get_loc('I_Inst')
        result = df.iloc[0, start_index:].tolist()
        for index1, re in enumerate(result):
            sheet.cell(row,start+index1+1).value = re

    # prepare the plot
    row = 1
    xvalues = Reference(sheet, min_col=start+1,max_col = start+14, min_row=row)

    for index, feeder in enumerate(feeder_list):


        row = row + 3
        # Create a scatter plot
        # chart = writer.book.add_chart({'type': 'scatter'})
        chart = ScatterChart('smoothMarker')
        chart.title = f'Incoming Feeder Protection Curve for {feeder}'
        chart.x_axis.title = "Time Window (s)"
        chart.y_axis.title = "Maximum RMS Current (A)"

        # Set chart size
        chart.height = 10  # Adjust width as needed
        chart.width = 25  # Adjust height as needed

        #xvalues = Reference(sheet, min_col=start+1,max_col = start+14, min_row=row)

        yvalues = Reference(sheet, min_col=start+1, max_col = start+14, min_row=row)
        series = Series(yvalues,xvalues,title_from_data=False,title = 'Maximum RMS Current')
        series.marker.symbol = 'circle'   # Choose the marker type you prefer
        series.marker.size = 5   # Choose the marker type you prefer
        # Set x-values directly for both series
        #series.data_points = [(x_val, None) for x_val in xvalues]
        chart.series.append(series)

        yvalues = Reference(sheet, min_col=start+1, max_col = start+14, min_row=row-1)
        series = Series(yvalues,xvalues,title_from_data=False,title = 'Protection Curve')
        series.marker.symbol = 'circle'   # Choose the marker type you prefer
        series.marker.size = 5   # Choose the marker type you prefer
        chart.series.append(series)

        chart.x_axis.scaling.min = 0       
        chart.x_axis.scaling.max = 120

        chart.legend.position = 't' # top legend position

        sheet.add_chart(chart,f"R{p_row}")
        p_row = p_row + 20

    return

# read the start tab and collect informaiton
def start_reading_process(simname, excel_file):
    feeder_list = []
    type_list = []
    p_type = []
    p_setting = []
    
    # File check
    print("Check Excel and Deleting Existing Contents ...")

    try:
        wb = load_workbook(excel_file)
        # Delete all sheets except 'Start'
        for sheet_name in wb.sheetnames:
            if sheet_name == 'Start':
                table_list = check_table_list(wb[sheet_name])
                
                if table_list is not False:
                    for row in table_list:
                        feeder_list.append(row[1])
                        type_list.append(row[2])
                        p_type.append(row[3])
                        if row[3] == "DT":
                            p_setting.append(row[4])
                        else:
                            p_setting.append([row[4],row[5]])

                else:
                    return False
            else:
                wb.remove(wb[sheet_name])
        
        # Save changes
        wb.save(excel_file)

    except Exception as e:
        SharedMethods.print_message(f"ERROR: (Close the Excel file and Start Again) Error: {e}","31")
        return False
        
    # read start up information 
    columns = ["Incoming Feeder","OSLO ID","OSLO Type","Protection Type","Pickup Current (A)","Time Multiplier (s)"]
    start_df = pd.DataFrame(table_list,columns=columns)
    # get supply list oslo id
    
    # print(feeder_list)
    # print(type_list)
    # print(p_type)
    # print(p_setting)
    
    # check duplication in OSLO id
    if not SharedMethods.find_duplicates(feeder_list): return False
    
    return feeder_list, type_list, p_type, p_setting, start_df

# check table list on start page
def check_table_list(sheet):
    print("Reading Configuration Setting ...")
    table_row = 11
    table_start_column = 1
    table_end_column = 6
    # create node and feeder list
    table_list = []

    # check feeder
    index = table_row
    column = table_start_column
    if sheet.cell(row=index, column=column+1).value is not None:
        while True:
            row_data = []
            for temp in range(table_start_column,table_end_column + 1):
                row_data.append(sheet.cell(row=index, column=temp).value)
            table_list.append(row_data)
            index += 1
            
            check = sheet.cell(row=index, column=column+1).value
            if check is None:
                break
    else:
        SharedMethods.print_message("ERROR: Wrong data format. No information at B11","31")
        return False

    return table_list

# create table 1 and table 2
def current_result(supply_list):
    columns = ['Instantanous','05 seconds','10 seconds','15 seconds', '20 seconds', '25 seconds','30 seconds','35 seconds','40 seconds', \
               '60 seconds', '120 seconds', '10 min','20 min','30 min']
    rows = []

    for feeder in supply_list:
        rows.append(feeder)
    df = pd.DataFrame(index=rows,columns=columns)

    return df

#create protection limit dataframe
def protection_limit(simname, feeder_list,p_type,p_setting,time_increment):
    limit_r_df = current_result(feeder_list) # create limit dataframe

    columns = limit_r_df.columns # get columns

    time_interval = [time_increment, 5, 10, 15, 20, 25, 30, 35, 40, 60, 120, 600, 1200, 1800]

    for index, item in enumerate(feeder_list):
        if p_type[index] == "DT":
            result = float(p_setting[index])
            for column in columns: # exclude the first column
                limit_r_df.at[item,column] = result
        
        if p_type[index] == "IDMT":
            pickup = float(p_setting[index][0])
            tm = float(p_setting[index][1])

            for index1, column in enumerate(columns):
                time = time_interval[index1]

                result = (((0.14 * tm / time) + 1) ** (1 / 0.02)) * pickup # IDMT threshold Calculation formula based on standard inverse curve

                limit_r_df.at[item,column] = result

    return limit_r_df

# read individual d4 file
def feeder_reading_process(simname, feeder_list,type_list,time_increment):

    i_r_df = current_result(feeder_list)

    # create dataframe
    d4dataframe = []
    sumdataframe = []

    for index, feeder in enumerate(feeder_list):
        print(f"Processing {feeder} ...")
        filename = simname + "_" + feeder +".osop.d4"
        if type_list [index] == "TR":
            delimiter = '\\s+'
            columns = ["TransID","Type","TR_Type","Time","P_inst","Q_inst","Voltage","V_angle","1st_Current","1st_I_angle","2nd_Current","2nd_I_angle"]
            dtype_mapping = {"Time": str,}
            df = pd.read_csv(filename, delimiter=delimiter, names = columns, skiprows = 11,dtype = dtype_mapping) 
            # Extracting parts from the string and formatting the 'Time' column
            df['Time'] = df['Time'].apply(lambda x: f"{int(x[1:3]):02d}:{int(x[3:5]):02d}:{int(x[5:]):02d}")
            # data frame process
            df = d4_file_tr_process(df,time_increment)
            df_sum = d4_find_max(df,f_type = 2)

            i_r_df = current_update(i_r_df, df_sum, feeder,f_type = 2)
        
        #elif type_list[index] == "SP":
        else:
            delimiter = '\\s+'
            columns = ["FeederID","Type","Time","P_inst","Q_inst","Voltage","V_angle","Current","I_angle"]
            dtype_mapping = {"Time": str,}
            df = pd.read_csv(filename, delimiter=delimiter, names = columns, skiprows = 11,dtype = dtype_mapping) 
            # Extracting parts from the string and formatting the 'Time' column
            df['Time'] = df['Time'].apply(lambda x: f"{int(x[1:3]):02d}:{int(x[3:5]):02d}:{int(x[5:]):02d}")
            # data frame process
            df = d4_file_sp_process(df,time_increment)
            # calculate sum value
            df_sum = d4_find_max(df,f_type = 1)
            #update summary result
            i_r_df = current_update(i_r_df, df_sum,feeder,f_type = 1)

        
        # calculate sum value
        sumdataframe.append(df_sum)
        d4dataframe.append(df)
        
    return i_r_df, d4dataframe, sumdataframe

# doing invididual d4 file process and calculatoin for supply points
def d4_file_sp_process(df, time_increment):

    df['I_Inst'] = df['Current'] # Current Duplicate

    for time_interval in [5, 10, 15, 20, 25, 30, 35, 40, 60, 120, 600, 1200, 1800]:
        column_name = f'I_{time_interval}s_RMS'
        df[column_name] = calculate_rolling_rms(df, 'Current', time_interval, time_increment)

    return df

# doing invididual d4 file process and calculatoin for transformer
def d4_file_tr_process(df, time_increment):

    df['I_Inst'] = df['2nd_Current'] # Current Duplicate

    for time_interval in [5, 10, 15, 20, 25, 30, 35, 40, 60, 120, 600, 1200, 1800]:
        column_name = f'I_{time_interval}s_RMS'
        df[column_name] = calculate_rolling_rms(df, '2nd_Current', time_interval, time_increment)

    return df

# calculate rms
def calculate_rolling_rms(data, column, window_size, time_increment):
    return data[column].rolling(window=int(window_size / time_increment)).apply(lambda x: np.sqrt(np.mean(x**2)), raw=True)

# find maximum value and time of each d4 file
def d4_find_max(df,f_type):
    # # Insert two empty rows above the first row
    # df = pd.concat([pd.DataFrame(index=range(2)), df], ignore_index=True)
    sum_df = pd.DataFrame(columns=df.columns,index=range(4))
    sum_df.iloc[0, 0] = "Maximum Value"
    sum_df.iloc[1, 0] = "Maximum Value at Time"
    sum_df.iloc[2, 0] = "Minimum Value"
    sum_df.iloc[3, 0] = "Minimum Value at Time"

    if f_type == 1:
        start = 9
    
    if f_type == 2:
        start = 12

    if df.empty:
        sum_df.iloc[0, 1] = "DATA FOR THIS FEEDER IS NOT AVAILABLE"

    else:
        for column in df.columns[start:]:
            if df[column].dtype in [int, float]:
                max_value = df[column].max()
                time_of_max = df.loc[df[column].idxmax(), 'Time']
                sum_df.at[0, column] = max_value
                sum_df.at[1, column] = time_of_max
                min_value = df[column].min()
                time_of_min = df.loc[df[column].idxmin(), 'Time']
                sum_df.at[2, column] = min_value
                sum_df.at[3, column] = time_of_min

    return sum_df

# update current based on output of d4 process
def current_update(target,df_sum,feeder,f_type):
    if f_type == 1:
        index = 9
    if f_type == 2:
        index = 12
    #index = 25 # from 25th column
    for column in target.columns:
        target.at[feeder,column] = df_sum.iloc[0,index]
        index = index + 1
    return target

# covert mxn time format to proper time format (ignore DAY)
def convert_time_format(time_str):

    if ':' not in time_str:
        return time_str  # Return the original string if ':' is not present

    # Split the time string into components
    days, time = time_str.split(':')
    
    # Split the time component into hours, minutes, and seconds
    hours, minutes, seconds = map(int, time.split('.'))

    # Convert to the desired format
    converted_time = f'{hours:02d}:{minutes:02d}:{seconds:02d}'

    return converted_time

def get_result_range(start,total,space):

    """
    Generates a list of range strings for different sections in a spreadsheet.

    Parameters:
    - start (int): Starting row index.
    - total (int): Total number of rows.
    - space (int): Space between sections.
    - fsnumber (int): Placeholder for fsnumber.

    Returns:
    List of range strings for different sections.
    """

    range_list = []
    # 0: table frame range (limit, result)
    # 1: table data range (limit, result)
    # 2: title range (2nd row)
    # 3: title range (1st row simulation info section1,2)
    # 4: title range (1st row current limit)
    # 5: title range (1st row current)
    
    # 6: threshold range (compare cell point)

    # 7: conditonal formatting range NPS (range, 1min, 10min, 30min)
    # 8: conditonal formatting range voltage ( min, max)


    
    # 0
    result_range = [f"B{start + 1}:U{start + total + 1}",f"B{start + total + space + 1}:U{start + total*2 + space + 1}"]
    range_list.append(result_range)
    # 1
    data_range = [f"H{start + 2}:U{start + total + 1}",f"H{start + total + space + 2}:U{start + total*2 + space + 1}"]
    range_list.append(data_range)
   
    # 2
    title_range = [f"B{start+1}:U{start+1}",f"B{start + total + space + 1}:U{start + total + space + 1}"]
    range_list.append(title_range)
    
    #3 4 5
    add_range_1 = [f"B{start}:G{start}",f"B{start + total + space}:G{start + total + space}"]
    range_list.append(add_range_1)

    add_range_2 = [f"H{start}:U{start}"]
    range_list.append(add_range_2)
    add_range_3 = [f"H{start + total + space}:U{start + total + space}"]
    range_list.append(add_range_3)

    #6 7 8
    condi_point= [f"H{start + 2}"]
    range_list.append(condi_point)

    return range_list

# Result table table formating
def table_formatting(simname,wb,range_list):
    # format the result table

    print("Formatting Process ...")
    # wb = load_workbook(excel_file)
    print("Information Collection ...")
    sheet = wb["Start"]
    project_name = sheet['B2'].value
    feeding_desp = sheet['B4'].value
    modeller = sheet['B5'].value
    date = sheet['B6'].value

    # Result Tab process
    sheet = wb["Result"]
    sheet['B2'].value = "Project Name:"
    sheet['C2'].value = project_name
    sheet['B3'].value = "Simulation Name:"
    sheet['C3'].value = simname
    sheet['B4'].value = "Feeding Arrangement:"
    sheet['C4'].value = feeding_desp
    sheet['B5'].value = "Result Created by:"
    sheet['C5'].value = modeller
    sheet['B6'].value = "Result Created at:"
    sheet['C6'].value = datetime.now().strftime("%d-%m-%Y %H:%M")

    sheet['H4'].value = "Protection Type:"
    sheet['H5'].value = "DT"
    sheet['H6'].value = "IDMT"

    sheet['K4'].value = "Threshold Calculation:"
    sheet['K5'].value = "Pickup Current --> i.e. Trip Current"
    sheet['K6'].value = "(((0.14*[time_multiplier]/[time_window])+1)^(1/0.02))*[pickup_current]"


    apply_border(sheet,range_list)
    apply_title(sheet,range_list)
    
    print("Apply Numbering Format ...")
    # Define a custom style for formatting with two decimal places
    for range_name in range_list[1]:
        for row in sheet[range_name]:
            for cell in row:
                #cell.style = NamedStyle(name='decimal_style', number_format='0.00')
                cell.number_format = '0.00'

    print("Apply Font ...")
    # Shade the range B11:Z11 with a light gray background color
    for range_name in range_list[2]:
        for row in sheet[range_name]:
            for cell in row:
                cell.font = Font(italic=True, size = 10)
    
    print("Apply Shading ...")
    # Shade the range B11:Z11 with a light gray background color
    for range_name in range_list[2]+range_list[3]+range_list[4]+range_list[5]:
        for row in sheet[range_name]:
            for cell in row:
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    condtional_formating(sheet,range_list)

    print("Apply Column Length ...")
    # Auto-adjust the width of column B based on the content in B2 to B6
    max_length = max(len(str(sheet.cell(row=i, column=2).value)) for i in range(2, 7))
    sheet.column_dimensions['B'].width = max_length + 2  # Add a little extra space

    # Auto-size columns after applying formatting
    for col_letter in ["C","D","E","F","G","H"]:
        sheet.column_dimensions[col_letter].auto_size = True

    print("Saving...")

    return

def apply_border(sheet,range_list):
    print("Apply Border ...")
    for range_name in range_list[0]+range_list[2]+range_list[3]+range_list[4]+range_list[5]:
        for row in sheet[range_name]:
            for cell in row:
                # Apply border to all sides of the cell
                cell.border = Border(left=Side(border_style='thin'),
                                    right=Side(border_style='thin'),
                                    top=Side(border_style='thin'),
                                    bottom=Side(border_style='thin'))

                # Align cell content to the middle
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    return

def apply_title(sheet,range_list):

    print("Apply Title ...")
    for range_name in range_list[3]:
        sheet.merge_cells(range_name)
        part = range_name.split(":")
        cell = sheet[part[0]]
        cell.value = "Site Information Summary"
        cell.font = Font(bold=True)

    sheet.merge_cells(range_list[4][0])
    part = range_list[4][0].split(":")
    cell = sheet[part[0]]
    cell.value = "Trip Current Threshold (A)"
    cell.font = Font(bold=True)

    sheet.merge_cells(range_list[5][0])
    part = range_list[5][0].split(":")
    cell = sheet[part[0]]
    cell.value = "Maximum Incoming Feeder RMS Current (A)"
    cell.font = Font(bold=True)

    return

def condtional_formating(sheet,range_list):
    print("Apply Conditional Formatting ...")
    # Compare values in columns I and J for each row and shade accordingly
    # set the pattern fill
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red (0% S)
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 255,152,51 (80% S)
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green (00% S)

    # power
    #yellow_line = str(float(sheet[range_list[6][0]])*0.9)
    sheet.conditional_formatting.add(range_list[1][1],CellIsRule(operator = 'lessThan',formula=[range_list[6][0]+'-300'],fill=green_fill))
    sheet.conditional_formatting.add(range_list[1][1],CellIsRule(operator = 'greaterThanOrEqual',formula=[range_list[6][0]],fill=red_fill))
    sheet.conditional_formatting.add(range_list[1][1],CellIsRule(operator = 'between',formula=[range_list[6][0]+'-300',range_list[6][0]],fill=yellow_fill))

    return


if __name__ == "__main__":
    # Add your debugging code here
    simname = "StraightLine1"  # Provide a simulation name or adjust as needed
    main_option = "1"  # Adjust as needed
    time_start = "0070000"  # Adjust as needed
    time_end = "0080000"  # Adjust as needed
    option_select = "1"  # Adjust as needed
    text_input = "output"  # Adjust as needed
    low_v = None  # Adjust as needed
    high_v = None  # Adjust as needed
    time_step = None  # Adjust as needed

    # Call your main function with the provided parameters
    main(simname, main_option, time_start, time_end, option_select, text_input, low_v, high_v, time_step)

