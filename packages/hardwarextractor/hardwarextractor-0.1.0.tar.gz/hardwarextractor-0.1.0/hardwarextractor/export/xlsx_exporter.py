"""XLSX exporter for HardwareXtractor."""

from __future__ import annotations

from pathlib import Path
from typing import TYPE_CHECKING

from hardwarextractor.export.base import BaseExporter, ExportResult

if TYPE_CHECKING:
    from hardwarextractor.engine.ficha_manager import FichaManager


# Origin colors (ARGB format for openpyxl)
ORIGIN_COLORS = {
    "OFICIAL": "FF90EE90",       # Light green
    "CATÁLOGO": "FF98FB98",      # Pale green
    "REFERENCIA": "FFFFE4B5",    # Light orange
    "CALCULADO": "FFADD8E6",     # Light blue
    "DESCONOCIDO": "FFD3D3D3",   # Light gray
    "": "FFFFFFFF",
}


class XLSXExporter(BaseExporter):
    """Export ficha to XLSX format with formatting and colors."""

    @property
    def format(self) -> str:
        return "xlsx"

    def export(self, ficha_manager: "FichaManager", path: Path) -> ExportResult:
        """Export the ficha to XLSX.

        Args:
            ficha_manager: The ficha manager with data
            path: Output file path

        Returns:
            ExportResult with export information
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            return ExportResult(
                path=Path(path),
                format="xlsx",
                rows=0,
                success=False,
                error="openpyxl is not installed. Install with: pip install openpyxl",
            )

        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "Ficha Técnica"

        header_info = self._get_header_info(ficha_manager)

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4472C4")
        warning_font = Font(bold=True, color="FF6600")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Header with metadata
        ws["A1"] = f"Ficha Técnica - {header_info['date']}"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:H1")

        ws["A2"] = f"Componentes: {header_info['component_count']}"

        start_row = 3

        # Warning banner if has reference data
        if header_info["has_reference"]:
            ws[f"A{start_row}"] = "ADVERTENCIA: Esta ficha contiene datos no oficiales (REFERENCE)"
            ws[f"A{start_row}"].font = warning_font
            ws.merge_cells(f"A{start_row}:H{start_row}")
            start_row += 1

        start_row += 1

        # Headers
        headers = ["Sección", "Campo", "Valor", "Unidad", "Origen", "Fuente", "URL"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        rows = ficha_manager.get_export_rows()
        current_section = None

        for i, row in enumerate(rows):
            data_row = start_row + 1 + i

            # Section (only show if changed)
            section = row["section"]
            section_display = section if section != current_section else ""
            current_section = section

            values = [
                section_display,
                row["field"],
                row["value"],
                row["unit"],
                row["origen"],
                row["source_name"],
                self._truncate_url(row["source_url"], 60),
            ]

            for col, value in enumerate(values, 1):
                cell = ws.cell(row=data_row, column=col, value=value)
                cell.border = border

                # Color the Origen column
                if col == 5:  # Origen column
                    origen_value = str(value) if value else ""
                    color = ORIGIN_COLORS.get(origen_value, "FFFFFFFF")
                    cell.fill = PatternFill("solid", fgColor=color[2:])  # Remove FF prefix

        # Adjust column widths
        column_widths = {
            "A": 20,  # Sección
            "B": 25,  # Campo
            "C": 20,  # Valor
            "D": 10,  # Unidad
            "E": 15,  # Origen
            "F": 15,  # Fuente
            "G": 50,  # URL
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Save
        wb.save(path)

        return ExportResult(
            path=path,
            format="xlsx",
            rows=len(rows),
            success=True,
        )
